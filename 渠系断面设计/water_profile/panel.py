# -*- coding: utf-8 -*-
"""
推求水面线面板 —— QWidget 版本

功能：
1. 基础设置（渠道名称、级别、起始水位、流量等）
2. 节点数据表格（输入/编辑/从批量计算导入）
3. 调用核心计算引擎进行水面线推求
4. 结果展示（结果表格 + 详细过程）
5. 导出Excel/Word

与批量计算模块的数据交互：
- 通过SharedDataManager从批量计算结果导入节点数据
- 导入后自动填充断面参数（底宽、水深、糙率等）
"""

import sys
import os
import math
import re
import datetime

_pkg_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# 推求水面线模块路径
_water_profile_dir = os.path.join(_pkg_root, '推求水面线')
if _water_profile_dir not in sys.path:
    sys.path.insert(0, _water_profile_dir)

# 渠系建筑物断面计算路径
_calc_dir = os.path.join(_pkg_root, '渠系建筑物断面计算')
if _calc_dir not in sys.path:
    sys.path.insert(0, _calc_dir)

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QGroupBox,
    QSplitter, QFrame, QTabWidget, QTextEdit, QFileDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QAbstractItemView, QScrollArea, QGridLayout, QFormLayout,
    QDialog, QDialogButtonBox, QToolTip, QCheckBox, QApplication
)
from PySide6.QtCore import Qt, QByteArray, Signal, QTimer, QRect, QPoint, QEvent, QObject
from PySide6.QtGui import QFont, QColor, QPixmap, QImage, QShortcut, QKeySequence, QCursor

from qfluentwidgets import (
    PushButton, PrimaryPushButton, LineEdit, ComboBox,
    InfoBar, InfoBarPosition
)

from 渠系断面设计.frozen_table import FrozenColumnTableWidget
from 渠系断面设计.styles import P, S, W, E, BG, CARD, BD, T1, T2, auto_resize_table, CollapsibleGroupBox, fluent_info, fluent_question, DIALOG_STYLE
from 渠系断面设计.export_utils import (
    WORD_EXPORT_AVAILABLE, ask_open_file,
    create_styled_doc, doc_add_h1, doc_add_h2, doc_add_body,
    doc_render_calc_text, doc_add_param_table, doc_add_result_table,
    doc_add_styled_table, doc_add_table_caption,
    create_engineering_report_doc, doc_add_eng_h, doc_add_eng_body,
    doc_render_calc_text_eng, update_doc_toc_via_com,
)
from 渠系断面设计.report_meta import (
    ExportConfirmDialog, build_calc_purpose, REFERENCES_BASE, load_meta
)
from 渠系断面设计.structure_type_selector import StructureTypeSelector
from 渠系断面设计.batch.panel import format_station_display, parse_station_input
from utils.pressure_pipe_result_helpers import (
    make_pressure_pipe_identity,
    empty_pressure_pipe_calc_records,
    normalize_pressure_pipe_calc_records,
    format_pressure_pipe_record_detail,
    append_pressure_pipe_calc_batch_text,
)

# 核心计算引擎
try:
    from models.data_models import ChannelNode, ProjectSettings
    from models.enums import StructureType, InOutType
    from core.calculator import WaterProfileCalculator
    CALCULATOR_AVAILABLE = True
except ImportError as _e:
    print(f"[水面线] 核心计算引擎加载失败: {_e}")
    CALCULATOR_AVAILABLE = False

# 共享数据管理器
try:
    from shared.shared_data_manager import get_shared_data_manager
    SHARED_DATA_AVAILABLE = True
except ImportError:
    SHARED_DATA_AVAILABLE = False

# 配置常量
try:
    from config.constants import (
        STRUCTURE_TYPE_OPTIONS, CHANNEL_LEVEL_OPTIONS,
        DEFAULT_ROUGHNESS, DEFAULT_SIPHON_ROUGHNESS, DEFAULT_TURN_RADIUS, DEFAULT_SIPHON_TURN_RADIUS_N, DEFAULT_GATE_HEAD_LOSS,
        TRANSITION_FORM_OPTIONS, SIPHON_TRANSITION_FORM_OPTIONS,
        TRANSITION_ZETA_COEFFICIENTS, SIPHON_TRANSITION_ZETA_COEFFICIENTS
    )
except ImportError:
    STRUCTURE_TYPE_OPTIONS = [
        "明渠-梯形", "明渠-矩形", "明渠-圆形", "明渠-U形",
        "渡槽-U形", "渡槽-矩形",
        "隧洞-圆形", "隧洞-圆弧直墙型", "隧洞-马蹄形Ⅰ型", "隧洞-马蹄形Ⅱ型",
        "矩形暗涵", "倒虹吸", "有压管道", "分水闸", "分水口", "节制闸", "泄水闸",
    ]
    CHANNEL_LEVEL_OPTIONS = ["总干渠", "总干管", "分干渠", "分干管", "干渠", "干管", "支渠", "支管", "分支渠", "分支管"]
    DEFAULT_ROUGHNESS = 0.014
    DEFAULT_SIPHON_ROUGHNESS = 0.014
    DEFAULT_TURN_RADIUS = 100.0
    DEFAULT_SIPHON_TURN_RADIUS_N = 3.0
    DEFAULT_GATE_HEAD_LOSS = 0.1
    TRANSITION_FORM_OPTIONS = ["曲线形反弯扭曲面", "直线形扭曲面", "圆弧直墙", "八字形", "直角形"]
    SIPHON_TRANSITION_FORM_OPTIONS = ["反弯扭曲面", "直线扭曲面", "1/4圆弧", "方头型"]
    TRANSITION_ZETA_COEFFICIENTS = {
        "进口": {"曲线形反弯扭曲面": 0.1, "圆弧直墙": 0.2, "八字形": 0.3, "直角形": 0.4},
        "出口": {"曲线形反弯扭曲面": 0.2, "圆弧直墙": 0.5, "八字形": 0.5, "直角形": 0.75},
    }
    SIPHON_TRANSITION_ZETA_COEFFICIENTS = {
        "进口": {"反弯扭曲面": 0.10, "直线扭曲面": 0.20, "1/4圆弧": 0.15, "方头型": 0.30},
        "出口": {"反弯扭曲面": 0.20, "直线扭曲面": 0.40, "1/4圆弧": 0.25, "方头型": 0.75},
    }

# 节点表列定义（与原版Tkinter ALL_COLUMNS保持完全一致的列顺序）
# 可编辑列索引集合（基础输入0-7 + 水力输入20-26 + 预留/过闸/倒虹吸或有压管道损失36,37,38）
EDITABLE_COLS = set(range(8)) | {20, 21, 22, 23, 24, 25, 26, 36, 37, 38}
# 第一行（水位起点）锁定的水头损失列：初始水位是用户输入的定值，不受水头损失影响
FIRST_ROW_LOCKED_LOSS_COLS = {36, 37, 38}

NODE_ALL_HEADERS = [
    # 基础输入列 (0-7) — 对应Tkinter INPUT_COLUMNS
    "流量段", "建筑物名称", "结构形式", "进出口判断", "IP",
    "X", "Y", "转弯半径",
    # 几何结果列 (8-19) — 对应Tkinter GEOMETRY_RESULT_COLUMNS
    "转角", "切线长", "弧长", "弯道长度", "IP直线间距",
    "IP点桩号", "弯前BC", "里程MC", "弯末EC",
    "复核弯前长度", "复核弯后长度", "复核总长度",
    # 水力输入列 (20-26) — 对应Tkinter HYDRAULIC_INPUT_COLUMNS
    "底宽B", "直径D", "半径R", "边坡系数m", "糙率n", "底坡1/i", "流量Q设计",
    # 水力结果列 (27-31) — 对应Tkinter HYDRAULIC_RESULT_COLUMNS
    "水深h设计", "过水断面面积A", "湿周X", "水力半径R", "流速v设计",
    # 水头损失列 (32-40) — 对应Tkinter HEAD_LOSS_COLUMNS
    "渐变段长度L", "渐变段水头损失", "弯道水头损失", "沿程水头损失",
    "预留水头损失", "过闸水头损失", "倒虹吸/有压管道水头损失",
    "总水头损失", "累计总水头损失",
    # 高程列 (41-43) — 对应Tkinter ELEVATION_COLUMNS
    "水位", "渠底高程", "渠顶高程",
]

# 导出Excel时使用的表头（与NODE_ALL_HEADERS一致）
NODE_EXPORT_HEADERS = NODE_ALL_HEADERS

# 节点数据表工具栏布局预设：
# compact（紧凑）/ balanced（平衡，默认）/ comfortable（宽松）
NODE_TOOLBAR_LAYOUT_PRESET = "balanced"


# ================================================================
# 渐变段参考系数表对话框（表K.1.2 + 表L.1.2）
# ================================================================

# 表K.1.2 数据
_K12_HEADERS = ["渐变段形式", "示意图", "进口ξ₁", "出口ξ₂"]
_K12_ROWS = [
    ["曲线形反弯扭曲面", "", "0.10", "0.20"],
    ["直线形扭曲面", "", "θ₁=15°~37°；ξ₁=0.05~0.30", "θ₂=10°~17°；ξ₂=0.30~0.50"],
    ["圆弧直墙", "", "0.20", "0.50"],
    ["八字形", "", "0.30", "0.50"],
    ["直角形", "", "0.40", "0.75"],
]

# 表L.1.2 数据
_L12_HEADERS = ["渐变段形式", "ξ₁", "ξ₂", "适用条件"]
_L12_ROWS = [
    ["反弯扭曲面", "0.10", "0.20", "θ₁,θ₂均≤12.5°"],
    ["1/4圆弧", "0.15", "0.25", "θ₁,θ₂均≤12.5°"],
    ["方头型", "0.30", "0.75", "θ₁,θ₂均≤12.5°"],
    ["直线扭曲面", "0.05~0.30", "0.30~0.50", "θ₁=15°~37°，θ₂=10°~17°"],
]


class TransitionReferenceDialog(QDialog):
    """渐变段参考系数表对话框（表K.1.2 + 表L.1.2）"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("渐变段局部损失系数参考表")
        self.setMinimumSize(780, 620)
        self.resize(820, 680)
        self._pixmap_refs = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # ── 表K.1.2 ──
        k12_label = QLabel("表K.1.2  进、出口水头损失系数（渡槽/隧洞渐变段）")
        k12_label.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        layout.addWidget(k12_label)

        self.k12_table = QTableWidget(len(_K12_ROWS), len(_K12_HEADERS))
        self.k12_table.setHorizontalHeaderLabels(_K12_HEADERS)
        self.k12_table.verticalHeader().setVisible(False)
        self.k12_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.k12_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.k12_table.horizontalHeader().setStretchLastSection(False)
        self.k12_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 加载示意图缩略图
        thumb_map = self._load_k12_thumbnails()

        for r, row_data in enumerate(_K12_ROWS):
            for c, val in enumerate(row_data):
                if c == 1:
                    # 示意图列
                    if r in thumb_map:
                        img_label = QLabel()
                        img_label.setPixmap(thumb_map[r])
                        img_label.setAlignment(Qt.AlignCenter)
                        img_label.setCursor(Qt.PointingHandCursor)
                        img_label.setToolTip("点击放大查看")
                        form_name = row_data[0]
                        img_label.mousePressEvent = lambda e, fn=form_name: self._show_k12_image(fn)
                        self.k12_table.setCellWidget(r, c, img_label)
                    else:
                        item = QTableWidgetItem("(无图)")
                        item.setTextAlignment(Qt.AlignCenter)
                        self.k12_table.setItem(r, c, item)
                else:
                    item = QTableWidgetItem(val)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.k12_table.setItem(r, c, item)

        for r in range(len(_K12_ROWS)):
            self.k12_table.setRowHeight(r, 60 if r in thumb_map else 32)

        self.k12_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        k12_h = sum(60 if r in thumb_map else 32 for r in range(len(_K12_ROWS))) + self.k12_table.horizontalHeader().height() + 4
        self.k12_table.setFixedHeight(k12_h)
        layout.addWidget(self.k12_table)

        # K.1.2 注释
        k12_note = QLabel("注：表中 θ₁ 表示进口渐变段水面收缩角；θ₂ 表示出口渐变段水面扩散角。点击示意图可放大查看。")
        k12_note.setWordWrap(True)
        k12_note.setStyleSheet("color: #555; font-size: 12px;")
        layout.addWidget(k12_note)

        # ── 表L.1.2 ──
        l12_label = QLabel("表L.1.2  渐变段局部损失系数表（倒虹吸渐变段）")
        l12_label.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        layout.addWidget(l12_label)

        self.l12_table = QTableWidget(len(_L12_ROWS), len(_L12_HEADERS))
        self.l12_table.setHorizontalHeaderLabels(_L12_HEADERS)
        self.l12_table.verticalHeader().setVisible(False)
        self.l12_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.l12_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.l12_table.horizontalHeader().setStretchLastSection(False)
        self.l12_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        for r, row_data in enumerate(_L12_ROWS):
            for c, val in enumerate(row_data):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                self.l12_table.setItem(r, c, item)

        self.l12_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        l12_h = sum(self.l12_table.rowHeight(r) for r in range(len(_L12_ROWS))) + self.l12_table.horizontalHeader().height() + 4
        self.l12_table.setFixedHeight(l12_h)
        layout.addWidget(self.l12_table)

        # L.1.2 注释
        l12_note = QLabel("注：θ₁ 为水面收敛角，θ₂ 为水面扩散角（灌排规范附录表L.1.2）")
        l12_note.setStyleSheet("color: #555; font-size: 12px;")
        layout.addWidget(l12_note)

        # 关闭按钮
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok)
        btn_box.button(QDialogButtonBox.Ok).setText("关闭")
        btn_box.accepted.connect(self.accept)
        layout.addWidget(btn_box)

    def _load_k12_thumbnails(self):
        """加载K.1.2示意图缩略图，返回 {行索引: QPixmap}"""
        thumb_map = {}
        try:
            from shared.k12_images_data import get_k12_image_bytes
        except ImportError:
            try:
                from 推求水面线.shared.k12_images_data import get_k12_image_bytes
            except ImportError:
                return thumb_map

        thumb_w, thumb_h = 150, 55
        for ri, row_data in enumerate(_K12_ROWS):
            form_name = row_data[0]
            try:
                img_bytes = get_k12_image_bytes(form_name)
                if not img_bytes:
                    continue
                qimg = QImage()
                qimg.loadFromData(QByteArray(img_bytes))
                if qimg.isNull():
                    continue
                pm = QPixmap.fromImage(qimg)
                pm = pm.scaled(thumb_w, thumb_h, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self._pixmap_refs.append(pm)
                thumb_map[ri] = pm
            except Exception:
                pass
        return thumb_map

    def _show_k12_image(self, form_name):
        """放大显示K.1.2示意图"""
        try:
            from shared.k12_images_data import get_k12_image_bytes
        except ImportError:
            try:
                from 推求水面线.shared.k12_images_data import get_k12_image_bytes
            except ImportError:
                return

        img_bytes = get_k12_image_bytes(form_name)
        if not img_bytes:
            return

        qimg = QImage()
        qimg.loadFromData(QByteArray(img_bytes))
        if qimg.isNull():
            return

        pm = QPixmap.fromImage(qimg)

        dlg = QDialog(self.window())
        dlg.setWindowTitle(f"K.1.2 示意图 — {form_name}")
        dlg.setStyleSheet(DIALOG_STYLE)
        lay = QVBoxLayout(dlg)

        # 等比缩放到合适大小
        max_w, max_h = 800, 500
        if pm.width() > max_w or pm.height() > max_h:
            pm = pm.scaled(max_w, max_h, Qt.KeepAspectRatio, Qt.SmoothTransformation)

        img_label = QLabel()
        img_label.setPixmap(pm)
        img_label.setAlignment(Qt.AlignCenter)
        lay.addWidget(img_label)

        btn = QDialogButtonBox(QDialogButtonBox.Ok)
        btn.button(QDialogButtonBox.Ok).setText("关闭")
        btn.accepted.connect(dlg.accept)
        lay.addWidget(btn)

        dlg.resize(pm.width() + 40, pm.height() + 80)
        dlg.exec()


# ================================================================
class _PopupClickFilter(QObject):
    """全局鼠标点击事件过滤器：点击弹窗和按钮外部时关闭弹窗。"""

    def __init__(self, popup, parent_btn):
        super().__init__()
        self._popup = popup
        self._parent_btn = parent_btn

    def eventFilter(self, obj, event):
        if event.type() == QEvent.MouseButtonPress:
            cursor_pos = QCursor.pos()
            popup_rect = self._popup.frameGeometry()
            btn_tl = self._parent_btn.mapToGlobal(self._parent_btn.rect().topLeft())
            btn_rect = QRect(btn_tl, self._parent_btn.rect().size())
            if not popup_rect.contains(cursor_pos) and not btn_rect.contains(cursor_pos):
                self._popup.close()
        return False


# 倒虹吸糙率展示组件（Badge 徽标按钮 + 弹出详情卡片）
# ================================================================
class SiphonRoughnessChipContainer(QWidget):
    """倒虹吸糙率展示 — 简洁按钮 + 数量徽标 + 点击弹出精致详情卡片"""

    _PRIMARY = "#0078D4"
    _PRIMARY_DARK = "#005A9E"

    # 管材参数映射表（与有压管道设计.py保持一致）
    PIPE_MATERIAL_PARAMS = {
        "HDPE管": {"f": 94800, "m": 1.77, "b": 4.77},
        "玻璃钢夹砂管": {"f": 94800, "m": 1.77, "b": 4.77},
        "球墨铸铁管": {"f": 223200, "m": 1.852, "b": 4.87},
        "预应力钢筒混凝土管": {"f": 1312000, "m": 2.0, "b": 5.33},  # n=0.013
        "预应力钢筒混凝土管_n014": {"f": 1516000, "m": 2.0, "b": 5.33},  # n=0.014
        "预应力钢筒混凝土管_n015": {"f": 1749000, "m": 2.0, "b": 5.33},  # n=0.015
        "钢管": {"f": 625000, "m": 1.9, "b": 5.1},
    }

    def __init__(self, parent=None, title_text="倒虹吸糙率详情", empty_text="导入后自动显示", label_prefix="n="):
        super().__init__(parent)
        self._pairs = []  # [(名称, 糙率), ...]
        self._title_text = title_text
        self._empty_text = empty_text
        self.label_prefix = label_prefix
        self._build_ui()

    def _build_ui(self):
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # 占位文字（无数据时显示）
        self._placeholder = QLabel(self._empty_text)
        self._placeholder.setStyleSheet("color: #555555; font-size: 12px;")
        lay.addWidget(self._placeholder)

        # 按钮（有数据时显示）
        self._btn = PushButton("点击查看")
        self._btn.setFixedHeight(28)
        from PySide6.QtWidgets import QSizePolicy
        self._btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self._btn.setStyleSheet(
            "PushButton {"
            "  font-size: 12px; padding: 3px 10px;"
            f"  border: 1px solid {self._PRIMARY}; border-radius: 6px;"
            f"  color: {self._PRIMARY}; background: white;"
            "}"
            "PushButton:hover {"
            f"  background: #E8F4FD; color: {self._PRIMARY_DARK};"
            f"  border-color: {self._PRIMARY_DARK};"
            "}"
        )
        self._btn.clicked.connect(self._show_popover)
        self._btn.setVisible(False)
        lay.addWidget(self._btn)

        # 数量徽标
        self._badge = QLabel("0")
        self._badge.setFixedSize(20, 20)
        self._badge.setAlignment(Qt.AlignCenter)
        self._badge.setStyleSheet(
            f"background: {self._PRIMARY}; color: white;"
            " border-radius: 10px; font-size: 11px; font-weight: bold;"
        )
        self._badge.setVisible(False)
        lay.addWidget(self._badge)

        lay.addStretch()

    def _show_popover(self):
        """点击按钮弹出糙率详情卡片。

        关闭方式：点击弹窗外部区域关闭，或再次点击按钮切换关闭。
        """
        if not self._pairs:
            return

        # 已有弹窗时切换关闭
        if hasattr(self, '_popup_win') and self._popup_win is not None:
            self._popup_win.close()
            return

        primary = self._PRIMARY

        popup = QFrame(None)
        popup.setWindowFlags(
            Qt.ToolTip | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint
        )
        popup.setAttribute(Qt.WA_DeleteOnClose)
        popup.setAttribute(Qt.WA_ShowWithoutActivating)  # 不抢父窗口焦点
        popup.setStyleSheet(
            "QFrame {"
            "  background: white;"
            "  border: 1px solid #E8ECF0;"
            "  border-radius: 12px;"
            "}"
        )

        def _on_destroyed():
            self._popup_win = None
            if hasattr(self, '_click_filter') and self._click_filter is not None:
                app = QApplication.instance()
                if app:
                    app.removeEventFilter(self._click_filter)
                self._click_filter = None

        popup.destroyed.connect(_on_destroyed)
        self._popup_win = popup

        self._click_filter = _PopupClickFilter(popup, self._btn)
        app = QApplication.instance()
        if app:
            app.installEventFilter(self._click_filter)

        card_lay = QVBoxLayout(popup)
        card_lay.setContentsMargins(0, 0, 0, 0)
        card_lay.setSpacing(0)

        # 标题栏
        header = QFrame()
        header.setObjectName("popoverHeader")
        header.setStyleSheet(
            "QFrame#popoverHeader {"
            "  background: white;"
            "  border-top-left-radius: 12px; border-top-right-radius: 12px;"
            "  border-bottom: 1px solid #F0F4F8;"
            "}"
        )
        header_lay = QHBoxLayout(header)
        header_lay.setContentsMargins(14, 11, 14, 9)
        header_lay.setSpacing(7)
        dot_lbl = QLabel("●")
        dot_lbl.setStyleSheet(f"color: {primary}; font-size: 8px; background: transparent;")
        header_lay.addWidget(dot_lbl)
        title = QLabel(self._title_text)
        title.setStyleSheet(
            "font-size: 13px; font-weight: 600; color: #2D3748; background: transparent;"
        )
        header_lay.addWidget(title)
        header_lay.addStretch()
        card_lay.addWidget(header)

        # 数据行
        for i, (name, n_val) in enumerate(self._pairs):
            row_w = QWidget()
            is_last = (i == len(self._pairs) - 1)
            radius_style = (
                "  border-bottom-left-radius: 12px; border-bottom-right-radius: 12px;"
                if is_last else ""
            )
            row_w.setStyleSheet(
                f"QWidget {{ background: white;{radius_style} }}"
                "QWidget:hover { background: #F7FAFC; }"
            )
            row_lay = QHBoxLayout(row_w)
            row_lay.setContentsMargins(14, 8, 14, 8)
            row_lay.setSpacing(16)

            lbl_name = QLabel(name)
            lbl_name.setStyleSheet("font-size: 12px; color: #4A5568; background: transparent;")

            val_lbl = QLabel(f"{self.label_prefix}{n_val}")
            val_lbl.setStyleSheet(
                f"font-size: 12px; font-weight: 700; color: {primary};"
                f" background: #EBF4FF; padding: 2px 8px;"
                f" border-radius: 10px;"
            )

            row_lay.addWidget(lbl_name)
            row_lay.addStretch()
            row_lay.addWidget(val_lbl)

            if not is_last:
                sep = QFrame()
                sep.setFrameShape(QFrame.HLine)
                sep.setStyleSheet("QFrame { background: #F7FAFC; border: none; max-height: 1px; }")
                card_lay.addWidget(row_w)
                card_lay.addWidget(sep)
            else:
                card_lay.addWidget(row_w)

        popup.setMinimumWidth(220)
        popup.adjustSize()
        # 定位：在按钮正下方，间距 0（消除鼠标真空带）
        btn_pos = self._btn.mapToGlobal(self._btn.rect().bottomLeft())
        popup.move(btn_pos.x(), btn_pos.y())
        popup.show()

    def set_pairs(self, pairs):
        """设置数据。pairs: [(名称, 糙率), ...]"""
        self._pairs = list(pairs) if pairs else []
        has_data = bool(self._pairs)
        self._placeholder.setVisible(not has_data)
        self._btn.setVisible(has_data)
        self._badge.setVisible(has_data)
        if has_data:
            n = len(self._pairs)
            self._badge.setText(str(n))
            btn_text = "点击查看"
            self._btn.setText(btn_text)
            text_w = self._btn.fontMetrics().horizontalAdvance(btn_text)
            self._btn.setMinimumWidth(max(88, text_w + 28))
            self._btn.adjustSize()
            self.updateGeometry()

    def set_siphon_data(self, pairs):
        """兼容旧接口"""
        self.set_pairs(pairs)

    def clear(self):
        self._pairs.clear()
        self._placeholder.setVisible(True)
        self._btn.setVisible(False)
        self._badge.setVisible(False)
        self._btn.setMinimumWidth(0)

    def text(self):
        """兼容旧接口"""
        return ""


class WaterProfilePanel(QWidget):
    """推求水面线面板"""

    # 数据变化信号（用于项目管理器追踪脏状态）
    data_changed = Signal()

    def __init__(self, parent=None, siphon_manager=None, pressure_pipe_manager=None):
        super().__init__(parent)
        self._siphon_manager = siphon_manager
        self._pressure_pipe_manager = pressure_pipe_manager
        self.nodes = []
        self.calculated_nodes = []
        self._settings = None
        self.btn_pressure_pipe_calc = None
        self._node_toolbar_layout_preset = NODE_TOOLBAR_LAYOUT_PRESET
        self._pressure_pipe_calc_done = {}
        self._pressure_pipe_calc_records = empty_pressure_pipe_calc_records()
        self._pressure_pipe_last_run_at = ""
        # 建筑物长度统计缓存
        self._last_building_lengths = []
        self._last_channel_total_length = 0.0
        self._last_type_summary = []
        # 纵断面文字导出设置（记住上次使用的参数）
        self._text_export_settings = {
            'y_bottom': 1, 'y_top': 31, 'y_water': 16,
            'text_height': 3.5, 'rotation': 90, 'elev_decimals': 3,
            'y_name': 112, 'y_slope': 102, 'y_ip': 77,
            'y_station': 47, 'y_line_height': 120,
        }
        # 缓存每行的结构高度（structure_height）
        # 该属性不显示在表格列中，但用于计算渠顶高程 = 渠底高程 + 结构高度
        # 在 _import_from_batch / _update_table_from_nodes_full 时存入，在 _build_nodes_from_table 时恢复
        self._node_structure_heights: dict = {}
        self._node_chamfer_params: dict = {}   # {row_idx: {'chamfer_angle': float, 'chamfer_length': float}}
        self._node_u_params: dict = {}         # {row_idx: {'theta_deg': float}}，明渠-U形的圆心角缓存
        self._node_velocity_increased: dict = {}  # {row_idx: float}，加大流速缓存（从批量计算导入）
        # 建筑物名称上平面图设置（记住上次使用的参数）
        self._plan_text_settings = {
            'offset': 10,
            'text_height': 10,
        }
        # CAD导出相关缓存（供"生成断面汇总表"与"导出全部DXF"互相复用）
        self._custom_pressurized_pipe_params = {"siphon": [], "pressure_pipe": []}
        self._custom_struct_thickness = {}
        self._custom_rock_lining = {}
        self._custom_tunnel_unified = {}
        # 防止 cellChanged 递归更新的守卫标志
        self._updating_cells = False
        # 表格编辑撤销栈（单元格编辑）
        self._loss_undo_stack = []
        self._loss_redo_stack = []
        self._pre_edit_cell_value = None  # (row, col, old_text)
        self._pre_edit_snapshot = None  # 编辑前的快照
        self._undo_group = 0  # 撤销分组计数器，用于批量操作时避免重复记录快照
        # 节点表行操作撤销栈（添加/删除/插入/复制/清空行）
        self._node_table_undo_stack = []
        self._node_table_redo_stack = []
        self._init_ui()

    # ================================================================
    # UI 构建
    # ================================================================
    def _init_ui(self):
        main_lay = QVBoxLayout(self)
        main_lay.setContentsMargins(8, 6, 8, 6)
        main_lay.setSpacing(4)

        self._splitter = QSplitter(Qt.Vertical)
        self._splitter.setChildrenCollapsible(False)
        main_lay.addWidget(self._splitter)

        # 上半区：设置 + 输入表
        top_w = QWidget()
        self._build_top_area(top_w)
        self._splitter.addWidget(top_w)

        # 下半区：结果
        bottom_w = QWidget()
        self._build_result_area(bottom_w)
        self._splitter.addWidget(bottom_w)

        self._splitter.setSizes([600, 340])

    def _build_top_area(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(6)

        # ────────────────────────────────────────────
        # 基础设置区（2行网格布局，可折叠）
        # ────────────────────────────────────────────
        settings_grp = CollapsibleGroupBox("基础设置")
        sg = QGridLayout(settings_grp.content_widget())
        sg.setHorizontalSpacing(10)
        sg.setVerticalSpacing(4)
        sg.setContentsMargins(8, 4, 8, 4)

        # 第1行：渠道基本信息
        r = 0
        sg.addWidget(QLabel("渠道名称:"), r, 0, Qt.AlignRight)
        self.channel_name_edit = LineEdit()
        self.channel_name_edit.setText("南峰寺")
        self.channel_name_edit.setMinimumWidth(90)
        sg.addWidget(self.channel_name_edit, r, 1)

        sg.addWidget(QLabel("级别:"), r, 2, Qt.AlignRight)
        self.channel_level_combo = ComboBox()
        self.channel_level_combo.addItems(CHANNEL_LEVEL_OPTIONS)
        self.channel_level_combo.setCurrentText("支渠")
        self.channel_level_combo.setMinimumWidth(75)
        sg.addWidget(self.channel_level_combo, r, 3)

        sg.addWidget(QLabel("起始水位(m):"), r, 4, Qt.AlignRight)
        self.start_wl_edit = LineEdit()
        self.start_wl_edit.setText("100.0")
        self.start_wl_edit.setMinimumWidth(70)
        self.start_wl_edit.setMaximumWidth(160)
        self.start_wl_edit.textChanged.connect(lambda t, e=self.start_wl_edit: e.setFixedWidth(max(70, min(160, e.fontMetrics().horizontalAdvance(t) + 24))))
        sg.addWidget(self.start_wl_edit, r, 5)

        sg.addWidget(QLabel("渠道糙率:"), r, 6, Qt.AlignRight)
        self.roughness_edit = LineEdit()
        self.roughness_edit.setText(str(DEFAULT_ROUGHNESS))
        self.roughness_edit.setFixedWidth(68)
        self.roughness_edit.setToolTip("渠道糙率：适用于明渠、渡槽、隧洞、暗涵等非倒虹吸建筑物")
        sg.addWidget(self.roughness_edit, r, 7)

        sg.addWidget(QLabel("倒虹吸糙率:"), r, 8, Qt.AlignRight)
        self.siphon_roughness_chips = SiphonRoughnessChipContainer()
        sg.addWidget(self.siphon_roughness_chips, r, 9)
        sg.addWidget(QLabel("有压管道参数:"), r, 10, Qt.AlignRight)
        self.pressure_pipe_roughness_chips = SiphonRoughnessChipContainer(
            title_text="有压管道参数详情",
            empty_text="导入后自动显示",
            label_prefix="管材: "
        )
        sg.addWidget(self.pressure_pipe_roughness_chips, r, 11)
        sg.setColumnMinimumWidth(11, 120)

        # 第2行：流量与高级参数
        r = 1
        sg.addWidget(QLabel("设计流量(m³/s):"), r, 0, Qt.AlignRight)
        self.design_flow_edit = LineEdit()
        self.design_flow_edit.setText("")
        self.design_flow_edit.setMinimumWidth(110)
        self.design_flow_edit.setPlaceholderText("多段用逗号分隔")
        sg.addWidget(self.design_flow_edit, r, 1)
        self.design_flow_edit.editingFinished.connect(self._on_design_flow_changed)

        sg.addWidget(QLabel("加大流量(m³/s):"), r, 2, Qt.AlignRight)
        self.max_flow_edit = LineEdit()
        self.max_flow_edit.setText("")
        self.max_flow_edit.setMinimumWidth(110)
        self.max_flow_edit.setPlaceholderText("自动计算")
        sg.addWidget(self.max_flow_edit, r, 3)

        sg.addWidget(QLabel("起始桩号(m):"), r, 4, Qt.AlignRight)
        self.start_station_edit = LineEdit()
        self.start_station_edit.setText("0+000.000")
        self.start_station_edit.setMinimumWidth(100)
        sg.addWidget(self.start_station_edit, r, 5)
        self.start_station_edit.editingFinished.connect(self._format_start_station)
        self.start_station_edit.installEventFilter(self)

        sg.addWidget(QLabel("转弯半径(m):"), r, 6, Qt.AlignRight)
        turn_r_box = QHBoxLayout()
        turn_r_box.setSpacing(4)
        self.turn_radius_edit = LineEdit()
        self.turn_radius_edit.setText(str(DEFAULT_TURN_RADIUS))
        self.turn_radius_edit.setFixedWidth(60)
        turn_r_box.addWidget(self.turn_radius_edit)
        btn_auto_r = PushButton("自动")
        btn_auto_r.setFixedWidth(52)
        btn_auto_r.setToolTip(
            "根据规范自动计算推荐转弯半径（取大值原则）\n"
            "• 隧洞：弯曲半径≥洞径(或洞宽)×5\n"
            "• 明渠：弯曲半径≥水面宽度×5\n"
            "• 渡槽：弯道半径≥连接明渠渠底宽度×5"
        )
        btn_auto_r.clicked.connect(self._auto_calc_turn_radius)
        turn_r_box.addWidget(btn_auto_r)
        turn_r_box.addStretch()
        sg.addLayout(turn_r_box, r, 7)

        # 列弹性
        for c in [1, 3, 5, 7]:
            sg.setColumnStretch(c, 1)
        sg.setColumnStretch(9, 1)
        sg.setColumnStretch(11, 2)
        settings_grp.toggled.connect(self._on_settings_toggled)
        lay.addWidget(settings_grp)

        # ────────────────────────────────────────────
        # 渐变段设置区（3行网格布局，按类型分行，可折叠）
        # ────────────────────────────────────────────
        trans_grp = CollapsibleGroupBox("渐变段设置")
        tg = QGridLayout(trans_grp.content_widget())
        tg.setHorizontalSpacing(8)
        tg.setVerticalSpacing(4)
        tg.setContentsMargins(8, 4, 8, 4)

        _cat_style = f"font-weight:bold; color:{P}; font-size:12px;"

        # 行0：渡槽/隧洞
        r = 0
        cat0 = QLabel("渡槽/隧洞")
        cat0.setStyleSheet(_cat_style)
        tg.addWidget(cat0, r, 0, Qt.AlignRight)
        tg.addWidget(QLabel("进口:"), r, 1, Qt.AlignRight)
        self.trans_inlet_combo = ComboBox()
        self.trans_inlet_combo.addItems(TRANSITION_FORM_OPTIONS)
        self.trans_inlet_combo.setMinimumWidth(130)
        tg.addWidget(self.trans_inlet_combo, r, 2)
        tg.addWidget(QLabel("ζ₁="), r, 3, Qt.AlignRight)
        self.trans_inlet_zeta = LineEdit()
        self.trans_inlet_zeta.setText("0.10")
        self.trans_inlet_zeta.setFixedWidth(65)
        tg.addWidget(self.trans_inlet_zeta, r, 4)
        tg.addWidget(QLabel("出口:"), r, 5, Qt.AlignRight)
        self.trans_outlet_combo = ComboBox()
        self.trans_outlet_combo.addItems(TRANSITION_FORM_OPTIONS)
        self.trans_outlet_combo.setMinimumWidth(130)
        tg.addWidget(self.trans_outlet_combo, r, 6)
        tg.addWidget(QLabel("ζ₂="), r, 7, Qt.AlignRight)
        self.trans_outlet_zeta = LineEdit()
        self.trans_outlet_zeta.setText("0.20")
        self.trans_outlet_zeta.setFixedWidth(65)
        tg.addWidget(self.trans_outlet_zeta, r, 8)

        # 行1：明渠
        r = 1
        cat1 = QLabel("明渠")
        cat1.setStyleSheet(_cat_style)
        tg.addWidget(cat1, r, 0, Qt.AlignRight)
        tg.addWidget(QLabel("型式:"), r, 1, Qt.AlignRight)
        self.oc_trans_combo = ComboBox()
        self.oc_trans_combo.addItems(TRANSITION_FORM_OPTIONS)
        self.oc_trans_combo.setMinimumWidth(130)
        tg.addWidget(self.oc_trans_combo, r, 2)
        tg.addWidget(QLabel("ζ="), r, 3, Qt.AlignRight)
        self.oc_trans_zeta = LineEdit()
        self.oc_trans_zeta.setText("0.10")
        self.oc_trans_zeta.setFixedWidth(65)
        tg.addWidget(self.oc_trans_zeta, r, 4)

        # 行2：倒虹吸
        r = 2
        cat2 = QLabel("倒虹吸")
        cat2.setStyleSheet(_cat_style)
        tg.addWidget(cat2, r, 0, Qt.AlignRight)
        tg.addWidget(QLabel("进口:"), r, 1, Qt.AlignRight)
        self.siphon_inlet_combo = ComboBox()
        self.siphon_inlet_combo.addItems(SIPHON_TRANSITION_FORM_OPTIONS)
        self.siphon_inlet_combo.setMinimumWidth(110)
        tg.addWidget(self.siphon_inlet_combo, r, 2)
        tg.addWidget(QLabel("ζ₁="), r, 3, Qt.AlignRight)
        self.siphon_inlet_zeta = LineEdit()
        self.siphon_inlet_zeta.setText("0.10")
        self.siphon_inlet_zeta.setFixedWidth(65)
        tg.addWidget(self.siphon_inlet_zeta, r, 4)
        tg.addWidget(QLabel("出口:"), r, 5, Qt.AlignRight)
        self.siphon_outlet_combo = ComboBox()
        self.siphon_outlet_combo.addItems(SIPHON_TRANSITION_FORM_OPTIONS)
        self.siphon_outlet_combo.setMinimumWidth(110)
        tg.addWidget(self.siphon_outlet_combo, r, 6)
        tg.addWidget(QLabel("ζ₂="), r, 7, Qt.AlignRight)
        self.siphon_outlet_zeta = LineEdit()
        self.siphon_outlet_zeta.setText("0.20")
        self.siphon_outlet_zeta.setFixedWidth(65)
        tg.addWidget(self.siphon_outlet_zeta, r, 8)

        # 参考系数表按钮（放在倒虹吸行末尾）
        btn_ref = PushButton("参考系数表")
        btn_ref.setToolTip("查看表K.1.2（渡槽/隧洞）和表L.1.2（倒虹吸）渐变段局部损失系数")
        btn_ref.clicked.connect(self._open_transition_reference)
        tg.addWidget(btn_ref, r, 9)

        # 列弹性
        tg.setColumnStretch(2, 1)
        tg.setColumnStretch(6, 1)
        trans_grp.toggled.connect(self._on_settings_toggled)
        lay.addWidget(trans_grp)

        # 渐变段型式变更 → 自动更新ζ系数
        self.trans_inlet_combo.currentTextChanged.connect(self._on_trans_inlet_form_changed)
        self.trans_outlet_combo.currentTextChanged.connect(self._on_trans_outlet_form_changed)
        self.oc_trans_combo.currentTextChanged.connect(self._on_oc_trans_form_changed)
        self.siphon_inlet_combo.currentTextChanged.connect(self._on_siphon_inlet_form_changed)
        self.siphon_outlet_combo.currentTextChanged.connect(self._on_siphon_outlet_form_changed)

        # ────────────────────────────────────────────
        # 工具栏（单行分组布局，优化间距与宽度）
        # ────────────────────────────────────────────
        toolbar_presets = {
            "compact": {
                "h_spacing": 6,
                "btn_height": 34,
                "extra_primary": 10,
                "extra_primary_long": 12,
                "extra_normal": 8,
                "extra_clear": 6,
            },
            "balanced": {
                "h_spacing": 8,
                "btn_height": 36,
                "extra_primary": 10,
                "extra_primary_long": 12,
                "extra_normal": 8,
                "extra_clear": 6,
            },
            "comfortable": {
                "h_spacing": 10,
                "btn_height": 38,
                "extra_primary": 12,
                "extra_primary_long": 14,
                "extra_normal": 10,
                "extra_clear": 8,
            },
        }
        self._node_toolbar_preset = toolbar_presets.get(self._node_toolbar_layout_preset, toolbar_presets["balanced"])

        tb = QHBoxLayout()
        tb.setContentsMargins(0, 0, 0, 0)
        tb.setSpacing(self._node_toolbar_preset["h_spacing"])
        tb.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        lbl = QLabel("节点数据表")
        lbl.setStyleSheet(f"font-size:13px; font-weight:bold; color:{T1};")
        from PySide6.QtWidgets import QSizePolicy
        lbl.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)
        tb.addWidget(lbl)
        tb.addSpacing(10)

        def _register_toolbar_button(btn, role):
            btn.ensurePolished()
            hint_w = btn.sizeHint().width()
            min_w = btn.minimumSizeHint().width() + 2
            if role == "primary_long":
                extra = self._node_toolbar_preset["extra_primary_long"]
            elif role == "primary":
                extra = self._node_toolbar_preset["extra_primary"]
            elif role == "clear":
                extra = self._node_toolbar_preset["extra_clear"]
            else:
                extra = self._node_toolbar_preset["extra_normal"]
            width = max(min_w, hint_w + extra)
            btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            btn.setFixedSize(width, self._node_toolbar_preset["btn_height"])

        # 数据导入组
        btn_import = PrimaryPushButton("从批量计算导入")
        btn_import.clicked.connect(self._import_from_batch)
        _register_toolbar_button(btn_import, "primary_long")
        tb.addWidget(btn_import)

        # 计算组（醒目按钮，紧跟在导入按钮之后）
        btn_transition = PrimaryPushButton("插入渐变段")
        btn_transition.clicked.connect(self._insert_transitions)
        _register_toolbar_button(btn_transition, "primary")
        
        btn_siphon = PrimaryPushButton("倒虹吸水力计算")
        btn_siphon.clicked.connect(self._open_siphon_calculator)
        _register_toolbar_button(btn_siphon, "primary_long")
        
        self.btn_pressure_pipe_calc = PrimaryPushButton("有压管道水力计算")
        self.btn_pressure_pipe_calc.clicked.connect(self._open_pressure_pipe_calculator)
        _register_toolbar_button(self.btn_pressure_pipe_calc, "primary_long")

        btn_calc = PrimaryPushButton("执行计算")
        btn_calc.clicked.connect(self._calculate)
        _register_toolbar_button(btn_calc, "primary")
        for w in [btn_transition, btn_siphon, self.btn_pressure_pipe_calc, btn_calc]:
            tb.addWidget(w)

        # 数据清理组
        btn_clear = PushButton("清空")
        btn_clear.setToolTip("清空表格中所有节点\n▶ 支持 Ctrl+Z 撤销")
        btn_clear.clicked.connect(self._clear_nodes)
        _register_toolbar_button(btn_clear, "clear")

        tb.addWidget(btn_clear)

        lay.addLayout(tb)

        # 统一节点表（输入+结果在同一个表格，与Tkinter版一致）
        self.node_table = FrozenColumnTableWidget(0, len(NODE_ALL_HEADERS), frozen_count=4)
        self.node_table.setHorizontalHeaderLabels(NODE_ALL_HEADERS)
        self.node_table.horizontalHeader().setStretchLastSection(False)
        self.node_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.node_table.horizontalHeader().setMinimumSectionSize(50)
        self.node_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.node_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.node_table.setAlternatingRowColors(True)
        self.node_table.setFont(QFont("Microsoft YaHei", 10))
        self.node_table.verticalHeader().setDefaultSectionSize(26)
        self.node_table.setMinimumHeight(180)
        self.node_table.cellDoubleClicked.connect(self._on_node_cell_double_clicked)
        self.node_table.cellChanged.connect(self._on_loss_cell_changed)
        self.node_table.currentCellChanged.connect(self._on_current_cell_changed)
        # Undo/Redo 通过 FrozenColumnTableWidget 的信号连接（表格 keyPressEvent 先处理按键）
        self.node_table.undoRequested.connect(self._undo_loss_edit)
        self.node_table.redoRequested.connect(self._redo_loss_edit)
        # Delete 键删除时记录快照
        self.node_table.deleteRequested.connect(self._push_undo_snapshot)
        lay.addWidget(self.node_table, stretch=1)
        
        # 面板级别的撤销/重做快捷键（无需选中单元格也能撤销）
        from PySide6.QtGui import QShortcut, QKeySequence
        undo_sc = QShortcut(QKeySequence.StandardKey.Undo, self)
        undo_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        undo_sc.activated.connect(self._undo_loss_edit)
        redo_sc = QShortcut(QKeySequence.StandardKey.Redo, self)
        redo_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        redo_sc.activated.connect(self._redo_loss_edit)
        self._setup_header_tooltips()
        self._refresh_pressure_pipe_controls()

    def _on_settings_toggled(self, collapsed):
        """折叠/展开设置区时，自动调整splitter分配，让底部图表获得释放的空间"""
        from PySide6.QtCore import QTimer
        QTimer.singleShot(0, self._adjust_splitter_for_settings)

    def _adjust_splitter_for_settings(self):
        """根据上半区实际需要的高度重新分配splitter空间"""
        top_w = self._splitter.widget(0)
        top_hint = top_w.minimumSizeHint().height()
        total = self._splitter.height()
        # 上半区取实际需要的高度（但不超过总高度的70%）
        top_h = min(top_hint, int(total * 0.7))
        bottom_h = total - top_h
        self._splitter.setSizes([top_h, bottom_h])

    def _setup_header_tooltips(self):
        """为表头设置悬浮提示（LaTeX公式渲染），使用自定义Fluent悬浮卡片"""
        from 渠系断面设计.water_profile.formula_dialog import COLUMN_FORMULAS, FormulaTooltipWidget

        self._formula_tooltip = FormulaTooltipWidget()
        self._formula_columns = set()
        for col_idx, col_name in enumerate(NODE_ALL_HEADERS):
            if col_name in COLUMN_FORMULAS:
                self._formula_columns.add(col_idx)

        header = self.node_table.horizontalHeader()
        header.setMouseTracking(True)
        self._node_header = header
        header.viewport().setMouseTracking(True)
        header.viewport().installEventFilter(self)

    def _on_node_cell_double_clicked(self, row, col):
        """双击单元格：结构形式列弹出选择面板；水头损失/高程列显示详细计算过程"""
        col_name = NODE_ALL_HEADERS[col] if col < len(NODE_ALL_HEADERS) else ""

        # 结构形式列：弹出分类选择面板
        if col == 2:
            current = ""
            item = self.node_table.item(row, col)
            if item:
                current = item.text()
            dlg = StructureTypeSelector(self)
            dlg.set_current(current)
            result = dlg.exec()
            if result == QDialog.DialogCode.Accepted and dlg.selected_type:
                type_item = QTableWidgetItem(dlg.selected_type)
                type_item.setTextAlignment(Qt.AlignCenter)
                self.node_table.setItem(row, col, type_item)
            # 无论选择还是 Esc 取消，都回到原表格单元格，保持操作连续性
            self.node_table.setCurrentCell(row, col)
            self.node_table.setFocus(Qt.OtherFocusReason)
            return

        # 水头损失/高程列：显示详细计算过程（与原版Tkinter _on_cell_double_click对齐）
        from 渠系断面设计.water_profile.formula_dialog import DOUBLE_CLICK_COLUMNS
        if col_name not in DOUBLE_CLICK_COLUMNS:
            return
        if not hasattr(self, 'calculated_nodes') or not self.calculated_nodes:
            return
        nodes = self.calculated_nodes
        if row < 0 or row >= len(nodes):
            return
        # 弹窗前强制从表格同步损失/水位/高程到 calculated_nodes，确保显示最新值
        self._sync_losses_from_table()
        node = nodes[row]

        if col_name == "渐变段长度L":
            self._show_transition_length_details(row, node)
        elif col_name == "弯道水头损失":
            self._show_bend_calc_details(row, node)
        elif col_name == "沿程水头损失":
            self._show_friction_calc_details(row, node)
        elif col_name == "渐变段水头损失":
            self._show_transition_calc_details(row, node)
        elif col_name == "总水头损失":
            self._show_total_calc_details(row, node, nodes)
        elif col_name == "累计总水头损失":
            self._show_cumulative_loss_details(row, node, nodes)
        elif col_name == "水位":
            self._show_water_level_details(row, node, nodes)
        elif col_name == "渠底高程":
            self._show_bottom_elevation_details(row, node, nodes)
        elif col_name == "渠顶高程":
            self._show_top_elevation_details(row, node)

    # ================================================================
    # 双击查看详细计算过程（与原版Tkinter data_table.py完全对齐）
    # ================================================================
    def _show_transition_length_details(self, row_idx, node):
        details = getattr(node, 'transition_length_calc_details', None)
        if not details:
            fluent_info(self, "提示", "该行没有渐变段长度计算数据")
            return
        from 渠系断面设计.water_profile.formula_dialog import show_transition_length_dialog
        show_transition_length_dialog(self, node.name or f"行{row_idx+1}", details)

    def _show_bend_calc_details(self, row_idx, node):
        if not getattr(node, 'bend_calc_details', None):
            fluent_info(self, "提示", "该行没有弯道水头损失计算数据")
            return
        from 渠系断面设计.water_profile.formula_dialog import show_bend_loss_dialog
        show_bend_loss_dialog(self, node.name or f"行{row_idx+1}", node.bend_calc_details)

    def _show_friction_calc_details(self, row_idx, node):
        if not getattr(node, 'friction_calc_details', None):
            fluent_info(self, "提示", "该行没有沿程水头损失计算数据")
            return
        from 渠系断面设计.water_profile.formula_dialog import show_friction_loss_dialog
        show_friction_loss_dialog(self, node.name or f"行{row_idx+1}", node.friction_calc_details)

    def _show_transition_calc_details(self, row_idx, node):
        if not getattr(node, 'is_transition', False):
            fluent_info(self, "提示", "该行不是渐变段，无法显示详细计算过程")
            return
        details = getattr(node, 'transition_calc_details', None)
        if not details:
            fluent_info(self, "提示", "该渐变段尚未计算水头损失")
            return
        from 渠系断面设计.water_profile.formula_dialog import show_transition_loss_dialog
        show_transition_loss_dialog(self, node.name or f"行{row_idx+1}", details)

    def _show_total_calc_details(self, row_idx, node, nodes):
        if getattr(node, 'is_transition', False):
            fluent_info(self, "提示", "渐变段行没有总水头损失，请双击渐变段水头损失列查看")
            return
        # 收集渐变段损失
        h_transition = 0.0
        for i in range(row_idx - 1, -1, -1):
            if nodes[i].is_transition:
                h_transition += nodes[i].head_loss_transition or 0.0
            elif not nodes[i].is_transition:
                break
        details = {
            'head_loss_bend': node.head_loss_bend or 0.0,
            'head_loss_transition': h_transition,
            'head_loss_friction': node.head_loss_friction or 0.0,
            'head_loss_local': getattr(node, 'head_loss_local', 0.0) or 0.0,
            'head_loss_reserve': getattr(node, 'head_loss_reserve', 0.0) or 0.0,
            'head_loss_gate': getattr(node, 'head_loss_gate', 0.0) or 0.0,
            'head_loss_siphon': getattr(node, 'head_loss_siphon', 0.0) or 0.0,
            'head_loss_total': node.head_loss_total or 0.0,
        }
        from 渠系断面设计.water_profile.formula_dialog import show_total_loss_dialog
        show_total_loss_dialog(self, node.name or f"行{row_idx+1}", details)

    def _show_cumulative_loss_details(self, row_idx, node, nodes):
        cumulative = 0.0
        lines = []
        for i, n in enumerate(nodes):
            if i > row_idx:
                break
            if n.is_transition:
                loss = n.head_loss_transition or 0.0
                if loss <= 0 and getattr(n, 'transition_calc_details', None):
                    loss = n.transition_calc_details.get('total', 0.0) or 0.0
                cumulative += loss
                lines.append(f"第{i+1}行(渐变段):  $h_{{tr}} = {loss:.4f}$ m，累计 $= {cumulative:.4f}$ m")
            else:
                loss = n.head_loss_total or 0.0
                cumulative += loss
                # 构建分项明细（含用户手动输入的预留/过闸/倒虹吸）
                parts = []
                hw = n.head_loss_bend or 0.0
                hf = n.head_loss_friction or 0.0
                hj = getattr(n, 'head_loss_local', 0.0) or 0.0
                hr = getattr(n, 'head_loss_reserve', 0.0) or 0.0
                hg = getattr(n, 'head_loss_gate', 0.0) or 0.0
                hs = getattr(n, 'head_loss_siphon', 0.0) or 0.0
                if hw: parts.append(f"弯道{hw:.4f}")
                if hf: parts.append(f"沿程{hf:.4f}")
                if hj: parts.append(f"局部{hj:.4f}")
                if hr: parts.append(f"预留{hr:.4f}")
                if hg: parts.append(f"过闸{hg:.4f}")
                if hs: parts.append(f"倒虹吸{hs:.4f}")
                detail = f"（{'＋'.join(parts)}）" if parts else ""
                lines.append(f"第{i+1}行(普通):  $h_{{\\Sigma}} = {loss:.4f}$ m{detail}，累计 $= {cumulative:.4f}$ m")
        from 渠系断面设计.water_profile.formula_dialog import show_cumulative_loss_dialog
        show_cumulative_loss_dialog(self, node.name or f"行{row_idx+1}",
                                    {"cumulative": cumulative, "rows_text": "\n".join(lines)})

    def _show_water_level_details(self, row_idx, node, nodes):
        if getattr(node, 'is_transition', False):
            fluent_info(self, "提示", "渐变段行不显示水位")
            return
        # 找第一个常规节点和上一个常规节点
        first_regular_idx = None
        for i, n in enumerate(nodes):
            if not n.is_transition:
                first_regular_idx = i
                break
        prev_idx = None
        for i in range(row_idx - 1, -1, -1):
            if not nodes[i].is_transition:
                prev_idx = i
                break
        is_first = (first_regular_idx == row_idx)
        is_gate = bool(getattr(node, 'is_diversion_gate', False))
        settings = self._build_settings()
        start_level = settings.start_water_level if settings else 0.0
        details = {
            "is_first": is_first,
            "is_gate": is_gate,
            "water_level": node.water_level or 0.0,
            "start_level": start_level,
            "cumulative": node.head_loss_cumulative or 0.0,
            "total_loss": node.head_loss_total or 0.0,
        }
        if is_first:
            pass
        elif prev_idx is not None:
            details["prev_level"] = nodes[prev_idx].water_level or 0.0
            if is_gate:
                details["head_loss_gate"] = getattr(node, 'head_loss_gate', 0.0) or 0.0
            else:
                details["hf"] = node.head_loss_friction or 0.0
                details["hj"] = getattr(node, 'head_loss_local', 0.0) or 0.0
                details["hw"] = node.head_loss_bend or 0.0
                # 收集渐变段损失
                h_tr = 0.0
                for j in range(prev_idx + 1, row_idx):
                    if nodes[j].is_transition:
                        h_tr += nodes[j].head_loss_transition or 0.0
                details["h_tr"] = h_tr
                details["h_reserve"] = getattr(node, 'head_loss_reserve', 0.0) or 0.0
                details["h_gate"] = getattr(node, 'head_loss_gate', 0.0) or 0.0
                details["h_siphon"] = getattr(node, 'head_loss_siphon', 0.0) or 0.0
        else:
            fluent_info(self, "提示", "该行无法获取上一节点水位")
            return
        from 渠系断面设计.water_profile.formula_dialog import show_water_level_dialog
        show_water_level_dialog(self, node.name or f"行{row_idx+1}", details)

    def _show_bottom_elevation_details(self, row_idx, node, nodes):
        if getattr(node, 'is_transition', False):
            fluent_info(self, "提示", "渐变段行不显示渠底高程")
            return
        # 倒虹吸出口节点：使用公式10.3.6专用弹窗
        try:
            from 推求水面线.models.data_models import StructureType as ST, InOutType as IOT
            if (node.structure_type == ST.INVERTED_SIPHON
                    and node.in_out == IOT.OUTLET
                    and getattr(node, 'siphon_outlet_elev_details', None)):
                from 渠系断面设计.water_profile.formula_dialog import show_siphon_outlet_elevation_dialog
                show_siphon_outlet_elevation_dialog(self, node.name or f"行{row_idx+1}", node.siphon_outlet_elev_details)
                return
            if (node.structure_type == ST.INVERTED_SIPHON
                    and node.in_out == IOT.INLET and node.bottom_elevation):
                fluent_info(self, "渠底高程说明",
                            f"倒虹吸进口渠底高程取自上游渠道末端的渠底高程：\n\n渠底高程 = {node.bottom_elevation:.3f} m")
                return
        except ImportError:
            pass
        wd = node.water_depth or 0.0
        if wd <= 0:
            fluent_info(self, "提示", "该行没有水深数据，无法计算渠底高程")
            return
        from 渠系断面设计.water_profile.formula_dialog import show_bottom_elevation_dialog
        show_bottom_elevation_dialog(self, node.name or f"行{row_idx+1}",
                                      {"water_level": node.water_level or 0.0,
                                       "water_depth": wd,
                                       "bottom_elevation": node.bottom_elevation or 0.0})

    def _show_top_elevation_details(self, row_idx, node):
        if getattr(node, 'is_transition', False):
            fluent_info(self, "提示", "渐变段行不显示渠顶高程")
            return
        sh = node.structure_height or 0.0
        if sh <= 0:
            fluent_info(self, "提示", "该行没有结构高度数据，无法计算渠顶高程")
            return
        be = node.bottom_elevation or 0.0
        te = node.top_elevation or 0.0
        if be == 0 and te == 0:
            fluent_info(self, "提示", "该行没有渠底高程数据，无法计算渠顶高程")
            return
        from 渠系断面设计.water_profile.formula_dialog import show_top_elevation_dialog
        show_top_elevation_dialog(self, node.name or f"行{row_idx+1}",
                                   {"bottom_elevation": be, "structure_height": sh, "top_elevation": te})

    def _build_result_area(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(4)

        # 导出工具栏
        tb = QHBoxLayout()
        lbl = QLabel("计算结果")
        lbl.setStyleSheet(f"font-size:13px;font-weight:bold;color:{T1};")
        tb.addWidget(lbl)
        tb.addStretch()
        btn_export_excel = PushButton("导出Excel"); btn_export_excel.clicked.connect(self._export_excel)
        btn_export_word = PushButton("导出Word"); btn_export_word.clicked.connect(self._export_word)
        for w in [btn_export_excel, btn_export_word]:
            tb.addWidget(w)
        lay.addLayout(tb)

        # CAD工具栏
        cad_tb = QHBoxLayout()
        cad_tb.setSpacing(6)
        cad_lbl = QLabel("CAD工具")
        cad_lbl.setStyleSheet(f"font-size:12px;font-weight:bold;color:{T2};")
        cad_tb.addWidget(cad_lbl)
        cad_tb.addStretch()
        btn_profile = PushButton("生成纵断面表格"); btn_profile.clicked.connect(self._cad_longitudinal_profile)
        btn_profile.setToolTip("导出上纵断面表格 DXF/TXT\n含线框、渠底/渠顶/水面折线、高程文字、桩号、建筑物名称、坡降、IP点名称")
        btn_summary = PushButton("生成断面汇总表"); btn_summary.clicked.connect(self._cad_section_summary)
        btn_summary.setToolTip("导出各类断面尺寸及水力要素汇总表 DXF\n含明渠/隧洞/渡槽/暗涵/倒虹吸等断面参数")
        btn_ip = PushButton("IP坐标及弯道参数表"); btn_ip.clicked.connect(self._cad_ip_table)
        btn_ip.setToolTip("导出IP坐标及弯道参数表 DXF/Excel\n含IP点坐标、桩号、转角、半径、切线长、弧长、底高程")
        btn_combined = PrimaryPushButton("导出全部DXF"); btn_combined.clicked.connect(self._cad_combined_dxf)
        btn_combined.setToolTip("一键合并导出：纵断面表格 + 断面汇总表 + IP坐标表\n三个表格输出到同一个DXF文件，分图层管理")
        btn_bzzh2 = PushButton("生成bzzh2命令内容"); btn_bzzh2.clicked.connect(self._cad_bzzh2)
        btn_bzzh2.setToolTip("生成ZDM用的bzzh2命令 TXT\n提取建筑物进出口数据")
        btn_plan = PushButton("建筑物名称上平面图"); btn_plan.clicked.connect(self._cad_building_plan)
        btn_plan.setToolTip("生成AutoCAD -TEXT命令并复制到剪贴板\n将建筑物名称平行于轴线放置在平面图上")
        for w in [btn_profile, btn_summary, btn_ip, btn_combined, btn_bzzh2, btn_plan]:
            cad_tb.addWidget(w)
        lay.addLayout(cad_tb)

        # 计算结果摘要面板（持久显示）
        self.summary_grp = QGroupBox("计算结果摘要")
        sg_lay = QHBoxLayout(self.summary_grp)
        sg_lay.setContentsMargins(8, 4, 8, 4)
        sg_lay.setSpacing(16)
        self.lbl_summary_info = QLabel("尚未计算")
        self.lbl_summary_info.setStyleSheet("font-size: 12px;")
        sg_lay.addWidget(self.lbl_summary_info, stretch=1)
        self.btn_building_stats = PushButton("建筑物长度统计")
        self.btn_building_stats.clicked.connect(self._show_building_length_dialog)
        self.btn_building_stats.setEnabled(False)
        sg_lay.addWidget(self.btn_building_stats)
        lay.addWidget(self.summary_grp)

        # 详细过程文本框
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        self.detail_text.setFont(QFont("Consolas", 10))
        lay.addWidget(self.detail_text)

        # 初始帮助
        self._show_help()

    def _show_help(self):
        lines = [
            "=" * 70,
            "  推求水面线 — 使用说明",
            "=" * 70, "",
            "操作步骤：",
            "  1. 填写基础设置（渠道名称、起始水位、流量等）",
            "  2. 输入节点数据（从批量计算导入）",
            "  3. 核对基础设置和渐变段设置",
            "  4. 点击「插入渐变段」",
            "  5.（如有倒虹吸）点击「倒虹吸水力计算」",
            "  6.（如有有压管道）点击「有压管道水力计算」",
            "  7. 点击「执行计算」",
            "  8. 查看结果表格和详细过程", "",
            "从批量计算导入：",
            "  - 先在「批量计算」模块完成计算",
            "  - 点击「从批量计算导入」自动填充节点表",
            "  - 导入后自动同步渠道名称、级别、起始水位等基础设置",
            "  - 自动填充多流量段设计流量和加大流量",
            "  - 自动计算推荐转弯半径（按规范取大值原则）", "",
            "多流量段支持：",
            "  - 设计流量和加大流量支持逗号分隔的多值输入",
            "  - 例如：5.0, 8.0, 10.0",
            "  - 修改设计流量后自动计算对应加大流量", "",
            "节点数据说明：",
            "  - 流量段：所属流量段编号（如1、2）",
            "  - 建筑物名称：节点名称（如隧洞1、渡槽2）",
            "  - 结构形式：双击选择断面类型（含闸·分水等）",
            "  - X/Y：平面坐标（用于几何计算）",
            "  - 转弯半径：弯道半径（m），留空使用全局设置",
            "  - 底宽B/直径D/半径R/边坡m：断面几何参数",
            "  - 糙率n/底坡1/i/流量Q：水力参数（留空使用全局设置）", "",
            "转弯半径取值规范：",
            "  - 隧洞：弯曲半径≥洞径(或洞宽)×5",
            "  - 明渠：弯曲半径≥水面宽度×5",
            "  - 渡槽：弯道半径≥连接明渠渠底宽度×5",
            "  - 倒虹吸：R = n × D（独立设置）",
            "=" * 70,
        ]
        self.detail_text.setPlainText("\n".join(lines))

    # ================================================================
    # 节点表操作
    # ================================================================
    def _sync_losses_from_table(self):
        """从表格读取最新的水头损失/水位/高程值，同步到 calculated_nodes。
        确保双击弹窗始终显示用户手动编辑后的最新数据。"""
        if not hasattr(self, 'calculated_nodes') or not self.calculated_nodes:
            return
        table = self.node_table
        row_count = table.rowCount()

        def _rf(r, c):
            item = table.item(r, c)
            if not item:
                return 0.0
            txt = item.text().strip()
            if not txt or txt == '-':
                return 0.0
            try:
                return float(txt)
            except ValueError:
                return 0.0

        for r in range(min(row_count, len(self.calculated_nodes))):
            node = self.calculated_nodes[r]
            if node.is_transition:
                node.head_loss_transition = _rf(r, 33) or node.head_loss_transition
                node.head_loss_cumulative = _rf(r, 40) or node.head_loss_cumulative
            else:
                # 可编辑损失列（用户可能手动修改）
                node.head_loss_reserve = _rf(r, 36)
                node.head_loss_gate = _rf(r, 37)
                node.head_loss_siphon = _rf(r, 38)
                # 联动计算列
                node.head_loss_total = _rf(r, 39) or node.head_loss_total
                node.head_loss_cumulative = _rf(r, 40) or node.head_loss_cumulative
                wl = _rf(r, 41)
                if wl:
                    node.water_level = wl
                be = _rf(r, 42)
                if be:
                    node.bottom_elevation = be
                te = _rf(r, 43)
                if te:
                    node.top_elevation = te

    def _on_current_cell_changed(self, row, col, prev_row, prev_col):
        """当用户切换单元格时，记录编辑前的快照（供撤销使用）"""
        if self._updating_cells:
            return
        # 重置批量操作标志（Delete 删除结束）
        self._undo_group = 0
        # 任何单元格切换时都记录快照（与批量计算面板保持一致）
        self._pre_edit_snapshot = self._snapshot_editable_cols()
        if col in EDITABLE_COLS:
            item = self.node_table.item(row, col)
            self._pre_edit_cell_value = (row, col, item.text() if item else "")

    def _snapshot_editable_cols(self):
        """保存所有可编辑列的快照（用于撤销）"""
        snapshot = {}
        for r in range(self.node_table.rowCount()):
            for c in EDITABLE_COLS:
                item = self.node_table.item(r, c)
                snapshot[(r, c)] = item.text() if item else ""
            # 也保存联动计算列（39-43）
            for c in range(39, 44):
                item = self.node_table.item(r, c)
                snapshot[(r, c)] = item.text() if item else ""
        return snapshot

    def _on_loss_cell_changed(self, row, col):
        """当用户编辑可编辑列时，记录撤销快照；若为水头损失列则联动更新"""
        if self._updating_cells:
            return
        # 仅对可编辑列触发
        if col not in EDITABLE_COLS:
            return
        
        # 触发数据变化信号（除非正在加载项目）
        if not getattr(self, '_loading_project', False):
            self.data_changed.emit()
        
        self._updating_cells = True
        try:
            # 如果在批量操作中（如 Delete 键删除），跳过快照记录。
            # _undo_group 记录本轮需要跳过的 cellChanged 次数，逐次递减归零。
            skip_snapshot = self._undo_group > 0
            if skip_snapshot:
                self._undo_group -= 1

            if not skip_snapshot:
                # 如果没有预先记录的快照，先生成一个
                if self._pre_edit_snapshot is None:
                    self._pre_edit_snapshot = self._snapshot_editable_cols()
                
                # 编辑单元格已经是新值，用 _pre_edit_cell_value 还原旧值
                if self._pre_edit_cell_value and self._pre_edit_cell_value[:2] == (row, col):
                    self._pre_edit_snapshot[(row, col)] = self._pre_edit_cell_value[2]
                
                self._append_loss_undo_snapshot(self._pre_edit_snapshot)
                self._pre_edit_snapshot = None

            # 对于水头损失列（36, 37, 38），触发联动计算
            if col in (36, 37, 38) and row > 0:
                self._recalc_downstream(row)
        finally:
            self._updating_cells = False
        # 更新 pre_edit 为当前新值，以便连续编辑同一单元格时也能撤销
        item = self.node_table.item(row, col)
        self._pre_edit_cell_value = (row, col, item.text() if item else "")

        self._refresh_pressure_pipe_controls()
        if col in (2, 24):
            nodes_for_view = self._build_nodes_from_table()
            self._update_pressure_pipe_roughness_overview(
                self._collect_pressure_pipe_roughness_pairs_from_nodes(nodes_for_view)
            )

    def _append_loss_undo_snapshot(self, snapshot):
        """将单元格编辑快照压入撤销栈（统一入口）。"""
        self._loss_undo_stack.append(snapshot)
        if len(self._loss_undo_stack) > 20:
            self._loss_undo_stack.pop(0)
        self._loss_redo_stack.clear()

    def _push_undo_snapshot(self):
        """记录当前表格状态到撤销栈（Delete 键删除前调用）"""
        snapshot = self._snapshot_editable_cols()
        self._append_loss_undo_snapshot(snapshot)
        self._pre_edit_snapshot = None
        # 设置标志，跳过本次 Delete 触发的 N 次 cellChanged 快照记录
        editable_count = 0
        for idx in self.node_table.selectedIndexes():
            item = self.node_table.item(idx.row(), idx.column())
            if item is not None and (item.flags() & Qt.ItemIsEditable):
                editable_count += 1
        self._undo_group = max(1, editable_count)

    def _undo_loss_edit(self):
        """Ctrl+Z 撤销：优先撤销行操作，其次撤销单元格编辑"""
        # 优先检查行操作撤销栈
        if self._node_table_undo_stack:
            self._undo_node_table()
            return
        # 再检查单元格编辑撤销栈
        if not self._loss_undo_stack:
            return
        self._updating_cells = True
        try:
            snapshot = self._loss_undo_stack.pop()
            # 保存当前状态到重做栈
            current = {}
            for (r, c) in snapshot.keys():
                item = self.node_table.item(r, c)
                current[(r, c)] = item.text() if item else ""
            self._loss_redo_stack.append(current)
            if len(self._loss_redo_stack) > 20:
                self._loss_redo_stack.pop(0)
            table = self.node_table
            for (r, c), text in snapshot.items():
                item = table.item(r, c)
                if item is None:
                    item = QTableWidgetItem("")
                    item.setTextAlignment(Qt.AlignCenter)
                    if c not in EDITABLE_COLS:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    table.setItem(r, c, item)
                item.setText(text)

            # 同步 calculated_nodes
            def _rf(r, c):
                item = table.item(r, c)
                if not item:
                    return 0.0
                txt = item.text().strip()
                if not txt or txt == '-':
                    return 0.0
                try:
                    return float(txt)
                except ValueError:
                    return 0.0

            if hasattr(self, 'calculated_nodes') and self.calculated_nodes:
                for r in range(min(table.rowCount(), len(self.calculated_nodes))):
                    node = self.calculated_nodes[r]
                    if node.is_transition:
                        node.head_loss_cumulative = _rf(r, 40)
                    else:
                        node.head_loss_reserve = _rf(r, 36)
                        node.head_loss_gate = _rf(r, 37)
                        node.head_loss_siphon = _rf(r, 38)
                        node.head_loss_total = _rf(r, 39)
                        node.head_loss_cumulative = _rf(r, 40)
                        node.water_level = _rf(r, 41)
                        be = _rf(r, 42)
                        if be:
                            node.bottom_elevation = be
                        te = _rf(r, 43)
                        if te:
                            node.top_elevation = te

            InfoBar.success("已撤销", "已恢复上一步操作",
                           parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
        finally:
            self._updating_cells = False

    def _redo_loss_edit(self):
        """Ctrl+Y 重做：优先重做行操作，其次重做单元格编辑"""
        # 优先检查行操作重做栈
        if self._node_table_redo_stack:
            self._redo_node_table()
            return
        # 再检查单元格编辑重做栈
        if not self._loss_redo_stack:
            return
        self._updating_cells = True
        try:
            snapshot = self._loss_redo_stack.pop()
            # 保存当前状态到撤销栈
            current = {}
            for (r, c) in snapshot.keys():
                item = self.node_table.item(r, c)
                current[(r, c)] = item.text() if item else ""
            self._loss_undo_stack.append(current)
            if len(self._loss_undo_stack) > 20:
                self._loss_undo_stack.pop(0)
            table = self.node_table
            for (r, c), text in snapshot.items():
                item = table.item(r, c)
                if item is None:
                    item = QTableWidgetItem("")
                    item.setTextAlignment(Qt.AlignCenter)
                    if c not in EDITABLE_COLS:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    table.setItem(r, c, item)
                item.setText(text)

            # 同步 calculated_nodes
            def _rf(r, c):
                item = table.item(r, c)
                if not item:
                    return 0.0
                txt = item.text().strip()
                if not txt or txt == '-':
                    return 0.0
                try:
                    return float(txt)
                except ValueError:
                    return 0.0

            if hasattr(self, 'calculated_nodes') and self.calculated_nodes:
                for r in range(min(table.rowCount(), len(self.calculated_nodes))):
                    node = self.calculated_nodes[r]
                    if node.is_transition:
                        node.head_loss_cumulative = _rf(r, 40)
                    else:
                        node.head_loss_reserve = _rf(r, 36)
                        node.head_loss_gate = _rf(r, 37)
                        node.head_loss_siphon = _rf(r, 38)
                        node.head_loss_total = _rf(r, 39)
                        node.head_loss_cumulative = _rf(r, 40)
                        node.water_level = _rf(r, 41)
                        be = _rf(r, 42)
                        if be:
                            node.bottom_elevation = be
                        te = _rf(r, 43)
                        if te:
                            node.top_elevation = te

            InfoBar.success("已重做", "已恢复上一步撤销的操作",
                           parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
        finally:
            self._updating_cells = False

    # ================================================================
    # 节点表行操作撤销（添加/删除/插入/复制/清空）
    # ================================================================
    def _snapshot_node_table(self):
        """保存完整节点表状态（用于行操作撤销）"""
        snapshot = {
            'rows': [],
            'row_meta': [],
            'structure_heights': dict(self._node_structure_heights),
            'chamfer_params': dict(self._node_chamfer_params),
            'u_params': dict(self._node_u_params),
            'velocity_increased': dict(self._node_velocity_increased),
        }
        for r in range(self.node_table.rowCount()):
            row_data = []
            for c in range(self.node_table.columnCount()):
                item = self.node_table.item(r, c)
                row_data.append(item.text() if item else "")
            snapshot['rows'].append(row_data)
            first_item = self.node_table.item(r, 0)
            snapshot['row_meta'].append(first_item.data(Qt.UserRole) if first_item else None)
        return snapshot

    def _restore_node_table(self, snapshot):
        """从快照恢复完整节点表状态"""
        self._updating_cells = True
        try:
            # 恢复表格内容
            self.node_table.setRowCount(0)
            for r, row_data in enumerate(snapshot['rows']):
                self.node_table.insertRow(r)
                for c, text in enumerate(row_data):
                    item = QTableWidgetItem(text)
                    item.setTextAlignment(Qt.AlignCenter)
                    if c not in EDITABLE_COLS:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    # 第一行锁定水头损失列
                    if r == 0 and c in FIRST_ROW_LOCKED_LOSS_COLS:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.node_table.setItem(r, c, item)
                row_meta = snapshot.get('row_meta', [])
                if r < len(row_meta):
                    first_item = self.node_table.item(r, 0)
                    if first_item is not None:
                        first_item.setData(Qt.UserRole, row_meta[r])
            # 恢复缓存字典
            self._node_structure_heights = dict(snapshot['structure_heights'])
            self._node_chamfer_params = dict(snapshot['chamfer_params'])
            self._node_u_params = dict(snapshot['u_params'])
            self._node_velocity_increased = dict(snapshot.get('velocity_increased', {}))
        finally:
            self._updating_cells = False
        self._refresh_pressure_pipe_controls()
        nodes_for_view = self._build_nodes_from_table()
        self._update_pressure_pipe_roughness_overview(
            self._collect_pressure_pipe_roughness_pairs_from_nodes(nodes_for_view)
        )
        self._refresh_pressure_pipe_controls()

    def _push_node_table_undo(self):
        """在行操作前记录快照到撤销栈"""
        self._node_table_undo_stack.append(self._snapshot_node_table())
        if len(self._node_table_undo_stack) > 20:
            self._node_table_undo_stack.pop(0)
        self._node_table_redo_stack.clear()

    def _undo_node_table(self):
        """撤销行操作"""
        if not self._node_table_undo_stack:
            return False
        # 保存当前状态到重做栈
        self._node_table_redo_stack.append(self._snapshot_node_table())
        if len(self._node_table_redo_stack) > 20:
            self._node_table_redo_stack.pop(0)
        # 恢复上一步状态
        self._restore_node_table(self._node_table_undo_stack.pop())
        InfoBar.success("已撤销", "已恢复上一步操作",
                       parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
        return True

    def _redo_node_table(self):
        """重做行操作"""
        if not self._node_table_redo_stack:
            return False
        # 保存当前状态到撤销栈
        self._node_table_undo_stack.append(self._snapshot_node_table())
        if len(self._node_table_undo_stack) > 20:
            self._node_table_undo_stack.pop(0)
        # 恢复下一步状态
        self._restore_node_table(self._node_table_redo_stack.pop())
        InfoBar.success("已重做", "已恢复上一步撤销的操作",
                       parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
        return True

    def _recalc_downstream(self, edited_row):
        """从 edited_row 开始，重算总水头损失 / 累计 / 水位 / 高程"""
        table = self.node_table
        row_count = table.rowCount()
        if row_count == 0:
            return

        def _rf(r, c):
            """读取单元格浮点值，'-' 或空视为 0"""
            item = table.item(r, c)
            if not item:
                return 0.0
            txt = item.text().strip()
            if not txt or txt == '-':
                return 0.0
            try:
                return float(txt)
            except ValueError:
                return 0.0

        def _set(r, c, val, fmt=".4f"):
            """写入单元格（保持居中对齐和只读标记）"""
            item = table.item(r, c)
            if item is None:
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                if c not in EDITABLE_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                table.setItem(r, c, item)
            item.setText(f"{val:{fmt}}" if val is not None else "-")

        def _is_transition_row(r):
            item = table.item(r, 2)
            return item and item.text().strip() == "渐变段"

        # ── 1. 重算编辑行的总水头损失 (col 39) ──
        if not _is_transition_row(edited_row):
            h_bend = _rf(edited_row, 34)
            h_friction = _rf(edited_row, 35)
            h_reserve = _rf(edited_row, 36)
            h_gate = _rf(edited_row, 37)
            h_siphon = _rf(edited_row, 38)
            new_total = h_bend + h_friction + h_reserve + h_gate + h_siphon
            _set(edited_row, 39, new_total)
            # 同步更新node对象，确保双击时读取到最新值
            if edited_row < len(self.calculated_nodes):
                node = self.calculated_nodes[edited_row]
                if not getattr(node, 'is_transition', False):
                    node.head_loss_total = new_total

        # ── 2. 从编辑行开始重算累计总水头损失 (col 40) ──
        cumulative = _rf(edited_row - 1, 40) if edited_row > 0 else 0.0
        for r in range(edited_row, row_count):
            if _is_transition_row(r):
                cumulative += _rf(r, 33)  # 渐变段水头损失
            else:
                cumulative += _rf(r, 39)  # 总水头损失
            _set(r, 40, cumulative)

        # ── 3. 从编辑行开始重算水位 (col 41)、渠底高程 (col 42)、渠顶高程 (col 43) ──
        start_wl = self._fval(self.start_wl_edit, 0.0)
        if start_wl <= 0:
            return  # 无起始水位，无法递推

        def _find_prev_regular_row(start_row):
            for rr in range(start_row - 1, -1, -1):
                if not _is_transition_row(rr):
                    return rr
            return -1

        prev_regular_row = _find_prev_regular_row(edited_row)
        prev_wl = _rf(prev_regular_row, 41) if prev_regular_row >= 0 else start_wl

        for r in range(edited_row, row_count):
            if _is_transition_row(r):
                continue

            if prev_regular_row < 0:
                # 第一个常规节点：水位 = 起始水位
                wl = start_wl
            else:
                # 后续常规节点：水位 = 上一常规节点水位 - 本行总损失 - 中间渐变段损失
                transition_loss = 0.0
                for j in range(prev_regular_row + 1, r):
                    if _is_transition_row(j):
                        transition_loss += _rf(j, 33)
                # 使用本行"增量总损失"(col39)，不能用累计值(col40)重复扣减
                total_drop = _rf(r, 39) + transition_loss
                wl = prev_wl - total_drop

            _set(r, 41, wl, ".3f")

            # 渠底高程 = 水位 - 水深
            h_depth = _rf(r, 27)
            if h_depth > 0:
                be = wl - h_depth
                _set(r, 42, be, ".3f")
                # 渠顶高程 = 渠底高程 + 结构高度
                sh = self._node_structure_heights.get(r, 0.0)
                if sh > 0:
                    _set(r, 43, be + sh, ".3f")

            prev_wl = wl
            prev_regular_row = r

        # 同步更新 calculated_nodes（如果存在），保证双击查看详情时数据一致
        if hasattr(self, 'calculated_nodes') and self.calculated_nodes:
            for r in range(min(row_count, len(self.calculated_nodes))):
                node = self.calculated_nodes[r]
                if node.is_transition:
                    node.head_loss_cumulative = _rf(r, 40)
                else:
                    node.head_loss_reserve = _rf(r, 36)
                    node.head_loss_gate = _rf(r, 37)
                    node.head_loss_siphon = _rf(r, 38)
                    node.head_loss_total = _rf(r, 39)
                    node.head_loss_cumulative = _rf(r, 40)
                    node.water_level = _rf(r, 41)
                    be = _rf(r, 42)
                    if be:
                        node.bottom_elevation = be
                    te = _rf(r, 43)
                    if te:
                        node.top_elevation = te

    # ================================================================
    def _add_node_row(self, data=None, _skip_undo=False):
        """添加一行节点，_skip_undo=True 时跳过撤销快照（内部调用用）"""
        if not _skip_undo:
            self._push_node_table_undo()
        row = self.node_table.rowCount()
        self.node_table.insertRow(row)
        total_cols = len(NODE_ALL_HEADERS)
        for col in range(total_cols):
            if data and isinstance(data, (list, tuple)) and col < len(data) and data[col]:
                item = QTableWidgetItem(str(data[col]))
            elif col == 2:
                item = QTableWidgetItem("明渠-梯形")
            else:
                item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            if col not in EDITABLE_COLS:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            # 第一行（水位起点）锁定水头损失列
            if row == 0 and col in FIRST_ROW_LOCKED_LOSS_COLS:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.node_table.setItem(row, col, item)
        self._refresh_pressure_pipe_controls()

    def _clear_nodes(self):
        if self.node_table.rowCount() > 0:
            self._push_node_table_undo()
        self.node_table.setRowCount(0)
        self._node_structure_heights.clear()
        self._node_chamfer_params.clear()
        self._node_u_params.clear()
        self._node_velocity_increased.clear()
        self.calculated_nodes = []
        self.nodes = []
        if hasattr(self, 'siphon_roughness_chips'):
            self.siphon_roughness_chips.clear()
        if hasattr(self, 'pressure_pipe_roughness_chips'):
            self.pressure_pipe_roughness_chips.clear()
        self._refresh_pressure_pipe_controls()

    def _get_node_row_data(self, row):
        data = []
        for col in range(self.node_table.columnCount()):
            item = self.node_table.item(row, col)
            data.append(item.text() if item else "")
        return data

    # ================================================================
    # 从批量计算导入
    # ================================================================
    def _import_from_batch(self):
        if not SHARED_DATA_AVAILABLE:
            InfoBar.warning("不可用", "SharedDataManager未加载，无法导入", parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        shared_data = get_shared_data_manager()
        results = shared_data.get_batch_results()
        if not results:
            InfoBar.warning("无数据", "批量计算模块尚无结果。请先在「批量计算」模块完成计算。",
                           parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            return

        # 同步批量计算面板中的渠道基础信息
        self._sync_batch_settings()

        self._updating_cells = True
        self._clear_nodes()
        imported = 0
        flow_segment_map = {}  # {流量段编号: 设计流量}
        general_roughness_vals = []   # 收集非倒虹吸行的糙率
        siphon_roughness_pairs = []   # 收集倒虹吸行的 (名称, 糙率)
        pressure_pipe_params_pairs = []  # 收集有压管道行的 (名称, 管材名称)

        # 映射结构形式名称（兼容简化名称 → 完整名称）
        struct_map = {
            "梯形": "明渠-梯形", "矩形": "明渠-矩形", "圆形": "明渠-圆形",
            "U形": "渡槽-U形", "U形渡槽": "渡槽-U形", "隧洞": "隧洞-圆形", "渡槽": "渡槽-U形",
            # 兜底：计算引擎返回的简化名（正常流程已在batch注册时修正）
            "圆拱直墙型": "隧洞-圆拱直墙型",
            "马蹄形标准Ⅰ型": "隧洞-马蹄形Ⅰ型", "马蹄形标准Ⅱ型": "隧洞-马蹄形Ⅱ型",
            "矩形暗涵": "矩形暗涵",
        }

        for sr in results:
            flow_section = str(getattr(sr, 'flow_section', ''))
            building_name = str(getattr(sr, 'building_name', ''))
            section_type = str(getattr(sr, 'section_type', ''))
            raw_result = getattr(sr, 'raw_result', {}) or {}
            x = getattr(sr, 'coord_X', 0)
            y = getattr(sr, 'coord_Y', 0)
            B = getattr(sr, 'B', None) or ""
            D = getattr(sr, 'D', None) or ""
            R = getattr(sr, 'R', None) or ""
            m_val = getattr(sr, 'm', None) or ""
            n_val = getattr(sr, 'n', 0) or ""
            slope_inv = getattr(sr, 'slope_inv', 0) or ""
            Q = getattr(sr, 'Q', 0) or ""
            pipe_material = str(
                getattr(sr, 'pipe_material', '') or raw_result.get('pipe_material', '')
            ).strip()
            local_loss_ratio = (
                getattr(sr, 'local_loss_ratio', 0.0)
                if getattr(sr, 'local_loss_ratio', None) is not None
                else raw_result.get('local_loss_ratio', 0.0)
            )
            in_out_raw = str(
                getattr(sr, 'in_out_raw', '') or raw_result.get('in_out_raw', '')
            ).strip()

            if section_type in struct_map:
                section_type = struct_map[section_type]
            # 模糊匹配增强
            elif "渡槽-U" in section_type or "U形渡槽" in section_type:
                section_type = "渡槽-U形"
            elif "渡槽-矩形" in section_type:
                section_type = "渡槽-矩形"
            elif "隧洞-圆拱直墙" in section_type:
                section_type = "隧洞-圆拱直墙型"
            elif "隧洞-马蹄形Ⅰ" in section_type:
                section_type = "隧洞-马蹄形Ⅰ型"
            elif "隧洞-马蹄形Ⅱ" in section_type:
                section_type = "隧洞-马蹄形Ⅱ型"
            elif "暗涵" in section_type:
                section_type = "矩形暗涵"

            # 收集糙率分类（倒虹吸 vs 一般建筑物）
            try:
                _n_float = float(n_val) if n_val and str(n_val).strip() else 0.0
            except (ValueError, TypeError):
                _n_float = 0.0
            if _n_float > 0:
                if "倒虹吸" in section_type:
                    siphon_roughness_pairs.append((building_name or f"倒虹吸{len(siphon_roughness_pairs)+1}", _n_float))
                elif "有压管道" in section_type:
                    # 有压管道收集管材信息而非糙率
                    if pipe_material:
                        pressure_pipe_params_pairs.append(
                            (building_name or f"有压管道{len(pressure_pipe_params_pairs)+1}", pipe_material)
                        )
                else:
                    # 排除闸类占位行（闸类无糙率意义）
                    if "闸" not in section_type and "分水" not in section_type:
                        general_roughness_vals.append(_n_float)

            # 收集流量段信息
            try:
                seg_num = int(flow_section)
            except (ValueError, TypeError):
                seg_num = 1
            q_val = float(Q) if Q and str(Q).strip() else 0.0
            if seg_num not in flow_segment_map and q_val > 0:
                flow_segment_map[seg_num] = q_val

            def fmt(v):
                if v is None or v == "" or v == 0: return ""
                if isinstance(v, float): return f"{v:.4f}" if v < 1 else f"{v:.3f}"
                return str(v)

            # 提取水力计算结果（与原版Tkinter _do_import_from_calc_result对齐）
            h_val = getattr(sr, 'h', None) or 0.0
            V_val = getattr(sr, 'V', 0) or 0.0
            V_max_val = getattr(sr, 'V_max', 0) or 0.0  # 加大流速
            A_val = getattr(sr, 'A', 0) or 0.0
            X_val = getattr(sr, 'X', 0) or 0.0
            R_hyd_val = getattr(sr, 'R_hydraulic', 0) or 0.0
            H_total = getattr(sr, 'H_total', 0) or 0.0

            tr_r = getattr(sr, 'turn_radius', 0.0) or 0.0
            row_data = [""] * len(NODE_ALL_HEADERS)
            row_data[0] = flow_section
            row_data[1] = building_name
            row_data[2] = section_type
            row_data[5] = fmt(x)
            row_data[6] = fmt(y)
            row_data[7] = fmt(tr_r) if tr_r > 0 else ""
            row_data[20] = fmt(B) if B else ""
            row_data[21] = fmt(D) if D else ""
            row_data[22] = fmt(R) if R else ""
            row_data[23] = fmt(m_val) if m_val else ""
            row_data[24] = fmt(n_val)
            row_data[25] = fmt(slope_inv)
            row_data[26] = fmt(Q)
            self._add_node_row(row_data, _skip_undo=True)

            # 写入水力结果到结果列（原版通过set_nodes写入water_depth/velocity等）
            cur_row = self.node_table.rowCount() - 1
            # 通过行级元数据透传有压管道专用参数（表格无专门列）
            if pipe_material or in_out_raw or (local_loss_ratio and float(local_loss_ratio) > 0):
                first_item = self.node_table.item(cur_row, 0)
                if first_item:
                    payload = first_item.data(Qt.UserRole)
                    if not isinstance(payload, dict):
                        payload = {}
                    if pipe_material:
                        payload['_pipe_material'] = pipe_material
                    if in_out_raw:
                        payload['_in_out_raw'] = in_out_raw
                    try:
                        llr = float(local_loss_ratio)
                        if llr > 0:
                            payload['_local_loss_ratio'] = llr
                    except (ValueError, TypeError):
                        pass
                    first_item.setData(Qt.UserRole, payload)
            if h_val and float(h_val) > 0:
                _item = QTableWidgetItem(f"{float(h_val):.3f}")
                _item.setTextAlignment(Qt.AlignCenter)
                _item.setFlags(_item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(cur_row, 27, _item)  # 水深h
            if A_val and float(A_val) > 0:
                _item = QTableWidgetItem(f"{float(A_val):.3f}")
                _item.setTextAlignment(Qt.AlignCenter)
                _item.setFlags(_item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(cur_row, 28, _item)  # 过水断面面积A
            if X_val and float(X_val) > 0:
                _item = QTableWidgetItem(f"{float(X_val):.3f}")
                _item.setTextAlignment(Qt.AlignCenter)
                _item.setFlags(_item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(cur_row, 29, _item)  # 湿周X
            if R_hyd_val and float(R_hyd_val) > 0:
                _item = QTableWidgetItem(f"{float(R_hyd_val):.3f}")
                _item.setTextAlignment(Qt.AlignCenter)
                _item.setFlags(_item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(cur_row, 30, _item)  # 水力半径R
            if V_val and float(V_val) > 0:
                _item = QTableWidgetItem(f"{float(V_val):.3f}")
                _item.setTextAlignment(Qt.AlignCenter)
                _item.setFlags(_item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(cur_row, 31, _item)  # 流速v
            # 缓存结构高度（与Tkinter版 data_table._node_structure_heights 对齐）
            if H_total and float(H_total) > 0:
                self._node_structure_heights[cur_row] = float(H_total)
            # 缓存倒角参数（渡槽-矩形专用，不占用表格列）
            if "渡槽-矩形" in section_type:
                _raw = getattr(sr, 'raw_result', {}) or {}
                _ca = _raw.get('chamfer_angle', 0) or 0
                _cl = _raw.get('chamfer_length', 0) or 0
                if _ca > 0 and _cl > 0:
                    self._node_chamfer_params[cur_row] = {'chamfer_angle': float(_ca), 'chamfer_length': float(_cl)}

            # 缓存明渠-U形的圆心角
            if "明渠-U形" in section_type:
                _raw_u = getattr(sr, 'raw_result', {}) or {}
                _theta = _raw_u.get('theta_deg', 0) or 0
                if _theta > 0:
                    self._node_u_params[cur_row] = {'theta_deg': float(_theta)}

            # 缓存加大流速（用于倒虹吸水力计算时自动填入v₁加大/v₃加大）
            if V_max_val and float(V_max_val) > 0:
                self._node_velocity_increased[cur_row] = float(V_max_val)

            imported += 1

        auto_resize_table(self.node_table)

        # 自动填充渠道糙率——值不一致时弹窗让用户选择
        chosen_n = self._choose_roughness_value(general_roughness_vals, "渠道糙率")
        if chosen_n is not None:
            self.roughness_edit.setText(f"{chosen_n:.4f}".rstrip('0').rstrip('.'))
        # 倒虹吸糙率只读概览（每个倒虹吸独立显示）
        self._update_siphon_roughness_overview(siphon_roughness_pairs)
        self._update_pressure_pipe_roughness_overview(pressure_pipe_params_pairs)

        # 自动填充多流量段设计流量和加大流量
        if flow_segment_map:
            sorted_segs = sorted(flow_segment_map.keys())
            design_flows = [flow_segment_map[s] for s in sorted_segs]
            flow_strs = []
            for q in design_flows:
                formatted = f"{q:.3f}".rstrip('0').rstrip('.')
                flow_strs.append(formatted)
            self.design_flow_edit.setText(", ".join(flow_strs))
            self._on_design_flow_changed()

        # 自动计算推荐转弯半径
        if CALCULATOR_AVAILABLE:
            nodes = self._build_nodes_from_table()
            recommended_r = self._calculate_recommended_turn_radius(nodes)
            if recommended_r > 0:
                self.turn_radius_edit.setText(f"{recommended_r:.1f}")

        # 检查是否包含倒虹吸
        has_siphon = False
        has_pressure_pipe = False
        if CALCULATOR_AVAILABLE:
            nodes = self._build_nodes_from_table()
            has_siphon = any(
                n.structure_type and "倒虹吸" in n.structure_type.value
                for n in nodes if n.structure_type
            )
            has_pressure_pipe = any(
                n.structure_type and "有压管道" in n.structure_type.value
                for n in nodes if n.structure_type
            )

        self._updating_cells = False

        # 触发几何计算（与原版Tkinter recalculate对齐）
        self._recalculate_geometry()
        self.nodes = self._build_nodes_from_table()

        next_steps = "请依次点击【插入渐变段】→"
        if has_siphon:
            next_steps += "【倒虹吸水力计算】→"
        if has_pressure_pipe:
            next_steps += "【有压管道水力计算】→"
        next_steps += "【执行计算】"

        InfoBar.success("导入成功",
                       f"已导入 {imported} 个节点，已自动填充流量和推荐转弯半径（全局）。{next_steps}",
                       parent=self._info_parent(), duration=6000, position=InfoBarPosition.TOP)

    def _recalculate_geometry(self):
        """
        导入数据后触发几何计算（与原版Tkinter recalculate对齐）

        流程：
        1. 填充转弯半径列：普通行用全局半径；倒虹吸行临时写 n×D（供几何算法用）；
           首行/闸类/分水口行不写（保持空白）
        2. 构建节点 → calculate_geometry → preprocess_nodes
        3. 回写几何结果列(8-19) + 进出口(3) + IP(4)；
           倒虹吸行清空临时 n×D（倒虹吸计算后由 import_losses_callback 写入真实值）；
           首行/闸类不写 col 7
        """
        self._updating_cells = True
        try:
            self._recalculate_geometry_impl()
        finally:
            self._updating_cells = False

    def _recalculate_geometry_impl(self):
        if not CALCULATOR_AVAILABLE:
            return

        # ---- 1. 填充转弯半径列 (col 7) ----
        # 优先级：col 7 已有非零值(导入/手动填写) → 保留；否则按规则填充
        # 规则：首行/闸类/分水口 → 不填（不参与弯道几何计算）
        #       倒虹吸/有压管道行 → 临时写 n×D 供几何计算用，Step3 中清空（写回值除外）
        #       普通行  → 使用全局转弯半径
        turn_radius = self._fval(self.turn_radius_edit, 0)
        siphon_n = DEFAULT_SIPHON_TURN_RADIUS_N
        # 记录倒虹吸/有压管道行中已有明确写回值的行（Step3 中保留，不清空）
        pressurized_rows_with_existing = set()
        for r in range(self.node_table.rowCount()):
            existing_r = 0.0
            ei = self.node_table.item(r, 7)
            if ei:
                try:
                    v = ei.text().strip()
                    existing_r = float(v) if v else 0.0
                except (ValueError, TypeError):
                    pass
            struct_item = self.node_table.item(r, 2)
            struct_text = struct_item.text().strip() if struct_item else ""
            _is_siphon = "倒虹吸" in struct_text
            _is_pressure_pipe = "有压管道" in struct_text
            _is_gate = "闸" in struct_text or "分水" in struct_text
            if existing_r > 0:
                if _is_siphon or _is_pressure_pipe:
                    pressurized_rows_with_existing.add(r)
                continue  # 保留导入/手动输入/写回的转弯半径
            if r == 0 or _is_gate:
                continue  # 首行/闸类：不填转弯半径
            if _is_siphon or _is_pressure_pipe:
                # 倒虹吸/有压管道行：临时写 n×D 供几何计算（Step3 会清空）
                d_item = self.node_table.item(r, 21)  # 直径D
                d_val = 0.0
                if d_item:
                    try: d_val = float(d_item.text())
                    except (ValueError, TypeError): pass
                r_val = siphon_n * d_val if d_val > 0 else turn_radius
                if r_val > 0:
                    item = QTableWidgetItem(f"{r_val:.1f}")
                    item.setTextAlignment(Qt.AlignCenter)
                    self.node_table.setItem(r, 7, item)
            elif turn_radius > 0:
                # 普通行：使用全局转弯半径
                item = QTableWidgetItem(f"{turn_radius:.1f}")
                item.setTextAlignment(Qt.AlignCenter)
                self.node_table.setItem(r, 7, item)

        # ---- 2. 构建节点 & 几何计算 ----
        nodes = self._build_nodes_from_table()
        if len(nodes) < 2:
            return

        settings = self._build_settings()
        if not settings:
            return

        calculator = WaterProfileCalculator(settings)
        # 先几何计算（此时in_out=NORMAL，所有转角都被计算），再设置进出口
        calculator.calculate_geometry(nodes)
        calculator.preprocess_nodes(nodes)

        # ---- 3. 回写几何结果到表格 ----
        prefix = settings.get_station_prefix() if hasattr(settings, 'get_station_prefix') else ""

        for r, node in enumerate(nodes):
            if r >= self.node_table.rowCount():
                break

            # 进出口判断 (col 3)
            in_out_str = node.get_in_out_str() if hasattr(node, 'get_in_out_str') else ""
            item = QTableWidgetItem(in_out_str)
            item.setTextAlignment(Qt.AlignCenter)
            self.node_table.setItem(r, 3, item)

            # IP编号 (col 4)
            ip_str = node.get_ip_str() if hasattr(node, 'get_ip_str') else ""
            item = QTableWidgetItem(ip_str)
            item.setTextAlignment(Qt.AlignCenter)
            self.node_table.setItem(r, 4, item)

            # 转弯半径 (col 7) — 按规则写回
            # 首行/闸类：不写（保持空白）
            # 倒虹吸/有压管道行：若已有写回值则保留；否则清空临时 n×D
            # 普通行：写回计算值（与原逻辑一致）
            _st_r3 = node.get_structure_type_str() if hasattr(node, 'get_structure_type_str') else ""
            _is_siphon_r3 = "倒虹吸" in _st_r3
            _is_pressure_pipe_r3 = "有压管道" in _st_r3
            _is_gate_r3 = "闸" in _st_r3 or "分水" in _st_r3
            if r == 0 or _is_gate_r3:
                pass  # 首行/闸类：不写 col 7
            elif _is_siphon_r3 or _is_pressure_pipe_r3:
                if r not in pressurized_rows_with_existing:
                    # 清空临时写入的 n×D，等待倒虹吸/有压管道计算后写回真实值
                    self.node_table.setItem(r, 7, QTableWidgetItem(""))
                # else: pressurized_rows_with_existing 中的行已有写回值，保留不动
            elif node.turn_radius and node.turn_radius > 0:
                item = QTableWidgetItem(f"{node.turn_radius:.1f}")
                item.setTextAlignment(Qt.AlignCenter)
                self.node_table.setItem(r, 7, item)

            # 几何结果列 (8-19) — 无条件格式化，0值也显示（与Tkinter一致）
            fmt_s = lambda s: ProjectSettings.format_station(s, prefix) if s is not None else ""
            geo_data = {
                8:  f"{node.turn_angle:.4f}",
                9:  f"{node.tangent_length:.6f}",
                10: f"{node.arc_length:.6f}",
                11: f"{node.curve_length:.6f}",
                12: f"{node.straight_distance:.6f}",
                13: fmt_s(node.station_ip),
                14: fmt_s(node.station_BC),
                15: fmt_s(getattr(node, 'station_MC', None)),
                16: fmt_s(node.station_EC),
                17: f"{getattr(node, 'check_pre_curve', 0):.3f}",
                18: f"{getattr(node, 'check_post_curve', 0):.3f}",
                19: f"{getattr(node, 'check_total_length', 0):.3f}",
            }
            for c, v in geo_data.items():
                item = QTableWidgetItem(str(v))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(r, c, item)

        auto_resize_table(self.node_table)

    def fill_turn_radius_for_geometry(self, nodes, n):
        """
        为有压管道节点填充临时转弯半径 R = n × D
        
        在几何计算前调用，为有压管道节点临时填充转弯半径值供几何计算使用。
        
        Args:
            nodes: ChannelNode 列表
            n: 转弯半径倍数（R = n × D）
            
        Requirements: 8.1, 8.2, 8.3
        """
        if not nodes:
            return
        
        for node in nodes:
            # 只处理有压管道节点
            if not node.is_pressure_pipe:
                continue
            
            # 如果已有非零转弯半径值，保留不变（可能是用户导入或手动输入的值）
            if node.turn_radius and node.turn_radius > 0:
                continue
            
            # 从 section_params 中获取管径 D
            diameter_D = node.section_params.get('D', 0.0)
            if diameter_D <= 0:
                continue
            
            # 计算临时转弯半径 R = n × D
            node.turn_radius = n * diameter_D
    
    def clear_temporary_turn_radius(self, nodes):
        """
        清空有压管道节点的临时转弯半径
        
        在几何计算后调用，清空临时写入的转弯半径值。
        但保留从水力计算回写的值（检查 head_loss_siphon/external_head_loss 字段）。
        
        Args:
            nodes: ChannelNode 列表
            
        Requirements: 8.4, 8.5
        """
        if not nodes:
            return
        
        for node in nodes:
            # 只处理有压管道节点
            if not node.is_pressure_pipe:
                continue
            
            # 已写回损失值（新字段或旧兼容字段）时，保留转弯半径不清空
            ext_loss = getattr(node, 'external_head_loss', None)
            siphon_loss = getattr(node, 'head_loss_siphon', 0.0) or 0.0
            if (ext_loss is not None and ext_loss > 0) or siphon_loss > 0:
                continue
            
            # 清空临时转弯半径
            node.turn_radius = 0.0

    def _sync_batch_settings(self):
        """从批量计算面板同步渠道基础信息"""
        try:
            main_win = self.window()
            if main_win and hasattr(main_win, 'batch_panel'):
                bp = main_win.batch_panel
                if hasattr(bp, 'channel_name_edit'):
                    name = bp.channel_name_edit.text().strip()
                    if name: self.channel_name_edit.setText(name)
                if hasattr(bp, 'channel_level_combo'):
                    level = bp.channel_level_combo.currentText()
                    if level:
                        idx = self.channel_level_combo.findText(level)
                        if idx >= 0: self.channel_level_combo.setCurrentIndex(idx)
                if hasattr(bp, 'start_wl_edit'):
                    wl = bp.start_wl_edit.text().strip()
                    if wl: self.start_wl_edit.setText(wl)
                if hasattr(bp, 'start_station_edit'):
                    st = bp.start_station_edit.text().strip()
                    if st: self.start_station_edit.setText(st)
        except Exception:
            pass

    # ================================================================
    # 计算
    # ================================================================
    def _build_settings(self):
        """从UI读取设置，构建ProjectSettings"""
        if not CALCULATOR_AVAILABLE:
            return None
        settings = ProjectSettings()
        settings.channel_name = self.channel_name_edit.text().strip() or "未命名渠道"
        settings.channel_level = self.channel_level_combo.currentText()
        settings.start_water_level = self._fval(self.start_wl_edit, 100.0)
        settings.start_station = parse_station_input(self.start_station_edit.text())
        # 多流量段支持
        design_flows = self._parse_flow_values(self.design_flow_edit.text())
        max_flows = self._parse_flow_values(self.max_flow_edit.text())
        settings.design_flows = design_flows
        settings.max_flows = max_flows
        settings.design_flow = design_flows[0] if design_flows else 0.0
        settings.max_flow = max_flows[0] if max_flows else 0.0
        settings.roughness = self._fval(self.roughness_edit, DEFAULT_ROUGHNESS)
        # siphon_roughness 不再从单一输入框读取（已改为只读概览），保留默认值
        # 每个倒虹吸的实际糙率从节点表格对应行读取
        settings.turn_radius = self._fval(self.turn_radius_edit, DEFAULT_TURN_RADIUS)
        # 渡槽/隧洞渐变段设置
        settings.transition_inlet_form = self.trans_inlet_combo.currentText()
        settings.transition_inlet_zeta = self._fval(self.trans_inlet_zeta, 0.10)
        settings.transition_outlet_form = self.trans_outlet_combo.currentText()
        settings.transition_outlet_zeta = self._fval(self.trans_outlet_zeta, 0.20)
        # 明渠渐变段设置
        settings.open_channel_transition_form = self.oc_trans_combo.currentText()
        settings.open_channel_transition_zeta = self._fval(self.oc_trans_zeta, 0.10)
        # 倒虹吸渐变段设置
        settings.siphon_transition_inlet_form = self.siphon_inlet_combo.currentText()
        settings.siphon_transition_inlet_zeta = self._fval(self.siphon_inlet_zeta, 0.10)
        settings.siphon_transition_outlet_form = self.siphon_outlet_combo.currentText()
        settings.siphon_transition_outlet_zeta = self._fval(self.siphon_outlet_zeta, 0.20)
        # 倒虹吸转弯半径倍数n
        settings.siphon_turn_radius_n = DEFAULT_SIPHON_TURN_RADIUS_N
        return settings

    def _build_nodes_from_table(self):
        """从节点表格构建ChannelNode列表（与原版Tkinter get_nodes完全对齐）"""
        if not CALCULATOR_AVAILABLE:
            return []

        # 辅助函数提到循环外避免每行重复定义（#16）
        table = self.node_table

        def _read_float(row, col):
            item = table.item(row, col)
            if item:
                try:
                    return float(item.text())
                except (ValueError, TypeError):
                    pass
            return 0.0

        def _read_text(row, col):
            item = table.item(row, col)
            if item:
                t = item.text().strip()
                if t and t != '-':
                    return t
            return ""

        def _parse_station(text):
            """解析格式化桩号文本（如 '南支0+123.456'）为浮点数"""
            if not text:
                return 0.0
            if '+' in text:
                parts = text.split('+')
                if len(parts) == 2:
                    km_digits = ''.join(c for c in parts[0] if c.isdigit())
                    try:
                        km = int(km_digits) if km_digits else 0
                        m = float(parts[1])
                        return km * 1000 + m
                    except (ValueError, TypeError):
                        pass
            try:
                return float(text)
            except (ValueError, TypeError):
                return 0.0

        def _parse_ip_number(text):
            """解析IP编号，支持复合格式如 'IP3 沪蓉倒进'（#11）"""
            if not text:
                return 0
            # 先尝试直接转int
            try:
                return int(text)
            except (ValueError, TypeError):
                pass
            # 从 "IPxx" 或 "IP xx ..." 中提取数字
            m = re.match(r'IP\s*(\d+)', text)
            if m:
                return int(m.group(1))
            return 0

        nodes = []
        _default_q = (self._parse_flow_values(self.design_flow_edit.text()) or [5.0])[0]

        for r in range(table.rowCount()):
            data = self._get_node_row_data(r)
            # data[0-7]: 流量段,建筑物名称,结构形式,进出口,IP,X,Y,转弯半径
            # data[8-19]: 几何结果列
            # data[20-26]: 底宽B,直径D,半径R,边坡m,糙率n,底坡1/i,流量Q
            # data[27-31]: 水力结果列
            # data[32-40]: 水头损失列
            # data[41-43]: 水位,渠底高程,渠顶高程
            node = ChannelNode()
            node.flow_section = str(data[0]).strip()
            node.name = str(data[1]).strip()

            # 结构形式 (col 2)
            struct_str = str(data[2]).strip()
            if struct_str:
                try:
                    node.structure_type = StructureType.from_string(struct_str)
                except ValueError:
                    pass
                if struct_str == "渐变段" or (node.structure_type and node.structure_type == StructureType.TRANSITION):
                    node.is_transition = True

            # 标记闸类结构（分水闸/分水口/节制闸/泄水闸等）
            if node.structure_type and StructureType.is_diversion_gate(node.structure_type):
                node.is_diversion_gate = True
            # 标记倒虹吸
            if node.structure_type and node.structure_type == StructureType.INVERTED_SIPHON:
                node.is_inverted_siphon = True
            # 标记有压管道
            if node.structure_type and node.structure_type == StructureType.PRESSURE_PIPE:
                node.is_pressure_pipe = True
            pipe_material = ""
            local_loss_ratio = None
            in_out_raw = ""
            # 恢复自动插入明渠段标记（通过UserRole存储）
            _first_item = table.item(r, 0)
            if _first_item:
                _ur = _first_item.data(Qt.UserRole)
                if isinstance(_ur, dict) and _ur.get('_auto_channel'):
                    node.is_auto_inserted_channel = True
                    node.x = float(_ur.get('_x', 0.0) or 0.0)
                    node.y = float(_ur.get('_y', 0.0) or 0.0)
                elif _ur == "auto_channel":  # 兼容旧格式
                    node.is_auto_inserted_channel = True
                # 恢复渐变段详细参数（#10）
                elif isinstance(_ur, dict) and _ur.get('_transition_data'):
                    td = _ur['_transition_data']
                    node.transition_type = td.get('transition_type', '')
                    node.transition_form = td.get('transition_form', '')
                    node.transition_zeta = td.get('transition_zeta', 0.0)
                    node.transition_theta = td.get('transition_theta', 0.0)
                if isinstance(_ur, dict):
                    _ext = _ur.get('_external_head_loss', None)
                    if _ext is not None and str(_ext).strip() != "":
                        try:
                            node.external_head_loss = float(_ext)
                        except (ValueError, TypeError):
                            node.external_head_loss = None
                    _pm = str(_ur.get('_pipe_material', '') or '').strip()
                    if _pm:
                        pipe_material = _pm
                    _ior = str(_ur.get('_in_out_raw', '') or '').strip()
                    if _ior:
                        in_out_raw = _ior
                    _llr = _ur.get('_local_loss_ratio', None)
                    if _llr is not None and str(_llr).strip() != "":
                        try:
                            local_loss_ratio = float(_llr)
                        except (ValueError, TypeError):
                            local_loss_ratio = None

            # 进出口 (col 3)
            _io_text = _read_text(r, 3)
            if _io_text:
                node.in_out = InOutType.from_string(_io_text)
            # IP编号 (col 4) — 支持复合格式 (#11)
            node.ip_number = _parse_ip_number(_read_text(r, 4))

            node.x = self._sf(data[5])
            node.y = self._sf(data[6])
            # 转弯半径 fallback 规则：
            #   首行/闸类/倒虹吸 → 0（不用默认全局半径，避免刷表时写入错误值）
            #   普通行           → 全局转弯半径（保持原有逻辑）
            _struct_for_r = str(data[2]).strip()
            _is_siphon_for_r = "倒虹吸" in _struct_for_r
            _is_gate_for_r = "闸" in _struct_for_r or "分水" in _struct_for_r
            if r == 0 or _is_siphon_for_r or _is_gate_for_r:
                node.turn_radius = self._sf(data[7], 0.0)
            else:
                node.turn_radius = self._sf(data[7], self._fval(self.turn_radius_edit, DEFAULT_TURN_RADIUS))

            # ===== 几何结果列 (8-19) =====
            # 转角 (col 8)
            _ta = _read_float(r, 8)
            if _ta > 0:
                node.turn_angle = _ta
            # 切线长 (col 9)
            _tl = _read_float(r, 9)
            if _tl > 0:
                node.tangent_length = _tl
            # 弧长 (col 10)
            _al = _read_float(r, 10)
            if _al > 0:
                node.arc_length = _al
            # 弯道长度 (col 11)
            _cl = _read_float(r, 11)
            if _cl > 0:
                node.curve_length = _cl
            # IP直线间距 (col 12)
            _sd = _read_float(r, 12)
            if _sd > 0:
                node.straight_distance = _sd
            # IP桩号 (col 13)
            node.station_ip = _parse_station(_read_text(r, 13))
            # 弯前BC (col 14)
            node.station_BC = _parse_station(_read_text(r, 14))
            # 里程MC (col 15)
            node.station_MC = _parse_station(_read_text(r, 15))
            # 弯末EC (col 16)
            node.station_EC = _parse_station(_read_text(r, 16))
            # 复核弯前 (col 17)
            _cpre = _read_float(r, 17)
            if _cpre != 0:
                node.check_pre_curve = _cpre
            # 复核弯后 (col 18)
            _cpost = _read_float(r, 18)
            if _cpost != 0:
                node.check_post_curve = _cpost
            # 复核总长 (col 19)
            _ctot = _read_float(r, 19)
            if _ctot != 0:
                node.check_total_length = _ctot

            # ===== 水力输入列 (20-26) =====
            B = self._sf(data[20])
            D = self._sf(data[21])
            R = self._sf(data[22])
            m_val = self._sf(data[23])
            # 糙率默认值：倒虹吸行用倒虹吸默认糙率常量，其他用渠道糙率输入框
            _default_n = (DEFAULT_SIPHON_ROUGHNESS
                          if struct_str and "倒虹吸" in struct_str
                          else self._fval(self.roughness_edit, DEFAULT_ROUGHNESS))
            n_val = self._sf(data[24], _default_n)
            slope_inv = self._sf(data[25])
            Q = self._sf(data[26], _default_q)

            # 与原版Tkinter get_nodes一致：始终写入B/D/R_circle/m（即使为0）
            # 原因：_estimate_transition_length中 section_params.get("D", 3.0)
            # 若D不在dict中会返回默认值3.0，导致隧洞渐变段长度偏大
            node.section_params['B'] = B
            node.section_params['D'] = D
            node.section_params['R_circle'] = R
            node.section_params['m'] = m_val
            if pipe_material:
                node.section_params['pipe_material'] = pipe_material
            if in_out_raw:
                node.section_params['in_out_raw'] = in_out_raw
            if local_loss_ratio is not None and local_loss_ratio > 0:
                node.section_params['local_loss_ratio'] = local_loss_ratio
            node.roughness = n_val
            if slope_inv > 0:
                node.slope_i = 1.0 / slope_inv
            node.flow = Q

            # 自动填充过闸水头损失
            if struct_str and ("闸" in struct_str or "分水" in struct_str):
                if not getattr(node, 'head_loss_gate', 0.0):
                    node.head_loss_gate = DEFAULT_GATE_HEAD_LOSS

            # ===== 水力结果列 (27-31) =====
            _h = _read_float(r, 27)
            if _h > 0:
                node.water_depth = _h
            # 断面面积A (col 28)
            _area = _read_float(r, 28)
            if _area > 0:
                node.section_params['A'] = _area
            # 湿周X (col 29)
            _wp = _read_float(r, 29)
            if _wp > 0:
                node.section_params['X'] = _wp
            # 水力半径R (col 30)
            _hr2 = _read_float(r, 30)
            if _hr2 > 0:
                node.section_params['R'] = _hr2
            # 流速v (col 31)
            _v = _read_float(r, 31)
            if _v > 0:
                node.velocity = _v

            # ===== 水头损失列 (32-40) =====
            # 渐变段长度 (col 32)
            _trl = _read_float(r, 32)
            if _trl > 0:
                node.transition_length = _trl
            # 渐变段损失 (col 33)
            _ht = _read_float(r, 33)
            if _ht != 0:
                node.head_loss_transition = _ht
            # 弯道损失 (col 34)
            _hb = _read_float(r, 34)
            if _hb != 0:
                node.head_loss_bend = _hb
            # 沿程损失 (col 35)
            _hf = _read_float(r, 35)
            if _hf != 0:
                node.head_loss_friction = _hf
            # 预留损失 (col 36)
            _hr = _read_float(r, 36)
            if _hr > 0:
                node.head_loss_reserve = _hr
            # 过闸损失 (col 37)
            _hg = _read_float(r, 37)
            if _hg > 0:
                node.head_loss_gate = _hg
            # 倒虹吸/有压管道损失 (col 38)
            _hs = _read_float(r, 38)
            if _hs > 0:
                node.head_loss_siphon = _hs
            elif (
                node.is_pressure_pipe
                and getattr(node, 'in_out', None) is not None
                and node.in_out.value == "出"
                and getattr(node, 'external_head_loss', None) is not None
            ):
                # 兼容旧数据：历史版本将有压管道损失只存于 external_head_loss
                try:
                    _ext_loss = float(node.external_head_loss)
                except (TypeError, ValueError):
                    _ext_loss = 0.0
                if _ext_loss > 0:
                    node.head_loss_siphon = _ext_loss
                # 迁移后清空旧字段，避免后续计算重复叠加
                node.external_head_loss = None
            # 总损失 (col 39)
            _htotal = _read_float(r, 39)
            if _htotal != 0:
                node.head_loss_total = _htotal
            # 累计损失 (col 40)
            _hcum = _read_float(r, 40)
            if _hcum != 0:
                node.head_loss_cumulative = _hcum

            # ===== 高程列 (41-43) =====
            _wl = _read_float(r, 41)
            if _wl != 0:
                node.water_level = _wl
            _be = _read_float(r, 42)
            if _be != 0:
                node.bottom_elevation = _be
            _te = _read_float(r, 43)
            if _te != 0:
                node.top_elevation = _te

            # 恢复结构高度（用于计算渠顶高程，与Tkinter版 data_table.py 对齐）
            if r in self._node_structure_heights:
                node.structure_height = self._node_structure_heights[r]

            # 恢复倒角参数（渡槽-矩形精确水力计算用）
            if r in self._node_chamfer_params:
                cp = self._node_chamfer_params[r]
                node.section_params['chamfer_angle'] = cp.get('chamfer_angle', 0)
                node.section_params['chamfer_length'] = cp.get('chamfer_length', 0)

            # 恢复明渠-U形圆心角
            if r in self._node_u_params:
                node.section_params['theta_deg'] = self._node_u_params[r].get('theta_deg', 0)

            # 恢复加大流速（从批量计算导入的加大流量工况流速）
            if r in self._node_velocity_increased:
                node.velocity_increased = self._node_velocity_increased[r]

            nodes.append(node)
        return nodes

    def _calculate(self):
        if not CALCULATOR_AVAILABLE:
            InfoBar.error("不可用", "核心计算引擎未加载，无法计算",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        if self.node_table.rowCount() < 2:
            InfoBar.warning("节点不足", "至少需要2个节点才能计算水面线",
                           parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        try:
            settings = self._build_settings()
            nodes = self._build_nodes_from_table()
            if not nodes or len(nodes) < 2:
                InfoBar.warning("数据不足", "有效节点不足", parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 前置检查：渐变段必须已插入
            has_transitions = any(getattr(n, 'is_transition', False) for n in nodes)
            if not has_transitions:
                InfoBar.warning("提示",
                               "请先点击工具栏的【插入渐变段】按钮，完成渐变段插入后再执行计算。",
                               parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
                return

            # 前置检查：倒虹吸水力计算
            has_siphon = any(
                n.structure_type and "倒虹吸" in n.structure_type.value
                for n in nodes if n.structure_type
            )
            has_siphon_loss = any(
                n.structure_type
                and "倒虹吸" in n.structure_type.value
                and getattr(n, 'in_out', None) is not None
                and n.in_out.value == "出"
                and (getattr(n, 'head_loss_siphon', 0.0) or 0.0) > 0
                for n in nodes if n.structure_type
            )
            if has_siphon and not has_siphon_loss:
                InfoBar.warning("提示",
                               "检测到表格中包含倒虹吸，但尚未执行水力计算。"
                               "请先点击【倒虹吸水力计算】按钮完成计算后，再点击【执行计算】。",
                               parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
                return

            # 前置检查：有压管道水力计算（阻断式）
            has_pressure_pipe = any(
                n.structure_type and "有压管道" in n.structure_type.value
                for n in nodes if n.structure_type
            )
            has_pressure_pipe_loss = any(
                n.structure_type
                and "有压管道" in n.structure_type.value
                and getattr(n, 'in_out', None) is not None
                and n.in_out.value == "出"
                and (
                    (getattr(n, 'head_loss_siphon', 0.0) or 0.0) > 0
                    or getattr(n, 'external_head_loss', None) is not None
                )
                for n in nodes if n.structure_type
            )
            if has_pressure_pipe and not has_pressure_pipe_loss:
                InfoBar.warning(
                    "提示",
                    "检测到表格中包含有压管道，但尚未执行水力计算。"
                    "请先点击【有压管道水力计算】按钮完成计算后，再点击【执行计算】。",
                    parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP
                )
                return

            calculator = WaterProfileCalculator(settings)

            # 验证输入
            is_valid, errors = calculator.validate_input(nodes)
            if not is_valid:
                InfoBar.error("输入错误", "\n".join(errors),
                             parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
                return

            calculated = calculator.calculate_all(nodes)
            self.calculated_nodes = calculated
            self._settings = settings

            self._display_results(calculated, settings)
            self._generate_detail_report(calculated, settings, calculator)

            # 计算摘要
            summary = calculator.get_calculation_summary(calculated)
            total_len = summary.get('总长度', 0.0)
            wl_drop = summary.get('水位落差', 0.0)

            # 更新建筑物长度统计缓存
            self._last_building_lengths = calculator.calculate_building_lengths(calculated)
            self._last_channel_total_length = total_len
            self._last_type_summary = calculator.calculate_comprehensive_type_summary(calculated)

            # 更新持久摘要面板
            self._update_summary_panel(calculated, total_len, wl_drop, summary)

            # 检查缺少结构高度（渠顶高程无法计算）的节点
            # 倒虹吸、渐变段、闸类/分水口不需要结构总高，跳过不提示
            missing_height_names = []
            for nd in calculated:
                if getattr(nd, 'is_transition', False):
                    continue
                _st = nd.structure_type.value if nd.structure_type else ""
                if "倒虹吸" in _st:
                    continue
                if "闸" in _st or "分水" in _st:
                    continue
                if nd.bottom_elevation and nd.bottom_elevation != 0 and (not nd.top_elevation or nd.top_elevation == 0):
                    missing_height_names.append(nd.name or "未命名")

            msg = f"共{len(calculated)}个节点，总长{total_len:.1f}m，水位落差{wl_drop:.3f}m"
            if missing_height_names:
                msg += f"\n⚠ 以下节点缺少结构总高，渠顶高程未计算: {', '.join(missing_height_names)}"
                msg += "\n请通过【从批量计算导入】获取正确的结构总高。"
                InfoBar.warning("计算完成（部分渠顶高程缺失）", msg,
                               parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
            else:
                InfoBar.success("计算完成", msg,
                               parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

        except Exception as e:
            InfoBar.error("计算错误", f"计算过程出错: {str(e)}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            import traceback
            self.detail_text.setPlainText(f"计算错误:\n{traceback.format_exc()}")
            self.result_notebook.setCurrentIndex(1)

    # ================================================================
    # 结果显示
    # ================================================================
    def _display_results(self, nodes, settings):
        """将计算结果填充到统一node_table的结果列（第13列起）"""
        prefix = settings.get_station_prefix() if settings else ""

        # 先用计算后的节点重建整个表格（输入+结果一体）
        self._update_table_from_nodes_full(nodes, prefix)

        auto_resize_table(self.node_table)

    def _update_table_from_nodes_full(self, nodes, prefix=""):
        """用计算后的完整节点数据重建统一表格（输入列+结果列）"""
        self._updating_cells = True
        try:
            self._update_table_from_nodes_full_impl(nodes, prefix)
        finally:
            self._updating_cells = False
        self._update_pressure_pipe_roughness_overview(
            self._collect_pressure_pipe_roughness_pairs_from_nodes(nodes)
        )
        self._refresh_pressure_pipe_controls()

    def _update_table_from_nodes_full_impl(self, nodes, prefix=""):
        # 更新结构高度缓存（计算完成后可能已重新计算）
        self._node_structure_heights.clear()
        for i, node in enumerate(nodes):
            if getattr(node, 'structure_height', 0) and node.structure_height > 0:
                self._node_structure_heights[i] = node.structure_height
        # 重建倒角参数缓存（节点计算往返后从 section_params 中恢复）
        self._node_chamfer_params.clear()
        for i, node in enumerate(nodes):
            sp = getattr(node, 'section_params', {}) or {}
            _ca = sp.get('chamfer_angle', 0) or 0
            _cl = sp.get('chamfer_length', 0) or 0
            if _ca > 0 and _cl > 0:
                self._node_chamfer_params[i] = {'chamfer_angle': float(_ca), 'chamfer_length': float(_cl)}
        # 重建明渠-U形圆心角缓存
        self._node_u_params.clear()
        for i, node in enumerate(nodes):
            sp = getattr(node, 'section_params', {}) or {}
            _th = sp.get('theta_deg', 0) or 0
            if _th > 0 and node.structure_type and 'U形' in node.structure_type.value and '明渠' in node.structure_type.value:
                self._node_u_params[i] = {'theta_deg': float(_th)}
        # 重建加大流速缓存（节点计算往返后从 velocity_increased 中恢复）
        self._node_velocity_increased.clear()
        for i, node in enumerate(nodes):
            _vi = getattr(node, 'velocity_increased', 0.0)
            if _vi and _vi > 0:
                self._node_velocity_increased[i] = float(_vi)
        self.node_table.setRowCount(0)
        for node in nodes:
            r = self.node_table.rowCount()
            self.node_table.insertRow(r)

            _is_trans = getattr(node, 'is_transition', False)
            _is_auto_ch = getattr(node, 'is_auto_inserted_channel', False)

            # 构建完整46列数据，按列索引直接赋值
            vals = [""] * len(NODE_ALL_HEADERS)

            # 基础输入列 (0-7)
            vals[0] = node.flow_section
            vals[1] = node.name
            _st_str = node.get_structure_type_str()
            vals[2] = f"{_st_str}(连接段)" if _is_auto_ch else _st_str
            if not _is_trans:
                vals[3] = node.get_in_out_str()
                vals[4] = "" if _is_auto_ch else node.get_ip_str()
                vals[5] = f"{node.x:.6f}" if (node.x and not _is_auto_ch) else ""
                vals[6] = f"{node.y:.6f}" if (node.y and not _is_auto_ch) else ""
                # 转弯半径 col 7：首行始终为空（起始节点无弯道意义）；
                # 闸类/分水口/倒虹吸 fallback=0（改动C），空 col7→0→显示"" 自然处理，
                # 不强制清空，允许用户手动填入的值保留显示
                vals[7] = "" if r == 0 else (f"{node.turn_radius:.1f}" if node.turn_radius else "")

            if not _is_trans:
                # 几何结果列 (8-19) — 无条件格式化，0值也显示（与Tkinter一致）
                _fmt_s = lambda s: ProjectSettings.format_station(s, prefix) if s is not None else "-"
                vals[8] = "" if _is_auto_ch else f"{node.turn_angle:.4f}"
                vals[9] = "" if _is_auto_ch else f"{node.tangent_length:.6f}"
                vals[10] = "" if _is_auto_ch else f"{node.arc_length:.6f}"
                vals[11] = "" if _is_auto_ch else f"{node.curve_length:.6f}"
                vals[12] = "" if _is_auto_ch else f"{node.straight_distance:.6f}"
                vals[13] = "" if _is_auto_ch else _fmt_s(node.station_ip)
                vals[14] = "" if _is_auto_ch else _fmt_s(node.station_BC)
                vals[15] = "" if _is_auto_ch else _fmt_s(getattr(node, 'station_MC', None))
                vals[16] = "" if _is_auto_ch else _fmt_s(node.station_EC)
                _skip_check = _is_auto_ch or getattr(node, 'is_inverted_siphon', False)
                vals[17] = "" if _skip_check else f"{getattr(node, 'check_pre_curve', 0):.3f}"
                vals[18] = "" if _skip_check else f"{getattr(node, 'check_post_curve', 0):.3f}"
                vals[19] = "" if _skip_check else f"{getattr(node, 'check_total_length', 0):.3f}"

                # 水力输入列 (20-26)
                _B = node.section_params.get('B', 0)
                _D = node.section_params.get('D', 0)
                _Rc = node.section_params.get('R_circle', 0)
                _m = node.section_params.get('m', 0)
                vals[20] = f"{_B:.3f}" if _B else ""
                vals[21] = f"{_D:.3f}" if _D else ""
                vals[22] = f"{_Rc:.3f}" if _Rc else ""
                vals[23] = f"{_m:.2f}" if _m else ""
                vals[24] = f"{node.roughness:.4f}" if node.roughness else ""
                vals[25] = f"{1.0/node.slope_i:.0f}" if node.slope_i and node.slope_i > 0 else ""
                vals[26] = f"{node.flow:.3f}" if node.flow else ""

                # 水力结果列 (27-31)
                _area = node.section_params.get('A', 0) if node.section_params else 0
                _peri = node.section_params.get('X', 0) if node.section_params else 0
                _hydr = node.section_params.get('R', 0) if node.section_params else 0
                vals[27] = f"{node.water_depth:.3f}" if node.water_depth else "-"
                vals[28] = f"{_area:.3f}" if _area else "-"
                vals[29] = f"{_peri:.3f}" if _peri else "-"
                vals[30] = f"{_hydr:.3f}" if _hydr else "-"
                vals[31] = f"{node.velocity:.3f}" if node.velocity else "-"

                # 水头损失列 (33-40) — 非渐变段行
                vals[33] = f"{node.head_loss_transition:.4f}" if node.head_loss_transition else "-"
                vals[34] = f"{node.head_loss_bend:.4f}" if node.head_loss_bend else "-"
                vals[35] = f"{node.head_loss_friction:.4f}" if node.head_loss_friction else "-"
                vals[36] = f"{getattr(node, 'head_loss_reserve', 0):.4f}" if getattr(node, 'head_loss_reserve', None) else "-"
                vals[37] = f"{node.head_loss_gate:.4f}" if node.head_loss_gate else "-"
                _h_sp = getattr(node, 'head_loss_siphon', 0.0) or 0.0
                _ext = getattr(node, 'external_head_loss', None)
                _is_pp_outlet = (
                    getattr(node, 'is_pressure_pipe', False)
                    and getattr(node, 'in_out', None) is not None
                    and getattr(node.in_out, 'value', '') == "出"
                )
                if _h_sp <= 0 and _is_pp_outlet and _ext is not None:
                    # 兼容旧项目：将 external_head_loss 迁移到统一展示列（col 38）
                    try:
                        _ext_f = float(_ext)
                    except (TypeError, ValueError):
                        _ext_f = 0.0
                    if _ext_f > 0:
                        _h_sp = _ext_f
                        node.head_loss_siphon = _ext_f
                    # 迁移后清空旧字段，避免下次构建节点时重复计损
                    node.external_head_loss = None
                vals[38] = f"{_h_sp:.4f}" if _h_sp else "-"
                vals[39] = f"{node.head_loss_total:.4f}" if node.head_loss_total else "-"
                vals[40] = f"{node.head_loss_cumulative:.4f}" if node.head_loss_cumulative else "-"

                # 高程列 (41-43)
                vals[41] = f"{node.water_level:.3f}" if node.water_level else "-"
                vals[42] = f"{node.bottom_elevation:.3f}" if node.bottom_elevation else "-"
                vals[43] = f"{node.top_elevation:.3f}" if node.top_elevation else "-"


            # 渐变段行特有数据
            if _is_trans:
                # 写入糙率/底坡/流量，确保通过表格读写循环不丢失
                vals[24] = f"{node.roughness:.4f}" if node.roughness else ""
                vals[25] = f"{1.0/node.slope_i:.0f}" if node.slope_i and node.slope_i > 0 else ""
                vals[26] = f"{node.flow:.3f}" if node.flow else ""
                vals[32] = f"{getattr(node, 'transition_length', 0):.3f}" if getattr(node, 'transition_length', None) else "-"
                vals[33] = f"{node.head_loss_transition:.4f}" if node.head_loss_transition else "-"
                vals[40] = f"{node.head_loss_cumulative:.4f}" if node.head_loss_cumulative else "-"

            # 渐变段长度（所有行通用）
            if not _is_trans:
                vals[32] = f"{getattr(node, 'transition_length', 0):.3f}" if getattr(node, 'transition_length', None) else "-"

            for c, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                item.setTextAlignment(Qt.AlignCenter)
                # 非可编辑列设为只读
                if c not in EDITABLE_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                # 第一行（水位起点）锁定水头损失列
                if r == 0 and c in FIRST_ROW_LOCKED_LOSS_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                # 渐变段行灰色
                if _is_trans:
                    item.setForeground(QColor("#9E9E9E"))
                # 自动插入明渠段绿色
                elif _is_auto_ch:
                    item.setForeground(QColor("#2E7D32"))
                    item.setToolTip("自动插入的明渠连接段，用于计算两个建筑物之间的沿程及弯道水头损失。\n几何列留空因为该行不是真实IP转折点。")
                # 倒虹吸蓝色
                elif getattr(node, 'is_inverted_siphon', False):
                    item.setForeground(QColor("#1565C0"))
                # 分水闸橙色
                elif getattr(node, 'is_diversion_gate', False):
                    item.setForeground(QColor("#E65100"))
                self.node_table.setItem(r, c, item)

            # 在行首单元格中存储标记（UserRole），供_build_nodes_from_table恢复
            first_item = self.node_table.item(r, 0)
            if first_item:
                payload = first_item.data(Qt.UserRole)
                if not isinstance(payload, dict):
                    payload = {}
                if _is_auto_ch:
                    payload.update({"_auto_channel": True, "_x": node.x, "_y": node.y})
                elif _is_trans and (node.transition_type or node.transition_form):
                    # 渐变段详细参数保存到UserRole（#10）
                    payload.update({
                        '_transition_data': {
                            'transition_type': getattr(node, 'transition_type', ''),
                            'transition_form': getattr(node, 'transition_form', ''),
                            'transition_zeta': getattr(node, 'transition_zeta', 0.0),
                            'transition_theta': getattr(node, 'transition_theta', 0.0),
                        }
                    })
                if getattr(node, 'external_head_loss', None) is not None:
                    payload['_external_head_loss'] = getattr(node, 'external_head_loss')
                # 持久化有压管道专用参数（表格无专门列，放在UserRole）
                _sp = getattr(node, 'section_params', {}) or {}
                _pm = str(_sp.get('pipe_material', '') or '').strip()
                _ior = str(_sp.get('in_out_raw', '') or '').strip()
                _llr = _sp.get('local_loss_ratio', None)
                if _pm:
                    payload['_pipe_material'] = _pm
                if _ior:
                    payload['_in_out_raw'] = _ior
                if _llr is not None and str(_llr).strip() != "":
                    try:
                        _llr_f = float(_llr)
                        if _llr_f > 0:
                            payload['_local_loss_ratio'] = _llr_f
                    except (ValueError, TypeError):
                        pass
                if payload:
                    first_item.setData(Qt.UserRole, payload)

    def _generate_detail_report(self, nodes, settings, calculator=None):
        """生成详细计算过程文本"""
        prefix = settings.get_station_prefix() if settings else ""
        lines = []
        lines.append("=" * 80)
        lines.append(f"  {settings.channel_name if settings else ''}推求水面线 — 详细计算结果")
        lines.append("=" * 80)
        lines.append(f"  计算时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"  渠道名称: {settings.channel_name if settings else '-'}")
        lines.append(f"  渠道级别: {settings.channel_level if settings else '-'}")
        lines.append(f"  起始水位: {settings.start_water_level if settings else '-'} m")
        lines.append(f"  起始桩号: {ProjectSettings.format_station(settings.start_station, prefix) if settings else '-'}")
        # 多流量段显示
        if settings and getattr(settings, 'design_flows', None):
            flows_str = ", ".join(f"{q:.3f}" for q in settings.design_flows)
            lines.append(f"  设计流量: {flows_str} m³/s")
        else:
            lines.append(f"  设计流量: {settings.design_flow if settings else '-'} m³/s")
        if settings and getattr(settings, 'max_flows', None):
            flows_str = ", ".join(f"{q:.3f}" for q in settings.max_flows)
            lines.append(f"  加大流量: {flows_str} m³/s")
        else:
            lines.append(f"  加大流量: {settings.max_flow if settings else '-'} m³/s")
        lines.append(f"  糙率: {settings.roughness if settings else '-'}")
        if settings and getattr(settings, 'siphon_roughness', None) is not None:
            lines.append(f"  倒虹吸糙率: {settings.siphon_roughness}")
        lines.append(f"  转弯半径: {settings.turn_radius if settings else '-'} m")
        lines.append(f"  总节点数: {len(nodes)}")
        # 渐变段设置
        if settings:
            lines.append(f"  渡槽/隧洞渐变段: 进口{settings.transition_inlet_form}(ζ={settings.transition_inlet_zeta:.2f}), "
                        f"出口{settings.transition_outlet_form}(ζ={settings.transition_outlet_zeta:.2f})")
            if getattr(settings, 'open_channel_transition_form', None):
                lines.append(f"  明渠渐变段: {settings.open_channel_transition_form}(ζ={settings.open_channel_transition_zeta:.2f})")
            lines.append(f"  倒虹吸渐变段: 进口{settings.siphon_transition_inlet_form}(ζ={settings.siphon_transition_inlet_zeta:.2f}), "
                        f"出口{settings.siphon_transition_outlet_form}(ζ={settings.siphon_transition_outlet_zeta:.2f})")
        lines.append("-" * 80)
        lines.append("")

        for i, node in enumerate(nodes):
            # 节点标题
            tag = ""
            if getattr(node, 'is_transition', False):
                tag = " [渐变段]"
            elif getattr(node, 'is_auto_inserted_channel', False):
                tag = " [自动插入明渠段]"
            elif getattr(node, 'is_inverted_siphon', False):
                tag = " [倒虹吸]"
            elif getattr(node, 'is_diversion_gate', False):
                tag = f" [{node.get_structure_type_str()}]"
            lines.append(f"--- 节点 {i+1}: {node.name} ({node.get_structure_type_str()}){tag} ---")
            lines.append(f"  IP编号: {node.get_ip_str()}")
            lines.append(f"  进出口: {node.get_in_out_str()}")
            lines.append(f"  流量段: {node.flow_section}")
            lines.append(f"  坐标: ({node.x:.3f}, {node.y:.3f})")
            if node.station_ip:
                lines.append(f"  桩号: {ProjectSettings.format_station(node.station_ip, prefix)}")
            if node.azimuth:
                lines.append(f"  方位角: {node.azimuth:.6f}°")
            if node.turn_angle:
                lines.append(f"  转角: {node.turn_angle:.6f}°")
            lines.append(f"  流量 Q = {node.flow:.3f} m³/s")
            lines.append(f"  糙率 n = {node.roughness}")
            if node.slope_i:
                lines.append(f"  底坡 i = {node.slope_i:.6f}")
            # 断面参数
            sp = node.section_params
            if sp:
                parts = []
                if sp.get('B'): parts.append(f"B={sp['B']:.3f}")
                if sp.get('D'): parts.append(f"D={sp['D']:.3f}")
                if sp.get('R_circle'): parts.append(f"R={sp['R_circle']:.3f}")
                if sp.get('m'): parts.append(f"m={sp['m']:.2f}")
                if parts:
                    lines.append(f"  断面参数: {', '.join(parts)}")
            # 水力结果
            if node.water_depth:
                lines.append(f"  水深 h = {node.water_depth:.3f} m")
            if node.velocity:
                lines.append(f"  流速 v = {node.velocity:.3f} m/s")
            if node.water_level:
                lines.append(f"  水位 Z = {node.water_level:.3f} m")
            if node.bottom_elevation:
                lines.append(f"  渠底高程 = {node.bottom_elevation:.3f} m")
            if node.top_elevation:
                lines.append(f"  渠顶高程 = {node.top_elevation:.3f} m")
            if node.structure_height:
                lines.append(f"  结构高度 = {node.structure_height:.3f} m")
            # 水头损失
            loss_parts = []
            if node.head_loss_friction:
                loss_parts.append(f"沿程={node.head_loss_friction:.4f}")
            if node.head_loss_bend:
                loss_parts.append(f"弯道={node.head_loss_bend:.4f}")
            if node.head_loss_transition:
                loss_parts.append(f"渐变段={node.head_loss_transition:.4f}")
            if node.head_loss_gate:
                loss_parts.append(f"过闸={node.head_loss_gate:.4f}")
            if node.head_loss_siphon:
                loss_parts.append(f"倒虹吸/有压管道={node.head_loss_siphon:.4f}")
            if node.head_loss_reserve:
                loss_parts.append(f"预留={node.head_loss_reserve:.4f}")
            if loss_parts:
                lines.append(f"  水头损失: {', '.join(loss_parts)}")
            if node.head_loss_total:
                lines.append(f"  总损失 = {node.head_loss_total:.4f} m")
            if node.head_loss_cumulative:
                lines.append(f"  累计损失 = {node.head_loss_cumulative:.4f} m")
            # 渐变段详情
            if getattr(node, 'is_transition', False) and node.transition_length:
                lines.append(f"  渐变段类型: {node.transition_type}")
                lines.append(f"  渐变段形式: {node.transition_form}")
                lines.append(f"  渐变段长度 L = {node.transition_length:.3f} m")
                if node.transition_zeta:
                    lines.append(f"  局部损失系数 ζ = {node.transition_zeta:.3f}")
            lines.append("")

        # 计算摘要
        if calculator:
            summary = calculator.get_calculation_summary(nodes)
            if summary:
                lines.append("=" * 80)
                lines.append("  计算摘要")
                lines.append("-" * 80)
                lines.append(f"  节点数量: {summary.get('节点数量', '-')}")
                s_start = summary.get('起点桩号', 0.0)
                s_end = summary.get('终点桩号', 0.0)
                lines.append(f"  起点桩号: {ProjectSettings.format_station(s_start, prefix)}")
                lines.append(f"  终点桩号: {ProjectSettings.format_station(s_end, prefix)}")
                lines.append(f"  总长度: {summary.get('总长度', 0.0):.3f} m")
                wl_s = summary.get('起点水位', 0.0)
                wl_e = summary.get('终点水位', 0.0)
                if wl_s and wl_e:
                    lines.append(f"  起点水位: {wl_s:.3f} m")
                    lines.append(f"  终点水位: {wl_e:.3f} m")
                    lines.append(f"  水位落差: {summary.get('水位落差', 0.0):.3f} m")
                lines.append("")

            # 建筑物长度汇总
            try:
                building_lengths = calculator.calculate_building_lengths(nodes)
                if building_lengths:
                    lines.append("=" * 80)
                    lines.append("  建筑物长度汇总")
                    lines.append("-" * 80)
                    lines.append(f"  {'序号':<4}  {'名称':<16}  {'结构形式':<12}  {'长度(m)':<10}  {'起始桩号':<16}  {'终止桩号':<16}")
                    lines.append("  " + "-" * 76)
                    for i, bl in enumerate(building_lengths, 1):
                        name = bl.get('name', '-')
                        stype = bl.get('structure_type', '-')
                        length = bl.get('length', 0.0)
                        s_s = bl.get('start_station', 0.0)
                        s_e = bl.get('end_station', 0.0)
                        lines.append(
                            f"  {i:<4}  {name:<16}  {stype:<12}  {length:<10.3f}  "
                            f"{ProjectSettings.format_station(s_s, prefix):<16}  "
                            f"{ProjectSettings.format_station(s_e, prefix):<16}"
                        )
                    total_length = sum(bl.get('length', 0.0) for bl in building_lengths)
                    lines.append("  " + "-" * 76)
                    lines.append(f"  {'合计':<22}  {'':<12}  {total_length:<10.3f}")
                    lines.append("")
            except Exception:
                pass

        lines.append("=" * 80)
        lines.append("  计算完毕")
        lines.append("=" * 80)
        self.detail_text.setPlainText("\n".join(lines))

    # ================================================================
    # 辅助
    # ================================================================
    def eventFilter(self, obj, event):
        """事件过滤器：起始桩号焦点 + 表头悬浮公式提示"""
        from PySide6.QtCore import QEvent, QPoint
        if obj is self.start_station_edit:
            if event.type() == QEvent.FocusIn:
                current = self.start_station_edit.text()
                value = parse_station_input(current)
                self.start_station_edit.setText(str(value))
        elif hasattr(self, '_node_header') and obj is self._node_header.viewport():
            header = self._node_header
            if event.type() == QEvent.Type.MouseMove:
                try:
                    pos = event.position().toPoint()
                except AttributeError:
                    pos = event.pos()
                logical_idx = header.logicalIndexAt(pos)
                if logical_idx >= 0 and logical_idx in self._formula_columns:
                    col_name = NODE_ALL_HEADERS[logical_idx]
                    vp_x = header.sectionViewportPosition(logical_idx)
                    sec_w = header.sectionSize(logical_idx)
                    gp = header.mapToGlobal(QPoint(vp_x + sec_w // 2, header.height()))
                    self._formula_tooltip.show_for_column(col_name, gp)
                else:
                    self._formula_tooltip.schedule_hide()
            elif event.type() == QEvent.Type.Leave:
                self._formula_tooltip.schedule_hide()
        return super().eventFilter(obj, event)

    def _format_start_station(self):
        """编辑完成后格式化起始桩号显示"""
        current = self.start_station_edit.text().strip()
        value = parse_station_input(current)
        formatted = format_station_display(value)
        self.start_station_edit.setText(formatted)

    def _update_siphon_roughness_overview(self, pairs):
        """更新倒虹吸糙率芯片展示。
        pairs: [(名称, 糙率), ...] 列表
        """
        if not pairs:
            self.siphon_roughness_chips.clear()
            return
        # 去重：同名倒虹吸只取第一个（批量计算中同一倒虹吸可能有多行）
        seen = {}
        for name, n_val in pairs:
            if name not in seen:
                seen[name] = n_val
        self.siphon_roughness_chips.set_siphon_data(list(seen.items()))

    def _update_pressure_pipe_roughness_overview(self, pairs):
        """更新有压管道参数芯片展示。pairs: [(名称, 管材名称), ...]"""
        if not hasattr(self, 'pressure_pipe_roughness_chips'):
            return
        if not pairs:
            self.pressure_pipe_roughness_chips.clear()
            return
        seen = {}
        default_idx = 1
        for name, material in pairs:
            display_name = str(name).strip() if name else ""
            if not display_name:
                display_name = f"有压管道{default_idx}"
                default_idx += 1
            if display_name not in seen:
                params = SiphonRoughnessChipContainer.PIPE_MATERIAL_PARAMS.get(material, {})
                if params:
                    param_str = f"{material} | f={params['f']}, m={params['m']}, b={params['b']}"
                else:
                    param_str = material if material else "未指定管材"
                seen[display_name] = param_str
        self.pressure_pipe_roughness_chips.set_pairs(list(seen.items()))

    def _collect_pressure_pipe_roughness_pairs_from_nodes(self, nodes):
        """从节点列表提取有压管道参数展示对。"""
        pairs = []
        if not nodes:
            return pairs
        default_idx = 1
        for node in nodes:
            st = node.structure_type.value if node.structure_type else ""
            if "有压管道" not in st:
                continue
            name = (node.name or "").strip()
            if not name:
                name = f"有压管道{default_idx}"
                default_idx += 1
            material = node.section_params.get('pipe_material', '') if hasattr(node, 'section_params') else ''
            pairs.append((name, material))
        return pairs

    def _refresh_pressure_pipe_controls(self):
        """刷新有压管道按钮提示状态（按钮始终可点击）。"""
        btn = getattr(self, "btn_pressure_pipe_calc", None)
        if btn is None:
            return
        table = getattr(self, "node_table", None)
        if table is None:
            btn.setEnabled(True)
            btn.setToolTip("执行有压管道水力计算并回写到\"倒虹吸/有压管道水头损失\"列")
            return
        has_ppipe = False
        has_transition = False
        for r in range(table.rowCount()):
            st_item = table.item(r, 2)
            st = st_item.text().strip() if st_item else ""
            if "有压管道" in st:
                has_ppipe = True
            if "渐变段" in st:
                has_transition = True
            if has_ppipe and has_transition:
                break
        # 交互优化：按钮始终保持可点击，具体前置校验在 _open_pressure_pipe_calculator 中处理并提示。
        btn.setEnabled(True)
        if not has_ppipe:
            self.btn_pressure_pipe_calc.setToolTip("尚未检测到有压管道节点。可先导入数据，点击按钮可查看前置提示")
        elif not has_transition:
            self.btn_pressure_pipe_calc.setToolTip("已检测到有压管道。请先插入渐变段后再执行有压管道水力计算")
        else:
            self.btn_pressure_pipe_calc.setToolTip("执行有压管道水力计算并回写到\"倒虹吸/有压管道水头损失\"列")

    def _update_pressure_pipe_last_result_button(self):
        """刷新有压管道计算相关控件状态（计算完成后调用）。"""
        self._refresh_pressure_pipe_controls()

    def _choose_roughness_value(self, values, label):
        """当同类建筑物糙率不一致时，弹窗让用户选择。
        values: 糙率值列表（已收集的同类建筑物糙率）
        label: 显示标签（如"渠道糙率"或"倒虹吸糙率"）
        返回选中的糙率值，若列表为空返回None。
        """
        if not values:
            return None
        from collections import Counter
        counter = Counter(values)
        unique_vals = sorted(counter.keys())
        # 所有值相同，直接返回
        if len(unique_vals) == 1:
            return unique_vals[0]
        # 构建选项列表
        options = []
        for v in unique_vals:
            cnt = counter[v]
            options.append(f"{v}    （出现 {cnt} 次）")
        from PySide6.QtWidgets import QInputDialog
        chosen, ok = QInputDialog.getItem(
            self, f"选择{label}",
            f"批量计算中不同建筑物的{label}值不一致，请选择一个作为全局{label}：",
            options, 0, False)
        if ok and chosen:
            # 从选项文本中提取数值
            val_str = chosen.split("（")[0].strip()
            try:
                return float(val_str)
            except ValueError:
                return unique_vals[0]
        # 用户取消，取众数
        return counter.most_common(1)[0][0]

    def _fval(self, edit, default=0.0):
        t = edit.text().strip()
        if not t: return default
        try: return float(t)
        except ValueError: return default

    def _sf(self, val, default=0.0):
        if not val: return default
        s = str(val).strip()
        if not s: return default
        try: return float(s)
        except ValueError: return default

    def _info_parent(self):
        w = self.window()
        return w if w else self

    def _calculate_recommended_turn_radius(self, nodes):
        """根据规范计算推荐的转弯半径（取大值原则）
        - 隧洞：弯曲半径≥洞径(或洞宽)×5
        - 明渠：弯曲半径≥水面宽度×5
        - 渡槽：弯道半径≥连接明渠渠底宽度×5
        """
        import math as _math
        max_r = 0.0
        for node in nodes:
            if not node.structure_type:
                continue
            sv = node.structure_type.value
            if "倒虹吸" in sv:
                continue
            b = node.section_params.get("B", 0)
            rc = node.section_params.get("R_circle", 0)
            m_s = node.section_params.get("m", 0)
            wd = node.water_depth or 0
            min_r = 0.0
            if "隧洞" in sv:
                if rc > 0:
                    min_r = rc * 2 * 5
                elif b > 0:
                    min_r = b * 5
            elif "明渠" in sv or sv == "矩形":
                if "U形" in sv and rc > 0:
                    theta_u_r = node.section_params.get('theta_deg', 0) or 0
                    if theta_u_r > 0 and wd > 0:
                        import math as _math_u
                        h0_u_r = rc * (1.0 - _math_u.cos(_math_u.radians(theta_u_r / 2.0)))
                        if wd <= h0_u_r:
                            wsw_u = 2 * _math_u.sqrt(max(0.0, rc ** 2 - (rc - wd) ** 2))
                        else:
                            b_arc_u_r = 2 * rc * _math_u.sin(_math_u.radians(theta_u_r / 2.0))
                            wsw_u = b_arc_u_r + 2 * m_s * (wd - h0_u_r)
                        min_r = wsw_u * 5
                    else:
                        min_r = rc * 2 * 5
                elif b > 0:
                    wsw = b + 2 * m_s * wd if m_s > 0 and wd > 0 else b
                    min_r = wsw * 5
            elif "渡槽" in sv:
                if b > 0:
                    min_r = b * 5
            elif "暗涵" in sv:
                if b > 0:
                    min_r = b * 5
            if min_r > max_r:
                max_r = min_r
        if max_r <= 0:
            max_r = DEFAULT_TURN_RADIUS
        return _math.ceil(max_r)

    def _on_trans_inlet_form_changed(self, form):
        """渡槽/隧洞进口渐变段形式变化时自动更新ζ系数"""
        zeta_table = TRANSITION_ZETA_COEFFICIENTS.get("进口", {})
        if form in zeta_table:
            self.trans_inlet_zeta.setText(f"{zeta_table[form]:.2f}")
        elif form == "直线形扭曲面":
            self.trans_inlet_zeta.setText("0.05")

    def _on_trans_outlet_form_changed(self, form):
        """渡槽/隧洞出口渐变段形式变化时自动更新ζ系数"""
        zeta_table = TRANSITION_ZETA_COEFFICIENTS.get("出口", {})
        if form in zeta_table:
            self.trans_outlet_zeta.setText(f"{zeta_table[form]:.2f}")
        elif form == "直线形扭曲面":
            self.trans_outlet_zeta.setText("0.14")

    def _on_oc_trans_form_changed(self, form):
        """明渠渐变段形式变化时自动更新ζ系数（使用进口系数）"""
        zeta_table = TRANSITION_ZETA_COEFFICIENTS.get("进口", {})
        if form in zeta_table:
            self.oc_trans_zeta.setText(f"{zeta_table[form]:.2f}")
        elif form == "直线形扭曲面":
            self.oc_trans_zeta.setText("0.05")

    def _on_siphon_inlet_form_changed(self, form):
        """倒虹吸进口渐变段型式变化时自动更新ζ系数"""
        zeta_table = SIPHON_TRANSITION_ZETA_COEFFICIENTS.get("进口", {})
        if form in zeta_table:
            self.siphon_inlet_zeta.setText(f"{zeta_table[form]:.2f}")

    def _on_siphon_outlet_form_changed(self, form):
        """倒虹吸出口渐变段型式变化时自动更新ζ系数"""
        zeta_table = SIPHON_TRANSITION_ZETA_COEFFICIENTS.get("出口", {})
        if form in zeta_table:
            self.siphon_outlet_zeta.setText(f"{zeta_table[form]:.2f}")

    def _auto_calc_turn_radius(self):
        """根据规范自动计算推荐转弯半径，并弹出详细计算过程"""
        if not CALCULATOR_AVAILABLE:
            InfoBar.warning("提示", "核心计算引擎未加载",
                           parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return
        nodes = self._build_nodes_from_table()
        if not nodes:
            InfoBar.info("提示", "请先导入节点数据",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        import math as _math
        details = []
        max_r = 0.0
        controlling_name = ""
        for node in nodes:
            if not node.structure_type:
                continue
            sv = node.structure_type.value
            if "倒虹吸" in sv:
                continue
            b = node.section_params.get("B", 0)
            rc = node.section_params.get("R_circle", 0)
            m_s = node.section_params.get("m", 0)
            wd = node.water_depth or 0
            min_r = 0.0
            basis = ""
            dim_str = ""
            if "隧洞" in sv:
                if rc > 0:
                    min_r = rc * 2 * 5
                    dim_str = f"洞径D={rc*2:.2f}m"
                    basis = f"R ≥ D×5 = {rc*2:.2f}×5 = {min_r:.1f}m"
                elif b > 0:
                    min_r = b * 5
                    dim_str = f"洞宽B={b:.2f}m"
                    basis = f"R ≥ B×5 = {b:.2f}×5 = {min_r:.1f}m"
            elif "明渠" in sv or sv == "矩形":
                if "U形" in sv and rc > 0:
                    theta_u_a = node.section_params.get('theta_deg', 0) or 0
                    if theta_u_a > 0 and wd > 0:
                        import math as _math_ua
                        h0_u_a = rc * (1.0 - _math_ua.cos(_math_ua.radians(theta_u_a / 2.0)))
                        if wd <= h0_u_a:
                            wsw_ua = 2 * _math_ua.sqrt(max(0.0, rc ** 2 - (rc - wd) ** 2))
                            basis_detail = f"水位在弧区 h={wd:.3f}m ≤ h_0={h0_u_a:.3f}m，B=2√(R²-(R-h)²)"
                        else:
                            b_arc_u_a = 2 * rc * _math_ua.sin(_math_ua.radians(theta_u_a / 2.0))
                            wsw_ua = b_arc_u_a + 2 * m_s * (wd - h0_u_a)
                            basis_detail = f"水位在直线段 h={wd:.3f}m > h_0={h0_u_a:.3f}m，B=b_arc+2m(h-h_0)"
                        min_r = wsw_ua * 5
                        dim_str = f"R={rc:.2f}m, θ={theta_u_a}°, m={m_s:.3f}, h={wd:.2f}m, 水面宽={wsw_ua:.3f}m"
                        basis = f"R ≥ 水面宽×5 = {wsw_ua:.3f}×5 = {min_r:.1f}m（{basis_detail}）"
                    else:
                        min_r = rc * 2 * 5
                        dim_str = f"R={rc:.2f}m（θ缺失，用D=2R近似）"
                        basis = f"R ≥ D×5 = {rc*2:.2f}×5 = {min_r:.1f}m"
                elif b > 0:
                    wsw = b + 2 * m_s * wd if m_s > 0 and wd > 0 else b
                    min_r = wsw * 5
                    if m_s > 0 and wd > 0:
                        dim_str = f"B={b:.2f}m, m={m_s}, h={wd:.2f}m, 水面宽={wsw:.2f}m"
                        basis = f"R ≥ 水面宽×5 = {wsw:.2f}×5 = {min_r:.1f}m"
                    else:
                        dim_str = f"B={b:.2f}m"
                        basis = f"R ≥ B×5 = {b:.2f}×5 = {min_r:.1f}m"
            elif "渡槽" in sv:
                if b > 0:
                    min_r = b * 5
                    dim_str = f"连接明渠底宽B={b:.2f}m"
                    basis = f"R ≥ B×5 = {b:.2f}×5 = {min_r:.1f}m"
            elif "暗涵" in sv:
                if b > 0:
                    min_r = b * 5
                    dim_str = f"涵宽B={b:.2f}m"
                    basis = f"R ≥ B×5 = {b:.2f}×5 = {min_r:.1f}m"
            if min_r > 0:
                name = getattr(node, 'name', '') or sv
                details.append((name, sv, dim_str, basis, min_r))
                if min_r > max_r:
                    max_r = min_r
                    controlling_name = name
        if max_r <= 0:
            max_r = DEFAULT_TURN_RADIUS

        rec_r = _math.ceil(max_r)
        self.turn_radius_edit.setText(f"{rec_r:.1f}")

        from 渠系断面设计.water_profile.water_profile_dialogs import TurnRadiusCalcDialog
        dlg = TurnRadiusCalcDialog(
            self, rec_r=rec_r, max_r=max_r,
            details=details, controlling_name=controlling_name
        )
        dlg.exec()

    def _open_transition_reference(self):
        """打开渐变段参考系数表对话框（表K.1.2 + 表L.1.2）"""
        dlg = TransitionReferenceDialog(self)
        dlg.exec()

    def _update_summary_panel(self, nodes, total_len=0.0, wl_drop=None, summary=None):
        """更新持久摘要面板信息"""
        if not nodes:
            self.lbl_summary_info.setText("尚未计算")
            self.btn_building_stats.setEnabled(False)
            return

        parts = [f"节点数: {len(nodes)}"]
        if total_len > 0:
            parts.append(f"总长度: {total_len:.1f}m")
        if wl_drop is not None:
            parts.append(f"水位落差: {wl_drop:.3f}m")

        # 起终点桩号和水位
        if summary:
            start_st = summary.get('起点桩号', None)
            end_st = summary.get('终点桩号', None)
            start_wl = summary.get('起点水位', None)
            end_wl = summary.get('终点水位', None)
            if start_st is not None:
                parts.append(f"桩号: {start_st:.3f}~{end_st:.3f}")
            if start_wl is not None:
                parts.append(f"水位: {start_wl:.3f}~{end_wl:.3f}")

        # 建筑物统计
        building_count = len(self._last_building_lengths)
        if building_count > 0:
            parts.append(f"建筑物: {building_count}段")

        self.lbl_summary_info.setText("    ".join(parts))
        self.btn_building_stats.setEnabled(len(self._last_building_lengths) > 0)

    def _show_building_length_dialog(self):
        """打开建筑物长度统计对话框"""
        if not self._last_building_lengths:
            InfoBar.info("提示", "暂无建筑物长度数据，请先执行计算",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return
        from 渠系断面设计.water_profile.water_profile_dialogs import BuildingLengthDialog
        prefix = self._settings.get_station_prefix() if self._settings else ""
        dlg = BuildingLengthDialog(
            self,
            building_lengths=self._last_building_lengths,
            channel_total_length=self._last_channel_total_length,
            type_summary=self._last_type_summary,
            station_prefix=prefix
        )
        dlg.exec()

    def _parse_flow_values(self, flow_str):
        """解析流量字符串为浮点数列表，支持逗号分隔的多流量段"""
        if not flow_str or not flow_str.strip():
            return []
        flow_str = flow_str.replace('，', ',')
        values = []
        for q_str in flow_str.split(','):
            q_str = q_str.strip()
            if q_str:
                try:
                    values.append(float(q_str))
                except ValueError:
                    continue
        return values

    def _on_design_flow_changed(self):
        """设计流量变化时自动计算加大流量"""
        design_flows = self._parse_flow_values(self.design_flow_edit.text())
        if not design_flows:
            return
        max_flows = []
        for q in design_flows:
            if q <= 0:
                max_flows.append(0.0)
                continue
            # 根据设计流量计算加大流量百分比（灌排规范）
            if q < 1:
                pct = 30
            elif q < 5:
                pct = 25
            elif q < 20:
                pct = 20
            elif q < 50:
                pct = 15
            elif q < 100:
                pct = 10
            else:
                pct = 5
            max_flows.append(round(q * (1 + pct / 100), 3))
        # 格式化输出
        strs = []
        for q in max_flows:
            formatted = f"{q:.3f}".rstrip('0').rstrip('.')
            strs.append(formatted)
        self.max_flow_edit.setText(", ".join(strs))

    def _insert_transitions(self):
        """插入渐变段"""
        if not CALCULATOR_AVAILABLE:
            InfoBar.error("不可用", "核心计算引擎未加载",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        if self.node_table.rowCount() < 2:
            InfoBar.warning("节点不足", "至少需要2个节点才能插入渐变段",
                           parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        try:
            settings = self._build_settings()
            if not settings:
                return

            # 检查流量参数
            design_flows = self._parse_flow_values(self.design_flow_edit.text())
            max_flows = self._parse_flow_values(self.max_flow_edit.text())
            if not design_flows or not max_flows or all(q <= 0 for q in design_flows):
                InfoBar.info("提示", "请先点击【从批量计算导入】导入数据后，再点击【插入渐变段】。",
                            parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
                return

            # 验证设置
            is_valid, error_msg = settings.validate()
            if not is_valid:
                InfoBar.error("参数错误", error_msg,
                             parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
                return

            nodes = self._build_nodes_from_table()
            if len(nodes) < 2:
                InfoBar.warning("数据不足", "至少需要2个节点",
                               parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 检查是否已经插入过渐变段或自动明渠段
            has_transitions = any(getattr(n, 'is_transition', False) for n in nodes)
            has_auto_channels = any(getattr(n, 'is_auto_inserted_channel', False) for n in nodes)
            if has_transitions or has_auto_channels:
                if not fluent_question(self, "提示",
                        "表格中已存在渐变段行。\n\n"
                        "是否清除已有渐变段并重新插入？\n"
                        "（选「否」则保留现有渐变段不做任何操作）"):
                    return
                # 同时清除渐变段行和自动插入的明渠段行，避免重复插入
                nodes = [n for n in nodes
                         if not getattr(n, 'is_transition', False)
                         and not getattr(n, 'is_auto_inserted_channel', False)]

            import copy
            calculator = WaterProfileCalculator(settings)

            # 与原版Tkinter一致：不调用calculate_geometry。
            # station_MC等几何参数已在导入时由_recalculate_geometry()计算完毕并写入表格，
            # _build_nodes_from_table()已从表格读取到正确的station_MC值。
            # 如果此处再调用calculate_geometry，会因为in_out已是INLET/OUTLET
            # 而跳过进/出节点转角计算，导致station_MC被覆盖为错误值。
            calculator.preprocess_nodes(nodes)

            # ===== 预扫描明渠段缺口 =====
            gaps = calculator.pre_scan_open_channels(nodes)

            # 批量处理状态
            batch_state = {
                'mode': 'manual',
                'current_index': 0,
                'total_count': len(gaps),
                'preset_params': {},
                'inserted_channels': []
            }

            # 若有多处（≥2）需要插入明渠段，弹出批量选择对话框
            from 渠系断面设计.water_profile.water_profile_dialogs import (
                BatchChannelConfirmDialog, OpenChannelDialog, OpenChannelParams
            )
            if len(gaps) >= 2:
                batch_dlg = BatchChannelConfirmDialog(self, len(gaps), gaps)
                batch_dlg.exec()
                batch_result = batch_dlg.get_result()

                if batch_result['mode'] == BatchChannelConfirmDialog.RESULT_CANCELLED:
                    return
                elif batch_result['mode'] == BatchChannelConfirmDialog.RESULT_TABLE_EDIT:
                    batch_state['mode'] = 'table_edit'
                    batch_state['preset_params'] = batch_result['params']

            # 创建明渠段参数获取回调
            def open_channel_callback(upstream_channel, available_length,
                                       prev_struct, next_struct, flow_section, flow):
                idx = batch_state['current_index']
                batch_state['current_index'] += 1

                def _track(params, source):
                    batch_state['inserted_channels'].append({
                        'gap_index': idx,
                        'prev_struct': prev_struct,
                        'next_struct': next_struct,
                        'available_length': available_length,
                        'params': params,
                        'source': source,
                    })

                # ① 表格编辑模式
                if batch_state['mode'] == 'table_edit' and idx in batch_state.get('preset_params', {}):
                    p = batch_state['preset_params'][idx]
                    _track(p, '表格编辑')
                    return p

                # ② 自动推荐模式
                if batch_state['mode'] == 'auto_recommend' and upstream_channel:
                    p = OpenChannelParams(
                        name="-",
                        structure_type=upstream_channel.get("structure_type", "明渠-梯形"),
                        bottom_width=upstream_channel.get("bottom_width", 0),
                        water_depth=upstream_channel.get("water_depth", 0),
                        side_slope=upstream_channel.get("side_slope", 0),
                        roughness=upstream_channel.get("roughness", 0.014),
                        slope_inv=upstream_channel.get("slope_inv", 3000),
                        flow=upstream_channel.get("flow", flow),
                        flow_section=upstream_channel.get("flow_section", flow_section),
                        structure_height=upstream_channel.get("structure_height", 0.0),
                        arc_radius=upstream_channel.get("arc_radius", 0.0),
                        theta_deg=upstream_channel.get("theta_deg", 0.0),
                    )
                    _track(p, '推荐')
                    return p

                # ③ 手动模式：逐一弹窗
                dlg = OpenChannelDialog(
                    self,
                    upstream_channel=upstream_channel,
                    available_length=available_length,
                    prev_structure=prev_struct,
                    next_structure=next_struct,
                    flow_section=flow_section,
                    flow=flow,
                    current_index=idx + 1,
                    total_count=batch_state['total_count']
                )
                if dlg.exec() == QDialog.DialogCode.Accepted:
                    result = dlg.get_result()
                    if result:
                        _track(result, '手动')
                    if dlg.apply_all_remaining:
                        batch_state['mode'] = 'auto_recommend'
                    return result
                return None

            # ===== 执行：预处理 + 插入渐变段 + 几何计算 =====
            prepared_nodes = calculator.prepare_transitions(nodes, open_channel_callback)

            # 更新表格显示（使用完整刷新，显示几何计算结果）
            prefix = settings.get_station_prefix() if settings else ""
            self._update_table_from_nodes_full(prepared_nodes, prefix)
            auto_resize_table(self.node_table)

            # 统计
            transition_count = sum(1 for n in prepared_nodes if getattr(n, 'is_transition', False))
            open_channel_count = len(batch_state.get('inserted_channels', []))
            original_count = len(prepared_nodes) - transition_count - open_channel_count

            # 统计建筑物长度（几何计算完成后即可统计）
            try:
                if len(prepared_nodes) >= 2 and getattr(prepared_nodes[-1], 'station_MC', 0):
                    building_lengths = calculator.calculate_building_lengths(prepared_nodes)
                    channel_total_length = prepared_nodes[-1].station_MC - prepared_nodes[0].station_MC
                    type_summary = calculator.calculate_comprehensive_type_summary(prepared_nodes)
                    self._last_building_lengths = building_lengths
                    self._last_channel_total_length = channel_total_length
                    self._last_type_summary = type_summary
                    self._update_summary_panel(prepared_nodes, channel_total_length)
            except Exception:
                pass  # 渐变段插入阶段统计失败不影响主流程

            # 检查是否有倒虹吸
            has_siphon = any(
                n.structure_type and "倒虹吸" in n.structure_type.value
                for n in prepared_nodes
                if n.structure_type and not getattr(n, 'is_transition', False)
            )
            has_pressure_pipe = any(
                n.structure_type and "有压管道" in n.structure_type.value
                for n in prepared_nodes
                if n.structure_type and not getattr(n, 'is_transition', False)
            )

            # 汇总信息（InfoBar非阻塞通知）
            summary = f"渐变段插入完成！共 {len(prepared_nodes)} 行（渐变段 {transition_count}，明渠段 {open_channel_count}）"
            if has_siphon or has_pressure_pipe:
                step_parts = []
                if has_siphon:
                    step_parts.append("【倒虹吸水力计算】")
                if has_pressure_pipe:
                    step_parts.append("【有压管道水力计算】")
                step_parts.append("【执行计算】")
                next_step = "下一步：请点击" + "→".join(step_parts)
            else:
                next_step = "下一步：请点击【执行计算】"
            InfoBar.success(summary, next_step,
                           parent=self._info_parent(), duration=6000, position=InfoBarPosition.TOP)

        except Exception as e:
            import traceback
            traceback.print_exc()
            InfoBar.error("插入渐变段失败", str(e),
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _update_table_from_nodes(self, nodes):
        """从节点列表更新输入表格（插入渐变段后刷新）"""
        self._updating_cells = True
        try:
            self._update_table_from_nodes_inner(nodes)
        finally:
            self._updating_cells = False

    def _update_table_from_nodes_inner(self, nodes):
        self.node_table.setRowCount(0)
        for node in nodes:
            r = self.node_table.rowCount()
            self.node_table.insertRow(r)
            _is_trans = getattr(node, 'is_transition', False)
            _is_auto_ch = getattr(node, 'is_auto_inserted_channel', False)
            vals = [""] * len(NODE_ALL_HEADERS)
            # 基础输入列 (0-7)
            vals[0] = node.flow_section
            vals[1] = node.name
            _st_str = node.get_structure_type_str()
            vals[2] = f"{_st_str}(连接段)" if _is_auto_ch else _st_str
            if not _is_trans:
                vals[5] = f"{node.x:.6f}" if (node.x and not _is_auto_ch) else ""
                vals[6] = f"{node.y:.6f}" if (node.y and not _is_auto_ch) else ""
                vals[7] = "" if r == 0 else (f"{node.turn_radius:.1f}" if node.turn_radius else "")
                # 水力输入列 (20-26)
                vals[20] = f"{node.section_params.get('B', '')}" if node.section_params.get('B') else ""
                vals[21] = f"{node.section_params.get('D', '')}" if node.section_params.get('D') else ""
                vals[22] = f"{node.section_params.get('R_circle', '')}" if node.section_params.get('R_circle') else ""
                vals[23] = f"{node.section_params.get('m', '')}" if node.section_params.get('m') else ""
                vals[24] = f"{node.roughness}" if node.roughness else ""
                vals[25] = f"{1.0/node.slope_i:.0f}" if node.slope_i and node.slope_i > 0 else ""
                vals[26] = f"{node.flow}" if node.flow else ""
            for c, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                item.setTextAlignment(Qt.AlignCenter)
                if c not in EDITABLE_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                # 第一行（水位起点）锁定水头损失列
                if r == 0 and c in FIRST_ROW_LOCKED_LOSS_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if _is_trans:
                    item.setForeground(QColor("#9E9E9E"))
                elif _is_auto_ch:
                    item.setForeground(QColor("#2E7D32"))
                    item.setToolTip("自动插入的明渠连接段，用于计算两个建筑物之间的沿程及弯道水头损失。\n几何列留空因为该行不是真实IP转折点。")
                self.node_table.setItem(r, c, item)
        auto_resize_table(self.node_table)

    def _open_siphon_calculator(self):
        """打开倒虹吸水力计算（PySide6 多标签页窗口）"""
        print("[DEBUG] _open_siphon_calculator 被调用")
        if not CALCULATOR_AVAILABLE:
            print("[DEBUG] CALCULATOR_AVAILABLE = False，返回")
            InfoBar.error("不可用", "核心计算引擎未加载",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        nodes = self._build_nodes_from_table()
        if not nodes:
            print("[DEBUG] nodes 为空，返回")
            InfoBar.info("提示", "表格中没有数据，请先导入断面参数",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        # 检查是否已插入渐变段
        has_transitions = any(getattr(n, 'is_transition', False) for n in nodes)
        if not has_transitions:
            print("[DEBUG] has_transitions = False，返回")
            InfoBar.warning("提示",
                           "请先点击工具栏的【插入渐变段】按钮，完成渐变段插入后再进行倒虹吸水力计算。\n"
                           "插入渐变段后，系统才能准确获取倒虹吸上下游流速、断面参数等信息。",
                           parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        # 检查是否有倒虹吸
        has_siphon = any(
            n.structure_type and "倒虹吸" in n.structure_type.value
            for n in nodes if n.structure_type
        )
        print(f"[DEBUG] has_siphon = {has_siphon}")
        if not has_siphon:
            print("[DEBUG] has_siphon = False，返回")
            InfoBar.info("提示", "表格中没有倒虹吸数据，请确保有结构形式为\"倒虹吸\"的行",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        print("[DEBUG] 开始导入模块和提取倒虹吸分组")
        try:
            from 渠系断面设计.siphon.multi_siphon_dialog import MultiSiphonDialog

            # 提取倒虹吸分组
            _water_profile_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                '推求水面线'
            )
            import sys as _sys
            if _water_profile_dir not in _sys.path:
                _sys.path.insert(0, _water_profile_dir)
            from utils.siphon_extractor import SiphonDataExtractor

            settings = self._build_settings()
            siphon_groups = SiphonDataExtractor.extract_siphons(nodes, settings=settings)
            if not siphon_groups:
                InfoBar.info("提示", "未找到倒虹吸数据组",
                            parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 使用共享的 SiphonManager（已绑定项目路径）
            if self._siphon_manager is not None:
                manager = self._siphon_manager
            else:
                from managers.siphon_manager import SiphonManager
                manager = SiphonManager()

            # 定义导入回调：将水头损失和平面转弯半径写回节点表格
            _panel = self
            def import_losses_callback(results):
                cur_nodes = _panel._build_nodes_from_table()
                cur_groups = SiphonDataExtractor.extract_siphons(cur_nodes)
                imported_count = 0
                has_radius_update = False
                for group in cur_groups:
                    if group.name in results and results[group.name] is not None:
                        result_data = results[group.name]
                        if isinstance(result_data, dict):
                            head_loss = result_data.get("head_loss", 0.0)
                            diameter = result_data.get("diameter", 0.0)
                            turn_radius = result_data.get("turn_radius", 0.0)
                        else:
                            head_loss = result_data
                            diameter = 0.0
                            turn_radius = 0.0
                        outlet_idx = group.outlet_row_index
                        if 0 <= outlet_idx < len(cur_nodes):
                            cur_nodes[outlet_idx].head_loss_siphon = head_loss
                            imported_count += 1
                        if diameter > 0:
                            for row_idx in group.row_indices:
                                if 0 <= row_idx < len(cur_nodes):
                                    if not hasattr(cur_nodes[row_idx], 'section_params') or not cur_nodes[row_idx].section_params:
                                        cur_nodes[row_idx].section_params = {}
                                    cur_nodes[row_idx].section_params["D"] = diameter
                        # 将平面转弯半径写回该倒虹吸所有行（进口+出口）
                        if turn_radius > 0:
                            for row_idx in group.row_indices:
                                if 0 <= row_idx < len(cur_nodes):
                                    cur_nodes[row_idx].turn_radius = turn_radius
                                    has_radius_update = True
                if imported_count > 0 or has_radius_update:
                    _panel._append_loss_undo_snapshot(_panel._snapshot_editable_cols())
                    _s = _panel._build_settings()
                    _pfx = _s.get_station_prefix() if _s else ""
                    _panel._update_table_from_nodes_full(cur_nodes, _pfx)
                    auto_resize_table(_panel.node_table)
                return imported_count

            siphon_n = DEFAULT_SIPHON_TURN_RADIUS_N

            # 打开PySide6多标签页倒虹吸计算窗口
            print(f"[DEBUG] 正在创建 MultiSiphonDialog，倒虹吸组数量: {len(siphon_groups)}")
            dlg = MultiSiphonDialog(
                self._info_parent(),
                siphon_groups,
                manager=manager,
                on_import_losses=import_losses_callback,
                siphon_turn_radius_n=siphon_n,
                show_case_management=False
            )
            print(f"[DEBUG] MultiSiphonDialog 创建完成，准备调用 exec()")
            result = dlg.exec()
            print(f"[DEBUG] dlg.exec() 返回值: {result}")

        except ImportError as e:
            import traceback
            traceback.print_exc()
            InfoBar.warning("提示", f"倒虹吸水力计算模块加载失败: {str(e)}",
                           parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
        except Exception as e:
            import traceback
            traceback.print_exc()
            InfoBar.error("错误", f"打开倒虹吸计算窗口失败: {str(e)}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _get_pressure_pipe_group_flow_section(self, group) -> str:
        rows = getattr(group, "rows", None) or []
        for node in rows:
            fs = str(getattr(node, "flow_section", "") or "").strip()
            if fs:
                return fs
        return "-"

    def _build_pressure_pipe_group_identity(self, group) -> str:
        return make_pressure_pipe_identity(
            self._get_pressure_pipe_group_flow_section(group),
            getattr(group, "name", "") or ""
        )

    def _apply_pressure_pipe_results(self, results_by_identity: dict, batch_data: dict):
        """将有压管道计算结果回写到表格"""
        try:
            from utils.pressure_pipe_extractor import PressurePipeDataExtractor

            settings = self._build_settings()
            cur_nodes = self._build_nodes_from_table()
            cur_groups = PressurePipeDataExtractor.extract_pipes(cur_nodes, settings=settings)
            imported_count = 0

            for group in cur_groups:
                identity = self._build_pressure_pipe_group_identity(group)
                record = results_by_identity.get(identity)
                if not record:
                    continue
                head_loss = record.get("total_head_loss")
                if head_loss is None:
                    continue
                outlet_idx = group.outlet_row_index
                if 0 <= outlet_idx < len(cur_nodes):
                    cur_nodes[outlet_idx].head_loss_siphon = float(head_loss)
                    cur_nodes[outlet_idx].external_head_loss = None
                    self._pressure_pipe_calc_done[identity] = True
                    imported_count += 1

            if imported_count > 0:
                self._append_loss_undo_snapshot(self._snapshot_editable_cols())
                _s = self._build_settings()
                _pfx = _s.get_station_prefix() if _s else ""
                self.nodes = cur_nodes
                self._update_table_from_nodes_full(cur_nodes, _pfx)
                auto_resize_table(self.node_table)

            summary = batch_data.get("summary", {})
            success_count = int(summary.get("success", 0))
            failed_count = int(summary.get("failed", 0))

            if success_count <= 0:
                InfoBar.warning(
                    "有压管道计算完成（全部失败）",
                    f"共 {summary.get('total', 0)} 条，全部失败。请查看\"有压管道计算结果汇总\"。",
                    parent=self._info_parent(), duration=7000, position=InfoBarPosition.TOP
                )
            elif failed_count > 0:
                InfoBar.warning(
                    "有压管道计算完成（部分成功）",
                    f"成功 {success_count} 条，失败 {failed_count} 条；已回写 {imported_count} 条到\"倒虹吸/有压管道水头损失\"列。",
                    parent=self._info_parent(), duration=7000, position=InfoBarPosition.TOP
                )
            else:
                InfoBar.success(
                    "有压管道计算完成",
                    f"已完成 {success_count} 条计算并回写 {imported_count} 条到\"倒虹吸/有压管道水头损失\"列。",
                    parent=self._info_parent(), duration=6000, position=InfoBarPosition.TOP
                )
        except Exception as e:
            import traceback
            traceback.print_exc()
            InfoBar.error("错误", f"回写数据失败: {str(e)}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _append_pressure_pipe_calc_details(self, batch_data: dict):
        data = normalize_pressure_pipe_calc_records(batch_data)
        if not data.get("records"):
            return
        token = f"【有压管道计算详情】  时间: {data.get('last_run_at', '-')}"
        old = self.detail_text.toPlainText() if hasattr(self, "detail_text") else ""
        if token in old:
            return
        self.detail_text.setPlainText(append_pressure_pipe_calc_batch_text(old, data, precision=4))

    def _show_pressure_pipe_calc_summary_dialog(self, batch_data: dict, results_by_identity: dict = None):
        data = normalize_pressure_pipe_calc_records(batch_data)
        records = data.get("records", [])
        if not records:
            InfoBar.info("提示", "暂无有压管道计算记录",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        dlg = QWidget()
        dlg.setWindowTitle("有压管道计算结果汇总（请确认是否应用）")
        dlg.setMinimumWidth(980)
        dlg.resize(1120, 620)
        dlg.setStyleSheet(DIALOG_STYLE)
        dlg.setWindowFlags(Qt.Window)

        from PySide6.QtGui import QIcon
        _res_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "resources")
        for _icon_name in ("logo.ico", "logo.svg"):
            _icon_path = os.path.join(_res_dir, _icon_name)
            if os.path.exists(_icon_path):
                dlg.setWindowIcon(QIcon(_icon_path))
                break

        # 标志：是否通过确认按钮关闭
        dlg._confirmed = False

        def _on_close_event(event):
            if dlg._confirmed:
                event.accept()
                return

            from 渠系断面设计.styles import fluent_question
            reply = fluent_question(
                dlg,
                "关闭确认",
                "是否将计算结果应用到水面线计算表格？\n\n"
                "点击「是」：应用结果并关闭\n"
                "点击「否」：放弃结果并关闭"
            )
            if reply:
                if results_by_identity:
                    self._apply_pressure_pipe_results(results_by_identity, data)
            event.accept()

        dlg.closeEvent = _on_close_event

        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(18, 14, 18, 14)
        lay.setSpacing(10)

        summary = data.get("summary", {})
        ts = data.get("last_run_at", "") or "-"
        lbl_summary = QLabel(
            f"本次总条数: {summary.get('total', 0)}  |  "
            f"成功: {summary.get('success', 0)}  |  "
            f"失败: {summary.get('failed', 0)}  |  "
            f"时间: {ts}"
        )
        lbl_summary.setStyleSheet(f"font-size:13px;font-weight:bold;color:{T1};")
        lay.addWidget(lbl_summary)

        headers = [
            "查看", "流量段", "名称", "状态", "数据模式", "总损失(m)", "沿程(m)",
            "弯头(m)", "进口渐变(m)", "出口渐变(m)", "备注",
            "下限总损失（m）", "Δ总损(m)"
        ]
        table = QTableWidget(len(records), len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setFont(QFont("Microsoft YaHei", 10))
        hh = table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.Stretch)
        for col in [3, 4, 5, 6, 7, 8, 9]:
            hh.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(10, QHeaderView.Stretch)
        hh.setSectionResizeMode(11, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(12, QHeaderView.ResizeToContents)

        has_sensitivity_data = any(
            rec.get("sensitivity_low_total_head_loss") is not None for rec in records
        )
        show_sensitivity = has_sensitivity_data

        # ---- 对比摘要卡片 ----
        compare_card = QFrame()
        compare_card.setStyleSheet(
            "QFrame#compareCard {"
            "  background: qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #f8fbff, stop:1 #f0f5fc);"
            "  border: 1px solid #c8dff5; border-radius: 6px;"
            "}"
        )
        compare_card.setObjectName("compareCard")
        compare_card_lay = QVBoxLayout(compare_card)
        compare_card_lay.setContentsMargins(14, 10, 14, 10)
        compare_card_lay.setSpacing(8)

        _card_title = QLabel("📊  球墨铸铁管 f 值对比摘要")
        _card_title.setStyleSheet("font-size:13px;font-weight:600;color:#0058a3;background:transparent;")
        compare_card_lay.addWidget(_card_title)

        compare_grid = QGridLayout()
        compare_grid.setSpacing(8)

        _compare_items_data = []
        for rec in records:
            if rec.get("status") != "success":
                continue
            main_val = rec.get("total_head_loss")
            low_val = rec.get("sensitivity_low_total_head_loss")
            delta_val = rec.get("sensitivity_delta_total_head_loss")
            if main_val is None or low_val is None:
                continue
            try:
                main_f = float(main_val)
                low_f = float(low_val)
                delta_f = float(delta_val) if delta_val is not None else (low_f - main_f)
                pct = (delta_f / main_f * 100) if main_f != 0 else 0
            except (TypeError, ValueError):
                continue
            _compare_items_data.append({
                "name": rec.get("name", "未命名"),
                "flow_section": rec.get("flow_section", "-"),
                "main": main_f, "low": low_f, "delta": delta_f, "pct": pct,
            })

        _item_style = (
            "QFrame { background: #fff; border: 1px solid #e0eaf5;"
            " border-radius: 4px; }"
        )
        for idx, citem in enumerate(_compare_items_data):
            row_label = QLabel(f"流量段 {citem['flow_section']}  —  {citem['name']}")
            row_label.setStyleSheet("font-size:12px;font-weight:600;color:#333;background:transparent;")
            compare_grid.addWidget(row_label, idx * 2, 0, 1, 3)

            for col, (label_t, val, color) in enumerate([
                (f"主值 (f=223200)", f"{citem['main']:.4f} m", "#1a1a1a"),
                (f"下限 (f=189900)", f"{citem['low']:.4f} m", "#1a1a1a"),
                (f"差值 (下限−主值)", f"{citem['delta']:+.4f} m  ({citem['pct']:+.1f}%)", "#0078d4"),
            ]):
                _f = QFrame()
                _f.setStyleSheet(_item_style)
                _fl = QVBoxLayout(_f)
                _fl.setContentsMargins(10, 6, 10, 6)
                _fl.setSpacing(2)
                _lbl = QLabel(label_t)
                _lbl.setStyleSheet("font-size:11px;color:#666;background:transparent;border:none;")
                _lbl.setAlignment(Qt.AlignCenter)
                _val = QLabel(val)
                _val.setStyleSheet(f"font-size:14px;font-weight:600;color:{color};background:transparent;border:none;")
                _val.setAlignment(Qt.AlignCenter)
                _fl.addWidget(_lbl)
                _fl.addWidget(_val)
                compare_grid.addWidget(_f, idx * 2 + 1, col)

        compare_card_lay.addLayout(compare_grid)

        if not _compare_items_data:
            _no_data = QLabel("暂无对比数据")
            _no_data.setStyleSheet("font-size:12px;color:#999;background:transparent;")
            _no_data.setAlignment(Qt.AlignCenter)
            compare_card_lay.addWidget(_no_data)

        compare_card.setVisible(show_sensitivity)

        def _set_sensitivity_columns_visible(visible: bool):
            table.setColumnHidden(11, not visible)
            table.setColumnHidden(12, not visible)
            compare_card.setVisible(visible)

        _set_sensitivity_columns_visible(show_sensitivity)

        preview = QTextEdit()
        preview.setReadOnly(True)
        preview.setFont(QFont("Consolas", 10))
        preview.setMinimumHeight(220)

        def _fmt(v):
            try:
                return f"{float(v):.4f}"
            except (TypeError, ValueError):
                return "-"

        def _show_record_detail(rec: dict):
            preview.setPlainText(format_pressure_pipe_record_detail(rec, precision=4))
            # 与下方过程框保持同步（若未写入则追加，已写入则不重复）
            self._append_pressure_pipe_calc_details(data)

        for i, rec in enumerate(records):
            btn = PushButton("查看详情")
            btn.clicked.connect(lambda checked=False, r=rec: _show_record_detail(r))
            table.setCellWidget(i, 0, btn)

            table.setItem(i, 1, QTableWidgetItem(str(rec.get("flow_section", "") or "-")))
            table.setItem(i, 2, QTableWidgetItem(str(rec.get("name", "") or "未命名")))

            status_ok = rec.get("status") == "success"
            status_text = "成功" if status_ok else "失败"
            status_item = QTableWidgetItem(status_text)
            status_item.setForeground(QColor("#2E7D32" if status_ok else "#C62828"))
            table.setItem(i, 3, status_item)
            table.setItem(i, 4, QTableWidgetItem(str(rec.get("data_mode", "") or ("平面模式" if status_ok else "-"))))

            if status_ok:
                table.setItem(i, 5, QTableWidgetItem(_fmt(rec.get("total_head_loss"))))
                table.setItem(i, 6, QTableWidgetItem(_fmt(rec.get("friction_loss"))))
                table.setItem(i, 7, QTableWidgetItem(_fmt(rec.get("total_bend_loss"))))
                table.setItem(i, 8, QTableWidgetItem(_fmt(rec.get("inlet_transition_loss"))))
                table.setItem(i, 9, QTableWidgetItem(_fmt(rec.get("outlet_transition_loss"))))
                note_text = (rec.get("note", "") or "").strip()
                table.setItem(i, 11, QTableWidgetItem(_fmt(rec.get("sensitivity_low_total_head_loss"))))
                table.setItem(i, 12, QTableWidgetItem(_fmt(rec.get("sensitivity_delta_total_head_loss"))))
            else:
                for col in [5, 6, 7, 8, 9, 11, 12]:
                    table.setItem(i, col, QTableWidgetItem("-"))
                note_text = (rec.get("error", "") or "").strip() or "计算失败"
            table.setItem(i, 10, QTableWidgetItem(note_text))

        lay.addWidget(table, 1)
        lay.addWidget(compare_card)
        lay.addWidget(QLabel("计算详情预览（标准深度）"))
        lay.addWidget(preview, 1)

        if records:
            _show_record_detail(records[0])

        btn_lay = QHBoxLayout()
        btn_lay.addStretch()

        btn_apply = PrimaryPushButton("关闭并将总水头损失返回至水面线计算表格")

        def _apply_and_close():
            if not results_by_identity:
                InfoBar.warning("提示", "没有可应用的计算结果",
                               parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 检查是否有成功的计算结果
            has_success = any(
                rec.get("status") == "success" and rec.get("total_head_loss") is not None
                for rec in records
            )
            if not has_success:
                InfoBar.warning("提示", "所有计算均失败，无法应用结果",
                               parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            self._apply_pressure_pipe_results(results_by_identity, data)
            dlg._confirmed = True
            dlg.close()

        btn_apply.clicked.connect(_apply_and_close)
        btn_lay.addWidget(btn_apply)

        btn_cancel = PushButton("取消")
        btn_cancel.clicked.connect(dlg.close)
        btn_lay.addWidget(btn_cancel)

        lay.addLayout(btn_lay)
        dlg.show()

    def _open_pressure_pipe_calculator(self):
        """打开有压管道水力计算窗口"""
        print("[DEBUG] _open_pressure_pipe_calculator 被调用")
        if not CALCULATOR_AVAILABLE:
            print("[DEBUG] CALCULATOR_AVAILABLE = False，返回")
            InfoBar.error("不可用", "核心计算引擎未加载",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        nodes = self._build_nodes_from_table()
        if not nodes:
            print("[DEBUG] nodes 为空，返回")
            InfoBar.info("提示", "表格中没有数据，请先导入断面参数",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        # 检查是否已插入渐变段
        has_transitions = any(getattr(n, 'is_transition', False) for n in nodes)
        if not has_transitions:
            print("[DEBUG] has_transitions = False，返回")
            InfoBar.warning("提示",
                           "请先点击工具栏的【插入渐变段】按钮，完成渐变段插入后再进行有压管道水力计算。\n"
                           "插入渐变段后，系统才能准确获取有压管道上下游流速、断面参数等信息。",
                           parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        # 检查是否有有压管道
        has_ppipe = any(
            n.structure_type and "有压管道" in n.structure_type.value
            for n in nodes if n.structure_type
        )
        print(f"[DEBUG] has_ppipe = {has_ppipe}")
        if not has_ppipe:
            print("[DEBUG] has_ppipe = False，返回")
            InfoBar.info("提示", "表格中没有有压管道数据，请确保有结构形式为\"有压管道\"的行",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        # 先提取pipe_groups和manager
        print("[DEBUG] 开始提取有压管道分组")
        try:
            _water_profile_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                '推求水面线'
            )
            import sys as _sys
            if _water_profile_dir not in _sys.path:
                _sys.path.insert(0, _water_profile_dir)
            from utils.pressure_pipe_extractor import PressurePipeDataExtractor

            settings = self._build_settings()
            pipe_groups = PressurePipeDataExtractor.extract_pipes(nodes, settings=settings)
            if not pipe_groups:
                InfoBar.info("提示", "未找到有压管道数据组",
                            parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 使用共享的 PressurePipeManager（已绑定项目路径）
            if self._pressure_pipe_manager is not None:
                manager = self._pressure_pipe_manager
            else:
                from managers.pressure_pipe_manager import PressurePipeManager
                manager = PressurePipeManager()
        except Exception as e:
            InfoBar.error("错误", f"初始化失败: {e}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        # 弹出配置对话框
        from 渠系断面设计.water_profile.water_profile_dialogs import PressurePipeConfigDialog
        config_dlg = PressurePipeConfigDialog(
            parent=self,
            pipe_groups=pipe_groups,
            manager=manager
        )
        if config_dlg.exec() != QDialog.Accepted:
            print("[DEBUG] 用户取消了配置对话框")
            return

        # 获取用户配置
        longitudinal_nodes_dict = config_dlg.get_longitudinal_nodes_dict()
        print(f"[DEBUG] 纵断面数据: {list(longitudinal_nodes_dict.keys())}")

        print("[DEBUG] 开始导入模块和提取有压管道分组")
        try:
            # 导入有压管道相关模块
            _water_profile_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                '推求水面线'
            )
            import sys as _sys
            if _water_profile_dir not in _sys.path:
                _sys.path.insert(0, _water_profile_dir)
            from utils.pressure_pipe_extractor import PressurePipeDataExtractor
            from core.pressure_pipe_calc import calc_total_head_loss, calc_total_head_loss_with_spatial, PIPE_MATERIALS

            settings = self._build_settings()
            pipe_groups = PressurePipeDataExtractor.extract_pipes(nodes, settings=settings)
            if not pipe_groups:
                InfoBar.info("提示", "未找到有压管道数据组",
                            parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 使用共享的 PressurePipeManager（已绑定项目路径）
            if self._pressure_pipe_manager is not None:
                manager = self._pressure_pipe_manager
            else:
                from managers.pressure_pipe_manager import PressurePipeManager
                manager = PressurePipeManager()

            # 逐条有压管道计算总水头损失并记录完整过程（标准深度）
            results_by_identity = {}
            records = []
            default_material = "预应力钢筒混凝土管"
            run_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            SENSITIVITY_LOW_F = 1.899e5

            for group in pipe_groups:
                flow_section = self._get_pressure_pipe_group_flow_section(group)
                pipe_name = (getattr(group, "name", "") or "").strip() or "未命名"
                identity = self._build_pressure_pipe_group_identity(group)
                base_record = {
                    "identity": identity,
                    "flow_section": flow_section,
                    "name": pipe_name,
                }

                if not group.is_valid():
                    msg = group.get_validation_message()
                    records.append({
                        **base_record,
                        "status": "failed",
                        "error": msg or f"{pipe_name}: 数据不完整，已跳过",
                    })
                    continue

                material_key = group.material_key if group.material_key in PIPE_MATERIALS else default_material
                note = ""
                if group.material_key not in PIPE_MATERIALS:
                    note = f"未识别管材\"{group.material_key}\"，已按\"{default_material}\"计算"

                try:
                    # 判断是否有纵断面数据
                    pipe_long_nodes = longitudinal_nodes_dict.get(pipe_name, [])

                    if pipe_long_nodes:
                        # 使用空间模式计算
                        calc_res = calc_total_head_loss_with_spatial(
                            name=pipe_name,
                            Q=group.design_flow,
                            D=group.diameter,
                            material_key=material_key,
                            ip_points=group.ip_points,
                            longitudinal_nodes=pipe_long_nodes,
                            upstream_velocity=group.upstream_velocity,
                            downstream_velocity=group.downstream_velocity,
                            inlet_transition_form=group.inlet_transition_form,
                            outlet_transition_form=group.outlet_transition_form,
                            inlet_transition_zeta=group.inlet_transition_zeta,
                            outlet_transition_zeta=group.outlet_transition_zeta,
                        )
                    else:
                        # 使用平面模式计算
                        calc_res = calc_total_head_loss(
                            name=pipe_name,
                            Q=group.design_flow,
                            D=group.diameter,
                            material_key=material_key,
                            ip_points=group.ip_points,
                            upstream_velocity=group.upstream_velocity,
                            downstream_velocity=group.downstream_velocity,
                            inlet_transition_form=group.inlet_transition_form,
                            outlet_transition_form=group.outlet_transition_form,
                            inlet_transition_zeta=group.inlet_transition_zeta,
                            outlet_transition_zeta=group.outlet_transition_zeta,
                        )
                except Exception as ex:
                    records.append({
                        **base_record,
                        "status": "failed",
                        "Q": group.design_flow,
                        "D": group.diameter,
                        "material_key": material_key,
                        "error": f"计算失败: {ex}",
                        "note": note,
                    })
                    continue

                record = {
                    **base_record,
                    "status": "success",
                    "Q": group.design_flow,
                    "D": group.diameter,
                    "material_key": material_key,
                    "total_length": calc_res.total_length,
                    "pipe_velocity": calc_res.pipe_velocity,
                    "friction_loss": calc_res.friction_loss,
                    "total_bend_loss": calc_res.total_bend_loss,
                    "inlet_transition_loss": calc_res.inlet_transition_loss,
                    "outlet_transition_loss": calc_res.outlet_transition_loss,
                    "total_head_loss": calc_res.total_head_loss,
                    "calc_steps": calc_res.calc_steps,
                    "data_mode": calc_res.data_mode,
                    "note": note,
                }
                if material_key == "球墨铸铁管":
                    fr = calc_res.friction_details or {}
                    main_f = fr.get("f")
                    q_m3h = fr.get("Q_m3h")
                    d_mm = fr.get("d_mm")
                    m_exp = fr.get("m")
                    b_exp = fr.get("b")
                    low_friction_loss = None
                    try:
                        if all(v is not None for v in [q_m3h, d_mm, m_exp, b_exp]) and float(d_mm) > 0:
                            low_friction_loss = (
                                SENSITIVITY_LOW_F * calc_res.total_length * (float(q_m3h) ** float(m_exp))
                                / (float(d_mm) ** float(b_exp))
                            )
                        elif main_f:
                            low_friction_loss = calc_res.friction_loss * (SENSITIVITY_LOW_F / float(main_f))
                    except Exception:
                        low_friction_loss = None

                    if low_friction_loss is not None:
                        low_total_head_loss = (
                            float(low_friction_loss)
                            + float(calc_res.total_bend_loss)
                            + float(calc_res.inlet_transition_loss)
                            + float(calc_res.outlet_transition_loss)
                        )
                        record.update({
                            "sensitivity_material": "球墨铸铁管",
                            "sensitivity_main_f": main_f,
                            "sensitivity_low_f": SENSITIVITY_LOW_F,
                            "sensitivity_low_friction_loss": low_friction_loss,
                            "sensitivity_low_total_head_loss": low_total_head_loss,
                            "sensitivity_delta_total_head_loss": low_total_head_loss - float(calc_res.total_head_loss),
                        })
                records.append(record)
                results_by_identity[identity] = record

                # 持久化计算结果，便于后续追溯
                manager.set_result(
                    pipe_name,
                    total_head_loss=calc_res.total_head_loss,
                    friction_loss=calc_res.friction_loss,
                    total_bend_loss=calc_res.total_bend_loss,
                    inlet_transition_loss=calc_res.inlet_transition_loss,
                    outlet_transition_loss=calc_res.outlet_transition_loss,
                    pipe_velocity=calc_res.pipe_velocity,
                    plan_total_length=calc_res.total_length,
                    data_mode=calc_res.data_mode,
                    longitudinal_nodes=pipe_long_nodes,
                )

            batch_data = normalize_pressure_pipe_calc_records({
                "last_run_at": run_at,
                "records": records,
            })
            self._pressure_pipe_calc_records = batch_data
            self._pressure_pipe_last_run_at = batch_data.get("last_run_at", "")
            self._update_pressure_pipe_last_result_button()

            # 追加结构化过程到下方详情框 + 弹出汇总对话框（不立即回写）
            self._append_pressure_pipe_calc_details(batch_data)
            self._show_pressure_pipe_calc_summary_dialog(batch_data, results_by_identity)

        except ImportError as e:
            import traceback
            traceback.print_exc()
            InfoBar.warning("提示", f"有压管道水力计算模块加载失败: {str(e)}",
                           parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
        except Exception as e:
            import traceback
            traceback.print_exc()
            InfoBar.error("错误", f"打开有压管道计算窗口失败: {str(e)}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    # ================================================================
    # 导出
    # ================================================================
    def _export_excel(self):
        if not self.calculated_nodes:
            InfoBar.warning("提示", "无结果可导出，请先执行计算", parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            InfoBar.warning("缺少依赖", "需要: pip install openpyxl", parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            return
        ch_name = self.channel_name_edit.text().strip()
        ch_level = self.channel_level_combo.currentText()
        auto_name = f"{ch_name}{ch_level}_水面线计算结果.xlsx" if ch_name else "水面线计算结果.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(self, "导出Excel", auto_name, "Excel文件 (*.xlsx)")
        if not filepath: return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "水面线计算结果"
            ncols = len(NODE_EXPORT_HEADERS)
            # 第1行：标题
            ws['A1'] = f"{ch_name}{ch_level} 水面线计算结果"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
            ws['A1'].alignment = Alignment(horizontal='center')
            # 第2行：基础信息
            ws['A2'] = "渠道名称"; ws['A2'].font = Font(bold=True); ws['B2'] = ch_name
            ws['C2'] = "渠道级别"; ws['C2'].font = Font(bold=True); ws['D2'] = ch_level
            ws['E2'] = "起始水位(m)"; ws['E2'].font = Font(bold=True); ws['F2'] = self.start_wl_edit.text()
            ws['G2'] = "起始桩号"; ws['G2'].font = Font(bold=True); ws['H2'] = self.start_station_edit.text()
            ws['I2'] = "设计流量"; ws['I2'].font = Font(bold=True); ws['J2'] = self.design_flow_edit.text()
            ws['K2'] = "加大流量"; ws['K2'].font = Font(bold=True); ws['L2'] = self.max_flow_edit.text()
            # 第3行：表头
            hdr_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            for c, h in enumerate(NODE_EXPORT_HEADERS, 1):
                cell = ws.cell(row=3, column=c, value=h)
                cell.font = Font(bold=True)
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
            # 第4行起：从统一表格读取数据
            for r in range(self.node_table.rowCount()):
                for c in range(self.node_table.columnCount()):
                    item = self.node_table.item(r, c)
                    ws.cell(row=r+4, column=c+1, value=item.text() if item else "")
            # 自动列宽
            for col_num in range(1, ncols + 1):
                max_len = len(str(NODE_EXPORT_HEADERS[col_num-1]))
                for row_num in range(4, ws.max_row + 1):
                    cv = ws.cell(row=row_num, column=col_num).value
                    if cv: max_len = max(max_len, len(str(cv)))
                ws.column_dimensions[get_column_letter(col_num)].width = min(max_len + 3, 30)
            wb.save(filepath)
            InfoBar.success("导出成功", f"已保存: {filepath}", parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            ask_open_file(filepath, self._info_parent())
        except PermissionError:
            InfoBar.error("文件被占用", f"无法写入文件，请先关闭已打开的同名文件（如Excel等），然后重新操作。", parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("导出失败", str(e), parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _export_word(self):
        if not WORD_EXPORT_AVAILABLE:
            InfoBar.warning("缺少依赖",
                "Word导出需要安装 python-docx、latex2mathml、lxml。请执行: pip install python-docx latex2mathml lxml",
                parent=self._info_parent(), duration=6000, position=InfoBarPosition.TOP)
            return
        if not self.calculated_nodes:
            InfoBar.warning("提示", "请先进行计算。", parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return
        ch_name = self.channel_name_edit.text().strip()
        meta = load_meta()
        auto_purpose = build_calc_purpose('water_profile', project=meta.project_name, name=ch_name, section_type='')
        dlg = ExportConfirmDialog('water_profile', '推求水面线计算书', auto_purpose, parent=self._info_parent())
        from PySide6.QtWidgets import QDialog
        if dlg.exec() != QDialog.Accepted:
            return
        self._word_export_meta = dlg.get_meta()
        self._word_export_purpose = dlg.get_calc_purpose()
        self._word_export_refs = dlg.get_references()
        ch_level = self.channel_level_combo.currentText()
        auto_name = f"{ch_name}_水面线计算书.docx" if ch_name else "水面线计算书.docx"
        filepath, _ = QFileDialog.getSaveFileName(self, "保存Word报告", auto_name, "Word文档 (*.docx);;所有文件 (*.*)")
        if not filepath: return
        try:
            self._build_word_report(filepath)
            InfoBar.success("导出成功", f"Word报告已保存到: {filepath}", parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            ask_open_file(filepath, self._info_parent())
        except PermissionError:
            InfoBar.error("文件被占用", "无法写入文件，请先关闭已打开的同名Word文档，然后重新操作。", parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("导出失败", f"Word导出失败: {str(e)}", parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _build_word_report(self, filepath):
        """构建水面线计算Word报告（工程产品运行卡格式）"""
        settings = self._settings
        nodes = self.calculated_nodes
        ch_name = settings.channel_name if settings else ""
        ch_level = settings.channel_level if settings else ""
        prefix = settings.get_station_prefix() if settings else ""
        meta = getattr(self, '_word_export_meta', load_meta())
        purpose = getattr(self, '_word_export_purpose', '')
        refs = getattr(self, '_word_export_refs', REFERENCES_BASE.get('water_profile', []))

        doc = create_engineering_report_doc(
            meta=meta,
            calc_title='推求水面线计算书',
            calc_content_desc=f'{ch_name}水面线推求计算' if ch_name else '水面线推求计算',
            calc_purpose=purpose,
            references=refs,
            calc_program_text=f'渠系建筑物水力计算系统 V1.0\n推求水面线',
        )
        doc.add_page_break()

        # 5. 基本计算参数
        doc_add_eng_h(doc, '5、基本计算参数')
        params = []
        params.append(("渠道名称", ch_name or "-"))
        params.append(("渠道级别", ch_level or "-"))
        if settings:
            params.append(("起始水位", f"{settings.start_water_level} m"))
            params.append(("起始桩号", ProjectSettings.format_station(settings.start_station, prefix) if settings.start_station else "-"))
            if getattr(settings, 'design_flows', None):
                params.append(("设计流量", ", ".join(f"{q:.3f}" for q in settings.design_flows) + " m³/s"))
            else:
                params.append(("设计流量", f"{settings.design_flow} m³/s"))
            if getattr(settings, 'max_flows', None):
                params.append(("加大流量", ", ".join(f"{q:.3f}" for q in settings.max_flows) + " m³/s"))
            else:
                params.append(("加大流量", f"{settings.max_flow} m³/s"))
            params.append(("糙率", str(settings.roughness)))
            if getattr(settings, 'siphon_roughness', None) is not None:
                params.append(("倒虹吸糙率", str(settings.siphon_roughness)))
            params.append(("转弯半径", f"{settings.turn_radius} m"))
            params.append(("渡槽/隧洞渐变段(进口)", f"{settings.transition_inlet_form}(ζ={settings.transition_inlet_zeta:.2f})"))
            params.append(("渡槽/隧洞渐变段(出口)", f"{settings.transition_outlet_form}(ζ={settings.transition_outlet_zeta:.2f})"))
            params.append(("倒虹吸渐变段(进口)", f"{settings.siphon_transition_inlet_form}(ζ={settings.siphon_transition_inlet_zeta:.2f})"))
            params.append(("倒虹吸渐变段(出口)", f"{settings.siphon_transition_outlet_form}(ζ={settings.siphon_transition_outlet_zeta:.2f})"))
        params.append(("总节点数", str(len(nodes))))
        doc_add_param_table(doc, params)

        # 6. 详细计算过程
        doc_add_eng_h(doc, '6、详细计算过程')
        self._append_pressure_pipe_calc_details(getattr(self, "_pressure_pipe_calc_records", None))
        calc_text = self.detail_text.toPlainText()
        doc_render_calc_text_eng(doc, calc_text, skip_title_keyword='详细计算结果')

        # 7. 建筑物长度汇总
        if hasattr(self, '_last_building_lengths') and self._last_building_lengths:
            doc_add_eng_h(doc, '7、建筑物长度汇总')
            headers = ['序号', '名称', '结构形式', '长度(m)', '起始桩号', '终止桩号']
            data = []
            for i, bl in enumerate(self._last_building_lengths, 1):
                s_s = bl.get('start_station', 0.0)
                s_e = bl.get('end_station', 0.0)
                data.append([
                    str(i),
                    bl.get('name', '-'),
                    bl.get('structure_type', '-'),
                    f"{bl.get('length', 0.0):.3f}",
                    ProjectSettings.format_station(s_s, prefix),
                    ProjectSettings.format_station(s_e, prefix),
                ])
            total_length = sum(bl.get('length', 0.0) for bl in self._last_building_lengths)
            data.append(['合计', '', '', f"{total_length:.3f}", '', ''])
            doc_add_table_caption(doc, '表 1  建筑物长度汇总表')
            doc_add_styled_table(doc, headers, data, with_full_border=True)

        doc.save(filepath)

    # ================================================================
    # CAD 工具
    # ================================================================
    def _cad_combined_dxf(self):
        """导出全部DXF（纵断面+断面汇总+IP表合并）"""
        try:
            from 渠系断面设计.water_profile.cad_tools import export_combined_dxf
            export_combined_dxf(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from 渠系断面设计.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"合并DXF导出时发生错误:\n{e}")

    def _cad_longitudinal_profile(self):
        """生成纵断面表格（DXF 文件，也支持 TXT）"""
        try:
            from 渠系断面设计.water_profile.cad_tools import export_longitudinal_profile_dxf
            export_longitudinal_profile_dxf(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from 渠系断面设计.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"生成纵断面表格时发生错误:\n{e}")

    def _cad_section_summary(self):
        """生成断面汇总表"""
        try:
            from 渠系断面设计.water_profile.cad_tools import open_section_summary_table
            open_section_summary_table(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from 渠系断面设计.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"生成断面汇总表时发生错误:\n{e}")

    def _cad_bzzh2(self):
        """生成bzzh2命令内容（ZDM用）"""
        try:
            from 渠系断面设计.water_profile.cad_tools import extract_bzzh2_data
            extract_bzzh2_data(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from 渠系断面设计.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"生成bzzh2命令时发生错误:\n{e}")

    def _cad_building_plan(self):
        """建筑物名称上平面图（AutoCAD -TEXT 命令）"""
        try:
            from 渠系断面设计.water_profile.cad_tools import export_building_name_plan
            export_building_name_plan(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from 渠系断面设计.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"生成建筑物平面图时发生错误:\n{e}")

    def _cad_ip_table(self):
        """IP坐标及弯道参数表导出DXF/Excel"""
        try:
            from 渠系断面设计.water_profile.cad_tools import export_ip_plan_table
            export_ip_plan_table(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from 渠系断面设计.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"生成IP坐标表时发生错误:\n{e}")

    def resizeEvent(self, event):
        super().resizeEvent(event)
        auto_resize_table(self.node_table)

    # ================================================================
    # 项目文件序列化/反序列化（用于 .qxproj 项目保存功能）
    # ================================================================
    def to_project_dict(self) -> dict:
        """
        将水面线推求面板数据序列化为字典
        
        用于 .qxproj 项目文件保存，包含所有设置、节点数据和计算结果。
        
        Returns:
            包含所有数据的字典
        """
        from 推求水面线.models.data_models import ChannelNode, ProjectSettings
        
        # 收集 UI 设置（控件当前值）
        ui_settings = {
            "channel_name": self.channel_name_edit.text().strip(),
            "channel_level": self.channel_level_combo.currentText(),
            "start_water_level": self.start_wl_edit.text().strip(),
            "design_flows_text": self.design_flow_edit.text().strip(),
            "max_flows_text": self.max_flow_edit.text().strip(),
            "start_station_text": self.start_station_edit.text().strip(),
            "roughness": self.roughness_edit.text().strip(),
            "turn_radius": self.turn_radius_edit.text().strip(),
            # 渐变段设置
            "trans_inlet_form": self.trans_inlet_combo.currentText(),
            "trans_inlet_zeta": self.trans_inlet_zeta.text().strip(),
            "trans_outlet_form": self.trans_outlet_combo.currentText(),
            "trans_outlet_zeta": self.trans_outlet_zeta.text().strip(),
            "oc_trans_form": self.oc_trans_combo.currentText(),
            "oc_trans_zeta": self.oc_trans_zeta.text().strip(),
            "siphon_inlet_form": self.siphon_inlet_combo.currentText(),
            "siphon_inlet_zeta": self.siphon_inlet_zeta.text().strip(),
            "siphon_outlet_form": self.siphon_outlet_combo.currentText(),
            "siphon_outlet_zeta": self.siphon_outlet_zeta.text().strip(),
        }
        
        # 序列化 ProjectSettings
        project_settings = {}
        if self._settings:
            project_settings = self._settings.to_dict()
        
        # 序列化节点列表
        nodes_data = []
        for node in self.nodes:
            if hasattr(node, 'to_project_dict'):
                nodes_data.append(node.to_project_dict())
        
        # 序列化计算结果节点列表
        calculated_nodes_data = []
        for node in self.calculated_nodes:
            if hasattr(node, 'to_project_dict'):
                calculated_nodes_data.append(node.to_project_dict())
        
        # 收集额外缓存数据（key 转为字符串，JSON 要求）
        _pp_cache = getattr(self, "_custom_pressurized_pipe_params", {}) or {}
        _pp_siphon = [
            [str(row[0]), str(row[1]), row[2]]
            for row in _pp_cache.get("siphon", [])
            if isinstance(row, (tuple, list)) and len(row) >= 3
        ]
        _pp_pressure = [
            [str(row[0]), str(row[1]), row[2]]
            for row in _pp_cache.get("pressure_pipe", [])
            if isinstance(row, (tuple, list)) and len(row) >= 3
        ]
        extra_caches = {
            "node_structure_heights": {str(k): v for k, v in self._node_structure_heights.items()},
            "node_chamfer_params": {str(k): v for k, v in self._node_chamfer_params.items()},
            "node_u_params": {str(k): v for k, v in self._node_u_params.items()},
            "node_velocity_increased": {str(k): v for k, v in self._node_velocity_increased.items()},
            "text_export_settings": self._text_export_settings.copy(),
            "plan_text_settings": self._plan_text_settings.copy(),
            "custom_pressurized_pipe_params": {
                "siphon": _pp_siphon,
                "pressure_pipe": _pp_pressure,
            },
            "custom_struct_thickness": dict(getattr(self, "_custom_struct_thickness", {}) or {}),
            "custom_rock_lining": dict(getattr(self, "_custom_rock_lining", {}) or {}),
            "custom_tunnel_unified": dict(getattr(self, "_custom_tunnel_unified", {}) or {}),
        }
        
        # 收集倒虹吸糙率数据
        siphon_roughness_data = []
        if hasattr(self.siphon_roughness_chips, '_pairs'):
            siphon_roughness_data = list(self.siphon_roughness_chips._pairs)
        pressure_pipe_roughness_data = []
        if hasattr(self, 'pressure_pipe_roughness_chips') and hasattr(self.pressure_pipe_roughness_chips, '_pairs'):
            pressure_pipe_roughness_data = list(self.pressure_pipe_roughness_chips._pairs)
        
        return {
            "version": "1.0",
            "ui_settings": ui_settings,
            "project_settings": project_settings,
            "nodes": nodes_data,
            "calculated_nodes": calculated_nodes_data,
            "extra_caches": extra_caches,
            "siphon_roughness_data": siphon_roughness_data,
            "pressure_pipe_roughness_data": pressure_pipe_roughness_data,
            "pressure_pipe_calc_records": normalize_pressure_pipe_calc_records(
                getattr(self, "_pressure_pipe_calc_records", None)
            ),
        }
    
    def from_project_dict(self, d: dict, skip_dirty_signal: bool = False):
        """
        从字典恢复水面线推求面板数据
        
        用于 .qxproj 项目文件加载。
        
        Args:
            d: 序列化的字典数据
            skip_dirty_signal: 是否跳过脏状态信号（加载时应为True）
        """
        from 推求水面线.models.data_models import ChannelNode, ProjectSettings
        
        # 设置守卫标志，防止加载时触发脏状态和 cellChanged
        old_updating = self._updating_cells
        old_loading = getattr(self, '_loading_project', False)
        self._updating_cells = True
        self._loading_project = True
        
        try:
            # 默认重置有压管道计算记录（兼容旧项目缺失字段）
            self._pressure_pipe_calc_records = empty_pressure_pipe_calc_records()
            self._pressure_pipe_last_run_at = ""

            # 恢复 UI 设置
            ui = d.get("ui_settings", {})
            
            if ui.get("channel_name"):
                self.channel_name_edit.setText(ui["channel_name"])
            
            if ui.get("channel_level"):
                idx = self.channel_level_combo.findText(ui["channel_level"])
                if idx >= 0:
                    self.channel_level_combo.setCurrentIndex(idx)
            
            if ui.get("start_water_level"):
                self.start_wl_edit.setText(ui["start_water_level"])
            
            if ui.get("design_flows_text"):
                self.design_flow_edit.setText(ui["design_flows_text"])
            
            if ui.get("max_flows_text"):
                self.max_flow_edit.setText(ui["max_flows_text"])
            
            if ui.get("start_station_text"):
                self.start_station_edit.setText(ui["start_station_text"])
            
            if ui.get("roughness"):
                self.roughness_edit.setText(ui["roughness"])
            
            if ui.get("turn_radius"):
                self.turn_radius_edit.setText(ui["turn_radius"])
            
            # 恢复渐变段设置
            if ui.get("trans_inlet_form"):
                idx = self.trans_inlet_combo.findText(ui["trans_inlet_form"])
                if idx >= 0:
                    self.trans_inlet_combo.setCurrentIndex(idx)
            if ui.get("trans_inlet_zeta"):
                self.trans_inlet_zeta.setText(ui["trans_inlet_zeta"])
            
            if ui.get("trans_outlet_form"):
                idx = self.trans_outlet_combo.findText(ui["trans_outlet_form"])
                if idx >= 0:
                    self.trans_outlet_combo.setCurrentIndex(idx)
            if ui.get("trans_outlet_zeta"):
                self.trans_outlet_zeta.setText(ui["trans_outlet_zeta"])
            
            if ui.get("oc_trans_form"):
                idx = self.oc_trans_combo.findText(ui["oc_trans_form"])
                if idx >= 0:
                    self.oc_trans_combo.setCurrentIndex(idx)
            if ui.get("oc_trans_zeta"):
                self.oc_trans_zeta.setText(ui["oc_trans_zeta"])
            
            if ui.get("siphon_inlet_form"):
                idx = self.siphon_inlet_combo.findText(ui["siphon_inlet_form"])
                if idx >= 0:
                    self.siphon_inlet_combo.setCurrentIndex(idx)
            if ui.get("siphon_inlet_zeta"):
                self.siphon_inlet_zeta.setText(ui["siphon_inlet_zeta"])
            
            if ui.get("siphon_outlet_form"):
                idx = self.siphon_outlet_combo.findText(ui["siphon_outlet_form"])
                if idx >= 0:
                    self.siphon_outlet_combo.setCurrentIndex(idx)
            if ui.get("siphon_outlet_zeta"):
                self.siphon_outlet_zeta.setText(ui["siphon_outlet_zeta"])
            
            # 恢复 ProjectSettings
            proj_settings = d.get("project_settings", {})
            if proj_settings:
                self._settings = ProjectSettings.from_dict(proj_settings)
            
            # 恢复节点列表
            nodes_data = d.get("nodes", [])
            self.nodes = []
            for nd in nodes_data:
                self.nodes.append(ChannelNode.from_project_dict(nd))
            
            # 恢复计算结果节点列表
            calc_nodes_data = d.get("calculated_nodes", [])
            self.calculated_nodes = []
            for nd in calc_nodes_data:
                self.calculated_nodes.append(ChannelNode.from_project_dict(nd))
            
            # 恢复额外缓存数据（key 从字符串转回 int）
            extra = d.get("extra_caches", {})
            
            struct_heights = extra.get("node_structure_heights", {})
            self._node_structure_heights = {int(k): v for k, v in struct_heights.items()}
            
            chamfer_params = extra.get("node_chamfer_params", {})
            self._node_chamfer_params = {int(k): v for k, v in chamfer_params.items()}
            
            u_params = extra.get("node_u_params", {})
            self._node_u_params = {int(k): v for k, v in u_params.items()}

            vi_params = extra.get("node_velocity_increased", {})
            self._node_velocity_increased = {int(k): v for k, v in vi_params.items()}

            text_settings = extra.get("text_export_settings", {})
            if text_settings:
                self._text_export_settings.update(text_settings)
            
            plan_settings = extra.get("plan_text_settings", {})
            if plan_settings:
                self._plan_text_settings.update(plan_settings)

            # 恢复 CAD 导出复用参数缓存
            pp_cache = extra.get("custom_pressurized_pipe_params", {})
            if isinstance(pp_cache, dict):
                def _restore_pressurized_rows(rows):
                    restored = []
                    for row in rows or []:
                        if not isinstance(row, (tuple, list)) or len(row) < 3:
                            continue
                        restored.append((str(row[0]), str(row[1]), row[2]))
                    return restored

                self._custom_pressurized_pipe_params = {
                    "siphon": _restore_pressurized_rows(pp_cache.get("siphon", [])),
                    "pressure_pipe": _restore_pressurized_rows(pp_cache.get("pressure_pipe", [])),
                }
            else:
                self._custom_pressurized_pipe_params = {"siphon": [], "pressure_pipe": []}

            struct_t = extra.get("custom_struct_thickness", {})
            self._custom_struct_thickness = dict(struct_t) if isinstance(struct_t, dict) else {}
            rock_lining = extra.get("custom_rock_lining", {})
            self._custom_rock_lining = dict(rock_lining) if isinstance(rock_lining, dict) else {}
            tunnel_unified = extra.get("custom_tunnel_unified", {})
            self._custom_tunnel_unified = dict(tunnel_unified) if isinstance(tunnel_unified, dict) else {}
            
            # 恢复倒虹吸糙率数据
            siphon_data = d.get("siphon_roughness_data", [])
            if siphon_data and hasattr(self.siphon_roughness_chips, 'set_siphon_data'):
                self.siphon_roughness_chips.set_siphon_data(siphon_data)
            ppipe_data = d.get("pressure_pipe_roughness_data", [])
            if ppipe_data and hasattr(self, 'pressure_pipe_roughness_chips'):
                self.pressure_pipe_roughness_chips.set_pairs(ppipe_data)

            # 恢复有压管道计算记录
            self._pressure_pipe_calc_records = normalize_pressure_pipe_calc_records(
                d.get("pressure_pipe_calc_records", None)
            )
            self._pressure_pipe_last_run_at = self._pressure_pipe_calc_records.get("last_run_at", "")
            
            # 刷新节点表格显示
            if self.calculated_nodes:
                # 有计算结果，显示计算后的数据
                self._update_table_from_nodes_full(self.calculated_nodes)
            elif self.nodes:
                # 无计算结果，显示原始节点数据
                self._update_table_from_nodes_full(self.nodes)
            else:
                # 清空表格
                self.node_table.setRowCount(0)
                self._refresh_pressure_pipe_controls()
            
        finally:
            self._updating_cells = old_updating
            self._loading_project = old_loading
