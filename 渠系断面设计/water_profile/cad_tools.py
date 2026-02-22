# -*- coding: utf-8 -*-
"""
水面线面板 CAD 工具集

移植自原版 TK 的工程辅助功能，包括：
1. 生成纵断面表格（AutoCAD pl + -text 命令）
2. 生成bzzh2命令内容（ZDM用）
3. 建筑物名称上平面图（AutoCAD -TEXT 命令）
4. IP坐标及弯道参数表导出Excel
5. 断面汇总表
"""

import os
import sys
import math

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QGroupBox, QTextEdit,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QFileDialog, QApplication, QScrollArea, QWidget, QComboBox, QFrame,
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QShortcut, QKeySequence

from qfluentwidgets import PushButton, PrimaryPushButton, LineEdit

from 渠系断面设计.styles import (
    auto_resize_table, DIALOG_STYLE,
    fluent_info, fluent_error, fluent_question,
)

# 确保推求水面线模块可用
_pkg_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_water_profile_dir = os.path.join(_pkg_root, '推求水面线')
if _water_profile_dir not in sys.path:
    sys.path.insert(0, _water_profile_dir)
if _pkg_root not in sys.path:
    sys.path.insert(0, _pkg_root)

try:
    from models.data_models import ProjectSettings
    from models.enums import StructureType, InOutType
    MODELS_AVAILABLE = True
except ImportError:
    MODELS_AVAILABLE = False


# ================================================================
# 辅助工具函数
# ================================================================

def _format_number(value):
    """格式化数值：保留完整精度，去除无意义的尾零"""
    return f"{value:.15g}"


def _get_building_display_name(node):
    """获取纵断面用的建筑物名称显示"""
    struct_str = node.get_structure_type_str() or ""
    if node.is_transition or struct_str == "渐变段":
        return ""
    if struct_str.startswith("明渠"):
        return struct_str
    if struct_str == "矩形暗涵":
        return struct_str
    if node.name:
        category = struct_str.split("-")[0]
        return f"{node.name}{category}"
    return struct_str.split("-")[0] if struct_str else ""


def _estimate_text_width(text, text_height):
    """估算 AutoCAD 中文字的总宽度（用于居中对齐）

    中文字符（CJK）宽度 ≈ text_height
    ASCII 字符（字母/数字/标点）宽度 ≈ text_height × 0.7
    """
    width = 0.0
    for ch in text:
        if ord(ch) > 127:
            width += text_height
        else:
            width += text_height * 0.7
    return width


def _format_slope_text(slope_i):
    """格式化坡降为显示文本"""
    if slope_i is not None and slope_i > 0:
        slope_inv = round(1.0 / slope_i)
        return f"1/{slope_inv}"
    return "/"


def _get_node_slope_text(node, next_node=None):
    """获取节点坡降文本（直接使用节点自身的 slope_i）。"""
    return _format_slope_text(getattr(node, 'slope_i', None))


def _struct_val(struct_type):
    """获取 StructureType 的字符串值（兼容双路径导入的 enum 实例）"""
    if struct_type is None:
        return ""
    return struct_type.value if hasattr(struct_type, 'value') else str(struct_type)


def _in_out_val(in_out):
    """获取 InOutType 的字符串值（兼容双路径导入的 enum 实例）"""
    if in_out is None:
        return ""
    return in_out.value if hasattr(in_out, 'value') else str(in_out)


def _is_special_structure_sv(struct_type):
    """判断是否为特殊建筑物（隧洞/倒虹吸/渡槽），使用字符串值比较

    注意：矩形暗涵不需要进出口标识，不属于特殊建筑物。
    避免双路径导入导致 enum 实例比较失败"""
    sv = _struct_val(struct_type)
    return any(k in sv for k in ("隧洞", "倒虹吸", "渡槽"))


def _is_gate_name(name):
    """判断建筑物显示名称是否为闸类点状建筑物（分水闸/分水口/节制闸/泄水闸等）"""
    if not name:
        return False
    return "闸" in name or "分水" in name


def _get_segment_slope_text(mc_list, mc_to_node):
    """从建筑物段的节点列表中提取坡降文本

    遍历段内所有 MC 对应的节点，取第一个有效的 slope_i 作为坡降。
    Args:
        mc_list: 该建筑物段包含的桩号列表
        mc_to_node: {station_MC: node} 查找表
    Returns:
        坡降文本（如 "1/3000"），无数据时返回 None
    """
    for mc in mc_list:
        node = mc_to_node.get(mc)
        if node:
            st = _format_slope_text(getattr(node, 'slope_i', None))
            if st != "/":
                return st
    return None


def _merge_segments_across_gates(segments, gate_mc_set=None):
    """合并被闸（点状建筑物）拆分的同名段落

    规则：如果段落 i 和段落 j 的值相同，且 i~j 之间的所有段落
    都是闸类点状建筑物，则将 j 的 MC 列表合并到 i，闸段保留不变。

    对建筑物名称段落：通过名称判断闸（_is_gate_name）。
    对坡降等段落：通过 gate_mc_set（闸节点的桩号集合）判断。

    Args:
        segments: [(value, [mc_list]), ...] 按位置排列的段落
        gate_mc_set: 闸节点桩号集合，仅在非名称场景下使用
    """
    if len(segments) <= 2:
        return segments

    def _is_gate_seg(val, mc_list):
        if gate_mc_set is not None:
            return all(mc in gate_mc_set for mc in mc_list)
        return _is_gate_name(val)

    merged = []
    i = 0
    while i < len(segments):
        val, mc_list = segments[i]

        if _is_gate_seg(val, mc_list):
            merged.append((val, list(mc_list)))
            i += 1
            continue

        # 非闸段：尝试向后合并同名段（跳过中间的闸段）
        mc_list = list(mc_list)
        j = i + 1
        while j + 1 < len(segments):
            mid_val, mid_mcs = segments[j]
            next_val, next_mcs = segments[j + 1]
            if _is_gate_seg(mid_val, mid_mcs) and next_val == val:
                merged.append((mid_val, list(mid_mcs)))  # 闸段保留
                mc_list.extend(next_mcs)
                j += 2
            else:
                break

        merged.append((val, mc_list))
        i = j if j > i + 1 else i + 1

    return merged


# ================================================================
# DXF 共享辅助工具
# ================================================================

class _OffsetMSP:
    """包装 ezdxf modelspace，自动为所有绘图操作添加坐标偏移。
    用于在同一 DXF 文件中将多个表格放置在不同位置。"""

    def __init__(self, msp, ox=0, oy=0):
        self._msp = msp
        self._ox = ox
        self._oy = oy

    def _p(self, pt):
        return (pt[0] + self._ox, pt[1] + self._oy)

    def add_line(self, start, end, dxfattribs=None):
        return self._msp.add_line(self._p(start), self._p(end),
                                   dxfattribs=dxfattribs or {})

    def add_lwpolyline(self, points, dxfattribs=None):
        return self._msp.add_lwpolyline(
            [self._p(p) for p in points], dxfattribs=dxfattribs or {})

    def add_text(self, text, dxfattribs=None):
        entity = self._msp.add_text(text, dxfattribs=dxfattribs or {})
        return _OffsetTextEntity(entity, self._ox, self._oy)


class _OffsetTextEntity:
    """包装 ezdxf text 实体，自动为 set_placement 添加坐标偏移。"""

    def __init__(self, entity, ox, oy):
        self._entity = entity
        self._ox = ox
        self._oy = oy

    def set_placement(self, point, align=None):
        p = (point[0] + self._ox, point[1] + self._oy)
        if align is not None:
            return self._entity.set_placement(p, align=align)
        return self._entity.set_placement(p)


def _setup_dxf_style(doc):
    """设置 DXF 文档的中文字体样式（仿宋，宽度因子0.7）。"""
    if "Standard" in doc.styles:
        _sty = doc.styles.get("Standard")
    else:
        _sty = doc.styles.add("Standard")
    _sty.dxf.font = ""
    _sty.dxf.width = 0.7
    try:
        if "ACAD" not in doc.appids:
            doc.appids.new("ACAD")
    except Exception:
        pass
    _sty.set_xdata("ACAD", [(1000, "仿宋"), (1071, 0)])


def _ensure_profile_layers(doc, layer_prefix=""):
    """确保纵断面所需的图层存在。layer_prefix 用于合并导出时区分组件。"""
    layer_defs = [
        ("表格线框", 7), ("渠底高程线", 3), ("渠顶高程线", 1),
        ("设计水位线", 5), ("文字标注", 7),
    ]
    for name, color in layer_defs:
        full = layer_prefix + name
        if full not in doc.layers:
            doc.layers.new(full, dxfattribs={"color": color})


def _compute_ip_preview_data(nodes, station_prefix):
    """从节点列表计算IP坐标及弯道参数表预览数据。
    返回 (preview_data, real_nodes)。"""
    real_nodes = [
        n for n in nodes
        if not getattr(n, 'is_transition', False)
        and not getattr(n, 'is_auto_inserted_channel', False)
    ]

    def _safe_float(val, default=0.0):
        if val is None:
            return default
        try:
            return float(val)
        except (TypeError, ValueError):
            return default

    def _format_ip_name(node):
        try:
            if _in_out_val(node.in_out) in ("进", "出"):
                struct_abbr = ""
                struct_str = _struct_val(node.structure_type)
                if struct_str:
                    if "隧洞" in struct_str: struct_abbr = "隧"
                    elif "倒虹吸" in struct_str: struct_abbr = "倒"
                    elif "渡槽" in struct_str: struct_abbr = "渡"
                    elif "暗涵" in struct_str: struct_abbr = "暗"
                in_out_str = "进" if _in_out_val(node.in_out) == "进" else "出"
                return f"{node.name}{struct_abbr}{in_out_str}"
        except Exception:
            pass
        return f"{station_prefix}IP{getattr(node, 'ip_number', 0)}"

    def _format_station(value):
        return ProjectSettings.format_station(_safe_float(value), station_prefix)

    preview_data = []
    for idx, node in enumerate(real_nodes):
        try:
            row = [
                _format_ip_name(node),
                f"{_safe_float(node.x):.6f}",
                f"{_safe_float(node.y):.6f}",
                _format_station(node.station_BC),
                _format_station(node.station_MC),
                _format_station(node.station_EC),
                f"{_safe_float(node.turn_angle):.3f}",
                f"{_safe_float(node.turn_radius):.3f}",
                f"{_safe_float(node.tangent_length):.3f}",
                f"{_safe_float(node.arc_length):.3f}",
                f"{_safe_float(node.bottom_elevation):.3f}" if _safe_float(node.bottom_elevation) != 0 else "-",
            ]
            preview_data.append(row)
        except Exception:
            preview_data.append([
                f"IP{getattr(node, 'ip_number', '?')}",
                "0.000000", "0.000000",
                "0+000.000", "0+000.000", "0+000.000",
                "0.000", "0.000", "0.000", "0.000", "-",
            ])
    return preview_data, real_nodes


# ================================================================
# 纵断面表格导出设置对话框
# ================================================================

class TextExportSettingsDialog(QDialog):
    """纵断面文字导出参数配置弹窗（QFluentWidgets 风格）"""

    def __init__(self, parent=None, defaults=None):
        super().__init__(parent)
        self.setWindowTitle("纵断面文字导出设置")
        self.setMinimumWidth(420)
        self.setStyleSheet(DIALOG_STYLE)
        self.result = None

        if defaults is None:
            defaults = {}
        self._defaults = {
            'y_bottom': defaults.get('y_bottom', 1),
            'y_top': defaults.get('y_top', 31),
            'y_water': defaults.get('y_water', 16),
            'text_height': defaults.get('text_height', 3.5),
            'rotation': defaults.get('rotation', 90),
            'elev_decimals': defaults.get('elev_decimals', 3),
            'y_name': defaults.get('y_name', 115),
            'y_slope': defaults.get('y_slope', 105),
            'y_ip': defaults.get('y_ip', 77),
            'y_station': defaults.get('y_station', 47),
            'y_line_height': defaults.get('y_line_height', 120),
            'scale_x': defaults.get('scale_x', 1),
            'scale_y': defaults.get('scale_y', 1),
        }

        self._entries = {}
        self._init_ui()

    def _init_ui(self):
        lay = QVBoxLayout(self)

        # Y坐标设置
        y_grp = QGroupBox("Y 坐标设置（CAD 表格行高）")
        y_form = QGridLayout(y_grp)
        for row, (label, key) in enumerate([
            ("渠底文字 Y 坐标:", 'y_bottom'),
            ("渠顶文字 Y 坐标:", 'y_top'),
            ("水面文字 Y 坐标:", 'y_water'),
        ]):
            y_form.addWidget(QLabel(label), row, 0)
            e = LineEdit(); e.setText(str(self._defaults[key])); e.setFixedWidth(100)
            y_form.addWidget(e, row, 1)
            self._entries[key] = e
        lay.addWidget(y_grp)

        # 文字样式
        style_grp = QGroupBox("文字样式")
        style_form = QGridLayout(style_grp)
        for row, (label, key) in enumerate([
            ("字高:", 'text_height'),
            ("旋转角度:", 'rotation'),
            ("高程小数位数:", 'elev_decimals'),
        ]):
            style_form.addWidget(QLabel(label), row, 0)
            e = LineEdit(); e.setText(str(self._defaults[key])); e.setFixedWidth(100)
            style_form.addWidget(e, row, 1)
            self._entries[key] = e
        lay.addWidget(style_grp)

        # 纵断面信息列
        info_grp = QGroupBox("纵断面信息列 Y 坐标")
        info_form = QGridLayout(info_grp)
        for row, (label, key) in enumerate([
            ("建筑物名称 Y 坐标:", 'y_name'),
            ("坡降 Y 坐标:", 'y_slope'),
            ("IP点名称 Y 坐标:", 'y_ip'),
            ("里程桩号 Y 坐标:", 'y_station'),
            ("整线竖线高度:", 'y_line_height'),
        ]):
            info_form.addWidget(QLabel(label), row, 0)
            e = LineEdit(); e.setText(str(self._defaults[key])); e.setFixedWidth(100)
            info_form.addWidget(e, row, 1)
            self._entries[key] = e
        lay.addWidget(info_grp)

        # 比例设置
        scale_grp = QGroupBox("比例设置")
        scale_form = QGridLayout(scale_grp)
        scale_form.addWidget(QLabel("X 方向 (1:N)，N ="), 0, 0)
        e = LineEdit(); e.setText(str(self._defaults['scale_x'])); e.setFixedWidth(100)
        scale_form.addWidget(e, 0, 1)
        scale_form.addWidget(QLabel("如 1:1000 则输入 1000"), 0, 2)
        self._entries['scale_x'] = e
        scale_form.addWidget(QLabel("Y 方向 (1:N)，N ="), 1, 0)
        e = LineEdit(); e.setText(str(self._defaults['scale_y'])); e.setFixedWidth(100)
        scale_form.addWidget(e, 1, 1)
        scale_form.addWidget(QLabel("如 1:100 则输入 100"), 1, 2)
        self._entries['scale_y'] = e
        lay.addWidget(scale_grp)

        # 预览
        preview_grp = QGroupBox("命令格式预览")
        preview_lay = QVBoxLayout(preview_grp)
        self._preview_label = QLabel()
        self._preview_label.setStyleSheet("color: #336699;")
        self._preview_label.setFont(QFont("Consolas", 9))
        preview_lay.addWidget(self._preview_label)
        lay.addWidget(preview_grp)
        self._update_preview()
        for entry in self._entries.values():
            entry.textChanged.connect(self._update_preview)

        # 按钮
        btn_lay = QHBoxLayout()
        btn_reset = PushButton("恢复默认"); btn_reset.clicked.connect(self._reset_defaults)
        btn_lay.addWidget(btn_reset); btn_lay.addStretch()
        btn_cancel = PushButton("取消"); btn_cancel.clicked.connect(self.reject)
        btn_ok = PrimaryPushButton("确定"); btn_ok.clicked.connect(self._on_confirm)
        btn_lay.addWidget(btn_cancel); btn_lay.addWidget(btn_ok)
        lay.addLayout(btn_lay)

        # 键盘快捷键
        QShortcut(QKeySequence(Qt.Key_Escape), self, self.reject)
        QShortcut(QKeySequence(Qt.Key_Return), self, self._on_confirm)

    def _update_preview(self):
        try:
            y_b = self._entries['y_bottom'].text().strip()
            h = self._entries['text_height'].text().strip()
            r = self._entries['rotation'].text().strip()
            d = self._entries['elev_decimals'].text().strip()
            try:
                decimals = int(float(d))
                sample = f"{431.666:.{decimals}f}"
            except Exception:
                sample = "431.666"
            self._preview_label.setText(f"-text 里程MC,{y_b} {h} {r} {sample} ")
        except Exception:
            self._preview_label.setText("-text 里程MC,Y 字高 角度 高程 ")

    def _reset_defaults(self):
        original = {
            'y_bottom': 1, 'y_top': 31, 'y_water': 16,
            'text_height': 3.5, 'rotation': 90, 'elev_decimals': 3,
            'y_name': 115, 'y_slope': 105, 'y_ip': 77,
            'y_station': 47, 'y_line_height': 120,
            'scale_x': 1, 'scale_y': 1,
        }
        for key, value in original.items():
            self._entries[key].setText(str(value))
        self._update_preview()

    def _on_confirm(self):
        try:
            result = {}
            for key, entry in self._entries.items():
                val_str = entry.text().strip()
                if not val_str:
                    raise ValueError("参数不能为空")
                val = float(val_str)
                if key == 'elev_decimals':
                    if val < 0 or val != int(val):
                        raise ValueError("高程小数位数必须为非负整数")
                    val = int(val)
                if key in ('scale_x', 'scale_y'):
                    if val <= 0:
                        raise ValueError("比例尺必须大于0")
                result[key] = val
            self.result = result
            self.accept()
        except ValueError as e:
            fluent_error(self, "输入错误", f"请输入有效的数值:\n{str(e)}")


# ================================================================
# 平面图参数设置对话框
# ================================================================

class PlanTextSettingsDialog(QDialog):
    """建筑物名称上平面图参数设置对话框"""

    def __init__(self, parent=None, defaults=None):
        super().__init__(parent)
        self.setWindowTitle("建筑物名称上平面图 - 参数设置")
        self.setMinimumWidth(380)
        self.setStyleSheet(DIALOG_STYLE)
        self.result = None
        if defaults is None:
            defaults = {}
        self._init_ui(defaults)

    def _init_ui(self, defaults):
        lay = QVBoxLayout(self)
        lay.addWidget(QLabel(
            "生成 AutoCAD -TEXT 命令，将建筑物名称平行于轴线放置。\n"
            "文字位于建筑物最中间两个IP点连线段的中点处。"
        ))

        form = QGridLayout()
        form.addWidget(QLabel("垂直偏移距离 (V):"), 0, 0)
        self.offset_edit = LineEdit(); self.offset_edit.setText(str(defaults.get('offset', 10)))
        self.offset_edit.setFixedWidth(100)
        form.addWidget(self.offset_edit, 0, 1)
        form.addWidget(QLabel("文字中心到轴线的距离"), 0, 2)

        form.addWidget(QLabel("文字高度:"), 1, 0)
        self.height_edit = LineEdit(); self.height_edit.setText(str(defaults.get('text_height', 10)))
        self.height_edit.setFixedWidth(100)
        form.addWidget(self.height_edit, 1, 1)
        form.addWidget(QLabel("AutoCAD -TEXT 字高"), 1, 2)
        lay.addLayout(form)

        # 预览
        preview_grp = QGroupBox("命令格式预览")
        preview_lay = QVBoxLayout(preview_grp)
        self._preview_label = QLabel()
        self._preview_label.setStyleSheet("color: gray;")
        preview_lay.addWidget(self._preview_label)
        lay.addWidget(preview_grp)
        self._update_preview()
        self.offset_edit.textChanged.connect(self._update_preview)
        self.height_edit.textChanged.connect(self._update_preview)

        btn_lay = QHBoxLayout()
        btn_lay.addStretch()
        btn_cancel = PushButton("取消"); btn_cancel.clicked.connect(self.reject)
        btn_ok = PrimaryPushButton("确定"); btn_ok.clicked.connect(self._on_confirm)
        btn_lay.addWidget(btn_cancel); btn_lay.addWidget(btn_ok)
        lay.addLayout(btn_lay)

        # 键盘快捷键
        QShortcut(QKeySequence(Qt.Key_Escape), self, self.reject)
        QShortcut(QKeySequence(Qt.Key_Return), self, self._on_confirm)

    def _update_preview(self):
        try:
            o = self.offset_edit.text().strip()
            h = self.height_edit.text().strip()
            self._preview_label.setText(
                f"-TEXT J MC x,y {h} 角度 建筑物名称\n"
                f"（文字中心偏移轴线 {o} 个单位）")
        except Exception:
            pass

    def _on_confirm(self):
        try:
            o = float(self.offset_edit.text().strip())
            h = float(self.height_edit.text().strip())
            if h <= 0:
                raise ValueError("文字高度必须大于0")
            self.result = {'offset': o, 'text_height': h}
            self.accept()
        except ValueError as e:
            fluent_error(self, "输入错误", f"请输入有效的数值:\n{e}")


# ================================================================
# 1. 生成纵断面表格 TXT
# ================================================================

def export_longitudinal_profile_txt(panel):
    """一键生成上纵断面表格 TXT（AutoCAD pl + -text 命令格式）

    输出格式为逐行 AutoCAD 命令（可直接全选复制粘贴到 AutoCAD 命令行自动执行）。
    *** 所有 pl 命令在前，-text 命令在后（避免命令交错导致解析异常） ***
    """
    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "没有数据可导出")
        return

    valid_nodes = [n for n in nodes if n.bottom_elevation or n.top_elevation or n.water_level]
    if not valid_nodes:
        fluent_info(panel.window(), "警告", "没有可用的高程数据，请先执行计算。")
        return

    # 弹出参数配置对话框
    dlg = TextExportSettingsDialog(panel.window(), panel._text_export_settings)
    if dlg.exec() != QDialog.Accepted or dlg.result is None:
        return

    panel._text_export_settings.update(dlg.result)
    settings = dlg.result

    y_bottom = settings['y_bottom']
    y_top = settings['y_top']
    y_water = settings['y_water']
    text_height = settings['text_height']
    rotation = settings['rotation']
    elev_decimals = int(settings.get('elev_decimals', 3))
    y_name = settings.get('y_name', 112)
    y_slope = settings.get('y_slope', 102)
    y_ip = settings.get('y_ip', 77)
    y_station = settings.get('y_station', 47)
    y_line_height = settings.get('y_line_height', 120)

    # 自动文件名
    try:
        ch_name = panel.channel_name_edit.text().strip()
        ch_level = panel.channel_level_combo.currentText()
        auto_name = f"{ch_name}{ch_level}_上纵断面表格.txt"
    except Exception:
        auto_name = "上纵断面表格.txt"

    file_path, _ = QFileDialog.getSaveFileName(
        panel, "保存上纵断面表格", auto_name,
        "文本文件 (*.txt);;所有文件 (*.*)")
    if not file_path:
        return

    try:
        fmt = _format_number
        s_y_bottom = fmt(y_bottom)
        s_y_top = fmt(y_top)
        s_y_water = fmt(y_water)
        s_height = fmt(text_height)
        s_rotation = fmt(rotation)
        s_y_name = fmt(y_name)
        s_y_slope = fmt(y_slope)
        s_y_ip = fmt(y_ip)
        s_y_station = fmt(y_station)
        s_y_line_height = fmt(y_line_height)

        # 第一个节点的文字偏移（避免与左侧竖线重叠）
        first_col_x_offset = text_height + 1.3

        try:
            proj_settings = panel._build_settings()
            station_prefix = proj_settings.get_station_prefix()
        except Exception:
            station_prefix = ""

        def fmt_elev(value):
            if value is None:
                return f"{0:.{elev_decimals}f}"
            return f"{value:.{elev_decimals}f}"

        lines = []

        # ======== 所有 pl 命令 ========

        # 第一部分：表头区域线框
        h_line_y_values = [0, 15, 30, 45, 75, 100, 110, 120]
        for hy in h_line_y_values:
            lines.append(f"pl 0,{hy} -40,{hy} ")
        lines.append(f"pl -40,0 -40,{s_y_line_height} ")
        lines.append("")

        # 第二部分：整线竖线
        tall_line_mcs = []
        for idx, node in enumerate(nodes):
            mc = fmt(node.station_MC)
            is_special = False
            if idx == 0 or idx == len(nodes) - 1:
                is_special = True
            elif (_is_special_structure_sv(node.structure_type) and
                  _in_out_val(node.in_out) in ("进", "出")):
                is_special = True
            if is_special:
                tall_line_mcs.append(node.station_MC)
            height = s_y_line_height if is_special else "100"
            lines.append(f"pl {mc},0 {mc},{height} ")
        lines.append("")

        # 第三部分：水平线
        last_mc = fmt(nodes[-1].station_MC)
        for hy in h_line_y_values:
            lines.append(f"pl 0,{hy} {last_mc},{hy} ")
        lines.append("")

        # 第四部分：渠底/渠顶/水面 pl
        for node in valid_nodes:
            if node.bottom_elevation:
                lines.append(f"pl {fmt(node.station_MC)},{fmt(node.bottom_elevation)}")
        lines.append("")
        for node in valid_nodes:
            if node.top_elevation:
                lines.append(f"pl {fmt(node.station_MC)},{fmt(node.top_elevation)}")
        lines.append("")
        for node in valid_nodes:
            if node.water_level:
                lines.append(f"pl {fmt(node.station_MC)},{fmt(node.water_level)}")
        lines.append("")

        # ======== 所有 -text 命令 ========

        # 第五部分：渠底/渠顶/水面文字
        for idx, node in enumerate(nodes):
            mc = fmt(node.station_MC)
            text_x = fmt(node.station_MC + first_col_x_offset) if idx == 0 else mc
            lines.append(
                f"-text {text_x},{s_y_bottom} {s_height} {s_rotation} {fmt_elev(node.bottom_elevation)} ")
        lines.append("")
        for idx, node in enumerate(nodes):
            mc = fmt(node.station_MC)
            text_x = fmt(node.station_MC + first_col_x_offset) if idx == 0 else mc
            lines.append(
                f"-text {text_x},{s_y_top} {s_height} {s_rotation} {fmt_elev(node.top_elevation)} ")
        lines.append("")
        for idx, node in enumerate(nodes):
            mc = fmt(node.station_MC)
            text_x = fmt(node.station_MC + first_col_x_offset) if idx == 0 else mc
            lines.append(
                f"-text {text_x},{s_y_water} {s_height} {s_rotation} {fmt_elev(node.water_level)} ")
        lines.append("")

        # 第六部分：里程桩号
        for idx, node in enumerate(nodes):
            mc = fmt(node.station_MC)
            text_x = fmt(node.station_MC + first_col_x_offset) if idx == 0 else mc
            station_text = ProjectSettings.format_station(node.station_MC, station_prefix)
            lines.append(
                f"-text {text_x},{s_y_station} {s_height} {s_rotation} {station_text} ")
        lines.append("")

        # 第七部分：表头文字
        header_col_width = 40
        header_single = [
            ("建筑物名称", 110, 120),
            ("坡 降", 100, 110),
            ("IP点名称", 75, 100),
            ("渠顶高程(m)", 30, 45),
            ("设计水位(m)", 15, 30),
            ("渠底高程(m)", 0, 15),
        ]
        header_cx = fmt(-40 + header_col_width / 2)
        for label, row_bot, row_top in header_single:
            cy = (row_bot + row_top) / 2
            lines.append(f"-text j mc {header_cx},{fmt(cy)} {s_height} 0 {label} ")
        line_spacing = text_height * 2.5
        block_h = line_spacing + text_height
        y_line_bottom_val = 45 + (30 - block_h) / 2 + text_height / 2
        y_line_top_val = y_line_bottom_val + line_spacing
        for label, y_pos in [("里程桩号", y_line_top_val), ("（千米+米）", y_line_bottom_val)]:
            lines.append(f"-text j mc {header_cx},{fmt(y_pos)} {s_height} 0 {label} ")
        lines.append("")

        # 第八部分：建筑物名称
        name_mc_pairs = []
        for node in nodes:
            building_name = _get_building_display_name(node)
            if building_name:
                name_mc_pairs.append((building_name, node.station_MC))
        building_segments = []
        for bname, bmc in name_mc_pairs:
            if building_segments and building_segments[-1][0] == bname:
                building_segments[-1][1].append(bmc)
            else:
                building_segments.append((bname, [bmc]))
        building_segments = _merge_segments_across_gates(building_segments)
        for bname, mc_list in building_segments:
            if _is_gate_name(bname):
                # 闸类点状建筑物：直接使用桩号 MC 定位
                mid_mc = mc_list[0]
            else:
                seg_start = mc_list[0]
                seg_end = mc_list[-1]
                left_bound = max((m for m in tall_line_mcs if m <= seg_start), default=seg_start)
                right_bound = min((m for m in tall_line_mcs if m >= seg_end), default=seg_end)
                mid_mc = (left_bound + right_bound) / 2.0
            lines.append(f"-text j mc {fmt(mid_mc)},115 {s_height} 0 {bname} ")
        lines.append("")

        # 第九部分：坡降（从建筑物名称分段派生，与名称行一一对齐）
        mc_to_node = {node.station_MC: node for node in nodes}
        for bname, mc_list in building_segments:
            # 闸/分水口：不显示坡降
            if _is_gate_name(bname):
                continue
            # 倒虹吸为有压流，不显示坡降
            if "倒虹吸" in bname:
                continue
            slope_text = _get_segment_slope_text(mc_list, mc_to_node)
            if not slope_text:
                continue
            # 居中位置与上方建筑物名称完全对齐
            seg_start = mc_list[0]
            seg_end = mc_list[-1]
            left_bound = max((m for m in tall_line_mcs if m <= seg_start), default=seg_start)
            right_bound = min((m for m in tall_line_mcs if m >= seg_end), default=seg_end)
            mid_mc = (left_bound + right_bound) / 2.0
            lines.append(f"-text j mc {fmt(mid_mc)},105 {s_height} 0 {slope_text} ")
        lines.append("")

        # 第十部分：IP点名称
        for idx, node in enumerate(nodes):
            struct_str = node.get_structure_type_str() or ""
            if node.is_transition or struct_str == "渐变段":
                continue
            if _is_special_structure_sv(node.structure_type):
                if _in_out_val(node.in_out) in ("进", "出"):
                    struct_abbr = ""
                    st_val = _struct_val(node.structure_type)
                    if "隧洞" in st_val:
                        struct_abbr = "隧"
                    elif "倒虹吸" in st_val:
                        struct_abbr = "倒"
                    elif "渡槽" in st_val:
                        struct_abbr = "渡"
                    in_out_str = "进" if _in_out_val(node.in_out) == "进" else "出"
                    ip_name = f"{node.name}{struct_abbr}{in_out_str}"
                else:
                    continue
            else:
                ip_name = f"{station_prefix}IP{getattr(node, 'ip_number', 0)}"
            mc = fmt(node.station_MC)
            text_x = fmt(node.station_MC + first_col_x_offset) if node.station_MC == nodes[0].station_MC else mc
            lines.append(
                f"-text {text_x},77 {s_height} {s_rotation} {ip_name} ")
        lines.append("")

        # 写入文件
        with open(file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")

        if fluent_question(panel.window(), "完成",
                f"上纵断面表格已生成（{len(nodes)} 个节点）:\n{file_path}\n\n是否立即打开该文件？"):
            os.startfile(file_path)

    except PermissionError:
        fluent_error(panel.window(), "文件被占用",
                     f"无法写入文件，该文件可能已被其他程序打开：\n\n{file_path}\n\n"
                     f"请先关闭该文件（如记事本、Word等），然后重新操作。")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "导出错误", f"生成上纵断面表格失败:\n{str(e)}")


# ================================================================
# 1b. 纵断面表格核心绘图
# ================================================================

def _draw_profile_on_msp(msp, nodes, valid_nodes, settings, station_prefix,
                         layer_prefix=""):
    """在 modelspace 上绘制纵断面表格（核心绘图逻辑）。

    msp 可以是真实的 ezdxf modelspace 或 _OffsetMSP 包装器。
    layer_prefix 用于合并导出时给图层名添加前缀以区分组件。
    返回 (width, height)。
    """
    import ezdxf

    text_height = settings['text_height']
    rotation = settings['rotation']
    elev_decimals = int(settings.get('elev_decimals', 3))
    y_bottom = settings['y_bottom']
    y_top = settings['y_top']
    y_water = settings['y_water']
    y_ip = settings.get('y_ip', 77)
    y_station = settings.get('y_station', 47)
    y_line_height = settings.get('y_line_height', 120)
    scale_x = settings.get('scale_x', 1)
    scale_y = settings.get('scale_y', 1)

    first_col_x_offset = text_height + 1.3

    def sx(mc):
        return mc / scale_x

    def sy(elev):
        return elev / scale_y

    def fmt_elev(value):
        if value is None:
            return f"{0:.{elev_decimals}f}"
        return f"{value:.{elev_decimals}f}"

    last_mc = nodes[-1].station_MC
    layer_grid = layer_prefix + "表格线框"
    layer_text = layer_prefix + "文字标注"

    # ======== 1. 表头区域线框 ========
    h_line_y_values = [0, 15, 30, 45, 75, 100, 110, 120]
    for hy in h_line_y_values:
        msp.add_line((-40, hy), (sx(0), hy), dxfattribs={"layer": layer_grid})
    msp.add_line((-40, 0), (-40, y_line_height), dxfattribs={"layer": layer_grid})

    # ======== 2. 节点竖线 ========
    tall_line_mcs = []
    for idx, node in enumerate(nodes):
        mc = node.station_MC
        is_special = (idx == 0 or idx == len(nodes) - 1)
        if not is_special and (_is_special_structure_sv(node.structure_type) and
                               _in_out_val(node.in_out) in ("进", "出")):
            is_special = True
        if is_special:
            tall_line_mcs.append(mc)
        height = y_line_height if is_special else 100
        msp.add_line((sx(mc), 0), (sx(mc), height), dxfattribs={"layer": layer_grid})

    # ======== 3. 全宽水平线 ========
    for hy in h_line_y_values:
        msp.add_line((sx(0), hy), (sx(last_mc), hy), dxfattribs={"layer": layer_grid})

    # ======== 4. 渠底/渠顶/水面折线 ========
    bottom_pts = [(sx(n.station_MC), sy(n.bottom_elevation))
                  for n in valid_nodes if n.bottom_elevation]
    top_pts = [(sx(n.station_MC), sy(n.top_elevation))
               for n in valid_nodes if n.top_elevation]
    water_pts = [(sx(n.station_MC), sy(n.water_level))
                 for n in valid_nodes if n.water_level]

    if len(bottom_pts) >= 2:
        msp.add_lwpolyline(bottom_pts, dxfattribs={"layer": layer_prefix + "渠底高程线"})
    if len(top_pts) >= 2:
        msp.add_lwpolyline(top_pts, dxfattribs={"layer": layer_prefix + "渠顶高程线"})
    if len(water_pts) >= 2:
        msp.add_lwpolyline(water_pts, dxfattribs={"layer": layer_prefix + "设计水位线"})

    # ======== 5. 渠底/渠顶/水面高程文字 ========
    for idx, node in enumerate(nodes):
        text_x = sx(node.station_MC) + first_col_x_offset if idx == 0 else sx(node.station_MC) - 1
        for elev_val, y_pos in [
            (node.bottom_elevation, y_bottom),
            (node.top_elevation, y_top),
            (node.water_level, y_water),
        ]:
            msp.add_text(
                fmt_elev(elev_val),
                dxfattribs={"layer": layer_text, "height": text_height,
                            "rotation": rotation, "width": 0.7, "style": "Standard"}
            ).set_placement((text_x, y_pos))

    # ======== 6. 里程桩号 ========
    for idx, node in enumerate(nodes):
        text_x = sx(node.station_MC) + first_col_x_offset if idx == 0 else sx(node.station_MC) - 1
        station_text = ProjectSettings.format_station(node.station_MC, station_prefix)
        msp.add_text(
            station_text,
            dxfattribs={"layer": layer_text, "height": text_height,
                        "rotation": rotation, "width": 0.7, "style": "Standard"}
        ).set_placement((text_x, y_station))

    # ======== 7. 表头文字 ========
    header_col_width = 40
    header_single = [
        ("建筑物名称", 110, 120),
        ("坡 降", 100, 110),
        ("IP点名称", 75, 100),
        ("渠顶高程(m)", 30, 45),
        ("设计水位(m)", 15, 30),
        ("渠底高程(m)", 0, 15),
    ]
    header_cx = -40 + header_col_width / 2
    for label, row_bot, row_top in header_single:
        cy = (row_bot + row_top) / 2
        msp.add_text(
            label,
            dxfattribs={"layer": layer_text, "height": text_height,
                        "width": 0.7, "style": "Standard"}
        ).set_placement((header_cx, cy),
                        align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)

    line_spacing = text_height * 2.5
    block_h = line_spacing + text_height
    y_line_bottom_val = 45 + (30 - block_h) / 2 + text_height / 2
    y_line_top_val = y_line_bottom_val + line_spacing
    for label, y_pos in [("里程桩号", y_line_top_val), ("（千米+米）", y_line_bottom_val)]:
        msp.add_text(
            label,
            dxfattribs={"layer": layer_text, "height": text_height,
                        "width": 0.7, "style": "Standard"}
        ).set_placement((header_cx, y_pos),
                        align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)

    # ======== 8. 建筑物名称 ========
    name_mc_pairs = []
    for node in nodes:
        building_name = _get_building_display_name(node)
        if building_name:
            name_mc_pairs.append((building_name, node.station_MC))
    building_segments = []
    for bname, bmc in name_mc_pairs:
        if building_segments and building_segments[-1][0] == bname:
            building_segments[-1][1].append(bmc)
        else:
            building_segments.append((bname, [bmc]))
    building_segments = _merge_segments_across_gates(building_segments)
    for bname, mc_list in building_segments:
        if _is_gate_name(bname):
            mid_mc = mc_list[0]
        else:
            seg_start = mc_list[0]
            seg_end = mc_list[-1]
            left_bound = max((m for m in tall_line_mcs if m <= seg_start), default=seg_start)
            right_bound = min((m for m in tall_line_mcs if m >= seg_end), default=seg_end)
            mid_mc = (left_bound + right_bound) / 2.0
        msp.add_text(
            bname,
            dxfattribs={"layer": layer_text, "height": text_height,
                        "width": 0.7, "style": "Standard"}
        ).set_placement((sx(mid_mc), (110 + 120) / 2),
                        align=ezdxf.enums.TextEntityAlignment.MIDDLE)

    # ======== 9. 坡降 ========
    mc_to_node = {node.station_MC: node for node in nodes}
    for bname, mc_list in building_segments:
        if _is_gate_name(bname):
            continue
        if "倒虹吸" in bname:
            continue
        slope_text = _get_segment_slope_text(mc_list, mc_to_node)
        if not slope_text:
            continue
        seg_start = mc_list[0]
        seg_end = mc_list[-1]
        left_bound = max((m for m in tall_line_mcs if m <= seg_start), default=seg_start)
        right_bound = min((m for m in tall_line_mcs if m >= seg_end), default=seg_end)
        mid_mc = (left_bound + right_bound) / 2.0
        msp.add_text(
            slope_text,
            dxfattribs={"layer": layer_text, "height": text_height,
                        "width": 0.7, "style": "Standard"}
        ).set_placement((sx(mid_mc), (100 + 110) / 2),
                        align=ezdxf.enums.TextEntityAlignment.MIDDLE)

    # ======== 10. IP点名称 ========
    for idx, node in enumerate(nodes):
        struct_str = node.get_structure_type_str() or ""
        if node.is_transition or struct_str == "渐变段":
            continue
        if _is_special_structure_sv(node.structure_type):
            if _in_out_val(node.in_out) in ("进", "出"):
                struct_abbr = ""
                st_val = _struct_val(node.structure_type)
                if "隧洞" in st_val:
                    struct_abbr = "隧"
                elif "倒虹吸" in st_val:
                    struct_abbr = "倒"
                elif "渡槽" in st_val:
                    struct_abbr = "渡"
                in_out_str = "进" if _in_out_val(node.in_out) == "进" else "出"
                ip_name = f"{node.name}{struct_abbr}{in_out_str}"
            else:
                continue
        else:
            ip_name = f"{station_prefix}IP{getattr(node, 'ip_number', 0)}"
        text_x = (sx(node.station_MC) + first_col_x_offset
                  if node.station_MC == nodes[0].station_MC else sx(node.station_MC) - 1)
        msp.add_text(
            ip_name,
            dxfattribs={"layer": layer_text, "height": text_height,
                        "rotation": rotation, "width": 0.7, "style": "Standard"}
        ).set_placement((text_x, y_ip))

    return 40 + sx(last_mc), y_line_height


# ================================================================
# 1b. 生成纵断面表格 DXF（直接生成 CAD 可打开的 DXF 文件）
# ================================================================

def export_longitudinal_profile_dxf(panel):
    """一键生成上纵断面表格 DXF

    直接生成 DXF 文件，包含表格线框、渠底/渠顶/水面折线、
    高程文字、里程桩号、建筑物名称、坡降、IP点名称等全部内容。
    双击即可在 AutoCAD / 浩辰CAD / 中望CAD 中打开。
    """
    import ezdxf

    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "没有数据可导出")
        return

    valid_nodes = [n for n in nodes if n.bottom_elevation or n.top_elevation or n.water_level]
    if not valid_nodes:
        fluent_info(panel.window(), "警告", "没有可用的高程数据，请先执行计算。")
        return

    # 弹出参数配置对话框（复用 TXT 版设置）
    dlg = TextExportSettingsDialog(panel.window(), panel._text_export_settings)
    if dlg.exec() != QDialog.Accepted or dlg.result is None:
        return

    panel._text_export_settings.update(dlg.result)
    settings = dlg.result

    # 自动文件名
    try:
        ch_name = panel.channel_name_edit.text().strip()
        ch_level = panel.channel_level_combo.currentText()
        auto_name = f"{ch_name}{ch_level}_上纵断面表格.dxf"
    except Exception:
        auto_name = "上纵断面表格.dxf"

    file_path, _ = QFileDialog.getSaveFileName(
        panel, "保存上纵断面表格", auto_name,
        "DXF 文件 (*.dxf);;文本文件 (*.txt);;所有文件 (*.*)")
    if not file_path:
        return

    # 如果用户选择了 .txt，走原有 TXT 导出逻辑
    if file_path.lower().endswith('.txt'):
        _export_longitudinal_txt_to_path(panel, nodes, valid_nodes, settings, file_path)
        return

    try:
        try:
            proj_settings = panel._build_settings()
            station_prefix = proj_settings.get_station_prefix()
        except Exception:
            station_prefix = ""

        doc = ezdxf.new("R2010")
        msp = doc.modelspace()
        _setup_dxf_style(doc)
        _ensure_profile_layers(doc)

        _draw_profile_on_msp(msp, nodes, valid_nodes, settings, station_prefix)

        doc.saveas(file_path)

        if fluent_question(panel.window(), "完成",
                f"上纵断面表格 DXF 已生成（{len(nodes)} 个节点）:\n{file_path}\n\n是否立即打开该文件？"):
            os.startfile(file_path)

    except PermissionError:
        fluent_error(panel.window(), "文件被占用",
                     f"无法写入文件，该文件可能已被其他程序打开：\n\n{file_path}\n\n"
                     f"请先关闭该文件（如CAD等），然后重新操作。")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "导出错误", f"生成上纵断面表格 DXF 失败:\n{str(e)}")


def _export_longitudinal_txt_to_path(panel, nodes, valid_nodes, settings, file_path):
    """内部方法：将纵断面表格以 TXT（AutoCAD 命令）格式写入指定路径"""
    fmt = _format_number

    y_bottom = settings['y_bottom']
    y_top = settings['y_top']
    y_water = settings['y_water']
    text_height = settings['text_height']
    rotation = settings['rotation']
    elev_decimals = int(settings.get('elev_decimals', 3))
    y_name = settings.get('y_name', 115)
    y_slope = settings.get('y_slope', 105)
    y_ip = settings.get('y_ip', 77)
    y_station = settings.get('y_station', 47)
    y_line_height = settings.get('y_line_height', 120)
    scale_x = settings.get('scale_x', 1)
    scale_y = settings.get('scale_y', 1)

    def sx(mc):
        return mc / scale_x

    def sy(elev):
        return elev / scale_y

    s_y_bottom = fmt(y_bottom)
    s_y_top = fmt(y_top)
    s_y_water = fmt(y_water)
    s_height = fmt(text_height)
    s_rotation = fmt(rotation)
    s_y_name = fmt(y_name)
    s_y_slope = fmt(y_slope)
    s_y_ip = fmt(y_ip)
    s_y_station = fmt(y_station)
    s_y_line_height = fmt(y_line_height)

    first_col_x_offset = text_height + 1.3

    try:
        proj_settings = panel._build_settings()
        station_prefix = proj_settings.get_station_prefix()
    except Exception:
        station_prefix = ""

    def fmt_elev(value):
        if value is None:
            return f"{0:.{elev_decimals}f}"
        return f"{value:.{elev_decimals}f}"

    lines = []

    # pl 命令
    h_line_y_values = [0, 15, 30, 45, 75, 100, 110, 120]
    for hy in h_line_y_values:
        lines.append(f"pl {fmt(sx(0))},{hy} -40,{hy} ")
    lines.append(f"pl -40,0 -40,{s_y_line_height} ")
    lines.append("")

    tall_line_mcs = []
    for idx, node in enumerate(nodes):
        mc_scaled = fmt(sx(node.station_MC))
        is_special = (idx == 0 or idx == len(nodes) - 1)
        if not is_special and (_is_special_structure_sv(node.structure_type) and
                               _in_out_val(node.in_out) in ("进", "出")):
            is_special = True
        if is_special:
            tall_line_mcs.append(node.station_MC)
        height = s_y_line_height if is_special else "100"
        lines.append(f"pl {mc_scaled},0 {mc_scaled},{height} ")
    lines.append("")

    last_mc_scaled = fmt(sx(nodes[-1].station_MC))
    for hy in h_line_y_values:
        lines.append(f"pl {fmt(sx(0))},{hy} {last_mc_scaled},{hy} ")
    lines.append("")

    for node in valid_nodes:
        if node.bottom_elevation:
            lines.append(f"pl {fmt(sx(node.station_MC))},{fmt(sy(node.bottom_elevation))}")
    lines.append("")
    for node in valid_nodes:
        if node.top_elevation:
            lines.append(f"pl {fmt(sx(node.station_MC))},{fmt(sy(node.top_elevation))}")
    lines.append("")
    for node in valid_nodes:
        if node.water_level:
            lines.append(f"pl {fmt(sx(node.station_MC))},{fmt(sy(node.water_level))}")
    lines.append("")

    # -text 命令
    for idx, node in enumerate(nodes):
        mc = fmt(sx(node.station_MC))
        text_x = fmt(sx(node.station_MC) + first_col_x_offset) if idx == 0 else mc
        lines.append(
            f"-text {text_x},{s_y_bottom} {s_height} {s_rotation} {fmt_elev(node.bottom_elevation)} ")
    lines.append("")
    for idx, node in enumerate(nodes):
        mc = fmt(sx(node.station_MC))
        text_x = fmt(sx(node.station_MC) + first_col_x_offset) if idx == 0 else mc
        lines.append(
            f"-text {text_x},{s_y_top} {s_height} {s_rotation} {fmt_elev(node.top_elevation)} ")
    lines.append("")
    for idx, node in enumerate(nodes):
        mc = fmt(sx(node.station_MC))
        text_x = fmt(sx(node.station_MC) + first_col_x_offset) if idx == 0 else mc
        lines.append(
            f"-text {text_x},{s_y_water} {s_height} {s_rotation} {fmt_elev(node.water_level)} ")
    lines.append("")

    for idx, node in enumerate(nodes):
        mc = fmt(sx(node.station_MC))
        text_x = fmt(sx(node.station_MC) + first_col_x_offset) if idx == 0 else mc
        station_text = ProjectSettings.format_station(node.station_MC, station_prefix)
        lines.append(
            f"-text {text_x},{s_y_station} {s_height} {s_rotation} {station_text} ")
    lines.append("")

    header_col_width = 40
    header_single = [
        ("建筑物名称", 110, 120),
        ("坡 降", 100, 110),
        ("IP点名称", 75, 100),
        ("渠顶高程(m)", 30, 45),
        ("设计水位(m)", 15, 30),
        ("渠底高程(m)", 0, 15),
    ]
    header_cx = fmt(-40 + header_col_width / 2)
    for label, row_bot, row_top in header_single:
        cy = (row_bot + row_top) / 2
        lines.append(f"-text j mc {header_cx},{fmt(cy)} {s_height} 0 {label} ")
    line_spacing = text_height * 2.5
    block_h = line_spacing + text_height
    y_line_bottom_val = 45 + (30 - block_h) / 2 + text_height / 2
    y_line_top_val = y_line_bottom_val + line_spacing
    for label, y_pos in [("里程桩号", y_line_top_val), ("（千米+米）", y_line_bottom_val)]:
        lines.append(f"-text j mc {header_cx},{fmt(y_pos)} {s_height} 0 {label} ")
    lines.append("")

    name_mc_pairs = []
    for node in nodes:
        building_name = _get_building_display_name(node)
        if building_name:
            name_mc_pairs.append((building_name, node.station_MC))
    building_segments = []
    for bname, bmc in name_mc_pairs:
        if building_segments and building_segments[-1][0] == bname:
            building_segments[-1][1].append(bmc)
        else:
            building_segments.append((bname, [bmc]))
    building_segments = _merge_segments_across_gates(building_segments)
    for bname, mc_list in building_segments:
        if _is_gate_name(bname):
            mid_mc = mc_list[0]
        else:
            seg_start = mc_list[0]
            seg_end = mc_list[-1]
            left_bound = max((m for m in tall_line_mcs if m <= seg_start), default=seg_start)
            right_bound = min((m for m in tall_line_mcs if m >= seg_end), default=seg_end)
            mid_mc = (left_bound + right_bound) / 2.0
        lines.append(f"-text j mc {fmt(sx(mid_mc))},115 {s_height} 0 {bname} ")
    lines.append("")

    # 坡降（从建筑物名称分段派生，与名称行一一对齐）
    mc_to_node = {node.station_MC: node for node in nodes}
    for bname, mc_list in building_segments:
        # 闸/分水口：不显示坡降
        if _is_gate_name(bname):
            continue
        # 倒虹吸为有压流，不显示坡降
        if "倒虹吸" in bname:
            continue
        slope_text = _get_segment_slope_text(mc_list, mc_to_node)
        if not slope_text:
            continue
        # 居中位置与上方建筑物名称完全对齐
        seg_start = mc_list[0]
        seg_end = mc_list[-1]
        left_bound = max((m for m in tall_line_mcs if m <= seg_start), default=seg_start)
        right_bound = min((m for m in tall_line_mcs if m >= seg_end), default=seg_end)
        mid_mc = (left_bound + right_bound) / 2.0
        lines.append(f"-text j mc {fmt(sx(mid_mc))},105 {s_height} 0 {slope_text} ")
    lines.append("")

    for idx, node in enumerate(nodes):
        struct_str = node.get_structure_type_str() or ""
        if node.is_transition or struct_str == "渐变段":
            continue
        if _is_special_structure_sv(node.structure_type):
            if _in_out_val(node.in_out) in ("进", "出"):
                struct_abbr = ""
                st_val = _struct_val(node.structure_type)
                if "隧洞" in st_val:
                    struct_abbr = "隧"
                elif "倒虹吸" in st_val:
                    struct_abbr = "倒"
                elif "渡槽" in st_val:
                    struct_abbr = "渡"
                in_out_str = "进" if _in_out_val(node.in_out) == "进" else "出"
                ip_name = f"{node.name}{struct_abbr}{in_out_str}"
            else:
                continue
        else:
            ip_name = f"{station_prefix}IP{getattr(node, 'ip_number', 0)}"
        mc = fmt(sx(node.station_MC))
        text_x = fmt(sx(node.station_MC) + first_col_x_offset) if node.station_MC == nodes[0].station_MC else mc
        lines.append(
            f"-text {text_x},77 {s_height} {s_rotation} {ip_name} ")
    lines.append("")

    try:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        if fluent_question(panel.window(), "完成",
                f"上纵断面表格已生成（{len(nodes)} 个节点）:\n{file_path}\n\n是否立即打开该文件？"):
            os.startfile(file_path)
    except PermissionError:
        fluent_error(panel.window(), "文件被占用",
                     f"无法写入文件，该文件可能已被其他程序打开：\n\n{file_path}\n\n"
                     f"请先关闭该文件（如记事本、Word等），然后重新操作。")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "导出错误", f"生成上纵断面表格失败:\n{str(e)}")


# ================================================================
# 2. 生成bzzh2命令内容
# ================================================================

def extract_bzzh2_data(panel):
    """bzzh2命令提取工具

    从计算结果中提取所有有进出口标识（进/出）的建筑物节点，
    按桩号排序，整理为制表符分隔的TXT文件，供ZDM的bzzh2命令使用。
    """
    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "表格中没有数据，请先导入或输入数据。")
        return

    try:
        proj_settings = panel._build_settings()
        station_prefix = proj_settings.get_station_prefix()
    except Exception:
        station_prefix = ""

    bzzh2_rows = []
    for node in nodes:
        try:
            in_out = getattr(node, 'in_out', None)
            if _in_out_val(in_out) not in ("进", "出"):
                continue
            if getattr(node, 'is_transition', False):
                continue

            station_mc = getattr(node, 'station_MC', 0.0)
            if not isinstance(station_mc, (int, float)):
                station_mc = 0.0
            station_str = ProjectSettings.format_station(station_mc, station_prefix)

            struct_name = ""
            struct_str = _struct_val(node.structure_type)
            if struct_str:
                if "隧洞" in struct_str:
                    struct_name = "隧洞"
                elif "倒虹吸" in struct_str:
                    struct_name = "倒虹吸"
                elif "渡槽" in struct_str:
                    struct_name = "渡槽"
                elif "暗涵" in struct_str:
                    struct_name = "暗涵"
                else:
                    struct_name = struct_str

            in_out_str = "进" if _in_out_val(in_out) == "进" else "出"
            name = getattr(node, 'name', '') or ''
            desc = f"{name}{struct_name}{in_out_str}"
            bzzh2_rows.append((station_str, desc))
        except Exception as node_err:
            import traceback; traceback.print_exc()
            print(f"[bzzh2] 跳过节点（处理异常）: {node_err}")
            continue

    if not bzzh2_rows:
        fluent_info(
            panel.window(), "无可提取数据",
            "未找到有进出口标识的建筑物节点。\n\n"
            "bzzh2命令需要隧洞、倒虹吸、渡槽等建筑物的进/出口数据。\n"
            "请确保表格中已有相关数据并完成计算。")
        return

    # 预览对话框
    preview_dlg = QDialog(panel.window())
    preview_dlg.setWindowTitle("预览 — bzzh2命令数据（ZDM用）")
    preview_dlg.setMinimumSize(600, 400)
    preview_dlg.setStyleSheet(DIALOG_STYLE)
    dlg_lay = QVBoxLayout(preview_dlg)

    dlg_lay.addWidget(QLabel(
        f"共 {len(bzzh2_rows)} 条建筑物进出口数据，请确认内容后点击「确认导出」保存为TXT文件。"))

    table = QTableWidget(len(bzzh2_rows), 2)
    table.setHorizontalHeaderLabels(["桩号", "说明"])
    table.horizontalHeader().setStretchLastSection(False)
    table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
    table.setEditTriggers(QAbstractItemView.NoEditTriggers)
    table.setAlternatingRowColors(True)
    for i, (s, d) in enumerate(bzzh2_rows):
        table.setItem(i, 0, QTableWidgetItem(s))
        table.setItem(i, 1, QTableWidgetItem(d))
    auto_resize_table(table)
    dlg_lay.addWidget(table)

    btn_lay = QHBoxLayout()
    btn_lay.addStretch()
    btn_cancel = PushButton("取消"); btn_cancel.clicked.connect(preview_dlg.reject)
    btn_ok = PrimaryPushButton("确认导出"); btn_ok.clicked.connect(preview_dlg.accept)
    btn_lay.addWidget(btn_cancel); btn_lay.addWidget(btn_ok)
    dlg_lay.addLayout(btn_lay)

    # 绑定 ESC 关闭 / Enter 确认
    QShortcut(QKeySequence(Qt.Key_Escape), preview_dlg, preview_dlg.reject)
    QShortcut(QKeySequence(Qt.Key_Return), preview_dlg, preview_dlg.accept)

    if preview_dlg.exec() != QDialog.Accepted:
        return

    # 保存文件
    try:
        ch_name = panel.channel_name_edit.text().strip()
        ch_level = panel.channel_level_combo.currentText()
        auto_name = f"{ch_name}{ch_level}_ZDM的bzzh2命令.txt"
    except Exception:
        auto_name = "ZDM的bzzh2命令.txt"

    file_path, _ = QFileDialog.getSaveFileName(
        panel, "保存bzzh2命令数据", auto_name,
        "文本文件 (*.txt);;所有文件 (*.*)")
    if not file_path:
        return

    try:
        with open(file_path, 'w', encoding='gbk', errors='replace') as f:
            for station_str, desc in bzzh2_rows:
                f.write(f"{station_str}\t{desc}\t\n")

        if fluent_question(panel.window(), "提取完成",
                f"bzzh2命令数据提取成功！\n\n"
                f"文件保存路径:\n{file_path}\n\n"
                f"导出数据行数: {len(bzzh2_rows)}\n\n"
                f"请使用ZDM的bzzh2命令完成建筑物进出口上平面图。\n\n"
                f"是否要立即打开该txt文件？"):
            try:
                os.startfile(file_path)
            except AttributeError:
                import subprocess
                subprocess.Popen(['xdg-open', file_path])
            except Exception:
                fluent_info(panel.window(), "打开文件",
                            f"无法自动打开文件，请手动打开:\n\n{file_path}")
    except PermissionError:
        fluent_error(panel.window(), "文件被占用",
                     f"无法写入文件，该文件可能已被其他程序打开：\n\n{file_path}\n\n"
                     f"请先关闭该文件（如记事本、Word等），然后重新操作。")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "提取失败",
                     f"bzzh2命令数据提取过程中发生错误:\n\n{str(e)}")


# ================================================================
# 3. 建筑物名称上平面图
# ================================================================

def export_building_name_plan(panel):
    """生成「平行于轴线的建筑物名称上平面图」AutoCAD -TEXT 命令

    对于进出口之间有多个IP点的建筑物，文字放置在最中间两个相邻
    IP点连线段的中点处（垂直偏移），方向角取该中间线段的方向。
    """
    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "表格中没有数据，请先导入或输入数据。")
        return

    try:
        # 按建筑物分组收集所有节点
        building_groups = {}
        building_order = []
        for node in nodes:
            if node.is_transition or getattr(node, 'is_auto_inserted_channel', False):
                continue
            if not node.name:
                continue
            key = (node.name, _struct_val(node.structure_type))
            if key not in building_groups:
                building_groups[key] = []
                building_order.append(key)
            building_groups[key].append(node)

        # 筛选有完整进出口且坐标有效的建筑物
        valid_buildings = []
        for key in building_order:
            group = building_groups[key]
            has_inlet = any(_in_out_val(n.in_out) == "进" for n in group)
            has_outlet = any(_in_out_val(n.in_out) == "出" for n in group)
            if not (has_inlet and has_outlet):
                continue
            coord_nodes = [n for n in group
                           if n.x is not None and n.y is not None and
                           (n.x != 0 or n.y != 0)]
            if len(coord_nodes) >= 2:
                valid_buildings.append((key, group, coord_nodes))

        if not valid_buildings:
            fluent_info(
                panel.window(), "无可提取数据",
                "未找到有效的建筑物进出口数据。\n\n"
                "需要隧洞、倒虹吸、渡槽等建筑物同时存在进口和出口，\n"
                "且节点具有有效的X、Y坐标。")
            return

        # 参数设置对话框
        dlg = PlanTextSettingsDialog(panel.window(), panel._plan_text_settings)
        if dlg.exec() != QDialog.Accepted or dlg.result is None:
            return

        panel._plan_text_settings.update(dlg.result)
        offset = dlg.result['offset']
        text_height = dlg.result['text_height']

        # 生成 -TEXT 命令
        text_commands = []
        for key, all_nodes, coord_nodes in valid_buildings:
            N = len(coord_nodes)
            mid_right = N // 2
            mid_left = mid_right - 1

            node_a = coord_nodes[mid_left]
            node_b = coord_nodes[mid_right]
            x1, y1 = node_a.x, node_a.y
            x2, y2 = node_b.x, node_b.y

            dx = x2 - x1
            dy = y2 - y1
            if abs(dx) < 1e-10 and abs(dy) < 1e-10:
                continue

            mx = (x1 + x2) / 2
            my = (y1 + y2) / 2
            angle_rad = math.atan2(dy, dx)
            angle_deg = math.degrees(angle_rad)

            text_x = mx - offset * math.sin(angle_rad)
            text_y = my + offset * math.cos(angle_rad)

            inlet_node = next(
                (n for n in all_nodes if _in_out_val(n.in_out) == "进"),
                all_nodes[0])
            struct_name = ""
            struct_str = _struct_val(inlet_node.structure_type)
            if struct_str:
                if "隧洞" in struct_str:
                    struct_name = "隧洞"
                elif "倒虹吸" in struct_str:
                    struct_name = "倒虹吸"
                elif "渡槽" in struct_str:
                    struct_name = "渡槽"
                elif "暗涵" in struct_str:
                    struct_name = "暗涵"
                else:
                    struct_name = struct_str

            building_name = f"{inlet_node.name or ''}{struct_name}"
            cmd = (f"-TEXT J MC {text_x},{text_y} "
                   f"{text_height} {angle_deg} {building_name}")
            text_commands.append((building_name, N, cmd))

        if not text_commands:
            fluent_info(panel.window(), "无有效数据",
                        "没有生成任何 -TEXT 命令，请检查建筑物坐标。")
            return

        # 显示预览
        all_cmds_text = "\n".join(cmd for _, _, cmd in text_commands)

        preview = QDialog(panel.window())
        preview.setWindowTitle(f"建筑物名称上平面图 — {len(text_commands)} 条命令")
        preview.setMinimumSize(700, 400)
        preview.setStyleSheet(DIALOG_STYLE)
        p_lay = QVBoxLayout(preview)

        p_lay.addWidget(QLabel(
            f"共 {len(text_commands)} 个建筑物  |  "
            f"偏移距离: {offset}  |  文字高度: {text_height}"))

        text_widget = QTextEdit()
        text_widget.setReadOnly(True)
        text_widget.setFont(QFont("Consolas", 10))
        html_parts = []
        for i, (name, node_count, cmd) in enumerate(text_commands):
            comment = f"' [{i+1}] {name}（{node_count}个IP点）"
            html_parts.append(f'<span style="color:gray">{comment}</span><br>')
            html_parts.append(f'{cmd}<br><br>')
        text_widget.setHtml('<pre style="font-family:Consolas;font-size:10pt">' +
                            ''.join(html_parts) + '</pre>')
        p_lay.addWidget(text_widget)

        btn_lay = QHBoxLayout()
        status_label = QLabel("")
        status_label.setStyleSheet("color: green;")
        btn_lay.addWidget(status_label)
        btn_lay.addStretch()

        def copy_commands_only():
            QApplication.clipboard().setText(all_cmds_text)
            status_label.setText("✓ 已复制纯命令到剪贴板，可直接粘贴到 AutoCAD")

        def copy_all_content():
            QApplication.clipboard().setText(text_widget.toPlainText())
            status_label.setText("✓ 已复制全部内容到剪贴板（含注释）")

        btn_copy_all = PushButton("复制全部内容")
        btn_copy_all.clicked.connect(copy_all_content)
        btn_copy_cmd = PrimaryPushButton("复制纯命令")
        btn_copy_cmd.clicked.connect(copy_commands_only)
        btn_close = PushButton("关闭")
        btn_close.clicked.connect(preview.close)
        btn_lay.addWidget(btn_copy_all)
        btn_lay.addWidget(btn_copy_cmd)
        btn_lay.addWidget(btn_close)
        p_lay.addLayout(btn_lay)

        preview.exec()

    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "生成失败",
                     f"建筑物名称上平面图生成过程中发生错误:\n\n{str(e)}")


# ================================================================
# 4. IP坐标及弯道参数表
# ================================================================

def _draw_ip_table_on_msp(msp, ox, oy, preview_data,
                          title="IP坐标及弯道参数表", layer="IP_TABLE"):
    """在 modelspace 上绘制IP坐标及弯道参数表。返回 (width, height)。"""
    import ezdxf

    ROW_H = 6.0
    HDR_ROW_H = 6.0
    TITLE_ROW_H = 7.0
    TEXT_H = 2.2
    HDR_TEXT_H = 2.5
    TITLE_TEXT_H = 3.0
    COL_PAD = 3.0

    sub_headers = [
        "IP点", "E（m）", "N（m）",
        "弯前(千米+米)", "弯中(千米+米)", "弯末(千米+米)",
        "转角", "半径", "切线长", "弧长", "底高程(m)",
    ]
    group_headers = [
        (0, 0, "IP点"),
        (1, 2, "坐标值"),
        (3, 5, "桩号"),
        (6, 9, "弯道参数"),
        (10, 10, "底高程(m)"),
    ]
    v_merged = {0, 10}
    ncols = 11
    nrows = len(preview_data)

    _wf = 1.0
    def _tw(text, h):
        if text is None:
            return 0.0
        return sum(h * _wf if ord(c) > 0x7F else h * 0.6 * _wf for c in str(text))

    col_w = [0.0] * ncols
    for ci, hdr in enumerate(sub_headers):
        col_w[ci] = _tw(hdr, HDR_TEXT_H)
    for sc, ec, gtxt in group_headers:
        span = ec - sc + 1
        gw_each = _tw(gtxt, HDR_TEXT_H) / span
        for ci in range(sc, ec + 1):
            col_w[ci] = max(col_w[ci], gw_each)
    for row in preview_data:
        for ci, val in enumerate(row):
            if ci < ncols:
                col_w[ci] = max(col_w[ci], _tw(val, TEXT_H))
    col_w = [w + COL_PAD for w in col_w]

    col_x = [ox]
    for w in col_w:
        col_x.append(col_x[-1] + w)
    total_w = col_x[-1] - col_x[0]
    x_left, x_right = col_x[0], col_x[-1]

    y_title_top = oy
    y_title_bot = y_title_top - TITLE_ROW_H
    y_hdr1_bot = y_title_bot - HDR_ROW_H
    y_hdr2_bot = y_hdr1_bot - HDR_ROW_H
    y_data_top = y_hdr2_bot
    row_y = [y_data_top]
    for _ in range(nrows):
        row_y.append(row_y[-1] - ROW_H)

    dxa = {"layer": layer}

    # === 标题行 ===
    msp.add_line((x_left, y_title_top), (x_right, y_title_top), dxfattribs=dxa)
    msp.add_line((x_left, y_title_bot), (x_right, y_title_bot), dxfattribs=dxa)
    msp.add_line((x_left, y_title_top), (x_left, y_title_bot), dxfattribs=dxa)
    msp.add_line((x_right, y_title_top), (x_right, y_title_bot), dxfattribs=dxa)
    msp.add_text(
        title,
        dxfattribs={"layer": layer, "height": TITLE_TEXT_H,
                    "width": 0.7, "style": "Standard"}
    ).set_placement(
        (x_left + total_w / 2, (y_title_top + y_title_bot) / 2),
        align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER,
    )

    # === 表头区边框 ===
    msp.add_line((x_left, y_hdr2_bot), (x_right, y_hdr2_bot), dxfattribs=dxa)
    msp.add_line((x_left, y_title_bot), (x_left, y_hdr2_bot), dxfattribs=dxa)
    msp.add_line((x_right, y_title_bot), (x_right, y_hdr2_bot), dxfattribs=dxa)

    for ci in range(ncols):
        if ci not in v_merged:
            msp.add_line((col_x[ci], y_hdr1_bot), (col_x[ci + 1], y_hdr1_bot),
                         dxfattribs=dxa)

    drawn_x = set()
    for sc, ec, _ in group_headers:
        for bx in (col_x[sc], col_x[ec + 1]):
            if bx not in drawn_x:
                msp.add_line((bx, y_title_bot), (bx, y_hdr2_bot), dxfattribs=dxa)
                drawn_x.add(bx)
        if sc != ec:
            for ci in range(sc + 1, ec + 1):
                msp.add_line((col_x[ci], y_hdr1_bot), (col_x[ci], y_hdr2_bot),
                             dxfattribs=dxa)

    for sc, ec, text in group_headers:
        cx = (col_x[sc] + col_x[ec + 1]) / 2
        cy = ((y_title_bot + y_hdr2_bot) / 2 if sc in v_merged
              else (y_title_bot + y_hdr1_bot) / 2)
        msp.add_text(
            text,
            dxfattribs={"layer": layer, "height": HDR_TEXT_H,
                        "width": 0.7, "style": "Standard"}
        ).set_placement((cx, cy), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)

    for ci, hdr in enumerate(sub_headers):
        if ci in v_merged:
            continue
        cx = (col_x[ci] + col_x[ci + 1]) / 2
        cy = (y_hdr1_bot + y_hdr2_bot) / 2
        msp.add_text(
            hdr,
            dxfattribs={"layer": layer, "height": HDR_TEXT_H,
                        "width": 0.7, "style": "Standard"}
        ).set_placement((cx, cy), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)

    # === 数据区 ===
    msp.add_line((x_left, y_data_top), (x_right, y_data_top), dxfattribs=dxa)
    if nrows > 0:
        msp.add_line((x_left, row_y[-1]), (x_right, row_y[-1]), dxfattribs=dxa)
    for ri in range(1, nrows):
        msp.add_line((x_left, row_y[ri]), (x_right, row_y[ri]), dxfattribs=dxa)

    y_bottom = row_y[-1] if nrows > 0 else y_data_top
    for x in col_x:
        msp.add_line((x, y_data_top), (x, y_bottom), dxfattribs=dxa)

    for ri, row_vals in enumerate(preview_data):
        for ci, val in enumerate(row_vals):
            if val is None or val == "":
                continue
            cx = (col_x[ci] + col_x[ci + 1]) / 2
            cy = (row_y[ri] + row_y[ri + 1]) / 2
            msp.add_text(
                str(val),
                dxfattribs={"layer": layer, "height": TEXT_H,
                            "width": 0.7, "style": "Standard"}
            ).set_placement((cx, cy), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)

    total_h = y_title_top - (row_y[-1] if nrows > 0 else y_data_top)
    return total_w, total_h


def _write_ip_table_dxf(file_path, preview_data, title="IP坐标及弯道参数表"):
    """将IP坐标及弯道参数表写入独立DXF文件。"""
    import ezdxf
    doc = ezdxf.new("R2010")
    msp = doc.modelspace()
    _setup_dxf_style(doc)
    _draw_ip_table_on_msp(msp, 0.0, 0.0, preview_data, title, "IP_TABLE")
    doc.saveas(file_path)


def export_ip_plan_table(panel):
    """导出IP坐标及弯道参数表DXF/Excel文件（含合并表头、桩号格式化）"""
    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "没有数据可导出，请先执行计算")
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        fluent_info(panel.window(), "缺少依赖",
                    "需要安装 openpyxl: pip install openpyxl")
        return

    try:
        try:
            proj_settings = panel._build_settings()
            station_prefix = proj_settings.get_station_prefix()
        except Exception:
            station_prefix = ""

        def _safe_float(val, default=0.0):
            if val is None:
                return default
            try:
                return float(val)
            except (TypeError, ValueError):
                return default

        def _format_ip_name(node):
            try:
                if _in_out_val(node.in_out) in ("进", "出"):
                    struct_abbr = ""
                    struct_str = _struct_val(node.structure_type)
                    if struct_str:
                        if "隧洞" in struct_str: struct_abbr = "隧"
                        elif "倒虹吸" in struct_str: struct_abbr = "倒"
                        elif "渡槽" in struct_str: struct_abbr = "渡"
                        elif "暗涵" in struct_str: struct_abbr = "暗"
                    in_out_str = "进" if _in_out_val(node.in_out) == "进" else "出"
                    return f"{node.name}{struct_abbr}{in_out_str}"
            except Exception:
                pass
            return f"{station_prefix}IP{getattr(node, 'ip_number', 0)}"

        def _format_station(value):
            return ProjectSettings.format_station(_safe_float(value), station_prefix)

        # 过滤节点
        real_nodes = [
            n for n in nodes
            if not getattr(n, 'is_transition', False)
            and not getattr(n, 'is_auto_inserted_channel', False)
        ]

        if not real_nodes:
            fluent_info(panel.window(), "警告", "没有有效的IP点数据可导出")
            return

        # 构建预览数据
        preview_headers = [
            "IP点", "E（m）", "N（m）",
            "弯前(千米+米)", "弯中(千米+米)", "弯末(千米+米)",
            "转角", "半径", "切线长", "弧长", "底高程\nm"
        ]
        preview_data = []
        for idx, node in enumerate(real_nodes):
            try:
                row = [
                    _format_ip_name(node),
                    f"{_safe_float(node.x):.6f}",
                    f"{_safe_float(node.y):.6f}",
                    _format_station(node.station_BC),
                    _format_station(node.station_MC),
                    _format_station(node.station_EC),
                    f"{_safe_float(node.turn_angle):.3f}",
                    f"{_safe_float(node.turn_radius):.3f}",
                    f"{_safe_float(node.tangent_length):.3f}",
                    f"{_safe_float(node.arc_length):.3f}",
                    f"{_safe_float(node.bottom_elevation):.3f}" if _safe_float(node.bottom_elevation) != 0 else "-",
                ]
                preview_data.append(row)
            except Exception as row_err:
                print(f"[警告] 第{idx}行数据格式化失败: {row_err}")
                import traceback; traceback.print_exc()
                preview_data.append([
                    f"IP{getattr(node, 'ip_number', '?')}",
                    "0.000000", "0.000000",
                    "0+000.000", "0+000.000", "0+000.000",
                    "0.000", "0.000", "0.000", "0.000", "-",
                ])

        # 预览对话框
        preview_dlg = QDialog(panel.window())
        preview_dlg.setWindowTitle("预览 — IP坐标及弯道参数表")
        preview_dlg.setMinimumSize(950, 450)
        preview_dlg.setStyleSheet(DIALOG_STYLE)
        dlg_lay = QVBoxLayout(preview_dlg)

        dlg_lay.addWidget(QLabel(
            f"共 {len(preview_data)} 条IP点数据，请确认内容后点击「确认导出」保存为DXF或Excel文件。"))

        table = QTableWidget(len(preview_data), len(preview_headers))
        table.setHorizontalHeaderLabels(preview_headers)
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        for r, row_data in enumerate(preview_data):
            for c, val in enumerate(row_data):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignCenter)
                table.setItem(r, c, item)
        auto_resize_table(table)
        dlg_lay.addWidget(table)

        btn_lay = QHBoxLayout()
        btn_lay.addStretch()
        btn_cancel = PushButton("取消"); btn_cancel.clicked.connect(preview_dlg.reject)
        btn_ok = PrimaryPushButton("确认导出"); btn_ok.clicked.connect(preview_dlg.accept)
        btn_lay.addWidget(btn_cancel); btn_lay.addWidget(btn_ok)
        dlg_lay.addLayout(btn_lay)

        # 绑定 ESC 关闭 / Enter 确认
        QShortcut(QKeySequence(Qt.Key_Escape), preview_dlg, preview_dlg.reject)
        QShortcut(QKeySequence(Qt.Key_Return), preview_dlg, preview_dlg.accept)

        if preview_dlg.exec() != QDialog.Accepted:
            return

        # 保存
        try:
            ch_name = panel.channel_name_edit.text().strip()
            ch_level = panel.channel_level_combo.currentText()
            auto_name = f"{ch_name}{ch_level}_IP坐标及弯道参数表.dxf"
        except Exception:
            auto_name = "IP坐标及弯道参数表.dxf"

        file_path, _ = QFileDialog.getSaveFileName(
            panel, "保存IP坐标及弯道参数表", auto_name,
            "DXF文件 (*.dxf);;Excel文件 (*.xlsx);;所有文件 (*.*)")
        if not file_path:
            return

        # DXF 导出（紧凑排版、自适应列宽、无底色）
        if file_path.lower().endswith('.dxf'):
            _write_ip_table_dxf(file_path, preview_data)
            if fluent_question(panel.window(), "导出完成",
                    f"IP坐标及弯道参数表DXF导出成功！\n\n"
                    f"文件保存路径:\n{file_path}\n\n"
                    f"导出IP点数量: {len(real_nodes)}\n\n"
                    f"是否要立即打开该文件？"):
                try:
                    os.startfile(file_path)
                except Exception:
                    pass
            return

        # openpyxl 写入
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IP点上平面图"

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(name='Microsoft YaHei', size=10, bold=True)
        data_font = Font(name='Microsoft YaHei', size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        ws.cell(row=1, column=1, value="IP点")
        ws.cell(row=1, column=2, value="坐标值")
        ws.cell(row=1, column=4, value="桩号")
        ws.cell(row=1, column=7, value="弯道参数")
        ws.cell(row=1, column=11, value="底高程\nm")
        ws.cell(row=2, column=2, value="E（m）")
        ws.cell(row=2, column=3, value="N（m）")
        ws.cell(row=2, column=4, value="弯前(千米+米)")
        ws.cell(row=2, column=5, value="弯中(千米+米)")
        ws.cell(row=2, column=6, value="弯末(千米+米)")
        ws.cell(row=2, column=7, value="转角")
        ws.cell(row=2, column=8, value="半径")
        ws.cell(row=2, column=9, value="切线长")
        ws.cell(row=2, column=10, value="弧长")

        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:C1')
        ws.merge_cells('D1:F1')
        ws.merge_cells('G1:J1')
        ws.merge_cells('K1:K2')

        for row in range(1, 3):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill

        for row_idx, node in enumerate(real_nodes, start=3):
            ws.cell(row=row_idx, column=1, value=_format_ip_name(node))
            cell_b = ws.cell(row=row_idx, column=2, value=_safe_float(node.x))
            cell_b.number_format = '0.000000'
            cell_c = ws.cell(row=row_idx, column=3, value=_safe_float(node.y))
            cell_c.number_format = '0.000000'
            ws.cell(row=row_idx, column=4, value=_format_station(node.station_BC))
            ws.cell(row=row_idx, column=5, value=_format_station(node.station_MC))
            ws.cell(row=row_idx, column=6, value=_format_station(node.station_EC))
            cell_g = ws.cell(row=row_idx, column=7,
                             value=round(_safe_float(node.turn_angle), 3))
            cell_g.number_format = '0.000'
            cell_h = ws.cell(row=row_idx, column=8,
                             value=round(_safe_float(node.turn_radius), 3))
            cell_h.number_format = '0.000'
            cell_i = ws.cell(row=row_idx, column=9,
                             value=round(_safe_float(node.tangent_length), 3))
            cell_i.number_format = '0.000'
            cell_j = ws.cell(row=row_idx, column=10,
                             value=round(_safe_float(node.arc_length), 3))
            cell_j.number_format = '0.000'
            _be_val = _safe_float(node.bottom_elevation)
            cell_k = ws.cell(row=row_idx, column=11,
                             value=round(_be_val, 3) if _be_val != 0 else "-")
            if _be_val != 0:
                cell_k.number_format = '0.000'

            for col in range(1, 12):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = data_font
                cell.border = thin_border
                if col == 1:
                    cell.alignment = left_align
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # 自适应列宽（根据实际内容计算，避免出现 ###）
        for col_idx in range(1, 12):
            col_letter = chr(64 + col_idx)  # A=1, B=2, ...
            max_len = 0
            for row_idx2 in range(1, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx2, column=col_idx).value
                if cell_val is not None:
                    s = str(cell_val)
                    # CJK字符算2个宽度单位
                    char_w = sum(2 if ord(c) > 0x7F else 1 for c in s)
                    max_len = max(max_len, char_w)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 8)

        wb.save(file_path)
        wb.close()

        if fluent_question(panel.window(), "导出完成",
                f"IP坐标及弯道参数表导出成功！\n\n"
                f"文件保存路径:\n{file_path}\n\n"
                f"导出IP点数量: {len(real_nodes)}\n\n"
                f"是否要立即打开该文件？"):
            try:
                os.startfile(file_path)
            except Exception:
                pass

    except PermissionError:
        fluent_error(panel.window(), "文件被占用",
                     f"无法写入文件，该文件可能已被其他程序打开：\n\n{file_path}\n\n"
                     f"请先关闭该文件（如Excel等），然后重新操作。")
    except Exception as e:
        import traceback
        traceback.print_exc()
        fluent_error(panel.window(), "导出失败",
                     f"IP坐标及弯道参数表导出失败，请检查以下可能的原因：\n\n"
                     f"1. 目标文件是否被其他程序占用（如AutoCAD、Excel等）\n"
                     f"2. 文件保存路径是否有写入权限\n"
                     f"3. 数据是否完整（坐标、桩号等）\n\n"
                     f"错误信息：{str(e)}\n\n"
                     f"如仍无法解决，请将以上信息反馈给技术支持。")


# ================================================================
# 6. 合并导出全部DXF（横向分区布局）
# ================================================================

def export_combined_dxf(panel):
    """将纵断面表格、断面汇总表、IP坐标表合并导出到一个DXF文件。

    布局：纵断面表格在上方（全宽），下方左侧放断面汇总表，右侧放IP表。
    """
    import ezdxf

    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "没有数据可导出，请先执行计算")
        return

    valid_nodes = [n for n in nodes if n.bottom_elevation or n.top_elevation or n.water_level]
    if not valid_nodes:
        fluent_info(panel.window(), "警告", "没有可用的高程数据，请先执行计算。")
        return

    # ---- 1. 纵断面参数设置 ----
    dlg = TextExportSettingsDialog(panel.window(), panel._text_export_settings)
    if dlg.exec() != QDialog.Accepted or dlg.result is None:
        return
    panel._text_export_settings.update(dlg.result)
    profile_settings = dlg.result

    # ---- 2. 获取项目设置 ----
    try:
        proj_settings = panel._build_settings()
        station_prefix = proj_settings.get_station_prefix()
    except Exception:
        proj_settings = None
        station_prefix = ""

    # ---- 3. 选择保存路径 ----
    try:
        ch_name = panel.channel_name_edit.text().strip()
        ch_level = panel.channel_level_combo.currentText()
        auto_name = f"{ch_name}{ch_level}_全部表格.dxf"
    except Exception:
        auto_name = "全部表格.dxf"

    file_path, _ = QFileDialog.getSaveFileName(
        panel, "保存合并DXF（纵断面+断面汇总+IP表）", auto_name,
        "DXF 文件 (*.dxf);;所有文件 (*.*)")
    if not file_path:
        return
    if not file_path.lower().endswith('.dxf'):
        file_path += '.dxf'

    try:
        # ---- 创建 DXF 文档 ----
        doc = ezdxf.new("R2010")
        msp = doc.modelspace()
        _setup_dxf_style(doc)

        # 三个组件使用独立图层（带前缀），便于在CAD中分别控制显示
        _PROF_PREFIX = "纵断面_"
        _SUMM_LAYER = "断面汇总表"
        _IP_LAYER = "IP坐标表"

        _ensure_profile_layers(doc, layer_prefix=_PROF_PREFIX)
        if _SUMM_LAYER not in doc.layers:
            doc.layers.new(_SUMM_LAYER, dxfattribs={"color": 7})   # 白色
        if _IP_LAYER not in doc.layers:
            doc.layers.new(_IP_LAYER, dxfattribs={"color": 7})     # 白色

        GAP = 20.0  # 各区域间距

        # ======== A. 纵断面表格（顶部，原点(0,0)） ========
        prof_w, prof_h = _draw_profile_on_msp(
            msp, nodes, valid_nodes, profile_settings, station_prefix,
            layer_prefix=_PROF_PREFIX)

        # 下方区域起始Y（纵断面底部再向下留间距）
        below_y = -GAP

        # ======== B. 断面汇总表（左下） ========
        summary_w = 0.0
        summary_h = 0.0
        try:
            from 渠系建筑物断面计算.生成断面汇总表 import (
                _extract_segment_defaults_from_nodes,
                _segment_name,
                _dxf_draw_table,
                _dxf_auto_col_widths,
                _DXF_TABLE_GAP,
                _DXF_BUILDERS,
                compute_rect_channel, compute_trapezoid_channel,
                compute_tunnel, compute_tunnel_circular, compute_tunnel_horseshoe,
                compute_aqueduct_u, compute_aqueduct_rect,
                compute_rect_culvert, compute_circular_pipe, compute_siphon,
                _default_segments_rect_channel, _default_segments_trap_channel,
                _default_segments_tunnel_arch, _default_segments_tunnel_circular,
                _default_segments_tunnel_horseshoe,
                _default_segments_aqueduct_u, _default_segments_aqueduct_rect,
                _default_segments_rect_culvert, _default_segments_circular_pipe,
                _default_segments_siphon,
                SIPHON_MATERIALS,
            )

            node_defaults, flow_qs = _extract_segment_defaults_from_nodes(nodes)

            # 确定流量段数和Q值
            counts = []
            if proj_settings and getattr(proj_settings, "design_flows", None):
                flows = [q for q in proj_settings.design_flows
                         if isinstance(q, (int, float)) and q > 0]
                if flows:
                    counts.append(len(flows))
            if flow_qs:
                counts.append(max(flow_qs.keys()))
            if node_defaults:
                for data in node_defaults.values():
                    if data:
                        counts.append(max(data.keys()))
            seg_count = max(1, max(counts)) if counts else 7

            # 构建Q值列表
            fallback_qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
            if proj_settings and getattr(proj_settings, "design_flows", None):
                flows = [q for q in proj_settings.design_flows
                         if isinstance(q, (int, float)) and q > 0]
                if flows:
                    qs = [flows[i] if i < len(flows) else flows[-1]
                          for i in range(seg_count)]
                else:
                    qs = fallback_qs[:seg_count]
            elif flow_qs:
                qs = []
                for i in range(1, seg_count + 1):
                    q = flow_qs.get(i, 0.0)
                    qs.append(q if q > 0 else
                              (fallback_qs[i-1] if i-1 < len(fallback_qs)
                               else fallback_qs[-1]))
            else:
                qs = (list(fallback_qs[:seg_count])
                      + [fallback_qs[-1]] * max(0, seg_count - len(fallback_qs)))

            has_source = bool(nodes) and any(node_defaults.values())

            def _make_segs(default_fn, overrides_by_idx=None):
                if has_source and overrides_by_idx is not None:
                    if not overrides_by_idx:
                        return []
                    pool = default_fn()
                    segs = []
                    for idx in sorted(overrides_by_idx.keys()):
                        base = dict(pool[0]) if pool else {}
                        base["name"] = _segment_name(idx)
                        base.update(overrides_by_idx[idx])
                        if 0 < idx <= len(qs):
                            base["Q"] = qs[idx - 1]
                        segs.append(base)
                    return segs
                segs = default_fn()
                if len(segs) < seg_count:
                    last = segs[-1] if segs else {}
                    for idx in range(len(segs) + 1, seg_count + 1):
                        new = dict(last)
                        new["name"] = _segment_name(idx)
                        segs.append(new)
                segs = segs[:seg_count]
                for i, seg in enumerate(segs):
                    if overrides_by_idx and (i+1) in overrides_by_idx:
                        seg.update(overrides_by_idx[i+1])
                    if i < len(qs):
                        seg["Q"] = qs[i]
                return segs

            # 构建各类型数据
            rc = _make_segs(_default_segments_rect_channel,
                            node_defaults.get("rect_channel"))
            tr = _make_segs(_default_segments_trap_channel,
                            node_defaults.get("trap_channel"))
            ta = _make_segs(_default_segments_tunnel_arch,
                            node_defaults.get("tunnel_arch"))
            tc = _make_segs(_default_segments_tunnel_circular,
                            node_defaults.get("tunnel_circular"))
            th = _make_segs(_default_segments_tunnel_horseshoe,
                            node_defaults.get("tunnel_horseshoe"))
            au = _make_segs(_default_segments_aqueduct_u,
                            node_defaults.get("aqueduct_u"))
            ar = _make_segs(_default_segments_aqueduct_rect,
                            node_defaults.get("aqueduct_rect"))
            rv = _make_segs(_default_segments_rect_culvert,
                            node_defaults.get("rect_culvert"))
            cp = _make_segs(_default_segments_circular_pipe,
                            node_defaults.get("circular_channel"))
            sp = _make_segs(_default_segments_siphon,
                            node_defaults.get("siphon"))

            # 读取用户自定义的构造参数（如果之前在断面汇总表对话框中设置过）
            _struct_t = getattr(panel, '_custom_struct_thickness', None)
            _rock_lining = getattr(panel, '_custom_rock_lining', None)

            # 将壁厚/衬砌参数注入各类型 segments
            if _struct_t:
                _st_rc = _struct_t.get('rect_channel', {})
                for seg in rc:
                    if 'wall_t' in _st_rc:
                        seg['wall_t'] = _st_rc['wall_t']
                    if 'tie_rod' in _st_rc:
                        seg['tie_rod'] = _st_rc['tie_rod']
                _st_tr = _struct_t.get('trap_channel', {})
                for seg in tr:
                    if 'wall_t' in _st_tr:
                        seg['wall_t'] = _st_tr['wall_t']
                    if 'tie_rod' in _st_tr:
                        seg['tie_rod'] = _st_tr['tie_rod']
                _st_au = _struct_t.get('aqueduct_u', {})
                for seg in au:
                    if 'wall_t' in _st_au:
                        seg['wall_t'] = _st_au['wall_t']
                _st_ar = _struct_t.get('aqueduct_rect', {})
                for seg in ar:
                    if 'wall_t' in _st_ar:
                        seg['wall_t'] = _st_ar['wall_t']
                _st_rv = _struct_t.get('rect_culvert', {})
                for seg in rv:
                    for k in ('t0', 't1', 't2'):
                        if k in _st_rv:
                            seg[k] = _st_rv[k]

            # 读取隧洞断面设计方式（复用断面汇总表对话框中的设置）
            _tu = getattr(panel, '_custom_tunnel_unified', {})
            _tu_arch = _tu.get('tunnel_arch', False)
            _tu_circ = _tu.get('tunnel_circular', False)
            _tu_horse = _tu.get('tunnel_horseshoe', False)

            # 计算
            d_rc = compute_rect_channel(rc) if rc else []
            d_tr = compute_trapezoid_channel(tr) if tr else []
            d_ta, _ = compute_tunnel(ta, _rock_lining, unified=_tu_arch) if ta else ([], {})
            d_tc, _ = compute_tunnel_circular(tc, _rock_lining, unified=_tu_circ) if tc else ([], {})
            d_th, d_th_info = compute_tunnel_horseshoe(th, rock_lining=_rock_lining, unified=_tu_horse) if th else ([], {})
            d_au = compute_aqueduct_u(au) if au else []
            d_ar = compute_aqueduct_rect(ar) if ar else []
            d_rv = compute_rect_culvert(rv) if rv else []
            d_cp = compute_circular_pipe(cp) if cp else []
            d_sp = compute_siphon(sp) if sp else []

            data_map = {
                "rect_channel": d_rc, "trap_channel": d_tr,
                "tunnel_arch": d_ta, "tunnel_circular": d_tc,
                "tunnel_horseshoe": d_th,
                "aqueduct_u": d_au, "aqueduct_rect": d_ar,
                "rect_culvert": d_rv, "circular_channel": d_cp,
                "siphon": d_sp,
            }
            table_order = ["rect_channel", "trap_channel",
                           "tunnel_arch", "tunnel_circular", "tunnel_horseshoe",
                           "aqueduct_u", "aqueduct_rect",
                           "rect_culvert", "circular_channel", "siphon"]

            cur_y = below_y
            max_table_w = 0.0
            for key in table_order:
                d = data_map.get(key)
                builder = _DXF_BUILDERS.get(key)
                if d and builder:
                    title_t, headers, col_widths, rows, merge = builder(d)
                    if key == "tunnel_horseshoe" and d_th_info:
                        st_name = d_th_info.get("section_type_name")
                        if st_name:
                            title_t = st_name + "隧洞断面尺寸及水力要素表"
                    h = _dxf_draw_table(
                        msp, 0.0, cur_y,
                        title_t, headers, col_widths, rows,
                        merge_groups=merge, layer=_SUMM_LAYER)
                    # 计算实际表宽（与 _dxf_draw_table 内部逻辑一致）
                    auto_w = _dxf_auto_col_widths(headers, rows)
                    ncols_t = len(headers)
                    actual_w = sum(
                        max(col_widths[ci], auto_w[ci]) if ci < len(col_widths) else auto_w[ci]
                        for ci in range(ncols_t))
                    max_table_w = max(max_table_w, actual_w)
                    summary_h += h + _DXF_TABLE_GAP
                    cur_y -= (h + _DXF_TABLE_GAP)

            summary_w = max_table_w

        except ImportError:
            print("[合并DXF] 断面汇总表模块不可用，跳过")
        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"[合并DXF] 断面汇总表生成失败: {e}")

        # ======== C. IP坐标及弯道参数表（右下） ========
        try:
            ip_preview, ip_nodes = _compute_ip_preview_data(nodes, station_prefix)
            if ip_preview:
                ip_ox = max(summary_w + GAP, 200.0)
                ip_oy = below_y
                _draw_ip_table_on_msp(
                    msp, ip_ox, ip_oy, ip_preview,
                    "IP坐标及弯道参数表", _IP_LAYER)
        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"[合并DXF] IP表生成失败: {e}")

        # ---- 保存 ----
        doc.saveas(file_path)

        if fluent_question(panel.window(), "导出完成",
                f"合并DXF已生成:\n{file_path}\n\n"
                f"包含：纵断面表格 + 断面汇总表 + IP坐标表\n"
                f"布局：纵断面在上，汇总表左下，IP表右下\n\n"
                f"是否立即打开该文件？"):
            try:
                os.startfile(file_path)
            except Exception:
                pass

    except PermissionError:
        fluent_error(panel.window(), "文件被占用",
                     f"无法写入文件：\n{file_path}\n\n请先关闭该文件后重试。")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "导出失败",
                     f"合并DXF导出失败:\n{str(e)}")


# ================================================================
# 5. 断面汇总表
# ================================================================

class SectionSummaryDialog(QDialog):
    """断面尺寸及水力要素汇总表生成对话框（纯 PySide6 版）"""

    def __init__(self, parent, nodes, proj_settings, auto_name="", panel=None):
        super().__init__(parent)
        self.setWindowTitle("断面尺寸及水力要素汇总表 — 生成器")
        self.setMinimumSize(520, 560)
        self.resize(640, 780)
        self.setStyleSheet(DIALOG_STYLE)

        self._nodes = nodes
        self._proj_settings = proj_settings
        self._auto_name = auto_name
        self._panel = panel

        # 导入计算模块
        from 渠系建筑物断面计算.生成断面汇总表 import (
            _extract_segment_defaults_from_nodes,
            _segment_name,
            SIPHON_MATERIALS,
            ROCK_CLASSES,
            ROCK_LINING_DEFAULT,
            generate_excel,
            generate_dxf,
            _default_segments_rect_channel,
            _default_segments_trap_channel,
            _default_segments_tunnel_arch,
            _default_segments_tunnel_circular,
            _default_segments_tunnel_horseshoe,
            _default_segments_aqueduct_u,
            _default_segments_aqueduct_rect,
            _default_segments_rect_culvert,
            _default_segments_circular_pipe,
        )
        self._ROCK_CLASSES = ROCK_CLASSES
        self._ROCK_LINING_DEFAULT = ROCK_LINING_DEFAULT
        self._generate_excel = generate_excel
        self._generate_dxf = generate_dxf
        self._segment_name = _segment_name
        self._SIPHON_MATERIALS = SIPHON_MATERIALS
        self._default_fns = {
            'rect_channel': _default_segments_rect_channel,
            'trap_channel': _default_segments_trap_channel,
            'tunnel_arch': _default_segments_tunnel_arch,
            'tunnel_circular': _default_segments_tunnel_circular,
            'tunnel_horseshoe': _default_segments_tunnel_horseshoe,
            'aqueduct_u': _default_segments_aqueduct_u,
            'aqueduct_rect': _default_segments_aqueduct_rect,
            'rect_culvert': _default_segments_rect_culvert,
            'circular_pipe': _default_segments_circular_pipe,
        }

        node_defaults, flow_qs = _extract_segment_defaults_from_nodes(nodes)
        self._node_defaults = node_defaults
        self._flow_qs = flow_qs

        # 提取倒虹吸分组（按名称，支持不同材质和管径）
        self._siphon_groups = self._extract_siphon_groups()

        # 确定流量段数
        self._segment_count = self._get_segment_count()
        default_qs = self._build_default_qs()

        self._build_ui(default_qs)

    # ---- 流量段数计算 ----
    def _get_segment_count(self):
        counts = []
        ps = self._proj_settings
        if ps is not None and getattr(ps, "design_flows", None):
            flows = [q for q in ps.design_flows if isinstance(q, (int, float)) and q > 0]
            if flows:
                counts.append(len(flows))
        if self._flow_qs:
            counts.append(max(self._flow_qs.keys()))
        if self._node_defaults:
            for data in self._node_defaults.values():
                if data:
                    counts.append(max(data.keys()))
        return max(1, max(counts)) if counts else 7

    def _build_default_qs(self):
        fallback = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
        sc = self._segment_count
        ps = self._proj_settings
        if ps is not None and getattr(ps, "design_flows", None):
            flows = [q for q in ps.design_flows if isinstance(q, (int, float)) and q > 0]
            if flows:
                return [flows[i] if i < len(flows) else flows[-1] for i in range(sc)]
        if self._flow_qs:
            out = []
            for i in range(1, sc + 1):
                q = self._flow_qs.get(i, 0.0)
                out.append(q if q > 0 else (fallback[i - 1] if i - 1 < len(fallback) else fallback[-1]))
            return out
        if sc <= len(fallback):
            return list(fallback[:sc])
        return list(fallback) + [fallback[-1]] * (sc - len(fallback))

    def _extract_siphon_groups(self):
        """从节点中提取不同名称的倒虹吸及其默认DN

        返回: OrderedDict-like list of tuples [(siphon_display_name, dn_mm), ...]
        """
        groups = {}  # {display_name: dn_mm}
        order = []   # 保持顺序
        if not self._nodes:
            return order

        for node in self._nodes:
            if getattr(node, 'is_transition', False) or getattr(node, 'is_auto_inserted_channel', False):
                continue

            # 判断是否为倒虹吸
            is_siphon = getattr(node, 'is_inverted_siphon', False)
            if not is_siphon:
                st_str = _struct_val(getattr(node, 'structure_type', None))
                if '倒虹吸' not in st_str:
                    continue

            # 建筑物名称
            raw_name = getattr(node, 'name', '') or ''
            display_name = raw_name.strip() if raw_name.strip() else '倒虹吸'

            if display_name not in groups:
                groups[display_name] = 0
                order.append(display_name)

            # 提取 DN（优先 section_params.D，其次 structure_height）
            params = getattr(node, 'section_params', {}) or {}
            d_val = 0.0
            for key in ('D', 'd'):
                try:
                    v = float(params.get(key, 0) or 0)
                except (TypeError, ValueError):
                    v = 0.0
                if v > 0:
                    d_val = v
                    break
            if d_val <= 0:
                try:
                    d_val = float(getattr(node, 'structure_height', 0) or 0)
                except (TypeError, ValueError):
                    d_val = 0.0

            if d_val > 0:
                dn_mm = d_val * 1000 if d_val < 20 else d_val
                groups[display_name] = max(groups[display_name], dn_mm)

        return [(name, groups[name]) for name in order]

    def showEvent(self, event):
        """确保对话框不超出屏幕可见区域"""
        super().showEvent(event)
        from PySide6.QtGui import QGuiApplication
        screen = QGuiApplication.primaryScreen()
        if screen:
            avail = screen.availableGeometry()
            geo = self.frameGeometry()
            # 如果窗口高度超出屏幕，缩小到屏幕高度
            if geo.height() > avail.height():
                self.resize(self.width(), avail.height() - 20)
                geo = self.frameGeometry()
            # 如果顶部超出屏幕，向下移动
            if geo.top() < avail.top():
                geo.moveTop(avail.top())
            # 如果底部超出屏幕，向上移动
            if geo.bottom() > avail.bottom():
                geo.moveBottom(avail.bottom())
            # 如果左侧超出屏幕
            if geo.left() < avail.left():
                geo.moveLeft(avail.left())
            self.move(geo.topLeft())

    # ---- UI 构建 ----
    def _build_ui(self, default_qs):
        outer_lay = QVBoxLayout(self)
        outer_lay.setContentsMargins(0, 0, 0, 0)
        outer_lay.setSpacing(0)

        # 用 QScrollArea 包裹全部内容，防止内容超高时顶部被截断
        from PySide6.QtWidgets import QScrollArea
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        content = QWidget()
        lay = QVBoxLayout(content)
        lay.setSpacing(10)

        # 提示文字
        desc = QLabel("本功能将自动计算并生成多种建筑物断面水力要素汇总表（Excel），\n"
                      "可直接用于 AutoCAD 制表。")
        desc.setWordWrap(True)
        desc.setStyleSheet("font-size:13px; color:#333;")
        lay.addWidget(desc)

        # ---- 流量段参数 ----
        q_group = QGroupBox("流量段设计流量 Q (m³/s)")
        q_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        q_lay = QVBoxLayout(q_group)

        q_grid = QGridLayout()
        q_grid.setSpacing(4)

        self._q_edits = []
        for i in range(self._segment_count):
            lbl = QLabel(self._segment_name(i + 1))
            lbl.setStyleSheet("font-size:12px;")
            edit = LineEdit()
            edit.setFixedWidth(100)
            edit.setText(str(default_qs[i] if i < len(default_qs) else default_qs[-1]))
            q_grid.addWidget(lbl, i, 0)
            q_grid.addWidget(edit, i, 1)
            self._q_edits.append(edit)

        q_grid.setColumnStretch(2, 1)
        q_lay.addLayout(q_grid)
        lay.addWidget(q_group)

        # ---- 倒虹吸管道参数（按名称分组） ----
        siphon_group = QGroupBox("倒虹吸管道参数")
        siphon_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        siphon_lay = QVBoxLayout(siphon_group)

        # 构建倒虹吸分组列表；无数据时提供一个默认行
        siphon_items = self._siphon_groups if self._siphon_groups else [('倒虹吸', 0)]

        # 表头
        hdr_grid = QGridLayout()
        hdr_grid.setSpacing(6)
        for ci, txt in enumerate(['倒虹吸名称', '管道材质', 'DN (mm)']):
            lbl = QLabel(txt)
            lbl.setStyleSheet("font-size:11px; color:#555; font-weight:bold;")
            hdr_grid.addWidget(lbl, 0, ci)
        hdr_grid.setColumnStretch(0, 2)
        hdr_grid.setColumnStretch(1, 3)
        hdr_grid.setColumnStretch(2, 2)
        siphon_lay.addLayout(hdr_grid)

        self._siphon_rows = []  # [(name, mat_combo, dn_edit), ...]
        sp_grid = QGridLayout()
        sp_grid.setSpacing(4)
        for ri, (sp_name, sp_dn) in enumerate(siphon_items):
            # 名称标签
            name_lbl = QLabel(sp_name)
            name_lbl.setStyleSheet("font-size:12px;")
            sp_grid.addWidget(name_lbl, ri, 0)

            # 材质下拉
            mat_combo = QComboBox()
            mat_combo.addItems(list(self._SIPHON_MATERIALS.keys()))
            mat_combo.setCurrentText("球墨铸铁管")
            mat_combo.setFixedWidth(160)
            sp_grid.addWidget(mat_combo, ri, 1)

            # DN 输入框
            dn_edit = LineEdit()
            dn_edit.setFixedWidth(100)
            dn_val = int(round(sp_dn)) if sp_dn > 0 else 1500
            dn_edit.setText(str(dn_val))
            sp_grid.addWidget(dn_edit, ri, 2)

            self._siphon_rows.append((sp_name, mat_combo, dn_edit))

        sp_grid.setColumnStretch(0, 2)
        sp_grid.setColumnStretch(1, 3)
        sp_grid.setColumnStretch(2, 2)
        siphon_lay.addLayout(sp_grid)

        dn_note = QLabel("（DN 从倒虹吸计算结果自动导入，也可手动修改）")
        dn_note.setStyleSheet("font-size:11px; color:#666;")
        siphon_lay.addWidget(dn_note)
        lay.addWidget(siphon_group)

        # ---- 构造参数设置（Tab页签） ----
        from PySide6.QtWidgets import QTabWidget
        struct_group = QGroupBox("构造参数设置（壁厚/衬砌厚度）")
        struct_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        struct_lay = QVBoxLayout(struct_group)

        struct_tabs = QTabWidget()
        struct_tabs.setStyleSheet("QTabWidget{font-size:11px;} QTabBar::tab{min-width:70px;}")

        # ---- Tab 1: 明渠类 ----
        tab_channel = QWidget()
        tc_lay = QVBoxLayout(tab_channel)
        tc_lay.setSpacing(6)

        tc_grid = QGridLayout()
        tc_grid.setSpacing(4)
        tc_grid.addWidget(QLabel(""), 0, 0)  # 空占位
        for ci, txt in enumerate(['壁厚 t (m)', '拉杆尺寸 (m)']):
            lbl = QLabel(txt)
            lbl.setStyleSheet("font-size:11px; color:#555; font-weight:bold;")
            tc_grid.addWidget(lbl, 0, ci + 1)

        def _tie_rod_pair(default_w=0.2, default_h=0.2):
            """创建拉杆尺寸 [宽] × [高] 组合控件，返回 (container, w_edit, h_edit)。"""
            container = QWidget()
            h_lay = QHBoxLayout(container)
            h_lay.setContentsMargins(0, 0, 0, 0)
            h_lay.setSpacing(3)
            w_lbl = QLabel("宽"); w_lbl.setStyleSheet("font-size:10px; color:#888;")
            w_edit = LineEdit(); w_edit.setFixedWidth(55)
            w_edit.setText(str(default_w)); w_edit.setPlaceholderText(str(default_w))
            x_lbl = QLabel("×"); x_lbl.setFixedWidth(12)
            x_lbl.setStyleSheet("font-size:12px;")
            h_lbl = QLabel("高"); h_lbl.setStyleSheet("font-size:10px; color:#888;")
            h_edit = LineEdit(); h_edit.setFixedWidth(55)
            h_edit.setText(str(default_h)); h_edit.setPlaceholderText(str(default_h))
            h_lay.addWidget(w_lbl)
            h_lay.addWidget(w_edit)
            h_lay.addWidget(x_lbl)
            h_lay.addWidget(h_lbl)
            h_lay.addWidget(h_edit)
            h_lay.addStretch()
            return container, w_edit, h_edit

        # 矩形明渠
        tc_grid.addWidget(QLabel("矩形明渠"), 1, 0)
        self._rect_ch_wall_t = LineEdit(); self._rect_ch_wall_t.setFixedWidth(90)
        self._rect_ch_wall_t.setText("0.3"); self._rect_ch_wall_t.setPlaceholderText("0.3")
        tc_grid.addWidget(self._rect_ch_wall_t, 1, 1)
        rc_tr_container, self._rect_ch_tie_w, self._rect_ch_tie_h = _tie_rod_pair()
        tc_grid.addWidget(rc_tr_container, 1, 2)

        # 梯形明渠
        tc_grid.addWidget(QLabel("梯形明渠"), 2, 0)
        self._trap_ch_wall_t = LineEdit(); self._trap_ch_wall_t.setFixedWidth(90)
        self._trap_ch_wall_t.setText("0.3"); self._trap_ch_wall_t.setPlaceholderText("0.3")
        tc_grid.addWidget(self._trap_ch_wall_t, 2, 1)
        tp_tr_container, self._trap_ch_tie_w, self._trap_ch_tie_h = _tie_rod_pair()
        tc_grid.addWidget(tp_tr_container, 2, 2)

        tc_grid.setColumnStretch(0, 1)
        tc_grid.setColumnStretch(1, 2)
        tc_grid.setColumnStretch(2, 3)
        tc_lay.addLayout(tc_grid)
        tc_lay.addStretch()
        struct_tabs.addTab(tab_channel, "明渠类")

        # ---- Tab 2: 渡槽类 ----
        tab_aqueduct = QWidget()
        ta_lay = QVBoxLayout(tab_aqueduct)
        ta_lay.setSpacing(6)

        ta_grid = QGridLayout()
        ta_grid.setSpacing(4)
        ta_grid.addWidget(QLabel(""), 0, 0)
        lbl_t = QLabel("壁厚 t (m)")
        lbl_t.setStyleSheet("font-size:11px; color:#555; font-weight:bold;")
        ta_grid.addWidget(lbl_t, 0, 1)

        # U形渡槽
        ta_grid.addWidget(QLabel("U形渡槽"), 1, 0)
        self._aq_u_wall_t = LineEdit(); self._aq_u_wall_t.setFixedWidth(90)
        self._aq_u_wall_t.setText("0.35"); self._aq_u_wall_t.setPlaceholderText("0.35")
        ta_grid.addWidget(self._aq_u_wall_t, 1, 1)

        # 矩形渡槽
        ta_grid.addWidget(QLabel("矩形渡槽"), 2, 0)
        self._aq_rect_wall_t = LineEdit(); self._aq_rect_wall_t.setFixedWidth(90)
        self._aq_rect_wall_t.setText("0.35"); self._aq_rect_wall_t.setPlaceholderText("0.35")
        ta_grid.addWidget(self._aq_rect_wall_t, 2, 1)

        ta_grid.setColumnStretch(0, 1)
        ta_grid.setColumnStretch(1, 2)
        ta_lay.addLayout(ta_grid)
        ta_lay.addStretch()
        struct_tabs.addTab(tab_aqueduct, "渡槽类")

        # ---- Tab 3: 暗涵 ----
        tab_culvert = QWidget()
        tv_lay = QVBoxLayout(tab_culvert)
        tv_lay.setSpacing(6)

        tv_grid = QGridLayout()
        tv_grid.setSpacing(4)
        tv_grid.addWidget(QLabel(""), 0, 0)
        for ci, txt in enumerate(['底板厚 t\u2080 (m)', '边墙厚 t\u2081 (m)', '顶板厚 t\u2082 (m)']):
            lbl = QLabel(txt)
            lbl.setStyleSheet("font-size:11px; color:#555; font-weight:bold;")
            tv_grid.addWidget(lbl, 0, ci + 1)

        tv_grid.addWidget(QLabel("矩形暗涵"), 1, 0)
        self._culvert_t0 = LineEdit(); self._culvert_t0.setFixedWidth(90)
        self._culvert_t0.setText("0.4"); self._culvert_t0.setPlaceholderText("0.4")
        tv_grid.addWidget(self._culvert_t0, 1, 1)
        self._culvert_t1 = LineEdit(); self._culvert_t1.setFixedWidth(90)
        self._culvert_t1.setText("0.4"); self._culvert_t1.setPlaceholderText("0.4")
        tv_grid.addWidget(self._culvert_t1, 1, 2)
        self._culvert_t2 = LineEdit(); self._culvert_t2.setFixedWidth(90)
        self._culvert_t2.setText("0.4"); self._culvert_t2.setPlaceholderText("0.4")
        tv_grid.addWidget(self._culvert_t2, 1, 3)

        tv_grid.setColumnStretch(0, 1)
        tv_grid.setColumnStretch(1, 2)
        tv_grid.setColumnStretch(2, 2)
        tv_grid.setColumnStretch(3, 2)
        tv_lay.addLayout(tv_grid)
        tv_lay.addStretch()
        struct_tabs.addTab(tab_culvert, "暗涵")

        # ---- Tab 4: 隧洞 ----
        tab_tunnel = QWidget()
        tt_lay = QVBoxLayout(tab_tunnel)
        tt_lay.setSpacing(6)

        tt_desc = QLabel("4种隧洞类型共用此设置（圆拱直墙型/圆形/马蹄形Ⅰ型/Ⅱ型）")
        tt_desc.setStyleSheet("font-size:11px; color:#666;")
        tt_lay.addWidget(tt_desc)

        tt_grid = QGridLayout()
        tt_grid.setSpacing(4)
        tt_grid.addWidget(QLabel(""), 0, 0)
        for ci, txt in enumerate(['底板厚 t\u2080 (m)', '边墙/顶拱/衬砌厚 t (m)']):
            lbl = QLabel(txt)
            lbl.setStyleSheet("font-size:11px; color:#555; font-weight:bold;")
            tt_grid.addWidget(lbl, 0, ci + 1)

        self._lining_edits = {}  # {rock_class: (t0_edit, t_edit)}
        for ri, rc in enumerate(self._ROCK_CLASSES):
            tt_grid.addWidget(QLabel(rc), ri + 1, 0)
            defaults = self._ROCK_LINING_DEFAULT[rc]
            t0_edit = LineEdit(); t0_edit.setFixedWidth(90)
            t0_edit.setText(str(defaults['t0'])); t0_edit.setPlaceholderText(str(defaults['t0']))
            tt_grid.addWidget(t0_edit, ri + 1, 1)
            t_edit = LineEdit(); t_edit.setFixedWidth(90)
            t_edit.setText(str(defaults['t'])); t_edit.setPlaceholderText(str(defaults['t']))
            tt_grid.addWidget(t_edit, ri + 1, 2)
            self._lining_edits[rc] = (t0_edit, t_edit)

        tt_grid.setColumnStretch(0, 1)
        tt_grid.setColumnStretch(1, 3)
        tt_grid.setColumnStretch(2, 3)
        tt_lay.addLayout(tt_grid)

        # ---- 隧洞断面设计方式 ----
        from PySide6.QtWidgets import QRadioButton, QButtonGroup, QHBoxLayout as _QHBox
        _tt_mode_row = QWidget()
        _tt_mode_hlay = _QHBox(_tt_mode_row)
        _tt_mode_hlay.setContentsMargins(0, 0, 0, 0)
        _tt_mode_hlay.setSpacing(4)
        tt_mode_lbl = QLabel("断面设计方式:")
        tt_mode_lbl.setStyleSheet("font-size:11px; color:#555; font-weight:bold; margin-top:6px;")
        _tt_mode_hlay.addWidget(tt_mode_lbl)
        _info_icon = QLabel("ⓘ")
        _info_icon.setStyleSheet(
            "font-size:13px; color:#1a73e8; font-weight:bold; margin-top:6px; cursor:pointer;"
        )
        _info_icon.setToolTip(
            "<b>统一断面</b>：按最大流量段设计统一断面尺寸，其余各流量段仅推求水深；<br>"
            "<b>独立断面</b>：每个流量段独立计算各自的断面尺寸。"
        )
        _tt_mode_hlay.addWidget(_info_icon)
        _tt_mode_hlay.addStretch()
        tt_lay.addWidget(_tt_mode_row)

        self._tunnel_mode_groups = {}  # {key: QButtonGroup}
        _tunnel_types = [
            ("tunnel_arch",      "圆拱直墙型"),
            ("tunnel_circular",  "圆形"),
            ("tunnel_horseshoe", "马蹄形（Ⅰ/Ⅱ型）"),
        ]
        tm_grid = QGridLayout()
        tm_grid.setSpacing(2)
        for ri, (tkey, tname) in enumerate(_tunnel_types):
            name_lbl = QLabel(tname)
            name_lbl.setStyleSheet("font-size:11px;")
            name_lbl.setFixedWidth(110)
            tm_grid.addWidget(name_lbl, ri, 0)
            rb_unified = QRadioButton("统一断面")
            rb_indep  = QRadioButton("独立断面")
            rb_unified.setStyleSheet("font-size:11px;")
            rb_indep.setStyleSheet("font-size:11px;")
            rb_indep.setChecked(True)
            bg = QButtonGroup(self)
            bg.addButton(rb_unified, 0)
            bg.addButton(rb_indep, 1)
            tm_grid.addWidget(rb_unified, ri, 1)
            tm_grid.addWidget(rb_indep, ri, 2)
            self._tunnel_mode_groups[tkey] = bg
        tt_lay.addLayout(tm_grid)

        tt_lay.addStretch()
        struct_tabs.addTab(tab_tunnel, "隧洞")

        struct_tabs.setFixedHeight(260)
        struct_lay.addWidget(struct_tabs)

        struct_note = QLabel('（不输入则使用默认值，修改后同时影响"生成断面汇总表"和"导出全部DXF"）')
        struct_note.setStyleSheet("font-size:11px; color:#666;")
        struct_lay.addWidget(struct_note)
        lay.addWidget(struct_group)

        # ---- 说明 ----
        note_group = QGroupBox("其他参数说明")
        note_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        note_lay = QVBoxLayout(note_group)
        note_lbl = QLabel(
            "• 各类构造参数可在上方按类型自定义\n"
            '• 隧洞断面设计方式可在"隧洞"选项卡中按类型分别设置\n'
            "• 圆管涵、倒虹吸无需设置壁厚")
        note_lbl.setWordWrap(True)
        note_lbl.setStyleSheet("font-size:11px; color:#555;")
        note_lay.addWidget(note_lbl)
        lay.addWidget(note_group)

        # ---- 导出格式选择 ----
        from PySide6.QtWidgets import QRadioButton, QButtonGroup
        fmt_group = QGroupBox("导出格式")
        fmt_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        fmt_lay = QHBoxLayout(fmt_group)
        self._radio_excel = QRadioButton("Excel (.xlsx)  — 多Sheet + 汇总Sheet")
        self._radio_dxf = QRadioButton("DXF (.dxf)  — 可直接导入AutoCAD")
        self._radio_excel.setStyleSheet("font-size:11px;")
        self._radio_dxf.setStyleSheet("font-size:11px;")
        self._radio_excel.setChecked(True)
        self._fmt_btn_group = QButtonGroup(self)
        self._fmt_btn_group.addButton(self._radio_excel, 0)
        self._fmt_btn_group.addButton(self._radio_dxf, 1)
        fmt_lay.addWidget(self._radio_excel)
        fmt_lay.addWidget(self._radio_dxf)
        fmt_lay.addStretch()
        lay.addWidget(fmt_group)

        scroll.setWidget(content)
        outer_lay.addWidget(scroll, 1)

        # ---- 按钮栏（固定在底部，不随滚动） ----
        btn_lay = QHBoxLayout()
        btn_lay.setContentsMargins(10, 6, 10, 6)
        btn_lay.addStretch()
        btn_cancel = PushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        btn_generate = PrimaryPushButton("生成汇总表")
        btn_generate.clicked.connect(self._on_generate)
        btn_lay.addWidget(btn_cancel)
        btn_lay.addWidget(btn_generate)
        outer_lay.addLayout(btn_lay)

    # ---- 读取构造参数 ----
    def _read_float(self, edit, default):
        """安全读取 LineEdit 的浮点值，空或非法返回默认值。"""
        t = edit.text().strip()
        if not t:
            return default
        try:
            return float(t)
        except ValueError:
            return default

    def _read_rock_lining(self):
        """从输入框读取用户自定义的围岩衬砌厚度。"""
        rock_lining = {}
        for rc in self._ROCK_CLASSES:
            t0_edit, t_edit = self._lining_edits[rc]
            defaults = self._ROCK_LINING_DEFAULT[rc]
            rock_lining[rc] = {
                't0': self._read_float(t0_edit, defaults['t0']),
                't':  self._read_float(t_edit,  defaults['t']),
            }
        return rock_lining

    def _read_tie_rod(self, w_edit, h_edit):
        """从拉杆宽/高输入框读取并组合为 'd1×d2' 字符串。"""
        w = self._read_float(w_edit, 0.2)
        h = self._read_float(h_edit, 0.2)
        return f"{w}×{h}"

    def _read_struct_thickness(self):
        """读取所有结构类型的用户自定义厚度参数，返回 dict。"""
        return {
            'rect_channel': {
                'wall_t':  self._read_float(self._rect_ch_wall_t, 0.3),
                'tie_rod': self._read_tie_rod(self._rect_ch_tie_w, self._rect_ch_tie_h),
            },
            'trap_channel': {
                'wall_t':  self._read_float(self._trap_ch_wall_t, 0.3),
                'tie_rod': self._read_tie_rod(self._trap_ch_tie_w, self._trap_ch_tie_h),
            },
            'aqueduct_u': {
                'wall_t': self._read_float(self._aq_u_wall_t, 0.35),
            },
            'aqueduct_rect': {
                'wall_t': self._read_float(self._aq_rect_wall_t, 0.35),
            },
            'rect_culvert': {
                't0': self._read_float(self._culvert_t0, 0.4),
                't1': self._read_float(self._culvert_t1, 0.4),
                't2': self._read_float(self._culvert_t2, 0.4),
            },
            'rock_lining': self._read_rock_lining(),
        }

    # ---- 生成 ----
    def _on_generate(self):
        from 渠系建筑物断面计算.生成断面汇总表 import (
            _default_segments_rect_channel,
            _default_segments_trap_channel,
            _default_segments_tunnel_arch,
            _default_segments_tunnel_circular,
            _default_segments_tunnel_horseshoe,
            _default_segments_aqueduct_u,
            _default_segments_aqueduct_rect,
            _default_segments_rect_culvert,
            _default_segments_circular_pipe,
            _segment_name,
        )

        # 读取 Q 值
        try:
            qs = [float(e.text()) for e in self._q_edits]
        except ValueError:
            fluent_error(self, "输入错误", "流量值必须为数字")
            return

        # 读取每个倒虹吸的材质和DN
        siphon_params = []  # [(name, material, dn), ...]
        for sp_name, mat_combo, dn_edit in self._siphon_rows:
            try:
                dn = int(dn_edit.text())
            except ValueError:
                fluent_error(self, "输入错误", f"{sp_name} 的 DN 必须为整数")
                return
            siphon_params.append((sp_name, mat_combo.currentText(), dn))

        segment_count = self._segment_count
        node_defaults = self._node_defaults

        # 判断导出格式
        export_dxf = self._radio_dxf.isChecked()
        if export_dxf:
            ext = ".dxf"
            filter_str = "DXF 文件 (*.dxf);;所有文件 (*.*)"
            auto_name = self._auto_name.replace('.xlsx', '.dxf') if self._auto_name else ""
        else:
            ext = ".xlsx"
            filter_str = "Excel 文件 (*.xlsx);;所有文件 (*.*)"
            auto_name = self._auto_name

        # 选择保存路径
        fp, _ = QFileDialog.getSaveFileName(
            self, "保存断面汇总表", auto_name, filter_str)
        if not fp:
            return
        if not fp.lower().endswith(ext):
            fp += ext

        # 构建各表参数
        has_source_data = bool(self._nodes) and any(node_defaults.values())

        def _make_segs(default_fn, overrides_by_idx=None):
            # 有源数据时，只生成有实际节点数据的流量段
            if has_source_data and overrides_by_idx is not None:
                if not overrides_by_idx:
                    return []
                defaults_pool = default_fn()
                segs = []
                for idx in sorted(overrides_by_idx.keys()):
                    # 用默认段作为基础模板
                    base = dict(defaults_pool[0]) if defaults_pool else {}
                    base["name"] = _segment_name(idx)
                    base.update(overrides_by_idx[idx])
                    if 0 < idx <= len(qs):
                        base["Q"] = qs[idx - 1]
                    segs.append(base)
                return segs
            # 无源数据时（独立运行），用默认值生成所有段
            segs = default_fn()
            if len(segs) < segment_count:
                last = segs[-1] if segs else {}
                for idx in range(len(segs) + 1, segment_count + 1):
                    new_seg = dict(last)
                    new_seg["name"] = _segment_name(idx)
                    segs.append(new_seg)
            segs = segs[:segment_count]
            for i, seg in enumerate(segs):
                if overrides_by_idx and (i + 1) in overrides_by_idx:
                    seg.update(overrides_by_idx[i + 1])
                if i < len(qs):
                    seg["Q"] = qs[i]
            return segs

        rc_segs = _make_segs(_default_segments_rect_channel, node_defaults.get("rect_channel"))
        tr_segs = _make_segs(_default_segments_trap_channel, node_defaults.get("trap_channel"))
        tn_arch_segs = _make_segs(_default_segments_tunnel_arch, node_defaults.get("tunnel_arch"))
        tn_circ_segs = _make_segs(_default_segments_tunnel_circular, node_defaults.get("tunnel_circular"))
        tn_horse_segs = _make_segs(_default_segments_tunnel_horseshoe, node_defaults.get("tunnel_horseshoe"))
        aq_u_segs = _make_segs(_default_segments_aqueduct_u, node_defaults.get("aqueduct_u"))
        aq_rect_segs = _make_segs(_default_segments_aqueduct_rect, node_defaults.get("aqueduct_rect"))
        rv_segs = _make_segs(_default_segments_rect_culvert, node_defaults.get("rect_culvert"))
        cp_segs = _make_segs(_default_segments_circular_pipe, node_defaults.get("circular_channel"))

        if not has_source_data:
            for segs_list in [rc_segs, tr_segs, tn_arch_segs, tn_circ_segs, tn_horse_segs,
                              aq_u_segs, aq_rect_segs, rv_segs, cp_segs]:
                for i, seg in enumerate(segs_list):
                    seg["name"] = _segment_name(i + 1)

        sp_overrides = node_defaults.get("siphon", {})
        sp_segs = []
        multi_siphon = len(siphon_params) > 1
        # 有源数据时只生成有实际数据的流量段
        sp_indices = sorted(sp_overrides.keys()) if (has_source_data and sp_overrides) else list(range(1, len(qs) + 1))
        for sp_name, sp_material, sp_dn in siphon_params:
            for idx in sp_indices:
                seg_name = (f"{sp_name}-{_segment_name(idx)}"
                            if multi_siphon else _segment_name(idx))
                seg = {"name": seg_name}
                if idx in sp_overrides:
                    seg.update(sp_overrides[idx])
                seg["DN_mm"] = sp_dn
                if 0 < idx <= len(qs):
                    seg["Q"] = qs[idx - 1]
                seg["pipe_material"] = sp_material
                sp_segs.append(seg)

        # 按结果决定表格类型
        _table_order = None
        if has_source_data:
            _table_order = []
            if node_defaults.get("rect_channel"):
                _table_order.append("rect_channel")
            if node_defaults.get("trap_channel"):
                _table_order.append("trap_channel")
            if node_defaults.get("tunnel_arch"):
                _table_order.append("tunnel_arch")
            if node_defaults.get("tunnel_circular"):
                _table_order.append("tunnel_circular")
            if node_defaults.get("tunnel_horseshoe"):
                _table_order.append("tunnel_horseshoe")
            if node_defaults.get("aqueduct_u"):
                _table_order.append("aqueduct_u")
            if node_defaults.get("aqueduct_rect"):
                _table_order.append("aqueduct_rect")
            if node_defaults.get("rect_culvert"):
                _table_order.append("rect_culvert")
            if node_defaults.get("circular_channel"):
                _table_order.append("circular_channel")
            if node_defaults.get("siphon"):
                _table_order.append("siphon")
            if not _table_order:
                _table_order = None

        # 读取所有用户自定义的构造参数（壁厚/衬砌厚度）
        struct_t = self._read_struct_thickness()
        rock_lining = struct_t['rock_lining']

        # 将壁厚/衬砌参数注入各类型 segments
        for seg in rc_segs:
            seg['wall_t'] = struct_t['rect_channel']['wall_t']
            seg['tie_rod'] = struct_t['rect_channel']['tie_rod']
        for seg in tr_segs:
            seg['wall_t'] = struct_t['trap_channel']['wall_t']
            seg['tie_rod'] = struct_t['trap_channel']['tie_rod']
        for seg in aq_u_segs:
            seg['wall_t'] = struct_t['aqueduct_u']['wall_t']
        for seg in aq_rect_segs:
            seg['wall_t'] = struct_t['aqueduct_rect']['wall_t']
        for seg in rv_segs:
            seg['t0'] = struct_t['rect_culvert']['t0']
            seg['t1'] = struct_t['rect_culvert']['t1']
            seg['t2'] = struct_t['rect_culvert']['t2']

        # 读取隧洞断面设计方式
        tunnel_unified = {}
        for tkey, bg in self._tunnel_mode_groups.items():
            tunnel_unified[tkey] = (bg.checkedId() == 0)  # 0=统一, 1=独立

        # 存储到 panel，供"导出全部DXF"复用
        if self._panel is not None:
            self._panel._custom_rock_lining = rock_lining
            self._panel._custom_struct_thickness = struct_t
            self._panel._custom_tunnel_unified = tunnel_unified

        gen_kwargs = dict(
            filepath=fp,
            rect_channel_segs=rc_segs,
            trap_channel_segs=tr_segs,
            tunnel_arch_segs=tn_arch_segs,
            tunnel_circular_segs=tn_circ_segs,
            tunnel_horseshoe_segs=tn_horse_segs,
            aqueduct_u_segs=aq_u_segs,
            aqueduct_rect_segs=aq_rect_segs,
            rect_culvert_segs=rv_segs,
            circular_pipe_segs=cp_segs,
            siphon_segs=sp_segs,
            siphon_material=siphon_params[0][1] if siphon_params else "球墨铸铁管",
            rock_lining=rock_lining,
            table_order=_table_order,
            tunnel_unified_arch=tunnel_unified.get("tunnel_arch", False),
            tunnel_unified_circular=tunnel_unified.get("tunnel_circular", False),
            tunnel_unified_horseshoe=tunnel_unified.get("tunnel_horseshoe", False),
        )

        try:
            self.setCursor(Qt.WaitCursor)
            QApplication.processEvents()
            if export_dxf:
                self._generate_dxf(**gen_kwargs)
            else:
                self._generate_excel(**gen_kwargs)
            self.unsetCursor()
            mat_summary = '、'.join(f"{n}({m})" for n, m, _ in siphon_params)
            fmt_name = "DXF" if export_dxf else "Excel"
            extra = "" if export_dxf else "\n表格数量以计算结果为准，另含 1 个汇总 Sheet。"
            if fluent_question(self, "完成",
                        f"断面汇总表已生成（{fmt_name}）：\n{fp}\n{extra}\n"
                        f"倒虹吸管道材质: {mat_summary}\n\n"
                        f"是否立即打开该文件？",
                        yes_text="打开", no_text="关闭"):
                try:
                    os.startfile(fp)
                except Exception:
                    pass
            self.accept()
        except PermissionError:
            self.unsetCursor()
            fluent_error(self, "文件被占用",
                         f"无法写入文件，该文件可能已被其他程序打开：\n\n{fp}\n\n"
                         f"请先关闭该文件，然后重新操作。")
        except Exception as e:
            self.unsetCursor()
            import traceback; traceback.print_exc()
            fluent_error(self, "生成失败", f"错误: {e}")


def open_section_summary_table(panel):
    """打开断面汇总表生成器（纯 PySide6 对话框）"""
    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "没有数据可用，请先执行计算。")
        return

    try:
        proj_settings = panel._build_settings()
    except Exception:
        proj_settings = panel._settings

    try:
        try:
            ch_name = panel.channel_name_edit.text().strip()
            ch_level = panel.channel_level_combo.currentText()
            auto_name = f"{ch_name}{ch_level}_断面汇总表.xlsx"
        except Exception:
            auto_name = "断面汇总表.xlsx"

        dlg = SectionSummaryDialog(panel.window(), nodes, proj_settings, auto_name, panel=panel)
        dlg.exec()
    except ImportError as e:
        fluent_error(
            panel.window(), "功能不可用",
            f"断面汇总表模块加载失败：\n{str(e)}")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "打开失败",
                     f"断面汇总表生成器打开失败：\n{str(e)}")
