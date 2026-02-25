# -*- coding: utf-8 -*-
"""
多流量段批量水力计算面板 —— QWidget 版本

支持：所有断面类型批量输入、计算、结果汇总、导出
功能：表格输入、新增/删除/复制行、批量计算、结果汇总表、详细过程、导出Excel/TXT
"""

import sys
import os
import math
import re
import datetime

_pkg_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(_pkg_root, "渠系建筑物断面计算"))

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QGroupBox,
    QSplitter, QFrame, QTabWidget, QTextEdit, QFileDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QAbstractItemView, QApplication, QMenu,
    QDialog, QDialogButtonBox, QGridLayout, QFormLayout,
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QColor, QShortcut, QKeySequence

from qfluentwidgets import (
    PushButton, PrimaryPushButton, CheckBox, InfoBar, InfoBarPosition,
    LineEdit, ComboBox,
)

from 渠系断面设计.styles import P, S, W, E, BG, CARD, BD, T1, T2, auto_resize_table, DIALOG_STYLE, fluent_info, fluent_error, fluent_question
from 渠系断面设计.frozen_table import FrozenColumnTableWidget
from 渠系断面设计.export_utils import (
    WORD_EXPORT_AVAILABLE, ask_open_file,
    create_styled_doc, doc_add_h1, doc_add_h2,
    doc_add_body, doc_add_styled_table, doc_add_table_caption,
    doc_add_param_table, doc_render_calc_text, update_doc_toc_via_com,
)


def format_station_display(value: float) -> str:
    """将桩号数值格式化为标准显示格式（如 12+111.222）"""
    if value < 0:
        value = abs(value)
    km = int(value // 1000)
    meters = value % 1000
    return f"{km}+{meters:07.3f}"


def parse_station_input(input_str: str) -> float:
    """解析桩号输入字符串为数值，支持纯数字和带加号格式"""
    input_str = str(input_str).strip()
    if not input_str:
        return 0.0
    if '+' in input_str:
        parts = input_str.split('+')
        if len(parts) == 2:
            try:
                km = int(parts[0])
                meters = float(parts[1])
                return km * 1000 + meters
            except ValueError:
                pass
    try:
        return float(input_str)
    except ValueError:
        return 0.0

from 渠系断面设计.structure_type_selector import StructureTypeSelector

# 共享数据管理器（用于与推求水面线模块交互）
_water_profile_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), '推求水面线')
if _water_profile_dir not in sys.path:
    sys.path.insert(0, _water_profile_dir)
try:
    from shared.shared_data_manager import get_shared_data_manager
    SHARED_DATA_AVAILABLE = True
except ImportError:
    SHARED_DATA_AVAILABLE = False

# 计算引擎导入
try:
    from 明渠设计 import (
        quick_calculate as mingqu_calculate,
        quick_calculate_circular as circular_calculate,
        quick_calculate_u_section as mingqu_u_calculate
    )
    MINGQU_AVAILABLE = True
except ImportError:
    MINGQU_AVAILABLE = False

try:
    from 渡槽设计 import (
        quick_calculate_u as ducao_u_calculate,
        quick_calculate_rect as ducao_rect_calculate
    )
    DUCAO_AVAILABLE = True
except ImportError:
    DUCAO_AVAILABLE = False

try:
    from 隧洞设计 import (
        quick_calculate_circular as suidong_circular_calculate,
        quick_calculate_horseshoe as suidong_horseshoe_calculate,
        quick_calculate_horseshoe_std as suidong_horseshoe_std_calculate
    )
    SUIDONG_AVAILABLE = True
except ImportError:
    SUIDONG_AVAILABLE = False

try:
    from 矩形暗涵设计 import (
        quick_calculate_rectangular_culvert as rect_culvert_calculate
    )
    RECT_CULVERT_AVAILABLE = True
except ImportError:
    RECT_CULVERT_AVAILABLE = False

# 断面类型列表
SECTION_TYPES = [
    "明渠-梯形", "明渠-矩形", "明渠-圆形", "明渠-U形",
    "渡槽-U形", "渡槽-矩形",
    "隧洞-圆形", "隧洞-圆拱直墙型", "隧洞-马蹄形Ⅰ型", "隧洞-马蹄形Ⅱ型",
    "矩形暗涵", "倒虹吸",
    "分水闸", "分水口", "节制闸", "泄水闸",
]

# 输入表列定义（与原版一致，含X/Y坐标列）
# 列索引: 0序号, 1流量段, 2建筑物名称, 3结构形式, 4X, 5Y, 6Q, 7糙率n, 8比降,
#          9边坡系数m, 10底宽B, 11明渠宽深比, 12半径R, 13直径D,
#          14矩形渡槽深宽比, 15倒角角度, 16倒角底边, 17圆心角, 18不淤流速, 19不冲流速
INPUT_HEADERS = [
    "序号", "流量段", "建筑物名称", "结构形式", "X", "Y",
    "Q(m³/s)", "糙率n", "比降(1/)",
    "边坡系数m", "底宽B(m)", "明渠宽深比",
    "半径R(m)", "直径D(m)",
    "矩形渡槽深宽比", "倒角角度(°)", "倒角底边(m)", "圆心角(°)",
    "不淤流速", "不冲流速",
]

# 输入表头悬浮提示（与原版一致）
_HEADER_TOOLTIPS = {
    "底宽B(m)": (
        "【底宽 B（单位：米）】\n\n"
        "定义：渠道或建筑物断面的底部宽度\n\n"
        "▶ 怎么填？\n"
        "  • 留空 → 程序自动搜索最优底宽\n"
        "  • 填一个数值 → 固定底宽，由程序推算水深\n\n"
        "▶ 哪些类型用到这个参数？\n"
        "  • 明渠-梯形 / 明渠-矩形：渠道底宽\n"
        "  • 渡槽-矩形：槽宽\n"
        "  • 隧洞-圆拱直墙型：洞宽\n"
        "  • 矩形暗涵：涵洞宽度\n\n"
        "▶ 与「明渠宽深比」的关系\n"
        "  明渠类型下，底宽B 与宽深比β 二选一填写，\n"
        "  都留空则程序自动寻优计算"
    ),
    "半径R(m)": (
        "【半径 R（单位：米）】\n\n"
        "定义：断面圆弧部分的半径\n\n"
        "▶ 怎么填？\n"
        "  • 留空 → 程序自动搜索最优半径\n"
        "  • 填一个数值 → 固定半径进行计算\n\n"
        "▶ 哪些类型用到这个参数？\n"
        "  • 渡槽-U形：底部半圆的半径 R\n"
        "  • 隧洞-马蹄形Ⅰ型 / Ⅱ型：内轮廓基准半径 r\n\n"
        "▶ 其他类型无需填写\n"
        "  明渠、圆形隧洞、圆拱直墙型、矩形暗涵\n"
        "  等类型不使用此参数"
    ),
    "直径D(m)": (
        "【直径 D（单位：米）】\n\n"
        "定义：圆形断面的管径 / 洞径\n\n"
        "▶ 怎么填？\n"
        "  • 留空 → 程序自动搜索最优直径\n"
        "  • 填一个数值 → 固定直径进行计算\n\n"
        "▶ 哪些类型用到这个参数？\n"
        "  • 明渠-圆形：圆形渠道的管径\n"
        "  • 隧洞-圆形：圆形隧洞的洞径\n\n"
        "▶ 规范要求\n"
        "  隧洞最小直径一般不小于 2.0m\n\n"
        "▶ 其他类型无需填写\n"
        "  梯形/矩形明渠、渡槽、马蹄形隧洞等\n"
        "  不使用此参数"
    ),
    "明渠宽深比": (
        "【明渠宽深比  β = B ÷ h】\n\n"
        "定义：渠道底宽 B 与设计水深 h 的比值\n"
        "公式：β = B / h\n\n"
        "▶ 怎么填？\n"
        "  • 留空 → 程序自动寻优，找到最经济的断面尺寸\n"
        "  • 填一个数值 → 固定底宽与水深的比例关系\n"
        "  • 与「底宽B(m)」二选一填写即可，都留空也行\n\n"
        "▶ 数值越大越小有什么区别？\n"
        "  • β 大（如 4~8）→ 渠道宽而浅，适合大流量、地形平坦\n"
        "  • β 小（如 1~2）→ 渠道窄而深，水力效率高、占地少\n\n"
        "▶ 适用范围\n"
        "  仅对「明渠-梯形」「明渠-矩形」类型生效"
    ),
    "倒角角度(°)": (
        "【倒角角度（单位：度）】\n\n"
        "定义：矩形渡槽底部两侧倒角的斜面与水平面的夹角；也用于明渠-U形的直线段外倾角α（°）\n\n"
        "▶ 怎么填？\n"
        "  • 留空或填 0 → 不设倒角，按纯矩形断面计算\n"
        "  • 填一个角度值 → 在底部两角切出斜面\n"
        "  • 需与「倒角底边」配合使用，两者同时填写才生效\n\n"
        "▶ 倒角有什么用？\n"
        "  减小底部直角处的应力集中，改善结构受力，\n"
        "  同时减小过水断面面积，会略微增大流速\n\n"
        "▶ 适用范围\n"
        "  仅对「渡槽-矩形」类型生效"
    ),
    "倒角底边(m)": (
        "【倒角底边（单位：米）】\n\n"
        "定义：矩形渡槽底部倒角在水平方向上的投影长度\n\n"
        "▶ 怎么填？\n"
        "  • 留空或填 0 → 不设倒角\n"
        "  • 填一个长度值 → 确定倒角的水平尺寸\n"
        "  • 需与「倒角角度」配合使用，两者同时填写才生效\n\n"
        "▶ 倒角高度如何确定？\n"
        "  倒角高度 = 倒角底边 × tan(倒角角度)\n"
        "  例如：底边 0.1m、角度 45° → 倒角高度 0.1m\n\n"
        "▶ 适用范围\n"
        "  仅对「渡槽-矩形」类型生效"
    ),
    "圆心角(°)": (
        "【圆心角  θ（单位：度）】\n\n"
        "定义：圆拱直墙型隧洞拱顶圆弧所对应的圆心角\n\n"
        "▶ 怎么填？\n"
        "  • 留空 → 默认按 180° 计算（即半圆拱）\n"
        "  • 填一个数值 → 按此角度确定拱顶形状\n\n"
        "▶ 取值范围：90° ~ 180°\n"
        "  • θ = 180° → 拱顶为半圆形，拱高最大\n"
        "  • θ = 120° → 拱顶弧度较平缓，拱高较低\n"
        "  • θ 越小 → 拱顶越平，直墙段越高\n\n"
        "▶ 适用范围\n"
        "  仅对「隧洞-圆拱直墙型」类型生效，\n"
        "  其他隧洞类型（圆形、马蹄形）无需填写"
    ),
    "矩形渡槽深宽比": (
        "【矩形渡槽深宽比  α = H ÷ B】\n\n"
        "定义：渡槽槽身总高度 H 与槽宽 B 的比值\n"
        "公式：α = H / B\n\n"
        "▶ 怎么填？\n"
        "  • 留空 → 默认按 0.8 计算\n"
        "  • 填一个数值 → 按此比例约束槽高与槽宽\n\n"
        "▶ 推荐取值：0.6 ~ 0.8\n"
        "  • α 大（如 0.8）→ 槽身偏深窄，结构刚度好\n"
        "  • α 小（如 0.6）→ 槽身偏浅宽，水面更开阔\n\n"
        "▶ 适用范围\n"
        "  仅对「渡槽-矩形」类型生效\n\n"
        "▶ 与明渠宽深比的区别\n"
        "  • 明渠宽深比：β = B / h，其中 h 为设计水深\n"
        "  • 渡槽深宽比：α = H / B，其中 H 为槽身总高度\n"
        "  两者的「深」含义不同，不要混淆"
    ),
}

# 结果表列定义
RESULT_HEADERS = [
    "序号", "流量段", "建筑物名称", "结构形式",
    "底宽B(m)", "直径D(m)", "半径R(m)",
    "h设计(m)", "V设计(m/s)", "A设计(m²)", "R水力(m)", "湿周χ(m)",
    "Q加大(m³/s)", "h加大(m)", "V加大(m/s)",
    "超高Fb(m)", "建筑物总高H(m)",
    "设计净空高度(m)", "加大净空高度(m)",
    "设计净空比例(%)", "加大净空比例(%)",
    "状态",
]


class BatchPanel(QWidget):
    """多流量段批量水力计算面板"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._main_window = None
        self.batch_results = []
        self._detail_text_cache = ""
        self._last_calc_snapshot = None
        self._last_calc_detail = None
        self._last_import_dir = None
        self._has_opened_template = False
        self._is_sample_data = False
        self._loading_sample = False
        self._load_user_prefs()
        self._init_ui()

    def _init_ui(self):
        main_lay = QVBoxLayout(self)
        main_lay.setContentsMargins(8, 6, 8, 6)
        main_lay.setSpacing(4)

        splitter = QSplitter(Qt.Vertical)
        splitter.setChildrenCollapsible(False)
        main_lay.addWidget(splitter)

        # 上半区：输入表格
        top_w = QWidget()
        self._build_input_area(top_w)
        splitter.addWidget(top_w)

        # 下半区：结果
        bottom_w = QWidget()
        self._build_result_area(bottom_w)
        splitter.addWidget(bottom_w)

        splitter.setSizes([540, 300])

        # 添加示例数据
        self._add_sample_data()

    # ================================================================
    # 输入区
    # ================================================================
    def _build_input_area(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(4)

        # 渠道基础信息栏
        info_grp = QGroupBox("渠道基础信息")
        ig = QHBoxLayout(info_grp)
        ig.setSpacing(8)
        ig.addWidget(QLabel("渠道名称:")); self.channel_name_edit = LineEdit(); self.channel_name_edit.setText("南峰寺"); self.channel_name_edit.setFixedWidth(100); ig.addWidget(self.channel_name_edit)
        ig.addWidget(QLabel("渠道类型:")); self.channel_level_combo = ComboBox(); self.channel_level_combo.addItems(["总干渠","总干管","分干渠","分干管","干渠","干管","支渠","支管","分支渠","分支管"]); self.channel_level_combo.setCurrentText("支渠"); self.channel_level_combo.setFixedWidth(80); ig.addWidget(self.channel_level_combo)
        ig.addWidget(QLabel("起始水位(m):")); self.start_wl_edit = LineEdit(); self.start_wl_edit.setFixedWidth(70); ig.addWidget(self.start_wl_edit)
        ig.addWidget(QLabel("起始桩号:")); self.start_station_edit = LineEdit(); self.start_station_edit.setText("0+000.000"); self.start_station_edit.setFixedWidth(100); ig.addWidget(self.start_station_edit)
        self.start_station_edit.editingFinished.connect(self._format_start_station_display)
        self.start_station_edit.focusInEvent = self._on_start_station_focus_in
        station_hint = QLabel("(如12+111.222,输入12111.222)")
        station_hint.setStyleSheet("font-size:10px;color:#0066CC;")
        ig.addWidget(station_hint)
        ig.addStretch()
        lay.addWidget(info_grp)

        # 流量段设置区域（与原版一致）
        flow_grp = QGroupBox("流量段设置（各流量段的设计流量Q）")
        fg = QHBoxLayout(flow_grp)
        fg.setSpacing(8)
        fg.addWidget(QLabel("流量值(m³/s):"))
        self.flow_segments_edit = LineEdit()
        self.flow_segments_edit.setText("5.0, 4.0, 3.0")
        self.flow_segments_edit.setMinimumWidth(200)
        fg.addWidget(self.flow_segments_edit)
        hint_lbl = QLabel("格式:Q1,Q2,Q3...其中Q1代表第一流量段流量、Q2代表第二流量段流量...")
        hint_lbl.setStyleSheet("font-size:10px;color:#0066CC;")
        fg.addWidget(hint_lbl)
        btn_apply_flow = PushButton("应用到表格")
        btn_apply_flow.clicked.connect(self._apply_flow_segments)
        fg.addWidget(btn_apply_flow)
        fg.addStretch()
        lay.addWidget(flow_grp)

        # 工具栏 —— 第一行：标题 + 核心按钮 + 复选框
        tb1 = QHBoxLayout()
        tb1.setSpacing(6)
        lbl = QLabel("输入参数表")
        lbl.setStyleSheet(f"font-size:13px;font-weight:bold;color:{T1};")
        tb1.addWidget(lbl)

        btn_full = PrimaryPushButton("一键全流程计算"); btn_full.clicked.connect(self._one_click_full_flow)
        btn_full.setToolTip("批量计算 → 导入 → 渐变段 → (倒虹吸) → 推求水面线")
        btn_calc = PrimaryPushButton("开始批量计算"); btn_calc.clicked.connect(self._batch_calculate)
        btn_sample = PushButton("示例数据"); btn_sample.clicked.connect(self._add_sample_data)
        btn_template = PushButton("打开Excel模板"); btn_template.clicked.connect(self._open_excel_template)
        btn_import = PrimaryPushButton("导入Excel"); btn_import.clicked.connect(self._import_from_excel)
        self.detail_cb = CheckBox("启用详细计算过程输出")
        self.detail_cb.setChecked(False)
        self.inc_cb = CheckBox("考虑加大流量比例系数")
        self.inc_cb.setChecked(True)

        for w in [btn_import, btn_calc, btn_sample, btn_template, btn_full]:
            tb1.addWidget(w)
        tb1.addStretch()
        tb1.addWidget(self.inc_cb)
        tb1.addWidget(self.detail_cb)
        lay.addLayout(tb1)

        # 工具栏 —— 第二行：表格操作（左对齐）+ 提示（右对齐）
        tb2 = QHBoxLayout()
        tb2.setSpacing(6)

        btn_add = PushButton("新增行"); btn_add.clicked.connect(self._add_row)
        btn_insert = PushButton("插入行"); btn_insert.clicked.connect(self._insert_row)
        btn_del = PushButton("删除行"); btn_del.clicked.connect(self._del_row)
        btn_copy = PushButton("复制行"); btn_copy.clicked.connect(self._copy_row)
        btn_clear = PushButton("清空输入"); btn_clear.clicked.connect(self._clear_input)
        btn_param = PushButton("参数设置"); btn_param.clicked.connect(self._open_parameter_dialog)

        for w in [btn_add, btn_insert, btn_del, btn_copy, btn_clear, btn_param]:
            tb2.addWidget(w)
        tb2.addStretch()
        hint_label = QLabel("提示: 双击参数列打开参数设置弹窗; 双击断面类型列可选择类型")
        hint_label.setStyleSheet("font-size:12px;font-weight:600;color:#0B5CAD;")
        tb2.addWidget(hint_label)
        lay.addLayout(tb2)

        # 输入表格
        self.input_table = FrozenColumnTableWidget(0, len(INPUT_HEADERS), frozen_count=4)
        self.input_table.setHorizontalHeaderLabels(INPUT_HEADERS)
        self.input_table.horizontalHeader().setStretchLastSection(False)
        self.input_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.input_table.horizontalHeader().setMinimumSectionSize(50)
        self.input_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.input_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.input_table.setAlternatingRowColors(True)
        self.input_table.setFont(QFont("Microsoft YaHei", 10))
        self.input_table.verticalHeader().setDefaultSectionSize(28)
        # 设置前4列固定宽度，与结果表保持一致
        self.input_table.setColumnWidth(0, 45)   # 序号
        self.input_table.setColumnWidth(1, 55)   # 流量段
        self.input_table.setColumnWidth(2, 90)   # 建筑物名称
        self.input_table.setColumnWidth(3, 110)  # 结构形式
        self.input_table.cellDoubleClicked.connect(self._on_cell_double_clicked)
        self.input_table.cellChanged.connect(self._on_cell_changed)
        # 右键上下文菜单
        self.input_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.input_table.customContextMenuRequested.connect(self._show_table_context_menu)
        # 为表头设置悬浮提示
        for i, h in enumerate(INPUT_HEADERS):
            if h in _HEADER_TOOLTIPS:
                item = self.input_table.horizontalHeaderItem(i)
                if item:
                    item.setToolTip(_HEADER_TOOLTIPS[h])
        lay.addWidget(self.input_table)

        # 快捷键绑定
        QShortcut(QKeySequence("Ctrl+C"), self.input_table, self._copy_to_clipboard)
        QShortcut(QKeySequence("Ctrl+V"), self.input_table, self._paste_from_clipboard)
        QShortcut(QKeySequence("Ctrl+D"), self.input_table, self._fill_down_input)
        QShortcut(QKeySequence("Ctrl+A"), self.input_table, self.input_table.selectAll)

        # 操作提示
        hint = QLabel("提示: Ctrl+C 复制 | Ctrl+V 粘贴 | Ctrl+A 全选 | Ctrl+D 填充向下 | 双击断面类型列可选择 | 右键菜单插入/删除行")
        hint.setStyleSheet(
            "font-size:12px;font-weight:600;color:#9A3412;"
            "background:#FFF4E5;border:1px solid #F3C88B;border-radius:4px;padding:4px 8px;"
        )
        lay.addWidget(hint)

    def _on_cell_changed(self, row, col):
        """单元格编辑完成后的回调（与原版_on_input_end_edit_cell一致）"""
        if not self._loading_sample:
            self._is_sample_data = False
        if col == 1:
            # 流量段列被编辑，自动同步Q值
            seg_item = self.input_table.item(row, 1)
            if seg_item:
                try:
                    segment_num = int(seg_item.text().strip())
                    new_Q = self._get_flow_for_segment(segment_num)
                    self.input_table.blockSignals(True)
                    self.input_table.setItem(row, 6, QTableWidgetItem(str(new_Q)))
                    self.input_table.blockSignals(False)
                except (ValueError, TypeError):
                    pass

    def _on_cell_double_clicked(self, row, col):
        """双击处理：结构形式列弹出选择面板，参数列(6-19)弹出参数设置弹窗（与原版一致）"""
        if col == 3:
            # 结构形式列
            current = ""
            item = self.input_table.item(row, col)
            if item:
                current = item.text()
            dlg = StructureTypeSelector(self)
            dlg.set_current(current)
            result = dlg.exec()
            if result == QDialog.DialogCode.Accepted and dlg.selected_type:
                new_type = dlg.selected_type
                old_type = current
                type_item = QTableWidgetItem(new_type)
                type_item.setTextAlignment(Qt.AlignCenter)
                self.input_table.setItem(row, col, type_item)
                if new_type != old_type:
                    self._apply_section_type_change(row, new_type)
            # 无论选择还是 Esc 取消，都回到原表格单元格，保持操作连续性
            self.input_table.setCurrentCell(row, col)
            self.input_table.setFocus(Qt.OtherFocusReason)
        elif 6 <= col <= 19:
            # 参数列(Q~不冲流速)，打开参数设置弹窗（与原版一致）
            self._open_parameter_dialog_for_row(row)

    def _show_table_context_menu(self, pos):
        """输入表格右键上下文菜单"""
        index = self.input_table.indexAt(pos)
        menu = QMenu(self.input_table)
        menu.setStyleSheet(
            "QMenu { font-family:'Microsoft YaHei'; font-size:13px; }"
            "QMenu::item { padding:6px 24px; }"
            "QMenu::item:selected { background:#E8F0FE; }"
        )

        act_copy_cells = menu.addAction("复制  Ctrl+C")
        act_copy_cells.triggered.connect(self._copy_to_clipboard)
        act_paste_cells = menu.addAction("粘贴  Ctrl+V")
        act_paste_cells.triggered.connect(self._paste_from_clipboard)
        act_select_all = menu.addAction("全选  Ctrl+A")
        act_select_all.triggered.connect(self.input_table.selectAll)
        menu.addSeparator()

        act_add = menu.addAction("新增行（末尾）")
        act_add.triggered.connect(self._add_row)

        if index.isValid():
            row = index.row()
            menu.addSeparator()
            act_insert_above = menu.addAction("在上方插入行")
            act_insert_above.triggered.connect(lambda: self._insert_row_at(row))
            act_insert_below = menu.addAction("在下方插入行")
            act_insert_below.triggered.connect(lambda: self._insert_row_at(row + 1))
            menu.addSeparator()
            act_copy = menu.addAction("复制选中行")
            act_copy.triggered.connect(self._copy_row)
            act_del = menu.addAction("删除选中行")
            act_del.triggered.connect(self._del_row)
            menu.addSeparator()

            # 参数设置（仅非闸/倒虹吸类型可用）
            section_item = self.input_table.item(row, 3)
            section_type = section_item.text().strip() if section_item else ""
            act_param = menu.addAction("打开参数设置...")
            if not section_type or "分水" in section_type or "闸" in section_type or "倒虹吸" in section_type:
                act_param.setEnabled(False)
            act_param.triggered.connect(lambda: self._open_parameter_dialog_for_row(row))

        menu.exec(self.input_table.viewport().mapToGlobal(pos))

    def _insert_row_at(self, insert_at):
        """在指定位置插入新行，继承上一行的流量段"""
        new_segment = 1
        if insert_at > 0:
            prev_seg_item = self.input_table.item(insert_at - 1, 1)
            if prev_seg_item:
                try:
                    new_segment = int(prev_seg_item.text().strip())
                except ValueError:
                    new_segment = 1
        new_Q = str(self._get_flow_for_segment(new_segment))
        self.input_table.insertRow(insert_at)
        self.input_table.setItem(insert_at, 0, QTableWidgetItem(str(insert_at + 1)))
        self.input_table.setItem(insert_at, 1, QTableWidgetItem(str(new_segment)))
        self.input_table.setItem(insert_at, 2, QTableWidgetItem("-"))
        type_item = QTableWidgetItem("明渠-梯形")
        type_item.setTextAlignment(Qt.AlignCenter)
        self.input_table.setItem(insert_at, 3, type_item)
        self.input_table.setItem(insert_at, 6, QTableWidgetItem(new_Q))
        self.input_table.setItem(insert_at, 7, QTableWidgetItem("0.014"))
        self.input_table.setItem(insert_at, 8, QTableWidgetItem("3000"))
        self.input_table.setItem(insert_at, 9, QTableWidgetItem("1.0"))
        self.input_table.setItem(insert_at, 18, QTableWidgetItem("0.1"))
        self.input_table.setItem(insert_at, 19, QTableWidgetItem("100"))
        self._renumber()
        self.input_table.selectRow(insert_at)

    def _add_row(self, data=None):
        """添加一行（无data时填充完整默认值，与原版_create_default_row一致）"""
        row = self.input_table.rowCount()
        self.input_table.insertRow(row)
        if data and isinstance(data, (list, tuple)):
            # 有数据时直接填充
            for col, val in enumerate(data):
                if col < self.input_table.columnCount():
                    item = QTableWidgetItem(str(val) if val else "")
                    if col == 3:
                        item.setTextAlignment(Qt.AlignCenter)
                    self.input_table.setItem(row, col, item)
        else:
            # 无数据时填充默认值：流量段编号在上一行基础上递增（与原版一致）
            if row > 0:
                prev_seg_item = self.input_table.item(row - 1, 1)
                try:
                    last_segment = int(prev_seg_item.text().strip()) if prev_seg_item else 1
                except ValueError:
                    last_segment = 1
                new_segment = last_segment + 1
            else:
                new_segment = 1
            new_Q = str(self._get_flow_for_segment(new_segment))
            self.input_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
            self.input_table.setItem(row, 1, QTableWidgetItem(str(new_segment)))
            self.input_table.setItem(row, 2, QTableWidgetItem("-"))
            type_item = QTableWidgetItem("明渠-梯形")
            type_item.setTextAlignment(Qt.AlignCenter)
            self.input_table.setItem(row, 3, type_item)
            self.input_table.setItem(row, 6, QTableWidgetItem(new_Q))
            self.input_table.setItem(row, 7, QTableWidgetItem("0.014"))
            self.input_table.setItem(row, 8, QTableWidgetItem("3000"))
            self.input_table.setItem(row, 9, QTableWidgetItem("1.0"))
            self.input_table.setItem(row, 18, QTableWidgetItem("0.1"))
            self.input_table.setItem(row, 19, QTableWidgetItem("100"))

    def _insert_row(self):
        """插入行 - 在选中行之前插入新行（与原版一致）"""
        selected = sorted(set(idx.row() for idx in self.input_table.selectedIndexes()))
        if not selected:
            InfoBar.warning("提示", "请先选择要在其之前插入行的位置", parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return
        self._insert_row_at(min(selected))

    def _open_excel_template(self):
        """使用默认程序打开Excel模板文件（与原版一致）"""
        if not fluent_question(self, "打开Excel模板",
            "打开Excel导入模板\n\n"
            "即将使用默认程序打开模板文件。\n\n"
            "使用说明：\n"
            "1. 打开后请【另存为】到您的目录\n"
            "2. 在副本中按模板格式填写数据\n"
            "3. 回到本界面点击【导入Excel】导入",
            yes_text="继续", no_text="取消"):
            return
        template_path = self._get_template_path()
        if not os.path.exists(template_path):
            fluent_info(self, "错误", f"未找到模板文件：\n{template_path}")
            return
        try:
            os.startfile(template_path)
            self._has_opened_template = True
            self._save_user_prefs()
        except Exception as e:
            fluent_error(self, "错误", f"打开模板文件失败：{str(e)}")

    def _get_template_path(self):
        """获取Excel模板文件路径（兼容开发环境和打包环境）"""
        template_name = "多流量段批量计算_导入Excel（模板）.xlsx"
        if getattr(sys, 'frozen', False):
            base = sys._MEIPASS
        else:
            # __file__ = 渠系断面设计/batch/panel.py → 需要向上3层到V1.0/
            base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        return os.path.join(base, "data", template_name)

    def _del_row(self):
        rows = sorted(set(idx.row() for idx in self.input_table.selectedIndexes()), reverse=True)
        if not rows:
            InfoBar.warning("提示", "请先选择要删除的行", parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return
        count = len(rows)
        if count == 1:
            name_item = self.input_table.item(rows[0], 2)
            name_text = name_item.text() if name_item else ""
            confirm_msg = f"确定要删除行 '{name_text}' 吗?"
        else:
            confirm_msg = f"确定要删除选中的 {count} 行吗?"
        if not fluent_question(self, "确认删除", confirm_msg):
            return
        for r in rows:
            self.input_table.removeRow(r)
        self._renumber()

    def _copy_row(self):
        rows = sorted(set(idx.row() for idx in self.input_table.selectedIndexes()))
        if not rows:
            InfoBar.warning("提示", "请先选择要复制的行", parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return
        for r in rows:
            data = self._get_row_data(r)
            self._add_row(data)
        self._renumber()

    def _clear_input(self, force=False):
        if self.input_table.rowCount() == 0:
            return
        if not force:
            if not fluent_question(self, "确认", "确定要清空所有输入数据吗?"):
                return
        self.input_table.setRowCount(0)
        self._last_calc_snapshot = None
        self._last_calc_detail = None
        if not self._loading_sample:
            self._is_sample_data = False
        # 同步清空结果（防止清空输入后仍能导出/查看旧结果）
        self._clear_results()

    def _copy_to_clipboard(self):
        """复制选中单元格到剪贴板（制表符分隔，可直接粘贴到Excel）"""
        selected = self.input_table.selectedIndexes()
        if not selected:
            return
        # 按行列排序
        rows = sorted(set(idx.row() for idx in selected))
        cols = sorted(set(idx.column() for idx in selected))
        selected_set = {(idx.row(), idx.column()) for idx in selected}
        # 构建制表符分隔文本
        lines = []
        for r in rows:
            cells = []
            for c in cols:
                if (r, c) in selected_set:
                    item = self.input_table.item(r, c)
                    cells.append(item.text() if item else "")
                else:
                    cells.append("")
            lines.append("\t".join(cells))
        text = "\n".join(lines)
        QApplication.clipboard().setText(text)
        cell_count = len(selected)
        InfoBar.success("已复制", f"已复制 {len(rows)} 行 × {len(cols)} 列（共 {cell_count} 个单元格）到剪贴板",
                       parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)

    def _paste_from_clipboard(self):
        """从剪贴板粘贴数据到输入表格（支持从Excel复制的多行多列制表符分隔数据）"""
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        if not text or not text.strip():
            return
        # 解析剪贴板文本（制表符分隔行/列）
        lines = text.strip().split('\n')
        paste_data = []
        for line in lines:
            cells = line.rstrip('\r').split('\t')
            paste_data.append(cells)
        if not paste_data:
            return
        # 确定粘贴起始位置
        selected = self.input_table.selectedIndexes()
        if selected:
            start_row = min(idx.row() for idx in selected)
            start_col = min(idx.column() for idx in selected)
        else:
            start_row = self.input_table.rowCount()
            start_col = 0
        # 扩展行数以容纳粘贴数据
        needed_rows = start_row + len(paste_data)
        while self.input_table.rowCount() < needed_rows:
            self.input_table.insertRow(self.input_table.rowCount())
        # 收集无效断面类型（与原版一致）
        invalid_types = []
        pasted_rows = 0
        # 填充数据
        for r_offset, row_data in enumerate(paste_data):
            if not any(c.strip() for c in row_data):
                continue
            for c_offset, val in enumerate(row_data):
                target_row = start_row + r_offset
                target_col = start_col + c_offset
                if target_col < self.input_table.columnCount():
                    cell_value = val.strip()
                    # 断面类型列(索引3)特殊处理：验证类型有效性（与原版一致）
                    if target_col == 3 and cell_value:
                        if cell_value not in SECTION_TYPES:
                            invalid_types.append(f"行{target_row + 1}: {cell_value}")
                    item = QTableWidgetItem(cell_value)
                    if target_col == 3:
                        item.setTextAlignment(Qt.AlignCenter)
                    self.input_table.setItem(target_row, target_col, item)
            pasted_rows += 1
        self._renumber()
        # 粘贴后校验：自动映射断面类型并确保比降有默认值（与原版_on_input_end_paste一致）
        self.input_table.blockSignals(True)
        for r_offset in range(len(paste_data)):
            target_row = start_row + r_offset
            if target_row >= self.input_table.rowCount():
                break
            # 断面类型映射（允许用户粘贴简写）
            type_item = self.input_table.item(target_row, 3)
            if type_item:
                raw_type = type_item.text().strip()
                if raw_type and raw_type not in SECTION_TYPES:
                    mapped = self._map_section_type(raw_type)
                    if mapped:
                        type_item.setText(mapped)
            # 确保比降有默认值
            slope_item = self.input_table.item(target_row, 8)
            if not slope_item or not slope_item.text().strip():
                self.input_table.setItem(target_row, 8, QTableWidgetItem("3000"))
        self.input_table.blockSignals(False)
        # 粘贴后自动检测流量段（与原版一致）
        self._auto_detect_flow_segments()
        # 显示结果（与原版一致：如有无效类型则警告）
        if invalid_types:
            error_msgs = ["【无效的类型】"]
            for msg in invalid_types[:5]:
                error_msgs.append(msg)
            if len(invalid_types) > 5:
                error_msgs.append(f"... 共{len(invalid_types)}条")
            error_msgs.append("")
            error_msgs.append("有效类型：" + "、".join(SECTION_TYPES))
            fluent_info(self, "断面类型错误",
                f"已粘贴 {pasted_rows} 行数据\n\n" + "\n".join(error_msgs))
        else:
            InfoBar.success("粘贴完成", f"已粘贴 {pasted_rows} 行数据",
                           parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)

    def _fill_down_input(self):
        """填充向下：将选中区域的第一行数据复制到下方所有选中行（与原版Ctrl+D功能一致）"""
        selected = self.input_table.selectedIndexes()
        if not selected:
            InfoBar.warning("提示", "请先选中要填充的区域", parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return
        # 按列分组
        col_rows = {}
        for idx in selected:
            col_rows.setdefault(idx.column(), []).append(idx.row())
        filled = 0
        for col, rows in col_rows.items():
            rows.sort()
            if len(rows) < 2:
                continue
            # 第一行的值作为源
            first_item = self.input_table.item(rows[0], col)
            source_val = first_item.text() if first_item else ""
            for r in rows[1:]:
                item = QTableWidgetItem(source_val)
                if col == 3:
                    item.setTextAlignment(Qt.AlignCenter)
                self.input_table.setItem(r, col, item)
                filled += 1
        if filled > 0:
            InfoBar.success("填充完成", f"已向下填充 {filled} 个单元格",
                           parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)

    def _renumber(self):
        for r in range(self.input_table.rowCount()):
            self.input_table.setItem(r, 0, QTableWidgetItem(str(r + 1)))

    def _get_row_data(self, row):
        """获取指定行的数据"""
        data = []
        for col in range(self.input_table.columnCount()):
            item = self.input_table.item(row, col)
            data.append(item.text() if item else "")
        return data

    def _get_all_input_data(self):
        """获取所有输入数据"""
        rows = []
        for r in range(self.input_table.rowCount()):
            rows.append(self._get_row_data(r))
        return rows

    def _add_sample_data(self):
        """加载示例数据 - 来自多流量段表格填写示例.xlsx的完整数据（与原版一致，44行）"""
        self._loading_sample = True
        # 列: 序号,流量段,建筑物名称,结构形式,X,Y,Q,n,比降,m,B,宽深比,R,D,渡槽深宽比,倒角角度,倒角底边,圆心角,不淤,不冲
        samples = [
            ["1",  "1", "-",      "明渠-矩形",       "649606.177086", "3377745.982674", "5", "0.014", "3000", "0",   "2",   "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["2",  "1", "-",      "明渠-矩形",       "649534.180449", "3377664.854614", "5", "0.014", "3000", "0",   "2",   "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["3",  "1", "-",      "明渠-矩形",       "649480.482814", "3377634.277101", "5", "0.014", "3000", "0",   "2",   "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["4",  "1", "土地垭", "隧洞-马蹄形Ⅰ型", "649478.323235", "3377610.807806", "5", "0.014", "2000", "",    "",    "", "1.5", "",    "",    "",    "",    "",    "0.1", "100"],
            ["5",  "1", "土地垭", "隧洞-马蹄形Ⅰ型", "649441.884821", "3377556.331275", "5", "0.014", "2000", "",    "",    "", "1.5", "",    "",    "",    "",    "",    "0.1", "100"],
            ["6",  "1", "-",      "明渠-梯形",       "649440.214195", "3377528.976904", "5", "0.014", "3000", "1",   "1.8", "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["7",  "1", "-",      "明渠-梯形",       "649419.568825", "3377522.66441",  "5", "0.014", "3000", "1",   "1.8", "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["8",  "1", "磨尔滩", "分水闸",           "649402.139216", "3377539.733849", "5", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["9",  "2", "-",      "明渠-梯形",       "649310.705602", "3377545.834305", "4", "0.014", "3000", "1",   "1.8", "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["10", "2", "沪蓉",   "倒虹吸",           "649264.563059", "3377548.912938", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["11", "2", "沪蓉",   "倒虹吸",           "649244.41293",  "3377550.257356", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["12", "2", "沪蓉",   "倒虹吸",           "649220.867829", "3377563.964679", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["13", "2", "沪蓉",   "倒虹吸",           "649184.732272", "3377556.518614", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["14", "2", "沪蓉",   "倒虹吸",           "649146.2872",   "3377588.1779",   "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["15", "2", "宋家沟", "隧洞-圆拱直墙型", "649104.399377", "3377613.995873", "4", "0.014", "2000", "",    "2.5", "", "",    "",    "",    "",    "",    "120", "0.1", "100"],
            ["16", "2", "宋家沟", "隧洞-圆拱直墙型", "649098.598741", "3377595.290122", "4", "0.014", "2000", "",    "2.5", "", "",    "",    "",    "",    "",    "120", "0.1", "100"],
            ["17", "2", "广岳",   "倒虹吸",           "649086.007282", "3377582.009467", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["18", "2", "广岳",   "倒虹吸",           "649066.369061", "3377577.838164", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["19", "2", "广岳",   "倒虹吸",           "649033.673293", "3377532.99707",  "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["20", "2", "广岳",   "倒虹吸",           "649018.983612", "3377482.347484", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["21", "2", "伍家沟", "隧洞-圆拱直墙型", "648991.829093", "3377453.363461", "4", "0.014", "2000", "",    "2.5", "", "",    "",    "",    "",    "",    "120", "0.1", "100"],
            ["22", "2", "伍家沟", "隧洞-圆拱直墙型", "648969.028473", "3377444.44637",  "4", "0.014", "2000", "",    "2.5", "", "",    "",    "",    "",    "",    "120", "0.1", "100"],
            ["23", "2", "伍家沟", "隧洞-圆拱直墙型", "648918.161636", "3377447.833438", "4", "0.014", "2000", "",    "2.5", "", "",    "",    "",    "",    "",    "120", "0.1", "100"],
            ["24", "2", "刘家沟", "渡槽-矩形",       "648879.873566", "3377424.400731", "4", "0.014", "2000", "",    "",    "", "",    "",    "0.8", "30",  "0.3", "120", "0.1", "100"],
            ["25", "2", "刘家沟", "渡槽-矩形",       "648873.319207", "3377389.201113", "4", "0.014", "2000", "",    "",    "", "",    "",    "0.8", "30",  "0.3", "120", "0.1", "100"],
            ["26", "2", "广高路", "倒虹吸",           "648875.83158",  "3377349.478728", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["27", "2", "广高路", "倒虹吸",           "648859.515404", "3377325.867714", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["28", "2", "广高路", "倒虹吸",           "648823.413217", "3377328.752934", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["29", "2", "广高路", "倒虹吸",           "648778.747964", "3377306.056947", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["30", "2", "广高路", "倒虹吸",           "648742.967801", "3377279.279514", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["31", "2", "广高路", "倒虹吸",           "648740.770589", "3377262.028944", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["32", "2", "广高路", "倒虹吸",           "648704.058844", "3377256.358551", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["33", "2", "-",      "明渠-圆形",       "648687.241348", "3377234.03888",  "4", "0.014", "3000", "0",   "",    "", "",    "2.8", "",    "",    "",    "",    "0.1", "100"],
            ["34", "2", "-",      "明渠-圆形",       "648677.298141", "3377230.295614", "4", "0.014", "3000", "0",   "",    "", "",    "2.8", "",    "",    "",    "",    "0.1", "100"],
            ["35", "2", "台儿沟", "隧洞-圆形",       "648610.458063", "3377205.132683", "4", "0.014", "2000", "",    "",    "", "",    "2.6", "",    "",    "",    "",    "0.1", "100"],
            ["36", "2", "美团沟", "分水闸",           "648588.193106", "3377182.717782", "4", "0.014", "",     "",    "",    "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["37", "3", "台儿沟", "隧洞-圆形",       "648359.767433", "3377105.690753", "3", "0.014", "2000", "",    "",    "", "",    "2.6", "",    "",    "",    "",    "0.1", "100"],
            ["38", "3", "台儿沟", "隧洞-圆形",       "648259.932461", "3376966.162254", "3", "0.014", "2000", "",    "",    "", "",    "2.6", "",    "",    "",    "",    "0.1", "100"],
            ["39", "3", "梨子园", "渡槽-U形",         "647962.330045", "3376909.650621", "3", "0.014", "1500", "",    "",    "", "1.6", "",    "",    "",    "",    "",    "0.1", "100"],
            ["40", "3", "梨子园", "渡槽-U形",         "647644.538898", "3376595.606329", "3", "0.014", "1500", "",    "",    "", "1.6", "",    "",    "",    "",    "",    "0.1", "100"],
            ["41", "3", "油房垭", "隧洞-圆拱直墙型", "647641.215709", "3376559.422576", "3", "0.014", "2000", "",    "2.5", "", "",    "",    "",    "",    "",    "120", "0.1", "100"],
            ["42", "3", "油房垭", "隧洞-圆拱直墙型", "647597.709292", "3376537.187663", "3", "0.014", "2000", "",    "2.5", "", "",    "",    "",    "",    "",    "120", "0.1", "100"],
            ["43", "3", "-",      "矩形暗涵",         "647506.778347", "3376513.531331", "3", "0.014", "3000", "0",   "2.4", "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["44", "3", "-",      "矩形暗涵",         "647387.9806",   "3376403.8971",   "3", "0.014", "3000", "",    "2.4", "", "",    "",    "",    "",    "",    "",    "0.1", "100"],
            ["45", "3", "-",      "明渠-U形",        "647350.0",      "3376350.0",      "3", "0.014", "3000", "",    "",    "", "0.8", "",    "",    "14",  "", "152", "0.1", "100"],
            ["46", "3", "-",      "明渠-U形",        "647280.0",      "3376280.0",      "3", "0.014", "3000", "",    "",    "", "0.8", "",    "",    "14",  "", "152", "0.1", "100"],
        ]
        self._clear_input(force=True)
        for row_data in samples:
            self._add_row(row_data)
        # 设置默认起始水位和流量段
        self.start_wl_edit.setText("400")
        self.flow_segments_edit.setText("5.0, 4.0, 3.0")
        self._auto_detect_flow_segments()
        auto_resize_table(self.input_table)
        self._loading_sample = False
        self._is_sample_data = True
        InfoBar.success("示例数据", f"已加载 {len(samples)} 行示例数据（含明渠/明渠-U形/隧洞/渡槽/暗涵/倒虹吸/分水闸）",
                       parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)

    # ================================================================
    # 结果区
    # ================================================================
    def _build_result_area(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(4)

        # 工具栏
        tb = QHBoxLayout()
        lbl = QLabel("计算结果")
        lbl.setStyleSheet(f"font-size:13px;font-weight:bold;color:{T1};")
        tb.addWidget(lbl)
        tb.addStretch()
        btn_excel = PushButton("导出Excel报告"); btn_excel.clicked.connect(self._export_excel)
        btn_word = PushButton("导出详细过程(Word)"); btn_word.clicked.connect(self._export_word)
        btn_clr = PushButton("清空结果"); btn_clr.clicked.connect(self._clear_results)
        for w in [btn_excel, btn_word, btn_clr]:
            tb.addWidget(w)
        lay.addLayout(tb)

        # 选项卡
        self.result_notebook = QTabWidget()
        lay.addWidget(self.result_notebook)

        # Tab1: 结果汇总表
        t1 = QWidget(); t1l = QVBoxLayout(t1); t1l.setContentsMargins(2, 2, 2, 2)
        self.result_table = FrozenColumnTableWidget(0, len(RESULT_HEADERS), frozen_count=4)
        self.result_table.setHorizontalHeaderLabels(RESULT_HEADERS)
        self.result_table.horizontalHeader().setStretchLastSection(False)
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.result_table.horizontalHeader().setMinimumSectionSize(50)
        self.result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setFont(QFont("Microsoft YaHei", 10))
        self.result_table.verticalHeader().setDefaultSectionSize(28)
        # 设置前4列固定宽度，与输入表保持一致
        self.result_table.setColumnWidth(0, 45)   # 序号
        self.result_table.setColumnWidth(1, 55)   # 流量段
        self.result_table.setColumnWidth(2, 90)   # 建筑物名称
        self.result_table.setColumnWidth(3, 110)  # 结构形式
        t1l.addWidget(self.result_table)
        self.result_notebook.addTab(t1, "结果汇总表")

        # Tab2: 详细过程
        t2 = QWidget(); t2l = QVBoxLayout(t2); t2l.setContentsMargins(2, 2, 2, 2)
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        self.detail_text.setFont(QFont("Consolas", 10))
        t2l.addWidget(self.detail_text)
        self.result_notebook.addTab(t2, "详细计算过程")

    # ================================================================
    # 跨面板引用
    # ================================================================
    def set_main_window(self, main_window):
        """由 MainWindow 在初始化后调用，保存引用以支持跨面板操作"""
        self._main_window = main_window

    def _one_click_full_flow(self):
        """一键全流程：批量计算 → 切换到推求水面线 → 导入 → 渐变段 → (倒虹吸) → 执行计算"""
        if not fluent_question(
            self, "一键全流程计算",
            "本功能将自动依次执行以下步骤：\n\n"
            "① 批量计算 —— 对表格中所有渠段进行水力计算\n"
            "② 导入水面线 —— 将计算结果自动导入「推求水面线」模块\n"
            "③ 渐变段计算 —— 自动计算各渠段间的渐变段水头损失\n"
            "④ 倒虹吸计算 —— 若存在倒虹吸，自动纳入计算\n"
            "⑤ 推求水面线 —— 执行全线水面线推算\n\n"
            "是否继续？",
            yes_text="开始执行", no_text="取消",
        ):
            return

        # 检查数据来源
        if self.input_table.rowCount() == 0:
            InfoBar.warning(
                "无输入数据",
                "请先准备输入数据再执行全流程：\n"
                "① 点击【导入Excel】导入您的数据文件\n"
                "② 或手动【新增行】逐行填写参数\n"
                "③ 或点击【示例数据】加载演示数据",
                parent=self._info_parent(), duration=6000,
                position=InfoBarPosition.TOP,
            )
            return

        if self._is_sample_data:
            if not fluent_question(
                self, "数据确认",
                "当前表格中的数据为【示例数据】，并非您的工程数据。\n\n"
                "如需使用自己的数据，请先点击【导入Excel】导入，\n"
                "或手动修改表格中的参数。\n\n"
                "是否仍要使用示例数据继续执行全流程？",
                yes_text="继续使用示例数据", no_text="取消",
            ):
                return

        # ① 批量计算
        self._batch_calculate()
        if not self.batch_results:
            return

        mw = self._main_window
        if mw is None:
            InfoBar.warning("提示", "未获取到主窗口引用，无法自动跳转",
                            parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        # ② 切换到推求水面线面板（index 6）
        mw._switch_to(6)
        wp = mw.water_profile_panel

        # ③ 从批量计算导入
        wp._import_from_batch()

        # ④ 插入渐变段（自动确认，跳过批量明渠段对话框）
        wp._insert_transitions(auto_confirm=True)

        # ⑤ 若有倒虹吸则弹出对话框（模态，关闭后自动继续）
        nodes = wp._build_nodes_from_table()
        has_siphon = any(
            getattr(n, 'structure_type', None) and "倒虹吸" in n.structure_type.value
            for n in nodes if getattr(n, 'structure_type', None)
        )
        if has_siphon:
            wp._open_siphon_calculator(auto_run=True)

        # ⑥ 执行计算
        wp._calculate()

    # ================================================================
    # 批量计算
    # ================================================================
    def _batch_calculate(self):
        input_rows = self._get_all_input_data()
        if not input_rows:
            InfoBar.warning("提示", "请先输入数据", parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return

        # 检查输入数据是否与上次计算时一致，若一致则无需重复计算
        current_snapshot = [[str(cell).strip() if cell is not None else "" for cell in row] for row in input_rows]
        current_snapshot.append(["__inc_cb__", str(self.inc_cb.isChecked())])
        if self._last_calc_snapshot is not None and current_snapshot == self._last_calc_snapshot:
            if self.detail_cb.isChecked() and not self._last_calc_detail and self.batch_results:
                # 从未勾选变为已勾选，且有缓存结果，补充生成详细输出
                self._regenerate_detail_report()
                self._last_calc_detail = True
                return
            InfoBar.info("提示", "输入数据未发生变化，已计算完毕，无需重复计算。\n如需重新计算，请先修改上方表格数据。",
                         parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        # 验证建筑物名称+结构形式是否重复（与原版一致）
        if not self._validate_duplicate_buildings(input_rows):
            return

        self._clear_results()
        self.batch_results = []
        success_count = 0; fail_count = 0; skip_count = 0; total_count = 0
        error_details = []  # 收集详细错误信息（与原版一致）
        result_rows = []
        detail_lines = []

        detail_lines.append("=" * 80)
        detail_lines.append("          多流量段批量水力计算 - 详细计算过程报告")
        detail_lines.append("=" * 80)
        detail_lines.append(f"计算时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        detail_lines.append(f"总计建筑物数量: {len(input_rows)}")
        detail_lines.append("-" * 80)
        detail_lines.append("")

        for values in input_rows:
            values = self._normalize_row(values, len(INPUT_HEADERS))
            section_type = str(values[3]).strip()
            if not section_type:
                continue
            total_count += 1
            seq = str(values[0]).strip()
            segment = str(values[1]).strip()
            building_name = str(values[2]).strip()

            try:
                # 倒虹吸占位行
                if section_type == "倒虹吸":
                    row_out = ["-"] * len(RESULT_HEADERS)
                    row_out[0] = seq; row_out[1] = segment; row_out[2] = building_name
                    row_out[3] = "倒虹吸"; row_out[-1] = "⏭ 占位行(不参与计算)"
                    result_rows.append(row_out)
                    siphon_result = {
                        'success': True, 'section_type': '倒虹吸', 'is_siphon': True,
                        'flow_section': segment, 'building_name': building_name,
                        'coord_X': self._sf(values[4], 0.0), 'coord_Y': self._sf(values[5], 0.0),
                        'Q': self._sf(values[6]), 'n': self._sf(values[7], 0.014), 'slope_inv': 0,
                    }
                    self.batch_results.append({'input': values, 'result': siphon_result})
                    skip_count += 1
                    if self.detail_cb.isChecked():
                        detail_lines.append(f"【项目 {total_count}】")
                        detail_lines.append(self._gen_detail_report(values, siphon_result))
                        detail_lines.append("\n" + "*" * 80 + "\n")
                    continue

                # 闸·分水类占位行（分水闸/分水口/节制闸/泄水闸等）
                if "闸" in section_type or "分水" in section_type:
                    row_out = ["-"] * len(RESULT_HEADERS)
                    row_out[0] = seq; row_out[1] = segment; row_out[2] = building_name
                    row_out[3] = section_type; row_out[-1] = f"⏭ {section_type}(不参与断面计算)"
                    result_rows.append(row_out)
                    gate_result = {
                        'success': True, 'section_type': section_type, 'is_diversion_gate': True,
                        'flow_section': segment, 'building_name': building_name,
                        'coord_X': self._sf(values[4], 0.0), 'coord_Y': self._sf(values[5], 0.0),
                        'Q': self._sf(values[6]), 'n': self._sf(values[7], 0.014), 'slope_inv': 0,
                    }
                    self.batch_results.append({'input': values, 'result': gate_result})
                    skip_count += 1
                    if self.detail_cb.isChecked():
                        detail_lines.append(f"【项目 {total_count}】")
                        detail_lines.append(self._gen_detail_report(values, gate_result))
                        detail_lines.append("\n" + "*" * 80 + "\n")
                    continue

                # 解析输入参数
                Q = self._sf(values[6])
                n = self._sf(values[7], 0.014)
                slope_inv = self._sf(values[8])
                m = self._sf(values[9], 0)
                b = self._sf(values[10], 0)
                beta = self._sf(values[11], 0)
                R = self._sf(values[12], 0)
                D = self._sf(values[13], 0)
                ducao_depth_ratio = self._sf(values[14], 0)
                chamfer_angle = self._sf(values[15], 0)
                chamfer_length = self._sf(values[16], 0)
                theta_deg = self._sf(values[17], 0)
                v_min = self._sf(values[18], 0.1)
                v_max = self._sf(values[19], 100)

                if Q <= 0: raise ValueError("流量Q必须大于0")
                if n <= 0: raise ValueError("糙率n必须大于0")
                if slope_inv <= 0: raise ValueError("比降必须大于0")

                # 计算分发
                use_inc = self.inc_cb.isChecked()
                result = self._calculate_single(
                    section_type, Q, n, slope_inv, v_min, v_max,
                    m=m, b=b, beta=beta, R=R, D=D,
                    ducao_depth_ratio=ducao_depth_ratio,
                    chamfer_angle=chamfer_angle, chamfer_length=chamfer_length,
                    theta_deg=theta_deg,
                    manual_increase_percent=0 if not use_inc else None
                )
                if result:
                    result['_use_increase'] = use_inc

                if result and result.get('success'):
                    row_out = self._extract_result_row(seq, segment, building_name, section_type, result,
                                                       use_increase=result.get('_use_increase', True))
                    result_rows.append(row_out)
                    self.batch_results.append({'input': values, 'result': result})
                    success_count += 1
                    # 详细报告
                    if self.detail_cb.isChecked():
                        detail_lines.append(f"【项目 {total_count}】 {building_name} ({section_type})")
                        detail_lines.append(self._gen_detail_report(values, result))
                        detail_lines.append("\n" + "*" * 80 + "\n")
                else:
                    err = result.get('error_message', '计算失败') if result else '计算失败'
                    if "计算失败" in err:
                        error_details.append(f"序号{seq} ({building_name}): {err}")
                    row_out = ["-"] * len(RESULT_HEADERS)
                    row_out[0] = seq; row_out[1] = segment; row_out[2] = building_name
                    row_out[3] = section_type; row_out[-1] = f"✗ {err}"
                    result_rows.append(row_out)
                    fail_count += 1

            except Exception as e:
                row_out = ["-"] * len(RESULT_HEADERS)
                row_out[0] = seq; row_out[1] = segment; row_out[2] = building_name
                row_out[3] = section_type; row_out[-1] = f"错误: {str(e)}"
                result_rows.append(row_out)
                fail_count += 1

        # 填充结果表格
        self.result_table.setRowCount(len(result_rows))
        for r, row in enumerate(result_rows):
            for c, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignCenter)
                # 状态列着色
                if c == len(RESULT_HEADERS) - 1:
                    if "✓" in str(val) or "成功" in str(val):
                        item.setForeground(QColor("#2E7D32"))
                    elif "✗" in str(val) or "失败" in str(val) or "错误" in str(val):
                        item.setForeground(QColor("#C62828"))
                    elif "占位" in str(val):
                        item.setForeground(QColor("#757575"))
                self.result_table.setItem(r, c, item)
        auto_resize_table(self.result_table)
        self._sync_common_columns()

        # 详细过程
        if self.detail_cb.isChecked():
            self.detail_text.setPlainText("\n".join(detail_lines))
            self.result_notebook.setCurrentIndex(1)
        else:
            self.result_notebook.setCurrentIndex(0)

        # 汇总
        msg = f"总计: {total_count}条\n成功: {success_count}条\n失败: {fail_count}条"
        if skip_count > 0:
            msg += f"\n跳过: {skip_count}条 (倒虹吸/闸类占位行不参与断面计算)"
        if error_details:
            detailed_errors = "\n\n" + "=" * 30 + "\n详细失败原因及建议：\n" + "\n\n".join(error_details)
            fluent_info(self, "批量计算完成 (存在异常)", msg + detailed_errors)
        elif fail_count == 0:
            InfoBar.success("批量计算完成", msg.replace('\n', ' | '), parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
        else:
            fluent_info(self, "批量计算完成", msg)

        # 保存当前输入数据快照，用于下次判断是否需要重新计算
        self._last_calc_snapshot = current_snapshot
        self._last_calc_detail = self.detail_cb.isChecked()

        # 注册批量计算结果到共享数据管理器，供推求水面线模块导入
        if SHARED_DATA_AVAILABLE and self.batch_results:
            try:
                shared_data = get_shared_data_manager()
                results_for_register = []
                for item in self.batch_results:
                    r = item['result']
                    v = item['input']
                    # 补充元数据（始终用输入表的完整类型名覆盖引擎简化名，
                    # 如 "隧洞-马蹄形Ⅰ型" 替代引擎返回的 "马蹄形标准Ⅰ型"）
                    input_section_type = str(v[3]).strip()
                    if input_section_type:
                        r['section_type'] = input_section_type
                    if 'building_name' not in r:
                        r['building_name'] = str(v[2]).strip()
                    if 'flow_section' not in r:
                        r['flow_section'] = str(v[1]).strip()
                    # 坐标始终从输入表获取（引擎不返回坐标信息）
                    r['coord_X'] = self._sf(v[4], 0.0)
                    r['coord_Y'] = self._sf(v[5], 0.0)
                    if 'Q' not in r:
                        r['Q'] = self._sf(v[6])
                    if 'n' not in r:
                        r['n'] = self._sf(v[7], 0.014)
                    if 'slope_inv' not in r:
                        r['slope_inv'] = self._sf(v[8])
                    if 'm' not in r:
                        r['m'] = self._sf(v[9])
                    results_for_register.append(r)
                count = shared_data.register_batch_results(results_for_register)
                if count > 0:
                    InfoBar.info("数据共享", f"已注册 {count} 条结果到共享数据管理器，可在推求水面线模块中导入",
                                 parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            except Exception as e:
                print(f"注册批量结果到共享数据管理器失败: {e}")

    # ================================================================
    # 计算分发
    # ================================================================
    def _calculate_single(self, section_type, Q, n, slope_inv, v_min, v_max, *,
                          m=0, b=0, beta=0, R=0, D=0,
                          ducao_depth_ratio=0, chamfer_angle=0, chamfer_length=0, theta_deg=0,
                          manual_increase_percent=None):
        """根据断面类型调用对应计算引擎"""
        _inc = manual_increase_percent
        if "明渠-梯形" in section_type:
            if not MINGQU_AVAILABLE: return {'success': False, 'error_message': '明渠计算模块未加载'}
            return mingqu_calculate(Q=Q, m=m, n=n, slope_inv=slope_inv,
                                    v_min=v_min, v_max=v_max,
                                    manual_b=b if b > 0 else None,
                                    manual_beta=beta if beta > 0 else None,
                                    manual_increase_percent=_inc)
        elif "明渠-矩形" in section_type:
            if not MINGQU_AVAILABLE: return {'success': False, 'error_message': '明渠计算模块未加载'}
            return mingqu_calculate(Q=Q, m=0, n=n, slope_inv=slope_inv,
                                    v_min=v_min, v_max=v_max,
                                    manual_b=b if b > 0 else None,
                                    manual_beta=beta if beta > 0 else None,
                                    manual_increase_percent=_inc)
        elif "明渠-圆形" in section_type:
            if not MINGQU_AVAILABLE: return {'success': False, 'error_message': '明渠计算模块未加载'}
            return circular_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                      v_min=v_min, v_max=v_max,
                                      manual_D=D if D > 0 else None,
                                      increase_percent=_inc)
        elif "明渠-U形" in section_type:
            if not MINGQU_AVAILABLE: return {'success': False, 'error_message': '明渠计算模块未加载'}
            if R <= 0:
                return {'success': False, 'error_message': '明渠-U形需要填写半径R（列索徕12）'}
            return mingqu_u_calculate(Q=Q, R=R, alpha_deg=chamfer_angle, theta_deg=theta_deg,
                                      n=n, slope_inv=slope_inv,
                                      v_min=v_min, v_max=v_max,
                                      manual_increase_percent=_inc)
        elif "渡槽-U形" in section_type:
            if not DUCAO_AVAILABLE: return {'success': False, 'error_message': '渡槽计算模块未加载'}
            return ducao_u_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                     v_min=v_min, v_max=v_max,
                                     manual_R=R if R > 0 else None,
                                     manual_increase_percent=_inc)
        elif "渡槽-矩形" in section_type:
            if not DUCAO_AVAILABLE: return {'success': False, 'error_message': '渡槽计算模块未加载'}
            dr = ducao_depth_ratio if ducao_depth_ratio > 0 else 0.8
            return ducao_rect_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                        v_min=v_min, v_max=v_max,
                                        depth_width_ratio=dr,
                                        chamfer_angle=chamfer_angle, chamfer_length=chamfer_length,
                                        manual_increase_percent=_inc)
        elif "隧洞-圆形" in section_type:
            if not SUIDONG_AVAILABLE: return {'success': False, 'error_message': '隙洞计算模块未加载'}
            return suidong_circular_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                              v_min=v_min, v_max=v_max,
                                              manual_D=D if D > 0 else None,
                                              manual_increase_percent=_inc)
        elif "隧洞-圆拱直墙型" in section_type:
            if not SUIDONG_AVAILABLE: return {'success': False, 'error_message': '隙洞计算模块未加载'}
            return suidong_horseshoe_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                               v_min=v_min, v_max=v_max,
                                               manual_B=b if b > 0 else None,
                                               theta_deg=theta_deg if theta_deg > 0 else None,
                                               manual_increase_percent=_inc)
        elif "隧洞-马蹄形" in section_type:
            if not SUIDONG_AVAILABLE: return {'success': False, 'error_message': '隙洞计算模块未加载'}
            st = 2 if "Ⅱ型" in section_type else 1
            return suidong_horseshoe_std_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                                    v_min=v_min, v_max=v_max,
                                                    manual_r=R if R > 0 else None,
                                                    section_type=st,
                                                    manual_increase_percent=_inc)
        elif "矩形暗涵" in section_type:
            if not RECT_CULVERT_AVAILABLE: return {'success': False, 'error_message': '矩形暗涵模块未加载'}
            return rect_culvert_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                          v_min=v_min, v_max=v_max,
                                          manual_B=b if b > 0 else None,
                                          target_BH_ratio=beta if beta > 0 else None,
                                          manual_increase_percent=_inc)
        else:
            return {'success': False, 'error_message': f'不支持的断面类型: {section_type}'}

    # ================================================================
    # 结果提取
    # ================================================================
    def _extract_result_row(self, seq, segment, building_name, section_type, result, use_increase=True):
        """从计算结果中提取关键数据为一行"""
        def fmt(val, default="-"):
            if val is None or val == 0: return default
            if isinstance(val, (int, float)): return f"{val:.3f}"
            return str(val)
        def fmt_pct(val, default="-"):
            if val is None or val == 0: return default
            if isinstance(val, (int, float)): return f"{val:.1f}"
            return str(val)

        B_val = D_val = R_val = "-"
        h_design = V_design = A_design = R_hyd = chi = "-"
        Q_inc = h_inc = V_inc = "-"
        Fb_surcharge = H_total_val = "-"
        Fb_cl_d = Fb_cl_i = Fb_pct_d = Fb_pct_i = "-"

        if "明渠" in section_type:
            if "圆形" in section_type:
                D_val = result.get('D_design', result.get('D', 0))
            elif "U形" in section_type:
                R_val = result.get('R', 0)
            else:
                B_val = result.get('b_design', result.get('B', 0))
            h_design = result.get('h_design', result.get('y_d', 0))
            V_design = result.get('V_design', result.get('V_d', 0))
            A_design = result.get('A_design', result.get('A_d', 0))
            R_hyd = result.get('R_design', result.get('R_d', 0))
            chi = result.get('X_design', result.get('P_d', 0))
            Q_inc = result.get('Q_increased', result.get('Q_inc', 0))
            h_inc = result.get('h_increased', result.get('y_i', 0))
            V_inc = result.get('V_increased', result.get('V_i', 0))
            if "圆形" in section_type:
                Fb_surcharge = result.get('FB_d', 0)
                Fb_cl_d = result.get('FB_d', 0)
                Fb_cl_i = result.get('FB_i', 0)
                Fb_pct_d = result.get('PA_d', "-")
                Fb_pct_i = result.get('PA_i', "-")
                H_total_val = result.get('D', result.get('D_design', 0))
            else:
                Fb_surcharge = result.get('Fb', result.get('FB_d', 0))
                H_total_val = result.get('h_prime', 0)

        elif "渡槽" in section_type:
            if "U形" in section_type:
                R_val = result.get('R', 0)
            else:
                B_val = result.get('B', 0)
            h_design = result.get('h_design', 0)
            V_design = result.get('V_design', 0)
            A_design = result.get('A_design', 0)
            R_hyd = result.get('R_hyd_design', 0)
            chi = result.get('P_design', 0)
            Q_inc = result.get('Q_increased', 0)
            h_inc = result.get('h_increased', 0)
            V_inc = result.get('V_increased', 0)
            Fb_surcharge = result.get('Fb', 0)
            H_total_val = result.get('H_total', 0)

        elif "隧洞" in section_type:
            if "圆形" in section_type:
                D_val = result.get('D', result.get('D_design', 0))
            elif "圆拱直墙型" in section_type:
                B_val = result.get('B', 0)
            elif "马蹄形" in section_type:
                R_val = result.get('r', 0)
            h_design = result.get('h_design', result.get('y_d', 0))
            V_design = result.get('V_design', result.get('V_d', 0))
            A_design = result.get('A_design', result.get('A_d', 0))
            R_hyd = result.get('R_hyd_design', result.get('R_d', 0))
            chi = result.get('P_design', result.get('P_d', 0))
            Q_inc = result.get('Q_increased', result.get('Q_inc', 0))
            h_inc = result.get('h_increased', result.get('y_i', 0))
            V_inc = result.get('V_increased', result.get('V_i', 0))
            if "圆形" in section_type:
                H_total_val = result.get('D', result.get('D_design', 0))
            elif "马蹄形" in section_type:
                r_v = result.get('r', 0)
                H_total_val = 2 * r_v if r_v else 0
            else:
                H_total_val = result.get('H_total', 0)
            Fb_cl_d = result.get('freeboard_hgt_design', result.get('FB_d', 0))
            Fb_cl_i = result.get('freeboard_hgt_inc', result.get('FB_i', 0))
            Fb_pct_d = result.get('freeboard_pct_design', 0)
            Fb_pct_i = result.get('freeboard_pct_inc', 0)

        elif "矩形暗涵" in section_type:
            B_val = result.get('B', 0)
            h_design = result.get('h_design', 0)
            V_design = result.get('V_design', 0)
            A_design = result.get('A_design', 0)
            R_hyd = result.get('R_hyd_design', 0)
            chi = result.get('P_design', 0)
            Q_inc = result.get('Q_increased', 0)
            h_inc = result.get('h_increased', 0)
            V_inc = result.get('V_increased', 0)
            H_total_val = result.get('H', result.get('H_total', 0))
            Fb_cl_d = result.get('freeboard_hgt_design', 0)
            Fb_cl_i = result.get('freeboard_hgt_inc', 0)
            Fb_pct_d = result.get('freeboard_pct_design', 0)
            Fb_pct_i = result.get('freeboard_pct_inc', 0)

        if not use_increase:
            Q_inc = h_inc = V_inc = "—"
            Fb_cl_i = Fb_pct_i = "—"

        return [seq, segment, building_name, section_type,
                fmt(B_val), fmt(D_val), fmt(R_val),
                fmt(h_design), fmt(V_design), fmt(A_design), fmt(R_hyd), fmt(chi),
                Q_inc if Q_inc == "—" else fmt(Q_inc),
                h_inc if h_inc == "—" else fmt(h_inc),
                V_inc if V_inc == "—" else fmt(V_inc),
                fmt(Fb_surcharge), fmt(H_total_val),
                fmt(Fb_cl_d),
                Fb_cl_i if Fb_cl_i == "—" else fmt(Fb_cl_i),
                fmt_pct(Fb_pct_d),
                Fb_pct_i if Fb_pct_i == "—" else fmt_pct(Fb_pct_i),
                "✓ 成功"]

    # ================================================================
    # 详细报告（与原版一致，按断面类型分发）
    # ================================================================
    def _gen_detail_report(self, input_vals, result):
        """生成单个建筑物的详细计算报告"""
        section_type = str(input_vals[3]).strip()
        building_name = str(input_vals[2]).strip()
        seq = str(input_vals[0]).strip()
        segment = str(input_vals[1]).strip()
        channel_name = self.channel_name_edit.text().strip()
        channel_level = self.channel_level_combo.currentText()
        header = f"  序号: {seq} | 建筑物: {building_name} | 类型: {section_type}"
        if channel_name or channel_level:
            header += f"\n  渠道: {channel_name} {channel_level} | 流量段: 第{segment}段"
        coord_x = str(input_vals[4]).strip() if len(input_vals) > 4 else ""
        coord_y = str(input_vals[5]).strip() if len(input_vals) > 5 else ""
        if coord_x and coord_y:
            header += f"\n  坐标: X={coord_x}, Y={coord_y}"
        header += "\n" + "-" * 50 + "\n"
        return header + self._gen_detail_body(input_vals, result)

    def _gen_detail_body(self, input_vals, result):
        """生成详细报告正文（不含标题头，供Word导出和文本显示共用）"""
        section_type = str(input_vals[3]).strip()
        try:
            if "倒虹吸" in section_type:
                return self._fmt_placeholder_report(input_vals, result, "倒虹吸")
            if "分水" in section_type or "闸" in section_type:
                return self._fmt_diversion_gate_report(input_vals, result)
            if "明渠" in section_type:
                return self._fmt_mingqu_report(input_vals, result)
            elif "渡槽" in section_type:
                return self._fmt_ducao_report(input_vals, result)
            elif "隧洞" in section_type or "矩形暗涵" in section_type:
                return self._fmt_suidong_report(input_vals, result)
            else:
                return "不支持的断面类型，无法生成详细报告。"
        except Exception as e:
            return f"生成详细报告时出错: {str(e)}"

    def _channel_info_lines(self, input_vals):
        """生成渠道信息行列表，用于详细报告输入参数部分"""
        lines = []
        channel_name = self.channel_name_edit.text().strip()
        channel_level = self.channel_level_combo.currentText()
        building_name = str(input_vals[2]).strip()
        segment = str(input_vals[1]).strip()
        if channel_name:
            lines.append(f"  渠道: {channel_name} {channel_level}")
        if building_name and building_name != "-":
            lines.append(f"  建筑物: {building_name}")
        lines.append(f"  流量段: 第{segment}段")
        coord_x = str(input_vals[4]).strip() if len(input_vals) > 4 else ""
        coord_y = str(input_vals[5]).strip() if len(input_vals) > 5 else ""
        if coord_x and coord_y:
            lines.append(f"  坐标: X={coord_x}, Y={coord_y}")
        return lines

    def _fmt_placeholder_report(self, input_vals, result, type_name):
        """格式化占位行报告（倒虹吸等）"""
        output = []
        output.append(f"  ⏭ {type_name}占位行，不参与断面计算。")
        channel_name = self.channel_name_edit.text().strip()
        channel_level = self.channel_level_combo.currentText()
        segment = str(input_vals[1]).strip() if len(input_vals) > 1 else ""
        if channel_name:
            output.append(f"  渠道: {channel_name} {channel_level}")
        if segment:
            output.append(f"  流量段: 第{segment}段")
        coord_x = input_vals[4] if len(input_vals) > 4 else ""
        coord_y = input_vals[5] if len(input_vals) > 5 else ""
        if coord_x and coord_y and str(coord_x).strip() and str(coord_y).strip():
            output.append(f"  坐标: X={coord_x}, Y={coord_y}")
        return "\n".join(output)

    def _fmt_diversion_gate_report(self, input_vals, result):
        """格式化分水闸/分水口详细报告"""
        section_type = str(input_vals[3]).strip()
        segment = str(input_vals[1]).strip()
        o = []
        o.append(f"  ⏭ {section_type} — 流量段分界点（不参与断面计算）")
        o.append("")
        o.append("【基本信息】")
        channel_name = self.channel_name_edit.text().strip()
        channel_level = self.channel_level_combo.currentText()
        if channel_name:
            o.append(f"  渠道: {channel_name} {channel_level}")
        o.append(f"  所在流量段: 第{segment}段")
        coord_x = input_vals[4] if len(input_vals) > 4 else ""
        coord_y = input_vals[5] if len(input_vals) > 5 else ""
        if coord_x and coord_y and str(coord_x).strip() and str(coord_y).strip():
            o.append(f"  坐标: X={coord_x}, Y={coord_y}")
        q_val = str(input_vals[6]).strip() if len(input_vals) > 6 and input_vals[6] else ""
        if q_val:
            o.append(f"  本段设计流量 Q = {q_val} m³/s")
        o.append("")
        o.append("【说明】")
        try:
            seg_int = int(segment)
            o.append(f"  {section_type}为流量段分界构筑物，其下游为新的流量段（流量段{seg_int+1}）。")
        except ValueError:
            o.append(f"  {section_type}为流量段分界构筑物。")
        o.append("  该行不参与断面尺寸计算，在推求水面线模块中将产生过闸水头损失（默认0.2m）。")
        return "\n".join(o)

    def _fmt_mingqu_report(self, input_vals, result):
        """格式化明渠报告（梯形/矩形/圆形，与原版一致）"""
        Q = self._sf(input_vals[6])
        n = self._sf(input_vals[7], 0.014)
        slope_inv = self._sf(input_vals[8])
        v_min = self._sf(input_vals[18], 0.1)
        v_max = self._sf(input_vals[19], 100)
        section_type = str(input_vals[3]).strip()
        i = 1.0 / slope_inv if slope_inv > 0 else 0
        o = []

        if "U形" in section_type and "明渠" in section_type:
            R_u = result.get('R', 0)
            alpha_u = result.get('alpha_deg', 0)
            theta_u = result.get('theta_deg', 0)
            m_u = result.get('m', 0)
            h0_u = result.get('h0', 0)
            b_arc_u = result.get('b_arc', 0)
            h_d = result.get('h_design', 0)
            V_d = result.get('V_design', 0)
            A_d = result.get('A_design', 0)
            chi_d = result.get('X_design', 0)
            Rh_d = result.get('R_design', 0)
            Q_inc = result.get('Q_increased', 0)
            h_i = result.get('h_increased', 0)
            V_i = result.get('V_increased', 0)
            A_i = result.get('A_increased', 0)
            Fb = result.get('Fb', 0)
            H = result.get('h_prime', 0)
            inc_pct = result.get('increase_percent', 0)
            o.append("【一、输入参数】")
            o.extend(self._channel_info_lines(input_vals))
            o.append(f"  断面类型 = 明渠-U形")
            o.append(f"  Q = {Q:.3f} m³/s,  n = {n},  坡度倒数 = {int(slope_inv)}")
            o.append(f"  R = {R_u:.3f} m,  α = {alpha_u}°,  θ = {theta_u}°")
            o.append(f"  不淤流速 = {v_min} m/s,  不冲流速 = {v_max} m/s")
            o.append("")
            o.append("【二、断面几何参数】")
            o.append(f"  m = tan(α) = {m_u:.4f},  h_0 = {h0_u:.3f} m,  b_{{arc}} = {b_arc_u:.3f} m")
            o.append("")
            o.append("【三、设计水深工况】")
            o.append(f"  h = {h_d:.3f} m,  A = {A_d:.3f} m²,  χ = {chi_d:.3f} m")
            o.append(f"  R_h = {Rh_d:.3f} m,  V = {V_d:.3f} m/s")
            o.append("")
            o.append("【四、加大流量工况】")
            o.append(f"  加大比例 = {inc_pct:.1f}%,  Q加大 = {Q_inc:.3f} m³/s")
            o.append(f"  h加大 = {h_i:.3f} m,  V加大 = {V_i:.3f} m/s")
            o.append(f"  Fb = {Fb:.3f} m,  H = {H:.3f} m")
            o.append("")
            o.append("【五、验证】")
            velocity_ok = v_min <= V_d <= v_max
            fb_ok = Fb >= (0.25 * h_i + 0.2 - 0.001) if h_i > 0 else True
            o.append(f"  1. 流速验证: {v_min} ≤ V={V_d:.3f} ≤ {v_max} m/s → {'通过 ✓' if velocity_ok else '未通过 ✗'}")
            o.append(f"  2. 超高验证: Fb={Fb:.3f}m → {'通过 ✓' if fb_ok else '未通过 ✗'}")
        elif "圆形" in section_type:
            D_design = result.get('D_design', result.get('D', 0))
            h_d = result.get('y_d', result.get('h_design', 0))
            V_d = result.get('V_d', result.get('V_design', 0))
            A_d = result.get('A_d', result.get('A_design', 0))
            P_d = result.get('P_d', result.get('P_design', 0))
            R_d = result.get('R_d', result.get('R_hyd_design', 0))
            PA_d = result.get('PA_d', 0)
            FB_d = result.get('FB_d', 0)
            o.append("【一、输入参数】")
            o.extend(self._channel_info_lines(input_vals))
            o.append(f"  断面类型 = 圆形")
            o.append(f"  设计流量 Q = {Q:.3f} m³/s")
            o.append(f"  糙率 n = {n}")
            o.append(f"  水力坡降 = 1/{int(slope_inv)}")
            o.append(f"  不淤流速 = {v_min} m/s")
            o.append(f"  不冲流速 = {v_max} m/s")
            o.append("")
            o.append("【二、管径确定】")
            o.append(f"  1. 设计管径: D = {D_design:.2f} m")
            import math
            pipe_area = math.pi * D_design**2 / 4
            o.append(f"  2. 管道总断面积: A总 = π×D²/4 = {pipe_area:.3f} m²")
            o.append("")
            o.append("【三、设计流量工况计算】")
            o.append(f"  设计水深 h_d = {h_d:.3f} m")
            o.append(f"  过水面积 A_d = {A_d:.3f} m²")
            o.append(f"  湿周 χ_d = {P_d:.3f} m")
            o.append(f"  水力半径 R_d = {R_d:.3f} m")
            o.append(f"  设计流速 V_d = {V_d:.3f} m/s")
            o.append(f"  净空高度 Fb_d = D - h_d = {D_design:.2f} - {h_d:.3f} = {FB_d:.3f} m")
            o.append(f"  净空面积百分比 PA_d = {PA_d:.1f}%")
            o.append("")
            # 加大流量
            Q_inc = result.get('Q_inc', result.get('Q_increased', 0))
            h_i = result.get('y_i', result.get('h_increased', 0))
            V_i = result.get('V_i', result.get('V_increased', 0))
            A_i = result.get('A_i', result.get('A_increased', 0))
            PA_i = result.get('PA_i', 0)
            FB_i = result.get('FB_i', 0)
            o.append("【四、加大流量工况计算】")
            o.append(f"  加大流量 Q_i = {Q_inc:.3f} m³/s")
            o.append(f"  加大水深 h_i = {h_i:.3f} m")
            o.append(f"  加大流速 V_i = {V_i:.3f} m/s")
            o.append(f"  净空高度 Fb_i = {FB_i:.3f} m")
            o.append(f"  净空面积百分比 PA_i = {PA_i:.1f}%")
            o.append("")
            o.append("【五、验证】")
            velocity_ok = v_min <= V_d <= v_max
            o.append(f"  1. 流速验证: {v_min} ≤ V ≤ {v_max} m/s → V = {V_d:.3f} → {'通过 ✓' if velocity_ok else '未通过 ✗'}")
            fb_ok = FB_i >= 0.4
            o.append(f"  2. 净空高度验证: Fb加大 = {FB_i:.3f}m ≥ 0.4m → {'通过 ✓' if fb_ok else '未通过 ✗'}")
            pa_ok = PA_i >= 15
            o.append(f"  3. 净空面积验证: PA加大 = {PA_i:.1f}% ≥ 15% → {'通过 ✓' if pa_ok else '未通过 ✗'}")
        else:
            # 梯形/矩形明渠
            m = float(input_vals[9]) if input_vals[9] else 0
            b = result.get('b_design', result.get('B', 0))
            h = result.get('h_design', 0)
            V = result.get('V_design', 0)
            A = result.get('A_design', 0)
            chi = result.get('X_design', result.get('P_design', 0))
            R = result.get('R_design', result.get('R_hyd_design', 0))
            beta = result.get('Beta_design', 0)
            import math
            o.append("【一、输入参数】")
            o.extend(self._channel_info_lines(input_vals))
            o.append(f"  断面类型 = {section_type}")
            o.append(f"  设计流量 Q = {Q:.3f} m³/s")
            if "梯形" in section_type:
                o.append(f"  边坡系数 m = {m}")
            o.append(f"  糙率 n = {n}")
            o.append(f"  水力坡降 i = 1/{int(slope_inv)} = {i:.6f}")
            o.append(f"  不淤流速 = {v_min} m/s")
            o.append(f"  不冲流速 = {v_max} m/s")
            o.append("")
            o.append("【二、设计方法】")
            o.append(f"  采用方法: {result.get('design_method', '自动计算')}")
            o.append("")
            o.append("【三、设计结果】")
            o.append(f"  1. 设计底宽 B = {b:.2f} m")
            o.append(f"  2. 设计水深 h = {h:.3f} m")
            o.append(f"  3. 宽深比 β = B/h = {b:.2f}/{h:.3f} = {beta:.3f}" if h > 0 else f"  3. 宽深比 β = {beta:.3f}")
            o.append(f"  4. 过水面积 A = (B + m×h)×h = ({b:.2f} + {m}×{h:.3f})×{h:.3f} = {A:.3f} m²")
            sqrt_1_m2 = math.sqrt(1 + m*m)
            o.append(f"  5. 湿周 χ = B + 2×h×√(1+m²) = {b:.2f} + 2×{h:.3f}×{sqrt_1_m2:.4f} = {chi:.3f} m")
            o.append(f"  6. 水力半径 R = A/χ = {A:.3f}/{chi:.3f} = {R:.3f} m")
            o.append(f"  7. 设计流速 V = (1/n)×R^(2/3)×i^(1/2) = {V:.3f} m/s")
            o.append(f"  8. 流量校核 Q计算 = V×A = {V:.3f}×{A:.3f} = {V*A:.3f} m³/s")
            if Q > 0:
                o.append(f"     误差 = {abs(V*A-Q)/Q*100:.2f}%")
            o.append("")
            o.append("【四、加大流量工况】")
            inc_pct = result.get('increase_percent', 0)
            Q_inc = result.get('Q_increased', 0)
            h_inc = result.get('h_increased', 0)
            V_inc = result.get('V_increased', 0)
            A_inc = result.get('A_increased', 0)
            Fb = result.get('Fb', 0)
            H = result.get('h_prime', 0)
            o.append(f"  加大比例 = {inc_pct:.1f}%")
            o.append(f"  1. 加大流量 Q加大 = {Q_inc:.3f} m³/s")
            o.append(f"  2. 加大水深 h加大 = {h_inc:.3f} m")
            o.append(f"  3. 加大流速 V加大 = {V_inc:.3f} m/s")
            o.append(f"  4. 超高 Fb = (1/4)×h加大 + 0.2 = {Fb:.3f} m")
            o.append(f"  5. 渠道高度 H = h加大 + Fb = {H:.3f} m")
            o.append("")
            o.append("【五、验证】")
            velocity_ok = v_min <= V <= v_max
            o.append(f"  1. 流速验证: {v_min} ≤ V ≤ {v_max} → V = {V:.3f} → {'通过 ✓' if velocity_ok else '未通过 ✗'}")
            fb_req = 0.25 * h_inc + 0.2
            fb_ok = Fb >= (fb_req - 0.001)
            o.append(f"  2. 超高复核（规范 6.4.8-2）: Fb = {Fb:.3f}m ≥ {fb_req:.3f}m → {'通过 ✓' if fb_ok else '未通过 ✗'}")
        return "\n".join(o)

    def _fmt_ducao_report(self, input_vals, result):
        """格式化渡槽报告（U形/矩形，与原版一致）"""
        Q = self._sf(input_vals[6])
        n = self._sf(input_vals[7], 0.014)
        slope_inv = self._sf(input_vals[8])
        section_type = str(input_vals[3]).strip()
        i = 1.0 / slope_inv if slope_inv > 0 else 0
        import math
        o = []
        v_min = self._sf(input_vals[18], 0.1)
        v_max = self._sf(input_vals[19], 100)
        o.append("【一、输入参数】")
        o.extend(self._channel_info_lines(input_vals))
        o.append(f"  设计流量 Q = {Q:.3f} m³/s")
        o.append(f"  糙率 n = {n}")
        o.append(f"  水力坡降 i = 1/{int(slope_inv)} = {i:.6f}")
        o.append(f"  不淤流速 = {v_min} m/s")
        o.append(f"  不冲流速 = {v_max} m/s")
        o.append("")

        H_total = result.get('H_total', 0)
        if "U形" in section_type:
            R_val = result.get('R', 0)
            f = result.get('f', 0)
            B = result.get('B', 0)
            f_R = result.get('f_R', 0)
            o.append("【二、断面尺寸】")
            o.append(f"  内半径 R = {R_val:.2f} m")
            o.append(f"  槽宽 B = 2×R = {B:.2f} m")
            o.append(f"  直段高度 f = {f:.2f} m, f/R = {f_R:.3f}")
            o.append(f"  槽身总高 H = R + f = {R_val:.2f} + {f:.2f} = {H_total:.2f} m")
        else:
            B = result.get('B', 0)
            ratio = result.get('depth_width_ratio', 0)
            o.append("【二、断面尺寸】")
            o.append(f"  槽宽 B = {B:.2f} m")
            o.append(f"  深宽比 = {ratio:.3f}")
            o.append(f"  槽高 H = B × 深宽比 = {B:.2f} × {ratio:.3f} = {H_total:.2f} m")
            if result.get('has_chamfer', False):
                o.append(f"  倒角参数: 角度 {result.get('chamfer_angle', 0):.1f}°, 底边 {result.get('chamfer_length', 0):.2f} m")

        o.append("")
        o.append("【三、设计流量工况】")
        h_design = result.get('h_design', 0)
        A_design = result.get('A_design', 0)
        P_design = result.get('P_design', 0)
        R_hyd = result.get('R_hyd_design', 0)
        V_design = result.get('V_design', 0)
        o.append(f"  设计水深 h = {h_design:.3f} m")
        o.append(f"  过水面积 A = {A_design:.3f} m²")
        o.append(f"  湿周 P = {P_design:.3f} m")
        o.append(f"  水力半径 R = A/P = {R_hyd:.3f} m")
        o.append(f"  设计流速 V = (1/n)×R^(2/3)×i^(1/2) = {V_design:.3f} m/s")
        o.append(f"  流量校核 Q计算 = A×V = {A_design*V_design:.3f} m³/s")

        o.append("")
        o.append("【四、加大流量工况】")
        inc_pct = result.get('increase_percent', 0)
        Q_inc = result.get('Q_increased', 0)
        h_inc = result.get('h_increased', 0)
        V_inc = result.get('V_increased', 0)
        A_inc = result.get('A_increased', 0)
        P_inc = result.get('P_increased', 0)
        R_inc = result.get('R_hyd_increased', 0)
        Fb = result.get('Fb', 0)
        o.append(f"  加大比例 = {inc_pct:.1f}%")
        o.append(f"  加大流量 Q加大 = {Q_inc:.3f} m³/s")
        o.append(f"  加大水深 h加大 = {h_inc:.3f} m")
        o.append(f"  加大过水面积 A加大 = {A_inc:.3f} m²")
        o.append(f"  加大湿周 P加大 = {P_inc:.3f} m")
        o.append(f"  加大水力半径 R加大 = {R_inc:.3f} m")
        o.append(f"  加大流速 V加大 = {V_inc:.3f} m/s")
        o.append(f"  流量校核 Q计算 = {A_inc*V_inc:.3f} m³/s")
        o.append(f"  超高 Fb = H - h加大 = {H_total:.2f} - {h_inc:.3f} = {Fb:.3f} m")

        o.append("")
        o.append("【五、验证】")
        v_rec_min, v_rec_max = 1.0, 2.5
        velocity_ok = v_rec_min <= V_design <= v_rec_max
        o.append(f"  1. 流速验证（规范 9.4.1-1）: 宜为 {v_rec_min}～{v_rec_max} m/s")
        o.append(f"     V = {V_design:.3f} m/s → {'通过 ✓' if velocity_ok else '超出推荐范围 ⚠'}")

        o.append(f"  2. 超高验证（规范 9.4.1-2）")
        if "U形" in section_type:
            R_val = result.get('R', 0)
            Fb_design = H_total - h_design
            Fb_design_min = R_val / 5
            Fb_inc_min = 0.10
            o.append(f"     设计流量超高: Fb_设计 = {Fb_design:.3f}m ≥ R/5 = {Fb_design_min:.3f}m → {'通过 ✓' if Fb_design >= Fb_design_min else '未通过 ✗'}")
            o.append(f"     加大流量超高: Fb_加大 = {Fb:.3f}m ≥ {Fb_inc_min:.2f}m → {'通过 ✓' if Fb >= Fb_inc_min else '未通过 ✗'}")
        else:
            Fb_design = H_total - h_design
            Fb_design_min = h_design / 12 + 0.05
            Fb_inc_min = 0.10
            o.append(f"     设计流量超高: Fb_设计 = {Fb_design:.3f}m ≥ h/12+0.05 = {Fb_design_min:.3f}m → {'通过 ✓' if Fb_design >= Fb_design_min else '未通过 ✗'}")
            o.append(f"     加大流量超高: Fb_加大 = {Fb:.3f}m ≥ {Fb_inc_min:.2f}m → {'通过 ✓' if Fb >= Fb_inc_min else '未通过 ✗'}")
        return "\n".join(o)

    def _fmt_suidong_report(self, input_vals, result):
        """格式化隧洞/矩形暗涵报告（与原版一致）"""
        Q = self._sf(input_vals[6])
        n = self._sf(input_vals[7], 0.014)
        slope_inv = self._sf(input_vals[8])
        v_min = self._sf(input_vals[18], 0.1)
        v_max = self._sf(input_vals[19], 100)
        section_type = str(input_vals[3]).strip()
        i = 1.0 / slope_inv if slope_inv > 0 else 0
        o = []
        o.append("【一、输入参数】")
        o.extend(self._channel_info_lines(input_vals))
        o.append(f"  设计流量 Q = {Q:.3f} m³/s")
        o.append(f"  糙率 n = {n}")
        o.append(f"  水力坡降 = 1/{int(slope_inv)}")
        o.append(f"  不淤流速 = {v_min} m/s")
        o.append(f"  不冲流速 = {v_max} m/s")
        o.append("")

        h_design = result.get('h_design', 0)
        A_design = result.get('A_design', 0)
        P_design = result.get('P_design', 0)
        R_hyd = result.get('R_hyd_design', 0)
        V_design = result.get('V_design', 0)
        fb_pct_design = result.get('freeboard_pct_design', 0)
        fb_hgt_design = result.get('freeboard_hgt_design', 0)
        A_total = result.get('A_total', 0)
        import math
        PI = math.pi

        o.append("【二、断面尺寸】")
        if "圆形" in section_type:
            D = result.get('D', 0)
            o.append(f"  设计直径: D = {D:.2f} m")
            o.append(f"  断面总面积: A总 = π×D²/4 = {A_total:.3f} m²")
        elif "圆拱直墙型" in section_type:
            B = result.get('B', 0)
            H = result.get('H_total', 0)
            theta_deg = result.get('theta_deg', 0)
            o.append(f"  设计宽度: B = {B:.2f} m")
            o.append(f"  设计高度: H = {H:.2f} m")
            o.append(f"  拱顶圆心角: θ = {theta_deg:.1f}°")
            o.append(f"  高宽比: H/B = {H/B:.3f}" if B > 0 else "  高宽比: N/A")
            o.append(f"  断面总面积: A总 = {A_total:.3f} m²")
        elif "马蹄形" in section_type:
            r = result.get('r', 0)
            o.append(f"  设计半径: R = {r:.2f} m")
            o.append(f"  等效直径: 2R = {2*r:.2f} m")
            o.append(f"  断面总面积: A总 = {A_total:.3f} m²")
        elif "矩形暗涵" in section_type:
            B = result.get('B', 0)
            H = result.get('H', 0)
            BH_ratio = result.get('BH_ratio', 0)
            HB_ratio_r = result.get('HB_ratio', H / B if B > 0 else 0)
            BH_box_r = B / H if H > 0 else 0
            hb_ok = result.get('hb_ratio_ok', True)
            bh_ok = result.get('bh_box_ratio_ok', True)
            o.append(f"  设计宽度: B = {B:.2f} m")
            o.append(f"  设计高度: H = {H:.2f} m")
            o.append(f"  宽深比 β = B/h_设计 = {BH_ratio:.3f}" if BH_ratio else "")
            o.append(f"  高宽比 H/B = {HB_ratio_r:.3f}" + ("" if hb_ok else "  ⚠ 超出建议值1.2"))
            o.append(f"  宽高比 B/H = {BH_box_r:.3f}" + ("" if bh_ok else "  ⚠ 超出建议值1.2"))
            o.append(f"  总断面积: A总 = B×H = {A_total:.3f} m²")

        o.append("")
        o.append("【三、设计流量工况】")
        o.append(f"  设计水深 h = {h_design:.3f} m")
        o.append(f"  过水面积 A = {A_design:.3f} m²")
        o.append(f"  湿周 χ = {P_design:.3f} m")
        o.append(f"  水力半径 R = A/χ = {R_hyd:.3f} m")
        o.append(f"  设计流速 V = (1/n)×R^(2/3)×i^(1/2) = {V_design:.3f} m/s")
        o.append(f"  流量校核 Q计算 = A×V = {A_design*V_design:.3f} m³/s")
        o.append(f"  净空高度 Fb = {fb_hgt_design:.3f} m")
        o.append(f"  净空面积比例 = {fb_pct_design:.1f}%")

        use_increase = result.get('_use_increase', True)
        inc_pct = result.get('increase_percent', 0)
        Q_inc = result.get('Q_increased', 0)
        h_inc = result.get('h_increased', 0)
        V_inc = result.get('V_increased', 0)
        A_inc = result.get('A_increased', 0)
        P_inc = result.get('P_increased', 0)
        R_inc = result.get('R_hyd_increased', 0)
        fb_pct_inc = result.get('freeboard_pct_inc', 0)
        fb_hgt_inc = result.get('freeboard_hgt_inc', 0)
        H_total_val = result.get('H_total', result.get('H', result.get('D', result.get('r', 0) * 2)))

        if use_increase:
            o.append("")
            o.append("【四、加大流量工况】")
            o.append(f"  加大比例 = {inc_pct:.1f}%")
            o.append(f"  加大流量 Q加大 = {Q_inc:.3f} m³/s")
            o.append(f"  加大水深 h加大 = {h_inc:.3f} m")
            o.append(f"  加大过水面积 A加大 = {A_inc:.3f} m²")
            o.append(f"  加大湿周 P加大 = {P_inc:.3f} m")
            o.append(f"  加大水力半径 R加大 = {R_inc:.3f} m")
            o.append(f"  加大流速 V加大 = {V_inc:.3f} m/s")
            o.append(f"  流量校核 Q计算 = {A_inc*V_inc:.3f} m³/s")
            o.append(f"  净空高度 Fb加大 = H - h加大 = {H_total_val:.2f} - {h_inc:.3f} = {fb_hgt_inc:.3f} m")
            o.append(f"  净空面积比例 = {fb_pct_inc:.1f}%")

        # 验证结论：不勾选加大流量时用设计工况数据
        fb_hgt_check = fb_hgt_inc if use_increase else fb_hgt_design
        fb_pct_check = fb_pct_inc if use_increase else fb_pct_design
        fb_label = "Fb加大" if use_increase else "Fb设计"

        o.append("")
        section_num = "五" if use_increase else "四"
        o.append(f"【{section_num}、验证结论】")
        velocity_ok = v_min <= V_design <= v_max
        o.append(f"  1. 流速验证: {v_min} ≤ {V_design:.3f} ≤ {v_max} → {'通过 ✓' if velocity_ok else '未通过 ✗'}")
        min_fb_hgt = 0.4
        is_culvert = "矩形暗涵" in section_type
        min_fb_pct = 10.0 if is_culvert else 15.0
        max_fb_pct = 30.0 if is_culvert else None
        fb_hgt_ok = fb_hgt_check >= min_fb_hgt
        fb_pct_ok = fb_pct_check >= min_fb_pct and (max_fb_pct is None or fb_pct_check <= max_fb_pct)
        o.append(f"  2. 净空高度验证: {fb_label} = {fb_hgt_check:.3f}m ≥ {min_fb_hgt}m → {'通过 ✓' if fb_hgt_ok else '需注意'}")
        if is_culvert:
            o.append(f"  3. 净空比例验证: {min_fb_pct}% ≤ {fb_pct_check:.1f}% ≤ {max_fb_pct}% → {'通过 ✓' if fb_pct_ok else '需注意'}")
        else:
            o.append(f"  3. 净空比例验证: {fb_pct_check:.1f}% ≥ {min_fb_pct}% → {'通过 ✓' if fb_pct_ok else '需注意'}")
        return "\n".join(o)

    # ================================================================
    # 参数设置弹窗（与原版一致）
    # ================================================================
    def _open_parameter_dialog(self):
        """打开参数设置弹窗（选中行或提示选择行）"""
        selected = self.input_table.selectedIndexes()
        if not selected:
            InfoBar.warning("提示", "请先选择要设置参数的行", parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return
        row_idx = selected[0].row()
        self._open_parameter_dialog_for_row(row_idx)

    def _open_parameter_dialog_for_row(self, row_idx):
        """为指定行打开参数设置弹窗（与原版一致）"""
        if row_idx < 0 or row_idx >= self.input_table.rowCount():
            return
        values = self._get_row_data(row_idx)
        values = self._normalize_row(values, len(INPUT_HEADERS))
        section_type = str(values[3]).strip()
        if not section_type or "分水" in section_type or "闸" in section_type or "倒虹吸" in section_type:
            InfoBar.info("提示", f"{section_type or '未设置类型'} 无需设置断面参数",
                        parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return
        # 列索引: 6Q,7n,8比降,9m,10B,11宽深比,12R,13D,14渡槽深宽比,15倒角角度,16倒角底边,17圆心角,18不淤,19不冲
        current_values = {
            "Q": values[6], "n": values[7], "slope_inv": values[8],
            "m": values[9], "b": values[10], "b_h_ratio": values[11],
            "R": values[12], "D": values[13],
            "ducao_depth_ratio": values[14], "chamfer_angle": values[15],
            "chamfer_length": values[16], "theta": values[17],
            "v_min": values[18], "v_max": values[19],
        }
        dlg = SectionParameterDialog(self, section_type, current_values)
        if dlg.exec() == QDialog.Accepted:
            result = dlg.get_result()
            if result is not None:
                self._update_table_row(row_idx, result, section_type)

    def _update_table_row(self, row_idx, params, section_type):
        """将弹窗参数回填到表格（与原版一致）"""
        values = self._get_row_data(row_idx)
        values = list(self._normalize_row(values, len(INPUT_HEADERS)))
        # 更新基础参数
        values[6] = str(params.get('Q', values[6]))
        values[7] = str(params.get('n', values[7]))
        values[8] = str(params.get('slope_inv', values[8]))
        values[18] = str(params.get('v_min', values[18]))
        values[19] = str(params.get('v_max', values[19]))
        # 清空所有可选参数列(9-17)
        for i in range(9, 18):
            values[i] = ""
        # 根据结构形式更新对应的参数列
        if "明渠-梯形" in section_type or "明渠-矩形" in section_type:
            m_val = params.get('m', "")
            values[9] = str(m_val) if m_val != "" else ""
            b_val = params.get('b', "")
            values[10] = str(b_val) if b_val != "" else ""
            ratio_val = params.get('b_h_ratio', "")
            values[11] = str(ratio_val) if ratio_val != "" else ""
        elif "明渠-圆形" in section_type:
            D_val = params.get('D', "")
            values[13] = str(D_val) if D_val != "" else ""
        elif "渡槽-U形" in section_type:
            R_val = params.get('R', "")
            values[12] = str(R_val) if R_val != "" else ""
        elif "渡槽-矩形" in section_type:
            h_b_ratio = params.get('h_b_ratio', "")
            values[14] = str(h_b_ratio) if h_b_ratio != "" else ""
            chamfer_angle = params.get('chamfer_angle', "")
            chamfer_length = params.get('chamfer_length', "")
            values[15] = str(chamfer_angle) if chamfer_angle != "" else ""
            values[16] = str(chamfer_length) if chamfer_length != "" else ""
        elif "隧洞-圆形" in section_type:
            D_val = params.get('D', "")
            values[13] = str(D_val) if D_val != "" else ""
        elif "隧洞-圆拱直墙型" in section_type:
            B_val = params.get('B', "")
            values[10] = str(B_val) if B_val != "" else ""
            theta_val = params.get('theta', "")
            values[17] = str(theta_val) if theta_val != "" else ""
        elif "隧洞-马蹄形" in section_type:
            r_val = params.get('r', "")
            values[12] = str(r_val) if r_val != "" else ""
        elif section_type == "矩形暗涵":
            BH_ratio_val = params.get('BH_ratio_rect', "")
            values[11] = str(BH_ratio_val) if BH_ratio_val != "" else ""
            B_rect_val = params.get('B_rect', "")
            values[10] = str(B_rect_val) if B_rect_val != "" else ""
        # 回填到表格
        self.input_table.blockSignals(True)
        for col, val in enumerate(values):
            if col < len(INPUT_HEADERS):
                item = QTableWidgetItem(str(val) if val else "")
                if col == 3:
                    item.setTextAlignment(Qt.AlignCenter)
                self.input_table.setItem(row_idx, col, item)
        self.input_table.blockSignals(False)

    # ================================================================
    # 辅助函数
    # ================================================================
    def _map_section_type(self, raw_type):
        """将粘贴的断面类型简写映射为有效类型（与原版一致）"""
        mapping = {
            "梯形": "明渠-梯形", "矩形": "明渠-矩形", "圆形": "明渠-圆形",
            "明渠U形": "明渠-U形", "U形明渠": "明渠-U形",
            "U形": "渡槽-U形", "U形渡槽": "渡槽-U形", "矩形渡槽": "渡槽-矩形",
            "圆形隧洞": "隧洞-圆形", "圆拱直墙": "隧洞-圆拱直墙型", "圆拱直墙型": "隧洞-圆拱直墙型",
            "马蹄形Ⅰ型": "隧洞-马蹄形Ⅰ型", "马蹄形Ⅱ型": "隧洞-马蹄形Ⅱ型",
            "马蹄形I型": "隧洞-马蹄形Ⅰ型", "马蹄形II型": "隧洞-马蹄形Ⅱ型",
            "暗涵": "矩形暗涵",
        }
        if raw_type in mapping:
            return mapping[raw_type]
        # 模糊匹配
        for key, val in mapping.items():
            if key in raw_type:
                return val
        return None

    def _apply_section_type_change(self, row, new_type):
        """结构形式变更后自动填充默认值"""
        name_item = self.input_table.item(row, 2)
        current_name = name_item.text().strip() if name_item else ""
        seq_item = self.input_table.item(row, 0)
        seq = seq_item.text().strip() if seq_item else str(row + 1)

        self.input_table.blockSignals(True)

        if "明渠" in new_type or new_type == "矩形暗涵":
            self.input_table.setItem(row, 2, QTableWidgetItem("-"))
            # 明渠-U形额外填充默认值 R=0.8, α=14°, θ=152°
            if new_type == "明渠-U形":
                for col, val in [(12, "0.8"), (15, "14"), (17, "152")]:
                    self.input_table.setItem(row, col, QTableWidgetItem(val))
        elif "分水" in new_type:
            if current_name == "-" or not current_name:
                self.input_table.setItem(row, 2, QTableWidgetItem(f"分水闸{seq}"))
        elif current_name == "-" or not current_name:
            self.input_table.setItem(row, 2, QTableWidgetItem(f"建筑物{seq}"))

        if "分水" in new_type:
            for i in range(8, 18):
                self.input_table.setItem(row, i, QTableWidgetItem(""))
            self.input_table.blockSignals(False)
            InfoBar.info("提示", f"序号{seq}已设为{new_type}（流量段分界点）\n请确认下一行的流量段编号已正确递增",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        for i in range(9, 18):
            self.input_table.setItem(row, i, QTableWidgetItem(""))

        slope_item = self.input_table.item(row, 8)
        if not slope_item or not slope_item.text().strip():
            self.input_table.setItem(row, 8, QTableWidgetItem("3000"))

        if "明渠-梯形" in new_type:
            self.input_table.setItem(row, 9, QTableWidgetItem("1.0"))

        self.input_table.blockSignals(False)

    def _validate_duplicate_buildings(self, input_rows=None):
        """验证建筑物重名（与原版一致：相邻同名同结构视为同一建筑物，不相邻则为重复）"""
        if input_rows is None:
            input_rows = self._get_all_input_data()
        if not input_rows:
            return True
        # 按相邻关系分组
        groups = []
        prev_key = None
        for values in input_rows:
            values = self._normalize_row(values, len(INPUT_HEADERS))
            if not any(str(v).strip() for v in values):
                prev_key = None
                continue
            building_name = str(values[2]).strip()
            section_type = str(values[3]).strip()
            seq = str(values[0]).strip() or "?"
            segment = str(values[1]).strip() or "?"
            if not building_name or building_name == "-":
                prev_key = None
                continue
            if not section_type:
                prev_key = None
                continue
            if "分水" in section_type:
                continue
            key = (building_name, section_type)
            row_desc = f"序号{seq}(流量段{segment})"
            if key == prev_key and groups:
                groups[-1][2].append(row_desc)
            else:
                groups.append((building_name, section_type, [row_desc]))
                prev_key = key
        # 检查重复
        key_to_groups = {}
        for i, (name, stype, _) in enumerate(groups):
            key = (name, stype)
            key_to_groups.setdefault(key, []).append(i)
        duplicates = {k: idxs for k, idxs in key_to_groups.items() if len(idxs) > 1}
        if not duplicates:
            return True
        # 构建提示
        dup_details = []
        for (name, stype), group_idxs in duplicates.items():
            locations = []
            for gi in group_idxs:
                rows_str = "、".join(groups[gi][2])
                locations.append(rows_str)
            all_locations = " ⟷ ".join(locations)
            dup_details.append(f"  「{name}」({stype}) — 在 {len(group_idxs)} 处出现：{all_locations}")
        dup_msg = "\n".join(dup_details)
        fluent_info(self, "建筑物重名",
                   f"检测到不同位置的建筑物使用了相同的名称和结构形式：\n\n{dup_msg}\n\n"
                   "相邻行的同名建筑物已视为同一建筑物，以上为不同位置的重名。\n"
                   "请修改建筑物名称以区分后重试。")
        return False

    def _validate_duplicate_buildings_warn(self):
        """导入后检查建筑物重名（仅警告提示，不阻止操作，与原版show_warning_only=True一致）"""
        input_rows = self._get_all_input_data()
        if not input_rows:
            return
        groups = []
        prev_key = None
        for values in input_rows:
            values = self._normalize_row(values, len(INPUT_HEADERS))
            if not any(str(v).strip() for v in values):
                prev_key = None
                continue
            building_name = str(values[2]).strip()
            section_type = str(values[3]).strip()
            seq = str(values[0]).strip() or "?"
            segment = str(values[1]).strip() or "?"
            if not building_name or building_name == "-":
                prev_key = None
                continue
            if not section_type:
                prev_key = None
                continue
            if "分水" in section_type:
                continue
            key = (building_name, section_type)
            row_desc = f"序号{seq}(流量段{segment})"
            if key == prev_key and groups:
                groups[-1][2].append(row_desc)
            else:
                groups.append((building_name, section_type, [row_desc]))
                prev_key = key
        key_to_groups = {}
        for i, (name, stype, _) in enumerate(groups):
            key = (name, stype)
            key_to_groups.setdefault(key, []).append(i)
        duplicates = {k: idxs for k, idxs in key_to_groups.items() if len(idxs) > 1}
        if not duplicates:
            return
        dup_details = []
        for (name, stype), group_idxs in duplicates.items():
            locations = [" 、".join(groups[gi][2]) for gi in group_idxs]
            dup_details.append(f"「{name}」({stype})")
        InfoBar.warning("建筑物重名提示",
                       f"检测到重名建筑物: {', '.join(dup_details)}\n请在批量计算前修改名称以区分。",
                       parent=self._info_parent(), duration=6000, position=InfoBarPosition.TOP)

    def _auto_detect_flow_segments(self):
        """自动从表格数据中识别流量段及其对应的流量值，并更新流量设置区域"""
        segment_flow_map = {}
        for r in range(self.input_table.rowCount()):
            seg_item = self.input_table.item(r, 1)
            q_item = self.input_table.item(r, 6)
            seg_val = seg_item.text().strip() if seg_item else ""
            q_val = q_item.text().strip() if q_item else ""
            if not seg_val or not q_val:
                continue
            try:
                segment = int(seg_val)
                if segment not in segment_flow_map:
                    segment_flow_map[segment] = float(q_val)
            except (ValueError, TypeError):
                continue
        if not segment_flow_map:
            return
        max_segment = max(segment_flow_map.keys())
        flow_values = []
        for seg in range(1, max_segment + 1):
            if seg in segment_flow_map:
                flow_values.append(str(segment_flow_map[seg]))
            else:
                flow_values.append(flow_values[-1] if flow_values else "5.0")
        self.flow_segments_edit.setText(", ".join(flow_values))

    def _get_flow_for_segment(self, segment: int) -> float:
        """获取指定流量段的流量值（与原版get_flow_for_segment一致）"""
        try:
            flow_str = self.flow_segments_edit.text().strip()
            if not flow_str:
                return 5.0
            flow_str = flow_str.replace('，', ',')
            flow_values = [float(q.strip()) for q in flow_str.split(',') if q.strip()]
            if 1 <= segment <= len(flow_values):
                return flow_values[segment - 1]
            else:
                return flow_values[-1] if flow_values else 5.0
        except Exception:
            return 5.0

    def _apply_flow_segments(self):
        """将流量段设置应用到表格（与原版一致）"""
        flow_str = self.flow_segments_edit.text().strip()
        if not flow_str:
            InfoBar.warning("提示", "请输入流量值", parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return
        try:
            flow_str = flow_str.replace('，', ',')
            flow_values = [float(q.strip()) for q in flow_str.split(',') if q.strip()]
            if not flow_values:
                InfoBar.warning("提示", "请输入有效的流量值", parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return
            updated_count = 0
            self.input_table.blockSignals(True)
            for r in range(self.input_table.rowCount()):
                seg_item = self.input_table.item(r, 1)
                seg_val = seg_item.text().strip() if seg_item else ""
                if not seg_val:
                    continue
                try:
                    segment = int(seg_val)
                    if 1 <= segment <= len(flow_values):
                        self.input_table.setItem(r, 6, QTableWidgetItem(str(flow_values[segment - 1])))
                        updated_count += 1
                except (ValueError, IndexError):
                    continue
            self.input_table.blockSignals(False)
            InfoBar.success("完成", f"已将流量值应用到 {updated_count} 行数据\n流量段1~{len(flow_values)}: {flow_values}",
                           parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
        except ValueError as e:
            InfoBar.error("格式错误", f"流量值格式错误: {str(e)}\n请使用逗号分隔的数字",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _on_start_station_focus_in(self, event):
        """起始桩号输入框获得焦点时转换为纯数字便于编辑（与原版一致）"""
        current = self.start_station_edit.text().strip()
        value = parse_station_input(current)
        self.start_station_edit.setText(str(value))
        LineEdit.focusInEvent(self.start_station_edit, event)

    def _format_start_station_display(self):
        """编辑完成后格式化起始桩号显示"""
        current = self.start_station_edit.text().strip()
        value = parse_station_input(current)
        formatted = format_station_display(value)
        self.start_station_edit.setText(formatted)

    def _get_start_station_value(self) -> float:
        """获取起始桩号的数值（米）"""
        return parse_station_input(self.start_station_edit.text())

    def _sf(self, val, default=0.0):
        if not val: return default
        s = str(val).strip()
        if not s: return default
        try: return float(s)
        except ValueError: return default

    def _normalize_row(self, row, length):
        row = list(row) if row else []
        while len(row) < length:
            row.append("")
        return row[:length]

    def _get_user_prefs_path(self):
        """获取用户偏好配置文件路径（与原版一致）"""
        import json
        appdata = os.path.join(os.path.expanduser("~"), ".canal_calc")
        if not os.path.isdir(appdata):
            os.makedirs(appdata, exist_ok=True)
        return os.path.join(appdata, "batch_prefs.json")

    def _load_user_prefs(self):
        """从本地配置文件加载用户偏好（与原版一致）"""
        try:
            import json
            path = self._get_user_prefs_path()
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    prefs = json.load(f)
                self._has_opened_template = prefs.get("has_opened_template", False)
                saved_dir = prefs.get("last_import_dir")
                if saved_dir and os.path.isdir(saved_dir):
                    self._last_import_dir = saved_dir
        except Exception:
            pass

    def _save_user_prefs(self):
        """将用户偏好保存到本地配置文件（与原版一致）"""
        try:
            import json
            prefs = {
                "has_opened_template": self._has_opened_template,
                "last_import_dir": self._last_import_dir,
            }
            with open(self._get_user_prefs_path(), 'w', encoding='utf-8') as f:
                json.dump(prefs, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _sync_common_columns(self):
        """同步输入表和结果表前4列(序号/流量段/建筑物名称/结构形式)的宽度，取较大值"""
        for col in range(4):
            w1 = self.input_table.columnWidth(col)
            w2 = self.result_table.columnWidth(col)
            max_w = max(w1, w2)
            self.input_table.setColumnWidth(col, max_w)
            self.result_table.setColumnWidth(col, max_w)

    def _info_parent(self):
        w = self.window()
        return w if w else self

    def _regenerate_detail_report(self):
        """根据已有的 batch_results 重新生成详细计算过程报告"""
        if not self.batch_results:
            return
        detail_lines = []
        detail_lines.append("=" * 80)
        detail_lines.append("          多流量段批量水力计算 - 详细计算过程报告")
        detail_lines.append("=" * 80)
        detail_lines.append(f"计算时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        detail_lines.append(f"总计建筑物数量: {len(self.batch_results)}")
        detail_lines.append("-" * 80)
        detail_lines.append("")
        for i, item_data in enumerate(self.batch_results, 1):
            input_vals = item_data['input']
            result_data = item_data['result']
            section_type = str(input_vals[3]).strip()
            building_name = str(input_vals[2]).strip()
            detail_lines.append(f"【项目 {i}】 {building_name} ({section_type})")
            detail_lines.append(self._gen_detail_report(input_vals, result_data))
            detail_lines.append("\n" + "*" * 80 + "\n")
        self.detail_text.setPlainText("\n".join(detail_lines))
        self.result_notebook.setCurrentIndex(1)

    def _clear_results(self):
        self.result_table.setRowCount(0)
        self.detail_text.clear()
        self.batch_results = []
        self._last_calc_snapshot = None
        self._last_calc_detail = None
        # 同时清除共享数据管理器中的批量结果
        if SHARED_DATA_AVAILABLE:
            try:
                get_shared_data_manager().clear_batch_results()
            except Exception:
                pass

    # ================================================================
    # Excel 导入
    # ================================================================
    def _import_from_excel(self):
        """从Excel文件导入数据（自动兼容模板格式和导出格式，与原版一致）"""
        try:
            import openpyxl
        except ImportError:
            InfoBar.warning("缺少依赖", "需要安装 openpyxl: pip install openpyxl", parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return
        # 首次导入且未查看过模板时，提示用户是否先查看模板（与原版一致）
        if not self._has_opened_template and self._last_import_dir is None:
            ret = fluent_question(self, "导入Excel",
                "检测到您尚未查看过Excel模板。\n\n"
                "建议先打开模板了解格式要求，再填写数据后导入。\n\n"
                "点击「查看模板」先打开模板\n"
                "点击「直接导入」选择文件导入",
                yes_text="查看模板", no_text="直接导入")
            self._has_opened_template = True
            self._save_user_prefs()
            if ret:
                self._open_excel_template()
                return
        # 确定文件对话框的初始目录（与原版一致：上次导入目录 > 模板目录 > 桌面）
        initial_dir = ""
        if self._last_import_dir and os.path.isdir(self._last_import_dir):
            initial_dir = self._last_import_dir
        else:
            template_dir = os.path.dirname(self._get_template_path())
            if os.path.isdir(template_dir):
                initial_dir = template_dir
            else:
                initial_dir = os.path.expanduser("~/Desktop")
        filepath, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", initial_dir, "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)")
        if not filepath: return
        # 记录本次导入的目录（与原版一致）
        self._last_import_dir = os.path.dirname(filepath)
        self._save_user_prefs()
        try:
            if filepath.lower().endswith('.xls') and not filepath.lower().endswith('.xlsx'):
                import xlrd
                _xls_book = xlrd.open_workbook(filepath)
                _xls_sheet = _xls_book.sheet_by_index(0)
                class _XlsCell:
                    def __init__(self, v):
                        self.value = v if v != '' else None
                class _XlsWs:
                    def __init__(self, sh):
                        self._sh = sh
                        self.max_row = sh.nrows
                    def cell(self, row, column):
                        try:
                            return _XlsCell(self._sh.cell_value(row - 1, column - 1))
                        except IndexError:
                            return _XlsCell(None)
                ws = _XlsWs(_xls_sheet)
            else:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active
            info_parts = []

            # 自动判断Excel格式：
            # 格式A（模板）：第1行=基础信息（A1含"渠道名称"），第2行=表头，第3行起=数据
            # 格式B（导出）：第1行=标题，第2行=基础信息（A2含"渠道名称"），第3行=表头，第4行起=数据
            a1_val = str(ws.cell(row=1, column=1).value or "")
            if "渠道名称" in a1_val:
                info_row = 1
                data_start_row = 3
            else:
                info_row = 2
                data_start_row = 4

            # 读取基础信息行
            def _read_cell(r, c):
                v = ws.cell(row=r, column=c).value
                return str(v).strip() if v is not None else ""

            label_a = _read_cell(info_row, 1)
            val_b = _read_cell(info_row, 2)
            if "渠道名称" in label_a and val_b:
                self.channel_name_edit.setText(val_b)
                info_parts.append(f"渠道名称: {val_b}")

            label_c = _read_cell(info_row, 3)
            val_d = _read_cell(info_row, 4)
            if ("渠道级别" in label_c or "渠道类型" in label_c) and val_d:
                idx = self.channel_level_combo.findText(val_d)
                if idx >= 0: self.channel_level_combo.setCurrentIndex(idx)
                info_parts.append(f"渠道类型: {val_d}")

            label_e = _read_cell(info_row, 5)
            val_f = _read_cell(info_row, 6)
            if "水位" in label_e and val_f:
                self.start_wl_edit.setText(val_f)
                info_parts.append(f"起始水位: {val_f}")

            val_h = _read_cell(info_row, 8)
            if val_h:
                station_value = parse_station_input(val_h)
                formatted_station = format_station_display(station_value)
                self.start_station_edit.setText(formatted_station)
                info_parts.append(f"起始桩号: {formatted_station}")

            # 读取数据行（前20列）
            data_rows = []
            for row_idx in range(data_start_row, ws.max_row + 1):
                row_data = []
                for col_idx in range(1, 21):
                    cv = ws.cell(row=row_idx, column=col_idx).value
                    row_data.append(str(cv) if cv is not None else "")
                if any(v.strip() for v in row_data):
                    data_rows.append(row_data)
            if not data_rows:
                InfoBar.warning("提示", "Excel文件中没有数据", parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 自动检测列映射：检查表头行第5列是否含"X"来判断是否有X/Y坐标列
            # 含X/Y列映射（20列）：0序号,1流量段,2名称,3结构形式,4X,5Y,6Q,7n,8比降,9m,10B,11宽深比,12R,13D,14渡槽深宽比,15倒角角度,16倒角底边,17圆心角,18不淤,19不冲
            # 无X/Y列映射（18列）：0序号,1流量段,2名称,3结构形式,4Q,5n,6比降,7m,8B,9宽深比,10R,11D,12渡槽深宽比,13倒角角度,14倒角底边,15圆心角,16不淤,17不冲
            header_row = data_start_row - 1
            h4_val = str(ws.cell(row=header_row, column=5).value or "")
            has_xy_cols = "X" in h4_val.upper() or "Q" not in h4_val.upper()
            # 如果第5列表头是"X"或不含"Q"，说明有X/Y列偏移
            if has_xy_cols and len(data_rows) > 0 and len(data_rows[0]) >= 8:
                # 进一步确认：检查第7列表头是否含"Q"
                h7_val = str(ws.cell(row=header_row, column=7).value or "")
                has_xy_cols = "Q" in h7_val.upper() or "流量" in h7_val

            if self.input_table.rowCount() > 0 and not self._is_sample_data:
                if not fluent_question(self, "确认覆盖",
                        f"当前表格已有 {self.input_table.rowCount()} 行数据，\n"
                        f"导入将覆盖全部现有数据。\n\n确定继续吗？",
                        yes_text="覆盖导入", no_text="取消"):
                    return
            self._clear_input(force=True)
            for rd in data_rows:
                rd = rd + [""] * 20
                if has_xy_cols:
                    mapped = rd[:20]
                else:
                    mapped = [
                        rd[0], rd[1], rd[2], rd[3],
                        "", "",
                        rd[4], rd[5], rd[6],
                        rd[7], rd[8], rd[9],
                        rd[10], rd[11],
                        rd[12], rd[13], rd[14], rd[15],
                        rd[16], rd[17],
                    ]
                self._add_row(mapped)
            self._auto_detect_flow_segments()
            auto_resize_table(self.input_table)
            info_msg = f"已成功导入 {len(data_rows)} 行数据"
            if info_parts:
                info_msg += " | " + ", ".join(info_parts)
            self._is_sample_data = False
            InfoBar.success("导入成功", info_msg, parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            # 导入后检查建筑物重名（仅警告，不阻止）
            self._validate_duplicate_buildings_warn()
        except Exception as e:
            InfoBar.error("导入失败", str(e), parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    # ================================================================
    # 导出
    # ================================================================
    def _export_excel(self):
        if self.result_table.rowCount() == 0:
            InfoBar.warning("提示", "没有计算结果可导出", parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return
        # 验证4个必填字段是否已输入
        missing_fields = []
        channel_name = self.channel_name_edit.text().strip()
        if not channel_name:
            missing_fields.append("渠道名称")
        channel_level = self.channel_level_combo.currentText()
        if not channel_level:
            missing_fields.append("渠道类型")
        start_station = self.start_station_edit.text().strip()
        if not start_station:
            missing_fields.append("起始桩号")
        start_water_level = self.start_wl_edit.text().strip()
        if not start_water_level:
            missing_fields.append("渠道起始水位")
        if missing_fields:
            fields_str = "、".join(missing_fields)
            fluent_info(self, "警告", f"以下必填项尚未填写，请补充后再导出Excel：\n\n  {fields_str}")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            InfoBar.warning("缺少依赖", "需要安装 openpyxl: pip install openpyxl", parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return
        auto_name = f"{channel_name}{channel_level}_多流量段批量计算结果.xlsx" if channel_name else "批量计算结果.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(self, "保存Excel报告", auto_name, "Excel文件 (*.xlsx)")
        if not filepath: return
        try:
            # 从输入表获取参数数据，建立映射（以序号为key）
            input_params_map = {}
            for r in range(self.input_table.rowCount()):
                seq_item = self.input_table.item(r, 0)
                if not seq_item:
                    continue
                seq_key = seq_item.text().strip()
                x_val = self.input_table.item(r, 4).text() if self.input_table.item(r, 4) else ""
                y_val = self.input_table.item(r, 5).text() if self.input_table.item(r, 5) else ""
                q_val = self.input_table.item(r, 6).text() if self.input_table.item(r, 6) else ""
                n_val = self.input_table.item(r, 7).text() if self.input_table.item(r, 7) else ""
                slope_val = self.input_table.item(r, 8).text() if self.input_table.item(r, 8) else ""
                m_val = self.input_table.item(r, 9).text() if self.input_table.item(r, 9) else ""
                input_params_map[seq_key] = (x_val, y_val, q_val, n_val, slope_val, m_val)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "批量计算结果"
            # 导出表头：在结果表头的结构形式后插入X/Y/Q/n/比降/边坡系数列（与原版一致）
            export_headers = list(RESULT_HEADERS[:4]) + ["X", "Y", "Q(m³/s)", "糙率n", "比降(1/)", "边坡系数m"] + list(RESULT_HEADERS[4:])
            # 第1行：标题
            ws['A1'] = "渠系建筑物多流量段批量水力计算系统结果报告"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells(f'A1:{get_column_letter(len(export_headers))}1')
            ws['A1'].alignment = Alignment(horizontal='center')
            # 第2行：基础信息
            ws['A2'] = "渠道名称"; ws['A2'].font = Font(bold=True)
            ws['B2'] = channel_name
            ws['C2'] = "渠道类型"; ws['C2'].font = Font(bold=True)
            ws['D2'] = channel_level
            ws['E2'] = "起始水位(m)"; ws['E2'].font = Font(bold=True)
            ws['F2'] = start_water_level
            ws['G2'] = "起始桩号"; ws['G2'].font = Font(bold=True)
            ws['H2'] = start_station
            # 第3行：表头
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            for c, h in enumerate(export_headers, 1):
                cell = ws.cell(row=3, column=c, value=h)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            # 第4行起：数据（插入输入参数列）
            for r in range(self.result_table.rowCount()):
                # 原结果行数据
                result_vals = []
                for c in range(self.result_table.columnCount()):
                    item = self.result_table.item(r, c)
                    result_vals.append(item.text() if item else "")
                seq_key = result_vals[0].strip() if result_vals else ""
                x_val, y_val, q_val, n_val, slope_val, m_val = input_params_map.get(seq_key, ("", "", "", "", "", ""))
                # 构建导出行：前4列 + X/Y/Q/n/比降/边坡系数 + 后续结果列
                export_row = result_vals[:4] + [x_val, y_val, q_val, n_val, slope_val, m_val] + result_vals[4:]
                for c, val in enumerate(export_row, 1):
                    ws.cell(row=r+4, column=c, value=val)
            # 自动列宽
            for col_num in range(1, len(export_headers) + 1):
                max_len = len(str(export_headers[col_num-1]))
                for row_num in range(4, ws.max_row + 1):
                    cv = ws.cell(row=row_num, column=col_num).value
                    if cv: max_len = max(max_len, len(str(cv)))
                ws.column_dimensions[get_column_letter(col_num)].width = min(max_len + 3, 30)
            wb.save(filepath)
            InfoBar.success("导出成功", f"Excel已保存到: {filepath}", parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
        except PermissionError:
            InfoBar.error("文件被占用", "无法写入文件，请先关闭已打开的同名文件（如Excel等），然后重新操作。", parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("导出失败", str(e), parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _export_word(self):
        """导出详细计算过程为Word文档"""
        if not WORD_EXPORT_AVAILABLE:
            fluent_info(self, "缺少依赖",
                "Word导出需要安装 python-docx、latex2mathml、lxml。\n"
                "请执行: pip install python-docx latex2mathml lxml")
            return
        if not self.batch_results:
            InfoBar.warning("提示", "请先进行批量计算", parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return
        channel_name = self.channel_name_edit.text().strip()
        channel_level = self.channel_level_combo.currentText()
        auto_name = f"{channel_name}{channel_level}_多流量段批量水力计算书.docx" if channel_name else "多流量段批量水力计算书.docx"
        filepath, _ = QFileDialog.getSaveFileName(self, "保存Word报告", auto_name, "Word文档 (*.docx);;所有文件 (*.*)")
        if not filepath: return
        try:
            self._build_word_report(filepath)
            InfoBar.success("导出成功", f"Word报告已保存到: {filepath}", parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            ask_open_file(filepath, self._info_parent())
        except PermissionError:
            InfoBar.error("文件被占用", "无法写入文件，请关闭已打开的同名Word文档，然后重新操作。", parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("导出失败", f"Word导出失败: {str(e)}", parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _build_word_report(self, filepath):
        """构建Word报告文档（与明渠设计模块统一的高端咨询报告风格）"""
        channel_name = self.channel_name_edit.text().strip()
        channel_level = self.channel_level_combo.currentText()
        start_station = self.start_station_edit.text().strip()
        start_wl = self.start_wl_edit.text().strip()
        flow_str = self.flow_segments_edit.text().strip()

        # 计算末尾桩号：从X/Y坐标累计距离（MC）得到
        end_station = ""
        try:
            import math
            start_val = parse_station_input(start_station)
            coords = []
            for item_data in self.batch_results:
                iv = self._normalize_row(item_data['input'], len(INPUT_HEADERS))
                x_s = str(iv[4]).strip()
                y_s = str(iv[5]).strip()
                if x_s and y_s:
                    coords.append((float(x_s), float(y_s)))
            if len(coords) >= 2:
                mc = 0.0
                for k in range(1, len(coords)):
                    dx = coords[k][0] - coords[k - 1][0]
                    dy = coords[k][1] - coords[k - 1][1]
                    mc += math.sqrt(dx * dx + dy * dy)
                end_station = format_station_display(start_val + mc)
        except Exception:
            pass

        doc = create_styled_doc(
            title='多流量段批量水力计算书',
            subtitle=f'{channel_name}  ·  {channel_level}' if channel_name else '',
            header_text=f'{channel_name}{channel_level} 多流量段批量水力计算书',
            channel_name=channel_name,
            channel_level=channel_level,
            start_station=start_station,
            end_station=end_station,
        )
        doc.add_page_break()

        # ===== 一、计算概述 =====
        doc_add_h1(doc, '一、计算概述')
        overview_params = []
        if channel_name:
            overview_params.append(("渠道名称", channel_name))
        if channel_level:
            overview_params.append(("渠道级别", channel_level))
        if start_station:
            overview_params.append(("起始桩号", start_station))
        if start_wl:
            overview_params.append(("起始水位", f"{start_wl} m"))
        if flow_str:
            overview_params.append(("流量段设置", f"{flow_str} m³/s"))
        total = len(self.batch_results)
        calc_count = sum(1 for it in self.batch_results
                         if it['result'].get('success') and not it['result'].get('is_siphon')
                         and not it['result'].get('is_diversion_gate'))
        siphon_count = sum(1 for it in self.batch_results if it['result'].get('is_siphon'))
        gate_count = sum(1 for it in self.batch_results if it['result'].get('is_diversion_gate'))
        overview_params.append(("建筑物总数", str(total)))
        if calc_count:
            overview_params.append(("参与计算", f"{calc_count} 条"))
        if siphon_count:
            overview_params.append(("倒虹吸占位", f"{siphon_count} 条"))
        if gate_count:
            overview_params.append(("闸类占位", f"{gate_count} 条"))
        overview_params.append(("计算时间", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        doc_add_param_table(doc, overview_params)

        # ===== 二、结果汇总表 =====
        doc_add_h1(doc, '二、结果汇总表')
        doc_add_table_caption(doc, '表 1  多流量段批量水力计算结果汇总')
        sum_headers = ["序号", "流量段", "建筑物", "结构形式", "尺寸",
                       "h设计(m)", "V设计(m/s)", "V加大(m/s)", "总高H(m)", "状态"]
        sum_data = []
        for r in range(self.result_table.rowCount()):
            rv = []
            for c in range(self.result_table.columnCount()):
                item = self.result_table.item(r, c)
                rv.append(item.text() if item else "-")
            B_v, D_v, R_v = rv[4], rv[5], rv[6]
            dim = "-"
            if B_v and B_v != "-":
                dim = f"B={B_v}"
            elif D_v and D_v != "-":
                dim = f"D={D_v}"
            elif R_v and R_v != "-":
                dim = f"R={R_v}"
            status = rv[-1].replace("✓", "OK").replace("✗", "NG").replace("⏭", ">>")
            sum_data.append([rv[0], rv[1], rv[2], rv[3], dim,
                             rv[7], rv[8], rv[14], rv[16], status])
        doc_add_styled_table(doc, sum_headers, sum_data, with_full_border=True)

        # ===== 三、详细计算过程 =====
        doc.add_page_break()
        doc_add_h1(doc, '三、详细计算过程')

        for i, item_data in enumerate(self.batch_results, 1):
            input_vals = self._normalize_row(item_data['input'], len(INPUT_HEADERS))
            result_data = item_data['result']
            section_type = str(input_vals[3]).strip()
            building_name = str(input_vals[2]).strip()

            if i > 1:
                doc.add_paragraph('')

            doc_add_h2(doc, f'项目 {i} — {building_name}（{section_type}）')

            # 渠道信息
            segment = str(input_vals[1]).strip()
            info_parts = []
            if channel_name or channel_level:
                info_parts.append(f"渠道: {channel_name} {channel_level}")
            info_parts.append(f"流量段: 第{segment}段")
            coord_x = str(input_vals[4]).strip()
            coord_y = str(input_vals[5]).strip()
            if coord_x and coord_y:
                info_parts.append(f"坐标: X={coord_x}, Y={coord_y}")
            for part in info_parts:
                doc_add_body(doc, part)
            doc.add_paragraph('')

            # 计算过程正文
            body = self._gen_detail_body(input_vals, result_data)
            doc_render_calc_text(doc, body)

        doc.save(filepath)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        auto_resize_table(self.input_table)
        if hasattr(self, 'result_table'):
            auto_resize_table(self.result_table)


# ============================================================
# 断面参数输入弹窗（与原版SectionParameterDialog一致）
# ============================================================
class SectionParameterDialog(QDialog):
    """断面参数输入弹窗 - 根据断面类型显示不同的参数输入界面"""

    def __init__(self, parent, section_type: str, current_values: dict):
        super().__init__(parent)
        self.section_type = section_type
        self.current_values = current_values
        self.result = None
        self.setWindowTitle(f"参数设置 - {section_type}")
        self.setMinimumWidth(420)
        self.setStyleSheet(DIALOG_STYLE)
        self._entries = {}
        self._build_ui()
        self._load_current_values()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # 基础参数区
        basic_grp = QGroupBox("基础参数")
        basic_form = QFormLayout()
        basic_form.setSpacing(8)
        self._entries['Q'] = LineEdit(); self._entries['Q'].setPlaceholderText("设计流量")
        basic_form.addRow("设计流量 Q (m³/s):", self._entries['Q'])
        self._entries['n'] = LineEdit(); self._entries['n'].setPlaceholderText("糙率")
        basic_form.addRow("糙率 n:", self._entries['n'])
        self._entries['slope_inv'] = LineEdit(); self._entries['slope_inv'].setPlaceholderText("比降倒数")
        basic_form.addRow("水力坡降 1/:", self._entries['slope_inv'])
        basic_grp.setLayout(basic_form)
        layout.addWidget(basic_grp)

        # 流速参数区
        vel_grp = QGroupBox("流速参数")
        vel_form = QFormLayout()
        vel_form.setSpacing(8)
        self._entries['v_min'] = LineEdit(); self._entries['v_min'].setPlaceholderText("不淤流速")
        vel_form.addRow("不淤流速 (m/s):", self._entries['v_min'])
        self._entries['v_max'] = LineEdit(); self._entries['v_max'].setPlaceholderText("不冲流速")
        vel_form.addRow("不冲流速 (m/s):", self._entries['v_max'])
        hint = QLabel("(一般情况下保持默认数值即可)")
        hint.setStyleSheet("font-size:10px; color:#424242;")
        vel_form.addRow(hint)
        vel_grp.setLayout(vel_form)
        layout.addWidget(vel_grp)

        # 可选参数区（根据断面类型）
        opt_grp = QGroupBox("可选参数")
        opt_form = QFormLayout()
        opt_form.setSpacing(8)
        self._build_optional_inputs(opt_form)
        opt_grp.setLayout(opt_form)
        layout.addWidget(opt_grp)

        # 按钮区
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.button(QDialogButtonBox.Ok).setText("确认")
        btn_box.button(QDialogButtonBox.Cancel).setText("取消")
        btn_box.accepted.connect(self._on_confirm)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _add_opt_entry(self, form, label, key, placeholder="", hint_text=""):
        self._entries[key] = LineEdit()
        self._entries[key].setPlaceholderText(placeholder)
        form.addRow(label, self._entries[key])
        if hint_text:
            h = QLabel(hint_text)
            h.setStyleSheet("font-size:10px; color:#424242;")
            form.addRow(h)

    def _build_optional_inputs(self, form):
        st = self.section_type
        if "明渠-梯形" in st:
            self._add_opt_entry(form, "边坡系数 m:", "m", "必填")
            self._add_opt_entry(form, "指定底宽 B (m):", "b", "留空自动计算")
            self._add_opt_entry(form, "宽深比 B/h:", "b_h_ratio", "留空自动计算",
                                "(梯形断面边坡系数必填; 底宽B和宽深比可选)")
        elif "明渠-矩形" in st:
            hint_label = QLabel("边坡系数 m = 0 (矩形断面)")
            hint_label.setStyleSheet("color:#555;")
            form.addRow(hint_label)
            self._add_opt_entry(form, "指定底宽 B (m):", "b", "留空自动计算")
            self._add_opt_entry(form, "宽深比 B/h:", "b_h_ratio", "留空自动计算",
                                "(底宽B和宽深比可选，留空则自动计算)")
        elif "明渠-圆形" in st:
            self._add_opt_entry(form, "指定直径 D (m):", "D", "留空自动计算",
                                "(留空则自动计算)")
        elif "渡槽-U形" in st:
            self._add_opt_entry(form, "指定内半径 R (m):", "R", "留空自动计算",
                                "(留空则自动计算)")
        elif "渡槽-矩形" in st:
            self._add_opt_entry(form, "深宽比 H/B:", "h_b_ratio", "推荐0.6~0.8",
                                "(推荐值0.6~0.8，留空则默认0.8)")
            self._add_opt_entry(form, "倒角角度 (度):", "chamfer_angle", "")
            self._add_opt_entry(form, "倒角底边 (m):", "chamfer_length", "",
                                "(倒角两者需同时填写或同时留空)")
        elif "隧洞-圆形" in st:
            self._add_opt_entry(form, "指定直径 D (m):", "D", "留空自动计算",
                                "(留空则自动计算)")
        elif "隧洞-圆拱直墙型" in st:
            self._add_opt_entry(form, "拱顶圆心角 (度):", "theta", "留空则采用180°",
                                "(留空则采用180°)")
            self._add_opt_entry(form, "指定底宽 B (m):", "B", "留空自动计算",
                                "(指定底宽B留空则自动计算)")
        elif "隧洞-马蹄形" in st:
            self._add_opt_entry(form, "指定半径 R (m):", "r", "留空自动计算",
                                "(留空则自动计算)")
        elif st == "矩形暗涵":
            self._add_opt_entry(form, "指定宽深比:", "BH_ratio_rect", "留空自动计算")
            self._add_opt_entry(form, "指定底宽 B (m):", "B_rect", "留空自动计算",
                                "(二选一输入，留空则自动计算)")
        else:
            no_param = QLabel("(无额外参数)")
            no_param.setStyleSheet("color:#424242;")
            form.addRow(no_param)

    def _load_current_values(self):
        try:
            cv = self.current_values
            for key in ['Q', 'n', 'slope_inv', 'v_min', 'v_max']:
                val = cv.get(key, '')
                if val and str(val).strip():
                    self._entries[key].setText(str(val).strip())
            st = self.section_type
            if "明渠-梯形" in st or "明渠-矩形" in st:
                for k in ['m', 'b', 'b_h_ratio']:
                    v = cv.get(k, '')
                    if k in self._entries and v and str(v).strip():
                        self._entries[k].setText(str(v).strip())
            elif "明渠-圆形" in st or "隧洞-圆形" in st:
                v = cv.get('D', '')
                if 'D' in self._entries and v and str(v).strip():
                    self._entries['D'].setText(str(v).strip())
            elif "渡槽-U形" in st:
                v = cv.get('R', '')
                if 'R' in self._entries and v and str(v).strip():
                    self._entries['R'].setText(str(v).strip())
            elif "渡槽-矩形" in st:
                for k, ek in [('ducao_depth_ratio', 'h_b_ratio'), ('chamfer_angle', 'chamfer_angle'), ('chamfer_length', 'chamfer_length')]:
                    v = cv.get(k, '')
                    if ek in self._entries and v and str(v).strip():
                        self._entries[ek].setText(str(v).strip())
            elif "隧洞-圆拱直墙型" in st:
                for k, ek in [('b', 'B'), ('theta', 'theta')]:
                    v = cv.get(k, '')
                    if ek in self._entries and v and str(v).strip():
                        self._entries[ek].setText(str(v).strip())
            elif "隧洞-马蹄形" in st:
                v = cv.get('R', '')
                if 'r' in self._entries and v and str(v).strip():
                    self._entries['r'].setText(str(v).strip())
            elif st == "矩形暗涵":
                for k, ek in [('b_h_ratio', 'BH_ratio_rect'), ('b', 'B_rect')]:
                    v = cv.get(k, '')
                    if ek in self._entries and v and str(v).strip():
                        self._entries[ek].setText(str(v).strip())
        except Exception:
            pass

    def _get_float(self, key, default=None):
        if key not in self._entries:
            return default
        s = self._entries[key].text().strip()
        if not s:
            return default if default is not None else ""
        return float(s)

    def _on_confirm(self):
        try:
            result = {}
            Q = self._get_float('Q')
            if Q == "" or Q is None or Q <= 0:
                raise ValueError("设计流量必须大于0")
            result['Q'] = Q
            n = self._get_float('n')
            if n == "" or n is None or n <= 0:
                raise ValueError("糙率必须大于0")
            result['n'] = n
            slope_inv = self._get_float('slope_inv')
            if slope_inv == "" or slope_inv is None or slope_inv <= 0:
                raise ValueError("水力坡降倒数必须大于0")
            result['slope_inv'] = slope_inv
            v_min = self._get_float('v_min', 0.1)
            v_max = self._get_float('v_max', 100)
            if isinstance(v_min, (int, float)) and isinstance(v_max, (int, float)) and v_min >= v_max:
                raise ValueError("不淤流速必须小于不冲流速")
            result['v_min'] = v_min
            result['v_max'] = v_max

            st = self.section_type
            if "明渠-梯形" in st:
                m_val = self._get_float('m')
                if m_val == "" or m_val is None:
                    raise ValueError("梯形断面必须填写边坡系数m")
                if m_val < 0:
                    raise ValueError("边坡系数不能为负")
                result['m'] = m_val
                result['b'] = self._get_float('b', "")
                result['b_h_ratio'] = self._get_float('b_h_ratio', "")
            elif "明渠-矩形" in st:
                result['m'] = 0
                result['b'] = self._get_float('b', "")
                result['b_h_ratio'] = self._get_float('b_h_ratio', "")
            elif "明渠-圆形" in st:
                result['D'] = self._get_float('D', "")
            elif "渡槽-U形" in st:
                result['R'] = self._get_float('R', "")
            elif "渡槽-矩形" in st:
                result['h_b_ratio'] = self._get_float('h_b_ratio', "")
                result['chamfer_angle'] = self._get_float('chamfer_angle', "")
                result['chamfer_length'] = self._get_float('chamfer_length', "")
            elif "隧洞-圆形" in st:
                result['D'] = self._get_float('D', "")
            elif "隧洞-圆拱直墙型" in st:
                theta = self._get_float('theta', "")
                if isinstance(theta, (int, float)) and (theta < 90 or theta > 180):
                    raise ValueError("圆心角必须在90~180度之间")
                result['theta'] = theta
                result['B'] = self._get_float('B', "")
            elif "隧洞-马蹄形" in st:
                result['r'] = self._get_float('r', "")
            elif st == "矩形暗涵":
                result['BH_ratio_rect'] = self._get_float('BH_ratio_rect', "")
                result['B_rect'] = self._get_float('B_rect', "")

            self.result = result
            self.accept()
        except ValueError as e:
            fluent_info(self, "输入错误", str(e))
        except Exception as e:
            fluent_error(self, "错误", f"参数处理错误: {str(e)}")

    def get_result(self):
        return self.result
