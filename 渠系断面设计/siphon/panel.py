# -*- coding: utf-8 -*-
"""
倒虹吸水力计算面板 —— QWidget 版本（完整复刻）

功能：
1. 管道可视化画布（纵断面/平面视图，缩放/平移）
2. 全局参数输入（流量、拟定流速、糙率、平面转弯半径倍数、水损阈值、渐变段型式/系数、v₁v₂v₃策略）
3. 结构段三区表格（通用构件/平面段/纵断面段，三色分区、双击编辑）
4. DXF导入纵断面
5. HydraulicCore 水力计算 + 计算回填
6. 结果展示（汇总 + 详细过程）
7. 导出Excel/TXT
8. 外部接口：set_params / get_result / to_dict / from_dict

依赖：倒虹吸水力计算系统/ 目录下的计算引擎
"""

import sys
import os
import math
import copy
import datetime
import traceback

_pkg_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# 倒虹吸计算引擎路径
_siphon_dir = os.path.join(_pkg_root, '倒虹吸水力计算系统')
if _siphon_dir not in sys.path:
    sys.path.insert(0, _siphon_dir)

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QGroupBox,
    QFrame, QTabWidget, QTextEdit, QFileDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QAbstractItemView, QGridLayout, QScrollArea, QSizePolicy,
    QDialog, QDialogButtonBox, QPushButton, QLineEdit as _QLineEdit
)
from PySide6.QtCore import Qt, Signal, QTimer
from PySide6.QtGui import QFont, QColor, QBrush, QIntValidator
from PySide6.QtWebEngineWidgets import QWebEngineView

from qfluentwidgets import (
    PushButton, PrimaryPushButton, LineEdit, ComboBox, CheckBox,
    InfoBar, InfoBarPosition
)

from 渠系断面设计.styles import P, S, W, E, BG, CARD, BD, T1, T2, auto_resize_table, fluent_info, fluent_error, fluent_question
from 渠系断面设计.export_utils import (
    WORD_EXPORT_AVAILABLE, ask_open_file,
    create_styled_doc, doc_add_h1, doc_add_h2,
    doc_add_formula, doc_add_body, doc_render_calc_text,
    doc_add_result_table, doc_add_param_table,
    create_engineering_report_doc, doc_add_eng_h, doc_add_eng_body,
    doc_render_calc_text_eng, update_doc_toc_via_com,
)
from 渠系断面设计.report_meta import (
    ExportConfirmDialog, build_calc_purpose, REFERENCES_BASE, load_meta
)

# 计算引擎导入
try:
    from siphon_models import (
        GlobalParameters, StructureSegment, CalculationResult,
        SegmentType, SegmentDirection, GradientType, V2Strategy,
        LongitudinalNode, TurnType, InletOutletShape,
        COMMON_SEGMENT_TYPES, is_common_type,
        INLET_SHAPE_COEFFICIENTS, PlanFeaturePoint, TrashRackParams
    )
    from siphon_hydraulics import HydraulicCore
    from siphon_coefficients import CoefficientService
    SIPHON_AVAILABLE = True
except ImportError as _e:
    print(f"[倒虹吸] 计算引擎加载失败: {_e}")
    SIPHON_AVAILABLE = False

# DXF解析器
try:
    from dxf_parser import DxfParser
    DXF_AVAILABLE = True
except ImportError:
    DXF_AVAILABLE = False

# 可视化画布
try:
    from 渠系断面设计.siphon.canvas_view import PipelineCanvas
    CANVAS_AVAILABLE = True
except ImportError:
    CANVAS_AVAILABLE = False

# 对话框
try:
    from 渠系断面设计.siphon.dialogs import (
        InletShapeDialog, OutletShapeDialog, TrashRackConfigDialog,
        SegmentEditDialog, InletSectionDialog,
        CommonSegmentAddDialog, CommonSegmentEditDialog, SimpleCommonEditDialog
    )
    DIALOGS_AVAILABLE = True
except ImportError:
    DIALOGS_AVAILABLE = False

# 渐变段型式
GRADIENT_TYPE_OPTIONS = ["无", "反弯扭曲面", "1/4圆弧", "方头型", "直线扭曲面"]
GRADIENT_TYPE_MAP = {}
if SIPHON_AVAILABLE:
    GRADIENT_TYPE_MAP = {
        "无": GradientType.NONE,
        "反弯扭曲面": GradientType.REVERSE_BEND,
        "1/4圆弧": GradientType.QUARTER_ARC,
        "方头型": GradientType.SQUARE_HEAD,
        "直线扭曲面": GradientType.LINEAR_TWIST,
    }
    GRADIENT_TYPE_MAP_REV = {v: k for k, v in GRADIENT_TYPE_MAP.items()}

# v₂策略
V2_STRATEGY_OPTIONS = ["自动（=管道流速）", "v₁ + 0.2", "断面参数计算", "指定输入"]
V2_STRATEGY_MAP = {}
if SIPHON_AVAILABLE:
    V2_STRATEGY_MAP = {
        "自动（=管道流速）": V2Strategy.AUTO_PIPE,
        "v₁ + 0.2": V2Strategy.V1_PLUS_02,
        "断面参数计算": V2Strategy.SECTION_CALC,
        "指定输入": V2Strategy.MANUAL,
    }

# ================================================================
# 表L.1.2 渐变段局部损失系数参考对话框（倒虹吸专用）
# ================================================================
_L12_HEADERS = ["渐变段形式", "ξ₁", "ξ₂", "适用条件"]
_L12_ROWS = [
    ["反弯扭曲面", "0.10", "0.20", "θ₁,θ₂均≤12.5°"],
    ["1/4圆弧",   "0.15", "0.25", "θ₁,θ₂均≤12.5°"],
    ["方头型",    "0.30", "0.75", "θ₁,θ₂均≤12.5°"],
    ["直线扭曲面", "0.05~0.30", "0.30~0.50", "θ₁=15°~37°，θ₂=10°~17°"],
]


class L12CoeffRefDialog(QDialog):
    """表L.1.2 倒虹吸渐变段局部损失系数参考表"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("渐变段局部损失系数参考表")
        self.setMinimumSize(680, 260)
        self.resize(720, 280)
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(8)

        title = QLabel("表L.1.2  渐变段局部损失系数表（倒虹吸渐变段）")
        title.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        lay.addWidget(title)

        tbl = QTableWidget(len(_L12_ROWS), len(_L12_HEADERS))
        tbl.setHorizontalHeaderLabels(_L12_HEADERS)
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl.setSelectionMode(QAbstractItemView.SingleSelection)
        tbl.horizontalHeader().setStretchLastSection(False)
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        for r, row_data in enumerate(_L12_ROWS):
            for c, val in enumerate(row_data):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                tbl.setItem(r, c, item)
        tbl.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        row_h = sum(tbl.rowHeight(r) for r in range(len(_L12_ROWS)))
        tbl.setFixedHeight(row_h + tbl.horizontalHeader().height() + 4)
        lay.addWidget(tbl)

        note = QLabel("注：θ₁ 为水面收敛角，θ₂ 为水面扩散角（灌排规范附录表L.1.2）")
        note.setStyleSheet("color:#555;font-size:12px;")
        lay.addWidget(note)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok)
        btn_box.button(QDialogButtonBox.Ok).setText("关闭")
        btn_box.accepted.connect(self.accept)
        lay.addWidget(btn_box)


# 结构段表头
SEG_HEADERS = ["序号", "分类", "类型", "方向", "长度(m)", "半径R(m)", "角度θ(°)",
               "起点高程", "终点高程", "空间长度", "局部系数", "锁定"]

# 纵断面节点表头
LONG_NODE_HEADERS = ["桩号(m)", "高程(m)", "竖曲线半径(m)", "转弯类型", "转角(°)"]


class _NumPipesWidget(QWidget):
    """[-] 数字 [+] 管道根数自定义控件"""
    valueChanged = Signal(int)
    editingFinished = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._val = 1
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)
        self.btn_minus = QPushButton("−")
        self.btn_minus.setFixedSize(30, 30)
        self.btn_minus.setCursor(Qt.PointingHandCursor)
        self.edit_val = _QLineEdit("1")
        self.edit_val.setFixedSize(36, 30)
        self.edit_val.setAlignment(Qt.AlignCenter)
        self.edit_val.setValidator(QIntValidator(1, 10))
        self.edit_val.editingFinished.connect(self._on_editing_finished)
        self.edit_val.textChanged.connect(self._on_text_changed)
        self.btn_plus = QPushButton("+")
        self.btn_plus.setFixedSize(30, 30)
        self.btn_plus.setCursor(Qt.PointingHandCursor)
        lay.addWidget(self.btn_minus)
        lay.addWidget(self.edit_val)
        lay.addWidget(self.btn_plus)
        self.btn_minus.clicked.connect(lambda: self._change(-1))
        self.btn_plus.clicked.connect(lambda: self._change(1))
        self.setFixedWidth(96)

    def value(self):
        return self._val

    def setValue(self, v):
        v = max(1, min(10, int(v)))
        if v != self._val:
            self._val = v
            self.edit_val.setText(str(v))
            self.valueChanged.emit(v)
        else:
            self.edit_val.setText(str(v))

    def setFocus(self):
        self.edit_val.setFocus()

    def selectAll(self):
        self.edit_val.selectAll()

    def _change(self, delta):
        self.setValue(self._val + delta)

    def _on_text_changed(self, text):
        try:
            v = int(text)
            v = max(1, min(10, v))
            if v != self._val:
                self._val = v
                self.valueChanged.emit(v)
        except (ValueError, TypeError):
            pass

    def _on_editing_finished(self):
        text = self.edit_val.text().strip()
        try:
            v = max(1, min(10, int(text)))
        except (ValueError, TypeError):
            v = self._val
        if v != self._val:
            self._val = v
        self.edit_val.setText(str(self._val))
        self.editingFinished.emit()


class SiphonPanel(QWidget):
    """倒虹吸水力计算面板（完整复刻版）"""

    def __init__(self, parent=None):
        super().__init__(parent)
        # 核心数据
        self.segments = []           # 所有结构段（通用+纵断面+平面）
        self.plan_segments = []      # 平面段
        self.plan_feature_points = []  # 平面IP特征点
        self.plan_total_length = 0.0
        self.longitudinal_nodes = [] # 纵断面变坡点
        self._longitudinal_is_example = True
        self._syncing = False
        self.calculation_result = None
        self._detail_text_cache = ""
        self._suppress_result_display = False
        self.on_result_callback = None
        self.show_detailed_process = True

        # 拟定流速确认标志（方案D：用户是否已手动输入或确认过流速）
        self._v_user_confirmed = False
        # 平面转弯半径倍数确认标志（方案B：温和提醒）
        self._turn_n_user_confirmed = False
        # 管道根数确认标志（用户需Enter/失焦/按钮确认）
        self._num_pipes_user_confirmed = False

        # 断面参数缓存（v₂策略=断面参数计算用）
        self._section_B = None
        self._section_h = None
        self._section_m = None

        # 倒虹吸平面转弯半径倍数 n（R = n × D）
        self._siphon_turn_radius_n = 0.0

        # 下游断面参数（出口系数计算用）
        self._downstream_params = {}
        self.calculation_result_increased = None
        self._inc_pct_used = 0.0

        self._init_ui()
        self._init_default_segments()

    # ================================================================
    # UI 构建
    # ================================================================
    def _init_ui(self):
        main_lay = QVBoxLayout(self)
        main_lay.setContentsMargins(6, 4, 6, 4)
        main_lay.setSpacing(3)

        # A: 可视化画布 + 工具栏
        self._build_canvas_area(main_lay)

        # B: 参数区（Notebook，含计算结果Tab）
        self._build_params_area(main_lay)

        # C: 操作栏
        self._build_operation_bar(main_lay)

    # ---- A: 可视化画布 ----
    def _build_canvas_area(self, parent_lay):
        canvas_frame = QFrame()
        canvas_frame.setFrameShape(QFrame.StyledPanel)
        canvas_frame.setStyleSheet(f"background:#14141E;border:1px solid {BD};border-radius:4px;")
        cl = QVBoxLayout(canvas_frame)
        cl.setContentsMargins(2, 2, 2, 2)
        cl.setSpacing(2)

        # 工具栏
        tb = QHBoxLayout()
        tb.setSpacing(4)

        self.btn_view_profile = PushButton("纵断面")
        self.btn_view_profile.setFixedSize(70, 26)
        self.btn_view_profile.clicked.connect(lambda: self._switch_view("profile"))
        tb.addWidget(self.btn_view_profile)

        self.btn_view_plan = PushButton("平面图")
        self.btn_view_plan.setFixedSize(70, 26)
        self.btn_view_plan.clicked.connect(lambda: self._switch_view("plan"))
        tb.addWidget(self.btn_view_plan)

        tb.addStretch()

        self.lbl_zoom = QLabel("100%")
        self.lbl_zoom.setStyleSheet("color:#555555;font-size:11px;background:transparent;border:none;")
        tb.addWidget(self.lbl_zoom)

        for text, slot in [("＋", "_zoom_in"), ("－", "_zoom_out"),
                           ("重置", "_zoom_reset")]:
            btn = PushButton(text)
            btn.setFixedHeight(26)
            btn.clicked.connect(getattr(self, slot))
            tb.addWidget(btn)

        cl.addLayout(tb)

        # 画布
        if CANVAS_AVAILABLE:
            self.canvas = PipelineCanvas()
            self.canvas.setMinimumHeight(140)
            self.canvas.setMaximumHeight(200)
            self.canvas.zoom_changed.connect(self._update_zoom_label)
            cl.addWidget(self.canvas)
        else:
            self.canvas = None
            lbl = QLabel("管道可视化组件未加载")
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("color:#424242;font-size:12px;background:transparent;border:none;")
            lbl.setMinimumHeight(100)
            cl.addWidget(lbl)

        # 状态提示
        self.lbl_data_status = QLabel("")
        self.lbl_data_status.setStyleSheet("color:#555555;font-size:12px;background:transparent;border:none;")
        cl.addWidget(self.lbl_data_status)

        parent_lay.addWidget(canvas_frame)

    def _switch_view(self, mode):
        if self.canvas:
            self.canvas.set_view_mode(mode)
            self._update_zoom_label()

    def _zoom_in(self):
        if self.canvas: self.canvas.zoom_in(); self._update_zoom_label()

    def _zoom_out(self):
        if self.canvas: self.canvas.zoom_out(); self._update_zoom_label()

    def _zoom_reset(self):
        if self.canvas: self.canvas.zoom_reset(); self._update_zoom_label()

    def _update_zoom_label(self, _zoom=None):
        if self.canvas:
            self.lbl_zoom.setText(f"{int(self.canvas._zoom * 100)}%")

    # ---- B: 参数区 ----
    def _build_params_area(self, parent_lay):
        self.params_notebook = QTabWidget()

        # Tab1: 基本参数
        t1 = QWidget()
        self._build_basic_params_tab(t1)
        self.params_notebook.addTab(t1, "基本参数")

        # Tab2: 结构段信息
        t2 = QWidget()
        self._build_segments_tab(t2)
        self.params_notebook.addTab(t2, "结构段信息")

        # Tab3: 纵断面节点
        t3 = QWidget()
        self._build_long_nodes_tab(t3)
        self.params_notebook.addTab(t3, "纵断面节点")

        # Tab4: 计算结果
        t4 = QWidget()
        self._build_result_tab(t4)
        self.params_notebook.addTab(t4, "计算结果")

        parent_lay.addWidget(self.params_notebook, 1)  # stretch=1，占满剩余空间

    def _build_basic_params_tab(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(4, 2, 4, 2)
        lay.setSpacing(6)

        # ========== Card 1: 全局水力参数 — 三栏布局（新方案，标签上方） ==========
        # 左栏(管道基础参数) | 中栏(管道运行参数) | 右栏(转弯控制参数)
        card1 = QGroupBox("全局水力参数")
        _c1_lay = QHBoxLayout(card1)
        _c1_lay.setContentsMargins(10, 8, 10, 8)
        _c1_lay.setSpacing(0)

        def _vsep():
            _s = QFrame(); _s.setFrameShape(QFrame.Shape.VLine); _s.setFrameShadow(QFrame.Shadow.Sunken)
            _s.setStyleSheet("color:#D0D5E0;"); return _s

        def _hrow(*widgets):
            _w = QWidget(); _l = QHBoxLayout(_w)
            _l.setContentsMargins(0, 0, 0, 0); _l.setSpacing(6)
            for ww in widgets: _l.addWidget(ww)
            _l.addStretch(); return _w

        def _star(color):
            _s = QLabel(f"<span style='color:{color};font-weight:bold;font-size:14px;'>*</span>")
            _s.setTextFormat(Qt.RichText); return _s

        def _lbl_with_star(text, color):
            _w = QWidget(); _l = QHBoxLayout(_w)
            _l.setContentsMargins(0, 0, 0, 0); _l.setSpacing(1)
            _l.addWidget(QLabel(text)); _l.addWidget(_star(color)); _l.addStretch(); return _w

        # ── 左栏：管道基础参数 ──
        col_l = QWidget()
        _ll = QVBoxLayout(col_l)
        _ll.setContentsMargins(0, 0, 12, 0)
        _ll.setSpacing(3)

        _ll.addWidget(QLabel("设计流量 Q (m³/s):"))
        self.edit_Q = LineEdit(); self.edit_Q.setText("10.0"); self.edit_Q.setFixedWidth(80)
        self.edit_Q.textChanged.connect(self._on_Qv_changed)
        self.lbl_Q_hint = QLabel("")
        self.lbl_Q_hint.setStyleSheet("color:#0066CC;font-size:12px;")
        _ll.addWidget(_hrow(self.edit_Q, self.lbl_Q_hint))
        _ll.addSpacing(4)

        _ll.addWidget(QLabel("糙率 n:"))
        self.edit_n = LineEdit(); self.edit_n.setText("0.014"); self.edit_n.setFixedWidth(60)
        self.lbl_n_hint = QLabel("")
        self.lbl_n_hint.setStyleSheet("color:#0066CC;font-size:12px;")
        _ll.addWidget(_hrow(self.edit_n, self.lbl_n_hint))
        _ll.addSpacing(4)

        _ll.addWidget(QLabel("管径 D:"))
        self.lbl_D_theory = QLabel("D = --")
        self.lbl_D_theory.setStyleSheet(f"color:{P};font-size:12px;font-weight:bold;")
        _ll.addWidget(self.lbl_D_theory)
        _ll.addSpacing(4)

        _d_container = QWidget()
        _d_vlay = QVBoxLayout(_d_container)
        _d_vlay.setContentsMargins(0, 0, 0, 0); _d_vlay.setSpacing(3)
        _d_cb_row = QWidget(); _d_cb_lay = QHBoxLayout(_d_cb_row)
        _d_cb_lay.setContentsMargins(0, 0, 0, 0); _d_cb_lay.setSpacing(6)
        self.cb_D_override = CheckBox("指定管径")
        self.cb_D_override.setChecked(False)
        self.cb_D_override.stateChanged.connect(self._on_D_override_toggled)
        _d_cb_lay.addWidget(self.cb_D_override)
        _btn_D_help = QPushButton("?")
        _btn_D_help.setFixedSize(18, 18)
        _btn_D_help.setStyleSheet(
            "QPushButton { border-radius:9px; border:1.5px solid #aaa; color:#aaa;"
            " background:transparent; font-size:10px; font-weight:bold; }"
            "QPushButton:hover { border-color:#1565C0; color:#1565C0; }"
        )
        _btn_D_help.setCursor(Qt.PointingHandCursor)
        _btn_D_help.setToolTip("点击查看说明")
        _d_cb_lay.addWidget(_btn_D_help)
        self.edit_D_override = LineEdit()
        self.edit_D_override.setPlaceholderText("输入管径(m)")
        self.edit_D_override.setFixedWidth(120)
        self.edit_D_override.setVisible(False)
        _d_cb_lay.addWidget(self.edit_D_override)
        _d_cb_lay.addStretch()
        _d_vlay.addWidget(_d_cb_row)
        self.lbl_D_help = QLabel(
            "当工程实际管径已由工程师自行确定时，勾选此项并输入实际管径，覆盖自动计算结果。\n"
            "注：覆盖后平面转弯半径 R=nD 将使用指定值重新计算。"
        )
        self.lbl_D_help.setWordWrap(True)
        self.lbl_D_help.setStyleSheet(
            "QLabel { background:#FFF8E1; border:1px solid #FFB74D; border-radius:4px;"
            " padding:5px 8px; color:#E65100; font-size:12px; }"
        )
        self.lbl_D_help.setVisible(False)
        _d_vlay.addWidget(self.lbl_D_help)
        _btn_D_help.clicked.connect(
            lambda: self.lbl_D_help.setVisible(not self.lbl_D_help.isVisible())
        )
        _ll.addWidget(_d_container)
        _ll.addStretch()

        _c1_lay.addWidget(col_l, 3)
        _c1_lay.addWidget(_vsep())

        # ── 中栏：管道运行参数 ──
        col_m = QWidget()
        _ml = QVBoxLayout(col_m)
        _ml.setContentsMargins(12, 0, 12, 0)
        _ml.setSpacing(3)

        _ml.addWidget(_lbl_with_star("拟定流速 v (m/s):", "#E53935"))
        self.edit_v = LineEdit(); self.edit_v.setText("2.0"); self.edit_v.setFixedWidth(70)
        self.edit_v.setStyleSheet("LineEdit { border: 2px dashed #E65100; background: #FFF8E1; }")
        self.edit_v.textChanged.connect(self._on_Qv_changed)
        self.edit_v.textChanged.connect(self._on_v_edited_by_user)
        self.edit_v.editingFinished.connect(self._on_v_confirmed)
        self.lbl_v_hint = QLabel("← 请输入拟定流速")
        self.lbl_v_hint.setStyleSheet("color:#E53935;font-size:12px;font-weight:bold;")
        _ml.addWidget(_hrow(self.edit_v, self.lbl_v_hint))
        _ml.addSpacing(4)

        _ml.addWidget(_lbl_with_star("管道根数 N (根):", "#E53935"))
        self.spin_num_pipes = _NumPipesWidget()
        self._apply_np_unconfirmed_style()
        self.spin_num_pipes.valueChanged.connect(self._on_num_pipes_value_changed)
        self.spin_num_pipes.btn_minus.clicked.connect(self._on_num_pipes_confirmed)
        self.spin_num_pipes.btn_plus.clicked.connect(self._on_num_pipes_confirmed)
        self.spin_num_pipes.editingFinished.connect(self._on_num_pipes_confirmed)
        self.lbl_num_pipes_hint = QLabel("← 请输入管道根数")
        self.lbl_num_pipes_hint.setStyleSheet("color:#E53935;font-size:12px;font-weight:bold;")
        _ml.addWidget(_hrow(self.spin_num_pipes, self.lbl_num_pipes_hint))
        _ml.addSpacing(4)

        _ml.addWidget(QLabel("计算目标:"))
        self.lbl_calc_target = QLabel("计算总水头损失")
        self.lbl_calc_target.setStyleSheet("color:#00796B;font-weight:bold;")
        _ml.addWidget(self.lbl_calc_target)
        _ml.addSpacing(4)

        self.inc_cb = CheckBox("考虑加大流量比例系数")
        self.inc_cb.setChecked(True)
        self.inc_cb.stateChanged.connect(self._on_inc_toggle)
        _ml.addWidget(self.inc_cb)
        _inc_r = QWidget(); _inc_r_lay = QHBoxLayout(_inc_r)
        _inc_r_lay.setContentsMargins(0, 0, 0, 0); _inc_r_lay.setSpacing(6)
        _inc_r_lay.addWidget(QLabel("加大比例(%):"))
        self.edit_inc = LineEdit()
        self.edit_inc.setPlaceholderText("留空自动计算")
        self.edit_inc.setFixedWidth(90)
        _inc_r_lay.addWidget(self.edit_inc)
        self.lbl_inc_hint = QLabel("(留空则按设计流量自动查表)")
        self.lbl_inc_hint.setStyleSheet("color:#0066CC;font-size:12px;")
        _inc_r_lay.addWidget(self.lbl_inc_hint)
        _inc_r_lay.addStretch()
        _ml.addWidget(_inc_r)
        _ml.addStretch()

        _c1_lay.addWidget(col_m, 4)
        _c1_lay.addWidget(_vsep())

        # ── 右栏：转弯控制参数 ──
        col_r = QWidget()
        _rl = QVBoxLayout(col_r)
        _rl.setContentsMargins(12, 0, 0, 0)
        _rl.setSpacing(3)

        _rl.addWidget(_lbl_with_star("平面转弯半径倍数 (R=nD):", "#1565C0"))
        self.edit_turn_n = LineEdit(); self.edit_turn_n.setText("3.0"); self.edit_turn_n.setFixedWidth(50)
        self.edit_turn_n.setStyleSheet("LineEdit { border: 1.5px dashed #1565C0; background: #E3F2FD; }")
        self.edit_turn_n.textChanged.connect(self._on_turn_n_edited_by_user)
        self.edit_turn_n.textChanged.connect(self._on_turn_n_changed)
        _lbl_tn_hint = QLabel("(请确认倍数)")
        _lbl_tn_hint.setStyleSheet("color:#FF6600;font-size:12px;")
        _rl.addWidget(_hrow(self.edit_turn_n, _lbl_tn_hint))
        _rl.addSpacing(4)

        _rl.addWidget(QLabel("平面转弯半径 R (m):"))
        self.edit_turn_R = LineEdit()
        self.edit_turn_R.setPlaceholderText("自动计算")
        self.edit_turn_R.setFixedWidth(70)
        self.edit_turn_R.setStyleSheet("LineEdit { border: 1px solid #90CAF9; background: #E3F2FD; }")
        self.edit_turn_R.textChanged.connect(self._on_turn_R_changed)
        self.lbl_turn_R_status = QLabel("← 可直接输入覆盖（修改R将反推n）")
        self.lbl_turn_R_status.setStyleSheet("color:#888;font-size:12px;")
        _rl.addWidget(_hrow(self.edit_turn_R, self.lbl_turn_R_status))

        self.lbl_turn_R = QLabel("R = --  (请确认倍数)")
        self.lbl_turn_R.setStyleSheet("color:#1565C0;font-size:12px;")
        self.lbl_turn_R.setWordWrap(True)
        _rl.addWidget(self.lbl_turn_R)
        _rl.addSpacing(4)

        _rl.addWidget(QLabel("水损阈值 (m):"))
        self.edit_threshold = LineEdit()
        self.edit_threshold.setPlaceholderText("如: 2.0")
        self.edit_threshold.setFixedWidth(75)
        _lbl_threshold_hint = QLabel("(ΔZ超此值将提醒调整参数)")
        _lbl_threshold_hint.setStyleSheet("color:#FF6600;font-size:12px;")
        _rl.addWidget(_hrow(self.edit_threshold, _lbl_threshold_hint))
        _rl.addStretch()

        _c1_lay.addWidget(col_r, 3)

        lay.addWidget(card1)

        # ========== Card 2: 渐变段与流速参数 (QGroupBox) ==========
        card2 = QGroupBox("渐变段与流速参数")
        b2 = QVBoxLayout(card2)
        b2.setContentsMargins(10, 6, 10, 6)
        b2.setSpacing(4)

        # ---- 进口 浅色底色框 ----
        inlet_box = QFrame()
        inlet_box.setObjectName("siphonInletBox")
        inlet_box.setStyleSheet(
            "QFrame#siphonInletBox { background:#F9FAFB;"
            " border:1px solid #ECEEF2; border-radius:4px; }"
        )
        ibl = QVBoxLayout(inlet_box)
        ibl.setContentsMargins(10, 6, 10, 6)
        ibl.setSpacing(2)

        lbl_inlet_sec = QLabel("  ▸ 进口")
        lbl_inlet_sec.setStyleSheet(f"font-weight:bold;color:{T1};font-size:11px;")
        ibl.addWidget(lbl_inlet_sec)

        # 进口 Row 1: 渐变段型式 + ξ₁ + 始端流速v₁
        inlet_r1 = QHBoxLayout(); inlet_r1.setSpacing(6)
        inlet_r1.addWidget(QLabel("渐变段型式:"))
        self.combo_inlet_type = ComboBox()
        self.combo_inlet_type.addItems(GRADIENT_TYPE_OPTIONS)
        self.combo_inlet_type.setCurrentText("无")
        self.combo_inlet_type.setFixedWidth(140)
        self.combo_inlet_type.currentTextChanged.connect(self._on_inlet_type_changed)
        inlet_r1.addWidget(self.combo_inlet_type)
        self.lbl_inlet_type_hint = QLabel("")
        self.lbl_inlet_type_hint.setStyleSheet("color:#0066CC;font-size:12px;")
        inlet_r1.addWidget(self.lbl_inlet_type_hint)
        btn_inlet_coeff_ref = PushButton("参考系数表")
        btn_inlet_coeff_ref.setMinimumWidth(110)
        btn_inlet_coeff_ref.setToolTip("查看表L.1.2 倒虹吸渐变段局部损失系数")
        btn_inlet_coeff_ref.clicked.connect(lambda: L12CoeffRefDialog(self).exec())
        inlet_r1.addWidget(btn_inlet_coeff_ref)
        inlet_r1.addWidget(QLabel("ξ₁:"))
        self.edit_xi_inlet = LineEdit(); self.edit_xi_inlet.setText("0.0"); self.edit_xi_inlet.setFixedWidth(80)
        inlet_r1.addWidget(self.edit_xi_inlet)
        self.lbl_xi_inlet_hint = QLabel("")
        self.lbl_xi_inlet_hint.setStyleSheet("color:#0066CC;font-size:12px;")
        inlet_r1.addWidget(self.lbl_xi_inlet_hint)
        inlet_r1.addWidget(QLabel("始端流速v₁(m/s):"))
        self.edit_v1 = LineEdit(); self.edit_v1.setPlaceholderText("留空=0"); self.edit_v1.setFixedWidth(80)
        self.edit_v1.textChanged.connect(self._on_v_channel_in_changed)
        inlet_r1.addWidget(self.edit_v1)
        self.lbl_v1_hint = QLabel("(上游渠道流速)")
        self.lbl_v1_hint.setStyleSheet("color:#FF6600;font-size:12px;")
        inlet_r1.addWidget(self.lbl_v1_hint)
        inlet_r1.addStretch()
        ibl.addLayout(inlet_r1)

        # 进口 Row 1.5: 直线扭曲面扭转角（仅选中直线扭曲面时显示）
        self._twist_in_row = QWidget()
        _ti_lay = QHBoxLayout(self._twist_in_row)
        _ti_lay.setContentsMargins(0, 0, 0, 0); _ti_lay.setSpacing(6)
        _ti_lay.addWidget(QLabel("扭转角θ₁(°):"))
        self.edit_twist_angle_inlet = LineEdit()
        self.edit_twist_angle_inlet.setFixedWidth(72)
        self.edit_twist_angle_inlet.setPlaceholderText("留空取均值")
        self.edit_twist_angle_inlet.textChanged.connect(self._update_inlet_twist_xi)
        _ti_lay.addWidget(self.edit_twist_angle_inlet)
        _ti_lbl = QLabel("(15°~37°，留空则取均值 0.20)")
        _ti_lbl.setStyleSheet("color:#0066CC;font-size:11px;")
        _ti_lay.addWidget(_ti_lbl)
        _ti_lay.addStretch()
        self._twist_in_row.setVisible(False)
        ibl.addWidget(self._twist_in_row)

        # 进口 Row 2: v₂策略 + 末端流速v₂
        inlet_r2 = QHBoxLayout(); inlet_r2.setSpacing(6)
        inlet_r2.addWidget(QLabel("v₂策略:"))
        self.combo_v2_strategy = ComboBox()
        self.combo_v2_strategy.addItems(V2_STRATEGY_OPTIONS)
        self.combo_v2_strategy.setFixedWidth(180)
        self.combo_v2_strategy.currentTextChanged.connect(self._on_v2_strategy_changed)
        inlet_r2.addWidget(self.combo_v2_strategy)
        self.lbl_v2_strategy_hint = QLabel("(推荐)")
        self.lbl_v2_strategy_hint.setStyleSheet("color:#00AA00;font-size:12px;")
        inlet_r2.addWidget(self.lbl_v2_strategy_hint)
        inlet_r2.addWidget(QLabel("末端流速v₂(m/s):"))
        self.edit_v2 = LineEdit(); self.edit_v2.setPlaceholderText("留空=管道流速"); self.edit_v2.setFixedWidth(120)
        self.edit_v2.setReadOnly(True)  # 自动策略下初始为readonly
        self.edit_v2.editingFinished.connect(self._validate_inlet_velocity)
        inlet_r2.addWidget(self.edit_v2)
        self.lbl_v2_hint = QLabel("(计算后自动填充: 管道流速)")
        self.lbl_v2_hint.setStyleSheet("color:#424242;font-size:12px;")
        inlet_r2.addWidget(self.lbl_v2_hint)
        inlet_r2.addStretch()
        ibl.addLayout(inlet_r2)

        b2.addWidget(inlet_box)

        # ---- 出口 浅色底色框 ----
        outlet_box = QFrame()
        outlet_box.setObjectName("siphonOutletBox")
        outlet_box.setStyleSheet(
            "QFrame#siphonOutletBox { background:#F9FAFB;"
            " border:1px solid #ECEEF2; border-radius:4px; }"
        )
        obl = QVBoxLayout(outlet_box)
        obl.setContentsMargins(10, 6, 10, 6)
        obl.setSpacing(2)

        lbl_outlet_sec = QLabel("  ▸ 出口")
        lbl_outlet_sec.setStyleSheet(f"font-weight:bold;color:{T1};font-size:11px;")
        obl.addWidget(lbl_outlet_sec)

        # 出口 Row 1: 渐变段型式 + ξ₂ + 始端流速v
        outlet_r1 = QHBoxLayout(); outlet_r1.setSpacing(6)
        outlet_r1.addWidget(QLabel("渐变段型式:"))
        self.combo_outlet_type = ComboBox()
        self.combo_outlet_type.addItems(GRADIENT_TYPE_OPTIONS)
        self.combo_outlet_type.setCurrentText("无")
        self.combo_outlet_type.setFixedWidth(140)
        self.combo_outlet_type.currentTextChanged.connect(self._on_outlet_type_changed)
        outlet_r1.addWidget(self.combo_outlet_type)
        self.lbl_outlet_type_hint = QLabel("")
        self.lbl_outlet_type_hint.setStyleSheet("color:#0066CC;font-size:12px;")
        outlet_r1.addWidget(self.lbl_outlet_type_hint)
        btn_outlet_coeff_ref = PushButton("参考系数表")
        btn_outlet_coeff_ref.setMinimumWidth(110)
        btn_outlet_coeff_ref.setToolTip("查看表L.1.2 倒虹吸渐变段局部损失系数")
        btn_outlet_coeff_ref.clicked.connect(lambda: L12CoeffRefDialog(self).exec())
        outlet_r1.addWidget(btn_outlet_coeff_ref)
        outlet_r1.addWidget(QLabel("ξ₂:"))
        self.edit_xi_outlet = LineEdit(); self.edit_xi_outlet.setText("0.0"); self.edit_xi_outlet.setFixedWidth(80)
        outlet_r1.addWidget(self.edit_xi_outlet)
        self.lbl_xi_outlet_hint = QLabel("")
        self.lbl_xi_outlet_hint.setStyleSheet("color:#0066CC;font-size:12px;")
        outlet_r1.addWidget(self.lbl_xi_outlet_hint)
        outlet_r1.addWidget(QLabel("始端流速v(m/s):"))
        self.edit_v_out = LineEdit(); self.edit_v_out.setPlaceholderText("计算后自动填充"); self.edit_v_out.setFixedWidth(120)
        outlet_r1.addWidget(self.edit_v_out)
        self.lbl_vout_hint = QLabel("(=管道流速，无需填写)")
        self.lbl_vout_hint.setStyleSheet("color:#424242;font-size:12px;")
        outlet_r1.addWidget(self.lbl_vout_hint)
        outlet_r1.addStretch()
        obl.addLayout(outlet_r1)

        # 出口 Row 1.5: 直线扭曲面扭转角（仅选中直线扭曲面时显示）
        self._twist_out_row = QWidget()
        _to_lay = QHBoxLayout(self._twist_out_row)
        _to_lay.setContentsMargins(0, 0, 0, 0); _to_lay.setSpacing(6)
        _to_lay.addWidget(QLabel("扭转角θ₂(°):"))
        self.edit_twist_angle_outlet = LineEdit()
        self.edit_twist_angle_outlet.setFixedWidth(72)
        self.edit_twist_angle_outlet.setPlaceholderText("留空取均值")
        self.edit_twist_angle_outlet.textChanged.connect(self._update_outlet_twist_xi)
        _to_lay.addWidget(self.edit_twist_angle_outlet)
        _to_lbl = QLabel("(10°~17°，留空则取均值 0.40)")
        _to_lbl.setStyleSheet("color:#0066CC;font-size:11px;")
        _to_lay.addWidget(_to_lbl)
        _to_lay.addStretch()
        self._twist_out_row.setVisible(False)
        obl.addWidget(self._twist_out_row)

        # 出口 Row 2: 末端流速v₃
        outlet_r2 = QHBoxLayout(); outlet_r2.setSpacing(6)
        outlet_r2.addWidget(QLabel("末端流速v₃(m/s):"))
        self.edit_v3 = LineEdit(); self.edit_v3.setPlaceholderText("留空=0"); self.edit_v3.setFixedWidth(80)
        outlet_r2.addWidget(self.edit_v3)
        self.lbl_v3_hint = QLabel("(下游渠道流速)")
        self.lbl_v3_hint.setStyleSheet("color:#FF6600;font-size:12px;")
        outlet_r2.addWidget(self.lbl_v3_hint)
        outlet_r2.addStretch()
        obl.addLayout(outlet_r2)

        b2.addWidget(outlet_box)

        lay.addWidget(card2)
        lay.addStretch(1)

    def _build_segments_tab(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(2, 2, 2, 2)
        lay.setSpacing(3)

        # 工具栏
        tb = QHBoxLayout()
        tb.setSpacing(4)
        lbl = QLabel("结构段列表")
        lbl.setStyleSheet(f"font-size:12px;font-weight:bold;color:{T1};")
        tb.addWidget(lbl)
        tb.addStretch()

        self.lbl_seg_status = QLabel("")
        self.lbl_seg_status.setStyleSheet(f"color:{T2};font-size:12px;")
        tb.addWidget(self.lbl_seg_status)

        btn_dxf2 = PushButton("导入DXF"); btn_dxf2.clicked.connect(self._import_dxf)
        btn_add_pipe = PushButton("添加管身段"); btn_add_pipe.clicked.connect(self._add_segment_dialog)
        btn_add_common = PushButton("添加通用构件"); btn_add_common.clicked.connect(self._add_common_segment_dialog)
        btn_add_ptrans = PushButton("管道渐变段"); btn_add_ptrans.clicked.connect(self._add_pipe_transition)
        btn_add_ptrans.setToolTip("插入压力管道渐变段 ξjb（收缩0.05/扩散0.10），双击可切换类型")
        btn_del = PushButton("删除"); btn_del.clicked.connect(self._del_segment)
        btn_up = PushButton("↑"); btn_up.setFixedWidth(30); btn_up.clicked.connect(self._move_seg_up)
        btn_dn = PushButton("↓"); btn_dn.setFixedWidth(30); btn_dn.clicked.connect(self._move_seg_down)
        btn_clr_long = PushButton("清空纵断面"); btn_clr_long.clicked.connect(self._clear_longitudinal)
        btn_default = PushButton("默认构件"); btn_default.clicked.connect(self._init_default_segments)
        for w in [btn_dxf2, btn_add_pipe, btn_add_common, btn_add_ptrans, btn_del, btn_up, btn_dn, btn_clr_long, btn_default]:
            tb.addWidget(w)
        lay.addLayout(tb)

        # 结构段表格
        self.seg_table = QTableWidget(0, len(SEG_HEADERS))
        self.seg_table.setHorizontalHeaderLabels(SEG_HEADERS)
        _hdr = self.seg_table.horizontalHeader()
        _hdr.setSectionResizeMode(QHeaderView.Stretch)   # 其余列均分剩余空间
        _hdr.setStretchLastSection(False)
        # 类型列固定140px（完整显示"进水口(进口稍微修圆)"），锁定列固定50px
        _hdr.setSectionResizeMode(2, QHeaderView.Fixed)
        self.seg_table.setColumnWidth(2, 140)
        _hdr.setSectionResizeMode(11, QHeaderView.Fixed)
        self.seg_table.setColumnWidth(11, 50)
        self.seg_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.seg_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.seg_table.setAlternatingRowColors(True)
        self.seg_table.setFont(QFont("Microsoft YaHei", 10))
        self.seg_table.verticalHeader().setDefaultSectionSize(26)
        self.seg_table.doubleClicked.connect(self._on_seg_double_click)
        lay.addWidget(self.seg_table)

        # 颜色图例
        legend_lay = QHBoxLayout()
        legend_lay.setSpacing(12)
        for label_text, color in [("■ 通用构件", "#B8860B"), ("■ 平面段", "#4169E1"),
                                   ("■ 纵断面段", "#228B22"), ("■ 示例数据", "#999999")]:
            lbl = QLabel(label_text)
            lbl.setStyleSheet(f"color:{color};font-size:12px;")
            legend_lay.addWidget(lbl)
        legend_lay.addStretch()
        lay.addLayout(legend_lay)

        # 操作说明面板
        info_grp = QGroupBox("操作说明")
        info_lay = QVBoxLayout(info_grp)
        info_lay.setContentsMargins(8, 4, 8, 4)
        info_text = (
            "1. 点击\"导入DXF\"可从CAD文件导入管道几何\n"
            "2. 点击\"添加管身段\"手动添加直管/弯管/折管\n"
            "3. 点击\"添加通用构件\"可自定义添加构件（如镇墩、排气阀等）\n"
            "4. 双击表格行可编辑该行数据\n"
            "5. 使用↑↓按钮可调整顺序（首末行除外）\n"
            "6. 表格分三区：通用构件(黄) → 平面段(蓝) → 纵断面段(绿)\n"
            "7. 通用构件：进水口、拦污栅、闸门槽、旁通管、其他、出水口（仅贡献ξ）\n"
            "8. 管身段：直管、弯管、折管（涉及几何线形和水头损失计算）\n"
            "9. 初始纵断面数据为示例（灰色显示），导入DXF或手动添加后将自动替换"
        )
        info_lbl = QLabel(info_text)
        info_lbl.setStyleSheet(f"color:{T2};font-size:12px;")
        info_lbl.setWordWrap(True)
        info_lay.addWidget(info_lbl)
        lay.addWidget(info_grp)

    def _build_long_nodes_tab(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(2, 2, 2, 2)
        lay.setSpacing(3)

        tb = QHBoxLayout()
        tb.setSpacing(4)
        lbl = QLabel("纵断面变坡点")
        lbl.setStyleSheet(f"font-size:12px;font-weight:bold;color:{T1};")
        tb.addWidget(lbl)
        tb.addStretch()
        btn_add = PushButton("添加"); btn_add.clicked.connect(self._add_long_node)
        btn_del = PushButton("删除"); btn_del.clicked.connect(self._del_long_node)
        btn_dxf = PushButton("导入DXF"); btn_dxf.clicked.connect(self._import_dxf)
        btn_clr = PushButton("清空"); btn_clr.clicked.connect(self._clear_long_nodes)
        for w in [btn_add, btn_del, btn_dxf, btn_clr]:
            tb.addWidget(w)
        lay.addLayout(tb)

        self.long_table = QTableWidget(0, len(LONG_NODE_HEADERS))
        self.long_table.setHorizontalHeaderLabels(LONG_NODE_HEADERS)
        self.long_table.horizontalHeader().setStretchLastSection(False)
        self.long_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.long_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.long_table.setAlternatingRowColors(True)
        self.long_table.setFont(QFont("Microsoft YaHei", 10))
        self.long_table.verticalHeader().setDefaultSectionSize(26)
        self.long_table.cellChanged.connect(self._on_long_table_edited)
        lay.addWidget(self.long_table)

    # ---- C: 操作栏 ----
    def _build_operation_bar(self, parent_lay):
        bar = QHBoxLayout()
        bar.setSpacing(6)

        bar.addWidget(QLabel("名称:"))
        self.edit_name = LineEdit(); self.edit_name.setPlaceholderText("倒虹吸"); self.edit_name.setFixedWidth(100)
        bar.addWidget(self.edit_name)

        bar.addStretch()

        self.detail_cb = CheckBox("输出详细过程")
        self.detail_cb.setChecked(True)
        bar.addWidget(self.detail_cb)

        btn_calc = PrimaryPushButton("执行计算")
        btn_calc.setFixedWidth(100)
        btn_calc.clicked.connect(self._execute_calculation)
        bar.addWidget(btn_calc)

        parent_lay.addLayout(bar)

    # ---- 计算结果Tab（与基本参数/结构段信息/纵断面节点并列） ----
    def _build_result_tab(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(2, 2, 2, 2)
        lay.setSpacing(2)

        # 工具栏：导出 + 清空
        tb = QHBoxLayout(); tb.setSpacing(4)
        tb.addStretch()
        btn_word = PushButton("导出Word"); btn_word.clicked.connect(self._export_word)
        btn_clr = PushButton("清空"); btn_clr.clicked.connect(self._clear_results)
        for w in [btn_word, btn_clr]:
            tb.addWidget(w)
        lay.addLayout(tb)

        # 结果子标签页（占满剩余空间）
        self.result_notebook = QTabWidget()
        lay.addWidget(self.result_notebook, 1)

        # Sub-Tab1: 结果汇总
        t1 = QWidget()
        t1l = QVBoxLayout(t1); t1l.setContentsMargins(2, 2, 2, 2)
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        self.summary_text.setFont(QFont("Consolas", 10))
        t1l.addWidget(self.summary_text)
        self.result_notebook.addTab(t1, "结果汇总")

        # Sub-Tab2: 详细计算过程
        t2 = QWidget()
        t2l = QVBoxLayout(t2); t2l.setContentsMargins(2, 2, 2, 2)
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        self.detail_text.setFont(QFont("Consolas", 10))
        t2l.addWidget(self.detail_text)
        self.result_notebook.addTab(t2, "详细计算过程")

        # Sub-Tab3: 计算公式（QWebEngineView + KaTeX 渲染）
        self.formula_view = QWebEngineView()
        self.formula_view.setHtml(self._build_formula_html())
        self.result_notebook.addTab(self.formula_view, "计算公式")

        self._show_help()

    @staticmethod
    def _build_formula_html():
        """生成计算公式页面HTML（方案A：WinUI3亚克力毛玻璃 + KaTeX渲染）"""
        return r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/katex@0.16.9/dist/katex.min.css">
<script src="https://cdn.jsdelivr.net/npm/katex@0.16.9/dist/katex.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/katex@0.16.9/dist/contrib/auto-render.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{
  font-family:"Segoe UI Variable","Segoe UI","Microsoft YaHei UI",sans-serif;
  background:linear-gradient(135deg,#e3f0ff 0%,#f0e6ff 100%);
  color:#1a1a1a; padding:20px 18px; min-height:100vh;
}
.group-header{
  display:flex;align-items:center;gap:8px;
  padding:10px 16px;margin-bottom:12px;
  background:linear-gradient(90deg,#0067c0,#005ba1);
  border-radius:6px;color:#fff;font-size:14px;font-weight:600;
}
.group-header svg{width:18px;height:18px;fill:#fff;}
.card{
  background:rgba(255,255,255,.72);
  backdrop-filter:blur(40px) saturate(1.6);
  -webkit-backdrop-filter:blur(40px) saturate(1.6);
  border:1px solid rgba(255,255,255,.85);
  border-radius:8px;padding:16px 20px;margin-bottom:10px;
  box-shadow:0 2px 8px rgba(0,0,0,.06),0 0 1px rgba(0,0,0,.08);
  transition:box-shadow .2s,transform .15s;
}
.card:hover{
  box-shadow:0 4px 16px rgba(0,0,0,.10),0 0 1px rgba(0,0,0,.12);
  transform:translateY(-1px);
}
.card-label{
  font-size:12px;font-weight:600;color:#0067c0;
  margin-bottom:6px;display:flex;align-items:center;gap:6px;
}
.card-label .dot{
  width:6px;height:6px;border-radius:50%;background:#0067c0;
}
.card .katex-display{margin:6px 0 0 0!important;}
.card .katex{font-size:1.15em;color:#1a1a1a;}
.card .note{
  font-size:13px;color:#333;margin-top:6px;
  padding-left:12px;border-left:2px solid #0067c0;
}
.group{margin-bottom:20px;}
</style>
</head>
<body>

<div class="group">
  <div class="group-header">
    <svg viewBox="0 0 24 24"><path d="M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm-1 17.93c-3.94-.49-7-3.85-7-7.93 0-.62.08-1.22.21-1.79L9 15v1c0 1.1.9 2 2 2v1.93z"/></svg>
    水力计算基本公式
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>谢才公式（Chézy）</div>
    $$C = \frac{1}{n} R^{\,1/6}$$
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>沿程水头损失</div>
    $$h_f = \frac{L \, v^2}{C^2 \, R}$$
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>局部水头损失</div>
    $$h_j = \sum \xi_i \, \frac{v^2}{2g}$$
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>总水头损失</div>
    $$\Delta Z = \Delta Z_1 + \Delta Z_2 - \Delta Z_3$$
    <div class="note">ΔZ₁ = 进口渐变段水面落差（进口局部损失 + 流速水头增加）；ΔZ₂ = 管身段总损失（沿程 + 管内局部）；ΔZ₃ = 出口渐变段净回升水头（动能回收，取减号）</div>
  </div>
</div>

<div class="group">
  <div class="group-header">
    <svg viewBox="0 0 24 24"><path d="M12 2l3.09 6.26L22 9.27l-5 4.87 1.18 6.88L12 17.77l-6.18 3.25L7 14.14 2 9.27l6.91-1.01L12 2z"/></svg>
    空间轴线合并算法（平面 + 纵断面 → 三维空间曲线）
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>三维单位切向量</div>
    $$\mathbf{T} = \begin{pmatrix} \cos\beta\,\cos\alpha \\[4pt] \cos\beta\,\sin\alpha \\[4pt] \sin\beta \end{pmatrix}$$
    <div class="note">α = 数学方位角（正东=0°，逆时针）；β = 纵断面坡角</div>
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>空间转角</div>
    $$\theta_{3D} = \arccos\!\Big(\mathbf{T}_{\text{before}} \cdot \mathbf{T}_{\text{after}}\Big)$$
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>空间长度</div>
    $$L_{\text{spatial}} = \sum_{i} \sqrt{\Delta s_i^{\,2} + \Delta Z_i^{\,2}}$$
    <div class="note">Δs = 桩号差（平面弧长参数增量），非 XY 弦长</div>
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>坡角计算</div>
    $$\beta = \arctan\!\left(\frac{\Delta Z}{\Delta s}\right)$$
    <div class="note">用桩号差 Δs（弧长参数）替代 XY 弦长，消除圆弧段系统性偏差</div>
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>曲率合成（重叠弯道 · 微分几何）</div>
    $$\kappa^2 = \frac{1}{R_v^2} + \frac{\cos^4\!\beta}{R_h^2}$$
    $$R_{3D} = \frac{R_h \, R_v}{\sqrt{R_h^2 + R_v^2 \cos^4\!\beta}}$$
    <div class="note">极限校核：$R_v \to \infty$ 时 $R_{3D} = R_h/\cos^2\!\beta$；$R_h \to \infty$ 时 $R_{3D} = R_v$</div>
  </div>
</div>

<div class="group">
  <div class="group-header">
    <svg viewBox="0 0 24 24"><path d="M19 3H5c-1.1 0-2 .9-2 2v14c0 1.1.9 2 2 2h14c1.1 0 2-.9 2-2V5c0-1.1-.9-2-2-2zm-7 14H7v-2h5v2zm3-4H7v-2h8v2zm0-4H7V7h8v2z"/></svg>
    局部阻力系数公式（附录 L）
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>弯管局部阻力系数（表 L.1.4-3 / L.1.4-4）</div>
    $$\xi = \xi_{90} \times \gamma$$
    <div class="note">$\xi_{90}$：查表 L.1.4-3，按 $R/D_0$ 线性插值；$\gamma$：查表 L.1.4-4，按弯管圆心角 $\theta$ 线性插值</div>
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>折管局部阻力系数</div>
    $$\zeta = 0.9457\sin^2\!\left(\frac{\theta}{2}\right) + 2.047\sin^4\!\left(\frac{\theta}{2}\right)$$
    <div class="note">$\theta$ 为折管折角（度）</div>
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>拦污栅局部阻力系数 — 无独立支墩（公式 L.1.4-2）</div>
    $$\xi_s = \beta_1 \left(\frac{s_1}{b_1}\right)^{4/3} \sin\alpha$$
    <div class="note">$\beta_1$：栅条形状系数（表 L.1.4-1）；$s_1$：栅条厚度；$b_1$：栅条间距；$\alpha$：栅面倾角</div>
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>拦污栅局部阻力系数 — 有独立支墩（公式 L.1.4-3）</div>
    $$\xi_s = \left[\beta_1 \left(\frac{s_1}{b_1}\right)^{4/3} + \beta_2 \left(\frac{s_2}{b_2}\right)^{4/3}\right] \sin\alpha$$
    <div class="note">$\beta_2$：支墩形状系数；$s_2$：支墩厚度；$b_2$：支墩净距</div>
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>出水口局部阻力系数</div>
    $$\xi_c = \left(1 - \frac{\omega_g}{\omega_q}\right)^2$$
    <div class="note">$\omega_g$：管道断面积 $= Q/v$；$\omega_q$：下游明渠断面积</div>
  </div>
  <div class="card">
    <div class="card-label"><span class="dot"></span>管道渐变段局部阻力系数</div>
    $$\xi_{jb} = \begin{cases} 0.05 & \text{收缩（方变圆 / 圆管收缩）} \\ 0.10 & \text{扩散（圆变方 / 圆管扩大，扩散角} \leq 10^{\circ} \text{）} \end{cases}$$
    <div class="note">GB 50288-2018 附录 L 第 4 条</div>
  </div>
</div>

<script>
document.addEventListener("DOMContentLoaded", function(){
  renderMathInElement(document.body, {
    delimiters:[{left:"$$",right:"$$",display:true},{left:"$",right:"$",display:false}],
    throwOnError:false
  });
});
</script>
</body>
</html>'''

    def _show_help(self):
        lines = [
            "=" * 60,
            "  倒虹吸水力计算 — 使用说明",
            "=" * 60, "",
            "操作步骤：",
            "  1. 设置全局参数（设计流量、拟定流速、糙率）",
            "  2. 设置渐变段型式和局部损失系数",
            "  3. 设置进出口流速（v₁、v₃）和v₂策略",
            "  4. 管理结构段（通用构件 + 管身段）",
            "  5. 输入纵断面变坡点（手动或导入DXF）",
            "  6. 点击「计算」执行水力计算", "",
            "结构段操作说明：",
            "  1. 点击\"导入DXF\"可从CAD文件导入管道几何",
            "  2. 点击\"添加管身段\"手动添加直管/弯管/折管",
            "  3. 点击\"添加通用构件\"可自定义添加构件（如镇墩、排气阀等）",
            "  4. 双击表格行可编辑该行数据",
            "  5. 使用↑↓按钮可调整顺序（首末行除外）",
            "  6. 表格分三区：通用构件(黄) → 平面段(蓝) → 纵断面段(绿)",
            "  7. 通用构件：进水口、拦污栅、闸门槽、旁通管、其他、出水口（仅贡献ξ）",
            "  8. 管身段：直管、弯管、折管（涉及几何线形和水头损失计算）",
            "  9. 初始纵断面数据为示例（灰色显示），导入DXF或手动添加后将自动替换", "",
            "功能特性：",
            "  - 管道可视化（纵断面/平面视图，支持缩放平移）",
            "  - 7种专业对话框（进出水口、拦污栅、弯管等）",
            "  - DXF导入纵断面多段线",
            "  - 计算完成自动回填流速值",
            "  - 水损阈值超限提醒",
            "  - 支持从推求水面线模块导入参数",
            "=" * 60,
        ]
        self.detail_text.setPlainText("\n".join(lines))

    # ================================================================
    # 外部接口（与推求水面线联动）
    # ================================================================
    def set_params(self, **kwargs):
        """从外部（推求水面线模块）设置参数"""
        if 'Q' in kwargs:
            self.edit_Q.setText(f"{kwargs['Q']:.4f}")
            self.edit_Q.setStyleSheet("color:#0066CC;")
            self.lbl_Q_hint.setText("(已从主表导入)")
        if 'n' in kwargs or 'roughness_n' in kwargs:
            n_val = kwargs.get('n') or kwargs.get('roughness_n')
            if n_val is not None:
                self.edit_n.setText(f"{n_val:.4f}")
                self.edit_n.setStyleSheet("color:#0066CC;")
                self.lbl_n_hint.setText("(已从主表导入)")
        if 'v_guess' in kwargs:
            self._syncing = True
            self.edit_v.setText(f"{kwargs['v_guess']:.4f}")
            self._syncing = False
            self.edit_v.setStyleSheet(f"LineEdit {{ border: 1.5px solid orange; }}")
            self.lbl_v_hint.setText("(已从主表导入, 请确认)")
            self.lbl_v_hint.setStyleSheet(f"color:#CC6600;font-size:12px;")

        # 渐变段型式（需在设置系数之前，因为型式变化会触发系数更新）
        if 'inlet_type' in kwargs and SIPHON_AVAILABLE:
            gt = kwargs['inlet_type']
            if isinstance(gt, str):
                self.combo_inlet_type.setCurrentText(gt)
            else:
                name = GRADIENT_TYPE_MAP_REV.get(gt, "无")
                self.combo_inlet_type.setCurrentText(name)
            self.lbl_inlet_type_hint.setText("(已从主表导入)")
        if 'outlet_type' in kwargs and SIPHON_AVAILABLE:
            gt = kwargs['outlet_type']
            if isinstance(gt, str):
                self.combo_outlet_type.setCurrentText(gt)
            else:
                name = GRADIENT_TYPE_MAP_REV.get(gt, "无")
                self.combo_outlet_type.setCurrentText(name)
            self.lbl_outlet_type_hint.setText("(已从主表导入)")

        # 系数（在型式之后设置，覆盖型式自动填充的值）
        if 'xi_inlet' in kwargs and kwargs['xi_inlet']:
            self.edit_xi_inlet.setText(f"{kwargs['xi_inlet']:.4f}")
            self.edit_xi_inlet.setStyleSheet("color:#0066CC;")
            self.lbl_xi_inlet_hint.setText("(已从主表导入)")
        if 'xi_outlet' in kwargs and kwargs['xi_outlet']:
            self.edit_xi_outlet.setText(f"{kwargs['xi_outlet']:.4f}")
            self.edit_xi_outlet.setStyleSheet("color:#0066CC;")
            self.lbl_xi_outlet_hint.setText("(已从主表导入)")

        # 进口断面参数（供断面参数计算v₂策略使用，仅B>0且h>0时设置）
        _sec_B = kwargs.get('inlet_section_B')
        _sec_h = kwargs.get('inlet_section_h')
        _sec_m = kwargs.get('inlet_section_m')
        if _sec_B is not None and _sec_B > 0 and _sec_h is not None and _sec_h > 0:
            self._section_B = _sec_B
            self._section_h = _sec_h
            self._section_m = _sec_m if _sec_m is not None else 0.0

        # 流速
        if 'v_channel_in' in kwargs and kwargs['v_channel_in']:
            self.edit_v1.setText(f"{kwargs['v_channel_in']:.4f}")
            self.edit_v1.setStyleSheet(f"color:#0066CC;")
            self.lbl_v1_hint.setText("(已导入上游渠道断面平均流速)")
            self.lbl_v1_hint.setStyleSheet(f"color:#0066CC;font-size:12px;")
            # 仅在 V1_PLUS_02 策略下联动更新 v₂
            self._on_v_channel_in_changed()
        if 'v_pipe_out' in kwargs and kwargs['v_pipe_out']:
            self.edit_v3.setText(f"{kwargs['v_pipe_out']:.4f}")
            self.edit_v3.setStyleSheet(f"color:#0066CC;")
            self.lbl_v3_hint.setText("(已导入下游渠道断面平均流速)")
            self.lbl_v3_hint.setStyleSheet(f"color:#0066CC;font-size:12px;")
        if 'v_channel_out' in kwargs and kwargs['v_channel_out']:
            self.edit_v_out.setText(f"{kwargs['v_channel_out']:.4f}")
            self.edit_v_out.setStyleSheet(f"color:#0066CC;")
            self.lbl_vout_hint.setText("(已从主表导入)")
            self.lbl_vout_hint.setStyleSheet(f"color:#0066CC;font-size:12px;")

        # 平面转弯半径倍数（同步内部变量，与原版一致）
        if 'siphon_turn_radius_n' in kwargs and kwargs['siphon_turn_radius_n']:
            n_val = float(kwargs['siphon_turn_radius_n'])
            self._siphon_turn_radius_n = n_val
            self._syncing = True
            self.edit_turn_n.setText(str(n_val))
            self._syncing = False
            self._update_turn_n_style()
            self._update_turn_R()

        # 下游断面参数（出水口系数计算用）
        for key in ('outlet_downstream_type', 'outlet_downstream_B',
                    'outlet_downstream_h', 'outlet_downstream_m',
                    'outlet_downstream_D', 'outlet_downstream_R'):
            if key in kwargs:
                self._downstream_params[key] = kwargs[key]
        if 'downstream_params' in kwargs:
            self._downstream_params.update(kwargs['downstream_params'])

        # 已构建的结构段列表
        if 'segments' in kwargs:
            self.segments = kwargs['segments']
            self._refresh_seg_table()

        # 平面段（字典列表 → StructureSegment）
        if 'plan_segments' in kwargs:
            data = kwargs['plan_segments']
            if data and isinstance(data[0], dict):
                self._set_plan_segments(data)
            else:
                self.plan_segments = data
        if 'plan_feature_points' in kwargs:
            data = kwargs['plan_feature_points']
            if data and isinstance(data[0], dict):
                self._set_plan_feature_points(data)
            else:
                self.plan_feature_points = data
        if 'plan_total_length' in kwargs:
            self.plan_total_length = kwargs['plan_total_length']

        # 纵断面节点
        if 'longitudinal_nodes' in kwargs:
            self.longitudinal_nodes = kwargs['longitudinal_nodes']
            self._longitudinal_is_example = False
            self._refresh_long_table()

        if 'siphon_name' in kwargs:
            self.edit_name.setText(kwargs['siphon_name'])

        # 更新平面转弯半径
        self._update_plan_bend_radius()
        # 自动计算出水口系数
        self._auto_compute_outlet_xi()

        self._update_canvas()
        self._update_data_status()
        self._update_D_theory()
        self._update_turn_R()

    def get_result(self):
        return self.calculation_result

    def get_plan_bend_radius(self) -> float:
        """获取平面转弯半径，用于写回水面线表格。
        优先取 plan_feature_points 中最大非零转弯半径；
        无有效 IP 点数据则取 edit_turn_R 输入框的值。
        """
        if self.plan_feature_points:
            radii = [fp.turn_radius for fp in self.plan_feature_points
                     if getattr(fp, 'turn_radius', 0) and fp.turn_radius > 0]
            if radii:
                return max(radii)
        return self._fval(self.edit_turn_R, 0.0)

    def get_total_head_loss(self):
        """获取总水头损失（便捷方法）"""
        if self.calculation_result:
            return self.calculation_result.total_head_loss
        return None

    # ---- set_params 辅助方法 ----

    def _set_plan_segments(self, plan_data):
        """从字典列表转换平面段为StructureSegment列表"""
        if not SIPHON_AVAILABLE:
            return
        self.plan_segments = []
        for item in plan_data:
            seg_type_str = item.get("segment_type", "直管")
            seg_type = SegmentType.STRAIGHT
            for st in SegmentType:
                if st.value == seg_type_str:
                    seg_type = st
                    break
            seg = StructureSegment(
                segment_type=seg_type,
                direction=SegmentDirection.PLAN,
                length=item.get("length", 0.0),
                radius=item.get("radius", 0.0),
                angle=item.get("angle", 0.0),
                locked=True,
                source_ip_index=item.get("source_ip_index"),
            )
            self.plan_segments.append(seg)
        self._update_plan_bend_radius()
        self._refresh_seg_table()

    def _set_plan_feature_points(self, fp_data):
        """从字典列表转换平面IP特征点"""
        if not SIPHON_AVAILABLE:
            return
        self.plan_feature_points = []
        for item in fp_data:
            tt = TurnType.NONE
            for t in TurnType:
                if t.value == item.get("turn_type", "无"):
                    tt = t
                    break
            fp = PlanFeaturePoint(
                chainage=item.get("chainage", 0.0),
                x=item.get("x", 0.0),
                y=item.get("y", 0.0),
                azimuth_meas_deg=item.get("azimuth", 0.0),
                turn_radius=item.get("turn_radius", 0.0),
                turn_angle=item.get("turn_angle", 0.0),
                turn_type=tt,
                ip_index=item.get("ip_index", 0),
            )
            self.plan_feature_points.append(fp)

    def _update_plan_bend_radius(self):
        """根据 n×D 自动更新平面弯管段的半径"""
        if not SIPHON_AVAILABLE:
            return
        n_mult = self._fval(self.edit_turn_n, 0)
        if n_mult <= 0:
            return
        Q = self._fval(self.edit_Q, 0)
        v = self._fval(self.edit_v, 0)
        if Q <= 0 or v <= 0:
            return
        N = max(1, self.spin_num_pipes.value()) if hasattr(self, 'spin_num_pipes') else 1
        Q_single = Q / N
        D_theory = math.sqrt(4 * Q_single / (math.pi * v))
        if D_theory <= 0:
            return
        D_design = HydraulicCore.round_diameter(D_theory)
        siphon_radius = round(n_mult * D_design, 2)

        updated = False
        for seg in self.plan_segments:
            if seg.segment_type == SegmentType.BEND:
                seg.radius = siphon_radius
                if seg.angle > 0:
                    seg.length = round(siphon_radius * math.radians(seg.angle), 3)
                if seg.angle > 0 and D_theory > 0 and seg.xi_user is None:
                    seg.xi_calc = CoefficientService.calculate_bend_coeff(
                        siphon_radius, D_theory, seg.angle, verbose=False)
                updated = True

        for fp in self.plan_feature_points:
            if fp.turn_angle > 0 and fp.turn_type != TurnType.NONE:
                fp.turn_radius = siphon_radius

        if updated:
            self._refresh_seg_table()

    def _auto_compute_outlet_xi(self):
        """自动计算出水口局部阻力系数"""
        if not SIPHON_AVAILABLE:
            return
        dp = self._downstream_params
        ds_type = dp.get('outlet_downstream_type', '')
        if not ds_type:
            return
        Q = self._fval(self.edit_Q, 0)
        v = self._fval(self.edit_v, 0)
        if Q <= 0 or v <= 0:
            return
        N = max(1, self.spin_num_pipes.value()) if hasattr(self, 'spin_num_pipes') else 1
        Q_single = Q / N
        omega_g = Q_single / v  # 单管断面积近似
        omega_q = 0.0

        # 根据下游类型计算断面积
        B = dp.get('outlet_downstream_B', 0) or 0
        h = dp.get('outlet_downstream_h', 0) or 0
        m = dp.get('outlet_downstream_m', 0) or 0
        D = dp.get('outlet_downstream_D', 0) or 0

        R_ds = dp.get('outlet_downstream_R', 0) or 0

        if '梯形' in ds_type and B > 0 and h > 0:
            omega_q = (B + m * h) * h
        elif '矩形' in ds_type and B > 0 and h > 0:
            omega_q = B * h
        elif '圆形' in ds_type and D > 0:
            r = D / 2.0
            if h >= D:
                omega_q = math.pi * r * r
            else:
                cos_val = max(-1.0, min(1.0, (r - h) / r))
                theta = math.acos(cos_val)
                omega_q = r * r * (theta - math.sin(theta) * math.cos(theta))
        elif 'U' in ds_type and R_ds > 0 and h > 0:
            if h <= R_ds:
                cos_val = max(-1.0, min(1.0, (R_ds - h) / R_ds))
                theta = math.acos(cos_val)
                omega_q = R_ds * R_ds * (theta - math.sin(theta) * math.cos(theta))
            else:
                omega_q = math.pi * R_ds * R_ds / 2.0 + 2 * R_ds * (h - R_ds)
        elif '圆拱直墙' in ds_type and B > 0 and R_ds > 0 and h > 0:
            wall_h = max(h - R_ds, 0)
            rect_part = B * wall_h
            if h >= wall_h + R_ds:
                arch_part = math.pi * R_ds * R_ds / 2.0
            else:
                arch_h = h - wall_h
                if arch_h > 0:
                    cos_val = max(-1.0, min(1.0, (R_ds - arch_h) / R_ds))
                    theta = math.acos(cos_val)
                    arch_part = R_ds * R_ds * (theta - math.sin(theta) * math.cos(theta))
                else:
                    arch_part = 0.0
            omega_q = rect_part + arch_part
        elif '马蹄' in ds_type and R_ds > 0 and h > 0:
            d_hs = 2 * R_ds
            if h >= d_hs:
                omega_q = math.pi * R_ds * R_ds
            else:
                cos_val = max(-1.0, min(1.0, (R_ds - h) / R_ds))
                theta = math.acos(cos_val)
                omega_q = R_ds * R_ds * (theta - math.sin(theta) * math.cos(theta))

        if omega_q > 0:
            xi_c = (1 - omega_g / omega_q) ** 2
            for seg in self.segments:
                if seg.segment_type == SegmentType.OUTLET:
                    seg.xi_calc = round(xi_c, 4)
            self._refresh_seg_table()

    def to_dict(self):
        """序列化为字典（项目保存用）"""
        d = {
            'Q': self._fval(self.edit_Q),
            'v_guess': self._fval(self.edit_v),
            'n': self._fval(self.edit_n),
            'turn_n': self._fval(self.edit_turn_n, 5),
            'threshold': self.edit_threshold.text().strip(),
            'inlet_type': self.combo_inlet_type.currentText(),
            'outlet_type': self.combo_outlet_type.currentText(),
            'xi_inlet': self._fval(self.edit_xi_inlet),
            'xi_outlet': self._fval(self.edit_xi_outlet),
            'v1': self.edit_v1.text().strip(),
            'v2': self.edit_v2.text().strip(),
            'v_out': self.edit_v_out.text().strip(),
            'v3': self.edit_v3.text().strip(),
            'v2_strategy': self.combo_v2_strategy.currentText(),
            'name': self.edit_name.text().strip(),
            'show_detail': self.detail_cb.isChecked(),
            'D_override': self.edit_D_override.text().strip(),
            'num_pipes': self.spin_num_pipes.value() if hasattr(self, 'spin_num_pipes') else 1,
        }
        if SIPHON_AVAILABLE:
            d['segments'] = [self._seg_to_dict(s) for s in self.segments]
            d['plan_segments'] = [self._seg_to_dict(s) for s in self.plan_segments]
            d['plan_total_length'] = self.plan_total_length
            d['plan_feature_points'] = [fp.to_dict() for fp in self.plan_feature_points]
            d['longitudinal_nodes'] = [
                {'chainage': nd.chainage, 'elevation': nd.elevation,
                 'vcr': nd.vertical_curve_radius,
                 'turn_type': nd.turn_type.value if hasattr(nd.turn_type, 'value') else str(nd.turn_type),
                 'turn_angle': nd.turn_angle}
                for nd in self.longitudinal_nodes
            ]
            d['longitudinal_is_example'] = self._longitudinal_is_example
        # 保存计算结果
        if self.calculation_result:
            d['total_head_loss'] = self.calculation_result.total_head_loss
            d['diameter'] = self.calculation_result.diameter
            d['velocity'] = self.calculation_result.velocity
        return d

    def from_dict(self, d):
        """从字典恢复状态（项目加载用）"""
        if 'Q' in d: self.edit_Q.setText(str(d['Q']))
        if 'v_guess' in d:
            self._syncing = True
            self.edit_v.setText(str(d['v_guess']))
            self._syncing = False
            self._update_v_style()
        # 兼容Tkinter版key: roughness_n → n
        n_val = d.get('n') or d.get('roughness_n')
        if n_val is not None: self.edit_n.setText(str(n_val))
        turn_n_val = d.get('turn_n') or d.get('siphon_turn_radius_n')
        if turn_n_val is not None:
            self._syncing = True
            self.edit_turn_n.setText(str(turn_n_val))
            self._syncing = False
            self._update_turn_n_style()
            self._update_turn_R()
        threshold_val = d.get('threshold') or d.get('head_loss_threshold')
        if threshold_val is not None: self.edit_threshold.setText(str(threshold_val))
        if 'D_override' in d:
            d_val = str(d['D_override']).strip()
            if d_val:
                self.cb_D_override.setChecked(True)
                self.edit_D_override.setText(d_val)
        if 'num_pipes' in d and hasattr(self, 'spin_num_pipes'):
            self._syncing = True
            self.spin_num_pipes.setValue(int(d['num_pipes']))
            self._syncing = False
            self._num_pipes_user_confirmed = True
            self._update_num_pipes_style()
        if 'inlet_type' in d: self.combo_inlet_type.setCurrentText(d['inlet_type'])
        if 'outlet_type' in d: self.combo_outlet_type.setCurrentText(d['outlet_type'])
        if 'xi_inlet' in d: self.edit_xi_inlet.setText(str(d['xi_inlet']))
        if 'xi_outlet' in d: self.edit_xi_outlet.setText(str(d['xi_outlet']))
        # 恢复v₂策略（需在恢复v₂值之前，触发UI状态同步）
        # 兼容Tkinter版枚举值与PySide6版选项文字的差异
        _V2_COMPAT = {
            "自动（= 管道流速）": "自动（=管道流速）",
            "由断面参数计算": "断面参数计算",
        }
        if 'v2_strategy' in d:
            v2s = _V2_COMPAT.get(d['v2_strategy'], d['v2_strategy'])
            self.combo_v2_strategy.setCurrentText(v2s)
            self._on_v2_strategy_changed(v2s)
        # 恢复流速（空值时保持placeholder，兼容Tkinter版key名）
        v1_val = d.get('v1') or d.get('v_channel_in')
        if v1_val:
            v1_str = str(v1_val).strip()
            if v1_str and v1_str != '0' and v1_str != '0.0':
                self.edit_v1.setText(v1_str)
        v2_val = d.get('v2') or d.get('v_pipe_in')
        if v2_val:
            v2_str = str(v2_val).strip()
            if v2_str:
                was_ro = self.edit_v2.isReadOnly()
                self.edit_v2.setReadOnly(False)
                self.edit_v2.setText(v2_str)
                self.edit_v2.setReadOnly(was_ro)
        v_out_val = d.get('v_out') or d.get('v_channel_out')
        if v_out_val:
            v_out_str = str(v_out_val).strip()
            if v_out_str and v_out_str != '0' and v_out_str != '0.0':
                self.edit_v_out.setText(v_out_str)
        v3_val = d.get('v3') or d.get('v_pipe_out')
        if v3_val:
            v3_str = str(v3_val).strip()
            if v3_str and v3_str != '0' and v3_str != '0.0':
                self.edit_v3.setText(v3_str)
        if 'name' in d: self.edit_name.setText(d['name'])
        if 'show_detail' in d: self.detail_cb.setChecked(d['show_detail'])
        if 'segments' in d and SIPHON_AVAILABLE:
            self.segments = [self._dict_to_seg(sd) for sd in d['segments']]
            self._refresh_seg_table()
        # 恢复平面段
        if 'plan_segments' in d and d['plan_segments'] and SIPHON_AVAILABLE:
            self.plan_segments = [self._dict_to_seg(sd) for sd in d['plan_segments']]
        if 'plan_total_length' in d:
            self.plan_total_length = d['plan_total_length']
        # 恢复平面IP特征点
        if 'plan_feature_points' in d and d['plan_feature_points'] and SIPHON_AVAILABLE:
            self.plan_feature_points = [
                PlanFeaturePoint.from_dict(fp) for fp in d['plan_feature_points']
            ]
        if 'longitudinal_nodes' in d and SIPHON_AVAILABLE:
            self.longitudinal_nodes = []
            for nd in d['longitudinal_nodes']:
                tt = TurnType.NONE
                tt_val = nd.get('turn_type', 'NONE')
                for t in TurnType:
                    if t.value == tt_val or t.name == tt_val:
                        tt = t
                        break
                self.longitudinal_nodes.append(LongitudinalNode(
                    chainage=nd.get('chainage', 0),
                    elevation=nd.get('elevation', 0),
                    vertical_curve_radius=nd.get('vcr', 0),
                    turn_type=tt,
                    turn_angle=nd.get('turn_angle', 0),
                ))
            self._refresh_long_table()
        # 恢复纵断面示例标志（兼容旧数据）
        if 'longitudinal_is_example' in d:
            self._longitudinal_is_example = d['longitudinal_is_example']
        else:
            self._longitudinal_is_example = len(self.longitudinal_nodes) < 2
        self._refresh_seg_table()
        self._update_canvas()
        self._update_data_status()
        self._update_D_theory()
        self._update_turn_R()

    def _seg_to_dict(self, seg):
        d = {
            'type': seg.segment_type.value,
            'direction': seg.direction.value if hasattr(seg.direction, 'value') else str(seg.direction),
            'length': seg.length, 'radius': seg.radius, 'angle': seg.angle,
            'xi_user': seg.xi_user, 'xi_calc': seg.xi_calc,
            'locked': seg.locked,
        }
        if hasattr(seg, 'custom_label') and seg.custom_label:
            d['custom_label'] = seg.custom_label
        if hasattr(seg, 'inlet_shape') and seg.inlet_shape:
            d['inlet_shape'] = seg.inlet_shape.value
        if hasattr(seg, 'outlet_shape') and seg.outlet_shape:
            d['outlet_shape'] = seg.outlet_shape.value
        if hasattr(seg, 'start_elevation') and seg.start_elevation is not None:
            d['start_elev'] = seg.start_elevation
        if hasattr(seg, 'end_elevation') and seg.end_elevation is not None:
            d['end_elev'] = seg.end_elevation
        if hasattr(seg, 'source_ip_index') and seg.source_ip_index is not None:
            d['source_ip_index'] = seg.source_ip_index
        if hasattr(seg, 'trash_rack_params') and seg.trash_rack_params is not None:
            d['trash_rack_params'] = seg.trash_rack_params.to_dict()
        return d

    def _dict_to_seg(self, d):
        st = SegmentType.OTHER
        for s in SegmentType:
            if s.value == d.get('type', ''):
                st = s
                break
        direction = SegmentDirection.COMMON
        dir_val = d.get('direction', '')
        for dd in SegmentDirection:
            if dd.value == dir_val or dd.name == dir_val:
                direction = dd
                break
        # 旧数据迁移：通用构件类型但 direction 不是 COMMON，自动修正
        if is_common_type(st) and direction != SegmentDirection.COMMON:
            direction = SegmentDirection.COMMON
        inlet_shape = None
        if 'inlet_shape' in d:
            for sh in InletOutletShape:
                if sh.value == d['inlet_shape']:
                    inlet_shape = sh
                    break
        outlet_shape = None
        if 'outlet_shape' in d:
            for sh in InletOutletShape:
                if sh.value == d['outlet_shape']:
                    outlet_shape = sh
                    break
        xi_user = d.get('xi_user')
        xi_calc = d.get('xi_calc')
        # 进水口系数校验：若系数不在该形状的合理范围内，自动修正为默认中值
        if st == SegmentType.INLET and inlet_shape and inlet_shape in INLET_SHAPE_COEFFICIENTS:
            coeff_range = INLET_SHAPE_COEFFICIENTS[inlet_shape]
            effective_xi = xi_user if xi_user is not None else xi_calc
            if effective_xi is not None and not (coeff_range[0] - 0.01 <= effective_xi <= coeff_range[1] + 0.01):
                xi_user = sum(coeff_range) / 2
                xi_calc = None
        # 旧数据迁移：通用构件不需要长度，自动清零
        length = d.get('length', 0)
        if st in (SegmentType.TRASH_RACK, SegmentType.GATE_SLOT,
                  SegmentType.BYPASS_PIPE, SegmentType.PIPE_TRANSITION,
                  SegmentType.OTHER) and length > 0:
            length = 0.0
        trash_rack_params = None
        if st == SegmentType.TRASH_RACK:
            if 'trash_rack_params' in d:
                try:
                    trash_rack_params = TrashRackParams.from_dict(d['trash_rack_params'])
                    if xi_calc is None and not trash_rack_params.manual_mode:
                        xi_calc = round(CoefficientService.calculate_trash_rack_xi(trash_rack_params), 4)
                except Exception:
                    pass
            # 旧数据迁移：无 trash_rack_params 时补上默认参数和计算值
            if trash_rack_params is None:
                trash_rack_params = TrashRackParams()
            if xi_user is None and xi_calc is None:
                xi_calc = round(CoefficientService.calculate_trash_rack_xi(trash_rack_params), 4)
        # 旧数据迁移：闸门槽无系数时补上默认 0.10
        if st == SegmentType.GATE_SLOT and xi_user is None and xi_calc is None:
            xi_user = 0.10
        return StructureSegment(
            segment_type=st, direction=direction,
            length=length, radius=d.get('radius', 0), angle=d.get('angle', 0),
            xi_user=xi_user, xi_calc=xi_calc,
            locked=d.get('locked', False),
            custom_label=d.get('custom_label', ''),
            inlet_shape=inlet_shape,
            outlet_shape=outlet_shape,
            start_elevation=d.get('start_elev'),
            end_elevation=d.get('end_elev'),
            source_ip_index=d.get('source_ip_index'),
            trash_rack_params=trash_rack_params,
        )

    # ================================================================
    # 拟定流速确认交互（方案D）
    # ================================================================
    def _on_v_edited_by_user(self):
        """用户手动编辑流速输入框时触发（重置确认状态）"""
        if self._syncing:
            return
        self._v_user_confirmed = False
        self._update_v_style()

    def _on_v_confirmed(self):
        """LineEdit editingFinished（Enter/失焦）时触发，视为用户已确认"""
        if self._syncing:
            return
        self._v_user_confirmed = True
        self._update_v_style()

    def _update_v_style(self):
        """根据确认状态动态更新流速输入框样式"""
        if self._v_user_confirmed:
            self.edit_v.setStyleSheet(
                f"LineEdit {{ border: 1.5px solid {S}; background: #F1F8E9; }}"
            )
            cur = self.lbl_v_hint.text()
            if '已从主表导入' in cur or '已导入' in cur:
                self.lbl_v_hint.setText("(已导入, ✓已确认)")
            elif '已参与计算' in cur:
                pass
            else:
                self.lbl_v_hint.setText("(✓已确认)")
            self.lbl_v_hint.setStyleSheet(f"color:{S};font-size:12px;font-weight:bold;")
        else:
            if self.edit_v.text().strip():
                self.edit_v.setStyleSheet(
                    "LineEdit { border: 1.5px solid orange; }"
                )
                self.lbl_v_hint.setText("(请确认流速)")
                self.lbl_v_hint.setStyleSheet("color:#CC6600;font-size:12px;font-weight:bold;")
            else:
                self.edit_v.setStyleSheet(
                    "LineEdit { border: 2px dashed #E65100; background: #FFF8E1; }"
                )
                self.lbl_v_hint.setText("← 请输入拟定流速")
                self.lbl_v_hint.setStyleSheet("color:#E53935;font-size:12px;font-weight:bold;")

    def _flash_v_field(self):
        """流速输入框边框闪烁3次红色警告"""
        # 停止已有的闪烁定时器，防止重复调用时多个定时器并行
        if hasattr(self, '_flash_timer') and self._flash_timer.isActive():
            self._flash_timer.stop()
            self._flash_timer.deleteLater()
        self._flash_count = 0
        self._flash_timer = QTimer(self)
        self._flash_timer.setInterval(250)

        def _do_flash():
            self._flash_count += 1
            if self._flash_count > 6:
                self._flash_timer.stop()
                self._update_v_style()
                return
            if self._flash_count % 2 == 1:
                self.edit_v.setStyleSheet(
                    "LineEdit { border: 3px solid #D50000; background: #FFCDD2; }"
                )
            else:
                self.edit_v.setStyleSheet(
                    "LineEdit { border: 2px dashed #E65100; background: #FFF8E1; }"
                )

        self._flash_timer.timeout.connect(_do_flash)
        self._flash_timer.start()

    def _validate_v_before_calc(self) -> bool:
        """计算前检查拟定流速是否已确认。返回True=通过，False=拦截"""
        if self._v_user_confirmed:
            return True
        # 自动跳转到基本参数Tab
        self.params_notebook.setCurrentIndex(0)
        # 聚焦输入框
        self.edit_v.setFocus()
        self.edit_v.selectAll()
        # 闪烁警告
        self._flash_v_field()
        # InfoBar提示
        InfoBar.error(
            "请先输入拟定流速",
            "“拟定流速 v”是必填参数，请根据工程实际输入流速值后再执行计算。",
            parent=self._info_parent(), duration=6000,
            position=InfoBarPosition.TOP
        )
        return False

    # ================================================================
    # 管道根数确认交互
    # ================================================================
    def _on_num_pipes_value_changed(self, value):
        """SpinBox 值变化时触发（重置确认状态，实时联动D理论值等）"""
        if self._syncing:
            return
        self._num_pipes_user_confirmed = False
        self._update_num_pipes_style()
        self._on_Qv_changed()

    def _on_num_pipes_confirmed(self):
        """[+]/[-] 按钮点击后触发，视为用户已确认"""
        if self._syncing:
            return
        self._num_pipes_user_confirmed = True
        self._update_num_pipes_style()
        self._on_Qv_changed()

    def _apply_np_btn_style(self, border_color, btn_bg, btn_color, btn_hover, val_bg, border_width='2px', border_type='dashed'):
        """统一设置 [-] val [+] 控件的样式"""
        w = self.spin_num_pipes
        btn_base = (f"QPushButton {{ border: {border_width} {border_type} {border_color}; "
                    f"background: {btn_bg}; color: {btn_color}; font-size: 18px; font-weight: bold; }} "
                    f"QPushButton:hover {{ background: {btn_hover}; color: #fff; }}")
        w.btn_minus.setStyleSheet(btn_base + f" QPushButton {{ border-right: none; border-top-left-radius: 5px; border-bottom-left-radius: 5px; }}")
        w.btn_plus.setStyleSheet(btn_base + f" QPushButton {{ border-left: none; border-top-right-radius: 5px; border-bottom-right-radius: 5px; }}")
        w.edit_val.setStyleSheet(f"QLineEdit {{ border-top: {border_width} {border_type} {border_color}; border-bottom: {border_width} {border_type} {border_color}; "
                                f"border-left: none; border-right: none; background: {val_bg}; font-size: 15px; font-weight: 600; }}")

    def _apply_np_unconfirmed_style(self):
        self._apply_np_btn_style('#E65100', '#FFE0B2', '#E65100', '#FFB74D', '#FFF8E1')

    def _apply_np_confirmed_style(self):
        self._apply_np_btn_style(S, '#C8E6C9', S, '#81C784', '#F1F8E9', '1.5px', 'solid')

    def _apply_np_flash_style(self):
        self._apply_np_btn_style('#D50000', '#FFCDD2', '#D50000', '#EF9A9A', '#FFCDD2', '3px', 'solid')

    def _update_num_pipes_style(self):
        """根据确认状态动态更新管道根数控件样式"""
        N = self.spin_num_pipes.value()
        if self._num_pipes_user_confirmed:
            self._apply_np_confirmed_style()
            if N > 1:
                self.lbl_num_pipes_hint.setText(f"(✓已确认，{N}管并联)")
            else:
                self.lbl_num_pipes_hint.setText("(✓已确认，单管)")
            self.lbl_num_pipes_hint.setStyleSheet(f"color:{S};font-size:12px;font-weight:bold;")
        else:
            self._apply_np_unconfirmed_style()
            self.lbl_num_pipes_hint.setText("← 请确认管道数")
            self.lbl_num_pipes_hint.setStyleSheet("color:#E53935;font-size:12px;font-weight:bold;")

    def _flash_num_pipes_field(self):
        """管道根数控件边框闪烁3次红色警告"""
        if hasattr(self, '_flash_np_timer') and self._flash_np_timer.isActive():
            self._flash_np_timer.stop()
            self._flash_np_timer.deleteLater()
        self._flash_np_count = 0
        self._flash_np_timer = QTimer(self)
        self._flash_np_timer.setInterval(250)

        def _do_flash():
            self._flash_np_count += 1
            if self._flash_np_count > 6:
                self._flash_np_timer.stop()
                self._update_num_pipes_style()
                return
            if self._flash_np_count % 2 == 1:
                self._apply_np_flash_style()
            else:
                self._apply_np_unconfirmed_style()

        self._flash_np_timer.timeout.connect(_do_flash)
        self._flash_np_timer.start()

    def _validate_num_pipes_before_calc(self) -> bool:
        """计算前检查管道根数是否已确认。返回True=通过，False=拦截"""
        if self._num_pipes_user_confirmed:
            return True
        self.params_notebook.setCurrentIndex(0)
        self.spin_num_pipes.setFocus()
        self.spin_num_pipes.selectAll()
        self._flash_num_pipes_field()
        InfoBar.error(
            "请先确认管道根数",
            "「管道根数 N」是必填参数，请输入后按 Enter 或点击 [+]/[−] 确认。",
            parent=self._info_parent(), duration=6000,
            position=InfoBarPosition.TOP
        )
        return False

    # ================================================================
    # 平面转弯半径倍数确认交互（方案B：温和提醒）
    # ================================================================
    def _on_turn_n_edited_by_user(self):
        """用户手动编辑平面转弯半径倍数时触发"""
        if self._syncing:
            return
        self._turn_n_user_confirmed = True
        self._update_turn_n_style()

    def _update_turn_n_style(self):
        """根据确认状态动态更新平面转弯半径倍数输入框样式"""
        if self._turn_n_user_confirmed:
            self.edit_turn_n.setStyleSheet(
                f"LineEdit {{ border: 1.5px solid {S}; background: #F1F8E9; }}"
            )
        else:
            self.edit_turn_n.setStyleSheet(
                "LineEdit { border: 1.5px dashed #1565C0; background: #E3F2FD; }"
            )

    def _warn_turn_n_if_needed(self):
        """计算时若平面转弯半径倍数未手动确认，弹出黄色警告（不拦截）"""
        if self._turn_n_user_confirmed or self._suppress_result_display:
            return
        n_val = self._fval(self.edit_turn_n, 3.0)
        InfoBar.warning(
            "请确认平面转弯半径倍数",
            f"平面转弯半径倍数当前为默认值 {n_val:.1f}，请确认是否符合工程实际。",
            parent=self._info_parent(), duration=5000,
            position=InfoBarPosition.TOP
        )

    # ================================================================
    # 参数联动
    # ================================================================
    def _on_Qv_changed(self):
        """Q或v参数改变（200ms防抖，与原版一致）"""
        if not hasattr(self, '_qv_timer'):
            self._qv_timer = QTimer(self)
            self._qv_timer.setSingleShot(True)
            self._qv_timer.timeout.connect(self._do_Qv_update)
        self._qv_timer.start(200)

    def _do_Qv_update(self):
        """Q/v变化后的实际更新逻辑"""
        self._update_turn_R()
        self._update_D_theory()
        self._update_plan_bend_radius()
        self._update_segment_coefficients()
        # Q变化后，断面参数策略下联动更新v₂
        strategy_text = self.combo_v2_strategy.currentText()
        if "断面" in strategy_text and self._section_B is not None:
            self._recalc_section_v2()
        # Q/v变化后，自动更新出水口局部阻力系数
        self._auto_compute_outlet_xi()

    def _update_D_theory(self):
        Q = self._fval(self.edit_Q, 0)
        v = self._fval(self.edit_v, 0)
        N = max(1, self.spin_num_pipes.value()) if hasattr(self, 'spin_num_pipes') else 1
        if Q > 0 and v > 0:
            Q_single = Q / N
            D = math.sqrt(4 * Q_single / (math.pi * v))
            if SIPHON_AVAILABLE:
                D_design = HydraulicCore.round_diameter(D)
                if N > 1:
                    self.lbl_D_theory.setText(
                        f"D设计 = {D_design:.4f} m（{N}管并联，每管 Q = {Q_single:.3f} m³/s）")
                else:
                    self.lbl_D_theory.setText(f"D设计 = {D_design:.4f} m（D理论 = {D:.4f} m）")
            else:
                self.lbl_D_theory.setText(f"D = {D:.4f} m")
        else:
            self.lbl_D_theory.setText("D = --")

    def _on_inc_toggle(self, _state):
        """考虑加大流量 CheckBox 切换"""
        enabled = self.inc_cb.isChecked()
        self.edit_inc.setVisible(enabled)
        self.lbl_inc_hint.setVisible(enabled)

    def _on_D_override_toggled(self, state):
        """指定管径 CheckBox 切换"""
        checked = bool(state)
        self.edit_D_override.setVisible(checked)
        if not checked:
            self.edit_D_override.clear()

    def _update_segment_coefficients(self):
        """Q/v变化后更新结构段系数（弯管/折管）"""
        if not SIPHON_AVAILABLE:
            return
        Q = self._fval(self.edit_Q, 0)
        v = self._fval(self.edit_v, 0)
        if Q <= 0 or v <= 0:
            return
        N = max(1, self.spin_num_pipes.value()) if hasattr(self, 'spin_num_pipes') else 1
        Q_single = Q / N
        D = math.sqrt(4 * Q_single / (math.pi * v))
        if D <= 0:
            return
        updated = False
        for seg in self.segments:
            if seg.xi_user is not None:
                continue
            if seg.segment_type == SegmentType.BEND and seg.radius > 0 and seg.angle > 0:
                seg.xi_calc = CoefficientService.calculate_bend_coeff(
                    seg.radius, D, seg.angle, verbose=False)
                updated = True
            elif seg.segment_type == SegmentType.FOLD and seg.angle > 0:
                seg.xi_calc = CoefficientService.calculate_fold_coeff(
                    seg.angle, verbose=False)
                updated = True
        for seg in self.plan_segments:
            if seg.xi_user is not None:
                continue
            if seg.segment_type == SegmentType.BEND and seg.radius > 0 and seg.angle > 0:
                seg.xi_calc = CoefficientService.calculate_bend_coeff(
                    seg.radius, D, seg.angle, verbose=False)
                updated = True
        if updated:
            self._refresh_seg_table()

    def _recalc_section_v2(self):
        """断面参数策略下，Q变化后重新计算v₂（支持梯形+圆形断面，与原版一致）"""
        if self._section_B is None or self._section_h is None:
            return
        Q = self._fval(self.edit_Q, 0)
        N = max(1, self.spin_num_pipes.value()) if hasattr(self, 'spin_num_pipes') else 1
        Q_single = Q / N
        B = self._section_B
        h = self._section_h
        m = self._section_m if self._section_m is not None else 0.0
        if Q_single <= 0 or B <= 0 or h <= 0:
            return

        # 圆形断面判断（与原版_calculate_trapezoidal_velocity一致）
        # m=0 且 h < B*0.95 时，B可能是直径D，尝试用圆形公式
        area = 0.0
        if m == 0 and h < B and h < B * 0.95:
            r = B / 2.0  # B视为直径D
            cos_arg = max(-1.0, min(1.0, (r - h) / r))
            theta = 2 * math.acos(cos_arg)
            area_circ = r * r * (theta - math.sin(theta)) / 2
            area_rect = B * h
            # 用较小的面积（圆形面积 < 矩形面积 当 h < D）
            if area_circ > 0 and area_circ < area_rect * 0.95:
                area = area_circ

        # 默认：梯形/矩形公式
        if area <= 0:
            area = (B + m * h) * h
        if area <= 0:
            return

        v2 = Q_single / area
        was_ro = self.edit_v2.isReadOnly()
        self.edit_v2.setReadOnly(False)
        self.edit_v2.setText(f"{v2:.4f}")
        self.edit_v2.setReadOnly(was_ro)

    def _on_turn_n_changed(self):
        # 同步内部变量（与原版_on_turn_radius_n_changed一致）
        n = self._fval(self.edit_turn_n, 0)
        if n > 0:
            self._siphon_turn_radius_n = n
        self._update_turn_R()
        # n变化后更新平面弯管段半径 R = n × D_design
        self._update_plan_bend_radius()

    def _update_turn_R(self, confirmed=False):
        Q = self._fval(self.edit_Q, 0)
        v = self._fval(self.edit_v, 0)
        n_mult = self._fval(self.edit_turn_n, 0)
        if Q <= 0 or v <= 0:
            self.lbl_turn_R.setText("R = n × D（请先输入Q和v）")
            return
        if n_mult <= 0:
            self.lbl_turn_R.setText("R = n × D（请输入n值）")
            return
        N = max(1, self.spin_num_pipes.value()) if hasattr(self, 'spin_num_pipes') else 1
        Q_single = Q / N
        D_theory = math.sqrt(4 * Q_single / (math.pi * v))
        D_design = HydraulicCore.round_diameter(D_theory)
        R = round(n_mult * D_design, 2)
        # 同步 R 输入框（n 为权威值，以n为准）
        if hasattr(self, 'edit_turn_R') and not self._syncing:
            self._syncing = True
            self.edit_turn_R.setText(f"{R:.2f}")
            self.edit_turn_R.setStyleSheet("LineEdit { border: 1px solid #90CAF9; background: #E3F2FD; }")
            self._syncing = False
        if hasattr(self, 'lbl_turn_R_status') and not self._syncing:
            self.lbl_turn_R_status.setText("← 可直接输入覆盖（修改R将反推n）")
            self.lbl_turn_R_status.setStyleSheet("color:#888;font-size:12px;")
        has_bends = any(seg.segment_type == SegmentType.BEND for seg in self.plan_segments)
        sync_hint = "，已同步至弯管段" if has_bends else ""
        if confirmed:
            self.lbl_turn_R.setText(f"D设计={D_design:.2f}m → R={R:.2f}m ✓已参与计算")
            self.lbl_turn_R.setStyleSheet(f"color:#008800;font-size:12px;")
        elif self._turn_n_user_confirmed:
            self.lbl_turn_R.setText(f"D设计={D_design:.2f}m → R={n_mult}×{D_design:.2f}={R:.2f}m{sync_hint} ✓已确认")
            self.lbl_turn_R.setStyleSheet(f"color:{S};font-size:12px;")
        else:
            self.lbl_turn_R.setText(f"D设计={D_design:.2f}m → R={n_mult}×{D_design:.2f}={R:.2f}m{sync_hint}  (请确认倍数)")
            self.lbl_turn_R.setStyleSheet("color:#1565C0;font-size:12px;")

    def _on_turn_R_changed(self):
        """用户直接修改 R 値时，反推 n = R / D设计"""
        if self._syncing:
            return
        R_val = self._fval(self.edit_turn_R, 0)
        if R_val <= 0:
            return
        Q = self._fval(self.edit_Q, 0)
        v = self._fval(self.edit_v, 0)
        if Q <= 0 or v <= 0 or not SIPHON_AVAILABLE:
            return
        N = max(1, self.spin_num_pipes.value()) if hasattr(self, 'spin_num_pipes') else 1
        D_theory = math.sqrt(4 * (Q / N) / (math.pi * v))
        D_design = HydraulicCore.round_diameter(D_theory)
        if D_design <= 0:
            return
        n_new = round(R_val / D_design, 3)
        self._syncing = True
        self.edit_turn_n.setText(str(n_new))
        self._syncing = False
        self._siphon_turn_radius_n = n_new
        self._turn_n_user_confirmed = True
        self._update_turn_n_style()
        self.lbl_turn_R.setText(
            f"D设计={D_design:.2f}m → R={n_new}×{D_design:.2f}={R_val:.2f}m ✓（R反推n）"
        )
        self.lbl_turn_R.setStyleSheet(f"color:{S};font-size:12px;")
        self.lbl_turn_R_status.setText("← 已反推，n已更新为主值")
        self.lbl_turn_R_status.setStyleSheet(f"color:{S};font-size:12px;")
        self.edit_turn_R.setStyleSheet(
            f"LineEdit {{ border: 1.5px solid {S}; background: #F1F8E9; }}"
        )
        self._update_plan_bend_radius()

    def _on_inlet_type_changed(self, text):
        """渐变段型式→进口系数自动联动"""
        if SIPHON_AVAILABLE:
            gt = GRADIENT_TYPE_MAP.get(text, GradientType.NONE)
            is_twist = (gt == GradientType.LINEAR_TWIST)
            self._twist_in_row.setVisible(is_twist)
            if is_twist:
                self._update_inlet_twist_xi()
            else:
                xi = CoefficientService.get_gradient_coeff(gt, True)
                self.edit_xi_inlet.setText(f"{xi:.4f}")

    def _on_outlet_type_changed(self, text):
        """渐变段型式→出口系数自动联动"""
        if SIPHON_AVAILABLE:
            gt = GRADIENT_TYPE_MAP.get(text, GradientType.NONE)
            is_twist = (gt == GradientType.LINEAR_TWIST)
            self._twist_out_row.setVisible(is_twist)
            if is_twist:
                self._update_outlet_twist_xi()
            else:
                xi = CoefficientService.get_gradient_coeff(gt, False)
                self.edit_xi_outlet.setText(f"{xi:.4f}")

    def _update_inlet_twist_xi(self):
        """直线扭曲面进口：有角度则插值，无角度则取均值"""
        if not SIPHON_AVAILABLE:
            return
        txt = self.edit_twist_angle_inlet.text().strip()
        if txt:
            try:
                xi = CoefficientService.calculate_linear_twist_coeff(float(txt), True)
                self.edit_xi_inlet.setText(f"{xi:.4f}")
            except ValueError:
                pass
        else:
            xi = CoefficientService.get_gradient_coeff(GradientType.LINEAR_TWIST, True)
            self.edit_xi_inlet.setText(f"{xi:.4f}")

    def _update_outlet_twist_xi(self):
        """直线扭曲面出口：有角度则插值，无角度则取均值"""
        if not SIPHON_AVAILABLE:
            return
        txt = self.edit_twist_angle_outlet.text().strip()
        if txt:
            try:
                xi = CoefficientService.calculate_linear_twist_coeff(float(txt), False)
                self.edit_xi_outlet.setText(f"{xi:.4f}")
            except ValueError:
                pass
        else:
            xi = CoefficientService.get_gradient_coeff(GradientType.LINEAR_TWIST, False)
            self.edit_xi_outlet.setText(f"{xi:.4f}")

    def _on_v_channel_in_changed(self):
        """进口始端流速v₁改变 —— 仅在 V1_PLUS_02 策略下联动更新 v₂"""
        strategy_text = self.combo_v2_strategy.currentText()
        if "v₁ + 0.2" in strategy_text:
            v1 = self._fval(self.edit_v1, 0)
            if v1 > 0:
                was_ro = self.edit_v2.isReadOnly()
                self.edit_v2.setReadOnly(False)
                self.edit_v2.setText(f"{v1 + 0.2:.4f}")
                self.edit_v2.setReadOnly(was_ro)

    def _on_v2_strategy_changed(self, text):
        """v₂策略变更 → 联动UI"""
        # 先重置可能被覆盖的 mousePressEvent 和光标
        self.lbl_v2_hint.setCursor(Qt.ArrowCursor)
        self.lbl_v2_hint.mousePressEvent = lambda e: None

        if "自动" in text:
            self.edit_v2.setReadOnly(True)
            self.edit_v2.clear()
            self.edit_v2.setPlaceholderText("留空=管道流速")
            self.lbl_v2_hint.setText("(计算后自动填充: 管道流速)")
            self.lbl_v2_hint.setStyleSheet("color:#424242;font-size:12px;")
            self.lbl_v2_strategy_hint.setText("(推荐)")
            self.lbl_v2_strategy_hint.setStyleSheet("color:#00AA00;font-size:12px;")
        elif "v₁ + 0.2" in text:
            self.edit_v2.setReadOnly(True)
            v1 = self._fval(self.edit_v1, 0)
            self.edit_v2.setText(f"{v1 + 0.2:.4f}" if v1 > 0 else "")
            self.lbl_v2_hint.setText("(自动计算: v₁ + 0.2)")
            self.lbl_v2_hint.setStyleSheet("color:#FF6600;font-size:12px;")
            self.lbl_v2_strategy_hint.setText("")
        elif "断面" in text:
            self.edit_v2.setReadOnly(True)
            # 若已有断面参数则自动计算v₂（与原版_apply_v2_strategy_ui一致）
            if self._section_B is not None and self._section_h is not None:
                self._recalc_section_v2()
            else:
                self.edit_v2.clear()
            self.lbl_v2_hint.setText("(双击设置断面参数)")
            self.lbl_v2_hint.setStyleSheet("color:#FF6600;font-size:12px;")
            self.lbl_v2_hint.setCursor(Qt.PointingHandCursor)
            self.lbl_v2_hint.mousePressEvent = self._open_v2_section_dialog
            self.lbl_v2_strategy_hint.setText("")
        elif "指定" in text:
            self.edit_v2.setReadOnly(False)
            self.edit_v2.setPlaceholderText("请输入")
            if not self.edit_v2.text().strip():
                self.edit_v2.setText("1.2")
            self.lbl_v2_hint.setText("(需输入)")
            self.lbl_v2_hint.setStyleSheet("color:#424242;font-size:12px;")
            self.lbl_v2_strategy_hint.setText("")

    def _open_v2_section_dialog(self, event=None):
        """双击v₂提示 → 打开断面参数设置对话框"""
        if not DIALOGS_AVAILABLE or not SIPHON_AVAILABLE:
            return
        Q = self._fval(self.edit_Q, 10)
        dlg = InletSectionDialog(self, Q, self._section_B, self._section_h, self._section_m)
        if dlg.exec() == QDialog.Accepted:
            self._section_B = dlg.result_B
            self._section_h = dlg.result_h
            self._section_m = dlg.result_m
            if dlg.result_velocity is not None:
                was_ro = self.edit_v2.isReadOnly()
                self.edit_v2.setReadOnly(False)
                self.edit_v2.setText(f"{dlg.result_velocity:.4f}")
                self.edit_v2.setReadOnly(was_ro)
                self.lbl_v2_hint.setText(f"(断面: B={self._section_B}, h={self._section_h}, m={self._section_m})")
            else:
                self.edit_v2.clear()
                self.lbl_v2_hint.setText("(双击设置断面参数)")

    def _validate_inlet_velocity(self):
        """验证进口流速（仅对非自动策略有意义）"""
        strategy_text = self.combo_v2_strategy.currentText()
        if "自动" in strategy_text:
            return  # 自动模式无需校验，计算核心会保证 v₂ > v₁
        v1 = self._fval(self.edit_v1, 0)
        v2 = self._fval(self.edit_v2, 0)
        if v2 > 0 and v2 <= v1:
            InfoBar.warning("警告",
                "进口末端流速v₂应大于始端流速v₁\n建议切换为\"自动（=管道流速）\"策略",
                parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    # ================================================================
    # 结构段管理
    # ================================================================
    def _init_default_segments(self):
        if not SIPHON_AVAILABLE:
            return
        self._longitudinal_is_example = True
        # 同步清空纵断面节点（防止残留旧数据）
        self.longitudinal_nodes.clear()
        if hasattr(self, 'long_table'):
            self.long_table.setRowCount(0)
        # 进水口默认使用"进口稍微修圆"，系数取中值
        inlet_shape = InletOutletShape.SLIGHTLY_ROUNDED
        inlet_xi = sum(INLET_SHAPE_COEFFICIENTS[inlet_shape]) / 2  # 取范围中值
        # 出水口默认系数设为0（待用户设置渠道参数后计算）
        self.segments = [
            # 通用构件（仅贡献局部阻力系数ξ）
            StructureSegment(segment_type=SegmentType.INLET, locked=True,
                             inlet_shape=inlet_shape, xi_calc=inlet_xi,
                             direction=SegmentDirection.COMMON),
            StructureSegment(segment_type=SegmentType.TRASH_RACK,
                             trash_rack_params=TrashRackParams(),
                             xi_calc=CoefficientService.calculate_trash_rack_xi(TrashRackParams()),
                             direction=SegmentDirection.COMMON),
            StructureSegment(segment_type=SegmentType.GATE_SLOT, xi_user=0.10,
                             direction=SegmentDirection.COMMON),
            StructureSegment(segment_type=SegmentType.BYPASS_PIPE, xi_user=0.1,
                             direction=SegmentDirection.COMMON),
            # 纵断面管身段（示例数据）
            StructureSegment(segment_type=SegmentType.FOLD, length=5.0),
            StructureSegment(segment_type=SegmentType.STRAIGHT, length=50.0),
            StructureSegment(segment_type=SegmentType.BEND, length=10.0, radius=5.0, angle=45.0),
            StructureSegment(segment_type=SegmentType.STRAIGHT, length=100.0),
            StructureSegment(segment_type=SegmentType.BEND, length=10.0, radius=5.0, angle=45.0),
            StructureSegment(segment_type=SegmentType.STRAIGHT, length=50.0),
            # 通用构件
            StructureSegment(segment_type=SegmentType.OTHER, xi_user=0.1,
                             direction=SegmentDirection.COMMON),
            StructureSegment(segment_type=SegmentType.OUTLET, locked=True, xi_calc=0.0,
                             direction=SegmentDirection.COMMON),
        ]
        self._refresh_seg_table()
        self._update_canvas()

    def _get_all_display_segments(self):
        """获取用于表格显示的所有段列表（通用构件 + 平面段 + 纵断面段）"""
        display = []
        # 1. 通用构件排在最前面
        for seg in self.segments:
            if seg.direction == SegmentDirection.COMMON:
                display.append((seg, 'common'))
        # 2. 平面段排在中间
        for seg in self.plan_segments:
            display.append((seg, 'plan'))
        # 3. 纵断面管身段排在最后
        for seg in self.segments:
            if seg.direction != SegmentDirection.COMMON:
                display.append((seg, 'longitudinal'))
        return display

    def _refresh_seg_table(self):
        """刷新结构段表格"""
        self.seg_table.setRowCount(0)
        # 颜色
        c_common = QColor(255, 248, 225)   # 浅黄色背景 - 通用构件
        c_plan = QColor(232, 240, 254)     # 浅蓝色背景 - 平面段
        c_long = QColor(232, 245, 233)     # 浅绿色背景 - 纵断面段
        c_example = QColor(240, 244, 240)  # 灰绿色背景 - 纵断面示例数据

        display_segments = self._get_all_display_segments()

        # 表格为空时显示引导提示
        if not display_segments:
            self.seg_table.insertRow(0)
            hint1 = QTableWidgetItem("暂无结构段数据")
            hint1.setForeground(QBrush(QColor(153, 153, 153)))
            self.seg_table.setItem(0, 2, hint1)
            self.seg_table.insertRow(1)
            hint2 = QTableWidgetItem("点击\"导入DXF\"从CAD导入，或手动\"添加管身段\"")
            hint2.setForeground(QBrush(QColor(153, 153, 153)))
            self.seg_table.setItem(1, 2, hint2)

        for i, (seg, source) in enumerate(display_segments):
            row = self.seg_table.rowCount()
            self.seg_table.insertRow(row)

            # 分类与颜色
            is_example_row = (source == 'longitudinal' and self._longitudinal_is_example)
            if source == 'common':
                cat = "通用"
                bg = c_common
            elif source == 'plan':
                cat = "平面"
                bg = c_plan
            else:
                cat = "纵断面(示例)" if is_example_row else "纵断面"
                bg = c_example if is_example_row else c_long

            dir_str = seg.direction.value if hasattr(seg.direction, 'value') else str(seg.direction)

            # 类型显示（进水口附加形状名）
            type_display = seg.segment_type.value
            if seg.segment_type == SegmentType.INLET and seg.inlet_shape:
                type_display = f"进水口({seg.inlet_shape.value})"
            elif seg.segment_type == SegmentType.OUTLET:
                type_display = "出水口"
            elif seg.segment_type == SegmentType.TRASH_RACK:
                type_display = "拦污栅(已配置)" if getattr(seg, 'trash_rack_params', None) else "拦污栅(未配置)"
            elif seg.segment_type == SegmentType.PIPE_TRANSITION:
                lbl = getattr(seg, 'custom_label', '')
                type_display = f"管道渐变段({'扩散' if lbl == '扩散' else '收缩'})"
            elif seg.segment_type == SegmentType.OTHER and getattr(seg, 'custom_label', ''):
                type_display = seg.custom_label

            # 局部系数显示
            xi = seg.xi_user if seg.xi_user is not None else (seg.xi_calc if seg.xi_calc is not None else None)
            xi_str = f"{xi:.4f}" if xi is not None else ""

            # 高程显示（仅纵断面直管段/折管段有意义）
            start_elev = f"{seg.start_elevation:.3f}" if getattr(seg, 'start_elevation', None) is not None else ""
            end_elev = f"{seg.end_elevation:.3f}" if getattr(seg, 'end_elevation', None) is not None else ""

            # 空间长度显示
            sp_len = getattr(seg, 'spatial_length', 0)
            spatial_display = f"{sp_len:.2f}" if sp_len and sp_len > 0 else ""

            vals = [str(i + 1), cat, type_display, dir_str,
                    f"{seg.length:.3f}" if seg.length > 0 else "",
                    f"{seg.radius:.3f}" if seg.radius > 0 else "",
                    f"{seg.angle:.3f}" if seg.angle > 0 else "",
                    start_elev, end_elev, spatial_display, xi_str,
                    "是" if seg.locked else "否"]

            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setBackground(QBrush(bg))
                if is_example_row:
                    item.setForeground(QBrush(QColor(153, 153, 153)))
                self.seg_table.setItem(row, col, item)

        # 更新状态（包含平面段计数）
        n_common = sum(1 for s in self.segments if SIPHON_AVAILABLE and is_common_type(s.segment_type))
        n_long = sum(1 for s in self.segments if s.direction != SegmentDirection.COMMON
                     and (not SIPHON_AVAILABLE or not is_common_type(s.segment_type)))
        n_plan = len(self.plan_segments)
        total = n_common + n_long + n_plan
        parts = []
        if n_common > 0: parts.append(f"通用:{n_common}")
        if n_plan > 0: parts.append(f"平面:{n_plan}")
        if n_long > 0: parts.append(f"纵断面:{n_long}")
        self.lbl_seg_status.setText(f"{total} ({', '.join(parts)})" if parts else "0")

        auto_resize_table(self.seg_table)
        self._update_data_status()
        # 自动切换画布视图（与原版_auto_select_canvas_view一致）
        if self.canvas:
            self.canvas.auto_select_view()

    def _add_segment_dialog(self):
        """通过对话框添加结构段"""
        if not SIPHON_AVAILABLE:
            return
        if DIALOGS_AVAILABLE:
            Q = self._fval(self.edit_Q, 10)
            v = self._fval(self.edit_v, 2)
            dlg = SegmentEditDialog(self, segment=None, Q=Q, v=v)
            if dlg.exec() == QDialog.Accepted and dlg.result:
                # 首次手动添加时，清除示例纵断面数据，仅保留通用构件
                if self._longitudinal_is_example:
                    self.segments = [s for s in self.segments if s.direction == SegmentDirection.COMMON]
                    self._longitudinal_is_example = False
                # 插入到出水口之前
                insert_idx = len(self.segments)
                for i, s in enumerate(self.segments):
                    if s.segment_type == SegmentType.OUTLET:
                        insert_idx = i
                        break
                self.segments.insert(insert_idx, dlg.result)
                self._refresh_seg_table()
                self._update_canvas()
                # 手动添加的管身段无高程数据，不做反向同步（避免污染节点表）
        else:
            InfoBar.warning("不可用", "对话框组件未加载", parent=self._info_parent(),
                           duration=3000, position=InfoBarPosition.TOP)

    def _del_segment(self):
        rows = sorted(set(idx.row() for idx in self.seg_table.selectedIndexes()), reverse=True)
        if not rows:
            InfoBar.warning("提示", "请先选择要删除的行", parent=self._info_parent(),
                           duration=2000, position=InfoBarPosition.TOP)
            return
        # 解析选中行对应的实际段
        display_segments = self._get_all_display_segments()
        for r in rows:
            if r < 0 or r >= len(display_segments):
                continue
            seg, source = display_segments[r]
            # 平面段不允许删除
            if source == 'plan':
                InfoBar.warning("提示", "平面段由推求水面线表格自动提取，不可手动删除",
                               parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return
            # 进出水口不允许删除
            if seg.segment_type in (SegmentType.INLET, SegmentType.OUTLET):
                InfoBar.warning("提示", "不能删除进水口或出水口",
                               parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return
        # 确认删除
        if not fluent_question(self, "确认", "确定要删除选中的结构段吗？"):
            return
        # 收集要删除的实际索引（在self.segments中）
        to_remove = []
        has_long_del = False
        for r in rows:
            if r < 0 or r >= len(display_segments):
                continue
            seg, source = display_segments[r]
            if source in ('common', 'longitudinal'):
                idx = self.segments.index(seg)
                to_remove.append(idx)
                if source == 'longitudinal':
                    has_long_del = True
        for idx in sorted(to_remove, reverse=True):
            if 0 <= idx < len(self.segments):
                self.segments.pop(idx)
        self._refresh_seg_table()
        self._update_canvas()
        if has_long_del and len(self.longitudinal_nodes) >= 2:
            self._sync_segments_to_nodes()

    def _add_common_segment_dialog(self):
        """通过专用对话框添加通用构件"""
        if not SIPHON_AVAILABLE or not DIALOGS_AVAILABLE:
            return
        dlg = CommonSegmentAddDialog(self)
        if dlg.exec() == QDialog.Accepted and dlg.result:
            label, xi = dlg.result
            seg = StructureSegment(
                segment_type=SegmentType.OTHER,
                direction=SegmentDirection.COMMON,
                custom_label=label, xi_user=xi)
            # 插入到出水口之前
            insert_idx = len(self.segments)
            for i, s in enumerate(self.segments):
                if s.segment_type == SegmentType.OUTLET:
                    insert_idx = i
                    break
            self.segments.insert(insert_idx, seg)
            self._refresh_seg_table()
            self._update_canvas()

    def _add_pipe_transition(self):
        """快速插入管道渐变段（默认收缩 ξjb=0.05），双击可修改"""
        if not SIPHON_AVAILABLE:
            return
        seg = StructureSegment(
            segment_type=SegmentType.PIPE_TRANSITION,
            direction=SegmentDirection.COMMON,
            custom_label='收缩',
            xi_user=CoefficientService.PIPE_TRANSITION_CONTRACT)
        insert_idx = len(self.segments)
        for i, s in enumerate(self.segments):
            if s.segment_type == SegmentType.OUTLET:
                insert_idx = i
                break
        self.segments.insert(insert_idx, seg)
        self._refresh_seg_table()
        self._update_canvas()

    def _move_seg_up(self):
        """上移选中的结构段（保护进出水口和平面段）"""
        rows = sorted(set(idx.row() for idx in self.seg_table.selectedIndexes()))
        if not rows or rows[0] <= 0:
            return
        display_segments = self._get_all_display_segments()
        for r in rows:
            if r < 0 or r >= len(display_segments):
                return
            seg, source = display_segments[r]
            # 平面段和进出水口不允许移动
            if source == 'plan' or seg.segment_type in (SegmentType.INLET, SegmentType.OUTLET):
                return
            # 目标位置也要检查
            tgt_seg, tgt_source = display_segments[r - 1]
            if tgt_source != source or tgt_seg.segment_type in (SegmentType.INLET, SegmentType.OUTLET):
                return
        # 在self.segments中执行交换
        for r in rows:
            seg, source = display_segments[r]
            tgt_seg, _ = display_segments[r - 1]
            try:
                i1 = self.segments.index(seg)
                i2 = self.segments.index(tgt_seg)
                self.segments[i1], self.segments[i2] = self.segments[i2], self.segments[i1]
            except ValueError:
                return
        self._refresh_seg_table()
        for r in rows:
            self.seg_table.selectRow(r - 1)
        if (len(self.longitudinal_nodes) >= 2
                and any(display_segments[r][1] == 'longitudinal' for r in rows
                        if 0 <= r < len(display_segments))):
            self._sync_segments_to_nodes()

    def _move_seg_down(self):
        """下移选中的结构段（保护进出水口和平面段）"""
        display_segments = self._get_all_display_segments()
        rows = sorted(set(idx.row() for idx in self.seg_table.selectedIndexes()), reverse=True)
        if not rows or rows[0] >= len(display_segments) - 1:
            return
        for r in rows:
            if r < 0 or r >= len(display_segments):
                return
            seg, source = display_segments[r]
            if source == 'plan' or seg.segment_type in (SegmentType.INLET, SegmentType.OUTLET):
                return
            tgt_seg, tgt_source = display_segments[r + 1]
            if tgt_source != source or tgt_seg.segment_type in (SegmentType.INLET, SegmentType.OUTLET):
                return
        for r in rows:
            seg, source = display_segments[r]
            tgt_seg, _ = display_segments[r + 1]
            try:
                i1 = self.segments.index(seg)
                i2 = self.segments.index(tgt_seg)
                self.segments[i1], self.segments[i2] = self.segments[i2], self.segments[i1]
            except ValueError:
                return
        self._refresh_seg_table()
        for r in rows:
            self.seg_table.selectRow(r + 1)
        if (len(self.longitudinal_nodes) >= 2
                and any(display_segments[r][1] == 'longitudinal' for r in rows
                        if 0 <= r < len(display_segments))):
            self._sync_segments_to_nodes()

    def _clear_longitudinal(self):
        """清空纵断面管身段数据（保留通用构件），同步清空纵断面节点"""
        if not fluent_question(self, "确认", "确定要清空所有纵断面管身段吗？"):
            return
        self.segments = [seg for seg in self.segments if seg.direction == SegmentDirection.COMMON]
        self._longitudinal_is_example = False
        self.longitudinal_nodes.clear()
        self.long_table.setRowCount(0)
        self._refresh_seg_table()
        self._update_canvas()

    def _clear_segments(self):
        self.segments.clear()
        self.longitudinal_nodes.clear()
        self.long_table.setRowCount(0)
        self._longitudinal_is_example = False
        self._refresh_seg_table()
        self._update_canvas()

    def _on_seg_double_click(self, index):
        """双击结构段行 → 打开对应编辑对话框"""
        row = index.row()
        if not DIALOGS_AVAILABLE or not SIPHON_AVAILABLE:
            return

        # 通过三区映射解析实际段
        display_segments = self._get_all_display_segments()
        if row < 0 or row >= len(display_segments):
            return

        seg, source = display_segments[row]

        # 平面段允许编辑
        if source == 'plan':
            Q = self._fval(self.edit_Q, 10)
            v = self._fval(self.edit_v, 2)
            dlg = SegmentEditDialog(self, segment=seg, Q=Q, v=v,
                                    direction=SegmentDirection.PLAN)
            if dlg.exec() == QDialog.Accepted and dlg.result:
                try:
                    real_idx = self.plan_segments.index(seg)
                    self.plan_segments[real_idx] = dlg.result
                    self.plan_segments[real_idx].direction = SegmentDirection.PLAN
                except ValueError:
                    pass
                self._refresh_seg_table()
                self._update_canvas()
            return

        # 锁定行编辑确认（进出水口等锁定构件除外，它们有专用对话框）
        if seg.locked and seg.segment_type not in (SegmentType.INLET, SegmentType.OUTLET):
            if not fluent_question(self, "提示", "该行已锁定，确定要编辑吗？"):
                return

        Q = self._fval(self.edit_Q, 10)
        v = self._fval(self.edit_v, 2)

        if seg.segment_type == SegmentType.INLET:
            dlg = InletShapeDialog(self, seg)
        elif seg.segment_type == SegmentType.OUTLET:
            # 转换为对话框期望的短键名
            dp_short = {
                'type': self._downstream_params.get('outlet_downstream_type', ''),
                'B': self._downstream_params.get('outlet_downstream_B'),
                'h': self._downstream_params.get('outlet_downstream_h'),
                'm': self._downstream_params.get('outlet_downstream_m'),
                'D': self._downstream_params.get('outlet_downstream_D'),
                'R': self._downstream_params.get('outlet_downstream_R'),
            }
            dlg = OutletShapeDialog(self, seg, Q=Q, v=v,
                                     downstream_params=dp_short)
        elif seg.segment_type == SegmentType.TRASH_RACK:
            dlg = TrashRackConfigDialog(self, seg.trash_rack_params)
            if dlg.exec() == QDialog.Accepted and dlg.result:
                try:
                    real_idx = self.segments.index(seg)
                    self.segments[real_idx].trash_rack_params = dlg.result
                    self.segments[real_idx].xi_calc = CoefficientService.calculate_trash_rack_xi(dlg.result)
                    self.segments[real_idx].xi_user = None
                except ValueError:
                    pass
                self._refresh_seg_table()
                self._update_canvas()
            return
        elif seg.segment_type in (SegmentType.GATE_SLOT, SegmentType.BYPASS_PIPE,
                                     SegmentType.PIPE_TRANSITION):
            dlg = SimpleCommonEditDialog(self, seg)
        elif seg.segment_type == SegmentType.OTHER and seg.direction == SegmentDirection.COMMON:
            dlg = CommonSegmentEditDialog(self, seg)
        else:
            dlg = SegmentEditDialog(self, segment=seg, Q=Q, v=v)

        if dlg.exec() == QDialog.Accepted and dlg.result:
            # 找到seg在self.segments中的真实索引
            try:
                real_idx = self.segments.index(seg)
                self.segments[real_idx] = dlg.result
            except ValueError:
                pass
            self._refresh_seg_table()
            self._update_canvas()
            # 反向同步：纵断面管身段编辑后更新节点表（仅当有有效节点数据时）
            if seg.direction != SegmentDirection.COMMON and len(self.longitudinal_nodes) >= 2:
                self._sync_segments_to_nodes()

    # ================================================================
    # 纵断面节点管理
    # ================================================================
    def _create_turn_type_combo(self):
        combo = QComboBox()
        combo.addItems(["无", "圆弧", "折线"])
        combo.setFont(QFont("Microsoft YaHei", 10))
        combo.currentTextChanged.connect(self._on_long_table_edited)
        return combo

    def _add_long_node(self, data=None):
        row = self.long_table.rowCount()
        self.long_table.insertRow(row)
        combo = self._create_turn_type_combo()
        self.long_table.setCellWidget(row, 3, combo)
        if data and isinstance(data, (list, tuple)):
            for col, val in enumerate(data):
                if col == 3:
                    idx = combo.findText(str(val))
                    if idx >= 0: combo.setCurrentIndex(idx)
                elif col < self.long_table.columnCount():
                    self.long_table.setItem(row, col, QTableWidgetItem(str(val) if val else ""))

    def _del_long_node(self):
        rows = sorted(set(idx.row() for idx in self.long_table.selectedIndexes()), reverse=True)
        if not rows:
            InfoBar.warning("提示", "请先选择要删除的行", parent=self._info_parent(),
                           duration=2000, position=InfoBarPosition.TOP)
            return
        for r in rows:
            self.long_table.removeRow(r)
        self._sync_nodes_to_segments()

    def _clear_long_nodes(self):
        if self.long_table.rowCount() > 0 or self.longitudinal_nodes:
            if not fluent_question(self, "确认", "确定要清空所有纵断面节点数据吗？"):
                return
        self.long_table.setRowCount(0)
        self.longitudinal_nodes.clear()
        self.segments = [seg for seg in self.segments if seg.direction == SegmentDirection.COMMON]
        self._longitudinal_is_example = False
        self._refresh_seg_table()
        self._update_canvas()

    def _refresh_long_table(self):
        """从 self.longitudinal_nodes 刷新表格"""
        old_syncing = self._syncing
        self._syncing = True
        try:
            self.long_table.setRowCount(0)
            for nd in self.longitudinal_nodes:
                tt_str = "无"
                if nd.turn_type == TurnType.ARC: tt_str = "圆弧"
                elif nd.turn_type == TurnType.FOLD: tt_str = "折线"
                data = [
                    f"{nd.chainage:.3f}",
                    f"{nd.elevation:.3f}",
                    f"{nd.vertical_curve_radius:.3f}" if nd.vertical_curve_radius > 0 else "",
                    tt_str,
                    f"{nd.turn_angle:.3f}" if nd.turn_angle > 0 else "",
                ]
                self._add_long_node(data)
            auto_resize_table(self.long_table)
        finally:
            self._syncing = old_syncing

    def _import_dxf(self):
        if not DXF_AVAILABLE:
            InfoBar.warning("不可用", "DXF解析器未加载", parent=self._info_parent(),
                           duration=3000, position=InfoBarPosition.TOP)
            return
        # 默认打开示例DXF所在的resources目录
        _res_dir = os.path.join(_siphon_dir, "resources")
        if not os.path.isdir(_res_dir):
            _res_dir = ""
        filepath, _ = QFileDialog.getOpenFileName(self, "选择纵断面DXF文件",
            _res_dir, "DXF文件 (*.dxf);;所有文件 (*.*)")
        if not filepath:
            return
        try:
            # 计算桩号偏移量：使多段线起点X对齐到进口MC桩号（有平面数据）或归零（无平面数据）
            chainage_offset = 0.0
            try:
                import ezdxf as _ezdxf
                _doc = _ezdxf.readfile(filepath)
                _msp = _doc.modelspace()
                _polys = list(_msp.query('LWPOLYLINE'))
                if not _polys:
                    _polys = list(_msp.query('POLYLINE'))
                if _polys:
                    _first_point = list(_polys[0].get_points(format='xyseb'))[0]
                    x_start = _first_point[0]
                    if self.plan_feature_points:
                        mc_inlet = self.plan_feature_points[0].chainage
                        chainage_offset = mc_inlet - x_start
                    else:
                        chainage_offset = -x_start
            except Exception:
                pass

            # 解析纵断面为变坡点节点表
            long_nodes, message = DxfParser.parse_longitudinal_profile(
                filepath, chainage_offset=chainage_offset)

            if not long_nodes:
                InfoBar.error("导入失败", message or "DXF文件中未找到纵断面数据",
                             parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
                return

            self._clear_long_nodes()
            self.longitudinal_nodes = long_nodes
            self._longitudinal_is_example = False
            self._refresh_long_table()

            # 同时用传统方式解析生成 StructureSegment（用于表格显示）
            try:
                segments, h_bottom, seg_msg = DxfParser.parse_dxf(filepath)
                if segments:
                    # 保留已有的通用构件，只替换管身段
                    old_common = [s for s in self.segments
                                  if s.direction == SegmentDirection.COMMON
                                  and s.segment_type not in (SegmentType.INLET, SegmentType.OUTLET)]
                    dxf_inlet = None
                    dxf_outlet = None
                    dxf_pipe = []
                    for s in segments:
                        if s.segment_type == SegmentType.INLET:
                            dxf_inlet = s
                        elif s.segment_type == SegmentType.OUTLET:
                            dxf_outlet = s
                        else:
                            dxf_pipe.append(s)
                    # 保留已有进水口的形状和系数设置
                    existing_inlet = next((s for s in self.segments
                                           if s.segment_type == SegmentType.INLET), None)
                    inlet_seg = existing_inlet if existing_inlet else dxf_inlet
                    existing_outlet = next((s for s in self.segments
                                            if s.segment_type == SegmentType.OUTLET), None)
                    outlet_seg = existing_outlet if existing_outlet else dxf_outlet
                    # 重新组装
                    new_segments = []
                    if inlet_seg:
                        new_segments.append(inlet_seg)
                    new_segments.extend(old_common)
                    if outlet_seg:
                        new_segments.append(outlet_seg)
                    new_segments.extend(dxf_pipe)
                    self.segments = new_segments
                    self._refresh_seg_table()
            except Exception:
                pass

            # DXF导入成功后，强制切换到纵断面视图
            if self.canvas:
                self.canvas.set_view_mode("profile")
                self._update_zoom_label()
            self._update_segment_coefficients()
            self._update_canvas()
            self._update_data_status()

            # 详细提示信息
            node_info = f"变坡点节点: {len(long_nodes)} 个"
            turns = sum(1 for nd in long_nodes if nd.turn_type != TurnType.NONE)
            if turns > 0:
                node_info += f"（其中转弯 {turns} 个）"
            spatial_info = ""
            if self.plan_feature_points:
                spatial_info = "\n已检测到平面数据，将使用三维空间合并计算"
            else:
                spatial_info = "\n未检测到平面数据，将使用纵断面独立计算模式"
            InfoBar.success("导入成功",
                f"{message}\n{node_info}{spatial_info}",
                parent=self._info_parent(), duration=6000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("导入失败", str(e), parent=self._info_parent(),
                         duration=5000, position=InfoBarPosition.TOP)

    def _build_longitudinal_nodes(self):
        """从纵断面节点表构建LongitudinalNode列表（并补全坡角信息）"""
        if not SIPHON_AVAILABLE:
            return []
        nodes = []
        for r in range(self.long_table.rowCount()):
            chainage = self._cell_float(self.long_table, r, 0)
            elevation = self._cell_float(self.long_table, r, 1)
            vcr = self._cell_float(self.long_table, r, 2)
            combo = self.long_table.cellWidget(r, 3)
            tt_str = combo.currentText() if combo else "无"
            angle = self._cell_float(self.long_table, r, 4)
            tt = TurnType.NONE
            if tt_str == "圆弧": tt = TurnType.ARC
            elif tt_str == "折线": tt = TurnType.FOLD
            nodes.append(LongitudinalNode(
                chainage=chainage, elevation=elevation,
                vertical_curve_radius=vcr, turn_type=tt, turn_angle=angle))
        # 补全坡角：从相邻节点的高程差推算 slope_before / slope_after
        n = len(nodes)
        for i in range(n):
            if i > 0:
                ds = nodes[i].chainage - nodes[i - 1].chainage
                dz = nodes[i].elevation - nodes[i - 1].elevation
                if abs(ds) > 1e-6:
                    nodes[i].slope_before = math.atan2(dz, abs(ds))
                    nodes[i - 1].slope_after = nodes[i].slope_before
            if i == 0 and n > 1:
                ds = nodes[1].chainage - nodes[0].chainage
                dz = nodes[1].elevation - nodes[0].elevation
                if abs(ds) > 1e-6:
                    nodes[0].slope_after = math.atan2(dz, abs(ds))
                    nodes[0].slope_before = nodes[0].slope_after
        if n >= 2:
            nodes[-1].slope_after = nodes[-1].slope_before
        return nodes

    # ================================================================
    # 双向同步：纵断面节点 ↔ 结构段
    # ================================================================
    def _on_long_table_edited(self, *_args):
        """节点表编辑（单元格值变更 / 转弯类型下拉框切换）后触发正向同步"""
        if self._syncing:
            return
        self._sync_nodes_to_segments()

    def _sync_nodes_to_segments(self):
        """正向同步：从纵断面节点表重建管身段（保留通用构件）"""
        if not SIPHON_AVAILABLE or self._syncing:
            return
        # 统计已填写桩号的有效行
        valid_rows = 0
        for r in range(self.long_table.rowCount()):
            item = self.long_table.item(r, 0)
            if item and item.text().strip():
                valid_rows += 1
        self._syncing = True
        try:
            nodes = self._build_longitudinal_nodes()
            common = [s for s in self.segments if s.direction == SegmentDirection.COMMON]
            if valid_rows < 2 or len(nodes) < 2:
                # 节点不足：清空纵断面管身段，保留通用构件
                self.segments = common
                self.longitudinal_nodes = nodes
                self._refresh_seg_table()
                self._update_canvas()
                return
            new_long_segs = self._nodes_to_segments(nodes)
            self.segments = common + new_long_segs
            self.longitudinal_nodes = nodes
            self._longitudinal_is_example = False
            self._update_segment_coefficients()
            self._refresh_seg_table()
            self._update_canvas()
        finally:
            self._syncing = False

    def _sync_segments_to_nodes(self):
        """反向同步：从纵断面管身段重建节点表"""
        if not SIPHON_AVAILABLE or self._syncing:
            return
        self._syncing = True
        try:
            long_segs = [s for s in self.segments
                         if s.direction != SegmentDirection.COMMON]
            if not long_segs:
                self.longitudinal_nodes.clear()
                self._refresh_long_table()
                return
            nodes = self._segments_to_nodes(long_segs)
            self.longitudinal_nodes = nodes
            self._refresh_long_table()
        finally:
            self._syncing = False

    def _nodes_to_segments(self, nodes):
        """将变坡点节点列表转换为纵断面管身段列表

        映射规则（与 DxfParser 产出一致）：
        - ARC 节点 → BEND 段（弧段从该节点到下一节点）
        - FOLD 节点 → FOLD 段（合并折点前后两段）
        - 其余相邻节点 → STRAIGHT 段
        """
        if len(nodes) < 2:
            return []
        segments = []
        i = 0
        while i < len(nodes) - 1:
            curr = nodes[i]
            nxt = nodes[i + 1]
            ds = nxt.chainage - curr.chainage
            dz = nxt.elevation - curr.elevation
            length = math.sqrt(ds ** 2 + dz ** 2) if (abs(ds) > 1e-9 or abs(dz) > 1e-9) else 0.0

            if curr.turn_type == TurnType.ARC and curr.vertical_curve_radius > 0:
                # 弧段：从 ARC 节点到下一节点
                segments.append(StructureSegment(
                    segment_type=SegmentType.BEND,
                    length=round(length, 4),
                    radius=curr.vertical_curve_radius,
                    angle=curr.turn_angle if curr.turn_angle > 0 else 0,
                    start_elevation=curr.elevation,
                    end_elevation=nxt.elevation,
                    direction=SegmentDirection.LONGITUDINAL,
                ))
                i += 1

            elif curr.turn_type == TurnType.FOLD and curr.turn_angle > 0:
                # 残余折点（上一轮 look-ahead 未消费）
                segments.append(StructureSegment(
                    segment_type=SegmentType.FOLD,
                    length=round(length, 4),
                    angle=curr.turn_angle,
                    start_elevation=curr.elevation,
                    end_elevation=nxt.elevation,
                    direction=SegmentDirection.LONGITUDINAL,
                ))
                i += 1

            elif nxt.turn_type == TurnType.FOLD:
                # 折点在 nxt：合并前后两段为一个 FOLD 段
                if i + 2 < len(nodes):
                    nxt2 = nodes[i + 2]
                    ds2 = nxt2.chainage - nxt.chainage
                    dz2 = nxt2.elevation - nxt.elevation
                    length2 = math.sqrt(ds2 ** 2 + dz2 ** 2) if (abs(ds2) > 1e-9 or abs(dz2) > 1e-9) else 0.0
                    segments.append(StructureSegment(
                        segment_type=SegmentType.FOLD,
                        length=round(length + length2, 4),
                        angle=nxt.turn_angle,
                        start_elevation=curr.elevation,
                        end_elevation=nxt2.elevation,
                        direction=SegmentDirection.LONGITUDINAL,
                    ))
                    i += 2
                else:
                    # 折点在末端，无后续段
                    segments.append(StructureSegment(
                        segment_type=SegmentType.FOLD,
                        length=round(length, 4),
                        angle=nxt.turn_angle,
                        start_elevation=curr.elevation,
                        end_elevation=nxt.elevation,
                        direction=SegmentDirection.LONGITUDINAL,
                    ))
                    i += 1

            else:
                # 直管段
                segments.append(StructureSegment(
                    segment_type=SegmentType.STRAIGHT,
                    length=round(length, 4),
                    start_elevation=curr.elevation,
                    end_elevation=nxt.elevation,
                    direction=SegmentDirection.LONGITUDINAL,
                ))
                i += 1
        return segments

    def _segments_to_nodes(self, long_segs):
        """将纵断面管身段列表转换为变坡点节点列表

        反向映射规则：
        - STRAIGHT → 起终点各一个 NONE 节点
        - BEND    → 起点 ARC 节点 + 终点 NONE 节点
        - FOLD    → 起点 NONE + 折点 FOLD（长度中点近似）+ 终点 NONE
        """
        if not long_segs:
            return []
        # 保留起始桩号（避免丢失 DXF 导入的偏移量）
        chainage = self.longitudinal_nodes[0].chainage if self.longitudinal_nodes else 0.0
        nodes = []

        for seg in long_segs:
            s_elev = seg.start_elevation if seg.start_elevation is not None else 0.0
            e_elev = seg.end_elevation if seg.end_elevation is not None else 0.0
            dh = e_elev - s_elev
            ds = math.sqrt(max(0.0, seg.length ** 2 - dh ** 2)) if seg.length > 0 else 0.0

            # 起点节点（与上一段终点合并）
            if not nodes or abs(nodes[-1].chainage - chainage) > 0.001:
                nodes.append(LongitudinalNode(
                    chainage=round(chainage, 3),
                    elevation=round(s_elev, 3),
                    turn_type=TurnType.NONE,
                ))

            if seg.segment_type == SegmentType.BEND:
                # 将起点升级为 ARC 节点
                nodes[-1].turn_type = TurnType.ARC
                nodes[-1].vertical_curve_radius = seg.radius
                nodes[-1].turn_angle = seg.angle
                chainage += ds
                nodes.append(LongitudinalNode(
                    chainage=round(chainage, 3),
                    elevation=round(e_elev, 3),
                    turn_type=TurnType.NONE,
                ))

            elif seg.segment_type == SegmentType.FOLD:
                # 折点按长度等分近似放在中点
                half_ds = ds / 2.0
                fold_elev = (s_elev + e_elev) / 2.0
                chainage += half_ds
                nodes.append(LongitudinalNode(
                    chainage=round(chainage, 3),
                    elevation=round(fold_elev, 3),
                    turn_type=TurnType.FOLD,
                    turn_angle=seg.angle,
                ))
                chainage += half_ds
                nodes.append(LongitudinalNode(
                    chainage=round(chainage, 3),
                    elevation=round(e_elev, 3),
                    turn_type=TurnType.NONE,
                ))

            else:
                # 直管段
                chainage += ds
                nodes.append(LongitudinalNode(
                    chainage=round(chainage, 3),
                    elevation=round(e_elev, 3),
                    turn_type=TurnType.NONE,
                ))

        # 补全坡角
        n = len(nodes)
        for i in range(n):
            if i > 0:
                d_s = nodes[i].chainage - nodes[i - 1].chainage
                d_z = nodes[i].elevation - nodes[i - 1].elevation
                if abs(d_s) > 1e-6:
                    slope = math.atan2(d_z, abs(d_s))
                    nodes[i].slope_before = slope
                    nodes[i - 1].slope_after = slope
        if n >= 1:
            nodes[0].slope_before = nodes[0].slope_after
        if n >= 2:
            nodes[-1].slope_after = nodes[-1].slope_before
        return nodes

    # ================================================================
    # 计算
    # ================================================================
    def _get_global_params(self):
        try:
            Q = float(self.edit_Q.text())
            v_guess = float(self.edit_v.text())
            n = float(self.edit_n.text())
            xi_inlet = float(self.edit_xi_inlet.text())
            xi_outlet = float(self.edit_xi_outlet.text())
        except (ValueError, TypeError) as e:
            InfoBar.error("输入错误", f"参数格式错误\n{str(e)}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return None

        v1 = self._fval(self.edit_v1, 0)
        v_out = self._fval(self.edit_v_out, 0)
        v3 = self._fval(self.edit_v3, 0)

        # v₂策略
        strategy_text = self.combo_v2_strategy.currentText()
        v2_strategy = V2_STRATEGY_MAP.get(strategy_text, V2Strategy.AUTO_PIPE)

        if v2_strategy == V2Strategy.AUTO_PIPE:
            v2 = 0.0
        elif v2_strategy == V2Strategy.V1_PLUS_02:
            v2 = v1 + 0.2
        else:
            v2 = self._fval(self.edit_v2, 0)

        # 渐变段型式
        inlet_type = GRADIENT_TYPE_MAP.get(self.combo_inlet_type.currentText(), GradientType.NONE)
        outlet_type = GRADIENT_TYPE_MAP.get(self.combo_outlet_type.currentText(), GradientType.NONE)

        num_pipes = max(1, self.spin_num_pipes.value()) if hasattr(self, 'spin_num_pipes') else 1

        return GlobalParameters(
            Q=Q, v_guess=v_guess, roughness_n=n,
            inlet_type=inlet_type, outlet_type=outlet_type,
            v_channel_in=v1, v_pipe_in=v2,
            v_channel_out=v_out, v_pipe_out=v3,
            xi_inlet=xi_inlet, xi_outlet=xi_outlet,
            v2_strategy=v2_strategy,
            num_pipes=num_pipes
        )

    def _execute_calculation(self):
        if not SIPHON_AVAILABLE:
            InfoBar.error("不可用", "计算引擎未加载", parent=self._info_parent(),
                         duration=5000, position=InfoBarPosition.TOP)
            return
        # 方案D：计算前验证拟定流速是否已确认
        if not self._validate_v_before_calc():
            return
        # 计算前验证管道根数是否已确认
        if not self._validate_num_pipes_before_calc():
            return
        try:
            params = self._get_global_params()
            if params is None:
                return
            if params.Q <= 0:
                InfoBar.error("输入错误", "设计流量必须大于0", parent=self._info_parent(),
                             duration=3000, position=InfoBarPosition.TOP)
                return
            if params.v_guess <= 0:
                InfoBar.error("输入错误", "拟定流速必须大于0", parent=self._info_parent(),
                             duration=3000, position=InfoBarPosition.TOP)
                return

            # 方案B：平面转弯半径倍数未确认时弹出黄色警告（不阻断）
            self._warn_turn_n_if_needed()

            # v₂验证（非阻断，仅警告）
            self._validate_inlet_velocity()

            # 同步纵断面节点
            self.longitudinal_nodes = self._build_longitudinal_nodes()

            # 自定义管径
            D_override_text = self.edit_D_override.text().strip()
            D_override = float(D_override_text) if D_override_text else None

            verbose = self.detail_cb.isChecked()

            # 确定加大比例（与其他4个面板保持一致：单次传入引擎）
            inc_pct_for_engine = None
            if self.inc_cb.isChecked():
                inc_text = self.edit_inc.text().strip()
                if inc_text:
                    try:
                        inc_pct_for_engine = float(inc_text)
                    except ValueError:
                        inc_pct_for_engine = None
                if inc_pct_for_engine is None:
                    try:
                        _calc_dir = os.path.join(_pkg_root, '渠系建筑物断面计算')
                        if _calc_dir not in sys.path:
                            sys.path.insert(0, _calc_dir)
                        from 明渠设计 import get_flow_increase_percent
                        inc_pct_for_engine = get_flow_increase_percent(params.Q)
                    except Exception:
                        inc_pct_for_engine = 20.0

            result = HydraulicCore.execute_calculation(
                params, self.segments,
                diameter_override=D_override,
                verbose=verbose,
                plan_segments=self.plan_segments,
                plan_total_length=self.plan_total_length,
                plan_feature_points=self.plan_feature_points,
                longitudinal_nodes=self.longitudinal_nodes,
                increase_percent=inc_pct_for_engine,
            )
            self.calculation_result = result
            self.calculation_result_increased = None  # 单次计算，结果在 result 本身
            self._inc_pct_used = inc_pct_for_engine or 0.0
            self._refresh_seg_table()

            # 回填流速
            self._fill_back_velocities(result)

            # 计算完成后刷新R值显示（追加✓确认标记）
            self._update_turn_R(confirmed=True)

            # 水损阈值检查
            threshold_str = self.edit_threshold.text().strip()
            if threshold_str:
                try:
                    threshold = float(threshold_str)
                    if threshold > 0 and result.total_head_loss > threshold:
                        InfoBar.warning("水损超限提醒",
                            f"计算总水面落差 ΔZ = {result.total_head_loss:.4f} m，"
                            f"已超过设定阈值 {threshold:.2f} m。"
                            f"建议调整拟定流速、管径或其他参数后重新计算。",
                            parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
                except ValueError:
                    pass

            # 显示结果（format_result 已内置加大工况展示）
            if not self._suppress_result_display:
                summary = HydraulicCore.format_result(result, show_steps=False)
                self.summary_text.setPlainText(summary)
                if verbose and result.calculation_steps:
                    detail = HydraulicCore.format_result(result, show_steps=True)
                    self.detail_text.setPlainText(detail)
                    self._detail_text_cache = detail
                    self.result_notebook.setCurrentIndex(1)
                else:
                    self.result_notebook.setCurrentIndex(0)

                self.params_notebook.setCurrentIndex(3)

                InfoBar.success("计算完成",
                    f"D={result.diameter:.3f}m  v={result.velocity:.3f}m/s  ΔZ={result.total_head_loss:.4f}m",
                    parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

            # 更新公式展示
            self._update_formula_display(result)

            # 更新画布和状态
            self._update_canvas()
            self._update_data_status()

            if self.on_result_callback:
                self.on_result_callback(result)

        except Exception as e:
            InfoBar.error("计算错误", f"计算过程发生错误: {str(e)}", parent=self._info_parent(),
                         duration=5000, position=InfoBarPosition.TOP)
            self.detail_text.setPlainText(f"计算过程发生错误:\n{traceback.format_exc()}")
            self.result_notebook.setCurrentIndex(1)
            self.params_notebook.setCurrentIndex(3)

    def _fill_back_velocities(self, result):
        """计算完成后回填实际使用的流速值（保留来源标注）"""
        green = f"color:{S};"

        # v（拟定流速）— 标记已参与计算（保持方案D的绿色边框样式）
        self.edit_v.setStyleSheet(
            f"LineEdit {{ border: 1.5px solid {S}; background: #F1F8E9; }}"
        )
        cur_v = self.lbl_v_hint.text()
        if '已从主表导入' in cur_v or '已导入' in cur_v:
            self.lbl_v_hint.setText("(已导入, ✓已参与计算)")
        else:
            self.lbl_v_hint.setText("(✓已参与计算)")
        self.lbl_v_hint.setStyleSheet(f"color:{S};font-size:12px;font-weight:bold;")

        # v₁ — 保留来源信息
        self.edit_v1.setText(f"{result.velocity_channel_in:.4f}")
        self.edit_v1.setStyleSheet(green)
        cur_v1 = self.lbl_v1_hint.text()
        if '已从主表导入' in cur_v1 or '已导入' in cur_v1:
            self.lbl_v1_hint.setText("(已导入, ✓已参与计算)")
        else:
            self.lbl_v1_hint.setText("(✓已参与计算)")
        self.lbl_v1_hint.setStyleSheet(f"color:{S};font-size:12px;")

        # v₂
        was_ro = self.edit_v2.isReadOnly()
        self.edit_v2.setReadOnly(False)
        self.edit_v2.setText(f"{result.velocity_pipe_in:.4f}")
        self.edit_v2.setStyleSheet(green)
        self.edit_v2.setReadOnly(was_ro)
        strategy = self.combo_v2_strategy.currentText()
        if "自动" in strategy:
            self.lbl_v2_hint.setText("(已计算: 取管道流速)")
        elif "v₁" in strategy:
            self.lbl_v2_hint.setText("(已计算: v₁ + 0.2)")
        elif "断面" in strategy:
            self.lbl_v2_hint.setText("(已计算: 断面反算)")
        else:
            self.lbl_v2_hint.setText("(已计算: 指定输入)")
        self.lbl_v2_hint.setStyleSheet(f"color:{S};font-size:12px;")

        # v_out — 出口渐变段始端流速（= 实际管道流速）
        self.edit_v_out.setText(f"{result.velocity_outlet_start:.4f}")
        self.edit_v_out.setStyleSheet(green)
        self.lbl_vout_hint.setText("(已计算: 管道流速)")
        self.lbl_vout_hint.setStyleSheet(f"color:{S};font-size:12px;")

        # v₃ — 保留来源信息
        self.edit_v3.setText(f"{result.velocity_channel_out:.4f}")
        self.edit_v3.setStyleSheet(green)
        cur_v3 = self.lbl_v3_hint.text()
        if '已从主表导入' in cur_v3 or '已导入' in cur_v3:
            self.lbl_v3_hint.setText("(已导入, ✓已参与计算)")
        else:
            self.lbl_v3_hint.setText("(✓已参与计算)")
        self.lbl_v3_hint.setStyleSheet(f"color:{S};font-size:12px;")

    def _update_formula_display(self, result):
        """计算完成后更新公式展示区（已迁移至 QWebEngineView + KaTeX 静态渲染，无需动态更新）"""
        pass

    # ================================================================
    # 画布更新
    # ================================================================
    def _update_canvas(self):
        """更新画布（100ms防抖，与原版一致）"""
        if not hasattr(self, '_canvas_timer'):
            self._canvas_timer = QTimer(self)
            self._canvas_timer.setSingleShot(True)
            self._canvas_timer.timeout.connect(self._do_update_canvas)
        self._canvas_timer.start(100)

    def _do_update_canvas(self):
        if self.canvas:
            self.canvas.set_data(
                segments=self.segments,
                plan_segments=self.plan_segments,
                plan_feature_points=self.plan_feature_points,
                plan_total_length=self.plan_total_length,
                longitudinal_nodes=self.longitudinal_nodes,
                longitudinal_is_example=self._longitudinal_is_example,
            )
            self.canvas.auto_select_view()
            self._update_zoom_label()

    def _update_data_status(self):
        """更新数据状态标签（含空间合并模式提示，与Tkinter版一致）"""
        has_plan_points = len(self.plan_feature_points) >= 2
        has_plan_segments = len(self.plan_segments) > 0 or self.plan_total_length > 0
        has_long_nodes = len(self.longitudinal_nodes) >= 2

        # 空间合并模式判断
        if has_plan_points and has_long_nodes:
            mode = "模式: 平面+纵断面（空间合并）"
            color = "#008800"
        elif has_plan_points and not has_long_nodes:
            if self._longitudinal_is_example:
                mode = "模式: 仅平面估算（纵断面为示例数据，可导入DXF替换）"
            else:
                mode = "模式: 仅平面估算（未导入纵断面）"
            color = "#CC6600"
        elif has_long_nodes and not has_plan_points:
            mode = "模式: 仅纵断面（未检测到平面数据）"
            color = "#CC6600"
        elif has_plan_segments:
            mode = "模式: 传统模式（仅平面总长度）"
            color = "#CC6600"
        elif self._longitudinal_is_example:
            mode = "模式: 纵断面为示例数据，可导入DXF或手动编辑"
            color = "#CC6600"
        else:
            mode = "模式: 无平面/纵断面数据"
            color = "#CC0000"

        # 附加数据计数
        counts = []
        n_seg = len(self.segments)
        n_long = len(self.longitudinal_nodes)
        n_plan = len(self.plan_feature_points)
        if n_seg: counts.append(f"结构段:{n_seg}")
        if n_long: counts.append(f"节点:{n_long}")
        if n_plan: counts.append(f"IP点:{n_plan}")
        if self.plan_total_length > 0:
            counts.append(f"平面长:{self.plan_total_length:.1f}m")

        text = mode
        if counts:
            text += "  |  " + ", ".join(counts)
        self.lbl_data_status.setText(text)
        self.lbl_data_status.setStyleSheet(f"color:{color};font-size:12px;background:transparent;border:none;")

    # ================================================================
    # 辅助函数
    # ================================================================
    def _fval(self, edit, default=0.0):
        t = edit.text().strip()
        if not t: return default
        try: return float(t)
        except ValueError: return default

    def _cell_float(self, table, row, col, default=0.0):
        item = table.item(row, col)
        if not item: return default
        t = item.text().strip()
        if not t: return default
        try: return float(t)
        except ValueError: return default

    def _info_parent(self):
        w = self.window()
        return w if w else self

    def _clear_results(self):
        self.summary_text.clear()
        self.detail_text.clear()
        self.calculation_result = None
        self._detail_text_cache = ""

    # ================================================================
    # 导出
    # ================================================================
    def _export_word(self):
        """导出Word计算书（工程产品运行卡格式）"""
        if not WORD_EXPORT_AVAILABLE:
            InfoBar.warning("缺少依赖",
                "Word导出需要安装 python-docx、latex2mathml、lxml。\n请执行: pip install python-docx latex2mathml lxml",
                parent=self._info_parent(), duration=6000, position=InfoBarPosition.TOP)
            return
        if not self.calculation_result:
            InfoBar.warning("提示", "请先执行计算后再导出。", parent=self._info_parent(),
                           duration=3000, position=InfoBarPosition.TOP)
            return
        # 若当前结果缺少详细步骤，以verbose模式重新计算一次
        if not self.calculation_result.calculation_steps and SIPHON_AVAILABLE:
            try:
                params = self._get_global_params()
                if params:
                    self.longitudinal_nodes = self._build_longitudinal_nodes()
                    D_override_text = self.edit_D_override.text().strip()
                    D_override = float(D_override_text) if D_override_text else None
                    result = HydraulicCore.execute_calculation(
                        params, self.segments,
                        diameter_override=D_override,
                        verbose=True,
                        plan_segments=self.plan_segments,
                        plan_total_length=self.plan_total_length,
                        plan_feature_points=self.plan_feature_points,
                        longitudinal_nodes=self.longitudinal_nodes,
                    )
                    self.calculation_result = result
            except Exception:
                pass
        name = self.edit_name.text().strip() or "倒虹吸"
        meta = load_meta()
        auto_purpose = build_calc_purpose('siphon', project=meta.project_name, name=name, section_type='')
        dlg = ExportConfirmDialog('siphon', '倒虹吸水力计算书', auto_purpose, parent=self._info_parent())
        from PySide6.QtWidgets import QDialog
        if dlg.exec() != QDialog.Accepted:
            return
        self._word_export_meta = dlg.get_meta()
        self._word_export_purpose = dlg.get_calc_purpose()
        self._word_export_refs = dlg.get_references()
        default_fn = f"{name}水力计算书.docx"
        filepath, _ = QFileDialog.getSaveFileName(self, "保存Word报告",
            default_fn, "Word文档 (*.docx);;所有文件 (*.*)")
        if not filepath:
            return
        try:
            self._build_word_report(filepath)
            InfoBar.success("导出成功", f"Word报告已保存到: {filepath}",
                           parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            ask_open_file(filepath, self._info_parent())
        except PermissionError:
            InfoBar.error("文件被占用", "无法写入文件，请先关闭已打开的同名Word文档，然后重新操作。",
                         parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("导出失败", f"Word导出失败: {str(e)}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _build_word_report(self, filepath):
        """构建Word报告文档（工程产品运行卡格式）"""
        r = self.calculation_result
        name = self.edit_name.text().strip() or "倒虹吸"
        meta = getattr(self, '_word_export_meta', load_meta())
        purpose = getattr(self, '_word_export_purpose', '')
        refs = getattr(self, '_word_export_refs', REFERENCES_BASE.get('siphon', []))

        doc = create_engineering_report_doc(
            meta=meta,
            calc_title='倒虹吸水力计算书',
            calc_content_desc=f'{name}倒虹吸水力计算',
            calc_purpose=purpose,
            references=refs,
            calc_program_text=f'渠系建筑物水力计算系统 V1.0\n倒虹吸水力计算',
        )
        doc.add_page_break()

        # 5. 基础公式
        doc_add_eng_h(doc, '5、基础公式')
        doc_add_formula(doc, r'D = \sqrt{\frac{4Q}{\pi v}}', '理论管径：')
        doc_add_formula(doc, r'A = \frac{\pi D^2}{4}', '断面积：')
        doc_add_formula(doc, r'R = \frac{D}{4}', '水力半径：')
        doc_add_formula(doc, r'C = \frac{1}{n} R^{1/6}', '谢才系数：')
        doc_add_formula(doc, r'h_f = \frac{L \cdot v^2}{C^2 \cdot R}', '沿程水头损失：')
        doc_add_formula(doc, r'h_j = \sum \xi \cdot \frac{v^2}{2g}', '局部水头损失：')
        doc_add_formula(doc, r'\Delta Z = \Delta Z_1 + \Delta Z_2 - \Delta Z_3', '总水面落差：')

        # 6. 设计参数
        doc_add_eng_h(doc, '6、设计参数')
        _np_word = self.spin_num_pipes.value() if hasattr(self, 'spin_num_pipes') else 1
        _Q_word = self._fval(self.edit_Q)
        params = [
            ("名称", name),
            ("设计流量 Q", f"{_Q_word:.4f} m³/s"),
            ("拟定流速 v", f"{self._fval(self.edit_v):.4f} m/s"),
            ("管道根数 N", f"{_np_word} 根"),
        ] + ([(f"单管流量 Q/N", f"{_Q_word/_np_word:.4f} m³/s")] if _np_word > 1 else []) + [
            ("糙率 n", f"{self._fval(self.edit_n):.4f}"),
            ("平面转弯半径倍数 n", f"{self._fval(self.edit_turn_n):.1f}"),
            ("进口渐变段型式", self.combo_inlet_type.currentText()),
            ("进口渐变段系数 ξ₁", f"{self._fval(self.edit_xi_inlet):.4f}"),
            ("出口渐变段型式", self.combo_outlet_type.currentText()),
            ("出口渐变段系数 ξ₂", f"{self._fval(self.edit_xi_outlet):.4f}"),
        ]
        # v₂策略
        strategy = self.combo_v2_strategy.currentText()
        params.append(("v₂策略", strategy))
        # 流速
        v1 = self._fval(self.edit_v1, 0)
        v2 = self._fval(self.edit_v2, 0)
        v_out = self._fval(self.edit_v_out, 0)
        v3 = self._fval(self.edit_v3, 0)
        if v1 > 0:
            params.append(("进口始端流速 v₁", f"{v1:.4f} m/s"))
        if v2 > 0:
            params.append(("进口末端流速 v₂", f"{v2:.4f} m/s"))
        if v_out > 0:
            params.append(("出口始端流速 v_out", f"{v_out:.4f} m/s"))
        if v3 > 0:
            params.append(("出口末端流速 v₃", f"{v3:.4f} m/s"))
        doc_add_param_table(doc, params)

        # 7. 计算结果
        doc_add_eng_h(doc, '7、计算结果')
        results = [
            ("理论管径 D理论", f"{r.diameter_theory:.4f} m"),
            ("设计管径 D", f"{r.diameter:.4f} m"),
            ("断面积 A", f"{r.area:.4f} m²"),
            ("管内流速 v", f"{r.velocity:.4f} m/s"),
            ("水力半径 Rh", f"{r.hydraulic_radius:.4f} m"),
            ("谢才系数 C", f"{r.chezy_c:.4f}"),
            ("进口渐变段落差 ΔZ₁", f"{r.loss_inlet:.4f} m"),
            ("管身段水头损失 ΔZ₂", f"{r.loss_pipe:.4f} m"),
            ("  其中沿程损失 hf", f"{r.loss_friction:.4f} m"),
            ("  其中局部损失 hj", f"{r.loss_local:.4f} m"),
            ("出口渐变段落差 ΔZ₃", f"{r.loss_outlet:.4f} m"),
            ("总水面落差 ΔZ", f"{r.total_head_loss:.4f} m"),
            ("管道总长", f"{r.total_length:.4f} m"),
        ]
        doc_add_result_table(doc, results)

        # 8. 详细计算过程
        if r.calculation_steps:
            steps_text = "\n".join(r.calculation_steps)
            doc_add_eng_h(doc, '8、详细计算过程')
            doc_render_calc_text_eng(doc, steps_text)
        else:
            summary = self.summary_text.toPlainText()
            if summary:
                doc_add_eng_h(doc, '8、计算过程')
                doc_render_calc_text_eng(doc, summary, skip_title_keyword='计算结果汇总')

        doc.save(filepath)

    def _export_excel(self):
        if not self.calculation_result:
            InfoBar.warning("提示", "无结果可导出", parent=self._info_parent(),
                           duration=2000, position=InfoBarPosition.TOP)
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            InfoBar.warning("缺少依赖", "需要: pip install openpyxl", parent=self._info_parent(),
                           duration=4000, position=InfoBarPosition.TOP)
            return
        filepath, _ = QFileDialog.getSaveFileName(self, "导出Excel",
            "倒虹吸水力计算结果.xlsx", "Excel文件 (*.xlsx)")
        if not filepath: return
        try:
            r = self.calculation_result
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "倒虹吸计算结果"
            name = self.edit_name.text().strip() or "倒虹吸"
            ws['A1'] = f"{name} — 水力计算结果报告"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:D1')
            ws['A1'].alignment = Alignment(horizontal='center')
            _np_xl = self.spin_num_pipes.value() if hasattr(self, 'spin_num_pipes') else 1
            _Q_xl = self._fval(self.edit_Q)
            _np_xl_rows = ([("单管流量 Q/N", f"{_Q_xl/_np_xl:.3f} m³/s")] if _np_xl > 1 else [])
            data_rows = [
                ("设计流量 Q", f"{_Q_xl:.3f} m³/s"),
                ("拟定流速 v", f"{self._fval(self.edit_v):.3f} m/s"),
                ("管道根数 N", f"{_np_xl} 根"),
            ] + _np_xl_rows + [
                ("糙率 n", f"{self._fval(self.edit_n):.4f}"),
                ("", ""),
                ("理论管径", f"{r.diameter_theory:.4f} m"),
                ("设计管径 D", f"{r.diameter:.4f} m"),
                ("断面积 A", f"{r.area:.4f} m²"),
                ("管内流速 v", f"{r.velocity:.4f} m/s"),
                ("水力半径 R_h", f"{r.hydraulic_radius:.4f} m"),
                ("谢才系数 C", f"{r.chezy_c:.4f}"),
                ("", ""),
                ("进口渐变段落差 ΔZ1", f"{r.loss_inlet:.4f} m"),
                ("管身段水头损失 ΔZ2", f"{r.loss_pipe:.4f} m"),
                ("  沿程损失 hf", f"{r.loss_friction:.4f} m"),
                ("  局部损失 hj", f"{r.loss_local:.4f} m"),
                ("出口渐变段落差 ΔZ3", f"{r.loss_outlet:.4f} m"),
                ("总水面落差 ΔZ", f"{r.total_head_loss:.4f} m"),
                ("管道总长", f"{r.total_length:.4f} m"),
            ]
            for i, (key, val) in enumerate(data_rows, 3):
                ws.cell(row=i, column=1, value=key)
                ws.cell(row=i, column=2, value=val)
                if key and not key.startswith(" "):
                    ws.cell(row=i, column=1).font = Font(bold=True)
            wb.save(filepath)
            InfoBar.success("导出成功", f"已保存: {filepath}", parent=self._info_parent(),
                           duration=4000, position=InfoBarPosition.TOP)
        except PermissionError:
            InfoBar.error("文件被占用", "无法写入文件，请先关闭已打开的同名文件（如Excel等），然后重新操作。",
                         parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("导出失败", str(e), parent=self._info_parent(),
                         duration=5000, position=InfoBarPosition.TOP)

    def _export_txt(self):
        txt = self.summary_text.toPlainText()
        detail = self.detail_text.toPlainText()
        if not txt and not detail:
            InfoBar.warning("提示", "无内容可导出", parent=self._info_parent(),
                           duration=2000, position=InfoBarPosition.TOP)
            return
        name = self.edit_name.text().strip() or "倒虹吸"
        default_fn = f"{name}水力计算报告.txt"
        filepath, _ = QFileDialog.getSaveFileName(self, "导出TXT",
            default_fn, "文本文件 (*.txt)")
        if not filepath: return
        try:
            content = f"名称: {name}\n\n{txt}"
            if detail:
                content += "\n\n" + detail
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(content)
            InfoBar.success("导出成功", f"已保存: {filepath}", parent=self._info_parent(),
                           duration=4000, position=InfoBarPosition.TOP)
            ask_open_file(filepath, self._info_parent())
        except PermissionError:
            InfoBar.error("文件被占用", "无法写入文件，请先关闭已打开的同名文件（如记事本等），然后重新操作。",
                         parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("导出失败", str(e), parent=self._info_parent(),
                         duration=5000, position=InfoBarPosition.TOP)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        auto_resize_table(self.seg_table)
        auto_resize_table(self.long_table)
