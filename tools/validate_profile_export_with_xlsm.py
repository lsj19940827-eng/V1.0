# -*- coding: utf-8 -*-
"""纵断面导出实数据验收脚本（基于 XLSM 关键行逐条核对 TXT/DXF 片段）。

用法：
    py -3 tools/validate_profile_export_with_xlsm.py
    py -3 tools/validate_profile_export_with_xlsm.py --xlsm "测试用_渠道表格...xlsm"
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from types import SimpleNamespace
from typing import Dict, List, Tuple

import ezdxf
from openpyxl import load_workbook


COL_TO_ROW_ID = {
    "BD": "bd_ip_before",
    "BE": "be_ip_text",
    "BF": "bf_ip_after",
    "BJ": "bj_station_before",
    "BK": "bk_station",
    "BL": "bl_station_after",
}


@dataclass
class CheckResult:
    name: str
    ok: bool
    detail: str


class _Node(SimpleNamespace):
    def get_structure_type_str(self):
        st = getattr(self, "structure_type", None)
        if st is None:
            return ""
        return st.value if hasattr(st, "value") else str(st)


def _load_cad_tools_module():
    root = Path(__file__).resolve().parents[1]
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))
    matches = list(root.glob("*/water_profile/cad_tools.py"))
    if not matches:
        raise FileNotFoundError("未找到 water_profile/cad_tools.py")
    target = matches[0]
    import importlib.util

    spec = importlib.util.spec_from_file_location("cad_tools_profile_validate_mod", target)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _discover_xlsm_path(user_path: str | None) -> Path:
    if user_path:
        p = Path(user_path)
        if p.exists():
            return p
        raise FileNotFoundError(f"指定 xlsm 不存在: {user_path}")
    candidates = [p for p in Path(".").glob("*.xlsm") if not p.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError("未找到可用 xlsm 文件")
    return candidates[0]


def _safe_float(v, default=0.0):
    try:
        if v is None:
            return default
        return float(v)
    except Exception:
        return default


def _extract_station_prefix(sheet) -> str:
    for row in range(9, sheet.max_row + 1):
        value = sheet[f"K{row}"].value
        if not value:
            continue
        text = str(value).strip()
        m = re.match(r"^(.*?)(\d+\+)", text)
        if m:
            return m.group(1)
    return ""


def _parse_b_label(label: str) -> Tuple[int, str, str, str]:
    """返回 (ip_number, name, structure, in_out)。"""
    text = str(label or "").strip()
    m = re.match(r"^IP(\d+)\s*(.*)$", text)
    ip_num = int(m.group(1)) if m else 0
    tail = (m.group(2) if m else text).strip()

    in_out = tail[-1] if tail.endswith(("进", "出")) else ""
    core = tail[:-1].strip() if in_out else tail

    full_map = [
        ("有压管道", "有压管道"),
        ("倒虹吸", "倒虹吸"),
        ("隧洞", "隧洞"),
        ("渡槽", "渡槽"),
        ("暗涵", "暗涵"),
    ]
    abbr_map = {"隧": "隧洞", "倒": "倒虹吸", "管": "有压管道", "渡": "渡槽", "暗": "暗涵"}

    structure = "明渠-矩形"
    name = ""
    special = False

    for keyword, full in full_map:
        if keyword in core:
            structure = full
            name = core.replace(keyword, "").strip()
            special = bool(in_out)
            break

    if not special and in_out and core:
        last = core[-1]
        if last in abbr_map:
            structure = abbr_map[last]
            name = core[:-1].strip()
            special = True

    if not special:
        name = ""
        structure = "明渠-矩形"
        in_out = ""

    return ip_num, name, structure, in_out


def _build_nodes_from_sheet(sheet) -> List[_Node]:
    nodes: List[_Node] = []
    for row in range(9, sheet.max_row + 1):
        b_val = sheet[f"B{row}"].value
        h_val = sheet[f"H{row}"].value
        if b_val is None or (not isinstance(h_val, (int, float))):
            continue

        ip_num, name, structure, in_out = _parse_b_label(str(b_val))
        node = _Node(
            src_row=row,
            ip_number=ip_num,
            name=name,
            structure_type=SimpleNamespace(value=structure),
            in_out=SimpleNamespace(value=in_out) if in_out else None,
            is_transition=False,
            station_BC=_safe_float(sheet[f"G{row}"].value, _safe_float(h_val)),
            station_MC=_safe_float(h_val),
            station_EC=_safe_float(sheet[f"I{row}"].value, _safe_float(h_val)),
            turn_angle=_safe_float(sheet[f"F{row}"].value, 0.0),
            bottom_elevation=_safe_float(sheet[f"AR{row}"].value, 0.0),
            top_elevation=_safe_float(sheet[f"AS{row}"].value, 0.0),
            water_level=_safe_float(sheet[f"AQ{row}"].value, 0.0),
            slope_i=1.0 / 2000.0,
        )
        nodes.append(node)
    return nodes


def _build_settings(cad_tools):
    return cad_tools._normalize_text_export_settings(
        {
            "text_height": 3.5,
            "rotation": 90,
            "elev_decimals": 3,
            "y_line_height": 120,
            "scale_x": 1,
            "scale_y": 1,
            "profile_row_items": [{"id": rid, "enabled": True} for rid in cad_tools._PROFILE_ROW_DEFAULT_ORDER],
        }
    )


def _parse_text_cmd(line: str):
    raw = (line or "").strip()
    m = re.match(r"^-text\s+([-\d\.eE]+),([-\d\.eE]+)\s+([-\d\.eE]+)\s+([-\d\.eE]+)\s+(.+)$", raw)
    if not m:
        return None
    return {
        "x": float(m.group(1)),
        "y": float(m.group(2)),
        "h": float(m.group(3)),
        "rot": float(m.group(4)),
        "text": m.group(5).strip(),
        "raw": raw,
    }


def _build_expected_from_records(cad_tools, nodes, station_prefix):
    records = cad_tools._build_ip_related_row_records(nodes, station_prefix)
    by_row: Dict[int, Dict[str, Dict[str, float | str]]] = {}
    for rid in COL_TO_ROW_ID.values():
        for rec in records[rid]:
            row = int(getattr(rec["node"], "src_row", -1))
            by_row.setdefault(row, {})
            by_row[row][rid] = {"x": float(rec["x"]), "text": str(rec["text"])}
    return by_row


def _collect_txt_rows_by_rid(txt_path: Path, row_layout: dict):
    parsed = []
    for line in txt_path.read_text(encoding="utf-8").splitlines():
        item = _parse_text_cmd(line)
        if item:
            parsed.append(item)

    rid_lines = {}
    for rid in COL_TO_ROW_ID.values():
        if rid not in row_layout:
            continue
        y_target = float(row_layout[rid]["text_y"])
        rid_lines[rid] = [
            item for item in parsed
            if abs(item["y"] - y_target) <= 1e-6 and abs(item["rot"] - 90.0) <= 1e-6
        ]
    return rid_lines


def _collect_dxf_rows_by_rid(dxf_path: Path, row_layout: dict):
    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()
    text_items = []
    for ent in msp.query("TEXT"):
        try:
            x, y, _z = ent.dxf.insert
            rot = float(getattr(ent.dxf, "rotation", 0.0))
            text_items.append({"x": float(x), "y": float(y), "rot": rot, "text": str(ent.dxf.text).strip()})
        except Exception:
            continue

    rid_lines = {}
    for rid in COL_TO_ROW_ID.values():
        if rid not in row_layout:
            continue
        y_target = float(row_layout[rid]["text_y"])
        rid_lines[rid] = [
            item for item in text_items
            if abs(item["y"] - y_target) <= 1e-6 and abs(item["rot"] - 90.0) <= 1e-6
        ]
    return rid_lines


def _find_line(lines, x, text):
    for item in lines:
        if abs(item["x"] - x) <= 1e-6 and item["text"].strip() == text.strip():
            return True
    return False


def _pick_key_rows(nodes):
    normal = [int(n.src_row) for n in nodes if not getattr(n, "in_out", None) and int(getattr(n, "ip_number", 0)) > 0][:3]
    special = [int(n.src_row) for n in nodes if getattr(n, "in_out", None) is not None][:3]
    merged = normal + [r for r in special if r not in normal]
    return merged


def _format_results(checks: List[CheckResult]) -> str:
    ok_count = sum(1 for c in checks if c.ok)
    bad_count = len(checks) - ok_count
    lines = [f"[SUMMARY] total={len(checks)} pass={ok_count} fail={bad_count}"]
    for c in checks:
        mark = "PASS" if c.ok else "FAIL"
        lines.append(f"[{mark}] {c.name} | {c.detail}")
    return "\n".join(lines)


def run_validation(xlsm_path: Path, out_dir: Path):
    cad_tools = _load_cad_tools_module()
    sheet = load_workbook(xlsm_path, data_only=True, keep_vba=True)["综合计算"]

    nodes = _build_nodes_from_sheet(sheet)
    if not nodes:
        raise RuntimeError("未从 xlsm 解析到节点数据")
    valid_nodes = [n for n in nodes if n.bottom_elevation or n.top_elevation or n.water_level]
    station_prefix = _extract_station_prefix(sheet)

    settings = _build_settings(cad_tools)
    enabled_ids, row_layout, _total_h, _line_h, _boundaries = cad_tools._build_profile_row_layout(settings)
    assert set(COL_TO_ROW_ID.values()).issubset(set(enabled_ids))

    class _Proj:
        def __init__(self, pfx):
            self._pfx = pfx

        def get_station_prefix(self):
            return self._pfx

    class _Panel:
        def __init__(self, pfx):
            self._pfx = pfx

        def _build_settings(self):
            return _Proj(self._pfx)

        def window(self):
            return None

    panel = _Panel(station_prefix)
    cad_tools.fluent_question = lambda *_a, **_k: False
    cad_tools.fluent_info = lambda *_a, **_k: None
    cad_tools.fluent_error = lambda *_a, **_k: None

    out_dir.mkdir(parents=True, exist_ok=True)
    txt_path = out_dir / "profile_acceptance_from_xlsm.txt"
    dxf_path = out_dir / "profile_acceptance_from_xlsm.dxf"

    cad_tools._export_longitudinal_txt_to_path(panel, nodes, valid_nodes, settings, str(txt_path))
    doc = ezdxf.new("R2010")
    msp = doc.modelspace()
    cad_tools._setup_dxf_style(doc)
    cad_tools._ensure_profile_layers(doc)
    cad_tools._draw_profile_on_msp(msp, nodes, valid_nodes, settings, station_prefix)
    doc.saveas(dxf_path)

    expected_by_row = _build_expected_from_records(cad_tools, nodes, station_prefix)
    txt_rows = _collect_txt_rows_by_rid(txt_path, row_layout)
    dxf_rows = _collect_dxf_rows_by_rid(dxf_path, row_layout)
    key_rows = _pick_key_rows(nodes)

    checks: List[CheckResult] = []

    # 1) TXT / DXF 与程序期望逐条核对（关键行）
    for row in key_rows:
        exp_row = expected_by_row.get(row, {})
        for col, rid in COL_TO_ROW_ID.items():
            exp = exp_row.get(rid)
            if not exp:
                checks.append(CheckResult(f"row{row}-{col}-expected", False, "未生成期望记录"))
                continue
            in_txt = _find_line(txt_rows.get(rid, []), exp["x"], exp["text"])
            in_dxf = _find_line(dxf_rows.get(rid, []), exp["x"], exp["text"])
            checks.append(CheckResult(f"row{row}-{col}-txt", in_txt, f"x={exp['x']:.6f}, text={exp['text']}"))
            checks.append(CheckResult(f"row{row}-{col}-dxf", in_dxf, f"x={exp['x']:.6f}, text={exp['text']}"))

    # 2) 普通IP行：与 xlsm 公式结果逐条比对
    row_to_node = {int(n.src_row): n for n in nodes}
    for row in key_rows:
        node = row_to_node.get(row)
        if not node:
            continue
        is_special = bool(getattr(node, "in_out", None))
        if is_special:
            continue
        exp_row = expected_by_row.get(row, {})
        for col, rid in COL_TO_ROW_ID.items():
            cmd = sheet[f"{col}{row}"].value
            parsed = _parse_text_cmd(str(cmd or ""))
            if not parsed:
                checks.append(CheckResult(f"row{row}-{col}-xlsm-parse", False, "无法解析 xlsm 命令"))
                continue
            exp = exp_row.get(rid)
            if not exp:
                checks.append(CheckResult(f"row{row}-{col}-xlsm-exp", False, "缺少程序期望"))
                continue
            x_ok = abs(parsed["x"] - exp["x"]) <= 1e-6
            t_ok = parsed["text"].strip() == exp["text"].strip()
            checks.append(CheckResult(f"row{row}-{col}-xlsm-x", x_ok, f"xlsm={parsed['x']:.6f}, actual={exp['x']:.6f}"))
            checks.append(CheckResult(f"row{row}-{col}-xlsm-text", t_ok, f"xlsm={parsed['text']} | actual={exp['text']}"))

    # 3) 特殊建筑行：业务规则核验（无弯前/弯后、结构全称）
    for row in key_rows:
        node = row_to_node.get(row)
        if not node or not getattr(node, "in_out", None):
            continue
        exp_row = expected_by_row.get(row, {})
        bd = exp_row.get("bd_ip_before", {}).get("text", "")
        be = exp_row.get("be_ip_text", {}).get("text", "")
        bf = exp_row.get("bf_ip_after", {}).get("text", "")
        no_suffix = ("弯前" not in bd) and ("弯后" not in bf)
        checks.append(CheckResult(f"row{row}-special-no-suffix", no_suffix, f"BD={bd} | BF={bf}"))

        has_full_struct = any(k in be for k in ("隧洞", "倒虹吸", "有压管道", "渡槽", "暗涵"))
        checks.append(CheckResult(f"row{row}-special-full-struct", has_full_struct, f"BE={be}"))

    report_text = _format_results(checks)
    report_json = {
        "xlsm": str(xlsm_path),
        "txt_output": str(txt_path),
        "dxf_output": str(dxf_path),
        "key_rows": key_rows,
        "summary": {
            "total": len(checks),
            "pass": sum(1 for c in checks if c.ok),
            "fail": sum(1 for c in checks if not c.ok),
        },
        "checks": [c.__dict__ for c in checks],
    }
    (out_dir / "profile_acceptance_report.txt").write_text(report_text, encoding="utf-8")
    (out_dir / "profile_acceptance_report.json").write_text(
        json.dumps(report_json, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(report_text)
    print(f"\n[REPORT] {out_dir / 'profile_acceptance_report.txt'}")
    print(f"[REPORT] {out_dir / 'profile_acceptance_report.json'}")
    return checks


def main():
    parser = argparse.ArgumentParser(description="纵断面导出实数据验收")
    parser.add_argument("--xlsm", default="", help="待验收 xlsm 路径（默认自动发现）")
    parser.add_argument(
        "--out-dir",
        default="tools/_tmp_test_outputs",
        help="输出目录（默认 tools/_tmp_test_outputs）",
    )
    args = parser.parse_args()

    xlsm_path = _discover_xlsm_path(args.xlsm or None)
    checks = run_validation(xlsm_path, Path(args.out_dir))
    failed = [c for c in checks if not c.ok]
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
