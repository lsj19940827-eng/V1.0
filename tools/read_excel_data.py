import openpyxl
import sys

# 设置UTF-8编码输出
sys.stdout.reconfigure(encoding='utf-8')

# 读取Excel文件
wb = openpyxl.load_workbook('多流量段表格填写示例.xlsx', data_only=True)
ws = wb.active

# 提取第1行的起始水位值
print('=' * 80)
print('第1行数据（起始水位值）')
print('=' * 80)
first_row = []
for cell in ws[1]:
    first_row.append(cell.value)
print(f'列标题: {first_row[0]}')
print(f'起始水位值: {first_row[1]}')
print()

# 提取第3行开始的所有数据行（共44行）
print('=' * 80)
print('第3-46行数据（共44行，每行20个字段）')
print('=' * 80)
print()
data_rows = []
for row_idx in range(3, 47):  # 从第3行到第46行，共44行
    row_data = []
    for cell in ws[row_idx]:
        val = cell.value
        # 将None转换为空字符串
        if val is None:
            row_data.append('')
        else:
            row_data.append(val)
    data_rows.append(tuple(row_data))

# 以Python格式输出
print('data_rows = [')
for i, row in enumerate(data_rows):
    if i < len(data_rows) - 1:
        print(f'    {row},')
    else:
        print(f'    {row}')
print(']')
print()
print('=' * 80)
print(f'统计信息: 共读取 {len(data_rows)} 行数据，每行 {len(data_rows[0])} 个字段')
print('=' * 80)
