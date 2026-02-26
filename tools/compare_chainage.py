import openpyxl

wb1 = openpyxl.load_workbook(r'data/示例二桩号比对，当做标准答案.xlsx')
wb2 = openpyxl.load_workbook(r'data/示例二程序计算的桩号.xlsx')
ws1 = wb1.active
ws2 = wb2.active

# --- 读取标准答案 (行1=列头, 行2=单位, 行3起=数据) ---
# 列0=名称, 列1=X, 列2=Y, 列3=R, 列4=转角°, 列5=切线长, 列6=弧长, 列7=BC, 列8=MC, 列9=EC
std_rows = []
for row in ws1.iter_rows(min_row=3, values_only=True):
    name = row[0]
    if name is None:
        continue
    std_rows.append({
        'name':  name,
        'x':     row[1],
        'y':     row[2],
        'r':     float(row[3]) if row[3] else 0.0,
        'delta': float(row[4]) if row[4] else 0.0,
        'tl':    float(row[5]) if row[5] else 0.0,
        'arc':   float(row[6]) if row[6] else 0.0,
        'bc':    float(row[7]) if row[7] is not None else 0.0,
        'mc':    float(row[8]) if row[8] is not None else 0.0,
        'ec':    float(row[9]) if row[9] is not None else 0.0,
    })

def parse_ch(s):
    """将 '总干10+160.646' 解析为 10160.646"""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip()
    if '+' in s:
        idx = s.rfind('+')
        km_str = s[:idx].replace('总干','').replace('总','').strip()
        m_str  = s[idx+1:].strip()
        return float(km_str) * 1000 + float(m_str)
    return float(s)

# --- 读取程序计算 (无列头, 每行一个节点) ---
# 列4=名称, 列7=R, 列8=转角, 列9=切线长, 列10=弧长
# 列13=IP桩号, 列14=BC, 列15=MC, 列16=EC
calc_rows = []
for row in ws2.iter_rows(values_only=True):
    if row[4] is None:
        continue
    calc_rows.append({
        'name':  row[4],
        'r':     row[7],
        'delta': row[8],
        'tl':    row[9],
        'arc':   row[10],
        'ip':    parse_ch(row[13]),
        'bc':    parse_ch(row[14]),
        'mc':    parse_ch(row[15]),
        'ec':    parse_ch(row[16]),
    })

THRESH = 0.005  # 5mm

print(f'共 {len(std_rows)} 行标准答案, {len(calc_rows)} 行程序计算')
print()
print(f"{'#':>3}  {'标准名称':<25} {'程序名称':<25}  "
      f"{'标准R':>5} {'程序R':>5}  "
      f"{'标准BC':>12} {'程序BC':>12} {'dBC':>9}  "
      f"{'标准EC':>12} {'程序EC':>12} {'dEC':>9}  状态")
print('-' * 145)

first_diff = True
n = min(len(std_rows), len(calc_rows))
for i in range(n):
    s = std_rows[i]
    c = calc_rows[i]

    d_bc = (c['bc'] - s['bc']) if c['bc'] is not None else None
    d_ec = (c['ec'] - s['ec']) if c['ec'] is not None else None

    has_diff = (d_bc is not None and abs(d_bc) > THRESH) or \
               (d_ec is not None and abs(d_ec) > THRESH)

    flag = ''
    if has_diff:
        if first_diff:
            flag = '  <<=== 首次偏差'
            first_diff = False
        else:
            flag = '  <<<'

    bc_c = f"{c['bc']:12.3f}" if c['bc'] is not None else f"{'None':>12}"
    ec_c = f"{c['ec']:12.3f}" if c['ec'] is not None else f"{'None':>12}"
    d_bc_s = f"{d_bc:+9.4f}" if d_bc is not None else f"{'N/A':>9}"
    d_ec_s = f"{d_ec:+9.4f}" if d_ec is not None else f"{'N/A':>9}"
    r_s = f"{s['r']:5.0f}" if s['r'] else f"{'?':>5}"
    r_c = f"{float(c['r']):5.0f}" if c['r'] is not None else f"{'?':>5}"

    print(f"{i+1:>3}  {s['name']:<25} {c['name']:<25}  "
          f"{r_s} {r_c}  "
          f"{s['bc']:12.3f} {bc_c} {d_bc_s}  "
          f"{s['ec']:12.3f} {ec_c} {d_ec_s}{flag}")

if len(std_rows) != len(calc_rows):
    print(f'\n行数不等: 标准={len(std_rows)}, 程序={len(calc_rows)}')

# ============================================================
# 第二部分：IP桩号 = 标准(BC+T) vs 程序col13
# ============================================================
print()
print('=== IP桩号 对比 (标准=BC+T, 程序=col13) ===')
print(f"{'#':>3}  {'标准名称':<25} {'程序名称':<25}  {'标准IP(BC+T)':>14} {'程序IP(col13)':>14} {'差值':>10}  状态")
print('-' * 105)

first_ip_diff = True
for i in range(n):
    s = std_rows[i]
    c = calc_rows[i]
    ip_std = s['bc'] + s['tl']   # IP桩号 = BC + T
    ip_calc = c['ip']
    if ip_calc is None:
        print(f"{i+1:>3}  {s['name']:<25} {c['name']:<25}  {ip_std:14.3f} {'None':>14} {'N/A':>10}")
        continue
    d = ip_calc - ip_std
    flag = ''
    if abs(d) > 0.01:
        if first_ip_diff:
            flag = '  <<=== 首次IP偏差'
            first_ip_diff = False
        else:
            flag = '  <<<'
    print(f"{i+1:>3}  {s['name']:<25} {c['name']:<25}  {ip_std:14.3f} {ip_calc:14.3f} {d:+10.4f}{flag}")

# ============================================================
# 第三部分：R / 转角delta / 切线长T / 弧长L 对比
# ============================================================
print()
print('=== 曲线参数对比 (R, 转角°, 切线长T, 弧长L) ===')
print(f"{'#':>3}  {'名称':<25}  {'stdR':>5} {'calR':>5}  {'std∆°':>8} {'cal∆°':>8} {'d∆':>7}  {'stdT':>8} {'calT':>8} {'dT':>7}  {'stdL':>8} {'calL':>8} {'dL':>7}  状态")
print('-' * 130)

first_param_diff = True
for i in range(n):
    s = std_rows[i]
    c = calc_rows[i]
    c_r     = float(c['r'])     if c['r']     is not None else 0.0
    c_delta = float(c['delta']) if c['delta'] is not None else 0.0
    c_tl    = float(c['tl'])    if c['tl']    is not None else 0.0
    c_arc   = float(c['arc'])   if c['arc']   is not None else 0.0

    d_r  = c_r     - s['r']
    d_d  = c_delta - s['delta']
    d_t  = c_tl    - s['tl']
    d_l  = c_arc   - s['arc']

    has_diff = abs(d_r) > 0.01 or abs(d_d) > 0.001 or abs(d_t) > 0.005 or abs(d_l) > 0.005
    flag = ''
    if has_diff:
        if first_param_diff:
            flag = '  <<=== 首次参数偏差'
            first_param_diff = False
        else:
            flag = '  <<<'
    print(f"{i+1:>3}  {s['name']:<25}  {s['r']:5.0f} {c_r:5.0f}  {s['delta']:8.4f} {c_delta:8.4f} {d_d:+7.4f}  {s['tl']:8.4f} {c_tl:8.4f} {d_t:+7.4f}  {s['arc']:8.4f} {c_arc:8.4f} {d_l:+7.4f}{flag}")
