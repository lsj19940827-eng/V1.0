# -*- coding: utf-8 -*-
"""有压管道水锤验算 Excel 导出辅助模块。"""

from __future__ import annotations

import re
from typing import Any, Dict, Iterable, List


SUMMARY_HEADERS = [
    "序号",
    "整线",
    "水锤段",
    "验算状态",
    "结论",
    "允许压力(MPa)",
    "允许压力水头(m)",
    "承压基准",
    "承压超限点数",
    "负压点数",
    "采样点数",
    "最危险桩号(m)",
    "承压余量(m)",
    "管顶最低余量(m)",
    "cp",
    "a0",
    "GB/T ΔH+(m)",
    "线性启闭 ΔH+(m)",
    "控制ΔH+(m)",
    "控制来源",
    "备注",
]

DETAIL_COLUMNS = [
    ("桩号(m)", "station_m"),
    ("所属成员", "member_label"),
    ("D(m)", "diameter_m"),
    ("v0(m/s)", "velocity_mps"),
    ("a(m/s)", "a"),
    ("cp", "pipe_coefficient_cp"),
    ("a0", "reinforcement_ratio_a0"),
    ("管顶高程(m)", "pipe_top_elevation_m"),
    ("管中心高程(m)", "centerline_elevation_m"),
    ("管底高程(m)", "pipe_bottom_elevation_m"),
    ("Hst(m)", "h_st_m"),
    ("Hmax(m)", "hmax_m"),
    ("Hmin(m)", "hmin_m"),
    ("GB/T ΔH+(m)", "gbt_positive_delta_h_m"),
    ("线性启闭 ΔH+(m)", "linear_positive_delta_h_m"),
    ("控制ΔH+(m)", "positive_delta_h_m"),
    ("允许压力水头(m)", "pressure_allow_head_m"),
    ("管底最大压力水头(m)", "pressure_head_max_m"),
    ("承压余量(m)", "pressure_margin_m"),
    ("负ΔH(m)", "negative_delta_h_m"),
    ("管顶最低压力水头(m)", "top_min_pressure_head_m"),
    ("控制来源", "positive_governing_method"),
    ("状态", "status"),
    ("负压状态", "negative_status"),
    ("分布说明", "distribution_note"),
    ("图1-3-3对照", "diagram_type_check"),
    ("流速来源", "velocity_source"),
    ("采用流量(m³/s)", "velocity_flow_m3s"),
]


def water_hammer_exportable_detail_count(segments: Iterable[Dict[str, Any]]) -> int:
    """统计当前可导出的水锤处理结果数量。"""
    count = 0
    for segment in segments:
        result = segment.get("result", {}) if isinstance(segment, dict) else {}
        details = result.get("details", []) if isinstance(result, dict) else []
        if details or _is_water_hammer_exempt_result(result):
            count += 1
    return count


def build_water_hammer_export_workbook(segments: List[Dict[str, Any]]):
    """根据水锤段结果构建 Excel 工作簿。"""
    openpyxl, styles, get_column_letter = _load_openpyxl()
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "汇总"
    _write_summary_sheet(summary_sheet, segments, styles, get_column_letter)

    used_names = {"汇总"}
    detail_index = 1
    for segment in segments:
        result = segment.get("result", {}) if isinstance(segment, dict) else {}
        details = result.get("details", []) if isinstance(result, dict) else []
        if not details:
            continue
        sheet_name = _unique_sheet_name(
            f"明细{detail_index}-{segment.get('segment_name', '')}",
            used_names,
        )
        used_names.add(sheet_name)
        sheet = workbook.create_sheet(sheet_name)
        member_lookup = segment.get("member_lookup", {}) if isinstance(segment, dict) else {}
        _write_detail_sheet(sheet, details, styles, get_column_letter, member_lookup=member_lookup)
        detail_index += 1
    return workbook


def save_water_hammer_export_workbook(filepath: str, segments: List[Dict[str, Any]]) -> None:
    """把水锤段结果保存为 Excel 文件。"""
    workbook = build_water_hammer_export_workbook(segments)
    workbook.save(filepath)


def _load_openpyxl():
    """延迟加载 openpyxl，便于界面给出缺依赖提示。"""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    styles = {
        "header_font": Font(bold=True, color="FFFFFF"),
        "header_fill": PatternFill("solid", fgColor="1F6FB2"),
        "header_alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "cell_alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }
    return openpyxl, styles, get_column_letter


def _write_summary_sheet(sheet, segments: List[Dict[str, Any]], styles: Dict[str, Any], get_column_letter) -> None:
    """写入水锤段汇总表。"""
    sheet.append(SUMMARY_HEADERS)
    _style_header_row(sheet, styles)
    for index, segment in enumerate(segments, start=1):
        result = segment.get("result", {}) if isinstance(segment, dict) else {}
        details = result.get("details", []) if isinstance(result, dict) else []
        is_calculated = bool(details)
        is_exempt = _is_water_hammer_exempt_result(result)
        is_processed = is_calculated or is_exempt
        critical = result.get("critical_point", {}) if isinstance(result, dict) else {}
        row = [
            index,
            segment.get("route_name", ""),
            segment.get("segment_name", ""),
            _format_summary_check_state(is_calculated=is_calculated, is_exempt=is_exempt),
            result.get("status", "") if is_processed else "",
            _number_or_blank(result.get("allowable_pressure_mpa")) if is_processed else "",
            _number_or_blank(result.get("pressure_allow_head_m")) if is_processed else "",
            result.get("pressure_check_basis_label", result.get("pressure_check_basis", "")) if is_processed else "",
            result.get("exceed_count", 0) if is_calculated else "",
            result.get("negative_pressure_risk_count", 0) if is_calculated else "",
            result.get("sample_count", 0) if is_calculated else "",
            _number_or_blank(critical.get("station_m")) if is_calculated and isinstance(critical, dict) else "",
            _number_or_blank(result.get("min_margin_m")) if is_calculated else "",
            _number_or_blank(result.get("min_negative_margin_m")) if is_calculated else "",
            _number_or_blank(_summary_result_value(result, "pipe_coefficient_cp")) if is_processed else "",
            _number_or_blank(_summary_result_value(result, "reinforcement_ratio_a0")) if is_processed else "",
            _number_or_blank(result.get("gbt_positive_delta_h")) if is_calculated else "",
            _number_or_blank(result.get("linear_positive_delta_h")) if is_calculated else "",
            _number_or_blank(result.get("positive_delta_h")) if is_calculated else "",
            result.get("positive_governing_method", "") if is_calculated else "",
            _summary_reason(result),
        ]
        sheet.append(row)
    _style_body(sheet, styles)
    _autosize_columns(sheet, get_column_letter)


def _summary_result_value(result: Dict[str, Any], key: str) -> Any:
    """从汇总、关键点或成员结果中读取展示值。"""
    if not isinstance(result, dict):
        return None
    value = result.get(key)
    if value is not None:
        return value
    critical = result.get("critical_point", {})
    if isinstance(critical, dict) and critical.get(key) is not None:
        return critical.get(key)
    for member in list(result.get("member_results", []) or []):
        if isinstance(member, dict) and member.get(key) is not None:
            return member.get(key)
    return None


def _summary_reason(result: Dict[str, Any]) -> str:
    """返回导出汇总备注，优先使用状态原因，其次使用cp简化提示。"""
    if not isinstance(result, dict):
        return ""
    return str(result.get("reason", "") or result.get("pipe_coefficient_note", "") or "")


def _write_detail_sheet(
    sheet,
    details: List[Dict[str, Any]],
    styles: Dict[str, Any],
    get_column_letter,
    *,
    member_lookup: Dict[str, Any] | None = None,
) -> None:
    """写入单个水锤段的采样明细表。"""
    sheet.append([label for label, _key in DETAIL_COLUMNS])
    _style_header_row(sheet, styles)
    lookup = member_lookup if isinstance(member_lookup, dict) else {}
    for item in details:
        row = []
        for _label, key in DETAIL_COLUMNS:
            value = item.get(key, "") if isinstance(item, dict) else ""
            if key == "member_label":
                value = _format_member_label(item, lookup)
            elif key == "diagram_type_check":
                value = _format_diagram_text(value)
            else:
                value = _number_or_blank(value)
            row.append(value)
        sheet.append(row)
    _style_body(sheet, styles)
    _autosize_columns(sheet, get_column_letter)


def _style_header_row(sheet, styles: Dict[str, Any]) -> None:
    """设置表头样式。"""
    for cell in sheet[1]:
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["header_alignment"]


def _style_body(sheet, styles: Dict[str, Any]) -> None:
    """设置普通单元格样式。"""
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = styles["cell_alignment"]


def _autosize_columns(sheet, get_column_letter) -> None:
    """按内容大致调整列宽。"""
    for column_cells in sheet.columns:
        column_index = column_cells[0].column
        max_len = 8
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 32))
        sheet.column_dimensions[get_column_letter(column_index)].width = max_len + 2


def _number_or_blank(value: Any):
    """数字保持为数值，空值保持为空字符串。"""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return round(float(value), 6)
    return value


def _is_water_hammer_exempt_result(result: Any) -> bool:
    """判断水锤结果是否为规范免验算。"""
    if not isinstance(result, dict):
        return False
    return bool(result.get("is_exempt")) or str(result.get("status", "") or "") == "可不验算"


def _format_summary_check_state(*, is_calculated: bool, is_exempt: bool) -> str:
    """返回汇总表中的处理状态文字。"""
    if is_calculated:
        return "已验算"
    if is_exempt:
        return "已处理"
    return "未验算"


def _format_diagram_text(value: Any) -> str:
    """格式化图1-3-3对照字段。"""
    if not isinstance(value, dict) or not value:
        return ""
    return str(value.get("positive_region", "") or value.get("negative_region", "") or "")


def _format_member_label(item: Dict[str, Any], member_lookup: Dict[str, Any]) -> str:
    """按 member_key 生成明细行所属成员名称。"""
    member_key = str((item or {}).get("member_key", "") or "")
    info = member_lookup.get(member_key, {}) if isinstance(member_lookup, dict) else {}
    if isinstance(info, dict):
        return str(info.get("label", "") or info.get("display_name", "") or member_key)
    return member_key


def _unique_sheet_name(raw_name: str, used_names: set[str]) -> str:
    """生成 Excel 可用且不重复的工作表名称。"""
    base = re.sub(r"[:\\/?*\[\]]+", "-", str(raw_name or "明细")).strip() or "明细"
    base = base[:31]
    name = base
    counter = 2
    while name in used_names:
        suffix = f"-{counter}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    return name
