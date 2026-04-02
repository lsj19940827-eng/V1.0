# -*- coding: utf-8 -*-
"""
水面线面板 CAD 工具集

移植自原版 TK 的工程辅助功能，包括：
1. 生成纵断面表格（AutoCAD pl + -text 命令）
2. 生成bzzh2命令内容（ZDM用）
3. 建筑物名称上平面图（AutoCAD -TEXT 命令）
4. IP坐标及弯道参数表导出Excel
5. 断面汇总表
"""

import os
import sys
import math
import copy
import json
import re

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGridLayout, QBoxLayout,
    QLabel, QGroupBox, QTextEdit,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QFileDialog, QApplication, QScrollArea, QWidget, QComboBox, QFrame,
    QSizePolicy, QMenu, QListWidget, QListWidgetItem, QLayout,
)
from PySide6.QtCore import Qt, Signal, QMimeData, QSettings, QSize, QEvent, QTimer
from PySide6.QtGui import QFont, QShortcut, QKeySequence, QDrag, QColor

from qfluentwidgets import (
    PushButton, PrimaryPushButton, LineEdit, SearchLineEdit,
    PopupTeachingTip, TeachingTipTailPosition, InfoBarIcon,
    ElevatedCardWidget, HeaderCardWidget, ListWidget, SegmentedWidget,
    ToolButton, FluentIcon, BodyLabel, CaptionLabel, InfoBar, InfoBarPosition, CheckBox,
)

from app_渠系计算前端.water_profile.text_export_settings_dialog import (
    create_text_export_settings_dialog,
)

from app_渠系计算前端.styles import (
    auto_resize_table, DIALOG_STYLE,
    fluent_info, fluent_error, fluent_question,
)

_XXPIPE_PROFILE_STATION_TOL = 1e-3
_XXPIPE_PROFILE_GEOMETRY_TOL = 1e-9

# 确保推求水面线模块可用
_pkg_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_water_profile_dir = os.path.join(_pkg_root, '推求水面线')
if _water_profile_dir not in sys.path:
    sys.path.insert(0, _water_profile_dir)
if _pkg_root not in sys.path:
    sys.path.insert(0, _pkg_root)

try:
    from models.data_models import ProjectSettings
    from models.enums import StructureType, InOutType
    MODELS_AVAILABLE = True
except ImportError:
    MODELS_AVAILABLE = False

from utils.pressure_pipe_result_helpers import make_pressure_pipe_identity

try:
    from config.constants import (
        XXPIPE_CHANNEL_LEVEL_OPTIONS,
        XXPIPE_ALLOWED_STRUCTURE_KEYWORDS,
        XXPIPE_ALLOWED_STRUCTURE_OPTIONS,
    )
except Exception:
    XXPIPE_CHANNEL_LEVEL_OPTIONS = ["总干管", "分干管", "干管", "支管", "分支管"]
    XXPIPE_ALLOWED_STRUCTURE_KEYWORDS = ()
    XXPIPE_ALLOWED_STRUCTURE_OPTIONS = [
        "有压管道",
        "定向钻",
        "顶管",
    ]

_SIPHON_PIPE_MATERIALS = [
    "PCCP管",
    "球墨铸铁管",
    "钢管",
    "钢筋混凝土管",
    "玻璃钢夹砂管",
]
_PRESSURE_PIPE_MATERIALS = [
    "HDPE管",
    "玻璃钢夹砂管",
    "球墨铸铁管",
    "PCCP管",
    "钢管",
    "钢筋混凝土管",
]
_PRESSURIZED_PIPE_MATERIALS = list(dict.fromkeys(_SIPHON_PIPE_MATERIALS + _PRESSURE_PIPE_MATERIALS))


def _safe_qt_parent(candidate):
    """Return a QWidget parent when available, otherwise None."""
    if isinstance(candidate, QWidget):
        return candidate
    if candidate is None:
        return None
    try:
        parent = candidate.window()
    except Exception:
        return None
    return parent if isinstance(parent, QWidget) else None


class _AutoWrapCaptionLabel(CaptionLabel):
    """CaptionLabel variant that reports wrapped height accurately for layouts."""

    def __init__(self, text="", parent=None):
        super().__init__(parent)
        self.setWordWrap(True)
        self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
        self.setText(text)
        self._sync_height_constraints()

    def hasHeightForWidth(self):
        return True

    def heightForWidth(self, width):
        margins = self.contentsMargins()
        available_width = max(1, int(width) - margins.left() - margins.right())
        text = self.text().strip()
        if not text:
            return margins.top() + margins.bottom()
        flags = int(self.alignment()) | int(Qt.TextWordWrap)
        rect = self.fontMetrics().boundingRect(0, 0, available_width, 32767, flags, text)
        return rect.height() + margins.top() + margins.bottom()

    def sizeHint(self):
        hint = super().sizeHint()
        hint.setHeight(self.heightForWidth(self._current_width_hint(hint.width())))
        return hint

    def minimumSizeHint(self):
        hint = super().minimumSizeHint()
        hint.setHeight(self.heightForWidth(self._current_width_hint(hint.width())))
        return hint

    def setText(self, text):
        super().setText(text)
        self._sync_height_constraints()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._sync_height_constraints()

    def event(self, event):
        if event.type() in (QEvent.FontChange, QEvent.StyleChange, QEvent.LayoutRequest):
            self._sync_height_constraints()
        return super().event(event)

    def _current_width_hint(self, default_width=0):
        width = self.width()
        if width > 0:
            return width
        parent = self.parentWidget()
        if parent is not None and parent.width() > 0:
            return parent.width()
        return max(1, int(default_width or 1))

    def _sync_height_constraints(self):
        target_height = self.heightForWidth(self._current_width_hint(super().sizeHint().width()))
        if self.minimumHeight() != target_height:
            self.setMinimumHeight(target_height)
        self.updateGeometry()


class _ProfileRowItemWidget(QWidget):
    """列表行组件：使用 qfluentwidgets.CheckBox 保持 Fluent 原生勾选框样式。"""

    clicked = Signal()
    doubleClicked = Signal()
    dragRequested = Signal()

    def __init__(self, title, subtitle, enabled, parent=None):
        super().__init__(parent)
        self._press_pos = None
        self._selected = False
        self._enabled = bool(enabled)

        self.setObjectName("profileRowItem")
        self.setAttribute(Qt.WA_StyledBackground, True)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 5, 10, 5)
        layout.setSpacing(10)

        self.checkbox = CheckBox("")
        self.checkbox.setChecked(bool(enabled))
        self.checkbox.setFixedWidth(36)
        self.checkbox.clicked.connect(self.clicked)
        layout.addWidget(self.checkbox, 0, Qt.AlignTop)

        text_col = QVBoxLayout()
        text_col.setContentsMargins(0, 0, 0, 0)
        text_col.setSpacing(1)

        self.title_label = QLabel()
        self.title_label.setTextInteractionFlags(Qt.NoTextInteraction)
        self.subtitle_label = QLabel()
        self.subtitle_label.setTextInteractionFlags(Qt.NoTextInteraction)

        text_col.addWidget(self.title_label)
        text_col.addWidget(self.subtitle_label)
        layout.addLayout(text_col, 1)

        for child in (self.title_label, self.subtitle_label):
            child.installEventFilter(self)

        self.set_content(title, subtitle, enabled)
        self.set_selected(False)

    def set_content(self, title, subtitle, enabled):
        self._enabled = bool(enabled)
        self.title_label.setText(title)
        self.subtitle_label.setText(subtitle)
        self._apply_visual_state()

    def set_selected(self, selected):
        self._selected = bool(selected)
        self._apply_visual_state()

    def _apply_visual_state(self):
        if self._selected:
            self.setStyleSheet(
                "QWidget#profileRowItem {"
                "background: rgba(210, 232, 255, 0.92);"
                "border: 1px solid rgba(0, 120, 212, 0.34);"
                "border-left: 4px solid #1596D1;"
                "border-radius: 8px;"
                "}"
            )
        else:
            self.setStyleSheet(
                "QWidget#profileRowItem {"
                "background: transparent;"
                "border: 1px solid transparent;"
                "border-radius: 8px;"
                "}"
            )
        if self._selected:
            self.title_label.setStyleSheet("color:#173A63; font-size:15px; font-weight:600;")
            self.subtitle_label.setStyleSheet("color:#34506E; font-size:13px;")
        elif self._enabled:
            self.title_label.setStyleSheet("color:#1E5EBE; font-size:15px;")
            self.subtitle_label.setStyleSheet("color:#2D5EAA; font-size:13px;")
        else:
            self.title_label.setStyleSheet("color:#6B7785; font-size:15px;")
            self.subtitle_label.setStyleSheet("color:#7E8A9C; font-size:13px;")

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._press_pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
            self.clicked.emit()
        super().mousePressEvent(event)

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.clicked.emit()
            self.doubleClicked.emit()
        super().mouseDoubleClickEvent(event)

    def mouseMoveEvent(self, event):
        if not (event.buttons() & Qt.LeftButton) or self._press_pos is None:
            super().mouseMoveEvent(event)
            return
        pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
        if (pos - self._press_pos).manhattanLength() >= QApplication.startDragDistance():
            self.dragRequested.emit()
            self._press_pos = None
        super().mouseMoveEvent(event)

    def eventFilter(self, obj, event):
        if obj in {self.title_label, self.subtitle_label}:
            if event.type() == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
                self._press_pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
                self.clicked.emit()
                return True
            if event.type() == QEvent.MouseButtonDblClick and event.button() == Qt.LeftButton:
                self.clicked.emit()
                self.doubleClicked.emit()
                return True
            if event.type() == QEvent.MouseMove and (event.buttons() & Qt.LeftButton) and self._press_pos is not None:
                pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
                if (pos - self._press_pos).manhattanLength() >= QApplication.startDragDistance():
                    self.dragRequested.emit()
                    self._press_pos = None
                    return True
        return super().eventFilter(obj, event)


class _ProfileRowListWidget(QListWidget):
    """单列表行配置控件：勾选启用，已启用项支持拖拽排序。"""

    enabledRowDropped = Signal(str, int)  # row_id, target_row

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.setDragEnabled(True)
        self.viewport().setAcceptDrops(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setDefaultDropAction(Qt.MoveAction)
        self.setDragDropMode(QAbstractItemView.DragDrop)
        self._set_drag_feedback(False)

    def start_drag_for_row_id(self, rid):
        rid = str(rid or "").strip()
        if not rid:
            return
        for row in range(self.count()):
            item = self.item(row)
            if item is not None and str(item.data(Qt.UserRole) or "").strip() == rid:
                self.setCurrentRow(row)
                self.startDrag(Qt.MoveAction)
                return

    def _set_drag_feedback(self, active: bool):
        self.setStyleSheet(
            "QListView { border: 1px solid rgba(0, 120, 212, 0.65); "
            "background: rgba(0, 120, 212, 0.06); }"
            if active else ""
        )

    def _enabled_count(self) -> int:
        count = 0
        for row in range(self.count()):
            item = self.item(row)
            if item is not None and bool(item.data(Qt.UserRole + 1)):
                count += 1
        return count

    def startDrag(self, supportedActions):
        item = self.currentItem()
        if item is None or not bool(item.data(Qt.UserRole + 1)):
            return
        rid = str(item.data(Qt.UserRole) or "").strip()
        if not rid:
            return
        mime = QMimeData()
        mime.setData("application/x-profile-enabled-row-id", rid.encode("utf-8"))
        drag = QDrag(self)
        drag.setMimeData(mime)
        self._set_drag_feedback(True)
        try:
            drag.exec(Qt.MoveAction)
        finally:
            self._set_drag_feedback(False)

    def dragEnterEvent(self, event):
        if event.mimeData().hasFormat("application/x-profile-enabled-row-id"):
            self._set_drag_feedback(True)
            event.acceptProposedAction()
            return
        super().dragEnterEvent(event)

    def dragLeaveEvent(self, event):
        self._set_drag_feedback(False)
        super().dragLeaveEvent(event)

    def dragMoveEvent(self, event):
        if event.mimeData().hasFormat("application/x-profile-enabled-row-id"):
            event.acceptProposedAction()
            return
        super().dragMoveEvent(event)

    def dropEvent(self, event):
        data = event.mimeData()
        if not data.hasFormat("application/x-profile-enabled-row-id"):
            super().dropEvent(event)
            return
        self._set_drag_feedback(False)
        try:
            rid = bytes(data.data("application/x-profile-enabled-row-id")).decode("utf-8").strip()
            if not rid:
                event.ignore()
                return
            enabled_count = self._enabled_count()
            pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
            row = self.indexAt(pos).row()
            if row < 0:
                row = enabled_count
            row = max(0, min(enabled_count, row))
            self.enabledRowDropped.emit(rid, row)
            event.acceptProposedAction()
        except Exception:
            event.ignore()


class _SingleListTextExportSettingsDialog(QDialog):
    """纵断面文字导出参数与行配置弹窗（单列表勾选 + 拖拽排序版）。"""

    _UI_SETTINGS_ORG = "SichuanShuifa"
    _UI_SETTINGS_APP = "HydroCalc"
    _UI_SIZE_W_KEY = "water_profile/text_export_dialog_width"
    _UI_SIZE_H_KEY = "water_profile/text_export_dialog_height"
    _ICON_COLLAPSED = None
    _ICON_EXPANDED = None
    _DESIGN_MIN_WIDTH = 960
    _DESIGN_MIN_HEIGHT = 500
    _MIN_SCREEN_MARGIN = 24

    def __init__(self, parent=None, defaults=None):
        super().__init__(parent)
        if self._ICON_COLLAPSED is None or self._ICON_EXPANDED is None:
            type(self)._ICON_COLLAPSED = _resolve_fluent_icon("CHEVRON_RIGHT_MED", "CHEVRON_RIGHT", "CHEVRON_DOWN_MED")
            type(self)._ICON_EXPANDED = _resolve_fluent_icon("CHEVRON_DOWN_MED", "CHEVRON_RIGHT_MED", "CHEVRON_RIGHT")
        self.setWindowTitle("纵断面文字导出设置")
        self._ui_settings = QSettings(self._UI_SETTINGS_ORG, self._UI_SETTINGS_APP)
        self._dialog_min_size = self._resolve_minimum_dialog_size()
        self.setMinimumSize(self._dialog_min_size)
        self._apply_initial_size()
        self.setSizeGripEnabled(True)
        self.setStyleSheet(DIALOG_STYLE + """
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f7f9fc, stop:1 #eef3fb);
            }
            QListView {
                border: 1px solid #d6dfef;
                border-radius: 10px;
                background: rgba(255,255,255,0.92);
                padding: 4px;
            }
            QListView::item {
                border-radius: 8px;
                padding: 7px 10px;
                margin: 1px 1px;
            }
            QListView::item:selected {
                background: rgba(0, 120, 212, 0.16);
                border: 1px solid rgba(0, 120, 212, 0.35);
            }
            QListView::item:hover {
                background: rgba(32, 97, 181, 0.08);
            }
        """)
        self.result = None
        self._row_updating = False

        defaults = _normalize_text_export_settings(defaults or {})
        self._defaults = dict(defaults)

        self._entries = {}
        self._ordered_row_ids = list(_PROFILE_ROW_VISIBLE_ORDER)
        self._enabled_row_ids = []
        self._row_widgets = {}
        self._compat_advanced_values = {
            key: defaults.get(key)
            for key in _PROFILE_RUNTIME_ADVANCED_KEYS
        }
        self._parameter_content_layout = None
        self._parameter_card = None
        self._parameter_left_section = None
        self._parameter_right_section = None
        self._runtime_rows_widget = None
        self._runtime_rows_layout = None
        self._runtime_row_labels = {}
        self._runtime_summary_label = None

        self._row_list = None
        self._body_scroll = None
        self._body_content_widget = None
        self._body_layout = None
        self._rows_card = None
        self._btn_reset = None
        self._btn_cancel = None
        self._btn_ok = None
        self._layout_refresh_pending = False

        self._init_ui()

    def _read_setting_bool(self, key, default=False):
        raw = self._ui_settings.value(key, default)
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, (int, float)):
            return bool(raw)
        text = str(raw).strip().lower()
        if text in {"1", "true", "yes", "on"}:
            return True
        if text in {"0", "false", "no", "off"}:
            return False
        return bool(default)

    def _read_setting_int(self, key, default_value):
        raw = self._ui_settings.value(key, default_value)
        try:
            return int(float(raw))
        except Exception:
            return int(default_value)

    def _available_geometry(self):
        screen = None
        parent_widget = self.parentWidget()
        if parent_widget is not None:
            parent_window = parent_widget.window()
            if parent_window is not None and parent_window.windowHandle() is not None:
                screen = parent_window.windowHandle().screen()
        if screen is None:
            app = QApplication.instance()
            if app is not None:
                screen = app.primaryScreen()
        return screen.availableGeometry() if screen is not None else None

    def _resolve_minimum_dialog_size(self):
        avail = self._available_geometry()
        if avail is None:
            return QSize(self._DESIGN_MIN_WIDTH, self._DESIGN_MIN_HEIGHT)

        width = min(self._DESIGN_MIN_WIDTH, max(640, avail.width() - self._MIN_SCREEN_MARGIN))
        height = min(self._DESIGN_MIN_HEIGHT, max(420, avail.height() - self._MIN_SCREEN_MARGIN))
        return QSize(width, height)

    def _apply_initial_size(self):
        avail = self._available_geometry()
        min_size = self._dialog_min_size
        if avail is not None:
            default_w = min(max(min_size.width(), int(avail.width() * 0.78)), 1360)
            default_h = min(max(min_size.height(), int(avail.height() * 0.72)), int(avail.height() * 0.92))
            max_w = max(min_size.width(), int(avail.width() * 0.96))
            max_h = max(min_size.height(), int(avail.height() * 0.92))
        else:
            default_w, default_h = 1160, 640
            max_w, max_h = 1400, 900

        width = self._read_setting_int(self._UI_SIZE_W_KEY, default_w)
        height = self._read_setting_int(self._UI_SIZE_H_KEY, default_h)
        width = max(min_size.width(), min(width, max_w))
        height = max(min_size.height(), min(height, max_h))
        self.resize(width, height)

    def minimumSizeHint(self):
        return QSize(self.minimumWidth(), self.minimumHeight())

    def sizeHint(self):
        avail = self._available_geometry()
        if avail is None:
            return QSize(max(self.minimumWidth(), 1160), max(self.minimumHeight(), 640))

        width = min(
            max(self.minimumWidth(), int(avail.width() * 0.78)),
            max(self.minimumWidth(), int(avail.width() * 0.96)),
        )
        height = min(
            max(self.minimumHeight(), int(avail.height() * 0.72)),
            max(self.minimumHeight(), int(avail.height() * 0.92)),
        )
        return QSize(width, height)

    def _persist_ui_state(self):
        size = self.size()
        self._ui_settings.setValue(self._UI_SIZE_W_KEY, int(size.width()))
        self._ui_settings.setValue(self._UI_SIZE_H_KEY, int(size.height()))

    def closeEvent(self, event):
        self._persist_ui_state()
        super().closeEvent(event)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._update_responsive_layout_mode()
        self._ensure_row_lists_visible_rows()
        self._request_dialog_layout_refresh(deferred=False)

    def _make_wrap_caption(self, text=""):
        return _AutoWrapCaptionLabel(text, self)

    def _init_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(14, 8, 14, 8)
        root.setSpacing(6)

        body_widget = QWidget(self)
        self._body_content_widget = body_widget
        self._body_layout = QVBoxLayout(body_widget)
        self._body_layout.setSizeConstraint(QLayout.SetMinAndMaxSize)
        self._body_layout.setContentsMargins(0, 0, 0, 0)
        self._body_layout.setSpacing(10)
        self._parameter_card = self._build_parameter_card()
        self._body_layout.addWidget(self._parameter_card, 0)
        self._rows_card = self._build_rows_card()
        self._body_layout.addWidget(self._rows_card, 0)
        self._body_layout.addStretch(1)

        self._body_scroll = QScrollArea(self)
        self._body_scroll.setWidgetResizable(True)
        self._body_scroll.setFrameShape(QFrame.NoFrame)
        self._body_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._body_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._body_scroll.setWidget(body_widget)
        root.addWidget(self._body_scroll, 1)

        btn_row = QHBoxLayout()
        self._btn_reset = PushButton("恢复默认")
        self._btn_reset.clicked.connect(self._reset_defaults)
        btn_row.addWidget(self._btn_reset)
        btn_row.addStretch(1)
        self._btn_cancel = PushButton("取消")
        self._btn_cancel.clicked.connect(self.reject)
        self._btn_ok = PrimaryPushButton("确定")
        self._btn_ok.clicked.connect(self._on_confirm)
        btn_row.addWidget(self._btn_cancel)
        btn_row.addWidget(self._btn_ok)
        root.addLayout(btn_row)

        self._load_rows(self._defaults.get("profile_row_items"))

        QShortcut(QKeySequence(Qt.Key_Escape), self, self.reject)
        QShortcut(QKeySequence(Qt.Key_Return), self, self._on_confirm)
        QShortcut(QKeySequence("Ctrl+Up"), self, lambda: self._move_selected_row(-1))
        QShortcut(QKeySequence("Ctrl+Down"), self, lambda: self._move_selected_row(1))
        QShortcut(QKeySequence("Ctrl+Home"), self, lambda: self._move_selected_row_to_edge(True))
        QShortcut(QKeySequence("Ctrl+End"), self, lambda: self._move_selected_row_to_edge(False))
        QShortcut(QKeySequence(Qt.Key_Delete), self, self._disable_selected_row)
        self._update_responsive_layout_mode()
        self._request_dialog_layout_refresh()

    def _run_dialog_layout_refresh(self):
        layouts = [
            self.layout(),
            self._body_layout,
            self._parameter_card.layout() if self._parameter_card is not None else None,
            self._rows_card.layout() if self._rows_card is not None else None,
            self._parameter_content_layout,
            self._runtime_rows_layout,
        ]
        widgets = [
            self,
            self._body_scroll,
            self._body_content_widget,
            self._parameter_card,
            self._rows_card,
            self._parameter_left_section,
            self._parameter_right_section,
            self._runtime_rows_widget,
            self._runtime_summary_label,
            self._row_list,
        ]

        for layout in layouts:
            if layout is not None:
                layout.invalidate()

        for widget in widgets:
            if widget is not None:
                widget.updateGeometry()

        if self._body_content_widget is not None:
            self._body_content_widget.adjustSize()
            layout = self._body_content_widget.layout()
            if layout is not None:
                layout.activate()

        if self._body_scroll is not None:
            self._body_scroll.updateGeometry()
            viewport = self._body_scroll.viewport()
            if viewport is not None:
                viewport.updateGeometry()

        for layout in layouts:
            if layout is not None:
                layout.activate()

    def _flush_deferred_dialog_layout_refresh(self):
        self._layout_refresh_pending = False
        self._run_dialog_layout_refresh()

    def _request_dialog_layout_refresh(self, *, deferred=True):
        self._run_dialog_layout_refresh()
        if deferred and not self._layout_refresh_pending:
            self._layout_refresh_pending = True
            QTimer.singleShot(0, self._flush_deferred_dialog_layout_refresh)

    def _update_responsive_layout_mode(self):
        if self._parameter_content_layout is None or self._body_scroll is None:
            return

        viewport_width = self._body_scroll.viewport().width()
        direction = QBoxLayout.TopToBottom if viewport_width and viewport_width < 1120 else QBoxLayout.LeftToRight
        if self._parameter_content_layout.direction() != direction:
            self._parameter_content_layout.setDirection(direction)

        if self._parameter_left_section is not None:
            self._parameter_left_section.setMaximumWidth(16777215 if direction == QBoxLayout.TopToBottom else 430)
        if self._parameter_content_layout.count() >= 2:
            self._parameter_content_layout.setStretch(0, 0 if direction == QBoxLayout.TopToBottom else 38)
            self._parameter_content_layout.setStretch(1, 1 if direction == QBoxLayout.TopToBottom else 62)

    def _build_parameter_card(self):
        card = ElevatedCardWidget(self)
        card_lay = QVBoxLayout(card)
        card_lay.setSizeConstraint(QLayout.SetMinAndMaxSize)
        card_lay.setContentsMargins(12, 10, 12, 10)
        card_lay.setSpacing(10)

        title = BodyLabel("参数设置")
        card_lay.addWidget(title)

        hint = self._make_wrap_caption("左侧设置基础参数，右侧实时查看当前已启用行的 Y 值与来源。")
        card_lay.addWidget(hint)

        content_widget = QWidget(self)
        self._parameter_content_layout = QBoxLayout(QBoxLayout.LeftToRight, content_widget)
        self._parameter_content_layout.setContentsMargins(0, 0, 0, 0)
        self._parameter_content_layout.setSpacing(10)

        self._parameter_left_section = QWidget(self)
        self._parameter_left_section.setObjectName("profileParameterSection")
        left_lay = QVBoxLayout(self._parameter_left_section)
        left_lay.setContentsMargins(12, 12, 12, 12)
        left_lay.setSpacing(8)
        left_lay.addWidget(BodyLabel("基础参数"))
        basic_form = QGridLayout()
        basic_form.setHorizontalSpacing(8)
        basic_form.setVerticalSpacing(8)
        basic_form.setColumnStretch(0, 0)
        basic_form.setColumnStretch(1, 0)
        basic_form.setColumnStretch(2, 1)
        self._add_entry_row(basic_form, 0, "字高", "text_height", "")
        self._add_entry_row(basic_form, 1, "旋转角度", "rotation", "")
        self._add_entry_row(basic_form, 2, "高程小数位数", "elev_decimals", "")
        self._add_entry_row(basic_form, 3, "X方向比例(1:N)", "scale_x", "如 1:1000 则输入 1000")
        self._add_entry_row(basic_form, 4, "Y方向比例(1:N)", "scale_y", "如 1:1000 则输入 1000")
        left_lay.addLayout(basic_form)
        left_lay.addStretch(1)

        self._parameter_right_section = QWidget(self)
        self._parameter_right_section.setObjectName("profileParameterSection")
        right_lay = QVBoxLayout(self._parameter_right_section)
        right_lay.setSizeConstraint(QLayout.SetMinAndMaxSize)
        right_lay.setContentsMargins(12, 12, 12, 12)
        right_lay.setSpacing(8)
        right_lay.addWidget(BodyLabel("启用行实时参数"))

        runtime_hint = self._make_wrap_caption("只显示当前已启用项；顺序、启停和拖拽变化会立即同步到这里。")
        right_lay.addWidget(runtime_hint)

        runtime_headers = QGridLayout()
        runtime_headers.setHorizontalSpacing(8)
        runtime_headers.setVerticalSpacing(4)
        hdr_name = CaptionLabel("行内容")
        hdr_value = CaptionLabel("实时Y")
        hdr_source = CaptionLabel("来源")
        runtime_headers.addWidget(hdr_name, 0, 0)
        runtime_headers.addWidget(hdr_value, 0, 1)
        runtime_headers.addWidget(hdr_source, 0, 2)
        runtime_headers.setColumnStretch(0, 3)
        runtime_headers.setColumnStretch(1, 0)
        runtime_headers.setColumnStretch(2, 2)
        right_lay.addLayout(runtime_headers)

        runtime_rows_widget = QWidget(self)
        self._runtime_rows_widget = runtime_rows_widget
        self._runtime_rows_layout = QGridLayout(runtime_rows_widget)
        self._runtime_rows_layout.setContentsMargins(0, 0, 0, 0)
        self._runtime_rows_layout.setHorizontalSpacing(8)
        self._runtime_rows_layout.setVerticalSpacing(6)
        self._runtime_rows_layout.setColumnStretch(0, 3)
        self._runtime_rows_layout.setColumnStretch(1, 0)
        self._runtime_rows_layout.setColumnStretch(2, 2)
        right_lay.addWidget(runtime_rows_widget)

        self._runtime_summary_label = self._make_wrap_caption("")
        right_lay.addWidget(self._runtime_summary_label)
        right_lay.addStretch(1)

        self._parameter_content_layout.addWidget(self._parameter_left_section, 38)
        self._parameter_content_layout.addWidget(self._parameter_right_section, 62)
        card_lay.addWidget(content_widget)

        card.setStyleSheet(card.styleSheet() + """
            QWidget#profileParameterSection {
                background: rgba(255,255,255,0.74);
                border: 1px solid rgba(209,219,231,0.85);
                border-radius: 14px;
            }
        """)
        return card

    def _build_rows_card(self):
        card = ElevatedCardWidget(self)
        card_lay = QVBoxLayout(card)
        card_lay.setSizeConstraint(QLayout.SetMinAndMaxSize)
        card_lay.setContentsMargins(12, 10, 12, 10)
        card_lay.setSpacing(6)

        title_row = QHBoxLayout()
        title_row.addWidget(BodyLabel(
            f"纵断面行内容（{len(_PROFILE_ROW_VISIBLE_ORDER)}项可选，勾选启用，已启用项可拖动排序）"
        ))
        title_row.addStretch(1)
        btn_preset = PushButton("应用亭子口二期项建/可研阶段模板")
        btn_preset.clicked.connect(self._apply_tingzikou_preset)
        title_row.addWidget(btn_preset)
        card_lay.addLayout(title_row)

        quick_row = QHBoxLayout()
        btn_enable_all = PushButton("全启用")
        btn_enable_all.clicked.connect(self._enable_all_rows)
        btn_disable_all = PushButton("全停用")
        btn_disable_all.clicked.connect(self._disable_all_rows)
        btn_restore_recommended = PushButton("恢复推荐")
        btn_restore_recommended.clicked.connect(self._restore_recommended_rows)
        quick_row.addWidget(btn_enable_all)
        quick_row.addWidget(btn_disable_all)
        quick_row.addWidget(btn_restore_recommended)
        quick_row.addStretch(1)
        card_lay.addLayout(quick_row)

        hint = self._make_wrap_caption(
            "操作说明：勾选即启用；拖动已启用项即可排序；右键支持启用/停用/置顶/置底；Ctrl+Up/Ctrl+Down 可微调顺序。"
        )
        card_lay.addWidget(hint)

        hidden_hint = self._make_wrap_caption(
            "本版本暂不显示：IP文字(BE)、桩号文字(BK)，避免与 IP点名称、里程桩号重复。"
        )
        card_lay.addWidget(hidden_hint)

        self._row_list = _ProfileRowListWidget(self)
        self._row_list.enabledRowDropped.connect(self._on_enabled_row_dropped)
        self._row_list.itemDoubleClicked.connect(lambda _item: self._toggle_current_row())
        self._row_list.currentItemChanged.connect(lambda _current, _previous: self._update_row_widget_selection())
        self._row_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self._row_list.customContextMenuRequested.connect(self._show_row_context_menu)
        card_lay.addWidget(self._row_list, 1)

        sort_row = QHBoxLayout()
        btn_up = PushButton("上移")
        btn_up.clicked.connect(lambda: self._move_selected_row(-1))
        btn_down = PushButton("下移")
        btn_down.clicked.connect(lambda: self._move_selected_row(1))
        btn_top = PushButton("置顶")
        btn_top.clicked.connect(lambda: self._move_selected_row_to_edge(True))
        btn_bottom = PushButton("置底")
        btn_bottom.clicked.connect(lambda: self._move_selected_row_to_edge(False))
        sort_row.addWidget(btn_up)
        sort_row.addWidget(btn_down)
        sort_row.addWidget(btn_top)
        sort_row.addWidget(btn_bottom)
        sort_row.addStretch(1)
        card_lay.addLayout(sort_row)
        return card

    def _add_entry_row(self, layout, row, label, key, hint):
        layout.addWidget(QLabel(f"{label}:"), row, 0)
        entry = LineEdit()
        entry.setText(str(self._defaults.get(key, "")))
        entry.setFixedWidth(130)
        layout.addWidget(entry, row, 1)
        layout.addWidget(CaptionLabel(hint), row, 2)
        self._entries[key] = entry

    def _build_runtime_view_input_settings(self):
        settings = dict(self._defaults)
        settings.update(self._compat_advanced_values)
        settings["profile_row_items"] = self._row_data_from_table()

        for key in ("text_height", "rotation", "elev_decimals", "scale_x", "scale_y"):
            entry = self._entries.get(key)
            if entry is None:
                continue
            txt = entry.text().strip()
            try:
                value = float(txt)
                if key == "elev_decimals":
                    value = int(value)
                settings[key] = value
            except Exception:
                # 输入中的临时非法值不打断实时预览，保留旧值。
                continue
        return settings

    def _clear_layout_widgets(self, layout):
        if layout is None:
            return
        while layout.count():
            item = layout.takeAt(0)
            child_layout = item.layout()
            child_widget = item.widget()
            if child_layout is not None:
                self._clear_layout_widgets(child_layout)
            if child_widget is not None:
                child_widget.deleteLater()

    def _make_runtime_value_chip(self, text):
        label = QLabel(text)
        label.setAlignment(Qt.AlignCenter)
        label.setMinimumWidth(76)
        label.setStyleSheet(
            "color:#173A63; background: rgba(255,255,255,0.94);"
            "border: 1px solid rgba(198,210,224,0.92); border-radius: 8px;"
            "padding: 3px 10px; font-family:'Consolas','Microsoft YaHei'; font-size:12px;"
        )
        return label

    def _refresh_runtime_advanced_view(self):
        if self._runtime_rows_layout is None:
            return
        runtime = _compute_runtime_advanced_parameter_view(self._build_runtime_view_input_settings())
        enabled_rows = list(runtime.get("enabled_runtime_rows") or [])
        self._runtime_row_labels = {}
        self._clear_layout_widgets(self._runtime_rows_layout)

        if not enabled_rows:
            empty_label = self._make_wrap_caption("\u5f53\u524d\u5c1a\u672a\u542f\u7528\u4efb\u4f55\u7eb5\u65ad\u9762\u884c\u3002")
            self._runtime_rows_layout.addWidget(empty_label, 0, 0, 1, 3)
        else:
            for row_index, row in enumerate(enabled_rows):
                display_label = QLabel(f"{row['order']:02d}. {row['label']}")
                display_label.setStyleSheet("color:#24384D; font-size:13px; font-weight:600;")
                value_label = self._make_runtime_value_chip(_format_number(row["text_y"]))
                source_label = self._make_wrap_caption(row.get("source_label", ""))
                self._runtime_rows_layout.addWidget(display_label, row_index, 0)
                self._runtime_rows_layout.addWidget(value_label, row_index, 1)
                self._runtime_rows_layout.addWidget(source_label, row_index, 2)
                self._runtime_row_labels[row["id"]] = {
                    "title": display_label,
                    "value": value_label,
                    "source": source_label,
                }

        line_row = len(enabled_rows)
        divider = QFrame(self)
        divider.setFrameShape(QFrame.HLine)
        divider.setFrameShadow(QFrame.Sunken)
        divider.setStyleSheet("color: rgba(210,218,229,0.95);")
        self._runtime_rows_layout.addWidget(divider, line_row, 0, 1, 3)

        line_height_row = line_row + 1
        line_height_title = QLabel("\u751f\u6548\u7ad6\u7ebf\u9ad8\u5ea6")
        line_height_title.setStyleSheet("color:#24384D; font-size:13px; font-weight:600;")
        line_height_value = self._make_runtime_value_chip(_format_number(runtime["line_height"]))
        line_height_source = self._make_wrap_caption("max(\u5185\u5bb9\u603b\u9ad8, \u6700\u5c0f\u7ad6\u7ebf\u53c2\u6570)")
        self._runtime_rows_layout.addWidget(line_height_title, line_height_row, 0)
        self._runtime_rows_layout.addWidget(line_height_value, line_height_row, 1)
        self._runtime_rows_layout.addWidget(line_height_source, line_height_row, 2)
        self._runtime_row_labels["y_line_height"] = {
            "title": line_height_title,
            "value": line_height_value,
            "source": line_height_source,
        }

        if self._runtime_summary_label is not None:
            self._runtime_summary_label.setText(
                f"\u5b9e\u65f6\u6c47\u603b\uff1a\u542f\u7528 {len(runtime['enabled_row_ids'])} \u9879 / "
                f"\u5185\u5bb9\u603b\u9ad8 {_format_number(runtime['total_height'])} / "
                f"\u751f\u6548\u7ad6\u7ebf\u9ad8\u5ea6 {_format_number(runtime['line_height'])} / "
                f"\u6700\u5c0f\u7ad6\u7ebf\u53c2\u6570 {_format_number(runtime['min_line_height'])}"
            )
        self._request_dialog_layout_refresh()

    def _selected_row_id(self):
        if not self._row_list:
            return ""
        item = self._row_list.currentItem()
        if item is None:
            return ""
        return str(item.data(Qt.UserRole) or "").strip()

    def _set_current_row_id(self, rid):
        if not self._row_list or not rid:
            return
        for row in range(self._row_list.count()):
            item = self._row_list.item(row)
            if item is not None and str(item.data(Qt.UserRole) or "").strip() == rid:
                self._row_list.setCurrentRow(row)
                return

    def _create_row_item(self, rid, order_index):
        enabled = rid in self._enabled_row_ids
        row_def = _PROFILE_ROW_DEF_MAP[rid]
        label = row_def["label"]
        if enabled:
            title = f"{order_index + 1:02d}. {label}  [拖动排序]"
            status = "已启用"
        else:
            title = f"--. {label}"
            status = "未启用"
        if rid in _PROFILE_RECOMMENDED_ROW_IDS:
            title += "  ★推荐"
        detail = row_def.get("hint", "")
        subtitle = f"{status} | {detail}" if detail else status
        item = QListWidgetItem()
        item.setData(Qt.UserRole, rid)
        item.setData(Qt.UserRole + 1, enabled)
        item.setData(Qt.UserRole + 2, title)
        item.setData(Qt.UserRole + 3, subtitle)
        flags = Qt.ItemIsSelectable | Qt.ItemIsEnabled
        if enabled:
            flags |= Qt.ItemIsDragEnabled
        item.setFlags(flags)
        item.setSizeHint(QSize(0, 56))
        if enabled:
            item.setForeground(QColor("#174EA6") if rid in _PROFILE_RECOMMENDED_ROW_IDS else QColor("#1F2D3D"))
        else:
            item.setForeground(QColor("#6B7785"))
        return item

    def _create_row_widget(self, item):
        rid = str(item.data(Qt.UserRole) or "").strip()
        enabled = bool(item.data(Qt.UserRole + 1))
        title = str(item.data(Qt.UserRole + 2) or "")
        subtitle = str(item.data(Qt.UserRole + 3) or "")

        widget = _ProfileRowItemWidget(title, subtitle, enabled, self._row_list)
        widget.checkbox.stateChanged.connect(
            lambda _state, row_item=item, row_id=rid: self._on_row_widget_checkbox_changed(row_item, row_id)
        )
        widget.clicked.connect(lambda row_id=rid: self._set_current_row_id(row_id))
        widget.doubleClicked.connect(self._toggle_current_row)
        widget.dragRequested.connect(lambda row_id=rid: self._row_list.start_drag_for_row_id(row_id))
        return widget

    def _on_row_widget_checkbox_changed(self, item, rid):
        if self._row_updating or item is None:
            return
        self._set_current_row_id(rid)
        widget = self._row_widgets.get(rid)
        if widget is None:
            return
        self._set_row_enabled(rid, widget.checkbox.isChecked(), show_feedback=True)

    def _update_row_widget_selection(self):
        if not self._row_list:
            return
        current = self._row_list.currentItem()
        current_rid = str(current.data(Qt.UserRole) or "").strip() if current is not None else ""
        for rid, widget in self._row_widgets.items():
            if widget is not None:
                widget.set_selected(rid == current_rid)

    def _normalize_row_model(self):
        enabled = [rid for rid in self._enabled_row_ids if rid in _PROFILE_ROW_VISIBLE_ID_SET]
        order = [rid for rid in self._ordered_row_ids if rid in _PROFILE_ROW_VISIBLE_ID_SET]
        for rid in _PROFILE_ROW_VISIBLE_ORDER:
            if rid not in order:
                order.append(rid)
        disabled = [rid for rid in order if rid not in enabled]
        self._enabled_row_ids = enabled
        self._ordered_row_ids = enabled + disabled

    def _refresh_row_list(self):
        if not self._row_list:
            return
        self._normalize_row_model()
        keep_current = self._selected_row_id()

        self._row_updating = True
        try:
            self._row_list.clear()
            self._row_widgets = {}
            enabled_index = 0
            for rid in self._ordered_row_ids:
                item = self._create_row_item(rid, enabled_index)
                self._row_list.addItem(item)
                widget = self._create_row_widget(item)
                self._row_widgets[rid] = widget
                self._row_list.setItemWidget(item, widget)
                if rid in self._enabled_row_ids:
                    enabled_index += 1
        finally:
            self._row_updating = False

        self._ensure_row_list_visible_rows()
        self._set_current_row_id(keep_current)
        self._update_row_widget_selection()
        self._refresh_runtime_advanced_view()
        self._request_dialog_layout_refresh()

    def _ensure_row_list_visible_rows(self):
        if not self._row_list:
            return
        row_h = self._row_list.sizeHintForRow(0)
        if row_h <= 0:
            row_h = 50
        visible_rows = min(max(10, len(_PROFILE_ROW_VISIBLE_ORDER)), len(_PROFILE_ROW_VISIBLE_ORDER))
        target_h = row_h * visible_rows + 12
        self._row_list.setMinimumHeight(target_h)
        self._row_list.setMaximumHeight(target_h)

    def _load_rows(self, row_items):
        normalized = _normalize_profile_row_items(row_items)
        self._ordered_row_ids = [item["id"] for item in normalized]
        self._enabled_row_ids = [item["id"] for item in normalized if item.get("enabled")]
        self._refresh_row_list()

    def _row_data_from_table(self):
        self._normalize_row_model()
        enabled = set(self._enabled_row_ids)
        return _normalize_profile_row_items([
            {"id": rid, "enabled": rid in enabled}
            for rid in self._ordered_row_ids
        ])

    def _set_row_enabled(self, rid, enabled, *, show_feedback=False):
        if rid not in _PROFILE_ROW_VISIBLE_ID_SET:
            return
        current_enabled = rid in self._enabled_row_ids
        if current_enabled == bool(enabled):
            return

        if enabled:
            self._enabled_row_ids = [row_id for row_id in self._enabled_row_ids if row_id != rid] + [rid]
        else:
            self._enabled_row_ids = [row_id for row_id in self._enabled_row_ids if row_id != rid]
        self._normalize_row_model()
        self._refresh_row_list()
        self._set_current_row_id(rid)

        if show_feedback:
            row_label = _PROFILE_ROW_DEF_MAP[rid]["label"]
            if enabled:
                InfoBar.success(
                    "已启用",
                    f"{row_label} 已加入导出",
                    parent=self,
                    position=InfoBarPosition.TOP_RIGHT,
                    duration=1200,
                )
            else:
                InfoBar.info(
                    "已停用",
                    f"{row_label} 已移出导出",
                    parent=self,
                    position=InfoBarPosition.TOP_RIGHT,
                    duration=1200,
                )

    def _toggle_current_row(self):
        rid = self._selected_row_id()
        if not rid:
            return
        self._set_row_enabled(rid, rid not in self._enabled_row_ids, show_feedback=True)

    def _on_row_item_changed(self, item):
        if self._row_updating or item is None:
            return
        rid = str(item.data(Qt.UserRole) or "").strip()
        enabled = item.checkState() == Qt.Checked
        self._set_row_enabled(rid, enabled, show_feedback=True)

    def _enable_all_rows(self):
        self._enabled_row_ids = list(_PROFILE_ROW_VISIBLE_ORDER)
        self._refresh_row_list()

    def _disable_all_rows(self):
        self._enabled_row_ids = []
        self._refresh_row_list()

    def _restore_recommended_rows(self):
        self._enabled_row_ids = [
            rid for rid in _PROFILE_ROW_VISIBLE_ORDER
            if rid in _PROFILE_RECOMMENDED_ROW_IDS
        ]
        self._refresh_row_list()

    def _apply_tingzikou_preset(self):
        ordered = list(_TINGZIKOU_TEMPLATE_ROW_IDS) + [
            rid for rid in _PROFILE_ROW_VISIBLE_ORDER if rid not in _TINGZIKOU_TEMPLATE_ROW_IDS
        ]
        self._ordered_row_ids = ordered
        self._enabled_row_ids = list(_TINGZIKOU_TEMPLATE_ROW_IDS)
        self._refresh_row_list()
        InfoBar.success(
            "模板已应用",
            "已切换为亭子口推荐顺序",
            parent=self,
            position=InfoBarPosition.TOP_RIGHT,
            duration=1500,
        )

    def _reorder_enabled_row(self, rid, target_row):
        enabled = list(self._enabled_row_ids)
        if rid not in enabled:
            return
        old_row = enabled.index(rid)
        enabled.pop(old_row)
        target_row = max(0, min(len(enabled), int(target_row)))
        if target_row > old_row:
            target_row -= 1
        enabled.insert(target_row, rid)
        self._enabled_row_ids = enabled
        self._refresh_row_list()
        self._set_current_row_id(rid)

    def _on_enabled_row_dropped(self, rid, target_row):
        if self._row_updating:
            return
        self._reorder_enabled_row(rid, target_row)

    def _move_selected_row(self, delta):
        rid = self._selected_row_id()
        if not rid or rid not in self._enabled_row_ids:
            return
        row = self._enabled_row_ids.index(rid)
        target = row + int(delta)
        if target < 0 or target >= len(self._enabled_row_ids):
            return
        # _reorder_enabled_row() 接收的是“原列表中的插入位置”；
        # 向下移动时需要插入到目标项之后，避免弹出后又插回原位。
        insertion_row = target + 1 if delta > 0 else target
        self._reorder_enabled_row(rid, insertion_row)

    def _move_selected_row_to_edge(self, to_top):
        rid = self._selected_row_id()
        if not rid or rid not in self._enabled_row_ids:
            return
        target = 0 if to_top else len(self._enabled_row_ids) - 1
        self._reorder_enabled_row(rid, target)

    def _disable_selected_row(self):
        rid = self._selected_row_id()
        if rid and rid in self._enabled_row_ids:
            self._set_row_enabled(rid, False, show_feedback=True)

    def _show_row_context_menu(self, pos):
        if not self._row_list:
            return
        item = self._row_list.itemAt(pos)
        if item is None:
            return
        self._row_list.setCurrentItem(item)
        rid = str(item.data(Qt.UserRole) or "").strip()
        enabled = rid in self._enabled_row_ids
        menu = QMenu(self)
        action_toggle = menu.addAction("停用" if enabled else "启用")
        action_up = action_down = action_top = action_bottom = None
        if enabled:
            menu.addSeparator()
            action_up = menu.addAction("上移")
            action_down = menu.addAction("下移")
            action_top = menu.addAction("置顶")
            action_bottom = menu.addAction("置底")

            row = self._enabled_row_ids.index(rid)
            action_up.setEnabled(row > 0)
            action_top.setEnabled(row > 0)
            action_down.setEnabled(row < len(self._enabled_row_ids) - 1)
            action_bottom.setEnabled(row < len(self._enabled_row_ids) - 1)

        chosen = menu.exec(self._row_list.viewport().mapToGlobal(pos))
        if chosen == action_toggle:
            self._set_row_enabled(rid, not enabled, show_feedback=True)
        elif chosen == action_up:
            self._move_selected_row(-1)
        elif chosen == action_down:
            self._move_selected_row(1)
        elif chosen == action_top:
            self._move_selected_row_to_edge(True)
        elif chosen == action_bottom:
            self._move_selected_row_to_edge(False)

    def _reset_defaults(self):
        original = {
            "y_bottom": 1, "y_top": 31, "y_water": 16,
            "text_height": 3.5, "rotation": 90, "elev_decimals": 3,
            "y_name": 115, "y_slope": 105, "y_ip": 77,
            "y_station": 47, "y_line_height": 120,
            "scale_x": 2000, "scale_y": 1000,
        }
        for key, value in original.items():
            if key in self._entries:
                self._entries[key].setText(str(value))
        for key in _PROFILE_RUNTIME_ADVANCED_KEYS:
            self._compat_advanced_values[key] = original.get(key)
        self._load_rows(_default_profile_row_items())

    def _focus_invalid_entry(self, key):
        entry = self._entries.get(key)
        if not entry:
            return
        entry.setFocus()
        entry.selectAll()

    def _on_confirm(self):
        try:
            parsed = {}
            ordered_keys = [
                "text_height", "rotation", "elev_decimals", "scale_x", "scale_y",
            ]
            labels = {
                "text_height": "字高",
                "rotation": "旋转角度",
                "elev_decimals": "高程小数位数",
                "scale_x": "X方向比例",
                "scale_y": "Y方向比例",
            }
            for key in ordered_keys:
                entry = self._entries[key]
                txt = entry.text().strip()
                if not txt:
                    self._focus_invalid_entry(key)
                    raise ValueError(f"{labels[key]}不能为空")
                try:
                    val = float(txt)
                except ValueError:
                    self._focus_invalid_entry(key)
                    raise ValueError(f"{labels[key]}必须为数值")
                if key == "elev_decimals":
                    if val < 0 or val != int(val):
                        self._focus_invalid_entry(key)
                        raise ValueError("高程小数位数必须为非负整数")
                    val = int(val)
                if key in ("scale_x", "scale_y") and val <= 0:
                    self._focus_invalid_entry(key)
                    raise ValueError("比例必须大于0")
                parsed[key] = val

            row_items = self._row_data_from_table()
            if not any(item.get("enabled") for item in row_items):
                if self._row_list is not None:
                    self._row_list.setFocus()
                raise ValueError("至少选择1项行内容")

            runtime_input = dict(self._defaults)
            runtime_input.update(self._compat_advanced_values)
            runtime_input.update(parsed)
            runtime_input["profile_row_items"] = row_items
            runtime = _compute_runtime_advanced_parameter_view(runtime_input)

            compatibility_values = {}
            runtime_values = runtime["legacy_writeback_values"]
            runtime_enabled_state = runtime["legacy_enabled_state"]
            for key in _PROFILE_RUNTIME_ADVANCED_KEYS:
                if key == "y_line_height":
                    compatibility_values[key] = float(runtime_values.get(key, self._compat_advanced_values.get(key, 120)))
                    continue
                if runtime_enabled_state.get(key):
                    compatibility_values[key] = float(runtime_values.get(key))
                else:
                    compatibility_values[key] = self._compat_advanced_values.get(key, self._defaults.get(key))

            result = dict(self._defaults)
            result.update(parsed)
            result.update(compatibility_values)
            result["profile_row_items"] = row_items
            self.result = _normalize_text_export_settings(result)
            self.accept()
        except ValueError as e:
            fluent_error(self, "输入错误", f"请输入有效的数值\n{str(e)}")

# ================================================================
# 辅助工具函数
# ================================================================

def _resolve_fluent_icon(*names):
    """按候选名称顺序获取可用 FluentIcon，避免版本差异导致属性不存在。"""
    for name in names:
        icon = getattr(FluentIcon, name, None)
        if icon is not None:
            return icon
    for fallback_name in ("CHEVRON_RIGHT", "CHEVRON_DOWN_MED", "ADD"):
        icon = getattr(FluentIcon, fallback_name, None)
        if icon is not None:
            return icon
    raise AttributeError("未找到可用的 FluentIcon 回退图标")


def _format_number(value):
    """格式化数值：保留完整精度，去除无意义的尾零"""
    return f"{value:.15g}"


def _get_building_display_name(node):
    """获取纵断面用的建筑物名称显示"""
    struct_str = node.get_structure_type_str() or ""
    if node.is_transition or struct_str == "渐变段":
        return ""
    if getattr(node, "is_auto_inserted_channel", False):
        return ""
    if struct_str.startswith("明渠"):
        return struct_str
    if struct_str == "矩形暗涵":
        return struct_str
    # 隧洞/倒虹吸/有压管道/渡槽 等特殊建筑物：只有进/出节点参与建筑物名称段的划定；
    # 内部 IP 节点不标注名称，否则 building_segments 会被碎化导致坡降行出现重叠。
    if _is_special_structure_sv(getattr(node, "structure_type", None)):
        if _in_out_val(getattr(node, "in_out", None)) not in ("进", "出"):
            return ""
    if node.name:
        category = struct_str.split("-")[0]
        return f"{node.name}{category}"
    return struct_str.split("-")[0] if struct_str else ""


def _estimate_text_width(text, text_height):
    """估算 AutoCAD 中文字的总宽度（用于居中对齐）

    中文字符（CJK）宽度 ≈ text_height
    ASCII 字符（字母/数字/标点）宽度 ≈ text_height × 0.7
    """
    width = 0.0
    for ch in text:
        if ord(ch) > 127:
            width += text_height
        else:
            width += text_height * 0.7
    return width


def _format_slope_text(slope_i):
    """格式化坡降为显示文本"""
    if slope_i is not None and slope_i > 0:
        slope_inv = round(1.0 / slope_i)
        return f"1/{slope_inv}"
    return "/"


def _get_node_slope_text(node, next_node=None):
    """获取节点坡降文本（直接使用节点自身的 slope_i）。"""
    return _format_slope_text(getattr(node, 'slope_i', None))


def _struct_val(struct_type):
    """获取 StructureType 的字符串值（兼容双路径导入的 enum 实例）"""
    if struct_type is None:
        return ""
    return struct_type.value if hasattr(struct_type, 'value') else str(struct_type)


def _allows_optional_blank_name(struct_type):
    """判断结构是否允许建筑物名称留空。"""
    if not MODELS_AVAILABLE:
        return False
    return StructureType.allows_empty_name(struct_type)


def _collect_optional_blank_name_rows(nodes):
    rows = []
    for idx, node in enumerate(nodes or [], start=1):
        if getattr(node, "is_transition", False) or getattr(node, "is_auto_inserted_channel", False):
            continue
        if not _allows_optional_blank_name(getattr(node, "structure_type", None)):
            continue
        if str(getattr(node, "name", "") or "").strip():
            continue
        rows.append((idx, _struct_val(getattr(node, "structure_type", None)) or "明渠"))
    return rows


def _build_optional_blank_name_notice(nodes, *, action_name):
    rows = _collect_optional_blank_name_rows(nodes)
    if not rows:
        return ""
    preview = "；".join(f"第{idx}行（{struct_name}）" for idx, struct_name in rows[:8])
    if len(rows) > 8:
        preview += f" 等{len(rows)}行"
    return (
        f"检测到部分建筑物名称为空，已按结构形式/占位符参与{action_name}，不影响本次处理：\n"
        f"{preview}"
    )


def _in_out_val(in_out):
    """获取 InOutType 的字符串值（兼容双路径导入的 enum 实例）"""
    if in_out is None:
        return ""
    return in_out.value if hasattr(in_out, 'value') else str(in_out)


def _is_special_structure_sv(struct_type):
    """判断是否为特殊建筑物（隧洞/倒虹吸/有压管道/渡槽/矩形暗涵），使用字符串值比较

    避免双路径导入导致 enum 实例比较失败"""
    sv = _struct_val(struct_type)
    return any(k in sv for k in ("隧洞", "倒虹吸", "有压管道", "渡槽", "暗涵", "定向钻", "顶管"))


_PROFILE_ROW_DEFS = [
    {
        "id": "building_name",
        "label": "建筑物名称",
        "hint": "按建筑物段居中标注",
        "header_lines": ["建筑物名称"],
        "height": 10.0,
        "anchor": "center",
    },
    {
        "id": "slope",
        "label": "坡降",
        "hint": "按建筑物段显示坡降",
        "header_lines": ["坡降"],
        "height": 10.0,
        "anchor": "center",
    },
    {
        "id": "ip_name",
        "label": "IP点名称",
        "hint": "IP节点名称（特殊建筑仅进/出点）",
        "header_lines": ["IP点名称"],
        "height": 40.0,
        "anchor": "bottom2",
    },
    {
        "id": "station",
        "label": "里程桩号(千米+米)",
        "hint": "显示格式：1+234.567",
        "header_lines": ["里程桩号", "（千米+米）"],
        "height": 30.0,
        "anchor": "bottom2",
    },
    {
        "id": "top_elev",
        "label": "渠顶高程(m)",
        "hint": "节点渠顶高程",
        "header_lines": ["渠顶高程(m)"],
        "height": 15.0,
        "anchor": "bottom1",
    },
    {
        "id": "water_elev",
        "label": "设计水位(m)",
        "hint": "节点设计水位",
        "header_lines": ["设计水位(m)"],
        "height": 15.0,
        "anchor": "bottom1",
    },
    {
        "id": "bottom_elev",
        "label": "渠底高程(m)",
        "hint": "节点渠底高程",
        "header_lines": ["渠底高程(m)"],
        "height": 15.0,
        "anchor": "bottom1",
    },
    {
        "id": "bd_ip_before",
        "label": "IP弯前(BD)",
        "hint": "IP文字弯前点（BC）",
        "header_lines": ["IP弯前"],
        "height": 40.0,
        "anchor": "bottom2",
    },
    {
        "id": "be_ip_text",
        "label": "IP文字(BE)",
        "hint": "IP文字中心点（MC）",
        "header_lines": ["IP文字"],
        "height": 30.0,
        "anchor": "bottom2",
    },
    {
        "id": "bf_ip_after",
        "label": "IP弯后(BF)",
        "hint": "IP文字弯后点（EC）",
        "header_lines": ["IP弯后"],
        "height": 40.0,
        "anchor": "bottom2",
    },
    {
        "id": "bj_station_before",
        "label": "桩号文字弯前(BJ)",
        "hint": "桩号文字弯前点（BC）",
        "header_lines": ["桩号文字弯前"],
        "height": 30.0,
        "anchor": "bottom2",
    },
    {
        "id": "bk_station",
        "label": "桩号文字(BK)",
        "hint": "桩号文字中心点（MC）",
        "header_lines": ["桩号文字"],
        "height": 25.0,
        "anchor": "bottom2",
    },
    {
        "id": "bl_station_after",
        "label": "桩号文字弯后(BL)",
        "hint": "桩号文字弯后点（EC）",
        "header_lines": ["桩号文字弯后"],
        "height": 30.0,
        "anchor": "bottom2",
    },
]
_PROFILE_ROW_DEF_MAP = {d["id"]: d for d in _PROFILE_ROW_DEFS}
_PROFILE_ROW_DEFAULT_ORDER = [d["id"] for d in _PROFILE_ROW_DEFS]
_PROFILE_ROW_VISIBLE_ORDER = [
    "building_name",
    "slope",
    "ip_name",
    "station",
    "top_elev",
    "water_elev",
    "bottom_elev",
    "bd_ip_before",
    # "be_ip_text",   # 暂停展示：与“IP点名称”语义重复
    "bf_ip_after",
    "bj_station_before",
    # "bk_station",   # 暂停展示：与“里程桩号”语义重复
    "bl_station_after",
]
_PROFILE_ROW_VISIBLE_ID_SET = frozenset(_PROFILE_ROW_VISIBLE_ORDER)
_PROFILE_ROW_HIDDEN_IDS = frozenset(
    rid for rid in _PROFILE_ROW_DEFAULT_ORDER if rid not in _PROFILE_ROW_VISIBLE_ID_SET
)
_TINGZIKOU_TEMPLATE_ROW_IDS = [
    "building_name", "slope", "ip_name", "station",
    "top_elev", "water_elev", "bottom_elev",
]
_PROFILE_RECOMMENDED_ROW_IDS = {
    "building_name", "slope", "top_elev", "water_elev", "bottom_elev"
}
_PROFILE_EXTENDED_ROW_IDS = [rid for rid in _PROFILE_ROW_VISIBLE_ORDER if rid not in _TINGZIKOU_TEMPLATE_ROW_IDS]
_PROFILE_RUNTIME_ADVANCED_ROW_BINDINGS = {
    "y_bottom": "bottom_elev",
    "y_top": "top_elev",
    "y_water": "water_elev",
    "y_name": "building_name",
    "y_slope": "slope",
    "y_ip": "ip_name",
    "y_station": "station",
}
_PROFILE_RUNTIME_ADVANCED_KEYS = (
    "y_bottom",
    "y_top",
    "y_water",
    "y_name",
    "y_slope",
    "y_ip",
    "y_station",
    "y_line_height",
)
_PROFILE_RUNTIME_ANCHOR_LABELS = {
    "center": "中线",
    "bottom1": "底+1",
    "bottom2": "底+2",
}
_XXPIPE_PROFILE_RUNTIME_ADVANCED_ROW_BINDINGS = {
    "y_name": "building_name",
    "y_ip": "ip_name",
    "y_station": "station",
}
_XXPIPE_PROFILE_ROW_DEFS = [
    {
        "id": "building_name",
        "label": "建筑物名称",
        "hint": "仅隧洞/定向钻/顶管段显示",
        "header_lines": ["建筑物名称"],
        "height": 20.0,
        "anchor": "center",
    },
    {
        "id": "ip_name",
        "label": "IP点名称",
        "hint": "复用现有 IP 名称文案",
        "header_lines": ["IP点名称"],
        "height": 40.0,
        "anchor": "bottom2",
    },
    {
        "id": "station",
        "label": "里程桩号（千米+米）",
        "hint": "显示格式：1+234.567",
        "header_lines": ["里程桩号", "（千米+米）"],
        "height": 30.0,
        "anchor": "bottom2",
    },
    {
        "id": "centerline_elev",
        "label": "管中心线高程（米）",
        "hint": "按导入轴线纵断面插值",
        "header_lines": ["管中心线高程（米）"],
        "height": 20.0,
        "anchor": "bottom1",
    },
    {
        "id": "pipe_material",
        "label": "管材（管径/米）",
        "hint": "只显示管材与管径",
        "header_lines": ["管材（管径/米）"],
        "height": 20.0,
        "anchor": "center",
    },
]
_XXPIPE_PROFILE_ROW_IDS = [row["id"] for row in _XXPIPE_PROFILE_ROW_DEFS]
_SPECIAL_STRUCTURE_FULLNAME_MAP = (
    ("隧洞", "隧洞"),
    ("倒虹吸", "倒虹吸"),
    ("有压管道", "有压管道"),
    ("定向钻", "定向钻"),
    ("顶管", "顶管"),
    ("渡槽", "渡槽"),
    ("暗涵", "暗涵"),
)
_SPECIAL_ANGLE_TOL_DEG = 0.01
_BC_ROW_IDS = frozenset({"bd_ip_before", "bj_station_before"})
_EC_ROW_IDS = frozenset({"bf_ip_after", "bl_station_after"})


def _default_profile_row_items():
    enabled_default = set(_TINGZIKOU_TEMPLATE_ROW_IDS)
    return [{"id": rid, "enabled": rid in enabled_default} for rid in _PROFILE_ROW_VISIBLE_ORDER]


def _normalize_profile_row_items(raw_items):
    enabled_default = set(_TINGZIKOU_TEMPLATE_ROW_IDS)
    order = []
    enabled_map = {}

    if isinstance(raw_items, list):
        for item in raw_items:
            if not isinstance(item, dict):
                continue
            rid = str(item.get("id", "")).strip()
            if rid not in _PROFILE_ROW_VISIBLE_ID_SET or rid in order:
                continue
            order.append(rid)
            enabled_map[rid] = bool(item.get("enabled", rid in enabled_default))

    for rid in _PROFILE_ROW_VISIBLE_ORDER:
        if rid not in order:
            order.append(rid)

    return [{"id": rid, "enabled": enabled_map.get(rid, rid in enabled_default)} for rid in order]


def _normalize_text_export_settings(settings):
    src = dict(settings or {})
    src["y_bottom"] = src.get("y_bottom", 1)
    src["y_top"] = src.get("y_top", 31)
    src["y_water"] = src.get("y_water", 16)
    src["text_height"] = src.get("text_height", 3.5)
    src["rotation"] = src.get("rotation", 90)
    src["elev_decimals"] = int(src.get("elev_decimals", 3))
    src["station_decimals"] = int(src.get("station_decimals", 2))
    src["xxpipe_centerline_elev_decimals"] = int(src.get("xxpipe_centerline_elev_decimals", 2))
    src["xxpipe_station_decimals"] = int(src.get("xxpipe_station_decimals", 2))
    src["y_name"] = src.get("y_name", 115)
    src["y_slope"] = src.get("y_slope", 105)
    src["y_ip"] = src.get("y_ip", 77)
    src["y_station"] = src.get("y_station", 47)
    src["y_line_height"] = src.get("y_line_height", 120)
    src["scale_x"] = src.get("scale_x", 2000)
    src["scale_y"] = src.get("scale_y", 1000)
    src["profile_row_items"] = _normalize_profile_row_items(src.get("profile_row_items"))
    return src


def _get_xxpipe_centerline_elev_decimals(settings):
    """读取 xx管 管中心线高程专用小数位数。"""
    normalized = _normalize_text_export_settings(settings)
    return int(normalized.get("xxpipe_centerline_elev_decimals", 2))


def _get_standard_station_decimals(settings):
    """读取普通模式导出桩号专用小数位数。"""
    normalized = _normalize_text_export_settings(settings)
    return int(normalized.get("station_decimals", 2))


def _get_xxpipe_station_decimals(settings):
    """读取 xx管 里程桩号专用小数位数。"""
    normalized = _normalize_text_export_settings(settings)
    return int(normalized.get("xxpipe_station_decimals", 2))


def _format_station_with_decimals(station_value, station_prefix="", *, decimals=3):
    """按指定精度格式化桩号文本。"""
    try:
        station_number = float(station_value)
    except (TypeError, ValueError):
        station_number = 0.0
    if station_number < 0:
        station_number = 0.0

    decimals = max(0, int(decimals))
    km = int(station_number / 1000)
    meters = station_number % 1000
    width = 3 if decimals == 0 else 4 + decimals
    meters_str = f"{meters:0{width}.{decimals}f}"
    return f"{station_prefix}{km}+{meters_str}"


def _format_xxpipe_station(station_value, station_prefix="", *, decimals=2):
    """按 xx管 专用精度格式化里程桩号。"""
    return _format_station_with_decimals(
        station_value,
        station_prefix,
        decimals=decimals,
    )


def _get_xxpipe_profile_row_defs():
    return [dict(row) for row in _XXPIPE_PROFILE_ROW_DEFS]


def _build_xxpipe_profile_row_layout(settings):
    normalized = _normalize_text_export_settings(settings)
    enabled_ids = list(_XXPIPE_PROFILE_ROW_IDS)
    total_height = sum(float(row["height"]) for row in _XXPIPE_PROFILE_ROW_DEFS)
    line_height = total_height

    row_layout = {}
    boundaries = {0.0, total_height}
    cursor_top = total_height
    for row_def in _XXPIPE_PROFILE_ROW_DEFS:
        rid = row_def["id"]
        height = float(row_def["height"])
        top = cursor_top
        bottom = top - height
        cursor_top = bottom
        anchor = row_def.get("anchor", "bottom2")
        if anchor == "center":
            text_y = (bottom + top) / 2.0
        elif anchor == "bottom1":
            text_y = bottom + 1.0
        else:
            text_y = bottom + 2.0
        row_layout[rid] = {
            "bottom": bottom,
            "top": top,
            "text_y": text_y,
            "height": height,
            "anchor": anchor,
            "header_lines": list(row_def.get("header_lines", [])),
            "label": row_def["label"],
        }
        boundaries.add(bottom)
        boundaries.add(top)
    return normalized, enabled_ids, row_layout, total_height, line_height, sorted(boundaries)


def _get_enabled_profile_row_ids(settings):
    normalized = _normalize_text_export_settings(settings)
    return [item["id"] for item in normalized["profile_row_items"] if item.get("enabled")]


def _build_profile_row_layout(settings):
    normalized = _normalize_text_export_settings(settings)
    enabled_ids = [item["id"] for item in normalized["profile_row_items"] if item.get("enabled")]
    if not enabled_ids:
        return [], {}, 0.0, float(normalized.get("y_line_height", 120)), [0.0]

    total_height = sum(float(_PROFILE_ROW_DEF_MAP[rid]["height"]) for rid in enabled_ids)
    min_line_height = float(normalized.get("y_line_height", 120))
    line_height = max(total_height, min_line_height)

    row_layout = {}
    boundaries = {0.0, total_height, line_height}
    cursor_top = total_height
    for rid in enabled_ids:
        row_def = _PROFILE_ROW_DEF_MAP[rid]
        height = float(row_def["height"])
        top = cursor_top
        bottom = top - height
        cursor_top = bottom

        if row_def["anchor"] == "center":
            text_y = (bottom + top) / 2.0
        elif row_def["anchor"] == "bottom1":
            text_y = bottom + 1.0
        else:
            text_y = bottom + 2.0

        row_layout[rid] = {
            "bottom": bottom,
            "top": top,
            "text_y": text_y,
            "height": height,
            "anchor": row_def.get("anchor", "bottom2"),
            "header_lines": list(row_def.get("header_lines", [])),
            "label": row_def["label"],
        }
        boundaries.add(bottom)
        boundaries.add(top)

    return enabled_ids, row_layout, total_height, line_height, sorted(boundaries)


def _compute_runtime_advanced_parameter_view(settings):
    """\u6839\u636e\u5f53\u524d\u884c\u914d\u7f6e\u8ba1\u7b97\u9ad8\u7ea7\u53c2\u6570\u5b9e\u65f6\u89c6\u56fe\u3002"""
    normalized = _normalize_text_export_settings(settings)
    enabled_ids, row_layout, total_height, line_height, boundaries = _build_profile_row_layout(normalized)
    legacy_writeback_values = {}
    legacy_enabled_state = {}
    for key, rid in _PROFILE_RUNTIME_ADVANCED_ROW_BINDINGS.items():
        if rid in row_layout:
            legacy_writeback_values[key] = float(row_layout[rid]["text_y"])
            legacy_enabled_state[key] = True
        else:
            legacy_writeback_values[key] = None
            legacy_enabled_state[key] = False
    legacy_writeback_values["y_line_height"] = float(line_height)
    legacy_enabled_state["y_line_height"] = True

    enabled_runtime_rows = []
    for order, rid in enumerate(enabled_ids, start=1):
        row_info = row_layout.get(rid, {})
        anchor = str(row_info.get("anchor", ""))
        enabled_runtime_rows.append({
            "order": order,
            "id": rid,
            "label": str(row_info.get("label", rid)),
            "text_y": float(row_info.get("text_y", 0.0)),
            "height": float(row_info.get("height", 0.0)),
            "anchor": anchor,
            "source_label": (
                f"{_PROFILE_RUNTIME_ANCHOR_LABELS.get(anchor, anchor or '--')} / \u884c\u9ad8 "
                f"{_format_number(float(row_info.get('height', 0.0)))}"
            ),
        })

    return {
        "enabled_row_ids": list(enabled_ids),
        "total_height": float(total_height),
        "line_height": float(line_height),
        "min_line_height": float(normalized.get("y_line_height", 120)),
        "boundaries": [float(v) for v in boundaries],
        "legacy_writeback_values": legacy_writeback_values,
        "legacy_enabled_state": legacy_enabled_state,
        "enabled_runtime_rows": enabled_runtime_rows,
        "row_details": enabled_runtime_rows,
        # keep aliases for older internal callers
        "compatibility_values": legacy_writeback_values,
        "enabled_state": legacy_enabled_state,
    }


def _compute_xxpipe_runtime_advanced_parameter_view(settings):
    """根据 xx管 固定 5 行模板计算实时参数视图。"""
    normalized, enabled_ids, row_layout, total_height, line_height, boundaries = _build_xxpipe_profile_row_layout(settings)
    legacy_writeback_values = {}
    legacy_enabled_state = {}

    for key in _PROFILE_RUNTIME_ADVANCED_KEYS:
        if key == "y_line_height":
            legacy_writeback_values[key] = float(line_height)
            legacy_enabled_state[key] = True
            continue

        rid = _XXPIPE_PROFILE_RUNTIME_ADVANCED_ROW_BINDINGS.get(key)
        if rid and rid in row_layout:
            legacy_writeback_values[key] = float(row_layout[rid]["text_y"])
            legacy_enabled_state[key] = True
        else:
            legacy_writeback_values[key] = None
            legacy_enabled_state[key] = False

    enabled_runtime_rows = []
    for order, rid in enumerate(enabled_ids, start=1):
        row_info = row_layout.get(rid, {})
        anchor = str(row_info.get("anchor", ""))
        enabled_runtime_rows.append({
            "order": order,
            "id": rid,
            "label": str(row_info.get("label", rid)),
            "text_y": float(row_info.get("text_y", 0.0)),
            "height": float(row_info.get("height", 0.0)),
            "anchor": anchor,
            "source_label": (
                f"{_PROFILE_RUNTIME_ANCHOR_LABELS.get(anchor, anchor or '--')} / 行高 "
                f"{_format_number(float(row_info.get('height', 0.0)))}"
            ),
        })

    return {
        "enabled_row_ids": list(enabled_ids),
        "total_height": float(total_height),
        "line_height": float(line_height),
        "min_line_height": float(normalized.get("y_line_height", 120)),
        "boundaries": [float(v) for v in boundaries],
        "legacy_writeback_values": legacy_writeback_values,
        "legacy_enabled_state": legacy_enabled_state,
        "enabled_runtime_rows": enabled_runtime_rows,
        "row_details": enabled_runtime_rows,
        "compatibility_values": legacy_writeback_values,
        "enabled_state": legacy_enabled_state,
    }



def _compute_node_vline_segments(node, row_layout, enabled_row_ids, v_top, tol=1e-9):
    """计算单个 IP 节点的竖线分段（按 BC/MC/EC x 坐标分组）。

    当 station_BC 或 station_EC 与 station_MC 不同时，将竖线按行的
    x 坐标组拆分，使每段竖线仅穿越属于同一 x 坐标组的行。

    返回 [(station_x, y_bottom, y_top), ...], station_x 为未缩放的原始桩号。
    """
    mc = float(getattr(node, "station_MC", 0) or 0.0)
    bc = float(getattr(node, "station_BC", mc) or mc)
    ec = float(getattr(node, "station_EC", mc) or mc)

    bc_differs = abs(bc - mc) > tol
    ec_differs = abs(ec - mc) > tol

    if not (bc_differs or ec_differs):
        return [(mc, 0.0, v_top)]

    bc_intervals = []
    ec_intervals = []
    for rid in enabled_row_ids:
        if rid not in row_layout:
            continue
        if rid in _BC_ROW_IDS and bc_differs:
            bc_intervals.append((row_layout[rid]["bottom"], row_layout[rid]["top"]))
        elif rid in _EC_ROW_IDS and ec_differs:
            ec_intervals.append((row_layout[rid]["bottom"], row_layout[rid]["top"]))

    if not bc_intervals and not ec_intervals:
        return [(mc, 0.0, v_top)]

    exclude = sorted(bc_intervals + ec_intervals)

    segments = []
    y_cursor = 0.0
    for exc_bot, exc_top in exclude:
        if exc_bot > y_cursor + tol and exc_bot <= v_top + tol:
            segments.append((mc, y_cursor, min(exc_bot, v_top)))
        y_cursor = max(y_cursor, exc_top)
    if y_cursor < v_top - tol:
        segments.append((mc, y_cursor, v_top))

    for bot, top in bc_intervals:
        eff_bot = max(bot, 0.0)
        eff_top = min(top, v_top)
        if eff_top > eff_bot + tol:
            segments.append((bc, eff_bot, eff_top))

    for bot, top in ec_intervals:
        eff_bot = max(bot, 0.0)
        eff_top = min(top, v_top)
        if eff_top > eff_bot + tol:
            segments.append((ec, eff_bot, eff_top))

    return segments


def _is_special_inout_node(node):
    if not _is_special_structure_sv(getattr(node, "structure_type", None)):
        return False
    return _in_out_val(getattr(node, "in_out", None)) in ("进", "出")


def _resolve_profile_vline_top(is_special, is_last_node, short_line_height, line_height):
    """普通节点保留顶部合并单元格，末列右边界和特殊进/出节点补齐整高。"""
    if is_special or is_last_node:
        return line_height
    return short_line_height


def _get_special_structure_full_name(struct_type):
    sv = _struct_val(struct_type)
    for key, full in _SPECIAL_STRUCTURE_FULLNAME_MAP:
        if key in sv:
            return full
    if "-" in sv:
        return sv.split("-")[0]
    return sv


def _merge_building_and_structure_name(building_name, structure_full):
    name = (building_name or "").strip()
    struct = (structure_full or "").strip()
    if not struct:
        return name
    if struct in name:
        return name
    return f"{name}{struct}" if name else struct


def _is_xxpipe_channel_level(level):
    return str(level or "").strip() in set(XXPIPE_CHANNEL_LEVEL_OPTIONS)


def _is_xxpipe_allowed_structure(struct_name):
    text = str(struct_name or "").strip()
    if not text:
        return False
    return text in {"有压管道", "定向钻", "顶管"}


def _is_xxpipe_named_structure(struct_name):
    text = str(struct_name or "").strip()
    if not text:
        return False
    return text in {"定向钻", "顶管"}


def _get_xxpipe_structure_display_name(struct_name):
    text = str(struct_name or "").strip()
    if not text:
        return ""
    return text


def _get_xxpipe_building_display_name(struct_name, building_name):
    if not _is_xxpipe_named_structure(struct_name):
        return ""
    return _merge_building_and_structure_name(
        building_name,
        _get_xxpipe_structure_display_name(struct_name),
    )


def _format_xxpipe_pipe_material_text(row):
    if not isinstance(row, dict):
        return ""
    material = str(
        row.get("pipe_material")
        or row.get("material")
        or row.get("material_key")
        or ""
    ).strip() or "球墨铸铁管"
    dn_mm = _normalize_dn_mm(
        row.get("DN_mm", row.get("dn_mm", row.get("dn", row.get("D", 1500)))),
        1500,
    )
    return f"{material} DN{dn_mm}"


def _build_profile_ip_base_text(node):
    ip_no = int(getattr(node, "ip_number", 0) or 0)
    ip_text = f"IP{ip_no}"
    if _is_special_inout_node(node):
        merged_name = _merge_building_and_structure_name(
            getattr(node, "name", ""),
            _get_special_structure_full_name(getattr(node, "structure_type", None)),
        )
        in_out = _in_out_val(getattr(node, "in_out", None))
        detail = f"{merged_name}{in_out}".strip()
        return f"{ip_text} {detail}".strip()
    return ip_text


def _iter_profile_ip_nodes(nodes):
    special_stations = set()
    for node in nodes:
        if _is_special_inout_node(node):
            special_stations.add(round(float(getattr(node, "station_MC", 0) or 0.0), 6))

    result = []
    for node in nodes:
        struct_str = node.get_structure_type_str() or ""
        if getattr(node, "is_transition", False) or struct_str == "渐变段":
            continue
        if getattr(node, "is_auto_inserted_channel", False):
            continue
        # 特殊建筑物（隧洞/倒虹吸/有压管道/渡槽/暗涵）内部的 IP 节点也需要显示（只显示 IPxx）；
        # 仅排除普通节点中与特殊建筑进/出口桩号重合的节点（避免双重标注）。
        if not _is_special_structure_sv(getattr(node, "structure_type", None)):
            if round(float(getattr(node, "station_MC", 0) or 0.0), 6) in special_stations:
                continue
        result.append(node)
    return result


def _build_ip_related_row_records(nodes, station_prefix, station_resolver=None, *, station_decimals=3):
    """构建 BD/BE/BF/BJ/BK/BL 六类文本记录。

    返回: {row_id: [{"x": float, "text": str, "node": node}, ...], ...}
    """
    row_ids = [
        "ip_name",
        "bd_ip_before", "be_ip_text", "bf_ip_after",
        "bj_station_before", "bk_station", "bl_station_after",
    ]
    records = {rid: [] for rid in row_ids}
    last_x_map = {rid: None for rid in row_ids}

    ip_nodes = _iter_profile_ip_nodes(nodes)
    for node in ip_nodes:
        base_text = _build_profile_ip_base_text(node)
        angle = abs(float(getattr(node, "turn_angle", 0) or 0.0))
        is_special = _is_special_inout_node(node)

        before_text = base_text if (is_special or angle <= 0) else f"{base_text}弯前"
        after_text = base_text if (is_special or angle <= 0) else f"{base_text}弯后"
        center_text = base_text

        station_mc = _resolve_profile_station_value(node, station_resolver)
        station_bc = float(getattr(node, "station_BC", station_mc) or 0.0)
        station_ec = float(getattr(node, "station_EC", station_mc) or station_mc)

        station_before = _format_station_with_decimals(
            station_bc,
            station_prefix,
            decimals=station_decimals,
        )
        station_center = _format_station_with_decimals(
            station_mc,
            station_prefix,
            decimals=station_decimals,
        )
        station_after = _format_station_with_decimals(
            station_ec,
            station_prefix,
            decimals=station_decimals,
        )

        row_payloads = [
            ("ip_name", station_mc, center_text),
            ("bd_ip_before", station_bc, before_text),
            ("be_ip_text", station_mc, center_text),
            ("bf_ip_after", station_ec, after_text),
            ("bj_station_before", station_bc, station_before),
            ("bk_station", station_mc, station_center),
            ("bl_station_after", station_ec, station_after),
        ]
        for rid, x_val, text_val in row_payloads:
            adjusted_x = float(x_val)
            prev_x = last_x_map.get(rid)
            if prev_x is not None and abs(prev_x - adjusted_x) <= 1e-9:
                adjusted_x += 6.0
            records[rid].append({
                "x": adjusted_x,
                "text": str(text_val),
                "node": node,
            })
            last_x_map[rid] = adjusted_x

    return records


def _build_special_angle_warning(nodes, tol_deg=_SPECIAL_ANGLE_TOL_DEG):
    near_msgs = []
    over_msgs = []
    for node in _iter_profile_ip_nodes(nodes):
        if not _is_special_inout_node(node):
            continue
        angle = abs(float(getattr(node, "turn_angle", 0) or 0.0))
        base_text = _build_profile_ip_base_text(node)
        if angle >= tol_deg:
            over_msgs.append(f"{base_text}: {angle:.6f}°")
        elif angle > 0:
            near_msgs.append(f"{base_text}: {angle:.6f}°")

    if not near_msgs and not over_msgs:
        return ""

    lines = ["检测到特殊建筑进/出点转角异常："]
    if near_msgs:
        lines.append("接近0（建议复核）:")
        lines.extend([f"  - {m}" for m in near_msgs])
    if over_msgs:
        lines.append(f"超过阈值 {tol_deg:.3f}°（建议重点复核）:")
        lines.extend([f"  - {m}" for m in over_msgs])
    lines.append("提示：本次仅提醒，不阻断导出。")
    return "\n".join(lines)


def _show_special_angle_warning(panel, nodes):
    msg = _build_special_angle_warning(nodes, tol_deg=_SPECIAL_ANGLE_TOL_DEG)
    if msg:
        fluent_info(panel.window(), "特殊建筑转角提示", msg)


def _parse_valid_station_value(value):
    """解析有效桩号值，无法使用时返回 None。"""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    if number < 0:
        return None
    return number


def _get_station_node_value(node, key, default=None):
    """兼容对象/字典两类节点取值。"""
    if isinstance(node, dict):
        return node.get(key, default)
    return getattr(node, key, default)


def _parse_plan_point(node):
    """解析节点平面坐标，无法使用时返回 None。"""
    try:
        x_val = float(_get_station_node_value(node, "x", None))
        y_val = float(_get_station_node_value(node, "y", None))
    except (TypeError, ValueError):
        return None
    if not math.isfinite(x_val) or not math.isfinite(y_val):
        return None
    return (x_val, y_val)


def _build_station_fallback_prefix(nodes):
    """为桩号回退计算准备距离前缀和。"""
    visible_nodes = list(nodes or [])
    segment_lengths = []
    prefix_lengths = [0.0]
    prefix_missing = [0]

    for index in range(1, len(visible_nodes)):
        start_point = _parse_plan_point(visible_nodes[index - 1])
        end_point = _parse_plan_point(visible_nodes[index])
        if start_point is None or end_point is None:
            distance = None
        else:
            distance = math.hypot(end_point[0] - start_point[0], end_point[1] - start_point[1])
        segment_lengths.append(distance)
        prefix_lengths.append(prefix_lengths[-1] + (distance or 0.0))
        prefix_missing.append(prefix_missing[-1] + (0 if distance is not None else 1))

    return segment_lengths, prefix_lengths, prefix_missing


def _segment_distance_between(prefix_lengths, prefix_missing, start_index: int, end_index: int):
    """计算节点区间的累计距离，中间存在缺失坐标时返回 None。"""
    if start_index == end_index:
        return 0.0
    if start_index > end_index:
        distance = _segment_distance_between(prefix_lengths, prefix_missing, end_index, start_index)
        return -distance if distance is not None else None
    if start_index < 0 or end_index >= len(prefix_lengths):
        return None
    missing_count = prefix_missing[end_index] - prefix_missing[start_index]
    if missing_count > 0:
        return None
    return prefix_lengths[end_index] - prefix_lengths[start_index]


def _build_station_resolution_label(node, index: int):
    """生成桩号解析提示标签。"""
    ip_no = _get_station_node_value(node, "ip_number", None)
    name = str(_get_station_node_value(node, "name", "") or "").strip()
    if isinstance(ip_no, (int, float)):
        base = f"IP{int(ip_no)}"
    else:
        base = f"第{int(index) + 1}个节点"
    return f"{base} {name}".strip()


def resolve_ordered_node_stations(nodes):
    """解析顺序节点桩号，个别缺失时按平面累计距离回退。"""
    ordered_nodes = list(nodes or [])
    if not ordered_nodes:
        return {"stations": [], "fallback_indices": [], "missing_items": []}

    _segment_lengths, prefix_lengths, prefix_missing = _build_station_fallback_prefix(ordered_nodes)
    explicit_stations = [
        _parse_valid_station_value(_get_station_node_value(node, "station_MC", _get_station_node_value(node, "station_mc", None)))
        for node in ordered_nodes
    ]
    explicit_indices = [index for index, value in enumerate(explicit_stations) if value is not None]
    stations = []
    fallback_indices = []
    missing_items = []

    for index, node in enumerate(ordered_nodes):
        station_mc = explicit_stations[index]
        used_fallback = False
        error_reason = ""

        if station_mc is None:
            if not explicit_indices:
                if prefix_missing[-1] == 0:
                    distance_from_start = _segment_distance_between(prefix_lengths, prefix_missing, 0, index)
                    if distance_from_start is not None:
                        station_mc = float(distance_from_start)
                        used_fallback = index > 0
                if station_mc is None:
                    error_reason = "缺少可用桩号锚点"
            else:
                prev_anchor = next(
                    (
                        anchor_index
                        for anchor_index in range(index - 1, -1, -1)
                        if explicit_stations[anchor_index] is not None
                        and _segment_distance_between(prefix_lengths, prefix_missing, anchor_index, index) is not None
                    ),
                    None,
                )
                next_anchor = next(
                    (
                        anchor_index
                        for anchor_index in range(index + 1, len(ordered_nodes))
                        if explicit_stations[anchor_index] is not None
                        and _segment_distance_between(prefix_lengths, prefix_missing, index, anchor_index) is not None
                    ),
                    None,
                )

                if prev_anchor is not None and next_anchor is not None:
                    prev_distance = _segment_distance_between(prefix_lengths, prefix_missing, prev_anchor, index)
                    next_distance = _segment_distance_between(prefix_lengths, prefix_missing, index, next_anchor)
                    span_distance = _segment_distance_between(prefix_lengths, prefix_missing, prev_anchor, next_anchor)
                    if (
                        prev_distance is not None
                        and next_distance is not None
                        and span_distance is not None
                        and abs(span_distance) > 1e-9
                    ):
                        station_mc = (
                            explicit_stations[prev_anchor]
                            + (explicit_stations[next_anchor] - explicit_stations[prev_anchor])
                            * (prev_distance / span_distance)
                        )
                        used_fallback = True
                elif prev_anchor is not None:
                    prev_distance = _segment_distance_between(prefix_lengths, prefix_missing, prev_anchor, index)
                    if prev_distance is not None:
                        station_mc = explicit_stations[prev_anchor] + prev_distance
                        used_fallback = True
                elif next_anchor is not None:
                    next_distance = _segment_distance_between(prefix_lengths, prefix_missing, index, next_anchor)
                    if next_distance is not None:
                        station_mc = explicit_stations[next_anchor] - next_distance
                        used_fallback = True

                if station_mc is None:
                    error_reason = "缺少可用于回退的连续平面坐标"

        stations.append(station_mc)
        if station_mc is None:
            missing_items.append(
                {
                    "index": index,
                    "label": _build_station_resolution_label(node, index),
                    "reason": error_reason or "缺少可用桩号",
                }
            )
        elif used_fallback:
            fallback_indices.append(index)

    return {
        "stations": stations,
        "fallback_indices": fallback_indices,
        "missing_items": missing_items,
    }


def resolve_xxpipe_profile_station_targets(nodes, station_prefix=""):
    """解析 xx管 纵断面导出节点桩号，个别缺失时按平面累计距离回退。"""
    visible_nodes = list(_iter_xxpipe_export_nodes(nodes))
    station_data = resolve_ordered_node_stations(visible_nodes)
    resolved_targets = []
    errors = list(station_data["missing_items"])

    for index, node in enumerate(visible_nodes):
        station_mc = station_data["stations"][index]
        used_fallback = index in set(station_data["fallback_indices"])

        node_label = _build_profile_ip_base_text(node)
        if not node_label:
            node_label = _make_xxpipe_identity_from_node(node)

        station_text = "-"
        if station_mc is not None:
            try:
                station_text = ProjectSettings.format_station(station_mc, station_prefix)
            except Exception:
                station_text = f"{station_mc:.3f}"

        resolved_targets.append(
            {
                "node": node,
                "identity": _make_xxpipe_identity_from_node(node),
                "label": node_label,
                "station_mc": station_mc,
                "station_text": station_text,
                "used_fallback": used_fallback,
            }
        )
    normalized_errors = []
    for item in errors:
        node = visible_nodes[item["index"]]
        normalized_errors.append(
            {
                "node": node,
                "identity": _make_xxpipe_identity_from_node(node),
                "label": item["label"],
                "reason": item["reason"],
            }
        )
    return resolved_targets, normalized_errors


def resolve_xxpipe_target_station_values(targets, station_prefix=""):
    """解析 xx管 整线覆盖校验目标桩号，缺失时按平面累计距离回退。"""
    ordered_targets = list(targets or [])
    station_data = resolve_ordered_node_stations(ordered_targets)
    resolved_targets = []
    fallback_indices = set(station_data["fallback_indices"])

    for index, target in enumerate(ordered_targets):
        station_mc = station_data["stations"][index]
        station_text = "-"
        if station_mc is not None:
            try:
                station_text = ProjectSettings.format_station(station_mc, station_prefix)
            except Exception:
                station_text = f"{station_mc:.3f}"

        payload = dict(target) if isinstance(target, dict) else {"value": target}
        payload["resolved_station_mc"] = station_mc
        payload["station_text"] = station_text
        payload["used_fallback"] = index in fallback_indices
        resolved_targets.append(payload)

    return resolved_targets, list(station_data["missing_items"])


def _profile_station_value(node):
    """提取纵断面导出用 station_MC 浮点值。"""
    value = _parse_valid_station_value(getattr(node, "station_MC", None))
    return float(value) if value is not None else 0.0


def _resolve_profile_station_value(node, station_resolver=None):
    """提取纵断面导出桩号，必要时使用外部回退结果。"""
    if isinstance(station_resolver, dict):
        value = station_resolver.get(id(node), None)
        if value is not None:
            return float(value)
    return _profile_station_value(node)


def _profile_elevation_score(node):
    """计算节点高程完整度分值（非零项越多越优先）。"""
    score = 0
    for attr in ("bottom_elevation", "top_elevation", "water_level"):
        try:
            val = float(getattr(node, attr, 0) or 0.0)
        except (TypeError, ValueError):
            val = 0.0
        if abs(val) > 1e-9:
            score += 1
    return score


def _is_profile_text_export_node(node):
    """判断节点是否属于纵断面文本行“真实节点”（排除渐变段与自动插入明渠段）。"""
    struct_str = node.get_structure_type_str() or ""
    if getattr(node, "is_transition", False) or struct_str == "渐变段":
        return False
    if getattr(node, "is_auto_inserted_channel", False):
        return False
    return True


def _build_profile_text_nodes(nodes, station_resolver=None):
    """构建纵断面四行文本输出节点（真实节点过滤 + 同桩号归并 + 冲突校验）。"""
    grouped_by_station = {}
    station_order = []
    for node in nodes:
        if not _is_profile_text_export_node(node):
            continue
        station_val = _resolve_profile_station_value(node, station_resolver)
        station_key = round(station_val, 9)
        if station_key not in grouped_by_station:
            grouped_by_station[station_key] = []
            station_order.append(station_key)
        grouped_by_station[station_key].append(node)

    def _node_label(node_obj):
        ip_no = getattr(node_obj, "ip_number", None)
        ip_label = f"IP{ip_no}" if ip_no is not None else "IP?"
        name = str(getattr(node_obj, "name", "") or "").strip()
        return f"{ip_label}({name})" if name else ip_label

    def _resolve_elev(group_nodes, attr_name, field_label, station_value, tol=1e-6):
        non_zero_values = []
        fallback_values = []
        for group_node in group_nodes:
            try:
                value = float(getattr(group_node, attr_name, 0) or 0.0)
            except (TypeError, ValueError):
                value = 0.0
            fallback_values.append(value)
            if abs(value) > tol:
                non_zero_values.append((value, group_node))

        unique_values = []
        for value, group_node in non_zero_values:
            if not any(abs(value - prev_value) <= tol for prev_value, _ in unique_values):
                unique_values.append((value, group_node))

        if len(unique_values) > 1:
            detail = "；".join(f"{val:.6f}@{_node_label(nd)}" for val, nd in unique_values)
            raise ValueError(
                f"纵断面导出检测到同桩号冲突：桩号 {station_value:.6f} 的{field_label}存在多个非零值（{detail}）"
            )

        if unique_values:
            return unique_values[0][0]
        return fallback_values[0] if fallback_values else 0.0

    merged_nodes = []
    field_labels = {
        "bottom_elevation": "渠底高程",
        "top_elevation": "渠顶高程",
        "water_level": "设计水位",
    }
    for station_key in station_order:
        group = grouped_by_station[station_key]
        representative = max(group, key=_profile_elevation_score)
        station_val = _profile_station_value(representative)

        merged = copy.copy(representative)
        merged.station_MC = station_val
        for attr_name, field_label in field_labels.items():
            setattr(merged, attr_name, _resolve_elev(group, attr_name, field_label, station_val))
        merged_nodes.append(merged)
    return merged_nodes


def _resolve_segment_mid_mc(seg_start, seg_end, boundary_mcs, tol=1e-9):
    """根据边界竖线计算段落中心MC；单点段优先取所在单元格几何中心。"""
    bounds = sorted({float(val) for val in boundary_mcs})
    if not bounds:
        return (seg_start + seg_end) / 2.0

    left_bound = max((val for val in bounds if val <= seg_start + tol), default=seg_start)
    right_bound = min((val for val in bounds if val >= seg_end - tol), default=seg_end)

    if right_bound - left_bound <= tol:
        pivot = seg_start if abs(seg_start - seg_end) <= tol else (seg_start + seg_end) / 2.0
        prev_bound = max((val for val in bounds if val < pivot - tol), default=None)
        next_bound = min((val for val in bounds if val > pivot + tol), default=None)
        if abs(seg_start - seg_end) <= tol:
            if next_bound is not None and abs(pivot - bounds[0]) <= tol:
                right_bound = next_bound
            elif prev_bound is not None and abs(pivot - bounds[-1]) <= tol:
                left_bound = prev_bound
            elif prev_bound is not None and next_bound is not None:
                left_bound, right_bound = prev_bound, next_bound
            elif next_bound is not None:
                right_bound = next_bound
            elif prev_bound is not None:
                left_bound = prev_bound

    return (left_bound + right_bound) / 2.0


def _is_gate_name(name):
    """判断建筑物显示名称是否为闸类点状建筑物（分水闸/分水口/节制闸/泄水闸等）"""
    if not name:
        return False
    return "闸" in name or "分水" in name


def _get_profile_slope_segment_identity(node):
    """返回坡降分段用的结构标识，避免与建筑物名称分段强绑定。"""
    struct_str = _struct_val(getattr(node, "structure_type", None))
    if not struct_str:
        return ""
    if struct_str.startswith("明渠") or struct_str == "矩形暗涵":
        return struct_str
    category = struct_str.split("-")[0]
    raw_name = str(getattr(node, "name", "") or "").strip()
    return _merge_building_and_structure_name(raw_name, category) if raw_name else category


def _is_profile_slope_placeholder_node(node):
    """判断坡降行是否应以 '-' 占位（倒虹吸/有压管道）。"""
    struct_type = getattr(node, "structure_type", None)
    struct_value = _struct_val(struct_type)
    struct_name = getattr(struct_type, "name", "")
    return bool(
        getattr(node, "is_inverted_siphon", False)
        or getattr(node, "is_pressure_pipe", False)
        or ("倒虹吸" in struct_value)
        or ("有压管道" in struct_value)
        or struct_name in ("INVERTED_SIPHON", "PRESSURE_PIPE")
    )


def _build_profile_slope_segments(nodes, profile_text_nodes=None):
    """按“当前节点作为区间终点”构建纵断面坡降区间，供 DXF/TXT 共用。"""
    visible_nodes = list(profile_text_nodes) if profile_text_nodes is not None else _build_profile_text_nodes(nodes or [])
    segments = []
    prev_visible_mc = None
    prev_merge_key = None

    for node in visible_nodes:
        current_mc = float(_profile_station_value(node))
        if prev_visible_mc is None:
            prev_visible_mc = current_mc
            prev_merge_key = None
            continue

        identity = _get_profile_slope_segment_identity(node)
        gate_hint = identity or _struct_val(getattr(node, "structure_type", None))
        if _is_gate_name(gate_hint):
            prev_visible_mc = current_mc
            prev_merge_key = None
            continue

        if _is_profile_slope_placeholder_node(node):
            text = "-"
            merge_key = ("placeholder", identity, text)
        else:
            text = _get_node_slope_text(node)
            if not text or text == "/":
                prev_visible_mc = current_mc
                prev_merge_key = None
                continue
            merge_key = ("slope", identity, text)

        if segments and prev_merge_key == merge_key:
            segments[-1]["end_mc"] = current_mc
        else:
            segments.append(
                {
                    "text": text,
                    "start_mc": prev_visible_mc,
                    "end_mc": current_mc,
                }
            )

        prev_visible_mc = current_mc
        prev_merge_key = merge_key

    return segments


def _collect_profile_slope_boundary_mcs(slope_segments, tol=1e-9):
    """提取坡降行需要补齐短竖线的区间边界。"""
    boundary_mcs = []
    for segment in slope_segments or []:
        for key in ("start_mc", "end_mc"):
            try:
                mc = float(segment.get(key, 0.0) or 0.0)
            except (TypeError, ValueError, AttributeError):
                continue
            if any(abs(mc - prev_mc) <= tol for prev_mc in boundary_mcs):
                continue
            boundary_mcs.append(mc)
    return boundary_mcs


def _merge_segments_across_gates(segments, gate_mc_set=None):
    """合并被闸（点状建筑物）拆分的同名段落

    规则：如果段落 i 和段落 j 的值相同，且 i~j 之间的所有段落
    都是闸类点状建筑物，则将 j 的 MC 列表合并到 i，闸段保留不变。

    对建筑物名称段落：通过名称判断闸（_is_gate_name）。
    对坡降等段落：通过 gate_mc_set（闸节点的桩号集合）判断。

    Args:
        segments: [(value, [mc_list]), ...] 按位置排列的段落
        gate_mc_set: 闸节点桩号集合，仅在非名称场景下使用
    """
    if len(segments) <= 2:
        return segments

    def _is_gate_seg(val, mc_list):
        if gate_mc_set is not None:
            return all(mc in gate_mc_set for mc in mc_list)
        return _is_gate_name(val)

    merged = []
    i = 0
    while i < len(segments):
        val, mc_list = segments[i]

        if _is_gate_seg(val, mc_list):
            merged.append((val, list(mc_list)))
            i += 1
            continue

        # 非闸段：尝试向后合并同名段（跳过中间的闸段）
        mc_list = list(mc_list)
        j = i + 1
        while j + 1 < len(segments):
            mid_val, mid_mcs = segments[j]
            next_val, next_mcs = segments[j + 1]
            if _is_gate_seg(mid_val, mid_mcs) and next_val == val:
                merged.append((mid_val, list(mid_mcs)))  # 闸段保留
                mc_list.extend(next_mcs)
                j += 2
            else:
                break

        merged.append((val, mc_list))
        i = j if j > i + 1 else i + 1

    return merged


# ================================================================
# DXF 共享辅助工具
# ================================================================

class _OffsetMSP:
    """包装 ezdxf modelspace，自动为所有绘图操作添加坐标偏移。
    用于在同一 DXF 文件中将多个表格放置在不同位置。"""

    def __init__(self, msp, ox=0, oy=0):
        self._msp = msp
        self._ox = ox
        self._oy = oy

    def _p(self, pt):
        return (pt[0] + self._ox, pt[1] + self._oy)

    def add_line(self, start, end, dxfattribs=None):
        return self._msp.add_line(self._p(start), self._p(end),
                                   dxfattribs=dxfattribs or {})

    def add_lwpolyline(self, points, dxfattribs=None):
        return self._msp.add_lwpolyline(
            [self._p(p) for p in points], dxfattribs=dxfattribs or {})

    def add_text(self, text, dxfattribs=None):
        entity = self._msp.add_text(text, dxfattribs=dxfattribs or {})
        return _OffsetTextEntity(entity, self._ox, self._oy)


class _OffsetTextEntity:
    """包装 ezdxf text 实体，自动为 set_placement 添加坐标偏移。"""

    def __init__(self, entity, ox, oy):
        self._entity = entity
        self._ox = ox
        self._oy = oy

    def set_placement(self, point, align=None):
        p = (point[0] + self._ox, point[1] + self._oy)
        if align is not None:
            return self._entity.set_placement(p, align=align)
        return self._entity.set_placement(p)


def _setup_dxf_style(doc):
    """设置 DXF 文档的中文字体样式（仿宋，宽度因子0.7）。"""
    if "Standard" in doc.styles:
        _sty = doc.styles.get("Standard")
    else:
        _sty = doc.styles.add("Standard")
    _sty.dxf.font = ""
    _sty.dxf.width = 0.7
    try:
        if "ACAD" not in doc.appids:
            doc.appids.new("ACAD")
    except Exception:
        pass
    _sty.set_xdata("ACAD", [(1000, "仿宋"), (1071, 0)])


def _setup_profile_dxf_document(doc):
    """初始化纵断面相关 DXF 文档的图纸单位与文字样式。"""
    _setup_dxf_style(doc)
    header = getattr(doc, "header", None)
    if header is not None:
        header["$INSUNITS"] = 4
        header["$MEASUREMENT"] = 1


def _profile_meters_to_paper_mm(value_m, scale_denom):
    """将纵断面源数据（米）按 1:N 比例换算为图纸单位（mm）。"""
    return float(value_m) * 1000.0 / float(scale_denom)


def _ensure_profile_layers(doc, layer_prefix=""):
    """确保纵断面所需的图层存在。layer_prefix 用于合并导出时区分组件。"""
    layer_defs = [
        ("表格线框", 7), ("渠底高程线", 3), ("渠顶高程线", 1),
        ("设计水位线", 5), ("管中心线", 2), ("文字标注", 7),
    ]
    for name, color in layer_defs:
        full = layer_prefix + name
        if full not in doc.layers:
            doc.layers.new(full, dxfattribs={"color": color})


_IP_TABLE_COLUMN_DEFS = [
    {"id": "ip_name", "label": "IP点", "group_label": "IP点", "merge_vertical": True},
    {"id": "x", "label": "E（m）", "group_label": "坐标值", "excel_number_format": "0.000000"},
    {"id": "y", "label": "N（m）", "group_label": "坐标值", "excel_number_format": "0.000000"},
    {"id": "station_bc", "label": "弯前(千米+米)", "group_label": "桩号"},
    {"id": "station_mc", "label": "里程(千米+米)", "group_label": "桩号"},
    {"id": "station_ec", "label": "弯末(千米+米)", "group_label": "桩号"},
    {"id": "turn_angle", "label": "转角", "group_label": "弯道参数", "excel_number_format": "0.000"},
    {"id": "turn_radius", "label": "半径", "group_label": "弯道参数", "excel_number_format": "0.000"},
    {"id": "tangent_length", "label": "切线长", "group_label": "弯道参数", "excel_number_format": "0.000"},
    {"id": "arc_length", "label": "弧长", "group_label": "弯道参数", "excel_number_format": "0.000"},
    {"id": "bottom_elevation", "label": "底高程(m)", "group_label": "底高程(m)", "merge_vertical": True, "excel_number_format": "0.000"},
    {"id": "water_level", "label": "设计水位(m)", "group_label": "设计水位(m)", "merge_vertical": True, "excel_number_format": "0.000"},
]


def _get_ip_table_preview_headers():
    """返回 IP 表预览/导出共用的列标题。"""
    return [col["label"] for col in _IP_TABLE_COLUMN_DEFS]


def _get_ip_table_group_headers():
    """按连续列分组生成 IP 表合并表头。"""
    groups = []
    start = 0
    current_group = _IP_TABLE_COLUMN_DEFS[0]["group_label"]
    for idx, col_def in enumerate(_IP_TABLE_COLUMN_DEFS[1:], start=1):
        if col_def["group_label"] == current_group:
            continue
        groups.append((start, idx - 1, current_group))
        start = idx
        current_group = col_def["group_label"]
    groups.append((start, len(_IP_TABLE_COLUMN_DEFS) - 1, current_group))
    return groups


def _get_ip_table_vertical_merged_columns():
    """返回需要纵向合并两行表头的列索引。"""
    return {
        idx
        for idx, col_def in enumerate(_IP_TABLE_COLUMN_DEFS)
        if col_def.get("merge_vertical")
    }


def _ip_table_safe_float(val, default=0.0):
    """安全转成浮点数，失败时回退默认值。"""
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _format_ip_table_name(node, station_prefix):
    """格式化 IP 点名称，保持与原导出口径一致。"""
    try:
        if _in_out_val(node.in_out) in ("进", "出"):
            struct_abbr = ""
            struct_str = _struct_val(node.structure_type)
            if struct_str:
                if "隧洞" in struct_str:
                    struct_abbr = "隧"
                elif "倒虹吸" in struct_str:
                    struct_abbr = "倒"
                elif "有压管道" in struct_str:
                    struct_abbr = "管"
                elif "渡槽" in struct_str:
                    struct_abbr = "渡"
                elif "暗涵" in struct_str:
                    struct_abbr = "暗"
            in_out_str = "进" if _in_out_val(node.in_out) == "进" else "出"
            return f"{node.name}{struct_abbr}{in_out_str}"
    except Exception:
        pass
    return f"{station_prefix}IP{getattr(node, 'ip_number', 0)}"


def _format_ip_table_station(value, station_prefix, *, station_decimals=2):
    """格式化桩号文本。"""
    return _format_station_with_decimals(
        _ip_table_safe_float(value),
        station_prefix,
        decimals=station_decimals,
    )


def _format_ip_table_optional_number_text(value, digits=3):
    """按 IP 表规则格式化可空数值文本，0 继续显示为横杠。"""
    number = _ip_table_safe_float(value)
    if abs(number) <= 1e-9:
        return "-"
    return f"{number:.{digits}f}"


def _format_ip_table_optional_number_excel(value, digits=3):
    """按 IP 表规则格式化 Excel 数值，0 继续显示为横杠。"""
    number = _ip_table_safe_float(value)
    if abs(number) <= 1e-9:
        return "-"
    return round(number, digits)


def _build_ip_table_row(node, station_prefix, settings=None, *, excel_mode=False):
    """按共享列定义构造一行 IP 表数据。"""
    station_decimals = _get_standard_station_decimals(settings)
    text_values = {
        "ip_name": _format_ip_table_name(node, station_prefix),
        "x": f"{_ip_table_safe_float(getattr(node, 'x', 0.0)):.6f}",
        "y": f"{_ip_table_safe_float(getattr(node, 'y', 0.0)):.6f}",
        "station_bc": _format_ip_table_station(
            getattr(node, "station_BC", 0.0),
            station_prefix,
            station_decimals=station_decimals,
        ),
        "station_mc": _format_ip_table_station(
            getattr(node, "station_MC", 0.0),
            station_prefix,
            station_decimals=station_decimals,
        ),
        "station_ec": _format_ip_table_station(
            getattr(node, "station_EC", 0.0),
            station_prefix,
            station_decimals=station_decimals,
        ),
        "turn_angle": f"{_ip_table_safe_float(getattr(node, 'turn_angle', 0.0)):.3f}",
        "turn_radius": f"{_ip_table_safe_float(getattr(node, 'turn_radius', 0.0)):.3f}",
        "tangent_length": f"{_ip_table_safe_float(getattr(node, 'tangent_length', 0.0)):.3f}",
        "arc_length": f"{_ip_table_safe_float(getattr(node, 'arc_length', 0.0)):.3f}",
        "bottom_elevation": _format_ip_table_optional_number_text(getattr(node, "bottom_elevation", 0.0)),
        "water_level": _format_ip_table_optional_number_text(getattr(node, "water_level", 0.0)),
    }
    if not excel_mode:
        return [text_values[col_def["id"]] for col_def in _IP_TABLE_COLUMN_DEFS]

    excel_values = {
        "ip_name": text_values["ip_name"],
        "x": _ip_table_safe_float(getattr(node, "x", 0.0)),
        "y": _ip_table_safe_float(getattr(node, "y", 0.0)),
        "station_bc": text_values["station_bc"],
        "station_mc": text_values["station_mc"],
        "station_ec": text_values["station_ec"],
        "turn_angle": round(_ip_table_safe_float(getattr(node, "turn_angle", 0.0)), 3),
        "turn_radius": round(_ip_table_safe_float(getattr(node, "turn_radius", 0.0)), 3),
        "tangent_length": round(_ip_table_safe_float(getattr(node, "tangent_length", 0.0)), 3),
        "arc_length": round(_ip_table_safe_float(getattr(node, "arc_length", 0.0)), 3),
        "bottom_elevation": _format_ip_table_optional_number_excel(getattr(node, "bottom_elevation", 0.0)),
        "water_level": _format_ip_table_optional_number_excel(getattr(node, "water_level", 0.0)),
    }
    return [excel_values[col_def["id"]] for col_def in _IP_TABLE_COLUMN_DEFS]


def _build_ip_table_fallback_row(node):
    """构造异常场景下的 IP 表兜底行。"""
    return [
        f"IP{getattr(node, 'ip_number', '?')}",
        "0.000000",
        "0.000000",
        "0+000.000",
        "0+000.000",
        "0+000.000",
        "0.000",
        "0.000",
        "0.000",
        "0.000",
        "-",
        "-",
    ]


def _compute_ip_preview_data(nodes, station_prefix, settings=None):
    """从节点列表计算IP坐标及弯道参数表预览数据。
    返回 (preview_data, real_nodes)。"""
    real_nodes = [
        n for n in nodes
        if not getattr(n, 'is_transition', False)
        and not getattr(n, 'is_auto_inserted_channel', False)
    ]

    preview_data = []
    for node in real_nodes:
        try:
            preview_data.append(_build_ip_table_row(node, station_prefix, settings))
        except Exception:
            preview_data.append(_build_ip_table_fallback_row(node))
    return preview_data, real_nodes


def _write_ip_table_excel_sheet(ws, real_nodes, station_prefix, settings=None):
    """按共享列定义写入 IP 表 Excel 内容。"""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_font = Font(name='Microsoft YaHei', size=10, bold=True)
    data_font = Font(name='Microsoft YaHei', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    group_headers = _get_ip_table_group_headers()
    ncols = len(_IP_TABLE_COLUMN_DEFS)
    v_merged = _get_ip_table_vertical_merged_columns()

    for start_idx, end_idx, text in group_headers:
        start_col = start_idx + 1
        end_col = end_idx + 1
        ws.cell(row=1, column=start_col, value=text)
        if start_idx in v_merged and start_idx == end_idx:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=2, end_column=end_col)
            continue
        if start_col != end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    for col_idx, header in enumerate(_get_ip_table_preview_headers(), start=1):
        if (col_idx - 1) in v_merged:
            continue
        ws.cell(row=2, column=col_idx, value=header)

    for row in range(1, 3):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill

    for row_idx, node in enumerate(real_nodes, start=3):
        row_values = _build_ip_table_row(
            node,
            station_prefix,
            settings,
            excel_mode=True,
        )
        for col_idx, col_def in enumerate(_IP_TABLE_COLUMN_DEFS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_values[col_idx - 1])
            number_format = col_def.get("excel_number_format")
            if number_format and isinstance(row_values[col_idx - 1], (int, float)):
                cell.number_format = number_format
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = left_align if col_idx == 1 else Alignment(horizontal='center', vertical='center')

    for col_idx in range(1, ncols + 1):
        col_letter = chr(64 + col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is None:
                continue
            text = str(cell_val)
            char_w = sum(2 if ord(ch) > 0x7F else 1 for ch in text)
            max_len = max(max_len, char_w)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 8)


def _parse_positive_dn(text):
    """解析 DN 输入，返回正整数；非法时返回 None。"""
    if text is None:
        return None
    t = str(text).strip()
    if not t:
        return None
    try:
        fv = float(t)
    except (TypeError, ValueError):
        return None
    if not fv.is_integer():
        return None
    dn = int(fv)
    return dn if dn > 0 else None


def _normalize_dn_mm(dn_value, default_dn=1500):
    """将 DN 归一化为正整数 mm。"""
    dn = _parse_positive_dn(dn_value)
    if dn is not None:
        return dn
    default = _parse_positive_dn(default_dn)
    return default if default is not None else 1500


def _extract_named_pressurized_groups(nodes, structure_kind):
    """提取按名称分组的有压流建筑物，返回 [(name, dn_mm), ...]。"""
    valid_rows, _ = _extract_pressurized_param_entities(nodes, structure_kind)
    return valid_rows
    groups = {}
    order = []
    if not nodes:
        return []

    is_siphon = (structure_kind == "siphon")
    default_name = "倒虹吸" if is_siphon else "有压管道"

    for node in nodes:
        if getattr(node, 'is_transition', False) or getattr(node, 'is_auto_inserted_channel', False):
            continue

        st_str = _struct_val(getattr(node, 'structure_type', None))
        if is_siphon:
            matched = bool(getattr(node, 'is_inverted_siphon', False) or ('倒虹吸' in st_str))
        else:
            matched = ('有压管道' in st_str)
        if not matched:
            continue

        raw_name = getattr(node, 'name', '') or ''
        display_name = raw_name.strip() if raw_name.strip() else default_name
        if display_name not in groups:
            groups[display_name] = 0
            order.append(display_name)

        params = getattr(node, 'section_params', {}) or {}
        d_val = 0.0
        for key in ('D', 'd'):
            try:
                v = float(params.get(key, 0) or 0)
            except (TypeError, ValueError):
                v = 0.0
            if v > 0:
                d_val = v
                break
        if d_val <= 0:
            try:
                d_val = float(getattr(node, 'structure_height', 0) or 0)
            except (TypeError, ValueError):
                d_val = 0.0

        if d_val > 0:
            dn_mm = d_val * 1000 if d_val < 20 else d_val
            groups[display_name] = max(groups[display_name], dn_mm)

    return [(name, groups[name]) for name in order]


_SEGMENT_LABELS = [
    "第一流量段",
    "第二流量段",
    "第三流量段",
    "第四流量段",
    "第五流量段",
    "第六流量段",
    "第七流量段",
    "第八流量段",
    "第九流量段",
    "第十流量段",
]


def _pressurized_structure_label(structure_kind):
    return "倒虹吸" if structure_kind == "siphon" else "有压管道"


def _segment_label_from_index(flow_section_idx):
    if isinstance(flow_section_idx, int) and flow_section_idx > 0:
        if flow_section_idx <= len(_SEGMENT_LABELS):
            return _SEGMENT_LABELS[flow_section_idx - 1]
        return f"第{flow_section_idx}流量段"
    return ""


def _parse_flow_section_index(flow_section):
    if flow_section is None:
        return None
    if isinstance(flow_section, int):
        return flow_section if flow_section > 0 else None
    if isinstance(flow_section, float):
        if flow_section.is_integer() and flow_section > 0:
            return int(flow_section)
        return None

    text = str(flow_section).strip()
    if not text:
        return None

    matched = re.search(r"\d+", text)
    if matched:
        idx = int(matched.group(0))
        return idx if idx > 0 else None

    for idx, label in enumerate(_SEGMENT_LABELS, start=1):
        if label in text:
            return idx
    return None


def _normalize_pressurized_name(name, structure_kind):
    text = str(name or "").strip()
    if text and text != "-":
        return text
    return f"未命名{_pressurized_structure_label(structure_kind)}"


def _make_pressurized_param_row(
    *,
    name,
    flow_section,
    structure_kind,
    pipe_material,
    dn_mm,
    display_name=None,
):
    flow_section_idx = _parse_flow_section_index(flow_section)
    base_name = _normalize_pressurized_name(name, structure_kind)
    parsed_dn_mm = _parse_positive_dn(dn_mm)
    fallback_display = (
        f"{base_name}-{_segment_label_from_index(flow_section_idx)}"
        if flow_section_idx
        else base_name
    )
    return {
        "name": base_name,
        "flow_section": flow_section_idx,
        "display_name": str(display_name or "").strip() or fallback_display,
        "pipe_material": str(pipe_material or "").strip() or "球墨铸铁管",
        "DN_mm": parsed_dn_mm if parsed_dn_mm is not None else _normalize_dn_mm(dn_mm, 1500),
        "_has_valid_dn_mm": parsed_dn_mm is not None,
        "structure_kind": structure_kind,
    }


def _extract_pressurized_dn_mm(node):
    params = getattr(node, "section_params", {}) or {}
    for key in ("D", "d"):
        try:
            value = float(params.get(key, 0) or 0)
        except (TypeError, ValueError):
            value = 0.0
        if value > 0:
            return int(round(value * 1000 if value < 20 else value))
    try:
        structure_height = float(getattr(node, "structure_height", 0) or 0)
    except (TypeError, ValueError):
        structure_height = 0.0
    if structure_height > 0:
        return int(round(structure_height * 1000 if structure_height < 20 else structure_height))
    return None


def _normalize_locked_velocity_value(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number) or number < 0:
        return None
    return round(number, 4)


def _normalize_positive_flow_value(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number) or number <= 0:
        return None
    return number


def _row_has_valid_pressurized_dn_mm(row):
    if not isinstance(row, dict):
        return False
    flag = row.get("_has_valid_dn_mm")
    if flag is not None:
        return bool(flag)
    return _parse_positive_dn(row.get("DN_mm", row.get("dn_mm", row.get("dn")))) is not None


def _backfill_pressure_pipe_velocity_from_q_and_dn(row):
    if not isinstance(row, dict):
        return None
    if _normalize_locked_velocity_value(row.get("V")) is not None:
        return _normalize_locked_velocity_value(row.get("V"))

    q_value = _normalize_positive_flow_value(row.get("Q"))
    dn_mm = _parse_positive_dn(row.get("DN_mm"))
    if q_value is None or dn_mm is None or not _row_has_valid_pressurized_dn_mm(row):
        return None

    try:
        from 推求水面线.core.pressure_pipe_calc import calc_pipe_velocity
    except ImportError:
        return None

    velocity = _normalize_locked_velocity_value(calc_pipe_velocity(q_value, dn_mm / 1000.0))
    if velocity is not None:
        row["V"] = velocity
    return velocity


def _extract_pressurized_pipe_material(node):
    params = getattr(node, "section_params", {}) or {}
    value = params.get("pipe_material", getattr(node, "pipe_material", ""))
    return str(value or "").strip() or "球墨铸铁管"


def _extract_pressurized_total_head_loss(node):
    value = getattr(node, "head_loss_siphon", None)
    try:
        number = float(value)
    except (TypeError, ValueError):
        number = None
    if number is None or not math.isfinite(number) or number <= 0:
        value = getattr(node, "external_head_loss", None)
    return _normalize_pressure_pipe_total_head_loss_value(value)


def _extract_pressurized_param_entities(nodes, structure_kind):
    entities = {}
    entity_order = []
    invalid = {}
    invalid_order = []
    if not nodes:
        return [], []

    is_siphon = (structure_kind == "siphon")
    for node in nodes:
        if getattr(node, "is_transition", False) or getattr(node, "is_auto_inserted_channel", False):
            continue

        st_str = _struct_val(getattr(node, "structure_type", None))
        if is_siphon:
            matched = bool(getattr(node, "is_inverted_siphon", False) or ("倒虹吸" in st_str))
        else:
            matched = ("有压管道" in st_str)
        if not matched:
            continue

        base_name = _normalize_pressurized_name(getattr(node, "name", ""), structure_kind)
        dn_mm = _extract_pressurized_dn_mm(node)
        flow_section_idx = _parse_flow_section_index(getattr(node, "flow_section", ""))

        if flow_section_idx is None:
            invalid_key = (base_name, structure_kind)
            if invalid_key not in invalid:
                invalid[invalid_key] = {
                    "name": base_name,
                    "display_name": base_name,
                    "structure_kind": structure_kind,
                }
                invalid_order.append(invalid_key)
            continue

        entity_key = (base_name, flow_section_idx, structure_kind)
        if entity_key not in entities:
            entities[entity_key] = {
                "name": base_name,
                "flow_section": flow_section_idx,
                "display_name": f"{base_name}-{_segment_label_from_index(flow_section_idx)}",
                "structure_kind": structure_kind,
                "DN_mm": 0,
                "pipe_material": "",
            }
            entity_order.append(entity_key)

        if dn_mm is not None:
            entities[entity_key]["DN_mm"] = max(entities[entity_key]["DN_mm"], dn_mm)
        if not entities[entity_key].get("pipe_material"):
            entities[entity_key]["pipe_material"] = _extract_pressurized_pipe_material(node)

        velocity = _normalize_locked_velocity_value(getattr(node, "velocity", None))
        if velocity is not None:
            entities[entity_key]["V"] = velocity

        total_head_loss = _extract_pressurized_total_head_loss(node)
        if total_head_loss is not None:
            entities[entity_key]["total_head_loss"] = total_head_loss

    entity_rows = []
    for entity_key in entity_order:
        item = dict(entities[entity_key])
        item["pipe_material"] = str(item.get("pipe_material") or "").strip() or "球墨铸铁管"
        item["DN_mm"] = _normalize_dn_mm(item.get("DN_mm"), 1500)
        entity_rows.append(item)
    invalid_rows = [invalid[key] for key in invalid_order]
    return entity_rows, invalid_rows


def _resolve_pressure_pipe_dialog_group_flow_section(group):
    """从窗口分组对象里尽量解析出流量段，保证导出按流量段逐行输出。"""
    rows = []
    candidates = []
    if isinstance(group, dict):
        rows = group.get("rows") or []
        candidates.extend(
            (
                group.get("flow_section"),
                group.get("identity"),
                group.get("storage_key"),
                group.get("display_name"),
            )
        )
    else:
        rows = getattr(group, "rows", []) or []
        candidates.extend(
            (
                getattr(group, "flow_section", None),
                getattr(group, "identity", None),
                getattr(group, "storage_key", None),
                getattr(group, "display_name", None),
            )
        )

    for candidate in candidates:
        flow_section_idx = _parse_flow_section_index(candidate)
        if flow_section_idx is not None:
            return flow_section_idx

    for node in rows:
        flow_section_idx = _parse_flow_section_index(getattr(node, "flow_section", ""))
        if flow_section_idx is not None:
            return flow_section_idx
    return None


def _normalize_pressure_pipe_dialog_group_dn_mm(group):
    """把窗口分组对象中的米制直径统一折算成导出所需的 DN(mm)。"""
    if isinstance(group, dict):
        raw_value = group.get("DN_mm", group.get("diameter"))
    else:
        raw_value = getattr(group, "DN_mm", getattr(group, "diameter", None))
    try:
        number = float(raw_value)
    except (TypeError, ValueError):
        number = None
    if number is not None and math.isfinite(number) and number > 0:
        raw_value = number * 1000 if number < 20 else number
    return _normalize_dn_mm(raw_value, 1500)


def _resolve_pressure_pipe_dialog_group_structure_type(group):
    """解析窗口分组的结构形式文本。"""
    candidates = []
    rows = []
    if isinstance(group, dict):
        candidates.extend(
            (
                group.get("pressure_pipe_structure_type"),
                group.get("structure_type"),
            )
        )
        rows = group.get("rows") or []
    else:
        candidates.extend(
            (
                getattr(group, "pressure_pipe_structure_type", None),
                getattr(group, "structure_type", None),
            )
        )
        rows = getattr(group, "rows", []) or []

    for candidate in candidates:
        value = getattr(candidate, "value", candidate)
        text = str(value or "").strip()
        if text:
            return text

    for node in rows:
        getter = getattr(node, "get_structure_type_str", None)
        if callable(getter):
            try:
                text = str(getter() or "").strip()
            except Exception:
                text = ""
            if text:
                return text
        struct_type = getattr(node, "structure_type", None)
        value = getattr(struct_type, "value", struct_type)
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _build_pressure_pipe_param_row_from_dialog_group(group):
    """把有压管道窗口分组对象转换成导出参数行，并保留稳定元数据。"""
    if isinstance(group, dict):
        source = group
    else:
        try:
            source = vars(group)
        except TypeError:
            source = {}
    if not isinstance(source, dict) or not source:
        return None

    flow_section_idx = _resolve_pressure_pipe_dialog_group_flow_section(group)
    if flow_section_idx is None:
        return None

    pipe_material = str(
        source.get("pipe_material", source.get("material_key", ""))
        or source.get("material_key", "")
        or ""
    ).strip() or "球墨铸铁管"
    base_row = _make_pressurized_param_row(
        name=source.get("name"),
        flow_section=flow_section_idx,
        structure_kind="pressure_pipe",
        pipe_material=pipe_material,
        dn_mm=_normalize_pressure_pipe_dialog_group_dn_mm(source),
        display_name=source.get("display_name"),
    )
    structure_type_text = _resolve_pressure_pipe_dialog_group_structure_type(group)
    if structure_type_text:
        base_row["pressure_pipe_structure_type"] = structure_type_text

    metadata = {}
    for key in (
        "group_mode",
        "storage_key",
        "identity",
        "target_row_index",
        "upstream_row_index",
        "route_key",
        "route_display_name",
        "route_start_row_index",
        "route_end_row_index",
        "route_start_mc",
        "route_end_mc",
        "route_ip_points",
        "route_member_keys",
        "segment_start_mc",
        "segment_end_mc",
        "ip_points",
        "plan_total_length",
        "upstream_velocity",
        "downstream_velocity",
        "inlet_transition_form",
        "outlet_transition_form",
        "inlet_transition_zeta",
        "outlet_transition_zeta",
        "has_inlet_transition",
        "has_outlet_transition",
        "inlet_transition_reason",
        "outlet_transition_reason",
        "local_loss_ratio",
    ):
        if key in source:
            metadata[key] = source.get(key)

    q_value = _normalize_positive_flow_value(source.get("Q"))
    if q_value is None:
        q_value = _normalize_positive_flow_value(source.get("design_flow"))
    if q_value is not None:
        metadata["Q"] = q_value

    pipe_material_key = str(source.get("material_key", "") or "").strip()
    if pipe_material_key:
        metadata["pipe_material_key"] = pipe_material_key

    _copy_pressurized_row_metadata(base_row, metadata)
    return base_row


def _xxpipe_longitudinal_node_get(node, key, default=None):
    if isinstance(node, dict):
        return node.get(key, default)
    return getattr(node, key, default)


def _xxpipe_longitudinal_node_float(node, key, *, required=True, default=None):
    value = _xxpipe_longitudinal_node_get(node, key, default)
    if value is None:
        if required:
            raise ValueError(f"xx管纵断面节点缺少有效字段: {key}")
        return None
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        if required:
            raise ValueError(f"xx管纵断面节点字段 {key} 不是有效数值: {value!r}") from exc
        return None
    if not math.isfinite(number):
        if required:
            raise ValueError(f"xx管纵断面节点字段 {key} 不是有限数值: {value!r}")
        return None
    return number


def _normalize_xxpipe_longitudinal_nodes(longitudinal_nodes):
    if not longitudinal_nodes or len(longitudinal_nodes) < 2:
        raise ValueError("xx管轴线高程采样至少需要 2 个纵断面节点")

    normalized = []
    for idx, node in enumerate(longitudinal_nodes):
        chainage = _xxpipe_longitudinal_node_float(node, "chainage")
        elevation = _xxpipe_longitudinal_node_float(node, "elevation")
        turn_type_raw = _xxpipe_longitudinal_node_get(node, "turn_type", "NONE")
        turn_type = getattr(turn_type_raw, "name", turn_type_raw)
        normalized.append(
            {
                "index": idx,
                "chainage": chainage,
                "elevation": elevation,
                "turn_type": str(turn_type or "NONE").strip().upper(),
                "vertical_curve_radius": _xxpipe_longitudinal_node_float(
                    node, "vertical_curve_radius", required=False, default=0.0
                ) or 0.0,
                "arc_center_s": _xxpipe_longitudinal_node_float(
                    node, "arc_center_s", required=False, default=None
                ),
                "arc_center_z": _xxpipe_longitudinal_node_float(
                    node, "arc_center_z", required=False, default=None
                ),
                "arc_end_chainage": _xxpipe_longitudinal_node_float(
                    node, "arc_end_chainage", required=False, default=None
                ),
                "arc_theta_rad": _xxpipe_longitudinal_node_float(
                    node, "arc_theta_rad", required=False, default=None
                ),
            }
        )

    normalized.sort(key=lambda item: (item["chainage"], item["index"]))
    return normalized


def _is_xxpipe_arc_segment_start(node):
    return (
        node["turn_type"] == "ARC"
        and node["vertical_curve_radius"] > 0
        and node["arc_center_s"] is not None
        and node["arc_center_z"] is not None
        and node["arc_end_chainage"] is not None
        and node["arc_theta_rad"] is not None
        and node["arc_theta_rad"] > 0
        and node["arc_end_chainage"] > node["chainage"]
    )


def _sample_xxpipe_arc_segment_elevation(node, station_mc):
    start_chainage = node["chainage"]
    start_elevation = node["elevation"]
    center_s = node["arc_center_s"]
    center_z = node["arc_center_z"]
    radius = node["vertical_curve_radius"]

    inside_start = max(0.0, radius ** 2 - (start_chainage - center_s) ** 2)
    root_start = math.sqrt(inside_start)
    eta = 1 if abs(center_z + root_start - start_elevation) <= abs(center_z - root_start - start_elevation) else -1

    inside_station = radius ** 2 - (station_mc - center_s) ** 2
    if inside_station < -1e-8:
        raise ValueError(
            f"station {station_mc:.6f} 超出 xx管轴线圆弧几何定义，无法根据纵断面节点求高程"
        )
    return center_z + eta * math.sqrt(max(0.0, inside_station))


def sample_xxpipe_centerline_elevation(longitudinal_nodes, station_mc):
    """按桩号求 xx管 管中心线高程；超出纵断面覆盖范围时拒绝外推。"""
    nodes = _normalize_xxpipe_longitudinal_nodes(longitudinal_nodes)
    station_value = _xxpipe_longitudinal_node_float({"station_mc": station_mc}, "station_mc")
    tol = _XXPIPE_PROFILE_STATION_TOL

    for node in nodes:
        if abs(node["chainage"] - station_value) <= tol:
            return node["elevation"]

    coverage_start = nodes[0]["chainage"]
    coverage_end = nodes[-1]["chainage"]
    if station_value < coverage_start - tol or station_value > coverage_end + tol:
        raise ValueError(
            f"station {station_value:.6f} 超出 xx管轴线高程覆盖范围 "
            f"[{coverage_start:.6f}, {coverage_end:.6f}]，不允许外推"
        )

    for idx, current in enumerate(nodes[:-1]):
        nxt = nodes[idx + 1]
        segment_start = current["chainage"]
        if _is_xxpipe_arc_segment_start(current):
            segment_end = current["arc_end_chainage"]
            if segment_start - tol <= station_value <= segment_end + tol:
                return _sample_xxpipe_arc_segment_elevation(current, station_value)
            continue

        segment_end = nxt["chainage"]
        if segment_start - tol <= station_value <= segment_end + tol:
            ds = segment_end - segment_start
            if abs(ds) <= _XXPIPE_PROFILE_GEOMETRY_TOL:
                return current["elevation"]
            ratio = (station_value - segment_start) / ds
            return current["elevation"] + (nxt["elevation"] - current["elevation"]) * ratio

    raise ValueError(
        f"station {station_value:.6f} 超出 xx管轴线高程覆盖范围 "
        f"[{coverage_start:.6f}, {coverage_end:.6f}]，不允许外推"
    )


def find_xxpipe_axis_elevation_coverage_gaps(longitudinal_nodes, station_mcs):
    """批量检查给定桩号是否全部落在 xx管纵断面覆盖范围内。"""
    missing = []
    for station_mc in station_mcs or []:
        try:
            sample_xxpipe_centerline_elevation(longitudinal_nodes, station_mc)
        except ValueError as exc:
            if "超出 xx管轴线高程覆盖范围" not in str(exc):
                raise
            missing.append(station_mc)
    return missing


def _sample_xxpipe_centerline_elevation(longitudinal_nodes, station_mc):
    return sample_xxpipe_centerline_elevation(longitudinal_nodes, station_mc)


def _validate_xxpipe_centerline_coverage(longitudinal_nodes, station_mcs):
    return find_xxpipe_axis_elevation_coverage_gaps(longitudinal_nodes, station_mcs)


def _extract_pressure_pipe_calc_contexts(nodes, proj_settings=None):
    """提取有压管道断面表所需的水头损失计算上下文。"""
    try:
        from 推求水面线.core.pressure_pipe_calc import calc_turn_angle
    except ImportError:
        return {}

    contexts = {}
    order = []
    if not nodes:
        return contexts

    for idx, node in enumerate(nodes):
        if getattr(node, "is_transition", False) or getattr(node, "is_auto_inserted_channel", False):
            continue
        st_str = _struct_val(getattr(node, "structure_type", None))
        if "有压管道" not in st_str and not getattr(node, "is_pressure_pipe", False):
            continue

        flow_section_idx = _parse_flow_section_index(getattr(node, "flow_section", ""))
        if flow_section_idx is None:
            continue

        base_name = _normalize_pressurized_name(getattr(node, "name", ""), "pressure_pipe")
        key = (base_name, flow_section_idx)
        if key not in contexts:
            contexts[key] = {
                "rows": [],
                "row_indices": [],
                "inlet_row_index": -1,
                "outlet_row_index": -1,
                "inlet_transition_form": str(getattr(proj_settings, "siphon_transition_inlet_form", "反弯扭曲面") or "反弯扭曲面"),
                "outlet_transition_form": str(getattr(proj_settings, "siphon_transition_outlet_form", "反弯扭曲面") or "反弯扭曲面"),
                "inlet_transition_zeta": float(getattr(proj_settings, "siphon_transition_inlet_zeta", 0.10) or 0.10),
                "outlet_transition_zeta": float(getattr(proj_settings, "siphon_transition_outlet_zeta", 0.20) or 0.20),
                "upstream_velocity": 0.0,
                "downstream_velocity": 0.0,
            }
            order.append(key)

        ctx = contexts[key]
        ctx["rows"].append(node)
        ctx["row_indices"].append(idx)

        params = getattr(node, "section_params", {}) or {}
        in_out_raw = str(params.get("in_out_raw", "") or "").strip()
        in_out = getattr(node, "in_out", None)
        if in_out_raw == "进" or str(in_out) == "InOutType.INLET":
            ctx["inlet_row_index"] = idx
        elif in_out_raw == "出" or str(in_out) == "InOutType.OUTLET":
            ctx["outlet_row_index"] = idx

    for key in order:
        ctx = contexts[key]
        rows = ctx["rows"]
        row_indices = ctx["row_indices"]
        if not rows:
            continue

        if ctx["inlet_row_index"] < 0:
            ctx["inlet_row_index"] = row_indices[0]
        if ctx["outlet_row_index"] < 0:
            ctx["outlet_row_index"] = row_indices[-1]

        ip_points = []
        for row in rows:
            ip_points.append({
                "x": getattr(row, "x", 0.0) or 0.0,
                "y": getattr(row, "y", 0.0) or 0.0,
                "turn_radius": getattr(row, "turn_radius", 0.0) or 0.0,
                "turn_angle": 0.0,
            })
        for i in range(1, len(ip_points) - 1):
            p_prev = (ip_points[i - 1]["x"], ip_points[i - 1]["y"])
            p_curr = (ip_points[i]["x"], ip_points[i]["y"])
            p_next = (ip_points[i + 1]["x"], ip_points[i + 1]["y"])
            try:
                ip_points[i]["turn_angle"] = calc_turn_angle(p_prev, p_curr, p_next)
            except Exception:
                ip_points[i]["turn_angle"] = 0.0
        ctx["ip_points"] = ip_points

        inlet_idx = ctx["inlet_row_index"]
        outlet_idx = ctx["outlet_row_index"]

        for i in range(inlet_idx - 1, -1, -1):
            upstream = nodes[i]
            if getattr(upstream, "is_transition", False):
                continue
            up_struct = _struct_val(getattr(upstream, "structure_type", None))
            if "有压管道" in up_struct or getattr(upstream, "is_pressure_pipe", False):
                continue
            ctx["upstream_velocity"] = float(getattr(upstream, "velocity", 0.0) or 0.0)
            break

        for i in range(outlet_idx + 1, len(nodes)):
            downstream = nodes[i]
            if getattr(downstream, "is_transition", False):
                continue
            down_struct = _struct_val(getattr(downstream, "structure_type", None))
            if "有压管道" in down_struct or getattr(downstream, "is_pressure_pipe", False):
                continue
            ctx["downstream_velocity"] = float(getattr(downstream, "velocity", 0.0) or 0.0)
            break

        ctx.pop("rows", None)
        ctx.pop("row_indices", None)

    return contexts


def _normalize_pressure_pipe_total_head_loss_value(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number) or number < 0:
        return None
    return round(number, 4)


def _normalize_pressure_pipe_total_length_value(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number) or number <= 0:
        return None
    return round(number, 4)


def _normalize_pressure_pipe_station_value(value):
    """标准化桩号值；允许 0，非法值返回 None。"""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return round(number, 6)


def _resolve_pressure_pipe_row_segment_length(row):
    """优先用原始分组起止桩号求长度，避免误用建筑物摘要长度。"""
    if not isinstance(row, dict):
        return None

    start_station = _normalize_pressure_pipe_station_value(row.get("segment_start_mc"))
    end_station = _normalize_pressure_pipe_station_value(row.get("segment_end_mc"))
    if start_station is None or end_station is None:
        start_station = _normalize_pressure_pipe_station_value(row.get("route_start_mc"))
        end_station = _normalize_pressure_pipe_station_value(row.get("route_end_mc"))
    if start_station is None or end_station is None:
        return None

    segment_length = end_station - start_station
    if segment_length <= 0:
        return None
    return round(segment_length, 4)


def _sum_pressure_pipe_flow_section_total_length(rows):
    """按流量段下全部原始分组累计总长度。"""
    total_length = 0.0
    seen_segments = set()

    for row in rows or []:
        segment_length = _resolve_pressure_pipe_row_segment_length(row)
        if segment_length is None:
            continue
        start_station = _normalize_pressure_pipe_station_value(row.get("segment_start_mc"))
        end_station = _normalize_pressure_pipe_station_value(row.get("segment_end_mc"))
        if start_station is None or end_station is None:
            start_station = _normalize_pressure_pipe_station_value(row.get("route_start_mc"))
            end_station = _normalize_pressure_pipe_station_value(row.get("route_end_mc"))
        segment_key = (
            str(row.get("identity", "") or "").strip(),
            start_station,
            end_station,
        )
        if segment_key in seen_segments:
            continue
        seen_segments.add(segment_key)
        total_length += segment_length

    if total_length <= 0:
        return None
    return round(total_length, 4)


def _get_panel_pressure_pipe_export_results(panel, rows):
    getter = getattr(panel, "get_pressure_pipe_export_results", None)
    if not callable(getter):
        return {}
    try:
        data = getter(rows)
    except TypeError:
        data = getter()
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def _get_panel_pressure_pipe_longitudinal_nodes_for_export(panel, rows):
    getter = getattr(panel, "get_pressure_pipe_longitudinal_nodes_for_export", None)
    if not callable(getter):
        return {}
    try:
        data = getter(rows)
    except TypeError:
        data = getter()
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def _get_panel_xxpipe_manager_config_by_identity(panel, rows):
    manager = getattr(panel, "_pressure_pipe_manager", None)
    to_dict = getattr(manager, "to_dict", None)
    if not callable(to_dict):
        return {}
    try:
        raw = to_dict() or {}
    except Exception:
        return {}
    pipes = raw.get("pipes", {}) if isinstance(raw, dict) else {}
    targets = {}
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        identity = str(row.get("identity", "") or "").strip()
        if not identity:
            identity = make_pressure_pipe_identity(row.get("flow_section"), row.get("name"))
        name = str(row.get("name", "") or "").strip() or "未命名"
        flow_section = str(row.get("flow_section", "") or "").strip()
        targets[identity] = {"name": name, "flow_section": flow_section}

    resolved = {}
    for key, pipe_data in pipes.items():
        if not isinstance(pipe_data, dict):
            continue
        key_text = str(key or "").strip()
        name = str(pipe_data.get("name", "") or "").strip()
        flow_section = str(pipe_data.get("flow_section", "") or "").strip()
        candidate_identities = []
        if key_text and "::" in key_text:
            candidate_identities.append(key_text)
            if not name:
                name = key_text.split("::", 1)[1].strip()
        candidate_identities.append(make_pressure_pipe_identity(flow_section, name or key_text))
        for identity, target in targets.items():
            if identity in candidate_identities:
                resolved[identity] = copy.deepcopy(pipe_data)
                break
            if (
                target["name"] == name
                and target["flow_section"] == flow_section
            ):
                resolved[identity] = copy.deepcopy(pipe_data)
                break
    return resolved


def _copy_pressurized_row_metadata(base_row, source_row, *, override_keys=()):
    """把原行中的稳定元数据补回标准化结果，避免导出链断字段。"""
    if not isinstance(base_row, dict) or not isinstance(source_row, dict):
        return base_row
    override_key_set = set(override_keys or ())
    for key in override_key_set:
        if key in source_row:
            base_row[key] = copy.deepcopy(source_row[key])
    for key, value in source_row.items():
        if key in override_key_set:
            continue
        if key not in base_row:
            base_row[key] = copy.deepcopy(value)
    return base_row


def _build_pressurized_output_row(source_row, pipe_material, dn_mm):
    """按当前界面输入生成标准化导出行，并保留原始稳定元数据。"""
    base_row = _make_pressurized_param_row(
        name=source_row.get("name"),
        flow_section=source_row.get("flow_section"),
        structure_kind=source_row.get("structure_kind"),
        pipe_material=pipe_material,
        dn_mm=dn_mm,
        display_name=source_row.get("display_name"),
    )
    skip_keys = {
        "pipe_material",
        "material",
        "material_key",
        "DN_mm",
        "dn_mm",
        "dn",
        "D",
        "_has_valid_dn_mm",
        "display_name",
        "flow_section",
        "structure_kind",
        "dialog_row_kind",
        "dialog_target_identities",
        "dialog_target_rows",
    }
    for key, value in (source_row or {}).items():
        if key in skip_keys:
            continue
        base_row[key] = copy.deepcopy(value)
    return base_row


def _attach_pressure_pipe_calc_contexts_to_rows(rows, calc_contexts):
    contexts = calc_contexts or {}
    for row in rows or []:
        key = (row.get("name"), row.get("flow_section"))
        context = contexts.get(key)
        if not context:
            continue
        for field in (
            "ip_points",
            "upstream_velocity",
            "downstream_velocity",
            "inlet_transition_form",
            "outlet_transition_form",
            "inlet_transition_zeta",
            "outlet_transition_zeta",
        ):
            if field in context:
                row[field] = copy.deepcopy(context[field])
    return rows


def _attach_pressure_pipe_export_results_to_rows(rows, panel=None):
    results_by_identity = _get_panel_pressure_pipe_export_results(panel, rows)
    if not results_by_identity:
        return rows
    for row in rows or []:
        row_identity = str(row.get("identity", "") or "").strip()
        legacy_identity = make_pressure_pipe_identity(row.get("flow_section"), row.get("name"))
        result = None
        if row_identity:
            result = results_by_identity.get(row_identity)
        if not isinstance(result, dict):
            result = results_by_identity.get(legacy_identity)
        if not isinstance(result, dict):
            continue
        velocity = _normalize_locked_velocity_value(
            result.get("pipe_velocity", result.get("velocity"))
        )
        if velocity is not None:
            row["V"] = velocity
        total_length = _normalize_pressure_pipe_total_length_value(
            result.get("total_length", result.get("plan_total_length"))
        )
        if total_length is not None:
            row["total_length"] = total_length
        total_head_loss = _normalize_pressure_pipe_total_head_loss_value(
            result.get("total_head_loss")
        )
        if total_head_loss is not None:
            row["total_head_loss"] = total_head_loss
    return rows


def _apply_pressure_pipe_length_fallbacks(rows):
    for row in rows or []:
        total_length = _normalize_pressure_pipe_total_length_value(row.get("total_length"))
        if total_length is not None:
            row["total_length"] = total_length
            continue
        fallback_length = _normalize_pressure_pipe_total_length_value(row.get("plan_total_length"))
        if fallback_length is not None:
            row["total_length"] = fallback_length
    return rows


def _apply_pressure_pipe_velocity_fallbacks(rows):
    for row in rows or []:
        _backfill_pressure_pipe_velocity_from_q_and_dn(row)
    return rows


def _normalize_pressurized_cache_rows(rows, structure_kind, default_material="球墨铸铁管"):
    normalized = []
    for row in rows or []:
        if isinstance(row, dict):
            row_kind = str(row.get("structure_kind") or structure_kind or "").strip() or structure_kind
            if not row_kind:
                continue
            base_row = _make_pressurized_param_row(
                name=row.get("name"),
                flow_section=row.get("flow_section"),
                structure_kind=row_kind,
                pipe_material=row.get("pipe_material", row.get("material", default_material)),
                dn_mm=row.get("DN_mm", row.get("dn_mm", row.get("dn", 1500))),
                display_name=row.get("display_name"),
            )
            if "_has_valid_dn_mm" in row:
                base_row["_has_valid_dn_mm"] = bool(row.get("_has_valid_dn_mm"))
            _copy_pressurized_row_metadata(base_row, row)
            normalized.append(base_row)
            continue

        if not isinstance(row, (tuple, list)) or len(row) < 3:
            continue
        normalized.append(
            _make_pressurized_param_row(
                name=row[0],
                flow_section=None,
                structure_kind=structure_kind,
                pipe_material=row[1],
                dn_mm=row[2],
                display_name=row[0],
            )
        )
    return normalized


def _serialize_pressurized_cache_rows(rows, structure_kind):
    serialized = []
    for row in _normalize_pressurized_cache_rows(rows, structure_kind):
        serialized.append(copy.deepcopy(row))
    return serialized


def _prepare_pressure_pipe_export_rows(rows, panel=None, calc_contexts=None):
    prepared_rows = _normalize_pressurized_cache_rows(rows, "pressure_pipe")
    _attach_pressure_pipe_calc_contexts_to_rows(prepared_rows, calc_contexts)
    _attach_pressure_pipe_export_results_to_rows(prepared_rows, panel=panel)
    _apply_pressure_pipe_velocity_fallbacks(prepared_rows)
    _apply_pressure_pipe_length_fallbacks(prepared_rows)
    return prepared_rows


def _normalize_pressure_pipe_flow_section_key(flow_section):
    idx = _parse_flow_section_index(flow_section)
    if idx is not None:
        return str(idx)
    return str(flow_section or "").strip()


def _resolve_pressure_pipe_characteristic_export_summary(panel, rows=None):
    getter = getattr(panel, "get_pressure_pipe_characteristic_export_summary", None)
    if not callable(getter):
        return {}
    try:
        result = getter(rows)
    except TypeError:
        result = getter()
    except Exception:
        return {}
    return result if isinstance(result, dict) else {}


def _merge_pressure_pipe_export_rows_by_flow_section(rows, panel=None):
    if not rows:
        return []

    summary_by_flow_section = _resolve_pressure_pipe_characteristic_export_summary(panel, rows)
    grouped_rows = {}
    row_order = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        flow_section = row.get("flow_section")
        if flow_section is None:
            continue
        if flow_section not in grouped_rows:
            grouped_rows[flow_section] = []
            row_order.append(flow_section)
        grouped_rows[flow_section].append(row)

    merged_rows = []
    show_building_characteristics = False
    preferred_keys = (
        "Q",
        "Q_inc",
        "pipe_material",
        "DN_mm",
        "V",
        "plan_total_length",
        "total_length",
        "total_head_loss",
        "friction_params",
        "pressure_f",
        "pressure_m",
        "pressure_b",
        "pipe_material_key",
        "ip_points",
        "upstream_velocity",
        "downstream_velocity",
        "inlet_transition_form",
        "outlet_transition_form",
        "inlet_transition_zeta",
        "outlet_transition_zeta",
    )

    for flow_section in row_order:
        items = grouped_rows[flow_section]
        base = copy.deepcopy(items[0])
        flow_section_key = _normalize_pressure_pipe_flow_section_key(flow_section)
        segment_label = _segment_label_from_index(flow_section) or flow_section_key or base.get("name", "")
        base["name"] = segment_label
        base["display_name"] = segment_label
        for key in preferred_keys:
            for item in items:
                value = item.get(key)
                if value in (None, "", "-"):
                    continue
                base[key] = copy.deepcopy(value)
                break

        summary = summary_by_flow_section.get(flow_section_key)
        if summary is None:
            summary = summary_by_flow_section.get(flow_section)
        summary_total_length = None
        if isinstance(summary, dict):
            for key in (
                "start_water_level",
                "end_water_level",
                "tunnel_count",
                "tunnel_length",
                "directional_drill_count",
                "directional_drill_length",
                "jacking_count",
                "jacking_length",
            ):
                if key in summary:
                    base[key] = copy.deepcopy(summary.get(key))
            summary_total_length = _normalize_pressure_pipe_total_length_value(summary.get("total_length"))

        segment_total_length = _sum_pressure_pipe_flow_section_total_length(items)
        if segment_total_length is None:
            segment_total_length = summary_total_length
        if segment_total_length is None:
            segment_total_length = _normalize_pressure_pipe_total_length_value(base.get("total_length"))
        if segment_total_length is None:
            segment_total_length = _normalize_pressure_pipe_total_length_value(base.get("plan_total_length"))
        if segment_total_length is not None:
            base["total_length"] = segment_total_length
            base["plan_total_length"] = segment_total_length

        tunnel_count = 0
        directional_drill_count = 0
        jacking_count = 0
        try:
            tunnel_count = max(int(float(base.get("tunnel_count", 0) or 0)), 0)
        except (TypeError, ValueError):
            tunnel_count = 0
        try:
            directional_drill_count = max(int(float(base.get("directional_drill_count", 0) or 0)), 0)
        except (TypeError, ValueError):
            directional_drill_count = 0
        try:
            jacking_count = max(int(float(base.get("jacking_count", 0) or 0)), 0)
        except (TypeError, ValueError):
            jacking_count = 0
        if tunnel_count > 0 or directional_drill_count > 0 or jacking_count > 0:
            show_building_characteristics = True
        merged_rows.append(base)

    for row in merged_rows:
        row["show_building_characteristics"] = show_building_characteristics
    return merged_rows


def _iter_xxpipe_export_nodes(nodes):
    for node in nodes or []:
        if getattr(node, "is_transition", False):
            continue
        if getattr(node, "is_auto_inserted_channel", False):
            continue
        yield node


def _make_xxpipe_identity_from_node(node):
    row_identity = str(getattr(node, "pressure_pipe_row_identity", "") or "").strip()
    if row_identity:
        return row_identity
    return make_pressure_pipe_identity(
        getattr(node, "flow_section", ""),
        getattr(node, "name", ""),
    )


def _build_xxpipe_segment_records(items):
    segments = []
    for item in items:
        text = str(item.get("text", "") or "").strip()
        if not text:
            continue
        station_mc = float(item.get("station_mc", 0.0) or 0.0)
        identity = str(item.get("identity", "") or "").strip()
        if (
            segments
            and segments[-1]["text"] == text
            and segments[-1]["identity"] == identity
        ):
            segments[-1]["mcs"].append(station_mc)
            continue
        segments.append({
            "text": text,
            "identity": identity,
            "mcs": [station_mc],
        })

    out = []
    for segment in segments:
        mc_list = segment["mcs"]
        start_mc = mc_list[0]
        end_mc = mc_list[-1]
        mid_mc = _resolve_segment_mid_mc(start_mc, end_mc, [])
        out.append({
            "text": segment["text"],
            "identity": segment["identity"],
            "start_mc": start_mc,
            "end_mc": end_mc,
            "mid_mc": mid_mc,
        })
    return out


def _build_xxpipe_material_segment_records(items):
    segments = []
    for item in items:
        text = str(item.get("text", "") or "").strip()
        if not text:
            continue
        station_mc = float(item.get("station_mc", 0.0) or 0.0)
        identity = str(item.get("identity", "") or "").strip()
        struct_name = str(item.get("structure_name", "") or "").strip()
        flow_section_key = _normalize_pressure_pipe_flow_section_key(item.get("flow_section"))
        merge_mode = "named_structure" if _is_xxpipe_named_structure(struct_name) else "plain_pressure_pipe"
        can_merge = False
        if segments:
            prev = segments[-1]
            if merge_mode == "plain_pressure_pipe":
                can_merge = (
                    prev.get("merge_mode") == "plain_pressure_pipe"
                    and prev.get("flow_section_key") == flow_section_key
                    and prev.get("text") == text
                )
            else:
                can_merge = (
                    prev.get("merge_mode") == "named_structure"
                    and prev.get("identity") == identity
                    and prev.get("text") == text
                )
        if can_merge:
            segments[-1]["mcs"].append(station_mc)
            continue
        segments.append({
            "text": text,
            "identity": identity,
            "mcs": [station_mc],
            "merge_mode": merge_mode,
            "flow_section_key": flow_section_key,
        })

    out = []
    for segment in segments:
        mc_list = segment["mcs"]
        start_mc = mc_list[0]
        end_mc = mc_list[-1]
        mid_mc = _resolve_segment_mid_mc(start_mc, end_mc, [])
        out.append({
            "text": segment["text"],
            "identity": segment["identity"],
            "start_mc": start_mc,
            "end_mc": end_mc,
            "mid_mc": mid_mc,
            "merge_mode": segment.get("merge_mode", ""),
            "flow_section_key": segment.get("flow_section_key", ""),
        })
    return out


def _build_xxpipe_profile_data(
    nodes,
    longitudinal_nodes_by_identity,
    *,
    station_prefix="",
    manager_config_by_identity=None,
):
    raw_visible_nodes = list(_iter_xxpipe_export_nodes(nodes))
    if not raw_visible_nodes:
        raise ValueError("xx管纵断面导出没有可用节点")

    invalid_nodes = []
    for node in raw_visible_nodes:
        struct_name = _struct_val(getattr(node, "structure_type", None))
        if _is_xxpipe_allowed_structure(struct_name):
            continue
        invalid_nodes.append(
            f"{getattr(node, 'name', '') or '未命名'}({struct_name or '未知结构'})"
        )
    if invalid_nodes:
        raise ValueError("xx管模式仅允许有压管道/定向钻/顶管，检测到冲突结构：\n" + "；".join(invalid_nodes))

    station_targets, station_errors = resolve_xxpipe_profile_station_targets(
        raw_visible_nodes,
        station_prefix=station_prefix,
    )
    if station_errors:
        raise ValueError(
            "以下节点缺少可用桩号，无法读取 xx管 轴线高程：\n"
            + "；".join(
                f"{item['label']}（{item['reason']}）"
                for item in station_errors
            )
        )

    visible_nodes = []
    for target in station_targets:
        node_copy = copy.copy(target["node"])
        setattr(node_copy, "station_MC", float(target["station_mc"]))
        visible_nodes.append(node_copy)

    profile_text_nodes = _build_profile_text_nodes(visible_nodes)
    ip_records = _build_ip_related_row_records(visible_nodes, station_prefix).get("ip_name", [])
    manager_map = manager_config_by_identity or {}
    long_map = longitudinal_nodes_by_identity or {}

    centerline_points = []
    centerline_records = []
    missing_axis = []
    for node in profile_text_nodes:
        station_mc = _profile_station_value(node)
        identity = _make_xxpipe_identity_from_node(node)
        long_nodes = long_map.get(identity) or []
        if not long_nodes:
            missing_axis.append(f"{identity} 缺少轴线纵断面")
            continue
        try:
            centerline_elev = sample_xxpipe_centerline_elevation(long_nodes, station_mc)
        except ValueError as exc:
            if "超出 xx管轴线高程覆盖范围" not in str(exc):
                raise
            try:
                station_text = ProjectSettings.format_station(station_mc, station_prefix)
            except Exception:
                station_text = f"{station_mc:.3f}"
            missing_axis.append(f"{identity}@{station_text}")
            continue
        centerline_points.append((station_mc, centerline_elev))
        centerline_records.append({
            "identity": identity,
            "station_mc": station_mc,
            "elevation": centerline_elev,
        })

    if missing_axis:
        raise ValueError("以下节点缺少可用的 xx管 轴线高程覆盖：\n" + "；".join(missing_axis))

    building_segments = _build_xxpipe_segment_records([
        {
            "identity": _make_xxpipe_identity_from_node(node),
            "station_mc": _profile_station_value(node),
            "text": _get_xxpipe_building_display_name(
                _struct_val(getattr(node, "structure_type", None)),
                getattr(node, "name", ""),
            ),
        }
        for node in visible_nodes
    ])

    material_segments = []
    for node in visible_nodes:
        identity = _make_xxpipe_identity_from_node(node)
        manager_row = manager_map.get(identity, {})
        struct_name = _struct_val(getattr(node, "structure_type", None))
        row = {
            "pipe_material": _extract_pressurized_pipe_material(node),
            "DN_mm": _extract_pressurized_dn_mm(node),
        }
        if not row["pipe_material"] and isinstance(manager_row, dict):
            row["pipe_material"] = manager_row.get("material_key", "")
        if row["DN_mm"] is None and isinstance(manager_row, dict):
            row["DN_mm"] = manager_row.get("D")
        material_segments.append({
            "identity": identity,
            "station_mc": _profile_station_value(node),
            "flow_section": getattr(node, "flow_section", ""),
            "structure_name": struct_name,
            "text": _format_xxpipe_pipe_material_text(row),
        })
    material_segments = _build_xxpipe_material_segment_records(material_segments)

    return {
        "profile_text_nodes": profile_text_nodes,
        "ip_records": ip_records,
        "centerline_points": centerline_points,
        "centerline_records": centerline_records,
        "building_segments": building_segments,
        "material_segments": material_segments,
    }


def _build_xxpipe_identity_rows(nodes):
    rows = []
    seen = set()
    for node in _iter_xxpipe_export_nodes(nodes):
        identity = _make_xxpipe_identity_from_node(node)
        if identity in seen:
            continue
        seen.add(identity)
        rows.append(
            {
                "name": getattr(node, "name", ""),
                "flow_section": getattr(node, "flow_section", ""),
                "identity": identity,
            }
        )
    return rows


def _get_panel_channel_level_text(panel):
    combo = getattr(panel, "channel_level_combo", None)
    current_text = getattr(combo, "currentText", None)
    if callable(current_text):
        try:
            text = current_text()
        except Exception:
            text = ""
        if str(text or "").strip():
            return str(text).strip()

    build_settings = getattr(panel, "_build_settings", None)
    if callable(build_settings):
        try:
            proj_settings = build_settings()
        except Exception:
            proj_settings = None
        text = getattr(proj_settings, "channel_level", "")
        if str(text or "").strip():
            return str(text).strip()

    settings_obj = getattr(panel, "_settings", None)
    text = getattr(settings_obj, "channel_level", "")
    return str(text or "").strip()


def _is_panel_xxpipe_mode(panel):
    return _is_xxpipe_channel_level(_get_panel_channel_level_text(panel))


def _resolve_xxpipe_export_source_nodes(panel, fallback_nodes=None):
    nodes, _source = _resolve_section_summary_source_nodes(panel, fallback_nodes=fallback_nodes)
    return list(nodes or [])


def _build_panel_xxpipe_profile_data(panel, nodes, station_prefix=""):
    rows = _build_xxpipe_identity_rows(nodes)
    longitudinal_nodes = _get_panel_pressure_pipe_longitudinal_nodes_for_export(panel, rows)
    manager_config = _get_panel_xxpipe_manager_config_by_identity(panel, rows)
    return _build_xxpipe_profile_data(
        nodes,
        longitudinal_nodes,
        station_prefix=station_prefix,
        manager_config_by_identity=manager_config,
    )


def _translate_xxpipe_export_error(exc):
    """将 xx管 纵断面导出的技术异常翻译成用户可执行的提示。"""
    message = str(exc or "").strip()
    if not message:
        return None
    if "缺少轴线纵断面" in message:
        return "对应整线还没有导入纵断面DXF，请先到表3的有压管道水力计算中导入后再导出。"
    if (
        "以下节点缺少可用的 xx管 轴线高程覆盖" in message
        or "超出 xx管轴线高程覆盖范围" in message
    ):
        return "已导入纵断面DXF，但未覆盖整线全部桩号，请重新导入完整纵断面后再导出。"
    return None


def _collect_xxpipe_full_height_boundary_mcs(profile_data):
    boundary_mcs = []

    def _add(mc):
        try:
            value = float(mc)
        except (TypeError, ValueError):
            return
        if not math.isfinite(value):
            return
        if any(abs(value - existing) <= 1e-9 for existing in boundary_mcs):
            return
        boundary_mcs.append(value)

    for segment in profile_data.get("building_segments", []) or []:
        _add(segment.get("start_mc"))
        _add(segment.get("end_mc"))

    for segment in profile_data.get("material_segments", []) or []:
        if str(segment.get("merge_mode", "") or "").strip() != "named_structure":
            continue
        _add(segment.get("start_mc"))
        _add(segment.get("end_mc"))

    records = profile_data.get("centerline_records", []) or []
    if records:
        _add(records[0].get("station_mc"))
        _add(records[-1].get("station_mc"))

    return sorted(boundary_mcs)


def _is_xxpipe_plain_pressure_pipe_profile_node(node):
    """判断当前文本节点是否为普通有压管道节点。"""
    return _struct_val(getattr(node, "structure_type", None)) == "有压管道"


def _is_xxpipe_named_boundary_profile_node(node):
    """判断当前文本节点是否为命名的定向钻/顶管进出口节点。"""
    if node is None:
        return False
    if not _is_xxpipe_named_structure(_struct_val(getattr(node, "structure_type", None))):
        return False
    if not str(getattr(node, "name", "") or "").strip():
        return False
    return _in_out_val(getattr(node, "in_out", None)) in ("进", "出")


def _collect_xxpipe_lower_half_vertical_line_mcs(profile_text_nodes):
    """收集需要改画下半段竖线的普通节点桩号。"""
    partial_mcs = []
    nodes = list(profile_text_nodes or [])

    def _add(mc):
        try:
            value = float(mc)
        except (TypeError, ValueError):
            return
        if not math.isfinite(value):
            return
        if any(abs(value - existing) <= 1e-9 for existing in partial_mcs):
            return
        partial_mcs.append(value)

    for idx, node in enumerate(nodes):
        if idx == 0 or idx == len(nodes) - 1:
            continue
        if not _is_xxpipe_plain_pressure_pipe_profile_node(node):
            continue

        prev_node = nodes[idx - 1] if idx > 0 else None
        next_node = nodes[idx + 1] if idx + 1 < len(nodes) else None

        if (
            _is_xxpipe_named_boundary_profile_node(next_node)
            and _in_out_val(getattr(next_node, "in_out", None)) == "进"
        ):
            _add(_profile_station_value(node))
            continue

        if (
            _is_xxpipe_named_boundary_profile_node(prev_node)
            and _in_out_val(getattr(prev_node, "in_out", None)) == "出"
        ):
            _add(_profile_station_value(node))

    return sorted(partial_mcs)


def _merge_pressurized_param_defaults(group_items, cached_rows, default_material="球墨铸铁管"):
    """按名称将历史配置与当前分组合并，返回 [(name, material, dn_mm), ...]。"""
    structure_kind = None
    if group_items:
        structure_kind = str(group_items[0].get("structure_kind") or "").strip() or None
    if not structure_kind:
        for row in cached_rows or []:
            if isinstance(row, dict):
                structure_kind = str(row.get("structure_kind") or "").strip() or None
                if structure_kind:
                    break
    structure_kind = structure_kind or "siphon"

    normalized_cache = _normalize_pressurized_cache_rows(
        cached_rows,
        structure_kind=structure_kind,
        default_material=default_material,
    )
    if not group_items:
        return normalized_cache

    exact_cache = {}
    legacy_by_name = {}
    for row in normalized_cache:
        exact_key = (row["name"], row.get("flow_section"), row["structure_kind"])
        if row.get("flow_section"):
            exact_cache[exact_key] = row
        legacy_by_name.setdefault((row["name"], row["structure_kind"]), row)

    merged_rows = []
    for item in group_items:
        base = _make_pressurized_param_row(
            name=item.get("name"),
            flow_section=item.get("flow_section"),
            structure_kind=item.get("structure_kind", structure_kind),
            pipe_material=item.get("pipe_material", default_material),
            dn_mm=item.get("DN_mm", 1500),
            display_name=item.get("display_name"),
        )
        # 先保留当前表格行自带的身份、长度、流量段补充信息。
        _copy_pressurized_row_metadata(base, item)
        cache_row = exact_cache.get(
            (base["name"], base.get("flow_section"), base["structure_kind"])
        ) or legacy_by_name.get((base["name"], base["structure_kind"]))
        if cache_row:
            # 再补回缓存里已有但当前行没有的结果字段，避免重新打开窗口后信息断链。
            _copy_pressurized_row_metadata(base, cache_row)
            base["pipe_material"] = cache_row["pipe_material"]
            base["DN_mm"] = _normalize_dn_mm(cache_row["DN_mm"], base["DN_mm"])
        merged_rows.append(base)
    return merged_rows
    cached_map = {}
    for row in cached_rows or []:
        if not isinstance(row, (tuple, list)) or len(row) < 3:
            continue
        name = str(row[0] or "").strip()
        if not name:
            continue
        mat = str(row[1] or "").strip() or default_material
        dn = _normalize_dn_mm(row[2], 1500)
        cached_map[name] = (mat, dn)


def _classify_pressure_pipe_dialog_bucket(row):
    """把有压同类结构归并成弹窗展示所需的类型桶。"""
    if not isinstance(row, dict):
        return "有压管道"
    structure_text = str(
        row.get("pressure_pipe_structure_type")
        or row.get("structure_type")
        or ""
    ).strip()
    if "顶管" in structure_text:
        return "顶管"
    if "定向钻" in structure_text:
        return "定向钻"
    return "有压管道"


def _resolve_pressure_pipe_dialog_special_base_label(row):
    """生成顶管/定向钻弹窗的基础展示名称。"""
    if not isinstance(row, dict):
        return "未命名"
    raw_name = str(row.get("name") or "").strip()
    if raw_name and raw_name not in {"-", "未命名有压管道"}:
        return raw_name
    display_name = str(row.get("display_name") or "").strip()
    if display_name:
        return display_name
    return "未命名"


def _resolve_pressure_pipe_dialog_special_label(row, duplicate_name_counts):
    """生成顶管/定向钻最终展示名；同名跨流量段时补上流量段。"""
    base_label = _resolve_pressure_pipe_dialog_special_base_label(row)
    if duplicate_name_counts.get(base_label, 0) <= 1:
        return base_label
    flow_section_idx = _parse_flow_section_index(row.get("flow_section"))
    flow_label = _segment_label_from_index(flow_section_idx)
    if flow_label:
        return f"{base_label}-{flow_label}"
    return base_label


def _pick_pressure_pipe_dialog_default_row(rows):
    """从同流量段普通有压管道里挑出弹窗默认材质与 DN。"""
    for row in rows or []:
        if _row_has_valid_pressurized_dn_mm(row):
            return row
    return (rows or [None])[0]


def _build_pressure_pipe_dialog_rows(group_items, cached_rows, default_material="球墨铸铁管"):
    """构造有压管道弹窗展示行。"""
    merged_rows = _merge_pressurized_param_defaults(
        group_items,
        cached_rows,
        default_material=default_material,
    )
    if not merged_rows:
        return []

    flow_section_order = []
    rows_by_flow_section = {}
    special_name_counts = {}
    for row in merged_rows:
        flow_section_idx = _parse_flow_section_index(row.get("flow_section"))
        if flow_section_idx not in rows_by_flow_section:
            rows_by_flow_section[flow_section_idx] = {
                "ordinary_rows": [],
                "special_rows": [],
            }
            flow_section_order.append(flow_section_idx)
        bucket = _classify_pressure_pipe_dialog_bucket(row)
        if bucket == "有压管道":
            rows_by_flow_section[flow_section_idx]["ordinary_rows"].append(row)
            continue
        rows_by_flow_section[flow_section_idx]["special_rows"].append(row)
        base_label = _resolve_pressure_pipe_dialog_special_base_label(row)
        special_name_counts[base_label] = special_name_counts.get(base_label, 0) + 1

    dialog_rows = []
    for flow_section_idx in flow_section_order:
        bucket_rows = rows_by_flow_section.get(flow_section_idx, {})
        ordinary_rows = bucket_rows.get("ordinary_rows", [])
        if ordinary_rows:
            default_row = _pick_pressure_pipe_dialog_default_row(ordinary_rows) or {}
            flow_label = _segment_label_from_index(flow_section_idx)
            dialog_row = _make_pressurized_param_row(
                name=flow_label or "有压管道",
                flow_section=flow_section_idx,
                structure_kind="pressure_pipe",
                pipe_material=default_row.get("pipe_material", default_material),
                dn_mm=default_row.get("DN_mm", 1500),
                display_name=flow_label or default_row.get("display_name") or "有压管道",
            )
            dialog_row["pressure_pipe_structure_type"] = "有压管道"
            dialog_row["dialog_row_kind"] = "flow_section_pressure_pipe"
            dialog_row["dialog_target_identities"] = [
                str(target.get("identity") or target.get("storage_key") or "").strip()
                for target in ordinary_rows
                if str(target.get("identity") or target.get("storage_key") or "").strip()
            ]
            dialog_row["dialog_target_rows"] = [copy.deepcopy(target) for target in ordinary_rows]
            dialog_rows.append(dialog_row)

        for special_row in bucket_rows.get("special_rows", []):
            display_name = _resolve_pressure_pipe_dialog_special_label(
                special_row,
                special_name_counts,
            )
            identity = str(
                special_row.get("identity") or special_row.get("storage_key") or ""
            ).strip()
            dialog_row = _make_pressurized_param_row(
                name=display_name,
                flow_section=special_row.get("flow_section"),
                structure_kind=special_row.get("structure_kind", "pressure_pipe"),
                pipe_material=special_row.get("pipe_material", default_material),
                dn_mm=special_row.get("DN_mm", 1500),
                display_name=display_name,
            )
            dialog_row["pressure_pipe_structure_type"] = str(
                special_row.get("pressure_pipe_structure_type") or ""
            ).strip()
            dialog_row["dialog_row_kind"] = "named_pressure_like_group"
            dialog_row["dialog_target_identities"] = [identity] if identity else []
            dialog_row["dialog_target_rows"] = [copy.deepcopy(special_row)]
            dialog_rows.append(dialog_row)
    return dialog_rows

    merged = []
    for name, dn_mm in group_items or []:
        base_dn = _normalize_dn_mm(dn_mm, 1500)
        mat, dn = cached_map.get(name, (default_material, base_dn))
        merged.append((name, mat, _normalize_dn_mm(dn, base_dn)))
    return merged


def _build_pressurized_segments(qs, overrides_by_idx, params, has_source_data, segment_name_fn):
    """基于分组参数构建倒虹吸/有压管道 segments。"""
    if not params:
        return []
    passthrough_keys = (
        "flow_section",
        "structure_kind",
        "ip_points",
        "upstream_velocity",
        "downstream_velocity",
        "inlet_transition_form",
        "outlet_transition_form",
        "inlet_transition_zeta",
        "outlet_transition_zeta",
        "V",
        "plan_total_length",
        "total_length",
        "total_head_loss",
        "friction_params",
        "start_water_level",
        "end_water_level",
        "tunnel_count",
        "tunnel_length",
        "directional_drill_count",
        "directional_drill_length",
        "jacking_count",
        "jacking_length",
        "show_building_characteristics",
    )

    structure_kind = None
    for row in params or []:
        if isinstance(row, dict):
            structure_kind = str(row.get("structure_kind") or "").strip() or None
            if structure_kind:
                break
    structure_kind = structure_kind or "siphon"

    normalized_params = _normalize_pressurized_cache_rows(params, structure_kind)
    if not normalized_params:
        return []

    overrides = overrides_by_idx or {}
    segs = []
    has_mapped_model = any(isinstance(row, dict) for row in params or [])

    mapped_params = []
    if has_source_data:
        for row in normalized_params:
            flow_section_idx = row.get("flow_section")
            if flow_section_idx is None:
                continue
            if overrides and flow_section_idx not in overrides:
                continue
            mapped_params.append(row)

    if has_source_data and has_mapped_model and not mapped_params:
        return []

    if has_source_data and mapped_params:
        params_by_segment = {}
        segment_order = []
        for row in mapped_params:
            flow_section_idx = row["flow_section"]
            if flow_section_idx not in params_by_segment:
                params_by_segment[flow_section_idx] = []
                segment_order.append(flow_section_idx)
            params_by_segment[flow_section_idx].append(row)

        for flow_section_idx in segment_order:
            base_override = {}
            if flow_section_idx in overrides and isinstance(overrides[flow_section_idx], dict):
                base_override = {
                    key: value
                    for key, value in overrides[flow_section_idx].items()
                    if key != "name"
                }

            candidates = []
            for row in params_by_segment[flow_section_idx]:
                seg = {
                    "name": row["display_name"],
                    "_struct_name": row.get("name"),
                    "flow_section": flow_section_idx,
                    "structure_kind": row.get("structure_kind"),
                }
                if base_override:
                    seg.update(base_override)
                if 0 < flow_section_idx <= len(qs):
                    seg["Q"] = qs[flow_section_idx - 1]
                seg["DN_mm"] = row["DN_mm"]
                seg["pipe_material"] = row["pipe_material"]
                for key in passthrough_keys:
                    if key in row:
                        seg[key] = copy.deepcopy(row[key])
                candidates.append(seg)

            # 有源建筑物数据时，始终保留“原名 + 流量段”的逐行展示，
            # 避免同流量段同参数的不同建筑物在弹窗与 DXF 中被合并后丢失名称。
            segs.extend(candidates)
        return segs

    overrides = overrides_by_idx or {}
    if has_source_data and overrides:
        indices = sorted(overrides.keys())
    else:
        indices = list(range(1, len(qs) + 1))
    if not indices:
        return []

    normalized_params = []
    for struct_name, pipe_material, dn_mm in params:
        normalized_params.append((
            str(struct_name or "").strip(),
            str(pipe_material or "").strip(),
            _normalize_dn_mm(dn_mm, 1500),
        ))

    segs = []
    for idx in indices:
        seg_label = segment_name_fn(idx)
        base_override = {}
        if idx in overrides and isinstance(overrides[idx], dict):
            base_override = {k: v for k, v in overrides[idx].items() if k != "name"}

        candidates = []
        for struct_name, pipe_material, dn_norm in normalized_params:
            seg = {
                "name": seg_label,
                "_struct_name": struct_name,
                "flow_section": idx,
                "structure_kind": structure_kind,
            }
            if base_override:
                seg.update(base_override)
            if 0 < idx <= len(qs):
                seg["Q"] = qs[idx - 1]
            seg["DN_mm"] = dn_norm
            seg["pipe_material"] = pipe_material
            candidates.append(seg)

        if len(candidates) <= 1:
            single = dict(candidates[0])
            single.pop("_struct_name", None)
            segs.append(single)
            continue

        grouped = {}
        for item in candidates:
            signature = (
                item.get("Q"),
                item.get("n"),
                item.get("DN_mm"),
                item.get("pipe_material"),
            )
            grouped.setdefault(signature, []).append(item)

        if len(grouped) == 1:
            merged = dict(candidates[0])
            merged["name"] = seg_label
            merged.pop("_struct_name", None)
            segs.append(merged)
            continue

        mergeable_group_count = sum(1 for items in grouped.values() if len(items) > 1)
        for items in grouped.values():
            merged = dict(items[0])
            anchor_name = str(merged.get("_struct_name") or "").strip()
            use_segment_name = mergeable_group_count == 1 and len(items) > 1
            if use_segment_name:
                merged["name"] = seg_label
            else:
                merged["name"] = f"{anchor_name}-{seg_label}" if anchor_name else seg_label
            merged.pop("_struct_name", None)
            segs.append(merged)
    return segs


# ================================================================
# 纵断面表格导出设置对话框
# ================================================================

class _LegacyTextExportSettingsDialog(QDialog):
    """旧版纵断面文字导出参数弹窗（保留作历史参考）。"""

    def __init__(self, parent=None, defaults=None):
        super().__init__(parent)
        self.setWindowTitle("纵断面文字导出设置")
        self.setMinimumWidth(420)
        self.setStyleSheet(DIALOG_STYLE)
        self.result = None

        if defaults is None:
            defaults = {}
        self._defaults = {
            'y_bottom': defaults.get('y_bottom', 1),
            'y_top': defaults.get('y_top', 31),
            'y_water': defaults.get('y_water', 16),
            'text_height': defaults.get('text_height', 3.5),
            'rotation': defaults.get('rotation', 90),
            'elev_decimals': defaults.get('elev_decimals', 3),
            'y_name': defaults.get('y_name', 115),
            'y_slope': defaults.get('y_slope', 105),
            'y_ip': defaults.get('y_ip', 77),
            'y_station': defaults.get('y_station', 47),
            'y_line_height': defaults.get('y_line_height', 120),
            'scale_x': defaults.get('scale_x', 2000),
            'scale_y': defaults.get('scale_y', 1000),
        }

        self._entries = {}
        self._init_ui()

    def _init_ui(self):
        lay = QVBoxLayout(self)

        # Y坐标设置
        y_grp = QGroupBox("Y 坐标设置（CAD 表格行高）")
        y_form = QGridLayout(y_grp)
        for row, (label, key) in enumerate([
            ("渠底文字 Y 坐标:", 'y_bottom'),
            ("渠顶文字 Y 坐标:", 'y_top'),
            ("水面文字 Y 坐标:", 'y_water'),
        ]):
            y_form.addWidget(QLabel(label), row, 0)
            e = LineEdit(); e.setText(str(self._defaults[key])); e.setFixedWidth(100)
            y_form.addWidget(e, row, 1)
            self._entries[key] = e
        lay.addWidget(y_grp)

        # 文字样式
        style_grp = QGroupBox("文字样式")
        style_form = QGridLayout(style_grp)
        for row, (label, key) in enumerate([
            ("字高:", 'text_height'),
            ("旋转角度:", 'rotation'),
            ("高程小数位数:", 'elev_decimals'),
        ]):
            style_form.addWidget(QLabel(label), row, 0)
            e = LineEdit(); e.setText(str(self._defaults[key])); e.setFixedWidth(100)
            style_form.addWidget(e, row, 1)
            self._entries[key] = e
        lay.addWidget(style_grp)

        # 纵断面信息列
        info_grp = QGroupBox("纵断面信息列 Y 坐标")
        info_form = QGridLayout(info_grp)
        for row, (label, key) in enumerate([
            ("建筑物名称 Y 坐标:", 'y_name'),
            ("坡降 Y 坐标:", 'y_slope'),
            ("IP点名称 Y 坐标:", 'y_ip'),
            ("里程桩号 Y 坐标:", 'y_station'),
            ("整线竖线高度:", 'y_line_height'),
        ]):
            info_form.addWidget(QLabel(label), row, 0)
            e = LineEdit(); e.setText(str(self._defaults[key])); e.setFixedWidth(100)
            info_form.addWidget(e, row, 1)
            self._entries[key] = e
        lay.addWidget(info_grp)

        # 比例设置
        scale_grp = QGroupBox("比例设置")
        scale_form = QGridLayout(scale_grp)
        scale_form.addWidget(QLabel("X 方向 (1:N)，N ="), 0, 0)
        e = LineEdit(); e.setText(str(self._defaults['scale_x'])); e.setFixedWidth(100)
        scale_form.addWidget(e, 0, 1)
        scale_form.addWidget(QLabel("如 1:1000 则输入 1000"), 0, 2)
        self._entries['scale_x'] = e
        scale_form.addWidget(QLabel("Y 方向 (1:N)，N ="), 1, 0)
        e = LineEdit(); e.setText(str(self._defaults['scale_y'])); e.setFixedWidth(100)
        scale_form.addWidget(e, 1, 1)
        scale_form.addWidget(QLabel("如 1:1000 则输入 1000"), 1, 2)
        self._entries['scale_y'] = e
        lay.addWidget(scale_grp)

        # 预览
        preview_grp = QGroupBox("命令格式预览")
        preview_lay = QVBoxLayout(preview_grp)
        self._preview_label = QLabel()
        self._preview_label.setStyleSheet("color: #336699;")
        self._preview_label.setFont(QFont("Consolas", 9))
        preview_lay.addWidget(self._preview_label)
        lay.addWidget(preview_grp)
        self._update_preview()
        for entry in self._entries.values():
            entry.textChanged.connect(self._update_preview)

        # 按钮
        btn_lay = QHBoxLayout()
        btn_reset = PushButton("恢复默认"); btn_reset.clicked.connect(self._reset_defaults)
        btn_lay.addWidget(btn_reset); btn_lay.addStretch()
        btn_cancel = PushButton("取消"); btn_cancel.clicked.connect(self.reject)
        btn_ok = PrimaryPushButton("确定"); btn_ok.clicked.connect(self._on_confirm)
        btn_lay.addWidget(btn_cancel); btn_lay.addWidget(btn_ok)
        lay.addLayout(btn_lay)

        # 键盘快捷键
        QShortcut(QKeySequence(Qt.Key_Escape), self, self.reject)
        QShortcut(QKeySequence(Qt.Key_Return), self, self._on_confirm)

    def _update_preview(self):
        try:
            y_b = self._entries['y_bottom'].text().strip()
            h = self._entries['text_height'].text().strip()
            r = self._entries['rotation'].text().strip()
            d = self._entries['elev_decimals'].text().strip()
            try:
                decimals = int(float(d))
                sample = f"{431.666:.{decimals}f}"
            except Exception:
                sample = "431.666"
            self._preview_label.setText(f"-text 里程MC,{y_b} {h} {r} {sample} ")
        except Exception:
            self._preview_label.setText("-text 里程MC,Y 字高 角度 高程 ")

    def _reset_defaults(self):
        original = {
            'y_bottom': 1, 'y_top': 31, 'y_water': 16,
            'text_height': 3.5, 'rotation': 90, 'elev_decimals': 3,
            'y_name': 115, 'y_slope': 105, 'y_ip': 77,
            'y_station': 47, 'y_line_height': 120,
            'scale_x': 2000, 'scale_y': 1000,
        }
        for key, value in original.items():
            self._entries[key].setText(str(value))
        self._update_preview()

    def _on_confirm(self):
        try:
            result = {}
            for key, entry in self._entries.items():
                val_str = entry.text().strip()
                if not val_str:
                    raise ValueError("参数不能为空")
                val = float(val_str)
                if key == 'elev_decimals':
                    if val < 0 or val != int(val):
                        raise ValueError("高程小数位数必须为非负整数")
                    val = int(val)
                if key in ('scale_x', 'scale_y'):
                    if val <= 0:
                        raise ValueError("比例尺必须大于0")
                result[key] = val
            self.result = result
            self.accept()
        except ValueError as e:
            fluent_error(self, "输入错误", f"请输入有效的数值:\n{str(e)}")


# ===================== 重构版：Win11 Fluent + 双列直拖 =====================

class _ProfileRowDragListWidget(ListWidget):
    """支持跨列拖放的纵断面行列表。"""

    rowsDropped = Signal(str, list, int, str)  # source_role, ids, row, target_role

    def __init__(self, role: str, parent=None):
        super().__init__(parent)
        self._role = role
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.setDragEnabled(True)
        self.viewport().setAcceptDrops(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setDefaultDropAction(Qt.MoveAction)
        self.setDragDropMode(QAbstractItemView.DragDrop)

    def _selected_ids(self):
        rows = sorted({self.row(item) for item in self.selectedItems()})
        out = []
        for row in rows:
            item = self.item(row)
            if item is None:
                continue
            rid = str(item.data(Qt.UserRole) or "").strip()
            if rid:
                out.append(rid)
        return out

    def startDrag(self, supportedActions):
        row_ids = self._selected_ids()
        if not row_ids:
            return
        payload = {"source": self._role, "ids": row_ids}
        mime = QMimeData()
        mime.setData(
            "application/x-profile-row-ids",
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        )
        drag = QDrag(self)
        drag.setMimeData(mime)
        drag.exec(Qt.MoveAction)

    def dragEnterEvent(self, event):
        if event.mimeData().hasFormat("application/x-profile-row-ids"):
            event.acceptProposedAction()
            return
        super().dragEnterEvent(event)

    def dragMoveEvent(self, event):
        if event.mimeData().hasFormat("application/x-profile-row-ids"):
            event.acceptProposedAction()
            return
        super().dragMoveEvent(event)

    def dropEvent(self, event):
        data = event.mimeData()
        if not data.hasFormat("application/x-profile-row-ids"):
            super().dropEvent(event)
            return
        try:
            payload = json.loads(bytes(data.data("application/x-profile-row-ids")).decode("utf-8"))
            source_role = str(payload.get("source", "")).strip()
            row_ids = [str(rid).strip() for rid in payload.get("ids", []) if str(rid).strip()]
            if not source_role or not row_ids:
                event.ignore()
                return
            pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
            row = self.indexAt(pos).row()
            if row < 0:
                row = self.count()
            self.rowsDropped.emit(source_role, row_ids, row, self._role)
            event.acceptProposedAction()
        except Exception:
            event.ignore()


class _LegacyTextExportSettingsDialogDualList(QDialog):
    """纵断面导出参数与行配置弹窗（Win11 Fluent 风格重构版）。"""

    _UI_SETTINGS_ORG = "SichuanShuifa"
    _UI_SETTINGS_APP = "HydroCalc"
    _UI_SIZE_W_KEY = "water_profile/text_export_dialog_width"
    _UI_SIZE_H_KEY = "water_profile/text_export_dialog_height"
    _UI_PREVIEW_EXPANDED_KEY = "water_profile/text_export_dialog_preview_expanded"
    _ICON_COLLAPSED = _resolve_fluent_icon("CHEVRON_RIGHT_MED", "CHEVRON_RIGHT", "CHEVRON_DOWN_MED")
    _ICON_EXPANDED = _resolve_fluent_icon("CHEVRON_DOWN_MED", "CHEVRON_RIGHT_MED", "CHEVRON_RIGHT")

    def __init__(self, parent=None, defaults=None):
        super().__init__(parent)
        self.setWindowTitle("纵断面文字导出设置")
        self.setMinimumSize(960, 500)
        self._ui_settings = QSettings(self._UI_SETTINGS_ORG, self._UI_SETTINGS_APP)
        self._preview_expanded = self._read_setting_bool(self._UI_PREVIEW_EXPANDED_KEY, True)
        self._apply_initial_size()
        self.setSizeGripEnabled(True)
        self.setStyleSheet(DIALOG_STYLE + """
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f7f9fc, stop:1 #eef3fb);
            }
            QListView {
                border: 1px solid #d6dfef;
                border-radius: 10px;
                background: rgba(255,255,255,0.92);
                padding: 4px;
            }
            QListView::item {
                border-radius: 8px;
                padding: 5px 10px;
                margin: 1px 1px;
            }
            QListView::item:selected {
                background: rgba(0, 120, 212, 0.16);
                border: 1px solid rgba(0, 120, 212, 0.35);
            }
            QListView::item:hover {
                background: rgba(32, 97, 181, 0.08);
            }
        """)
        self.result = None
        self._row_updating = False
        self._segment_key = "all"

        defaults = _normalize_text_export_settings(defaults or {})
        self._defaults = dict(defaults)

        self._entries = {}
        self._ordered_row_ids = list(_PROFILE_ROW_DEFAULT_ORDER)
        self._enabled_row_ids = []

        self._candidate_search = None
        self._candidate_segment = None
        self._candidate_list = None
        self._enabled_list = None
        self._advanced_body = None
        self._advanced_toggle_btn = None
        self._preview_label = None
        self._preview_body = None
        self._preview_toggle_btn = None

        self._init_ui()

    def _read_setting_bool(self, key, default=False):
        raw = self._ui_settings.value(key, default)
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, (int, float)):
            return bool(raw)
        text = str(raw).strip().lower()
        if text in {"1", "true", "yes", "on"}:
            return True
        if text in {"0", "false", "no", "off"}:
            return False
        return bool(default)

    def _read_setting_int(self, key, default_value):
        raw = self._ui_settings.value(key, default_value)
        try:
            return int(float(raw))
        except Exception:
            return int(default_value)

    def _available_geometry(self):
        screen = None
        parent_widget = self.parentWidget()
        if parent_widget is not None:
            parent_window = parent_widget.window()
            if parent_window is not None and parent_window.windowHandle() is not None:
                screen = parent_window.windowHandle().screen()
        if screen is None:
            app = QApplication.instance()
            if app is not None:
                screen = app.primaryScreen()
        return screen.availableGeometry() if screen is not None else None

    def _apply_initial_size(self):
        avail = self._available_geometry()
        if avail is not None:
            default_w = min(max(self.minimumWidth(), int(avail.width() * 0.78)), 1360)
            default_h = min(max(self.minimumHeight(), int(avail.height() * 0.72)), int(avail.height() * 0.92))
            max_w = max(self.minimumWidth(), int(avail.width() * 0.96))
            max_h = max(self.minimumHeight(), int(avail.height() * 0.92))
        else:
            default_w, default_h = 1160, 640
            max_w, max_h = 1400, 900

        width = self._read_setting_int(self._UI_SIZE_W_KEY, default_w)
        height = self._read_setting_int(self._UI_SIZE_H_KEY, default_h)
        width = max(self.minimumWidth(), min(width, max_w))
        height = max(self.minimumHeight(), min(height, max_h))
        self.resize(width, height)

    def _persist_ui_state(self):
        size = self.size()
        self._ui_settings.setValue(self._UI_SIZE_W_KEY, int(size.width()))
        self._ui_settings.setValue(self._UI_SIZE_H_KEY, int(size.height()))
        self._ui_settings.setValue(self._UI_PREVIEW_EXPANDED_KEY, bool(self._preview_expanded))

    def closeEvent(self, event):
        self._persist_ui_state()
        super().closeEvent(event)

    def _init_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(14, 8, 14, 8)
        root.setSpacing(6)

        body_row = QHBoxLayout()
        body_row.setSpacing(8)

        left_col = QVBoxLayout()
        left_col.setSpacing(8)
        left_col.addWidget(self._build_basic_card())
        left_col.addWidget(self._build_advanced_card())
        left_col.addStretch(0)

        right_col = QVBoxLayout()
        right_col.setSpacing(8)
        right_col.addWidget(self._build_rows_card(), 0)
        right_col.addWidget(self._build_preview_card(), 0)
        right_col.addStretch(1)

        body_row.addLayout(left_col, 38)
        body_row.addLayout(right_col, 62)
        body_row.setAlignment(left_col, Qt.AlignTop)
        body_row.setAlignment(right_col, Qt.AlignTop)
        root.addLayout(body_row, 1)

        btn_row = QHBoxLayout()
        btn_reset = PushButton("恢复默认")
        btn_reset.clicked.connect(self._reset_defaults)
        btn_row.addWidget(btn_reset)
        btn_row.addStretch(1)
        btn_cancel = PushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        btn_ok = PrimaryPushButton("确定")
        btn_ok.clicked.connect(self._on_confirm)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        root.addLayout(btn_row)

        self._load_rows(self._defaults.get("profile_row_items"))
        for entry in self._entries.values():
            entry.textChanged.connect(self._update_preview)
        self._update_preview()

        QShortcut(QKeySequence(Qt.Key_Escape), self, self.reject)
        QShortcut(QKeySequence(Qt.Key_Return), self, self._on_confirm)
        QShortcut(QKeySequence("Ctrl+Up"), self, lambda: self._move_selected_row(-1))
        QShortcut(QKeySequence("Ctrl+Down"), self, lambda: self._move_selected_row(1))
        QShortcut(QKeySequence(Qt.Key_Delete), self, self._remove_selected_rows)
        QShortcut(QKeySequence("Ctrl+Right"), self, self._enable_selected_rows)
        QShortcut(QKeySequence("Ctrl+Left"), self, self._remove_selected_rows)

    def _build_basic_card(self):
        card = ElevatedCardWidget(self)
        card_lay = QVBoxLayout(card)
        card_lay.setContentsMargins(12, 10, 12, 10)
        card_lay.setSpacing(8)

        card_lay.addWidget(BodyLabel("基础参数"))
        form = QGridLayout()
        form.setHorizontalSpacing(8)
        form.setVerticalSpacing(8)
        form.setColumnStretch(0, 0)
        form.setColumnStretch(1, 0)
        form.setColumnStretch(2, 1)
        self._add_entry_row(form, 0, "字高", "text_height", "")
        self._add_entry_row(form, 1, "旋转角度", "rotation", "")
        self._add_entry_row(form, 2, "高程小数位数", "elev_decimals", "")
        self._add_entry_row(form, 3, "X方向比例(1:N)", "scale_x", "如 1:1000 则输入 1000")
        self._add_entry_row(form, 4, "Y方向比例(1:N)", "scale_y", "如 1:1000 则输入 1000")
        card_lay.addLayout(form)
        return card

    def _build_advanced_card(self):
        card = ElevatedCardWidget(self)
        card_lay = QVBoxLayout(card)
        card_lay.setContentsMargins(12, 10, 12, 10)
        card_lay.setSpacing(8)

        row = QHBoxLayout()
        row.setSpacing(6)
        row.addWidget(BodyLabel("高级参数（旧版Y坐标）"))
        self._advanced_toggle_btn = ToolButton(self._ICON_COLLAPSED)
        self._advanced_toggle_btn.clicked.connect(self._toggle_advanced)
        row.addStretch(1)
        row.addWidget(self._advanced_toggle_btn)
        card_lay.addLayout(row)

        self._advanced_body = QWidget()
        adv_form = QGridLayout(self._advanced_body)
        adv_form.setHorizontalSpacing(8)
        adv_form.setVerticalSpacing(6)
        adv_form.setColumnStretch(2, 1)
        self._add_entry_row(adv_form, 0, "渠底文字Y", "y_bottom", "")
        self._add_entry_row(adv_form, 1, "渠顶文字Y", "y_top", "")
        self._add_entry_row(adv_form, 2, "水面文字Y", "y_water", "")
        self._add_entry_row(adv_form, 3, "建筑物名称Y", "y_name", "兼容旧项目")
        self._add_entry_row(adv_form, 4, "坡降Y", "y_slope", "兼容旧项目")
        self._add_entry_row(adv_form, 5, "IP点名称Y", "y_ip", "兼容旧项目")
        self._add_entry_row(adv_form, 6, "里程桩号Y", "y_station", "兼容旧项目")
        self._add_entry_row(adv_form, 7, "最小竖线高度", "y_line_height", "最小值 > 0")
        self._advanced_body.setVisible(False)
        card_lay.addWidget(self._advanced_body)
        return card

    def _build_rows_card(self):
        card = ElevatedCardWidget(self)
        card_lay = QVBoxLayout(card)
        card_lay.setContentsMargins(12, 10, 12, 10)
        card_lay.setSpacing(6)

        title_row = QHBoxLayout()
        title_row.addWidget(BodyLabel("纵断面行内容（13项可选，可排序）"))
        title_row.addStretch(1)
        btn_preset = PushButton("应用亭子口二期项建/可研阶段模板")
        btn_preset.clicked.connect(self._apply_tingzikou_preset)
        title_row.addWidget(btn_preset)
        card_lay.addLayout(title_row)

        quick_row = QHBoxLayout()
        btn_enable_all = PushButton("全启用")
        btn_enable_all.clicked.connect(self._enable_all_rows)
        btn_disable_all = PushButton("全停用")
        btn_disable_all.clicked.connect(self._disable_all_rows)
        btn_restore_recommended = PushButton("恢复推荐")
        btn_restore_recommended.clicked.connect(self._restore_recommended_rows)
        quick_row.addWidget(btn_enable_all)
        quick_row.addWidget(btn_disable_all)
        quick_row.addWidget(btn_restore_recommended)
        quick_row.addStretch(1)
        card_lay.addLayout(quick_row)

        list_row = QHBoxLayout()
        list_row.setSpacing(8)

        candidate_col = QVBoxLayout()
        candidate_col.setSpacing(6)
        candidate_col.addWidget(BodyLabel("可选项"))
        self._candidate_segment = SegmentedWidget(self)
        self._candidate_segment.addItem("all", "全部", onClick=lambda: self._set_candidate_segment("all"))
        self._candidate_segment.addItem("recommended", "推荐", onClick=lambda: self._set_candidate_segment("recommended"))
        self._candidate_segment.addItem("extended", "扩展", onClick=lambda: self._set_candidate_segment("extended"))
        self._candidate_segment.setCurrentItem("all")
        candidate_col.addWidget(self._candidate_segment)
        self._candidate_search = SearchLineEdit()
        self._candidate_search.setPlaceholderText("搜索行内容（中文包含匹配）")
        self._candidate_search.textChanged.connect(self._refresh_row_lists)
        candidate_col.addWidget(self._candidate_search)
        self._candidate_list = _ProfileRowDragListWidget("candidate", self)
        self._candidate_list.rowsDropped.connect(self._on_rows_dropped)
        self._candidate_list.itemDoubleClicked.connect(lambda _item: self._enable_selected_rows())
        candidate_col.addWidget(self._candidate_list, 1)

        action_col = QVBoxLayout()
        action_col.setSpacing(6)
        action_col.addStretch(1)
        btn_add = PushButton("添加 ->")
        btn_add.clicked.connect(self._enable_selected_rows)
        btn_remove = PushButton("<- 移除")
        btn_remove.clicked.connect(self._remove_selected_rows)
        action_col.addWidget(btn_add)
        action_col.addWidget(btn_remove)
        action_col.addStretch(1)

        enabled_col = QVBoxLayout()
        enabled_col.setSpacing(6)
        enabled_col.addWidget(BodyLabel("已启用项（支持拖拽排序）"))
        self._enabled_list = _ProfileRowDragListWidget("enabled", self)
        self._enabled_list.rowsDropped.connect(self._on_rows_dropped)
        self._enabled_list.itemDoubleClicked.connect(lambda _item: self._remove_selected_rows())
        enabled_col.addWidget(self._enabled_list, 1)
        sort_row = QHBoxLayout()
        btn_up = PushButton("上移")
        btn_up.clicked.connect(lambda: self._move_selected_row(-1))
        btn_down = PushButton("下移")
        btn_down.clicked.connect(lambda: self._move_selected_row(1))
        sort_row.addWidget(btn_up)
        sort_row.addWidget(btn_down)
        sort_row.addStretch(1)
        enabled_col.addLayout(sort_row)

        list_row.addLayout(candidate_col, 46)
        list_row.addLayout(action_col, 12)
        list_row.addLayout(enabled_col, 42)
        card_lay.addLayout(list_row, 0)
        return card

    def _build_preview_card(self):
        card = ElevatedCardWidget(self)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(12, 10, 12, 10)
        lay.setSpacing(6)
        row = QHBoxLayout()
        row.addWidget(BodyLabel("当前配置预览"))
        row.addStretch(1)
        self._preview_toggle_btn = ToolButton(self._ICON_EXPANDED)
        self._preview_toggle_btn.clicked.connect(self._toggle_preview)
        row.addWidget(self._preview_toggle_btn)
        lay.addLayout(row)

        self._preview_body = QWidget()
        preview_lay = QVBoxLayout(self._preview_body)
        preview_lay.setContentsMargins(0, 0, 0, 0)
        preview_lay.setSpacing(0)

        self._preview_label = QLabel()
        self._preview_label.setWordWrap(True)
        self._preview_label.setStyleSheet("color:#245A9B; font-family:'Consolas','Microsoft YaHei';")
        preview_lay.addWidget(self._preview_label)
        lay.addWidget(self._preview_body)
        self._set_preview_expanded(self._preview_expanded)
        return card

    def _set_preview_expanded(self, expanded):
        self._preview_expanded = bool(expanded)
        if self._preview_body is not None:
            self._preview_body.setVisible(self._preview_expanded)
        if self._preview_toggle_btn is not None:
            self._preview_toggle_btn.setIcon(self._ICON_EXPANDED if self._preview_expanded else self._ICON_COLLAPSED)

    def _toggle_preview(self):
        self._set_preview_expanded(not self._preview_expanded)

    def _add_entry_row(self, layout, row, label, key, hint):
        layout.addWidget(QLabel(f"{label}:"), row, 0)
        entry = LineEdit()
        entry.setText(str(self._defaults.get(key, "")))
        entry.setFixedWidth(130)
        layout.addWidget(entry, row, 1)
        layout.addWidget(CaptionLabel(hint), row, 2)
        self._entries[key] = entry

    def _toggle_advanced(self):
        visible = not self._advanced_body.isVisible()
        self._advanced_body.setVisible(visible)
        self._advanced_toggle_btn.setIcon(self._ICON_EXPANDED if visible else self._ICON_COLLAPSED)

    def _set_candidate_segment(self, key):
        self._segment_key = key
        self._refresh_row_lists()

    def _selected_ids(self, list_widget):
        rows = sorted({list_widget.row(item) for item in list_widget.selectedItems()})
        ids = []
        for row in rows:
            item = list_widget.item(row)
            if item is None:
                continue
            rid = str(item.data(Qt.UserRole) or "").strip()
            if rid in _PROFILE_ROW_DEF_MAP:
                ids.append(rid)
        return ids

    def _create_row_item(self, rid):
        from PySide6.QtWidgets import QListWidgetItem
        from PySide6.QtCore import QSize
        from PySide6.QtGui import QColor

        row_def = _PROFILE_ROW_DEF_MAP[rid]
        badge = "  ★推荐" if rid in _PROFILE_RECOMMENDED_ROW_IDS else ""
        title = f"{row_def['label']}{badge}"
        hint = row_def.get("hint", "")
        list_item = QListWidgetItem(f"{title}\n{hint}")
        list_item.setData(Qt.UserRole, rid)
        if rid in _PROFILE_RECOMMENDED_ROW_IDS:
            list_item.setForeground(QColor("#174EA6"))
        list_item.setSizeHint(QSize(0, 44))
        return list_item

    def _normalize_row_model(self):
        enabled = [rid for rid in self._enabled_row_ids if rid in _PROFILE_ROW_DEF_MAP]
        order = [rid for rid in self._ordered_row_ids if rid in _PROFILE_ROW_DEF_MAP]
        for rid in _PROFILE_ROW_DEFAULT_ORDER:
            if rid not in order:
                order.append(rid)
        disabled = [rid for rid in order if rid not in enabled]
        self._enabled_row_ids = enabled
        self._ordered_row_ids = enabled + disabled

    def _candidate_visible(self, rid):
        if rid in self._enabled_row_ids:
            return False
        if self._segment_key == "recommended" and rid not in _PROFILE_RECOMMENDED_ROW_IDS:
            return False
        if self._segment_key == "extended" and rid not in _PROFILE_EXTENDED_ROW_IDS:
            return False
        q = (self._candidate_search.text() if self._candidate_search else "").strip()
        if q and q not in _PROFILE_ROW_DEF_MAP[rid]["label"]:
            return False
        return True

    def _refresh_row_lists(self, *_args):
        if not self._candidate_list or not self._enabled_list:
            return
        self._normalize_row_model()
        keep_candidate = set(self._selected_ids(self._candidate_list))
        keep_enabled = set(self._selected_ids(self._enabled_list))

        self._row_updating = True
        try:
            self._candidate_list.clear()
            for rid in self._ordered_row_ids:
                if not self._candidate_visible(rid):
                    continue
                item = self._create_row_item(rid)
                self._candidate_list.addItem(item)
                if rid in keep_candidate:
                    item.setSelected(True)

            self._enabled_list.clear()
            for rid in self._enabled_row_ids:
                item = self._create_row_item(rid)
                self._enabled_list.addItem(item)
                if rid in keep_enabled:
                    item.setSelected(True)
        finally:
            self._row_updating = False
        self._ensure_row_lists_visible_rows()
        self._update_preview()

    def _ensure_row_lists_visible_rows(self):
        if not self._enabled_list or not self._candidate_list:
            return
        row_h = max(self._enabled_list.sizeHintForRow(0), self._candidate_list.sizeHintForRow(0))
        if row_h <= 0:
            row_h = 44
        visible_rows = 8
        target_h = row_h * visible_rows + 12
        for list_widget in (self._candidate_list, self._enabled_list):
            list_widget.setMinimumHeight(target_h)
            list_widget.setMaximumHeight(target_h)

    def _load_rows(self, row_items):
        normalized = _normalize_profile_row_items(row_items)
        self._ordered_row_ids = [item["id"] for item in normalized]
        self._enabled_row_ids = [item["id"] for item in normalized if item.get("enabled")]
        self._refresh_row_lists()

    def _row_data_from_table(self):
        self._normalize_row_model()
        enabled = set(self._enabled_row_ids)
        return _normalize_profile_row_items([
            {"id": rid, "enabled": rid in enabled}
            for rid in self._ordered_row_ids
        ])

    def _apply_drop(self, source_role, row_ids, row, target_role):
        ids = [rid for rid in row_ids if rid in _PROFILE_ROW_DEF_MAP]
        if not ids:
            return

        enabled = list(self._enabled_row_ids)
        if source_role == "enabled" and target_role == "enabled":
            old_pos = [enabled.index(rid) for rid in ids if rid in enabled]
            remaining = [rid for rid in enabled if rid not in ids]
            row_adj = int(row) - sum(1 for p in old_pos if p < int(row))
            row_adj = max(0, min(len(remaining), row_adj))
            enabled = remaining[:row_adj] + ids + remaining[row_adj:]
        elif source_role == "candidate" and target_role == "enabled":
            insert_pos = max(0, min(len(enabled), int(row)))
            existing = [rid for rid in enabled if rid not in ids]
            enabled = existing[:insert_pos] + ids + existing[insert_pos:]
            InfoBar.success(
                "已启用",
                f"已添加 {len(ids)} 项",
                parent=self,
                position=InfoBarPosition.TOP_RIGHT,
                duration=1200,
            )
        elif source_role == "enabled" and target_role == "candidate":
            enabled = [rid for rid in enabled if rid not in ids]
            InfoBar.info(
                "已停用",
                f"已移除 {len(ids)} 项",
                parent=self,
                position=InfoBarPosition.TOP_RIGHT,
                duration=1200,
            )
        else:
            return

        self._enabled_row_ids = enabled
        self._normalize_row_model()
        self._refresh_row_lists()

    def _on_rows_dropped(self, source_role, row_ids, row, target_role):
        if self._row_updating:
            return
        self._apply_drop(source_role, row_ids, row, target_role)

    def _enable_selected_rows(self):
        ids = self._selected_ids(self._candidate_list)
        if not ids:
            return
        enabled = [rid for rid in self._enabled_row_ids if rid not in ids]
        insert_pos = self._enabled_list.currentRow()
        if insert_pos < 0:
            insert_pos = len(enabled)
        enabled[insert_pos:insert_pos] = ids
        self._enabled_row_ids = enabled
        self._refresh_row_lists()

    def _remove_selected_rows(self):
        ids = set(self._selected_ids(self._enabled_list))
        if not ids:
            return
        self._enabled_row_ids = [rid for rid in self._enabled_row_ids if rid not in ids]
        self._refresh_row_lists()

    def _enable_all_rows(self):
        self._enabled_row_ids = list(_PROFILE_ROW_DEFAULT_ORDER)
        self._refresh_row_lists()

    def _disable_all_rows(self):
        self._enabled_row_ids = []
        self._refresh_row_lists()

    def _restore_recommended_rows(self):
        self._enabled_row_ids = [
            rid for rid in _PROFILE_ROW_DEFAULT_ORDER
            if rid in _PROFILE_RECOMMENDED_ROW_IDS
        ]
        self._refresh_row_lists()

    def _apply_tingzikou_preset(self):
        ordered = list(_TINGZIKOU_TEMPLATE_ROW_IDS) + [
            rid for rid in _PROFILE_ROW_DEFAULT_ORDER if rid not in _TINGZIKOU_TEMPLATE_ROW_IDS
        ]
        self._ordered_row_ids = ordered
        self._enabled_row_ids = list(_TINGZIKOU_TEMPLATE_ROW_IDS)
        self._refresh_row_lists()
        InfoBar.success(
            "模板已应用",
            "已切换为亭子口推荐顺序",
            parent=self,
            position=InfoBarPosition.TOP_RIGHT,
            duration=1500,
        )

    def _move_selected_row(self, delta):
        row = self._enabled_list.currentRow()
        if row < 0:
            return
        target = row + int(delta)
        if target < 0 or target >= len(self._enabled_row_ids):
            return
        rid = self._enabled_row_ids.pop(row)
        self._enabled_row_ids.insert(target, rid)
        self._refresh_row_lists()
        self._enabled_list.setCurrentRow(target)

    def _update_preview(self):
        try:
            enabled = [item for item in self._row_data_from_table() if item.get("enabled")]
            labels = [_PROFILE_ROW_DEF_MAP[item["id"]]["label"] for item in enabled[:6]]
            summary = "、".join(labels) if labels else "无"
            if len(enabled) > 6:
                summary += f" ...（共{len(enabled)}行）"
            self._preview_label.setText(
                f"已启用行：{summary}\n"
                f"示例：-text X,Y {self._entries['text_height'].text().strip()} "
                f"{self._entries['rotation'].text().strip()} 文本"
            )
        except Exception:
            self._preview_label.setText("预览不可用")

    def _reset_defaults(self):
        original = {
            "y_bottom": 1, "y_top": 31, "y_water": 16,
            "text_height": 3.5, "rotation": 90, "elev_decimals": 3,
            "y_name": 115, "y_slope": 105, "y_ip": 77,
            "y_station": 47, "y_line_height": 120,
            "scale_x": 2000, "scale_y": 1000,
        }
        for key, value in original.items():
            if key in self._entries:
                self._entries[key].setText(str(value))
        self._load_rows(_default_profile_row_items())
        self._update_preview()

    def _focus_invalid_entry(self, key):
        entry = self._entries.get(key)
        if not entry:
            return
        if key in {"y_bottom", "y_top", "y_water", "y_name", "y_slope", "y_ip", "y_station", "y_line_height"}:
            if self._advanced_body and not self._advanced_body.isVisible():
                self._toggle_advanced()
        entry.setFocus()
        entry.selectAll()

    def _on_confirm(self):
        try:
            parsed = {}
            ordered_keys = [
                "text_height", "rotation", "elev_decimals", "scale_x", "scale_y",
                "y_bottom", "y_top", "y_water",
                "y_name", "y_slope", "y_ip", "y_station", "y_line_height",
            ]
            labels = {
                "text_height": "字高",
                "rotation": "旋转角度",
                "elev_decimals": "高程小数位数",
                "scale_x": "X方向比例",
                "scale_y": "Y方向比例",
                "y_bottom": "渠底文字Y",
                "y_top": "渠顶文字Y",
                "y_water": "水面文字Y",
                "y_name": "建筑物名称Y",
                "y_slope": "坡降Y",
                "y_ip": "IP点名称Y",
                "y_station": "里程桩号Y",
                "y_line_height": "最小竖线高度",
            }
            for key in ordered_keys:
                entry = self._entries[key]
                txt = entry.text().strip()
                if not txt:
                    self._focus_invalid_entry(key)
                    raise ValueError(f"{labels[key]}不能为空")
                try:
                    val = float(txt)
                except ValueError:
                    self._focus_invalid_entry(key)
                    raise ValueError(f"{labels[key]}必须为数值")
                if key == "elev_decimals":
                    if val < 0 or val != int(val):
                        self._focus_invalid_entry(key)
                        raise ValueError("高程小数位数必须为非负整数")
                    val = int(val)
                if key in ("scale_x", "scale_y", "y_line_height") and val <= 0:
                    self._focus_invalid_entry(key)
                    raise ValueError("比例与最小竖线高度必须大于0")
                parsed[key] = val

            row_items = self._row_data_from_table()
            if not any(item.get("enabled") for item in row_items):
                self._enabled_list.setFocus()
                raise ValueError("至少选择1项行内容")

            result = dict(self._defaults)
            result.update(parsed)
            result["profile_row_items"] = row_items
            self.result = _normalize_text_export_settings(result)
            self.accept()
        except ValueError as e:
            fluent_error(self, "输入错误", f"请输入有效的数值:\n{str(e)}")
# ================================================================
# 有压流参数配置对话框
# ================================================================

class _FluentProfileDragHandle(QLabel):
    """Drag handle that only starts sorting when the user drags the grip."""

    dragRequested = Signal()

    def __init__(self, parent=None):
        super().__init__(":::")
        self._press_pos = None
        self.setAlignment(Qt.AlignCenter)
        self.setCursor(Qt.OpenHandCursor)
        self.setFixedWidth(20)
        self.setStyleSheet("color:#7F8B99; font-size:16px; font-weight:600; letter-spacing:1px;")

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._press_pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
            self.setCursor(Qt.ClosedHandCursor)
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if not (event.buttons() & Qt.LeftButton) or self._press_pos is None:
            super().mouseMoveEvent(event)
            return
        pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
        if (pos - self._press_pos).manhattanLength() >= QApplication.startDragDistance():
            self.dragRequested.emit()
            self._press_pos = None
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._press_pos = None
        self.setCursor(Qt.OpenHandCursor)
        super().mouseReleaseEvent(event)


class _FluentProfileRowItemWidget(QWidget):
    """Two-line Fluent row with checkbox, subtle recommendation badge and drag handle."""

    clicked = Signal()
    doubleClicked = Signal()
    dragRequested = Signal()

    def __init__(self, title, subtitle, enabled, recommended=False, parent=None):
        super().__init__(parent)
        self._selected = False
        self._enabled = bool(enabled)
        self._recommended = bool(recommended)

        self.setObjectName("profileRowItemFluent")
        self.setAttribute(Qt.WA_StyledBackground, True)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 5, 10, 5)
        layout.setSpacing(8)

        self.checkbox = CheckBox("")
        self.checkbox.setFixedWidth(36)
        self.checkbox.clicked.connect(self.clicked)
        layout.addWidget(self.checkbox, 0, Qt.AlignTop)

        text_col = QVBoxLayout()
        text_col.setContentsMargins(0, 0, 0, 0)
        text_col.setSpacing(2)

        title_row = QHBoxLayout()
        title_row.setContentsMargins(0, 0, 0, 0)
        title_row.setSpacing(6)

        self.title_label = QLabel()
        self.title_label.setTextInteractionFlags(Qt.NoTextInteraction)
        title_row.addWidget(self.title_label, 1)

        self.badge_label = QLabel("推荐")
        self.badge_label.setTextInteractionFlags(Qt.NoTextInteraction)
        self.badge_label.setVisible(False)
        title_row.addWidget(self.badge_label, 0, Qt.AlignVCenter)
        title_row.addStretch(0)

        self.subtitle_label = QLabel()
        self.subtitle_label.setTextInteractionFlags(Qt.NoTextInteraction)
        self.subtitle_label.setWordWrap(False)

        text_col.addLayout(title_row)
        text_col.addWidget(self.subtitle_label)
        layout.addLayout(text_col, 1)

        self.drag_handle = _FluentProfileDragHandle(self)
        self.drag_handle.dragRequested.connect(self.dragRequested)
        layout.addWidget(self.drag_handle, 0, Qt.AlignVCenter)

        for child in (self.title_label, self.subtitle_label, self.badge_label):
            child.installEventFilter(self)

        self.set_content(title, subtitle, enabled, recommended)
        self.set_selected(False)

    def set_content(self, title, subtitle, enabled, recommended=False):
        self._enabled = bool(enabled)
        self._recommended = bool(recommended)
        self.title_label.setText(title)
        self.subtitle_label.setText(subtitle)
        self.checkbox.setChecked(bool(enabled))
        self.drag_handle.setVisible(bool(enabled))
        self.badge_label.setVisible(self._recommended)
        self._apply_visual_state()

    def set_selected(self, selected):
        self._selected = bool(selected)
        self._apply_visual_state()

    def _apply_visual_state(self):
        if self._selected:
            self.setStyleSheet(
                "QWidget#profileRowItemFluent {"
                "background: rgba(230, 238, 248, 0.96);"
                "border: 1px solid rgba(0, 120, 212, 0.28);"
                "border-radius: 10px;"
                "}"
            )
        else:
            self.setStyleSheet(
                "QWidget#profileRowItemFluent {"
                "background: rgba(255, 255, 255, 0.88);"
                "border: 1px solid rgba(198, 210, 224, 0.32);"
                "border-radius: 10px;"
                "}"
            )

        if self._selected:
            title_style = "color:#173A63; font-size:13px; font-weight:600;"
            subtitle_style = "color:#43617E; font-size:11px;"
        elif self._enabled:
            title_style = "color:#24384D; font-size:13px; font-weight:600;"
            subtitle_style = "color:#5C6E81; font-size:11px;"
        else:
            title_style = "color:#2F4457; font-size:13px; font-weight:500;"
            subtitle_style = "color:#697B8D; font-size:11px;"

        self.title_label.setStyleSheet(title_style)
        self.subtitle_label.setStyleSheet(subtitle_style)
        self.badge_label.setStyleSheet(
            "color:#5F6F82; background: rgba(225, 231, 238, 0.92);"
            "border: 1px solid rgba(198, 210, 224, 0.85); border-radius: 9px;"
            "padding: 1px 7px; font-size:11px;"
        )

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.doubleClicked.emit()
        super().mouseDoubleClickEvent(event)

    def eventFilter(self, obj, event):
        if obj in {self.title_label, self.subtitle_label, self.badge_label}:
            if event.type() == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
                self.clicked.emit()
                return True
            if event.type() == QEvent.MouseButtonDblClick and event.button() == Qt.LeftButton:
                self.doubleClicked.emit()
                return True
        return super().eventFilter(obj, event)


class _FluentProfileRowListWidget(QListWidget):
    """Single-selection list used by the grouped text export dialog."""

    enabledRowDropped = Signal(str, int)
    toggleRequested = Signal(str)

    def __init__(self, allow_reorder=False, parent=None):
        super().__init__(parent)
        self._allow_reorder = bool(allow_reorder)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        if self._allow_reorder:
            self.setDragEnabled(True)
            self.viewport().setAcceptDrops(True)
            self.setAcceptDrops(True)
            self.setDropIndicatorShown(True)
            self.setDefaultDropAction(Qt.MoveAction)
            self.setDragDropMode(QAbstractItemView.DragDrop)
        else:
            self.setDragEnabled(False)
            self.setAcceptDrops(False)
            self.setDropIndicatorShown(False)
            self.setDragDropMode(QAbstractItemView.NoDragDrop)
        self._set_drag_feedback(False)

    def current_row_id(self):
        item = self.currentItem()
        if item is None:
            return ""
        return str(item.data(Qt.UserRole) or "").strip()

    def start_drag_for_row_id(self, rid):
        if not self._allow_reorder:
            return
        rid = str(rid or "").strip()
        if not rid:
            return
        for row in range(self.count()):
            item = self.item(row)
            if item is not None and str(item.data(Qt.UserRole) or "").strip() == rid:
                self.setCurrentRow(row)
                self.startDrag(Qt.MoveAction)
                return

    def _set_drag_feedback(self, active: bool):
        if active and self._allow_reorder:
            self.setStyleSheet(
                "QListView { border: 1px solid rgba(0, 120, 212, 0.45); "
                "background: rgba(0, 120, 212, 0.05); border-radius: 12px; }"
            )
        else:
            self.setStyleSheet("")

    def startDrag(self, supportedActions):
        if not self._allow_reorder:
            return
        rid = self.current_row_id()
        if not rid:
            return
        mime = QMimeData()
        mime.setData("application/x-profile-enabled-row-id", rid.encode("utf-8"))
        drag = QDrag(self)
        drag.setMimeData(mime)
        self._set_drag_feedback(True)
        try:
            drag.exec(Qt.MoveAction)
        finally:
            self._set_drag_feedback(False)

    def dragEnterEvent(self, event):
        if self._allow_reorder and event.mimeData().hasFormat("application/x-profile-enabled-row-id"):
            self._set_drag_feedback(True)
            event.acceptProposedAction()
            return
        super().dragEnterEvent(event)

    def dragLeaveEvent(self, event):
        self._set_drag_feedback(False)
        super().dragLeaveEvent(event)

    def dragMoveEvent(self, event):
        if self._allow_reorder and event.mimeData().hasFormat("application/x-profile-enabled-row-id"):
            event.acceptProposedAction()
            return
        super().dragMoveEvent(event)

    def dropEvent(self, event):
        if not self._allow_reorder:
            super().dropEvent(event)
            return
        data = event.mimeData()
        if not data.hasFormat("application/x-profile-enabled-row-id"):
            super().dropEvent(event)
            return
        self._set_drag_feedback(False)
        try:
            rid = bytes(data.data("application/x-profile-enabled-row-id")).decode("utf-8").strip()
            if not rid:
                event.ignore()
                return
            pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
            row = self.indexAt(pos).row()
            if row < 0:
                row = self.count()
            row = max(0, min(self.count(), row))
            self.enabledRowDropped.emit(rid, row)
            event.acceptProposedAction()
        except Exception:
            event.ignore()

    def keyPressEvent(self, event):
        rid = self.current_row_id()
        if rid and event.key() in (Qt.Key_Space, Qt.Key_Return, Qt.Key_Enter):
            self.toggleRequested.emit(rid)
            event.accept()
            return
        super().keyPressEvent(event)


class _GroupedFluentTextExportSettingsDialog(_SingleListTextExportSettingsDialog):
    """Grouped Fluent variant: enabled rows are sortable, candidates stay lightweight."""

    def __init__(self, parent=None, defaults=None):
        self._enabled_list = None
        self._candidate_list = None
        self._enabled_caption_label = None
        self._candidate_caption_label = None
        self._candidate_body = None
        self._candidate_search = None
        self._candidate_toggle_btn = None
        self._candidate_expanded = False
        self._selection_syncing = False
        self._active_list_role = "enabled"
        super().__init__(parent=parent, defaults=defaults)
        self.setStyleSheet(self.styleSheet() + """
            QWidget#profileRowsSection {
                background: rgba(255,255,255,0.66);
                border: 1px solid rgba(209,219,231,0.8);
                border-radius: 14px;
            }
        """)

    def _build_rows_card(self):
        card = ElevatedCardWidget(self)
        card_lay = QVBoxLayout(card)
        card_lay.setSizeConstraint(QLayout.SetMinAndMaxSize)
        card_lay.setContentsMargins(12, 10, 12, 10)
        card_lay.setSpacing(10)

        title_row = QHBoxLayout()
        title_row.addWidget(BodyLabel("纵断面行内容"))
        title_row.addStretch(1)
        btn_preset = PushButton("应用亭子口二期项建/可研阶段模板")
        btn_preset.clicked.connect(self._apply_tingzikou_preset)
        title_row.addWidget(btn_preset)
        card_lay.addLayout(title_row)

        hint = self._make_wrap_caption(
            "已启用区优先完整展示；勾选即可启用；拖动右侧手柄排序；Ctrl+Up/Ctrl+Down 可微调顺序。"
        )
        card_lay.addWidget(hint)

        hidden_hint = self._make_wrap_caption("本版本暂不显示：IP文字(BE)、桩号文字(BK)。")
        card_lay.addWidget(hidden_hint)

        enabled_section = QWidget(self)
        enabled_section.setObjectName("profileRowsSection")
        enabled_lay = QVBoxLayout(enabled_section)
        enabled_lay.setSizeConstraint(QLayout.SetMinAndMaxSize)
        enabled_lay.setContentsMargins(10, 10, 10, 10)
        enabled_lay.setSpacing(6)
        enabled_section.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        enabled_header = QHBoxLayout()
        enabled_header.addWidget(BodyLabel("已启用"))
        self._enabled_caption_label = CaptionLabel("")
        enabled_header.addWidget(self._enabled_caption_label)
        enabled_header.addStretch(1)
        btn_up = PushButton("上移")
        btn_up.clicked.connect(lambda: self._move_selected_row(-1))
        btn_down = PushButton("下移")
        btn_down.clicked.connect(lambda: self._move_selected_row(1))
        btn_top = PushButton("置顶")
        btn_top.clicked.connect(lambda: self._move_selected_row_to_edge(True))
        btn_bottom = PushButton("置底")
        btn_bottom.clicked.connect(lambda: self._move_selected_row_to_edge(False))
        enabled_header.addWidget(btn_up)
        enabled_header.addWidget(btn_down)
        enabled_header.addWidget(btn_top)
        enabled_header.addWidget(btn_bottom)
        enabled_lay.addLayout(enabled_header)

        enabled_desc = self._make_wrap_caption("参与导出的项目会尽量一次性完整展示在这里，拖动右侧手柄即可排序。")
        enabled_lay.addWidget(enabled_desc)

        self._enabled_list = _FluentProfileRowListWidget(allow_reorder=True, parent=self)
        self._enabled_list.enabledRowDropped.connect(self._on_enabled_row_dropped)
        self._enabled_list.toggleRequested.connect(lambda rid: self._toggle_current_row(rid, show_feedback=True))
        self._enabled_list.itemDoubleClicked.connect(lambda _item: self._toggle_current_row(show_feedback=True))
        self._enabled_list.currentItemChanged.connect(
            lambda current, previous: self._on_list_current_changed("enabled", current, previous)
        )
        self._enabled_list.customContextMenuRequested.connect(
            lambda pos: self._show_grouped_row_context_menu("enabled", pos)
        )
        self._enabled_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._enabled_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._row_list = self._enabled_list
        enabled_lay.addWidget(self._enabled_list, 0)
        card_lay.addWidget(enabled_section, 0)

        candidate_section = QWidget(self)
        candidate_section.setObjectName("profileRowsSection")
        candidate_lay = QVBoxLayout(candidate_section)
        candidate_lay.setSizeConstraint(QLayout.SetMinAndMaxSize)
        candidate_lay.setContentsMargins(10, 10, 10, 10)
        candidate_lay.setSpacing(6)
        candidate_section.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        candidate_header = QHBoxLayout()
        candidate_header.addWidget(BodyLabel("可选项"))
        self._candidate_caption_label = CaptionLabel("")
        candidate_header.addWidget(self._candidate_caption_label)
        candidate_header.addStretch(1)
        self._candidate_toggle_btn = ToolButton(self._ICON_EXPANDED)
        self._candidate_toggle_btn.clicked.connect(self._toggle_candidate_section)
        candidate_header.addWidget(self._candidate_toggle_btn)
        candidate_lay.addLayout(candidate_header)

        self._candidate_body = QWidget(self)
        candidate_body_lay = QVBoxLayout(self._candidate_body)
        candidate_body_lay.setSizeConstraint(QLayout.SetMinAndMaxSize)
        candidate_body_lay.setContentsMargins(0, 0, 0, 0)
        candidate_body_lay.setSpacing(6)

        candidate_desc = self._make_wrap_caption("展开后可搜索并勾选加入；推荐顺序被打乱时，新项会追加到已启用末尾。")
        candidate_body_lay.addWidget(candidate_desc)

        search_row = QHBoxLayout()
        self._candidate_search = SearchLineEdit(self)
        self._candidate_search.setPlaceholderText("搜索可选项")
        self._candidate_search.textChanged.connect(lambda _text: self._refresh_all_row_lists())
        search_row.addWidget(self._candidate_search, 1)
        btn_enable_all = PushButton("全启用")
        btn_enable_all.clicked.connect(self._enable_all_rows)
        btn_disable_all = PushButton("全停用")
        btn_disable_all.clicked.connect(self._disable_all_rows)
        btn_restore_recommended = PushButton("恢复推荐")
        btn_restore_recommended.clicked.connect(self._restore_recommended_rows)
        search_row.addWidget(btn_enable_all)
        search_row.addWidget(btn_disable_all)
        search_row.addWidget(btn_restore_recommended)
        candidate_body_lay.addLayout(search_row)

        self._candidate_list = _FluentProfileRowListWidget(allow_reorder=False, parent=self)
        self._candidate_list.toggleRequested.connect(lambda rid: self._toggle_current_row(rid, show_feedback=True))
        self._candidate_list.itemDoubleClicked.connect(lambda _item: self._toggle_current_row(show_feedback=True))
        self._candidate_list.currentItemChanged.connect(
            lambda current, previous: self._on_list_current_changed("candidate", current, previous)
        )
        self._candidate_list.customContextMenuRequested.connect(
            lambda pos: self._show_grouped_row_context_menu("candidate", pos)
        )
        self._candidate_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._candidate_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        candidate_body_lay.addWidget(self._candidate_list, 0)

        candidate_lay.addWidget(self._candidate_body, 1)
        card_lay.addWidget(candidate_section, 0)
        self._set_candidate_expanded(False)
        return card

    def _set_candidate_expanded(self, expanded):
        self._candidate_expanded = bool(expanded)
        if self._candidate_body is not None:
            self._candidate_body.setVisible(self._candidate_expanded)
        if self._candidate_toggle_btn is not None:
            self._candidate_toggle_btn.setIcon(self._ICON_EXPANDED if self._candidate_expanded else self._ICON_COLLAPSED)
        self._ensure_row_lists_visible_rows()
        self._request_dialog_layout_refresh()

    def _toggle_candidate_section(self):
        self._set_candidate_expanded(not self._candidate_expanded)

    def _row_display(self, rid, enabled, order_index=None):
        row_def = _PROFILE_ROW_DEF_MAP[rid]
        title = row_def["label"]
        if enabled and order_index is not None:
            title = f"{order_index + 1:02d}. {title}"
        subtitle_parts = ["已启用" if enabled else "可选项"]
        hint = str(row_def.get("hint", "") or "").strip()
        if hint:
            subtitle_parts.append(hint)
        return title, " | ".join(subtitle_parts), rid in _PROFILE_RECOMMENDED_ROW_IDS

    def _create_row_item(self, rid, enabled, order_index=None):
        item = QListWidgetItem()
        item.setData(Qt.UserRole, rid)
        item.setData(Qt.UserRole + 1, bool(enabled))
        item.setSizeHint(QSize(0, 58))
        widget = self._create_row_widget_for_state(rid, enabled, order_index)
        return item, widget

    def _create_row_widget_for_state(self, rid, enabled, order_index=None):
        title, subtitle, recommended = self._row_display(rid, enabled, order_index)
        widget = _FluentProfileRowItemWidget(title, subtitle, enabled, recommended)
        widget.checkbox.stateChanged.connect(
            lambda _state, row_id=rid: self._on_row_widget_checkbox_changed(None, row_id)
        )
        widget.clicked.connect(
            lambda row_id=rid, prefer_enabled=bool(enabled): self._set_current_row_id(row_id, prefer_enabled=prefer_enabled)
        )
        widget.doubleClicked.connect(lambda row_id=rid: self._toggle_current_row(row_id, show_feedback=True))
        if enabled:
            widget.drag_handle.dragRequested.connect(lambda row_id=rid: self._enabled_list.start_drag_for_row_id(row_id))
        return widget

    def _refresh_row_widget_content(self, rid):
        widget = self._row_widgets.get(rid)
        if widget is None:
            return
        enabled = rid in self._enabled_row_ids
        order_index = self._enabled_row_ids.index(rid) if enabled else None
        title, subtitle, recommended = self._row_display(rid, enabled, order_index)
        widget.set_content(title, subtitle, enabled, recommended)

    def _on_row_widget_checkbox_changed(self, item, rid):
        if self._row_updating:
            return
        widget = self._row_widgets.get(rid)
        if widget is None:
            return
        self._set_current_row_id(rid, prefer_enabled=(rid in self._enabled_row_ids))
        self._set_row_enabled(rid, widget.checkbox.isChecked(), show_feedback=True)

    def _candidate_row_ids(self):
        row_ids = [rid for rid in self._ordered_row_ids if rid not in self._enabled_row_ids]
        query = ""
        if self._candidate_search is not None:
            query = self._candidate_search.text().strip().lower()
        if not query:
            return row_ids
        filtered = []
        for rid in row_ids:
            row_def = _PROFILE_ROW_DEF_MAP.get(rid, {})
            haystack = " ".join([
                rid,
                str(row_def.get("label", "")),
                str(row_def.get("hint", "")),
            ]).lower()
            if query in haystack:
                filtered.append(rid)
        return filtered

    def _refresh_section_labels(self):
        enabled_count = len(self._enabled_row_ids)
        candidate_count = len(self._candidate_row_ids())
        if self._enabled_caption_label is not None:
            self._enabled_caption_label.setText(f"{enabled_count} 项")
        if self._candidate_caption_label is not None:
            self._candidate_caption_label.setText(f"{candidate_count} 项")

    def _ensure_row_lists_visible_rows(self):
        if not self._enabled_list or not self._candidate_list:
            return
        enabled_row_h = self._enabled_list.sizeHintForRow(0)
        candidate_row_h = self._candidate_list.sizeHintForRow(0)
        if enabled_row_h <= 0:
            enabled_row_h = 50
        if candidate_row_h <= 0:
            candidate_row_h = 50

        enabled_visible = max(1, len(self._enabled_row_ids))
        candidate_count = len(self._candidate_row_ids())
        candidate_visible = max(1, candidate_count) if candidate_count else 0
        if not self._candidate_expanded:
            candidate_visible = 0

        enabled_height = enabled_row_h * enabled_visible + 12
        candidate_height = candidate_row_h * candidate_visible + 12 if candidate_visible else 0
        self._enabled_list.setMinimumHeight(enabled_height)
        self._enabled_list.setMaximumHeight(enabled_height)
        self._candidate_list.setMinimumHeight(candidate_height)
        self._candidate_list.setMaximumHeight(candidate_height)

    def _find_item_in_list(self, list_widget, rid):
        if list_widget is None:
            return None, -1
        for row in range(list_widget.count()):
            item = list_widget.item(row)
            if item is not None and str(item.data(Qt.UserRole) or "").strip() == rid:
                return item, row
        return None, -1

    def _insert_row_into_list(self, list_widget, rid, enabled, row):
        row = max(0, min(list_widget.count(), int(row)))
        item, widget = self._create_row_item(rid, enabled, None)
        item.setData(Qt.UserRole, rid)
        item.setData(Qt.UserRole + 1, bool(enabled))
        list_widget.insertItem(row, item)
        list_widget.setItemWidget(item, widget)
        self._row_widgets[rid] = widget
        self._refresh_row_widget_content(rid)

    def _refresh_all_row_lists(self, selected_rid="", prefer_enabled=True):
        if not self._enabled_list or not self._candidate_list:
            return
        self._normalize_row_model()
        self._row_updating = True
        try:
            self._enabled_list.clear()
            self._candidate_list.clear()
            self._row_widgets = {}
            for row, rid in enumerate(self._enabled_row_ids):
                self._insert_row_into_list(self._enabled_list, rid, True, row)
            for row, rid in enumerate(self._candidate_row_ids()):
                self._insert_row_into_list(self._candidate_list, rid, False, row)
        finally:
            self._row_updating = False

        self._refresh_section_labels()
        self._ensure_row_lists_visible_rows()
        if selected_rid:
            self._set_current_row_id(selected_rid, prefer_enabled=prefer_enabled)
        else:
            self._update_row_widget_selection()
        self._refresh_runtime_advanced_view()

    def _load_rows(self, row_items):
        normalized = _normalize_profile_row_items(row_items)
        self._ordered_row_ids = [item["id"] for item in normalized]
        self._enabled_row_ids = [item["id"] for item in normalized if item.get("enabled")]
        candidate_ids = [item["id"] for item in normalized if not item.get("enabled")]
        prefer_enabled = bool(self._enabled_row_ids)
        selected_rid = self._enabled_row_ids[0] if prefer_enabled else (candidate_ids[0] if candidate_ids else "")
        self._refresh_all_row_lists(selected_rid=selected_rid, prefer_enabled=prefer_enabled)

    def _selected_row_id(self):
        primary = self._enabled_list if self._active_list_role == "enabled" else self._candidate_list
        secondary = self._candidate_list if primary is self._enabled_list else self._enabled_list
        for list_widget in (primary, secondary):
            if list_widget is None:
                continue
            rid = list_widget.current_row_id()
            if rid:
                return rid
        return ""

    def _set_current_row_id(self, rid, prefer_enabled=None):
        rid = str(rid or "").strip()
        if not rid:
            return
        if prefer_enabled is True:
            roles = [("enabled", self._enabled_list), ("candidate", self._candidate_list)]
        elif prefer_enabled is False:
            roles = [("candidate", self._candidate_list), ("enabled", self._enabled_list)]
        else:
            roles = [(self._active_list_role, self._enabled_list if self._active_list_role == "enabled" else self._candidate_list)]
            roles.append((
                "candidate" if self._active_list_role == "enabled" else "enabled",
                self._candidate_list if self._active_list_role == "enabled" else self._enabled_list,
            ))

        for role, list_widget in roles:
            item, _row = self._find_item_in_list(list_widget, rid)
            if item is None:
                continue
            self._selection_syncing = True
            try:
                other = self._candidate_list if role == "enabled" else self._enabled_list
                if other is not None:
                    other.clearSelection()
                    other.setCurrentItem(None)
                list_widget.setCurrentItem(item)
                list_widget.scrollToItem(item, QAbstractItemView.PositionAtCenter)
                list_widget.setFocus()
            finally:
                self._selection_syncing = False
            self._active_list_role = role
            self._update_row_widget_selection()
            return

    def _on_list_current_changed(self, role, current, previous):
        if self._selection_syncing or current is None:
            self._update_row_widget_selection()
            return
        other = self._candidate_list if role == "enabled" else self._enabled_list
        self._selection_syncing = True
        try:
            if other is not None:
                other.clearSelection()
                other.setCurrentItem(None)
        finally:
            self._selection_syncing = False
        self._active_list_role = role
        self._update_row_widget_selection()

    def _update_row_widget_selection(self):
        current_rid = self._selected_row_id()
        for rid, widget in self._row_widgets.items():
            if widget is not None:
                widget.set_selected(rid == current_rid)



    def _show_grouped_row_context_menu(self, role, pos):
        list_widget = self._enabled_list if role == "enabled" else self._candidate_list
        if list_widget is None:
            return
        item = list_widget.itemAt(pos)
        if item is None:
            return
        rid = str(item.data(Qt.UserRole) or "").strip()
        self._set_current_row_id(rid, prefer_enabled=(role == "enabled"))

        menu = QMenu(self)
        if role == "enabled":
            action_toggle = menu.addAction("停用")
            menu.addSeparator()
            action_up = menu.addAction("上移")
            action_down = menu.addAction("下移")
            action_top = menu.addAction("置顶")
            action_bottom = menu.addAction("置底")
            row = self._enabled_row_ids.index(rid)
            action_up.setEnabled(row > 0)
            action_top.setEnabled(row > 0)
            action_down.setEnabled(row < len(self._enabled_row_ids) - 1)
            action_bottom.setEnabled(row < len(self._enabled_row_ids) - 1)
        else:
            action_toggle = menu.addAction("启用")
            action_up = action_down = action_top = action_bottom = None

        chosen = menu.exec(list_widget.viewport().mapToGlobal(pos))
        if chosen == action_toggle:
            self._set_row_enabled(rid, role != "enabled", show_feedback=True)
        elif chosen == action_up:
            self._move_selected_row(-1)
        elif chosen == action_down:
            self._move_selected_row(1)
        elif chosen == action_top:
            self._move_selected_row_to_edge(True)
        elif chosen == action_bottom:
            self._move_selected_row_to_edge(False)

    def _row_data_from_table(self):
        self._normalize_row_model()
        enabled = set(self._enabled_row_ids)
        return _normalize_profile_row_items([
            {"id": rid, "enabled": rid in enabled}
            for rid in self._ordered_row_ids
        ])

    def _get_recommended_insert_row(self, rid):
        if rid not in _PROFILE_RECOMMENDED_ROW_IDS:
            return len(self._enabled_row_ids)

        current_recommended = [row_id for row_id in self._enabled_row_ids if row_id in _PROFILE_RECOMMENDED_ROW_IDS]
        expected_recommended = [
            row_id for row_id in _PROFILE_ROW_VISIBLE_ORDER
            if row_id in _PROFILE_RECOMMENDED_ROW_IDS and row_id in current_recommended
        ]
        if current_recommended != expected_recommended:
            return len(self._enabled_row_ids)

        rid_index = _PROFILE_ROW_VISIBLE_ORDER.index(rid)
        for row, row_id in enumerate(self._enabled_row_ids):
            if row_id in _PROFILE_RECOMMENDED_ROW_IDS and _PROFILE_ROW_VISIBLE_ORDER.index(row_id) > rid_index:
                return row

        previous_recommended = [
            row_id for row_id in self._enabled_row_ids
            if row_id in _PROFILE_RECOMMENDED_ROW_IDS and _PROFILE_ROW_VISIBLE_ORDER.index(row_id) < rid_index
        ]
        if previous_recommended:
            return self._enabled_row_ids.index(previous_recommended[-1]) + 1
        return len(self._enabled_row_ids)

    def _set_row_enabled(self, rid, enabled, *, show_feedback=False):
        if rid not in _PROFILE_ROW_VISIBLE_ID_SET:
            return
        current_enabled = rid in self._enabled_row_ids
        enabled = bool(enabled)
        if current_enabled == enabled:
            return

        if enabled:
            insert_row = self._get_recommended_insert_row(rid)
            self._enabled_row_ids = [row_id for row_id in self._enabled_row_ids if row_id != rid]
            self._enabled_row_ids.insert(insert_row, rid)
        else:
            self._enabled_row_ids = [row_id for row_id in self._enabled_row_ids if row_id != rid]

        self._normalize_row_model()
        self._refresh_all_row_lists(selected_rid=rid, prefer_enabled=enabled)

        if show_feedback:
            row_label = _PROFILE_ROW_DEF_MAP[rid]["label"]
            if enabled:
                InfoBar.success(
                    "已启用",
                    f"{row_label} 已加入导出。",
                    parent=self,
                    position=InfoBarPosition.TOP_RIGHT,
                    duration=1200,
                )
            else:
                InfoBar.info(
                    "已停用",
                    f"{row_label} 已移回可选项。",
                    parent=self,
                    position=InfoBarPosition.TOP_RIGHT,
                    duration=1200,
                )

    def _toggle_current_row(self, rid=None, *, show_feedback=False):
        rid = str(rid or self._selected_row_id() or "").strip()
        if not rid:
            return
        self._set_row_enabled(rid, rid not in self._enabled_row_ids, show_feedback=show_feedback)

    def _reorder_enabled_row(self, rid, target_row):
        enabled = list(self._enabled_row_ids)
        if rid not in enabled:
            return
        old_row = enabled.index(rid)
        enabled.pop(old_row)
        target_row = max(0, min(len(enabled), int(target_row)))
        if target_row > old_row:
            target_row -= 1
        enabled.insert(target_row, rid)
        self._enabled_row_ids = enabled
        self._normalize_row_model()
        self._refresh_all_row_lists(selected_rid=rid, prefer_enabled=True)

    def _enable_all_rows(self):
        self._enabled_row_ids = list(_PROFILE_ROW_VISIBLE_ORDER)
        self._refresh_all_row_lists(selected_rid=self._enabled_row_ids[0] if self._enabled_row_ids else "", prefer_enabled=True)
        InfoBar.success("已全启用", "所有可选行已加入导出。", parent=self, position=InfoBarPosition.TOP_RIGHT, duration=1200)

    def _disable_all_rows(self):
        self._enabled_row_ids = []
        candidate_ids = [item["id"] for item in _default_profile_row_items()]
        self._refresh_all_row_lists(selected_rid=candidate_ids[0] if candidate_ids else "", prefer_enabled=False)
        InfoBar.info("已全停用", "当前没有启用任何导出行。", parent=self, position=InfoBarPosition.TOP_RIGHT, duration=1200)

    def _restore_recommended_rows(self):
        self._enabled_row_ids = [
            rid for rid in _PROFILE_ROW_VISIBLE_ORDER
            if rid in _PROFILE_RECOMMENDED_ROW_IDS
        ]
        selected_rid = self._enabled_row_ids[0] if self._enabled_row_ids else ""
        self._refresh_all_row_lists(selected_rid=selected_rid, prefer_enabled=True)
        InfoBar.success("已恢复推荐", "已切换到推荐的启用项组合。", parent=self, position=InfoBarPosition.TOP_RIGHT, duration=1200)

    def _apply_tingzikou_preset(self):
        ordered = list(_TINGZIKOU_TEMPLATE_ROW_IDS) + [
            rid for rid in _PROFILE_ROW_VISIBLE_ORDER if rid not in _TINGZIKOU_TEMPLATE_ROW_IDS
        ]
        self._ordered_row_ids = ordered
        self._enabled_row_ids = list(_TINGZIKOU_TEMPLATE_ROW_IDS)
        self._refresh_all_row_lists(selected_rid=self._enabled_row_ids[0], prefer_enabled=True)
        InfoBar.success(
            "模板已应用",
            "已切换为亭子口推荐顺序。",
            parent=self,
            position=InfoBarPosition.TOP_RIGHT,
            duration=1500,
        )

TextExportSettingsDialog = create_text_export_settings_dialog(globals())


class PressurizedPipeConfigDialog(QDialog):
    """倒虹吸/有压管道参数配置对话框（导出全部DXF专用）。"""

    def __init__(self, parent=None, siphon_rows=None, pressure_pipe_rows=None, materials=None):
        super().__init__(parent)
        self.setWindowTitle("有压流建筑物参数设置")
        self.setMinimumSize(520, 420)
        self.setStyleSheet(DIALOG_STYLE)
        self.result = None

        self._materials = list(materials or _PRESSURIZED_PIPE_MATERIALS)
        if not self._materials:
            self._materials = ["球墨铸铁管"]
        self._siphon_rows = []
        self._pressure_pipe_rows = []

        lay = QVBoxLayout(self)
        desc = QLabel(
            "请确认倒虹吸/有压管道导出参数。\n"
            "倒虹吸按材质确定糙率 n；有压管道按材质自动派生摩阻参数 f/m/b，并结合 DN 计算设计流速。"
        )
        desc.setWordWrap(True)
        desc.setStyleSheet("font-size:12px; color:#333;")
        lay.addWidget(desc)

        if siphon_rows:
            self._build_group(
                parent_layout=lay,
                group_title="倒虹吸参数",
                name_header="倒虹吸名称（含流量段）",
                source_rows=siphon_rows,
                target_rows=self._siphon_rows,
                structure_kind="siphon",
            )
        if pressure_pipe_rows:
            self._build_group(
                parent_layout=lay,
                group_title="有压管道参数",
                name_header="有压管道名称（含流量段）",
                source_rows=pressure_pipe_rows,
                target_rows=self._pressure_pipe_rows,
                structure_kind="pressure_pipe",
            )

        btn_lay = QHBoxLayout()
        btn_lay.addStretch()
        btn_cancel = PushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        btn_ok = PrimaryPushButton("确认")
        btn_ok.clicked.connect(self._on_confirm)
        btn_lay.addWidget(btn_cancel)
        btn_lay.addWidget(btn_ok)
        lay.addLayout(btn_lay)

        QShortcut(QKeySequence(Qt.Key_Escape), self, self.reject)
        QShortcut(QKeySequence(Qt.Key_Return), self, self._on_confirm)

    def _build_group(self, parent_layout, group_title, name_header, source_rows, target_rows, structure_kind):
        group = QGroupBox(group_title)
        group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        glay = QVBoxLayout(group)

        hdr = QGridLayout()
        hdr.setSpacing(6)
        for ci, txt in enumerate([name_header, "管道材质", "DN (mm)"]):
            lbl = QLabel(txt)
            lbl.setStyleSheet("font-size:11px; color:#555; font-weight:bold;")
            hdr.addWidget(lbl, 0, ci)
        hdr.setColumnStretch(0, 2)
        hdr.setColumnStretch(1, 3)
        hdr.setColumnStretch(2, 2)
        glay.addLayout(hdr)

        grid = QGridLayout()
        grid.setSpacing(4)
        normalized_rows = _normalize_pressurized_cache_rows(source_rows, structure_kind)
        for ri, row in enumerate(normalized_rows):
            name_lbl = QLabel(row["display_name"])
            name_lbl.setStyleSheet("font-size:12px;")
            grid.addWidget(name_lbl, ri, 0)

            mat_combo = QComboBox()
            mat_combo.addItems(self._materials)
            mat_combo.setCurrentText(row["pipe_material"] if row["pipe_material"] in self._materials else self._materials[0])
            mat_combo.setFixedWidth(160)
            grid.addWidget(mat_combo, ri, 1)

            dn_edit = LineEdit()
            dn_edit.setFixedWidth(100)
            dn_edit.setText(str(_normalize_dn_mm(row["DN_mm"], 1500)))
            grid.addWidget(dn_edit, ri, 2)

            target_rows.append((dict(row), mat_combo, dn_edit))

        grid.setColumnStretch(0, 2)
        grid.setColumnStretch(1, 3)
        grid.setColumnStretch(2, 2)
        glay.addLayout(grid)
        parent_layout.addWidget(group)

    def _read_rows(self, rows, title_prefix):
        out = []
        for row, mat_combo, dn_edit in rows:
            dn = _parse_positive_dn(dn_edit.text())
            row_name = row.get("display_name") or row.get("name") or title_prefix
            if dn is None:
                fluent_error(self, "输入错误", f"{row_name} 的 DN 必须为正整数")
                return None
            out.append(
                _make_pressurized_param_row(
                    name=row.get("name"),
                    flow_section=row.get("flow_section"),
                    structure_kind=row.get("structure_kind"),
                    pipe_material=mat_combo.currentText(),
                    dn_mm=dn,
                    display_name=row.get("display_name"),
                )
            )
        return out

    def _on_confirm(self):
        siphon = self._read_rows(self._siphon_rows, "倒虹吸")
        if siphon is None:
            return
        pressure_pipe = self._read_rows(self._pressure_pipe_rows, "有压管道")
        if pressure_pipe is None:
            return
        self.result = {
            "siphon": siphon,
            "pressure_pipe": pressure_pipe,
        }
        self.accept()


# ================================================================
# 平面图参数设置对话框
# ================================================================

class PlanTextSettingsDialog(QDialog):
    """建筑物名称上平面图参数设置对话框"""

    def __init__(self, parent=None, defaults=None):
        super().__init__(parent)
        self.setWindowTitle("建筑物名称上平面图 - 参数设置")
        self.setMinimumWidth(380)
        self.setStyleSheet(DIALOG_STYLE)
        self.result = None
        if defaults is None:
            defaults = {}
        self._init_ui(defaults)

    def _init_ui(self, defaults):
        lay = QVBoxLayout(self)
        lay.addWidget(QLabel(
            "生成 AutoCAD -TEXT 命令，将建筑物名称平行于轴线放置。\n"
            "文字位于建筑物最中间两个IP点连线段的中点处。"
        ))

        form = QGridLayout()
        form.addWidget(QLabel("垂直偏移距离 (V):"), 0, 0)
        self.offset_edit = LineEdit(); self.offset_edit.setText(str(defaults.get('offset', 10)))
        self.offset_edit.setFixedWidth(100)
        form.addWidget(self.offset_edit, 0, 1)
        form.addWidget(QLabel("文字中心到轴线的距离"), 0, 2)

        form.addWidget(QLabel("文字高度:"), 1, 0)
        self.height_edit = LineEdit(); self.height_edit.setText(str(defaults.get('text_height', 10)))
        self.height_edit.setFixedWidth(100)
        form.addWidget(self.height_edit, 1, 1)
        form.addWidget(QLabel("AutoCAD -TEXT 字高"), 1, 2)
        lay.addLayout(form)

        # 预览
        preview_grp = QGroupBox("命令格式预览")
        preview_lay = QVBoxLayout(preview_grp)
        self._preview_label = QLabel()
        self._preview_label.setStyleSheet("color: gray;")
        preview_lay.addWidget(self._preview_label)
        lay.addWidget(preview_grp)
        self._update_preview()
        self.offset_edit.textChanged.connect(self._update_preview)
        self.height_edit.textChanged.connect(self._update_preview)

        btn_lay = QHBoxLayout()
        btn_lay.addStretch()
        btn_cancel = PushButton("取消"); btn_cancel.clicked.connect(self.reject)
        btn_ok = PrimaryPushButton("确定"); btn_ok.clicked.connect(self._on_confirm)
        btn_lay.addWidget(btn_cancel); btn_lay.addWidget(btn_ok)
        lay.addLayout(btn_lay)

        # 键盘快捷键
        QShortcut(QKeySequence(Qt.Key_Escape), self, self.reject)
        QShortcut(QKeySequence(Qt.Key_Return), self, self._on_confirm)

    def _update_preview(self):
        try:
            o = self.offset_edit.text().strip()
            h = self.height_edit.text().strip()
            self._preview_label.setText(
                f"-TEXT J MC x,y {h} 角度 建筑物名称\n"
                f"（文字中心偏移轴线 {o} 个单位）")
        except Exception:
            pass

    def _on_confirm(self):
        try:
            o = float(self.offset_edit.text().strip())
            h = float(self.height_edit.text().strip())
            if h <= 0:
                raise ValueError("文字高度必须大于0")
            self.result = {'offset': o, 'text_height': h}
            self.accept()
        except ValueError as e:
            fluent_error(self, "输入错误", f"请输入有效的数值:\n{e}")


# ================================================================
# 1. 生成纵断面表格 TXT
# ================================================================

def export_longitudinal_profile_txt(panel):
    """Generate longitudinal profile TXT in AutoCAD command format."""
    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    export_mode = "xxpipe" if _is_panel_xxpipe_mode(panel) else None
    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "没有数据可导出")
        return

    if export_mode == "xxpipe":
        nodes = _resolve_xxpipe_export_source_nodes(panel, fallback_nodes=nodes)
        if not nodes:
            fluent_info(panel.window(), "警告", "没有可用于 xx管 纵断面导出的节点数据。")
            return
        valid_nodes = list(nodes)
    else:
        valid_nodes = [n for n in nodes if n.bottom_elevation or n.top_elevation or n.water_level]
        if not valid_nodes:
            fluent_info(panel.window(), "警告", "没有可用的高程数据，请先执行计算。")
            return

    dlg = TextExportSettingsDialog(
        panel.window(),
        panel._text_export_settings,
        mode=export_mode or "standard",
    )
    if dlg.exec() != QDialog.Accepted or dlg.result is None:
        return

    panel._text_export_settings.update(dlg.result)
    settings = dlg.result

    try:
        ch_name = panel.channel_name_edit.text().strip()
        ch_level = panel.channel_level_combo.currentText()
        auto_name = f"{ch_name}{ch_level}_上纵断面表格.txt"
    except Exception:
        auto_name = "上纵断面表格.txt"

    file_path, _ = QFileDialog.getSaveFileName(
        panel, "保存上纵断面表格", auto_name,
        "文本文件 (*.txt);;所有文件 (*.*)")
    if not file_path:
        return

    _export_longitudinal_txt_to_path(
        panel,
        nodes,
        valid_nodes,
        settings,
        file_path,
        export_mode=export_mode,
    )


def _draw_xxpipe_profile_on_msp(
    msp,
    nodes,
    settings,
    station_prefix,
    *,
    xxpipe_profile_data,
    layer_prefix="",
):
    import ezdxf

    if not isinstance(xxpipe_profile_data, dict):
        raise ValueError("xx管纵断面绘制缺少可用的 profile_data")

    settings, enabled_row_ids, row_layout, _total_height, line_height, h_line_y_values = _build_xxpipe_profile_row_layout(
        settings
    )
    text_height = settings["text_height"]
    rotation = settings["rotation"]
    elev_decimals = _get_xxpipe_centerline_elev_decimals(settings)
    station_decimals = _get_xxpipe_station_decimals(settings)
    scale_x = settings.get("scale_x", 1)
    scale_y = settings.get("scale_y", 1)
    first_col_x_offset = text_height + 1.3

    profile_text_nodes = list(xxpipe_profile_data.get("profile_text_nodes", []) or [])
    if not profile_text_nodes:
        raise ValueError("xx管纵断面导出没有可用于绘制的节点")

    centerline_records = list(xxpipe_profile_data.get("centerline_records", []) or [])
    centerline_points = list(xxpipe_profile_data.get("centerline_points", []) or [])
    ip_records = list(xxpipe_profile_data.get("ip_records", []) or [])
    building_segments = list(xxpipe_profile_data.get("building_segments", []) or [])
    material_segments = list(xxpipe_profile_data.get("material_segments", []) or [])

    def sx(mc):
        return _profile_meters_to_paper_mm(mc, scale_x)

    def sy(elev):
        return _profile_meters_to_paper_mm(elev, scale_y)

    def fmt_elev(value):
        if value is None:
            return f"{0:.{elev_decimals}f}"
        return f"{value:.{elev_decimals}f}"

    last_mc = _profile_station_value(profile_text_nodes[-1])
    layer_grid = layer_prefix + "表格线框"
    layer_text = layer_prefix + "文字标注"
    layer_centerline = layer_prefix + "管中心线"

    for hy in h_line_y_values:
        msp.add_line((-40, hy), (sx(0), hy), dxfattribs={"layer": layer_grid})
    msp.add_line((-40, 0), (-40, line_height), dxfattribs={"layer": layer_grid})
    msp.add_line((0, 0), (0, line_height), dxfattribs={"layer": layer_grid})

    top_merge_bottom = row_layout["building_name"]["bottom"]
    bottom_merge_top = row_layout["pipe_material"]["top"]
    lower_half_top = row_layout["ip_name"]["top"]
    full_height_boundary_mcs = {
        round(float(mc), 9)
        for mc in _collect_xxpipe_full_height_boundary_mcs(xxpipe_profile_data)
    }
    lower_half_vertical_mcs = {
        round(float(mc), 9)
        for mc in _collect_xxpipe_lower_half_vertical_line_mcs(profile_text_nodes)
    }

    for node in profile_text_nodes:
        station_mc = _profile_station_value(node)
        station_key = round(station_mc, 9)
        if station_key in full_height_boundary_mcs or top_merge_bottom <= bottom_merge_top:
            y0, y1 = 0.0, line_height
        elif station_key in lower_half_vertical_mcs:
            y0, y1 = 0.0, lower_half_top
        else:
            y0, y1 = bottom_merge_top, top_merge_bottom
        msp.add_line((sx(station_mc), y0), (sx(station_mc), y1), dxfattribs={"layer": layer_grid})

    for hy in h_line_y_values:
        msp.add_line((sx(0), hy), (sx(last_mc), hy), dxfattribs={"layer": layer_grid})

    if len(centerline_points) >= 2:
        msp.add_lwpolyline(
            [(sx(mc), sy(elev)) for mc, elev in centerline_points],
            dxfattribs={"layer": layer_centerline},
        )

    text_attr_rot = {"layer": layer_text, "height": text_height, "rotation": rotation, "width": 0.7, "style": "Standard"}
    text_attr_no_rot = {"layer": layer_text, "height": text_height, "width": 0.7, "style": "Standard"}
    centerline_elev_by_station = {
        round(float(record.get("station_mc", 0.0) or 0.0), 9): record.get("elevation")
        for record in centerline_records
    }

    for rid in enabled_row_ids:
        y_pos = row_layout[rid]["text_y"]
        if rid == "building_name":
            for segment in building_segments:
                msp.add_text(segment["text"], dxfattribs=text_attr_no_rot).set_placement(
                    (sx(segment["mid_mc"]), y_pos),
                    align=ezdxf.enums.TextEntityAlignment.MIDDLE,
                )
            continue

        if rid == "pipe_material":
            for segment in material_segments:
                msp.add_text(segment["text"], dxfattribs=text_attr_no_rot).set_placement(
                    (sx(segment["mid_mc"]), y_pos),
                    align=ezdxf.enums.TextEntityAlignment.MIDDLE,
                )
            continue

        if rid == "ip_name":
            for idx, rec in enumerate(ip_records):
                text_x = sx(rec["x"]) + first_col_x_offset if idx == 0 else sx(rec["x"]) - 1
                msp.add_text(rec["text"], dxfattribs=text_attr_rot).set_placement((text_x, y_pos))
            continue

        if rid == "station":
            for idx, node in enumerate(profile_text_nodes):
                station_mc = _profile_station_value(node)
                text_x = sx(station_mc) + first_col_x_offset if idx == 0 else sx(station_mc) - 1
                text = _format_xxpipe_station(
                    station_mc,
                    station_prefix,
                    decimals=station_decimals,
                )
                msp.add_text(text, dxfattribs=text_attr_rot).set_placement((text_x, y_pos))
            continue

        if rid == "centerline_elev":
            for idx, node in enumerate(profile_text_nodes):
                station_mc = _profile_station_value(node)
                text_x = sx(station_mc) + first_col_x_offset if idx == 0 else sx(station_mc) - 1
                text = fmt_elev(centerline_elev_by_station.get(round(station_mc, 9)))
                msp.add_text(text, dxfattribs=text_attr_rot).set_placement((text_x, y_pos))
            continue

    header_cx = -40 + 20
    for rid in enabled_row_ids:
        row_info = row_layout[rid]
        labels = row_info.get("header_lines", [])
        if not labels:
            continue
        if len(labels) == 1:
            msp.add_text(
                labels[0],
                dxfattribs=text_attr_no_rot,
            ).set_placement(
                (header_cx, (row_info["bottom"] + row_info["top"]) / 2.0),
                align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER,
            )
            continue
        line_spacing = text_height * 2.5
        block_h = line_spacing + text_height
        y_bottom_line = row_info["bottom"] + (row_info["height"] - block_h) / 2.0 + text_height / 2.0
        y_top_line = y_bottom_line + line_spacing
        msp.add_text(labels[0], dxfattribs=text_attr_no_rot).set_placement(
            (header_cx, y_top_line), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER
        )
        msp.add_text(labels[1], dxfattribs=text_attr_no_rot).set_placement(
            (header_cx, y_bottom_line), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER
        )

    return 40 + sx(last_mc), line_height


# ================================================================
# 1b. ?????????
# ================================================================

def _draw_profile_on_msp(
    msp,
    nodes,
    valid_nodes,
    settings,
    station_prefix,
    layer_prefix="",
    export_mode=None,
    xxpipe_profile_data=None,
):
    """在 modelspace 上绘制纵断面表格（核心绘图逻辑）。

    msp 可以是真实的 ezdxf modelspace 或 _OffsetMSP 包装器。
    layer_prefix 用于合并导出时给图层名添加前缀以区分组件。
    返回 (width, height)。
    """
    if export_mode == "xxpipe":
        return _draw_xxpipe_profile_on_msp(
            msp,
            nodes,
            settings,
            station_prefix,
            xxpipe_profile_data=xxpipe_profile_data,
            layer_prefix=layer_prefix,
        )

    import ezdxf

    settings = _normalize_text_export_settings(settings)
    text_height = settings["text_height"]
    rotation = settings["rotation"]
    elev_decimals = int(settings.get("elev_decimals", 3))
    station_decimals = _get_standard_station_decimals(settings)
    scale_x = settings.get("scale_x", 1)
    scale_y = settings.get("scale_y", 1)
    enabled_row_ids, row_layout, _total_height, line_height, h_line_y_values = _build_profile_row_layout(settings)
    first_col_x_offset = text_height + 1.3

    def sx(mc):
        return _profile_meters_to_paper_mm(mc, scale_x)

    def sy(elev):
        return _profile_meters_to_paper_mm(elev, scale_y)

    def fmt_elev(value):
        if value is None:
            return f"{0:.{elev_decimals}f}"
        return f"{value:.{elev_decimals}f}"

    last_mc = nodes[-1].station_MC
    layer_grid = layer_prefix + "表格线框"
    layer_text = layer_prefix + "文字标注"

    # ======== 1. 表头区域线框 ========
    for hy in h_line_y_values:
        msp.add_line((-40, hy), (sx(0), hy), dxfattribs={"layer": layer_grid})
    msp.add_line((-40, 0), (-40, line_height), dxfattribs={"layer": layer_grid})
    msp.add_line((0, 0), (0, line_height), dxfattribs={"layer": layer_grid})

    # ======== 2. 节点竖线 ========
    if "slope" in row_layout:
        short_line_height = row_layout["slope"]["bottom"]
    elif "building_name" in row_layout:
        short_line_height = row_layout["building_name"]["bottom"]
    else:
        short_line_height = line_height

    has_bc_ec_rows = any(rid in row_layout for rid in _BC_ROW_IDS) or \
        any(rid in row_layout for rid in _EC_ROW_IDS)
    ip_segment_map = {}
    if has_bc_ec_rows:
        for n in _iter_profile_ip_nodes(nodes):
            _mc = float(getattr(n, "station_MC", 0) or 0.0)
            _bc = float(getattr(n, "station_BC", _mc) or _mc)
            _ec = float(getattr(n, "station_EC", _mc) or _mc)
            if abs(_bc - _mc) > 1e-9 or abs(_ec - _mc) > 1e-9:
                ip_segment_map[round(_mc, 6)] = n

    tall_line_mcs = []
    full_vline_mcs = set()
    for idx, node in enumerate(nodes):
        mc = node.station_MC
        is_special = _is_special_inout_node(node)
        if is_special:
            tall_line_mcs.append(mc)
        is_last_node = idx == len(nodes) - 1
        v_top = _resolve_profile_vline_top(
            is_special=is_special,
            is_last_node=is_last_node,
            short_line_height=short_line_height,
            line_height=line_height,
        )
        if v_top > short_line_height + 1e-9:
            full_vline_mcs.add(round(float(mc), 9))

        ip_ref = ip_segment_map.get(round(float(mc), 6)) if has_bc_ec_rows else None
        if ip_ref is not None:
            for seg_x, seg_y0, seg_y1 in _compute_node_vline_segments(
                    ip_ref, row_layout, enabled_row_ids, v_top):
                msp.add_line((sx(seg_x), seg_y0), (sx(seg_x), seg_y1),
                             dxfattribs={"layer": layer_grid})
        else:
            msp.add_line((sx(mc), 0), (sx(mc), v_top), dxfattribs={"layer": layer_grid})

    # ======== 3. 全宽水平线 ========
    for hy in h_line_y_values:
        msp.add_line((sx(0), hy), (sx(last_mc), hy), dxfattribs={"layer": layer_grid})

    # ======== 4. 渠底/渠顶/水面折线 ========
    bottom_pts = [(sx(n.station_MC), sy(n.bottom_elevation))
                  for n in valid_nodes if n.bottom_elevation]
    top_pts = [(sx(n.station_MC), sy(n.top_elevation))
               for n in valid_nodes if n.top_elevation]
    water_pts = [(sx(n.station_MC), sy(n.water_level))
                 for n in valid_nodes if n.water_level]

    if len(bottom_pts) >= 2:
        msp.add_lwpolyline(bottom_pts, dxfattribs={"layer": layer_prefix + "渠底高程线"})
    if len(top_pts) >= 2:
        msp.add_lwpolyline(top_pts, dxfattribs={"layer": layer_prefix + "渠顶高程线"})
    if len(water_pts) >= 2:
        msp.add_lwpolyline(water_pts, dxfattribs={"layer": layer_prefix + "设计水位线"})

    # ======== 5. 建筑物/坡降分段 ========
    name_mc_pairs = []
    for node in nodes:
        building_name = _get_building_display_name(node)
        if building_name:
            name_mc_pairs.append((building_name, node.station_MC))
    building_segments = []
    for bname, bmc in name_mc_pairs:
        if building_segments and building_segments[-1][0] == bname:
            building_segments[-1][1].append(bmc)
        else:
            building_segments.append((bname, [bmc]))
    building_segments = _merge_segments_across_gates(building_segments)

    # ======== 6. 各行文本 ========
    profile_text_nodes = _build_profile_text_nodes(nodes)
    slope_segments = _build_profile_slope_segments(nodes, profile_text_nodes=profile_text_nodes)
    ip_records = _build_ip_related_row_records(
        nodes,
        station_prefix,
        station_decimals=station_decimals,
    )
    text_attr_rot = {"layer": layer_text, "height": text_height, "rotation": rotation, "width": 0.7, "style": "Standard"}
    text_attr_no_rot = {"layer": layer_text, "height": text_height, "width": 0.7, "style": "Standard"}

    if "slope" in row_layout:
        slope_top = row_layout["slope"]["top"]
        slope_bottom = row_layout["slope"]["bottom"]
        for boundary_mc in _collect_profile_slope_boundary_mcs(slope_segments):
            if round(float(boundary_mc), 9) in full_vline_mcs:
                continue
            msp.add_line(
                (sx(boundary_mc), slope_bottom),
                (sx(boundary_mc), slope_top),
                dxfattribs={"layer": layer_grid},
            )

    for rid in enabled_row_ids:
        y_pos = row_layout[rid]["text_y"]
        if rid in ("bottom_elev", "top_elev", "water_elev", "station"):
            for idx, node in enumerate(profile_text_nodes):
                station_mc = _profile_station_value(node)
                text_x = sx(station_mc) + first_col_x_offset if idx == 0 else sx(station_mc) - 1
                if rid == "bottom_elev":
                    text = fmt_elev(node.bottom_elevation)
                elif rid == "top_elev":
                    text = fmt_elev(node.top_elevation)
                elif rid == "water_elev":
                    text = fmt_elev(node.water_level)
                else:
                    text = _format_station_with_decimals(
                        station_mc,
                        station_prefix,
                        decimals=station_decimals,
                    )
                msp.add_text(text, dxfattribs=text_attr_rot).set_placement((text_x, y_pos))
            continue

        if rid == "building_name":
            for bname, mc_list in building_segments:
                if _is_gate_name(bname):
                    mid_mc = mc_list[0]
                else:
                    seg_start = mc_list[0]
                    seg_end = mc_list[-1]
                    mid_mc = _resolve_segment_mid_mc(seg_start, seg_end, tall_line_mcs)
                msp.add_text(bname, dxfattribs=text_attr_no_rot).set_placement(
                    (sx(mid_mc), y_pos), align=ezdxf.enums.TextEntityAlignment.MIDDLE
                )
            continue

        if rid == "slope":
            for segment in slope_segments:
                mid_mc = (segment["start_mc"] + segment["end_mc"]) / 2.0
                msp.add_text(segment["text"], dxfattribs=text_attr_no_rot).set_placement(
                    (sx(mid_mc), y_pos), align=ezdxf.enums.TextEntityAlignment.MIDDLE
                )
            continue

        if rid in ip_records:
            for idx, rec in enumerate(ip_records[rid]):
                text_x = sx(rec["x"]) + first_col_x_offset if idx == 0 else sx(rec["x"]) - 1
                msp.add_text(rec["text"], dxfattribs=text_attr_rot).set_placement((text_x, y_pos))
            continue

    # ======== 7. 表头文字 ========
    header_cx = -40 + 20
    for rid in enabled_row_ids:
        row_info = row_layout[rid]
        labels = row_info.get("header_lines", [])
        if not labels:
            continue
        if len(labels) == 1:
            msp.add_text(
                labels[0], dxfattribs=text_attr_no_rot
            ).set_placement(
                (header_cx, (row_info["bottom"] + row_info["top"]) / 2.0),
                align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER
            )
            continue
        line_spacing = text_height * 2.5
        block_h = line_spacing + text_height
        y_bottom_line = row_info["bottom"] + (row_info["height"] - block_h) / 2.0 + text_height / 2.0
        y_top_line = y_bottom_line + line_spacing
        msp.add_text(labels[0], dxfattribs=text_attr_no_rot).set_placement(
            (header_cx, y_top_line), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER
        )
        msp.add_text(labels[1], dxfattribs=text_attr_no_rot).set_placement(
            (header_cx, y_bottom_line), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER
        )

    return 40 + sx(last_mc), line_height


# ================================================================
# 1b. 生成纵断面表格 DXF（直接生成 CAD 可打开的 DXF 文件）
# ================================================================

def export_longitudinal_profile_dxf(panel):
    """一键生成上纵断面表格 DXF

    直接生成 DXF 文件，包含表格线框、渠底/渠顶/水面折线、
    高程文字、里程桩号、建筑物名称、坡降、IP点名称等全部内容。
    双击即可在 AutoCAD / 浩辰CAD / 中望CAD 中打开。
    """
    import ezdxf

    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    export_mode = "xxpipe" if _is_panel_xxpipe_mode(panel) else None
    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "没有数据可导出")
        return

    if export_mode == "xxpipe":
        nodes = _resolve_xxpipe_export_source_nodes(panel, fallback_nodes=nodes)
        if not nodes:
            fluent_info(panel.window(), "警告", "没有可用于 xx管 纵断面导出的节点数据。")
            return
        valid_nodes = list(nodes)
    else:
        valid_nodes = [n for n in nodes if n.bottom_elevation or n.top_elevation or n.water_level]
        if not valid_nodes:
            fluent_info(panel.window(), "警告", "没有可用的高程数据，请先执行计算。")
            return

    # 弹出参数配置对话框（复用 TXT 版设置）
    dlg = TextExportSettingsDialog(
        panel.window(),
        panel._text_export_settings,
        mode=export_mode or "standard",
    )
    if dlg.exec() != QDialog.Accepted or dlg.result is None:
        return

    settings = _normalize_text_export_settings(dlg.result)
    panel._text_export_settings.update(settings)
    if export_mode != "xxpipe" and not _get_enabled_profile_row_ids(settings):
        fluent_error(panel.window(), "导出失败", "至少选择1项行内容后再导出。")
        return

    if export_mode != "xxpipe":
        _show_special_angle_warning(panel, nodes)

    # 自动文件名
    try:
        ch_name = panel.channel_name_edit.text().strip()
        ch_level = panel.channel_level_combo.currentText()
        auto_name = f"{ch_name}{ch_level}_上纵断面表格.dxf"
    except Exception:
        auto_name = "上纵断面表格.dxf"

    file_path, _ = QFileDialog.getSaveFileName(
        panel, "保存上纵断面表格", auto_name,
        "DXF 文件 (*.dxf);;文本文件 (*.txt);;所有文件 (*.*)")
    if not file_path:
        return

    # 如果用户选择了 .txt，走原有 TXT 导出逻辑
    if file_path.lower().endswith('.txt'):
        _export_longitudinal_txt_to_path(
            panel,
            nodes,
            valid_nodes,
            settings,
            file_path,
            export_mode=export_mode,
        )
        return

    try:
        try:
            proj_settings = panel._build_settings()
            station_prefix = proj_settings.get_station_prefix()
        except Exception:
            station_prefix = ""

        xxpipe_profile_data = None
        if export_mode == "xxpipe":
            xxpipe_profile_data = _build_panel_xxpipe_profile_data(panel, nodes, station_prefix=station_prefix)

        doc = ezdxf.new("R2010")
        msp = doc.modelspace()
        _setup_profile_dxf_document(doc)
        _ensure_profile_layers(doc)

        _draw_profile_on_msp(
            msp,
            nodes,
            valid_nodes,
            settings,
            station_prefix,
            export_mode=export_mode,
            xxpipe_profile_data=xxpipe_profile_data,
        )

        doc.saveas(file_path)

        if fluent_question(panel.window(), "完成",
                f"上纵断面表格 DXF 已生成（{len(nodes)} 个节点）:\n{file_path}\n\n是否立即打开该文件？"):
            os.startfile(file_path)

    except PermissionError:
        fluent_error(panel.window(), "文件被占用",
                     f"无法写入文件，该文件可能已被其他程序打开：\n\n{file_path}\n\n"
                     f"请先关闭该文件（如CAD等），然后重新操作。")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "导出错误", f"生成上纵断面表格 DXF 失败:\n{str(e)}")


def _export_xxpipe_longitudinal_txt_to_path(
    panel,
    nodes,
    settings,
    file_path,
    *,
    station_prefix="",
    xxpipe_profile_data=None,
):
    fmt = _format_number
    settings, enabled_row_ids, row_layout, _total_height, line_height, h_line_y_values = _build_xxpipe_profile_row_layout(
        settings
    )
    text_height = settings["text_height"]
    rotation = settings["rotation"]
    elev_decimals = _get_xxpipe_centerline_elev_decimals(settings)
    station_decimals = _get_xxpipe_station_decimals(settings)
    scale_x = settings.get("scale_x", 1)
    scale_y = settings.get("scale_y", 1)
    first_col_x_offset = text_height + 1.3

    if xxpipe_profile_data is None:
        xxpipe_profile_data = _build_panel_xxpipe_profile_data(panel, nodes, station_prefix=station_prefix)

    profile_text_nodes = list(xxpipe_profile_data.get("profile_text_nodes", []) or [])
    if not profile_text_nodes:
        raise ValueError("xx管纵断面导出没有可用于绘制的节点")

    centerline_records = list(xxpipe_profile_data.get("centerline_records", []) or [])
    centerline_points = list(xxpipe_profile_data.get("centerline_points", []) or [])
    ip_records = list(xxpipe_profile_data.get("ip_records", []) or [])
    building_segments = list(xxpipe_profile_data.get("building_segments", []) or [])
    material_segments = list(xxpipe_profile_data.get("material_segments", []) or [])

    def sx(mc):
        return _profile_meters_to_paper_mm(mc, scale_x)

    def sy(elev):
        return _profile_meters_to_paper_mm(elev, scale_y)

    def fmt_elev(value):
        if value is None:
            return f"{0:.{elev_decimals}f}"
        return f"{value:.{elev_decimals}f}"

    lines = []
    s_height = fmt(text_height)
    s_rotation = fmt(rotation)
    last_mc = _profile_station_value(profile_text_nodes[-1])

    for hy in h_line_y_values:
        hy_fmt = fmt(hy)
        lines.append(f"pl {fmt(sx(0))},{hy_fmt} -40,{hy_fmt} ")
    lines.append(f"pl -40,0 -40,{fmt(line_height)} ")
    lines.append(f"pl 0,0 0,{fmt(line_height)} ")
    lines.append("")

    top_merge_bottom = row_layout["building_name"]["bottom"]
    bottom_merge_top = row_layout["pipe_material"]["top"]
    lower_half_top = row_layout["ip_name"]["top"]
    full_height_boundary_mcs = {
        round(float(mc), 9)
        for mc in _collect_xxpipe_full_height_boundary_mcs(xxpipe_profile_data)
    }
    lower_half_vertical_mcs = {
        round(float(mc), 9)
        for mc in _collect_xxpipe_lower_half_vertical_line_mcs(profile_text_nodes)
    }
    for node in profile_text_nodes:
        station_mc = _profile_station_value(node)
        station_key = round(station_mc, 9)
        if station_key in full_height_boundary_mcs or top_merge_bottom <= bottom_merge_top:
            y0, y1 = 0.0, line_height
        elif station_key in lower_half_vertical_mcs:
            y0, y1 = 0.0, lower_half_top
        else:
            y0, y1 = bottom_merge_top, top_merge_bottom
        lines.append(f"pl {fmt(sx(station_mc))},{fmt(y0)} {fmt(sx(station_mc))},{fmt(y1)} ")
    lines.append("")

    for hy in h_line_y_values:
        lines.append(f"pl {fmt(sx(0))},{fmt(hy)} {fmt(sx(last_mc))},{fmt(hy)} ")
    lines.append("")

    for station_mc, elevation in centerline_points:
        lines.append(f"pl {fmt(sx(station_mc))},{fmt(sy(elevation))}")
    lines.append("")

    centerline_elev_by_station = {
        round(float(record.get("station_mc", 0.0) or 0.0), 9): record.get("elevation")
        for record in centerline_records
    }

    for rid in enabled_row_ids:
        y_pos = row_layout[rid]["text_y"]
        if rid == "building_name":
            for segment in building_segments:
                lines.append(f"-text j mc {fmt(sx(segment['mid_mc']))},{fmt(y_pos)} {s_height} 0 {segment['text']} ")
            lines.append("")
            continue

        if rid == "pipe_material":
            for segment in material_segments:
                lines.append(f"-text j mc {fmt(sx(segment['mid_mc']))},{fmt(y_pos)} {s_height} 0 {segment['text']} ")
            lines.append("")
            continue

        if rid == "ip_name":
            for idx, rec in enumerate(ip_records):
                text_x = sx(rec["x"]) + first_col_x_offset if idx == 0 else sx(rec["x"])
                lines.append(f"-text {fmt(text_x)},{fmt(y_pos)} {s_height} {s_rotation} {rec['text']} ")
            lines.append("")
            continue

        if rid == "station":
            for idx, node in enumerate(profile_text_nodes):
                station_mc = _profile_station_value(node)
                text_x = sx(station_mc) + first_col_x_offset if idx == 0 else sx(station_mc)
                text = _format_xxpipe_station(
                    station_mc,
                    station_prefix,
                    decimals=station_decimals,
                )
                lines.append(f"-text {fmt(text_x)},{fmt(y_pos)} {s_height} {s_rotation} {text} ")
            lines.append("")
            continue

        if rid == "centerline_elev":
            for idx, node in enumerate(profile_text_nodes):
                station_mc = _profile_station_value(node)
                text_x = sx(station_mc) + first_col_x_offset if idx == 0 else sx(station_mc)
                text = fmt_elev(centerline_elev_by_station.get(round(station_mc, 9)))
                lines.append(f"-text {fmt(text_x)},{fmt(y_pos)} {s_height} {s_rotation} {text} ")
            lines.append("")
            continue

    header_cx = fmt(-40 + 20)
    for rid in enabled_row_ids:
        row_info = row_layout[rid]
        labels = row_info.get("header_lines", [])
        if not labels:
            continue
        if len(labels) == 1:
            center_y = (row_info["bottom"] + row_info["top"]) / 2.0
            lines.append(f"-text j mc {header_cx},{fmt(center_y)} {s_height} 0 {labels[0]} ")
            continue
        line_spacing = text_height * 2.5
        block_h = line_spacing + text_height
        y_bottom_line = row_info["bottom"] + (row_info["height"] - block_h) / 2.0 + text_height / 2.0
        y_top_line = y_bottom_line + line_spacing
        lines.append(f"-text j mc {header_cx},{fmt(y_top_line)} {s_height} 0 {labels[0]} ")
        lines.append(f"-text j mc {header_cx},{fmt(y_bottom_line)} {s_height} 0 {labels[1]} ")
    lines.append("")

    with open(file_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    if fluent_question(panel.window(), "完成", f"上纵断面表格已生成（{len(nodes)} 个节点）：{file_path}"):
        os.startfile(file_path)


def _export_longitudinal_txt_to_path(
    panel,
    nodes,
    valid_nodes,
    settings,
    file_path,
    export_mode=None,
    xxpipe_profile_data=None,
):
    """Internal helper: export longitudinal profile as AutoCAD TXT commands."""
    fmt = _format_number

    settings = _normalize_text_export_settings(settings)
    text_height = settings["text_height"]
    rotation = settings["rotation"]
    elev_decimals = int(settings.get("elev_decimals", 3))
    station_decimals = _get_standard_station_decimals(settings)
    scale_x = settings.get("scale_x", 1)
    scale_y = settings.get("scale_y", 1)
    enabled_row_ids, row_layout, _total_height, line_height, h_line_y_values = _build_profile_row_layout(settings)
    first_col_x_offset = text_height + 1.3

    def sx(mc):
        return _profile_meters_to_paper_mm(mc, scale_x)

    def sy(elev):
        return _profile_meters_to_paper_mm(elev, scale_y)

    def fmt_elev(value):
        if value is None:
            return f"{0:.{elev_decimals}f}"
        return f"{value:.{elev_decimals}f}"

    try:
        proj_settings = panel._build_settings()
        station_prefix = proj_settings.get_station_prefix()
    except Exception:
        station_prefix = ""

    try:
        if export_mode == "xxpipe":
            return _export_xxpipe_longitudinal_txt_to_path(
                panel,
                nodes,
                settings,
                file_path,
                station_prefix=station_prefix,
                xxpipe_profile_data=xxpipe_profile_data,
            )

        lines = []
        s_height = fmt(text_height)
        s_rotation = fmt(rotation)

        # ======== 1. ?????? ========
        for hy in h_line_y_values:
            hy_fmt = fmt(hy)
            lines.append(f"pl {fmt(sx(0))},{hy_fmt} -40,{hy_fmt} ")
        lines.append(f"pl -40,0 -40,{fmt(line_height)} ")
        lines.append(f"pl 0,0 0,{fmt(line_height)} ")
        lines.append("")

        # ======== 2. ???? ========
        if "slope" in row_layout:
            short_line_height = row_layout["slope"]["bottom"]
        elif "building_name" in row_layout:
            short_line_height = row_layout["building_name"]["bottom"]
        else:
            short_line_height = line_height

        has_bc_ec_rows = any(rid in row_layout for rid in _BC_ROW_IDS) or \
            any(rid in row_layout for rid in _EC_ROW_IDS)
        ip_segment_map = {}
        if has_bc_ec_rows:
            for n in _iter_profile_ip_nodes(nodes):
                _mc = float(getattr(n, "station_MC", 0) or 0.0)
                _bc = float(getattr(n, "station_BC", _mc) or _mc)
                _ec = float(getattr(n, "station_EC", _mc) or _mc)
                if abs(_bc - _mc) > 1e-9 or abs(_ec - _mc) > 1e-9:
                    ip_segment_map[round(_mc, 6)] = n

        tall_line_mcs = []
        full_vline_mcs = set()
        for idx, node in enumerate(nodes):
            station_mc = float(getattr(node, "station_MC", 0) or 0.0)
            is_special = _is_special_inout_node(node)
            if is_special:
                tall_line_mcs.append(station_mc)
            is_last_node = idx == len(nodes) - 1
            v_top_val = _resolve_profile_vline_top(
                is_special=is_special,
                is_last_node=is_last_node,
                short_line_height=short_line_height,
                line_height=line_height,
            )
            if v_top_val > short_line_height + 1e-9:
                full_vline_mcs.add(round(station_mc, 9))

            ip_ref = ip_segment_map.get(round(station_mc, 6)) if has_bc_ec_rows else None
            if ip_ref is not None:
                for seg_x, seg_y0, seg_y1 in _compute_node_vline_segments(
                        ip_ref, row_layout, enabled_row_ids, v_top_val):
                    lines.append(f"pl {fmt(sx(seg_x))},{fmt(seg_y0)} {fmt(sx(seg_x))},{fmt(seg_y1)} ")
            else:
                station_text = fmt(sx(station_mc))
                v_top = fmt(v_top_val)
                lines.append(f"pl {station_text},0 {station_text},{v_top} ")
        lines.append("")

        # ======== 3. ????? ========
        last_mc_scaled = fmt(sx(nodes[-1].station_MC))
        for hy in h_line_y_values:
            lines.append(f"pl {fmt(sx(0))},{fmt(hy)} {last_mc_scaled},{fmt(hy)} ")
        lines.append("")

        # ======== 4. ??/??/???? ========
        for node in valid_nodes:
            if node.bottom_elevation:
                lines.append(f"pl {fmt(sx(node.station_MC))},{fmt(sy(node.bottom_elevation))}")
        lines.append("")
        for node in valid_nodes:
            if node.top_elevation:
                lines.append(f"pl {fmt(sx(node.station_MC))},{fmt(sy(node.top_elevation))}")
        lines.append("")
        for node in valid_nodes:
            if node.water_level:
                lines.append(f"pl {fmt(sx(node.station_MC))},{fmt(sy(node.water_level))}")
        lines.append("")

        # ======== 5. ???? ========
        profile_text_nodes = _build_profile_text_nodes(nodes)
        ip_records = _build_ip_related_row_records(
            nodes,
            station_prefix,
            station_decimals=station_decimals,
        )

        name_mc_pairs = []
        for node in nodes:
            building_name = _get_building_display_name(node)
            if building_name:
                name_mc_pairs.append((building_name, node.station_MC))
        building_segments = []
        for bname, bmc in name_mc_pairs:
            if building_segments and building_segments[-1][0] == bname:
                building_segments[-1][1].append(bmc)
            else:
                building_segments.append((bname, [bmc]))
        building_segments = _merge_segments_across_gates(building_segments)
        slope_segments = _build_profile_slope_segments(nodes, profile_text_nodes=profile_text_nodes)

        if "slope" in row_layout:
            slope_top = row_layout["slope"]["top"]
            slope_bottom = row_layout["slope"]["bottom"]
            for boundary_mc in _collect_profile_slope_boundary_mcs(slope_segments):
                if round(float(boundary_mc), 9) in full_vline_mcs:
                    continue
                lines.append(f"pl {fmt(sx(boundary_mc))},{fmt(slope_bottom)} {fmt(sx(boundary_mc))},{fmt(slope_top)} ")
            lines.append("")

        for rid in enabled_row_ids:
            y_pos = row_layout[rid]["text_y"]
            if rid in ("bottom_elev", "top_elev", "water_elev", "station"):
                for idx, node in enumerate(profile_text_nodes):
                    station_mc = _profile_station_value(node)
                    text_x = sx(station_mc) + first_col_x_offset if idx == 0 else sx(station_mc)
                    if rid == "bottom_elev":
                        text = fmt_elev(node.bottom_elevation)
                    elif rid == "top_elev":
                        text = fmt_elev(node.top_elevation)
                    elif rid == "water_elev":
                        text = fmt_elev(node.water_level)
                    else:
                        text = _format_station_with_decimals(
                            station_mc,
                            station_prefix,
                            decimals=station_decimals,
                        )
                    lines.append(f"-text {fmt(text_x)},{fmt(y_pos)} {s_height} {s_rotation} {text} ")
                lines.append("")
                continue

            if rid == "building_name":
                for bname, mc_list in building_segments:
                    if _is_gate_name(bname):
                        mid_mc = mc_list[0]
                    else:
                        seg_start = mc_list[0]
                        seg_end = mc_list[-1]
                        mid_mc = _resolve_segment_mid_mc(seg_start, seg_end, tall_line_mcs)
                    lines.append(f"-text j mc {fmt(sx(mid_mc))},{fmt(y_pos)} {s_height} 0 {bname} ")
                lines.append("")
                continue

            if rid == "slope":
                for segment in slope_segments:
                    mid_mc = (segment["start_mc"] + segment["end_mc"]) / 2.0
                    lines.append(f"-text j mc {fmt(sx(mid_mc))},{fmt(y_pos)} {s_height} 0 {segment['text']} ")
                lines.append("")
                continue

            if rid in ip_records:
                for idx, rec in enumerate(ip_records[rid]):
                    text_x = sx(rec["x"]) + first_col_x_offset if idx == 0 else sx(rec["x"])
                    lines.append(f"-text {fmt(text_x)},{fmt(y_pos)} {s_height} {s_rotation} {rec['text']} ")
                lines.append("")
                continue

        # ======== 6. ???? ========
        header_cx = fmt(-40 + 20)
        for rid in enabled_row_ids:
            row_info = row_layout[rid]
            labels = row_info.get("header_lines", [])
            if not labels:
                continue
            if len(labels) == 1:
                center_y = (row_info["bottom"] + row_info["top"]) / 2.0
                lines.append(f"-text j mc {header_cx},{fmt(center_y)} {s_height} 0 {labels[0]} ")
                continue
            line_spacing = text_height * 2.5
            block_h = line_spacing + text_height
            y_bottom_line = row_info["bottom"] + (row_info["height"] - block_h) / 2.0 + text_height / 2.0
            y_top_line = y_bottom_line + line_spacing
            lines.append(f"-text j mc {header_cx},{fmt(y_top_line)} {s_height} 0 {labels[0]} ")
            lines.append(f"-text j mc {header_cx},{fmt(y_bottom_line)} {s_height} 0 {labels[1]} ")
        lines.append("")

        with open(file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")

        if fluent_question(panel.window(), "完成", f"上纵断面表格已生成（{len(nodes)} 个节点）：{file_path}"):
            os.startfile(file_path)

    except PermissionError:
        fluent_error(panel.window(), "文件被占用", f"无法写入文件，请先关闭该文件：{file_path}")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "导出错误", f"生成上纵断面表格失败：{str(e)}")


# ================================================================
# 2. ??bzzh2????
# ================================================================

def _collect_bzzh2_rows(nodes, station_prefix, settings=None):
    """收集 bzzh2 导出所需的桩号与说明文本。"""
    station_decimals = _get_standard_station_decimals(settings)
    rows = []
    for node in nodes:
        try:
            in_out = getattr(node, 'in_out', None)
            if _in_out_val(in_out) not in ("进", "出"):
                continue
            if getattr(node, 'is_transition', False):
                continue

            station_mc = getattr(node, 'station_MC', 0.0)
            if not isinstance(station_mc, (int, float)):
                station_mc = 0.0
            station_str = _format_station_with_decimals(
                station_mc,
                station_prefix,
                decimals=station_decimals,
            )

            struct_name = ""
            struct_str = _struct_val(node.structure_type)
            if struct_str:
                if "隧洞" in struct_str:
                    struct_name = "隧洞"
                elif "倒虹吸" in struct_str:
                    struct_name = "倒虹吸"
                elif "有压管道" in struct_str:
                    struct_name = "有压管道"
                elif "渡槽" in struct_str:
                    struct_name = "渡槽"
                elif "暗涵" in struct_str:
                    struct_name = "暗涵"
                else:
                    struct_name = struct_str

            in_out_str = "进" if _in_out_val(in_out) == "进" else "出"
            name = getattr(node, 'name', '') or ''
            desc = f"{name}{struct_name}{in_out_str}"
            rows.append((station_str, desc))
        except Exception as node_err:
            import traceback; traceback.print_exc()
            print(f"[bzzh2] 跳过节点（处理异常）: {node_err}")
            continue
    return rows


def extract_bzzh2_data(panel):
    """bzzh2命令提取工具

    从计算结果中提取所有有进出口标识（进/出）的建筑物节点，
    按桩号排序，整理为制表符分隔的TXT文件，供ZDM的bzzh2命令使用。
    """
    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "表格中没有数据，请先导入或输入数据。")
        return

    try:
        proj_settings = panel._build_settings()
        station_prefix = proj_settings.get_station_prefix()
    except Exception:
        station_prefix = ""

    bzzh2_rows = _collect_bzzh2_rows(
        nodes,
        station_prefix,
        getattr(panel, "_text_export_settings", {}),
    )

    if not bzzh2_rows:
        fluent_info(
            panel.window(), "无可提取数据",
            "未找到有进出口标识的建筑物节点。\n\n"
            "bzzh2命令需要隧洞、倒虹吸、有压管道、渡槽等建筑物的进/出口数据。\n"
            "请确保表格中已有相关数据并完成计算。")
        return

    # 预览对话框
    preview_dlg = QDialog(panel.window())
    preview_dlg.setWindowTitle("预览 — bzzh2命令数据（ZDM用）")
    preview_dlg.setMinimumSize(600, 400)
    preview_dlg.setStyleSheet(DIALOG_STYLE)
    dlg_lay = QVBoxLayout(preview_dlg)

    dlg_lay.addWidget(QLabel(
        f"共 {len(bzzh2_rows)} 条建筑物进出口数据，请确认内容后点击「确认导出」保存为TXT文件。"))

    table = QTableWidget(len(bzzh2_rows), 2)
    table.setHorizontalHeaderLabels(["桩号", "说明"])
    table.horizontalHeader().setStretchLastSection(False)
    table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
    table.setEditTriggers(QAbstractItemView.NoEditTriggers)
    table.setAlternatingRowColors(True)
    for i, (s, d) in enumerate(bzzh2_rows):
        table.setItem(i, 0, QTableWidgetItem(s))
        table.setItem(i, 1, QTableWidgetItem(d))
    auto_resize_table(table)
    dlg_lay.addWidget(table)

    btn_lay = QHBoxLayout()
    btn_lay.addStretch()
    btn_cancel = PushButton("取消"); btn_cancel.clicked.connect(preview_dlg.reject)
    btn_ok = PrimaryPushButton("确认导出"); btn_ok.clicked.connect(preview_dlg.accept)
    btn_lay.addWidget(btn_cancel); btn_lay.addWidget(btn_ok)
    dlg_lay.addLayout(btn_lay)

    # 绑定 ESC 关闭 / Enter 确认
    QShortcut(QKeySequence(Qt.Key_Escape), preview_dlg, preview_dlg.reject)
    QShortcut(QKeySequence(Qt.Key_Return), preview_dlg, preview_dlg.accept)

    if preview_dlg.exec() != QDialog.Accepted:
        return

    # 保存文件
    try:
        ch_name = panel.channel_name_edit.text().strip()
        ch_level = panel.channel_level_combo.currentText()
        auto_name = f"{ch_name}{ch_level}_ZDM的bzzh2命令.txt"
    except Exception:
        auto_name = "ZDM的bzzh2命令.txt"

    file_path, _ = QFileDialog.getSaveFileName(
        panel, "保存bzzh2命令数据", auto_name,
        "文本文件 (*.txt);;所有文件 (*.*)")
    if not file_path:
        return

    try:
        with open(file_path, 'w', encoding='gbk', errors='replace') as f:
            for station_str, desc in bzzh2_rows:
                f.write(f"{station_str}\t{desc}\t\n")

        if fluent_question(panel.window(), "提取完成",
                f"bzzh2命令数据提取成功！\n\n"
                f"文件保存路径:\n{file_path}\n\n"
                f"导出数据行数: {len(bzzh2_rows)}\n\n"
                f"请使用ZDM的bzzh2命令完成建筑物进出口上平面图。\n\n"
                f"是否要立即打开该txt文件？"):
            try:
                os.startfile(file_path)
            except AttributeError:
                import subprocess
                subprocess.Popen(['xdg-open', file_path])
            except Exception:
                fluent_info(panel.window(), "打开文件",
                            f"无法自动打开文件，请手动打开:\n\n{file_path}")
    except PermissionError:
        fluent_error(panel.window(), "文件被占用",
                     f"无法写入文件，该文件可能已被其他程序打开：\n\n{file_path}\n\n"
                     f"请先关闭该文件（如记事本、Word等），然后重新操作。")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "提取失败",
                     f"bzzh2命令数据提取过程中发生错误:\n\n{str(e)}")


# ================================================================
# 3. 建筑物名称上平面图
# ================================================================

def export_building_name_plan(panel):
    """生成「平行于轴线的建筑物名称上平面图」AutoCAD -TEXT 命令

    对于进出口之间有多个IP点的建筑物，文字放置在最中间两个相邻
    IP点连线段的中点处（垂直偏移），方向角取该中间线段的方向。
    """
    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "表格中没有数据，请先导入或输入数据。")
        return

    try:
        # 按建筑物分组收集所有节点
        building_groups = {}
        building_order = []
        for node in nodes:
            if node.is_transition or getattr(node, 'is_auto_inserted_channel', False):
                continue
            if not node.name:
                continue
            key = (node.name, _struct_val(node.structure_type))
            if key not in building_groups:
                building_groups[key] = []
                building_order.append(key)
            building_groups[key].append(node)

        # 筛选有完整进出口且坐标有效的建筑物
        valid_buildings = []
        for key in building_order:
            group = building_groups[key]
            has_inlet = any(_in_out_val(n.in_out) == "进" for n in group)
            has_outlet = any(_in_out_val(n.in_out) == "出" for n in group)
            if not (has_inlet and has_outlet):
                continue
            coord_nodes = [n for n in group
                           if n.x is not None and n.y is not None and
                           (n.x != 0 or n.y != 0)]
            if len(coord_nodes) >= 2:
                valid_buildings.append((key, group, coord_nodes))

        if not valid_buildings:
            fluent_info(
                panel.window(), "无可提取数据",
                "未找到有效的建筑物进出口数据。\n\n"
                "需要隧洞、倒虹吸、有压管道、渡槽等建筑物同时存在进口和出口，\n"
                "且节点具有有效的X、Y坐标。")
            return

        # 参数设置对话框
        dlg = PlanTextSettingsDialog(panel.window(), panel._plan_text_settings)
        if dlg.exec() != QDialog.Accepted or dlg.result is None:
            return

        panel._plan_text_settings.update(dlg.result)
        offset = dlg.result['offset']
        text_height = dlg.result['text_height']

        # 生成 -TEXT 命令
        text_commands = []
        for key, all_nodes, coord_nodes in valid_buildings:
            N = len(coord_nodes)
            mid_right = N // 2
            mid_left = mid_right - 1

            node_a = coord_nodes[mid_left]
            node_b = coord_nodes[mid_right]
            # 建筑物名称上平面图与表3 / IP坐标表保持同一坐标口径：
            # node.x -> CAD X，node.y -> CAD Y。
            x1, y1 = node_a.x, node_a.y
            x2, y2 = node_b.x, node_b.y

            dx = x2 - x1
            dy = y2 - y1
            if abs(dx) < 1e-10 and abs(dy) < 1e-10:
                continue

            mx = (x1 + x2) / 2
            my = (y1 + y2) / 2
            angle_rad = math.atan2(dy, dx)
            angle_deg = math.degrees(angle_rad)

            text_x = mx - offset * math.sin(angle_rad)
            text_y = my + offset * math.cos(angle_rad)

            inlet_node = next(
                (n for n in all_nodes if _in_out_val(n.in_out) == "进"),
                all_nodes[0])
            struct_name = ""
            struct_str = _struct_val(inlet_node.structure_type)
            if struct_str:
                if "隧洞" in struct_str:
                    struct_name = "隧洞"
                elif "倒虹吸" in struct_str:
                    struct_name = "倒虹吸"
                elif "有压管道" in struct_str:
                    struct_name = "有压管道"
                elif "渡槽" in struct_str:
                    struct_name = "渡槽"
                elif "暗涵" in struct_str:
                    struct_name = "暗涵"
                else:
                    struct_name = struct_str

            building_name = f"{inlet_node.name or ''}{struct_name}"
            cmd = (f"-TEXT J MC {text_x},{text_y} "
                   f"{text_height} {angle_deg} {building_name}")
            text_commands.append((building_name, N, cmd))

        if not text_commands:
            fluent_info(panel.window(), "无有效数据",
                        "没有生成任何 -TEXT 命令，请检查建筑物坐标。")
            return

        # 显示预览
        all_cmds_text = "\n".join(cmd for _, _, cmd in text_commands)

        preview = QDialog(panel.window())
        preview.setWindowTitle(f"建筑物名称上平面图 — {len(text_commands)} 条命令")
        preview.setMinimumSize(700, 400)
        preview.setStyleSheet(DIALOG_STYLE)
        p_lay = QVBoxLayout(preview)

        p_lay.addWidget(QLabel(
            f"共 {len(text_commands)} 个建筑物  |  "
            f"偏移距离: {offset}  |  文字高度: {text_height}"))

        text_widget = QTextEdit()
        text_widget.setReadOnly(True)
        text_widget.setFont(QFont("Consolas", 10))
        html_parts = []
        for i, (name, node_count, cmd) in enumerate(text_commands):
            comment = f"' [{i+1}] {name}（{node_count}个IP点）"
            html_parts.append(f'<span style="color:gray">{comment}</span><br>')
            html_parts.append(f'{cmd}<br><br>')
        text_widget.setHtml('<pre style="font-family:Consolas;font-size:10pt">' +
                            ''.join(html_parts) + '</pre>')
        p_lay.addWidget(text_widget)

        btn_lay = QHBoxLayout()
        status_label = QLabel("")
        status_label.setStyleSheet("color: green;")
        btn_lay.addWidget(status_label)
        btn_lay.addStretch()

        def copy_commands_only():
            QApplication.clipboard().setText(all_cmds_text)
            status_label.setText("✓ 已复制纯命令到剪贴板，可直接粘贴到 AutoCAD")

        def copy_all_content():
            QApplication.clipboard().setText(text_widget.toPlainText())
            status_label.setText("✓ 已复制全部内容到剪贴板（含注释）")

        btn_copy_all = PushButton("复制全部内容")
        btn_copy_all.clicked.connect(copy_all_content)
        btn_copy_cmd = PrimaryPushButton("复制纯命令")
        btn_copy_cmd.clicked.connect(copy_commands_only)
        btn_close = PushButton("关闭")
        btn_close.clicked.connect(preview.close)
        btn_lay.addWidget(btn_copy_all)
        btn_lay.addWidget(btn_copy_cmd)
        btn_lay.addWidget(btn_close)
        p_lay.addLayout(btn_lay)

        preview.exec()

    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "生成失败",
                     f"建筑物名称上平面图生成过程中发生错误:\n\n{str(e)}")


# ================================================================
# 4. IP坐标及弯道参数表
# ================================================================

def _draw_ip_table_on_msp(msp, ox, oy, preview_data,
                          title="IP坐标及弯道参数表", layer="IP_TABLE"):
    """在 modelspace 上绘制IP坐标及弯道参数表。返回 (width, height)。"""
    import ezdxf

    ROW_H = 6.0
    HDR_ROW_H = 6.0
    TITLE_ROW_H = 7.0
    TEXT_H = 2.2
    HDR_TEXT_H = 2.5
    TITLE_TEXT_H = 3.0
    COL_PAD = 3.0

    sub_headers = _get_ip_table_preview_headers()
    group_headers = _get_ip_table_group_headers()
    v_merged = _get_ip_table_vertical_merged_columns()
    ncols = len(_IP_TABLE_COLUMN_DEFS)
    nrows = len(preview_data)

    _wf = 1.0
    def _tw(text, h):
        if text is None:
            return 0.0
        return sum(h * _wf if ord(c) > 0x7F else h * 0.6 * _wf for c in str(text))

    col_w = [0.0] * ncols
    for ci, hdr in enumerate(sub_headers):
        col_w[ci] = _tw(hdr, HDR_TEXT_H)
    for sc, ec, gtxt in group_headers:
        span = ec - sc + 1
        gw_each = _tw(gtxt, HDR_TEXT_H) / span
        for ci in range(sc, ec + 1):
            col_w[ci] = max(col_w[ci], gw_each)
    for row in preview_data:
        for ci, val in enumerate(row):
            if ci < ncols:
                col_w[ci] = max(col_w[ci], _tw(val, TEXT_H))
    col_w = [w + COL_PAD for w in col_w]

    col_x = [ox]
    for w in col_w:
        col_x.append(col_x[-1] + w)
    total_w = col_x[-1] - col_x[0]
    x_left, x_right = col_x[0], col_x[-1]

    y_title_top = oy
    y_title_bot = y_title_top - TITLE_ROW_H
    y_hdr1_bot = y_title_bot - HDR_ROW_H
    y_hdr2_bot = y_hdr1_bot - HDR_ROW_H
    y_data_top = y_hdr2_bot
    row_y = [y_data_top]
    for _ in range(nrows):
        row_y.append(row_y[-1] - ROW_H)

    dxa = {"layer": layer}

    # === 标题行 ===
    msp.add_line((x_left, y_title_top), (x_right, y_title_top), dxfattribs=dxa)
    msp.add_line((x_left, y_title_bot), (x_right, y_title_bot), dxfattribs=dxa)
    msp.add_line((x_left, y_title_top), (x_left, y_title_bot), dxfattribs=dxa)
    msp.add_line((x_right, y_title_top), (x_right, y_title_bot), dxfattribs=dxa)
    msp.add_text(
        title,
        dxfattribs={"layer": layer, "height": TITLE_TEXT_H,
                    "width": 0.7, "style": "Standard"}
    ).set_placement(
        (x_left + total_w / 2, (y_title_top + y_title_bot) / 2),
        align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER,
    )

    # === 表头区边框 ===
    msp.add_line((x_left, y_hdr2_bot), (x_right, y_hdr2_bot), dxfattribs=dxa)
    msp.add_line((x_left, y_title_bot), (x_left, y_hdr2_bot), dxfattribs=dxa)
    msp.add_line((x_right, y_title_bot), (x_right, y_hdr2_bot), dxfattribs=dxa)

    for ci in range(ncols):
        if ci not in v_merged:
            msp.add_line((col_x[ci], y_hdr1_bot), (col_x[ci + 1], y_hdr1_bot),
                         dxfattribs=dxa)

    drawn_x = set()
    for sc, ec, _ in group_headers:
        for bx in (col_x[sc], col_x[ec + 1]):
            if bx not in drawn_x:
                msp.add_line((bx, y_title_bot), (bx, y_hdr2_bot), dxfattribs=dxa)
                drawn_x.add(bx)
        if sc != ec:
            for ci in range(sc + 1, ec + 1):
                msp.add_line((col_x[ci], y_hdr1_bot), (col_x[ci], y_hdr2_bot),
                             dxfattribs=dxa)

    for sc, ec, text in group_headers:
        cx = (col_x[sc] + col_x[ec + 1]) / 2
        cy = ((y_title_bot + y_hdr2_bot) / 2 if sc in v_merged
              else (y_title_bot + y_hdr1_bot) / 2)
        msp.add_text(
            text,
            dxfattribs={"layer": layer, "height": HDR_TEXT_H,
                        "width": 0.7, "style": "Standard"}
        ).set_placement((cx, cy), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)

    for ci, hdr in enumerate(sub_headers):
        if ci in v_merged:
            continue
        cx = (col_x[ci] + col_x[ci + 1]) / 2
        cy = (y_hdr1_bot + y_hdr2_bot) / 2
        msp.add_text(
            hdr,
            dxfattribs={"layer": layer, "height": HDR_TEXT_H,
                        "width": 0.7, "style": "Standard"}
        ).set_placement((cx, cy), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)

    # === 数据区 ===
    msp.add_line((x_left, y_data_top), (x_right, y_data_top), dxfattribs=dxa)
    if nrows > 0:
        msp.add_line((x_left, row_y[-1]), (x_right, row_y[-1]), dxfattribs=dxa)
    for ri in range(1, nrows):
        msp.add_line((x_left, row_y[ri]), (x_right, row_y[ri]), dxfattribs=dxa)

    y_bottom = row_y[-1] if nrows > 0 else y_data_top
    for x in col_x:
        msp.add_line((x, y_data_top), (x, y_bottom), dxfattribs=dxa)

    for ri, row_vals in enumerate(preview_data):
        for ci, val in enumerate(row_vals):
            if val is None or val == "":
                continue
            cx = (col_x[ci] + col_x[ci + 1]) / 2
            cy = (row_y[ri] + row_y[ri + 1]) / 2
            msp.add_text(
                str(val),
                dxfattribs={"layer": layer, "height": TEXT_H,
                            "width": 0.7, "style": "Standard"}
            ).set_placement((cx, cy), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)

    total_h = y_title_top - (row_y[-1] if nrows > 0 else y_data_top)
    return total_w, total_h


def _write_ip_table_dxf(file_path, preview_data, title="IP坐标及弯道参数表"):
    """将IP坐标及弯道参数表写入独立DXF文件。"""
    import ezdxf
    doc = ezdxf.new("R2010")
    msp = doc.modelspace()
    _setup_profile_dxf_document(doc)
    _draw_ip_table_on_msp(msp, 0.0, 0.0, preview_data, title, "IP_TABLE")
    doc.saveas(file_path)


def export_ip_plan_table(panel):
    """导出IP坐标及弯道参数表DXF/Excel文件（含合并表头、桩号格式化）"""
    if not MODELS_AVAILABLE:
        fluent_info(panel.window(), "不可用", "核心模型未加载")
        return

    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(panel.window(), "警告", "没有数据可导出，请先执行计算")
        return

    try:
        import openpyxl
    except ImportError:
        fluent_info(panel.window(), "缺少依赖",
                    "需要安装 openpyxl: pip install openpyxl")
        return

    try:
        try:
            proj_settings = panel._build_settings()
            station_prefix = proj_settings.get_station_prefix()
        except Exception:
            station_prefix = ""
        preview_settings = _normalize_text_export_settings(getattr(panel, "_text_export_settings", {}))
        preview_data, real_nodes = _compute_ip_preview_data(
            nodes,
            station_prefix,
            preview_settings,
        )
        if not real_nodes:
            fluent_info(panel.window(), "警告", "没有有效的IP点数据可导出")
            return

        preview_headers = _get_ip_table_preview_headers()

        # 预览对话框
        preview_dlg = QDialog(panel.window())
        preview_dlg.setWindowTitle("预览 — IP坐标及弯道参数表")
        preview_dlg.setMinimumSize(950, 450)
        preview_dlg.setStyleSheet(DIALOG_STYLE)
        dlg_lay = QVBoxLayout(preview_dlg)

        dlg_lay.addWidget(QLabel(
            f"共 {len(preview_data)} 条IP点数据，请确认内容后点击「确认导出」保存为DXF或Excel文件。"))

        table = QTableWidget(len(preview_data), len(preview_headers))
        table.setHorizontalHeaderLabels(preview_headers)
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        for r, row_data in enumerate(preview_data):
            for c, val in enumerate(row_data):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignCenter)
                table.setItem(r, c, item)
        auto_resize_table(table)
        dlg_lay.addWidget(table)

        btn_lay = QHBoxLayout()
        btn_lay.addStretch()
        btn_cancel = PushButton("取消"); btn_cancel.clicked.connect(preview_dlg.reject)
        btn_ok = PrimaryPushButton("确认导出"); btn_ok.clicked.connect(preview_dlg.accept)
        btn_lay.addWidget(btn_cancel); btn_lay.addWidget(btn_ok)
        dlg_lay.addLayout(btn_lay)

        # 绑定 ESC 关闭 / Enter 确认
        QShortcut(QKeySequence(Qt.Key_Escape), preview_dlg, preview_dlg.reject)
        QShortcut(QKeySequence(Qt.Key_Return), preview_dlg, preview_dlg.accept)

        if preview_dlg.exec() != QDialog.Accepted:
            return

        # 保存
        try:
            ch_name = panel.channel_name_edit.text().strip()
            ch_level = panel.channel_level_combo.currentText()
            auto_name = f"{ch_name}{ch_level}_IP坐标及弯道参数表.dxf"
        except Exception:
            auto_name = "IP坐标及弯道参数表.dxf"

        file_path, _ = QFileDialog.getSaveFileName(
            panel, "保存IP坐标及弯道参数表", auto_name,
            "DXF文件 (*.dxf);;Excel文件 (*.xlsx);;所有文件 (*.*)")
        if not file_path:
            return

        # DXF 导出（紧凑排版、自适应列宽、无底色）
        if file_path.lower().endswith('.dxf'):
            _write_ip_table_dxf(file_path, preview_data)
            if fluent_question(panel.window(), "导出完成",
                    f"IP坐标及弯道参数表DXF导出成功！\n\n"
                    f"文件保存路径:\n{file_path}\n\n"
                    f"导出IP点数量: {len(real_nodes)}\n\n"
                    f"是否要立即打开该文件？"):
                try:
                    os.startfile(file_path)
                except Exception:
                    pass
            return

        # openpyxl 写入
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IP点上平面图"
        _write_ip_table_excel_sheet(ws, real_nodes, station_prefix, preview_settings)

        wb.save(file_path)
        wb.close()

        if fluent_question(panel.window(), "导出完成",
                f"IP坐标及弯道参数表导出成功！\n\n"
                f"文件保存路径:\n{file_path}\n\n"
                f"导出IP点数量: {len(real_nodes)}\n\n"
                f"是否要立即打开该文件？"):
            try:
                os.startfile(file_path)
            except Exception:
                pass

    except PermissionError:
        fluent_error(panel.window(), "文件被占用",
                     f"无法写入文件，该文件可能已被其他程序打开：\n\n{file_path}\n\n"
                     f"请先关闭该文件（如Excel等），然后重新操作。")
    except Exception as e:
        import traceback
        traceback.print_exc()
        fluent_error(panel.window(), "导出失败",
                     f"IP坐标及弯道参数表导出失败，请检查以下可能的原因：\n\n"
                     f"1. 目标文件是否被其他程序占用（如AutoCAD、Excel等）\n"
                     f"2. 文件保存路径是否有写入权限\n"
                     f"3. 数据是否完整（坐标、桩号等）\n\n"
                     f"错误信息：{str(e)}\n\n"
                     f"如仍无法解决，请将以上信息反馈给技术支持。")


# ================================================================
# 6. 合并导出全部DXF（横向分区布局）
# ================================================================

def _draw_section_summary_on_msp(
    panel,
    msp,
    nodes,
    proj_settings,
    pressurized_params,
    below_y,
    summary_layer,
):
    """Draw section summary tables onto modelspace.

    Returns:
        tuple[float, float, int]: (summary_width, summary_height, drawn_table_count)
    """
    from calc_渠系计算算法内核.生成断面汇总表 import (
        _extract_segment_defaults_from_nodes,
        _segment_name,
        _dxf_draw_table,
        _dxf_auto_col_widths,
        _DXF_TABLE_GAP,
        _DXF_BUILDERS,
        _build_horseshoe_export_entries,
        _expand_horseshoe_table_order,
        compute_rect_channel,
        compute_trapezoid_channel,
        compute_u_channel,
        compute_tunnel,
        compute_tunnel_circular,
        compute_aqueduct_u,
        compute_aqueduct_rect,
        compute_rect_culvert,
        compute_circular_pipe,
        compute_siphon,
        compute_pressure_pipe,
        _default_segments_rect_channel,
        _default_segments_trap_channel,
        _default_segments_u_channel,
        _default_segments_tunnel_arch,
        _default_segments_tunnel_circular,
        _default_segments_tunnel_horseshoe,
        _default_segments_aqueduct_u,
        _default_segments_aqueduct_rect,
        _default_segments_rect_culvert,
        _default_segments_circular_pipe,
    )

    node_defaults, flow_qs = _extract_segment_defaults_from_nodes(nodes)

    counts = []
    if proj_settings and getattr(proj_settings, "design_flows", None):
        flows = [q for q in proj_settings.design_flows if isinstance(q, (int, float)) and q > 0]
        if flows:
            counts.append(len(flows))
    if flow_qs:
        counts.append(max(flow_qs.keys()))
    if node_defaults:
        for data in node_defaults.values():
            if data:
                counts.append(max(data.keys()))
    seg_count = max(1, max(counts)) if counts else 7

    fallback_qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    if flow_qs:
        qs = []
        for i in range(1, seg_count + 1):
            q = flow_qs.get(i, 0.0)
            qs.append(q if q > 0 else (fallback_qs[i - 1] if i - 1 < len(fallback_qs) else fallback_qs[-1]))
    elif proj_settings and getattr(proj_settings, "design_flows", None):
        flows = [q for q in proj_settings.design_flows if isinstance(q, (int, float)) and q > 0]
        if flows:
            qs = [flows[i] if i < len(flows) else flows[-1] for i in range(seg_count)]
        else:
            qs = fallback_qs[:seg_count]
    else:
        qs = list(fallback_qs[:seg_count]) + [fallback_qs[-1]] * max(0, seg_count - len(fallback_qs))

    has_source = bool(nodes) and any(node_defaults.values())

    def _make_segs(default_fn, overrides_by_idx=None):
        if has_source and overrides_by_idx is not None:
            if not overrides_by_idx:
                return []
            pool = default_fn()
            segs = []
            for idx in sorted(overrides_by_idx.keys()):
                base = dict(pool[0]) if pool else {}
                base["name"] = _segment_name(idx)
                base.update(overrides_by_idx[idx])
                if 0 < idx <= len(qs):
                    base["Q"] = qs[idx - 1]
                segs.append(base)
            return segs

        segs = default_fn()
        if len(segs) < seg_count:
            last = segs[-1] if segs else {}
            for idx in range(len(segs) + 1, seg_count + 1):
                new = dict(last)
                new["name"] = _segment_name(idx)
                segs.append(new)
        segs = segs[:seg_count]
        for i, seg in enumerate(segs):
            if overrides_by_idx and (i + 1) in overrides_by_idx:
                seg.update(overrides_by_idx[i + 1])
            if i < len(qs):
                seg["Q"] = qs[i]
        return segs

    rc = _make_segs(_default_segments_rect_channel, node_defaults.get("rect_channel"))
    tr = _make_segs(_default_segments_trap_channel, node_defaults.get("trap_channel"))
    uc = _make_segs(_default_segments_u_channel, node_defaults.get("u_channel"))
    ta = _make_segs(_default_segments_tunnel_arch, node_defaults.get("tunnel_arch"))
    tc = _make_segs(_default_segments_tunnel_circular, node_defaults.get("tunnel_circular"))
    th = _make_segs(_default_segments_tunnel_horseshoe, node_defaults.get("tunnel_horseshoe"))
    au = _make_segs(_default_segments_aqueduct_u, node_defaults.get("aqueduct_u"))
    ar = _make_segs(_default_segments_aqueduct_rect, node_defaults.get("aqueduct_rect"))
    rv = _make_segs(_default_segments_rect_culvert, node_defaults.get("rect_culvert"))
    cp = _make_segs(_default_segments_circular_pipe, node_defaults.get("circular_channel"))
    pressure_pipe_params = _prepare_pressure_pipe_export_rows(
        pressurized_params.get("pressure_pipe", []),
        panel=panel,
        calc_contexts=_extract_pressure_pipe_calc_contexts(nodes, proj_settings),
    )
    pressure_pipe_export_rows = _merge_pressure_pipe_export_rows_by_flow_section(
        pressure_pipe_params,
        panel=panel,
    )

    sp = _build_pressurized_segments(
        qs=qs,
        overrides_by_idx=node_defaults.get("siphon", {}),
        params=pressurized_params.get("siphon", []),
        has_source_data=has_source,
        segment_name_fn=_segment_name,
    )
    pp = _build_pressurized_segments(
        qs=qs,
        overrides_by_idx=node_defaults.get("pressure_pipe", {}),
        params=pressure_pipe_export_rows,
        has_source_data=has_source,
        segment_name_fn=_segment_name,
    )

    _struct_t = getattr(panel, "_custom_struct_thickness", None)
    _rock_lining = getattr(panel, "_custom_rock_lining", None)
    if _struct_t:
        _st_rc = _struct_t.get("rect_channel", {})
        for seg in rc:
            if "wall_t" in _st_rc:
                seg["wall_t"] = _st_rc["wall_t"]
            if "tie_rod" in _st_rc:
                seg["tie_rod"] = _st_rc["tie_rod"]
        _st_tr = _struct_t.get("trap_channel", {})
        for seg in tr:
            if "wall_t" in _st_tr:
                seg["wall_t"] = _st_tr["wall_t"]
            if "tie_rod" in _st_tr:
                seg["tie_rod"] = _st_tr["tie_rod"]
        _st_au = _struct_t.get("aqueduct_u", {})
        for seg in au:
            if "wall_t" in _st_au:
                seg["wall_t"] = _st_au["wall_t"]
        _st_ar = _struct_t.get("aqueduct_rect", {})
        for seg in ar:
            if "wall_t" in _st_ar:
                seg["wall_t"] = _st_ar["wall_t"]
        _st_rv = _struct_t.get("rect_culvert", {})
        for seg in rv:
            for key in ("t0", "t1", "t2"):
                if key in _st_rv:
                    seg[key] = _st_rv[key]

    _tu = getattr(panel, "_custom_tunnel_unified", {})
    _tu_arch = _tu.get("tunnel_arch", False)
    _tu_circ = _tu.get("tunnel_circular", False)
    _tu_horse = _tu.get("tunnel_horseshoe", False)
    if has_source:
        _tu_arch = False
        _tu_circ = False
        _tu_horse = False

    d_rc = compute_rect_channel(rc) if rc else []
    d_tr = compute_trapezoid_channel(tr) if tr else []
    d_uc = compute_u_channel(uc) if uc else []
    d_ta, _ = compute_tunnel(ta, _rock_lining, unified=_tu_arch) if ta else ([], {})
    d_tc, _ = compute_tunnel_circular(tc, _rock_lining, unified=_tu_circ) if tc else ([], {})
    horseshoe_entries = _build_horseshoe_export_entries(
        th,
        rock_lining=_rock_lining,
        unified=_tu_horse,
    ) if th else []
    d_au = compute_aqueduct_u(au) if au else []
    d_ar = compute_aqueduct_rect(ar) if ar else []
    d_rv = compute_rect_culvert(rv) if rv else []
    d_cp = compute_circular_pipe(cp) if cp else []
    d_sp = compute_siphon(sp) if sp else []
    d_pp = compute_pressure_pipe(pp) if pp else []

    data_map = {
        "rect_channel": d_rc,
        "trap_channel": d_tr,
        "u_channel": d_uc,
        "tunnel_arch": d_ta,
        "tunnel_circular": d_tc,
        "aqueduct_u": d_au,
        "aqueduct_rect": d_ar,
        "rect_culvert": d_rv,
        "circular_channel": d_cp,
        "siphon": d_sp,
        "pressure_pipe": d_pp,
    }
    horseshoe_titles = {}
    horseshoe_keys = []
    for entry in horseshoe_entries:
        data_map[entry["key"]] = entry["rows"]
        horseshoe_titles[entry["key"]] = entry["title"]
        horseshoe_keys.append(entry["key"])
    table_order = [
        "rect_channel",
        "trap_channel",
        "u_channel",
        "tunnel_arch",
        "tunnel_circular",
        "tunnel_horseshoe",
        "aqueduct_u",
        "aqueduct_rect",
        "rect_culvert",
        "circular_channel",
        "siphon",
        "pressure_pipe",
    ]
    table_order = _expand_horseshoe_table_order(table_order, horseshoe_keys)

    cur_y = below_y
    max_table_w = 0.0
    summary_h = 0.0
    drawn_table_count = 0
    for key in table_order:
        data_rows = data_map.get(key)
        builder = _DXF_BUILDERS.get(key)
        if data_rows and builder:
            title_arg = horseshoe_titles.get(key)
            title_t, headers, col_widths, rows, merge = (
                builder(data_rows, title=title_arg)
                if title_arg is not None
                else builder(data_rows)
            )
            table_h = _dxf_draw_table(
                msp,
                0.0,
                cur_y,
                title_t,
                headers,
                col_widths,
                rows,
                merge_groups=merge,
                layer=summary_layer,
            )
            auto_w = _dxf_auto_col_widths(headers, rows)
            col_count = len(headers)
            actual_w = sum(
                max(col_widths[col_idx], auto_w[col_idx]) if col_idx < len(col_widths) else auto_w[col_idx]
                for col_idx in range(col_count)
            )
            max_table_w = max(max_table_w, actual_w)
            summary_h += table_h + _DXF_TABLE_GAP
            cur_y -= (table_h + _DXF_TABLE_GAP)
            drawn_table_count += 1

    return max_table_w, summary_h, drawn_table_count


def _resolve_section_summary_source_nodes(panel, fallback_nodes=None):
    """返回断面汇总导出的节点列表及其来源描述。"""
    build_nodes = getattr(panel, "_build_nodes_from_table", None)
    if callable(build_nodes):
        try:
            current_nodes = build_nodes()
        except Exception:
            current_nodes = None
        if current_nodes:
            return current_nodes, "current_table_snapshot"

    if fallback_nodes:
        return fallback_nodes, "fallback_nodes"

    return list(getattr(panel, "calculated_nodes", None) or []), "calculated_nodes"


def _record_section_summary_runtime_debug(panel, nodes, nodes_source):
    """记录断面汇总运行态来源，便于排查源码与导出链路是否一致。"""
    summary_module_file = ""
    try:
        from calc_渠系计算算法内核 import 生成断面汇总表 as summary_module

        summary_module_file = os.path.abspath(getattr(summary_module, "__file__", "") or "")
    except Exception:
        summary_module_file = ""

    try:
        panel._last_section_summary_runtime_debug = {
            "summary_nodes_source": str(nodes_source or ""),
            "summary_node_count": len(nodes or []),
            "summary_module_file": summary_module_file,
        }
    except Exception:
        pass


def _get_section_summary_source_nodes(panel, fallback_nodes=None):
    """断面汇总导出优先使用当前表3快照，避免吃到过期 calculated_nodes。"""
    nodes, _source = _resolve_section_summary_source_nodes(panel, fallback_nodes=fallback_nodes)
    return nodes


def export_combined_dxf(panel):
    """将纵断面表格、断面汇总表、IP坐标表合并导出到一个DXF文件。

    布局：纵断面表格在上方（全宽），下方左侧放断面汇总表，右侧放IP表。
    """
    import ezdxf
    parent_window = _safe_qt_parent(panel)

    if not MODELS_AVAILABLE:
        fluent_info(parent_window, "不可用", "核心模型未加载")
        return

    export_mode = "xxpipe" if _is_panel_xxpipe_mode(panel) else None
    nodes = panel.calculated_nodes
    if not nodes:
        fluent_info(parent_window, "警告", "没有数据可导出，请先执行计算")
        return

    if export_mode == "xxpipe":
        valid_nodes = list(nodes)
    else:
        valid_nodes = [n for n in nodes if n.bottom_elevation or n.top_elevation or n.water_level]
        if not valid_nodes:
            fluent_info(parent_window, "警告", "没有可用的高程数据，请先执行计算。")
            return

        optional_blank_name_notice = _build_optional_blank_name_notice(nodes, action_name="导出")
        if optional_blank_name_notice:
            fluent_info(parent_window, "提示", optional_blank_name_notice)

    # ---- 1. 纵断面参数设置 ----
    dlg = TextExportSettingsDialog(
        parent_window,
        panel._text_export_settings,
        mode=export_mode or "standard",
    )
    if dlg.exec() != QDialog.Accepted or dlg.result is None:
        return
    panel._text_export_settings.update(dlg.result)
    profile_settings = dlg.result

    # ---- 2. 获取项目设置 ----
    try:
        proj_settings = panel._build_settings()
        station_prefix = proj_settings.get_station_prefix()
    except Exception:
        proj_settings = None
        station_prefix = ""
    summary_nodes, summary_nodes_source = _resolve_section_summary_source_nodes(panel, fallback_nodes=nodes)
    _record_section_summary_runtime_debug(panel, summary_nodes, summary_nodes_source)
    profile_nodes = summary_nodes if export_mode == "xxpipe" else nodes
    profile_valid_nodes = profile_nodes if export_mode == "xxpipe" else valid_nodes
    ip_source_nodes = summary_nodes if export_mode == "xxpipe" else nodes
    xxpipe_profile_data = None
    if export_mode == "xxpipe":
        if not profile_nodes:
            fluent_info(parent_window, "警告", "没有可用于 xx管 纵断面导出的节点数据。")
            return
        try:
            xxpipe_profile_data = _build_panel_xxpipe_profile_data(panel, profile_nodes, station_prefix=station_prefix)
        except Exception as exc:
            friendly_message = _translate_xxpipe_export_error(exc)
            if friendly_message:
                fluent_error(parent_window, "导出失败", friendly_message)
                return
            raise

    # ---- 3. 断面汇总表参数设置（构造参数、有压流参数等）----
    try:
        ch_name_for_dlg = panel.channel_name_edit.text().strip()
        ch_level_for_dlg = panel.channel_level_combo.currentText()
        auto_name_dlg = f"{ch_name_for_dlg}{ch_level_for_dlg}_断面汇总表.xlsx"
    except Exception:
        auto_name_dlg = "断面汇总表.xlsx"
    if parent_window is not None:
        summary_dlg = SectionSummaryDialog(
            parent_window, summary_nodes, proj_settings, auto_name_dlg,
            panel=panel, config_only=True,
        )
        if summary_dlg.exec() != QDialog.Accepted:
            return
    cached_pressurized = getattr(panel, "_custom_pressurized_pipe_params", {}) or {}
    pressurized_params = {
        "siphon": cached_pressurized.get("siphon", []),
        "pressure_pipe": cached_pressurized.get("pressure_pipe", []),
    }

    # ---- 4. 选择保存路径 ----
    try:
        ch_name = panel.channel_name_edit.text().strip()
        ch_level = panel.channel_level_combo.currentText()
        auto_name = f"{ch_name}{ch_level}_全部表格.dxf"
    except Exception:
        auto_name = "全部表格.dxf"

    file_path, _ = QFileDialog.getSaveFileName(
        parent_window, "保存合并DXF（纵断面+断面汇总+IP表）", auto_name,
        "DXF 文件 (*.dxf);;所有文件 (*.*)")
    if not file_path:
        return
    if not file_path.lower().endswith('.dxf'):
        file_path += '.dxf'

    try:
        # ---- 创建 DXF 文档 ----
        doc = ezdxf.new("R2010")
        msp = doc.modelspace()
        _setup_profile_dxf_document(doc)

        # 三个组件使用独立图层（带前缀），便于在CAD中分别控制显示
        _PROF_PREFIX = "纵断面_"
        _SUMM_LAYER = "断面汇总表"
        _IP_LAYER = "IP坐标表"

        _ensure_profile_layers(doc, layer_prefix=_PROF_PREFIX)
        if _SUMM_LAYER not in doc.layers:
            doc.layers.new(_SUMM_LAYER, dxfattribs={"color": 7})   # 白色
        if _IP_LAYER not in doc.layers:
            doc.layers.new(_IP_LAYER, dxfattribs={"color": 7})     # 白色

        GAP = 20.0  # 各区域间距

        # ======== A. 纵断面表格（顶部，原点(0,0)） ========
        try:
            prof_w, prof_h = _draw_profile_on_msp(
                msp,
                profile_nodes,
                profile_valid_nodes,
                profile_settings,
                station_prefix,
                layer_prefix=_PROF_PREFIX,
                export_mode=export_mode,
                xxpipe_profile_data=xxpipe_profile_data,
            )
        except Exception as exc:
            if export_mode == "xxpipe":
                friendly_message = _translate_xxpipe_export_error(exc)
                if friendly_message:
                    fluent_error(parent_window, "导出失败", friendly_message)
                    return
            raise

        # 下方区域起始Y（纵断面底部再向下留间距）
        below_y = -GAP

        try:
            summary_w, summary_h, drawn_table_count = _draw_section_summary_on_msp(
                panel=panel,
                msp=msp,
                nodes=summary_nodes,
                proj_settings=proj_settings,
                pressurized_params=pressurized_params,
                below_y=below_y,
                summary_layer=_SUMM_LAYER,
            )
        except Exception as e:
            import traceback; traceback.print_exc()
            fluent_error(
                parent_window,
                "导出失败",
                f"断面汇总表生成失败，已取消“导出全部DXF”。\n{e}",
            )
            return

        if drawn_table_count <= 0:
            fluent_error(
                parent_window,
                "导出失败",
                "断面汇总表无可导出内容，已取消“导出全部DXF”。",
            )
            return

        try:
            ip_preview_settings = _normalize_text_export_settings(getattr(panel, "_text_export_settings", {}))
            ip_preview, ip_nodes = _compute_ip_preview_data(
                ip_source_nodes,
                station_prefix,
                ip_preview_settings,
            )
        except Exception as e:
            import traceback; traceback.print_exc()
            fluent_error(
                parent_window,
                "导出失败",
                f"IP坐标及弯道参数表生成失败，已取消“导出全部DXF”。\n{e}",
            )
            return

        if not ip_preview:
            fluent_error(
                parent_window,
                "导出失败",
                "IP坐标及弯道参数表无可导出内容，已取消“导出全部DXF”。",
            )
            return

        try:
            ip_ox = max(summary_w + GAP, 200.0)
            ip_oy = below_y
            _draw_ip_table_on_msp(
                msp,
                ip_ox,
                ip_oy,
                ip_preview,
                "IP坐标及弯道参数表",
                _IP_LAYER,
            )
        except Exception as e:
            import traceback; traceback.print_exc()
            fluent_error(
                parent_window,
                "导出失败",
                f"IP坐标及弯道参数表绘制失败，已取消“导出全部DXF”。\n{e}",
            )
            return

        doc.saveas(file_path)
        if fluent_question(
            parent_window,
            "导出完成",
            f"合并DXF已生成：\n{file_path}\n\n"
            f"包含：纵断面表格 + 断面汇总表 + IP坐标表\n"
            f"断面汇总表: {drawn_table_count} 张可用表格，IP表: {len(ip_preview)} 行数据\n"
            f"是否立即打开该文件？",
        ):
            try:
                os.startfile(file_path)
            except Exception:
                pass
        return

    except PermissionError:
        fluent_error(parent_window, "文件被占用",
                     f"无法写入文件：\n{file_path}\n\n请先关闭该文件后重试。")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(parent_window, "导出失败",
                     f"合并DXF导出失败:\n{str(e)}")


# ================================================================
# 5. 断面汇总表
# ================================================================

class _MultiLineElidedLabel(QLabel):
    """Wrap text and elide only the last visible line."""

    def __init__(self, text="", tooltip_text="", max_lines=3, parent=None):
        super().__init__(parent)
        self._full_text = ""
        self._tooltip_text = ""
        self._max_lines = max(1, int(max_lines or 1))
        self.setWordWrap(True)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.set_full_text(text, tooltip_text=tooltip_text)

    def hasHeightForWidth(self):
        return True

    def heightForWidth(self, width):
        return self._height_for_width(width)

    def sizeHint(self):
        hint = super().sizeHint()
        hint.setHeight(self._height_for_width(self._available_width()))
        return hint

    def minimumSizeHint(self):
        hint = super().minimumSizeHint()
        hint.setHeight(self._height_for_width(self._available_width()))
        return hint

    def setStyleSheet(self, style_sheet):
        super().setStyleSheet(style_sheet)
        self._sync_height_constraints()

    def set_full_text(self, text, tooltip_text=""):
        self._full_text = str(text or "")
        self._tooltip_text = str(tooltip_text or "")
        self.setToolTip(self._tooltip_text)
        self._refresh_visible_text()
        self._sync_height_constraints()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._refresh_visible_text()
        self._sync_height_constraints()

    def event(self, event):
        if event.type() in (QEvent.FontChange, QEvent.LayoutRequest, QEvent.StyleChange):
            self._refresh_visible_text()
            self._sync_height_constraints()
        return super().event(event)

    def _line_height(self):
        self.ensurePolished()
        return max(1, self.fontMetrics().lineSpacing())

    def _available_width(self):
        width = self.contentsRect().width()
        if width > 0:
            return width
        width = self.width()
        if width > 0:
            return width
        return self.fontMetrics().horizontalAdvance(self._full_text or " ")

    def _wrapped_lines(self, width):
        text = self._full_text
        if not text:
            return [""]
        metrics = self.fontMetrics()
        lines = []
        for paragraph in text.splitlines() or [""]:
            if not paragraph:
                lines.append("")
                continue
            current = ""
            tokens = re.findall(r"[^,，]+(?:[,，]\s*)?|[,，]\s*", paragraph) or [paragraph]
            for token in tokens:
                token = token if current else token.lstrip()
                candidate = f"{current}{token}"
                if current and metrics.horizontalAdvance(candidate) <= width:
                    current = candidate
                    continue

                if current:
                    lines.append(current.rstrip())
                    current = ""
                    token = token.lstrip()

                if metrics.horizontalAdvance(token) <= width:
                    current = token
                    continue

                for ch in token:
                    candidate = f"{current}{ch}"
                    if current and metrics.horizontalAdvance(candidate) > width:
                        lines.append(current.rstrip())
                        current = ch.lstrip()
                    else:
                        current = candidate
            if current:
                lines.append(current.rstrip())
        return lines or [text]

    def _refresh_visible_text(self):
        text = self._full_text
        if not text:
            super().setText("")
            self.updateGeometry()
            return

        width = max(1, self._available_width())
        wrapped = self._wrapped_lines(width)
        if len(wrapped) <= self._max_lines:
            visible_text = "\n".join(wrapped)
        else:
            prefix_lines = wrapped[: self._max_lines - 1]
            last_line_source = "".join(wrapped[self._max_lines - 1 :])
            last_line = self.fontMetrics().elidedText(last_line_source, Qt.ElideRight, width)
            visible_text = "\n".join(prefix_lines + [last_line])

        if visible_text != self.text():
            super().setText(visible_text)
            self.updateGeometry()

    def _visible_line_count_for_width(self, width):
        wrapped = self._wrapped_lines(max(1, width))
        return max(1, min(self._max_lines, len(wrapped)))

    def _height_for_width(self, width):
        visible_lines = self._visible_line_count_for_width(width)
        margins = self.contentsMargins()
        return visible_lines * self._line_height() + margins.top() + margins.bottom()

    def _sync_height_constraints(self):
        target_height = self._height_for_width(self._available_width())
        if self.minimumHeight() != target_height:
            self.setMinimumHeight(target_height)
        if self.maximumHeight() != target_height:
            self.setMaximumHeight(target_height)


class SectionSummaryDialog(QDialog):
    """断面尺寸及水力要素汇总表生成对话框（纯 PySide6 版）"""

    def __init__(self, parent, nodes, proj_settings, auto_name="", panel=None,
                 config_only=False):
        super().__init__(parent)
        self._config_only = config_only
        if config_only:
            self.setWindowTitle("断面尺寸及水力要素汇总表 — 参数设置")
        else:
            self.setWindowTitle("断面尺寸及水力要素汇总表 — 生成器")
        self.setMinimumSize(520, 560)
        self.resize(640, 780)
        self.setStyleSheet(DIALOG_STYLE)

        self._nodes = nodes
        self._proj_settings = proj_settings
        self._auto_name = auto_name
        self._panel = panel
        self._cached_pressurized = {}
        if panel is not None:
            self._cached_pressurized = getattr(panel, "_custom_pressurized_pipe_params", {}) or {}

        # 导入计算模块
        from calc_渠系计算算法内核.生成断面汇总表 import (
            _extract_segment_defaults_from_nodes,
            _segment_name,
            PRESSURE_PIPE_MATERIALS,
            SIPHON_MATERIALS,
            ROCK_CLASSES,
            ROCK_LINING_DEFAULT,
            normalize_pressure_pipe_material_key,
            get_pressure_pipe_material_display_name,
            compute_pressure_pipe,
            generate_excel,
            generate_dxf,
            _default_segments_rect_channel,
            _default_segments_trap_channel,
            _default_segments_u_channel,
            _default_segments_tunnel_arch,
            _default_segments_tunnel_circular,
            _default_segments_tunnel_horseshoe,
            _default_segments_aqueduct_u,
            _default_segments_aqueduct_rect,
            _default_segments_rect_culvert,
            _default_segments_circular_pipe,
        )
        self._ROCK_CLASSES = ROCK_CLASSES
        self._ROCK_LINING_DEFAULT = ROCK_LINING_DEFAULT
        self._generate_excel = generate_excel
        self._generate_dxf = generate_dxf
        self._segment_name = _segment_name
        self._SIPHON_MATERIALS = SIPHON_MATERIALS
        self._PRESSURE_PIPE_MATERIALS = PRESSURE_PIPE_MATERIALS
        self._normalize_pressure_pipe_material_key = normalize_pressure_pipe_material_key
        self._get_pressure_pipe_material_display_name = get_pressure_pipe_material_display_name
        self._compute_pressure_pipe = compute_pressure_pipe
        self._default_fns = {
            'rect_channel': _default_segments_rect_channel,
            'trap_channel': _default_segments_trap_channel,
            'u_channel': _default_segments_u_channel,
            'tunnel_arch': _default_segments_tunnel_arch,
            'tunnel_circular': _default_segments_tunnel_circular,
            'tunnel_horseshoe': _default_segments_tunnel_horseshoe,
            'aqueduct_u': _default_segments_aqueduct_u,
            'aqueduct_rect': _default_segments_aqueduct_rect,
            'rect_culvert': _default_segments_rect_culvert,
            'circular_pipe': _default_segments_circular_pipe,
        }
        self._ui_name_column_min_width = 220
        self._ui_material_column_width = 200
        self._ui_dn_column_width = 108
        self._ui_q_value_column_width = 108
        self._ui_numeric_column_width = 96
        self._ui_grid_spacing = 8
        self._ui_group_margins = (14, 12, 14, 12)

        node_defaults, flow_qs = _extract_segment_defaults_from_nodes(nodes)
        self._node_defaults = node_defaults
        self._flow_qs = flow_qs
        self._has_source_data = bool(self._nodes) and any(self._node_defaults.values())

        # 提取实际的有压流实体，并记录缺失流量段的条目以便提示后跳过
        self._siphon_groups, self._invalid_siphon_groups = self._extract_siphon_groups()
        self._pressure_pipe_groups, self._invalid_pressure_pipe_groups = self._extract_pressure_pipe_groups()
        self._pressure_pipe_calc_contexts = _extract_pressure_pipe_calc_contexts(self._nodes, self._proj_settings)
        self._invalid_pressurized_notice_shown = False

        # 确定流量段数
        self._segment_count = self._get_segment_count()
        default_qs = self._build_default_qs()

        self._build_ui(default_qs)

    # ---- 流量段数计算 ----
    def _get_segment_count(self):
        counts = []
        ps = self._proj_settings
        if ps is not None and getattr(ps, "design_flows", None):
            flows = [q for q in ps.design_flows if isinstance(q, (int, float)) and q > 0]
            if flows:
                counts.append(len(flows))
        if self._flow_qs:
            counts.append(max(self._flow_qs.keys()))
        if self._node_defaults:
            for data in self._node_defaults.values():
                if data:
                    counts.append(max(data.keys()))
        return max(1, max(counts)) if counts else 7

    def _build_default_qs(self):
        fallback = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
        sc = self._segment_count
        ps = self._proj_settings
        if self._flow_qs:
            out = []
            for i in range(1, sc + 1):
                q = self._flow_qs.get(i, 0.0)
                out.append(q if q > 0 else (fallback[i - 1] if i - 1 < len(fallback) else fallback[-1]))
            return out
        if ps is not None and getattr(ps, "design_flows", None):
            flows = [q for q in ps.design_flows if isinstance(q, (int, float)) and q > 0]
            if flows:
                return [flows[i] if i < len(flows) else flows[-1] for i in range(sc)]
        if sc <= len(fallback):
            return list(fallback[:sc])
        return list(fallback) + [fallback[-1]] * (sc - len(fallback))

    def _extract_siphon_groups(self):
        return _extract_pressurized_param_entities(self._nodes, "siphon")
    
    def _extract_pressure_pipe_groups(self):
        fallback_groups, invalid_groups = _extract_pressurized_param_entities(
            self._nodes,
            "pressure_pipe",
        )
        panel = getattr(self, "_panel", None)
        extractor = getattr(panel, "_extract_pressure_pipe_dialog_groups", None)
        if not callable(extractor):
            return fallback_groups, invalid_groups

        try:
            dialog_groups = extractor(
                self._nodes,
                settings=getattr(self, "_proj_settings", None),
            )
        except TypeError:
            try:
                dialog_groups = extractor(self._nodes)
            except Exception:
                return fallback_groups, invalid_groups
        except Exception:
            return fallback_groups, invalid_groups

        converted_groups = []
        for group in dialog_groups or []:
            row = _build_pressure_pipe_param_row_from_dialog_group(group)
            if isinstance(row, dict):
                converted_groups.append(row)
        return (converted_groups or fallback_groups), invalid_groups

    def _build_q_segment_structure_names(self):
        segment_names = {idx: [] for idx in range(1, self._segment_count + 1)}
        seen = {idx: set() for idx in range(1, self._segment_count + 1)}
        for node in self._nodes or []:
            if getattr(node, "is_transition", False) or getattr(node, "is_auto_inserted_channel", False):
                continue
            st_str = _struct_val(getattr(node, "structure_type", None))
            if getattr(node, "is_inverted_siphon", False) or ("倒虹吸" in st_str):
                structure_kind = "siphon"
            elif "有压管道" in st_str or getattr(node, "is_pressure_pipe", False):
                structure_kind = "pressure_pipe"
            else:
                continue

            flow_section_idx = _parse_flow_section_index(getattr(node, "flow_section", ""))
            if flow_section_idx is None or flow_section_idx not in segment_names:
                continue

            base_name = _normalize_pressurized_name(getattr(node, "name", ""), structure_kind)
            entity_key = (base_name, structure_kind)
            if entity_key in seen[flow_section_idx]:
                continue
            seen[flow_section_idx].add(entity_key)
            segment_names[flow_section_idx].append({
                "base_name": base_name,
                "structure_kind": structure_kind,
            })

        formatted_names = {}
        for flow_section_idx, items in segment_names.items():
            name_counts = {}
            for item in items:
                base_name = item["base_name"]
                name_counts[base_name] = name_counts.get(base_name, 0) + 1

            formatted_names[flow_section_idx] = []
            for item in items:
                base_name = item["base_name"]
                if name_counts.get(base_name, 0) > 1:
                    suffix = "倒虹吸" if item["structure_kind"] == "siphon" else "有压管道"
                    formatted_names[flow_section_idx].append(f"{base_name}（{suffix}）")
                else:
                    formatted_names[flow_section_idx].append(base_name)
        return formatted_names

    def _build_q_segment_label(self, flow_section_idx):
        base_label = self._segment_name(flow_section_idx)
        names = self._q_segment_structure_names.get(flow_section_idx, [])
        if not names:
            return base_label, ""
        if len(names) <= 2:
            summary = "、".join(names)
        else:
            summary = f"{names[0]}、{names[1]}等{len(names)}项"
        return f"{base_label}（{summary}）", "、".join(names)

    def _pressure_pipe_material_key(self, material_value):
        normalizer = getattr(self, "_normalize_pressure_pipe_material_key", None)
        if callable(normalizer):
            return normalizer(material_value)
        return str(material_value or "").strip() or "球墨铸铁管"

    def _pressure_pipe_material_display_name(self, material_value):
        display_getter = getattr(self, "_get_pressure_pipe_material_display_name", None)
        if callable(display_getter):
            return display_getter(material_value)
        return str(material_value or "").strip() or "球墨铸铁管"

    def _styled_form_header_label(self, text):
        lbl = QLabel(text)
        lbl.setStyleSheet("font-size:11px; color:#555; font-weight:bold;")
        lbl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        return lbl

    def _styled_name_value_label(self, text, tooltip_text="", max_lines=3):
        lbl = _MultiLineElidedLabel(text, tooltip_text=tooltip_text, max_lines=max_lines)
        lbl.setStyleSheet("font-size:12px;")
        return lbl

    def _build_q_segment_label(self, flow_section_idx):
        base_label = self._segment_name(flow_section_idx)
        names = self._q_segment_structure_names.get(flow_section_idx, [])
        if not names:
            return base_label, ""
        summary = ", ".join(names)
        return f"{base_label}（{summary}）", summary

    def _styled_name_value_label(self, text, tooltip_text="", max_lines=None):
        if max_lines is not None:
            lbl = _MultiLineElidedLabel(text, tooltip_text=tooltip_text, max_lines=max_lines)
        else:
            lbl = QLabel(text)
            if tooltip_text:
                lbl.setToolTip(tooltip_text)
            lbl.setWordWrap(True)
            lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            lbl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        lbl.setStyleSheet("font-size:12px;")
        return lbl

    def _styled_q_segment_title_label(self, text):
        lbl = QLabel(text)
        lbl.setStyleSheet("font-size:12px;")
        lbl.setWordWrap(False)
        lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        lbl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        return lbl

    def _build_q_segment_label(self, flow_section_idx):
        segment_title = self._segment_name(flow_section_idx)
        names = self._q_segment_structure_names.get(flow_section_idx, [])
        names_text = ", ".join(names)
        return {
            "segment_title": segment_title,
            "names_text": names_text,
            "tooltip_text": names_text,
        }

    def _build_q_segment_row_widget(self, flow_section_idx):
        parts = self._build_q_segment_label(flow_section_idx)
        container = QWidget()
        container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(2)

        title_label = self._styled_q_segment_title_label(parts["segment_title"])
        layout.addWidget(title_label)

        names_label = None
        if parts["names_text"]:
            names_label = self._styled_name_value_label(
                parts["names_text"],
                tooltip_text=parts["tooltip_text"],
                max_lines=2,
            )
            layout.addWidget(names_label)
            container.setToolTip(parts["tooltip_text"])

        container._segment_title_label = title_label
        container._segment_names_label = names_label
        return container

    def _make_fixed_line_edit(self, width, text="", placeholder_text=""):
        edit = LineEdit()
        edit.setFixedWidth(width)
        if text:
            edit.setText(str(text))
        if placeholder_text:
            edit.setPlaceholderText(str(placeholder_text))
        return edit

    def _apply_group_body_layout(self, layout):
        left, top, right, bottom = self._ui_group_margins
        layout.setContentsMargins(left, top, right, bottom)
        layout.setSpacing(self._ui_grid_spacing)

    def _configure_q_grid(self, grid):
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(8)
        grid.setColumnMinimumWidth(0, self._ui_name_column_min_width)
        grid.setColumnStretch(0, 1)
        grid.setColumnMinimumWidth(1, self._ui_q_value_column_width)
        grid.setColumnStretch(1, 0)

    def _configure_pressurized_form_grid(self, grid):
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(8)
        grid.setColumnMinimumWidth(0, self._ui_name_column_min_width)
        grid.setColumnStretch(0, 1)
        grid.setColumnMinimumWidth(1, self._ui_material_column_width)
        grid.setColumnStretch(1, 0)
        grid.setColumnMinimumWidth(2, self._ui_dn_column_width)
        grid.setColumnStretch(2, 0)

    def _configure_struct_form_grid(self, grid, value_column_count):
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(8)
        grid.setColumnMinimumWidth(0, 120)
        grid.setColumnStretch(0, 1)
        for ci in range(1, value_column_count + 1):
            grid.setColumnMinimumWidth(ci, self._ui_numeric_column_width)
            grid.setColumnStretch(ci, 0)

    def _set_grid_first_column_labels(self, grid, labels):
        for row, label in labels:
            label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            label.setMinimumWidth(120)
            grid.addWidget(label, row, 0, alignment=Qt.AlignLeft | Qt.AlignVCenter)

    def _populate_pressure_pipe_material_combo(self, combo):
        combo.clear()
        for material_key in self._PRESSURE_PIPE_MATERIALS.keys():
            combo.addItem(self._pressure_pipe_material_display_name(material_key), material_key)

    def _set_pressure_pipe_material_combo_value(self, combo, material_value):
        material_key = self._pressure_pipe_material_key(material_value)
        idx = combo.findData(material_key)
        if idx < 0:
            idx = 0
        combo.setCurrentIndex(idx)

    def _current_pressure_pipe_material_value(self, combo):
        material_key = combo.currentData()
        if material_key:
            return str(material_key)
        return self._pressure_pipe_material_key(combo.currentText())

    def showEvent(self, event):
        """确保对话框不超出屏幕可见区域"""
        super().showEvent(event)
        from PySide6.QtGui import QGuiApplication
        screen = QGuiApplication.primaryScreen()
        if screen:
            avail = screen.availableGeometry()
            geo = self.frameGeometry()
            # 如果窗口高度超出屏幕，缩小到屏幕高度
            if geo.height() > avail.height():
                self.resize(self.width(), avail.height() - 20)
                geo = self.frameGeometry()
            # 如果顶部超出屏幕，向下移动
            if geo.top() < avail.top():
                geo.moveTop(avail.top())
            # 如果底部超出屏幕，向上移动
            if geo.bottom() > avail.bottom():
                geo.moveBottom(avail.bottom())
            # 如果左侧超出屏幕
            if geo.left() < avail.left():
                geo.moveLeft(avail.left())
            self.move(geo.topLeft())

    # ---- UI 构建 ----
    def _build_ui(self, default_qs):
        outer_lay = QVBoxLayout(self)
        outer_lay.setContentsMargins(0, 0, 0, 0)
        outer_lay.setSpacing(0)

        # 用 QScrollArea 包裹全部内容，防止内容超高时顶部被截断
        from PySide6.QtWidgets import QScrollArea
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        content = QWidget()
        lay = QVBoxLayout(content)
        lay.setSpacing(10)

        # 提示文字
        desc = QLabel("本功能将自动计算并生成多种建筑物断面水力要素汇总表（Excel），\n"
                      "可直接用于 AutoCAD 制表。")
        desc.setWordWrap(True)
        desc.setStyleSheet("font-size:13px; color:#333;")
        lay.addWidget(desc)

        invalid_notice = self._build_invalid_pressurized_notice()
        if invalid_notice:
            invalid_lbl = QLabel(invalid_notice)
            invalid_lbl.setWordWrap(True)
            invalid_lbl.setStyleSheet(
                "font-size:11px; color:#8a4b00; background:#fff4e5; border:1px solid #f3d19c; "
                "border-radius:6px; padding:8px;"
            )
            lay.addWidget(invalid_lbl)

        # ---- 流量段参数 ----
        q_group = QGroupBox("流量段设计流量 Q (m³/s)")
        q_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        q_lay = QVBoxLayout(q_group)
        self._apply_group_body_layout(q_lay)

        q_grid = QGridLayout()
        self._configure_q_grid(q_grid)
        self._q_form_grid = q_grid

        self._q_segment_structure_names = self._build_q_segment_structure_names()
        self._q_edits = []
        for i in range(self._segment_count):
            row_widget = self._build_q_segment_row_widget(i + 1)
            edit = self._make_fixed_line_edit(
                self._ui_q_value_column_width,
                text=default_qs[i] if i < len(default_qs) else default_qs[-1],
            )
            if self._has_source_data:
                edit.setReadOnly(True)
                edit.setToolTip("当前导出严格复用表2/表3结果，流量不允许在导出阶段改写。")
            q_grid.addWidget(row_widget, i, 0)
            q_grid.addWidget(edit, i, 1, alignment=Qt.AlignLeft | Qt.AlignVCenter)
            self._q_edits.append(edit)

        q_lay.addLayout(q_grid)
        lay.addWidget(q_group)

        # ---- 倒虹吸管道参数（按建筑物 + 流量段） ----
        siphon_group = QGroupBox("倒虹吸管道参数")
        siphon_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        siphon_lay = QVBoxLayout(siphon_group)
        self._apply_group_body_layout(siphon_lay)

        if self._siphon_groups:
            siphon_items = _merge_pressurized_param_defaults(
                self._siphon_groups,
                self._cached_pressurized.get("siphon", []),
            )
        else:
            siphon_items = [] if self._nodes else [
                _make_pressurized_param_row(
                    name="倒虹吸",
                    flow_section=None,
                    structure_kind="siphon",
                    pipe_material="球墨铸铁管",
                    dn_mm=1500,
                    display_name="倒虹吸",
                )
            ]

        self._siphon_rows = []  # [(row_dict, mat_combo, dn_edit), ...]
        sp_grid = QGridLayout()
        self._configure_pressurized_form_grid(sp_grid)
        self._siphon_form_grid = sp_grid
        for ci, txt in enumerate(['倒虹吸名称（含流量段）', '管道材质', 'DN (mm)']):
            sp_grid.addWidget(
                self._styled_form_header_label(txt),
                0,
                ci,
                alignment=Qt.AlignLeft | Qt.AlignVCenter,
            )
        for ri, sp_row in enumerate(siphon_items):
            row_index = ri + 1
            name_lbl = self._styled_name_value_label(sp_row["display_name"])
            sp_grid.addWidget(name_lbl, row_index, 0, alignment=Qt.AlignLeft | Qt.AlignVCenter)

            mat_combo = QComboBox()
            mat_combo.addItems(list(self._SIPHON_MATERIALS.keys()))
            mat_combo.setCurrentText(
                sp_row["pipe_material"] if sp_row["pipe_material"] in self._SIPHON_MATERIALS else "球墨铸铁管"
            )
            mat_combo.setFixedWidth(self._ui_material_column_width)
            if self._has_source_data:
                mat_combo.setEnabled(False)
            sp_grid.addWidget(mat_combo, row_index, 1, alignment=Qt.AlignLeft | Qt.AlignVCenter)

            dn_edit = self._make_fixed_line_edit(self._ui_dn_column_width)
            dn_val = _normalize_dn_mm(sp_row["DN_mm"], 1500)
            dn_edit.setText(str(dn_val))
            if self._has_source_data:
                dn_edit.setReadOnly(True)
                dn_edit.setToolTip("当前导出严格复用表2/表3结果，DN 不允许在导出阶段改写。")
            sp_grid.addWidget(dn_edit, row_index, 2, alignment=Qt.AlignLeft | Qt.AlignVCenter)

            self._siphon_rows.append((dict(sp_row), mat_combo, dn_edit))

        siphon_lay.addLayout(sp_grid)

        dn_note_text = (
            "（检测到表2/表3结果时，倒虹吸 DN 与相关结果将严格复用当前已确认值，导出阶段不可改写）"
            if self._has_source_data
            else "（DN 从倒虹吸计算结果自动导入，可在缺少源结果时作为补全参数）"
        )
        dn_note = QLabel(dn_note_text)
        dn_note.setStyleSheet("font-size:11px; color:#666;")
        siphon_lay.addWidget(dn_note)
        lay.addWidget(siphon_group)

        # ---- 有压管道参数（按建筑物 + 流量段） ----
        pressure_pipe_group = QGroupBox("有压管道参数")
        pressure_pipe_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        pp_lay = QVBoxLayout(pressure_pipe_group)
        self._apply_group_body_layout(pp_lay)

        if self._pressure_pipe_groups:
            pp_items = _build_pressure_pipe_dialog_rows(
                self._pressure_pipe_groups,
                self._cached_pressurized.get("pressure_pipe", []),
            )
        else:
            pp_items = [] if self._nodes else [
                _make_pressurized_param_row(
                    name="有压管道",
                    flow_section=None,
                    structure_kind="pressure_pipe",
                    pipe_material="球墨铸铁管",
                    dn_mm=1500,
                    display_name="有压管道",
                )
            ]

        self._pressure_pipe_rows = []  # [(row_dict, mat_combo, dn_edit), ...]
        pp_grid = QGridLayout()
        self._configure_pressurized_form_grid(pp_grid)
        self._pressure_pipe_form_grid = pp_grid
        for ci, txt in enumerate(['有压管道名称（含流量段）', '管道材质', 'DN (mm)']):
            pp_grid.addWidget(
                self._styled_form_header_label(txt),
                0,
                ci,
                alignment=Qt.AlignLeft | Qt.AlignVCenter,
            )
        for ri, pp_row in enumerate(pp_items):
            row_index = ri + 1
            name_lbl = self._styled_name_value_label(pp_row["display_name"])
            pp_grid.addWidget(name_lbl, row_index, 0, alignment=Qt.AlignLeft | Qt.AlignVCenter)

            mat_combo = QComboBox()
            self._populate_pressure_pipe_material_combo(mat_combo)
            self._set_pressure_pipe_material_combo_value(mat_combo, pp_row.get("pipe_material"))
            mat_combo.setFixedWidth(self._ui_material_column_width)
            if self._has_source_data:
                mat_combo.setEnabled(False)
            pp_grid.addWidget(mat_combo, row_index, 1, alignment=Qt.AlignLeft | Qt.AlignVCenter)

            dn_edit = self._make_fixed_line_edit(self._ui_dn_column_width)
            dn_val = _normalize_dn_mm(pp_row["DN_mm"], 1500)
            dn_edit.setText(str(dn_val))
            if self._has_source_data:
                dn_edit.setReadOnly(True)
                dn_edit.setToolTip("当前导出严格复用表2/表3结果，DN 不允许在导出阶段改写。")
            pp_grid.addWidget(dn_edit, row_index, 2, alignment=Qt.AlignLeft | Qt.AlignVCenter)

            self._pressure_pipe_rows.append((dict(pp_row), mat_combo, dn_edit))

        pp_lay.addLayout(pp_grid)

        pp_note_text = (
            "（检测到表2/表3结果时，有压管道 DN、材质及相关结果将严格复用当前已确认值，导出阶段不可改写）"
            if self._has_source_data
            else "（DN 从有压管道计算结果自动导入，可在缺少源结果时作为补全参数）"
        )
        pp_note = QLabel(pp_note_text)
        pp_note.setStyleSheet("font-size:11px; color:#666;")
        pp_lay.addWidget(pp_note)
        lay.addWidget(pressure_pipe_group)

        # ---- 构造参数设置（Tab页签） ----
        from PySide6.QtWidgets import QTabWidget
        struct_group = QGroupBox("构造参数设置（壁厚/衬砌厚度）")
        struct_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        struct_lay = QVBoxLayout(struct_group)

        struct_tabs = QTabWidget()
        struct_tabs.setStyleSheet("QTabWidget{font-size:11px;} QTabBar::tab{min-width:70px;}")

        # ---- Tab 1: 明渠类 ----
        tab_channel = QWidget()
        tc_lay = QVBoxLayout(tab_channel)
        tc_lay.setSpacing(6)

        tc_grid = QGridLayout()
        self._configure_struct_form_grid(tc_grid, 2)
        self._struct_channel_grid = tc_grid
        for ci, txt in enumerate(['壁厚 t (m)', '拉杆尺寸 (m)']):
            tc_grid.addWidget(
                self._styled_form_header_label(txt),
                0,
                ci + 1,
                alignment=Qt.AlignLeft | Qt.AlignVCenter,
            )

        def _tie_rod_pair(default_w=0.2, default_h=0.2):
            """创建拉杆尺寸 [宽] × [高] 组合控件，返回 (container, w_edit, h_edit)。"""
            container = QWidget()
            h_lay = QHBoxLayout(container)
            h_lay.setContentsMargins(0, 0, 0, 0)
            h_lay.setSpacing(3)
            w_lbl = QLabel("宽"); w_lbl.setStyleSheet("font-size:10px; color:#424242;")
            w_edit = LineEdit(); w_edit.setFixedWidth(55)
            w_edit.setText(str(default_w)); w_edit.setPlaceholderText(str(default_w))
            x_lbl = QLabel("×"); x_lbl.setFixedWidth(12)
            x_lbl.setStyleSheet("font-size:12px;")
            h_lbl = QLabel("高"); h_lbl.setStyleSheet("font-size:10px; color:#424242;")
            h_edit = LineEdit(); h_edit.setFixedWidth(55)
            h_edit.setText(str(default_h)); h_edit.setPlaceholderText(str(default_h))
            h_lay.addWidget(w_lbl)
            h_lay.addWidget(w_edit)
            h_lay.addWidget(x_lbl)
            h_lay.addWidget(h_lbl)
            h_lay.addWidget(h_edit)
            h_lay.addStretch()
            return container, w_edit, h_edit

        # 矩形明渠
        self._set_grid_first_column_labels(
            tc_grid,
            [
                (1, QLabel("矩形明渠")),
                (2, QLabel("梯形明渠")),
            ],
        )
        self._rect_ch_wall_t = self._make_fixed_line_edit(self._ui_numeric_column_width, "0.3", "0.3")
        tc_grid.addWidget(self._rect_ch_wall_t, 1, 1, alignment=Qt.AlignLeft | Qt.AlignVCenter)
        rc_tr_container, self._rect_ch_tie_w, self._rect_ch_tie_h = _tie_rod_pair()
        tc_grid.addWidget(rc_tr_container, 1, 2, alignment=Qt.AlignLeft | Qt.AlignVCenter)

        # 梯形明渠
        self._trap_ch_wall_t = self._make_fixed_line_edit(self._ui_numeric_column_width, "0.3", "0.3")
        tc_grid.addWidget(self._trap_ch_wall_t, 2, 1, alignment=Qt.AlignLeft | Qt.AlignVCenter)
        tp_tr_container, self._trap_ch_tie_w, self._trap_ch_tie_h = _tie_rod_pair()
        tc_grid.addWidget(tp_tr_container, 2, 2, alignment=Qt.AlignLeft | Qt.AlignVCenter)
        tc_lay.addLayout(tc_grid)
        tc_lay.addStretch()
        struct_tabs.addTab(tab_channel, "明渠类")

        # ---- Tab 2: 渡槽类 ----
        tab_aqueduct = QWidget()
        ta_lay = QVBoxLayout(tab_aqueduct)
        ta_lay.setSpacing(6)

        ta_grid = QGridLayout()
        self._configure_struct_form_grid(ta_grid, 1)
        self._struct_aqueduct_grid = ta_grid
        ta_grid.addWidget(self._styled_form_header_label("壁厚 t (m)"), 0, 1, alignment=Qt.AlignLeft | Qt.AlignVCenter)

        # U形渡槽
        self._set_grid_first_column_labels(
            ta_grid,
            [
                (1, QLabel("U形渡槽")),
                (2, QLabel("矩形渡槽")),
            ],
        )
        self._aq_u_wall_t = self._make_fixed_line_edit(self._ui_numeric_column_width, "0.35", "0.35")
        ta_grid.addWidget(self._aq_u_wall_t, 1, 1, alignment=Qt.AlignLeft | Qt.AlignVCenter)

        # 矩形渡槽
        self._aq_rect_wall_t = self._make_fixed_line_edit(self._ui_numeric_column_width, "0.35", "0.35")
        ta_grid.addWidget(self._aq_rect_wall_t, 2, 1, alignment=Qt.AlignLeft | Qt.AlignVCenter)
        ta_lay.addLayout(ta_grid)
        ta_lay.addStretch()
        struct_tabs.addTab(tab_aqueduct, "渡槽类")

        # ---- Tab 3: 暗涵 ----
        tab_culvert = QWidget()
        tv_lay = QVBoxLayout(tab_culvert)
        tv_lay.setSpacing(6)

        tv_grid = QGridLayout()
        self._configure_struct_form_grid(tv_grid, 3)
        self._struct_culvert_grid = tv_grid
        for ci, txt in enumerate(['底板厚 t\u2080 (m)', '边墙厚 t\u2081 (m)', '顶板厚 t\u2082 (m)']):
            tv_grid.addWidget(
                self._styled_form_header_label(txt),
                0,
                ci + 1,
                alignment=Qt.AlignLeft | Qt.AlignVCenter,
            )

        self._set_grid_first_column_labels(tv_grid, [(1, QLabel("矩形暗涵"))])
        self._culvert_t0 = self._make_fixed_line_edit(self._ui_numeric_column_width, "0.4", "0.4")
        tv_grid.addWidget(self._culvert_t0, 1, 1, alignment=Qt.AlignLeft | Qt.AlignVCenter)
        self._culvert_t1 = self._make_fixed_line_edit(self._ui_numeric_column_width, "0.4", "0.4")
        tv_grid.addWidget(self._culvert_t1, 1, 2, alignment=Qt.AlignLeft | Qt.AlignVCenter)
        self._culvert_t2 = self._make_fixed_line_edit(self._ui_numeric_column_width, "0.4", "0.4")
        tv_grid.addWidget(self._culvert_t2, 1, 3, alignment=Qt.AlignLeft | Qt.AlignVCenter)
        tv_lay.addLayout(tv_grid)
        tv_lay.addStretch()
        struct_tabs.addTab(tab_culvert, "暗涵")

        # ---- Tab 4: 隧洞 ----
        tab_tunnel = QWidget()
        tt_lay = QVBoxLayout(tab_tunnel)
        tt_lay.setSpacing(6)

        tt_desc = QLabel("4种隧洞类型共用此设置（圆拱直墙型/圆形/马蹄形Ⅰ型/Ⅱ型）")
        tt_desc.setStyleSheet("font-size:11px; color:#666;")
        tt_lay.addWidget(tt_desc)

        tt_grid = QGridLayout()
        self._configure_struct_form_grid(tt_grid, 2)
        self._struct_tunnel_grid = tt_grid
        for ci, txt in enumerate(['底板厚 t\u2080 (m)', '边墙/顶拱/衬砌厚 t (m)']):
            tt_grid.addWidget(
                self._styled_form_header_label(txt),
                0,
                ci + 1,
                alignment=Qt.AlignLeft | Qt.AlignVCenter,
            )

        self._lining_edits = {}  # {rock_class: (t0_edit, t_edit)}
        for ri, rc in enumerate(self._ROCK_CLASSES):
            name_lbl = QLabel(rc)
            name_lbl.setMinimumWidth(120)
            tt_grid.addWidget(name_lbl, ri + 1, 0, alignment=Qt.AlignLeft | Qt.AlignVCenter)
            defaults = self._ROCK_LINING_DEFAULT[rc]
            t0_edit = self._make_fixed_line_edit(self._ui_numeric_column_width, str(defaults['t0']), str(defaults['t0']))
            tt_grid.addWidget(t0_edit, ri + 1, 1, alignment=Qt.AlignLeft | Qt.AlignVCenter)
            t_edit = self._make_fixed_line_edit(self._ui_numeric_column_width, str(defaults['t']), str(defaults['t']))
            tt_grid.addWidget(t_edit, ri + 1, 2, alignment=Qt.AlignLeft | Qt.AlignVCenter)
            self._lining_edits[rc] = (t0_edit, t_edit)

        tt_lay.addLayout(tt_grid)

        # ---- 隧洞断面设计方式 ----
        from PySide6.QtWidgets import QRadioButton, QButtonGroup, QHBoxLayout as _QHBox
        _tt_mode_row = QWidget()
        _tt_mode_hlay = _QHBox(_tt_mode_row)
        _tt_mode_hlay.setContentsMargins(0, 0, 0, 0)
        _tt_mode_hlay.setSpacing(4)
        tt_mode_lbl = QLabel("断面设计方式:")
        tt_mode_lbl.setStyleSheet("font-size:11px; color:#555; font-weight:bold; margin-top:6px;")
        _tt_mode_hlay.addWidget(tt_mode_lbl)
        _info_icon = QLabel("ⓘ")
        _info_icon.setStyleSheet(
            "font-size:13px; color:#1a73e8; font-weight:bold; margin-top:6px; cursor:pointer;"
        )
        _info_icon.setCursor(Qt.PointingHandCursor)
        _dialog_self = self
        _info_icon.mousePressEvent = lambda e: PopupTeachingTip.create(
            target=_info_icon,
            icon=InfoBarIcon.INFORMATION,
            title='断面设计方式',
            content='统一断面：按最大流量段设计统一断面尺寸，其余各流量段仅推求水深；\n'
                    '独立断面：每个流量段独立计算各自的断面尺寸。',
            isClosable=False,
            tailPosition=TeachingTipTailPosition.BOTTOM,
            duration=-1,
            parent=_dialog_self,
        )
        _tt_mode_hlay.addWidget(_info_icon)
        _tt_mode_hlay.addStretch()
        tt_lay.addWidget(_tt_mode_row)

        self._tunnel_mode_groups = {}  # {key: QButtonGroup}
        _tunnel_types = [
            ("tunnel_arch",      "圆拱直墙型"),
            ("tunnel_circular",  "圆形"),
            ("tunnel_horseshoe", "马蹄形（Ⅰ/Ⅱ型）"),
        ]
        tm_grid = QGridLayout()
        tm_grid.setSpacing(2)
        for ri, (tkey, tname) in enumerate(_tunnel_types):
            name_lbl = QLabel(tname)
            name_lbl.setStyleSheet("font-size:11px;")
            name_lbl.setFixedWidth(110)
            tm_grid.addWidget(name_lbl, ri, 0)
            rb_unified = QRadioButton("统一断面")
            rb_indep  = QRadioButton("独立断面")
            rb_unified.setStyleSheet("font-size:11px;")
            rb_indep.setStyleSheet("font-size:11px;")
            rb_indep.setChecked(True)
            bg = QButtonGroup(self)
            bg.addButton(rb_unified, 0)
            bg.addButton(rb_indep, 1)
            tm_grid.addWidget(rb_unified, ri, 1)
            tm_grid.addWidget(rb_indep, ri, 2)
            self._tunnel_mode_groups[tkey] = bg
        tt_lay.addLayout(tm_grid)

        tt_lay.addStretch()
        struct_tabs.addTab(tab_tunnel, "隧洞")

        struct_tabs.setFixedHeight(260)
        struct_lay.addWidget(struct_tabs)

        struct_note = QLabel('（不输入则使用默认值，修改后同时影响"生成断面汇总表"和"导出全部DXF"）')
        struct_note.setStyleSheet("font-size:11px; color:#666;")
        struct_lay.addWidget(struct_note)
        lay.addWidget(struct_group)

        # ---- 说明 ----
        note_group = QGroupBox("其他参数说明")
        note_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        note_lay = QVBoxLayout(note_group)
        note_lbl = QLabel(
            "• 各类构造参数可在上方按类型自定义\n"
            '• 隧洞断面设计方式可在"隧洞"选项卡中按类型分别设置\n'
            "• 圆管涵、倒虹吸无需设置壁厚")
        note_lbl.setWordWrap(True)
        note_lbl.setStyleSheet("font-size:11px; color:#555;")
        note_lay.addWidget(note_lbl)
        lay.addWidget(note_group)

        # ---- 导出格式选择 ----
        from PySide6.QtWidgets import QRadioButton, QButtonGroup
        fmt_group = QGroupBox("导出格式")
        fmt_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:12px;}")
        fmt_lay = QHBoxLayout(fmt_group)
        self._radio_excel = QRadioButton("Excel (.xlsx)  — 多Sheet + 汇总Sheet")
        self._radio_dxf = QRadioButton("DXF (.dxf)  — 可直接导入AutoCAD")
        self._radio_excel.setStyleSheet("font-size:11px;")
        self._radio_dxf.setStyleSheet("font-size:11px;")
        self._radio_excel.setChecked(True)
        self._fmt_btn_group = QButtonGroup(self)
        self._fmt_btn_group.addButton(self._radio_excel, 0)
        self._fmt_btn_group.addButton(self._radio_dxf, 1)
        fmt_lay.addWidget(self._radio_excel)
        fmt_lay.addWidget(self._radio_dxf)
        fmt_lay.addStretch()
        lay.addWidget(fmt_group)
        if self._config_only:
            fmt_group.setVisible(False)

        scroll.setWidget(content)
        outer_lay.addWidget(scroll, 1)

        # ---- 按钮栏（固定在底部，不随滚动） ----
        btn_lay = QHBoxLayout()
        btn_lay.setContentsMargins(10, 6, 10, 6)
        btn_lay.addStretch()
        btn_cancel = PushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        btn_text = "确认参数" if self._config_only else "生成汇总表"
        btn_generate = PrimaryPushButton(btn_text)
        btn_generate.clicked.connect(self._on_generate)
        btn_lay.addWidget(btn_cancel)
        btn_lay.addWidget(btn_generate)
        outer_lay.addLayout(btn_lay)

    # ---- 读取构造参数 ----
    def _invalid_pressurized_items(self):
        items = []
        for row in self._invalid_siphon_groups:
            items.append(f"倒虹吸：{row.get('display_name') or row.get('name') or '倒虹吸'}")
        for row in self._invalid_pressure_pipe_groups:
            items.append(f"有压管道：{row.get('display_name') or row.get('name') or '有压管道'}")
        return items

    def _build_invalid_pressurized_notice(self):
        items = self._invalid_pressurized_items()
        if not items:
            return ""
        return "以下有压流建筑物缺少有效流量段，当前无法导出，请先补全后再试：\n" + "；".join(items)

    def _block_invalid_pressurized_export(self):
        notice = self._build_invalid_pressurized_notice()
        if notice:
            self._invalid_pressurized_notice_shown = True
            fluent_error(self, "无法导出", notice)
            return True
        return False

    def _read_pressurized_rows(self, rows, title_prefix):
        out = []
        for row, mat_combo, dn_edit in rows:
            dn = _parse_positive_dn(dn_edit.text())
            row_name = row.get("display_name") or row.get("name") or title_prefix
            if dn is None:
                fluent_error(self, "输入错误", f"{row_name} 的 DN 必须为正整数")
                return None
            pipe_material = mat_combo.currentText()
            if title_prefix == "有压管道":
                pipe_material = self._current_pressure_pipe_material_value(mat_combo)
            target_rows = row.get("dialog_target_rows") if isinstance(row, dict) else None
            if title_prefix == "有压管道" and isinstance(target_rows, list) and target_rows:
                for target_row in target_rows:
                    if not isinstance(target_row, dict):
                        continue
                    out.append(
                        _build_pressurized_output_row(
                            target_row,
                            pipe_material,
                            dn,
                        )
                    )
                continue
            out.append(_build_pressurized_output_row(row, pipe_material, dn))
        return out

    def _attach_pressure_pipe_calc_contexts(self, rows):
        return _attach_pressure_pipe_calc_contexts_to_rows(
            rows,
            getattr(self, "_pressure_pipe_calc_contexts", None),
        )

    @staticmethod
    def _normalize_pressure_pipe_total_head_loss_value(value):
        return _normalize_pressure_pipe_total_head_loss_value(value)

    def _get_panel_pressure_pipe_export_results(self, rows):
        return _get_panel_pressure_pipe_export_results(getattr(self, "_panel", None), rows)

    def _attach_pressure_pipe_export_results(self, rows):
        return _attach_pressure_pipe_export_results_to_rows(rows, panel=getattr(self, "_panel", None))

    def _summarize_pressurized_materials(self, rows):
        parts = []
        for row in rows or []:
            label = row.get("display_name") or row.get("name")
            material = row.get("pipe_material")
            if row.get("structure_kind") == "pressure_pipe" and material:
                material = self._pressure_pipe_material_display_name(material)
            if label and material:
                parts.append(f"{label}({material})")
        return "、".join(parts)

    def _collect_pressure_pipe_missing_total_head_loss_labels(self, rows):
        if not rows:
            return []
        return [
            row.get("name") or ""
            for row in rows
            if self._normalize_pressure_pipe_total_head_loss_value(row.get("total_head_loss")) is None
        ]

    def _collect_pressure_pipe_missing_velocity_labels(self, rows):
        if not rows:
            return []
        labels = []
        for row in rows:
            if _normalize_locked_velocity_value(row.get("V")) is not None:
                continue
            q_value = _normalize_positive_flow_value(row.get("Q"))
            dn_ready = _parse_positive_dn(row.get("DN_mm")) is not None and _row_has_valid_pressurized_dn_mm(row)
            if q_value is not None and dn_ready:
                continue
            label = row.get("display_name") or row.get("name") or ""
            if label:
                labels.append(label)
        return labels

    def _collect_siphon_missing_velocity_labels(self, rows):
        if not rows:
            return []
        return [
            row.get("name") or ""
            for row in rows
            if _normalize_locked_velocity_value(row.get("V")) is None
        ]

    def _warn_siphon_missing_velocity(self, rows):
        items = [label for label in self._collect_siphon_missing_velocity_labels(rows) if label]
        if not items:
            return
        fluent_info(
            self,
            "提示",
            "以下倒虹吸缺少已算流速结果，本次导出将以“-”显示：\n" + "；".join(items),
        )

    def _warn_pressure_pipe_missing_velocity(self, rows):
        items = [label for label in self._collect_pressure_pipe_missing_velocity_labels(rows) if label]
        if not items:
            return
        fluent_info(
            self,
            "提示",
            "以下有压管道缺少已算流速结果，且无法根据 Q/DN 补算，本次导出将以“-”显示：\n" + "；".join(items),
        )

    def _warn_pressure_pipe_missing_total_head_loss(self, rows):
        # 新版导出表不再展示总水头损失，因此这里保留检测逻辑但不弹提示。
        _ = rows
        return

    def _read_float(self, edit, default):
        """安全读取 LineEdit 的浮点值，空或非法返回默认值。"""
        t = edit.text().strip()
        if not t:
            return default
        try:
            return float(t)
        except ValueError:
            return default

    def _read_rock_lining(self):
        """从输入框读取用户自定义的围岩衬砌厚度。"""
        rock_lining = {}
        for rc in self._ROCK_CLASSES:
            t0_edit, t_edit = self._lining_edits[rc]
            defaults = self._ROCK_LINING_DEFAULT[rc]
            rock_lining[rc] = {
                't0': self._read_float(t0_edit, defaults['t0']),
                't':  self._read_float(t_edit,  defaults['t']),
            }
        return rock_lining

    def _read_tie_rod(self, w_edit, h_edit):
        """从拉杆宽/高输入框读取并组合为 'd1×d2' 字符串。"""
        w = self._read_float(w_edit, 0.2)
        h = self._read_float(h_edit, 0.2)
        return f"{w}×{h}"

    def _read_struct_thickness(self):
        """读取所有结构类型的用户自定义厚度参数，返回 dict。"""
        return {
            'rect_channel': {
                'wall_t':  self._read_float(self._rect_ch_wall_t, 0.3),
                'tie_rod': self._read_tie_rod(self._rect_ch_tie_w, self._rect_ch_tie_h),
            },
            'trap_channel': {
                'wall_t':  self._read_float(self._trap_ch_wall_t, 0.3),
                'tie_rod': self._read_tie_rod(self._trap_ch_tie_w, self._trap_ch_tie_h),
            },
            'aqueduct_u': {
                'wall_t': self._read_float(self._aq_u_wall_t, 0.35),
            },
            'aqueduct_rect': {
                'wall_t': self._read_float(self._aq_rect_wall_t, 0.35),
            },
            'rect_culvert': {
                't0': self._read_float(self._culvert_t0, 0.4),
                't1': self._read_float(self._culvert_t1, 0.4),
                't2': self._read_float(self._culvert_t2, 0.4),
            },
            'rock_lining': self._read_rock_lining(),
        }

    # ---- 生成 ----
    def _on_generate(self):
        from calc_渠系计算算法内核.生成断面汇总表 import (
            _default_segments_rect_channel,
            _default_segments_trap_channel,
            _default_segments_u_channel,
            _default_segments_tunnel_arch,
            _default_segments_tunnel_circular,
            _default_segments_tunnel_horseshoe,
            _default_segments_aqueduct_u,
            _default_segments_aqueduct_rect,
            _default_segments_rect_culvert,
            _default_segments_circular_pipe,
            _segment_name,
        )

        # 读取 Q 值
        try:
            qs = [float(e.text()) for e in self._q_edits]
        except ValueError:
            fluent_error(self, "输入错误", "流量值必须为数字")
            return

        if self._has_source_data:
            qs = self._build_default_qs()

        if self._block_invalid_pressurized_export():
            return

        siphon_params = self._read_pressurized_rows(self._siphon_rows, "倒虹吸")
        if siphon_params is None:
            return

        pressure_pipe_params = self._read_pressurized_rows(self._pressure_pipe_rows, "有压管道")
        if pressure_pipe_params is None:
            return
        pressure_pipe_params = _prepare_pressure_pipe_export_rows(
            pressure_pipe_params,
            panel=getattr(self, "_panel", None),
            calc_contexts=getattr(self, "_pressure_pipe_calc_contexts", None),
        )
        pressure_pipe_export_rows = _merge_pressure_pipe_export_rows_by_flow_section(
            pressure_pipe_params,
            panel=getattr(self, "_panel", None),
        )

        # config_only 模式：只读取并缓存参数，不生成文件
        if self._config_only:
            struct_t = self._read_struct_thickness()
            rock_lining = struct_t['rock_lining']
            tunnel_unified = {}
            for tkey, bg in self._tunnel_mode_groups.items():
                tunnel_unified[tkey] = (bg.checkedId() == 0)
            if self._panel is not None:
                self._panel._custom_rock_lining = rock_lining
                self._panel._custom_struct_thickness = struct_t
                self._panel._custom_tunnel_unified = tunnel_unified
                self._panel._custom_pressurized_pipe_params = {
                    "siphon": _serialize_pressurized_cache_rows(siphon_params, "siphon"),
                    "pressure_pipe": _serialize_pressurized_cache_rows(pressure_pipe_params, "pressure_pipe"),
                }
            self.accept()
            return

        segment_count = self._segment_count
        node_defaults = self._node_defaults

        # 判断导出格式
        export_dxf = self._radio_dxf.isChecked()
        if export_dxf:
            ext = ".dxf"
            filter_str = "DXF 文件 (*.dxf);;所有文件 (*.*)"
            auto_name = self._auto_name.replace('.xlsx', '.dxf') if self._auto_name else ""
        else:
            ext = ".xlsx"
            filter_str = "Excel 文件 (*.xlsx);;所有文件 (*.*)"
            auto_name = self._auto_name

        # 选择保存路径
        fp, _ = QFileDialog.getSaveFileName(
            self, "保存断面汇总表", auto_name, filter_str)
        if not fp:
            return
        if not fp.lower().endswith(ext):
            fp += ext

        # 构建各表参数
        has_source_data = self._has_source_data

        def _make_segs(default_fn, overrides_by_idx=None):
            # 有源数据时，只生成有实际节点数据的流量段
            if has_source_data and overrides_by_idx is not None:
                if not overrides_by_idx:
                    return []
                defaults_pool = default_fn()
                segs = []
                for idx in sorted(overrides_by_idx.keys()):
                    # 用默认段作为基础模板
                    base = dict(defaults_pool[0]) if defaults_pool else {}
                    base["name"] = _segment_name(idx)
                    base.update(overrides_by_idx[idx])
                    if 0 < idx <= len(qs):
                        base["Q"] = qs[idx - 1]
                    segs.append(base)
                return segs
            # 无源数据时（独立运行），用默认值生成所有段
            segs = default_fn()
            if len(segs) < segment_count:
                last = segs[-1] if segs else {}
                for idx in range(len(segs) + 1, segment_count + 1):
                    new_seg = dict(last)
                    new_seg["name"] = _segment_name(idx)
                    segs.append(new_seg)
            segs = segs[:segment_count]
            for i, seg in enumerate(segs):
                if overrides_by_idx and (i + 1) in overrides_by_idx:
                    seg.update(overrides_by_idx[i + 1])
                if i < len(qs):
                    seg["Q"] = qs[i]
            return segs

        rc_segs = _make_segs(_default_segments_rect_channel, node_defaults.get("rect_channel"))
        tr_segs = _make_segs(_default_segments_trap_channel, node_defaults.get("trap_channel"))
        uc_segs = _make_segs(_default_segments_u_channel, node_defaults.get("u_channel"))
        tn_arch_segs = _make_segs(_default_segments_tunnel_arch, node_defaults.get("tunnel_arch"))
        tn_circ_segs = _make_segs(_default_segments_tunnel_circular, node_defaults.get("tunnel_circular"))
        tn_horse_segs = _make_segs(_default_segments_tunnel_horseshoe, node_defaults.get("tunnel_horseshoe"))
        aq_u_segs = _make_segs(_default_segments_aqueduct_u, node_defaults.get("aqueduct_u"))
        aq_rect_segs = _make_segs(_default_segments_aqueduct_rect, node_defaults.get("aqueduct_rect"))
        rv_segs = _make_segs(_default_segments_rect_culvert, node_defaults.get("rect_culvert"))
        cp_segs = _make_segs(_default_segments_circular_pipe, node_defaults.get("circular_channel"))

        if not has_source_data:
            for segs_list in [rc_segs, tr_segs, uc_segs, tn_arch_segs, tn_circ_segs, tn_horse_segs,
                              aq_u_segs, aq_rect_segs, rv_segs, cp_segs]:
                for i, seg in enumerate(segs_list):
                    seg["name"] = _segment_name(i + 1)

        sp_overrides = node_defaults.get("siphon", {})
        sp_segs = _build_pressurized_segments(
            qs=qs,
            overrides_by_idx=sp_overrides,
            params=siphon_params,
            has_source_data=has_source_data,
            segment_name_fn=_segment_name,
        )

        # 按结果决定表格类型
        _table_order = None
        if has_source_data:
            _table_order = []
            if node_defaults.get("rect_channel"):
                _table_order.append("rect_channel")
            if node_defaults.get("trap_channel"):
                _table_order.append("trap_channel")
            if node_defaults.get("u_channel"):
                _table_order.append("u_channel")
            if node_defaults.get("tunnel_arch"):
                _table_order.append("tunnel_arch")
            if node_defaults.get("tunnel_circular"):
                _table_order.append("tunnel_circular")
            if node_defaults.get("tunnel_horseshoe"):
                _table_order.append("tunnel_horseshoe")
            if node_defaults.get("aqueduct_u"):
                _table_order.append("aqueduct_u")
            if node_defaults.get("aqueduct_rect"):
                _table_order.append("aqueduct_rect")
            if node_defaults.get("rect_culvert"):
                _table_order.append("rect_culvert")
            if node_defaults.get("circular_channel"):
                _table_order.append("circular_channel")
            if node_defaults.get("siphon"):
                _table_order.append("siphon")
            if node_defaults.get("pressure_pipe"):
                _table_order.append("pressure_pipe")
            if not _table_order:
                _table_order = None

        # 读取所有用户自定义的构造参数（壁厚/衬砌厚度）
        struct_t = self._read_struct_thickness()
        rock_lining = struct_t['rock_lining']

        # 将壁厚/衬砌参数注入各类型 segments
        for seg in rc_segs:
            seg['wall_t'] = struct_t['rect_channel']['wall_t']
            seg['tie_rod'] = struct_t['rect_channel']['tie_rod']
        for seg in tr_segs:
            seg['wall_t'] = struct_t['trap_channel']['wall_t']
            seg['tie_rod'] = struct_t['trap_channel']['tie_rod']
        for seg in aq_u_segs:
            seg['wall_t'] = struct_t['aqueduct_u']['wall_t']
        for seg in aq_rect_segs:
            seg['wall_t'] = struct_t['aqueduct_rect']['wall_t']
        for seg in rv_segs:
            seg['t0'] = struct_t['rect_culvert']['t0']
            seg['t1'] = struct_t['rect_culvert']['t1']
            seg['t2'] = struct_t['rect_culvert']['t2']

        # 读取隧洞断面设计方式
        tunnel_unified = {}
        for tkey, bg in self._tunnel_mode_groups.items():
            tunnel_unified[tkey] = (bg.checkedId() == 0)  # 0=统一, 1=独立

        # 存储到 panel，供"导出全部DXF"复用
        if self._panel is not None:
            self._panel._custom_rock_lining = rock_lining
            self._panel._custom_struct_thickness = struct_t
            self._panel._custom_tunnel_unified = tunnel_unified
            self._panel._custom_pressurized_pipe_params = {
                "siphon": _serialize_pressurized_cache_rows(siphon_params, "siphon"),
                "pressure_pipe": _serialize_pressurized_cache_rows(pressure_pipe_params, "pressure_pipe"),
            }

        # 构建有压管道 segments（与倒虹吸类似）
        pp_overrides = node_defaults.get("pressure_pipe", {})
        pp_segs = _build_pressurized_segments(
            qs=qs,
            overrides_by_idx=pp_overrides,
            params=pressure_pipe_export_rows,
            has_source_data=has_source_data,
            segment_name_fn=_segment_name,
        )

        self._warn_siphon_missing_velocity(sp_segs)
        self._warn_pressure_pipe_missing_velocity(pressure_pipe_params)

        gen_kwargs = dict(
            filepath=fp,
            rect_channel_segs=rc_segs,
            trap_channel_segs=tr_segs,
            u_channel_segs=uc_segs,
            tunnel_arch_segs=tn_arch_segs,
            tunnel_circular_segs=tn_circ_segs,
            tunnel_horseshoe_segs=tn_horse_segs,
            aqueduct_u_segs=aq_u_segs,
            aqueduct_rect_segs=aq_rect_segs,
            rect_culvert_segs=rv_segs,
            circular_pipe_segs=cp_segs,
            siphon_segs=sp_segs,
            siphon_material=siphon_params[0]["pipe_material"] if siphon_params else "球墨铸铁管",
            pressure_pipe_segs=pp_segs,
            pressure_pipe_material=pressure_pipe_params[0]["pipe_material"] if pressure_pipe_params else "球墨铸铁管",
            rock_lining=rock_lining,
            table_order=_table_order,
            tunnel_unified_arch=False if has_source_data else tunnel_unified.get("tunnel_arch", False),
            tunnel_unified_circular=False if has_source_data else tunnel_unified.get("tunnel_circular", False),
            tunnel_unified_horseshoe=False if has_source_data else tunnel_unified.get("tunnel_horseshoe", False),
        )

        try:
            self.setCursor(Qt.WaitCursor)
            QApplication.processEvents()
            if export_dxf:
                self._generate_dxf(**gen_kwargs)
            else:
                self._generate_excel(**gen_kwargs)
            self.unsetCursor()
            mat_summary = self._summarize_pressurized_materials(siphon_params)
            pp_mat_summary = self._summarize_pressurized_materials(pressure_pipe_params)
            fmt_name = "DXF" if export_dxf else "Excel"
            extra = "" if export_dxf else "\n表格数量以计算结果为准，另含 1 个汇总 Sheet。"
            msg_parts = [f"断面汇总表已生成（{fmt_name}）：\n{fp}\n{extra}"]
            if mat_summary:
                msg_parts.append(f"倒虹吸管道材质：{mat_summary}")
            if pp_mat_summary:
                msg_parts.append(f"有压管道材质：{pp_mat_summary}")
            msg_parts.append("\n是否立即打开该文件？")
            if fluent_question(self, "完成", "\n".join(msg_parts), yes_text="打开", no_text="关闭"):
                try:
                    os.startfile(fp)
                except Exception:
                    pass
            self.accept()
        except PermissionError:
            self.unsetCursor()
            fluent_error(self, "文件被占用",
                         f"无法写入文件，该文件可能已被其他程序打开：\n\n{fp}\n\n"
                         f"请先关闭该文件，然后重新操作。")
        except Exception as e:
            self.unsetCursor()
            import traceback; traceback.print_exc()
            fluent_error(self, "生成失败", f"错误: {e}")


def open_section_summary_table(panel):
    """打开断面汇总表生成器（纯 PySide6 对话框）"""
    nodes, nodes_source = _resolve_section_summary_source_nodes(
        panel,
        fallback_nodes=getattr(panel, "calculated_nodes", None) or [],
    )
    _record_section_summary_runtime_debug(panel, nodes, nodes_source)
    if not nodes:
        fluent_info(panel.window(), "警告", "没有数据可用，请先执行计算。")
        return

    try:
        proj_settings = panel._build_settings()
    except Exception:
        proj_settings = panel._settings

    try:
        try:
            ch_name = panel.channel_name_edit.text().strip()
            ch_level = panel.channel_level_combo.currentText()
            auto_name = f"{ch_name}{ch_level}_断面汇总表.xlsx"
        except Exception:
            auto_name = "断面汇总表.xlsx"

        dlg = SectionSummaryDialog(panel.window(), nodes, proj_settings, auto_name, panel=panel)
        dlg.exec()
    except ImportError as e:
        fluent_error(
            panel.window(), "功能不可用",
            f"断面汇总表模块加载失败：\n{str(e)}")
    except Exception as e:
        import traceback; traceback.print_exc()
        fluent_error(panel.window(), "打开失败",
                     f"断面汇总表生成器打开失败：\n{str(e)}")

