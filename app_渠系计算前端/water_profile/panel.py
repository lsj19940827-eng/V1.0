# -*- coding: utf-8 -*-
"""
推求水面线面板 —— QWidget 版本

功能：
1. 基础设置（渠道名称、级别、起始水位、流量等）
2. 节点数据表格（输入/编辑/断面计算同步）
3. 调用核心计算引擎进行水面线推求
4. 结果展示（结果表格 + 详细过程）
5. 导出Excel/Word

与断面计算子流程的数据交互：
- 断面批量计算结果通过 SharedDataManager 同步到水面线计算表
- 同步后自动填充断面参数（底宽、水深、糙率等）
"""

import sys
import os
import math
import re
import copy
import datetime
from contextlib import contextmanager

_pkg_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _pkg_root not in sys.path:
    sys.path.insert(0, _pkg_root)

# 推求水面线模块路径
_water_profile_dir = os.path.join(_pkg_root, '推求水面线')
if _water_profile_dir not in sys.path:
    sys.path.insert(0, _water_profile_dir)

# calc_渠系计算算法内核路径
_calc_dir = os.path.join(_pkg_root, 'calc_渠系计算算法内核')
if _calc_dir not in sys.path:
    sys.path.insert(0, _calc_dir)

import 推求水面线.utils  # noqa: F401  # 建立顶层 utils 兼容别名，避免打包环境同名包冲突

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QGroupBox,
    QSplitter, QFrame, QTabWidget, QTextEdit, QFileDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QAbstractItemView, QGridLayout, QFormLayout, QSizePolicy,
    QDialog, QDialogButtonBox, QToolTip, QCheckBox, QApplication
)
from PySide6.QtCore import Qt, QByteArray, Signal, QTimer, QRect, QPoint, QEvent, QObject, QSignalBlocker
from PySide6.QtGui import QFont, QColor, QPixmap, QImage, QShortcut, QKeySequence, QCursor, QBrush

try:
    from PySide6.QtWidgets import QDoubleSpinBox
except ImportError:
    class QDoubleSpinBox(QWidget):
        """兼容精简 Qt stub 的占位控件；真实运行环境会使用原生 QDoubleSpinBox。"""

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._value = 0.0

        def setDecimals(self, *_args, **_kwargs):
            pass

        def setRange(self, *_args, **_kwargs):
            pass

        def setSingleStep(self, *_args, **_kwargs):
            pass

        def setSuffix(self, *_args, **_kwargs):
            pass

        def setToolTip(self, *_args, **_kwargs):
            pass

        def setValue(self, value):
            try:
                self._value = float(value)
            except (TypeError, ValueError):
                self._value = 0.0

        def value(self):
            return float(self._value)

        def setEnabled(self, *_args, **_kwargs):
            pass

try:
    from qfluentwidgets import (
        PushButton, PrimaryPushButton, LineEdit, ComboBox,
        InfoBar, InfoBarIcon, InfoBarPosition, DropDownPushButton, RoundMenu, Action, MessageBox
    )
except ImportError:
    from qfluentwidgets import (
        PushButton, PrimaryPushButton, LineEdit, ComboBox,
        InfoBar, InfoBarPosition, DropDownPushButton, RoundMenu, Action, MessageBox
    )

    class _InfoBarIconFallback:
        SUCCESS = None

    InfoBarIcon = _InfoBarIconFallback()

from app_渠系计算前端.frozen_table import FrozenColumnTableWidget
from app_渠系计算前端.styles import P, S, W, E, BG, CARD, BD, T1, T2, auto_resize_table, CollapsibleGroupBox, fluent_info, fluent_question, DIALOG_STYLE
from app_渠系计算前端.export_utils import (
    WORD_EXPORT_AVAILABLE, ask_open_file,
    create_styled_doc, doc_add_h1, doc_add_h2, doc_add_body,
    doc_render_calc_text, doc_add_param_table, doc_add_result_table,
    doc_add_styled_table, doc_add_table_caption,
    create_engineering_report_doc, doc_add_eng_h, doc_add_eng_body,
    doc_render_calc_text_eng, update_doc_toc_via_com,
)
from app_渠系计算前端.report_meta import (
    ExportConfirmDialog, build_calc_purpose, REFERENCES_BASE, load_meta
)
from app_渠系计算前端.structure_type_selector import StructureTypeSelector
from app_渠系计算前端.case_manager import FlowLayout as _FlowLayout
from app_渠系计算前端.batch.panel import BatchPanel, format_station_display, parse_station_input
from app_渠系计算前端.debug_utils import debug_print
try:
    from utils.pressure_pipe_result_helpers import (
        make_pressure_pipe_identity,
        empty_pressure_pipe_calc_records,
        normalize_pressure_pipe_calc_records,
        format_pressure_pipe_record_detail,
        format_pressure_pipe_chain_summary,
        append_pressure_pipe_calc_batch_text,
        build_pressure_pipe_transition_note,
    )
except ImportError:
    from utils.pressure_pipe_result_helpers import (
        make_pressure_pipe_identity,
        empty_pressure_pipe_calc_records,
        normalize_pressure_pipe_calc_records,
        format_pressure_pipe_record_detail,
        format_pressure_pipe_chain_summary,
        append_pressure_pipe_calc_batch_text,
    )

    def build_pressure_pipe_transition_note(*_args, **_kwargs):
        """兼容精简测试桩，缺少新工具函数时返回空备注。"""
        return ""

try:
    from utils.pressure_pipe_result_helpers import split_pressure_pipe_records
except (ImportError, AttributeError):
    def split_pressure_pipe_records(records):
        """兼容旧测试桩：按默认回写口径拆分正式、参考和失败记录。"""
        writeback_records = []
        reference_records = []
        failed_records = []
        for record in records or []:
            if not isinstance(record, dict):
                continue
            if record.get("status") != "success":
                failed_records.append(record)
            elif record.get("writeback_enabled", True):
                writeback_records.append(record)
            else:
                reference_records.append(record)
        return {
            "writeback": writeback_records,
            "reference": reference_records,
            "failed": failed_records,
        }

# 核心计算引擎
try:
    from models.data_models import ChannelNode, ProjectSettings, TransitionLengthRule
    from models.enums import StructureType, InOutType
    from core.calculator import WaterProfileCalculator
    CALCULATOR_AVAILABLE = True
except ImportError as _e:
    print(f"[水面线] 核心计算引擎加载失败: {_e}")
    CALCULATOR_AVAILABLE = False
    TransitionLengthRule = None

# 共享数据管理器
try:
    from shared.shared_data_manager import (
        get_shared_data_manager,
        normalize_section_type_name,
    )
    SHARED_DATA_AVAILABLE = True
except ImportError:
    SHARED_DATA_AVAILABLE = False

    def normalize_section_type_name(section_type):
        return str(section_type or "").strip()

# 配置常量
try:
    from config.constants import (
        STRUCTURE_TYPE_OPTIONS, CHANNEL_LEVEL_OPTIONS, XXPIPE_CHANNEL_LEVEL_OPTIONS,
        DEFAULT_ROUGHNESS, DEFAULT_SIPHON_ROUGHNESS, DEFAULT_TURN_RADIUS, DEFAULT_AUTO_TURN_RADIUS, DEFAULT_SIPHON_TURN_RADIUS_N, DEFAULT_GATE_HEAD_LOSS,
        TRANSITION_FORM_OPTIONS, SIPHON_TRANSITION_FORM_OPTIONS,
        TRANSITION_ZETA_COEFFICIENTS, SIPHON_TRANSITION_ZETA_COEFFICIENTS,
        ZERO_TOLERANCE,
    )
except ImportError:
    STRUCTURE_TYPE_OPTIONS = [
        "明渠-梯形", "明渠-矩形", "明渠-圆形", "明渠-U形",
        "渡槽-U形", "渡槽-矩形",
        "隧洞-圆形", "隧洞-圆弧直墙型", "隧洞-马蹄形Ⅰ型", "隧洞-马蹄形Ⅱ型",
        "矩形暗涵", "倒虹吸", "有压管道", "定向钻", "顶管",
        "分水闸", "分水口", "节制闸", "泄水闸", "退水闸",
    ]
    CHANNEL_LEVEL_OPTIONS = ["总干渠", "总干管", "分干渠", "分干管", "干渠", "干管", "支渠", "支管", "分支渠", "分支管", "充水渠", "泄水渠"]
    DEFAULT_ROUGHNESS = 0.014
    DEFAULT_SIPHON_ROUGHNESS = 0.014
    DEFAULT_TURN_RADIUS = 100.0
    DEFAULT_AUTO_TURN_RADIUS = 20.0
    DEFAULT_SIPHON_TURN_RADIUS_N = 3.0
    DEFAULT_GATE_HEAD_LOSS = 0.1
    TRANSITION_FORM_OPTIONS = ["曲线形反弯扭曲面", "直线形扭曲面", "圆弧直墙", "八字形", "直角形"]
    SIPHON_TRANSITION_FORM_OPTIONS = ["反弯扭曲面", "直线扭曲面", "1/4圆弧", "方头型"]
    XXPIPE_CHANNEL_LEVEL_OPTIONS = ["总干管", "分干管", "干管", "支管", "分支管"]
    TRANSITION_ZETA_COEFFICIENTS = {
        "进口": {"曲线形反弯扭曲面": 0.1, "圆弧直墙": 0.2, "八字形": 0.3, "直角形": 0.4},
        "出口": {"曲线形反弯扭曲面": 0.2, "圆弧直墙": 0.5, "八字形": 0.5, "直角形": 0.75},
    }
    SIPHON_TRANSITION_ZETA_COEFFICIENTS = {
        "进口": {"反弯扭曲面": 0.10, "直线扭曲面": 0.20, "1/4圆弧": 0.15, "方头型": 0.30},
        "出口": {"反弯扭曲面": 0.20, "直线扭曲面": 0.40, "1/4圆弧": 0.25, "方头型": 0.75},
    }
    ZERO_TOLERANCE = 1e-9

# 节点表列定义（与原版Tkinter ALL_COLUMNS保持完全一致的列顺序）
# 可编辑列索引集合（基础输入0-7 + 水力输入20-26 + 预留/过闸/倒虹吸或有压管道损失36,37,38）
EDITABLE_COLS = set(range(8)) | {20, 21, 22, 23, 24, 25, 26, 32, 36, 37, 38}
# 表1同步来源行锁定列（需要回到表1修改后重同步）
TABLE1_SOURCE_LOCKED_COLS = set(range(8)) | set(range(20, 27))
TRANSITION_PREPARATION_RELEVANT_COLS = set(range(8)) | set(range(20, 27))
# 第一行（水位起点）锁定的水头损失列：初始水位是用户输入的定值，不受水头损失影响
FIRST_ROW_LOCKED_LOSS_COLS = {36, 37, 38}
PRESSURE_PIPE_LIKE_STRUCTURE_TEXTS = {"有压管道", "定向钻", "顶管"}
CULVERT_FAMILY_TYPE_KEY = "culvert_family_type"
ARCH_CULVERT_THETA_ROLE_KEY = "_arch_culvert_theta_deg"
ARCH_CULVERT_SOURCE_ROLE_KEY = "_arch_culvert_source_allowed"
ARCH_CULVERT_H_STRAIGHT_ROLE_KEY = "_arch_culvert_H_straight"
ARCH_CULVERT_MANUAL_H_STRAIGHT_ROLE_KEY = "_arch_culvert_manual_H_straight"
ARCH_CULVERT_USED_MANUAL_H_STRAIGHT_ROLE_KEY = "_arch_culvert_used_manual_H_straight"
TUNNEL_ARCH_THETA_ROLE_KEY = "_tunnel_arch_theta_deg"
TUNNEL_ARCH_H_STRAIGHT_ROLE_KEY = "_tunnel_arch_H_straight"
TUNNEL_ARCH_MANUAL_H_STRAIGHT_ROLE_KEY = "_tunnel_arch_manual_H_straight"
TUNNEL_ARCH_USED_MANUAL_H_STRAIGHT_ROLE_KEY = "_tunnel_arch_used_manual_H_straight"
RECT_CULVERT_FAMILY_TEXT = "暗涵-矩形"
ARCH_CULVERT_FAMILY_TEXT = "暗涵-圆拱直墙型"

NODE_ALL_HEADERS = [
    # 基础输入列 (0-7) — 对应Tkinter INPUT_COLUMNS
    "流量段", "建筑物名称", "结构形式", "进出口判断", "IP",
    "X", "Y", "转弯半径",
    # 几何结果列 (8-19) — 对应Tkinter GEOMETRY_RESULT_COLUMNS
    "转角", "切线长", "弧长", "弯道长度", "IP直线间距",
    "IP点桩号", "弯前BC", "里程MC", "弯末EC",
    "复核弯前长度", "复核弯后长度", "复核总长度",
    # 水力输入列 (20-26) — 对应Tkinter HYDRAULIC_INPUT_COLUMNS
    "底宽B", "直径D", "半径R", "边坡系数m", "糙率n", "底坡1/i", "流量Q设计",
    # 水力结果列 (27-31) — 对应Tkinter HYDRAULIC_RESULT_COLUMNS
    "水深h设计", "过水断面面积A", "湿周X", "水力半径R", "流速v设计",
    # 水头损失列 (32-40) — 对应Tkinter HEAD_LOSS_COLUMNS
    "渐变段长度L", "渐变段水头损失", "弯道水头损失", "沿程水头损失",
    "预留水头损失", "过闸水头损失", "倒虹吸/有压管道水头损失",
    "总水头损失", "累计总水头损失",
    # 高程列 (41-43) — 对应Tkinter ELEVATION_COLUMNS
    "水位", "渠底高程", "渠顶高程",
]

# 导出Excel时使用的表头（与NODE_ALL_HEADERS一致）
NODE_EXPORT_HEADERS = NODE_ALL_HEADERS
CURVE_CHECK_COLUMN_LABELS = {
    17: "复核弯前长度",
    18: "复核弯后长度",
    19: "复核总长度",
}
CURVE_CHECK_ATTRS = (
    ("check_pre_curve", "复核弯前长度"),
    ("check_post_curve", "复核弯后长度"),
    ("check_total_length", "复核总长度"),
)
CURVE_CHECK_NEGATIVE_COLOR = "#C62828"
CURVE_CHECK_NOTICE_LIMIT = 5

# 节点数据表工具栏布局预设：
# compact（紧凑）/ balanced（平衡，默认）/ comfortable（宽松）
NODE_TOOLBAR_LAYOUT_PRESET = "balanced"

SOURCE_COORD_X_ROLE_KEY = "_source_x_text"
SOURCE_COORD_Y_ROLE_KEY = "_source_y_text"
FLAT_BOTTOM_TUNNEL_SOURCE_ROLE_KEY = "_flat_bottom_tunnel_source_allowed"
USE_INCREASE_ROLE_KEY = "_use_increase"
PRESSURE_PIPE_ROW_ID_ROLE_KEY = "_pressure_pipe_row_identity"
PRESSURE_PIPE_WINDOW_OVERRIDE_ROLE_KEY = "_pressure_pipe_window_override"
PRESSURE_PIPE_NAMED_GROUP_RESULT_ROLE_KEY = "_pressure_pipe_named_group_result"
PRESSURE_PIPE_LOSS_OVERRIDE_ROLE_KEY = "_pressure_pipe_loss_override_m"
GATE_HEAD_LOSS_USER_SET_ROLE_KEY = "_gate_head_loss_user_set"
GATE_HEAD_LOSS_USER_SET_PARAM_KEY = "gate_head_loss_user_set"
COMPOUND_TRAPEZOID_PARAMS_ROLE_KEY = "_compound_trapezoid_params"
TRANSITION_LENGTH_RULE_STEP_DEFAULT = 1.0
TRANSITION_LENGTH_RULE_MODE_OPTIONS = (
    ("公式值", "formula"),
    ("向上修约", "step_up"),
    ("固定值", "fixed"),
)
COMPOUND_TRAPEZOID_PARAM_KEYS = ("m1", "B1", "m2", "B2", "m3", "h1")


def normalize_culvert_family_type_name(structure_type) -> str:
    """统一暗涵家族子类型名称。"""
    text = normalize_section_type_name(structure_type)
    if not text:
        return ""
    if "暗涵" not in text and "暗渠" not in text and text != "矩形暗涵":
        return ""
    if "圆拱直墙" in text or "圆弧直墙" in text:
        return ARCH_CULVERT_FAMILY_TEXT
    return RECT_CULVERT_FAMILY_TEXT


def resolve_node_effective_structure_type_text(node) -> str:
    """读取节点的有效结构类型，暗涵优先返回家族子类型。"""
    section_params = getattr(node, "section_params", None)
    if isinstance(section_params, dict):
        normalized_family = normalize_culvert_family_type_name(
            section_params.get(CULVERT_FAMILY_TYPE_KEY, "")
        )
        if normalized_family:
            return normalized_family
    getter = getattr(node, "get_structure_type_str", None)
    if callable(getter):
        try:
            text = str(getter() or "").strip()
            normalized_family = normalize_culvert_family_type_name(text)
            if normalized_family:
                return normalized_family
            if text:
                return text
        except Exception:
            pass
    struct_type = getattr(node, "structure_type", None)
    value = getattr(struct_type, "value", struct_type)
    return str(value or "").strip()


for _culvert_option in (RECT_CULVERT_FAMILY_TEXT, ARCH_CULVERT_FAMILY_TEXT):
    if _culvert_option not in STRUCTURE_TYPE_OPTIONS:
        STRUCTURE_TYPE_OPTIONS.append(_culvert_option)


def normalize_use_increase_flag(value, default=True):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if not text:
        return default
    if text in {"true", "1", "yes", "y", "on"}:
        return True
    if text in {"false", "0", "no", "n", "off"}:
        return False
    return bool(value)


def normalize_compound_trapezoid_params(source):
    """统一整理复式梯形隐藏参数，避免在表格刷新时丢失。"""
    if not isinstance(source, dict):
        return {}

    params = {}
    for key in COMPOUND_TRAPEZOID_PARAM_KEYS:
        value = source.get(key, None)
        if value is None or str(value).strip() == "":
            continue
        try:
            params[key] = float(value)
        except (TypeError, ValueError):
            continue
    return params


def parse_flow_values_text(flow_text: str) -> list:
    """把逗号分隔流量文本解析为浮点列表。"""
    if not flow_text or not str(flow_text).strip():
        return []
    text = str(flow_text).replace("，", ",")
    values = []
    for item in text.split(","):
        piece = str(item or "").strip()
        if not piece:
            continue
        try:
            values.append(float(piece))
        except (TypeError, ValueError):
            continue
    return values


def format_flow_value(value) -> str:
    """统一格式化单个流量值。"""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    return f"{number:.3f}".rstrip("0").rstrip(".")


def format_flow_values_text(values) -> str:
    """把流量列表格式化为兼容旧链路的逗号文本。"""
    return ", ".join(format_flow_value(value) for value in list(values or []) if format_flow_value(value))


def format_flow_segment_label(index: int) -> str:
    """把流量段索引转换成展示名称。"""
    chinese_digits = {
        1: "一",
        2: "二",
        3: "三",
        4: "四",
        5: "五",
        6: "六",
        7: "七",
        8: "八",
        9: "九",
        10: "十",
    }
    order = int(index) + 1
    label = chinese_digits.get(order, str(order))
    return f"第{label}流量段"


def calculate_max_flow_values(design_flows) -> list:
    """按设计流量规则计算对应的加大流量。"""
    max_flows = []
    for q in list(design_flows or []):
        try:
            q_value = float(q)
        except (TypeError, ValueError):
            continue
        if q_value <= 0:
            max_flows.append(0.0)
            continue
        if q_value < 1:
            pct = 30
        elif q_value < 5:
            pct = 25
        elif q_value < 20:
            pct = 20
        elif q_value < 50:
            pct = 15
        elif q_value < 100:
            pct = 10
        else:
            pct = 5
        max_flows.append(round(q_value * (1 + pct / 100), 3))
    return max_flows


def calculate_final_max_flow_values(design_flows, preferred_max_flows=None) -> list:
    """优先使用已有加大流量，缺失时再按设计流量规则补齐。"""
    final_max_flows = []
    preferred_values = list(preferred_max_flows or [])
    auto_values = calculate_max_flow_values(design_flows)
    for index, auto_value in enumerate(auto_values):
        preferred_value = preferred_values[index] if index < len(preferred_values) else None
        try:
            preferred_number = float(preferred_value)
        except (TypeError, ValueError):
            preferred_number = 0.0
        if preferred_number > 0:
            final_max_flows.append(round(preferred_number, 3))
            continue
        final_max_flows.append(auto_value)
    return final_max_flows


class FlowSegmentSelectorField(QWidget):
    """只读流量段选择控件：主界面只负责查看和切换当前流量段。"""

    editingFinished = Signal()
    segmentSelected = Signal(int)

    def __init__(self, placeholder_text="未设置流量", parent=None):
        super().__init__(parent)
        self._placeholder_text = str(placeholder_text or "未设置流量")
        self._raw_text = ""
        self._values = []
        self._current_index = 0

        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        self._summary_btn = DropDownPushButton(self._placeholder_text)
        self._summary_btn.setMinimumWidth(180)
        self._summary_btn.setFixedHeight(36)
        self._summary_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self._summary_btn.setStyleSheet(
            """
            DropDownPushButton {
                background: #ffffff;
                border: 1px solid #d7e3f7;
                border-radius: 8px;
                padding: 0 12px;
                text-align: left;
                font-size: 13px;
                font-weight: 600;
                color: #1f4f96;
            }
            DropDownPushButton:hover {
                background: #f7fbff;
                border-color: #aac3ec;
            }
            DropDownPushButton:pressed {
                background: #eef5ff;
            }
            DropDownPushButton:disabled {
                color: #94a3b8;
                background: #f8fafc;
                border-color: #d7dee8;
            }
            """
        )
        layout.addWidget(self._summary_btn, 1)

        self._inline_hint_label = QLabel("")
        self._inline_hint_label.setAlignment(Qt.AlignCenter)
        self._inline_hint_label.setMinimumHeight(24)
        self._inline_hint_label.setStyleSheet(
            "QLabel {"
            "  min-width: 44px;"
            "  padding: 0 8px;"
            "  border: 1px solid #d7e3f7;"
            "  border-radius: 8px;"
            "  background: #f5f9ff;"
            "  color: #6d84a3;"
            "  font-size: 11px;"
            "}"
        )
        self._inline_hint_label.hide()
        layout.addWidget(self._inline_hint_label, 0, Qt.AlignVCenter)

        self._refresh_menu_and_summary()

    def text(self):
        """返回兼容旧逻辑的原始流量文本。"""
        return self._raw_text

    def setText(self, text):
        """设置原始文本并刷新展示。"""
        self._raw_text = str(text or "").strip()
        self._values = parse_flow_values_text(self._raw_text)
        max_index = max(len(self._values) - 1, 0)
        self._current_index = min(max(self._current_index, 0), max_index)
        self._refresh_menu_and_summary()

    def clear(self):
        """兼容旧逻辑的清空接口。"""
        self.setText("")

    def setPlaceholderText(self, text):
        """兼容旧控件占位文案接口。"""
        self._placeholder_text = str(text or "").strip() or "未设置流量"
        self._refresh_menu_and_summary()

    def values(self) -> list:
        """返回当前解析后的流量列表。"""
        return list(self._values)

    def set_current_segment_index(self, index: int):
        """切换当前显示的流量段。"""
        if not self._values:
            self._current_index = 0
        else:
            self._current_index = min(max(int(index), 0), len(self._values) - 1)
        self._refresh_summary()

    def summary_text(self) -> str:
        """返回当前摘要文字，供测试和序列化校验使用。"""
        return self._summary_btn.text()

    def count_text(self) -> str:
        """兼容旧测试入口：不再对外暴露独立的第二行弱提示。"""
        return ""

    def _refresh_menu_and_summary(self):
        """刷新下拉菜单和当前摘要。"""
        menu = RoundMenu(parent=self)
        for index, value in enumerate(self._values):
            label = f"{format_flow_segment_label(index)}：{format_flow_value(value)}"
            menu.addAction(
                Action(
                    label,
                    triggered=lambda checked=False, idx=index: self._on_segment_action_triggered(idx),
                )
            )
        self._summary_btn.setMenu(menu)
        self._refresh_summary()

    def _refresh_summary(self):
        """刷新按钮摘要和字段内轻提示。"""
        if not self._values:
            self._summary_btn.setEnabled(False)
            self._summary_btn.setText(self._placeholder_text)
            self._summary_btn.setToolTip("")
            self._inline_hint_label.setText("")
            self._inline_hint_label.hide()
            self.updateGeometry()
            return

        self._summary_btn.setEnabled(True)
        current_index = min(max(self._current_index, 0), len(self._values) - 1)
        label = format_flow_segment_label(current_index)
        value = format_flow_value(self._values[current_index])
        self._summary_btn.setText(f"{label} · {value}")
        self._summary_btn.setToolTip(f"{label}：{value} m³/s")
        # 主界面不再展示“x段”提示，避免首次切换多流量段时触发布局突增。
        self._inline_hint_label.setText("")
        self._inline_hint_label.hide()
        self.updateGeometry()

    def _on_segment_action_triggered(self, index: int):
        """处理下拉菜单选中的流量段。"""
        self.set_current_segment_index(index)
        self.segmentSelected.emit(int(index))


# ================================================================
# 渐变段参考系数表对话框（表K.1.2 + 表L.1.2）
# ================================================================

# 表K.1.2 数据
_K12_HEADERS = ["渐变段形式", "示意图", "进口ξ₁", "出口ξ₂"]
_K12_ROWS = [
    ["曲线形反弯扭曲面", "", "0.10", "0.20"],
    ["直线形扭曲面", "", "θ₁=15°~37°；ξ₁=0.05~0.30", "θ₂=10°~17°；ξ₂=0.30~0.50"],
    ["圆弧直墙", "", "0.20", "0.50"],
    ["八字形", "", "0.30", "0.50"],
    ["直角形", "", "0.40", "0.75"],
]

# 表L.1.2 数据
_L12_HEADERS = ["渐变段形式", "ξ₁", "ξ₂", "适用条件"]
_L12_ROWS = [
    ["反弯扭曲面", "0.10", "0.20", "θ₁,θ₂均≤12.5°"],
    ["1/4圆弧", "0.15", "0.25", "θ₁,θ₂均≤12.5°"],
    ["方头型", "0.30", "0.75", "θ₁,θ₂均≤12.5°"],
    ["直线扭曲面", "0.05~0.30", "0.30~0.50", "θ₁=15°~37°，θ₂=10°~17°"],
]


class TransitionReferenceDialog(QDialog):
    """渐变段参考系数表对话框（表K.1.2 + 表L.1.2）"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("渐变段局部损失系数参考表")
        self.setMinimumSize(780, 620)
        self.resize(820, 680)
        self._pixmap_refs = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # ── 表K.1.2 ──
        k12_label = QLabel("表K.1.2  进、出口水头损失系数（渡槽/隧洞渐变段）")
        k12_label.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        layout.addWidget(k12_label)

        self.k12_table = QTableWidget(len(_K12_ROWS), len(_K12_HEADERS))
        self.k12_table.setHorizontalHeaderLabels(_K12_HEADERS)
        self.k12_table.verticalHeader().setVisible(False)
        self.k12_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.k12_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.k12_table.horizontalHeader().setStretchLastSection(False)
        self.k12_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 加载示意图缩略图
        thumb_map = self._load_k12_thumbnails()

        for r, row_data in enumerate(_K12_ROWS):
            for c, val in enumerate(row_data):
                if c == 1:
                    # 示意图列
                    if r in thumb_map:
                        img_label = QLabel()
                        img_label.setPixmap(thumb_map[r])
                        img_label.setAlignment(Qt.AlignCenter)
                        img_label.setCursor(Qt.PointingHandCursor)
                        img_label.setToolTip("点击放大查看")
                        form_name = row_data[0]
                        img_label.mousePressEvent = lambda e, fn=form_name: self._show_k12_image(fn)
                        self.k12_table.setCellWidget(r, c, img_label)
                    else:
                        item = QTableWidgetItem("(无图)")
                        item.setTextAlignment(Qt.AlignCenter)
                        self.k12_table.setItem(r, c, item)
                else:
                    item = QTableWidgetItem(val)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.k12_table.setItem(r, c, item)

        for r in range(len(_K12_ROWS)):
            self.k12_table.setRowHeight(r, 60 if r in thumb_map else 32)

        self.k12_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        k12_h = sum(60 if r in thumb_map else 32 for r in range(len(_K12_ROWS))) + self.k12_table.horizontalHeader().height() + 4
        self.k12_table.setFixedHeight(k12_h)
        layout.addWidget(self.k12_table)

        # K.1.2 注释
        k12_note = QLabel("注：表中 θ₁ 表示进口渐变段水面收缩角；θ₂ 表示出口渐变段水面扩散角。点击示意图可放大查看。")
        k12_note.setWordWrap(True)
        k12_note.setStyleSheet("color: #555; font-size: 12px;")
        layout.addWidget(k12_note)

        # ── 表L.1.2 ──
        l12_label = QLabel("表L.1.2  渐变段局部损失系数表（倒虹吸渐变段）")
        l12_label.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        layout.addWidget(l12_label)

        self.l12_table = QTableWidget(len(_L12_ROWS), len(_L12_HEADERS))
        self.l12_table.setHorizontalHeaderLabels(_L12_HEADERS)
        self.l12_table.verticalHeader().setVisible(False)
        self.l12_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.l12_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.l12_table.horizontalHeader().setStretchLastSection(False)
        self.l12_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        for r, row_data in enumerate(_L12_ROWS):
            for c, val in enumerate(row_data):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                self.l12_table.setItem(r, c, item)

        self.l12_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        l12_h = sum(self.l12_table.rowHeight(r) for r in range(len(_L12_ROWS))) + self.l12_table.horizontalHeader().height() + 4
        self.l12_table.setFixedHeight(l12_h)
        layout.addWidget(self.l12_table)

        # L.1.2 注释
        l12_note = QLabel("注：θ₁ 为水面收敛角，θ₂ 为水面扩散角（灌排规范附录表L.1.2）")
        l12_note.setStyleSheet("color: #555; font-size: 12px;")
        layout.addWidget(l12_note)

        # 关闭按钮
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok)
        btn_box.button(QDialogButtonBox.Ok).setText("关闭")
        btn_box.accepted.connect(self.accept)
        layout.addWidget(btn_box)

    def _load_k12_thumbnails(self):
        """加载K.1.2示意图缩略图，返回 {行索引: QPixmap}"""
        thumb_map = {}
        try:
            from shared.k12_images_data import get_k12_image_bytes
        except ImportError:
            try:
                from 推求水面线.shared.k12_images_data import get_k12_image_bytes
            except ImportError:
                return thumb_map

        thumb_w, thumb_h = 150, 55
        for ri, row_data in enumerate(_K12_ROWS):
            form_name = row_data[0]
            try:
                img_bytes = get_k12_image_bytes(form_name)
                if not img_bytes:
                    continue
                qimg = QImage()
                qimg.loadFromData(QByteArray(img_bytes))
                if qimg.isNull():
                    continue
                pm = QPixmap.fromImage(qimg)
                pm = pm.scaled(thumb_w, thumb_h, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self._pixmap_refs.append(pm)
                thumb_map[ri] = pm
            except Exception:
                pass
        return thumb_map

    def _show_k12_image(self, form_name):
        """放大显示K.1.2示意图"""
        try:
            from shared.k12_images_data import get_k12_image_bytes
        except ImportError:
            try:
                from 推求水面线.shared.k12_images_data import get_k12_image_bytes
            except ImportError:
                return

        img_bytes = get_k12_image_bytes(form_name)
        if not img_bytes:
            return

        qimg = QImage()
        qimg.loadFromData(QByteArray(img_bytes))
        if qimg.isNull():
            return

        pm = QPixmap.fromImage(qimg)

        dlg = QDialog(self.window())
        dlg.setWindowTitle(f"K.1.2 示意图 — {form_name}")
        dlg.setStyleSheet(DIALOG_STYLE)
        lay = QVBoxLayout(dlg)

        # 等比缩放到合适大小
        max_w, max_h = 800, 500
        if pm.width() > max_w or pm.height() > max_h:
            pm = pm.scaled(max_w, max_h, Qt.KeepAspectRatio, Qt.SmoothTransformation)

        img_label = QLabel()
        img_label.setPixmap(pm)
        img_label.setAlignment(Qt.AlignCenter)
        lay.addWidget(img_label)

        btn = QDialogButtonBox(QDialogButtonBox.Ok)
        btn.button(QDialogButtonBox.Ok).setText("关闭")
        btn.accepted.connect(dlg.accept)
        lay.addWidget(btn)

        dlg.resize(pm.width() + 40, pm.height() + 80)
        dlg.exec()


# ================================================================
class TransitionLengthRuleDialog(QDialog):
    """按当前工程组合编辑渐变段长度规则。"""

    def __init__(self, rule_rows, parent=None):
        super().__init__(parent)
        self._rule_rows = list(rule_rows or [])
        self._result_rules = {}
        self.setWindowTitle("渐变段长度规则")
        self.setStyleSheet(DIALOG_STYLE)
        self.setMinimumSize(860, 420)
        self.resize(920, 500)
        self._build_ui()

    @staticmethod
    def _format_hit_scope_text(row: dict) -> str:
        hit_count = int(row.get("hit_count", row.get("count", 0)) or 0)
        return f"当前命中 {hit_count} 条"

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        intro = QLabel(
            "仅展示表3中已插入的渐变段实例组合。可选择保留公式值、按步长向上修约，或直接指定固定长度。"
        )
        intro.setWordWrap(True)
        intro.setStyleSheet("color: #555; font-size: 12px;")
        layout.addWidget(intro)

        self.table = QTableWidget(len(self._rule_rows), 8)
        self.table.setHorizontalHeaderLabels(
            ["上游结构", "下游结构", "渐变段类型", "命中情况", "出现次数", "规则模式", "步长(m)", "固定值(m)"]
        )
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)

        self._row_editors = {}
        for row_idx, row in enumerate(self._rule_rows):
            for col_idx, key in enumerate(("upstream_structure_type", "downstream_structure_type", "transition_type")):
                item = QTableWidgetItem(str(row.get(key, "") or ""))
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, col_idx, item)

            hit_item = QTableWidgetItem(self._format_hit_scope_text(row))
            hit_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 3, hit_item)

            count_item = QTableWidgetItem(str(int(row.get("count", 0) or 0)))
            count_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 4, count_item)

            mode_combo = QComboBox(self.table)
            for label, value in TRANSITION_LENGTH_RULE_MODE_OPTIONS:
                mode_combo.addItem(label, value)
            current_mode = str(row.get("rule_mode", "formula") or "formula")
            current_index = 0
            for idx in range(mode_combo.count()):
                if str(mode_combo.itemData(idx) or "") == current_mode:
                    current_index = idx
                    break
            mode_combo.setCurrentIndex(current_index)
            self.table.setCellWidget(row_idx, 5, mode_combo)

            step_box = QDoubleSpinBox(self.table)
            step_box.setDecimals(3)
            step_box.setRange(0.0, 9999.0)
            step_box.setSingleStep(0.1)
            step_box.setSuffix(" m")
            step_box.setValue(float(row.get("step_size_m", TRANSITION_LENGTH_RULE_STEP_DEFAULT) or 0.0))
            step_box.setToolTip("仅在“向上修约”模式下生效；0 表示停用步长修约。")
            self.table.setCellWidget(row_idx, 6, step_box)

            fixed_box = QDoubleSpinBox(self.table)
            fixed_box.setDecimals(3)
            fixed_box.setRange(0.0, 9999.0)
            fixed_box.setSingleStep(0.1)
            fixed_box.setSuffix(" m")
            fixed_box.setValue(float(row.get("fixed_length_m", 0.0) or 0.0))
            fixed_box.setToolTip("仅在“固定值”模式下生效。")
            self.table.setCellWidget(row_idx, 7, fixed_box)

            self._row_editors[row_idx] = {
                "mode_combo": mode_combo,
                "step_box": step_box,
                "fixed_box": fixed_box,
            }
            mode_combo.currentIndexChanged.connect(
                lambda _idx, row_idx=row_idx: self._sync_row_mode_state(row_idx)
            )
            self._sync_row_mode_state(row_idx)

        layout.addWidget(self.table, 1)

        footer = QHBoxLayout()
        footer.addStretch(1)
        reset_formula_btn = PushButton("全部恢复公式值")
        reset_formula_btn.clicked.connect(self._reset_formula_rules)
        footer.addWidget(reset_formula_btn)
        reset_step_btn = PushButton("全部默认 1.0 m 修约")
        reset_step_btn.clicked.connect(self._reset_default_steps)
        footer.addWidget(reset_step_btn)
        layout.addLayout(footer)

        note = QLabel(
            "说明：向上修整先计算目标整数长度，再按当前物理可用长度裁剪；若当前拓扑不足，最终长度可能不是整数。"
        )
        note.setWordWrap(True)
        note.setStyleSheet("color: #6B7280; font-size: 12px;")
        layout.addWidget(note)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.button(QDialogButtonBox.Ok).setText("保存并应用")
        btn_box.button(QDialogButtonBox.Cancel).setText("取消")
        btn_box.accepted.connect(self._accept_rules)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _sync_row_mode_state(self, row_idx: int):
        editors = self._row_editors.get(row_idx, {})
        mode_combo = editors.get("mode_combo")
        step_box = editors.get("step_box")
        fixed_box = editors.get("fixed_box")
        mode_value = str(mode_combo.currentData() or "formula") if mode_combo is not None else "formula"
        if step_box is not None:
            step_box.setEnabled(mode_value == "step_up")
        if fixed_box is not None:
            fixed_box.setEnabled(mode_value == "fixed")

    def _reset_default_steps(self):
        for row_idx, editors in self._row_editors.items():
            mode_combo = editors.get("mode_combo")
            step_box = editors.get("step_box")
            fixed_box = editors.get("fixed_box")
            if mode_combo is not None:
                for idx in range(mode_combo.count()):
                    if str(mode_combo.itemData(idx) or "") == "step_up":
                        mode_combo.setCurrentIndex(idx)
                        break
            if step_box is not None:
                step_box.setValue(TRANSITION_LENGTH_RULE_STEP_DEFAULT)
            if fixed_box is not None:
                fixed_box.setValue(0.0)
            self._sync_row_mode_state(row_idx)

    def _reset_formula_rules(self):
        for row_idx, editors in self._row_editors.items():
            mode_combo = editors.get("mode_combo")
            if mode_combo is not None:
                mode_combo.setCurrentIndex(0)
            self._sync_row_mode_state(row_idx)

    def _accept_rules(self):
        result = {}
        for row_idx, row in enumerate(self._rule_rows):
            editors = self._row_editors.get(row_idx, {})
            mode_combo = editors.get("mode_combo")
            step_box = editors.get("step_box")
            fixed_box = editors.get("fixed_box")
            rule_mode = str(mode_combo.currentData() or "formula") if mode_combo is not None else "formula"
            step_size_m = float(step_box.value()) if step_box is not None else TRANSITION_LENGTH_RULE_STEP_DEFAULT
            fixed_length_m = float(fixed_box.value()) if fixed_box is not None else 0.0
            key = str(row.get("rule_key", "") or "").strip()
            if not key:
                continue
            result[key] = {
                "upstream_structure_type": str(row.get("upstream_structure_type", "") or "").strip(),
                "downstream_structure_type": str(row.get("downstream_structure_type", "") or "").strip(),
                "transition_type": str(row.get("transition_type", "") or "").strip(),
                "rule_mode": rule_mode,
                "step_size_m": step_size_m,
                "fixed_length_m": fixed_length_m,
            }
        self._result_rules = result
        self.accept()

    def get_rules(self):
        return dict(self._result_rules)


# ================================================================
class _PopupClickFilter(QObject):
    """全局鼠标点击事件过滤器：点击弹窗和按钮外部时关闭弹窗。"""

    def __init__(self, popup, parent_btn):
        super().__init__()
        self._popup = popup
        self._parent_btn = parent_btn

    def eventFilter(self, obj, event):
        if event.type() == QEvent.MouseButtonPress:
            cursor_pos = QCursor.pos()
            popup_rect = self._popup.frameGeometry()
            btn_tl = self._parent_btn.mapToGlobal(self._parent_btn.rect().topLeft())
            btn_rect = QRect(btn_tl, self._parent_btn.rect().size())
            if not popup_rect.contains(cursor_pos) and not btn_rect.contains(cursor_pos):
                self._popup.close()
        return False


# 倒虹吸糙率展示组件（Badge 徽标按钮 + 弹出详情卡片）
# ================================================================
class SiphonRoughnessChipContainer(QWidget):
    """倒虹吸糙率展示 — 简洁按钮 + 数量徽标 + 点击弹出精致详情卡片"""

    _PRIMARY = "#0078D4"
    _PRIMARY_DARK = "#005A9E"

    # 管材参数映射表（与有压管道设计.py保持一致）
    PIPE_MATERIAL_PARAMS = {
        "HDPE管": {"f": 94800, "m": 1.77, "b": 4.77},
        "玻璃钢夹砂管": {"f": 94800, "m": 1.77, "b": 4.77},
        "球墨铸铁管": {"f": 223200, "m": 1.852, "b": 4.87},
        "预应力钢筒混凝土管": {"f": 1312000, "m": 2.0, "b": 5.33},  # n=0.013
        "预应力钢筒混凝土管_n014": {"f": 1516000, "m": 2.0, "b": 5.33},  # n=0.014
        "预应力钢筒混凝土管_n015": {"f": 1749000, "m": 2.0, "b": 5.33},  # n=0.015
        "钢管": {"f": 625000, "m": 1.9, "b": 5.1},
    }

    def __init__(self, parent=None, title_text="倒虹吸糙率详情", empty_text="导入后自动显示", label_prefix="n="):
        super().__init__(parent)
        self._pairs = []  # [(名称, 糙率), ...]
        self._title_text = title_text
        self._empty_text = empty_text
        self.label_prefix = label_prefix
        self._compact_min_width = 0
        self._build_ui()

    def _build_ui(self):
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # 占位文字（无数据时显示）
        self._placeholder = QLabel(self._empty_text)
        self._placeholder.setStyleSheet("color: #555555; font-size: 12px;")
        lay.addWidget(self._placeholder)

        # 按钮（有数据时显示）
        self._btn = PushButton("点击查看")
        self._btn.setFixedHeight(28)
        from PySide6.QtWidgets import QSizePolicy
        self._btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self._btn.setStyleSheet(
            "PushButton {"
            "  font-size: 12px; padding: 3px 10px;"
            f"  border: 1px solid {self._PRIMARY}; border-radius: 6px;"
            f"  color: {self._PRIMARY}; background: white;"
            "}"
            "PushButton:hover {"
            f"  background: #E8F4FD; color: {self._PRIMARY_DARK};"
            f"  border-color: {self._PRIMARY_DARK};"
            "}"
        )
        self._btn.clicked.connect(self._show_popover)
        self._btn.setVisible(False)
        lay.addWidget(self._btn)

        # 数量徽标
        self._badge = QLabel("0")
        self._badge.setFixedSize(20, 20)
        self._badge.setAlignment(Qt.AlignCenter)
        self._badge.setStyleSheet(
            f"background: {self._PRIMARY}; color: white;"
            " border-radius: 10px; font-size: 11px; font-weight: bold;"
        )
        self._badge.setVisible(False)
        lay.addWidget(self._badge)
        self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self._apply_width_policy()

    def set_compact_min_width(self, width: int):
        """设置有数据时的紧凑最小宽度；无数据时会自动放宽以完整显示占位文案。"""
        self._compact_min_width = max(0, int(width))
        self._apply_width_policy()

    def _placeholder_min_width(self) -> int:
        return self._placeholder.fontMetrics().horizontalAdvance(self._empty_text) + 12

    def _apply_width_policy(self):
        has_data = bool(self._pairs)
        if has_data:
            target_min_w = self._compact_min_width
        else:
            target_min_w = max(self._compact_min_width, self._placeholder_min_width())
        self.setMinimumWidth(target_min_w)
        self.updateGeometry()

    def _show_popover(self):
        """点击按钮弹出糙率详情卡片。

        关闭方式：点击弹窗外部区域关闭，或再次点击按钮切换关闭。
        """
        if not self._pairs:
            return

        # 已有弹窗时切换关闭
        if hasattr(self, '_popup_win') and self._popup_win is not None:
            self._popup_win.close()
            return

        primary = self._PRIMARY

        popup = QFrame(None)
        popup.setWindowFlags(
            Qt.ToolTip | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint
        )
        popup.setAttribute(Qt.WA_DeleteOnClose)
        popup.setAttribute(Qt.WA_ShowWithoutActivating)  # 不抢父窗口焦点
        popup.setStyleSheet(
            "QFrame {"
            "  background: white;"
            "  border: 1px solid #E8ECF0;"
            "  border-radius: 12px;"
            "}"
        )

        def _on_destroyed():
            self._popup_win = None
            if hasattr(self, '_click_filter') and self._click_filter is not None:
                app = QApplication.instance()
                if app:
                    app.removeEventFilter(self._click_filter)
                self._click_filter = None

        popup.destroyed.connect(_on_destroyed)
        self._popup_win = popup

        self._click_filter = _PopupClickFilter(popup, self._btn)
        app = QApplication.instance()
        if app:
            app.installEventFilter(self._click_filter)

        card_lay = QVBoxLayout(popup)
        card_lay.setContentsMargins(0, 0, 0, 0)
        card_lay.setSpacing(0)

        # 标题栏
        header = QFrame()
        header.setObjectName("popoverHeader")
        header.setStyleSheet(
            "QFrame#popoverHeader {"
            "  background: white;"
            "  border-top-left-radius: 12px; border-top-right-radius: 12px;"
            "  border-bottom: 1px solid #F0F4F8;"
            "}"
        )
        header_lay = QHBoxLayout(header)
        header_lay.setContentsMargins(14, 11, 14, 9)
        header_lay.setSpacing(7)
        dot_lbl = QLabel("●")
        dot_lbl.setStyleSheet(f"color: {primary}; font-size: 8px; background: transparent;")
        header_lay.addWidget(dot_lbl)
        title = QLabel(self._title_text)
        title.setStyleSheet(
            "font-size: 13px; font-weight: 600; color: #2D3748; background: transparent;"
        )
        header_lay.addWidget(title)
        header_lay.addStretch()
        card_lay.addWidget(header)

        # 数据行
        for i, (name, n_val) in enumerate(self._pairs):
            row_w = QWidget()
            is_last = (i == len(self._pairs) - 1)
            radius_style = (
                "  border-bottom-left-radius: 12px; border-bottom-right-radius: 12px;"
                if is_last else ""
            )
            row_w.setStyleSheet(
                f"QWidget {{ background: white;{radius_style} }}"
                "QWidget:hover { background: #F7FAFC; }"
            )
            row_lay = QHBoxLayout(row_w)
            row_lay.setContentsMargins(14, 8, 14, 8)
            row_lay.setSpacing(16)

            lbl_name = QLabel(name)
            lbl_name.setStyleSheet("font-size: 12px; color: #4A5568; background: transparent;")

            val_lbl = QLabel(f"{self.label_prefix}{n_val}")
            val_lbl.setStyleSheet(
                f"font-size: 12px; font-weight: 700; color: {primary};"
                f" background: #EBF4FF; padding: 2px 8px;"
                f" border-radius: 10px;"
            )

            row_lay.addWidget(lbl_name)
            row_lay.addStretch()
            row_lay.addWidget(val_lbl)

            if not is_last:
                sep = QFrame()
                sep.setFrameShape(QFrame.HLine)
                sep.setStyleSheet("QFrame { background: #F7FAFC; border: none; max-height: 1px; }")
                card_lay.addWidget(row_w)
                card_lay.addWidget(sep)
            else:
                card_lay.addWidget(row_w)

        popup.setMinimumWidth(220)
        popup.adjustSize()
        # 定位：在按钮正下方，间距 0（消除鼠标真空带）
        btn_pos = self._btn.mapToGlobal(self._btn.rect().bottomLeft())
        popup.move(btn_pos.x(), btn_pos.y())
        popup.show()

    def set_pairs(self, pairs):
        """设置数据。pairs: [(名称, 糙率), ...]"""
        self._pairs = list(pairs) if pairs else []
        has_data = bool(self._pairs)
        self._placeholder.setVisible(not has_data)
        self._btn.setVisible(has_data)
        self._badge.setVisible(has_data)
        if has_data:
            n = len(self._pairs)
            self._badge.setText(str(n))
            btn_text = "点击查看"
            self._btn.setText(btn_text)
            text_w = self._btn.fontMetrics().horizontalAdvance(btn_text)
            self._btn.setMinimumWidth(max(88, text_w + 28))
            self._btn.adjustSize()
        self._apply_width_policy()

    def set_siphon_data(self, pairs):
        """兼容旧接口"""
        self.set_pairs(pairs)

    def clear(self):
        self._pairs.clear()
        self._placeholder.setVisible(True)
        self._btn.setVisible(False)
        self._badge.setVisible(False)
        self._btn.setMinimumWidth(0)
        self._apply_width_policy()

    def text(self):
        """兼容旧接口"""
        return ""


class WaterProfilePanel(QWidget):
    """推求水面线面板"""

    # 数据变化信号（用于项目管理器追踪脏状态）
    data_changed = Signal()

    def __init__(self, parent=None, siphon_manager=None, pressure_pipe_manager=None):
        super().__init__(parent)
        self._siphon_manager = siphon_manager
        self._pressure_pipe_manager = pressure_pipe_manager
        self.nodes = []
        self.calculated_nodes = []
        self._settings = None
        self.btn_pressure_pipe_calc = None
        self.btn_pressure_pipe_water_hammer = None
        self._node_toolbar_layout_preset = NODE_TOOLBAR_LAYOUT_PRESET
        self._pressure_pipe_calc_done = {}
        self._pressure_pipe_calc_records = empty_pressure_pipe_calc_records()
        self._pressure_pipe_last_run_at = ""
        self._pressure_pipe_summary_dialog = None
        self._pressure_turn_radius_fallback_groups = set()
        # 建筑物长度统计缓存
        self._last_building_lengths = []
        self._last_channel_total_length = 0.0
        self._last_type_summary = []
        self._transition_length_rules = {}
        self._length_rule_nudge_seen = False
        self._length_rule_nudge_bar = None
        # 纵断面文字导出设置（记住上次使用的参数）
        self._text_export_settings = {
            'y_bottom': 1, 'y_top': 31, 'y_water': 16,
            'text_height': 3.5, 'rotation': 90, 'elev_decimals': 3,
            'station_decimals': 2,
            'xxpipe_centerline_elev_decimals': 2,
            'xxpipe_station_decimals': 2,
            'y_name': 115, 'y_slope': 105, 'y_ip': 77,
            'y_station': 47, 'y_line_height': 120,
            'scale_x': 2000, 'scale_y': 1000,
            'profile_row_items': [
                {"id": "building_name", "enabled": True},
                {"id": "slope", "enabled": True},
                {"id": "ip_name", "enabled": True},
                {"id": "station", "enabled": True},
                {"id": "top_elev", "enabled": True},
                {"id": "water_elev", "enabled": True},
                {"id": "bottom_elev", "enabled": True},
                {"id": "bd_ip_before", "enabled": False},
                # {"id": "be_ip_text", "enabled": False},  # 暂停：与 IP点名称 重复
                {"id": "bf_ip_after", "enabled": False},
                {"id": "bj_station_before", "enabled": False},
                # {"id": "bk_station", "enabled": False},  # 暂停：与 里程桩号 重复
                {"id": "bl_station_after", "enabled": False},
            ],
        }
        # 缓存每行的结构高度（structure_height）
        # 该属性不显示在表格列中，但用于计算渠顶高程 = 渠底高程 + 结构高度
        # 在 _import_from_batch / _update_table_from_nodes_full 时存入，在 _build_nodes_from_table 时恢复
        self._node_structure_heights: dict = {}
        self._node_chamfer_params: dict = {}   # {row_idx: {'chamfer_angle': float, 'chamfer_length': float}}
        self._node_u_params: dict = {}         # {row_idx: {'theta_deg': float}}，明渠-U形的圆心角缓存
        self._node_velocity_increased: dict = {}  # {row_idx: float}，加大流速缓存（从批量计算导入）
        # 建筑物名称上平面图设置（记住上次使用的参数）
        self._plan_text_settings = {
            'offset': 10,
            'text_height': 10,
        }
        # CAD导出相关缓存（供"生成断面汇总表"与"导出全部DXF"互相复用）
        self._custom_pressurized_pipe_params = {"siphon": [], "pressure_pipe": []}
        self._custom_struct_thickness = {}
        self._custom_rock_lining = {}
        self._custom_tunnel_unified = {}
        # 防止 cellChanged 递归更新的守卫标志
        self._updating_cells = False
        # 表格编辑撤销栈（单元格编辑）
        self._loss_undo_stack = []
        self._loss_redo_stack = []
        self._pre_edit_cell_value = None  # (row, col, old_text)
        self._pre_edit_snapshot = None  # 编辑前的快照
        self._undo_group = 0  # 撤销分组计数器，用于批量操作时避免重复记录快照
        # 节点表行操作撤销栈（添加/删除/插入/复制/清空行）
        self._node_table_undo_stack = []
        self._node_table_redo_stack = []
        # 合并面板：断面批量计算后端（复用 BatchPanel 的计算与校验能力）
        self._batch_backend = BatchPanel(self)
        self._batch_backend.set_info_parent(lambda: self)
        self._batch_backend.hide()
        try:
            self._batch_backend._clear_input(force=True, clear_shared=False)
            # 这里只重置嵌入式批量面板自身状态，避免把外部刚算好的共享结果一并清空。
            self._batch_backend._clear_results(clear_shared=False)
        except Exception:
            pass
        self._section_sync_ready = False
        self._transition_topology_prepared = False
        self._section_first_success_switched = False
        self._section_failure_auto_expanded_once = False
        self._section_failure_records = []
        self._flow_segment_current_index = 0
        self._design_flow_group_widget = None
        self._max_flow_group_widget = None
        self._flow_pair_group_widget = None
        self._settings_group = None
        self._transition_group = None
        self._siphon_pressure_group = None
        self._last_layout_width = 0
        self._section_status_bar = None
        self._section_state_icon = None
        self._section_state_label = None
        self._section_status_kind = "neutral"
        self._btn_section_failure_summary = None
        self._section_failure_panel = None
        self._section_failure_table = None
        self._btn_section_failure_copy = None
        self._btn_section_failure_locate = None
        self._btn_section_failure_collapse = None
        self._btn_section_calc = None
        self._btn_section_clear = None
        self._workspace_tabs = None
        self._process_tabs = None
        self._tab_section_input = None
        self._tab_section_result = None
        self._tab_water_profile = None
        self._tab_output = None
        self._btn_section_import_excel = None
        self._btn_section_sample = None
        self._btn_section_template = None
        self._btn_section_add = None
        self._btn_section_insert = None
        self._btn_section_delete = None
        self._btn_section_copy = None
        self._btn_section_params = None
        self._btn_section_export_excel = None
        self._btn_section_export_word = None
        self._btn_table3_clear = None
        self._btn_transition = None
        self._btn_siphon = None
        self._btn_calc = None
        self._section_input_table = None
        self._section_result_table = None
        self._init_ui()

    # ================================================================
    # UI 构建
    # ================================================================
    def _init_ui(self):
        main_lay = QVBoxLayout(self)
        main_lay.setContentsMargins(8, 6, 8, 6)
        main_lay.setSpacing(4)

        self._splitter = QSplitter(Qt.Vertical)
        self._splitter.setChildrenCollapsible(False)
        main_lay.addWidget(self._splitter)

        # 上半区：设置 + 输入表
        top_w = QWidget()
        self._build_top_area(top_w)
        self._splitter.addWidget(top_w)

        # 下半区：流程工作区（表1/表2/表3/结果与导出）
        bottom_w = QWidget()
        self._build_workspace_area(bottom_w)
        self._splitter.addWidget(bottom_w)

        self._splitter.setSizes([260, 760])

    def _build_top_area(self, parent):
        lay = QVBoxLayout(parent)
        self._top_lay = lay
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(4)
        lay.setAlignment(Qt.AlignTop)

        # ────────────────────────────────────────────
        # 基础设置区（连续自适应重排，可折叠）
        # ────────────────────────────────────────────
        settings_grp = CollapsibleGroupBox("基础设置")
        settings_grp.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self._settings_group = settings_grp
        settings_content_lay = QVBoxLayout(settings_grp.content_widget())
        settings_content_lay.setContentsMargins(8, 6, 8, 6)
        settings_content_lay.setSpacing(6)
        row1_flow = _FlowLayout(spacing=8)
        row1_flow.setContentsMargins(0, 0, 0, 0)
        settings_content_lay.addLayout(row1_flow)
        row2_flow = _FlowLayout(spacing=8)
        row2_flow.setContentsMargins(0, 0, 0, 0)
        settings_content_lay.addLayout(row2_flow)

        def _make_field_group(label_text: str, widgets: list, min_w: int = 0):
            group_w = QWidget()
            group_lay = QHBoxLayout(group_w)
            group_lay.setContentsMargins(0, 0, 0, 0)
            group_lay.setSpacing(6)
            label = QLabel(label_text)
            label_w = max(label.fontMetrics().horizontalAdvance(label_text) + 8, 56)
            label.setMinimumWidth(label_w)
            label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)
            group_lay.addWidget(label)
            for widget in widgets:
                group_lay.addWidget(widget)
            group_w.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Fixed)
            if min_w > 0:
                group_w.setMinimumWidth(min_w)
            return group_w

        self.channel_name_edit = LineEdit()
        self.channel_name_edit.setText("南峰寺")
        # 给 5 个汉字名称留出稳定显示空间，利用“级别”前的空白避免截断。
        self.channel_name_edit.setFixedWidth(112)
        row1_flow.addWidget(_make_field_group("渠道名称:", [self.channel_name_edit], min_w=114))

        self.channel_level_combo = ComboBox()
        self.channel_level_combo.addItems(CHANNEL_LEVEL_OPTIONS)
        self.channel_level_combo.setCurrentText("支渠")
        self.channel_level_combo.setMinimumWidth(90)
        row1_flow.addWidget(_make_field_group("级别:", [self.channel_level_combo], min_w=150))

        self.start_wl_edit = LineEdit()
        self.start_wl_edit.setText("100.0")
        self.start_wl_edit.setMinimumWidth(62)
        self.start_wl_edit.setMaximumWidth(94)
        self.start_wl_edit.textChanged.connect(
            lambda t, e=self.start_wl_edit: e.setFixedWidth(max(62, min(94, e.fontMetrics().horizontalAdvance(t) + 34)))
        )
        self.start_wl_edit.setFixedWidth(
            max(62, min(94, self.start_wl_edit.fontMetrics().horizontalAdvance(self.start_wl_edit.text()) + 34))
        )
        row1_flow.addWidget(_make_field_group("起始水位(m):", [self.start_wl_edit], min_w=108))

        self.design_flow_edit = FlowSegmentSelectorField("未设置流量段")
        self.design_flow_edit.setText("")
        self.design_flow_edit.setMinimumWidth(196)
        self.design_flow_edit.setPlaceholderText("未设置流量段")
        self.design_flow_edit.segmentSelected.connect(self._set_flow_segment_current_index)
        self._design_flow_group_widget = _make_field_group("设计流量(m³/s):", [self.design_flow_edit], min_w=248)

        self.max_flow_edit = FlowSegmentSelectorField("自动计算")
        self.max_flow_edit.setText("")
        self.max_flow_edit.setMinimumWidth(196)
        self.max_flow_edit.setPlaceholderText("自动计算")
        self.max_flow_edit.segmentSelected.connect(self._set_flow_segment_current_index)
        self._max_flow_group_widget = _make_field_group("加大流量(m³/s):", [self.max_flow_edit], min_w=248)
        self._flow_pair_group_widget = QWidget()
        flow_pair_lay = QHBoxLayout(self._flow_pair_group_widget)
        flow_pair_lay.setContentsMargins(0, 0, 0, 0)
        flow_pair_lay.setSpacing(8)
        flow_pair_lay.addWidget(self._design_flow_group_widget)
        flow_pair_lay.addWidget(self._max_flow_group_widget)
        self._flow_pair_group_widget.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Fixed)
        self._flow_pair_group_widget.setMinimumWidth(512)
        row1_flow.addWidget(self._flow_pair_group_widget)

        self.start_station_edit = LineEdit()
        self.start_station_edit.setText("0+000.000")
        self.start_station_edit.setMinimumWidth(130)
        row2_flow.addWidget(_make_field_group("起始桩号(m):", [self.start_station_edit], min_w=230))
        self.start_station_edit.editingFinished.connect(self._format_start_station)
        self.start_station_edit.installEventFilter(self)

        turn_r_box_w = QWidget()
        turn_r_box = QHBoxLayout(turn_r_box_w)
        turn_r_box.setContentsMargins(0, 0, 0, 0)
        turn_r_box.setSpacing(6)
        self.turn_radius_edit = LineEdit()
        self.turn_radius_edit.setText("")
        self.turn_radius_edit.setPlaceholderText("待应用值")
        self.turn_radius_edit.setFixedWidth(84)
        turn_r_box.addWidget(self.turn_radius_edit)
        btn_apply_r = PushButton("应用")
        btn_apply_r.setFixedWidth(58)
        btn_apply_r.setToolTip("将当前栏位中的转弯半径统一应用到表3所有真实导入行")
        btn_apply_r.clicked.connect(self._apply_pending_turn_radius_to_source_rows)
        turn_r_box.addWidget(btn_apply_r)
        btn_auto_r = PushButton("自动")
        btn_auto_r.setFixedWidth(58)
        btn_auto_r.setToolTip(
            "根据规范自动计算推荐转弯半径（取大值原则）\n"
            "仅填入当前栏位，不会自动改写表格各行\n"
            "• 隧洞：弯曲半径≥洞径(或洞宽)×5\n"
            "• 明渠：弯曲半径≥水面宽度×5\n"
            "• 渡槽：弯道半径≥连接明渠渠底宽度×5"
        )
        btn_auto_r.clicked.connect(self._auto_calc_turn_radius)
        turn_r_box.addWidget(btn_auto_r)

        self._sync_flow_segment_widgets(reset_index=True)
        row2_flow.addWidget(_make_field_group("转弯半径(m):", [turn_r_box_w], min_w=230))

        self.roughness_edit = LineEdit()
        self.roughness_edit.setText(str(DEFAULT_ROUGHNESS))
        self.roughness_edit.setFixedWidth(78)
        self.roughness_edit.setToolTip("渠道糙率：适用于明渠、渡槽、隧洞、暗涵等非倒虹吸建筑物")
        row2_flow.addWidget(_make_field_group("渠道糙率:", [self.roughness_edit], min_w=160))

        # 倒虹吸与有压管道糙率字段放在第2行末尾，按同一显隐规则同步显示
        self.siphon_roughness_chips = SiphonRoughnessChipContainer()
        self.siphon_roughness_chips.set_compact_min_width(96)
        self.siphon_roughness_chips.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        siphon_group = _make_field_group("倒虹吸糙率:", [self.siphon_roughness_chips], min_w=180)

        self.pressure_pipe_roughness_chips = SiphonRoughnessChipContainer(
            title_text="有压管道糙率详情",
            empty_text="导入后自动显示",
            label_prefix="管材: "
        )
        self.pressure_pipe_roughness_chips.set_compact_min_width(108)
        self.pressure_pipe_roughness_chips.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        pressure_pipe_group = _make_field_group("有压管道糙率:", [self.pressure_pipe_roughness_chips], min_w=200)

        siphon_pressure_pair = QWidget()
        siphon_pressure_pair_lay = QHBoxLayout(siphon_pressure_pair)
        siphon_pressure_pair_lay.setContentsMargins(0, 0, 0, 0)
        siphon_pressure_pair_lay.setSpacing(4)
        siphon_pressure_pair_lay.addWidget(siphon_group)
        siphon_pressure_pair_lay.addWidget(pressure_pipe_group)
        siphon_pressure_pair.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Fixed)

        self._siphon_pressure_group = siphon_pressure_pair
        row2_flow.addWidget(siphon_pressure_pair)
        self._refresh_roughness_overview_visibility()

        settings_grp.toggled.connect(self._on_settings_toggled)
        lay.addWidget(settings_grp)

        # ────────────────────────────────────────────
        # 渐变段设置区（3行网格布局，按类型分行，可折叠）
        # ────────────────────────────────────────────
        trans_grp = CollapsibleGroupBox("渐变段设置")
        trans_grp.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self._transition_group = trans_grp
        tg = QGridLayout(trans_grp.content_widget())
        tg.setHorizontalSpacing(10)
        tg.setVerticalSpacing(8)
        tg.setContentsMargins(8, 8, 8, 8)

        _cat_style = f"font-weight:bold; color:{P}; font-size:12px;"
        _transition_field_h = max(34, self.fontMetrics().height() + 16)

        def _setup_transition_combo(combo: ComboBox, min_width: int):
            combo.setMinimumWidth(min_width)
            combo.setMinimumHeight(_transition_field_h)
            combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        def _setup_transition_zeta(edit: LineEdit, default_text: str):
            edit.setText(default_text)
            edit.setFixedWidth(72)
            edit.setMinimumHeight(_transition_field_h)
            edit.setAlignment(Qt.AlignCenter)
            edit.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        # 行0：渡槽/隧洞
        r = 0
        cat0 = QLabel("渡槽/隧洞")
        cat0.setStyleSheet(_cat_style)
        tg.addWidget(cat0, r, 0, Qt.AlignRight)
        tg.addWidget(QLabel("进口:"), r, 1, Qt.AlignRight)
        self.trans_inlet_combo = ComboBox()
        self.trans_inlet_combo.addItems(TRANSITION_FORM_OPTIONS)
        _setup_transition_combo(self.trans_inlet_combo, 220)
        tg.addWidget(self.trans_inlet_combo, r, 2)
        tg.addWidget(QLabel("ζ₁="), r, 3, Qt.AlignRight)
        self.trans_inlet_zeta = LineEdit()
        _setup_transition_zeta(self.trans_inlet_zeta, "0.10")
        tg.addWidget(self.trans_inlet_zeta, r, 4)
        tg.addWidget(QLabel("出口:"), r, 5, Qt.AlignRight)
        self.trans_outlet_combo = ComboBox()
        self.trans_outlet_combo.addItems(TRANSITION_FORM_OPTIONS)
        _setup_transition_combo(self.trans_outlet_combo, 220)
        tg.addWidget(self.trans_outlet_combo, r, 6)
        tg.addWidget(QLabel("ζ₂="), r, 7, Qt.AlignRight)
        self.trans_outlet_zeta = LineEdit()
        _setup_transition_zeta(self.trans_outlet_zeta, "0.20")
        tg.addWidget(self.trans_outlet_zeta, r, 8)

        # 行1：明渠
        r = 1
        cat1 = QLabel("明渠")
        cat1.setStyleSheet(_cat_style)
        tg.addWidget(cat1, r, 0, Qt.AlignRight)
        tg.addWidget(QLabel("型式:"), r, 1, Qt.AlignRight)
        self.oc_trans_combo = ComboBox()
        self.oc_trans_combo.addItems(TRANSITION_FORM_OPTIONS)
        _setup_transition_combo(self.oc_trans_combo, 220)
        tg.addWidget(self.oc_trans_combo, r, 2)
        tg.addWidget(QLabel("ζ="), r, 3, Qt.AlignRight)
        self.oc_trans_zeta = LineEdit()
        _setup_transition_zeta(self.oc_trans_zeta, "0.10")
        tg.addWidget(self.oc_trans_zeta, r, 4)

        # 行2：倒虹吸
        r = 2
        cat2 = QLabel("倒虹吸")
        cat2.setStyleSheet(_cat_style)
        tg.addWidget(cat2, r, 0, Qt.AlignRight)
        tg.addWidget(QLabel("进口:"), r, 1, Qt.AlignRight)
        self.siphon_inlet_combo = ComboBox()
        self.siphon_inlet_combo.addItems(SIPHON_TRANSITION_FORM_OPTIONS)
        _setup_transition_combo(self.siphon_inlet_combo, 220)
        tg.addWidget(self.siphon_inlet_combo, r, 2)
        tg.addWidget(QLabel("ζ₁="), r, 3, Qt.AlignRight)
        self.siphon_inlet_zeta = LineEdit()
        _setup_transition_zeta(self.siphon_inlet_zeta, "0.10")
        tg.addWidget(self.siphon_inlet_zeta, r, 4)
        tg.addWidget(QLabel("出口:"), r, 5, Qt.AlignRight)
        self.siphon_outlet_combo = ComboBox()
        self.siphon_outlet_combo.addItems(SIPHON_TRANSITION_FORM_OPTIONS)
        _setup_transition_combo(self.siphon_outlet_combo, 220)
        tg.addWidget(self.siphon_outlet_combo, r, 6)
        tg.addWidget(QLabel("ζ₂="), r, 7, Qt.AlignRight)
        self.siphon_outlet_zeta = LineEdit()
        _setup_transition_zeta(self.siphon_outlet_zeta, "0.20")
        tg.addWidget(self.siphon_outlet_zeta, r, 8)

        # 参考系数表按钮（放在倒虹吸行末尾）
        btn_ref = PushButton("参考系数表")
        btn_ref.setToolTip("查看表K.1.2（渡槽/隧洞）和表L.1.2（倒虹吸）渐变段局部损失系数")
        btn_ref.setMinimumWidth(104)
        btn_ref.setMinimumHeight(_transition_field_h)
        btn_ref.clicked.connect(self._open_transition_reference)
        tg.addWidget(btn_ref, r, 9)

        btn_rules = PushButton("长度规则")
        btn_rules.setToolTip("按当前工程已出现的上游/下游结构组合设置项目级渐变段长度规则（可选）")
        btn_rules.setMinimumWidth(104)
        btn_rules.setMinimumHeight(_transition_field_h)
        btn_rules.clicked.connect(self._open_transition_length_rules)
        tg.addWidget(btn_rules, r, 10)

        # 列弹性
        tg.setColumnStretch(2, 1)
        tg.setColumnStretch(6, 1)
        tg.setColumnMinimumWidth(4, 72)
        tg.setColumnMinimumWidth(8, 72)
        for row_idx in range(3):
            tg.setRowMinimumHeight(row_idx, _transition_field_h)
        trans_grp.toggled.connect(self._on_settings_toggled)
        lay.addWidget(trans_grp)
        lay.addStretch(1)

        # 渐变段型式变更 → 自动更新ζ系数
        self.trans_inlet_combo.currentTextChanged.connect(self._on_trans_inlet_form_changed)
        self.trans_outlet_combo.currentTextChanged.connect(self._on_trans_outlet_form_changed)
        self.oc_trans_combo.currentTextChanged.connect(self._on_oc_trans_form_changed)
        self.siphon_inlet_combo.currentTextChanged.connect(self._on_siphon_inlet_form_changed)
        self.siphon_outlet_combo.currentTextChanged.connect(self._on_siphon_outlet_form_changed)

    def _on_settings_toggled(self, collapsed):
        """折叠/展开设置区时，自动调整splitter分配，让底部图表获得释放的空间"""
        self._schedule_adjust_splitter_for_settings()

    def _schedule_adjust_splitter_for_settings(self):
        """延后重算顶部 splitter 高度，兼容控件刚刷新的尺寸变化。"""
        splitter = getattr(self, "_splitter", None)
        if splitter is None or splitter.count() <= 0:
            return
        QTimer.singleShot(0, self._adjust_splitter_for_settings)
        QTimer.singleShot(80, self._adjust_splitter_for_settings)

    def _refresh_top_layout_for_measurement(self):
        """测量 splitter 前，强制更新顶部折叠区布局缓存。"""
        top_w = self._splitter.widget(0)
        top_layout = getattr(self, "_top_lay", None)
        if top_layout is None:
            top_layout = top_w.layout()

        for group in (
            getattr(self, "_settings_group", None),
            getattr(self, "_transition_group", None),
        ):
            if group is None:
                continue

            content_layout = group.content_layout()
            if content_layout is not None:
                content_layout.invalidate()
                content_layout.activate()

            content_widget = group.content_widget()
            if content_widget is not None:
                content_widget.adjustSize()
                content_widget.updateGeometry()

            group.updateGeometry()

        if top_layout is not None:
            top_layout.invalidate()
            top_layout.activate()

        top_w.updateGeometry()
        return top_w, top_layout

    def _adjust_splitter_for_settings(self):
        """根据上半区实际需求高度重新分配 splitter 空间。"""
        top_w, top_layout = self._refresh_top_layout_for_measurement()

        if top_layout is not None:
            margins = top_layout.contentsMargins()
            current_width = max(1, top_w.width() - margins.left() - margins.right())
            hfw_hint = top_layout.totalHeightForWidth(current_width)
            size_hint = top_layout.totalSizeHint().height()
            top_hint = hfw_hint if hfw_hint > 0 else size_hint
            top_hint = max(top_hint, top_layout.totalMinimumSize().height())
        else:
            top_hint = max(top_w.sizeHint().height(), top_w.minimumSizeHint().height())

        top_hint = max(54, int(top_hint))
        total = max(1, self._splitter.height())
        min_bottom_h = 180
        top_cap = max(int(total * 0.78), total - min_bottom_h)
        top_h = min(top_hint, max(0, top_cap))
        bottom_h = total - top_h
        self._splitter.setSizes([top_h, bottom_h])

    def _build_workspace_area(self, parent):
        """构建下半区流程化标签页工作区。"""
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(4)

        self._workspace_tabs = QTabWidget(parent)
        self._workspace_tabs.setDocumentMode(True)

        self._tab_section_input = QWidget()
        self._tab_section_result = QWidget()
        self._tab_water_profile = QWidget()
        self._tab_output = QWidget()

        self._workspace_tabs.addTab(self._tab_section_input, "表1：基本参数输入")
        self._workspace_tabs.addTab(self._tab_section_result, "表2：断面计算结果")
        self._workspace_tabs.addTab(self._tab_water_profile, "表3：水面线计算表")
        self._workspace_tabs.addTab(self._tab_output, "结果与导出")
        self._workspace_tabs.tabBar().setUsesScrollButtons(True)
        self._workspace_tabs.tabBar().setElideMode(Qt.ElideNone)
        for tab_idx in range(self._workspace_tabs.count()):
            self._workspace_tabs.tabBar().setTabToolTip(tab_idx, self._workspace_tabs.tabText(tab_idx))

        self._build_tab_section_input(self._tab_section_input)
        self._build_tab_section_result(self._tab_section_result)
        self._build_tab_water_profile(self._tab_water_profile)
        self._build_tab_output(self._tab_output)
        self._workspace_tabs.setCurrentWidget(self._tab_section_input)

        lay.addWidget(self._workspace_tabs)

    def _switch_workspace_tab(self, target_widget):
        if self._workspace_tabs and target_widget:
            self._workspace_tabs.setCurrentWidget(target_widget)

    def _switch_to_output_process_tab(self, process_index: int = 1):
        """切到“结果与导出”页，并定位到详细过程子标签。"""
        self._switch_workspace_tab(self._tab_output)
        if not self._process_tabs:
            return
        max_index = self._process_tabs.count() - 1
        if max_index < 0:
            return
        self._process_tabs.setCurrentIndex(max(0, min(process_index, max_index)))

    def _build_tab_section_input(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(8, 6, 8, 6)
        lay.setSpacing(6)

        # 第一行：批量工具链
        tb1 = QHBoxLayout()
        tb1.setSpacing(6)
        lbl = QLabel("表1：基本参数输入")
        lbl.setStyleSheet(f"font-size:13px;font-weight:bold;color:{T1};")
        tb1.addWidget(lbl)
        tb1.addSpacing(8)

        tb1.addWidget(QLabel("流量段设置(m³/s):"))
        self._section_flow_segments_edit = LineEdit()
        self._section_flow_segments_edit.setText(
            self._batch_backend.flow_segments_edit.text().strip() or "5.0, 4.0, 3.0"
        )
        self._section_flow_segments_edit.setMinimumWidth(220)
        tb1.addWidget(self._section_flow_segments_edit)

        self._btn_section_apply_flow = PushButton("应用到表1")
        self._btn_section_apply_flow.clicked.connect(self._apply_flow_values_to_section_table)
        tb1.addWidget(self._btn_section_apply_flow)

        # A1：复选框位于顶行右侧（在“应用到表1”之后）
        has_option_toggle = False
        if hasattr(self._batch_backend, "inc_cb") and self._batch_backend.inc_cb:
            self._batch_backend.inc_cb.setParent(parent)
            self._batch_backend.inc_cb.setText("考虑加大流量")
            tb1.addWidget(self._batch_backend.inc_cb)
            has_option_toggle = True
        if hasattr(self._batch_backend, "detail_cb") and self._batch_backend.detail_cb:
            self._batch_backend.detail_cb.setParent(parent)
            self._batch_backend.detail_cb.setText("展示详细计算过程")
            tb1.addWidget(self._batch_backend.detail_cb)
            has_option_toggle = True
        if not has_option_toggle:
            tb1.addSpacing(12)
        tb1.addStretch()
        lay.addLayout(tb1)

        # 第一行下方：批量工具按钮（导入/计算/示例/模板）
        button_row = QHBoxLayout()
        button_row.setContentsMargins(0, 0, 0, 0)
        button_row.setSpacing(6)
        self._btn_section_import_excel = PrimaryPushButton("导入Excel")
        self._btn_section_import_excel.clicked.connect(self._import_section_excel)
        button_row.addWidget(self._btn_section_import_excel)

        self._btn_section_calc = PrimaryPushButton("开始批量计算")
        self._btn_section_calc.clicked.connect(self._run_section_batch_calculate)
        button_row.addWidget(self._btn_section_calc)

        sample_menu = RoundMenu(parent=self)
        sample_menu.addAction(Action("示例一（综合演示）", triggered=self._load_section_sample_1))
        sample_menu.addAction(Action("示例二（龙塘马坝河分干渠）", triggered=self._load_section_sample_2))
        sample_menu.addAction(Action("示例三（罗寂寺支渠）", triggered=self._load_section_sample_3))
        sample_menu.addAction(Action("示例四（飞龙分干渠）", triggered=self._load_section_sample_4))
        sample_menu.addAction(Action("示例五（茶亭支渠）", triggered=self._load_section_sample_5))
        sample_menu.addAction(Action("示例六（合作干渠）", triggered=self._load_section_sample_6))
        sample_menu.addAction(Action("示例七（甘家沟充水渠）", triggered=self._load_section_sample_7))
        sample_menu.addAction(Action("示例八（江家坝支管）", triggered=self._load_section_sample_8))
        sample_menu.addAction(Action("示例九（赛金支渠）", triggered=self._load_section_sample_9))
        sample_menu.addAction(Action("示例十（苏溪支渠）", triggered=self._load_section_sample_10))
        self._btn_section_sample = DropDownPushButton("示例数据")
        self._btn_section_sample.setMenu(sample_menu)
        button_row.addWidget(self._btn_section_sample)

        template_menu = RoundMenu(parent=self)
        template_menu.addAction(Action("示例一（综合演示）", triggered=lambda: self._open_section_excel_template("blank")))
        template_menu.addAction(Action("示例二（龙塘马坝河分干渠）", triggered=lambda: self._open_section_excel_template("longtang")))
        template_menu.addAction(Action("示例三（罗寂寺支渠）", triggered=lambda: self._open_section_excel_template("luojisi")))
        template_menu.addAction(Action("示例四（飞龙分干渠）", triggered=lambda: self._open_section_excel_template("feilong")))
        template_menu.addAction(Action("示例五（茶亭支渠）", triggered=lambda: self._open_section_excel_template("chating")))
        template_menu.addAction(Action("示例六（合作干渠）", triggered=lambda: self._open_section_excel_template("hezuo")))
        template_menu.addAction(Action("示例七（甘家沟充水渠）", triggered=lambda: self._open_section_excel_template("ganjiagou")))
        template_menu.addAction(Action("示例八（江家坝支管）", triggered=lambda: self._open_section_excel_template("jiangjiaba")))
        template_menu.addAction(Action("示例九（赛金支渠）", triggered=lambda: self._open_section_excel_template("saijin")))
        template_menu.addAction(Action("示例十（苏溪支渠）", triggered=lambda: self._open_section_excel_template("suxi")))
        self._btn_section_template = DropDownPushButton("打开Excel模板")
        self._btn_section_template.setMenu(template_menu)
        button_row.addWidget(self._btn_section_template)
        button_row.addStretch()

        lay.addLayout(button_row)

        # 下一行：行操作
        tb2 = QHBoxLayout()
        tb2.setSpacing(6)
        self._btn_section_add = PushButton("新增行")
        self._btn_section_add.clicked.connect(lambda: self._on_section_input_action(self._batch_backend._add_row))
        tb2.addWidget(self._btn_section_add)

        self._btn_section_insert = PushButton("插入行")
        self._btn_section_insert.clicked.connect(lambda: self._on_section_input_action(self._batch_backend._insert_row))
        tb2.addWidget(self._btn_section_insert)

        self._btn_section_delete = PushButton("删除行")
        self._btn_section_delete.clicked.connect(lambda: self._on_section_input_action(self._batch_backend._del_row))
        tb2.addWidget(self._btn_section_delete)

        self._btn_section_copy = PushButton("复制行")
        self._btn_section_copy.clicked.connect(lambda: self._on_section_input_action(self._batch_backend._copy_row))
        tb2.addWidget(self._btn_section_copy)

        self._btn_section_clear = PushButton("清空表1/表2")
        self._btn_section_clear.setToolTip("清空表1输入与表2结果\n▶ 表1可 Ctrl+Z 撤销；表2需重新计算生成")
        self._btn_section_clear.clicked.connect(self._clear_section_tables)
        tb2.addWidget(self._btn_section_clear)

        self._btn_section_params = PushButton("参数设置")
        self._btn_section_params.clicked.connect(self._batch_backend._open_parameter_dialog)
        tb2.addWidget(self._btn_section_params)
        hint = QLabel("提示: 双击参数列打开参数设置弹窗; 双击断面类型列可选择类型")
        hint.setStyleSheet("font-size:12px;font-weight:600;color:#0B5CAD;")
        tb2.addWidget(hint)
        tb2.addStretch()
        lay.addLayout(tb2)

        # 第三行：独立状态条（避免与按钮抢宽度）
        self._section_status_bar = QFrame(parent)
        self._section_status_bar.setObjectName("sectionStatusBar")
        self._section_status_bar.setStyleSheet(
            "QFrame#sectionStatusBar{border:1px solid #D9E2EF;border-radius:6px;background:#F7FAFE;}"
        )
        status_lay = QHBoxLayout(self._section_status_bar)
        status_lay.setContentsMargins(10, 6, 10, 6)
        status_lay.setSpacing(8)
        self._section_state_icon = QLabel("ℹ")
        self._section_state_icon.setStyleSheet("font-size:14px;color:#546E7A;font-weight:600;")
        self._section_state_icon.setFixedWidth(18)
        status_lay.addWidget(self._section_state_icon, alignment=Qt.AlignTop)
        self._section_state_label = QLabel("状态：未执行断面批量计算")
        self._section_state_label.setWordWrap(True)
        self._section_state_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self._section_state_label.setStyleSheet("font-size:12px;color:#455A64;")
        status_lay.addWidget(self._section_state_label, stretch=1)
        self._btn_section_failure_summary = PushButton("失败原因汇总")
        self._btn_section_failure_summary.setVisible(False)
        self._btn_section_failure_summary.clicked.connect(self._show_section_failure_summary_from_status)
        status_lay.addWidget(self._btn_section_failure_summary)
        lay.addWidget(self._section_status_bar)

        # 表1输入表
        self._section_input_table = self._batch_backend.input_table
        self._section_input_table.setParent(parent)
        self._section_input_table.cellChanged.connect(self._on_section_input_changed)
        lay.addWidget(self._section_input_table, stretch=1)

    def _build_tab_section_result(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(8, 6, 8, 6)
        lay.setSpacing(6)

        tb = QHBoxLayout()
        tb.setSpacing(6)
        lbl = QLabel("表2：断面计算结果汇总")
        lbl.setStyleSheet(f"font-size:13px;font-weight:bold;color:{T1};")
        tb.addWidget(lbl)
        tb.addStretch()

        self._btn_section_export_excel = PushButton("导出Excel报告")
        self._btn_section_export_excel.clicked.connect(self._batch_backend._export_excel)
        tb.addWidget(self._btn_section_export_excel)

        self._btn_section_export_word = PushButton("导出详细过程(Word)")
        self._btn_section_export_word.clicked.connect(self._batch_backend._export_word)
        tb.addWidget(self._btn_section_export_word)
        lay.addLayout(tb)

        # 失败锁定提示条（复用批量面板）
        if hasattr(self._batch_backend, "_error_lock_label") and self._batch_backend._error_lock_label:
            self._batch_backend._error_lock_label.setParent(parent)
            lay.addWidget(self._batch_backend._error_lock_label)

        self._section_result_table = self._batch_backend.result_table
        self._section_result_table.setParent(parent)
        lay.addWidget(self._section_result_table, stretch=1)

        # 失败原因汇总面板（默认收起，失败时可一键展开）
        self._section_failure_panel = QFrame(parent)
        self._section_failure_panel.setObjectName("sectionFailurePanel")
        self._section_failure_panel.setStyleSheet(
            "QFrame#sectionFailurePanel{border:1px solid #F1D0D0;border-radius:6px;background:#FFF8F8;}"
        )
        fail_lay = QVBoxLayout(self._section_failure_panel)
        fail_lay.setContentsMargins(10, 8, 10, 8)
        fail_lay.setSpacing(6)
        head_lay = QHBoxLayout()
        head_lay.setSpacing(8)
        fail_title = QLabel("失败原因汇总")
        fail_title.setStyleSheet("font-size:12px;font-weight:bold;color:#C62828;")
        head_lay.addWidget(fail_title)
        head_lay.addStretch()
        self._btn_section_failure_copy = PushButton("复制全部原因")
        self._btn_section_failure_copy.clicked.connect(self._copy_section_failure_reasons)
        head_lay.addWidget(self._btn_section_failure_copy)
        self._btn_section_failure_locate = PushButton("定位到选中行")
        self._btn_section_failure_locate.clicked.connect(self._locate_selected_section_failure)
        head_lay.addWidget(self._btn_section_failure_locate)
        self._btn_section_failure_collapse = PushButton("收起")
        self._btn_section_failure_collapse.clicked.connect(lambda: self._toggle_section_failure_panel(False))
        head_lay.addWidget(self._btn_section_failure_collapse)
        fail_lay.addLayout(head_lay)

        self._section_failure_table = QTableWidget(0, 3, self._section_failure_panel)
        self._section_failure_table.setHorizontalHeaderLabels(["序号", "建筑物", "失败原因"])
        self._section_failure_table.verticalHeader().setVisible(False)
        self._section_failure_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._section_failure_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self._section_failure_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._section_failure_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self._section_failure_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self._section_failure_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self._section_failure_table.itemDoubleClicked.connect(lambda _item: self._locate_selected_section_failure())
        fail_lay.addWidget(self._section_failure_table)
        self._section_failure_panel.setVisible(False)
        lay.addWidget(self._section_failure_panel)

    def _build_tab_water_profile(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(8, 6, 8, 6)
        lay.setSpacing(6)

        toolbar_presets = {
            "compact": {"h_spacing": 6, "btn_height": 34, "extra_primary": 10, "extra_primary_long": 12, "extra_normal": 8, "extra_clear": 6},
            "balanced": {"h_spacing": 8, "btn_height": 36, "extra_primary": 10, "extra_primary_long": 12, "extra_normal": 8, "extra_clear": 6},
            "comfortable": {"h_spacing": 10, "btn_height": 38, "extra_primary": 12, "extra_primary_long": 14, "extra_normal": 10, "extra_clear": 8},
        }
        self._node_toolbar_preset = toolbar_presets.get(self._node_toolbar_layout_preset, toolbar_presets["balanced"])

        tb = QHBoxLayout()
        tb.setContentsMargins(0, 0, 0, 0)
        tb.setSpacing(self._node_toolbar_preset["h_spacing"])
        tb.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        lbl = QLabel("表3：水面线计算表")
        lbl.setStyleSheet(f"font-size:13px; font-weight:bold; color:{T1};")
        from PySide6.QtWidgets import QSizePolicy
        lbl.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)
        tb.addWidget(lbl)
        tb.addSpacing(10)

        def _register_toolbar_button(btn, role):
            btn.ensurePolished()
            hint_w = btn.sizeHint().width()
            min_w = btn.minimumSizeHint().width() + 2
            if role == "primary_long":
                extra = self._node_toolbar_preset["extra_primary_long"]
            elif role == "primary":
                extra = self._node_toolbar_preset["extra_primary"]
            elif role == "clear":
                extra = self._node_toolbar_preset["extra_clear"]
            else:
                extra = self._node_toolbar_preset["extra_normal"]
            width = max(min_w, hint_w + extra)
            btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            btn.setFixedSize(width, self._node_toolbar_preset["btn_height"])

        self._btn_transition = PrimaryPushButton("插入渐变段")
        self._btn_transition.clicked.connect(self._insert_transitions)
        _register_toolbar_button(self._btn_transition, "primary")

        self._btn_siphon = PrimaryPushButton("倒虹吸水力计算")
        self._btn_siphon.clicked.connect(self._open_siphon_calculator)
        _register_toolbar_button(self._btn_siphon, "primary_long")

        self.btn_pressure_pipe_calc = PrimaryPushButton("有压管道水力计算")
        self.btn_pressure_pipe_calc.clicked.connect(self._open_pressure_pipe_calculator)
        _register_toolbar_button(self.btn_pressure_pipe_calc, "primary_long")

        self._btn_calc = PrimaryPushButton("执行计算")
        self._btn_calc.clicked.connect(self._calculate)
        _register_toolbar_button(self._btn_calc, "primary")

        self.btn_pressure_pipe_water_hammer = PrimaryPushButton("有压管道水锤验算")
        self.btn_pressure_pipe_water_hammer.clicked.connect(self._open_pressure_pipe_water_hammer_checker)
        _register_toolbar_button(self.btn_pressure_pipe_water_hammer, "primary_long")

        for w in [
            self._btn_transition,
            self._btn_siphon,
            self.btn_pressure_pipe_calc,
            self._btn_calc,
            self.btn_pressure_pipe_water_hammer,
        ]:
            tb.addWidget(w)

        self._btn_table3_clear = PushButton("清空表3")
        self._btn_table3_clear.setToolTip("清空表3节点并清理倒虹吸/有压管道配置\n▶ 表3可 Ctrl+Z 撤销；配置清理不可撤销")
        self._btn_table3_clear.clicked.connect(self._on_clear_table3_clicked)
        _register_toolbar_button(self._btn_table3_clear, "clear")
        tb.addWidget(self._btn_table3_clear)
        lay.addLayout(tb)

        self.node_table = FrozenColumnTableWidget(0, len(NODE_ALL_HEADERS), frozen_count=4)
        self.node_table.setHorizontalHeaderLabels(NODE_ALL_HEADERS)
        self.node_table.horizontalHeader().setStretchLastSection(False)
        self.node_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.node_table.horizontalHeader().setMinimumSectionSize(50)
        self.node_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.node_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.node_table.setAlternatingRowColors(True)
        self.node_table.setEditTriggers(
            QAbstractItemView.EditKeyPressed | QAbstractItemView.AnyKeyPressed
        )
        self.node_table.setFont(QFont("Microsoft YaHei", 10))
        self.node_table.verticalHeader().setDefaultSectionSize(26)
        self.node_table.setMinimumHeight(180)
        self.node_table.cellDoubleClicked.connect(self._on_node_cell_double_clicked)
        self.node_table.cellChanged.connect(self._on_loss_cell_changed)
        self.node_table.currentCellChanged.connect(self._on_current_cell_changed)
        self.node_table.undoRequested.connect(self._undo_loss_edit)
        self.node_table.redoRequested.connect(self._redo_loss_edit)
        self.node_table.deleteRequested.connect(self._push_undo_snapshot)
        self.node_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.node_table.customContextMenuRequested.connect(self._show_node_table_context_menu)
        self.node_table.installEventFilter(self)
        self.node_table.viewport().installEventFilter(self)
        lay.addWidget(self.node_table, stretch=1)

        undo_sc = QShortcut(QKeySequence.StandardKey.Undo, self)
        undo_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        undo_sc.activated.connect(self._undo_loss_edit)
        redo_sc = QShortcut(QKeySequence.StandardKey.Redo, self)
        redo_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        redo_sc.activated.connect(self._redo_loss_edit)

        self._setup_header_tooltips()
        self._set_downstream_actions_enabled(False, state_text="状态：断面结果未就绪，请先执行断面批量计算")
        self._refresh_pressure_pipe_controls()

    def _build_tab_output(self, parent):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(8, 6, 8, 6)
        lay.setSpacing(4)

        # 导出工具栏
        tb = QHBoxLayout()
        lbl = QLabel("计算结果")
        lbl.setStyleSheet(f"font-size:13px;font-weight:bold;color:{T1};")
        tb.addWidget(lbl)
        tb.addStretch()
        btn_export_excel = PushButton("导出Excel")
        btn_export_excel.clicked.connect(self._export_excel)
        btn_export_word = PushButton("导出Word")
        btn_export_word.clicked.connect(self._export_word)
        for w in [btn_export_excel, btn_export_word]:
            tb.addWidget(w)
        lay.addLayout(tb)

        # CAD工具栏
        cad_tb = QHBoxLayout()
        cad_tb.setSpacing(6)
        cad_lbl = QLabel("CAD工具")
        cad_lbl.setStyleSheet(f"font-size:12px;font-weight:bold;color:{T2};")
        cad_tb.addWidget(cad_lbl)
        cad_tb.addStretch()
        btn_profile = PushButton("生成纵断面表格"); btn_profile.clicked.connect(self._cad_longitudinal_profile)
        btn_profile.setToolTip("导出上纵断面表格 DXF/TXT\n含线框、渠底/渠顶/水面折线、高程文字、桩号、建筑物名称、坡降、IP点名称")
        btn_summary = PushButton("生成断面汇总表"); btn_summary.clicked.connect(self._cad_section_summary)
        btn_summary.setToolTip("导出各类断面尺寸及水力要素汇总表 DXF\n含明渠/隧洞/渡槽/暗涵/倒虹吸等断面参数")
        btn_ip = PushButton("IP坐标及弯道参数表"); btn_ip.clicked.connect(self._cad_ip_table)
        btn_ip.setToolTip("导出IP坐标及弯道参数表 DXF/Excel\n含IP点坐标、桩号、转角、半径、切线长、弧长、底高程")
        btn_combined = PrimaryPushButton("导出全部DXF"); btn_combined.clicked.connect(self._cad_combined_dxf)
        btn_combined.setToolTip("一键合并导出：纵断面表格 + 断面汇总表 + IP坐标表\n三个表格输出到同一个DXF文件，分图层管理")
        btn_bzzh2 = PushButton("生成bzzh2命令内容"); btn_bzzh2.clicked.connect(self._cad_bzzh2)
        btn_bzzh2.setToolTip("生成ZDM用的bzzh2命令 TXT\n提取建筑物进出口数据")
        btn_plan = PushButton("建筑物名称上平面图"); btn_plan.clicked.connect(self._cad_building_plan)
        btn_plan.setToolTip("生成AutoCAD -TEXT命令并复制到剪贴板\n将建筑物名称平行于轴线放置在平面图上")
        for w in [btn_profile, btn_summary, btn_ip, btn_combined, btn_bzzh2, btn_plan]:
            cad_tb.addWidget(w)
        lay.addLayout(cad_tb)

        # 计算结果摘要面板（持久显示）
        self.summary_grp = QGroupBox("计算结果摘要")
        sg_lay = QHBoxLayout(self.summary_grp)
        sg_lay.setContentsMargins(8, 4, 8, 4)
        sg_lay.setSpacing(16)
        self.lbl_summary_info = QLabel("尚未计算")
        self.lbl_summary_info.setStyleSheet("font-size: 12px;")
        sg_lay.addWidget(self.lbl_summary_info, stretch=1)
        self.btn_building_stats = PushButton("建筑物长度统计")
        self.btn_building_stats.clicked.connect(self._show_building_length_dialog)
        self.btn_building_stats.setEnabled(False)
        sg_lay.addWidget(self.btn_building_stats)
        lay.addWidget(self.summary_grp)

        # 详细过程双子标签
        self._process_tabs = QTabWidget(parent)
        tab_batch_detail = QWidget()
        batch_detail_lay = QVBoxLayout(tab_batch_detail)
        batch_detail_lay.setContentsMargins(2, 2, 2, 2)
        self._batch_backend.detail_text.setParent(tab_batch_detail)
        self._batch_backend.detail_text.setReadOnly(True)
        batch_detail_lay.addWidget(self._batch_backend.detail_text)
        self._process_tabs.addTab(tab_batch_detail, "表2断面计算结果详细过程")

        tab_water_detail = QWidget()
        water_detail_lay = QVBoxLayout(tab_water_detail)
        water_detail_lay.setContentsMargins(2, 2, 2, 2)
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        self.detail_text.setFont(QFont("Consolas", 10))
        water_detail_lay.addWidget(self.detail_text)
        self._process_tabs.addTab(tab_water_detail, "水面线详细过程")
        self._process_tabs.tabBar().setUsesScrollButtons(True)
        self._process_tabs.tabBar().setElideMode(Qt.ElideNone)
        for tab_idx in range(self._process_tabs.count()):
            self._process_tabs.tabBar().setTabToolTip(tab_idx, self._process_tabs.tabText(tab_idx))
        self._process_tabs.setCurrentIndex(1)
        lay.addWidget(self._process_tabs, stretch=1)

        # 初始帮助
        self._show_help()

    def _import_section_excel(self):
        self._batch_backend._import_from_excel()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _load_section_sample_1(self):
        self._batch_backend._add_sample_data()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _load_section_sample_2(self):
        self._batch_backend._add_sample_data_2()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _load_section_sample_3(self):
        self._batch_backend._add_sample_data_3()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _load_section_sample_4(self):
        self._batch_backend._add_sample_data_4()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _load_section_sample_5(self):
        self._batch_backend._add_sample_data_5()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _load_section_sample_6(self):
        self._batch_backend._add_sample_data_6()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _load_section_sample_7(self):
        self._batch_backend._add_sample_data_7()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _load_section_sample_8(self):
        self._batch_backend._add_sample_data_8()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _load_section_sample_9(self):
        self._batch_backend._add_sample_data_9()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _load_section_sample_10(self):
        self._batch_backend._add_sample_data_10()
        self._sync_batch_settings()
        self._switch_workspace_tab(self._tab_section_input)
        if self._section_input_table and self._section_input_table.rowCount() > 0:
            self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _open_section_excel_template(self, template_key: str):
        if not self._batch_backend:
            return
        self._batch_backend._open_excel_template_file(template_key, dialog_parent=self)

    def _on_section_input_action(self, action):
        if not action:
            return
        action()
        self._switch_workspace_tab(self._tab_section_input)
        self._mark_section_results_stale("状态：表1已更新，请重新执行断面批量计算")

    def _ask_destructive_confirm(self, title: str, content: str, yes_text: str = "确认", no_text: str = "取消") -> bool:
        box = MessageBox(title, content, self)
        box.yesButton.setText(yes_text)
        box.cancelButton.setText(no_text)
        try:
            box.yesButton.setAutoDefault(False)
            box.yesButton.setDefault(False)
            box.cancelButton.setAutoDefault(True)
            box.cancelButton.setDefault(True)
        except Exception:
            pass
        box.cancelButton.setFocus(Qt.TabFocusReason)
        return bool(box.exec())

    def _clear_section_results_only(self):
        self._batch_backend._clear_results()
        if SHARED_DATA_AVAILABLE:
            try:
                get_shared_data_manager().clear_batch_results()
            except Exception:
                pass
        self._switch_workspace_tab(self._tab_section_result)
        self._mark_section_results_stale("状态：表2结果已清空，请重新执行断面批量计算")
        self._refresh_section_failure_feedback(auto_focus=False, auto_expand_once=False)

    def _apply_flow_values_to_section_table(self):
        """将流量段设置应用到表1。"""
        if not self._batch_backend:
            return
        self._batch_backend.flow_segments_edit.setText(self._section_flow_segments_edit.text().strip())
        self._batch_backend._apply_flow_segments()

    def _copy_global_settings_to_batch_backend(self):
        """同步渠道基础设置到断面计算后端。"""
        if not self._batch_backend:
            return
        self._batch_backend.channel_name_edit.setText(self.channel_name_edit.text().strip())
        self._batch_backend.channel_level_combo.setCurrentText(self.channel_level_combo.currentText())
        self._batch_backend.start_wl_edit.setText(self.start_wl_edit.text().strip())
        self._batch_backend.start_station_edit.setText(self.start_station_edit.text().strip())
        flow_text = self._section_flow_segments_edit.text().strip()
        if not flow_text:
            flow_text = self.design_flow_edit.text().strip()
        self._batch_backend.flow_segments_edit.setText(flow_text)

    def _count_section_calc_failures(self) -> int:
        """统计表2中的失败行数量。"""
        return len(self._collect_section_failures())

    def _collect_section_failures(self):
        failures = []
        table = self._section_result_table
        if not table:
            return failures
        status_col = table.columnCount() - 1
        for row in range(table.rowCount()):
            status_item = table.item(row, status_col)
            status_text = (status_item.text() if status_item else "").strip()
            if not any(tag in status_text for tag in ("✗", "失败", "错误")):
                continue
            seq_item = table.item(row, 0)
            name_item = table.item(row, 2)
            failures.append({
                "row": row,
                "seq": (seq_item.text().strip() if seq_item else str(row + 1)) or str(row + 1),
                "name": (name_item.text().strip() if name_item else "") or "-",
                "reason": self._normalize_section_failure_reason(status_text),
                "status": status_text,
            })
        return failures

    @staticmethod
    def _normalize_section_failure_reason(status_text: str) -> str:
        text = (status_text or "").strip()
        if text.startswith("✗"):
            text = text.lstrip("✗").strip()
        text = re.sub(r"^(失败|错误)\s*[:：]?\s*", "", text)
        return text or "计算失败"

    def _highlight_section_failure_rows(self, failures):
        table = self._section_result_table
        if not table:
            return
        failed_rows = {f.get("row", -1) for f in failures}
        for row in range(table.rowCount()):
            is_failed = row in failed_rows
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if not item:
                    continue
                if is_failed:
                    item.setBackground(QColor("#FDEBEC"))
                else:
                    item.setBackground(QBrush())

    def _focus_section_result_row(self, row: int):
        table = self._section_result_table
        if not table or row < 0 or row >= table.rowCount():
            return
        table.clearSelection()
        table.selectRow(row)
        focus_item = table.item(row, 0) or table.item(row, table.columnCount() - 1)
        if focus_item:
            table.setCurrentItem(focus_item)
            table.scrollToItem(focus_item, QAbstractItemView.PositionAtCenter)

    def _refresh_section_failure_panel(self, failures):
        if not self._section_failure_table:
            return
        self._section_failure_table.setRowCount(len(failures))
        for idx, info in enumerate(failures):
            seq_item = QTableWidgetItem(str(info.get("seq", "")))
            seq_item.setTextAlignment(Qt.AlignCenter)
            seq_item.setData(Qt.UserRole, info.get("row", -1))
            name_item = QTableWidgetItem(str(info.get("name", "")))
            reason_item = QTableWidgetItem(str(info.get("reason", "")))
            reason_item.setToolTip(str(info.get("status", "")))
            self._section_failure_table.setItem(idx, 0, seq_item)
            self._section_failure_table.setItem(idx, 1, name_item)
            self._section_failure_table.setItem(idx, 2, reason_item)
        if failures:
            self._section_failure_table.selectRow(0)

    def _toggle_section_failure_panel(self, visible=None, switch_to_result_tab: bool = True):
        if not self._section_failure_panel:
            return
        if visible is None:
            visible = not self._section_failure_panel.isVisible()
        if visible and not self._section_failure_records:
            visible = False
        self._section_failure_panel.setVisible(bool(visible))
        if self._btn_section_failure_summary:
            self._btn_section_failure_summary.setVisible(bool(self._section_failure_records))
            self._btn_section_failure_summary.setText("收起失败汇总" if visible else "失败原因汇总")
        if visible and switch_to_result_tab:
            self._switch_workspace_tab(self._tab_section_result)

    def _show_section_failure_summary_from_status(self):
        if not self._section_failure_records:
            return
        self._toggle_section_failure_panel(True, switch_to_result_tab=True)
        self._focus_section_result_row(self._section_failure_records[0]["row"])

    def _copy_section_failure_reasons(self):
        if not self._section_failure_records:
            return
        lines = [f"序号{f['seq']}（{f['name']}）：{f['reason']}" for f in self._section_failure_records]
        QApplication.clipboard().setText("\n".join(lines))
        InfoBar.success(
            "已复制",
            f"已复制 {len(lines)} 条失败原因到剪贴板。",
            parent=self._info_parent(), duration=2500, position=InfoBarPosition.TOP
        )

    def _locate_selected_section_failure(self):
        if not self._section_failure_table or not self._section_failure_records:
            return
        row = self._section_failure_table.currentRow()
        if row < 0:
            row = 0
        item = self._section_failure_table.item(row, 0)
        target_row = int(item.data(Qt.UserRole)) if item else -1
        self._toggle_section_failure_panel(True, switch_to_result_tab=True)
        self._focus_section_result_row(target_row)

    def _refresh_section_failure_feedback(self, auto_focus: bool = False, auto_expand_once: bool = False):
        failures = self._collect_section_failures()
        self._section_failure_records = failures
        self._highlight_section_failure_rows(failures)
        self._refresh_section_failure_panel(failures)
        if self._btn_section_failure_summary:
            self._btn_section_failure_summary.setVisible(bool(failures))
            self._btn_section_failure_summary.setText("失败原因汇总")
        if not failures:
            self._toggle_section_failure_panel(False, switch_to_result_tab=False)
            return failures
        if auto_focus:
            self._focus_section_result_row(failures[0]["row"])
        if auto_expand_once and not self._section_failure_auto_expanded_once:
            self._toggle_section_failure_panel(True, switch_to_result_tab=True)
            self._section_failure_auto_expanded_once = True
        return failures

    def _set_section_status(self, state_text: str, status_kind: str = "neutral"):
        if not self._section_state_label:
            return
        palette = {
            "success": {"bg": "#ECF8F0", "bd": "#B9E0C8", "fg": "#2E7D32", "icon": "✔"},
            "error": {"bg": "#FFF1F0", "bd": "#F1C1BE", "fg": "#C62828", "icon": "✗"},
            "warning": {"bg": "#FFF8E8", "bd": "#F3D9A7", "fg": "#B76E00", "icon": "⚠"},
            "neutral": {"bg": "#F7FAFE", "bd": "#D9E2EF", "fg": "#455A64", "icon": "ℹ"},
        }
        final_kind = status_kind if status_kind in palette else "neutral"
        self._section_status_kind = final_kind
        token = palette[final_kind]
        if self._section_status_bar:
            self._section_status_bar.setStyleSheet(
                "QFrame#sectionStatusBar{"
                f"border:1px solid {token['bd']};border-radius:6px;background:{token['bg']};"
                "}"
            )
        if self._section_state_icon:
            self._section_state_icon.setText(token["icon"])
            self._section_state_icon.setStyleSheet(f"font-size:14px;color:{token['fg']};font-weight:700;")
        self._section_state_label.setText(state_text or "状态：未执行断面批量计算")
        self._section_state_label.setStyleSheet(f"font-size:12px;color:{token['fg']};font-weight:600;")

    def get_first_success_auto_jump_marker(self) -> bool:
        return bool(getattr(self, "_section_first_success_switched", False))

    def set_first_success_auto_jump_marker(self, marked: bool):
        self._section_first_success_switched = bool(marked)

    def reset_first_success_auto_jump_marker(self):
        self._section_first_success_switched = False

    def _run_section_batch_calculate(self):
        """执行表1断面批量计算，并在全成功后自动同步到表3。"""
        if not self._batch_backend:
            return
        self._copy_global_settings_to_batch_backend()
        self._batch_backend._batch_calculate()
        self._switch_workspace_tab(self._tab_section_result)

        result_rows = self._section_result_table.rowCount() if self._section_result_table else 0
        fail_count = self._count_section_calc_failures()
        if result_rows == 0:
            self._mark_section_results_stale("状态：断面计算未生成结果")
            self._refresh_section_failure_feedback(auto_focus=False, auto_expand_once=False)
            return
        if fail_count > 0:
            self._mark_section_results_stale(
                f"状态：断面计算存在 {fail_count} 条失败，已锁定下游操作",
                status_kind="error"
            )
            self._refresh_section_failure_feedback(auto_focus=True, auto_expand_once=True)
            InfoBar.warning(
                "断面计算未通过",
                "存在失败行，已锁定插入渐变段/倒虹吸/有压管道/执行计算，请修复表1后重算。",
                parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP
            )
            return

        if not self._sync_to_water_profile_table():
            self._mark_section_results_stale("状态：同步到表3失败，请重试", status_kind="error")
            return

        self._refresh_section_failure_feedback(auto_focus=False, auto_expand_once=False)
        self._section_sync_ready = True
        self._set_downstream_actions_enabled(True, state_text="状态：断面全成功，表1+表2已同步到表3")
        if not self._section_first_success_switched:
            self._section_first_success_switched = True
            self._switch_workspace_tab(self._tab_water_profile)

    def _sync_to_water_profile_table(self) -> bool:
        """表1+表2自动同步到表3（仅断面全成功后调用）。"""
        if not SHARED_DATA_AVAILABLE:
            InfoBar.warning(
                "不可用", "SharedDataManager未加载，无法同步到表3",
                parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP
            )
            return False
        try:
            shared_data = get_shared_data_manager()
            if not shared_data.get_batch_results():
                InfoBar.warning(
                    "无法同步",
                    "未检测到可同步的断面结果，请先执行断面批量计算且确保全部成功。",
                    parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP
                )
                return False
            self._import_from_batch()
            return self.node_table.rowCount() > 0
        except Exception as e:
            InfoBar.error(
                "同步失败", f"表1/表2同步到表3失败：{e}",
                parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP
            )
            return False

    def _clear_section_tables(self):
        """清空表1与表2（不清空表3）。"""
        if not self._batch_backend:
            return
        input_rows = self._section_input_table.rowCount() if self._section_input_table else 0
        result_rows = self._section_result_table.rowCount() if self._section_result_table else 0
        if input_rows <= 0 and result_rows <= 0:
            self._mark_section_results_stale("状态：表1/表2已为空，请先执行断面批量计算", status_kind="warning")
            self._refresh_section_failure_feedback(auto_focus=False, auto_expand_once=False)
            InfoBar.info(
                "已为空",
                "表1与表2已为空，无需清空；下游操作已保持锁定，请先执行断面批量计算。",
                parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP
            )
            return
        confirm_msg = (
            f"将清空表1与表2数据：\n"
            f"• 表1输入行数：{input_rows}\n"
            f"• 表2结果行数：{result_rows}\n\n"
            "影响：将锁定下游操作，需重新执行断面批量计算后同步到表3。\n"
            "恢复说明：表1可按 Ctrl+Z 撤销，表2需重新计算生成。"
        )
        if not self._ask_destructive_confirm("确认清空表1+表2", confirm_msg, yes_text="确认清空", no_text="取消"):
            return
        if input_rows > 0:
            self._batch_backend._clear_input(force=True)
        elif result_rows > 0:
            self._batch_backend._clear_results()
        self._mark_section_results_stale("状态：表1/表2已清空，请重新执行断面批量计算", status_kind="warning")
        self._refresh_section_failure_feedback(auto_focus=False, auto_expand_once=False)
        InfoBar.success(
            "清空完成",
            f"已清空表1 {input_rows} 行、表2 {result_rows} 行。表1可按 Ctrl+Z 撤销，表2需重新计算生成。",
            parent=self._info_parent(), duration=2500, position=InfoBarPosition.TOP
        )

    def _on_section_input_changed(self, row, col):
        """表1任意编辑后，立即使下游链路失效并锁定。"""
        _ = (row, col)
        if getattr(self, "_loading_project", False):
            return
        if self._section_sync_ready:
            self._mark_section_results_stale("状态：表1已变更，结果已失效，请重新执行断面批量计算")

    def _mark_section_results_stale(self, state_text: str, status_kind: str = ""):
        self._section_sync_ready = False
        self._transition_topology_prepared = False
        self._set_downstream_actions_enabled(False, state_text=state_text, status_kind=status_kind)
        self._refresh_pressure_pipe_controls()

    def _has_transition_topology_ready(self, nodes=None) -> bool:
        source_nodes = nodes
        if source_nodes is None:
            source_nodes = self.calculated_nodes or self.nodes
        has_transition_rows = any(getattr(node, 'is_transition', False) for node in (source_nodes or []))
        return has_transition_rows or bool(getattr(self, "_transition_topology_prepared", False))

    def _set_downstream_actions_enabled(self, enabled: bool, state_text: str = "", status_kind: str = ""):
        buttons = [
            getattr(self, "_btn_transition", None),
            getattr(self, "_btn_siphon", None),
            getattr(self, "btn_pressure_pipe_calc", None),
            getattr(self, "_btn_calc", None),
            getattr(self, "btn_pressure_pipe_water_hammer", None),
        ]
        for btn in buttons:
            if btn is not None:
                btn.setEnabled(bool(enabled))
        if state_text:
            final_kind = status_kind
            if not final_kind:
                if enabled:
                    final_kind = "success"
                elif any(k in state_text for k in ("失败", "错误", "锁定")):
                    final_kind = "error"
                elif any(k in state_text for k in ("未生成", "未就绪", "请先", "请重新")):
                    final_kind = "warning"
                else:
                    final_kind = "neutral"
            self._set_section_status(state_text, final_kind)

    def _ensure_downstream_ready(self, action_name: str) -> bool:
        if self._section_sync_ready:
            return True
        InfoBar.info(
            "操作已锁定",
            f"{action_name}前请先完成断面批量计算并确保全成功（自动同步到表3后解锁）。",
            parent=self._info_parent(), duration=4500, position=InfoBarPosition.TOP
        )
        return False

    def _is_table1_source_row(self, row: int) -> bool:
        table = self.node_table
        if not table or row < 0 or row >= table.rowCount():
            return False
        first_item = table.item(row, 0)
        payload = first_item.data(Qt.UserRole) if first_item else None
        if isinstance(payload, dict) and "_from_table1_source" in payload:
            return bool(payload.get("_from_table1_source"))
        if isinstance(payload, dict) and payload.get("_auto_channel"):
            return False
        struct_item = table.item(row, 2)
        struct_text = struct_item.text().strip() if struct_item else ""
        if "渐变段" in struct_text or "(连接段)" in struct_text:
            return False
        basic_text = ""
        for col in (0, 1, 2):
            item = table.item(row, col)
            if item and item.text().strip():
                basic_text = item.text().strip()
                break
        return bool(basic_text)

    def _is_pressure_pipe_row(self, row: int) -> bool:
        table = self.node_table
        if not table or row < 0 or row >= table.rowCount():
            return False
        struct_item = table.item(row, 2)
        struct_text = struct_item.text().strip() if struct_item else ""
        return self._is_pressure_pipe_like_structure_text(struct_text)

    @staticmethod
    def _is_pressure_pipe_like_structure_text(struct_text: str) -> bool:
        text = normalize_section_type_name(struct_text)
        if not text:
            return False
        structure_type_cls = globals().get("StructureType")
        helper = getattr(structure_type_cls, "is_pressure_pipe_like_str", None)
        if callable(helper):
            try:
                return bool(helper(text))
            except Exception:
                pass
        return text in PRESSURE_PIPE_LIKE_STRUCTURE_TEXTS

    @staticmethod
    def _is_channel_roughness_structure_text(struct_text: str) -> bool:
        """判断结构类型是否应纳入“渠道糙率”候选白名单。"""
        text = normalize_section_type_name(struct_text)
        if not text:
            return False
        if text.startswith(("明渠-", "渡槽-", "隧洞-")):
            return True
        return bool(normalize_culvert_family_type_name(text))

    @staticmethod
    def _resolve_batch_import_section_type(raw_section_type: str, raw_result: dict | None = None) -> tuple[str, str]:
        """统一表1批量结果导入到表3前的结构类型名称。"""
        raw_result = raw_result or {}
        section_type = normalize_section_type_name(raw_section_type)
        culvert_family_type = normalize_culvert_family_type_name(raw_section_type or section_type)
        struct_map = {
            "梯形": "明渠-梯形", "矩形": "明渠-矩形", "圆形": "明渠-圆形",
            "U形": "渡槽-U形", "U形渡槽": "渡槽-U形", "隧洞": "隧洞-圆形", "渡槽": "渡槽-U形",
            # 兜底：计算引擎返回的简化名（正常流程已在batch注册时修正）
            "圆拱直墙型": "隧洞-圆拱直墙型",
            "马蹄形标准Ⅰ型": "隧洞-马蹄形Ⅰ型", "马蹄形标准Ⅱ型": "隧洞-马蹄形Ⅱ型",
            "矩形暗涵": RECT_CULVERT_FAMILY_TEXT,
            "暗渠": RECT_CULVERT_FAMILY_TEXT,
            "矩形暗渠": RECT_CULVERT_FAMILY_TEXT,
            "暗涵-矩形": RECT_CULVERT_FAMILY_TEXT,
            "暗涵-圆拱直墙型": ARCH_CULVERT_FAMILY_TEXT,
            "退水闸": "退水闸",
        }
        if (
            not culvert_family_type
            and ("暗涵" in section_type or "暗渠" in section_type or section_type == "矩形暗涵")
            and float(raw_result.get("theta_deg", 0) or 0) > 0
        ):
            culvert_family_type = ARCH_CULVERT_FAMILY_TEXT

        if culvert_family_type:
            section_type = culvert_family_type
        elif section_type in struct_map:
            section_type = struct_map[section_type]
        elif "渡槽-U" in section_type or "U形渡槽" in section_type:
            section_type = "渡槽-U形"
        elif "渡槽-矩形" in section_type:
            section_type = "渡槽-矩形"
        elif "隧洞-圆拱直墙" in section_type:
            section_type = "隧洞-圆拱直墙型"
        elif "隧洞-马蹄形Ⅰ" in section_type:
            section_type = "隧洞-马蹄形Ⅰ型"
        elif "隧洞-马蹄形Ⅱ" in section_type:
            section_type = "隧洞-马蹄形Ⅱ型"
        elif "暗涵" in section_type or "暗渠" in section_type:
            section_type = normalize_culvert_family_type_name(section_type) or RECT_CULVERT_FAMILY_TEXT
        return section_type, culvert_family_type

    def _prepare_batch_import_results(self, results):
        """预扫表1批量结果，先确定渠道糙率候选与各类概览信息。"""
        prepared_results = []
        general_roughness_vals = []
        siphon_roughness_pairs = []
        pressure_pipe_params_pairs = []

        for sr in results:
            raw_result = getattr(sr, 'raw_result', {}) or {}
            section_type, culvert_family_type = self._resolve_batch_import_section_type(
                str(getattr(sr, 'section_type', '') or ''),
                raw_result,
            )
            building_name = str(getattr(sr, 'building_name', '') or '')
            pipe_material = str(
                getattr(sr, 'pipe_material', '') or raw_result.get('pipe_material', '')
            ).strip()
            n_val = getattr(sr, 'n', 0) or ""
            try:
                roughness_value = float(n_val) if n_val and str(n_val).strip() else 0.0
            except (ValueError, TypeError):
                roughness_value = 0.0

            is_channel_roughness_row = False
            if self._is_pressure_pipe_like_structure_text(section_type) and building_name.strip():
                pressure_pipe_params_pairs.append((building_name, pipe_material))
            elif roughness_value > 0:
                if "倒虹吸" in section_type:
                    siphon_roughness_pairs.append(
                        (building_name or f"倒虹吸{len(siphon_roughness_pairs) + 1}", roughness_value)
                    )
                elif self._is_channel_roughness_structure_text(section_type):
                    general_roughness_vals.append(roughness_value)
                    is_channel_roughness_row = True

            prepared_results.append({
                "result": sr,
                "raw_result": raw_result,
                "section_type": section_type,
                "culvert_family_type": culvert_family_type,
                "pipe_material": pipe_material,
                "is_channel_roughness_row": is_channel_roughness_row,
            })

        return (
            prepared_results,
            general_roughness_vals,
            siphon_roughness_pairs,
            pressure_pipe_params_pairs,
        )

    @classmethod
    def _is_pressure_pipe_like_node(cls, node) -> bool:
        if getattr(node, "is_pressure_pipe", False):
            return True
        struct_type = getattr(node, "structure_type", None)
        structure_type_cls = globals().get("StructureType")
        helper = getattr(structure_type_cls, "is_pressure_pipe_like", None)
        if callable(helper):
            try:
                return bool(helper(struct_type))
            except Exception:
                pass
        value = getattr(struct_type, "value", struct_type)
        return cls._is_pressure_pipe_like_structure_text(value)

    @classmethod
    def _get_node_structure_type_text(cls, node) -> str:
        getter = getattr(node, "get_structure_type_str", None)
        if callable(getter):
            try:
                text = normalize_section_type_name(getter())
                if text:
                    return text
            except Exception:
                pass
        struct_type = getattr(node, "structure_type", None)
        value = getattr(struct_type, "value", struct_type)
        return normalize_section_type_name(value)

    @classmethod
    def _should_warn_missing_structure_height(cls, node) -> bool:
        """判断当前节点是否需要纳入“缺少结构总高”提示。"""
        if getattr(node, "is_transition", False):
            return False
        # 自动插入补段只是辅助连通行，不应参与结构总高缺失提示。
        if getattr(node, "is_auto_inserted_channel", False):
            return False

        structure_text = cls._get_node_structure_type_text(node)
        if not structure_text:
            return False
        if "倒虹吸" in structure_text:
            return False
        if "闸" in structure_text or "分水" in structure_text:
            return False
        if cls._is_pressure_pipe_like_node(node):
            return False

        bottom_elevation = cls._coerce_pressure_pipe_finite_float(
            getattr(node, "bottom_elevation", None)
        )
        if bottom_elevation is None or bottom_elevation <= 0:
            return False

        top_elevation = cls._coerce_pressure_pipe_finite_float(
            getattr(node, "top_elevation", None)
        )
        return top_elevation is None or top_elevation <= 0

    @classmethod
    def _collect_missing_structure_height_names(cls, nodes) -> list[str]:
        """收集真正需要结构总高但仍缺少渠顶高程的节点名称。"""
        missing_names = []
        for node in list(nodes or []):
            if not cls._should_warn_missing_structure_height(node):
                continue
            missing_names.append(str(getattr(node, "name", "") or "未命名"))
        return missing_names

    @classmethod
    def _collect_missing_siphon_name_rows(cls, nodes) -> list[int]:
        """收集倒虹吸结构中缺少建筑物名称的真实表3行号。"""
        rows = []
        for idx, node in enumerate(list(nodes or []), start=1):
            if getattr(node, "is_transition", False) or getattr(node, "is_auto_inserted_channel", False):
                continue
            structure_text = cls._get_node_structure_type_text(node)
            if "倒虹吸" not in structure_text:
                continue
            if str(getattr(node, "name", "") or "").strip():
                continue
            rows.append(idx)
        return rows

    @staticmethod
    def _build_missing_siphon_name_notice(row_indices: list[int]) -> str:
        """生成倒虹吸名称缺失提示文案。"""
        if not row_indices:
            return ""
        preview = "、".join(f"第{idx}行" for idx in row_indices[:8])
        if len(row_indices) > 8:
            preview += f" 等{len(row_indices)}行"
        return (
            "检测到倒虹吸行未填写建筑物名称，当前无法识别倒虹吸分组。\n"
            "请先回到“建筑物名称”列，为同一段倒虹吸填写相同名称后，再执行【倒虹吸水力计算】：\n"
            f"{preview}"
        )

    @staticmethod
    def _is_xxpipe_channel_level_text(channel_level: str | None) -> bool:
        return str(channel_level or "").strip() in XXPIPE_CHANNEL_LEVEL_OPTIONS

    def _get_current_channel_level_text(self, settings=None) -> str:
        level_text = str(getattr(settings, "channel_level", "") or "").strip()
        if level_text:
            return level_text

        combo = getattr(self, "channel_level_combo", None)
        current_text = getattr(combo, "currentText", None)
        if callable(current_text):
            try:
                level_text = str(current_text() or "").strip()
            except Exception:
                level_text = ""
            if level_text:
                return level_text

        builder = getattr(self, "_build_settings", None)
        if callable(builder):
            try:
                settings = builder()
            except Exception:
                settings = None
            level_text = str(getattr(settings, "channel_level", "") or "").strip()
            if level_text:
                return level_text
        return ""

    @staticmethod
    def _get_settings_station_prefix(settings=None) -> str:
        """读取当前工程桩号前缀。"""
        getter = getattr(settings, "get_station_prefix", None)
        if callable(getter):
            try:
                return str(getter() or "").strip()
            except Exception:
                return ""
        return str(getattr(settings, "station_prefix", "") or "").strip()

    @classmethod
    def _is_regular_pressure_pipe_node(cls, node) -> bool:
        return cls._get_node_structure_type_text(node) == "有压管道"

    @classmethod
    def _is_unnamed_pressure_pipe_row_node(cls, node, channel_level: str | None = None) -> bool:
        if not cls._is_regular_pressure_pipe_node(node):
            return False
        if str(getattr(node, "name", "") or "").strip():
            return False
        if cls._is_xxpipe_channel_level_text(channel_level):
            return True
        return bool(cls._get_pressure_pipe_window_override(node))

    @classmethod
    def _is_named_pressure_pipe_group_node(cls, node) -> bool:
        if not cls._is_pressure_pipe_like_node(node):
            return False
        return bool(str(getattr(node, "name", "") or "").strip())

    @staticmethod
    def _build_pressure_pipe_row_identity_from_flow_section(flow_section, row_index: int) -> str:
        flow_section = str(flow_section or "").strip()
        row_part = f"row{int(row_index) + 1}"
        if flow_section:
            return f"flow{flow_section}-{row_part}"
        return row_part

    @classmethod
    def _build_pressure_pipe_row_identity(cls, node, row_index: int) -> str:
        return cls._build_pressure_pipe_row_identity_from_flow_section(
            getattr(node, "flow_section", ""),
            row_index,
        )

    @classmethod
    def _ensure_pressure_pipe_row_identity(cls, node, row_index: int | None = None) -> str:
        identity = str(getattr(node, "pressure_pipe_row_identity", "") or "").strip()
        if not identity and row_index is not None:
            identity = cls._build_pressure_pipe_row_identity(node, row_index)
            setattr(node, "pressure_pipe_row_identity", identity)
        return identity

    @classmethod
    def _should_persist_pressure_pipe_row_identity(
        cls,
        node,
        channel_level: str | None = None,
    ) -> bool:
        """判断当前节点是否需要把承压行稳定身份落到表3元数据。"""
        if not cls._is_pressure_pipe_like_node(node):
            return False
        if str(getattr(node, "pressure_pipe_row_identity", "") or "").strip():
            return True
        return cls._is_pressure_pipe_row_override_node(node, channel_level)

    @staticmethod
    def _get_pressure_pipe_group_storage_key(group) -> str:
        """返回有压管道窗口分组稳定存储键。"""
        storage_key = str(getattr(group, "storage_key", "") or "").strip()
        if storage_key:
            return storage_key
        identity = str(getattr(group, "identity", "") or "").strip()
        if identity:
            return identity
        return str(getattr(group, "name", "") or "").strip()

    @staticmethod
    def _get_pressure_pipe_group_display_name(group) -> str:
        """返回有压管道窗口分组展示名称。"""
        display_name = str(getattr(group, "display_name", "") or "").strip()
        if display_name:
            return display_name
        return str(getattr(group, "name", "") or "").strip() or "未命名有压管道"

    @staticmethod
    def _get_pressure_pipe_group_structure_text(group) -> str:
        """返回分组结构形式文本。"""
        value = getattr(group, "structure_type", "")
        value = getattr(value, "value", value)
        return str(value or "").strip()

    @classmethod
    def _is_pressure_pipe_group_tunnel_segment(cls, group) -> bool:
        """判断分组是否为隧洞子段。"""
        return "隧洞" in cls._get_pressure_pipe_group_structure_text(group)

    @staticmethod
    def _get_pressure_pipe_group_route_key(group) -> str:
        """返回有压管道分组所属整线键。"""
        return str(getattr(group, "route_key", "") or "").strip()

    @staticmethod
    def _get_pressure_pipe_group_route_display_name(group) -> str:
        """返回有压管道分组所属整线展示名称。"""
        return str(getattr(group, "route_display_name", "") or "").strip()

    @classmethod
    def _get_pressure_pipe_group_longitudinal_storage_key(cls, group) -> str:
        """返回纵断面优先存储键。"""
        route_key = cls._get_pressure_pipe_group_route_key(group)
        if route_key:
            return route_key
        return cls._get_pressure_pipe_group_storage_key(group)

    @classmethod
    def _collect_pressure_pipe_group_export_identity_aliases(cls, group) -> list[str]:
        """收集当前分组可用于纵断面导出匹配的稳定身份别名。"""
        aliases = []
        seen = set()

        def _append(value) -> None:
            text = str(value or "").strip()
            if not text or text in seen:
                return
            seen.add(text)
            aliases.append(text)

        _append(cls._build_pressure_pipe_group_identity(group))
        _append(cls._get_pressure_pipe_group_storage_key(group))

        for node in list(getattr(group, "rows", []) or []):
            _append(getattr(node, "pressure_pipe_row_identity", ""))
        for node in list(getattr(group, "rows", []) or []):
            override = cls._get_pressure_pipe_window_override(node)
            if isinstance(override, dict):
                _append(override.get("identity", ""))
        for node in list(getattr(group, "rows", []) or []):
            _append(
                make_pressure_pipe_identity(
                    getattr(node, "flow_section", ""),
                    getattr(node, "name", ""),
                )
            )

        return aliases

    @staticmethod
    def _normalize_pressure_pipe_profile_segments(profile_segments) -> list[dict]:
        """清洗 route 级分段纵断面数据。"""
        normalized = []
        for raw in list(profile_segments or []):
            if not isinstance(raw, dict):
                continue
            segment = copy.deepcopy(raw)
            segment["segment_identity"] = str(segment.get("segment_identity", "") or "").strip()
            segment["source_kind"] = str(segment.get("source_kind", "") or "").strip()
            segment["structure_type"] = str(segment.get("structure_type", "") or "").strip()
            segment["warnings"] = list(segment.get("warnings", []) or [])
            normalized.append(segment)
        return normalized

    @classmethod
    def _find_pressure_pipe_profile_segment_for_group(
        cls,
        group,
        profile_segments,
        *,
        identity: str = "",
        storage_key: str = "",
    ) -> dict | None:
        """优先按稳定标识，再按桩号范围匹配 route 级分段。"""
        normalized_segments = cls._normalize_pressure_pipe_profile_segments(profile_segments)
        if not normalized_segments:
            return None

        candidate_keys = []
        for key in (
            identity,
            storage_key,
            getattr(group, "identity", ""),
            getattr(group, "storage_key", ""),
        ):
            key_text = str(key or "").strip()
            if key_text and key_text not in candidate_keys:
                candidate_keys.append(key_text)

        for segment in normalized_segments:
            if str(segment.get("segment_identity", "") or "").strip() in candidate_keys:
                return segment

        start_mc = getattr(group, "segment_start_mc", None)
        end_mc = getattr(group, "segment_end_mc", None)
        try:
            start_value = float(start_mc)
            end_value = float(end_mc)
        except (TypeError, ValueError):
            return None

        for segment in normalized_segments:
            try:
                seg_start = float(segment.get("start_mc"))
                seg_end = float(segment.get("end_mc"))
            except (TypeError, ValueError):
                continue
            if abs(seg_start - start_value) <= 1e-6 and abs(seg_end - end_value) <= 1e-6:
                return segment
        return None

    def _resolve_pressure_pipe_group_longitudinal_nodes(
        self,
        group,
        longitudinal_nodes_dict,
        route_profile_segments_by_key=None,
    ) -> tuple:
        """解析分组可用的整线/子段纵断面数据。"""
        route_key = self._get_pressure_pipe_group_route_key(group)
        storage_key = self._get_pressure_pipe_group_storage_key(group)
        identity = self._build_pressure_pipe_group_identity(group)
        source_key = self._get_pressure_pipe_group_longitudinal_storage_key(group)
        route_nodes = copy.deepcopy((longitudinal_nodes_dict or {}).get(source_key, []) or [])
        if not route_nodes and route_key:
            route_nodes = copy.deepcopy((longitudinal_nodes_dict or {}).get(storage_key, []) or [])
        if not route_nodes:
            route_nodes = []

        route_profile_segments = []
        if isinstance(route_profile_segments_by_key, dict):
            route_profile_segments = list(route_profile_segments_by_key.get(route_key, []) or [])
        if route_profile_segments:
            matched_segment = self._find_pressure_pipe_profile_segment_for_group(
                group,
                route_profile_segments,
                identity=identity,
                storage_key=storage_key,
            )
            if matched_segment is not None:
                segment_nodes = copy.deepcopy(matched_segment.get("longitudinal_nodes", []) or [])
                warnings = [
                    str(item or "").strip()
                    for item in list(matched_segment.get("warnings", []) or [])
                    if str(item or "").strip()
                ]
                return route_nodes, segment_nodes, "；".join(warnings)
        if not route_nodes:
            return [], [], ""

        segment_start = getattr(group, "segment_start_mc", None)
        segment_end = getattr(group, "segment_end_mc", None)
        if route_key and segment_start is not None and segment_end is not None:
            try:
                from utils.pressure_pipe_longitudinal_utils import clip_longitudinal_nodes_to_range

                segment_nodes = clip_longitudinal_nodes_to_range(route_nodes, float(segment_start), float(segment_end))
                return route_nodes, segment_nodes, ""
            except ValueError as exc:
                return route_nodes, [], str(exc)
            except Exception:
                return route_nodes, [], "整线纵断面裁切失败，已回退到平面长度"

        return route_nodes, copy.deepcopy(route_nodes), ""

    @staticmethod
    def _coerce_pressure_pipe_row_index(value, default: int = -1) -> int:
        """将有压管道相关行索引安全转换为整数，保留合法的 0。"""
        if value is None:
            return default
        if isinstance(value, str) and not value.strip():
            return default
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _normalize_pressure_pipe_window_override(value) -> dict:
        """规范化匿名有压管道窗口覆盖结构。"""
        if not isinstance(value, dict):
            return {}
        enabled = bool(value.get("enabled", False))
        identity = str(value.get("identity", "") or "").strip()
        if not enabled or not identity:
            return {}
        normalized = {
            "enabled": True,
            "identity": identity,
            "storage_key": str(value.get("storage_key", "") or identity).strip() or identity,
            "display_name": str(value.get("display_name", "") or "").strip(),
            "group_mode": str(value.get("group_mode", "") or "unnamed_row_segment").strip(),
            "data_mode": str(value.get("data_mode", "") or "").strip(),
            "applied_at": str(value.get("applied_at", "") or "").strip(),
            "calc_steps": str(value.get("calc_steps", "") or "").strip(),
            "target_row_index": WaterProfilePanel._coerce_pressure_pipe_row_index(value.get("target_row_index", -1)),
            "upstream_row_index": WaterProfilePanel._coerce_pressure_pipe_row_index(value.get("upstream_row_index", -1)),
        }
        numeric_fields = (
            "Q", "D", "total_length", "pipe_velocity", "friction_loss", "total_bend_loss",
            "local_loss", "inlet_transition_loss", "outlet_transition_loss", "total_head_loss",
        )
        for field_name in numeric_fields:
            try:
                normalized[field_name] = float(value.get(field_name, 0.0) or 0.0)
            except (TypeError, ValueError):
                normalized[field_name] = 0.0
        manual_total = value.get("manual_total_head_loss", None)
        if manual_total is not None and str(manual_total).strip() != "":
            try:
                normalized["manual_total_head_loss"] = float(manual_total)
            except (TypeError, ValueError):
                pass
        manual_source = str(value.get("manual_override_source", "") or "").strip()
        if manual_source:
            normalized["manual_override_source"] = manual_source
        manual_updated_at = str(value.get("manual_override_updated_at", "") or "").strip()
        if manual_updated_at:
            normalized["manual_override_updated_at"] = manual_updated_at
        for dict_field in ("friction_details", "bend_details", "local_details"):
            dict_value = value.get(dict_field, {})
            normalized[dict_field] = copy.deepcopy(dict_value) if isinstance(dict_value, dict) else {}
        return normalized

    @staticmethod
    def _normalize_pressure_pipe_named_group_result(value) -> dict:
        """规范化命名承压组隐藏结果元数据。"""
        if not isinstance(value, dict):
            return {}
        identity = str(value.get("identity", "") or "").strip()
        storage_key = str(value.get("storage_key", "") or identity).strip() or identity
        display_name = str(value.get("display_name", "") or "").strip()
        structure_type = str(value.get("structure_type", "") or "").strip()
        applied_at = str(value.get("applied_at", "") or "").strip()
        calc_steps = str(value.get("calc_steps", "") or "").strip()
        target_row_index = WaterProfilePanel._coerce_pressure_pipe_row_index(
            value.get("target_row_index", -1)
        )
        total_head_loss_raw = value.get("total_head_loss", None)
        total_head_loss = None
        if total_head_loss_raw is not None and str(total_head_loss_raw).strip() != "":
            try:
                total_head_loss = float(total_head_loss_raw)
            except (TypeError, ValueError):
                total_head_loss = None
        if not any((identity, storage_key, display_name, structure_type, applied_at, calc_steps)) and total_head_loss is None:
            return {}
        return {
            "identity": identity,
            "storage_key": storage_key,
            "display_name": display_name,
            "structure_type": structure_type,
            "total_head_loss": total_head_loss,
            "applied_at": applied_at,
            "calc_steps": calc_steps,
            "target_row_index": target_row_index,
        }

    @classmethod
    def _get_pressure_pipe_window_override(cls, node) -> dict:
        """读取匿名有压管道窗口覆盖。"""
        override = cls._normalize_pressure_pipe_window_override(
            getattr(node, "pressure_pipe_window_override", {})
        )
        if not override:
            section_params = getattr(node, "section_params", {}) or {}
            override = cls._normalize_pressure_pipe_window_override(
                section_params.get("pressure_pipe_window_override", {})
            )
        if override:
            setattr(node, "pressure_pipe_window_override", copy.deepcopy(override))
        return override

    @staticmethod
    def _normalize_pressure_pipe_loss_override_value(value):
        """标准化表3第38列人工采用值。"""
        if value in ("", None):
            return None
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            return None
        if not math.isfinite(numeric_value) or numeric_value < 0:
            return None
        return numeric_value

    @classmethod
    def _get_pressure_pipe_loss_override(cls, node):
        """读取表3第38列人工采用值。"""
        override_value = cls._normalize_pressure_pipe_loss_override_value(
            getattr(node, "pressure_pipe_loss_override_m", None)
        )
        if override_value is not None:
            setattr(node, "pressure_pipe_loss_override_m", override_value)
        return override_value

    @classmethod
    def _set_pressure_pipe_loss_override(cls, node, value):
        """写入或清空表3第38列人工采用值。"""
        override_value = cls._normalize_pressure_pipe_loss_override_value(value)
        setattr(node, "pressure_pipe_loss_override_m", override_value)
        if override_value is None and hasattr(node, "_pressure_pipe_display_loss"):
            setattr(node, "_pressure_pipe_display_loss", 0.0)
        return override_value

    @staticmethod
    def _is_gate_like_structure_text(structure_text) -> bool:
        """判断结构形式是否属于需要默认过闸损失的闸/分水口。"""
        text = str(structure_text or "").strip()
        return bool(text and ("闸" in text or "分水" in text))

    @classmethod
    def _is_gate_loss_user_set_node(cls, node) -> bool:
        """判断节点是否已有用户明确设置过闸损失的标记。"""
        section_params = getattr(node, "section_params", {}) or {}
        return bool(section_params.get(GATE_HEAD_LOSS_USER_SET_PARAM_KEY, False))

    @classmethod
    def _set_gate_loss_user_set(cls, node, user_set: bool):
        """写入或清空过闸损失用户设置标记。"""
        section_params = getattr(node, "section_params", None)
        if not isinstance(section_params, dict):
            section_params = {}
            setattr(node, "section_params", section_params)
        if user_set:
            section_params[GATE_HEAD_LOSS_USER_SET_PARAM_KEY] = True
        else:
            section_params.pop(GATE_HEAD_LOSS_USER_SET_PARAM_KEY, None)

    def _get_row_payload(self, row: int) -> dict:
        """读取表3行首隐藏元数据。"""
        table = getattr(self, "node_table", None)
        if not table or row < 0 or row >= table.rowCount():
            return {}
        first_item = table.item(row, 0)
        payload = first_item.data(Qt.UserRole) if first_item else None
        return dict(payload) if isinstance(payload, dict) else {}

    def _set_row_payload_value(self, row: int, key: str, value):
        """更新表3行首隐藏元数据。"""
        table = getattr(self, "node_table", None)
        if not table or row < 0 or row >= table.rowCount():
            return
        first_item = table.item(row, 0)
        if first_item is None:
            first_item = QTableWidgetItem("")
            first_item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row, 0, first_item)
        payload = first_item.data(Qt.UserRole)
        if not isinstance(payload, dict):
            payload = {}
        if value in (None, False):
            payload.pop(key, None)
        else:
            payload[key] = value
        first_item.setData(Qt.UserRole, payload if payload else None)

    def _is_gate_loss_user_set_for_row(self, row: int, node=None) -> bool:
        """判断表3某行过闸损失是否由用户明确设置。"""
        payload = self._get_row_payload(row)
        if GATE_HEAD_LOSS_USER_SET_ROLE_KEY in payload:
            return bool(payload.get(GATE_HEAD_LOSS_USER_SET_ROLE_KEY))
        table = getattr(self, "node_table", None)
        item = table.item(row, 37) if table and 0 <= row < table.rowCount() else None
        raw_text = str(item.text() if item else "").strip()
        if raw_text and raw_text != "-":
            try:
                return abs(float(raw_text)) <= ZERO_TOLERANCE
            except (TypeError, ValueError):
                return False
        if node is not None:
            return self._is_gate_loss_user_set_node(node)
        return False

    def _mark_gate_loss_user_set_for_row(self, row: int, user_set: bool = True):
        """标记表3某行过闸损失已由用户明确设置。"""
        self._set_row_payload_value(
            row,
            GATE_HEAD_LOSS_USER_SET_ROLE_KEY,
            True if user_set else None,
        )
        for nodes in (
            getattr(self, "calculated_nodes", None),
            getattr(self, "nodes", None),
        ):
            if nodes and 0 <= row < len(nodes):
                self._set_gate_loss_user_set(nodes[row], bool(user_set))

    @classmethod
    def _get_pressure_pipe_named_group_result(cls, node) -> dict:
        """读取命名承压组隐藏结果元数据。"""
        result = cls._normalize_pressure_pipe_named_group_result(
            getattr(node, "pressure_pipe_named_group_result", {})
        )
        if not result:
            section_params = getattr(node, "section_params", {}) or {}
            result = cls._normalize_pressure_pipe_named_group_result(
                section_params.get("pressure_pipe_named_group_result", {})
            )
        if result:
            setattr(node, "pressure_pipe_named_group_result", copy.deepcopy(result))
            section_params = getattr(node, "section_params", None)
            if isinstance(section_params, dict):
                section_params["pressure_pipe_named_group_result"] = copy.deepcopy(result)
        return result

    @classmethod
    def _set_pressure_pipe_named_group_result(cls, node, result) -> dict:
        """写入命名承压组隐藏结果元数据。"""
        normalized = cls._normalize_pressure_pipe_named_group_result(result)
        setattr(node, "pressure_pipe_named_group_result", copy.deepcopy(normalized))
        section_params = getattr(node, "section_params", None)
        if isinstance(section_params, dict):
            if normalized:
                section_params["pressure_pipe_named_group_result"] = copy.deepcopy(normalized)
            else:
                section_params.pop("pressure_pipe_named_group_result", None)
        return normalized

    @classmethod
    def _is_named_pressure_pipe_outlet_with_hidden_result(cls, node) -> bool:
        """判断是否为带隐藏结果元数据的命名承压组出口行。"""
        if not cls._is_named_pressure_pipe_group_node(node):
            return False
        in_out = getattr(node, "in_out", None)
        if getattr(in_out, "value", "") != "出":
            return False
        return bool(cls._get_pressure_pipe_named_group_result(node))

    @classmethod
    def _get_named_pressure_pipe_group_total_head_loss(cls, node):
        """读取命名承压组整组总损失。"""
        result = cls._get_pressure_pipe_named_group_result(node)
        if not result:
            return None
        return result.get("total_head_loss", None)

    @classmethod
    def _get_named_pressure_pipe_outlet_display_loss(cls, node) -> float:
        """计算命名承压组出口行在表3列38的显示值。"""
        total_head_loss = cls._get_named_pressure_pipe_group_total_head_loss(node)
        stored_display = float(getattr(node, "_pressure_pipe_display_loss", 0.0) or 0.0)
        if stored_display > 0:
            return stored_display

        siphon_loss = float(getattr(node, "head_loss_siphon", 0.0) or 0.0)
        if siphon_loss > 0 and (
            total_head_loss is None or abs(siphon_loss - float(total_head_loss)) > 1e-6
        ):
            setattr(node, "_pressure_pipe_display_loss", siphon_loss)
            return siphon_loss

        hydraulic_loss = (
            float(getattr(node, "head_loss_bend", 0.0) or 0.0)
            + float(getattr(node, "head_loss_friction", 0.0) or 0.0)
            + float(getattr(node, "head_loss_local", 0.0) or 0.0)
        )
        if hydraulic_loss > 0:
            setattr(node, "_pressure_pipe_display_loss", hydraulic_loss)
            return hydraulic_loss
        return 0.0

    @classmethod
    def _rebuild_named_pressure_pipe_outlet_total_loss(cls, node) -> float:
        """按逐行口径重建命名承压组出口行总损失。"""
        display_loss = cls._get_named_pressure_pipe_outlet_display_loss(node)
        total_loss = (
            display_loss
            + float(getattr(node, "head_loss_reserve", 0.0) or 0.0)
            + float(getattr(node, "head_loss_gate", 0.0) or 0.0)
        )
        node.head_loss_total = total_loss
        return total_loss

    @classmethod
    def _build_pressure_pipe_named_group_result_payload(
        cls,
        *,
        node=None,
        group=None,
        record: dict | None = None,
        row_index: int = -1,
        total_head_loss=None,
        calc_steps: str = "",
        applied_at: str = "",
    ) -> dict:
        """构造命名承压组隐藏结果元数据。"""
        record = record or {}
        identity = str(record.get("identity", "") or "").strip()
        if not identity and group is not None:
            identity = cls._build_pressure_pipe_group_identity(group)
        if not identity and node is not None:
            identity = make_pressure_pipe_identity(
                getattr(node, "flow_section", ""),
                getattr(node, "name", ""),
            )

        storage_key = str(record.get("storage_key", "") or "").strip()
        if not storage_key and group is not None:
            storage_key = cls._get_pressure_pipe_group_storage_key(group)
        storage_key = storage_key or identity

        display_name = str(record.get("display_name", "") or "").strip()
        if not display_name and group is not None:
            display_name = cls._get_pressure_pipe_group_display_name(group)
        if not display_name and node is not None:
            display_name = str(getattr(node, "name", "") or "").strip()

        structure_type = str(record.get("structure_type", "") or "").strip()
        if not structure_type and group is not None:
            structure_type = cls._get_pressure_pipe_group_structure_text(group)
        if not structure_type and node is not None:
            structure_type = cls._get_node_structure_type_text(node)

        if total_head_loss is None:
            total_head_loss = record.get("total_head_loss", None)
        if total_head_loss is not None and str(total_head_loss).strip() != "":
            try:
                total_head_loss = float(total_head_loss)
            except (TypeError, ValueError):
                total_head_loss = None
        else:
            total_head_loss = None

        if not applied_at:
            applied_at = str(record.get("applied_at", "") or "").strip()
        if not applied_at:
            applied_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not calc_steps:
            calc_steps = str(record.get("calc_steps", "") or "").strip()

        target_row_index = cls._coerce_pressure_pipe_row_index(
            record.get(
                "target_row_index",
                getattr(group, "target_row_index", getattr(group, "outlet_row_index", row_index))
                if group is not None else row_index,
            )
        )

        return cls._normalize_pressure_pipe_named_group_result({
            "identity": identity,
            "storage_key": storage_key,
            "display_name": display_name,
            "structure_type": structure_type,
            "total_head_loss": total_head_loss,
            "applied_at": applied_at,
            "calc_steps": calc_steps,
            "target_row_index": target_row_index,
        })

    @classmethod
    def _migrate_named_pressure_pipe_outlet_visible_group_loss(
        cls,
        node,
        row_index: int = -1,
    ) -> bool:
        """把旧版命名承压组出口行的可见整组损失迁入隐藏元数据。"""
        if not cls._is_named_pressure_pipe_group_node(node):
            return False
        in_out = getattr(node, "in_out", None)
        if getattr(in_out, "value", "") != "出":
            return False

        changed = False
        hidden_result = cls._get_pressure_pipe_named_group_result(node)
        hidden_total = hidden_result.get("total_head_loss", None) if hidden_result else None

        external_loss = getattr(node, "external_head_loss", None)
        legacy_total = None
        if external_loss is not None and str(external_loss).strip() != "":
            try:
                legacy_total = float(external_loss)
            except (TypeError, ValueError):
                legacy_total = None

        if hidden_total is None:
            siphon_loss = float(getattr(node, "head_loss_siphon", 0.0) or 0.0)
            if legacy_total is None and siphon_loss > 0:
                legacy_total = siphon_loss
            if legacy_total is not None and legacy_total > 0:
                payload = cls._build_pressure_pipe_named_group_result_payload(
                    node=node,
                    row_index=row_index,
                    total_head_loss=legacy_total,
                    calc_steps="legacy-visible-migration",
                )
                if payload:
                    cls._set_pressure_pipe_named_group_result(node, payload)
                    hidden_result = payload
                    hidden_total = payload.get("total_head_loss", None)
                    changed = True

        if external_loss is not None:
            node.external_head_loss = None
            changed = True

        if hidden_total is not None:
            siphon_loss = float(getattr(node, "head_loss_siphon", 0.0) or 0.0)
            if siphon_loss > 0 and abs(siphon_loss - float(hidden_total)) <= 1e-6:
                node.head_loss_siphon = 0.0
                changed = True
            stored_display = float(getattr(node, "_pressure_pipe_display_loss", 0.0) or 0.0)
            if stored_display > 0 and abs(stored_display - float(hidden_total)) <= 1e-6:
                setattr(node, "_pressure_pipe_display_loss", 0.0)
                changed = True

        if hidden_result:
            cls._rebuild_named_pressure_pipe_outlet_total_loss(node)
        return changed

    @classmethod
    def _migrate_named_pressure_pipe_group_results(cls, nodes) -> bool:
        """批量迁移命名承压组旧版出口写回数据。"""
        changed = False
        for row_index, node in enumerate(nodes or []):
            if cls._migrate_named_pressure_pipe_outlet_visible_group_loss(node, row_index=row_index):
                changed = True
        return changed

    @classmethod
    def _has_pressure_pipe_window_override(cls, node) -> bool:
        """判断节点是否存在匿名有压管道窗口覆盖。"""
        return bool(cls._get_pressure_pipe_window_override(node))

    @staticmethod
    def _get_pressure_pipe_row_override_group_modes() -> set[str]:
        """返回按 row override 展示/回写的有压结果模式。"""
        return {
            "unnamed_row_segment",
            "named_row_segment",
            "chain_row_member",
            "chain_tunnel_member",
            "chain_prefix_member",
        }

    @classmethod
    def _is_pressure_pipe_row_override_node(cls, node, channel_level: str | None = None) -> bool:
        """判断当前行是否应按 row override 口径展示有压结果。"""
        if cls._is_unnamed_pressure_pipe_row_node(node, channel_level):
            return True
        if not cls._is_pressure_pipe_like_node(node):
            return False
        override = cls._get_pressure_pipe_window_override(node)
        if not override:
            return False
        group_mode = str(override.get("group_mode", "") or "").strip()
        return group_mode in cls._get_pressure_pipe_row_override_group_modes()

    @classmethod
    def _has_pressure_pipe_row_override_result(
        cls,
        node,
        *,
        identity: str = "",
        expected_modes=None,
    ) -> bool:
        """判断当前节点是否已写回指定 row override 结果。"""
        override = cls._get_pressure_pipe_window_override(node)
        if not override:
            return False
        override_identity = str(override.get("identity", "") or "").strip()
        if identity and override_identity != identity:
            return False
        expected_mode_set = {
            str(mode or "").strip()
            for mode in list(expected_modes or [])
            if str(mode or "").strip()
        }
        group_mode = str(override.get("group_mode", "") or "").strip()
        if expected_mode_set and group_mode not in expected_mode_set:
            return False
        return True

    @staticmethod
    def _is_pressure_pipe_row_segment_group(group) -> bool:
        """判断窗口分组是否为逐行承压段。"""
        return str(getattr(group, "group_mode", "") or "").strip() in {
            "unnamed_row_segment",
            "named_row_segment",
        }

    @classmethod
    def _is_pressure_pipe_group_split_to_row_members(cls, group) -> bool:
        """判断命名有压组是否改为逐段链成员回写。"""
        if bool(getattr(group, "split_to_row_members", False)):
            return True
        split_ids = list(getattr(group, "split_row_member_identities", []) or [])
        return any(str(identity or "").strip() for identity in split_ids)

    @classmethod
    def _set_pressure_pipe_window_override(cls, node, override: dict | None):
        """写入或清空匿名有压管道窗口覆盖。"""
        normalized = cls._normalize_pressure_pipe_window_override(override or {})
        section_params = getattr(node, "section_params", None)
        if not isinstance(section_params, dict):
            section_params = {}
            node.section_params = section_params
        if normalized:
            setattr(node, "pressure_pipe_window_override", copy.deepcopy(normalized))
            section_params["pressure_pipe_window_override"] = copy.deepcopy(normalized)
        else:
            setattr(node, "pressure_pipe_window_override", {})
            section_params.pop("pressure_pipe_window_override", None)

    @classmethod
    def _get_pressure_pipe_row_manual_override_loss(cls, override: dict | None):
        """读取逐行承压手动采用值。"""
        if not isinstance(override, dict):
            return None
        raw_value = override.get("manual_total_head_loss", None)
        if raw_value is None or str(raw_value).strip() == "":
            return None
        try:
            return float(raw_value)
        except (TypeError, ValueError):
            return None

    @classmethod
    def _set_pressure_pipe_row_manual_override(
        cls,
        node,
        value,
        *,
        clear: bool = False,
        source: str = "pressure_pipe_loss_dialog",
        updated_at: str | None = None,
    ) -> bool:
        """写入或清空逐行承压手动采用值。"""
        override = cls._get_pressure_pipe_window_override(node)
        if not override:
            return False

        normalized = copy.deepcopy(override)
        if clear:
            changed = any(
                key in normalized
                for key in ("manual_total_head_loss", "manual_override_source", "manual_override_updated_at")
            )
            normalized.pop("manual_total_head_loss", None)
            normalized.pop("manual_override_source", None)
            normalized.pop("manual_override_updated_at", None)
            if changed:
                cls._set_pressure_pipe_window_override(node, normalized)
            return changed

        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            return False
        if numeric_value < 0:
            return False
        if not updated_at:
            updated_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        normalized["manual_total_head_loss"] = numeric_value
        normalized["manual_override_source"] = str(source or "pressure_pipe_loss_dialog")
        normalized["manual_override_updated_at"] = str(updated_at or "")
        cls._set_pressure_pipe_window_override(node, normalized)
        return True

    @classmethod
    def _is_pressure_pipe_display_locked_node(cls, node, channel_level: str | None = None) -> bool:
        """判断倒虹吸/有压管道损失单元格是否应锁定。"""
        return (
            cls._is_pressure_pipe_row_override_node(node, channel_level)
            and cls._has_pressure_pipe_window_override(node)
        )

    @classmethod
    def _get_unnamed_pressure_pipe_row_display_loss(cls, node) -> float:
        override = cls._get_pressure_pipe_window_override(node)
        if override:
            manual_total = cls._get_pressure_pipe_row_manual_override_loss(override)
            if manual_total is not None:
                setattr(node, "_pressure_pipe_display_loss", manual_total)
                return manual_total
            override_total = float(override.get("total_head_loss", 0.0) or 0.0)
            if override_total > 0:
                setattr(node, "_pressure_pipe_display_loss", override_total)
                return override_total

        display_loss = getattr(node, "_pressure_pipe_display_loss", None)
        if display_loss is not None:
            try:
                display_loss = float(display_loss)
            except (TypeError, ValueError):
                display_loss = 0.0
            if display_loss > 0:
                setattr(node, "_pressure_pipe_display_loss", display_loss)
                return display_loss

        siphon_loss = float(getattr(node, "head_loss_siphon", 0.0) or 0.0)
        if siphon_loss > 0:
            setattr(node, "_pressure_pipe_display_loss", siphon_loss)
            return siphon_loss

        hydraulic_loss = (
            float(getattr(node, "head_loss_bend", 0.0) or 0.0)
            + float(getattr(node, "head_loss_friction", 0.0) or 0.0)
            + float(getattr(node, "head_loss_local", 0.0) or 0.0)
        )
        if hydraulic_loss > 0:
            setattr(node, "_pressure_pipe_display_loss", hydraulic_loss)
            return hydraulic_loss

        total_loss = float(getattr(node, "head_loss_total", 0.0) or 0.0)
        reserve_loss = float(getattr(node, "head_loss_reserve", 0.0) or 0.0)
        gate_loss = float(getattr(node, "head_loss_gate", 0.0) or 0.0)
        display_loss = max(total_loss - reserve_loss - gate_loss, 0.0)
        setattr(node, "_pressure_pipe_display_loss", display_loss)
        return display_loss

    @classmethod
    def _rebuild_pressure_pipe_row_override_total_loss(
        cls,
        node,
        fallback_display_loss: float | None = None,
    ) -> float:
        """把逐行承压覆盖结果同步回正式损失字段。"""
        override = cls._get_pressure_pipe_window_override(node)
        manual_override = cls._get_pressure_pipe_loss_override(node)
        row_override_modes = cls._get_pressure_pipe_row_override_group_modes()
        friction_loss = float(getattr(node, "head_loss_friction", 0.0) or 0.0)
        bend_loss = float(getattr(node, "head_loss_bend", 0.0) or 0.0)
        local_loss = float(getattr(node, "head_loss_local", 0.0) or 0.0)
        display_loss = None
        manual_override_applied = False

        if override and str(override.get("group_mode", "") or "").strip() in row_override_modes:
            friction_loss = float(override.get("friction_loss", 0.0) or 0.0)
            bend_loss = float(override.get("total_bend_loss", 0.0) or 0.0)
            local_loss = float(override.get("local_loss", 0.0) or 0.0)
            if local_loss <= ZERO_TOLERANCE:
                local_loss = (
                    float(override.get("inlet_transition_loss", 0.0) or 0.0)
                    + float(override.get("outlet_transition_loss", 0.0) or 0.0)
                )
            manual_total = cls._get_pressure_pipe_row_manual_override_loss(override)
            if manual_total is not None:
                display_loss = manual_total
                manual_override_applied = True
            elif manual_override is None:
                display_loss = float(override.get("total_head_loss", 0.0) or 0.0)
            node.head_loss_friction = friction_loss
            node.head_loss_bend = bend_loss
            node.head_loss_local = local_loss

        if manual_override is not None:
            display_loss = manual_override
            manual_override_applied = True
        elif display_loss is None:
            try:
                display_loss = float(fallback_display_loss or 0.0)
            except (TypeError, ValueError):
                display_loss = 0.0
        if display_loss <= ZERO_TOLERANCE and not manual_override_applied:
            display_loss = max(friction_loss + bend_loss + local_loss, 0.0)

        node.head_loss_siphon = 0.0
        node.external_head_loss = None
        setattr(node, "_pressure_pipe_display_loss", display_loss)
        node.head_loss_total = (
            display_loss
            + float(getattr(node, "head_loss_reserve", 0.0) or 0.0)
            + float(getattr(node, "head_loss_gate", 0.0) or 0.0)
        )
        return display_loss

    @classmethod
    def _get_pressure_pipe_loss_display_value(
        cls,
        node,
        row_index: int | None = None,
        channel_level: str | None = None,
    ) -> float:
        manual_override = cls._get_pressure_pipe_loss_override(node)
        if manual_override is not None:
            return manual_override
        return cls._get_pressure_pipe_loss_calculated_value(
            node,
            row_index=row_index,
            channel_level=channel_level,
        )

    @classmethod
    def _get_pressure_pipe_loss_calculated_value(
        cls,
        node,
        row_index: int | None = None,
        channel_level: str | None = None,
    ) -> float:
        if cls._is_pressure_pipe_row_override_node(node, channel_level):
            cls._ensure_pressure_pipe_row_identity(node, row_index)
            return cls._get_unnamed_pressure_pipe_row_display_loss(node)

        if cls._is_regular_pressure_pipe_node(node) and not str(getattr(node, "name", "") or "").strip():
            return 0.0

        if cls._is_named_pressure_pipe_outlet_with_hidden_result(node):
            return cls._get_named_pressure_pipe_outlet_display_loss(node)

        display_loss = float(getattr(node, "head_loss_siphon", 0.0) or 0.0)
        external_loss = getattr(node, "external_head_loss", None)
        is_named_outlet = (
            cls._is_named_pressure_pipe_group_node(node)
            and getattr(node, "in_out", None) is not None
            and getattr(node.in_out, "value", "") == "出"
        )
        if display_loss <= 0 and is_named_outlet and external_loss is not None:
            try:
                display_loss = float(external_loss)
            except (TypeError, ValueError):
                display_loss = 0.0
            if display_loss > 0:
                node.head_loss_siphon = display_loss
            node.external_head_loss = None
        return display_loss

    @classmethod
    def _apply_pressure_pipe_loss_cell_to_node(
        cls,
        node,
        value,
        channel_level: str | None = None,
    ) -> float:
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            numeric_value = 0.0

        manual_override = cls._get_pressure_pipe_loss_override(node)
        effective_value = manual_override if manual_override is not None else numeric_value

        if cls._is_pressure_pipe_row_override_node(node, channel_level):
            cls._ensure_pressure_pipe_row_identity(node)
            return cls._rebuild_pressure_pipe_row_override_total_loss(
                node,
                fallback_display_loss=effective_value,
            )

        if cls._is_regular_pressure_pipe_node(node) and not str(getattr(node, "name", "") or "").strip():
            node.head_loss_siphon = 0.0
            setattr(node, "_pressure_pipe_display_loss", 0.0)
            return 0.0

        if cls._is_named_pressure_pipe_outlet_with_hidden_result(node):
            node.head_loss_siphon = 0.0
            node.external_head_loss = None
            if manual_override is None:
                setattr(node, "_pressure_pipe_display_loss", numeric_value)
            cls._rebuild_named_pressure_pipe_outlet_total_loss(node)
            return effective_value

        if manual_override is None and hasattr(node, "_pressure_pipe_display_loss"):
            setattr(node, "_pressure_pipe_display_loss", 0.0)
        if manual_override is None:
            node.head_loss_siphon = numeric_value
        if manual_override is None and numeric_value > 0 and getattr(node, "external_head_loss", None) is not None:
            node.external_head_loss = None
        node.head_loss_total = (
            float(getattr(node, "head_loss_bend", 0.0) or 0.0)
            + float(getattr(node, "head_loss_friction", 0.0) or 0.0)
            + float(getattr(node, "head_loss_local", 0.0) or 0.0)
            + float(getattr(node, "head_loss_reserve", 0.0) or 0.0)
            + float(getattr(node, "head_loss_gate", 0.0) or 0.0)
            + effective_value
        )
        return effective_value

    def _get_pressure_pipe_display_context(self, node, row_index: int | None = None) -> dict:
        channel_level = self._get_current_channel_level_text()
        calculated_loss = self._get_pressure_pipe_loss_calculated_value(
            node,
            row_index=row_index,
            channel_level=channel_level,
        )
        display_loss = self._get_pressure_pipe_loss_display_value(
            node,
            row_index=row_index,
            channel_level=channel_level,
        )
        manual_override = self._get_pressure_pipe_loss_override(node)
        row_manual_override = None
        is_row_sum = self._is_pressure_pipe_row_override_node(node, channel_level)
        if is_row_sum:
            row_manual_override = self._get_pressure_pipe_row_manual_override_loss(
                self._get_pressure_pipe_window_override(node)
            )
        effective_manual_override = (
            manual_override if manual_override is not None else row_manual_override
        )
        is_named_group_outlet = self._is_named_pressure_pipe_outlet_with_hidden_result(node)
        is_display_only = is_row_sum or is_named_group_outlet
        return {
            "channel_level": channel_level,
            "calculated_loss": calculated_loss,
            "display_loss": display_loss,
            "is_row_sum": is_row_sum,
            "is_display_only": is_display_only,
            "display_mode": "named_group_outlet" if is_named_group_outlet else ("row_sum" if is_row_sum else "normal"),
            "named_group_total_loss": self._get_named_pressure_pipe_group_total_head_loss(node)
            if is_named_group_outlet else None,
            "has_manual_override": effective_manual_override is not None,
            "manual_override_value": effective_manual_override,
            "formula_term_loss": display_loss if manual_override is not None else (0.0 if is_display_only else display_loss),
        }

    def _has_pressure_pipe_loss_details(self, node, row_index: int | None = None, pressure_pipe_ctx=None) -> bool:
        """判断当前行是否存在可展示的第38列详情。"""
        pressure_pipe_ctx = pressure_pipe_ctx or self._get_pressure_pipe_display_context(node, row_index)
        if pressure_pipe_ctx["is_row_sum"]:
            return (
                bool(self._get_pressure_pipe_window_override(node))
                or abs(float(pressure_pipe_ctx["calculated_loss"] or 0.0)) > ZERO_TOLERANCE
                or abs(float(getattr(node, "head_loss_friction", 0.0) or 0.0)) > ZERO_TOLERANCE
                or abs(float(getattr(node, "head_loss_bend", 0.0) or 0.0)) > ZERO_TOLERANCE
                or abs(float(getattr(node, "head_loss_local", 0.0) or 0.0)) > ZERO_TOLERANCE
            )
        if self._is_named_pressure_pipe_outlet_with_hidden_result(node):
            return True
        if self._is_pressure_pipe_like_node(node):
            return bool(self._get_pressure_pipe_named_group_result(node))
        return False

    @classmethod
    def _collect_named_pressure_pipe_groups(cls, nodes, settings=None):
        if not nodes:
            return []
        from utils.pressure_pipe_extractor import PressurePipeDataExtractor
        return PressurePipeDataExtractor.extract_pipes(nodes, settings=settings)

    @classmethod
    def _collect_dialog_pressure_pipe_groups(cls, nodes, settings=None):
        """提取窗口计算专用分组，优先包含匿名普通有压管道段。"""
        if not nodes:
            return []
        from utils.pressure_pipe_extractor import PressurePipeDataExtractor

        extractor = getattr(PressurePipeDataExtractor, "extract_dialog_pipe_groups", None)
        if callable(extractor):
            return extractor(nodes, settings=settings)
        return PressurePipeDataExtractor.extract_pipes(nodes, settings=settings)

    @staticmethod
    def _has_named_pressure_pipe_group_result(node) -> bool:
        """判断命名有压组落点是否已经有可用结果。"""
        hidden_result = WaterProfilePanel._get_pressure_pipe_named_group_result(node)
        if hidden_result:
            return True
        total_head_loss = getattr(node, "head_loss_siphon", None)
        if total_head_loss is not None:
            try:
                if float(total_head_loss) > 0:
                    return True
            except (TypeError, ValueError):
                pass
        return getattr(node, "external_head_loss", None) is not None

    def _build_pressure_pipe_execute_record_map(self) -> dict:
        """把最近一次有压计算结果整理成 identity -> record 的查询表。"""
        data = normalize_pressure_pipe_calc_records(getattr(self, "_pressure_pipe_calc_records", None))
        if not isinstance(data, dict):
            return {}
        return {
            str(record.get("identity", "") or "").strip(): record
            for record in data.get("records", [])
            if str(record.get("identity", "") or "").strip()
        }

    def _is_pressure_pipe_execute_member_ready(
        self,
        nodes,
        *,
        identity: str,
        target_row_index: int,
        display_name: str,
        record_map: dict,
        expected_modes=None,
    ) -> bool:
        """判断执行水面线前，某个承压成员是否已完成必要的写回。"""
        record = record_map.get(identity, {})
        status = str(record.get("status", "") or "").strip().lower()
        if status == "failed":
            return False
        if status == "success" and not bool(record.get("writeback_enabled", True)):
            return True
        if not (0 <= target_row_index < len(nodes)):
            return False
        node = nodes[target_row_index]
        if expected_modes:
            return self._has_pressure_pipe_row_override_result(
                node,
                identity=identity,
                expected_modes=expected_modes,
            )
        return self._has_named_pressure_pipe_group_result(node)

    def _collect_pending_pressure_pipe_execute_members(self, nodes, settings=None) -> list[str]:
        """收集执行计算前仍缺少有压结果的成员展示名。"""
        dialog_context = self._prepare_pressure_pipe_dialog_context(
            nodes,
            settings=settings,
            show_xxpipe_warning=False,
        )
        pipe_groups = list(dialog_context.get("pipe_groups", []) or [])
        chain_descriptors = list(dialog_context.get("chain_descriptors", []) or [])
        if not pipe_groups:
            # 与“有压管道水力计算”按钮入口保持同一口径：
            # 当前表里没有真正的有压管道同类分组时，不应让隧洞链残留成员单独拦住执行计算。
            return []
        record_map = self._build_pressure_pipe_execute_record_map()
        chain_source_lookup = self._build_pressure_chain_source_lookup(chain_descriptors)

        chain_member_lookup = {}
        for descriptor in chain_descriptors:
            for member in descriptor.get("members", []) or []:
                identity = self._get_pressure_chain_member_identity(member)
                if identity:
                    chain_member_lookup[identity] = member

        missing = []
        handled_identities = set()

        for group in pipe_groups:
            identity = self._build_pressure_pipe_group_identity(group)
            if not identity or identity in handled_identities:
                continue
            split_members = list(chain_source_lookup.get(identity, []) or [])
            member = chain_member_lookup.get(identity)
            record = record_map.get(identity, {})
            display_name = (
                str(getattr(member, "display_name", "") or "").strip()
                if member is not None
                else ""
            ) or self._get_pressure_pipe_group_display_name(group)

            if split_members or self._is_pressure_pipe_group_split_to_row_members(group):
                handled_identities.add(identity)
                continue

            if member is not None and self._is_pressure_chain_anchor_member(member):
                handled_identities.add(identity)
                continue

            expected_modes = None
            if member is not None and self._is_pressure_chain_prefix_member(member):
                target_row_index = self._coerce_pressure_pipe_row_index(
                    getattr(member, "prefix_end_row_index", getattr(member, "target_row_index", -1))
                )
                expected_modes = {"chain_prefix_member"}
            elif self._is_pressure_pipe_row_segment_group(group):
                target_row_index = self._coerce_pressure_pipe_row_index(
                    getattr(group, "target_row_index", -1)
                )
                expected_modes = {str(getattr(group, "group_mode", "") or "unnamed_row_segment").strip()}
            else:
                row_override_mode = str(
                    record.get("group_mode", "") or getattr(member, "group_mode", "") or ""
                ).strip()
                if row_override_mode in self._get_pressure_pipe_row_override_group_modes():
                    # 链内命名承压成员（如隧洞/定向钻/顶管）已按 row override 回写后，
                    # 执行前检查也要按同一口径识别，避免把已完成结果误判为未计算。
                    target_row_index = self._coerce_pressure_pipe_row_index(
                        record.get("target_row_index", getattr(member, "target_row_index", -1))
                    )
                    expected_modes = {row_override_mode}
                elif member is not None and self._is_pressure_chain_single_row_member(member):
                    target_row_index = self._coerce_pressure_pipe_row_index(
                        getattr(member, "target_row_index", -1)
                    )
                    structure_type = str(getattr(member, "structure_type", "") or "").strip()
                    expected_modes = {"chain_tunnel_member" if "隧洞" in structure_type else "chain_row_member"}
                else:
                    target_row_index = self._coerce_pressure_pipe_row_index(
                        getattr(group, "outlet_row_index", -1)
                    )

            if not self._is_pressure_pipe_execute_member_ready(
                nodes,
                identity=identity,
                target_row_index=target_row_index,
                display_name=display_name,
                record_map=record_map,
                expected_modes=expected_modes,
            ):
                missing.append(display_name)
            handled_identities.add(identity)

        for descriptor in chain_descriptors:
            for member in descriptor.get("members", []) or []:
                identity = self._get_pressure_chain_member_identity(member)
                if not identity or identity in handled_identities:
                    continue
                display_name = str(getattr(member, "display_name", "") or "").strip() or "未命名成员"
                if self._is_pressure_chain_anchor_member(member):
                    handled_identities.add(identity)
                    continue

                expected_modes = None
                if self._is_pressure_chain_prefix_member(member):
                    target_row_index = self._coerce_pressure_pipe_row_index(
                        getattr(member, "prefix_end_row_index", getattr(member, "target_row_index", -1))
                    )
                    expected_modes = {"chain_prefix_member"}
                else:
                    target_row_index = self._coerce_pressure_pipe_row_index(
                        getattr(member, "target_row_index", -1)
                    )
                    structure_type = str(getattr(member, "structure_type", "") or "").strip()
                    expected_modes = {"chain_tunnel_member" if "隧洞" in structure_type else "chain_row_member"}

                if not self._is_pressure_pipe_execute_member_ready(
                    nodes,
                    identity=identity,
                    target_row_index=target_row_index,
                    display_name=display_name,
                    record_map=record_map,
                    expected_modes=expected_modes,
                ):
                    missing.append(display_name)
                handled_identities.add(identity)

        return missing

    def _is_table1_source_locked_cell(self, row: int, col: int) -> bool:
        if col == 7 and self._is_pressure_pipe_row(row):
            return True
        if col not in TABLE1_SOURCE_LOCKED_COLS:
            return False
        return self._is_table1_source_row(row)

    def _is_forbidden_manual_flat_bottom_circle_row(self, row: int, struct_text: str) -> bool:
        """判断表3当前行是否是被禁止的手工平底圆形输入。"""
        if normalize_section_type_name(struct_text) != "隧洞-平底圆形":
            return False
        return not self._is_flat_bottom_circle_source_row(row)

    def _is_forbidden_manual_arch_culvert_row(self, row: int, struct_text: str) -> bool:
        """判断表3当前行是否是被禁止的手工圆拱直墙型暗涵输入。"""
        if normalize_section_type_name(struct_text) != ARCH_CULVERT_FAMILY_TEXT:
            return False
        return not self._is_arch_culvert_source_row(row)

    def _structure_selector_excluded_items_for_row(self, row: int) -> set[str]:
        """返回表3当前行打开结构选择器时应隐藏的结构类型。"""
        excluded_items = {"隧洞-平底圆形"}
        if not self._is_arch_culvert_source_row(row):
            excluded_items.add(ARCH_CULVERT_FAMILY_TEXT)
        return excluded_items

    @staticmethod
    def _is_flat_bottom_circle_source_payload(payload) -> bool:
        """读取平底圆形来源标记，缺失时按不允许处理。"""
        if not isinstance(payload, dict):
            return False
        if FLAT_BOTTOM_TUNNEL_SOURCE_ROLE_KEY in payload:
            return bool(payload.get(FLAT_BOTTOM_TUNNEL_SOURCE_ROLE_KEY))
        if "_from_table1_source" in payload:
            return bool(payload.get("_from_table1_source"))
        return False

    def _is_flat_bottom_circle_source_row(self, row: int) -> bool:
        """仅认带显式来源标记的平底圆形行，避免手输/粘贴被误判为合法来源。"""
        table = self.node_table
        if not table or row < 0 or row >= table.rowCount():
            return False
        first_item = table.item(row, 0)
        payload = first_item.data(Qt.UserRole) if first_item else None
        return self._is_flat_bottom_circle_source_payload(payload)

    @staticmethod
    def _is_arch_culvert_source_payload(payload) -> bool:
        """读取圆拱直墙型暗涵来源标记，兼容旧带入行。"""
        if not isinstance(payload, dict):
            return False
        if ARCH_CULVERT_SOURCE_ROLE_KEY in payload:
            return bool(payload.get(ARCH_CULVERT_SOURCE_ROLE_KEY))
        family_type = normalize_culvert_family_type_name(payload.get(CULVERT_FAMILY_TYPE_KEY, ""))
        if family_type == ARCH_CULVERT_FAMILY_TEXT and "_from_table1_source" in payload:
            return bool(payload.get("_from_table1_source"))
        return False

    def _is_arch_culvert_source_row(self, row: int) -> bool:
        """仅认带入链路明确标记的圆拱直墙型暗涵行。"""
        table = self.node_table
        if not table or row < 0 or row >= table.rowCount():
            return False
        first_item = table.item(row, 0)
        payload = first_item.data(Qt.UserRole) if first_item else None
        return self._is_arch_culvert_source_payload(payload)

    def _restore_cell_text_from_pre_edit(self, row: int, col: int) -> str:
        """按编辑前快照恢复单元格内容，并返回恢复后的文本。"""
        fallback_text = ""
        if self._pre_edit_cell_value and self._pre_edit_cell_value[:2] == (row, col):
            fallback_text = str(self._pre_edit_cell_value[2] or "")
        item = self.node_table.item(row, col)
        if item is not None:
            item.setText(fallback_text)
        return fallback_text

    def _warn_manual_flat_bottom_circle_not_allowed(self):
        """提示表3不允许手工新建平底圆形。"""
        InfoBar.warning(
            "不支持手工新选",
            "平底圆形仅允许通过表1同步或共享结果导入，表3不允许手工新选。",
            parent=self._info_parent(),
            duration=3200,
            position=InfoBarPosition.TOP,
        )

    def _warn_manual_arch_culvert_not_allowed(self):
        """提示表3不允许手工新建圆拱直墙型暗涵。"""
        InfoBar.warning(
            "不支持手工新选",
            "暗涵-圆拱直墙型仅允许通过表1同步或共享结果导入，表3不允许手工新选。",
            parent=self._info_parent(),
            duration=3200,
            position=InfoBarPosition.TOP,
        )

    def _show_table1_source_lock_hint(self):
        InfoBar.info(
            "来源列已锁定",
            "该列数据来自表1，请在“表1：基本参数输入”修改后重新同步。",
            parent=self._info_parent(), duration=2600, position=InfoBarPosition.TOP
        )

    def _is_transition_row(self, row: int, source_nodes=None) -> bool:
        nodes = source_nodes if source_nodes is not None else None
        if nodes and 0 <= row < len(nodes):
            return bool(getattr(nodes[row], 'is_transition', False))
        table = self.node_table
        if not table or row < 0 or row >= table.rowCount():
            return False
        item = table.item(row, 2)
        if not item:
            return False
        return "渐变段" in str(item.text() or "").strip()

    def _is_transition_length_editable_cell(self, row: int, col: int, source_nodes=None) -> bool:
        return col == 32 and self._is_transition_row(row, source_nodes)

    @staticmethod
    def _make_transition_length_rule_key(upstream_structure_type: str,
                                         downstream_structure_type: str,
                                         transition_type: str) -> str:
        upstream = str(upstream_structure_type or "").strip()
        downstream = str(downstream_structure_type or "").strip()
        trans_type = str(transition_type or "").strip()
        return f"{upstream}|{downstream}|{trans_type}"

    def _normalize_transition_length_rule(self, rule, *, key: str = "",
                                          upstream_structure_type: str = "",
                                          downstream_structure_type: str = "",
                                          transition_type: str = "") -> dict:
        if isinstance(rule, TransitionLengthRule):
            src = rule.to_dict()
        elif hasattr(rule, "to_dict") and callable(getattr(rule, "to_dict")):
            try:
                src = dict(rule.to_dict() or {})
            except Exception:
                src = {}
        elif hasattr(rule, "upstream_structure_type") or hasattr(rule, "rule_mode"):
            src = {
                "upstream_structure_type": getattr(rule, "upstream_structure_type", ""),
                "downstream_structure_type": getattr(rule, "downstream_structure_type", ""),
                "transition_type": getattr(rule, "transition_type", ""),
                "rule_mode": getattr(rule, "rule_mode", "formula"),
                "step_size_m": getattr(rule, "step_size_m", 0.0),
                "fixed_length_m": getattr(rule, "fixed_length_m", 0.0),
            }
        elif isinstance(rule, dict):
            src = dict(rule or {})
        else:
            src = {}
        upstream = str(src.get("upstream_structure_type", upstream_structure_type) or "").strip()
        downstream = str(src.get("downstream_structure_type", downstream_structure_type) or "").strip()
        trans_type = str(src.get("transition_type", transition_type) or "").strip()
        resolved_key = str(key or src.get("rule_key") or "").strip()
        if not resolved_key:
            resolved_key = self._make_transition_length_rule_key(upstream, downstream, trans_type)
        try:
            step_size_m = float(src.get("step_size_m", src.get("step_up", TRANSITION_LENGTH_RULE_STEP_DEFAULT)))
        except (TypeError, ValueError):
            step_size_m = TRANSITION_LENGTH_RULE_STEP_DEFAULT
        try:
            fixed_length_m = float(src.get("fixed_length_m", 0.0) or 0.0)
        except (TypeError, ValueError):
            fixed_length_m = 0.0
        rule_mode = str(src.get("rule_mode", "formula") or "formula").strip()
        if rule_mode not in {"formula", "step_up", "fixed"}:
            rule_mode = "formula"
        step_size_m = max(0.0, step_size_m)
        fixed_length_m = max(0.0, fixed_length_m)
        return {
            "rule_key": resolved_key,
            "upstream_structure_type": upstream,
            "downstream_structure_type": downstream,
            "transition_type": trans_type,
            "rule_mode": rule_mode,
            "step_size_m": step_size_m,
            "fixed_length_m": fixed_length_m,
        }

    def _normalize_transition_length_rule_map(self, rules) -> dict:
        normalized = {}
        if isinstance(rules, dict):
            iterable = rules.items()
        elif isinstance(rules, list):
            iterable = [(None, raw_rule) for raw_rule in rules]
        else:
            iterable = []
        for raw_key, raw_rule in iterable:
            rule = self._normalize_transition_length_rule(raw_rule, key=str(raw_key or "").strip())
            if rule["rule_key"]:
                normalized[rule["rule_key"]] = rule
        return normalized

    def _serialize_transition_length_rules(self) -> list:
        rules = []
        for key in sorted(self._transition_length_rules.keys()):
            normalized = self._normalize_transition_length_rule(
                self._transition_length_rules.get(key, {}),
                key=key,
            )
            if not normalized["rule_key"]:
                continue
            rules.append(
                TransitionLengthRule(
                    upstream_structure_type=normalized["upstream_structure_type"],
                    downstream_structure_type=normalized["downstream_structure_type"],
                    transition_type=normalized["transition_type"] or "出口",
                    rule_mode=normalized["rule_mode"],
                    step_size_m=normalized["step_size_m"],
                    fixed_length_m=normalized["fixed_length_m"],
                )
            )
        return rules

    def _load_transition_length_rules(self, rules) -> None:
        self._transition_length_rules = self._normalize_transition_length_rule_map(rules)

    def _has_customized_transition_length_rules(self) -> bool:
        normalized_rules = self._normalize_transition_length_rule_map(self._transition_length_rules)
        if (not normalized_rules) and getattr(self, "_settings", None):
            normalized_rules = self._normalize_transition_length_rule_map(
                getattr(self._settings, "transition_length_rules", []) or []
            )
        for rule in normalized_rules.values():
            if str(rule.get("rule_mode", "formula") or "formula").strip() != "formula":
                return True
        return False

    def _has_existing_transition_insertions(self) -> bool:
        try:
            nodes = self._build_nodes_from_table()
        except Exception:
            return False
        if not nodes:
            return False
        return any(
            getattr(node, "is_transition", False) or getattr(node, "is_auto_inserted_channel", False)
            for node in nodes
        )

    def _has_current_transition_instances(self, source_nodes=None) -> bool:
        nodes, _ = self._get_transition_nodes_for_editing(source_nodes)
        if not nodes:
            return False
        return any(getattr(node, "is_transition", False) for node in nodes)

    def _should_show_transition_length_rule_nudge(self) -> bool:
        return False

    def _clear_transition_length_rule_nudge(self) -> None:
        info_bar = getattr(self, "_length_rule_nudge_bar", None)
        self._length_rule_nudge_bar = None
        if info_bar is None:
            return
        try:
            info_bar.close()
        except Exception:
            pass

    def _show_transition_length_rule_nudge(self) -> None:
        self._show_transition_length_rules_insert_first_dialog()

    def _show_transition_length_rules_insert_first_dialog(self) -> None:
        box = MessageBox(
            "请先插入渐变段",
            "长度规则需要基于已插入的渐变段实例设置。\n\n当前还没有渐变段实例，请先执行“插入渐变段”。",
            self,
        )
        box.yesButton.setText("去插入渐变段")
        box.cancelButton.setText("关闭")
        try:
            box.yesButton.setAutoDefault(False)
            box.yesButton.setDefault(False)
            box.cancelButton.setAutoDefault(True)
            box.cancelButton.setDefault(True)
        except Exception:
            pass
        box.cancelButton.setFocus(Qt.TabFocusReason)
        if box.exec():
            self._insert_transitions()

    def _show_transition_length_rules_post_insert_bar(self, summary: str, next_step: str) -> None:
        self._clear_transition_length_rule_nudge()
        info_bar = InfoBar.new(
            InfoBarIcon.SUCCESS,
            summary,
            next_step,
            isClosable=True,
            duration=6000,
            position=InfoBarPosition.TOP,
            parent=self._info_parent(),
        )
        action_btn = PushButton("继续设置长度规则")
        action_btn.clicked.connect(
            lambda: (
                self._clear_transition_length_rule_nudge(),
                self._open_transition_length_rules(),
            )
        )
        info_bar.addWidget(action_btn)
        info_bar.destroyed.connect(lambda *_args: setattr(self, "_length_rule_nudge_bar", None))
        self._length_rule_nudge_bar = info_bar

    @staticmethod
    def _get_transition_length_source_kind(source: str) -> str:
        source = str(source or "").strip()
        if source == "override":
            return "single_override"
        if source.startswith("rule:"):
            return "combo_rule"
        if source == "formula":
            return "formula"
        return "other"

    @staticmethod
    def _get_transition_length_source_label(source: str) -> str:
        source = str(source or "").strip()
        if source == "override":
            return "单条覆盖"
        if source == "rule:step_up":
            return "组合规则-向上修约"
        if source == "rule:fixed":
            return "组合规则-固定值"
        if source == "rule:formula":
            return "组合规则-公式值"
        if source == "formula":
            return "公式/规范"
        return "当前采用值"

    @staticmethod
    def _round_up_transition_length(length: float, step_up: float) -> float:
        try:
            value = float(length)
        except (TypeError, ValueError):
            return 0.0
        try:
            step = float(step_up)
        except (TypeError, ValueError):
            step = 0.0
        if value <= 0 or step <= 0:
            return round(max(0.0, value), 3)
        multiple = math.ceil((value - 1e-9) / step)
        return round(multiple * step, 3)

    def _get_transition_nodes_for_editing(self, source_nodes=None):
        if source_nodes is not None:
            uses_calculated = bool(getattr(self, 'calculated_nodes', None) and source_nodes is self.calculated_nodes)
            return source_nodes, uses_calculated
        nodes = self._build_nodes_from_table()
        self.nodes = nodes
        return nodes, bool(getattr(self, 'calculated_nodes', None))

    @staticmethod
    def _get_node_stat_length_value(node) -> float:
        try:
            value = float(getattr(node, "stat_length", 0.0) or 0.0)
        except (TypeError, ValueError):
            value = 0.0
        return max(0.0, value)

    @staticmethod
    def _get_node_station_mc_value(node) -> float:
        try:
            value = float(getattr(node, "station_MC", 0.0) or 0.0)
        except (TypeError, ValueError):
            value = 0.0
        return value

    def _get_transition_length_override_upper_bound(self, ctx) -> float:
        if not ctx:
            return 0.0

        node = ctx["node"]
        nodes = ctx["nodes"]
        row_idx = ctx["row_idx"]
        prev_node = ctx.get("prev_node")
        next_node = ctx.get("next_node")
        prev_idx = int(ctx.get("prev_idx", -1))
        next_idx = int(ctx.get("next_idx", -1))
        base_length = 0.0
        try:
            base_length = float(getattr(node, "transition_length", 0.0) or 0.0)
        except (TypeError, ValueError):
            base_length = 0.0

        extra_slack = 0.0
        transition_type = str(getattr(node, "transition_type", "") or "").strip()
        if transition_type == "出口":
            cursor = row_idx + 1
            while cursor < len(nodes) and getattr(nodes[cursor], "is_auto_inserted_channel", False):
                extra_slack += self._get_node_stat_length_value(nodes[cursor])
                cursor += 1
        elif transition_type == "进口":
            cursor = row_idx - 1
            while cursor >= 0 and getattr(nodes[cursor], "is_auto_inserted_channel", False):
                extra_slack += self._get_node_stat_length_value(nodes[cursor])
                cursor -= 1

        # 直接夹在两个真实节点之间的渐变段，没有显式自动连接段可借时，
        # 仍应按前后真实节点的 gap 里程判断物理上限。
        if (
            extra_slack <= 1e-9
            and prev_idx >= 0
            and next_idx >= 0
            and prev_node is not None
            and next_node is not None
            and not getattr(prev_node, "is_auto_inserted_channel", False)
            and not getattr(next_node, "is_auto_inserted_channel", False)
        ):
            has_other_auxiliary = any(
                idx != row_idx
                and (
                    getattr(nodes[idx], "is_transition", False)
                    or getattr(nodes[idx], "is_auto_inserted_channel", False)
                )
                for idx in range(prev_idx + 1, next_idx)
            )
            if not has_other_auxiliary:
                gap_distance = (
                    self._get_node_station_mc_value(next_node)
                    - self._get_node_station_mc_value(prev_node)
                )
                if gap_distance > 0:
                    return max(base_length, gap_distance)

        return max(0.0, base_length + extra_slack)

    def _reject_transition_length_override(self, row_idx: int, fallback_length: float, message: str) -> bool:
        table = getattr(self, "node_table", None)
        if table and 0 <= row_idx < table.rowCount():
            item = table.item(row_idx, 32)
            if item is None:
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row_idx, 32, item)
            item.setText(f"{max(0.0, float(fallback_length or 0.0)):.3f}")
        InfoBar.warning(
            "需要重新插入渐变段",
            message,
            parent=self._info_parent(),
            duration=3500,
            position=InfoBarPosition.TOP,
        )
        return False

    def _get_transition_context_for_row(self, row_idx: int, source_nodes=None):
        nodes, uses_calculated = self._get_transition_nodes_for_editing(source_nodes)
        if not nodes or row_idx < 0 or row_idx >= len(nodes):
            return None
        node = nodes[row_idx]
        if not getattr(node, 'is_transition', False):
            return None

        prev_node = None
        next_node = None
        prev_idx = -1
        next_idx = -1
        for idx in range(row_idx - 1, -1, -1):
            if not getattr(nodes[idx], 'is_transition', False):
                prev_node = nodes[idx]
                prev_idx = idx
                break
        for idx in range(row_idx + 1, len(nodes)):
            if not getattr(nodes[idx], 'is_transition', False):
                next_node = nodes[idx]
                next_idx = idx
                break
        if prev_node is None or next_node is None:
            return None

        upstream_structure_type = str(
            getattr(node, "transition_rule_upstream_structure_type", "") or ""
        ).strip()
        downstream_structure_type = str(
            getattr(node, "transition_rule_downstream_structure_type", "") or ""
        ).strip()
        if not upstream_structure_type:
            upstream_structure_type = prev_node.get_structure_type_str() if hasattr(prev_node, "get_structure_type_str") else str(getattr(prev_node, "structure_type", "") or "")
        if not downstream_structure_type:
            downstream_structure_type = next_node.get_structure_type_str() if hasattr(next_node, "get_structure_type_str") else str(getattr(next_node, "structure_type", "") or "")
        transition_type = str(getattr(node, "transition_type", "") or "").strip()
        rule_key = self._make_transition_length_rule_key(
            upstream_structure_type,
            downstream_structure_type,
            transition_type,
        )
        return {
            "nodes": nodes,
            "uses_calculated_nodes": uses_calculated,
            "node": node,
            "row_idx": row_idx,
            "prev_node": prev_node,
            "next_node": next_node,
            "prev_idx": prev_idx,
            "next_idx": next_idx,
            "upstream_structure_type": upstream_structure_type,
            "downstream_structure_type": downstream_structure_type,
            "transition_type": transition_type,
            "rule_key": rule_key,
        }

    def _collect_transition_length_rule_rows(self, source_nodes=None):
        nodes, _ = self._get_transition_nodes_for_editing(source_nodes)
        if not nodes:
            return []

        def _get_struct_type_name(node):
            if not node or not getattr(node, "structure_type", None):
                return ""
            structure_type = node.structure_type
            return structure_type.value if hasattr(structure_type, "value") else str(structure_type or "")

        rows_by_key = {}

        def _append_rule_hit(upstream_structure_type, downstream_structure_type, transition_type, hit_scope):
            upstream_structure_type = str(upstream_structure_type or "").strip()
            downstream_structure_type = str(downstream_structure_type or "").strip()
            transition_type = str(transition_type or "").strip()
            if not transition_type or not (upstream_structure_type or downstream_structure_type):
                return
            key = self._make_transition_length_rule_key(
                upstream_structure_type,
                downstream_structure_type,
                transition_type,
            )
            entry = rows_by_key.setdefault(key, {
                "rule_key": key,
                "upstream_structure_type": upstream_structure_type,
                "downstream_structure_type": downstream_structure_type,
                "transition_type": transition_type,
                "count": 0,
                "hit_count": 0,
                "hit_scope": hit_scope,
            })
            entry["count"] += 1
            entry["hit_count"] += 1
            entry["hit_scope"] = hit_scope

        has_current_transitions = any(getattr(node, "is_transition", False) for node in nodes)
        if not has_current_transitions:
            return []

        for row_idx, node in enumerate(nodes):
            if not getattr(node, "is_transition", False):
                continue
            details = getattr(node, "transition_length_calc_details", None) or {}
            upstream_structure_type = str(
                getattr(node, "transition_rule_upstream_structure_type", "") or details.get("upstream_structure_type", "")
            ).strip()
            downstream_structure_type = str(
                getattr(node, "transition_rule_downstream_structure_type", "") or details.get("downstream_structure_type", "")
            ).strip()
            transition_type = str(getattr(node, "transition_type", "") or details.get("transition_type", "")).strip()

            if not upstream_structure_type:
                for idx in range(row_idx - 1, -1, -1):
                    candidate = nodes[idx]
                    if getattr(candidate, "is_transition", False) or getattr(candidate, "is_auto_inserted_channel", False):
                        continue
                    upstream_structure_type = _get_struct_type_name(candidate)
                    break
            if not downstream_structure_type:
                for idx in range(row_idx + 1, len(nodes)):
                    candidate = nodes[idx]
                    if getattr(candidate, "is_transition", False) or getattr(candidate, "is_auto_inserted_channel", False):
                        continue
                    downstream_structure_type = _get_struct_type_name(candidate)
                    break

            _append_rule_hit(
                upstream_structure_type,
                downstream_structure_type,
                transition_type,
                "current",
            )

        result = []
        for key in sorted(rows_by_key.keys()):
            entry = rows_by_key[key]
            stored_rule = self._normalize_transition_length_rule(
                self._transition_length_rules.get(key, {}),
                key=key,
                upstream_structure_type=entry["upstream_structure_type"],
                downstream_structure_type=entry["downstream_structure_type"],
                transition_type=entry["transition_type"],
            )
            entry["rule_mode"] = stored_rule["rule_mode"]
            entry["step_size_m"] = stored_rule["step_size_m"]
            entry["fixed_length_m"] = stored_rule["fixed_length_m"]
            result.append(entry)
        return result

    @staticmethod
    def _get_transition_length_source_kind(details) -> str:
        if not isinstance(details, dict):
            return ""
        source = str(details.get("source", details.get("length_source_kind", "")) or "").strip()
        if source in {"override", "single_override"}:
            return "single_override"
        if source.startswith("rule:"):
            return source
        return "formula"

    @classmethod
    def _get_transition_length_source_label(cls, details) -> str:
        source_kind = cls._get_transition_length_source_kind(details)
        if source_kind == "single_override":
            return "单条覆盖"
        if source_kind == "rule:step_up":
            return "组合规则-向上修约"
        if source_kind == "rule:fixed":
            return "组合规则-固定值"
        if source_kind == "rule:formula":
            return "组合规则-公式值"
        return "公式/规范"

    @classmethod
    def _build_transition_length_rule_key_from_details(cls, details) -> str:
        if not isinstance(details, dict):
            return ""
        explicit_key = str(details.get("length_rule_key", "") or "").strip()
        if explicit_key:
            return explicit_key
        upstream = str(details.get("upstream_structure_type", "") or "").strip()
        downstream = str(details.get("downstream_structure_type", "") or "").strip()
        transition_type = str(details.get("transition_type", "") or "").strip()
        if not (upstream or downstream or transition_type):
            return ""
        return cls._make_transition_length_rule_key(upstream, downstream, transition_type)

    def _build_transition_length_rule_objects(self):
        rule_objects = []
        if TransitionLengthRule is None:
            return rule_objects
        normalized_rules = self._normalize_transition_length_rule_map(self._transition_length_rules)
        if (not normalized_rules) and getattr(self, "_settings", None):
            normalized_rules = self._normalize_transition_length_rule_map(
                getattr(self._settings, "transition_length_rules", []) or []
            )
            if normalized_rules:
                self._transition_length_rules = dict(normalized_rules)
        for rule in normalized_rules.values():
            rule_objects.append(
                TransitionLengthRule(
                    upstream_structure_type=rule["upstream_structure_type"],
                    downstream_structure_type=rule["downstream_structure_type"],
                    transition_type="进口" if rule["transition_type"] == "进口" else "出口",
                    rule_mode=rule.get("rule_mode", "formula") or "formula",
                    step_size_m=max(0.0, float(rule.get("step_size_m", 0.0) or 0.0)),
                    fixed_length_m=max(0.0, float(rule.get("fixed_length_m", 0.0) or 0.0)),
                )
            )
        return rule_objects

    def _rebuild_calculation_summary_state(self, nodes):
        """在不重新执行总计算的情况下刷新摘要面板。"""
        if not nodes:
            self._update_summary_panel([])
            return

        total_len = float(getattr(self, "_last_channel_total_length", 0.0) or 0.0)
        if total_len <= 0:
            regular_nodes = [node for node in nodes if not getattr(node, "is_transition", False)]
            if len(regular_nodes) >= 2:
                start_mc = float(getattr(regular_nodes[0], "station_MC", 0.0) or 0.0)
                end_mc = float(getattr(regular_nodes[-1], "station_MC", 0.0) or 0.0)
                total_len = max(0.0, end_mc - start_mc)

        summary = None
        regular_nodes = [node for node in nodes if not getattr(node, "is_transition", False)]
        if regular_nodes:
            first_node = regular_nodes[0]
            last_node = regular_nodes[-1]
            summary = {
                "起点桩号": float(getattr(first_node, "station_MC", 0.0) or 0.0),
                "终点桩号": float(getattr(last_node, "station_MC", 0.0) or 0.0),
                "起点水位": float(getattr(first_node, "water_level", 0.0) or 0.0),
                "终点水位": float(getattr(last_node, "water_level", 0.0) or 0.0),
            }
        wl_drop = None
        if summary:
            wl_drop = summary["起点水位"] - summary["终点水位"]

        self._update_summary_panel(nodes, total_len=total_len, wl_drop=wl_drop, summary=summary)

    @staticmethod
    def _should_display_transition_length_value(node) -> bool:
        if getattr(node, "transition_length_override_m", None) is not None:
            return True
        if getattr(node, "transition_length_calc_details", None):
            return True
        return bool(getattr(node, "transition_length", 0.0))

    @staticmethod
    def _should_display_transition_loss_value(node) -> bool:
        if getattr(node, "transition_calc_details", None):
            return True
        return bool(getattr(node, "head_loss_transition", 0.0))

    def _persist_transition_calc_payload_for_row(self, row_idx: int, node=None):
        table = getattr(self, "node_table", None)
        if not table or row_idx < 0 or row_idx >= table.rowCount():
            return
        first_item = table.item(row_idx, 0)
        if first_item is None:
            first_item = QTableWidgetItem("")
            first_item.setTextAlignment(Qt.AlignCenter)
            if 0 not in EDITABLE_COLS:
                first_item.setFlags(first_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_idx, 0, first_item)
        payload = first_item.data(Qt.UserRole)
        if not isinstance(payload, dict):
            payload = {}
        node_obj = node
        if node_obj is None and getattr(self, "calculated_nodes", None) and row_idx < len(self.calculated_nodes):
            node_obj = self.calculated_nodes[row_idx]
        if node_obj is None:
            first_item.setData(Qt.UserRole, payload)
            return
        override_value = getattr(node_obj, "transition_length_override_m", None)
        if override_value is None or str(override_value).strip() == "":
            payload.pop("_transition_length_override_m", None)
        else:
            payload["_transition_length_override_m"] = float(override_value)
        payload["_transition_length_source"] = str(getattr(node_obj, "transition_length_source", "") or "")
        payload["_transition_length_warning"] = str(getattr(node_obj, "transition_length_warning", "") or "")
        payload["_transition_rule_upstream_structure_type"] = str(
            getattr(node_obj, "transition_rule_upstream_structure_type", "") or ""
        )
        payload["_transition_rule_downstream_structure_type"] = str(
            getattr(node_obj, "transition_rule_downstream_structure_type", "") or ""
        )
        payload["_transition_length_calc_details"] = copy.deepcopy(
            getattr(node_obj, "transition_length_calc_details", {}) or {}
        )
        payload["_transition_loss_calc_details"] = copy.deepcopy(
            getattr(node_obj, "transition_calc_details", {}) or {}
        )
        first_item.setData(Qt.UserRole, payload)

    def _persist_pressure_pipe_override_payload_for_row(self, row_idx: int, node=None):
        """把第38列采用值与逐行承压 override 一并写回当前行元数据。"""
        table = getattr(self, "node_table", None)
        if not table or row_idx < 0 or row_idx >= table.rowCount():
            return
        first_item = table.item(row_idx, 0)
        if first_item is None:
            first_item = QTableWidgetItem("")
            first_item.setTextAlignment(Qt.AlignCenter)
            if 0 not in EDITABLE_COLS:
                first_item.setFlags(first_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_idx, 0, first_item)
        payload = first_item.data(Qt.UserRole)
        if not isinstance(payload, dict):
            payload = {}
        node_obj = node
        if node_obj is None and getattr(self, "calculated_nodes", None) and row_idx < len(self.calculated_nodes):
            node_obj = self.calculated_nodes[row_idx]
        if node_obj is None:
            first_item.setData(Qt.UserRole, payload)
            return
        override_value = self._get_pressure_pipe_loss_override(node_obj)
        if override_value is None:
            payload.pop(PRESSURE_PIPE_LOSS_OVERRIDE_ROLE_KEY, None)
        else:
            payload[PRESSURE_PIPE_LOSS_OVERRIDE_ROLE_KEY] = float(override_value)

        current_channel_level = self._get_current_channel_level_text()
        if self._should_persist_pressure_pipe_row_identity(node_obj, current_channel_level):
            payload[PRESSURE_PIPE_ROW_ID_ROLE_KEY] = self._ensure_pressure_pipe_row_identity(node_obj, row_idx)
        else:
            payload.pop(PRESSURE_PIPE_ROW_ID_ROLE_KEY, None)

        override = self._get_pressure_pipe_window_override(node_obj)
        if override:
            payload[PRESSURE_PIPE_WINDOW_OVERRIDE_ROLE_KEY] = copy.deepcopy(override)
        else:
            payload.pop(PRESSURE_PIPE_WINDOW_OVERRIDE_ROLE_KEY, None)
        first_item.setData(Qt.UserRole, payload)

    def _persist_pressure_pipe_loss_override_payload_for_row(self, row_idx: int, node=None):
        """兼容旧调用，复用统一的第38列持久化入口。"""
        self._persist_pressure_pipe_override_payload_for_row(row_idx, node)

    def _build_transition_length_tooltip(self, details):
        if not isinstance(details, dict):
            return ""
        formula_length = details.get("formula_length", details.get("L_result", 0.0) or 0.0) or 0.0
        requested_length = details.get("requested_length", details.get("selected_length", formula_length) or 0.0) or 0.0
        physical_limit = details.get("physical_limit", 0.0) or 0.0
        actual_length = details.get("actual_length", details.get("L_result", 0.0) or 0.0) or 0.0
        lines = [
            f"来源：{self._get_transition_length_source_label(details)}",
            f"公式/规范长度：{formula_length:.3f} m",
            f"规则目标长度：{requested_length:.3f} m",
        ]
        if physical_limit > 0:
            lines.append(f"物理上限：{physical_limit:.3f} m")
        lines.append(f"最终采用长度：{actual_length:.3f} m")
        warning = str(details.get("warning", details.get("length_warning", "")) or "").strip()
        if warning:
            lines.append(f"警告：{warning}")
        rule_key = self._build_transition_length_rule_key_from_details(details)
        if rule_key:
            lines.append(f"组合键：{rule_key}")
        return "\n".join(lines)

    def _refresh_transition_length_item_presentation(self, row_idx: int, source_nodes=None):
        table = self.node_table
        if not table or row_idx < 0 or row_idx >= table.rowCount():
            return
        nodes = source_nodes if source_nodes is not None else getattr(self, 'calculated_nodes', None)
        details = None
        if nodes and row_idx < len(nodes):
            details = getattr(nodes[row_idx], 'transition_length_calc_details', None)
        tooltip = self._build_transition_length_tooltip(details)
        warning = str(details.get("warning", details.get("length_warning", "")) or "").strip() if isinstance(details, dict) else ""
        source_kind = self._get_transition_length_source_kind(details)
        for col in (32, 33):
            item = table.item(row_idx, col)
            if not item:
                continue
            item.setToolTip(tooltip)
            if warning:
                item.setBackground(QColor("#FFF4CE"))
            elif source_kind == "single_override":
                item.setBackground(QColor("#E8F4FD"))
            elif source_kind.startswith("rule:"):
                item.setBackground(QColor("#EEF7E8"))
            else:
                item.setBackground(QBrush())

    def _refresh_all_transition_length_presentations(self, source_nodes=None):
        table = getattr(self, "node_table", None)
        if not table:
            return
        for row_idx in range(table.rowCount()):
            if self._is_transition_row(row_idx, source_nodes):
                self._refresh_transition_length_item_presentation(row_idx, source_nodes)

    def _apply_table1_source_row_lock_flags(self):
        table = self.node_table
        if not table:
            return
        current_channel_level = self._get_current_channel_level_text()
        try:
            nodes = self._build_nodes_from_table()
        except Exception:
            nodes = []
        for row in range(table.rowCount()):
            is_source_row = self._is_table1_source_row(row)
            node = nodes[row] if row < len(nodes) else None
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if not item:
                    continue
                editable = col in EDITABLE_COLS
                if col == 32:
                    editable = self._is_transition_length_editable_cell(row, col)
                if row == 0 and col in FIRST_ROW_LOCKED_LOSS_COLS:
                    editable = False
                if col == 7 and self._is_pressure_pipe_row(row):
                    editable = False
                if is_source_row and col in TABLE1_SOURCE_LOCKED_COLS:
                    editable = False
                if col == 38 and node is not None and self._is_pressure_pipe_display_locked_node(node, current_channel_level):
                    editable = False
                flags = item.flags()
                if editable:
                    new_flags = flags | Qt.ItemIsEditable
                else:
                    new_flags = flags & ~Qt.ItemIsEditable
                if new_flags != flags:
                    item.setFlags(new_flags)

    def _begin_transition_length_edit(self, row_idx: int) -> bool:
        table = getattr(self, "node_table", None)
        if not table or not self._is_transition_length_editable_cell(row_idx, 32):
            return False
        item = table.item(row_idx, 32)
        if item is None:
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row_idx, 32, item)
            self._apply_table1_source_row_lock_flags()
        table.setCurrentCell(row_idx, 32)
        table.editItem(item)
        return True

    def _begin_pressure_pipe_loss_edit(self, row_idx: int) -> bool:
        """进入第38列单元格编辑。"""
        table = getattr(self, "node_table", None)
        if not table:
            return False
        item = table.item(row_idx, 38)
        if item is None:
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row_idx, 38, item)
            self._apply_table1_source_row_lock_flags()
        if not (item.flags() & Qt.ItemIsEditable):
            return False
        table.setCurrentCell(row_idx, 38)
        table.editItem(item)
        return True

    def _begin_inline_loss_edit(self, row_idx: int, col_idx: int) -> bool:
        """进入可手动填写的水头损失单元格编辑。"""
        table = getattr(self, "node_table", None)
        if not table or row_idx <= 0 or col_idx not in (36, 37):
            return False
        item = table.item(row_idx, col_idx)
        if item is None:
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row_idx, col_idx, item)
            self._apply_table1_source_row_lock_flags()
        if not (item.flags() & Qt.ItemIsEditable):
            return False
        table.setCurrentCell(row_idx, col_idx)
        table.editItem(item)
        return True

    def _begin_gate_loss_edit(self, row_idx: int) -> bool:
        """兼容旧调用：进入过闸水头损失单元格编辑。"""
        return self._begin_inline_loss_edit(row_idx, 37)

    def _show_node_table_context_menu(self, pos):
        table = getattr(self, "node_table", None)
        if not table:
            return
        index = table.indexAt(pos)
        if not index.isValid():
            return
        row = index.row()
        col = index.column()
        table.setCurrentCell(row, col)
        if col != 32 or not self._is_transition_row(row):
            return

        menu = RoundMenu(parent=self)
        menu.addAction(Action("查看渐变段长度详情", triggered=lambda: self._on_node_cell_double_clicked(row, 32)))
        if self._is_transition_length_editable_cell(row, 32):
            menu.addAction(Action("编辑渐变段长度", triggered=lambda: self._begin_transition_length_edit(row)))
        ctx = self._get_transition_context_for_row(row)
        node = ctx["node"] if ctx else None
        if node is not None and getattr(node, "transition_length_override_m", None) is not None:
            menu.addAction(
                Action(
                    "恢复公式/规则结果",
                    triggered=lambda: self._apply_transition_length_override(
                        row,
                        clear_override=True,
                        mark_dirty=True,
                    ),
                )
            )
        menu.exec(table.viewport().mapToGlobal(pos))

    @staticmethod
    def _get_transition_length_update_signature(node):
        return (
            round(float(getattr(node, "transition_length", 0.0) or 0.0), 6),
            round(float(getattr(node, "head_loss_transition", 0.0) or 0.0), 6),
        )

    def _open_transition_length_rules(self):
        self._clear_transition_length_rule_nudge()
        nodes_for_apply, uses_calculated_nodes = self._get_transition_nodes_for_editing()
        if not nodes_for_apply or not any(getattr(node, "is_transition", False) for node in nodes_for_apply):
            self._show_transition_length_rules_insert_first_dialog()
            return

        rule_rows = self._collect_transition_length_rule_rows(nodes_for_apply)
        if not rule_rows:
            fluent_info(self, "提示", "当前未找到可配置的渐变段实例")
            return
        dialog = TransitionLengthRuleDialog(rule_rows, self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        self._length_rule_nudge_seen = True
        self._transition_length_rules = self._normalize_transition_length_rule_map(dialog.get_rules())
        self._settings = self._build_settings()

        changed_rows = []
        skipped_override_rows = 0
        clamped_rows = 0
        if nodes_for_apply:
            with self._table_batch_update(self.node_table):
                for row_idx, node in enumerate(nodes_for_apply):
                    if not getattr(node, "is_transition", False):
                        continue
                    if getattr(node, "transition_length_override_m", None) is not None:
                        skipped_override_rows += 1
                        self._refresh_transition_length_item_presentation(row_idx, nodes_for_apply)
                        continue
                    before_signature = self._get_transition_length_update_signature(node)
                    if self._apply_transition_length_override(
                        row_idx,
                        source_nodes=nodes_for_apply,
                        trigger_downstream=False,
                        refresh_summary=False,
                        mark_dirty=False,
                    ):
                        details = getattr(nodes_for_apply[row_idx], "transition_length_calc_details", None) or {}
                        after_signature = self._get_transition_length_update_signature(nodes_for_apply[row_idx])
                        if before_signature != after_signature:
                            changed_rows.append(row_idx)
                        if bool(details.get("distance_clamped", False)):
                            clamped_rows += 1
                if changed_rows:
                    self._recalc_downstream(min(changed_rows))
            self._refresh_all_transition_length_presentations(nodes_for_apply)

        if uses_calculated_nodes:
            self.calculated_nodes = nodes_for_apply
        else:
            self.nodes = nodes_for_apply
        self._rebuild_calculation_summary_state(nodes_for_apply)
        if not getattr(self, "_loading_project", False):
            self.data_changed.emit()
        detail_parts = []
        if skipped_override_rows > 0:
            detail_parts.append(f"{skipped_override_rows} 条因单条覆盖未改")
        if clamped_rows > 0:
            detail_parts.append(
                f"其中 {clamped_rows} 条受物理极限约束，未能修整到目标整数；如需更长长度，请重新插入渐变段或调整结构布置"
            )
        detail_text = "；".join(detail_parts) if detail_parts else "已按当前实例重新计算并应用长度规则。"
        InfoBar.success(
            "已应用长度规则",
            f"已更新 {len(changed_rows)} 条渐变段。{detail_text}",
            parent=self._info_parent(),
            duration=4500,
            position=InfoBarPosition.TOP,
        )

    def _apply_transition_length_override(
        self,
        row_idx: int,
        manual_length=None,
        *,
        clear_override: bool = False,
        source_nodes=None,
        trigger_downstream: bool = True,
        refresh_summary: bool = True,
        mark_dirty: bool = False,
    ) -> bool:
        if not CALCULATOR_AVAILABLE:
            return False
        ctx = self._get_transition_context_for_row(row_idx, source_nodes)
        if not ctx:
            return False

        node = ctx["node"]
        nodes = ctx["nodes"]
        prev_node = ctx["prev_node"]
        next_node = ctx["next_node"]

        override_value = None
        if manual_length is not None:
            try:
                override_value = float(manual_length)
            except (TypeError, ValueError):
                InfoBar.warning(
                    "输入无效",
                    "渐变段长度请输入大于等于 0 的数值。",
                    parent=self._info_parent(),
                    duration=2500,
                    position=InfoBarPosition.TOP,
                )
                return False
            if override_value < 0:
                InfoBar.warning(
                    "输入无效",
                    "渐变段长度请输入大于等于 0 的数值。",
                    parent=self._info_parent(),
                    duration=2500,
                    position=InfoBarPosition.TOP,
                )
                return False

        settings = self._build_settings()
        self._settings = settings

        try:
            from core.hydraulic_calc import HydraulicCalculator
            hyd_calc = HydraulicCalculator(settings)
        except Exception:
            return False

        physical_limit = self._get_transition_length_override_upper_bound(ctx)
        if physical_limit <= 0:
            physical_limit = None

        if clear_override:
            node.transition_length_override_m = None
        elif override_value is not None:
            current_length = 0.0
            try:
                current_length = float(getattr(node, "transition_length", 0.0) or 0.0)
            except (TypeError, ValueError):
                current_length = 0.0
            upper_bound = physical_limit or 0.0
            if override_value > upper_bound + 1e-6:
                return self._reject_transition_length_override(
                    row_idx,
                    current_length,
                    "输入的渐变段长度超过当前可用里程，会改变现有拓扑。请重新插入渐变段或调整结构布置。",
                )
            node.transition_length_override_m = override_value

        old_updating = self._updating_cells
        self._updating_cells = True
        try:
            hyd_calc.ensure_transition_length_details(
                node,
                prev_node,
                next_node,
                nodes,
                actual_length=override_value,
                preserve_existing_length=False,
                physical_limit=physical_limit,
            )
            hyd_calc.ensure_transition_loss_details(
                node,
                prev_node,
                next_node,
                nodes,
                actual_length=override_value,
                preserve_existing_length=False,
                physical_limit=physical_limit,
            )

            length_details = getattr(node, "transition_length_calc_details", None)
            if not isinstance(length_details, dict):
                length_details = {}
                node.transition_length_calc_details = length_details
            length_details["length_rule_key"] = ctx["rule_key"]
            length_details["source"] = str(
                length_details.get("source", getattr(node, "transition_length_source", "formula")) or "formula"
            )
            length_details["warning"] = str(
                length_details.get("warning", getattr(node, "transition_length_warning", "")) or ""
            )

            loss_details = getattr(node, "transition_calc_details", None)
            if not isinstance(loss_details, dict):
                loss_details = {}
                node.transition_calc_details = loss_details
            loss_details["length_rule_key"] = ctx["rule_key"]
            loss_details["length_details"] = copy.deepcopy(length_details)

            table = self.node_table
            for col_idx in (32, 33):
                item = table.item(row_idx, col_idx)
                if item is None:
                    item = QTableWidgetItem("")
                    item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(row_idx, col_idx, item)
            table.item(row_idx, 32).setText(f"{float(getattr(node, 'transition_length', 0.0) or 0.0):.3f}")
            table.item(row_idx, 33).setText(f"{float(getattr(node, 'head_loss_transition', 0.0) or 0.0):.4f}")
            self._persist_transition_calc_payload_for_row(row_idx, node)
            self._apply_table1_source_row_lock_flags()
            self._refresh_transition_length_item_presentation(row_idx, nodes)
        finally:
            self._updating_cells = old_updating

        if ctx["uses_calculated_nodes"]:
            self.calculated_nodes = nodes
        else:
            self.nodes = nodes

        if trigger_downstream:
            self._recalc_downstream(row_idx)
        if refresh_summary:
            self._rebuild_calculation_summary_state(nodes)
        if mark_dirty and not getattr(self, "_loading_project", False):
            self.data_changed.emit()
        return True

    def _apply_pressure_pipe_loss_override(
        self,
        row_idx: int,
        manual_loss=None,
        *,
        clear_override: bool = False,
        trigger_downstream: bool = True,
        refresh_summary: bool = True,
        mark_dirty: bool = False,
    ) -> bool:
        """应用表3第38列人工采用值，并联动刷新总损失、累计损失和水位。"""
        table = getattr(self, "node_table", None)
        if not table or row_idx < 0 or row_idx >= table.rowCount():
            return False

        source_nodes = getattr(self, "calculated_nodes", None)
        if not source_nodes or row_idx >= len(source_nodes):
            source_nodes = getattr(self, "nodes", None)
        if not source_nodes or row_idx >= len(source_nodes):
            return False

        node = source_nodes[row_idx]
        override_value = None
        if not clear_override:
            try:
                override_value = float(manual_loss)
            except (TypeError, ValueError):
                InfoBar.warning(
                    "输入无效",
                    "倒虹吸/有压管道水头损失请输入大于等于 0 的数值。",
                    parent=self._info_parent(),
                    duration=2500,
                    position=InfoBarPosition.TOP,
                )
                return False
            if not math.isfinite(override_value) or override_value < 0:
                InfoBar.warning(
                    "输入无效",
                    "倒虹吸/有压管道水头损失请输入大于等于 0 的数值。",
                    parent=self._info_parent(),
                    duration=2500,
                    position=InfoBarPosition.TOP,
                )
                return False

        self._set_pressure_pipe_loss_override(node, None if clear_override else override_value)
        channel_level = self._get_current_channel_level_text()
        display_loss = self._get_pressure_pipe_loss_display_value(
            node,
            row_index=row_idx,
            channel_level=channel_level,
        )

        old_updating = self._updating_cells
        self._updating_cells = True
        try:
            item = table.item(row_idx, 38)
            if item is None:
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row_idx, 38, item)
                self._apply_table1_source_row_lock_flags()
            item.setText("-" if abs(float(display_loss or 0.0)) <= ZERO_TOLERANCE else f"{float(display_loss):.4f}")
            self._persist_pressure_pipe_loss_override_payload_for_row(row_idx, node)
        finally:
            self._updating_cells = old_updating

        if trigger_downstream:
            self._recalc_downstream(row_idx)
        if refresh_summary:
            self._rebuild_calculation_summary_state(source_nodes)
        if mark_dirty and not getattr(self, "_loading_project", False):
            self.data_changed.emit()
        return True

    def _setup_header_tooltips(self):
        """为表头设置悬浮提示（LaTeX公式渲染），使用自定义Fluent悬浮卡片"""
        from app_渠系计算前端.water_profile.formula_dialog import COLUMN_FORMULAS, FormulaTooltipWidget

        self._formula_tooltip = FormulaTooltipWidget()
        self._formula_columns = set()
        for col_idx, col_name in enumerate(NODE_ALL_HEADERS):
            if col_name in COLUMN_FORMULAS:
                self._formula_columns.add(col_idx)

        header = self.node_table.horizontalHeader()
        header.setMouseTracking(True)
        self._node_header = header
        header.viewport().setMouseTracking(True)
        header.viewport().installEventFilter(self)

    def _on_node_cell_double_clicked(self, row, col):
        """双击单元格：结构形式列弹出选择面板；水头损失/高程列显示详细计算过程"""
        if self._is_table1_source_locked_cell(row, col):
            self._show_table1_source_lock_hint()
            return
        col_name = NODE_ALL_HEADERS[col] if col < len(NODE_ALL_HEADERS) else ""

        # 结构形式列：弹出分类选择面板
        if col == 2:
            current = ""
            item = self.node_table.item(row, col)
            if item:
                current = item.text()
            dlg = StructureTypeSelector(self, excluded_items=self._structure_selector_excluded_items_for_row(row))
            dlg.set_current(current)
            result = dlg.exec()
            if result == QDialog.DialogCode.Accepted and dlg.selected_type:
                type_item = QTableWidgetItem(dlg.selected_type)
                type_item.setTextAlignment(Qt.AlignCenter)
                self.node_table.setItem(row, col, type_item)
            # 无论选择还是 Esc 取消，都回到原表格单元格，保持操作连续性
            self.node_table.setCurrentCell(row, col)
            self.node_table.setFocus(Qt.OtherFocusReason)
            return

        # 水头损失/高程列：显示详细计算过程（与原版Tkinter _on_cell_double_click对齐）
        from app_渠系计算前端.water_profile.formula_dialog import DOUBLE_CLICK_COLUMNS
        if col_name == "预留水头损失":
            self._begin_inline_loss_edit(row, 36)
            return
        if col_name == "过闸水头损失":
            self._begin_inline_loss_edit(row, 37)
            return
        if col_name not in DOUBLE_CLICK_COLUMNS:
            return
        if col_name == "渐变段长度L":
            ctx = self._get_transition_context_for_row(row)
            if not ctx:
                return
            if ctx["uses_calculated_nodes"]:
                self._sync_losses_from_table()
                self._sync_transition_lengths_from_table(ctx["nodes"])
            self._show_transition_length_details(row, ctx["node"], ctx["nodes"])
            return

        if not hasattr(self, 'calculated_nodes') or not self.calculated_nodes:
            return
        nodes = self.calculated_nodes
        if row < 0 or row >= len(nodes):
            return
        # 弹窗前强制从表格同步损失/水位/高程到 calculated_nodes，确保显示最新值
        self._sync_losses_from_table()
        self._sync_transition_lengths_from_table(nodes)
        node = nodes[row]

        if col_name == "渐变段长度L":
            self._show_transition_length_details(row, node, nodes)
        elif col_name == "弯道水头损失":
            self._show_bend_calc_details(row, node)
        elif col_name == "沿程水头损失":
            self._show_friction_calc_details(row, node)
        elif col_name == "渐变段水头损失":
            self._show_transition_calc_details(row, node)
        elif col_name == "倒虹吸/有压管道水头损失":
            if not self._show_pressure_pipe_loss_details(row, node):
                self._begin_pressure_pipe_loss_edit(row)
        elif col_name == "总水头损失":
            self._show_total_calc_details(row, node, nodes)
        elif col_name == "累计总水头损失":
            self._show_cumulative_loss_details(row, node, nodes)
        elif col_name == "水位":
            self._show_water_level_details(row, node, nodes)
        elif col_name == "渠底高程":
            self._show_bottom_elevation_details(row, node, nodes)
        elif col_name == "渠顶高程":
            self._show_top_elevation_details(row, node)

    # ================================================================
    # 双击查看详细计算过程（与原版Tkinter data_table.py完全对齐）
    # ================================================================
    def _show_transition_length_details(self, row_idx, node, source_nodes=None):
        # 双击时优先按当前表格与规则刷新一次，避免旧详情缓存继续把过期采用值带回弹窗。
        details = self._repair_transition_length_details_for_row(row_idx, source_nodes)
        if not details:
            details = getattr(node, 'transition_length_calc_details', None)
        if not details:
            fluent_info(self, "提示", "该行没有渐变段长度计算数据")
            return
        from app_渠系计算前端.water_profile.formula_dialog import show_transition_length_dialog
        try:
            show_transition_length_dialog(
                self,
                node.name or f"行{row_idx+1}",
                details,
                on_save_override=lambda value: self._apply_transition_length_override(
                    row_idx,
                    manual_length=value,
                    source_nodes=source_nodes,
                    mark_dirty=True,
                ),
                on_clear_override=lambda: self._apply_transition_length_override(
                    row_idx,
                    clear_override=True,
                    source_nodes=source_nodes,
                    mark_dirty=True,
                ),
            )
        except TypeError as exc:
            message = str(exc)
            if "unexpected keyword argument" not in message and "positional arguments" not in message:
                raise
            show_transition_length_dialog(
                self,
                node.name or f"行{row_idx+1}",
                details,
            )

    def _sync_transition_lengths_from_table(self, source_nodes=None):
        """将表格中的渐变段长度文本同步回节点模型。"""
        nodes = source_nodes if source_nodes is not None else getattr(self, 'calculated_nodes', None)
        if not nodes or not getattr(self, 'node_table', None):
            return

        row_count = min(self.node_table.rowCount(), len(nodes))
        for row_idx in range(row_count):
            item = self.node_table.item(row_idx, 32)
            if not item:
                continue
            raw_text = str(item.text() or "").strip()
            if raw_text in ("", "-"):
                nodes[row_idx].transition_length = 0.0
                continue
            try:
                length_val = float(raw_text)
            except (TypeError, ValueError):
                continue
            if length_val >= 0:
                nodes[row_idx].transition_length = length_val

    def _get_transition_length_cell_value(self, row_idx):
        """读取表格中的渐变段长度，返回 (是否显式数值, 数值)。"""
        if not getattr(self, 'node_table', None):
            return False, None
        item = self.node_table.item(row_idx, 32)
        if not item:
            return False, None
        raw_text = str(item.text() or "").strip()
        if raw_text in ("", "-"):
            return False, 0.0
        try:
            return True, float(raw_text)
        except (TypeError, ValueError):
            return False, None

    def _repair_transition_length_details_for_row(self, row_idx, source_nodes=None):
        """按当前表格/节点上下文补建或刷新某一行的渐变段长度详情。"""
        if not CALCULATOR_AVAILABLE:
            return None

        nodes = source_nodes if source_nodes is not None else getattr(self, 'calculated_nodes', None)
        if not nodes or row_idx < 0 or row_idx >= len(nodes):
            return None

        self._sync_transition_lengths_from_table(nodes)
        node = nodes[row_idx]
        if not getattr(node, 'is_transition', False):
            return None

        has_explicit_length, cell_length = self._get_transition_length_cell_value(row_idx)
        actual_length = getattr(node, 'transition_length', 0.0) or 0.0
        if has_explicit_length and cell_length is not None and cell_length >= 0:
            actual_length = cell_length
        if actual_length < 0:
            return None
        if actual_length == 0 and not has_explicit_length:
            return None

        prev_node = None
        next_node = None
        for idx in range(row_idx - 1, -1, -1):
            if not getattr(nodes[idx], 'is_transition', False):
                prev_node = nodes[idx]
                break
        for idx in range(row_idx + 1, len(nodes)):
            if not getattr(nodes[idx], 'is_transition', False):
                next_node = nodes[idx]
                break
        if prev_node is None or next_node is None:
            return None

        try:
            settings = self._build_settings()
        except Exception:
            settings = ProjectSettings()

        try:
            from core.hydraulic_calc import HydraulicCalculator
            hyd_calc = HydraulicCalculator(settings)
            details = hyd_calc.ensure_transition_length_details(
                node,
                prev_node,
                next_node,
                nodes,
                actual_length=actual_length,
                preserve_existing_length=True,
            )
            repaired_length = float(details.get("actual_length", getattr(node, "transition_length", 0.0) or 0.0) or 0.0)
            node.transition_length = repaired_length
            if getattr(self, "node_table", None):
                item = self.node_table.item(row_idx, 32)
                if item is not None:
                    item.setText(f"{repaired_length:.3f}" if repaired_length > 0 else "-")
            self._persist_transition_calc_payload_for_row(row_idx, node)
            self._refresh_transition_length_item_presentation(row_idx, nodes)
            return details
        except Exception:
            return None

    def _repair_missing_transition_length_details(self, source_nodes=None):
        """批量补建缺失的渐变段长度详情，兼容旧项目或表格回填场景。"""
        nodes = source_nodes if source_nodes is not None else getattr(self, 'calculated_nodes', None)
        if not nodes:
            return False

        repaired = False
        self._sync_transition_lengths_from_table(nodes)
        for row_idx, node in enumerate(nodes):
            if not getattr(node, 'is_transition', False):
                continue
            has_explicit_length, _ = self._get_transition_length_cell_value(row_idx)
            actual_length = getattr(node, 'transition_length', 0.0) or 0.0
            if actual_length < 0:
                continue
            if actual_length == 0 and not has_explicit_length:
                continue
            if getattr(node, 'transition_length_calc_details', None):
                continue
            if self._repair_transition_length_details_for_row(row_idx, nodes):
                repaired = True
        return repaired

    def _transition_loss_details_complete(self, details):
        required_keys = ("R1", "R2", "n", "hydraulic_slope_i", "length_details")
        return isinstance(details, dict) and all(key in details for key in required_keys)

    def _repair_transition_loss_details_for_row(self, row_idx, source_nodes=None):
        """按当前表格/节点上下文补建某一行的渐变段水头损失详情。"""
        nodes = source_nodes if source_nodes is not None else getattr(self, 'calculated_nodes', None)
        details = None
        if nodes and 0 <= row_idx < len(nodes):
            details = getattr(nodes[row_idx], 'transition_calc_details', None)

        if not CALCULATOR_AVAILABLE:
            return details
        if not nodes or row_idx < 0 or row_idx >= len(nodes):
            return details

        self._sync_transition_lengths_from_table(nodes)
        node = nodes[row_idx]
        details = getattr(node, 'transition_calc_details', None)
        if self._transition_loss_details_complete(details):
            return details
        if not getattr(node, 'is_transition', False):
            return details
        if getattr(node, 'transition_skip_loss', False):
            return details

        has_explicit_length, cell_length = self._get_transition_length_cell_value(row_idx)
        actual_length = getattr(node, 'transition_length', 0.0) or 0.0
        if has_explicit_length and cell_length is not None and cell_length >= 0:
            actual_length = cell_length
        if actual_length < 0:
            return details
        if actual_length == 0 and not has_explicit_length and not details:
            return None

        prev_node = None
        next_node = None
        for idx in range(row_idx - 1, -1, -1):
            if not getattr(nodes[idx], 'is_transition', False):
                prev_node = nodes[idx]
                break
        for idx in range(row_idx + 1, len(nodes)):
            if not getattr(nodes[idx], 'is_transition', False):
                next_node = nodes[idx]
                break
        if prev_node is None or next_node is None:
            return details

        try:
            settings = self._build_settings()
        except Exception:
            settings = ProjectSettings()

        try:
            from core.hydraulic_calc import HydraulicCalculator
            hyd_calc = HydraulicCalculator(settings)
            repaired = hyd_calc.ensure_transition_loss_details(
                node,
                prev_node,
                next_node,
                nodes,
                actual_length=actual_length if (has_explicit_length or actual_length > 0) else None,
                preserve_existing_length=has_explicit_length or actual_length > 0,
            )
            if has_explicit_length and cell_length is not None and cell_length >= 0:
                node.transition_length = cell_length
            return repaired
        except Exception:
            return details

    def _repair_missing_transition_loss_details(self, source_nodes=None):
        """批量补建缺失或旧版的渐变段水头损失详情。"""
        nodes = source_nodes if source_nodes is not None else getattr(self, 'calculated_nodes', None)
        if not nodes:
            return False

        repaired = False
        self._sync_transition_lengths_from_table(nodes)
        for row_idx, node in enumerate(nodes):
            if not getattr(node, 'is_transition', False):
                continue
            if getattr(node, 'transition_skip_loss', False):
                continue
            details = getattr(node, 'transition_calc_details', None)
            if self._transition_loss_details_complete(details):
                continue
            result = self._repair_transition_loss_details_for_row(row_idx, nodes)
            if self._transition_loss_details_complete(result):
                repaired = True
        return repaired

    def _repair_bend_loss_details_for_row(self, row_idx, source_nodes=None):
        """按当前节点上下文补建某一行的弯道水头损失详情。"""
        nodes = source_nodes if source_nodes is not None else getattr(self, 'calculated_nodes', None)
        details = None
        if nodes and 0 <= row_idx < len(nodes):
            details = getattr(nodes[row_idx], 'bend_calc_details', None)

        if not CALCULATOR_AVAILABLE:
            return details
        if not nodes or row_idx < 0 or row_idx >= len(nodes):
            return details

        node = nodes[row_idx]
        details = getattr(node, 'bend_calc_details', None)
        if details:
            return details

        existing_loss = float(getattr(node, 'head_loss_bend', 0.0) or 0.0)
        if existing_loss <= 0 and float(getattr(node, 'arc_length', 0.0) or 0.0) <= 0:
            return details

        try:
            settings = self._build_settings()
        except Exception:
            settings = ProjectSettings()

        try:
            from core.hydraulic_calc import HydraulicCalculator
            hyd_calc = HydraulicCalculator(settings)
            hyd_calc.calculate_bend_loss(node)
            details = getattr(node, 'bend_calc_details', None)
            if details and existing_loss > 0:
                node.head_loss_bend = existing_loss
                details['hw'] = existing_loss
            return details
        except Exception:
            return details

    def _show_bend_calc_details(self, row_idx, node):
        details = getattr(node, 'bend_calc_details', None)
        if not details:
            repaired = self._repair_bend_loss_details_for_row(row_idx)
            if repaired:
                details = repaired
        if not details:
            fluent_info(self, "提示", "该行没有弯道水头损失计算数据")
            return
        from app_渠系计算前端.water_profile.formula_dialog import show_bend_loss_dialog
        show_bend_loss_dialog(self, node.name or f"行{row_idx+1}", details)

    def _show_friction_calc_details(self, row_idx, node):
        if not getattr(node, 'friction_calc_details', None):
            fluent_info(self, "提示", "该行没有沿程水头损失计算数据")
            return
        from app_渠系计算前端.water_profile.formula_dialog import show_friction_loss_dialog
        show_friction_loss_dialog(self, node.name or f"行{row_idx+1}", node.friction_calc_details)

    def _show_transition_calc_details(self, row_idx, node):
        if not getattr(node, 'is_transition', False):
            fluent_info(self, "提示", "该行不是渐变段，无法显示详细计算过程")
            return
        details = getattr(node, 'transition_calc_details', None)
        if not self._transition_loss_details_complete(details):
            repaired = self._repair_transition_loss_details_for_row(row_idx)
            if repaired:
                details = repaired
        if not details:
            fluent_info(self, "提示", "该渐变段尚未计算水头损失")
            return
        from app_渠系计算前端.water_profile.formula_dialog import show_transition_loss_dialog
        show_transition_loss_dialog(self, node.name or f"行{row_idx+1}", details)

    def _show_total_calc_details(self, row_idx, node, nodes):
        if getattr(node, 'is_transition', False):
            self._show_transition_calc_details(row_idx, node)
            return
        pressure_pipe_ctx = self._get_pressure_pipe_display_context(node, row_idx)
        details = {
            'head_loss_bend': node.head_loss_bend or 0.0,
            'head_loss_transition': 0.0,
            'head_loss_friction': node.head_loss_friction or 0.0,
            'head_loss_local': getattr(node, 'head_loss_local', 0.0) or 0.0,
            'head_loss_reserve': getattr(node, 'head_loss_reserve', 0.0) or 0.0,
            'head_loss_gate': getattr(node, 'head_loss_gate', 0.0) or 0.0,
            'head_loss_siphon': pressure_pipe_ctx['formula_term_loss'],
            'pressure_pipe_calc_loss': pressure_pipe_ctx['calculated_loss'],
            'pressure_pipe_display_loss': pressure_pipe_ctx['display_loss'],
            'pressure_pipe_display_is_row_sum': pressure_pipe_ctx['is_row_sum'],
            'pressure_pipe_display_is_display_only': pressure_pipe_ctx['is_display_only'],
            'pressure_pipe_display_mode': pressure_pipe_ctx['display_mode'],
            'pressure_pipe_named_group_total': pressure_pipe_ctx['named_group_total_loss'],
            'pressure_pipe_display_has_manual_override': pressure_pipe_ctx['has_manual_override'],
            'pressure_pipe_manual_override_value': pressure_pipe_ctx['manual_override_value'],
            'head_loss_total': node.head_loss_total or 0.0,
        }
        from app_渠系计算前端.water_profile.formula_dialog import show_total_loss_dialog
        show_total_loss_dialog(self, node.name or f"行{row_idx+1}", details)

    def _show_cumulative_loss_details(self, row_idx, node, nodes):
        cumulative = 0.0
        lines = []
        channel_level = self._get_current_channel_level_text()
        for i, n in enumerate(nodes):
            if i > row_idx:
                break
            if n.is_transition:
                loss = n.head_loss_transition or 0.0
                if loss <= 0 and getattr(n, 'transition_calc_details', None):
                    loss = n.transition_calc_details.get('total', 0.0) or 0.0
                cumulative += loss
                lines.append(f"第{i+1}行(渐变段):  $h_{{tr}} = {loss:.4f}$ m，累计 $= {cumulative:.4f}$ m")
            else:
                loss = n.head_loss_total or 0.0
                cumulative += loss
                # 构建分项明细（含用户手动输入的预留/过闸/倒虹吸）
                parts = []
                hw = n.head_loss_bend or 0.0
                hf = n.head_loss_friction or 0.0
                hj = getattr(n, 'head_loss_local', 0.0) or 0.0
                hr = getattr(n, 'head_loss_reserve', 0.0) or 0.0
                hg = getattr(n, 'head_loss_gate', 0.0) or 0.0
                hs = self._get_pressure_pipe_loss_display_value(
                    n,
                    row_index=i,
                    channel_level=channel_level,
                )
                pressure_pipe_ctx = self._get_pressure_pipe_display_context(n, i)
                if hw: parts.append(f"弯道{hw:.4f}")
                if hf: parts.append(f"沿程{hf:.4f}")
                if hj: parts.append(f"局部{hj:.4f}")
                if hr: parts.append(f"预留{hr:.4f}")
                if hg: parts.append(f"过闸{hg:.4f}")
                if hs and pressure_pipe_ctx["formula_term_loss"] > 0:
                    parts.append(f"倒虹吸{hs:.4f}")
                detail = f"（{'＋'.join(parts)}）" if parts else ""
                lines.append(f"第{i+1}行(普通):  $h_{{\\Sigma}} = {loss:.4f}$ m{detail}，累计 $= {cumulative:.4f}$ m")
        from app_渠系计算前端.water_profile.formula_dialog import show_cumulative_loss_dialog
        show_cumulative_loss_dialog(self, node.name or f"行{row_idx+1}",
                                    {"cumulative": cumulative, "rows_text": "\n".join(lines)})

    @staticmethod
    def _get_loss_value_for_cumulative_detail(node) -> float:
        if getattr(node, 'is_transition', False):
            loss = float(getattr(node, 'head_loss_transition', 0.0) or 0.0)
            if loss <= 0 and getattr(node, 'transition_calc_details', None):
                loss = float(node.transition_calc_details.get('total', 0.0) or 0.0)
            return loss
        return float(getattr(node, 'head_loss_total', 0.0) or 0.0)

    def _calculate_exact_cumulative_loss_for_row(self, nodes, row_idx: int) -> float:
        if not nodes or row_idx < 0:
            return 0.0
        exact_cumulative = 0.0
        for i in range(min(row_idx, len(nodes) - 1) + 1):
            exact_cumulative += self._get_loss_value_for_cumulative_detail(nodes[i])
        return round(exact_cumulative, 6)

    def _show_water_level_details(self, row_idx, node, nodes):
        if getattr(node, 'is_transition', False):
            fluent_info(self, "提示", "渐变段行不显示水位")
            return
        # 找第一个常规节点和上一个常规节点
        first_regular_idx = None
        for i, n in enumerate(nodes):
            if not n.is_transition:
                first_regular_idx = i
                break
        prev_idx = None
        for i in range(row_idx - 1, -1, -1):
            if not nodes[i].is_transition:
                prev_idx = i
                break
        is_first = (first_regular_idx == row_idx)
        is_gate = bool(getattr(node, 'is_diversion_gate', False))
        settings = self._build_settings()
        start_level = settings.start_water_level if settings else 0.0
        details = {
            "is_first": is_first,
            "is_gate": is_gate,
            "water_level": node.water_level or 0.0,
            "start_level": start_level,
            "cumulative": node.head_loss_cumulative or 0.0,
            "total_loss": node.head_loss_total or 0.0,
            "start_level_exact": round(float(start_level or 0.0), 6),
        }
        if is_first:
            details["cumulative_exact"] = self._calculate_exact_cumulative_loss_for_row(nodes, row_idx)
            details["water_level_exact"] = round(float(start_level or 0.0), 6)
        elif prev_idx is not None:
            details["prev_level"] = nodes[prev_idx].water_level or 0.0
            prev_cumulative_exact = self._calculate_exact_cumulative_loss_for_row(nodes, prev_idx)
            prev_level_exact = round(float(start_level or 0.0) - prev_cumulative_exact, 6)
            details["prev_level_exact"] = prev_level_exact
            if is_gate:
                head_loss_gate = getattr(node, 'head_loss_gate', 0.0) or 0.0
                details["head_loss_gate"] = head_loss_gate
                details["head_loss_gate_exact"] = round(float(head_loss_gate or 0.0), 6)
                cumulative_exact = self._calculate_exact_cumulative_loss_for_row(nodes, row_idx)
                details["cumulative_exact"] = cumulative_exact
                details["water_level_exact"] = round(float(start_level or 0.0) - cumulative_exact, 6)
            else:
                hf = node.head_loss_friction or 0.0
                hj = getattr(node, 'head_loss_local', 0.0) or 0.0
                hw = node.head_loss_bend or 0.0
                h_reserve = getattr(node, 'head_loss_reserve', 0.0) or 0.0
                h_gate = getattr(node, 'head_loss_gate', 0.0) or 0.0
                pressure_pipe_ctx = self._get_pressure_pipe_display_context(node, row_idx)
                row_total_loss = node.head_loss_total or 0.0

                transition_step_loss = 0.0
                for j in range(prev_idx + 1, row_idx):
                    if nodes[j].is_transition:
                        transition_step_loss += nodes[j].head_loss_transition or 0.0

                details["hf"] = hf
                details["hj"] = hj
                details["hw"] = hw
                details["h_tr"] = transition_step_loss
                details["h_reserve"] = h_reserve
                details["h_gate"] = h_gate
                details["h_siphon"] = pressure_pipe_ctx["formula_term_loss"]
                details["pressure_pipe_calc_loss"] = pressure_pipe_ctx["calculated_loss"]
                details["pressure_pipe_display_loss"] = pressure_pipe_ctx["display_loss"]
                details["pressure_pipe_display_is_row_sum"] = pressure_pipe_ctx["is_row_sum"]
                details["pressure_pipe_display_is_display_only"] = pressure_pipe_ctx["is_display_only"]
                details["pressure_pipe_display_mode"] = pressure_pipe_ctx["display_mode"]
                details["pressure_pipe_named_group_total"] = pressure_pipe_ctx["named_group_total_loss"]
                details["pressure_pipe_display_has_manual_override"] = pressure_pipe_ctx["has_manual_override"]
                details["pressure_pipe_manual_override_value"] = pressure_pipe_ctx["manual_override_value"]
                details["total_loss"] = round(row_total_loss, 4)
                details["transition_step_loss"] = round(transition_step_loss, 4)
                details["step_drop"] = round(row_total_loss + transition_step_loss, 4)
                details["total_loss_exact"] = round(float(row_total_loss or 0.0), 6)
                details["transition_step_loss_exact"] = round(float(transition_step_loss or 0.0), 6)
                details["step_drop_exact"] = round(float(row_total_loss or 0.0) + float(transition_step_loss or 0.0), 6)
                cumulative_exact = self._calculate_exact_cumulative_loss_for_row(nodes, row_idx)
                details["cumulative_exact"] = cumulative_exact
                details["water_level_exact"] = round(float(start_level or 0.0) - cumulative_exact, 6)
        else:
            fluent_info(self, "提示", "该行无法获取上一节点水位")
            return
        from app_渠系计算前端.water_profile.formula_dialog import show_water_level_dialog
        show_water_level_dialog(self, node.name or f"行{row_idx+1}", details)

    def _apply_pressure_pipe_row_manual_override_change(
        self,
        row_idx: int,
        node,
        *,
        manual_value=None,
        clear: bool = False,
        mark_dirty: bool = True,
    ) -> bool:
        """将锁定逐行承压行的手动采用值写回表3并触发后续递推。"""
        channel_level = self._get_current_channel_level_text()
        if not self._is_pressure_pipe_display_locked_node(node, channel_level):
            return False

        changed = self._set_pressure_pipe_row_manual_override(
            node,
            manual_value,
            clear=clear,
            source="pressure_pipe_loss_dialog",
        )
        if not changed:
            return False
        # 锁定逐行承压行统一走窗口 override，避免旧的普通人工采用值继续抢优先级。
        self._set_pressure_pipe_loss_override(node, None)

        display_loss = self._get_pressure_pipe_loss_display_value(
            node,
            row_index=row_idx,
            channel_level=channel_level,
        )

        if hasattr(self, "_snapshot_editable_cols") and hasattr(self, "_append_loss_undo_snapshot"):
            self._append_loss_undo_snapshot(self._snapshot_editable_cols())

        table = getattr(self, "node_table", None)
        old_updating = getattr(self, "_updating_cells", False)
        self._updating_cells = True
        try:
            if table is not None and row_idx >= 0 and row_idx < table.rowCount():
                item = table.item(row_idx, 38)
                if item is None:
                    item = QTableWidgetItem("")
                    item.setTextAlignment(Qt.AlignCenter)
                    if 38 not in EDITABLE_COLS:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    table.setItem(row_idx, 38, item)
                item.setText(f"{float(display_loss or 0.0):.4f}")
                if self._is_pressure_pipe_display_locked_node(node, channel_level):
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self._persist_pressure_pipe_override_payload_for_row(row_idx, node)
        finally:
            self._updating_cells = old_updating

        self._apply_pressure_pipe_loss_cell_to_node(
            node,
            display_loss,
            channel_level=channel_level,
        )
        self._recalc_downstream(row_idx)
        self._refresh_pressure_pipe_controls()
        if mark_dirty and not getattr(self, "_loading_project", False):
            self.data_changed.emit()
        return True

    def _show_pressure_pipe_loss_details(self, row_idx, node):
        pressure_pipe_ctx = self._get_pressure_pipe_display_context(node, row_idx)
        if not self._has_pressure_pipe_loss_details(node, row_idx, pressure_pipe_ctx):
            return False
        channel_level = self._get_current_channel_level_text()
        editable_override = self._is_pressure_pipe_display_locked_node(node, channel_level)
        manual_override_value = None
        calculated_loss = float(pressure_pipe_ctx.get("calculated_loss", pressure_pipe_ctx.get("display_loss", 0.0)) or 0.0)
        display_loss = float(pressure_pipe_ctx.get("display_loss", calculated_loss) or 0.0)
        has_manual_override = bool(pressure_pipe_ctx.get("has_manual_override", False))
        ctx_manual_override_value = pressure_pipe_ctx.get("manual_override_value", None)
        if editable_override:
            manual_override_value = self._get_pressure_pipe_row_manual_override_loss(
                self._get_pressure_pipe_window_override(node)
            )
        details = {
            "head_loss_bend": node.head_loss_bend or 0.0,
            "head_loss_friction": node.head_loss_friction or 0.0,
            "head_loss_local": getattr(node, "head_loss_local", 0.0) or 0.0,
            "head_loss_siphon": calculated_loss,
            "pressure_pipe_calc_loss": calculated_loss,
            "pressure_pipe_display_loss": display_loss,
            "pressure_pipe_display_is_row_sum": bool(pressure_pipe_ctx.get("is_row_sum", False)),
            "pressure_pipe_display_is_display_only": bool(pressure_pipe_ctx.get("is_display_only", False)),
            "pressure_pipe_display_mode": str(pressure_pipe_ctx.get("display_mode", "normal") or "normal"),
            "pressure_pipe_named_group_total": pressure_pipe_ctx.get("named_group_total_loss"),
            "pressure_pipe_display_has_manual_override": (
                has_manual_override or manual_override_value is not None
            ),
            "pressure_pipe_manual_override_value": (
                manual_override_value
                if manual_override_value is not None
                else ctx_manual_override_value
            ),
        }
        from app_渠系计算前端.water_profile.formula_dialog import show_pressure_pipe_loss_dialog
        show_pressure_pipe_loss_dialog(
            self,
            node.name or f"行{row_idx+1}",
            details,
            editable_override=editable_override,
            manual_override_value=manual_override_value,
            on_save_override=(
                lambda value: self._apply_pressure_pipe_row_manual_override_change(
                    row_idx,
                    node,
                    manual_value=value,
                    mark_dirty=True,
                )
            ) if editable_override else (
                lambda value: self._apply_pressure_pipe_loss_override(
                    row_idx,
                    manual_loss=value,
                    mark_dirty=True,
                )
            ),
            on_clear_override=(
                lambda: self._apply_pressure_pipe_row_manual_override_change(
                    row_idx,
                    node,
                    clear=True,
                    mark_dirty=True,
                )
            ) if editable_override else (
                lambda: self._apply_pressure_pipe_loss_override(
                    row_idx,
                    clear_override=True,
                    mark_dirty=True,
                )
            ),
        )
        return True

    def _collect_terminal_gate_backfill_records(self, nodes=None):
        """收集末尾闸行高程回推记录。"""
        source_nodes = nodes if nodes is not None else (self.calculated_nodes or [])
        records = []
        for node in source_nodes:
            details = getattr(node, "terminal_gate_backfill_details", None) or {}
            if details.get("attempted"):
                records.append(details)
        return records

    @staticmethod
    def _format_terminal_gate_backfill_target(details):
        row = details.get("target_row", 0)
        name = details.get("target_name", f"行{row}") or f"行{row}"
        struct_type = details.get("target_structure_type", "闸类")
        return f"第 {row} 行 {name}（{struct_type}）"

    @staticmethod
    def _format_terminal_gate_backfill_source(details):
        row = details.get("donor_row")
        if not row:
            return "未找到有效参考断面"
        name = details.get("donor_name", f"行{row}") or f"行{row}"
        struct_type = details.get("donor_structure_type", "参考断面")
        return f"第 {row} 行 {name}（{struct_type}）"

    def _build_terminal_gate_backfill_notice_lines(self, nodes=None):
        """构建计算完成后的末尾闸行回推提示。"""
        records = self._collect_terminal_gate_backfill_records(nodes)
        lines = []
        for details in records:
            target = self._format_terminal_gate_backfill_target(details)
            source = self._format_terminal_gate_backfill_source(details)
            bottom_ok = bool((details.get("bottom") or {}).get("success"))
            top_ok = bool((details.get("top") or {}).get("success"))
            if details.get("status") == "success":
                fields = []
                if bottom_ok:
                    fields.append("渠底高程")
                if top_ok:
                    fields.append("渠顶高程")
                lines.append(
                    f"末尾闸行 {target} 已按回推规则补齐{'、'.join(fields)}，参考来源：{source}"
                )
                continue
            if details.get("status") == "partial":
                fields = []
                if bottom_ok:
                    fields.append("渠底高程")
                if top_ok:
                    fields.append("渠顶高程")
                failure_reason = details.get("failure_reason") or (
                    (details.get("bottom") or {}).get("failure_reason")
                    or (details.get("top") or {}).get("failure_reason")
                    or "未记录失败原因"
                )
                lines.append(
                    f"末尾闸行 {target} 已部分回推{'、'.join(fields)}，参考来源：{source}；未完成部分原因：{failure_reason}"
                )
                continue
            failure_reason = details.get("failure_reason") or (
                (details.get("bottom") or {}).get("failure_reason")
                or (details.get("top") or {}).get("failure_reason")
                or "未记录失败原因"
            )
            lines.append(
                f"末尾闸行 {target} 未能完成高程回推，参考搜索范围为同流量段上游明渠/矩形暗涵；原因：{failure_reason}"
            )
        return lines

    def _build_terminal_gate_backfill_report_lines(self, nodes=None):
        """构建详细过程/导出中的末尾闸行回推说明。"""
        records = self._collect_terminal_gate_backfill_records(nodes)
        if not records:
            return []

        lines = [
            "=" * 80,
            "  末尾闸行高程回推说明",
            "-" * 80,
        ]
        for details in records:
            target = self._format_terminal_gate_backfill_target(details)
            source = self._format_terminal_gate_backfill_source(details)
            lines.append(f"  目标闸行: {target}")
            lines.append(
                f"  搜索范围: 同流量段 {details.get('target_flow_section') or '-'} 上游明渠/矩形暗涵"
            )
            lines.append(f"  参考来源: {source}")

            bottom = details.get("bottom") or {}
            if bottom.get("attempted"):
                if bottom.get("success"):
                    lines.append(
                        "  渠底高程: "
                        f"Zb = {details.get('target_water_level', 0.0):.3f} - "
                        f"{details.get('donor_water_depth', 0.0):.3f} = "
                        f"{bottom.get('result', 0.0):.3f} m"
                    )
                else:
                    lines.append(
                        f"  渠底高程: 未回推。原因：{bottom.get('failure_reason') or '未记录失败原因'}"
                    )

            top = details.get("top") or {}
            if top.get("attempted"):
                if top.get("success"):
                    lines.append(
                        "  渠顶高程: "
                        f"Zt = {top.get('base_bottom_elevation', 0.0):.3f} + "
                        f"{details.get('donor_structure_height', 0.0):.3f} = "
                        f"{top.get('result', 0.0):.3f} m"
                    )
                else:
                    lines.append(
                        f"  渠顶高程: 未回推。原因：{top.get('failure_reason') or '未记录失败原因'}"
                    )
            lines.append("")
        return lines

    def _show_bottom_elevation_details(self, row_idx, node, nodes):
        if getattr(node, 'is_transition', False):
            fluent_info(self, "提示", "渐变段行不显示渠底高程")
            return
        gate_backfill = getattr(node, 'terminal_gate_backfill_details', None) or {}
        if gate_backfill.get("attempted") and (gate_backfill.get("bottom") or {}).get("attempted"):
            from app_渠系计算前端.water_profile.formula_dialog import show_terminal_gate_backfill_bottom_dialog
            show_terminal_gate_backfill_bottom_dialog(self, node.name or f"行{row_idx+1}", gate_backfill)
            return
        # 倒虹吸出口节点：使用公式10.3.6专用弹窗
        try:
            from 推求水面线.models.data_models import StructureType as ST, InOutType as IOT
            if (node.structure_type == ST.INVERTED_SIPHON
                    and node.in_out == IOT.OUTLET
                    and getattr(node, 'siphon_outlet_elev_details', None)):
                from app_渠系计算前端.water_profile.formula_dialog import show_siphon_outlet_elevation_dialog
                show_siphon_outlet_elevation_dialog(self, node.name or f"行{row_idx+1}", node.siphon_outlet_elev_details)
                return
            if (node.structure_type == ST.INVERTED_SIPHON
                    and node.in_out == IOT.INLET and node.bottom_elevation):
                fluent_info(self, "渠底高程说明",
                            f"倒虹吸进口渠底高程取自上游渠道末端的渠底高程：\n\n渠底高程 = {node.bottom_elevation:.3f} m")
                return
        except ImportError:
            pass
        wd = node.water_depth or 0.0
        if wd <= 0:
            fluent_info(self, "提示", "该行没有水深数据，无法计算渠底高程")
            return
        from app_渠系计算前端.water_profile.formula_dialog import show_bottom_elevation_dialog
        show_bottom_elevation_dialog(self, node.name or f"行{row_idx+1}",
                                      {"water_level": node.water_level or 0.0,
                                       "water_depth": wd,
                                       "bottom_elevation": node.bottom_elevation or 0.0})

    def _show_top_elevation_details(self, row_idx, node):
        if getattr(node, 'is_transition', False):
            fluent_info(self, "提示", "渐变段行不显示渠顶高程")
            return
        gate_backfill = getattr(node, 'terminal_gate_backfill_details', None) or {}
        if gate_backfill.get("attempted") and (gate_backfill.get("top") or {}).get("attempted"):
            from app_渠系计算前端.water_profile.formula_dialog import show_terminal_gate_backfill_top_dialog
            show_terminal_gate_backfill_top_dialog(self, node.name or f"行{row_idx+1}", gate_backfill)
            return
        struct_text = node.structure_type.value if getattr(node, "structure_type", None) else ""
        io_text = node.in_out.value if getattr(node, "in_out", None) else ""
        if struct_text == "倒虹吸" and io_text in ("进", "出") and getattr(node, "top_elevation", 0.0):
            source = "上游渠道末端" if io_text == "进" else "下游渠道起始"
            io_label = "进口" if io_text == "进" else "出口"
            fluent_info(
                self,
                "渠顶高程说明",
                f"倒虹吸{io_label}渠顶高程取自{source}的渠顶高程：\n\n"
                f"渠顶高程 = {node.top_elevation:.3f} m",
            )
            return
        sh = node.structure_height or 0.0
        if sh <= 0:
            fluent_info(self, "提示", "该行没有结构高度数据，无法计算渠顶高程")
            return
        be = node.bottom_elevation or 0.0
        te = node.top_elevation or 0.0
        if be == 0 and te == 0:
            fluent_info(self, "提示", "该行没有渠底高程数据，无法计算渠顶高程")
            return
        from app_渠系计算前端.water_profile.formula_dialog import show_top_elevation_dialog
        show_top_elevation_dialog(self, node.name or f"行{row_idx+1}",
                                   {"bottom_elevation": be, "structure_height": sh, "top_elevation": te})

    def _build_result_area(self, parent):
        """兼容旧调用：v3.2 起统一走流程标签工作区。"""
        self._build_workspace_area(parent)

    def _show_help(self):
        lines = [
            "=" * 70,
            "  推求水面线 — 使用说明",
            "=" * 70, "",
            "操作步骤：",
            "  1. 填写基础设置（渠道名称、起始水位、流量等）",
            "  2. 在表1填写断面参数并点击「断面批量计算」",
            "  3. 断面全成功后，系统自动同步表1+表2到表3",
            "  4. 核对基础设置和渐变段设置",
            "  5. 点击「插入渐变段」",
            "  6.（如有倒虹吸）点击「倒虹吸水力计算」",
            "  7.（如有有压管道）点击「有压管道水力计算」",
            "  8. 点击「执行计算」",
            "  9. 查看结果表格和详细过程", "",
            "同步与锁定规则：",
            "  - 仅断面全成功时允许自动同步到表3",
            "  - 表1存在失败行时，后续操作保持锁定",
            "  - 表1任意编辑后，表3结果立即失效并重新锁定", "",
            "多流量段支持：",
            "  - 设计流量和加大流量支持逗号分隔的多值输入",
            "  - 例如：5.0, 8.0, 10.0",
            "  - 修改设计流量后自动计算对应加大流量", "",
            "节点数据说明：",
            "  - 流量段：所属流量段编号（如1、2）",
            "  - 建筑物名称：节点名称（如隧洞1、渡槽2）",
            "  - 结构形式：双击选择断面类型（含闸·分水等）",
            "  - X/Y：平面坐标（用于几何计算）",
            "  - 转弯半径：行内留空或填 0 时按 0 处理",
            "  - 顶部转弯半径栏位：仅在点击「应用」后才统一覆盖真实导入行",
            "  - 底宽B/直径D/半径R/边坡m：断面几何参数",
            "  - 糙率n/底坡1/i/流量Q：水力参数（留空使用全局设置）", "",
            "转弯半径取值规范：",
            "  - 隧洞：弯曲半径≥洞径(或洞宽)×5",
            "  - 明渠：弯曲半径≥水面宽度×5",
            "  - 渡槽：弯道半径≥连接明渠渠底宽度×5",
            "  - 倒虹吸：R = n × D（独立设置）",
            "=" * 70,
        ]
        self.detail_text.setPlainText("\n".join(lines))

    # ================================================================
    # 节点表操作
    # ================================================================
    def _sync_losses_from_table(self):
        """从表格读取最新的水头损失/水位/高程值，同步到 calculated_nodes。
        确保双击弹窗始终显示用户手动编辑后的最新数据。"""
        if not hasattr(self, 'calculated_nodes') or not self.calculated_nodes:
            return
        table = self.node_table
        row_count = table.rowCount()
        channel_level = self._get_current_channel_level_text()

        def _rf(r, c):
            item = table.item(r, c)
            if not item:
                return 0.0
            txt = item.text().strip()
            if not txt or txt == '-':
                return 0.0
            try:
                return float(txt)
            except ValueError:
                return 0.0

        for r in range(min(row_count, len(self.calculated_nodes))):
            node = self.calculated_nodes[r]
            if node.is_transition:
                node.head_loss_transition = _rf(r, 33) or node.head_loss_transition
                node.head_loss_cumulative = _rf(r, 40) or node.head_loss_cumulative
            else:
                # 可编辑损失列（用户可能手动修改）
                node.head_loss_reserve = _rf(r, 36)
                node.head_loss_gate = _rf(r, 37)
                self._set_gate_loss_user_set(node, self._is_gate_loss_user_set_for_row(r, node))
                self._apply_pressure_pipe_loss_cell_to_node(node, _rf(r, 38), channel_level=channel_level)
                # 联动计算列
                node.head_loss_total = _rf(r, 39) or node.head_loss_total
                node.head_loss_cumulative = _rf(r, 40) or node.head_loss_cumulative
                wl = _rf(r, 41)
                if wl:
                    node.water_level = wl
                be = _rf(r, 42)
                if be:
                    node.bottom_elevation = be
                te = _rf(r, 43)
                if te:
                    node.top_elevation = te

    def _on_current_cell_changed(self, row, col, prev_row, prev_col):
        """当用户切换单元格时，记录编辑前的快照（供撤销使用）"""
        if self._updating_cells:
            return
        # 重置批量操作标志（Delete 删除结束）
        self._undo_group = 0
        # 任何单元格切换时都记录快照（与批量计算面板保持一致）
        self._pre_edit_snapshot = self._snapshot_editable_cols()
        if col in EDITABLE_COLS:
            item = self.node_table.item(row, col)
            self._pre_edit_cell_value = (row, col, item.text() if item else "")

    def _snapshot_editable_cols(self):
        """保存所有可编辑列的快照（用于撤销）"""
        snapshot = {}
        for r in range(self.node_table.rowCount()):
            for c in EDITABLE_COLS:
                item = self.node_table.item(r, c)
                snapshot[(r, c)] = item.text() if item else ""
            # 也保存联动计算列（39-43）
            for c in range(39, 44):
                item = self.node_table.item(r, c)
                snapshot[(r, c)] = item.text() if item else ""
        return snapshot

    def _on_loss_cell_changed(self, row, col):
        """当用户编辑可编辑列时，记录撤销快照；若为水头损失列则联动更新"""
        if self._updating_cells:
            return
        # 仅对可编辑列触发
        if col not in EDITABLE_COLS:
            return
        
        # 触发数据变化信号（除非正在加载项目）
        if not getattr(self, '_loading_project', False):
            self.data_changed.emit()
        
        self._updating_cells = True
        try:
            # 如果在批量操作中（如 Delete 键删除），跳过快照记录。
            # _undo_group 记录本轮需要跳过的 cellChanged 次数，逐次递减归零。
            skip_snapshot = self._undo_group > 0
            if skip_snapshot:
                self._undo_group -= 1

            if not skip_snapshot:
                # 如果没有预先记录的快照，先生成一个
                if self._pre_edit_snapshot is None:
                    self._pre_edit_snapshot = self._snapshot_editable_cols()
                
                # 编辑单元格已经是新值，用 _pre_edit_cell_value 还原旧值
                if self._pre_edit_cell_value and self._pre_edit_cell_value[:2] == (row, col):
                    self._pre_edit_snapshot[(row, col)] = self._pre_edit_cell_value[2]
                
                self._append_loss_undo_snapshot(self._pre_edit_snapshot)
                self._pre_edit_snapshot = None

            if col == 2:
                item = self.node_table.item(row, col)
                struct_text = str(item.text() if item else "").strip()
                if self._is_forbidden_manual_flat_bottom_circle_row(row, struct_text):
                    self._restore_cell_text_from_pre_edit(row, col)
                    self._warn_manual_flat_bottom_circle_not_allowed()
                    return
                if self._is_forbidden_manual_arch_culvert_row(row, struct_text):
                    self._restore_cell_text_from_pre_edit(row, col)
                    self._warn_manual_arch_culvert_not_allowed()
                    return
            if col == 32:
                if not self._is_transition_length_editable_cell(row, col):
                    return
                item = self.node_table.item(row, col)
                raw_text = str(item.text() if item else "").strip()
                if raw_text in ("", "-"):
                    self._apply_transition_length_override(
                        row,
                        clear_override=True,
                        mark_dirty=False,
                    )
                else:
                    try:
                        manual_length = float(raw_text)
                    except (TypeError, ValueError):
                        fallback_text = ""
                        if self._pre_edit_cell_value and self._pre_edit_cell_value[:2] == (row, col):
                            fallback_text = self._pre_edit_cell_value[2]
                        if item is not None:
                            item.setText(fallback_text)
                        InfoBar.warning(
                            "输入无效",
                            "渐变段长度请输入大于等于 0 的数值。",
                            parent=self._info_parent(),
                            duration=2500,
                            position=InfoBarPosition.TOP,
                        )
                        return
                    if manual_length < 0:
                        fallback_text = ""
                        if self._pre_edit_cell_value and self._pre_edit_cell_value[:2] == (row, col):
                            fallback_text = self._pre_edit_cell_value[2]
                        if item is not None:
                            item.setText(fallback_text)
                        InfoBar.warning(
                            "输入无效",
                            "渐变段长度请输入大于等于 0 的数值。",
                            parent=self._info_parent(),
                            duration=2500,
                            position=InfoBarPosition.TOP,
                        )
                        return
                    self._apply_transition_length_override(
                        row,
                        manual_length=manual_length,
                        mark_dirty=False,
                    )
            # 对于水头损失列（36, 37, 38），触发联动计算
            elif col in (36, 37, 38) and row > 0:
                if col == 37:
                    self._mark_gate_loss_user_set_for_row(row, True)
                    self._recalc_downstream(row)
                elif col == 38:
                    item = self.node_table.item(row, col)
                    raw_text = str(item.text() if item else "").strip()
                    manual_loss = 0.0
                    if raw_text not in ("", "-"):
                        try:
                            manual_loss = float(raw_text)
                        except (TypeError, ValueError):
                            fallback_text = ""
                            if self._pre_edit_cell_value and self._pre_edit_cell_value[:2] == (row, col):
                                fallback_text = self._pre_edit_cell_value[2]
                            if item is not None:
                                item.setText(fallback_text)
                            InfoBar.warning(
                                "输入无效",
                                "倒虹吸/有压管道水头损失请输入大于等于 0 的数值。",
                                parent=self._info_parent(),
                                duration=2500,
                                position=InfoBarPosition.TOP,
                            )
                            return
                        if not math.isfinite(manual_loss) or manual_loss < 0:
                            fallback_text = ""
                            if self._pre_edit_cell_value and self._pre_edit_cell_value[:2] == (row, col):
                                fallback_text = self._pre_edit_cell_value[2]
                            if item is not None:
                                item.setText(fallback_text)
                            InfoBar.warning(
                                "输入无效",
                                "倒虹吸/有压管道水头损失请输入大于等于 0 的数值。",
                                parent=self._info_parent(),
                                duration=2500,
                                position=InfoBarPosition.TOP,
                            )
                            return
                    self._apply_pressure_pipe_loss_override(
                        row,
                        manual_loss=manual_loss,
                        mark_dirty=False,
                    )
                else:
                    self._recalc_downstream(row)
        finally:
            self._updating_cells = False
        # 更新 pre_edit 为当前新值，以便连续编辑同一单元格时也能撤销
        item = self.node_table.item(row, col)
        self._pre_edit_cell_value = (row, col, item.text() if item else "")
        if col in TRANSITION_PREPARATION_RELEVANT_COLS:
            self._transition_topology_prepared = False

        self._refresh_pressure_pipe_controls()
        if col in (2, 24):
            nodes_for_view = self._build_nodes_from_table()
            self._update_pressure_pipe_roughness_overview(
                self._collect_pressure_pipe_roughness_pairs_from_nodes(nodes_for_view)
            )

    def _append_loss_undo_snapshot(self, snapshot):
        """将单元格编辑快照压入撤销栈（统一入口）。"""
        self._loss_undo_stack.append(snapshot)
        if len(self._loss_undo_stack) > 20:
            self._loss_undo_stack.pop(0)
        self._loss_redo_stack.clear()

    def _push_undo_snapshot(self):
        """记录当前表格状态到撤销栈（Delete 键删除前调用）"""
        snapshot = self._snapshot_editable_cols()
        self._append_loss_undo_snapshot(snapshot)
        self._pre_edit_snapshot = None
        # 设置标志，跳过本次 Delete 触发的 N 次 cellChanged 快照记录
        editable_count = 0
        for idx in self.node_table.selectedIndexes():
            item = self.node_table.item(idx.row(), idx.column())
            if item is not None and (item.flags() & Qt.ItemIsEditable):
                editable_count += 1
        self._undo_group = max(1, editable_count)

    def _undo_loss_edit(self):
        """Ctrl+Z 撤销：优先撤销行操作，其次撤销单元格编辑"""
        # 优先检查行操作撤销栈
        if self._node_table_undo_stack:
            self._undo_node_table()
            return
        # 再检查单元格编辑撤销栈
        if not self._loss_undo_stack:
            return
        self._updating_cells = True
        try:
            snapshot = self._loss_undo_stack.pop()
            # 保存当前状态到重做栈
            current = {}
            for (r, c) in snapshot.keys():
                item = self.node_table.item(r, c)
                current[(r, c)] = item.text() if item else ""
            self._loss_redo_stack.append(current)
            if len(self._loss_redo_stack) > 20:
                self._loss_redo_stack.pop(0)
            table = self.node_table
            for (r, c), text in snapshot.items():
                item = table.item(r, c)
                if item is None:
                    item = QTableWidgetItem("")
                    item.setTextAlignment(Qt.AlignCenter)
                    if c not in EDITABLE_COLS:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    table.setItem(r, c, item)
                item.setText(text)

            # 同步 calculated_nodes
            def _rf(r, c):
                item = table.item(r, c)
                if not item:
                    return 0.0
                txt = item.text().strip()
                if not txt or txt == '-':
                    return 0.0
                try:
                    return float(txt)
                except ValueError:
                    return 0.0

            if hasattr(self, 'calculated_nodes') and self.calculated_nodes:
                channel_level = self._get_current_channel_level_text()
                for r in range(min(table.rowCount(), len(self.calculated_nodes))):
                    node = self.calculated_nodes[r]
                    if node.is_transition:
                        node.head_loss_cumulative = _rf(r, 40)
                    else:
                        node.head_loss_reserve = _rf(r, 36)
                        node.head_loss_gate = _rf(r, 37)
                        self._set_gate_loss_user_set(node, self._is_gate_loss_user_set_for_row(r, node))
                        self._apply_pressure_pipe_loss_cell_to_node(node, _rf(r, 38), channel_level=channel_level)
                        node.head_loss_total = _rf(r, 39)
                        node.head_loss_cumulative = _rf(r, 40)
                        node.water_level = _rf(r, 41)
                        be = _rf(r, 42)
                        if be:
                            node.bottom_elevation = be
                        te = _rf(r, 43)
                        if te:
                            node.top_elevation = te

            InfoBar.success("已撤销", "已恢复上一步操作",
                           parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
        finally:
            self._updating_cells = False

    def _redo_loss_edit(self):
        """Ctrl+Y 重做：优先重做行操作，其次重做单元格编辑"""
        # 优先检查行操作重做栈
        if self._node_table_redo_stack:
            self._redo_node_table()
            return
        # 再检查单元格编辑重做栈
        if not self._loss_redo_stack:
            return
        self._updating_cells = True
        try:
            snapshot = self._loss_redo_stack.pop()
            # 保存当前状态到撤销栈
            current = {}
            for (r, c) in snapshot.keys():
                item = self.node_table.item(r, c)
                current[(r, c)] = item.text() if item else ""
            self._loss_undo_stack.append(current)
            if len(self._loss_undo_stack) > 20:
                self._loss_undo_stack.pop(0)
            table = self.node_table
            for (r, c), text in snapshot.items():
                item = table.item(r, c)
                if item is None:
                    item = QTableWidgetItem("")
                    item.setTextAlignment(Qt.AlignCenter)
                    if c not in EDITABLE_COLS:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    table.setItem(r, c, item)
                item.setText(text)

            # 同步 calculated_nodes
            def _rf(r, c):
                item = table.item(r, c)
                if not item:
                    return 0.0
                txt = item.text().strip()
                if not txt or txt == '-':
                    return 0.0
                try:
                    return float(txt)
                except ValueError:
                    return 0.0

            if hasattr(self, 'calculated_nodes') and self.calculated_nodes:
                channel_level = self._get_current_channel_level_text()
                for r in range(min(table.rowCount(), len(self.calculated_nodes))):
                    node = self.calculated_nodes[r]
                    if node.is_transition:
                        node.head_loss_cumulative = _rf(r, 40)
                    else:
                        node.head_loss_reserve = _rf(r, 36)
                        node.head_loss_gate = _rf(r, 37)
                        self._set_gate_loss_user_set(node, self._is_gate_loss_user_set_for_row(r, node))
                        self._apply_pressure_pipe_loss_cell_to_node(node, _rf(r, 38), channel_level=channel_level)
                        node.head_loss_total = _rf(r, 39)
                        node.head_loss_cumulative = _rf(r, 40)
                        node.water_level = _rf(r, 41)
                        be = _rf(r, 42)
                        if be:
                            node.bottom_elevation = be
                        te = _rf(r, 43)
                        if te:
                            node.top_elevation = te

            InfoBar.success("已重做", "已恢复上一步撤销的操作",
                           parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
        finally:
            self._updating_cells = False

    # ================================================================
    # 节点表行操作撤销（添加/删除/插入/复制/清空）
    # ================================================================
    def _snapshot_node_table(self):
        """保存完整节点表状态（用于行操作撤销）"""
        snapshot = {
            'rows': [],
            'row_meta': [],
            'structure_heights': dict(self._node_structure_heights),
            'chamfer_params': dict(self._node_chamfer_params),
            'u_params': dict(self._node_u_params),
            'velocity_increased': dict(self._node_velocity_increased),
            'section_sync_ready': bool(getattr(self, "_section_sync_ready", False)),
            'transition_topology_prepared': bool(getattr(self, "_transition_topology_prepared", False)),
            'section_state_text': (self._section_state_label.text().strip() if self._section_state_label else ""),
            'section_status_kind': str(getattr(self, "_section_status_kind", "neutral") or "neutral"),
            'siphon_manager_config': self._snapshot_manager_config(getattr(self, "_siphon_manager", None)),
            'pressure_pipe_manager_config': self._snapshot_manager_config(getattr(self, "_pressure_pipe_manager", None)),
        }
        for r in range(self.node_table.rowCount()):
            row_data = []
            for c in range(self.node_table.columnCount()):
                item = self.node_table.item(r, c)
                row_data.append(item.text() if item else "")
            snapshot['rows'].append(row_data)
            first_item = self.node_table.item(r, 0)
            snapshot['row_meta'].append(copy.deepcopy(first_item.data(Qt.UserRole)) if first_item else None)
        return snapshot

    @staticmethod
    def _snapshot_manager_config(manager):
        if manager is None:
            return None
        to_dict = getattr(manager, "to_dict", None)
        if not callable(to_dict):
            return None
        try:
            import copy
            return copy.deepcopy(to_dict())
        except Exception:
            return None

    @staticmethod
    def _restore_manager_config(manager, config_data):
        if manager is None or not isinstance(config_data, dict):
            return
        from_dict = getattr(manager, "from_dict", None)
        if not callable(from_dict):
            return
        try:
            from_dict(config_data)
        except Exception:
            return
        save_fn = getattr(manager, "save_config", None)
        if callable(save_fn):
            try:
                save_fn()
            except Exception:
                pass

    def _restore_section_gate_from_snapshot(self, snapshot):
        if not isinstance(snapshot, dict):
            return
        sync_ready = bool(snapshot.get('section_sync_ready', getattr(self, "_section_sync_ready", False)))
        self._transition_topology_prepared = bool(
            snapshot.get('transition_topology_prepared', getattr(self, "_transition_topology_prepared", False))
        )
        state_text = str(snapshot.get('section_state_text', "") or "").strip()
        status_kind = str(snapshot.get('section_status_kind', "") or "").strip()
        if not state_text:
            state_text = (
                "状态：断面全成功，表1+表2已同步到表3"
                if sync_ready else
                "状态：断面结果未就绪，请先执行断面批量计算"
            )
        if status_kind not in ("success", "error", "warning", "neutral"):
            status_kind = ""
        self._section_sync_ready = sync_ready
        self._set_downstream_actions_enabled(sync_ready, state_text=state_text, status_kind=status_kind)

    @staticmethod
    def _resolve_loaded_section_gate_state(
        merged_section,
        node_row_count: int,
        *,
        calculated_node_count: int = 0,
        result_row_count: int = 0,
    ) -> tuple[bool, bool]:
        """兼容旧项目恢复下游门禁状态。"""
        if int(node_row_count or 0) <= 0:
            return False, False
        if not isinstance(merged_section, dict) or not merged_section:
            return True, True

        state_text = str(merged_section.get("state_text", "") or "").strip()
        has_restored_results = (
            int(calculated_node_count or 0) > 0
            or int(result_row_count or 0) > 0
        )
        stale_markers = ("未就绪", "请先", "请重新", "已变更", "已更新", "锁定", "失败", "未生成")

        if "sync_ready" in merged_section:
            sync_ready = bool(merged_section.get("sync_ready", False))
            if sync_ready:
                return True, False
            if has_restored_results and not any(marker in state_text for marker in stale_markers):
                return True, True
            return False, False
        return True, True

    def _restore_node_table(self, snapshot):
        """从快照恢复完整节点表状态"""
        self._updating_cells = True
        try:
            # 恢复表格内容
            self.node_table.setRowCount(0)
            for r, row_data in enumerate(snapshot['rows']):
                self.node_table.insertRow(r)
                for c, text in enumerate(row_data):
                    item = QTableWidgetItem(text)
                    item.setTextAlignment(Qt.AlignCenter)
                    if c not in EDITABLE_COLS:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    # 第一行锁定水头损失列
                    if r == 0 and c in FIRST_ROW_LOCKED_LOSS_COLS:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.node_table.setItem(r, c, item)
                row_meta = snapshot.get('row_meta', [])
                if r < len(row_meta):
                    first_item = self.node_table.item(r, 0)
                    if first_item is not None:
                        first_item.setData(Qt.UserRole, row_meta[r])
            # 恢复缓存字典
            self._node_structure_heights = dict(snapshot['structure_heights'])
            self._node_chamfer_params = dict(snapshot['chamfer_params'])
            self._node_u_params = dict(snapshot['u_params'])
            self._node_velocity_increased = dict(snapshot.get('velocity_increased', {}))
        finally:
            self._updating_cells = False
        self._restore_manager_config(getattr(self, "_siphon_manager", None), snapshot.get('siphon_manager_config'))
        self._restore_manager_config(getattr(self, "_pressure_pipe_manager", None), snapshot.get('pressure_pipe_manager_config'))
        self._restore_section_gate_from_snapshot(snapshot)
        nodes_for_view = self._build_nodes_from_table()
        self._apply_table1_source_row_lock_flags()
        self._refresh_all_transition_length_presentations(nodes_for_view)
        self._update_pressure_pipe_roughness_overview(
            self._collect_pressure_pipe_roughness_pairs_from_nodes(nodes_for_view)
        )
        self._refresh_pressure_pipe_controls()

    def _push_node_table_undo(self):
        """在行操作前记录快照到撤销栈"""
        self._node_table_undo_stack.append(self._snapshot_node_table())
        if len(self._node_table_undo_stack) > 20:
            self._node_table_undo_stack.pop(0)
        self._node_table_redo_stack.clear()

    def _undo_node_table(self):
        """撤销行操作"""
        if not self._node_table_undo_stack:
            return False
        # 保存当前状态到重做栈
        self._node_table_redo_stack.append(self._snapshot_node_table())
        if len(self._node_table_redo_stack) > 20:
            self._node_table_redo_stack.pop(0)
        # 恢复上一步状态
        self._restore_node_table(self._node_table_undo_stack.pop())
        InfoBar.success("已撤销", "已恢复上一步操作",
                       parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
        return True

    def _redo_node_table(self):
        """重做行操作"""
        if not self._node_table_redo_stack:
            return False
        # 保存当前状态到撤销栈
        self._node_table_undo_stack.append(self._snapshot_node_table())
        if len(self._node_table_undo_stack) > 20:
            self._node_table_undo_stack.pop(0)
        # 恢复下一步状态
        self._restore_node_table(self._node_table_redo_stack.pop())
        InfoBar.success("已重做", "已恢复上一步撤销的操作",
                       parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
        return True

    def _recalc_downstream(self, edited_row):
        """从 edited_row 开始，重算总水头损失 / 累计 / 水位 / 高程"""
        table = self.node_table
        row_count = table.rowCount()
        if row_count == 0:
            return

        def _rf(r, c):
            """读取单元格浮点值，'-' 或空视为 0"""
            item = table.item(r, c)
            if not item:
                return 0.0
            txt = item.text().strip()
            if not txt or txt == '-':
                return 0.0
            try:
                return float(txt)
            except ValueError:
                return 0.0

        def _set(r, c, val, fmt=".4f"):
            """写入单元格（保持居中对齐和只读标记）"""
            item = table.item(r, c)
            if item is None:
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                if c not in EDITABLE_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                table.setItem(r, c, item)
            item.setText(f"{val:{fmt}}" if val is not None else "-")

        def _is_transition_row(r):
            item = table.item(r, 2)
            return item and item.text().strip() == "渐变段"

        # ── 1. 重算编辑行的总水头损失 (col 39) ──
        if not _is_transition_row(edited_row):
            h_bend = _rf(edited_row, 34)
            h_friction = _rf(edited_row, 35)
            h_reserve = _rf(edited_row, 36)
            h_gate = _rf(edited_row, 37)
            h_siphon = _rf(edited_row, 38)
            new_total = h_bend + h_friction + h_reserve + h_gate + h_siphon
            _set(edited_row, 39, new_total)
            # 同步更新node对象，确保双击时读取到最新值
            if edited_row < len(self.calculated_nodes):
                node = self.calculated_nodes[edited_row]
                if not getattr(node, 'is_transition', False):
                    node.head_loss_total = new_total

        # ── 2. 从编辑行开始重算累计总水头损失 (col 40) ──
        cumulative = _rf(edited_row - 1, 40) if edited_row > 0 else 0.0
        for r in range(edited_row, row_count):
            if _is_transition_row(r):
                cumulative += _rf(r, 33)  # 渐变段水头损失
            else:
                cumulative += _rf(r, 39)  # 总水头损失
            _set(r, 40, cumulative)

        # ── 3. 从编辑行开始重算水位 (col 41)、渠底高程 (col 42)、渠顶高程 (col 43) ──
        start_wl = self._fval(self.start_wl_edit, 0.0)
        if start_wl <= 0:
            return  # 无起始水位，无法递推

        def _find_prev_regular_row(start_row):
            for rr in range(start_row - 1, -1, -1):
                if not _is_transition_row(rr):
                    return rr
            return -1

        prev_regular_row = _find_prev_regular_row(edited_row)
        prev_wl = _rf(prev_regular_row, 41) if prev_regular_row >= 0 else start_wl

        for r in range(edited_row, row_count):
            if _is_transition_row(r):
                continue

            if prev_regular_row < 0:
                # 第一个常规节点：水位 = 起始水位
                wl = start_wl
            else:
                # 后续常规节点：水位 = 上一常规节点水位 - 本行总损失 - 中间渐变段损失
                transition_loss = 0.0
                for j in range(prev_regular_row + 1, r):
                    if _is_transition_row(j):
                        transition_loss += _rf(j, 33)
                # 使用本行"增量总损失"(col39)，不能用累计值(col40)重复扣减
                total_drop = _rf(r, 39) + transition_loss
                wl = prev_wl - total_drop

            _set(r, 41, wl, ".3f")

            # 渠底高程 = 水位 - 水深
            h_depth = _rf(r, 27)
            if h_depth > 0:
                be = wl - h_depth
                _set(r, 42, be, ".3f")
                # 渠顶高程 = 渠底高程 + 结构高度
                sh = self._node_structure_heights.get(r, 0.0)
                if sh > 0:
                    _set(r, 43, be + sh, ".3f")

            prev_wl = wl
            prev_regular_row = r

        # 同步更新 calculated_nodes（如果存在），保证双击查看详情时数据一致
        if hasattr(self, 'calculated_nodes') and self.calculated_nodes:
            channel_level = self._get_current_channel_level_text()
            for r in range(min(row_count, len(self.calculated_nodes))):
                node = self.calculated_nodes[r]
                if node.is_transition:
                    node.head_loss_cumulative = _rf(r, 40)
                else:
                    node.head_loss_reserve = _rf(r, 36)
                    node.head_loss_gate = _rf(r, 37)
                    self._set_gate_loss_user_set(node, self._is_gate_loss_user_set_for_row(r, node))
                    self._apply_pressure_pipe_loss_cell_to_node(node, _rf(r, 38), channel_level=channel_level)
                    node.head_loss_total = _rf(r, 39)
                    node.head_loss_cumulative = _rf(r, 40)
                    node.water_level = _rf(r, 41)
                    be = _rf(r, 42)
                    if be:
                        node.bottom_elevation = be
                    te = _rf(r, 43)
                    if te:
                        node.top_elevation = te

    # ================================================================
    def _add_node_row(self, data=None, _skip_undo=False, _from_table1_source=False, _defer_controls_refresh=False):
        """添加一行节点，_skip_undo=True 时跳过撤销快照（内部调用用）"""
        if not _skip_undo:
            self._push_node_table_undo()
        self._transition_topology_prepared = False
        row = self.node_table.rowCount()
        self.node_table.insertRow(row)
        struct_text_in_data = ""
        if isinstance(data, (list, tuple)) and len(data) > 2:
            struct_text_in_data = str(data[2] or "").strip()
        explicit_gate_loss_text = ""
        if isinstance(data, (list, tuple)) and len(data) > 37:
            explicit_gate_loss_text = str(data[37] or "").strip()
        row_is_pressure_pipe = self._is_pressure_pipe_like_structure_text(struct_text_in_data)
        total_cols = len(NODE_ALL_HEADERS)
        for col in range(total_cols):
            if data and isinstance(data, (list, tuple)) and col < len(data) and data[col]:
                item = QTableWidgetItem(str(data[col]))
            elif col == 2:
                item = QTableWidgetItem("明渠-梯形")
            else:
                item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            if col not in EDITABLE_COLS:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            # 第一行（水位起点）锁定水头损失列
            if row == 0 and col in FIRST_ROW_LOCKED_LOSS_COLS:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if _from_table1_source and col in TABLE1_SOURCE_LOCKED_COLS:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if row_is_pressure_pipe and col == 7:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.node_table.setItem(row, col, item)
        first_item = self.node_table.item(row, 0)
        if first_item:
            payload = first_item.data(Qt.UserRole)
            if not isinstance(payload, dict):
                payload = {}
            payload["_from_table1_source"] = bool(_from_table1_source)
            payload[FLAT_BOTTOM_TUNNEL_SOURCE_ROLE_KEY] = bool(_from_table1_source)
            payload[ARCH_CULVERT_SOURCE_ROLE_KEY] = bool(
                _from_table1_source
                and normalize_culvert_family_type_name(struct_text_in_data) == ARCH_CULVERT_FAMILY_TEXT
            )
            if _from_table1_source:
                x_text = ""
                y_text = ""
                if isinstance(data, (list, tuple)):
                    if len(data) > 5:
                        x_text = self._normalize_coord_text(data[5])
                    if len(data) > 6:
                        y_text = self._normalize_coord_text(data[6])
                payload[SOURCE_COORD_X_ROLE_KEY] = x_text
                payload[SOURCE_COORD_Y_ROLE_KEY] = y_text
            else:
                payload.pop(SOURCE_COORD_X_ROLE_KEY, None)
                payload.pop(SOURCE_COORD_Y_ROLE_KEY, None)
            if explicit_gate_loss_text in ("", "-"):
                payload.pop(GATE_HEAD_LOSS_USER_SET_ROLE_KEY, None)
            first_item.setData(Qt.UserRole, payload)
        if not _defer_controls_refresh:
            self._refresh_pressure_pipe_controls()

    def _clear_nodes(self, _skip_undo=False):
        if not _skip_undo and self.node_table.rowCount() > 0:
            self._push_node_table_undo()
        self.node_table.setRowCount(0)
        self._transition_topology_prepared = False
        self._node_structure_heights.clear()
        self._node_chamfer_params.clear()
        self._node_u_params.clear()
        self._node_velocity_increased.clear()
        self.calculated_nodes = []
        self.nodes = []
        if hasattr(self, 'siphon_roughness_chips'):
            self.siphon_roughness_chips.clear()
        if hasattr(self, 'pressure_pipe_roughness_chips'):
            self.pressure_pipe_roughness_chips.clear()
        self._refresh_roughness_overview_visibility()
        self._refresh_pressure_pipe_controls()

    def _count_table3_related_config_entries(self):
        siphon_count = 0
        pressure_count = 0
        manager = getattr(self, "_siphon_manager", None)
        if manager is not None:
            getter = getattr(manager, "get_siphon_names", None)
            if callable(getter):
                try:
                    siphon_count = len([name for name in getter() if str(name).strip()])
                except Exception:
                    siphon_count = 0
        manager = getattr(self, "_pressure_pipe_manager", None)
        if manager is not None:
            getter = getattr(manager, "get_all_pipe_names", None)
            if callable(getter):
                try:
                    pressure_count = len([name for name in getter() if str(name).strip()])
                except Exception:
                    pressure_count = 0
        return siphon_count, pressure_count

    def _clear_table3_related_configs(self):
        self._close_pressure_pipe_summary_dialog(force=True)
        cleared_siphon = 0
        cleared_pressure = 0
        manager = getattr(self, "_siphon_manager", None)
        if manager is not None:
            getter = getattr(manager, "get_siphon_names", None)
            clear_fn = getattr(manager, "clear_all", None)
            save_fn = getattr(manager, "save_config", None)
            try:
                if callable(getter):
                    cleared_siphon = len([name for name in getter() if str(name).strip()])
                if callable(clear_fn):
                    clear_fn()
                if callable(save_fn):
                    save_fn()
            except Exception:
                cleared_siphon = 0
        manager = getattr(self, "_pressure_pipe_manager", None)
        if manager is not None:
            getter = getattr(manager, "get_all_pipe_names", None)
            clear_fn = getattr(manager, "clear_all", None)
            try:
                if callable(getter):
                    cleared_pressure = len([name for name in getter() if str(name).strip()])
                if callable(clear_fn):
                    clear_fn()
            except Exception:
                cleared_pressure = 0
        return cleared_siphon, cleared_pressure

    def _on_clear_table3_clicked(self):
        node_rows = self.node_table.rowCount() if self.node_table else 0
        siphon_count, pressure_count = self._count_table3_related_config_entries()
        if node_rows <= 0 and siphon_count <= 0 and pressure_count <= 0:
            self._mark_section_results_stale("状态：表3已为空，请先执行断面批量计算", status_kind="warning")
            InfoBar.info(
                "已为空",
                "表3已为空，无需清空；下游操作已保持锁定，请先执行断面批量计算。",
                parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP
            )
            return
        confirm_msg = (
            f"将清空表3并清理联动配置：\n"
            f"• 表3节点行数：{node_rows}\n"
            f"• 倒虹吸配置：{siphon_count} 组\n"
            f"• 有压管道配置：{pressure_count} 组\n\n"
            "影响：将锁定下游操作，需重新执行断面批量计算后同步到表3。\n"
            "恢复说明：表3行可按 Ctrl+Z 撤销；配置清理不可撤销。"
        )
        if not self._ask_destructive_confirm("确认清空表3", confirm_msg, yes_text="确认清空", no_text="取消"):
            return
        self._push_node_table_undo()
        self._clear_nodes(_skip_undo=True)
        cleared_siphon, cleared_pressure = self._clear_table3_related_configs()
        self._mark_section_results_stale("状态：表3已清空并清理联动配置，请先执行断面批量计算", status_kind="warning")
        InfoBar.success(
            "清空完成",
            f"已清空表3 {node_rows} 行，清理倒虹吸 {cleared_siphon} 组、有压管道 {cleared_pressure} 组。表3可按 Ctrl+Z 撤销，配置清理不可撤销。",
            parent=self._info_parent(), duration=2500, position=InfoBarPosition.TOP
        )

    def _get_node_row_data(self, row):
        data = []
        for col in range(self.node_table.columnCount()):
            item = self.node_table.item(row, col)
            data.append(item.text() if item else "")
        return data

    def _get_node_row_payload(self, row):
        table = getattr(self, "node_table", None)
        if not table or row < 0 or row >= table.rowCount():
            return {}
        first_item = table.item(row, 0)
        payload = first_item.data(Qt.UserRole) if first_item else None
        return payload if isinstance(payload, dict) else {}

    def _normalize_coord_text(self, value) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if not text or text == "-":
            return ""
        return text

    def _coord_decimal_places(self, value) -> int:
        text = self._normalize_coord_text(value)
        if not text:
            return 0
        if text[0] in "+-":
            text = text[1:]
        if "e" in text.lower() or "." not in text:
            return 0
        return len(text.split(".", 1)[1])

    def _node_is_source_row(self, node) -> bool:
        if bool(getattr(node, "from_table1_source", False)):
            return True
        return not getattr(node, "is_transition", False) and not getattr(node, "is_auto_inserted_channel", False)

    def _get_node_source_coord_text(self, node, axis: str) -> str:
        attr = "source_x_text" if axis == "x" else "source_y_text"
        return self._normalize_coord_text(getattr(node, attr, ""))

    def _get_table_turn_radius_text(self, row: int) -> str:
        """读取表格中的转弯半径原始文本，显式填写 0 也视为有效输入。"""
        table = getattr(self, "node_table", None)
        if table is None or row < 0 or row >= table.rowCount():
            return ""
        item = table.item(row, 7)
        if item is None:
            return ""
        text = str(item.text() if hasattr(item, "text") else item).strip()
        return "" if text == "-" else text

    def _try_parse_turn_radius_text(self, text):
        """尝试解析转弯半径文本，解析失败返回 None。"""
        raw = str(text or "").strip()
        if not raw or raw == "-":
            return None
        try:
            return float(raw)
        except (TypeError, ValueError):
            return None

    def _get_source_turn_radius_entry_state(self) -> dict:
        """统计真实导入行的转弯半径状态，用于决定顶部栏位显示。"""
        table = getattr(self, "node_table", None)
        if table is None:
            return {"has_source_rows": False, "uniform_positive": None, "mixed_positive": False}

        has_source_rows = False
        positive_values = []
        for row in range(table.rowCount()):
            if not self._is_table1_source_row(row):
                continue
            has_source_rows = True
            value = self._try_parse_turn_radius_text(self._get_table_turn_radius_text(row))
            if value is None or value <= 0:
                continue
            positive_values.append(float(value))

        if not positive_values:
            return {"has_source_rows": has_source_rows, "uniform_positive": None, "mixed_positive": False}

        first_value = positive_values[0]
        is_uniform = all(abs(value - first_value) <= 1e-9 for value in positive_values[1:])
        return {
            "has_source_rows": has_source_rows,
            "uniform_positive": first_value if is_uniform else None,
            "mixed_positive": not is_uniform,
        }

    def _sync_turn_radius_entry_from_source_rows(self):
        """按真实导入行的状态更新顶部转弯半径栏位。"""
        edit = getattr(self, "turn_radius_edit", None)
        if edit is None or not hasattr(edit, "setText"):
            return
        state = self._get_source_turn_radius_entry_state()
        uniform_positive = state.get("uniform_positive")
        if uniform_positive is None:
            edit.setText("")
            return
        edit.setText(f"{uniform_positive:.1f}")

    def _node_has_explicit_turn_radius(self, node) -> bool:
        """判断节点是否带有用户显式填写的转弯半径。"""
        text = str(getattr(node, "turn_radius_text", "") or "").strip()
        if text and text != "-":
            return True
        return bool(getattr(node, "turn_radius_is_explicit", False))

    def _should_write_siphon_turn_radius_result(
        self,
        group,
        *,
        row_position: int,
        node,
        turn_radius_overrode_excel: bool,
    ) -> bool:
        """判断倒虹吸计算结果能否写回某个中间 IP 转弯半径。"""
        row_indices = list(getattr(group, "row_indices", []) or [])
        if row_position <= 0 or row_position >= len(row_indices) - 1:
            return False
        try:
            if float(getattr(node, "turn_angle", 0.0) or 0.0) <= 0:
                return False
        except (TypeError, ValueError):
            return False
        if turn_radius_overrode_excel:
            return True
        return not self._node_has_explicit_turn_radius(node)

    def _format_node_turn_radius_display_text(self, node, row_index: int) -> str:
        """生成节点回写到表格时的转弯半径文本。"""
        text = str(getattr(node, "turn_radius_text", "") or "").strip()
        if text and text != "-":
            return text
        try:
            turn_radius = float(getattr(node, "turn_radius", 0.0) or 0.0)
        except (TypeError, ValueError):
            turn_radius = 0.0
        if self._node_has_explicit_turn_radius(node):
            if abs(turn_radius) <= 1e-12:
                return "0"
            return f"{turn_radius:.1f}"
        if turn_radius > 0:
            return f"{turn_radius:.1f}"
        if self._node_is_source_row(node):
            return "0"
        return ""

    def _describe_turn_radius_entry_usage(self, settings=None) -> str:
        """生成人类可读的顶部转弯半径栏位状态说明。"""
        edit = getattr(self, "turn_radius_edit", None)
        text = str(edit.text().strip()) if edit is not None and hasattr(edit, "text") else ""
        if text:
            state = self._get_source_turn_radius_entry_state()
            value = self._try_parse_turn_radius_text(text)
            uniform_positive = state.get("uniform_positive")
            if (
                value is not None
                and uniform_positive is not None
                and abs(value - uniform_positive) <= 1e-9
                and not state.get("mixed_positive", False)
            ):
                return f"已统一为 {text} m"
            return f"待应用值 {text} m"
        if settings is not None:
            try:
                if float(getattr(settings, "turn_radius", 0.0) or 0.0) > 0:
                    return f"{float(getattr(settings, 'turn_radius', 0.0)):.1f} m"
            except (TypeError, ValueError):
                pass
        return "未统一应用（按各行值）"

    def _apply_pending_turn_radius_to_source_rows(self):
        """将顶部栏位中的转弯半径统一应用到所有真实导入行。"""
        edit = getattr(self, "turn_radius_edit", None)
        raw_text = str(edit.text().strip()) if edit is not None and hasattr(edit, "text") else ""
        if not raw_text:
            InfoBar.warning(
                "提示",
                "请先在顶部栏位填写要应用的转弯半径。",
                parent=self._info_parent(),
                duration=3000,
                position=InfoBarPosition.TOP,
            )
            return 0

        radius_value = self._try_parse_turn_radius_text(raw_text)
        if radius_value is None or radius_value < 0:
            InfoBar.warning(
                "提示",
                "转弯半径只能填写 0 或正数。",
                parent=self._info_parent(),
                duration=3000,
                position=InfoBarPosition.TOP,
            )
            return 0

        source_rows = [row for row in range(self.node_table.rowCount()) if self._is_table1_source_row(row)]
        if not source_rows:
            InfoBar.warning(
                "提示",
                "当前没有可应用的真实导入行。",
                parent=self._info_parent(),
                duration=3000,
                position=InfoBarPosition.TOP,
            )
            return 0

        confirm_message = (
            f"点击确认后，会把当前栏位中的转弯半径 {raw_text} m 作用于所有真实导入行。\n\n"
            "如果希望不同的行使用不同的转弯半径，建议在导入 Excel 前就先在 Excel 中分别填好，"
            "不要在程序里再统一覆盖。"
        )
        if not self._ask_destructive_confirm("确认应用转弯半径", confirm_message, yes_text="确认应用", no_text="取消"):
            return 0

        changed_cells = 0
        self._push_node_table_undo()
        for row in source_rows:
            if self._get_table_turn_radius_text(row) == raw_text:
                continue
            self._set_table_cell_text_preserve_flags(row, 7, raw_text)
            changed_cells += 1

        if changed_cells <= 0:
            return 0

        self._apply_table1_source_row_lock_flags()
        self._recalculate_geometry()
        self._refresh_pressure_pipe_controls()
        self._sync_turn_radius_entry_from_source_rows()
        InfoBar.success(
            "应用成功",
            f"已将顶部转弯半径应用到 {changed_cells} 行真实导入行。",
            parent=self._info_parent(),
            duration=3000,
            position=InfoBarPosition.TOP,
        )
        return changed_cells

    def _resolve_neighbor_coord_precision(self, nodes, index: int, axis: str, default: int = 6) -> int:
        precisions = []
        for direction in (-1, 1):
            cursor = index + direction
            while 0 <= cursor < len(nodes):
                candidate = nodes[cursor]
                if self._node_is_source_row(candidate):
                    text = self._get_node_source_coord_text(candidate, axis)
                    if text:
                        precisions.append(self._coord_decimal_places(text))
                    break
                cursor += direction
        if precisions:
            return max(precisions)
        current_text = self._get_node_source_coord_text(nodes[index], axis) if 0 <= index < len(nodes) else ""
        if current_text:
            return self._coord_decimal_places(current_text)
        return default

    def _format_coord_value(self, value, decimals: int) -> str:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return ""
        if abs(number) < 1e-15:
            return ""
        return f"{number:.{max(0, int(decimals))}f}"

    def _resolve_node_coord_display_text(self, nodes, index: int, node, axis: str, default: int = 6) -> str:
        if self._node_is_source_row(node):
            source_text = self._get_node_source_coord_text(node, axis)
            if source_text:
                return source_text
        decimals = self._resolve_neighbor_coord_precision(nodes, index, axis, default=default)
        return self._format_coord_value(getattr(node, axis, 0.0), decimals)

    def _refresh_source_coord_payloads_from_table(self):
        table = getattr(self, "node_table", None)
        if not table:
            return
        for row in range(table.rowCount()):
            payload = self._get_node_row_payload(row)
            if not payload and not self._is_table1_source_row(row):
                continue
            x_text = self._normalize_coord_text(table.item(row, 5).text() if table.item(row, 5) else "")
            y_text = self._normalize_coord_text(table.item(row, 6).text() if table.item(row, 6) else "")
            if self._is_table1_source_row(row):
                payload[SOURCE_COORD_X_ROLE_KEY] = x_text
                payload[SOURCE_COORD_Y_ROLE_KEY] = y_text
            else:
                payload.pop(SOURCE_COORD_X_ROLE_KEY, None)
                payload.pop(SOURCE_COORD_Y_ROLE_KEY, None)
            first_item = table.item(row, 0)
            if first_item and payload:
                first_item.setData(Qt.UserRole, payload)

    # ================================================================
    # 从断面计算结果同步到表3
    # ================================================================
    def _import_from_batch(self):
        if not SHARED_DATA_AVAILABLE:
            InfoBar.warning("不可用", "SharedDataManager未加载，无法导入", parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        shared_data = get_shared_data_manager()
        results = shared_data.get_batch_results()
        if not results:
            InfoBar.warning("无数据", "断面批量计算尚无可同步结果。请先完成表1断面批量计算。",
                           parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            return

        # 同步批量计算面板中的渠道基础信息
        self._sync_batch_settings()
        (
            prepared_results,
            general_roughness_vals,
            siphon_roughness_pairs,
            pressure_pipe_params_pairs,
        ) = self._prepare_batch_import_results(results)
        chosen_n = self._choose_roughness_value(general_roughness_vals, "渠道糙率")
        if chosen_n is not None:
            self.roughness_edit.setText(f"{chosen_n:.4f}".rstrip('0').rstrip('.'))

        self._updating_cells = True
        table_signal_blocker = QSignalBlocker(self.node_table)
        self._clear_nodes()
        imported = 0
        flow_segment_map = {}  # {流量段编号: 设计流量}
        max_flow_segment_map = {}  # {流量段编号: 最终加大流量}
        for prepared in prepared_results:
            sr = prepared["result"]
            flow_section = str(getattr(sr, 'flow_section', ''))
            building_name = str(getattr(sr, 'building_name', ''))
            section_type = prepared["section_type"]
            culvert_family_type = prepared["culvert_family_type"]
            raw_result = prepared["raw_result"]
            x = getattr(sr, 'coord_X', 0)
            y = getattr(sr, 'coord_Y', 0)
            B = getattr(sr, 'B', None) or ""
            D = getattr(sr, 'D', None) or ""
            R = getattr(sr, 'R', None) or ""
            m_val = getattr(sr, 'm', None) or ""
            n_val = (
                chosen_n
                if prepared["is_channel_roughness_row"] and chosen_n is not None
                else getattr(sr, 'n', 0) or ""
            )
            slope_inv = getattr(sr, 'slope_inv', 0) or ""
            Q = getattr(sr, 'Q', 0) or ""
            pipe_material = prepared["pipe_material"]
            compound_params = normalize_compound_trapezoid_params(raw_result)
            local_loss_ratio = (
                getattr(sr, 'local_loss_ratio', 0.0)
                if getattr(sr, 'local_loss_ratio', None) is not None
                else raw_result.get('local_loss_ratio', 0.0)
            )
            in_out_raw = str(
                getattr(sr, 'in_out_raw', '') or raw_result.get('in_out_raw', '')
            ).strip()
            use_increase = normalize_use_increase_flag(
                getattr(sr, 'use_increase', raw_result.get('use_increase', raw_result.get('_use_increase', True))),
                default=True,
            )

            # 收集流量段信息
            try:
                seg_num = int(flow_section)
            except (ValueError, TypeError):
                seg_num = 1
            q_val = float(Q) if Q and str(Q).strip() else 0.0
            if seg_num not in flow_segment_map and q_val > 0:
                flow_segment_map[seg_num] = q_val
            q_max_val = getattr(sr, 'Q_max', 0.0) or 0.0
            if not q_max_val:
                q_max_val = raw_result.get('Q_increased', raw_result.get('Q_max', raw_result.get('Q_inc', 0.0))) or 0.0
            try:
                q_max_val = float(q_max_val)
            except (TypeError, ValueError):
                q_max_val = 0.0
            if seg_num not in max_flow_segment_map and q_max_val > 0:
                max_flow_segment_map[seg_num] = q_max_val

            x_text = self._normalize_coord_text(raw_result.get("coord_X_text", ""))
            y_text = self._normalize_coord_text(raw_result.get("coord_Y_text", ""))
            if not x_text:
                x_text = self._normalize_coord_text(x)
            if not y_text:
                y_text = self._normalize_coord_text(y)

            def fmt(v):
                if v is None or v == "" or v == 0: return ""
                if isinstance(v, float): return f"{v:.4f}" if v < 1 else f"{v:.3f}"
                return str(v)

            # 提取水力计算结果（与原版Tkinter _do_import_from_calc_result对齐）
            h_val = getattr(sr, 'h', None) or 0.0
            V_val = getattr(sr, 'V', 0) or 0.0
            V_max_val = getattr(sr, 'V_max', 0) or 0.0  # 加大流速
            A_val = getattr(sr, 'A', 0) or 0.0
            X_val = getattr(sr, 'X', 0) or 0.0
            R_hyd_val = getattr(sr, 'R_hydraulic', 0) or 0.0
            H_total = getattr(sr, 'H_total', 0) or 0.0

            tr_r = getattr(sr, 'turn_radius', 0.0) or 0.0
            _raw_result = getattr(sr, 'raw_result', {}) or {}
            tr_r_text = str(getattr(sr, 'turn_radius_text', '') or _raw_result.get('turn_radius_text', '') or '').strip()
            row_data = [""] * len(NODE_ALL_HEADERS)
            row_data[0] = flow_section
            row_data[1] = building_name
            row_data[2] = section_type
            row_data[5] = x_text
            row_data[6] = y_text
            row_data[7] = tr_r_text if tr_r_text and tr_r_text != "-" else (fmt(tr_r) if tr_r > 0 else "0")
            row_data[20] = fmt(B) if B else ""
            row_data[21] = fmt(D) if D else ""
            row_data[22] = fmt(R) if R else ""
            row_data[23] = fmt(m_val) if m_val else ""
            row_data[24] = fmt(n_val)
            row_data[25] = fmt(slope_inv)
            row_data[26] = fmt(Q)
            self._add_node_row(
                row_data,
                _skip_undo=True,
                _from_table1_source=True,
                _defer_controls_refresh=True
            )

            # 写入水力结果到结果列（原版通过set_nodes写入water_depth/velocity等）
            cur_row = self.node_table.rowCount() - 1
            first_item = self.node_table.item(cur_row, 0)
            if first_item:
                payload = first_item.data(Qt.UserRole)
                if not isinstance(payload, dict):
                    payload = {}
                # use_increase 是自由水面导出链路的重要运行态字段，不能只在有压流参数存在时才落盘。
                payload[USE_INCREASE_ROLE_KEY] = use_increase
                if culvert_family_type:
                    payload[CULVERT_FAMILY_TYPE_KEY] = culvert_family_type
                    theta_deg = float(raw_result.get("theta_deg", 0) or 0)
                    if theta_deg > 0:
                        payload[ARCH_CULVERT_THETA_ROLE_KEY] = theta_deg
                    if culvert_family_type == ARCH_CULVERT_FAMILY_TEXT:
                        H_straight = raw_result.get("H_straight", getattr(sr, "H_straight", None))
                        manual_H_straight = raw_result.get("manual_H_straight", None)
                        used_manual_H_straight = raw_result.get("used_manual_H_straight", None)
                        if H_straight not in (None, ""):
                            payload[ARCH_CULVERT_H_STRAIGHT_ROLE_KEY] = H_straight
                        if manual_H_straight not in (None, ""):
                            payload[ARCH_CULVERT_MANUAL_H_STRAIGHT_ROLE_KEY] = manual_H_straight
                        if used_manual_H_straight is not None:
                            payload[ARCH_CULVERT_USED_MANUAL_H_STRAIGHT_ROLE_KEY] = bool(used_manual_H_straight)
                payload[ARCH_CULVERT_SOURCE_ROLE_KEY] = bool(
                    culvert_family_type == ARCH_CULVERT_FAMILY_TEXT
                )
                if section_type == "隧洞-圆拱直墙型":
                    theta_deg = float(raw_result.get("theta_deg", 0) or 0)
                    H_straight = raw_result.get("H_straight", getattr(sr, "H_straight", None))
                    manual_H_straight = raw_result.get("manual_H_straight", None)
                    used_manual_H_straight = raw_result.get("used_manual_H_straight", None)
                    if theta_deg > 0:
                        payload[TUNNEL_ARCH_THETA_ROLE_KEY] = theta_deg
                    if H_straight not in (None, ""):
                        payload[TUNNEL_ARCH_H_STRAIGHT_ROLE_KEY] = H_straight
                    if manual_H_straight not in (None, ""):
                        payload[TUNNEL_ARCH_MANUAL_H_STRAIGHT_ROLE_KEY] = manual_H_straight
                    if used_manual_H_straight is not None:
                        payload[TUNNEL_ARCH_USED_MANUAL_H_STRAIGHT_ROLE_KEY] = bool(used_manual_H_straight)
                # 表1导入到表3时，给承压类行补齐稳定身份，后续导出优先按真实身份匹配。
                if self._is_pressure_pipe_like_structure_text(section_type):
                    payload[PRESSURE_PIPE_ROW_ID_ROLE_KEY] = self._build_pressure_pipe_row_identity_from_flow_section(
                        flow_section,
                        cur_row,
                    )
                first_item.setData(Qt.UserRole, payload)
            # 通过行级元数据透传有压管道专用参数（表格无专门列）
            if pipe_material or in_out_raw or (local_loss_ratio and float(local_loss_ratio) > 0):
                if first_item:
                    payload = first_item.data(Qt.UserRole)
                    if not isinstance(payload, dict):
                        payload = {}
                    if pipe_material:
                        payload['_pipe_material'] = pipe_material
                    if in_out_raw:
                        payload['_in_out_raw'] = in_out_raw
                    try:
                        llr = float(local_loss_ratio)
                        if llr > 0:
                            payload['_local_loss_ratio'] = llr
                    except (ValueError, TypeError):
                        pass
                    first_item.setData(Qt.UserRole, payload)
            if section_type == "明渠-复式梯形" and compound_params and first_item:
                payload = first_item.data(Qt.UserRole)
                if not isinstance(payload, dict):
                    payload = {}
                payload[COMPOUND_TRAPEZOID_PARAMS_ROLE_KEY] = copy.deepcopy(compound_params)
                first_item.setData(Qt.UserRole, payload)
            if h_val and float(h_val) > 0:
                _item = QTableWidgetItem(f"{float(h_val):.3f}")
                _item.setTextAlignment(Qt.AlignCenter)
                _item.setFlags(_item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(cur_row, 27, _item)  # 水深h
            if A_val and float(A_val) > 0:
                _item = QTableWidgetItem(f"{float(A_val):.3f}")
                _item.setTextAlignment(Qt.AlignCenter)
                _item.setFlags(_item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(cur_row, 28, _item)  # 过水断面面积A
            if X_val and float(X_val) > 0:
                _item = QTableWidgetItem(f"{float(X_val):.3f}")
                _item.setTextAlignment(Qt.AlignCenter)
                _item.setFlags(_item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(cur_row, 29, _item)  # 湿周X
            if R_hyd_val and float(R_hyd_val) > 0:
                _item = QTableWidgetItem(f"{float(R_hyd_val):.3f}")
                _item.setTextAlignment(Qt.AlignCenter)
                _item.setFlags(_item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(cur_row, 30, _item)  # 水力半径R
            if V_val and float(V_val) > 0:
                _item = QTableWidgetItem(f"{float(V_val):.3f}")
                _item.setTextAlignment(Qt.AlignCenter)
                _item.setFlags(_item.flags() & ~Qt.ItemIsEditable)
                self.node_table.setItem(cur_row, 31, _item)  # 流速v
            # 缓存结构高度（与Tkinter版 data_table._node_structure_heights 对齐）
            if H_total and float(H_total) > 0:
                self._node_structure_heights[cur_row] = float(H_total)
            # 缓存倒角参数（渡槽-矩形专用，不占用表格列）
            if "渡槽-矩形" in section_type:
                _raw = getattr(sr, 'raw_result', {}) or {}
                _ca = _raw.get('chamfer_angle', 0) or 0
                _cl = _raw.get('chamfer_length', 0) or 0
                if _ca > 0 and _cl > 0:
                    self._node_chamfer_params[cur_row] = {'chamfer_angle': float(_ca), 'chamfer_length': float(_cl)}

            # 缓存明渠-U形的圆心角
            if "明渠-U形" in section_type:
                _raw_u = getattr(sr, 'raw_result', {}) or {}
                _theta = _raw_u.get('theta_deg', 0) or 0
                if _theta > 0:
                    self._node_u_params[cur_row] = {'theta_deg': float(_theta)}

            # 缓存加大流速（用于倒虹吸水力计算时自动填入v₁加大/v₃加大）
            if V_max_val and float(V_max_val) > 0:
                self._node_velocity_increased[cur_row] = float(V_max_val)

            imported += 1

        auto_resize_table(self.node_table)

        # 倒虹吸糙率只读概览（每个倒虹吸独立显示）
        self._update_siphon_roughness_overview(siphon_roughness_pairs)
        self._update_pressure_pipe_roughness_overview(pressure_pipe_params_pairs)

        # 自动填充多流量段设计流量和加大流量
        if flow_segment_map:
            sorted_segs = sorted(flow_segment_map.keys())
            design_flows = [flow_segment_map[s] for s in sorted_segs]
            preferred_max_flows = [max_flow_segment_map.get(s, 0.0) for s in sorted_segs]
            final_max_flow_text = format_flow_values_text(
                calculate_final_max_flow_values(design_flows, preferred_max_flows)
            )
            if hasattr(self, "design_flow_edit") and hasattr(self.design_flow_edit, "setText"):
                self.design_flow_edit.setText(format_flow_values_text(design_flows))
            if hasattr(self, "max_flow_edit") and hasattr(self.max_flow_edit, "setText"):
                self.max_flow_edit.setText(final_max_flow_text)
                self._sync_flow_segment_widgets(reset_index=True)

        # 检查是否包含倒虹吸
        has_siphon = False
        has_pressure_pipe = False
        if CALCULATOR_AVAILABLE:
            nodes = self._build_nodes_from_table()
            has_siphon = any(
                n.structure_type and "倒虹吸" in n.structure_type.value
                for n in nodes if n.structure_type
            )
            has_pressure_pipe = any(
                self._is_pressure_pipe_like_node(n)
                for n in nodes if n.structure_type
            )

        self._updating_cells = False

        # 触发几何计算（与原版Tkinter recalculate对齐）
        self._recalculate_geometry()
        self.nodes = self._build_nodes_from_table()
        self._sync_turn_radius_entry_from_source_rows()

        next_steps = "请依次点击【插入渐变段】→"
        if has_siphon:
            next_steps += "【倒虹吸水力计算】→"
        if has_pressure_pipe:
            next_steps += "【有压管道水力计算】→"
        next_steps += "【执行计算】"
        self._apply_table1_source_row_lock_flags()
        self._refresh_pressure_pipe_controls()
        del table_signal_blocker

        InfoBar.success("导入成功",
                       f"已导入 {imported} 个节点，已自动填充流量。转弯半径按各行导入值保留；如需统一，请在顶部填写后点击“应用”。{next_steps}",
                       parent=self._info_parent(), duration=6000, position=InfoBarPosition.TOP)

    def _recalculate_geometry(self):
        """
        导入数据后触发几何计算（与原版Tkinter recalculate对齐）

        流程：
        1. 保留表格中的行级转弯半径；留空或填 0 的行按 0 参与几何计算；
           顶部栏位仅作为待应用统一值，不在这里自动介入
        2. 构建节点 → calculate_geometry → preprocess_nodes
        3. 回写几何结果列(8-19) + 进出口(3) + IP(4)；转弯半径列按行级原始语义保留
        """
        self._updating_cells = True
        try:
            self._recalculate_geometry_impl()
        finally:
            self._updating_cells = False

    def _recalculate_geometry_impl(self):
        if not CALCULATOR_AVAILABLE:
            return

        # ---- 1. 记录显式填写的转弯半径行 ----
        # 新规则：留空统一按 0 处理，不再为任何结构自动填充全局/临时半径。
        siphon_rows_with_existing = set()
        pressure_rows_with_existing = set()
        for r in range(self.node_table.rowCount()):
            existing_r_text = self._get_table_turn_radius_text(r)
            has_explicit_turn_radius = bool(existing_r_text)
            struct_item = self.node_table.item(r, 2)
            struct_text = struct_item.text().strip() if struct_item else ""
            _is_siphon = "倒虹吸" in struct_text
            _is_pressure_pipe = self._is_pressure_pipe_like_structure_text(struct_text)
            _is_gate = "闸" in struct_text or "分水" in struct_text
            if has_explicit_turn_radius:
                if _is_siphon:
                    siphon_rows_with_existing.add(r)
                if _is_pressure_pipe:
                    pressure_rows_with_existing.add(r)

        # ---- 2. 构建节点 & 几何计算 ----
        nodes = self._build_nodes_from_table()
        if len(nodes) < 2:
            return

        settings = self._build_settings()
        if not settings:
            return

        calculator = WaterProfileCalculator(settings)
        # 先清掉表格遗留的进/出标记，再按纯几何口径重算；否则特殊节点会被旧标记提前清零。
        for node in nodes:
            node.in_out = InOutType.NORMAL
        calculator.calculate_geometry(nodes)
        calculator.preprocess_nodes(nodes)

        # ---- 3. 回写几何结果到表格 ----
        prefix = settings.get_station_prefix() if hasattr(settings, 'get_station_prefix') else ""

        for r, node in enumerate(nodes):
            if r >= self.node_table.rowCount():
                break

            # 进出口判断 (col 3)
            in_out_str = node.get_in_out_str() if hasattr(node, 'get_in_out_str') else ""
            item = QTableWidgetItem(in_out_str)
            item.setTextAlignment(Qt.AlignCenter)
            self.node_table.setItem(r, 3, item)

            # IP编号 (col 4)
            ip_str = node.get_ip_str() if hasattr(node, 'get_ip_str') else ""
            item = QTableWidgetItem(ip_str)
            item.setTextAlignment(Qt.AlignCenter)
            item.setData(Qt.UserRole, {
                "_raw_ip_number": getattr(node, "ip_number", 0),
                "_display_ip_number": getattr(node, "display_ip_number", None),
            })
            self.node_table.setItem(r, 4, item)

            # 转弯半径 (col 7) — 新规则：源行留空按 0 显示，不再自动套全局半径。
            turn_radius_text = self._format_node_turn_radius_display_text(node, r)
            item = QTableWidgetItem(turn_radius_text)
            item.setTextAlignment(Qt.AlignCenter)
            self.node_table.setItem(r, 7, item)

            # 几何结果列 (8-19) — 无条件格式化，0值也显示（与Tkinter一致）
            fmt_s = lambda s: ProjectSettings.format_station(s, prefix) if s is not None else ""
            geo_data = {
                8:  f"{node.turn_angle:.4f}",
                9:  f"{node.tangent_length:.6f}",
                10: f"{node.arc_length:.6f}",
                11: f"{node.curve_length:.6f}",
                12: f"{node.straight_distance:.6f}",
                13: fmt_s(node.station_ip),
                14: fmt_s(node.station_BC),
                15: fmt_s(getattr(node, 'station_MC', None)),
                16: fmt_s(node.station_EC),
                17: f"{getattr(node, 'check_pre_curve', 0):.3f}",
                18: f"{getattr(node, 'check_post_curve', 0):.3f}",
                19: f"{getattr(node, 'check_total_length', 0):.3f}",
            }
            for c, v in geo_data.items():
                item = QTableWidgetItem(str(v))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self._apply_curve_check_item_style(item, c)
                self.node_table.setItem(r, c, item)

        auto_resize_table(self.node_table)

    def fill_turn_radius_for_geometry(self, nodes, n):
        """
        为有压管道节点填充临时转弯半径 R = n × D
        
        在几何计算前调用，为有压管道节点临时填充转弯半径值供几何计算使用。
        
        Args:
            nodes: ChannelNode 列表
            n: 转弯半径倍数（R = n × D）
            
        Requirements: 8.1, 8.2, 8.3
        """
        if not nodes:
            return
        
        for node in nodes:
            # 只处理有压管道节点
            if not node.is_pressure_pipe:
                continue
            
            # 用户显式填写的值（包含0）一律保留，不参与临时半径填充
            if self._node_has_explicit_turn_radius(node):
                continue
            
            # 如果已有非零转弯半径值，保留不变（可能是用户导入或手动输入的值）
            if node.turn_radius and node.turn_radius > 0:
                continue
            
            # 从 section_params 中获取管径 D
            diameter_D = node.section_params.get('D', 0.0)
            if diameter_D <= 0:
                continue
            
            # 计算临时转弯半径 R = n × D
            node.turn_radius = n * diameter_D
    
    def clear_temporary_turn_radius(self, nodes):
        """
        清空有压管道节点的临时转弯半径
        
        在几何计算后调用，清空临时写入的转弯半径值。
        但保留从水力计算回写的值（检查 head_loss_siphon/external_head_loss 字段）。
        
        Args:
            nodes: ChannelNode 列表
            
        Requirements: 8.4, 8.5
        """
        if not nodes:
            return
        
        for node in nodes:
            # 只处理有压管道节点
            if not node.is_pressure_pipe:
                continue
            
            # 已写回损失值（新字段或旧兼容字段）时，保留转弯半径不清空
            ext_loss = getattr(node, 'external_head_loss', None)
            siphon_loss = getattr(node, 'head_loss_siphon', 0.0) or 0.0
            named_group_result = self._get_pressure_pipe_named_group_result(node)
            if (ext_loss is not None and ext_loss > 0) or siphon_loss > 0 or named_group_result:
                continue
            
            # 清空临时转弯半径
            node.turn_radius = 0.0

    def _sync_batch_settings(self):
        """从断面计算输入区同步渠道基础信息"""
        try:
            bp = getattr(self, "_batch_backend", None)
            if not bp:
                main_win = self.window()
                if main_win and hasattr(main_win, 'batch_panel'):
                    bp = main_win.batch_panel
            if not bp:
                return
            if hasattr(bp, 'channel_name_edit'):
                name = bp.channel_name_edit.text().strip()
                if name:
                    self.channel_name_edit.setText(name)
            if hasattr(bp, 'channel_level_combo'):
                level = bp.channel_level_combo.currentText()
                if level:
                    idx = self.channel_level_combo.findText(level)
                    if idx >= 0:
                        self.channel_level_combo.setCurrentIndex(idx)
            if hasattr(bp, 'start_wl_edit'):
                wl = bp.start_wl_edit.text().strip()
                if wl:
                    self.start_wl_edit.setText(wl)
            if hasattr(bp, 'start_station_edit'):
                st = bp.start_station_edit.text().strip()
                if st:
                    self.start_station_edit.setText(st)
            if hasattr(bp, 'flow_segments_edit'):
                flow_segments = bp.flow_segments_edit.text().strip()
                if flow_segments:
                    manual_qmax_by_segment = getattr(bp, "_manual_qmax_by_segment", {}) or {}
                    if hasattr(self, '_section_flow_segments_edit') and self._section_flow_segments_edit:
                        self._section_flow_segments_edit.setText(flow_segments)
                    if hasattr(self, 'design_flow_edit') and self.design_flow_edit:
                        self.design_flow_edit.setText(flow_segments)
                        design_flows = self._parse_flow_values(flow_segments)
                        if hasattr(self, '_on_design_flow_changed'):
                            self._on_design_flow_changed()
                        elif design_flows and hasattr(self, 'max_flow_edit') and self.max_flow_edit:
                            self.max_flow_edit.setText(
                                format_flow_values_text(calculate_final_max_flow_values(design_flows))
                            )
                            if hasattr(self, '_sync_flow_segment_widgets'):
                                self._sync_flow_segment_widgets(reset_index=False)
                        if manual_qmax_by_segment and design_flows and hasattr(self, 'max_flow_edit') and self.max_flow_edit:
                            preferred_max_flows = []
                            for segment_index in range(len(design_flows)):
                                raw_value = manual_qmax_by_segment.get(segment_index + 1, 0.0)
                                try:
                                    preferred_max_flows.append(float(raw_value))
                                except (TypeError, ValueError):
                                    preferred_max_flows.append(0.0)
                            self.max_flow_edit.setText(
                                format_flow_values_text(
                                    calculate_final_max_flow_values(design_flows, preferred_max_flows)
                                )
                            )
                            if hasattr(self, '_sync_flow_segment_widgets'):
                                self._sync_flow_segment_widgets(reset_index=False)
                        if hasattr(self, '_reset_flow_segment_current_index'):
                            self._reset_flow_segment_current_index()
        except Exception:
            pass

    # ================================================================
    # 计算
    # ================================================================
    def _get_current_start_station_value(self):
        """统一读取当前项目级起始桩号，供表3各类重算入口复用。"""
        text = self.start_station_edit.text() if self.start_station_edit else ""
        return parse_station_input(text)

    def _build_settings(self):
        """从UI读取设置，构建ProjectSettings"""
        if not CALCULATOR_AVAILABLE:
            return None
        settings = ProjectSettings()
        settings.channel_name = self.channel_name_edit.text().strip() or "未命名渠道"
        settings.channel_level = self.channel_level_combo.currentText()
        settings.start_water_level = self._fval(self.start_wl_edit, 100.0)
        settings.start_station = self._get_current_start_station_value()
        settings._start_station_from_ui = True
        # 多流量段支持
        design_flows = self._parse_flow_values(self.design_flow_edit.text())
        max_flows = self._parse_flow_values(self.max_flow_edit.text())
        settings.design_flows = design_flows
        settings.max_flows = max_flows
        settings.design_flow = design_flows[0] if design_flows else 0.0
        settings.max_flow = max_flows[0] if max_flows else 0.0
        settings.roughness = self._fval(self.roughness_edit, DEFAULT_ROUGHNESS)
        # siphon_roughness 不再从单一输入框读取（已改为只读概览），保留默认值
        # 每个倒虹吸的实际糙率从节点表格对应行读取
        settings.turn_radius = self._fval(self.turn_radius_edit, 0.0)
        # 渡槽/隧洞渐变段设置
        settings.transition_inlet_form = self.trans_inlet_combo.currentText()
        settings.transition_inlet_zeta = self._fval(self.trans_inlet_zeta, 0.10)
        settings.transition_outlet_form = self.trans_outlet_combo.currentText()
        settings.transition_outlet_zeta = self._fval(self.trans_outlet_zeta, 0.20)
        # 明渠渐变段设置
        settings.open_channel_transition_form = self.oc_trans_combo.currentText()
        settings.open_channel_transition_zeta = self._fval(self.oc_trans_zeta, 0.10)
        # 倒虹吸渐变段设置
        settings.siphon_transition_inlet_form = self.siphon_inlet_combo.currentText()
        settings.siphon_transition_inlet_zeta = self._fval(self.siphon_inlet_zeta, 0.10)
        settings.siphon_transition_outlet_form = self.siphon_outlet_combo.currentText()
        settings.siphon_transition_outlet_zeta = self._fval(self.siphon_outlet_zeta, 0.20)
        # 倒虹吸转弯半径倍数n
        settings.siphon_turn_radius_n = DEFAULT_SIPHON_TURN_RADIUS_N
        settings.transition_length_rules = self._build_transition_length_rule_objects()
        return settings

    def _build_nodes_from_table(self):
        """从节点表格构建ChannelNode列表（与原版Tkinter get_nodes完全对齐）"""
        if not CALCULATOR_AVAILABLE:
            return []

        # 辅助函数提到循环外避免每行重复定义（#16）
        table = self.node_table

        def _read_float(row, col):
            item = table.item(row, col)
            if item:
                try:
                    return float(item.text())
                except (ValueError, TypeError):
                    pass
            return 0.0

        def _read_text(row, col):
            item = table.item(row, col)
            if item:
                t = item.text().strip()
                if t and t != '-':
                    return t
            return ""

        def _parse_station(text):
            """解析格式化桩号文本（如 '南支0+123.456'）为浮点数"""
            if not text:
                return 0.0
            if '+' in text:
                parts = text.split('+')
                if len(parts) == 2:
                    km_digits = ''.join(c for c in parts[0] if c.isdigit())
                    try:
                        km = int(km_digits) if km_digits else 0
                        m = float(parts[1])
                        return km * 1000 + m
                    except (ValueError, TypeError):
                        pass
            try:
                return float(text)
            except (ValueError, TypeError):
                return 0.0

        def _read_ip_metadata(row):
            """读取 IP 单元格保存的原始/显示编号元数据。"""
            item = table.item(row, 4)
            if not item:
                return {}
            data = item.data(Qt.UserRole)
            return data if isinstance(data, dict) else {}

        def _parse_ip_number(row, text):
            """解析IP编号，支持复合格式如 'IP3 沪蓉倒进'（#11）"""
            meta = _read_ip_metadata(row)
            raw_ip_number = meta.get("_raw_ip_number", None)
            if raw_ip_number not in ("", None):
                try:
                    return int(raw_ip_number)
                except (ValueError, TypeError):
                    pass
            if not text:
                return 0
            # 先尝试直接转int
            try:
                return int(text)
            except (ValueError, TypeError):
                pass
            # 从 "IPxx" 或 "IP xx ..." 中提取数字
            m = re.match(r'IP\s*(\d+)', text)
            if m:
                return int(m.group(1))
            return 0

        def _parse_display_ip_number(row, text):
            """解析显示用 IP 编号；特殊建筑进出口允许为空。"""
            meta = _read_ip_metadata(row)
            display_ip_number = meta.get("_display_ip_number", None)
            if display_ip_number not in ("", None):
                try:
                    return int(display_ip_number)
                except (ValueError, TypeError):
                    pass
            if not text:
                return None
            m = re.match(r'IP\s*(\d+)', text)
            if m:
                return int(m.group(1))
            return None

        nodes = []
        channel_level = self._get_current_channel_level_text()
        _default_q = (self._parse_flow_values(self.design_flow_edit.text()) or [5.0])[0]
        self._pressure_turn_radius_fallback_groups = set()

        for r in range(table.rowCount()):
            data = self._get_node_row_data(r)
            # data[0-7]: 流量段,建筑物名称,结构形式,进出口,IP,X,Y,转弯半径
            # data[8-19]: 几何结果列
            # data[20-26]: 底宽B,直径D,半径R,边坡m,糙率n,底坡1/i,流量Q
            # data[27-31]: 水力结果列
            # data[32-40]: 水头损失列
            # data[41-43]: 水位,渠底高程,渠顶高程
            node = ChannelNode()
            node.flow_section = str(data[0]).strip()
            node.name = str(data[1]).strip()

            # 结构形式 (col 2)
            raw_struct_str = normalize_section_type_name(str(data[2]).strip())
            culvert_family_type = normalize_culvert_family_type_name(raw_struct_str)
            struct_str = raw_struct_str
            if culvert_family_type or struct_str in {"暗渠", "矩形暗渠", "矩形暗涵"}:
                struct_str = "矩形暗涵"
            if struct_str:
                try:
                    node.structure_type = StructureType.from_string(struct_str)
                except ValueError:
                    pass
                if struct_str == "渐变段" or (node.structure_type and node.structure_type == StructureType.TRANSITION):
                    node.is_transition = True

            # 标记闸类结构（分水闸/分水口/节制闸/泄水闸等）
            if node.structure_type and StructureType.is_diversion_gate(node.structure_type):
                node.is_diversion_gate = True
            # 标记倒虹吸
            if node.structure_type and node.structure_type == StructureType.INVERTED_SIPHON:
                node.is_inverted_siphon = True
            # 标记有压管道
            if node.structure_type and self._is_pressure_pipe_like_structure_text(node.structure_type.value):
                node.is_pressure_pipe = True
            pipe_material = ""
            local_loss_ratio = None
            in_out_raw = ""
            use_increase = True
            from_table1_source = False
            explicit_table1_source_flag = False
            flat_bottom_source_allowed = False
            arch_culvert_H_straight = None
            arch_culvert_manual_H_straight = None
            arch_culvert_used_manual_H_straight = None
            tunnel_arch_theta_deg = 0.0
            tunnel_arch_H_straight = None
            tunnel_arch_manual_H_straight = None
            tunnel_arch_used_manual_H_straight = None
            pressure_pipe_row_identity = ""
            compound_trapezoid_params = {}
            # 恢复自动插入补段标记（通过UserRole存储）
            _first_item = table.item(r, 0)
            if _first_item:
                _ur = _first_item.data(Qt.UserRole)
                if isinstance(_ur, dict) and _ur.get('_auto_channel'):
                    node.is_auto_inserted_channel = True
                    _auto_channel_struct = normalize_section_type_name(
                        str(_ur.get('_auto_channel_structure_type', '') or '')
                    )
                    if _auto_channel_struct and not node.structure_type:
                        try:
                            node.structure_type = StructureType.from_string(_auto_channel_struct)
                        except ValueError:
                            pass
                    try:
                        node.stat_length = float(_ur.get('_stat_length', 0.0) or 0.0)
                    except (TypeError, ValueError):
                        node.stat_length = 0.0
                elif _ur == "auto_channel":  # 兼容旧格式
                    node.is_auto_inserted_channel = True
                # 恢复渐变段详细参数（#10）
                if isinstance(_ur, dict) and _ur.get('_transition_data'):
                    td = _ur['_transition_data']
                    node.transition_type = td.get('transition_type', '')
                    node.transition_form = td.get('transition_form', '')
                    node.transition_zeta = td.get('transition_zeta', 0.0)
                    node.transition_theta = td.get('transition_theta', 0.0)
                    if not node.structure_type:
                        node.structure_type = StructureType.TRANSITION
                        node.is_transition = True
                if isinstance(_ur, dict):
                    if _ur.get('_aux_coords') or _ur.get('_auto_channel'):
                        try:
                            node.x = float(_ur.get('_x', 0.0) or 0.0)
                            node.y = float(_ur.get('_y', 0.0) or 0.0)
                        except (TypeError, ValueError):
                            node.x = 0.0
                            node.y = 0.0
                    _override = _ur.get("_transition_length_override_m", None)
                    if _override is not None and str(_override).strip() != "":
                        try:
                            node.transition_length_override_m = float(_override)
                        except (TypeError, ValueError):
                            node.transition_length_override_m = None
                    node.transition_length_source = str(_ur.get("_transition_length_source", "formula") or "formula")
                    node.transition_length_warning = str(_ur.get("_transition_length_warning", "") or "")
                    node.transition_rule_upstream_structure_type = str(
                        _ur.get("_transition_rule_upstream_structure_type", "") or ""
                    )
                    node.transition_rule_downstream_structure_type = str(
                        _ur.get("_transition_rule_downstream_structure_type", "") or ""
                    )
                    _length_details = _ur.get("_transition_length_calc_details", None)
                    if isinstance(_length_details, dict):
                        node.transition_length_calc_details = copy.deepcopy(_length_details)
                        if not node.transition_rule_upstream_structure_type:
                            node.transition_rule_upstream_structure_type = str(
                                _length_details.get("upstream_structure_type", "") or ""
                            )
                        if not node.transition_rule_downstream_structure_type:
                            node.transition_rule_downstream_structure_type = str(
                                _length_details.get("downstream_structure_type", "") or ""
                            )
                    _loss_details = _ur.get("_transition_loss_calc_details", None)
                    if isinstance(_loss_details, dict):
                        node.transition_calc_details = copy.deepcopy(_loss_details)
                    if "_from_table1_source" in _ur:
                        explicit_table1_source_flag = True
                        from_table1_source = bool(_ur.get("_from_table1_source"))
                    flat_bottom_source_allowed = self._is_flat_bottom_circle_source_payload(_ur)
                    arch_culvert_source_allowed = self._is_arch_culvert_source_payload(_ur)
                    pressure_pipe_row_identity = str(
                        _ur.get(PRESSURE_PIPE_ROW_ID_ROLE_KEY, "") or ""
                    ).strip()
                    pressure_pipe_window_override = self._normalize_pressure_pipe_window_override(
                        _ur.get(PRESSURE_PIPE_WINDOW_OVERRIDE_ROLE_KEY, {})
                    )
                    if pressure_pipe_window_override:
                        self._set_pressure_pipe_window_override(node, pressure_pipe_window_override)
                    pressure_pipe_named_group_result = self._normalize_pressure_pipe_named_group_result(
                        _ur.get(PRESSURE_PIPE_NAMED_GROUP_RESULT_ROLE_KEY, {})
                    )
                    if pressure_pipe_named_group_result:
                        self._set_pressure_pipe_named_group_result(node, pressure_pipe_named_group_result)
                    pressure_pipe_loss_override = self._normalize_pressure_pipe_loss_override_value(
                        _ur.get(PRESSURE_PIPE_LOSS_OVERRIDE_ROLE_KEY, None)
                    )
                    self._set_pressure_pipe_loss_override(node, pressure_pipe_loss_override)
                    _ext = _ur.get('_external_head_loss', None)
                    if _ext is not None and str(_ext).strip() != "":
                        try:
                            node.external_head_loss = float(_ext)
                        except (ValueError, TypeError):
                            node.external_head_loss = None
                    _pm = str(_ur.get('_pipe_material', '') or '').strip()
                    if _pm:
                        pipe_material = _pm
                    stored_culvert_family_type = normalize_culvert_family_type_name(
                        _ur.get(CULVERT_FAMILY_TYPE_KEY, "")
                    )
                    if stored_culvert_family_type:
                        culvert_family_type = stored_culvert_family_type
                    _ior = str(_ur.get('_in_out_raw', '') or '').strip()
                    if _ior:
                        in_out_raw = _ior
                    _llr = _ur.get('_local_loss_ratio', None)
                    if _llr is not None and str(_llr).strip() != "":
                        try:
                            local_loss_ratio = float(_llr)
                        except (ValueError, TypeError):
                            local_loss_ratio = None
                    if USE_INCREASE_ROLE_KEY in _ur:
                        use_increase = normalize_use_increase_flag(
                            _ur.get(USE_INCREASE_ROLE_KEY),
                            default=True,
                        )
                    compound_trapezoid_params = normalize_compound_trapezoid_params(
                        _ur.get(COMPOUND_TRAPEZOID_PARAMS_ROLE_KEY, {})
                    )
                    arch_culvert_theta_deg = self._sf(_ur.get(ARCH_CULVERT_THETA_ROLE_KEY, 0), 0.0)
                    if ARCH_CULVERT_H_STRAIGHT_ROLE_KEY in _ur:
                        arch_culvert_H_straight = self._sf(_ur.get(ARCH_CULVERT_H_STRAIGHT_ROLE_KEY, 0), 0.0)
                    if ARCH_CULVERT_MANUAL_H_STRAIGHT_ROLE_KEY in _ur:
                        arch_culvert_manual_H_straight = self._sf(_ur.get(ARCH_CULVERT_MANUAL_H_STRAIGHT_ROLE_KEY, 0), 0.0)
                    if ARCH_CULVERT_USED_MANUAL_H_STRAIGHT_ROLE_KEY in _ur:
                        arch_culvert_used_manual_H_straight = bool(_ur.get(ARCH_CULVERT_USED_MANUAL_H_STRAIGHT_ROLE_KEY))
                    tunnel_arch_theta_deg = self._sf(_ur.get(TUNNEL_ARCH_THETA_ROLE_KEY, 0), 0.0)
                    if TUNNEL_ARCH_H_STRAIGHT_ROLE_KEY in _ur:
                        tunnel_arch_H_straight = self._sf(_ur.get(TUNNEL_ARCH_H_STRAIGHT_ROLE_KEY, 0), 0.0)
                    if TUNNEL_ARCH_MANUAL_H_STRAIGHT_ROLE_KEY in _ur:
                        tunnel_arch_manual_H_straight = self._sf(_ur.get(TUNNEL_ARCH_MANUAL_H_STRAIGHT_ROLE_KEY, 0), 0.0)
                    if TUNNEL_ARCH_USED_MANUAL_H_STRAIGHT_ROLE_KEY in _ur:
                        tunnel_arch_used_manual_H_straight = bool(_ur.get(TUNNEL_ARCH_USED_MANUAL_H_STRAIGHT_ROLE_KEY))
                else:
                    arch_culvert_source_allowed = False
                    arch_culvert_theta_deg = 0.0
                    arch_culvert_H_straight = None
                    arch_culvert_manual_H_straight = None
                    arch_culvert_used_manual_H_straight = None
                    tunnel_arch_theta_deg = 0.0
                    tunnel_arch_H_straight = None
                    tunnel_arch_manual_H_straight = None
                    tunnel_arch_used_manual_H_straight = None
            else:
                arch_culvert_source_allowed = False
                arch_culvert_theta_deg = 0.0
                arch_culvert_H_straight = None
                arch_culvert_manual_H_straight = None
                arch_culvert_used_manual_H_straight = None
                tunnel_arch_theta_deg = 0.0
                tunnel_arch_H_straight = None
                tunnel_arch_manual_H_straight = None
                tunnel_arch_used_manual_H_straight = None
            if (
                not explicit_table1_source_flag
                and not from_table1_source
                and not getattr(node, 'is_transition', False)
                and not getattr(node, 'is_auto_inserted_channel', False)
            ):
                from_table1_source = True
            if struct_str == "隧洞-平底圆形":
                from_table1_source = bool(flat_bottom_source_allowed)
            node.from_table1_source = from_table1_source
            if struct_str == "隧洞-平底圆形" and not from_table1_source:
                node.structure_type = None
            if culvert_family_type == ARCH_CULVERT_FAMILY_TEXT and not arch_culvert_source_allowed:
                node.structure_type = None
                culvert_family_type = ""

            source_x_text = ""
            source_y_text = ""
            if _first_item and isinstance(_ur, dict):
                source_x_text = self._normalize_coord_text(_ur.get(SOURCE_COORD_X_ROLE_KEY, ""))
                source_y_text = self._normalize_coord_text(_ur.get(SOURCE_COORD_Y_ROLE_KEY, ""))

            x_text = self._normalize_coord_text(data[5])
            y_text = self._normalize_coord_text(data[6])
            if x_text:
                node.x = self._sf(x_text)
            if y_text:
                node.y = self._sf(y_text)
            if self._node_is_source_row(node):
                node.source_x_text = x_text or source_x_text
                node.source_y_text = y_text or source_y_text
            else:
                node.source_x_text = ""
                node.source_y_text = ""
            if pressure_pipe_row_identity:
                node.pressure_pipe_row_identity = pressure_pipe_row_identity
            elif self._is_pressure_pipe_row_override_node(node, channel_level):
                self._ensure_pressure_pipe_row_identity(node, r)

            # 进出口 (col 3)
            _io_text = _read_text(r, 3)
            if _io_text:
                node.in_out = InOutType.from_string(_io_text)
            # IP编号 (col 4) — 支持复合格式 (#11)
            ip_text = _read_text(r, 4)
            node.ip_number = _parse_ip_number(r, ip_text)
            node.display_ip_number = _parse_display_ip_number(r, ip_text)
            # 记录转弯半径是否来自用户/表格的显式输入；"0" 也算显式输入，不能再当成空白兜底。
            _turn_radius_text = _read_text(r, 7)
            _turn_radius_is_explicit = bool(_turn_radius_text)
            node.turn_radius_is_explicit = _turn_radius_is_explicit
            node.turn_radius_text = _turn_radius_text if _turn_radius_is_explicit else ""
            _row_turn = self._sf(_turn_radius_text, 0.0) if _turn_radius_is_explicit else 0.0
            # 新规则：转弯半径单元格留空也按 0 处理，不再回退到全局半径。
            if _turn_radius_is_explicit:
                node.turn_radius = _row_turn
            else:
                node.turn_radius = 0.0
            # 有压管道仅保留表格中的显式值；留空统一按 0 处理。
            try:
                _struct_for_turn = str(data[2]).strip()
                _is_pressure_turn = self._is_pressure_pipe_like_structure_text(_struct_for_turn)
                _is_siphon_turn = "倒虹吸" in _struct_for_turn
                if _is_pressure_turn:
                    if _turn_radius_is_explicit:
                        node.turn_radius = _row_turn
                    else:
                        node.turn_radius = 0.0
                elif _is_siphon_turn and _turn_radius_is_explicit:
                    node.turn_radius = _row_turn
            except Exception:
                pass

            # ===== 几何结果列 (8-19) =====
            # 转角 (col 8)
            _ta = _read_float(r, 8)
            if _ta > 0:
                node.turn_angle = _ta
            # 切线长 (col 9)
            _tl = _read_float(r, 9)
            if _tl > 0:
                node.tangent_length = _tl
            # 弧长 (col 10)
            _al = _read_float(r, 10)
            if _al > 0:
                node.arc_length = _al
            # 弯道长度 (col 11)
            _cl = _read_float(r, 11)
            if _cl > 0:
                node.curve_length = _cl
            # IP直线间距 (col 12)
            _sd = _read_float(r, 12)
            if _sd > 0:
                node.straight_distance = _sd
            # IP桩号 (col 13)
            node.station_ip = _parse_station(_read_text(r, 13))
            # 弯前BC (col 14)
            node.station_BC = _parse_station(_read_text(r, 14))
            # 里程MC (col 15)
            node.station_MC = _parse_station(_read_text(r, 15))
            # 弯末EC (col 16)
            node.station_EC = _parse_station(_read_text(r, 16))
            # 复核弯前 (col 17)
            _cpre = _read_float(r, 17)
            if _cpre != 0:
                node.check_pre_curve = _cpre
            # 复核弯后 (col 18)
            _cpost = _read_float(r, 18)
            if _cpost != 0:
                node.check_post_curve = _cpost
            # 复核总长 (col 19)
            _ctot = _read_float(r, 19)
            if _ctot != 0:
                node.check_total_length = _ctot

            # ===== 水力输入列 (20-26) =====
            B = self._sf(data[20])
            D = self._sf(data[21])
            R = self._sf(data[22])
            m_val = self._sf(data[23])
            # 糙率默认值：倒虹吸行用倒虹吸默认糙率常量，其他用渠道糙率输入框
            _default_n = (DEFAULT_SIPHON_ROUGHNESS
                          if struct_str and "倒虹吸" in struct_str
                          else self._fval(self.roughness_edit, DEFAULT_ROUGHNESS))
            n_val = self._sf(data[24], _default_n)
            slope_inv = self._sf(data[25])
            Q = self._sf(data[26], _default_q)

            # 与原版Tkinter get_nodes一致：始终写入B/D/R_circle/m（即使为0）
            # 原因：_estimate_transition_length中 section_params.get("D", 3.0)
            # 若D不在dict中会返回默认值3.0，导致隧洞渐变段长度偏大
            node.section_params['B'] = B
            node.section_params['D'] = D
            node.section_params['R_circle'] = R
            node.section_params['m'] = m_val
            node.section_params['use_increase'] = use_increase
            if culvert_family_type:
                node.section_params[CULVERT_FAMILY_TYPE_KEY] = culvert_family_type
            if struct_str == "明渠-复式梯形" and compound_trapezoid_params:
                node.section_params.update(copy.deepcopy(compound_trapezoid_params))
                if not node.section_params.get('B', 0) and compound_trapezoid_params.get('B2', 0):
                    node.section_params['B'] = compound_trapezoid_params['B2']
            node.use_increase = use_increase
            if pipe_material:
                node.section_params['pipe_material'] = pipe_material
            if in_out_raw:
                node.section_params['in_out_raw'] = in_out_raw
            if local_loss_ratio is not None and local_loss_ratio > 0:
                node.section_params['local_loss_ratio'] = local_loss_ratio
            node.roughness = n_val
            if slope_inv > 0:
                node.slope_i = 1.0 / slope_inv
            node.flow = Q

            # ===== 水力结果列 (27-31) =====
            _h = _read_float(r, 27)
            if _h > 0:
                node.water_depth = _h
            # 断面面积A (col 28)
            _area = _read_float(r, 28)
            if _area > 0:
                node.section_params['A'] = _area
            # 湿周X (col 29)
            _wp = _read_float(r, 29)
            if _wp > 0:
                node.section_params['X'] = _wp
            # 水力半径R (col 30)
            _hr2 = _read_float(r, 30)
            if _hr2 > 0:
                node.section_params['R'] = _hr2
            # 流速v (col 31)
            _v = _read_float(r, 31)
            if _v > 0:
                node.velocity = _v

            # ===== 水头损失列 (32-40) =====
            # 渐变段长度 (col 32)
            _trl_text = _read_text(r, 32)
            _trl = _read_float(r, 32)
            if _trl_text:
                node.transition_length = _trl
            # 渐变段损失 (col 33)
            _ht = _read_float(r, 33)
            if _ht != 0:
                node.head_loss_transition = _ht
            # 弯道损失 (col 34)
            _hb = _read_float(r, 34)
            if _hb != 0:
                node.head_loss_bend = _hb
            # 沿程损失 (col 35)
            _hf = _read_float(r, 35)
            if _hf != 0:
                node.head_loss_friction = _hf
            # 预留损失 (col 36)
            _hr = _read_float(r, 36)
            if _hr > 0:
                node.head_loss_reserve = _hr
            # 过闸损失 (col 37)
            _hg = _read_float(r, 37)
            gate_loss_user_set = self._is_gate_loss_user_set_for_row(r, node)
            if gate_loss_user_set:
                node.head_loss_gate = _hg
                self._set_gate_loss_user_set(node, True)
            elif _hg > 0:
                node.head_loss_gate = _hg
            elif self._is_gate_like_structure_text(struct_str):
                node.head_loss_gate = DEFAULT_GATE_HEAD_LOSS
            # 倒虹吸/有压管道损失 (col 38)
            _hs = _read_float(r, 38)
            self._apply_pressure_pipe_loss_cell_to_node(node, _hs, channel_level=channel_level)
            # 总损失 (col 39)
            _htotal = _read_float(r, 39)
            if _htotal != 0:
                node.head_loss_total = _htotal
            # 累计损失 (col 40)
            _hcum = _read_float(r, 40)
            if _hcum != 0:
                node.head_loss_cumulative = _hcum

            # ===== 高程列 (41-43) =====
            _wl = _read_float(r, 41)
            if _wl != 0:
                node.water_level = _wl
            _be = _read_float(r, 42)
            if _be != 0:
                node.bottom_elevation = _be
            _te = _read_float(r, 43)
            if _te != 0:
                node.top_elevation = _te

            # 恢复结构高度（用于计算渠顶高程，与Tkinter版 data_table.py 对齐）
            if r in self._node_structure_heights:
                node.structure_height = self._node_structure_heights[r]
            if culvert_family_type and node.structure_height > 0:
                node.section_params['H_total'] = node.structure_height
            if culvert_family_type == ARCH_CULVERT_FAMILY_TEXT:
                if arch_culvert_theta_deg > 0:
                    node.section_params['theta_deg'] = arch_culvert_theta_deg
                if arch_culvert_H_straight is not None:
                    node.section_params['H_straight'] = arch_culvert_H_straight
                if arch_culvert_manual_H_straight is not None:
                    node.section_params['manual_H_straight'] = arch_culvert_manual_H_straight
                if arch_culvert_used_manual_H_straight is not None:
                    node.section_params['used_manual_H_straight'] = arch_culvert_used_manual_H_straight
            if struct_str == "隧洞-圆拱直墙型":
                if node.structure_height > 0:
                    node.section_params['H_total'] = node.structure_height
                if tunnel_arch_theta_deg > 0:
                    node.section_params['theta_deg'] = tunnel_arch_theta_deg
                if tunnel_arch_H_straight is not None:
                    node.section_params['H_straight'] = tunnel_arch_H_straight
                if tunnel_arch_manual_H_straight is not None:
                    node.section_params['manual_H_straight'] = tunnel_arch_manual_H_straight
                if tunnel_arch_used_manual_H_straight is not None:
                    node.section_params['used_manual_H_straight'] = tunnel_arch_used_manual_H_straight

            # 恢复倒角参数（渡槽-矩形精确水力计算用）
            if r in self._node_chamfer_params:
                cp = self._node_chamfer_params[r]
                node.section_params['chamfer_angle'] = cp.get('chamfer_angle', 0)
                node.section_params['chamfer_length'] = cp.get('chamfer_length', 0)

            # 恢复明渠-U形圆心角
            if r in self._node_u_params:
                node.section_params['theta_deg'] = self._node_u_params[r].get('theta_deg', 0)
            # 恢复加大流速（从批量计算导入的加大流量工况流速）
            if r in self._node_velocity_increased:
                node.velocity_increased = self._node_velocity_increased[r]

            self._migrate_named_pressure_pipe_outlet_visible_group_loss(node, row_index=r)

            nodes.append(node)
        self._pressure_turn_radius_fallback_groups = set()
        return nodes

    def _collect_optional_blank_name_rows(self, nodes, channel_level: str | None = None):
        rows = []
        for idx, node in enumerate(nodes, start=1):
            if getattr(node, "is_transition", False) or getattr(node, "is_auto_inserted_channel", False):
                continue
            if not StructureType.warns_on_empty_name(getattr(node, "structure_type", None)):
                continue
            if str(getattr(node, "name", "") or "").strip():
                continue
            if self._is_unnamed_pressure_pipe_row_node(node, channel_level):
                continue
            rows.append((idx, node.get_structure_type_str() or "明渠"))
        return rows

    def _show_optional_blank_name_notice(self, nodes, *, action_name):
        channel_level = self._get_current_channel_level_text()
        rows = self._collect_optional_blank_name_rows(nodes, channel_level=channel_level)
        if not rows:
            return
        preview = "；".join(f"第{idx}行（{struct_name}）" for idx, struct_name in rows[:8])
        if len(rows) > 8:
            preview += f" 等{len(rows)}行"
        InfoBar.info(
            "提示",
            f"检测到部分暗涵未填写建筑物名称，建议补充名称便于识别；本次{action_name}不会中断：\n{preview}",
            parent=self._info_parent(),
            duration=5000,
            position=InfoBarPosition.TOP,
        )

    def _calculate(self):
        if not self._ensure_downstream_ready("执行计算"):
            return
        if not CALCULATOR_AVAILABLE:
            InfoBar.error("不可用", "核心计算引擎未加载，无法计算",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        if self.node_table.rowCount() < 2:
            InfoBar.warning("节点不足", "至少需要2个节点才能计算水面线",
                           parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        try:
            settings = self._build_settings()
            nodes = self._build_nodes_from_table()
            if not nodes or len(nodes) < 2:
                InfoBar.warning("数据不足", "有效节点不足", parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 前置检查：渐变段必须已插入
            if not self._has_transition_topology_ready(nodes):
                InfoBar.warning("提示",
                               "请先点击工具栏的【插入渐变段】按钮，完成渐变段插入后再执行计算。",
                               parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
                return

            # 前置检查：倒虹吸水力计算
            has_siphon = any(
                n.structure_type and "倒虹吸" in n.structure_type.value
                for n in nodes if n.structure_type
            )
            has_siphon_loss = any(
                n.structure_type
                and "倒虹吸" in n.structure_type.value
                and getattr(n, 'in_out', None) is not None
                and n.in_out.value == "出"
                and (getattr(n, 'head_loss_siphon', 0.0) or 0.0) > 0
                for n in nodes if n.structure_type
            )
            if has_siphon and not has_siphon_loss:
                InfoBar.warning("提示",
                               "检测到表格中包含倒虹吸，但尚未执行水力计算。"
                               "请先点击【倒虹吸水力计算】按钮完成计算后，再点击【执行计算】。",
                               parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
                return

            # 前置检查：只校验当前表3中确实需要写回的承压成员，锚点成员不再误报“未计算”。
            missing_pressure_pipe_groups = self._collect_pending_pressure_pipe_execute_members(
                nodes,
                settings=settings,
            )
            if missing_pressure_pipe_groups:
                InfoBar.warning(
                    "提示",
                    "检测到表格中包含有压管道同类结构（有压管道/定向钻/顶管），但尚未执行水力计算。"
                    "请先点击【有压管道水力计算】按钮完成计算后，再点击【执行计算】。",
                    parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP
                )
                return

            calculator = WaterProfileCalculator(settings)

            # 验证输入
            is_valid, errors = calculator.validate_input(nodes)
            if not is_valid:
                InfoBar.error("输入错误", "\n".join(errors),
                             parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
                return

            self._show_optional_blank_name_notice(nodes, action_name="计算")

            calculated = calculator.calculate_all(nodes)
            self.calculated_nodes = calculated
            self._settings = settings

            self._display_results(calculated, settings)
            self._generate_detail_report(calculated, settings, calculator)

            # 计算摘要
            summary = calculator.get_calculation_summary(calculated)
            total_len = summary.get('总长度', 0.0)
            wl_drop = summary.get('水位落差', 0.0)

            # 更新建筑物长度统计缓存
            self._last_building_lengths = calculator.calculate_building_lengths(calculated)
            self._last_channel_total_length = total_len
            self._last_type_summary = calculator.calculate_comprehensive_type_summary(calculated)

            # 更新持久摘要面板
            self._update_summary_panel(calculated, total_len, wl_drop, summary)

            # 检查缺少结构高度（渠顶高程无法计算）的节点。
            # 承压类节点已有独立导出与展示口径，这里只提示真正依赖结构总高的普通渠道/隧洞节点。
            missing_height_names = self._collect_missing_structure_height_names(calculated)

            msg = f"共{len(calculated)}个节点，总长{total_len:.1f}m，水位落差{wl_drop:.3f}m"
            gate_backfill_notice_lines = self._build_terminal_gate_backfill_notice_lines(calculated)
            if gate_backfill_notice_lines:
                msg += "\n" + "\n".join(gate_backfill_notice_lines)

            has_gate_backfill_issue = any(
                details.get("status") != "success"
                for details in self._collect_terminal_gate_backfill_records(calculated)
            )
            self._show_calculation_completion_notice(
                msg,
                calculated,
                missing_height_names=missing_height_names,
                has_gate_backfill_issue=has_gate_backfill_issue,
            )

            self._switch_workspace_tab(self._tab_output)

        except Exception as e:
            InfoBar.error("计算错误", f"计算过程出错: {str(e)}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            import traceback
            self.detail_text.setPlainText(f"计算错误:\n{traceback.format_exc()}")
            self._switch_to_output_process_tab(process_index=1)

    def _recalculate_silent(self):
        """静默重新推求水面线（跳过前置检查和InfoBar提示）。

        在倒虹吸/有压管道回写损失后自动调用，确保累计损失、
        水位、渠底/渠顶高程等下游列与新的损失值保持一致。
        """
        if not CALCULATOR_AVAILABLE:
            return
        try:
            settings = self._build_settings()
            nodes = self._build_nodes_from_table()
            if not nodes or len(nodes) < 2:
                return

            calculator = WaterProfileCalculator(settings)
            calculated = self._calculate_all_for_roundtrip_refresh(calculator, nodes)
            self.calculated_nodes = calculated
            self._settings = settings

            self._display_results(calculated, settings)
            self._generate_detail_report(calculated, settings, calculator)

            summary = calculator.get_calculation_summary(calculated)
            total_len = summary.get('总长度', 0.0)
            wl_drop = summary.get('水位落差', 0.0)
            self._last_building_lengths = calculator.calculate_building_lengths(calculated)
            self._last_channel_total_length = total_len
            self._last_type_summary = calculator.calculate_comprehensive_type_summary(calculated)
            self._update_summary_panel(calculated, total_len, wl_drop, summary)
        except Exception:
            import traceback
            traceback.print_exc()

    def _calculate_all_for_roundtrip_refresh(self, calculator, nodes):
        """按“回写后刷新”口径重算，避免进/出标记把既有转角清零。"""
        has_auxiliary_nodes = calculator._has_auxiliary_geometry_nodes(nodes)

        calculator.preprocess_nodes(nodes)
        if not has_auxiliary_nodes:
            nodes = calculator.identify_and_insert_transitions(nodes)

        calculator._calculate_geometry_preserving_special_turns(nodes)
        calculator.calculate_hydraulics(nodes)
        calculator.calculate_transition_losses(nodes)
        calculator._update_total_head_loss(nodes)
        calculator.hyd_calc.recalculate_water_levels_with_transition_losses(nodes)
        calculator.hyd_calc.apply_siphon_outlet_elevation(nodes)
        calculator.hyd_calc.apply_terminal_gate_elevation_backfill(nodes)
        calculator._calculate_cumulative_head_loss(nodes)
        calculator._validate_real_node_station_conflicts(nodes)
        return nodes

    # ================================================================
    # 结果显示
    # ================================================================
    def _display_results(self, nodes, settings):
        """将计算结果填充到统一node_table的结果列（第13列起）"""
        prefix = settings.get_station_prefix() if settings else ""

        # 先用计算后的节点重建整个表格（输入+结果一体）
        self._update_table_from_nodes_full(nodes, prefix)

        auto_resize_table(self.node_table)

    def _update_table_from_nodes_full(self, nodes, prefix=""):
        """用计算后的完整节点数据重建统一表格（输入列+结果列）"""
        with self._table_batch_update(self.node_table):
            self._update_table_from_nodes_full_impl(nodes, prefix)
            self._apply_table1_source_row_lock_flags()
        self._update_pressure_pipe_roughness_overview(
            self._collect_pressure_pipe_roughness_pairs_from_nodes(nodes)
        )
        self._refresh_pressure_pipe_controls()

    def _update_table_from_nodes_full_impl(self, nodes, prefix=""):
        current_channel_level = self._get_current_channel_level_text()
        # 更新结构高度缓存（计算完成后可能已重新计算）
        self._node_structure_heights.clear()
        for i, node in enumerate(nodes):
            if getattr(node, 'structure_height', 0) and node.structure_height > 0:
                self._node_structure_heights[i] = node.structure_height
        # 重建倒角参数缓存（节点计算往返后从 section_params 中恢复）
        self._node_chamfer_params.clear()
        for i, node in enumerate(nodes):
            sp = getattr(node, 'section_params', {}) or {}
            _ca = sp.get('chamfer_angle', 0) or 0
            _cl = sp.get('chamfer_length', 0) or 0
            if _ca > 0 and _cl > 0:
                self._node_chamfer_params[i] = {'chamfer_angle': float(_ca), 'chamfer_length': float(_cl)}
        # 重建明渠-U形圆心角缓存
        self._node_u_params.clear()
        for i, node in enumerate(nodes):
            sp = getattr(node, 'section_params', {}) or {}
            _th = sp.get('theta_deg', 0) or 0
            if _th > 0 and node.structure_type and 'U形' in node.structure_type.value and '明渠' in node.structure_type.value:
                self._node_u_params[i] = {'theta_deg': float(_th)}
        # 重建加大流速缓存（节点计算往返后从 velocity_increased 中恢复）
        self._node_velocity_increased.clear()
        for i, node in enumerate(nodes):
            _vi = getattr(node, 'velocity_increased', 0.0)
            if _vi and _vi > 0:
                self._node_velocity_increased[i] = float(_vi)
        self.node_table.setRowCount(0)
        for node in nodes:
            r = self.node_table.rowCount()
            self.node_table.insertRow(r)

            _is_trans = getattr(node, 'is_transition', False)
            _is_auto_ch = getattr(node, 'is_auto_inserted_channel', False)
            _is_source_row = self._node_is_source_row(node)

            # 构建完整46列数据，按列索引直接赋值
            vals = [""] * len(NODE_ALL_HEADERS)

            # 基础输入列 (0-7)
            vals[0] = node.flow_section
            vals[1] = node.name
            _st_str = node.get_structure_type_str()
            vals[2] = f"{_st_str}(连接段)" if _is_auto_ch else _st_str
            if not _is_trans:
                vals[3] = node.get_in_out_str()
                vals[4] = "" if _is_auto_ch else node.get_ip_str()
                vals[5] = "" if _is_auto_ch else self._resolve_node_coord_display_text(nodes, r, node, "x")
                vals[6] = "" if _is_auto_ch else self._resolve_node_coord_display_text(nodes, r, node, "y")
                # 转弯半径 col 7：显式填写值优先保留；只有真正留空时才按原规则显示。
                vals[7] = self._format_node_turn_radius_display_text(node, r)

            if not _is_trans:
                # 几何结果列 (8-19) — 无条件格式化，0值也显示（与Tkinter一致）
                _fmt_s = lambda s: ProjectSettings.format_station(s, prefix) if s is not None else "-"
                vals[8] = "" if _is_auto_ch else f"{node.turn_angle:.4f}"
                vals[9] = "" if _is_auto_ch else f"{node.tangent_length:.6f}"
                vals[10] = "" if _is_auto_ch else f"{node.arc_length:.6f}"
                vals[11] = "" if _is_auto_ch else f"{node.curve_length:.6f}"
                vals[12] = "" if _is_auto_ch else f"{node.straight_distance:.6f}"
                vals[13] = "" if _is_auto_ch else _fmt_s(node.station_ip)
                vals[14] = "" if _is_auto_ch else _fmt_s(node.station_BC)
                vals[15] = "" if _is_auto_ch else _fmt_s(getattr(node, 'station_MC', None))
                vals[16] = "" if _is_auto_ch else _fmt_s(node.station_EC)
                _skip_check = _is_auto_ch or getattr(node, 'is_inverted_siphon', False)
                vals[17] = "" if _skip_check else f"{getattr(node, 'check_pre_curve', 0):.3f}"
                vals[18] = "" if _skip_check else f"{getattr(node, 'check_post_curve', 0):.3f}"
                vals[19] = "" if _skip_check else f"{getattr(node, 'check_total_length', 0):.3f}"

                # 水力输入列 (20-26)
                _B = node.section_params.get('B', 0)
                _D = node.section_params.get('D', 0)
                _Rc = node.section_params.get('R_circle', 0)
                _m = node.section_params.get('m', 0)
                vals[20] = f"{_B:.3f}" if _B else ""
                vals[21] = f"{_D:.3f}" if _D else ""
                vals[22] = f"{_Rc:.3f}" if _Rc else ""
                vals[23] = f"{_m:.2f}" if _m else ""
                vals[24] = f"{node.roughness:.4f}" if node.roughness else ""
                vals[25] = f"{1.0/node.slope_i:.0f}" if node.slope_i and node.slope_i > 0 else ""
                vals[26] = f"{node.flow:.3f}" if node.flow else ""

                # 水力结果列 (27-31)
                _area = node.section_params.get('A', 0) if node.section_params else 0
                _peri = node.section_params.get('X', 0) if node.section_params else 0
                _hydr = node.section_params.get('R', 0) if node.section_params else 0
                vals[27] = f"{node.water_depth:.3f}" if node.water_depth else "-"
                vals[28] = f"{_area:.3f}" if _area else "-"
                vals[29] = f"{_peri:.3f}" if _peri else "-"
                vals[30] = f"{_hydr:.3f}" if _hydr else "-"
                vals[31] = f"{node.velocity:.3f}" if node.velocity else "-"

                # 水头损失列 (33-40) — 非渐变段行
                vals[33] = f"{node.head_loss_transition:.4f}" if node.head_loss_transition else "-"
                vals[34] = f"{node.head_loss_bend:.4f}" if node.head_loss_bend else "-"
                vals[35] = f"{node.head_loss_friction:.4f}" if node.head_loss_friction else "-"
                vals[36] = f"{getattr(node, 'head_loss_reserve', 0):.4f}" if getattr(node, 'head_loss_reserve', None) else "-"
                vals[37] = f"{node.head_loss_gate:.4f}" if node.head_loss_gate else "-"
                _h_sp = self._get_pressure_pipe_loss_display_value(
                    node,
                    row_index=r,
                    channel_level=current_channel_level,
                )
                vals[38] = f"{_h_sp:.4f}" if _h_sp else "-"
                vals[39] = f"{node.head_loss_total:.4f}" if node.head_loss_total else "-"
                vals[40] = f"{node.head_loss_cumulative:.4f}" if node.head_loss_cumulative else "-"

                # 高程列 (41-43)
                vals[41] = f"{node.water_level:.3f}" if node.water_level else "-"
                vals[42] = f"{node.bottom_elevation:.3f}" if node.bottom_elevation else "-"
                vals[43] = f"{node.top_elevation:.3f}" if node.top_elevation else "-"


            # 渐变段行特有数据
            if _is_trans:
                # 写入糙率/底坡/流量，确保通过表格读写循环不丢失
                vals[24] = f"{node.roughness:.4f}" if node.roughness else ""
                vals[25] = f"{1.0/node.slope_i:.0f}" if node.slope_i and node.slope_i > 0 else ""
                vals[26] = f"{node.flow:.3f}" if node.flow else ""
                vals[32] = f"{getattr(node, 'transition_length', 0):.3f}" if getattr(node, 'transition_length', None) else "-"
                vals[33] = f"{node.head_loss_transition:.4f}" if node.head_loss_transition else "-"
                vals[39] = f"{node.head_loss_transition:.4f}" if node.head_loss_transition else "-"
                vals[40] = f"{node.head_loss_cumulative:.4f}" if node.head_loss_cumulative else "-"

            # 渐变段长度（所有行通用）
            if not _is_trans:
                vals[32] = f"{getattr(node, 'transition_length', 0):.3f}" if getattr(node, 'transition_length', None) else "-"

            for c, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                item.setTextAlignment(Qt.AlignCenter)
                if c == 4:
                    item.setData(Qt.UserRole, {
                        "_raw_ip_number": getattr(node, "ip_number", 0),
                        "_display_ip_number": getattr(node, "display_ip_number", None),
                    })
                # 非可编辑列设为只读
                if c not in EDITABLE_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                # 第一行（水位起点）锁定水头损失列
                if r == 0 and c in FIRST_ROW_LOCKED_LOSS_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if _is_source_row and c in TABLE1_SOURCE_LOCKED_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if c == 7 and (
                    self._is_pressure_pipe_like_structure_text(_st_str)
                    or getattr(node, 'is_pressure_pipe', False)
                ):
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if c == 38 and self._is_pressure_pipe_display_locked_node(node, current_channel_level):
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                # 渐变段行灰色
                if _is_trans:
                    item.setForeground(QColor("#9E9E9E"))
                # 自动插入补段绿色
                elif _is_auto_ch:
                    item.setForeground(QColor("#2E7D32"))
                    item.setToolTip("自动插入的补段，用于计算两个建筑物之间的沿程及弯道水头损失。\n几何列留空因为该行不是真实IP转折点。")
                # 倒虹吸蓝色
                elif getattr(node, 'is_inverted_siphon', False):
                    item.setForeground(QColor("#1565C0"))
                # 分水闸橙色
                elif getattr(node, 'is_diversion_gate', False):
                    item.setForeground(QColor("#E65100"))
                self._apply_curve_check_item_style(item, c)
                self.node_table.setItem(r, c, item)

            # 在行首单元格中存储标记（UserRole），供_build_nodes_from_table恢复
            first_item = self.node_table.item(r, 0)
            if first_item:
                payload = first_item.data(Qt.UserRole)
                if not isinstance(payload, dict):
                    payload = {}
                if _is_auto_ch:
                    payload.update({
                        "_auto_channel": True,
                        "_auto_channel_structure_type": _st_str,
                        "_x": node.x,
                        "_y": node.y,
                        "_stat_length": float(getattr(node, "stat_length", 0.0) or 0.0),
                        "_aux_coords": True,
                    })
                elif _is_trans and (node.transition_type or node.transition_form):
                    # 渐变段详细参数保存到UserRole（#10）
                    payload.update({
                        '_transition_data': {
                            'transition_type': getattr(node, 'transition_type', ''),
                            'transition_form': getattr(node, 'transition_form', ''),
                            'transition_zeta': getattr(node, 'transition_zeta', 0.0),
                            'transition_theta': getattr(node, 'transition_theta', 0.0),
                        }
                    })
                    payload["_transition_rule_upstream_structure_type"] = str(
                        getattr(node, "transition_rule_upstream_structure_type", "") or ""
                    )
                    payload["_transition_rule_downstream_structure_type"] = str(
                        getattr(node, "transition_rule_downstream_structure_type", "") or ""
                    )
                if _is_trans:
                    payload.update({
                        "_x": node.x,
                        "_y": node.y,
                        "_aux_coords": True,
                    })
                if getattr(node, 'external_head_loss', None) is not None:
                    payload['_external_head_loss'] = getattr(node, 'external_head_loss')
                elif '_external_head_loss' in payload:
                    payload.pop('_external_head_loss', None)
                named_group_result = self._get_pressure_pipe_named_group_result(node)
                if named_group_result:
                    payload[PRESSURE_PIPE_NAMED_GROUP_RESULT_ROLE_KEY] = copy.deepcopy(named_group_result)
                else:
                    payload.pop(PRESSURE_PIPE_NAMED_GROUP_RESULT_ROLE_KEY, None)
                pressure_pipe_loss_override = self._get_pressure_pipe_loss_override(node)
                if pressure_pipe_loss_override is not None:
                    payload[PRESSURE_PIPE_LOSS_OVERRIDE_ROLE_KEY] = float(pressure_pipe_loss_override)
                else:
                    payload.pop(PRESSURE_PIPE_LOSS_OVERRIDE_ROLE_KEY, None)
                if self._is_gate_loss_user_set_node(node):
                    payload[GATE_HEAD_LOSS_USER_SET_ROLE_KEY] = True
                else:
                    payload.pop(GATE_HEAD_LOSS_USER_SET_ROLE_KEY, None)
                # 持久化有压管道专用参数（表格无专门列，放在UserRole）
                _sp = getattr(node, 'section_params', {}) or {}
                _pm = str(_sp.get('pipe_material', '') or '').strip()
                _ior = str(_sp.get('in_out_raw', '') or '').strip()
                _llr = _sp.get('local_loss_ratio', None)
                culvert_family_type = normalize_culvert_family_type_name(
                    _sp.get(CULVERT_FAMILY_TYPE_KEY, "")
                )
                if _pm:
                    payload['_pipe_material'] = _pm
                if _ior:
                    payload['_in_out_raw'] = _ior
                payload[USE_INCREASE_ROLE_KEY] = normalize_use_increase_flag(
                    _sp.get('use_increase', getattr(node, 'use_increase', True)),
                    default=True,
                )
                if _llr is not None and str(_llr).strip() != "":
                    try:
                        _llr_f = float(_llr)
                        if _llr_f > 0:
                            payload['_local_loss_ratio'] = _llr_f
                    except (ValueError, TypeError):
                        pass
                if self._should_persist_pressure_pipe_row_identity(node, current_channel_level):
                    payload[PRESSURE_PIPE_ROW_ID_ROLE_KEY] = self._ensure_pressure_pipe_row_identity(node, r)
                else:
                    payload.pop(PRESSURE_PIPE_ROW_ID_ROLE_KEY, None)
                payload[ARCH_CULVERT_SOURCE_ROLE_KEY] = bool(
                    culvert_family_type == ARCH_CULVERT_FAMILY_TEXT and getattr(node, "from_table1_source", False)
                )
                if culvert_family_type == ARCH_CULVERT_FAMILY_TEXT:
                    if _sp.get('theta_deg'):
                        payload[ARCH_CULVERT_THETA_ROLE_KEY] = _sp.get('theta_deg')
                    if _sp.get('H_straight') is not None:
                        payload[ARCH_CULVERT_H_STRAIGHT_ROLE_KEY] = _sp.get('H_straight')
                    if _sp.get('manual_H_straight') is not None:
                        payload[ARCH_CULVERT_MANUAL_H_STRAIGHT_ROLE_KEY] = _sp.get('manual_H_straight')
                    if _sp.get('used_manual_H_straight') is not None:
                        payload[ARCH_CULVERT_USED_MANUAL_H_STRAIGHT_ROLE_KEY] = bool(_sp.get('used_manual_H_straight'))
                override = self._get_pressure_pipe_window_override(node)
                if override:
                    payload[PRESSURE_PIPE_WINDOW_OVERRIDE_ROLE_KEY] = copy.deepcopy(override)
                else:
                    payload.pop(PRESSURE_PIPE_WINDOW_OVERRIDE_ROLE_KEY, None)
                compound_trapezoid_params = {}
                if _st_str == "明渠-复式梯形":
                    compound_trapezoid_params = normalize_compound_trapezoid_params(_sp)
                if compound_trapezoid_params:
                    payload[COMPOUND_TRAPEZOID_PARAMS_ROLE_KEY] = copy.deepcopy(compound_trapezoid_params)
                else:
                    payload.pop(COMPOUND_TRAPEZOID_PARAMS_ROLE_KEY, None)
                if _is_source_row:
                    payload[SOURCE_COORD_X_ROLE_KEY] = self._normalize_coord_text(vals[5])
                    payload[SOURCE_COORD_Y_ROLE_KEY] = self._normalize_coord_text(vals[6])
                else:
                    payload.pop(SOURCE_COORD_X_ROLE_KEY, None)
                    payload.pop(SOURCE_COORD_Y_ROLE_KEY, None)
                payload["_from_table1_source"] = bool(_is_source_row)
                if payload:
                    first_item.setData(Qt.UserRole, payload)
                if _is_trans:
                    self._persist_transition_calc_payload_for_row(r, node)

    def _generate_detail_report(self, nodes, settings, calculator=None):
        """生成详细计算过程文本"""
        prefix = settings.get_station_prefix() if settings else ""
        lines = []
        lines.append("=" * 80)
        lines.append(f"  {settings.channel_name if settings else ''}推求水面线 — 详细计算结果")
        lines.append("=" * 80)
        lines.append(f"  计算时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"  渠道名称: {settings.channel_name if settings else '-'}")
        lines.append(f"  渠道级别: {settings.channel_level if settings else '-'}")
        lines.append(f"  起始水位: {settings.start_water_level if settings else '-'} m")
        lines.append(f"  起始桩号: {ProjectSettings.format_station(settings.start_station, prefix) if settings else '-'}")
        # 多流量段显示
        if settings and getattr(settings, 'design_flows', None):
            flows_str = ", ".join(f"{q:.3f}" for q in settings.design_flows)
            lines.append(f"  设计流量: {flows_str} m³/s")
        else:
            lines.append(f"  设计流量: {settings.design_flow if settings else '-'} m³/s")
        if settings and getattr(settings, 'max_flows', None):
            flows_str = ", ".join(f"{q:.3f}" for q in settings.max_flows)
            lines.append(f"  加大流量: {flows_str} m³/s")
        else:
            lines.append(f"  加大流量: {settings.max_flow if settings else '-'} m³/s")
        lines.append(f"  糙率: {settings.roughness if settings else '-'}")
        if settings and getattr(settings, 'siphon_roughness', None) is not None:
            lines.append(f"  倒虹吸糙率: {settings.siphon_roughness}")
        lines.append(f"  转弯半径: {self._describe_turn_radius_entry_usage(settings)}")
        lines.append(f"  总节点数: {len(nodes)}")
        # 渐变段设置
        if settings:
            lines.append(f"  渡槽/隧洞渐变段: 进口{settings.transition_inlet_form}(ζ={settings.transition_inlet_zeta:.2f}), "
                        f"出口{settings.transition_outlet_form}(ζ={settings.transition_outlet_zeta:.2f})")
            if getattr(settings, 'open_channel_transition_form', None):
                lines.append(f"  明渠渐变段: {settings.open_channel_transition_form}(ζ={settings.open_channel_transition_zeta:.2f})")
            lines.append(f"  倒虹吸渐变段: 进口{settings.siphon_transition_inlet_form}(ζ={settings.siphon_transition_inlet_zeta:.2f}), "
                        f"出口{settings.siphon_transition_outlet_form}(ζ={settings.siphon_transition_outlet_zeta:.2f})")
        lines.append("-" * 80)
        lines.append("")

        for i, node in enumerate(nodes):
            # 节点标题
            tag = ""
            if getattr(node, 'is_transition', False):
                tag = " [渐变段]"
            elif getattr(node, 'is_auto_inserted_channel', False):
                tag = " [自动插入补段]"
            elif getattr(node, 'is_inverted_siphon', False):
                tag = " [倒虹吸]"
            elif getattr(node, 'is_diversion_gate', False):
                tag = f" [{node.get_structure_type_str()}]"
            lines.append(f"--- 节点 {i+1}: {node.name} ({node.get_structure_type_str()}){tag} ---")
            lines.append(f"  IP编号: {node.get_ip_str()}")
            lines.append(f"  进出口: {node.get_in_out_str()}")
            lines.append(f"  流量段: {node.flow_section}")
            lines.append(f"  坐标: ({node.x:.3f}, {node.y:.3f})")
            if node.station_ip:
                lines.append(f"  桩号: {ProjectSettings.format_station(node.station_ip, prefix)}")
            if node.azimuth:
                lines.append(f"  方位角: {node.azimuth:.6f}°")
            if node.turn_angle:
                lines.append(f"  转角: {node.turn_angle:.6f}°")
            lines.append(f"  流量 Q = {node.flow:.3f} m³/s")
            lines.append(f"  糙率 n = {node.roughness}")
            if node.slope_i:
                lines.append(f"  底坡 i = {node.slope_i:.6f}")
            # 断面参数
            sp = node.section_params
            if sp:
                parts = []
                if sp.get('B'): parts.append(f"B={sp['B']:.3f}")
                if sp.get('D'): parts.append(f"D={sp['D']:.3f}")
                if sp.get('R_circle'): parts.append(f"R={sp['R_circle']:.3f}")
                if sp.get('m'): parts.append(f"m={sp['m']:.2f}")
                if parts:
                    lines.append(f"  断面参数: {', '.join(parts)}")
            # 水力结果
            if node.water_depth:
                lines.append(f"  水深 h = {node.water_depth:.3f} m")
            if node.velocity:
                lines.append(f"  流速 v = {node.velocity:.3f} m/s")
            if node.water_level:
                lines.append(f"  水位 Z = {node.water_level:.3f} m")
            if node.bottom_elevation:
                lines.append(f"  渠底高程 = {node.bottom_elevation:.3f} m")
            if node.top_elevation:
                lines.append(f"  渠顶高程 = {node.top_elevation:.3f} m")
            if node.structure_height:
                lines.append(f"  结构高度 = {node.structure_height:.3f} m")
            # 水头损失
            loss_parts = []
            if node.head_loss_friction:
                loss_parts.append(f"沿程={node.head_loss_friction:.4f}")
            if node.head_loss_bend:
                loss_parts.append(f"弯道={node.head_loss_bend:.4f}")
            if node.head_loss_transition:
                loss_parts.append(f"渐变段={node.head_loss_transition:.4f}")
            if node.head_loss_gate:
                loss_parts.append(f"过闸={node.head_loss_gate:.4f}")
            if node.head_loss_siphon:
                loss_parts.append(f"倒虹吸/有压管道={node.head_loss_siphon:.4f}")
            if node.head_loss_reserve:
                loss_parts.append(f"预留={node.head_loss_reserve:.4f}")
            if loss_parts:
                lines.append(f"  水头损失: {', '.join(loss_parts)}")
            if node.head_loss_total:
                lines.append(f"  总损失 = {node.head_loss_total:.4f} m")
            if node.head_loss_cumulative:
                lines.append(f"  累计损失 = {node.head_loss_cumulative:.4f} m")
            # 渐变段详情
            if getattr(node, 'is_transition', False) and node.transition_length:
                lines.append(f"  渐变段类型: {node.transition_type}")
                lines.append(f"  渐变段形式: {node.transition_form}")
                lines.append(f"  渐变段长度 L = {node.transition_length:.3f} m")
                if node.transition_zeta:
                    lines.append(f"  局部损失系数 ζ = {node.transition_zeta:.3f}")
            lines.append("")

        # 计算摘要
        if calculator:
            summary = calculator.get_calculation_summary(nodes)
            if summary:
                lines.append("=" * 80)
                lines.append("  计算摘要")
                lines.append("-" * 80)
                lines.append(f"  节点数量: {summary.get('节点数量', '-')}")
                s_start = summary.get('起点桩号', 0.0)
                s_end = summary.get('终点桩号', 0.0)
                lines.append(f"  起点桩号: {ProjectSettings.format_station(s_start, prefix)}")
                lines.append(f"  终点桩号: {ProjectSettings.format_station(s_end, prefix)}")
                lines.append(f"  总长度: {summary.get('总长度', 0.0):.3f} m")
                wl_s = summary.get('起点水位', 0.0)
                wl_e = summary.get('终点水位', 0.0)
                if wl_s and wl_e:
                    lines.append(f"  起点水位: {wl_s:.3f} m")
                    lines.append(f"  终点水位: {wl_e:.3f} m")
                    lines.append(f"  水位落差: {summary.get('水位落差', 0.0):.3f} m")
                lines.append("")

            # 建筑物长度汇总
            try:
                building_lengths = calculator.calculate_building_lengths(nodes)
                if building_lengths:
                    lines.append("=" * 80)
                    lines.append("  建筑物长度汇总")
                    lines.append("-" * 80)
                    lines.append(f"  {'序号':<4}  {'名称':<16}  {'结构形式':<12}  {'长度(m)':<10}  {'起始桩号':<16}  {'终止桩号':<16}")
                    lines.append("  " + "-" * 76)
                    for i, bl in enumerate(building_lengths, 1):
                        name = bl.get('name', '-')
                        stype = bl.get('structure_type', '-')
                        length = bl.get('length', 0.0)
                        s_s = bl.get('start_station', 0.0)
                        s_e = bl.get('end_station', 0.0)
                        lines.append(
                            f"  {i:<4}  {name:<16}  {stype:<12}  {length:<10.3f}  "
                            f"{ProjectSettings.format_station(s_s, prefix):<16}  "
                            f"{ProjectSettings.format_station(s_e, prefix):<16}"
                        )
                    total_length = sum(bl.get('length', 0.0) for bl in building_lengths)
                    lines.append("  " + "-" * 76)
                    lines.append(f"  {'合计':<22}  {'':<12}  {total_length:<10.3f}")
                    lines.append("")
            except Exception:
                pass

        gate_backfill_lines = self._build_terminal_gate_backfill_report_lines(nodes)
        if gate_backfill_lines:
            lines.extend(gate_backfill_lines)

        lines.append("=" * 80)
        lines.append("  计算完毕")
        lines.append("=" * 80)
        self.detail_text.setPlainText("\n".join(lines))

    # ================================================================
    # 辅助
    # ================================================================
    def eventFilter(self, obj, event):
        """事件过滤器：起始桩号焦点 + 表头悬浮公式提示"""
        from PySide6.QtCore import QEvent, QPoint
        if obj is self.start_station_edit:
            if event.type() == QEvent.FocusIn:
                current = self.start_station_edit.text()
                value = parse_station_input(current)
                self.start_station_edit.setText(str(value))
        elif hasattr(self, '_node_header') and obj is self._node_header.viewport():
            header = self._node_header
            if event.type() == QEvent.Type.MouseMove:
                try:
                    pos = event.position().toPoint()
                except AttributeError:
                    pos = event.pos()
                logical_idx = header.logicalIndexAt(pos)
                if logical_idx >= 0 and logical_idx in self._formula_columns:
                    col_name = NODE_ALL_HEADERS[logical_idx]
                    vp_x = header.sectionViewportPosition(logical_idx)
                    sec_w = header.sectionSize(logical_idx)
                    gp = header.mapToGlobal(QPoint(vp_x + sec_w // 2, header.height()))
                    self._formula_tooltip.show_for_column(col_name, gp)
                else:
                    self._formula_tooltip.schedule_hide()
            elif event.type() == QEvent.Type.Leave:
                self._formula_tooltip.schedule_hide()
        elif hasattr(self, "node_table") and self.node_table and obj in (self.node_table, self.node_table.viewport()):
            if obj is self.node_table.viewport() and event.type() == QEvent.Type.MouseButtonDblClick:
                try:
                    pos = event.position().toPoint()
                except AttributeError:
                    pos = event.pos()
                index = self.node_table.indexAt(pos)
                if index.isValid() and index.column() == 32 and self._is_transition_row(index.row()):
                    self.node_table.setCurrentCell(index.row(), index.column())
                    self._on_node_cell_double_clicked(index.row(), index.column())
                    return True
            if event.type() == QEvent.Type.KeyPress:
                key = event.key()
                modifiers = event.modifiers()
                row = self.node_table.currentRow()
                col = self.node_table.currentColumn()
                if col == 32 and self._is_transition_length_editable_cell(row, col):
                    if key in {Qt.Key_Return, Qt.Key_Enter, Qt.Key_F2} and not (modifiers & Qt.ControlModifier):
                        self._begin_transition_length_edit(row)
                        return True
                is_edit_key = bool(event.text().strip()) or key in {
                    Qt.Key_Backspace, Qt.Key_Delete, Qt.Key_Return, Qt.Key_Enter, Qt.Key_F2
                }
                is_paste = bool(modifiers & Qt.ControlModifier) and key in {Qt.Key_V, Qt.Key_X}
                if is_edit_key or is_paste:
                    if self._is_table1_source_locked_cell(row, col):
                        self._show_table1_source_lock_hint()
                        return True
        return super().eventFilter(obj, event)

    def _format_start_station(self):
        """编辑完成后格式化起始桩号显示"""
        current = self.start_station_edit.text().strip()
        value = parse_station_input(current)
        formatted = format_station_display(value)
        self.start_station_edit.setText(formatted)

    def _get_flow_values_from_widget(self, widget) -> list:
        """兼容旧控件与新控件，统一读取流量列表。"""
        if widget is None:
            return []
        values_getter = getattr(widget, "values", None)
        if callable(values_getter):
            try:
                return list(values_getter())
            except Exception:
                pass
        text_getter = getattr(widget, "text", None)
        if callable(text_getter):
            try:
                return parse_flow_values_text(text_getter())
            except Exception:
                return []
        return []

    def _sync_flow_segment_widgets(self, reset_index: bool = False):
        """按共享当前段刷新设计流量和加大流量的紧凑摘要。"""
        if reset_index:
            self._flow_segment_current_index = 0

        design_values = self._get_flow_values_from_widget(getattr(self, "design_flow_edit", None))
        max_values = self._get_flow_values_from_widget(getattr(self, "max_flow_edit", None))
        total_count = max(len(design_values), len(max_values))
        if total_count <= 0:
            self._flow_segment_current_index = 0
        else:
            self._flow_segment_current_index = min(max(int(self._flow_segment_current_index), 0), total_count - 1)

        for widget in (getattr(self, "design_flow_edit", None), getattr(self, "max_flow_edit", None)):
            setter = getattr(widget, "set_current_segment_index", None)
            if callable(setter):
                setter(self._flow_segment_current_index)

        if self.isVisible():
            self._schedule_adjust_splitter_for_settings()

    def _set_flow_segment_current_index(self, index: int):
        """设置共享当前流量段，并同步两个摘要控件。"""
        self._flow_segment_current_index = max(0, int(index))
        self._sync_flow_segment_widgets(reset_index=False)

    def _reset_flow_segment_current_index(self):
        """把当前流量段重置回第一段。"""
        self._set_flow_segment_current_index(0)

    def _refresh_roughness_overview_visibility(self):
        groups = getattr(self, "_siphon_pressure_group", None)
        if groups is None:
            return
        if not isinstance(groups, (list, tuple)):
            groups = [groups]
        siphon_pairs = getattr(getattr(self, "siphon_roughness_chips", None), "_pairs", [])
        ppipe_pairs = getattr(getattr(self, "pressure_pipe_roughness_chips", None), "_pairs", [])
        show_group = bool(siphon_pairs) or bool(ppipe_pairs)
        changed = False
        for group in groups:
            if group is None:
                continue
            if group.isVisible() != show_group:
                group.setVisible(show_group)
                changed = True
        if changed and self.isVisible():
            QTimer.singleShot(0, self._adjust_splitter_for_settings)

    def _update_siphon_roughness_overview(self, pairs):
        """更新倒虹吸糙率芯片展示。
        pairs: [(名称, 糙率), ...] 列表
        """
        if not pairs:
            self.siphon_roughness_chips.clear()
            self._refresh_roughness_overview_visibility()
            return
        # 去重：同名倒虹吸只取第一个（批量计算中同一倒虹吸可能有多行）
        seen = {}
        for name, n_val in pairs:
            if name not in seen:
                seen[name] = n_val
        self.siphon_roughness_chips.set_siphon_data(list(seen.items()))
        self._refresh_roughness_overview_visibility()

    def _update_pressure_pipe_roughness_overview(self, pairs):
        """更新有压管道参数芯片展示。pairs: [(名称, 管材名称), ...]"""
        if not hasattr(self, 'pressure_pipe_roughness_chips'):
            return
        if not pairs:
            self.pressure_pipe_roughness_chips.clear()
            self._refresh_roughness_overview_visibility()
            return
        seen = {}
        default_idx = 1
        for name, material in pairs:
            display_name = str(name).strip() if name else ""
            if not display_name:
                display_name = f"有压管道{default_idx}"
                default_idx += 1
            if display_name not in seen:
                params = SiphonRoughnessChipContainer.PIPE_MATERIAL_PARAMS.get(material, {})
                if params:
                    param_str = f"{material} | f={params['f']}, m={params['m']}, b={params['b']}"
                else:
                    param_str = material if material else "未指定管材"
                seen[display_name] = param_str
        self.pressure_pipe_roughness_chips.set_pairs(list(seen.items()))
        self._refresh_roughness_overview_visibility()

    def _collect_pressure_pipe_roughness_pairs_from_nodes(self, nodes):
        """从节点列表提取有压管道参数展示对。"""
        pairs = []
        if not nodes:
            return pairs
        for node in nodes:
            if not self._is_named_pressure_pipe_group_node(node):
                continue
            name = (node.name or "").strip()
            material = node.section_params.get('pipe_material', '') if hasattr(node, 'section_params') else ''
            pairs.append((name, material))
        return pairs

    def _refresh_pressure_pipe_controls(self):
        """刷新有压管道按钮提示状态；启停状态由统一下游门禁控制。"""
        btn = getattr(self, "btn_pressure_pipe_calc", None)
        if btn is None:
            self._refresh_pressure_pipe_water_hammer_controls()
            return
        if not getattr(self, "_section_sync_ready", False):
            btn.setToolTip("")
            self._refresh_pressure_pipe_water_hammer_controls()
            return
        table = getattr(self, "node_table", None)
        if table is None:
            btn.setToolTip("执行有压管道水力计算并回写到\"倒虹吸/有压管道水头损失\"列")
            self._refresh_pressure_pipe_water_hammer_controls()
            return
        has_ppipe = False
        has_named_ppipe_group = False
        has_transition = False
        for r in range(table.rowCount()):
            st_item = table.item(r, 2)
            st = st_item.text().strip() if st_item else ""
            if self._is_pressure_pipe_like_structure_text(st):
                has_ppipe = True
                name_item = table.item(r, 1)
                name_text = name_item.text().strip() if name_item else ""
                if name_text:
                    has_named_ppipe_group = True
            if "渐变段" in st:
                has_transition = True
            if has_named_ppipe_group and has_transition:
                break
        transition_topology_ready = has_transition or bool(getattr(self, "_transition_topology_prepared", False))
        if not has_ppipe:
            self.btn_pressure_pipe_calc.setToolTip("尚未检测到有压管道同类节点。")
        elif not has_named_ppipe_group:
            self.btn_pressure_pipe_calc.setToolTip("仅检测到匿名有压管道行；当前渠道级别满足条件时可直接打开有压管道水力计算窗口。")
        elif not transition_topology_ready:
            self.btn_pressure_pipe_calc.setToolTip("已检测到有压管道同类结构。请先插入渐变段后再执行有压管道水力计算")
        else:
            self.btn_pressure_pipe_calc.setToolTip("执行有压管道水力计算并回写到\"倒虹吸/有压管道水头损失\"列")
        self._refresh_pressure_pipe_water_hammer_controls()

    @staticmethod
    def _is_valid_pressure_pipe_water_hammer_level(value) -> bool:
        """判断表3节点水位是否可作为水锤验算基础水位。"""
        try:
            return float(value) > ZERO_TOLERANCE
        except (TypeError, ValueError):
            return False

    def _collect_pressure_pipe_water_hammer_nodes(self):
        """读取当前可用于水锤验算的表3节点，优先使用最近一次计算结果。"""
        nodes = list(getattr(self, "calculated_nodes", None) or [])
        if nodes:
            return nodes
        try:
            return list(self._build_nodes_from_table() or [])
        except Exception:
            return []

    def _has_pressure_pipe_like_table_row_for_water_hammer(self) -> bool:
        """兜底检查表3中是否存在有压管道同类行。"""
        table = getattr(self, "node_table", None)
        if table is None:
            return False
        for row in range(table.rowCount()):
            item = table.item(row, 2)
            text = item.text().strip() if item else ""
            if self._is_pressure_pipe_like_structure_text(text):
                return True
        return False

    def _resolve_pressure_pipe_water_hammer_button_state(self) -> tuple[bool, str]:
        """返回水锤验算按钮是否可用及对应提示原因。"""
        if not getattr(self, "_section_sync_ready", False):
            return False, "请先完成断面批量计算并同步到表3后再进行水锤验算。"

        nodes = self._collect_pressure_pipe_water_hammer_nodes()
        has_pressure_pipe = any(self._is_pressure_pipe_like_node(node) for node in nodes)
        if not has_pressure_pipe:
            has_pressure_pipe = self._has_pressure_pipe_like_table_row_for_water_hammer()
        if not has_pressure_pipe:
            return False, "尚未检测到有压管道、定向钻或顶管。"

        try:
            settings = self._build_settings()
        except Exception:
            settings = None
        try:
            pending = self._collect_pending_pressure_pipe_execute_members(nodes, settings=settings)
        except Exception:
            return False, "请先完成有压管道水力计算并应用成功结果后再进行水锤验算。"
        if pending:
            return False, "请先完成有压管道水力计算并应用成功结果后再进行水锤验算。"

        has_water_level = any(
            self._is_pressure_pipe_like_node(node)
            and self._is_valid_pressure_pipe_water_hammer_level(getattr(node, "water_level", None))
            for node in nodes
        )
        if not has_water_level:
            return False, "请先点击【执行计算】生成有效表3水位后再进行水锤验算。"

        return True, "执行有压管道水锤验算"

    def _refresh_pressure_pipe_water_hammer_controls(self):
        """按前置条件刷新有压管道水锤验算按钮。"""
        btn = getattr(self, "btn_pressure_pipe_water_hammer", None)
        if btn is None:
            return
        enabled, reason = self._resolve_pressure_pipe_water_hammer_button_state()
        btn.setEnabled(bool(enabled))
        btn.setToolTip(reason or "")

    def _update_pressure_pipe_last_result_button(self):
        """刷新有压管道计算相关控件状态（计算完成后调用）。"""
        self._refresh_pressure_pipe_controls()

    def _choose_roughness_value(self, values, label):
        """当同类建筑物糙率不一致时，弹窗让用户选择。
        values: 糙率值列表（已收集的同类建筑物糙率）
        label: 显示标签（如"渠道糙率"或"倒虹吸糙率"）
        返回选中的糙率值，若列表为空返回None。
        """
        if not values:
            return None
        from collections import Counter
        counter = Counter(values)
        unique_vals = sorted(counter.keys())
        # 所有值相同，直接返回
        if len(unique_vals) == 1:
            return unique_vals[0]
        # 构建选项列表
        options = []
        for v in unique_vals:
            cnt = counter[v]
            options.append(f"{v}    （出现 {cnt} 次）")
        from PySide6.QtWidgets import QInputDialog
        dialog = QInputDialog(self)
        dialog.setWindowTitle(f"选择{label}")
        dialog.setLabelText(f"同步到表3时，不同建筑物的{label}值不一致，请选择一个用于本次导入及后续计算：")
        dialog.setComboBoxItems(options)
        dialog.setComboBoxEditable(False)
        dialog.setTextValue(options[0])
        dialog.setOkButtonText("确定")
        dialog.setCancelButtonText("取消")
        if dialog.exec():
            chosen = dialog.textValue() or options[0]
            # 从选项文本中提取数值
            val_str = chosen.split("（")[0].strip()
            try:
                return float(val_str)
            except ValueError:
                return unique_vals[0]
        # 用户取消，保留现有差异
        return None

    def _fval(self, edit, default=0.0):
        t = edit.text().strip()
        if not t: return default
        try: return float(t)
        except ValueError: return default

    def _sf(self, val, default=0.0):
        if not val: return default
        s = str(val).strip()
        if not s: return default
        try: return float(s)
        except ValueError: return default

    def _is_negative_curve_check_display_text(self, text) -> bool:
        """按表格三位显示值判断复核长度是否为负数。"""
        try:
            return float(str(text).strip()) < 0.0
        except (TypeError, ValueError):
            return False

    def _apply_curve_check_item_style(self, item, col: int):
        """把复核长度列中显示为负数的单元格标红。"""
        if col not in CURVE_CHECK_COLUMN_LABELS:
            return
        if item is None or not self._is_negative_curve_check_display_text(item.text()):
            return
        item.setForeground(QColor(CURVE_CHECK_NEGATIVE_COLOR))

    def _skip_curve_check_notice_for_node(self, node) -> bool:
        """判断节点是否不参与复核长度负数提示。"""
        if getattr(node, "is_transition", False):
            return True
        if getattr(node, "is_auto_inserted_channel", False):
            return True
        if getattr(node, "is_inverted_siphon", False):
            return True
        structure_value = str(getattr(getattr(node, "structure_type", None), "value", "") or "")
        return structure_value == "倒虹吸"

    def _collect_negative_curve_check_entries(self, nodes):
        """收集复核长度显示值为负数的行列信息。"""
        entries = []
        for row_idx, node in enumerate(nodes or [], start=1):
            if self._skip_curve_check_notice_for_node(node):
                continue
            for attr_name, label in CURVE_CHECK_ATTRS:
                text = f"{float(getattr(node, attr_name, 0.0) or 0.0):.3f}"
                if self._is_negative_curve_check_display_text(text):
                    entries.append({
                        "row": row_idx,
                        "label": label,
                        "value": text,
                    })
        return entries

    def _format_negative_curve_check_notice(self, nodes) -> str:
        """生成复核长度负数的非阻断提示文本。"""
        entries = self._collect_negative_curve_check_entries(nodes)
        if not entries:
            return ""

        grouped_lines = []
        grouped = {}
        for entry in entries[:CURVE_CHECK_NOTICE_LIMIT]:
            grouped.setdefault(entry["row"], []).append(f'{entry["label"]} {entry["value"]}')
        for row, parts in grouped.items():
            grouped_lines.append(f"第{row}行：" + "、".join(parts))

        if len(entries) > CURVE_CHECK_NOTICE_LIMIT:
            grouped_lines.append(f"等{len(entries)}处复核长度为负数")

        return "复核长度存在负数，请复核弯道半径或相邻IP间距：\n" + "\n".join(grouped_lines)

    def _show_calculation_completion_notice(
        self,
        msg: str,
        calculated,
        *,
        missing_height_names=None,
        has_gate_backfill_issue: bool = False,
    ):
        """统一显示执行计算完成后的成功或非阻断警告提示。"""
        final_msg = str(msg or "")
        missing_height_names = list(missing_height_names or [])
        if missing_height_names:
            final_msg += f"\n⚠ 以下节点缺少结构总高，渠顶高程未计算: {', '.join(missing_height_names)}"
            final_msg += "\n请通过【断面批量计算】并自动同步后获取正确的结构总高。"

        negative_notice = self._format_negative_curve_check_notice(calculated)
        if negative_notice:
            final_msg += "\n" + negative_notice

        if missing_height_names:
            InfoBar.warning(
                "计算完成（部分渠顶高程缺失）",
                final_msg,
                parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP,
            )
        elif has_gate_backfill_issue:
            InfoBar.warning(
                "计算完成（末尾闸行回推未完成）",
                final_msg,
                parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP,
            )
        elif negative_notice:
            InfoBar.warning(
                "计算完成（复核长度需复核）",
                final_msg,
                parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP,
            )
        else:
            InfoBar.success(
                "计算完成",
                final_msg,
                parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP,
            )

    def _info_parent(self):
        return self

    @contextmanager
    def _table_batch_update(self, table):
        """批量更新表格：屏蔽信号 + 暂停重绘 + 兜底维护 _updating_cells。"""
        if table is None:
            yield
            return
        prev_updating = self._updating_cells
        prev_updates_enabled = table.updatesEnabled()
        self._updating_cells = True
        blocker = QSignalBlocker(table)
        if prev_updates_enabled:
            table.setUpdatesEnabled(False)
        try:
            yield
        finally:
            if prev_updates_enabled:
                table.setUpdatesEnabled(True)
            del blocker
            self._updating_cells = prev_updating

    def _calculate_recommended_turn_radius(self, nodes):
        """根据规范计算推荐的转弯半径（取大值原则）
        - 隧洞：弯曲半径≥洞径(或洞宽)×5
        - 明渠：弯曲半径≥水面宽度×5
        - 渡槽：弯道半径≥连接明渠渠底宽度×5
        """
        import math as _math
        max_r = 0.0
        for node in nodes:
            if not node.structure_type:
                continue
            sv = resolve_node_effective_structure_type_text(node)
            if "倒虹吸" in sv:
                continue
            b = node.section_params.get("B", 0)
            rc = node.section_params.get("R_circle", 0)
            m_s = node.section_params.get("m", 0)
            wd = node.water_depth or 0
            min_r = 0.0
            if "隧洞" in sv:
                if rc > 0:
                    min_r = rc * 2 * 5
                elif b > 0:
                    min_r = b * 5
            elif "明渠" in sv or sv == "矩形":
                if "U形" in sv and rc > 0:
                    theta_u_r = node.section_params.get('theta_deg', 0) or 0
                    if theta_u_r > 0 and wd > 0:
                        import math as _math_u
                        h0_u_r = rc * (1.0 - _math_u.cos(_math_u.radians(theta_u_r / 2.0)))
                        if wd <= h0_u_r:
                            wsw_u = 2 * _math_u.sqrt(max(0.0, rc ** 2 - (rc - wd) ** 2))
                        else:
                            b_arc_u_r = 2 * rc * _math_u.sin(_math_u.radians(theta_u_r / 2.0))
                            wsw_u = b_arc_u_r + 2 * m_s * (wd - h0_u_r)
                        min_r = wsw_u * 5
                    else:
                        min_r = rc * 2 * 5
                elif b > 0:
                    wsw = b + 2 * m_s * wd if m_s > 0 and wd > 0 else b
                    min_r = wsw * 5
            elif "渡槽" in sv:
                if b > 0:
                    min_r = b * 5
            elif "暗涵" in sv:
                if b > 0:
                    min_r = b * 5
            if min_r > max_r:
                max_r = min_r
        if max_r <= 0:
            max_r = DEFAULT_AUTO_TURN_RADIUS
        return _math.ceil(max_r)

    def _on_trans_inlet_form_changed(self, form):
        """渡槽/隧洞进口渐变段形式变化时自动更新ζ系数"""
        zeta_table = TRANSITION_ZETA_COEFFICIENTS.get("进口", {})
        if form in zeta_table:
            self.trans_inlet_zeta.setText(f"{zeta_table[form]:.2f}")
        elif form == "直线形扭曲面":
            self.trans_inlet_zeta.setText("0.05")

    def _on_trans_outlet_form_changed(self, form):
        """渡槽/隧洞出口渐变段形式变化时自动更新ζ系数"""
        zeta_table = TRANSITION_ZETA_COEFFICIENTS.get("出口", {})
        if form in zeta_table:
            self.trans_outlet_zeta.setText(f"{zeta_table[form]:.2f}")
        elif form == "直线形扭曲面":
            self.trans_outlet_zeta.setText("0.14")

    def _on_oc_trans_form_changed(self, form):
        """明渠渐变段形式变化时自动更新ζ系数（使用进口系数）"""
        zeta_table = TRANSITION_ZETA_COEFFICIENTS.get("进口", {})
        if form in zeta_table:
            self.oc_trans_zeta.setText(f"{zeta_table[form]:.2f}")
        elif form == "直线形扭曲面":
            self.oc_trans_zeta.setText("0.05")

    def _on_siphon_inlet_form_changed(self, form):
        """倒虹吸进口渐变段型式变化时自动更新ζ系数"""
        zeta_table = SIPHON_TRANSITION_ZETA_COEFFICIENTS.get("进口", {})
        if form in zeta_table:
            self.siphon_inlet_zeta.setText(f"{zeta_table[form]:.2f}")

    def _on_siphon_outlet_form_changed(self, form):
        """倒虹吸出口渐变段型式变化时自动更新ζ系数"""
        zeta_table = SIPHON_TRANSITION_ZETA_COEFFICIENTS.get("出口", {})
        if form in zeta_table:
            self.siphon_outlet_zeta.setText(f"{zeta_table[form]:.2f}")

    def _auto_calc_turn_radius(self):
        """根据规范自动计算推荐转弯半径，并弹出详细计算过程"""
        if not CALCULATOR_AVAILABLE:
            InfoBar.warning("提示", "核心计算引擎未加载",
                           parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return
        nodes = self._build_nodes_from_table()
        if not nodes:
            InfoBar.info("提示", "请先导入节点数据",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        import math as _math
        details = []
        max_r = 0.0
        controlling_name = ""
        for node in nodes:
            if not node.structure_type:
                continue
            sv = resolve_node_effective_structure_type_text(node)
            if "倒虹吸" in sv:
                continue
            b = node.section_params.get("B", 0)
            rc = node.section_params.get("R_circle", 0)
            m_s = node.section_params.get("m", 0)
            wd = node.water_depth or 0
            min_r = 0.0
            basis = ""
            dim_str = ""
            if "隧洞" in sv:
                if rc > 0:
                    min_r = rc * 2 * 5
                    dim_str = f"洞径D={rc*2:.2f}m"
                    basis = f"R ≥ D×5 = {rc*2:.2f}×5 = {min_r:.1f}m"
                elif b > 0:
                    min_r = b * 5
                    dim_str = f"洞宽B={b:.2f}m"
                    basis = f"R ≥ B×5 = {b:.2f}×5 = {min_r:.1f}m"
            elif "明渠" in sv or sv == "矩形":
                if "U形" in sv and rc > 0:
                    theta_u_a = node.section_params.get('theta_deg', 0) or 0
                    if theta_u_a > 0 and wd > 0:
                        import math as _math_ua
                        h0_u_a = rc * (1.0 - _math_ua.cos(_math_ua.radians(theta_u_a / 2.0)))
                        if wd <= h0_u_a:
                            wsw_ua = 2 * _math_ua.sqrt(max(0.0, rc ** 2 - (rc - wd) ** 2))
                            basis_detail = f"水位在弧区 h={wd:.3f}m ≤ h_0={h0_u_a:.3f}m，B=2√(R²-(R-h)²)"
                        else:
                            b_arc_u_a = 2 * rc * _math_ua.sin(_math_ua.radians(theta_u_a / 2.0))
                            wsw_ua = b_arc_u_a + 2 * m_s * (wd - h0_u_a)
                            basis_detail = f"水位在直线段 h={wd:.3f}m > h_0={h0_u_a:.3f}m，B=b_arc+2m(h-h_0)"
                        min_r = wsw_ua * 5
                        dim_str = f"R={rc:.2f}m, θ={theta_u_a}°, m={m_s:.3f}, h={wd:.2f}m, 水面宽={wsw_ua:.3f}m"
                        basis = f"R ≥ 水面宽×5 = {wsw_ua:.3f}×5 = {min_r:.1f}m（{basis_detail}）"
                    else:
                        min_r = rc * 2 * 5
                        dim_str = f"R={rc:.2f}m（θ缺失，用D=2R近似）"
                        basis = f"R ≥ D×5 = {rc*2:.2f}×5 = {min_r:.1f}m"
                elif b > 0:
                    wsw = b + 2 * m_s * wd if m_s > 0 and wd > 0 else b
                    min_r = wsw * 5
                    if m_s > 0 and wd > 0:
                        dim_str = f"B={b:.2f}m, m={m_s}, h={wd:.2f}m, 水面宽={wsw:.2f}m"
                        basis = f"R ≥ 水面宽×5 = {wsw:.2f}×5 = {min_r:.1f}m"
                    else:
                        dim_str = f"B={b:.2f}m"
                        basis = f"R ≥ B×5 = {b:.2f}×5 = {min_r:.1f}m"
            elif "渡槽" in sv:
                if b > 0:
                    min_r = b * 5
                    dim_str = f"连接明渠底宽B={b:.2f}m"
                    basis = f"R ≥ B×5 = {b:.2f}×5 = {min_r:.1f}m"
            elif "暗涵" in sv:
                if b > 0:
                    min_r = b * 5
                    dim_str = f"涵宽B={b:.2f}m"
                    basis = f"R ≥ B×5 = {b:.2f}×5 = {min_r:.1f}m"
            if min_r > 0:
                name = getattr(node, 'name', '') or sv
                details.append((name, sv, dim_str, basis, min_r))
                if min_r > max_r:
                    max_r = min_r
                    controlling_name = name
        if max_r <= 0:
            max_r = DEFAULT_AUTO_TURN_RADIUS

        rec_r = _math.ceil(max_r)
        self.turn_radius_edit.setText(f"{rec_r:.1f}")

        from app_渠系计算前端.water_profile.water_profile_dialogs import TurnRadiusCalcDialog
        dlg = TurnRadiusCalcDialog(
            self, rec_r=rec_r, max_r=max_r,
            details=details, controlling_name=controlling_name
        )
        dlg.exec()

    def _open_transition_reference(self):
        """打开渐变段参考系数表对话框（表K.1.2 + 表L.1.2）"""
        dlg = TransitionReferenceDialog(self)
        dlg.exec()

    def _update_summary_panel(self, nodes, total_len=0.0, wl_drop=None, summary=None):
        """更新持久摘要面板信息"""
        if not nodes:
            self.lbl_summary_info.setText("尚未计算")
            self.btn_building_stats.setEnabled(False)
            return

        parts = [f"节点数: {len(nodes)}"]
        if total_len > 0:
            parts.append(f"总长度: {total_len:.1f}m")
        if wl_drop is not None:
            parts.append(f"水位落差: {wl_drop:.3f}m")

        # 起终点桩号和水位
        if summary:
            start_st = summary.get('起点桩号', None)
            end_st = summary.get('终点桩号', None)
            start_wl = summary.get('起点水位', None)
            end_wl = summary.get('终点水位', None)
            if start_st is not None:
                parts.append(f"桩号: {start_st:.3f}~{end_st:.3f}")
            if start_wl is not None:
                parts.append(f"水位: {start_wl:.3f}~{end_wl:.3f}")

        # 建筑物统计
        building_count = len(self._last_building_lengths)
        if building_count > 0:
            parts.append(f"建筑物: {building_count}段")

        self.lbl_summary_info.setText("    ".join(parts))
        self.btn_building_stats.setEnabled(len(self._last_building_lengths) > 0)

    def _show_building_length_dialog(self):
        """打开建筑物长度统计对话框"""
        if not self._last_building_lengths:
            InfoBar.info("提示", "暂无建筑物长度数据，请先执行计算",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return
        from app_渠系计算前端.water_profile.water_profile_dialogs import BuildingLengthDialog
        prefix = self._settings.get_station_prefix() if self._settings else ""
        dlg = BuildingLengthDialog(
            self,
            building_lengths=self._last_building_lengths,
            channel_total_length=self._last_channel_total_length,
            type_summary=self._last_type_summary,
            station_prefix=prefix
        )
        dlg.exec()

    def _parse_flow_values(self, flow_str):
        """解析流量字符串为浮点数列表，支持逗号分隔的多流量段"""
        return parse_flow_values_text(flow_str)

    def _on_design_flow_changed(self):
        """设计流量变化时自动计算加大流量"""
        design_flows = self._parse_flow_values(self.design_flow_edit.text())
        if not design_flows:
            self.max_flow_edit.setText("")
            self._sync_flow_segment_widgets(reset_index=False)
            return
        self.max_flow_edit.setText(
            format_flow_values_text(calculate_final_max_flow_values(design_flows))
        )
        self._sync_flow_segment_widgets(reset_index=False)

    def _insert_transitions(self):
        """插入渐变段"""
        if not self._ensure_downstream_ready("插入渐变段"):
            return
        if not CALCULATOR_AVAILABLE:
            InfoBar.error("不可用", "核心计算引擎未加载",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        if self.node_table.rowCount() < 2:
            InfoBar.warning("节点不足", "至少需要2个节点才能插入渐变段",
                           parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        try:
            # 检查流量参数
            design_flows = self._parse_flow_values(self.design_flow_edit.text())
            max_flows = self._parse_flow_values(self.max_flow_edit.text())
            if not design_flows or not max_flows or all(q <= 0 for q in design_flows):
                InfoBar.info("提示", "请先执行【断面批量计算】并完成自动同步后，再点击【插入渐变段】。",
                            parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
                return

            settings = self._build_settings()
            if not settings:
                return

            # 验证设置
            is_valid, error_msg = settings.validate()
            if not is_valid:
                InfoBar.error("参数错误", error_msg,
                             parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
                return

            nodes = self._build_nodes_from_table()
            if len(nodes) < 2:
                InfoBar.warning("数据不足", "至少需要2个节点",
                               parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 检查是否已经插入过渐变段或自动补段
            has_transitions = any(getattr(n, 'is_transition', False) for n in nodes)
            has_auto_channels = any(getattr(n, 'is_auto_inserted_channel', False) for n in nodes)
            if has_transitions or has_auto_channels:
                if not fluent_question(self, "提示",
                        "表格中已存在渐变段行。\n\n"
                        "是否清除已有渐变段并重新插入？\n"
                        "（选「否」则保留现有渐变段不做任何操作）"):
                    return
                # 同时清除渐变段行和自动插入的补段行，避免重复插入
                nodes = [n for n in nodes
                         if not getattr(n, 'is_transition', False)
                         and not getattr(n, 'is_auto_inserted_channel', False)]

            import copy
            calculator = WaterProfileCalculator(settings)

            # 与原版Tkinter一致：不调用calculate_geometry。
            # station_MC等几何参数已在导入时由_recalculate_geometry()计算完毕并写入表格，
            # _build_nodes_from_table()已从表格读取到正确的station_MC值。
            # 如果此处再调用calculate_geometry，会因为in_out已是INLET/OUTLET
            # 而跳过进/出节点转角计算，导致station_MC被覆盖为错误值。
            calculator.preprocess_nodes(nodes)

            # ===== 预扫描补段缺口 =====
            gaps = calculator.pre_scan_open_channels(nodes)

            # 批量处理状态
            batch_state = {
                'mode': 'manual',
                'current_index': 0,
                'total_count': len(gaps),
                'preset_params': {},
                'inserted_channels': []
            }

            # 若有多处（≥2）需要插入补段，弹出批量选择对话框
            from app_渠系计算前端.water_profile.water_profile_dialogs import (
                BatchChannelConfirmDialog, OpenChannelDialog, OpenChannelParams
            )
            if len(gaps) >= 2:
                batch_dlg = BatchChannelConfirmDialog(self, len(gaps), gaps)
                batch_dlg.exec()
                batch_result = batch_dlg.get_result()

                if batch_result['mode'] == BatchChannelConfirmDialog.RESULT_CANCELLED:
                    return
                elif batch_result['mode'] == BatchChannelConfirmDialog.RESULT_TABLE_EDIT:
                    batch_state['mode'] = 'table_edit'
                    batch_state['preset_params'] = batch_result['params']

            # 创建补段参数获取回调
            def open_channel_callback(reference_segment, available_length,
                                       prev_struct, next_struct, flow_section, flow):
                idx = batch_state['current_index']
                batch_state['current_index'] += 1

                def _track(params, source):
                    batch_state['inserted_channels'].append({
                        'gap_index': idx,
                        'prev_struct': prev_struct,
                        'next_struct': next_struct,
                        'available_length': available_length,
                        'params': params,
                        'source': source,
                    })

                # ① 表格编辑模式
                if batch_state['mode'] == 'table_edit' and idx in batch_state.get('preset_params', {}):
                    p = batch_state['preset_params'][idx]
                    _track(p, '表格编辑')
                    return p

                # ② 自动推荐模式
                if batch_state['mode'] == 'auto_recommend' and reference_segment:
                    p = calculator._build_open_channel_params_from_reference(
                        reference_segment,
                        flow_section,
                        flow,
                    )
                    _track(p, '推荐')
                    return p

                # ③ 手动模式：逐一弹窗
                dlg = OpenChannelDialog(
                    self,
                    upstream_channel=reference_segment,
                    available_length=available_length,
                    prev_structure=prev_struct,
                    next_structure=next_struct,
                    flow_section=flow_section,
                    flow=flow,
                    current_index=idx + 1,
                    total_count=batch_state['total_count']
                )
                if dlg.exec() == QDialog.DialogCode.Accepted:
                    result = dlg.get_result()
                    if result:
                        _track(result, '手动')
                    if dlg.apply_all_remaining:
                        batch_state['mode'] = 'auto_recommend'
                    return result
                return None

            # ===== 执行：预处理 + 插入渐变段 + 几何计算 =====
            prepared_nodes = calculator.prepare_transitions(nodes, open_channel_callback)

            # 更新表格显示（使用完整刷新，显示几何计算结果）
            prefix = settings.get_station_prefix() if settings else ""
            self._update_table_from_nodes_full(prepared_nodes, prefix)
            auto_resize_table(self.node_table)
            self.calculated_nodes = []
            self.nodes = list(prepared_nodes or [])
            self._transition_topology_prepared = True

            # 统计
            transition_count = sum(1 for n in prepared_nodes if getattr(n, 'is_transition', False))
            open_channel_count = len(batch_state.get('inserted_channels', []))
            original_count = len(prepared_nodes) - transition_count - open_channel_count

            # 统计建筑物长度（几何计算完成后即可统计）
            try:
                if len(prepared_nodes) >= 2 and getattr(prepared_nodes[-1], 'station_MC', 0):
                    building_lengths = calculator.calculate_building_lengths(prepared_nodes)
                    channel_total_length = prepared_nodes[-1].station_MC - prepared_nodes[0].station_MC
                    type_summary = calculator.calculate_comprehensive_type_summary(prepared_nodes)
                    self._last_building_lengths = building_lengths
                    self._last_channel_total_length = channel_total_length
                    self._last_type_summary = type_summary
                    self._update_summary_panel(prepared_nodes, channel_total_length)
            except Exception:
                pass  # 渐变段插入阶段统计失败不影响主流程

            # 检查是否有倒虹吸
            has_siphon = any(
                n.structure_type and "倒虹吸" in n.structure_type.value
                for n in prepared_nodes
                if n.structure_type and not getattr(n, 'is_transition', False)
            )
            has_pressure_pipe = any(
                self._is_pressure_pipe_like_node(n)
                for n in prepared_nodes
                if n.structure_type and not getattr(n, 'is_transition', False)
            )

            # 汇总信息（InfoBar非阻塞通知）
            if transition_count == 0 and open_channel_count == 0:
                summary = f"当前拓扑无需新增渐变段/补段，已完成拓扑检查：共 {len(prepared_nodes)} 行"
            else:
                summary = f"渐变段插入完成！共 {len(prepared_nodes)} 行（渐变段 {transition_count}，补段 {open_channel_count}）"
            if has_siphon or has_pressure_pipe:
                step_parts = []
                if has_siphon:
                    step_parts.append("【倒虹吸水力计算】")
                if has_pressure_pipe:
                    step_parts.append("【有压管道水力计算】")
                step_parts.append("【执行计算】")
                next_step = "下一步：请点击" + "→".join(step_parts)
            else:
                next_step = "下一步：请点击【执行计算】"
            self._show_transition_length_rules_post_insert_bar(summary, next_step)

        except Exception as e:
            import traceback
            traceback.print_exc()
            InfoBar.error("插入渐变段失败", str(e),
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _update_table_from_nodes(self, nodes):
        """从节点列表更新输入表格（插入渐变段后刷新）"""
        self._updating_cells = True
        try:
            self._update_table_from_nodes_inner(nodes)
        finally:
            self._updating_cells = False

    def _update_table_from_nodes_inner(self, nodes):
        self.node_table.setRowCount(0)
        for node in nodes:
            r = self.node_table.rowCount()
            self.node_table.insertRow(r)
            _is_trans = getattr(node, 'is_transition', False)
            _is_auto_ch = getattr(node, 'is_auto_inserted_channel', False)
            vals = [""] * len(NODE_ALL_HEADERS)
            # 基础输入列 (0-7)
            vals[0] = node.flow_section
            vals[1] = node.name
            _st_str = node.get_structure_type_str()
            vals[2] = f"{_st_str}(连接段)" if _is_auto_ch else _st_str
            if not _is_trans:
                vals[5] = f"{node.x:.6f}" if (node.x and not _is_auto_ch) else ""
                vals[6] = f"{node.y:.6f}" if (node.y and not _is_auto_ch) else ""
                vals[7] = self._format_node_turn_radius_display_text(node, r)
                # 水力输入列 (20-26)
                vals[20] = f"{node.section_params.get('B', '')}" if node.section_params.get('B') else ""
                vals[21] = f"{node.section_params.get('D', '')}" if node.section_params.get('D') else ""
                vals[22] = f"{node.section_params.get('R_circle', '')}" if node.section_params.get('R_circle') else ""
                vals[23] = f"{node.section_params.get('m', '')}" if node.section_params.get('m') else ""
                vals[24] = f"{node.roughness}" if node.roughness else ""
                vals[25] = f"{1.0/node.slope_i:.0f}" if node.slope_i and node.slope_i > 0 else ""
                vals[26] = f"{node.flow}" if node.flow else ""
            for c, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                item.setTextAlignment(Qt.AlignCenter)
                if c not in EDITABLE_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                # 第一行（水位起点）锁定水头损失列
                if r == 0 and c in FIRST_ROW_LOCKED_LOSS_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if c == 7 and (self._is_pressure_pipe_like_structure_text(_st_str) or getattr(node, 'is_pressure_pipe', False)):
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if _is_trans:
                    item.setForeground(QColor("#9E9E9E"))
                elif _is_auto_ch:
                    item.setForeground(QColor("#2E7D32"))
                    item.setToolTip("自动插入的明渠连接段，用于计算两个建筑物之间的沿程及弯道水头损失。\n几何列留空因为该行不是真实IP转折点。")
                self.node_table.setItem(r, c, item)
            first_item = self.node_table.item(r, 0)
            if first_item:
                payload = first_item.data(Qt.UserRole)
                if not isinstance(payload, dict):
                    payload = {}
                compound_trapezoid_params = {}
                if _st_str == "明渠-复式梯形":
                    compound_trapezoid_params = normalize_compound_trapezoid_params(
                        getattr(node, 'section_params', {}) or {}
                    )
                if compound_trapezoid_params:
                    payload[COMPOUND_TRAPEZOID_PARAMS_ROLE_KEY] = copy.deepcopy(compound_trapezoid_params)
                if _st_str == "隧洞-圆拱直墙型":
                    section_params = getattr(node, 'section_params', {}) or {}
                    if section_params.get('theta_deg'):
                        payload[TUNNEL_ARCH_THETA_ROLE_KEY] = section_params.get('theta_deg')
                    if section_params.get('H_straight') is not None:
                        payload[TUNNEL_ARCH_H_STRAIGHT_ROLE_KEY] = section_params.get('H_straight')
                    if section_params.get('manual_H_straight') is not None:
                        payload[TUNNEL_ARCH_MANUAL_H_STRAIGHT_ROLE_KEY] = section_params.get('manual_H_straight')
                    if section_params.get('used_manual_H_straight') is not None:
                        payload[TUNNEL_ARCH_USED_MANUAL_H_STRAIGHT_ROLE_KEY] = bool(section_params.get('used_manual_H_straight'))
                section_params = getattr(node, 'section_params', {}) or {}
                if normalize_culvert_family_type_name(section_params.get(CULVERT_FAMILY_TYPE_KEY, "")) == ARCH_CULVERT_FAMILY_TEXT:
                    if section_params.get('theta_deg'):
                        payload[ARCH_CULVERT_THETA_ROLE_KEY] = section_params.get('theta_deg')
                    if section_params.get('H_straight') is not None:
                        payload[ARCH_CULVERT_H_STRAIGHT_ROLE_KEY] = section_params.get('H_straight')
                    if section_params.get('manual_H_straight') is not None:
                        payload[ARCH_CULVERT_MANUAL_H_STRAIGHT_ROLE_KEY] = section_params.get('manual_H_straight')
                    if section_params.get('used_manual_H_straight') is not None:
                        payload[ARCH_CULVERT_USED_MANUAL_H_STRAIGHT_ROLE_KEY] = bool(section_params.get('used_manual_H_straight'))
                first_item.setData(Qt.UserRole, payload)
        auto_resize_table(self.node_table)

    def _open_siphon_calculator(self):
        """打开倒虹吸水力计算（PySide6 多标签页窗口）"""
        debug_print("[DEBUG] _open_siphon_calculator 被调用")
        if not self._ensure_downstream_ready("倒虹吸水力计算"):
            return
        if not CALCULATOR_AVAILABLE:
            debug_print("[DEBUG] CALCULATOR_AVAILABLE = False，返回")
            InfoBar.error("不可用", "核心计算引擎未加载",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        nodes = self._build_nodes_from_table()
        if not nodes:
            debug_print("[DEBUG] nodes 为空，返回")
            InfoBar.info("提示", "表格中没有数据，请先导入断面参数",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        # 检查是否已插入渐变段
        transition_topology_ready = self._has_transition_topology_ready(nodes)
        if not transition_topology_ready:
            debug_print("[DEBUG] transition_topology_ready = False，返回")
            InfoBar.warning("提示",
                           "请先点击工具栏的【插入渐变段】按钮，完成渐变段插入后再进行倒虹吸水力计算。\n"
                           "插入渐变段后，系统才能准确获取倒虹吸上下游流速、断面参数等信息。",
                           parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        # 检查是否有倒虹吸
        has_siphon = any(
            n.structure_type and "倒虹吸" in n.structure_type.value
            for n in nodes if n.structure_type
        )
        debug_print(f"[DEBUG] has_siphon = {has_siphon}")
        if not has_siphon:
            debug_print("[DEBUG] has_siphon = False，返回")
            InfoBar.info("提示", "表格中没有倒虹吸数据，请确保有结构形式为\"倒虹吸\"的行",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        missing_siphon_name_rows = self._collect_missing_siphon_name_rows(nodes)
        if missing_siphon_name_rows:
            InfoBar.warning(
                "提示",
                self._build_missing_siphon_name_notice(missing_siphon_name_rows),
                parent=self._info_parent(),
                duration=5000,
                position=InfoBarPosition.TOP,
            )
            return

        debug_print("[DEBUG] 开始导入模块和提取倒虹吸分组")
        try:
            from app_渠系计算前端.siphon.multi_siphon_dialog import MultiSiphonDialog

            # 提取倒虹吸分组
            _water_profile_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                '推求水面线'
            )
            import sys as _sys
            if _water_profile_dir not in _sys.path:
                _sys.path.insert(0, _water_profile_dir)
            from utils.siphon_extractor import SiphonDataExtractor

            settings = self._build_settings()
            siphon_groups = SiphonDataExtractor.extract_siphons(nodes, settings=settings)
            if not siphon_groups:
                InfoBar.info("提示", "未找到倒虹吸数据组",
                            parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 使用共享的 SiphonManager（已绑定项目路径）
            if self._siphon_manager is not None:
                manager = self._siphon_manager
            else:
                from managers.siphon_manager import SiphonManager
                manager = SiphonManager()

            # 定义导入回调：将水头损失和平面转弯半径写回节点表格
            _panel = self
            def import_losses_callback(results):
                cur_nodes = _panel._build_nodes_from_table()
                cur_groups = SiphonDataExtractor.extract_siphons(cur_nodes)
                imported_count = 0
                has_radius_update = False
                has_velocity_update = False
                for group in cur_groups:
                    if group.name in results and results[group.name] is not None:
                        result_data = results[group.name]
                        if isinstance(result_data, dict):
                            head_loss = result_data.get("head_loss", 0.0)
                            diameter = result_data.get("diameter", 0.0)
                            velocity = result_data.get("velocity")
                            turn_radius = result_data.get("turn_radius", 0.0)
                            turn_radius_overrode_excel = bool(result_data.get("turn_radius_overrode_excel", False))
                        else:
                            head_loss = result_data
                            diameter = 0.0
                            velocity = None
                            turn_radius = 0.0
                            turn_radius_overrode_excel = False
                        outlet_idx = group.outlet_row_index
                        if 0 <= outlet_idx < len(cur_nodes):
                            _nd = cur_nodes[outlet_idx]
                            _nd.head_loss_siphon = head_loss
                            _nd.head_loss_total = (
                                (_nd.head_loss_bend or 0.0)
                                + (_nd.head_loss_friction or 0.0)
                                + (getattr(_nd, 'head_loss_local', 0.0) or 0.0)
                                + (getattr(_nd, 'head_loss_reserve', 0.0) or 0.0)
                                + (getattr(_nd, 'head_loss_gate', 0.0) or 0.0)
                                + (float(head_loss) if head_loss else 0.0)
                            )
                            imported_count += 1
                        if diameter > 0:
                            for row_idx in group.row_indices:
                                if 0 <= row_idx < len(cur_nodes):
                                    if not hasattr(cur_nodes[row_idx], 'section_params') or not cur_nodes[row_idx].section_params:
                                        cur_nodes[row_idx].section_params = {}
                                    cur_nodes[row_idx].section_params["D"] = diameter
                        try:
                            velocity_value = float(velocity)
                        except (TypeError, ValueError):
                            velocity_value = None
                        if velocity_value is not None and velocity_value > 0:
                            for row_idx in group.row_indices:
                                if 0 <= row_idx < len(cur_nodes):
                                    cur_nodes[row_idx].velocity = velocity_value
                                    has_velocity_update = True
                        # 平面转弯半径只写回倒虹吸中间 IP 转弯行；Excel 显式值未确认覆盖时保留。
                        if turn_radius > 0:
                            for row_position, row_idx in enumerate(group.row_indices):
                                if 0 <= row_idx < len(cur_nodes):
                                    node = cur_nodes[row_idx]
                                    if not _panel._should_write_siphon_turn_radius_result(
                                        group,
                                        row_position=row_position,
                                        node=node,
                                        turn_radius_overrode_excel=turn_radius_overrode_excel,
                                    ):
                                        continue
                                    node.turn_radius = turn_radius
                                    node.turn_radius_is_explicit = True
                                    node.turn_radius_text = f"{float(turn_radius):.2f}"
                                    has_radius_update = True
                if imported_count > 0 or has_radius_update or has_velocity_update:
                    _panel._append_loss_undo_snapshot(_panel._snapshot_editable_cols())
                    _s = _panel._build_settings()
                    _pfx = _s.get_station_prefix() if _s else ""
                    _panel._update_table_from_nodes_full(cur_nodes, _pfx)
                    auto_resize_table(_panel.node_table)
                    _panel._recalculate_silent()
                return imported_count

            siphon_n = DEFAULT_SIPHON_TURN_RADIUS_N

            # 打开PySide6多标签页倒虹吸计算窗口
            debug_print(f"[DEBUG] 正在创建 MultiSiphonDialog，倒虹吸组数量: {len(siphon_groups)}")
            dlg = MultiSiphonDialog(
                self._info_parent(),
                siphon_groups,
                manager=manager,
                on_import_losses=import_losses_callback,
                siphon_turn_radius_n=siphon_n,
                show_case_management=False
            )
            debug_print("[DEBUG] MultiSiphonDialog 创建完成，准备调用 exec()")
            result = dlg.exec()
            debug_print(f"[DEBUG] dlg.exec() 返回值: {result}")

        except ImportError as e:
            import traceback
            traceback.print_exc()
            InfoBar.warning("提示", f"倒虹吸水力计算模块加载失败: {str(e)}",
                           parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
        except Exception as e:
            import traceback
            traceback.print_exc()
            InfoBar.error("错误", f"打开倒虹吸计算窗口失败: {str(e)}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    @staticmethod
    def _get_pressure_pipe_group_flow_section(group) -> str:
        rows = getattr(group, "rows", None) or []
        for node in rows:
            fs = str(getattr(node, "flow_section", "") or "").strip()
            if fs:
                return fs
        return "-"

    @classmethod
    def _build_pressure_pipe_group_identity(cls, group) -> str:
        identity = str(getattr(group, "identity", "") or "").strip()
        if identity:
            return identity
        return make_pressure_pipe_identity(
            cls._get_pressure_pipe_group_flow_section(group),
            getattr(group, "name", "") or ""
        )

    @staticmethod
    def _is_valid_pressure_pipe_total_head_loss(value) -> bool:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return False
        return math.isfinite(number) and number >= 0

    @staticmethod
    def _normalize_pressure_pipe_export_number(value, *, allow_zero=False) -> float | None:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None
        if not math.isfinite(number):
            return None
        if allow_zero:
            return number if number >= 0 else None
        return number if number > 0 else None

    @classmethod
    def _collect_pressure_pipe_export_target_identity_aliases(
        cls,
        name,
        flow_section,
        identity: str = "",
        metadata=None,
    ) -> list[str]:
        """收集导出目标可识别的承压身份别名，优先使用真实行身份。"""
        aliases = []
        seen = set()

        def _append(value) -> None:
            text = str(value or "").strip()
            if not text or text in seen:
                return
            seen.add(text)
            aliases.append(text)

        if isinstance(metadata, dict):
            _append(metadata.get("pressure_pipe_row_identity", ""))
            _append(metadata.get("row_identity", ""))
            _append(metadata.get(PRESSURE_PIPE_ROW_ID_ROLE_KEY, ""))
            for alias in list(metadata.get("identity_aliases", []) or []):
                _append(alias)

        _append(identity)
        _append(make_pressure_pipe_identity(str(flow_section or "").strip(), str(name or "").strip() or "未命名"))
        return aliases

    @staticmethod
    def _get_pressure_pipe_export_target_identities(target: dict) -> list[str]:
        aliases = []
        seen = set()
        for value in list(target.get("identity_aliases", []) or []):
            text = str(value or "").strip()
            if not text or text in seen:
                continue
            seen.add(text)
            aliases.append(text)
        identity = str(target.get("identity", "") or "").strip()
        if identity and identity not in seen:
            aliases.append(identity)
        return aliases

    @classmethod
    def _make_pressure_pipe_export_target(cls, name, flow_section, identity: str = "", metadata=None) -> dict:
        flow_section_text = str(flow_section or "").strip()
        name_text = str(name or "").strip() or "未命名"
        identity_aliases = cls._collect_pressure_pipe_export_target_identity_aliases(
            name_text,
            flow_section_text,
            identity,
            metadata=metadata,
        )
        identity_text = identity_aliases[0] if identity_aliases else make_pressure_pipe_identity(flow_section_text, name_text)
        target = {
            "identity": identity_text,
            "identity_aliases": identity_aliases or [identity_text],
            "flow_section": flow_section_text,
            "name": name_text,
        }
        if isinstance(metadata, dict):
            for key in (
                "segment_start_mc",
                "segment_end_mc",
                "route_start_mc",
                "route_end_mc",
                "target_row_index",
                "upstream_row_index",
                "route_key",
                "route_display_name",
                "station_mc",
                "station_text",
                "node_label",
                "is_tunnel",
            ):
                if key in metadata:
                    target[key] = copy.deepcopy(metadata.get(key))
        return target

    def _collect_pressure_pipe_export_targets(self, rows=None) -> list:
        targets = []
        if rows is not None:
            for row in rows or []:
                if not isinstance(row, dict):
                    continue
                targets.append(
                    self._make_pressure_pipe_export_target(
                        row.get("name", ""),
                        row.get("flow_section", ""),
                        row.get("identity", ""),
                        metadata=row,
                    )
                )
            return targets

        try:
            settings = self._build_settings()
            cur_nodes = self._build_nodes_from_table()
            cur_groups = self._extract_pressure_pipe_dialog_groups(cur_nodes, settings=settings)
        except Exception:
            return targets

        for group in cur_groups or []:
            metadata = {
                "segment_start_mc": getattr(group, "segment_start_mc", None),
                "segment_end_mc": getattr(group, "segment_end_mc", None),
                "route_start_mc": getattr(group, "route_start_mc", None),
                "route_end_mc": getattr(group, "route_end_mc", None),
                "target_row_index": getattr(group, "target_row_index", -1),
                "upstream_row_index": getattr(group, "upstream_row_index", -1),
                "route_key": getattr(group, "route_key", ""),
                "route_display_name": getattr(group, "route_display_name", ""),
                "identity_aliases": self._collect_pressure_pipe_group_export_identity_aliases(group),
            }
            targets.append(
                self._make_pressure_pipe_export_target(
                    self._get_pressure_pipe_group_display_name(group),
                    self._get_pressure_pipe_group_flow_section(group),
                    self._build_pressure_pipe_group_identity(group),
                    metadata=metadata,
                )
            )
        return targets

    def _index_pressure_pipe_calc_records_for_export(self) -> tuple:
        exact = {}
        plain_name_candidates = {}
        data = normalize_pressure_pipe_calc_records(getattr(self, "_pressure_pipe_calc_records", None))
        for record in data.get("records", []):
            if str(record.get("status", "")).strip().lower() != "success":
                continue
            total_head_loss = self._normalize_pressure_pipe_export_number(
                record.get("total_head_loss"), allow_zero=True
            )
            pipe_velocity = self._normalize_pressure_pipe_export_number(record.get("pipe_velocity"))
            total_length = self._normalize_pressure_pipe_export_number(record.get("total_length"))
            if total_head_loss is None and pipe_velocity is None and total_length is None:
                continue
            name = str(record.get("name", "") or "").strip() or "未命名"
            flow_section = str(record.get("flow_section", "") or "").strip()
            identity = str(record.get("identity", "") or "").strip()
            if not identity:
                identity = make_pressure_pipe_identity(flow_section, name)
            payload = {
                "identity": identity,
                "flow_section": flow_section,
                "name": name,
                "source": "calc_records",
            }
            if total_head_loss is not None:
                payload["total_head_loss"] = float(total_head_loss)
            if pipe_velocity is not None:
                payload["pipe_velocity"] = float(pipe_velocity)
            if total_length is not None:
                payload["total_length"] = float(total_length)
            exact[identity] = payload
            plain_name_candidates.setdefault(name, []).append(payload)

        plain_name = {}
        for name, items in plain_name_candidates.items():
            if len({item["identity"] for item in items}) == 1:
                plain_name[name] = items[-1]
        return exact, plain_name

    def _index_pressure_pipe_table_results_for_export(self, target_identities=None) -> tuple:
        exact = {}
        plain_name_candidates = {}
        try:
            from utils.pressure_pipe_extractor import PressurePipeDataExtractor

            settings = self._build_settings()
            cur_nodes = self._build_nodes_from_table()
            cur_groups = PressurePipeDataExtractor.extract_pipes(cur_nodes, settings=settings)
        except Exception:
            return exact, {}

        target_identity_set = set(target_identities or [])
        for group in cur_groups or []:
            identity = self._build_pressure_pipe_group_identity(group)
            identity_aliases = self._collect_pressure_pipe_group_export_identity_aliases(group)
            if not identity_aliases:
                identity_aliases = [identity]
            matched_aliases = (
                [alias for alias in identity_aliases if alias in target_identity_set]
                if target_identity_set
                else list(identity_aliases)
            )
            if target_identity_set and not matched_aliases:
                continue

            outlet_idx_raw = getattr(group, "outlet_row_index", -1)
            try:
                outlet_idx = int(outlet_idx_raw)
            except (TypeError, ValueError):
                outlet_idx = -1
            if outlet_idx < 0 or outlet_idx >= len(cur_nodes):
                continue

            outlet_node = cur_nodes[outlet_idx]
            total_head_loss = getattr(outlet_node, "head_loss_siphon", None)
            if (
                not self._is_valid_pressure_pipe_total_head_loss(total_head_loss)
                or float(total_head_loss) <= 0
            ):
                total_head_loss = getattr(outlet_node, "external_head_loss", None)
            total_head_loss = self._normalize_pressure_pipe_export_number(total_head_loss, allow_zero=True)
            pipe_velocity = self._normalize_pressure_pipe_export_number(getattr(outlet_node, "velocity", None))
            total_length = self._normalize_pressure_pipe_export_number(
                getattr(group, "plan_total_length", None)
            )
            if total_head_loss is None and pipe_velocity is None and total_length is None:
                continue

            flow_section = self._get_pressure_pipe_group_flow_section(group)
            name = str(getattr(group, "name", "") or "").strip() or "未命名"
            payload = {
                "identity": identity,
                "flow_section": flow_section,
                "name": name,
                "source": "table3",
            }
            if total_head_loss is not None:
                payload["total_head_loss"] = float(total_head_loss)
            if pipe_velocity is not None:
                payload["pipe_velocity"] = float(pipe_velocity)
            if total_length is not None:
                payload["total_length"] = float(total_length)
            for alias in matched_aliases:
                alias_payload = dict(payload)
                alias_payload["identity"] = alias
                exact[alias] = alias_payload
            plain_name_candidates.setdefault(name, []).append(payload)

        plain_name = {}
        for name, items in plain_name_candidates.items():
            if len({item["identity"] for item in items}) == 1:
                plain_name[name] = items[-1]
        return exact, plain_name

    def _index_pressure_pipe_manager_results_for_export(self, target_identities=None) -> tuple:
        exact = {}
        plain_name_candidates = {}
        manager = getattr(self, "_pressure_pipe_manager", None)
        to_dict = getattr(manager, "to_dict", None)
        if not callable(to_dict):
            return exact, {}

        try:
            raw = to_dict() or {}
        except Exception:
            return exact, {}
        segments = raw.get("segments", {}) if isinstance(raw, dict) else {}
        identity_set = set(target_identities or [])
        for key, segment_data in segments.items():
            row = segment_data if isinstance(segment_data, dict) else {}
            total_head_loss = self._normalize_pressure_pipe_export_number(
                row.get("total_loss", row.get("total_head_loss")),
                allow_zero=True,
            )
            pipe_velocity = self._normalize_pressure_pipe_export_number(row.get("pipe_velocity"))
            total_length = self._normalize_pressure_pipe_export_number(
                row.get("plan_total_length", row.get("total_length"))
            )
            if total_head_loss is None and pipe_velocity is None and total_length is None:
                continue

            resolved_identity = str(row.get("identity", "") or key or "").strip()
            if identity_set and resolved_identity not in identity_set:
                continue
            name = str(
                row.get("member_display_name", "")
                or row.get("base_name", "")
                or row.get("dxf_display_name", "")
                or row.get("name", "")
                or resolved_identity
            ).strip() or "未命名"
            payload = {
                "identity": resolved_identity,
                "flow_section": str(row.get("flow_section", "") or "").strip(),
                "name": name,
                "source": "manager_segment",
            }
            if total_head_loss is not None:
                payload["total_head_loss"] = float(total_head_loss)
            if pipe_velocity is not None:
                payload["pipe_velocity"] = float(pipe_velocity)
            if total_length is not None:
                payload["total_length"] = float(total_length)
            exact[resolved_identity] = payload
            plain_name_candidates.setdefault(name, []).append(payload)
        pipes = raw.get("pipes", {}) if isinstance(raw, dict) else {}

        for key, pipe_data in pipes.items():
            row = pipe_data if isinstance(pipe_data, dict) else {}
            total_head_loss = self._normalize_pressure_pipe_export_number(
                row.get("total_head_loss"), allow_zero=True
            )
            pipe_velocity = self._normalize_pressure_pipe_export_number(row.get("pipe_velocity"))
            total_length = self._normalize_pressure_pipe_export_number(
                row.get("plan_total_length", row.get("total_length"))
            )
            if total_head_loss is None and pipe_velocity is None and total_length is None:
                continue

            key_text = str(key or "").strip()
            name = str(row.get("name", "") or "").strip()
            flow_section = str(row.get("flow_section", "") or "").strip()

            candidate_identities = []
            if key_text:
                candidate_identities.append(key_text)
            if key_text and "::" in key_text:
                if not name:
                    name = key_text.split("::", 1)[1].strip()
            if flow_section or name:
                candidate_identities.append(make_pressure_pipe_identity(flow_section, name or key_text))

            resolved_identity = ""
            for candidate in candidate_identities:
                if candidate and candidate in identity_set:
                    resolved_identity = candidate
                    break
            if not resolved_identity and candidate_identities and candidate_identities[0].count("::") > 0:
                resolved_identity = candidate_identities[0]

            if not name:
                if key_text and "::" in key_text:
                    name = key_text.split("::", 1)[1].strip() or "未命名"
                else:
                    name = key_text or "未命名"

            payload = {
                "identity": resolved_identity or "",
                "flow_section": flow_section,
                "name": name,
                "source": "manager",
            }
            if total_head_loss is not None:
                payload["total_head_loss"] = float(total_head_loss)
            if pipe_velocity is not None:
                payload["pipe_velocity"] = float(pipe_velocity)
            if total_length is not None:
                payload["total_length"] = float(total_length)
            if resolved_identity:
                exact[resolved_identity] = payload
            plain_name_candidates.setdefault(name, []).append(payload)

        plain_name = {}
        for name, items in plain_name_candidates.items():
            identities = {item.get("identity", "") for item in items}
            identities.discard("")
            if len(identities) <= 1:
                plain_name[name] = items[-1]
        return exact, plain_name

    @staticmethod
    def _has_exportable_pressure_pipe_longitudinal_nodes(value) -> bool:
        # 单点只够表示边界占位，中心线高程采样至少要两个纵断面点。
        return isinstance(value, list) and len(value) >= 2

    @classmethod
    def _pressure_pipe_longitudinal_nodes_cover_station(cls, longitudinal_nodes, station_mc, tol: float = 1e-3) -> bool:
        """判断当前纵断面是否覆盖指定桩号。"""
        target_station = cls._coerce_pressure_pipe_finite_float(station_mc)
        if target_station is None:
            return cls._has_exportable_pressure_pipe_longitudinal_nodes(longitudinal_nodes)
        if not cls._has_exportable_pressure_pipe_longitudinal_nodes(longitudinal_nodes):
            return False

        chainages = []
        for item in list(longitudinal_nodes or []):
            if not isinstance(item, dict):
                continue
            chainage = cls._coerce_pressure_pipe_finite_float(item.get("chainage"))
            if chainage is None:
                continue
            chainages.append(chainage)
        if len(chainages) < 2:
            return False
        start_mc = min(chainages[0], chainages[-1])
        end_mc = max(chainages[0], chainages[-1])
        return (start_mc - tol) <= target_station <= (end_mc + tol)

    def _index_pressure_pipe_manager_route_longitudinal_nodes(self) -> dict:
        """按 route_key 收集整线纵断面，供导出时做 route 级兜底。"""
        manager = getattr(self, "_pressure_pipe_manager", None)
        to_dict = getattr(manager, "to_dict", None)
        if not callable(to_dict):
            return {}

        try:
            raw = to_dict() or {}
        except Exception:
            return {}

        routes = raw.get("routes", {}) if isinstance(raw, dict) else {}
        resolved = {}
        for route_key, route_data in routes.items():
            route_key_text = str(route_key or "").strip()
            if not route_key_text or not isinstance(route_data, dict):
                continue
            longitudinal_nodes = copy.deepcopy(route_data.get("longitudinal_nodes", []) or [])
            if not self._has_exportable_pressure_pipe_longitudinal_nodes(longitudinal_nodes):
                continue
            resolved[route_key_text] = {
                "route_key": route_key_text,
                "route_display_name": str(route_data.get("display_name", "") or route_key_text).strip(),
                "longitudinal_nodes": longitudinal_nodes,
            }
        return resolved

    def _index_pressure_pipe_manager_route_raw_profile_polylines(self) -> dict:
        """按 route_key 收集导入原线几何。"""
        manager = getattr(self, "_pressure_pipe_manager", None)
        to_dict = getattr(manager, "to_dict", None)
        if not callable(to_dict):
            return {}

        try:
            raw = to_dict() or {}
        except Exception:
            return {}

        routes = raw.get("routes", {}) if isinstance(raw, dict) else {}
        resolved = {}
        for route_key, route_data in routes.items():
            route_key_text = str(route_key or "").strip()
            if not route_key_text or not isinstance(route_data, dict):
                continue
            raw_profile_polyline = copy.deepcopy(route_data.get("raw_profile_polyline", {}) or {})
            if not raw_profile_polyline:
                continue
            resolved[route_key_text] = {
                "route_key": route_key_text,
                "route_display_name": str(route_data.get("display_name", "") or route_key_text).strip(),
                "raw_profile_polyline": raw_profile_polyline,
            }
        return resolved

    def _index_pressure_pipe_manager_longitudinal_nodes_for_export(self, target_identities=None) -> tuple:
        exact = {}
        plain_name_candidates = {}
        manager = getattr(self, "_pressure_pipe_manager", None)
        get_pipe_config = getattr(manager, "get_pipe_config", None)
        to_dict = getattr(manager, "to_dict", None)
        identity_set = set(target_identities or [])
        raw_manager = {}
        route_buckets = {}

        if callable(to_dict):
            try:
                raw_manager = to_dict() or {}
            except Exception:
                raw_manager = {}
            route_buckets = raw_manager.get("routes", {}) if isinstance(raw_manager, dict) else {}
            segments = raw_manager.get("segments", {}) if isinstance(raw_manager, dict) else {}
            for key, segment_data in segments.items():
                row = segment_data if isinstance(segment_data, dict) else {}
                longitudinal_nodes = copy.deepcopy(row.get("longitudinal_nodes", []) or [])
                if not self._has_exportable_pressure_pipe_longitudinal_nodes(longitudinal_nodes):
                    continue
                resolved_identity = str(row.get("identity", "") or key or "").strip()
                if identity_set and resolved_identity not in identity_set:
                    continue
                name = str(
                    row.get("member_display_name", "")
                    or row.get("base_name", "")
                    or row.get("dxf_display_name", "")
                    or row.get("name", "")
                    or resolved_identity
                ).strip() or "未命名"
                payload = {
                    "identity": resolved_identity,
                    "flow_section": str(row.get("flow_section", "") or "").strip(),
                    "name": name,
                    "longitudinal_nodes": longitudinal_nodes,
                    "route_key": str(row.get("route_key", "") or "").strip(),
                    "route_display_name": str(row.get("route_display_name", "") or "").strip(),
                }
                exact[resolved_identity] = payload
                plain_name_candidates.setdefault(name, []).append(payload)

        if callable(get_pipe_config):
            try:
                settings = self._build_settings()
                cur_nodes = self._build_nodes_from_table()
                cur_groups = self._extract_pressure_pipe_dialog_groups(cur_nodes, settings=settings)
            except Exception:
                cur_groups = []

            for group in cur_groups or []:
                identity = self._build_pressure_pipe_group_identity(group)
                identity_aliases = self._collect_pressure_pipe_group_export_identity_aliases(group)
                matched_aliases = (
                    [alias for alias in identity_aliases if alias in identity_set]
                    if identity_set
                    else [identity]
                )
                if identity_set and not matched_aliases:
                    continue
                storage_key = self._get_pressure_pipe_group_storage_key(group)
                config = get_pipe_config(storage_key)
                longitudinal_nodes = []
                route_longitudinal_nodes = []
                route_profile_segments = []
                if config is not None:
                    longitudinal_nodes = copy.deepcopy(getattr(config, "longitudinal_nodes", []) or [])
                    route_profile_segments = copy.deepcopy(getattr(config, "profile_segments", []) or [])

                route_key = self._get_pressure_pipe_group_route_key(group)
                if route_key:
                    route_data = route_buckets.get(route_key, {}) if isinstance(route_buckets, dict) else {}
                    if isinstance(route_data, dict):
                        if not route_profile_segments:
                            route_profile_segments = copy.deepcopy(route_data.get("profile_segments", []) or [])
                        route_longitudinal_nodes = copy.deepcopy(route_data.get("longitudinal_nodes", []) or [])
                if route_profile_segments:
                    matched_segment = self._find_pressure_pipe_profile_segment_for_group(
                        group,
                        route_profile_segments,
                        identity=identity,
                        storage_key=storage_key,
                    )
                    if matched_segment is not None:
                        matched_longitudinal_nodes = copy.deepcopy(
                            matched_segment.get("longitudinal_nodes", []) or []
                        )
                        source_kind = str(matched_segment.get("source_kind", "") or "").strip()
                        # 普通子段若只剩单点缓存，优先回退整线 DXF；隧洞生成段继续沿用自己的分段结果。
                        if self._has_exportable_pressure_pipe_longitudinal_nodes(matched_longitudinal_nodes):
                            longitudinal_nodes = matched_longitudinal_nodes
                        elif (
                            source_kind != "generated_tunnel"
                            and self._has_exportable_pressure_pipe_longitudinal_nodes(route_longitudinal_nodes)
                        ):
                            longitudinal_nodes = copy.deepcopy(route_longitudinal_nodes)
                        else:
                            longitudinal_nodes = matched_longitudinal_nodes
                elif not self._has_exportable_pressure_pipe_longitudinal_nodes(longitudinal_nodes) and route_key:
                    longitudinal_nodes = copy.deepcopy(route_longitudinal_nodes)
                if not self._has_exportable_pressure_pipe_longitudinal_nodes(longitudinal_nodes):
                    continue

                segment_start = getattr(group, "segment_start_mc", None)
                segment_end = getattr(group, "segment_end_mc", None)
                is_route_anchor_group = self._is_pressure_pipe_route_anchor_group(group)
                if (
                    is_route_anchor_group
                    and self._has_exportable_pressure_pipe_longitudinal_nodes(route_longitudinal_nodes)
                    and not route_profile_segments
                ):
                    # xx管整线起点只对应一个桩号点，导出仍需整线纵断面来取起点高程。
                    longitudinal_nodes = copy.deepcopy(route_longitudinal_nodes)
                elif (
                    self._is_pressure_pipe_cross_flow_boundary_group(group, cur_nodes)
                    and self._has_exportable_pressure_pipe_longitudinal_nodes(route_longitudinal_nodes)
                    and not route_profile_segments
                ):
                    # 连续承压整线跨流量段时，新流量段首个普通行会退化成单点边界；
                    # 导出仍要继承整线 DXF，不能再按单点范围裁切。
                    longitudinal_nodes = copy.deepcopy(route_longitudinal_nodes)
                elif (
                    route_key
                    and segment_start is not None
                    and segment_end is not None
                    and not route_profile_segments
                ):
                    try:
                        from utils.pressure_pipe_longitudinal_utils import clip_longitudinal_nodes_to_range

                        longitudinal_nodes = clip_longitudinal_nodes_to_range(
                            longitudinal_nodes,
                            float(segment_start),
                            float(segment_end),
                        )
                    except ValueError:
                        continue
                    except Exception:
                        continue

                name = self._get_pressure_pipe_group_display_name(group)
                payload_base = {
                    "flow_section": self._get_pressure_pipe_group_flow_section(group),
                    "name": name,
                    "longitudinal_nodes": longitudinal_nodes,
                    "route_key": route_key,
                    "route_display_name": str(getattr(group, "route_display_name", "") or "").strip(),
                }
                if not matched_aliases:
                    matched_aliases = [identity]
                for alias in matched_aliases:
                    payload = dict(payload_base)
                    payload["identity"] = alias
                    exact[alias] = payload
                primary_payload = dict(payload_base)
                primary_payload["identity"] = identity
                plain_name_candidates.setdefault(name, []).append(primary_payload)

            if exact or plain_name_candidates:
                plain_name = {}
                for name, items in plain_name_candidates.items():
                    identities = {item.get("identity", "") for item in items}
                    identities.discard("")
                    if len(identities) <= 1:
                        plain_name[name] = items[-1]
                return exact, plain_name

        to_dict = getattr(manager, "to_dict", None)
        if not callable(to_dict):
            return exact, {}

        raw = raw_manager if isinstance(raw_manager, dict) and raw_manager else {}
        if not raw:
            try:
                raw = to_dict() or {}
            except Exception:
                return exact, {}
        pipes = raw.get("pipes", {}) if isinstance(raw, dict) else {}
        for key, pipe_data in pipes.items():
            row = pipe_data if isinstance(pipe_data, dict) else {}
            longitudinal_nodes = row.get("longitudinal_nodes")
            if not self._has_exportable_pressure_pipe_longitudinal_nodes(longitudinal_nodes):
                continue

            key_text = str(key or "").strip()
            name = str(row.get("name", "") or "").strip()
            flow_section = str(row.get("flow_section", "") or "").strip()

            candidate_identities = []
            if key_text:
                candidate_identities.append(key_text)
            if key_text and "::" in key_text:
                if not name:
                    name = key_text.split("::", 1)[1].strip()
            if flow_section or name:
                candidate_identities.append(make_pressure_pipe_identity(flow_section, name or key_text))

            resolved_identity = ""
            for candidate in candidate_identities:
                if candidate and candidate in identity_set:
                    resolved_identity = candidate
                    break
            if not resolved_identity and candidate_identities and candidate_identities[0].count("::") > 0:
                resolved_identity = candidate_identities[0]

            if not name:
                if key_text and "::" in key_text:
                    name = key_text.split("::", 1)[1].strip() or "未命名"
                else:
                    name = key_text or "未命名"

            payload = {
                "identity": resolved_identity or "",
                "flow_section": flow_section,
                "name": name,
                "longitudinal_nodes": copy.deepcopy(longitudinal_nodes),
            }
            if resolved_identity:
                exact[resolved_identity] = payload
            plain_name_candidates.setdefault(name, []).append(payload)

        plain_name = {}
        for name, items in plain_name_candidates.items():
            identities = {item.get("identity", "") for item in items}
            identities.discard("")
            if len(identities) <= 1:
                plain_name[name] = items[-1]
        return exact, plain_name

    @staticmethod
    def _normalize_pressure_pipe_summary_flow_section(flow_section) -> str:
        text = str(flow_section or "").strip()
        if not text or text == "-":
            return ""
        try:
            number = float(text)
        except (TypeError, ValueError):
            return text
        if not math.isfinite(number):
            return text
        if abs(number - round(number)) <= 1e-9:
            return str(int(round(number)))
        return text

    @staticmethod
    def _get_pressure_pipe_summary_structure_type(node) -> str:
        return resolve_node_effective_structure_type_text(node)

    @classmethod
    def _classify_pressure_pipe_summary_bucket(cls, node) -> str | None:
        if getattr(node, "is_transition", False) or getattr(node, "is_auto_inserted_channel", False):
            return None
        structure_text = cls._get_pressure_pipe_summary_structure_type(node)
        if "暗涵" in structure_text:
            return None
        if "隧洞" in structure_text:
            return "隧洞"
        if "定向钻" in structure_text:
            return "定向钻"
        if "顶管" in structure_text:
            return "顶管"
        if cls._is_pressure_pipe_like_structure_text(structure_text) or getattr(node, "is_pressure_pipe", False):
            return "有压管道"
        return None

    def _iter_pressure_pipe_continuous_flow_boundaries(self, nodes) -> list[dict]:
        """识别相邻承压行跨流量段时共用的切段水位。"""
        if not isinstance(nodes, list) or len(nodes) < 2:
            return []
        pressure_gate_buckets = {"有压管道", "定向钻", "顶管"}
        boundaries = []
        for idx in range(len(nodes) - 1):
            current = nodes[idx]
            nxt = nodes[idx + 1]
            current_flow_section_text = self._normalize_pressure_pipe_summary_flow_section(
                getattr(current, "flow_section", "")
            )
            next_flow_section_text = self._normalize_pressure_pipe_summary_flow_section(
                getattr(nxt, "flow_section", "")
            )
            if (
                not current_flow_section_text
                or not next_flow_section_text
                or current_flow_section_text == next_flow_section_text
            ):
                continue
            current_bucket = self._classify_pressure_pipe_summary_bucket(current)
            next_bucket = self._classify_pressure_pipe_summary_bucket(nxt)
            if current_bucket not in pressure_gate_buckets or next_bucket not in pressure_gate_buckets:
                continue
            boundary_water_level = self._normalize_pressure_pipe_export_number(
                getattr(nxt, "water_level", None),
                allow_zero=True,
            )
            if boundary_water_level is None:
                continue
            boundaries.append(
                {
                    "current_flow_section": current_flow_section_text,
                    "next_flow_section": next_flow_section_text,
                    "boundary_water_level": boundary_water_level,
                    "current_index": idx,
                    "next_index": idx + 1,
                }
            )
        return boundaries

    @staticmethod
    def _get_pressure_pipe_summary_in_out_text(node) -> str:
        """读取摘要统计需要的进出口标记。"""
        getter = getattr(node, "get_in_out_str", None)
        if callable(getter):
            try:
                text = str(getter() or "").strip()
                if text and text != "IP":
                    return text
            except Exception:
                pass
        section_params = getattr(node, "section_params", None)
        if isinstance(section_params, dict):
            text = str(section_params.get("in_out_raw", "") or "").strip()
            if text and text != "IP":
                return text
        in_out = getattr(node, "in_out", None)
        value = getattr(in_out, "value", in_out)
        text = str(value or "").strip()
        return "" if text == "IP" else text

    def _get_pressure_pipe_summary_source_nodes(self, nodes=None) -> list:
        """统一压力管道摘要取值的节点来源。"""
        if isinstance(nodes, list) and nodes:
            return nodes

        build_nodes = getattr(self, "_build_nodes_from_table", None)
        if callable(build_nodes):
            try:
                current_nodes = build_nodes()
            except Exception:
                current_nodes = None
            if isinstance(current_nodes, list) and current_nodes:
                return current_nodes

        for attr_name in ("nodes", "calculated_nodes"):
            source_nodes = getattr(self, attr_name, None)
            if isinstance(source_nodes, list) and source_nodes:
                return source_nodes
        return []

    def _get_pressure_pipe_summary_waterline(self, nodes) -> tuple[float | None, float | None]:
        start_water_level = None
        end_water_level = None
        if nodes and CALCULATOR_AVAILABLE:
            try:
                calculator = WaterProfileCalculator(self._build_settings())
                summary = calculator.get_calculation_summary(nodes)
            except Exception:
                summary = {}
            if isinstance(summary, dict):
                start_water_level = self._normalize_pressure_pipe_export_number(
                    summary.get("起点水位"), allow_zero=True
                )
                end_water_level = self._normalize_pressure_pipe_export_number(
                    summary.get("终点水位"), allow_zero=True
                )
        if start_water_level is None and nodes:
            start_water_level = self._normalize_pressure_pipe_export_number(
                getattr(nodes[0], "water_level", None), allow_zero=True
            )
        if end_water_level is None and nodes:
            end_water_level = self._normalize_pressure_pipe_export_number(
                getattr(nodes[-1], "water_level", None), allow_zero=True
            )
        return start_water_level, end_water_level

    @classmethod
    def _resolve_pressure_pipe_target_station_range(cls, target) -> tuple[float | None, float | None]:
        """优先读取有压管道子段边界桩号，旧数据再回退整线范围。"""
        if not isinstance(target, dict):
            return None, None
        start_station = cls._coerce_pressure_pipe_finite_float(target.get("segment_start_mc"))
        end_station = cls._coerce_pressure_pipe_finite_float(target.get("segment_end_mc"))
        if start_station is None or end_station is None:
            start_station = cls._coerce_pressure_pipe_finite_float(target.get("route_start_mc"))
            end_station = cls._coerce_pressure_pipe_finite_float(target.get("route_end_mc"))
        if start_station is None or end_station is None or end_station < start_station:
            return None, None
        return float(start_station), float(end_station)

    def _resolve_pressure_pipe_target_boundary_water_levels(self, nodes, target) -> tuple[float | None, float | None]:
        """按有压管道起终点行读取导出摘要所需的边界水位。"""
        if not isinstance(nodes, list) or not nodes or not isinstance(target, dict):
            return None, None
        start_water_level = None
        end_water_level = None
        upstream_idx = self._coerce_pressure_pipe_row_index(target.get("upstream_row_index", -1))
        target_idx = self._coerce_pressure_pipe_row_index(target.get("target_row_index", -1))
        if 0 <= upstream_idx < len(nodes):
            start_water_level = self._normalize_pressure_pipe_export_number(
                getattr(nodes[upstream_idx], "water_level", None),
                allow_zero=True,
            )
        if 0 <= target_idx < len(nodes):
            end_water_level = self._normalize_pressure_pipe_export_number(
                getattr(nodes[target_idx], "water_level", None),
                allow_zero=True,
            )
        return start_water_level, end_water_level

    def _build_pressure_pipe_characteristic_boundary_summary_from_targets(self, nodes, targets) -> dict:
        """按有压管道自身边界汇总主长度与首末水位。"""
        summary = {}
        for target in targets or []:
            flow_section_text = self._normalize_pressure_pipe_summary_flow_section(
                target.get("flow_section", "")
            )
            if not flow_section_text:
                continue

            start_station, end_station = self._resolve_pressure_pipe_target_station_range(target)
            if start_station is None or end_station is None:
                continue

            start_water_level, end_water_level = self._resolve_pressure_pipe_target_boundary_water_levels(
                nodes,
                target,
            )
            entry = summary.setdefault(
                flow_section_text,
                {
                    "start_station": None,
                    "end_station": None,
                    "start_water_level": None,
                    "end_water_level": None,
                },
            )

            if (
                entry["start_station"] is None
                or start_station < entry["start_station"] - 1e-9
                or (
                    abs(start_station - entry["start_station"]) <= 1e-9
                    and entry["start_water_level"] is None
                    and start_water_level is not None
                )
            ):
                entry["start_station"] = start_station
                entry["start_water_level"] = start_water_level
            elif (
                abs(start_station - entry["start_station"]) <= 1e-9
                and start_water_level is not None
            ):
                entry["start_water_level"] = start_water_level

            if (
                entry["end_station"] is None
                or end_station > entry["end_station"] + 1e-9
                or (
                    abs(end_station - entry["end_station"]) <= 1e-9
                    and entry["end_water_level"] is None
                    and end_water_level is not None
                )
            ):
                entry["end_station"] = end_station
                entry["end_water_level"] = end_water_level
            elif (
                abs(end_station - entry["end_station"]) <= 1e-9
                and end_water_level is not None
            ):
                entry["end_water_level"] = end_water_level

        resolved = {}
        for flow_section_text, entry in summary.items():
            start_station = entry.get("start_station")
            end_station = entry.get("end_station")
            if start_station is None or end_station is None or end_station < start_station:
                continue
            payload = {
                "total_length": round(float(end_station) - float(start_station), 6),
            }
            if entry.get("start_water_level") is not None:
                payload["start_water_level"] = entry.get("start_water_level")
            if entry.get("end_water_level") is not None:
                payload["end_water_level"] = entry.get("end_water_level")
            resolved[flow_section_text] = payload
        return resolved

    def _build_pressure_pipe_characteristic_export_summary_from_nodes(self, nodes) -> tuple[dict, float | None, float | None]:
        settings = None
        build_settings = getattr(self, "_build_settings", None)
        if callable(build_settings):
            try:
                settings = build_settings()
            except Exception:
                settings = None
        get_current_channel_level = getattr(self, "_get_current_channel_level_text", None)
        if callable(get_current_channel_level):
            try:
                channel_level = str(get_current_channel_level(settings) or "").strip()
            except Exception:
                channel_level = str(getattr(settings, "channel_level", "") or "").strip()
        else:
            channel_level = str(getattr(settings, "channel_level", "") or "").strip()
        is_xxpipe_channel = channel_level in XXPIPE_CHANNEL_LEVEL_OPTIONS

        def _make_entry(flow_section_text: str) -> dict:
            return {
                "flow_section": flow_section_text,
                "total_length": 0.0,
                "prefix_extension_length": 0.0,
                "start_water_level": None,
                "end_water_level": None,
                "tunnel_count": 0,
                "tunnel_length": 0.0,
                "directional_drill_count": 0,
                "directional_drill_length": 0.0,
                "jacking_count": 0,
                "jacking_length": 0.0,
            }

        def _resolve_station_mc(node) -> float | None:
            """提取节点里程桩号。"""
            try:
                number = float(getattr(node, "station_MC", 0.0) or 0.0)
            except (TypeError, ValueError):
                return None
            if not math.isfinite(number):
                return None
            return number

        summary_by_flow_section = {}
        if not nodes:
            return summary_by_flow_section, None, None

        count_map = {
            "隧洞": "tunnel_count",
            "定向钻": "directional_drill_count",
            "顶管": "jacking_count",
        }
        length_map = {
            "隧洞": "tunnel_length",
            "定向钻": "directional_drill_length",
            "顶管": "jacking_length",
        }
        building_buckets = set(count_map.keys())
        active_building_starts = {}
        pressure_gate_buckets = {"有压管道", "定向钻", "顶管"}
        xxqu_total_length_enabled_indexes = set()
        xxqu_tunnel_enabled_indexes = set()
        xxpipe_boundary_extension_lengths = {}

        def _flush_xxqu_pressure_chain(chain_entries: list[tuple[int, str]], flow_section_text: str) -> None:
            """按连续承压链标记主长度与 xx渠 隧洞统计口径。"""
            if not chain_entries:
                return
            if not any(bucket in pressure_gate_buckets for _, bucket in chain_entries):
                return

            has_pressure_before = {}
            pressure_seen = False
            for row_index, bucket in chain_entries:
                has_pressure_before[row_index] = pressure_seen
                if bucket in pressure_gate_buckets:
                    pressure_seen = True
                    xxqu_total_length_enabled_indexes.add(row_index)

            has_pressure_after = {}
            pressure_seen = False
            for row_index, bucket in reversed(chain_entries):
                has_pressure_after[row_index] = pressure_seen
                if bucket in pressure_gate_buckets:
                    pressure_seen = True

            for row_index, bucket in chain_entries:
                if bucket != "隧洞":
                    continue
                if not has_pressure_before.get(row_index) or not has_pressure_after.get(row_index):
                    continue
                xxqu_total_length_enabled_indexes.add(row_index)
                if not is_xxpipe_channel:
                    xxqu_tunnel_enabled_indexes.add(row_index)

            if not is_xxpipe_channel or not flow_section_text:
                return
            first_pressure_pos = next(
                (pos for pos, (_, bucket) in enumerate(chain_entries) if bucket in pressure_gate_buckets),
                None,
            )
            if first_pressure_pos in (None, 0):
                return
            prefix_entries = chain_entries[:first_pressure_pos]
            if not prefix_entries or any(bucket != "隧洞" for _, bucket in prefix_entries):
                return
            prefix_start_station = _resolve_station_mc(nodes[prefix_entries[0][0]])
            first_pressure_station = _resolve_station_mc(nodes[chain_entries[first_pressure_pos][0]])
            if prefix_start_station is None or first_pressure_station is None:
                return
            extra_length = first_pressure_station - prefix_start_station
            if extra_length <= 0:
                return
            xxpipe_boundary_extension_lengths[flow_section_text] = round(
                xxpipe_boundary_extension_lengths.get(flow_section_text, 0.0) + extra_length,
                6,
            )
        current_chain_entries = []
        current_chain_flow_section = ""
        for row_index, node in enumerate(nodes):
            flow_section_text = self._normalize_pressure_pipe_summary_flow_section(
                getattr(node, "flow_section", "")
            )
            bucket = self._classify_pressure_pipe_summary_bucket(node)
            if flow_section_text and bucket:
                if current_chain_entries and flow_section_text == current_chain_flow_section:
                    current_chain_entries.append((row_index, bucket))
                else:
                    _flush_xxqu_pressure_chain(current_chain_entries, current_chain_flow_section)
                    current_chain_entries = [(row_index, bucket)]
                    current_chain_flow_section = flow_section_text
                continue
            _flush_xxqu_pressure_chain(current_chain_entries, current_chain_flow_section)
            current_chain_entries = []
            current_chain_flow_section = ""
        _flush_xxqu_pressure_chain(current_chain_entries, current_chain_flow_section)
        if is_xxpipe_channel:
            flow_section_groups = {}
            flow_section_order = []
            for node in nodes:
                flow_section_text = self._normalize_pressure_pipe_summary_flow_section(
                    getattr(node, "flow_section", "")
                )
                if not flow_section_text:
                    continue
                if flow_section_text not in flow_section_groups:
                    flow_section_groups[flow_section_text] = []
                    flow_section_order.append(flow_section_text)
                flow_section_groups[flow_section_text].append(node)

            for flow_section_text in flow_section_order:
                prefix_start_station = None
                first_pressure_station = None
                for node in flow_section_groups.get(flow_section_text, []):
                    bucket = self._classify_pressure_pipe_summary_bucket(node)
                    if bucket is None:
                        continue
                    station_mc = _resolve_station_mc(node)
                    if station_mc is None:
                        continue
                    if bucket in pressure_gate_buckets:
                        first_pressure_station = station_mc
                        break
                    if bucket == "隧洞" and prefix_start_station is None:
                        prefix_start_station = station_mc
                if prefix_start_station is None or first_pressure_station is None:
                    continue
                extra_length = first_pressure_station - prefix_start_station
                if extra_length <= 0:
                    continue
                xxpipe_boundary_extension_lengths[flow_section_text] = round(
                    max(
                        xxpipe_boundary_extension_lengths.get(flow_section_text, 0.0),
                        extra_length,
                    ),
                    6,
                )

        for row_index, node in enumerate(nodes):
            flow_section_text = self._normalize_pressure_pipe_summary_flow_section(
                getattr(node, "flow_section", "")
            )
            bucket = self._classify_pressure_pipe_summary_bucket(node)
            if flow_section_text and bucket:
                entry = summary_by_flow_section.setdefault(flow_section_text, _make_entry(flow_section_text))
                water_level = self._normalize_pressure_pipe_export_number(
                    getattr(node, "water_level", None),
                    allow_zero=True,
                )
                if water_level is not None:
                    # 先按各流量段自己的首末有效水位记默认值；
                    # 若后续识别到连续切段点，再用边界点水位统一覆盖两侧。
                    if entry["start_water_level"] is None:
                        entry["start_water_level"] = water_level
                    entry["end_water_level"] = water_level
            if not flow_section_text or bucket not in building_buckets:
                continue
            # xx渠 只统计夹在有压类结构中间的隧洞；xx管 保持原有整线口径。
            if (
                not is_xxpipe_channel
                and bucket == "隧洞"
                and row_index not in xxqu_tunnel_enabled_indexes
            ):
                continue

            in_out_text = self._get_pressure_pipe_summary_in_out_text(node)
            if in_out_text == "进":
                entry = summary_by_flow_section.setdefault(flow_section_text, _make_entry(flow_section_text))
                entry[count_map[bucket]] += 1
                station_mc = _resolve_station_mc(node)
                if station_mc is not None:
                    key = (flow_section_text, bucket)
                    active_building_starts.setdefault(key, []).append(station_mc)
                continue

            if in_out_text != "出":
                continue

            station_mc = _resolve_station_mc(node)
            if station_mc is None:
                continue
            key = (flow_section_text, bucket)
            start_stations = active_building_starts.get(key)
            if not start_stations:
                continue
            start_station = start_stations.pop()
            if not start_stations:
                active_building_starts.pop(key, None)
            seg_len = station_mc - start_station
            if seg_len <= 0:
                continue
            entry = summary_by_flow_section.setdefault(flow_section_text, _make_entry(flow_section_text))
            entry[length_map[bucket]] = round(entry[length_map[bucket]] + seg_len, 6)

        for boundary in self._iter_pressure_pipe_continuous_flow_boundaries(nodes):
            current_flow_section_text = boundary.get("current_flow_section")
            next_flow_section_text = boundary.get("next_flow_section")
            boundary_water_level = boundary.get("boundary_water_level")
            current_entry = summary_by_flow_section.setdefault(
                current_flow_section_text,
                _make_entry(current_flow_section_text),
            )
            next_entry = summary_by_flow_section.setdefault(
                next_flow_section_text,
                _make_entry(next_flow_section_text),
            )
            # 切段点取后一流量段首个承压节点水位，并同时保护前后两段。
            current_entry["end_water_level"] = boundary_water_level
            next_entry["start_water_level"] = boundary_water_level

        for idx in range(len(nodes) - 1):
            current = nodes[idx]
            nxt = nodes[idx + 1]
            flow_section_text = self._normalize_pressure_pipe_summary_flow_section(
                getattr(current, "flow_section", "")
            )
            bucket = self._classify_pressure_pipe_summary_bucket(current)
            if not flow_section_text:
                continue
            if bucket is None:
                continue
            try:
                seg_len = float(getattr(nxt, "station_MC", 0.0) or 0.0) - float(getattr(current, "station_MC", 0.0) or 0.0)
            except (TypeError, ValueError):
                continue
            if seg_len <= 0:
                continue
            # 主长度统一按“首个真正有压段到末个真正有压段”的承压链累计；
            # xx渠 只统计夹在承压链里的隧洞，xx管 额外把前置无压隧洞和空档补回。
            if idx not in xxqu_total_length_enabled_indexes:
                continue
            current_in_out = self._get_pressure_pipe_summary_in_out_text(current)
            if current_in_out == "出" and (idx + 1) not in xxqu_total_length_enabled_indexes:
                continue
            # 跨流量段时，下一流量段首行桩号仍是上一流量段的终点，
            # 这段边界长度应归入当前行所属的上一流量段。
            entry = summary_by_flow_section.setdefault(flow_section_text, _make_entry(flow_section_text))
            entry["total_length"] = round(entry["total_length"] + seg_len, 6)

        if is_xxpipe_channel:
            for flow_section_text, extra_length in xxpipe_boundary_extension_lengths.items():
                if extra_length <= 0:
                    continue
                entry = summary_by_flow_section.setdefault(flow_section_text, _make_entry(flow_section_text))
                entry["prefix_extension_length"] = round(
                    entry["prefix_extension_length"] + extra_length,
                    6,
                )
                entry["total_length"] = round(entry["total_length"] + extra_length, 6)

        return summary_by_flow_section, None, None

    def get_pressure_pipe_characteristic_export_summary(self, rows=None, nodes=None) -> dict:
        targets = self._collect_pressure_pipe_export_targets(rows)
        if not targets:
            return {}

        settings = None
        build_settings = getattr(self, "_build_settings", None)
        if callable(build_settings):
            try:
                settings = build_settings()
            except Exception:
                settings = None
        get_current_channel_level = getattr(self, "_get_current_channel_level_text", None)
        if callable(get_current_channel_level):
            try:
                channel_level = str(get_current_channel_level(settings) or "").strip()
            except Exception:
                channel_level = str(getattr(settings, "channel_level", "") or "").strip()
        else:
            channel_level = str(getattr(settings, "channel_level", "") or "").strip()
        is_xxpipe_channel = channel_level in XXPIPE_CHANNEL_LEVEL_OPTIONS
        nodes = self._get_pressure_pipe_summary_source_nodes(nodes)
        boundary_summary_by_flow_section = self._build_pressure_pipe_characteristic_boundary_summary_from_targets(
            nodes,
            targets,
        )
        summary_by_flow_section, start_water_level, end_water_level = (
            self._build_pressure_pipe_characteristic_export_summary_from_nodes(nodes)
        )

        protected_boundary_water_keys = {}
        for boundary in self._iter_pressure_pipe_continuous_flow_boundaries(nodes):
            current_flow_section_text = boundary.get("current_flow_section")
            next_flow_section_text = boundary.get("next_flow_section")
            protected_boundary_water_keys.setdefault(current_flow_section_text, set()).add(
                "end_water_level"
            )
            protected_boundary_water_keys.setdefault(next_flow_section_text, set()).add(
                "start_water_level"
            )

        resolved = {}
        for target in targets:
            flow_section_text = self._normalize_pressure_pipe_summary_flow_section(
                target.get("flow_section", "")
            )
            if not flow_section_text:
                continue
            payload = copy.deepcopy(summary_by_flow_section.get(flow_section_text) or {})
            boundary_payload = boundary_summary_by_flow_section.get(flow_section_text) or {}
            if isinstance(boundary_payload, dict) and boundary_payload:
                for key in ("start_water_level", "end_water_level"):
                    if (
                        key in boundary_payload
                        and key not in protected_boundary_water_keys.get(flow_section_text, set())
                    ):
                        payload[key] = copy.deepcopy(boundary_payload.get(key))
                if "total_length" in boundary_payload:
                    boundary_total_length = self._normalize_pressure_pipe_export_number(
                        boundary_payload.get("total_length"),
                        allow_zero=True,
                    )
                    if boundary_total_length is not None:
                        prefix_extension_length = 0.0
                        if is_xxpipe_channel:
                            prefix_extension_length = (
                                self._normalize_pressure_pipe_export_number(
                                    payload.get("prefix_extension_length"),
                                    allow_zero=True,
                                )
                                or 0.0
                            )
                        payload["total_length"] = round(boundary_total_length + prefix_extension_length, 6)
            payload.setdefault("flow_section", flow_section_text)
            payload.setdefault("start_water_level", start_water_level)
            payload.setdefault("end_water_level", end_water_level)
            payload.setdefault("tunnel_count", 0)
            payload.setdefault("tunnel_length", 0.0)
            payload.setdefault("directional_drill_count", 0)
            payload.setdefault("directional_drill_length", 0.0)
            payload.setdefault("jacking_count", 0)
            payload.setdefault("jacking_length", 0.0)
            payload.pop("prefix_extension_length", None)
            resolved[flow_section_text] = payload
        return resolved

    def get_pressure_pipe_export_results(self, rows=None) -> dict:
        """
        返回断面汇总导出可直接复用的有压管道结果映射。

        优先级固定为：
        1. 当前表3出口行已写回的损失值
        2. 当前批次 _pressure_pipe_calc_records
        3. 共享 PressurePipeManager 持久化结果
        """
        targets = self._collect_pressure_pipe_export_targets(rows)
        if not targets:
            return {}

        target_name_counts = {}
        target_identities = []
        for target in targets:
            target_identities.extend(self._get_pressure_pipe_export_target_identities(target))
            target_name_counts[target["name"]] = target_name_counts.get(target["name"], 0) + 1

        table_exact, table_by_name = self._index_pressure_pipe_table_results_for_export(target_identities)
        calc_exact, calc_by_name = self._index_pressure_pipe_calc_records_for_export()
        manager_exact, manager_by_name = self._index_pressure_pipe_manager_results_for_export(target_identities)

        resolved = {}
        for target in targets:
            identity = target["identity"]
            name = target["name"]
            candidate_identities = self._get_pressure_pipe_export_target_identities(target)
            result = next((table_exact.get(candidate) for candidate in candidate_identities if table_exact.get(candidate) is not None), None)
            if result is None and target_name_counts.get(name, 0) == 1:
                result = table_by_name.get(name)
            if result is None:
                result = next((calc_exact.get(candidate) for candidate in candidate_identities if calc_exact.get(candidate) is not None), None)
            if result is None and target_name_counts.get(name, 0) == 1:
                result = calc_by_name.get(name)
            if result is None:
                result = next((manager_exact.get(candidate) for candidate in candidate_identities if manager_exact.get(candidate) is not None), None)
            if result is None and target_name_counts.get(name, 0) == 1:
                result = manager_by_name.get(name)
            if result is not None:
                resolved[identity] = dict(result)
        return resolved

    def get_pressure_pipe_longitudinal_nodes_for_export(self, rows=None) -> dict:
        """
        返回断面汇总导出可直接复用的有压管道纵断面节点映射。

        优先按 identity 精确匹配；仅在目标名称唯一时才允许名称兜底，
        以避免同名不同 flow_section 的数据串用。
        """
        targets = self._collect_pressure_pipe_export_targets(rows)
        if not targets:
            return {}

        target_name_counts = {}
        target_identities = []
        for target in targets:
            target_identities.extend(self._get_pressure_pipe_export_target_identities(target))
            target_name_counts[target["name"]] = target_name_counts.get(target["name"], 0) + 1

        manager_exact, manager_by_name = self._index_pressure_pipe_manager_longitudinal_nodes_for_export(
            target_identities
        )
        route_longitudinal_nodes = self._index_pressure_pipe_manager_route_longitudinal_nodes()

        resolved = {}
        for target in targets:
            identity = target["identity"]
            name = target["name"]
            candidate_identities = self._get_pressure_pipe_export_target_identities(target)
            result = next((manager_exact.get(candidate) for candidate in candidate_identities if manager_exact.get(candidate) is not None), None)
            if result is None and target_name_counts.get(name, 0) == 1:
                result = manager_by_name.get(name)
            route_key = str(target.get("route_key", "") or "").strip()
            if not route_key and isinstance(result, dict):
                route_key = str(result.get("route_key", "") or "").strip()
            station_mc = self._coerce_pressure_pipe_finite_float(target.get("station_mc"))
            longitudinal_nodes = list(result.get("longitudinal_nodes", []) or []) if isinstance(result, dict) else []
            route_result = route_longitudinal_nodes.get(route_key) if route_key else None
            route_nodes = (
                list(route_result.get("longitudinal_nodes", []) or [])
                if isinstance(route_result, dict)
                else []
            )
            if route_key and (
                not self._has_exportable_pressure_pipe_longitudinal_nodes(longitudinal_nodes)
                or not self._pressure_pipe_longitudinal_nodes_cover_station(longitudinal_nodes, station_mc)
            ):
                result = route_result
                longitudinal_nodes = list(route_nodes)
            if not self._has_exportable_pressure_pipe_longitudinal_nodes(longitudinal_nodes):
                continue
            resolved[identity] = copy.deepcopy(longitudinal_nodes)
        return resolved

    def get_pressure_pipe_raw_profile_polylines_for_export(self, rows=None) -> dict:
        """返回 xx管 导出直接画线使用的 route 原线几何。"""
        targets = self._collect_pressure_pipe_export_targets(rows)
        if not targets:
            return {}

        route_keys = {
            str(target.get("route_key", "") or "").strip()
            for target in targets
            if str(target.get("route_key", "") or "").strip()
        }
        indexed = self._index_pressure_pipe_manager_route_raw_profile_polylines()
        if not route_keys:
            return {
                key: copy.deepcopy(payload.get("raw_profile_polyline", {}) or {})
                for key, payload in indexed.items()
            }

        resolved = {}
        for route_key in route_keys:
            payload = indexed.get(route_key, {})
            raw_profile_polyline = copy.deepcopy(payload.get("raw_profile_polyline", {}) or {})
            if not raw_profile_polyline:
                continue
            resolved[route_key] = raw_profile_polyline
        return resolved

    @staticmethod
    def _parse_item_float(item) -> float:
        if item is None:
            return 0.0
        try:
            txt = item.text().strip()
            return float(txt) if txt else 0.0
        except Exception:
            return 0.0

    def _set_table_cell_text_preserve_flags(self, row: int, col: int, text: str):
        if not self.node_table:
            return
        old_item = self.node_table.item(row, col)
        flags = old_item.flags() if old_item else (Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(flags)
        self.node_table.setItem(row, col, item)

    def _extract_pressure_pipe_dialog_groups(self, nodes, settings=None):
        """提取有压管道窗口专用分组。"""
        from utils.pressure_pipe_extractor import PressurePipeDataExtractor

        extractor = getattr(PressurePipeDataExtractor, "extract_dialog_pipe_groups", None)
        if callable(extractor):
            return extractor(nodes, settings=settings)
        return PressurePipeDataExtractor.extract_pipes(nodes, settings=settings)

    def _extract_pressure_pipe_dialog_chains(self, nodes, settings=None):
        """提取 xx管 连续承压链。"""
        from utils.pressure_pipe_extractor import PressurePipeDataExtractor

        extractor = getattr(PressurePipeDataExtractor, "extract_continuous_pressure_chains", None)
        if callable(extractor):
            return extractor(nodes, settings=settings)
        return []

    def _extract_pressure_pipe_routes(self, nodes, settings=None):
        """提取连续承压整线对象。"""
        from utils.pressure_pipe_extractor import PressurePipeDataExtractor

        extractor = getattr(PressurePipeDataExtractor, "extract_pressure_routes", None)
        if callable(extractor):
            return extractor(nodes, settings=settings)
        return []

    @staticmethod
    def _get_pressure_chain_member_identity(member) -> str:
        """返回连续承压链成员稳定标识。"""
        identity = str(getattr(member, "identity", "") or "").strip()
        if identity:
            return identity
        if bool(getattr(member, "split_from_named_group", False)) or str(
            getattr(member, "member_type", "") or ""
        ).strip() == "single_row":
            return ""
        group = getattr(member, "group", None)
        if group is not None:
            identity = str(getattr(group, "identity", "") or "").strip()
            if identity:
                return identity
        return ""

    @staticmethod
    def _get_pressure_chain_member_source_identity_aliases(member) -> list[str]:
        """返回逐段链成员对应的原命名组身份键列表。"""
        aliases = []
        for candidate in list(getattr(member, "source_identity_aliases", []) or []):
            candidate_text = str(candidate or "").strip()
            if candidate_text and candidate_text not in aliases:
                aliases.append(candidate_text)
        for candidate in (
            getattr(member, "parent_group_identity", ""),
            getattr(member, "parent_group_storage_key", ""),
        ):
            candidate_text = str(candidate or "").strip()
            if candidate_text and candidate_text not in aliases:
                aliases.append(candidate_text)
        return aliases

    @classmethod
    def _build_pressure_chain_source_lookup(cls, chain_descriptors) -> dict[str, list]:
        """按原命名组身份键索引逐段链成员。"""
        lookup: dict[str, list] = {}
        for descriptor in chain_descriptors or []:
            for member in descriptor.get("members", []) or []:
                aliases = cls._get_pressure_chain_member_source_identity_aliases(member)
                if not aliases:
                    continue
                for alias in aliases:
                    lookup.setdefault(alias, []).append(member)
        return lookup

    @staticmethod
    def _is_pressure_chain_anchor_member(member) -> bool:
        """判断是否为仅用于拓扑定位的链锚点成员。"""
        return bool(getattr(member, "is_anchor_member", False)) or not bool(
            getattr(member, "should_generate_row_loss", True)
        )

    @staticmethod
    def _is_pressure_chain_single_row_member(member) -> bool:
        """判断是否为单行链成员。"""
        return str(getattr(member, "member_type", "") or "").strip() == "single_row"

    @staticmethod
    def _is_pressure_chain_prefix_member(member) -> bool:
        """判断是否为需要单独计损的链前缀段。"""
        return str(getattr(member, "member_role", "") or "").strip() == "prefix_segment"

    def _build_pressure_pipe_chain_summary(self, descriptor, record_map: dict) -> dict:
        """汇总单条连续承压链，未完整成功时隐藏整线总损失。"""
        member_results = []
        success_count = 0
        failed_count = 0
        total_head_loss = 0.0

        for member in descriptor.get("members", []) or []:
            identity = self._get_pressure_chain_member_identity(member)
            record = record_map.get(identity)
            display_name = str(
                getattr(member, "display_name", "") or (record or {}).get("display_name", "") or ""
            ).strip()
            structure_type = str(
                getattr(member, "structure_type", "") or (record or {}).get("structure_type", "") or ""
            ).strip()
            member_role = str(
                getattr(member, "member_role", "") or (record or {}).get("member_role", "") or ""
            ).strip()

            if record is None:
                if self._is_pressure_chain_anchor_member(member):
                    status = "success"
                    writeback_enabled = False
                    member_total_head_loss = None
                    note = "链起点锚点，本行不写回"
                    error = ""
                else:
                    status = "failed"
                    writeback_enabled = False
                    member_total_head_loss = None
                    note = ""
                    error = f"{display_name or '未命名成员'}: 未匹配到本成员计算记录"
            else:
                status = str(record.get("status", "failed") or "failed").strip().lower()
                writeback_enabled = bool(record.get("writeback_enabled", True))
                member_total_head_loss = record.get("total_head_loss")
                note = str(record.get("note", "") or "").strip()
                error = str(record.get("error", "") or "").strip()

            if status == "success":
                success_count += 1
                if writeback_enabled and member_total_head_loss is not None:
                    total_head_loss += float(member_total_head_loss or 0.0)
            else:
                failed_count += 1
            member_results.append({
                "identity": identity,
                "display_name": display_name,
                "structure_type": structure_type,
                "member_role": member_role,
                "status": status,
                "writeback_enabled": writeback_enabled,
                "total_head_loss": member_total_head_loss,
                "note": note,
                "error": error,
            })

        chain_complete = failed_count <= 0
        return {
            "chain_id": descriptor.get("chain_id", ""),
            "flow_section": descriptor.get("flow_section", ""),
            "display_name": descriptor.get("display_name", ""),
            "chain_complete": chain_complete,
            "chain_status": "complete" if chain_complete else "incomplete",
            "total_head_loss": total_head_loss if chain_complete else None,
            "member_count": len(member_results),
            "success_count": success_count,
            "failed_count": failed_count,
            "member_results": member_results,
        }

    @staticmethod
    def _summarize_pressure_pipe_chain_flow_sections(chain) -> str:
        """汇总连续承压链覆盖的流量段文本。"""
        flow_sections = []
        for member in list(getattr(chain, "members", []) or []):
            flow_section = str(getattr(member, "flow_section", "") or "").strip()
            if flow_section and flow_section not in flow_sections:
                flow_sections.append(flow_section)
        if flow_sections:
            return "、".join(flow_sections)
        return str(getattr(chain, "flow_section", "") or "").strip() or "-"

    def _build_pressure_pipe_chain_descriptors(self, chains) -> list[dict]:
        """将连续承压链转换为界面与结果汇总可复用的描述结构。"""
        descriptors = []
        for chain_index, chain in enumerate(chains or [], start=1):
            flow_section = self._summarize_pressure_pipe_chain_flow_sections(chain)
            start_row_index = self._coerce_pressure_pipe_row_index(getattr(chain, "start_row_index", -1))
            end_row_index = self._coerce_pressure_pipe_row_index(
                getattr(chain, "end_row_index", start_row_index),
                start_row_index,
            )
            descriptors.append({
                "chain_id": f"chain{chain_index}-r{start_row_index + 1}-{end_row_index + 1}",
                "flow_section": flow_section,
                "display_name": f"连续承压链{chain_index}",
                "members": list(getattr(chain, "members", []) or []),
                "start_row_index": start_row_index,
                "end_row_index": end_row_index,
            })
        return descriptors

    @staticmethod
    def _build_pressure_route_segment_index(pressure_routes) -> tuple[dict, dict]:
        """按 identity / route_key 索引连续承压整线与子段。"""
        segment_by_identity = {}
        route_by_key = {}
        for route in list(pressure_routes or []):
            route_key = str(getattr(route, "route_key", "") or "").strip()
            if route_key:
                route_by_key[route_key] = route
            for segment in list(getattr(route, "segments", []) or []):
                identity = str(getattr(segment, "identity", "") or "").strip()
                if identity and identity not in segment_by_identity:
                    segment_by_identity[identity] = segment
        return route_by_key, segment_by_identity

    @staticmethod
    def _resolve_pressure_route_profile_state(route_key: str, route_profiles: dict, route_profile_segments_by_key: dict) -> str:
        """归一化整线纵断面覆盖状态。"""
        route_nodes = list((route_profiles or {}).get(route_key, []) or [])
        if isinstance(route_key, str) and route_key and len(route_nodes) >= 2:
            return "ok"
        route_segments = list((route_profile_segments_by_key or {}).get(route_key, []) or [])
        if route_segments:
            return "ok"
        return "not_imported"

    @classmethod
    def _build_pressure_route_payloads(cls, pressure_routes, route_profiles, route_profile_segments_by_key) -> list[dict]:
        """把 Route 对象翻译成正式保存所需的轻量结构。"""
        payloads = []
        for route in list(pressure_routes or []):
            route_key = str(getattr(route, "route_key", "") or "").strip()
            if not route_key:
                continue
            payloads.append(
                {
                    "route_key": route_key,
                    "route_display_name": str(getattr(route, "route_display_name", "") or route_key).strip(),
                    "channel_level": str(getattr(route, "channel_level", "") or "").strip(),
                    "start_row_index": getattr(route, "start_row_index", -1),
                    "end_row_index": getattr(route, "end_row_index", -1),
                    "start_mc": getattr(route, "start_mc", 0.0),
                    "end_mc": getattr(route, "end_mc", 0.0),
                    "entered_pressurized_at_row": getattr(route, "entered_pressurized_at_row", -1),
                    "profile_state": cls._resolve_pressure_route_profile_state(
                        route_key,
                        route_profiles,
                        route_profile_segments_by_key,
                    ),
                    "segment_identities": [
                        str(getattr(segment, "identity", "") or "").strip()
                        for segment in list(getattr(route, "segments", []) or [])
                        if str(getattr(segment, "identity", "") or "").strip()
                    ],
                }
            )
        return payloads

    @classmethod
    def _resolve_pressure_segment_saved_profile(cls, segment_meta, route_profiles, route_profile_segments_by_key) -> tuple[list[dict], str, str]:
        """为正式 segments 存储解析子段纵断面与覆盖状态。"""
        if segment_meta is None:
            return [], "not_imported", ""

        identity = str(getattr(segment_meta, "identity", "") or "").strip()
        route_key = str(getattr(segment_meta, "route_key", "") or "").strip()
        route_segments = list((route_profile_segments_by_key or {}).get(route_key, []) or [])
        for item in route_segments:
            if not isinstance(item, dict):
                continue
            if str(item.get("segment_identity", "") or "").strip() != identity:
                continue
            long_nodes = copy.deepcopy(item.get("longitudinal_nodes", []) or [])
            if cls._has_exportable_pressure_pipe_longitudinal_nodes(long_nodes):
                return long_nodes, "ok", str(item.get("source_kind", "") or "segment_profile").strip() or "segment_profile"

        route_nodes = copy.deepcopy((route_profiles or {}).get(route_key, []) or [])
        if not route_nodes:
            return [], "not_imported", ""

        start_mc = cls._coerce_pressure_pipe_finite_float(getattr(segment_meta, "start_mc", None))
        end_mc = cls._coerce_pressure_pipe_finite_float(getattr(segment_meta, "end_mc", None))
        if start_mc is None or end_mc is None:
            if cls._has_exportable_pressure_pipe_longitudinal_nodes(route_nodes):
                return route_nodes, "ok", "route_profile"
            return [], "coverage_missing", "route_profile"

        try:
            from utils.pressure_pipe_longitudinal_utils import clip_longitudinal_nodes_to_range

            clipped_nodes = clip_longitudinal_nodes_to_range(route_nodes, float(start_mc), float(end_mc))
        except ValueError:
            return [], "coverage_missing", "route_profile"
        except Exception:
            return [], "identity_unmatched", "route_profile"

        if cls._has_exportable_pressure_pipe_longitudinal_nodes(clipped_nodes):
            return clipped_nodes, "ok", "route_profile"
        if clipped_nodes:
            return clipped_nodes, "coverage_missing", "route_profile"
        return [], "coverage_missing", "route_profile"

    @classmethod
    def _build_pressure_segment_result_payload(
        cls,
        record: dict,
        segment_meta,
        route_by_key: dict,
        route_profiles: dict,
        route_profile_segments_by_key: dict,
    ) -> dict:
        """把本次计算记录翻译成正式保存的 segment 结果。"""
        identity = str(record.get("identity", "") or "").strip()
        route_key = str(getattr(segment_meta, "route_key", "") or record.get("route_key", "") or "").strip()
        route_meta = route_by_key.get(route_key)
        long_nodes, profile_state, profile_source = cls._resolve_pressure_segment_saved_profile(
            segment_meta,
            route_profiles,
            route_profile_segments_by_key,
        )
        if segment_meta is None:
            base_name = str(record.get("name", "") or record.get("display_name", "") or identity).strip()
            member_display_name = str(record.get("display_name", "") or base_name).strip() or base_name
            dxf_display_name = base_name
            structure_type = str(record.get("structure_type", "") or "").strip()
            member_role = str(record.get("member_role", "") or "").strip()
            start_row_index = int(record.get("target_row_index", -1) or -1)
            end_row_index = start_row_index
            target_row_index = int(record.get("target_row_index", -1) or -1)
            upstream_row_index = int(record.get("upstream_row_index", -1) or -1)
            start_mc = None
            end_mc = None
        else:
            base_name = str(getattr(segment_meta, "base_name", "") or "").strip()
            member_display_name = str(getattr(segment_meta, "member_display_name", "") or base_name).strip() or base_name
            dxf_display_name = str(getattr(segment_meta, "dxf_display_name", "") or base_name).strip() or base_name
            structure_type = str(getattr(segment_meta, "structure_type", "") or "").strip()
            member_role = str(getattr(segment_meta, "member_role", "") or "").strip()
            start_row_index = int(getattr(segment_meta, "start_row_index", -1) or -1)
            end_row_index = int(getattr(segment_meta, "end_row_index", -1) or -1)
            target_row_index = int(getattr(segment_meta, "target_row_index", -1) or -1)
            upstream_row_index = int(getattr(segment_meta, "upstream_row_index", -1) or -1)
            start_mc = getattr(segment_meta, "start_mc", None)
            end_mc = getattr(segment_meta, "end_mc", None)

        if not long_nodes and record.get("status") == "success":
            profile_state = profile_state or "not_imported"

        return {
            "identity": identity,
            "route_key": route_key,
            "route_display_name": str(
                getattr(route_meta, "route_display_name", "")
                or record.get("route_display_name", "")
                or route_key
            ).strip(),
            "base_name": base_name,
            "member_display_name": member_display_name,
            "dxf_display_name": dxf_display_name,
            "structure_type": structure_type,
            "member_role": member_role,
            "start_row_index": start_row_index,
            "end_row_index": end_row_index,
            "target_row_index": target_row_index,
            "upstream_row_index": upstream_row_index,
            "start_mc": start_mc,
            "end_mc": end_mc,
            "is_pressurized_tail_member": True,
            "status": str(record.get("status", "") or "").strip(),
            "friction_loss": record.get("friction_loss"),
            "bend_loss": record.get("total_bend_loss"),
            "local_loss": record.get("local_loss"),
            "total_loss": record.get("total_head_loss"),
            "applied_to_row_index": int(record.get("target_row_index", -1) or -1),
            "note": str(record.get("note", "") or "").strip(),
            "computed_from_profile_source": profile_source,
            "longitudinal_nodes": long_nodes,
            "profile_state": profile_state,
            "flow_section": str(record.get("flow_section", "") or "").strip(),
            "plan_total_length": record.get("total_length"),
            "pipe_velocity": record.get("pipe_velocity"),
            "data_mode": str(record.get("data_mode", "") or "").strip(),
            "material_key": str(record.get("resolved_material_key", "") or record.get("material_key", "") or "").strip(),
            "D": record.get("D"),
        }

    @staticmethod
    def _get_pressure_pipe_node_structure_text(node) -> str:
        """提取节点结构形式文本。"""
        struct_type = getattr(node, "structure_type", None)
        value = getattr(struct_type, "value", struct_type)
        return str(value or "").strip()

    @classmethod
    def _build_xxpipe_route_target_label(cls, node, row_index: int) -> str:
        """生成 xx管 整线覆盖校验目标的展示文本。"""
        name = str(getattr(node, "name", "") or "").strip()
        struct_text = cls._get_pressure_pipe_node_structure_text(node) or "未知结构"
        base_text = name or struct_text
        return f"第{int(row_index) + 1}行 {base_text}"

    @classmethod
    def _resolve_xxpipe_route_import_anchor_station_mc(cls, route_nodes) -> float | None:
        """解析 xx管 整线纵断面导入锚点，优先取首个非隧洞节点的实际或回退桩号。"""
        ordered_nodes = list(route_nodes or [])
        if not ordered_nodes:
            return None

        first_target_index = next(
            (
                index
                for index, node in enumerate(ordered_nodes)
                if "隧洞" not in cls._get_pressure_pipe_node_structure_text(node)
            ),
            None,
        )
        if first_target_index is None:
            return None

        station_value = cls._coerce_pressure_pipe_finite_float(
            getattr(ordered_nodes[first_target_index], "station_MC", None)
        )
        if station_value is not None:
            return station_value

        try:
            from app_渠系计算前端.water_profile.cad_tools import (
                resolve_xxpipe_profile_station_targets,
            )

            station_targets, _station_errors = resolve_xxpipe_profile_station_targets(
                ordered_nodes,
                station_prefix="",
            )
        except Exception:
            station_targets = []

        if first_target_index >= len(station_targets):
            return None
        return cls._coerce_pressure_pipe_finite_float(
            station_targets[first_target_index].get("station_mc", None)
        )

    @staticmethod
    def _coerce_pressure_pipe_finite_float(value):
        """将任意值安全转换为有限浮点数。"""
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None
        if not math.isfinite(number):
            return None
        return number

    @classmethod
    def _apply_pressure_pipe_manager_tunnel_config_to_group(cls, group, config):
        """把缓存里的隧洞参数补回当前分组对象。"""
        if config is None or not cls._is_pressure_pipe_group_tunnel_segment(group):
            return

        text_fields = ("segment_geometry_source", "tunnel_section_type", "tunnel_profile_mode")
        float_fields = (
            "tunnel_invert_inlet",
            "tunnel_slope_i",
            "tunnel_invert_outlet_check",
            "tunnel_roughness_n",
        )
        positive_float_fields = {"tunnel_slope_i", "tunnel_roughness_n"}

        for field_name in text_fields:
            current_value = str(getattr(group, field_name, "") or "").strip()
            if current_value:
                continue
            value = str(getattr(config, field_name, "") or "").strip()
            if value:
                setattr(group, field_name, value)
        for field_name in float_fields:
            current_value = cls._coerce_pressure_pipe_finite_float(getattr(group, field_name, None))
            if current_value is not None:
                if field_name in positive_float_fields:
                    if current_value > 0:
                        continue
                else:
                    continue
            value = getattr(config, field_name, None)
            config_value = cls._coerce_pressure_pipe_finite_float(value)
            if config_value is None:
                continue
            if field_name in positive_float_fields and config_value <= 0:
                continue
            setattr(group, field_name, config_value)

        current_params = getattr(group, "tunnel_section_params", {}) or {}
        params = getattr(config, "tunnel_section_params", {}) or {}
        if (not isinstance(current_params, dict) or not current_params) and isinstance(params, dict) and params:
            setattr(group, "tunnel_section_params", copy.deepcopy(params))
        roughness_n = cls._coerce_pressure_pipe_finite_float(getattr(group, "tunnel_roughness_n", None))
        if roughness_n is not None and roughness_n > 0:
            setattr(group, "roughness", float(roughness_n))
            for node in list(getattr(group, "rows", []) or []):
                try:
                    node.roughness = float(roughness_n)
                except Exception:
                    continue

    def _hydrate_pressure_pipe_groups_from_manager(self, pipe_groups, manager):
        """把对话框缓存的隧洞参数回填到新提取的分组对象。"""
        get_pipe_config = getattr(manager, "get_pipe_config", None)
        if not callable(get_pipe_config):
            return

        for group in pipe_groups or []:
            if not self._is_pressure_pipe_group_tunnel_segment(group):
                continue
            storage_key = self._get_pressure_pipe_group_storage_key(group)
            config = get_pipe_config(storage_key)
            if config is None:
                legacy_name = str(getattr(group, "name", "") or "").strip()
                if legacy_name and legacy_name != storage_key:
                    config = get_pipe_config(legacy_name)
            self._apply_pressure_pipe_manager_tunnel_config_to_group(group, config)

    @classmethod
    def _resolve_tunnel_display_bottom_elevation(cls, node) -> float | None:
        """优先读取当前节点可用于展示的隧洞底高。"""
        bottom = cls._coerce_pressure_pipe_finite_float(getattr(node, "bottom_elevation", None))
        if bottom is not None:
            return bottom
        water_level = cls._coerce_pressure_pipe_finite_float(getattr(node, "water_level", None))
        water_depth = cls._coerce_pressure_pipe_finite_float(getattr(node, "water_depth", None))
        if water_level is None or water_depth is None:
            return None
        return float(water_level - water_depth)

    @classmethod
    def _build_generated_tunnel_profile_segment(cls, group) -> tuple[list[dict], list[str]]:
        """按当前水力结果反推底线，必要时兼容旧版底高输入生成隧洞纵断面。"""
        start_mc = cls._coerce_pressure_pipe_finite_float(getattr(group, "segment_start_mc", None))
        end_mc = cls._coerce_pressure_pipe_finite_float(getattr(group, "segment_end_mc", None))
        if start_mc is None or end_mc is None:
            return [], ["隧洞参数不完整，无法生成纵断面"]

        rows = list(getattr(group, "rows", []) or [])
        start_node = rows[0] if rows else None
        end_node = rows[-1] if rows else None
        start_bottom = cls._resolve_tunnel_display_bottom_elevation(start_node)
        end_bottom = cls._resolve_tunnel_display_bottom_elevation(end_node)
        if start_bottom is not None and end_bottom is not None:
            return [
                {
                    "chainage": float(start_mc),
                    "elevation": float(start_bottom),
                    "turn_type": "NONE",
                    "turn_angle": 0.0,
                    "vertical_curve_radius": 0.0,
                },
                {
                    "chainage": float(end_mc),
                    "elevation": float(end_bottom),
                    "turn_type": "NONE",
                    "turn_angle": 0.0,
                    "vertical_curve_radius": 0.0,
                },
            ], []

        invert_inlet = cls._coerce_pressure_pipe_finite_float(getattr(group, "tunnel_invert_inlet", None))
        slope_i = cls._coerce_pressure_pipe_finite_float(getattr(group, "tunnel_slope_i", None))
        if invert_inlet is None or slope_i is None:
            return [], ["隧洞参数不完整，无法生成纵断面"]

        outlet_invert = invert_inlet - slope_i * (end_mc - start_mc)
        warnings = []
        outlet_check = cls._coerce_pressure_pipe_finite_float(
            getattr(group, "tunnel_invert_outlet_check", None)
        )
        if outlet_check is not None and abs(outlet_check - outlet_invert) > 0.01:
            warnings.append(
                f"出口底高校核值与坡降推算值偏差 {abs(outlet_check - outlet_invert):.3f}m"
            )

        return [
            {
                "chainage": float(start_mc),
                "elevation": float(invert_inlet),
                "turn_type": "NONE",
                "turn_angle": 0.0,
                "vertical_curve_radius": 0.0,
            },
            {
                "chainage": float(end_mc),
                "elevation": float(outlet_invert),
                "turn_type": "NONE",
                "turn_angle": 0.0,
                "vertical_curve_radius": 0.0,
            },
        ], warnings

    @classmethod
    def _build_pressure_pipe_route_profile_segments(cls, pipe_groups, longitudinal_nodes_dict) -> dict:
        """按 route 组装整线分段纵断面，供 mixed route 计算和导出复用。"""
        route_groups = {}
        for group in list(pipe_groups or []):
            route_key = cls._get_pressure_pipe_group_route_key(group)
            if not route_key:
                continue
            route_groups.setdefault(route_key, []).append(group)

        route_profile_segments = {}
        for route_key, groups in route_groups.items():
            route_nodes = copy.deepcopy((longitudinal_nodes_dict or {}).get(route_key, []) or [])
            ordered_groups = sorted(
                groups,
                key=lambda item: (
                    cls._coerce_pressure_pipe_finite_float(getattr(item, "segment_start_mc", None)) is None,
                    cls._coerce_pressure_pipe_finite_float(getattr(item, "segment_start_mc", None)) or 0.0,
                ),
            )
            segments = []
            for group in ordered_groups:
                segment_identity = cls._build_pressure_pipe_group_identity(group) or cls._get_pressure_pipe_group_storage_key(group)
                structure_text = cls._get_pressure_pipe_group_structure_text(group)
                start_mc = cls._coerce_pressure_pipe_finite_float(getattr(group, "segment_start_mc", None))
                end_mc = cls._coerce_pressure_pipe_finite_float(getattr(group, "segment_end_mc", None))
                warnings = []
                source_kind = "non_tunnel_dxf"
                long_nodes = []

                if cls._is_pressure_pipe_group_tunnel_segment(group):
                    source_kind = "generated_tunnel"
                    long_nodes, warnings = cls._build_generated_tunnel_profile_segment(group)
                elif cls._has_exportable_pressure_pipe_longitudinal_nodes(route_nodes):
                    if start_mc is not None and end_mc is not None:
                        try:
                            from utils.pressure_pipe_longitudinal_utils import clip_longitudinal_nodes_to_range

                            long_nodes = clip_longitudinal_nodes_to_range(
                                route_nodes,
                                float(start_mc),
                                float(end_mc),
                            )
                        except Exception as exc:
                            warnings.append(str(exc))
                    else:
                        long_nodes = copy.deepcopy(route_nodes)
                else:
                    warnings.append("整线未导入可用纵断面DXF")

                if not long_nodes and not warnings:
                    warnings.append("当前子段缺少可用纵断面")
                segments.append(
                    {
                        "segment_identity": segment_identity,
                        "structure_type": structure_text,
                        "source_kind": source_kind,
                        "start_mc": start_mc,
                        "end_mc": end_mc,
                        "longitudinal_nodes": long_nodes,
                        "warnings": warnings,
                    }
                )

            for idx in range(len(segments) - 1):
                current = segments[idx]
                nxt = segments[idx + 1]
                current_nodes = list(current.get("longitudinal_nodes", []) or [])
                next_nodes = list(nxt.get("longitudinal_nodes", []) or [])
                if not current_nodes or not next_nodes:
                    continue
                current_end = cls._coerce_pressure_pipe_finite_float(current_nodes[-1].get("elevation"))
                next_start = cls._coerce_pressure_pipe_finite_float(next_nodes[0].get("elevation"))
                if current_end is None or next_start is None:
                    continue
                mismatch = abs(current_end - next_start)
                if mismatch <= 0.01:
                    continue
                warning_text = f"相邻子段接点高差 {mismatch:.3f}m，已保留但建议复核"
                current.setdefault("warnings", []).append(warning_text)
                nxt.setdefault("warnings", []).append(warning_text)

            route_profile_segments[route_key] = segments

        return route_profile_segments

    @staticmethod
    def _resolve_pressure_route_profile_state(route_key, route_profiles, route_profile_segments_by_key) -> str:
        """根据当前整线和分段纵断面情况生成统一覆盖状态。"""
        from utils.pressure_pipe_extractor import ProfileCoverageState

        route_nodes = list((route_profiles or {}).get(route_key, []) or [])
        segment_profiles = list((route_profile_segments_by_key or {}).get(route_key, []) or [])
        if len(route_nodes) >= 2:
            return ProfileCoverageState.OK
        if segment_profiles:
            if any(
                len(list(segment.get("longitudinal_nodes", []) or [])) >= 2
                for segment in segment_profiles
                if isinstance(segment, dict)
            ):
                if any(list(segment.get("warnings", []) or []) for segment in segment_profiles if isinstance(segment, dict)):
                    return ProfileCoverageState.COVERAGE_MISSING
                return ProfileCoverageState.OK
            return ProfileCoverageState.COVERAGE_MISSING
        return ProfileCoverageState.NOT_IMPORTED

    @classmethod
    def _build_pressure_route_persist_payloads(cls, nodes, settings, route_profiles, route_profile_segments_by_key):
        """把当前整线识别结果整理成可持久化的 route 载荷。"""
        from utils.pressure_pipe_extractor import PressurePipeDataExtractor

        pressure_routes = list(
            PressurePipeDataExtractor.extract_pressure_routes(nodes, settings=settings) or []
        )
        payloads = []
        for route in pressure_routes:
            segment_identities = [
                str(getattr(segment, "identity", "") or "").strip()
                for segment in list(getattr(route, "segments", []) or [])
                if str(getattr(segment, "identity", "") or "").strip()
            ]
            payloads.append({
                "route_key": str(getattr(route, "route_key", "") or "").strip(),
                "route_display_name": str(getattr(route, "route_display_name", "") or "").strip(),
                "channel_level": str(getattr(route, "channel_level", "") or "").strip(),
                "start_row_index": int(getattr(route, "start_row_index", -1) or -1),
                "end_row_index": int(getattr(route, "end_row_index", -1) or -1),
                "start_mc": cls._coerce_pressure_pipe_finite_float(getattr(route, "start_mc", None)),
                "end_mc": cls._coerce_pressure_pipe_finite_float(getattr(route, "end_mc", None)),
                "entered_pressurized_at_row": int(getattr(route, "entered_pressurized_at_row", -1) or -1),
                "profile_state": cls._resolve_pressure_route_profile_state(
                    str(getattr(route, "route_key", "") or "").strip(),
                    route_profiles,
                    route_profile_segments_by_key,
                ),
                "segment_identities": segment_identities,
            })
        return pressure_routes, payloads

    @classmethod
    def _resolve_pressure_segment_profile_payload(cls, segment, route_profiles, route_profile_segments_by_key):
        """优先读取子段纵断面，其次回退整线纵断面。"""
        route_key = str(getattr(segment, "route_key", "") or "").strip()
        identity = str(getattr(segment, "identity", "") or "").strip()
        segment_profiles = list((route_profile_segments_by_key or {}).get(route_key, []) or [])
        for profile in segment_profiles:
            if not isinstance(profile, dict):
                continue
            if str(profile.get("segment_identity", "") or "").strip() != identity:
                continue
            return (
                copy.deepcopy(profile.get("longitudinal_nodes", []) or []),
                "segment_profile",
            )
        return (
            copy.deepcopy((route_profiles or {}).get(route_key, []) or []),
            "route_profile",
        )

    @classmethod
    def _build_pressure_segment_persist_payloads(
        cls,
        pressure_routes,
        record_map,
        route_profiles,
        route_profile_segments_by_key,
        route_payloads_by_key,
    ):
        """把批量计算结果整理成正式的 segment 持久化载荷。"""
        from utils.pressure_pipe_extractor import ProfileCoverageState

        payloads = []
        for route in list(pressure_routes or []):
            route_key = str(getattr(route, "route_key", "") or "").strip()
            route_display_name = str(getattr(route, "route_display_name", "") or "").strip()
            route_profile_state = str(
                (route_payloads_by_key.get(route_key, {}) or {}).get("profile_state", "")
                or ProfileCoverageState.NOT_IMPORTED
            ).strip()
            for segment in list(getattr(route, "segments", []) or []):
                identity = str(getattr(segment, "identity", "") or "").strip()
                if not identity:
                    continue
                record = copy.deepcopy(record_map.get(identity) or {})
                longitudinal_nodes, profile_source = cls._resolve_pressure_segment_profile_payload(
                    segment,
                    route_profiles,
                    route_profile_segments_by_key,
                )
                local_loss = record.get("local_loss")
                if local_loss is None:
                    local_loss = (
                        float(record.get("inlet_transition_loss", 0.0) or 0.0)
                        + float(record.get("outlet_transition_loss", 0.0) or 0.0)
                    )
                payloads.append({
                    "identity": identity,
                    "route_key": route_key,
                    "route_display_name": route_display_name,
                    "base_name": str(getattr(segment, "base_name", "") or "").strip(),
                    "member_display_name": str(getattr(segment, "member_display_name", "") or "").strip(),
                    "dxf_display_name": str(getattr(segment, "dxf_display_name", "") or "").strip(),
                    "structure_type": str(getattr(segment, "structure_type", "") or "").strip(),
                    "member_role": str(getattr(segment, "member_role", "") or "").strip(),
                    "start_row_index": int(getattr(segment, "start_row_index", -1) or -1),
                    "end_row_index": int(getattr(segment, "end_row_index", -1) or -1),
                    "target_row_index": int(getattr(segment, "target_row_index", -1) or -1),
                    "upstream_row_index": int(getattr(segment, "upstream_row_index", -1) or -1),
                    "start_mc": cls._coerce_pressure_pipe_finite_float(getattr(segment, "start_mc", None)),
                    "end_mc": cls._coerce_pressure_pipe_finite_float(getattr(segment, "end_mc", None)),
                    "is_pressurized_tail_member": bool(getattr(segment, "is_pressurized_tail_member", False)),
                    "status": str(record.get("status", "") or "").strip(),
                    "friction_loss": float(record.get("friction_loss", 0.0) or 0.0),
                    "bend_loss": float(
                        record.get("total_bend_loss", record.get("bend_loss", 0.0)) or 0.0
                    ),
                    "local_loss": float(local_loss or 0.0),
                    "total_loss": float(
                        record.get("total_head_loss", record.get("total_loss", 0.0)) or 0.0
                    ),
                    "applied_to_row_index": int(
                        record.get("target_row_index", getattr(segment, "target_row_index", -1)) or -1
                    ),
                    "note": str(record.get("note", "") or "").strip(),
                    "computed_from_profile_source": str(
                        record.get("computed_from_profile_source", "") or profile_source
                    ).strip(),
                    "longitudinal_nodes": longitudinal_nodes,
                    "profile_state": (
                        ProfileCoverageState.OK
                        if len(longitudinal_nodes) >= 2
                        else route_profile_state
                    ),
                })
        return payloads

    @classmethod
    def _collect_xxpipe_route_context_map(cls, nodes, pipe_groups) -> dict:
        """按整线汇总 xx管 路由信息，识别是否夹带隧洞。"""
        route_map = {}
        total_nodes = len(nodes or [])

        for group in pipe_groups or []:
            route_key = cls._get_pressure_pipe_group_route_key(group)
            if not route_key or route_key in route_map:
                continue

            display_name = cls._get_pressure_pipe_group_route_display_name(group) or route_key
            route_indices = []
            start_idx = getattr(group, "route_start_row_index", None)
            end_idx = getattr(group, "route_end_row_index", None)
            if isinstance(start_idx, int) and isinstance(end_idx, int) and total_nodes > 0:
                start = max(0, min(int(start_idx), total_nodes - 1))
                end = max(start, min(int(end_idx), total_nodes - 1))
                route_indices = list(range(start, end + 1))
            else:
                for row_idx in list(getattr(group, "row_indices", []) or []):
                    if isinstance(row_idx, int) and 0 <= row_idx < total_nodes:
                        route_indices.append(row_idx)
                target_row_index = getattr(group, "target_row_index", None)
                if isinstance(target_row_index, int) and 0 <= target_row_index < total_nodes:
                    route_indices.append(target_row_index)
                route_indices = sorted(set(route_indices))

            route_nodes = []
            route_targets = []
            contains_tunnel = False
            for row_idx in route_indices:
                node = nodes[row_idx]
                if getattr(node, "is_transition", False) or getattr(node, "is_auto_inserted_channel", False):
                    continue
                struct_text = cls._get_pressure_pipe_node_structure_text(node)
                if "隧洞" in struct_text:
                    contains_tunnel = True
                route_nodes.append(node)
                route_targets.append({
                    "row_index": int(row_idx),
                    "label": cls._build_xxpipe_route_target_label(node, row_idx),
                    "station_mc": getattr(node, "station_MC", None),
                    "x": getattr(node, "x", None),
                    "y": getattr(node, "y", None),
                })
            import_anchor_station_mc = cls._resolve_xxpipe_route_import_anchor_station_mc(route_nodes)

            route_map[route_key] = {
                "route_key": route_key,
                "display_name": display_name,
                "contains_tunnel": contains_tunnel,
                "import_anchor_station_mc": import_anchor_station_mc,
                "targets": route_targets,
                "nodes": route_nodes,
            }

        return route_map

    def _prepare_pressure_pipe_dialog_context(self, nodes, settings=None, *, show_xxpipe_warning: bool = True) -> dict:
        """提取并规范化有压管道弹窗上下文。"""
        pipe_groups = list(self._extract_pressure_pipe_dialog_groups(nodes, settings=settings) or [])
        pressure_chains = list(self._extract_pressure_pipe_dialog_chains(nodes, settings=settings) or [])
        pressure_routes = list(self._extract_pressure_pipe_routes(nodes, settings=settings) or [])
        chain_descriptors = self._build_pressure_pipe_chain_descriptors(pressure_chains)

        channel_level = self._get_current_channel_level_text(settings)
        xxpipe_route_mode = self._should_enable_pressure_pipe_route_mode(
            channel_level,
            pipe_groups,
            pressure_chains,
        )
        if not xxpipe_route_mode:
            return {
                "pipe_groups": pipe_groups,
                "chain_descriptors": chain_descriptors,
                "pressure_routes": pressure_routes,
                "xxpipe_route_mode": False,
                "route_import_targets": {},
                "blocked_route_names": [],
            }

        route_map = self._collect_xxpipe_route_context_map(nodes, pipe_groups)
        station_prefix = self._get_settings_station_prefix(settings)
        route_import_targets = {
            route_key: {
                "display_name": str(meta.get("display_name", "") or route_key).strip(),
                "station_prefix": station_prefix,
                "import_anchor_station_mc": meta.get("import_anchor_station_mc"),
                "targets": list(meta.get("targets", []) or []),
                "nodes": list(meta.get("nodes", []) or []),
            }
            for route_key, meta in route_map.items()
        }
        return {
            "pipe_groups": pipe_groups,
            "chain_descriptors": chain_descriptors,
            "pressure_routes": pressure_routes,
            "xxpipe_route_mode": True,
            "route_import_targets": route_import_targets,
            "blocked_route_names": [],
        }

    @staticmethod
    def _is_supported_continuous_pressure_chain(chain) -> bool:
        """判断当前承压链是否满足连续承压整线入口条件。"""
        members = list(getattr(chain, "members", []) or [])
        return len(members) >= 2

    @classmethod
    def _should_enable_pressure_pipe_route_mode(cls, channel_level: str | None, pipe_groups, pressure_chains) -> bool:
        """统一判断是否启用连续承压整线入口。"""
        if cls._is_xxpipe_channel_level_text(channel_level):
            return True
        if not any(cls._get_pressure_pipe_group_route_key(group) for group in (pipe_groups or [])):
            return False
        return any(
            cls._is_supported_continuous_pressure_chain(chain)
            for chain in (pressure_chains or [])
        )

    @staticmethod
    def _find_next_regular_row_index(nodes, start_index: int) -> int:
        """查找当前行之后最近的普通行。"""
        for idx in range(start_index + 1, len(nodes)):
            node = nodes[idx]
            if getattr(node, "is_transition", False):
                continue
            if getattr(node, "is_auto_inserted_channel", False):
                continue
            return idx
        return -1

    @staticmethod
    def _build_pressure_pipe_effective_length_context(nodes, upstream_idx: int, target_idx: int) -> dict:
        """复用表3现有口径，计算匿名段有效长度。"""
        if upstream_idx < 0 or target_idx < 0 or upstream_idx >= len(nodes) or target_idx >= len(nodes):
            return {
                "L_mc": 0.0,
                "transition_length": 0.0,
                "arc1_half": 0.0,
                "arc2_half": 0.0,
                "effective_length": 0.0,
            }
        upstream_node = nodes[upstream_idx]
        target_node = nodes[target_idx]
        transition_length = 0.0
        for idx in range(upstream_idx + 1, target_idx):
            mid_node = nodes[idx]
            if getattr(mid_node, "is_transition", False):
                transition_length += float(getattr(mid_node, "transition_length", 0.0) or 0.0)
        L_mc = float(getattr(target_node, "station_MC", 0.0) or 0.0) - float(
            getattr(upstream_node, "station_MC", 0.0) or 0.0
        )
        arc1_half = float(getattr(upstream_node, "arc_length", 0.0) or 0.0) / 2.0
        arc2_half = float(getattr(target_node, "arc_length", 0.0) or 0.0) / 2.0
        effective_length = max(0.0, L_mc - transition_length - arc1_half - arc2_half)
        return {
            "L_mc": L_mc,
            "transition_length": transition_length,
            "arc1_half": arc1_half,
            "arc2_half": arc2_half,
            "effective_length": effective_length,
        }

    @staticmethod
    def _calc_pressure_pipe_segment_spatial_length(group, longitudinal_nodes) -> float:
        """兼容旧函数名：返回纵断面实长，不再做平纵三维空间合并。"""
        _ = group
        return WaterProfilePanel._calc_pressure_pipe_longitudinal_path_length(longitudinal_nodes)

    def _build_pressure_chain_member_base_record(self, member, group_mode: str) -> dict:
        """构造连续承压链成员记录骨架。"""
        flow_section = str(getattr(member, "flow_section", "") or "").strip()
        display_name = str(getattr(member, "display_name", "") or "").strip() or "未命名成员"
        identity = self._get_pressure_chain_member_identity(member)
        storage_key = str(getattr(member, "storage_key", "") or identity).strip() or identity
        structure_type = str(getattr(member, "structure_type", "") or "").strip()
        member_role = str(getattr(member, "member_role", "") or "").strip()
        return {
            "identity": identity,
            "storage_key": storage_key,
            "display_name": display_name,
            "flow_section": flow_section,
            "name": display_name,
            "structure_type": structure_type,
            "member_role": member_role,
            "group_mode": group_mode,
            "target_row_index": self._coerce_pressure_pipe_row_index(getattr(member, "target_row_index", -1)),
            "upstream_row_index": self._coerce_pressure_pipe_row_index(getattr(member, "upstream_row_index", -1)),
        }

    def _build_pressure_chain_anchor_record(self, member) -> dict:
        """构造连续承压链起点锚点记录。"""
        group_mode = "chain_row_member"
        structure_type = str(getattr(member, "structure_type", "") or "")
        if "隧洞" in structure_type:
            group_mode = "chain_tunnel_member"
        record = self._build_pressure_chain_member_base_record(member, group_mode)
        record.update({
            "status": "success",
            "writeback_enabled": False,
            "Q": None,
            "D": None,
            "material_key": "",
            "resolved_material_key": "",
            "pipe_velocity": None,
            "total_length": None,
            "friction_loss": None,
            "total_bend_loss": None,
            "local_loss": None,
            "inlet_transition_loss": None,
            "outlet_transition_loss": None,
            "total_head_loss": None,
            "note": "链起点锚点，本行不写回",
            "calc_steps": "【连续承压链锚点】\n本成员仅用于确定链起点，不生成本行损失。",
        })
        return record

    @classmethod
    def _is_pressure_pipe_cross_flow_boundary_group(cls, group, nodes) -> bool:
        """判断跨流量段连续整线在新流量段首行形成的单点边界行。"""
        if not cls._is_pressure_pipe_row_segment_group(group):
            return False
        if not cls._get_pressure_pipe_group_route_key(group):
            return False

        segment_start = cls._coerce_pressure_pipe_finite_float(getattr(group, "segment_start_mc", None))
        segment_end = cls._coerce_pressure_pipe_finite_float(getattr(group, "segment_end_mc", None))
        if segment_start is None or segment_end is None:
            return False
        if abs(segment_start - segment_end) > 1e-6:
            return False

        target_idx = cls._coerce_pressure_pipe_row_index(getattr(group, "target_row_index", -1))
        upstream_idx = cls._coerce_pressure_pipe_row_index(getattr(group, "upstream_row_index", -1))
        route_start_idx = cls._coerce_pressure_pipe_row_index(getattr(group, "route_start_row_index", -1))
        if target_idx < 0 or upstream_idx < 0:
            return False
        if route_start_idx >= 0 and target_idx == route_start_idx:
            return False
        if not isinstance(nodes, list):
            return False
        if target_idx >= len(nodes) or upstream_idx >= len(nodes):
            return False

        target_flow_section = str(
            getattr(nodes[target_idx], "flow_section", "") or getattr(group, "flow_section", "") or ""
        ).strip()
        upstream_flow_section = str(getattr(nodes[upstream_idx], "flow_section", "") or "").strip()
        if not target_flow_section or not upstream_flow_section:
            return False
        return target_flow_section != upstream_flow_section

    def _is_pressure_pipe_route_anchor_group(self, group) -> bool:
        """判断 xx管 整线首行是否应仅作为起点锚点。"""
        if not self._is_pressure_pipe_row_segment_group(group):
            return False
        target_idx = self._coerce_pressure_pipe_row_index(getattr(group, "target_row_index", -1))
        route_start_idx = self._coerce_pressure_pipe_row_index(getattr(group, "route_start_row_index", -1))
        upstream_idx = self._coerce_pressure_pipe_row_index(getattr(group, "upstream_row_index", -1))
        return target_idx >= 0 and route_start_idx >= 0 and target_idx == route_start_idx and upstream_idx < 0

    def _build_pressure_pipe_route_anchor_record(self, group) -> dict:
        """构造 xx管 整线首行起点记录，避免误报为失败。"""
        target_idx = self._coerce_pressure_pipe_row_index(getattr(group, "target_row_index", -1))
        upstream_idx = self._coerce_pressure_pipe_row_index(getattr(group, "upstream_row_index", -1))
        display_name = self._get_pressure_pipe_group_display_name(group)
        return {
            "identity": self._build_pressure_pipe_group_identity(group),
            "storage_key": self._get_pressure_pipe_group_storage_key(group),
            "display_name": display_name,
            "flow_section": self._get_pressure_pipe_group_flow_section(group),
            "name": display_name,
            "group_mode": "unnamed_row_segment",
            "status": "success",
            "writeback_enabled": False,
            "data_mode": "起点行",
            "Q": None,
            "D": None,
            "material_key": "",
            "resolved_material_key": "",
            "pipe_velocity": None,
            "total_length": None,
            "friction_loss": None,
            "total_bend_loss": None,
            "local_loss": None,
            "inlet_transition_loss": None,
            "outlet_transition_loss": None,
            "total_head_loss": None,
            "target_row_index": target_idx,
            "upstream_row_index": upstream_idx,
            "note": "整线起点，不计算本行水头损失",
            "calc_steps": "【xx管整线起点】\n本行位于整线起点，仅作为后续线路的起算位置，不生成本行水头损失。",
        }

    @staticmethod
    def _build_transition_length_between_nodes(nodes, upstream_idx: int, target_idx: int) -> float:
        """统计两普通行之间的渐变段总长度。"""
        transition_length = 0.0
        if upstream_idx < 0 or target_idx < 0:
            return transition_length
        for idx in range(upstream_idx + 1, target_idx):
            mid_node = nodes[idx]
            if getattr(mid_node, "is_transition", False):
                transition_length += float(getattr(mid_node, "transition_length", 0.0) or 0.0)
        return transition_length

    @staticmethod
    def _calc_pressure_pipe_longitudinal_path_length(longitudinal_nodes) -> float:
        """按纵断面节点累计前缀段实长。"""
        total_length = 0.0
        prev_chainage = None
        prev_elevation = None
        for item in list(longitudinal_nodes or []):
            if not isinstance(item, dict):
                continue
            chainage = WaterProfilePanel._coerce_pressure_pipe_finite_float(item.get("chainage"))
            elevation = WaterProfilePanel._coerce_pressure_pipe_finite_float(item.get("elevation"))
            if chainage is None or elevation is None:
                continue
            if prev_chainage is not None and prev_elevation is not None:
                total_length += math.hypot(chainage - prev_chainage, elevation - prev_elevation)
            prev_chainage = chainage
            prev_elevation = elevation
        return float(total_length or 0.0)

    @staticmethod
    def _get_pressure_chain_member_base_name(member) -> str:
        """返回链成员未带后缀的基础名称。"""
        return str(
            getattr(member, "base_display_name", "")
            or getattr(member, "display_name", "")
            or ""
        ).strip()

    def _find_pressure_chain_prefix_fallback_member(self, member, chain_members) -> object | None:
        """查找链前缀段可继承参数的同名普通有压后段。"""
        base_name = self._get_pressure_chain_member_base_name(member)
        if not base_name:
            return None
        for other in list(chain_members or []):
            if other is member:
                continue
            if self._is_pressure_chain_anchor_member(other):
                continue
            if self._is_pressure_chain_prefix_member(other):
                continue
            if str(getattr(other, "structure_type", "") or "").strip() != "有压管道":
                continue
            if self._get_pressure_chain_member_base_name(other) != base_name:
                continue
            return other
        return None

    def _resolve_pressure_chain_prefix_member_inputs(self, member, nodes, chain_members) -> dict:
        """解析链前缀段的流量、管径和管材参数。"""
        start_idx = self._coerce_pressure_pipe_row_index(
            getattr(member, "prefix_target_row_index", getattr(member, "target_row_index", -1))
        )
        start_node = nodes[start_idx] if 0 <= start_idx < len(nodes) else None
        section_params = getattr(start_node, "section_params", {}) or {}
        group = getattr(member, "group", None)
        fallback_member = self._find_pressure_chain_prefix_fallback_member(member, chain_members)
        fallback_group = getattr(fallback_member, "group", None) if fallback_member is not None else None
        fallback_node = None
        if fallback_member is not None:
            fallback_row_indices = list(getattr(fallback_member, "row_indices", []) or [])
            fallback_row_index = fallback_row_indices[0] if fallback_row_indices else getattr(fallback_member, "start_row_index", -1)
            fallback_row_index = self._coerce_pressure_pipe_row_index(fallback_row_index)
            if 0 <= fallback_row_index < len(nodes):
                fallback_node = nodes[fallback_row_index]

        flow = self._coerce_pressure_pipe_finite_float(getattr(start_node, "flow", None))
        if flow is None or flow <= 0:
            flow = self._coerce_pressure_pipe_finite_float(getattr(group, "design_flow", None))
        if (flow is None or flow <= 0) and fallback_group is not None:
            flow = self._coerce_pressure_pipe_finite_float(getattr(fallback_group, "design_flow", None))
        if (flow is None or flow <= 0) and fallback_node is not None:
            flow = self._coerce_pressure_pipe_finite_float(getattr(fallback_node, "flow", None))

        diameter = (
            self._coerce_pressure_pipe_finite_float(section_params.get("D"))
            or self._coerce_pressure_pipe_finite_float(section_params.get("直径D"))
        )
        if diameter is None or diameter <= 0:
            diameter = self._coerce_pressure_pipe_finite_float(getattr(group, "diameter", None))
        if (diameter is None or diameter <= 0) and fallback_group is not None:
            diameter = self._coerce_pressure_pipe_finite_float(getattr(fallback_group, "diameter", None))
        if (diameter is None or diameter <= 0) and fallback_node is not None:
            fallback_params = getattr(fallback_node, "section_params", {}) or {}
            diameter = (
                self._coerce_pressure_pipe_finite_float(fallback_params.get("D"))
                or self._coerce_pressure_pipe_finite_float(fallback_params.get("直径D"))
            )

        material_key = str(section_params.get("pipe_material", "") or "").strip()
        if not material_key:
            material_key = str(getattr(group, "material_key", "") or "").strip()
        if not material_key and fallback_group is not None:
            material_key = str(getattr(fallback_group, "material_key", "") or "").strip()
        if not material_key and fallback_node is not None:
            fallback_params = getattr(fallback_node, "section_params", {}) or {}
            material_key = str(fallback_params.get("pipe_material", "") or "").strip()

        return {
            "start_row_index": start_idx,
            "start_node": start_node,
            "flow": float(flow or 0.0),
            "diameter": float(diameter or 0.0),
            "material_key": material_key,
        }

    def _resolve_pressure_chain_prefix_length(self, member, nodes, longitudinal_nodes_dict, route_profile_segments_by_key) -> dict:
        """解析链前缀段长度，优先取纵断面裁切后的实长。"""
        start_idx = self._coerce_pressure_pipe_row_index(
            getattr(member, "prefix_target_row_index", getattr(member, "target_row_index", -1))
        )
        end_idx = self._coerce_pressure_pipe_row_index(
            getattr(member, "prefix_end_row_index", -1)
        )
        if not (0 <= start_idx < len(nodes) and 0 <= end_idx < len(nodes)):
            return {"length": 0.0, "data_mode": "链前缀段", "note": "缺少有效起终点行"}

        start_node = nodes[start_idx]
        end_node = nodes[end_idx]
        start_mc = self._coerce_pressure_pipe_finite_float(getattr(start_node, "station_MC", None))
        end_mc = self._coerce_pressure_pipe_finite_float(getattr(end_node, "station_MC", None))
        if start_mc is None or end_mc is None:
            return {"length": 0.0, "data_mode": "链前缀段", "note": "缺少有效桩号"}

        group = getattr(member, "group", None)
        route_nodes = []
        spatial_note = ""
        if group is not None:
            try:
                route_nodes, _, spatial_note = self._resolve_pressure_pipe_group_longitudinal_nodes(
                    group,
                    longitudinal_nodes_dict or {},
                    route_profile_segments_by_key=route_profile_segments_by_key,
                )
            except Exception as exc:
                spatial_note = str(exc)

        if route_nodes:
            try:
                from utils.pressure_pipe_longitudinal_utils import clip_longitudinal_nodes_to_range

                clipped_nodes = clip_longitudinal_nodes_to_range(route_nodes, float(start_mc), float(end_mc))
                profile_length = self._calc_pressure_pipe_longitudinal_path_length(clipped_nodes)
                if profile_length > 1e-6:
                    return {
                        "length": float(profile_length),
                        "data_mode": "链前缀段（纵断面）",
                        "note": spatial_note,
                    }
            except Exception as exc:
                spatial_note = str(exc)

        effective_context = self._build_pressure_pipe_effective_length_context(nodes, start_idx, end_idx)
        effective_length = float(effective_context.get("effective_length", 0.0) or 0.0)
        if effective_length > 1e-6:
            return {
                "length": effective_length,
                "data_mode": "链前缀段（桩号）",
                "note": spatial_note,
            }

        station_length = abs(float(end_mc - start_mc))
        return {
            "length": station_length,
            "data_mode": "链前缀段（桩号）",
            "note": spatial_note,
        }

    def _build_pressure_chain_prefix_anchor_record(self, member, *, start_row_index: int, target_row_index: int) -> dict:
        """当前缀长度无效时，退回不写回的锚点记录。"""
        record = self._build_pressure_chain_member_base_record(member, "chain_prefix_member")
        record.update({
            "status": "success",
            "writeback_enabled": False,
            "target_row_index": target_row_index,
            "upstream_row_index": start_row_index,
            "Q": None,
            "D": None,
            "material_key": "",
            "resolved_material_key": "",
            "pipe_velocity": None,
            "total_length": None,
            "friction_loss": None,
            "total_bend_loss": None,
            "local_loss": None,
            "inlet_transition_loss": None,
            "outlet_transition_loss": None,
            "total_head_loss": None,
            "note": "前缀段长度无效，已按链起点锚点处理",
            "calc_steps": "【连续承压链前缀段】\n前缀段长度无效，本次不写回水头损失。",
        })
        return record

    def _calculate_pressure_chain_prefix_member_result(
        self,
        member,
        nodes,
        settings,
        *,
        chain_members=None,
        longitudinal_nodes_dict=None,
        route_profile_segments_by_key=None,
    ) -> dict:
        """计算链首前缀段，并把结果写回下一特殊承压段进口行。"""
        base_record = self._build_pressure_chain_member_base_record(member, "chain_prefix_member")
        inputs = self._resolve_pressure_chain_prefix_member_inputs(member, nodes, chain_members or [])
        start_idx = self._coerce_pressure_pipe_row_index(inputs.get("start_row_index", -1))
        target_idx = self._coerce_pressure_pipe_row_index(
            getattr(member, "prefix_end_row_index", -1)
        )
        base_record["target_row_index"] = target_idx
        base_record["upstream_row_index"] = start_idx

        if (
            start_idx < 0
            or target_idx < 0
            or start_idx >= len(nodes)
            or target_idx >= len(nodes)
        ):
            return {
                **base_record,
                "status": "failed",
                "writeback_enabled": False,
                "error": f"{base_record['display_name']}: 缺少有效起终点行",
            }

        length_context = self._resolve_pressure_chain_prefix_length(
            member,
            nodes,
            longitudinal_nodes_dict or {},
            route_profile_segments_by_key or {},
        )
        total_length = float(length_context.get("length", 0.0) or 0.0)
        if total_length <= 1e-6:
            return self._build_pressure_chain_prefix_anchor_record(
                member,
                start_row_index=start_idx,
                target_row_index=target_idx,
            )

        Q = float(inputs.get("flow", 0.0) or 0.0)
        D = float(inputs.get("diameter", 0.0) or 0.0)
        raw_material_key = str(inputs.get("material_key", "") or "").strip()
        if Q <= 0 or D <= 0 or not raw_material_key:
            issues = []
            if Q <= 0:
                issues.append("设计流量无效")
            if D <= 0:
                issues.append("管径无效")
            if not raw_material_key:
                issues.append("未指定管材")
            return {
                **base_record,
                "status": "failed",
                "writeback_enabled": False,
                "error": f"{base_record['display_name']}: " + "，".join(issues),
            }

        from core.pressure_pipe_calc import PIPE_MATERIALS, calc_friction_loss, calc_pipe_velocity
        from utils.pressure_pipe_common import resolve_pressure_pipe_material

        default_material = "预应力钢筒混凝土管"
        material_info = resolve_pressure_pipe_material(
            raw_material_key,
            PIPE_MATERIALS,
            default_material=default_material,
        )
        material_key = str(material_info.get("canonical_key", default_material) or default_material)
        display_material = str(material_info.get("display_value", raw_material_key or material_key) or material_key)
        note_parts = []
        if raw_material_key and bool(material_info.get("used_default")):
            note_parts.append(f"未识别管材\"{raw_material_key}\"，已按\"{default_material}\"计算")
        if str(length_context.get("note", "") or "").strip():
            note_parts.append(str(length_context.get("note", "") or "").strip())

        friction_loss, friction_details = calc_friction_loss(Q, D, total_length, material_key)
        pipe_velocity = calc_pipe_velocity(Q, D)
        target_node = nodes[target_idx]
        start_node = nodes[start_idx]
        note_parts.append(f"结果已写入{str(getattr(target_node, 'name', '') or f'第{target_idx + 1}行').strip()}进口行")

        calc_steps = [
            "【连续承压链前缀段计算】",
            f"对象: {base_record['display_name']}",
            f"起点行: 第{start_idx + 1}行",
            f"落点行: 第{target_idx + 1}行",
            f"长度口径: {length_context.get('data_mode', '链前缀段')}",
            f"起点名称: {str(getattr(start_node, 'name', '') or '-').strip() or '-'}",
            f"落点名称: {str(getattr(target_node, 'name', '') or '-').strip() or '-'}",
            f"Q = {Q:.4f} m³/s",
            f"D = {D:.4f} m",
            f"L = {total_length:.4f} m",
            f"沿程损失 hf = {float(friction_loss or 0.0):.4f} m",
            "本段按前缀段仅计沿程损失，不额外计渐变、弯头和接头局部损失。",
        ]

        return {
            **base_record,
            "status": "success",
            "writeback_enabled": True,
            "target_row_index": target_idx,
            "upstream_row_index": start_idx,
            "Q": Q,
            "D": D,
            "material_key": display_material,
            "resolved_material_key": material_key,
            "pipe_velocity": float(pipe_velocity or 0.0),
            "total_length": total_length,
            "friction_loss": float(friction_loss or 0.0),
            "total_bend_loss": 0.0,
            "local_loss": 0.0,
            "inlet_transition_loss": 0.0,
            "outlet_transition_loss": 0.0,
            "total_head_loss": float(friction_loss or 0.0),
            "friction_details": copy.deepcopy(friction_details or {}),
            "bend_details": {},
            "local_details": {
                "method": "chain_prefix_member",
                "hj": 0.0,
            },
            "data_mode": str(length_context.get("data_mode", "") or "链前缀段"),
            "note": "；".join(item for item in note_parts if item),
            "calc_steps": "\n".join(calc_steps),
        }

    @staticmethod
    def _rewrite_pressure_chain_member_error_header(error_text, display_name: str) -> str:
        """把逐段成员失败原因的抬头改成当前链成员名称。"""
        member_name = str(display_name or "").strip() or "未命名成员"
        text = str(error_text or "").strip()
        if not text:
            return f"{member_name}: 计算失败"

        separators = [idx for idx in (text.find(":"), text.find("：")) if idx >= 0]
        if not separators:
            return f"{member_name}: {text}"

        separator_index = min(separators)
        reason = text[separator_index + 1:].strip()
        if not reason:
            return f"{member_name}: 计算失败"
        return f"{member_name}: {reason}"

    def _calculate_pressure_chain_single_row_member_result(self, member, nodes, settings) -> dict:
        """计算连续承压链中的单行成员，逐段承压成员沿用正式承压口径。"""
        structure_type = str(getattr(member, "structure_type", "") or "")
        group_mode = "chain_tunnel_member" if "隧洞" in structure_type else "chain_row_member"
        base_record = self._build_pressure_chain_member_base_record(member, group_mode)
        member_group = getattr(member, "group", None)
        member_group_mode = str(getattr(member_group, "group_mode", "") or "").strip()

        if self._is_pressure_chain_anchor_member(member):
            return self._build_pressure_chain_anchor_record(member)

        if member_group is not None and member_group_mode in {"named_row_segment", "unnamed_row_segment"}:
            longitudinal_nodes_dict = getattr(self, "_pressure_pipe_calc_longitudinal_nodes_dict", {}) or {}
            route_profile_segments_by_key = (
                getattr(self, "_pressure_pipe_calc_route_profile_segments_by_key", {}) or {}
            )
            try:
                _, pipe_long_nodes, spatial_fallback_reason = self._resolve_pressure_pipe_group_longitudinal_nodes(
                    member_group,
                    longitudinal_nodes_dict,
                    route_profile_segments_by_key=route_profile_segments_by_key,
                )
            except Exception as exc:
                pipe_long_nodes = []
                spatial_fallback_reason = str(exc)

            record = self._calculate_unnamed_pressure_pipe_group_result(
                member_group,
                nodes,
                pipe_long_nodes,
                spatial_fallback_reason=spatial_fallback_reason,
            )
            record.update({
                "identity": base_record["identity"],
                "storage_key": base_record["storage_key"],
                "display_name": base_record["display_name"],
                "flow_section": base_record["flow_section"],
                "name": base_record["name"],
                "structure_type": base_record["structure_type"],
                "member_role": base_record["member_role"],
                "group_mode": group_mode,
                "target_row_index": base_record["target_row_index"],
                "upstream_row_index": base_record["upstream_row_index"],
            })
            if record.get("status") != "success":
                record["writeback_enabled"] = False
                if member_group_mode == "named_row_segment":
                    record["error"] = self._rewrite_pressure_chain_member_error_header(
                        record.get("error", ""),
                        base_record["display_name"],
                    )
            return record

        target_idx = self._coerce_pressure_pipe_row_index(getattr(member, "target_row_index", -1))
        upstream_idx = self._coerce_pressure_pipe_row_index(getattr(member, "upstream_row_index", -1))
        if (
            upstream_idx < 0
            or target_idx < 0
            or upstream_idx >= len(nodes)
            or target_idx >= len(nodes)
        ):
            return {
                **base_record,
                "status": "failed",
                "writeback_enabled": False,
                "error": f"{base_record['display_name']}: 缺少有效上下游普通行",
            }

        from core.hydraulic_calc import HydraulicCalculator

        upstream_node = copy.deepcopy(nodes[upstream_idx])
        target_node = copy.deepcopy(nodes[target_idx])
        calc = HydraulicCalculator(settings)
        for node in (upstream_node, target_node):
            try:
                calc.fill_section_params(node)
            except Exception:
                pass

        if float(getattr(upstream_node, "velocity", 0.0) or 0.0) <= 0:
            try:
                upstream_node.velocity = calc.calculate_velocity(upstream_node)
            except Exception:
                upstream_node.velocity = float(getattr(upstream_node, "velocity", 0.0) or 0.0)
        if float(getattr(target_node, "velocity", 0.0) or 0.0) <= 0:
            try:
                target_node.velocity = calc.calculate_velocity(target_node)
            except Exception:
                target_node.velocity = float(getattr(target_node, "velocity", 0.0) or 0.0)

        transition_length = self._build_transition_length_between_nodes(nodes, upstream_idx, target_idx)
        friction_loss = float(calc.calculate_friction_loss(upstream_node, target_node, transition_length) or 0.0)
        bend_loss = float(calc.calculate_bend_loss(target_node) or 0.0)
        local_loss = float(calc.calculate_local_loss(target_node) or 0.0)
        total_head_loss = friction_loss + bend_loss + local_loss
        note = f"按{structure_type or '原结构'}原有规则计算"
        section_params = getattr(target_node, "section_params", {}) or {}
        diameter = section_params.get("D", 0.0) or section_params.get("直径D", 0.0) or 0.0

        return {
            **base_record,
            "status": "success",
            "writeback_enabled": True,
            "Q": float(getattr(target_node, "flow", 0.0) or 0.0),
            "D": float(diameter or 0.0),
            "material_key": str(section_params.get("pipe_material", "") or ""),
            "pipe_velocity": float(getattr(target_node, "velocity", 0.0) or 0.0),
            "total_length": float(getattr(target_node, "station_MC", 0.0) or 0.0) - float(
                getattr(upstream_node, "station_MC", 0.0) or 0.0
            ),
            "friction_loss": friction_loss,
            "total_bend_loss": bend_loss,
            "local_loss": local_loss,
            "inlet_transition_loss": 0.0,
            "outlet_transition_loss": 0.0,
            "total_head_loss": total_head_loss,
            "friction_details": copy.deepcopy(getattr(target_node, "friction_calc_details", {}) or {}),
            "bend_details": copy.deepcopy(getattr(target_node, "bend_calc_details", {}) or {}),
            "local_details": {
                "method": "existing_structure_rule",
                "structure_type": structure_type,
                "hj": local_loss,
            },
            "data_mode": "链成员模式",
            "note": note,
            "calc_steps": (
                "【连续承压链成员计算】\n"
                f"对象: {base_record['display_name']}\n"
                f"结构: {structure_type or '-'}\n"
                f"沿程损失 hf = {friction_loss:.4f} m\n"
                f"弯头损失 hw = {bend_loss:.4f} m\n"
                f"局部损失 hj = {local_loss:.4f} m\n"
                f"总损失 ΔH = {total_head_loss:.4f} m"
            ),
        }

    def _calculate_unnamed_pressure_pipe_group_result(self, group, nodes, longitudinal_nodes, spatial_fallback_reason: str = ""):
        """计算逐段承压成员专项结果。"""
        from core.pressure_pipe_calc import (
            PIPE_MATERIALS,
            calc_bend_local_loss,
            calc_fold_local_loss,
            calc_friction_loss,
            calc_pipe_velocity,
            calc_transition_loss,
        )
        from utils.pressure_pipe_common import resolve_pressure_pipe_material

        flow_section = self._get_pressure_pipe_group_flow_section(group)
        display_name = self._get_pressure_pipe_group_display_name(group)
        storage_key = self._get_pressure_pipe_group_storage_key(group)
        identity = self._build_pressure_pipe_group_identity(group)
        group_mode = str(getattr(group, "group_mode", "") or "unnamed_row_segment").strip()
        segment_label = "逐段承压成员" if group_mode == "named_row_segment" else "匿名有压管道段"
        base_record = {
            "identity": identity,
            "storage_key": storage_key,
            "display_name": display_name,
            "flow_section": flow_section,
            "name": display_name,
            "group_mode": group_mode,
        }

        if not group.is_valid():
            return {
                **base_record,
                "status": "failed",
                "error": group.get_validation_message() or f"{display_name}: 数据不完整，已跳过",
            }

        target_idx = self._coerce_pressure_pipe_row_index(getattr(group, "target_row_index", -1))
        upstream_idx = self._coerce_pressure_pipe_row_index(getattr(group, "upstream_row_index", -1))
        target_node = nodes[target_idx]
        next_regular_idx = self._find_next_regular_row_index(nodes, target_idx)
        next_regular_node = nodes[next_regular_idx] if 0 <= next_regular_idx < len(nodes) else None

        default_material = "预应力钢筒混凝土管"
        raw_material_key = str(getattr(group, "material_key", "") or "").strip()
        material_info = resolve_pressure_pipe_material(
            raw_material_key,
            PIPE_MATERIALS,
            default_material=default_material,
        )
        material_key = str(material_info.get("canonical_key", default_material) or default_material)
        display_material = str(material_info.get("display_value", raw_material_key or material_key) or material_key)
        note_parts = []
        if raw_material_key and bool(material_info.get("used_default")):
            note_parts.append(f"未识别管材“{raw_material_key}”，已按“{default_material}”计算")

        Q = float(getattr(group, "design_flow", 0.0) or 0.0)
        D = float(getattr(group, "diameter", 0.0) or 0.0)
        V_pipe = calc_pipe_velocity(Q, D)

        length_ctx = self._build_pressure_pipe_effective_length_context(nodes, upstream_idx, target_idx)
        profile_length = self._calc_pressure_pipe_segment_spatial_length(group, longitudinal_nodes)
        total_length = profile_length if profile_length > 0 else float(length_ctx.get("effective_length", 0.0) or 0.0)
        data_mode = "平面+纵断面（独立叠加）" if profile_length > 0 else "仅平面（独立计算）"
        if spatial_fallback_reason and profile_length <= 0:
            note_parts.append(spatial_fallback_reason)

        friction_loss = 0.0
        friction_details = {}
        if total_length > 0:
            friction_loss, friction_details = calc_friction_loss(Q, D, total_length, material_key)
            friction_details = {
                **(friction_details or {}),
                "method": "pressure_pipe_fmb",
                "length_source": "longitudinal_profile" if profile_length > 0 else "effective_length",
                "L_effective": float(length_ctx.get("effective_length", 0.0) or 0.0),
                "L_mc": float(length_ctx.get("L_mc", 0.0) or 0.0),
                "L_transition": float(length_ctx.get("transition_length", 0.0) or 0.0),
                "arc1_half": float(length_ctx.get("arc1_half", 0.0) or 0.0),
                "arc2_half": float(length_ctx.get("arc2_half", 0.0) or 0.0),
                "L_window": total_length,
                "hf": friction_loss,
            }
        else:
            note_parts.append("有效长度为 0，沿程损失按 0 处理")

        bend_loss = 0.0
        bend_details = {}
        bend_loss_note = ""
        turn_angle = float(getattr(target_node, "turn_angle", 0.0) or 0.0)
        turn_radius = float(getattr(target_node, "turn_radius", 0.0) or 0.0)
        if D > 0 and V_pipe > 0 and 0.1 <= turn_angle < 180:
            if turn_radius > 0:
                xi_bend, bend_loss, bend_calc_details = calc_bend_local_loss(
                    D,
                    turn_radius,
                    turn_angle,
                    V_pipe,
                )
                bend_details = {
                    "D_m": D,
                    "turn_radius_m": turn_radius,
                    "turn_angle_deg": turn_angle,
                    "V_m_s": V_pipe,
                    **(bend_calc_details or {}),
                    "method": "pressure_pipe_bend",
                    "xi_bend": xi_bend,
                    "hj": bend_loss,
                    "hw": bend_loss,
                }
            else:
                xi_bend, bend_loss, bend_calc_details = calc_fold_local_loss(turn_angle, V_pipe)
                bend_details = {
                    "D_m": D,
                    "turn_radius_m": turn_radius,
                    "turn_angle_deg": turn_angle,
                    "V_m_s": V_pipe,
                    **(bend_calc_details or {}),
                    "method": "pressure_pipe_fold",
                    "note": "未设置转弯半径，按折管计算",
                    "xi_bend": xi_bend,
                    "hj": bend_loss,
                    "hw": bend_loss,
                }
                bend_loss_note = "（未设置转弯半径，按折管计算）"

        inlet_transition_loss = 0.0
        outlet_transition_loss = 0.0
        local_details = {
            "method": "pressure_pipe_window_transition",
            "inlet": {},
            "outlet": {},
        }

        upstream_node = nodes[upstream_idx] if 0 <= upstream_idx < len(nodes) else None
        upstream_structure = self._get_node_structure_type_text(upstream_node)
        if (
            upstream_node is not None
            and not self._is_pressure_pipe_like_structure_text(upstream_structure)
            and float(getattr(upstream_node, "velocity", 0.0) or 0.0) > 0
            and float(getattr(group, "inlet_transition_zeta", 0.0) or 0.0) > 0
        ):
            inlet_transition_loss, inlet_details = calc_transition_loss(
                V_pipe,
                float(getattr(upstream_node, "velocity", 0.0) or 0.0),
                float(getattr(group, "inlet_transition_zeta", 0.0) or 0.0),
                is_inlet=True,
            )
            local_details["inlet"] = {
                **(inlet_details or {}),
                "form": getattr(group, "inlet_transition_form", "") or "",
                "zeta": float(getattr(group, "inlet_transition_zeta", 0.0) or 0.0),
                "reference_structure": upstream_structure,
            }

        downstream_structure = self._get_node_structure_type_text(next_regular_node)
        if (
            next_regular_node is not None
            and not self._is_pressure_pipe_like_structure_text(downstream_structure)
            and float(getattr(next_regular_node, "velocity", 0.0) or 0.0) > 0
            and float(getattr(group, "outlet_transition_zeta", 0.0) or 0.0) > 0
        ):
            outlet_transition_loss, outlet_details = calc_transition_loss(
                V_pipe,
                float(getattr(next_regular_node, "velocity", 0.0) or 0.0),
                float(getattr(group, "outlet_transition_zeta", 0.0) or 0.0),
                is_inlet=False,
            )
            local_details["outlet"] = {
                **(outlet_details or {}),
                "form": getattr(group, "outlet_transition_form", "") or "",
                "zeta": float(getattr(group, "outlet_transition_zeta", 0.0) or 0.0),
                "reference_structure": downstream_structure,
            }

        local_loss = inlet_transition_loss + outlet_transition_loss
        total_head_loss = friction_loss + bend_loss + local_loss
        note = "；".join([part for part in note_parts if str(part).strip()])
        steps = [
            f"【{segment_label}水头损失计算】",
            f"对象: {display_name}",
            (
                f"Q = {Q:.4f} m³/s, D = {D:.4f} m, 管材 = {display_material}"
                if display_material == material_key
                else f"Q = {Q:.4f} m³/s, D = {D:.4f} m, 管材 = {display_material}（按 {material_key} 计算）"
            ),
            f"长度 = {total_length:.2f} m（{'纵断面实长' if profile_length > 0 else '表3有效长度'}）",
            f"沿程损失 hf = {friction_loss:.4f} m",
            f"弯头损失 hw = {bend_loss:.4f} m{bend_loss_note}",
            f"局部损失 hj = {local_loss:.4f} m（进口 {inlet_transition_loss:.4f} m，出口 {outlet_transition_loss:.4f} m）",
            f"总损失 ΔH = {total_head_loss:.4f} m",
        ]
        if note:
            steps.append(f"备注: {note}")

        return {
            **base_record,
            "status": "success",
            "writeback_enabled": True,
            "Q": Q,
            "D": D,
            "material_key": display_material,
            "resolved_material_key": material_key,
            "pipe_velocity": V_pipe,
            "total_length": total_length,
            "friction_loss": friction_loss,
            "total_bend_loss": bend_loss,
            "local_loss": local_loss,
            "inlet_transition_loss": inlet_transition_loss,
            "outlet_transition_loss": outlet_transition_loss,
            "total_head_loss": total_head_loss,
            "friction_details": friction_details,
            "bend_details": bend_details,
            "local_details": local_details,
            "data_mode": data_mode,
            "target_row_index": target_idx,
            "upstream_row_index": upstream_idx,
            "note": note,
            "calc_steps": "\n".join(steps),
        }

    def _build_pressure_pipe_window_override_payload(self, group, record: dict) -> dict:
        """构造匿名有压管道段窗口覆盖载荷。"""
        payload = {
            "enabled": True,
            "identity": str(record.get("identity", "") or self._build_pressure_pipe_group_identity(group)).strip(),
            "storage_key": str(record.get("storage_key", "") or self._get_pressure_pipe_group_storage_key(group)).strip(),
            "display_name": str(record.get("display_name", "") or self._get_pressure_pipe_group_display_name(group)).strip(),
            "route_key": str(record.get("route_key", "") or self._get_pressure_pipe_group_route_key(group)).strip(),
            "route_display_name": str(
                record.get("route_display_name", "")
                or self._get_pressure_pipe_group_route_display_name(group)
            ).strip(),
            "group_mode": str(record.get("group_mode", "") or "unnamed_row_segment").strip(),
            "data_mode": str(record.get("data_mode", "") or "").strip(),
            "applied_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "calc_steps": str(record.get("calc_steps", "") or "").strip(),
            "target_row_index": self._coerce_pressure_pipe_row_index(
                record.get("target_row_index", getattr(group, "target_row_index", -1))
            ),
            "upstream_row_index": self._coerce_pressure_pipe_row_index(
                record.get("upstream_row_index", getattr(group, "upstream_row_index", -1))
            ),
            "Q": float(record.get("Q", 0.0) or 0.0),
            "D": float(record.get("D", 0.0) or 0.0),
            "total_length": float(record.get("total_length", 0.0) or 0.0),
            "pipe_velocity": float(record.get("pipe_velocity", 0.0) or 0.0),
            "friction_loss": float(record.get("friction_loss", 0.0) or 0.0),
            "total_bend_loss": float(record.get("total_bend_loss", 0.0) or 0.0),
            "local_loss": float(record.get("local_loss", 0.0) or 0.0),
            "inlet_transition_loss": float(record.get("inlet_transition_loss", 0.0) or 0.0),
            "outlet_transition_loss": float(record.get("outlet_transition_loss", 0.0) or 0.0),
            "total_head_loss": float(record.get("total_head_loss", 0.0) or 0.0),
            "friction_details": copy.deepcopy(record.get("friction_details", {}) or {}),
            "bend_details": copy.deepcopy(record.get("bend_details", {}) or {}),
            "local_details": copy.deepcopy(record.get("local_details", {}) or {}),
        }
        return self._normalize_pressure_pipe_window_override(payload)

    def _apply_pressure_pipe_member_result(self, node, group, record: dict) -> bool:
        """将单个链成员或分组结果写回节点。"""
        status = str(record.get("status", "") or "").strip().lower()
        if status != "success":
            return False
        if not bool(record.get("writeback_enabled", True)):
            return False

        head_loss = record.get("total_head_loss")
        if head_loss is None:
            return False

        identity = str(record.get("identity", "") or "").strip()
        group_mode = str(record.get("group_mode", "") or getattr(group, "group_mode", "") or "").strip()
        row_override_modes = self._get_pressure_pipe_row_override_group_modes()

        if group_mode in row_override_modes:
            target_row_index = self._coerce_pressure_pipe_row_index(
                record.get("target_row_index", getattr(group, "target_row_index", -1))
            )
            if target_row_index >= 0 and hasattr(self, "_ensure_pressure_pipe_row_identity"):
                try:
                    self._ensure_pressure_pipe_row_identity(node, target_row_index)
                except Exception:
                    pass
            friction_loss = float(record.get("friction_loss", 0.0) or 0.0)
            bend_loss = float(record.get("total_bend_loss", 0.0) or 0.0)
            local_loss = float(
                record.get(
                    "local_loss",
                    (record.get("inlet_transition_loss", 0.0) or 0.0)
                    + (record.get("outlet_transition_loss", 0.0) or 0.0),
                ) or 0.0
            )
            self._set_pressure_pipe_window_override(
                node,
                self._build_pressure_pipe_window_override_payload(group, record),
            )
            self._set_pressure_pipe_named_group_result(node, None)
            node.head_loss_friction = friction_loss
            node.head_loss_bend = bend_loss
            node.head_loss_local = local_loss
            node.head_loss_siphon = 0.0
            node.external_head_loss = None
            setattr(node, "_pressure_pipe_display_loss", float(head_loss))
            node.head_loss_total = (
                friction_loss
                + bend_loss
                + local_loss
                + float(getattr(node, "head_loss_reserve", 0.0) or 0.0)
                + float(getattr(node, "head_loss_gate", 0.0) or 0.0)
            )
            if isinstance(record.get("friction_details"), dict):
                node.friction_calc_details = copy.deepcopy(record.get("friction_details") or {})
            if isinstance(record.get("bend_details"), dict):
                node.bend_calc_details = copy.deepcopy(record.get("bend_details") or {})
            if isinstance(record.get("local_details"), dict):
                node.transition_calc_details = copy.deepcopy(record.get("local_details") or {})
        else:
            self._set_pressure_pipe_window_override(node, None)
            if self._is_named_pressure_pipe_group_node(node):
                payload = self._build_pressure_pipe_named_group_result_payload(
                    node=node,
                    group=group,
                    record=record,
                    total_head_loss=head_loss,
                )
                self._set_pressure_pipe_named_group_result(node, payload)
                hidden_total = payload.get("total_head_loss", None)
                current_siphon_loss = float(getattr(node, "head_loss_siphon", 0.0) or 0.0)
                if hidden_total is not None and abs(current_siphon_loss - float(hidden_total)) <= 1e-6:
                    node.head_loss_siphon = 0.0
                node.external_head_loss = None
                display_loss = self._get_named_pressure_pipe_outlet_display_loss(node)
                setattr(node, "_pressure_pipe_display_loss", display_loss)
                self._rebuild_named_pressure_pipe_outlet_total_loss(node)
            else:
                self._set_pressure_pipe_named_group_result(node, None)
                node.head_loss_siphon = float(head_loss)
                node.external_head_loss = None
                node.head_loss_total = (
                    (getattr(node, "head_loss_bend", 0.0) or 0.0)
                    + (getattr(node, "head_loss_friction", 0.0) or 0.0)
                    + (getattr(node, "head_loss_local", 0.0) or 0.0)
                    + (getattr(node, "head_loss_reserve", 0.0) or 0.0)
                    + (getattr(node, "head_loss_gate", 0.0) or 0.0)
                    + float(head_loss)
                )

        if identity:
            self._pressure_pipe_calc_done[identity] = True
        return True

    def _apply_pressure_pipe_d_override_payload(self, pipe_groups, d_payload: dict) -> int:
        if not d_payload:
            return 0
        changed = 0
        for group in pipe_groups or []:
            d_val = d_payload.get(self._get_pressure_pipe_group_storage_key(group))
            if d_val is None:
                continue
            try:
                d_num = float(d_val)
            except (TypeError, ValueError):
                continue
            if d_num <= 0:
                continue
            d_text = f"{d_num:.3f}"
            if self._is_pressure_pipe_row_segment_group(group):
                row_indices = [self._coerce_pressure_pipe_row_index(getattr(group, "target_row_index", -1))]
            else:
                row_indices = list(getattr(group, "row_indices", []) or [])
            for row_idx in row_indices:
                if row_idx < 0 or row_idx >= self.node_table.rowCount():
                    continue
                cur_val = self._parse_item_float(self.node_table.item(row_idx, 21))
                if abs(cur_val - d_num) <= 1e-9:
                    continue
                self._set_table_cell_text_preserve_flags(row_idx, 21, d_text)
                changed += 1
        return changed

    def _apply_pressure_pipe_turn_radius_payload(self, pipe_groups, radius_payload: dict) -> dict:
        result = {"changed_cells": 0, "force_groups": 0, "fill_groups": 0}
        if not radius_payload:
            return result
        for group in pipe_groups or []:
            group_name = self._get_pressure_pipe_group_storage_key(group)
            cfg = radius_payload.get(group_name)
            if not isinstance(cfg, dict):
                continue
            turn_r = float(cfg.get("turn_R", 0.0) or 0.0)
            if turn_r <= 0:
                continue
            force_override = bool(cfg.get("force_override", False))
            if force_override:
                result["force_groups"] += 1
            else:
                result["fill_groups"] += 1
            radius_text = f"{turn_r:.2f}"
            if self._is_pressure_pipe_row_segment_group(group):
                row_indices = [self._coerce_pressure_pipe_row_index(getattr(group, "target_row_index", -1))]
            else:
                row_indices = list(getattr(group, "row_indices", []) or [])
            for row_idx in row_indices:
                if row_idx < 0 or row_idx >= self.node_table.rowCount():
                    continue
                cur_text = self._get_table_turn_radius_text(row_idx)
                cur_val = self._parse_item_float(self.node_table.item(row_idx, 7))
                if (not force_override) and cur_text:
                    continue
                if abs(cur_val - turn_r) <= 1e-9:
                    continue
                self._set_table_cell_text_preserve_flags(row_idx, 7, radius_text)
                result["changed_cells"] += 1
        return result

    def _apply_pressure_pipe_dialog_payloads(self, pipe_groups, radius_payload: dict, d_payload: dict) -> dict:
        d_changed = self._apply_pressure_pipe_d_override_payload(pipe_groups, d_payload)
        radius_result = self._apply_pressure_pipe_turn_radius_payload(pipe_groups, radius_payload)
        total_changed = d_changed + int(radius_result.get("changed_cells", 0))
        if total_changed > 0:
            self._apply_table1_source_row_lock_flags()
        return {
            "changed_cells": total_changed,
            "d_changed": d_changed,
            "radius_changed": int(radius_result.get("changed_cells", 0)),
            "force_groups": int(radius_result.get("force_groups", 0)),
            "fill_groups": int(radius_result.get("fill_groups", 0)),
        }

    def _show_pressure_turn_radius_fallback_notice_if_needed(self):
        groups = sorted({str(g).strip() for g in (self._pressure_turn_radius_fallback_groups or set()) if str(g).strip()})
        if not groups:
            return
        preview = " / ".join(groups[:6])
        suffix = "" if len(groups) <= 6 else f" (+{len(groups)} groups)"
        msg = (
            "Pressure-pipe planar R had blank rows; this run used global radius fallback for: "
            f"{preview}{suffix}."
        )
        InfoBar.info(
            "Planar R fallback",
            msg,
            parent=self._info_parent(),
            duration=4800,
            position=InfoBarPosition.TOP,
        )
        self._pressure_turn_radius_fallback_groups = set()
        return
        """
        preview = "、".join(groups[:6])
        suffix = "" if len(groups) <= 6 else f" 等{len(groups)}组"
        msg = f"有压管道平面R存在空值，已在本次计算中回退顶部全局半径：{preview}{suffix}。"
        InfoBar.info(
            "平面R回退提示",
            msg,
            parent=self._info_parent(),
            duration=4800,
            position=InfoBarPosition.TOP,
        )

        """

    def _repair_pressure_pipe_open_state_from_table3(self, nodes=None) -> bool:
        """按当前表3现场状态修正有压管道打开前的旧门禁标记。"""
        source_nodes = nodes if isinstance(nodes, list) else []
        if not source_nodes:
            build_nodes = getattr(self, "_build_nodes_from_table", None)
            if callable(build_nodes):
                try:
                    source_nodes = build_nodes() or []
                except Exception:
                    source_nodes = []
        if not source_nodes:
            return False

        has_pressure_pipe = any(
            self._is_pressure_pipe_like_node(node)
            for node in source_nodes
            if getattr(node, "structure_type", None)
        )
        if not has_pressure_pipe:
            return False

        has_transition_rows = any(getattr(node, "is_transition", False) for node in source_nodes)
        repaired = False

        if has_transition_rows and not bool(getattr(self, "_transition_topology_prepared", False)):
            self._transition_topology_prepared = True
            repaired = True

        topology_ready = has_transition_rows or bool(getattr(self, "_transition_topology_prepared", False))
        if topology_ready and not bool(getattr(self, "_section_sync_ready", False)):
            self._section_sync_ready = True
            repaired = True

        if repaired:
            refresh_controls = getattr(self, "_refresh_pressure_pipe_controls", None)
            if callable(refresh_controls):
                try:
                    refresh_controls()
                except Exception:
                    pass
        return repaired

    def _clear_pressure_pipe_summary_dialog_ref(self, dialog=None):
        """在汇总窗销毁后清空面板侧引用。"""
        current = getattr(self, "_pressure_pipe_summary_dialog", None)
        if dialog is None or current is dialog:
            self._pressure_pipe_summary_dialog = None

    def _close_pressure_pipe_summary_dialog(self, force: bool = False):
        """关闭现有的有压管道结果汇总窗，并回收引用。"""
        dialog = getattr(self, "_pressure_pipe_summary_dialog", None)
        if dialog is None:
            return
        if force and hasattr(dialog, "_confirmed"):
            dialog._confirmed = True
        try:
            dialog.close()
        except Exception:
            pass
        self._clear_pressure_pipe_summary_dialog_ref(dialog)

    def _build_pressure_pipe_apply_target_map(self, cur_groups, chain_descriptors) -> dict:
        """建立 identity 到当前表3目标节点的映射，供本轮结果清理和写回复用。"""
        target_map = {}

        for descriptor in chain_descriptors or []:
            for member in descriptor.get("members", []) or []:
                identity = self._get_pressure_chain_member_identity(member)
                if not identity or identity in target_map:
                    continue
                target_map[identity] = {
                    "target_row_index": self._coerce_pressure_pipe_row_index(
                        getattr(member, "target_row_index", -1)
                    ),
                    "storage_key": str(getattr(member, "storage_key", "") or identity).strip() or identity,
                    "group_mode": str(getattr(member, "group_mode", "") or "").strip(),
                }

        for group in cur_groups or []:
            identity = self._build_pressure_pipe_group_identity(group)
            if not identity or identity in target_map:
                continue
            if self._is_pressure_pipe_row_segment_group(group):
                target_row_index = self._coerce_pressure_pipe_row_index(
                    getattr(group, "target_row_index", -1)
                )
            else:
                target_row_index = self._coerce_pressure_pipe_row_index(
                    getattr(group, "outlet_row_index", -1)
                )
            target_map[identity] = {
                "target_row_index": target_row_index,
                "storage_key": self._get_pressure_pipe_group_storage_key(group),
                "group_mode": str(getattr(group, "group_mode", "") or "").strip(),
            }

        return target_map

    def _clear_pressure_pipe_member_result(
        self,
        node,
        *,
        identity: str = "",
        storage_keys=None,
        clear_hydraulic_components: bool = False,
    ) -> bool:
        """清理当前节点上一轮残留的有压管道结果。"""
        changed = False

        if node is not None:
            if self._get_pressure_pipe_window_override(node):
                changed = True
            if self._get_pressure_pipe_named_group_result(node):
                changed = True

            self._set_pressure_pipe_window_override(node, None)
            self._set_pressure_pipe_named_group_result(node, None)

            display_loss = getattr(node, "_pressure_pipe_display_loss", 0.0)
            try:
                display_loss = float(display_loss or 0.0)
            except (TypeError, ValueError):
                display_loss = 0.0
            if abs(display_loss) > 1e-9:
                changed = True
            setattr(node, "_pressure_pipe_display_loss", 0.0)

            if getattr(node, "external_head_loss", None) is not None:
                changed = True
            node.external_head_loss = None

            siphon_loss = float(getattr(node, "head_loss_siphon", 0.0) or 0.0)
            if abs(siphon_loss) > 1e-9:
                changed = True
            node.head_loss_siphon = 0.0

            if clear_hydraulic_components:
                for attr_name in ("head_loss_friction", "head_loss_bend", "head_loss_local"):
                    current_value = float(getattr(node, attr_name, 0.0) or 0.0)
                    if abs(current_value) > 1e-9:
                        changed = True
                    setattr(node, attr_name, 0.0)
                for detail_attr in ("friction_calc_details", "bend_calc_details", "transition_calc_details"):
                    if getattr(node, detail_attr, None):
                        changed = True
                    setattr(node, detail_attr, {})

            total_loss = (
                float(getattr(node, "head_loss_friction", 0.0) or 0.0)
                + float(getattr(node, "head_loss_bend", 0.0) or 0.0)
                + float(getattr(node, "head_loss_local", 0.0) or 0.0)
                + float(getattr(node, "head_loss_reserve", 0.0) or 0.0)
                + float(getattr(node, "head_loss_gate", 0.0) or 0.0)
            )
            current_total = float(getattr(node, "head_loss_total", 0.0) or 0.0)
            if abs(current_total - total_loss) > 1e-9:
                changed = True
            node.head_loss_total = total_loss

        cleanup_keys = set()
        if identity:
            cleanup_keys.add(str(identity).strip())
        for key in list(storage_keys or []):
            text = str(key or "").strip()
            if text:
                cleanup_keys.add(text)

        for key in cleanup_keys:
            if key in self._pressure_pipe_calc_done:
                self._pressure_pipe_calc_done.pop(key, None)
                changed = True

        manager = getattr(self, "_pressure_pipe_manager", None)
        remove_fn = getattr(manager, "remove_pipe", None)
        if callable(remove_fn):
            for key in cleanup_keys:
                try:
                    remove_fn(key)
                    changed = True
                except Exception:
                    continue

        return changed

    def _apply_pressure_pipe_results(self, results_by_identity: dict, batch_data: dict):
        """将有压管道计算结果回写到表格"""
        try:
            settings = self._build_settings()
            cur_nodes = self._build_nodes_from_table()
            cur_groups = self._extract_pressure_pipe_dialog_groups(cur_nodes, settings=settings)
            cur_chains = self._extract_pressure_pipe_dialog_chains(cur_nodes, settings=settings)
            chain_descriptors = self._build_pressure_pipe_chain_descriptors(cur_chains)
            batch_records = normalize_pressure_pipe_calc_records(batch_data).get("records", [])
            target_map = self._build_pressure_pipe_apply_target_map(cur_groups, chain_descriptors)
            record_map = {
                str(record.get("identity", "") or "").strip(): record
                for record in batch_records
                if str(record.get("identity", "") or "").strip()
            }
            imported_count = 0
            cleared_count = 0
            handled_identities = set()
            row_override_modes = self._get_pressure_pipe_row_override_group_modes()

            for record in batch_records:
                identity = str(record.get("identity", "") or "").strip()
                if not identity:
                    continue
                target_meta = target_map.get(identity, {})
                target_idx = self._coerce_pressure_pipe_row_index(
                    record.get("target_row_index", target_meta.get("target_row_index", -1))
                )
                target_node = cur_nodes[target_idx] if 0 <= target_idx < len(cur_nodes) else None
                group_mode = str(
                    record.get("group_mode", "") or target_meta.get("group_mode", "") or ""
                ).strip()
                storage_keys = [
                    record.get("storage_key", ""),
                    target_meta.get("storage_key", ""),
                ]
                if self._clear_pressure_pipe_member_result(
                    target_node,
                    identity=identity,
                    storage_keys=storage_keys,
                    clear_hydraulic_components=group_mode in row_override_modes,
                ):
                    cleared_count += 1

            for descriptor in chain_descriptors:
                for member in descriptor.get("members", []) or []:
                    identity = self._get_pressure_chain_member_identity(member)
                    if not identity or identity in handled_identities:
                        continue
                    record = results_by_identity.get(identity) or record_map.get(identity)
                    if not record:
                        continue
                    handled_identities.add(identity)
                    target_idx = self._coerce_pressure_pipe_row_index(
                        record.get("target_row_index", getattr(member, "target_row_index", -1))
                    )
                    if not (0 <= target_idx < len(cur_nodes)):
                        continue
                    _nd = cur_nodes[target_idx]
                    group_mode = str(
                        record.get("group_mode", "") or getattr(member, "group_mode", "") or ""
                    ).strip()
                    if group_mode in self._get_pressure_pipe_row_override_group_modes():
                        self._ensure_pressure_pipe_row_identity(_nd, target_idx)
                    if self._apply_pressure_pipe_member_result(_nd, member, record):
                        imported_count += 1

            for group in cur_groups:
                identity = self._build_pressure_pipe_group_identity(group)
                if not identity or identity in handled_identities:
                    continue
                record = results_by_identity.get(identity) or record_map.get(identity)
                if not record:
                    continue
                handled_identities.add(identity)
                target_idx = (
                    self._coerce_pressure_pipe_row_index(getattr(group, "target_row_index", -1))
                    if self._is_pressure_pipe_row_segment_group(group)
                    else self._coerce_pressure_pipe_row_index(getattr(group, "outlet_row_index", -1))
                )
                if not (0 <= target_idx < len(cur_nodes)):
                    continue
                if self._apply_pressure_pipe_member_result(cur_nodes[target_idx], group, record):
                    imported_count += 1

            if imported_count > 0 or cleared_count > 0:
                self._append_loss_undo_snapshot(self._snapshot_editable_cols())
                _s = self._build_settings()
                _pfx = _s.get_station_prefix() if _s else ""
                self.nodes = cur_nodes
                self._update_table_from_nodes_full(cur_nodes, _pfx)
                auto_resize_table(self.node_table)
                self._recalculate_silent()

            summary = batch_data.get("summary", {})
            success_count = int(summary.get("success", 0))
            failed_count = int(summary.get("failed", 0))

            if success_count <= 0:
                InfoBar.warning(
                    "有压管道计算完成（全部失败）",
                    f"共 {summary.get('total', 0)} 条，全部失败。请查看\"有压管道计算结果汇总\"。",
                    parent=self._info_parent(), duration=7000, position=InfoBarPosition.TOP
                )
            elif failed_count > 0:
                InfoBar.warning(
                    "有压管道计算完成（部分成功）",
                    f"成功 {success_count} 条，失败 {failed_count} 条；已回写 {imported_count} 条到\"倒虹吸/有压管道水头损失\"列。",
                    parent=self._info_parent(), duration=7000, position=InfoBarPosition.TOP
                )
            else:
                InfoBar.success(
                    "有压管道计算完成",
                    f"已完成 {success_count} 条计算并回写 {imported_count} 条到\"倒虹吸/有压管道水头损失\"列。",
                    parent=self._info_parent(), duration=6000, position=InfoBarPosition.TOP
                )
        except Exception as e:
            import traceback
            traceback.print_exc()
            InfoBar.error("错误", f"回写数据失败: {str(e)}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _append_pressure_pipe_calc_details(self, batch_data: dict):
        data = normalize_pressure_pipe_calc_records(batch_data)
        if not data.get("records"):
            return
        token = f"【有压管道计算详情】  时间: {data.get('last_run_at', '-')}"
        old = self.detail_text.toPlainText() if hasattr(self, "detail_text") else ""
        if token in old:
            return
        self.detail_text.setPlainText(append_pressure_pipe_calc_batch_text(old, data, precision=4))

    def _show_pressure_pipe_calc_summary_dialog(self, batch_data: dict, results_by_identity: dict = None):
        data = normalize_pressure_pipe_calc_records(batch_data)
        records = data.get("records", [])
        if not records:
            InfoBar.info("提示", "暂无有压管道计算记录",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return
        record_groups = split_pressure_pipe_records(records)
        official_records = record_groups.get("writeback", [])
        reference_records = record_groups.get("reference", [])
        table_records = [
            rec for rec in records
            if not (
                rec.get("status") == "success"
                and not rec.get("writeback_enabled", True)
            )
        ]
        official_keys = {
            str(value or "")
            for rec in official_records
            for value in (rec.get("identity"), rec.get("storage_key"))
            if str(value or "")
        }
        apply_results_by_identity = {
            str(key): value
            for key, value in (results_by_identity or {}).items()
            if str(key) in official_keys
        }

        self._close_pressure_pipe_summary_dialog(force=True)
        dlg = QWidget()
        dlg.setWindowTitle("有压管道计算结果汇总（请确认是否应用）")
        dlg.setMinimumWidth(980)
        dlg.setMinimumHeight(680)
        dlg.resize(1120, 760)
        dlg.setStyleSheet(DIALOG_STYLE)
        dlg.setWindowFlags(Qt.Window)
        delete_on_close_attr = getattr(Qt, "WA_DeleteOnClose", None)
        if delete_on_close_attr is None and hasattr(Qt, "WidgetAttribute"):
            delete_on_close_attr = getattr(Qt.WidgetAttribute, "WA_DeleteOnClose", None)
        if delete_on_close_attr is not None:
            dlg.setAttribute(delete_on_close_attr, True)
        self._pressure_pipe_summary_dialog = dlg
        dlg.destroyed.connect(
            lambda *_args, _dlg=dlg: self._clear_pressure_pipe_summary_dialog_ref(_dlg)
        )

        from PySide6.QtGui import QIcon
        _res_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "resources")
        for _icon_name in ("logo.ico", "logo.svg"):
            _icon_path = os.path.join(_res_dir, _icon_name)
            if os.path.exists(_icon_path):
                dlg.setWindowIcon(QIcon(_icon_path))
                break

        # 标志：是否通过确认按钮关闭
        dlg._confirmed = False

        def _on_close_event(event):
            if dlg._confirmed:
                event.accept()
                return

            from app_渠系计算前端.styles import fluent_question
            reply = fluent_question(
                dlg,
                "关闭确认",
                "是否将计算结果应用到水面线计算表格？\n\n"
                "点击「是」：应用结果并关闭\n"
                "点击「否」：放弃结果并关闭"
            )
            if reply:
                if apply_results_by_identity:
                    self._apply_pressure_pipe_results(apply_results_by_identity, data)
            event.accept()

        dlg.closeEvent = _on_close_event

        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(18, 14, 18, 14)
        lay.setSpacing(10)

        summary = data.get("summary", {})
        ts = data.get("last_run_at", "") or "-"
        lbl_summary = QLabel(
            f"本次总条数: {summary.get('total', 0)}  |  "
            f"正式计入: {summary.get('writeback_success', len(official_records))}  |  "
            f"参考结果: {summary.get('reference_success', len(reference_records))}  |  "
            f"失败: {summary.get('failed', 0)}  |  "
            f"时间: {ts}"
        )
        lbl_summary.setObjectName("pressurePipeSummaryHeaderLabel")
        lbl_summary.setStyleSheet(f"font-size:13px;font-weight:bold;color:{T1};")
        lay.addWidget(lbl_summary)

        tabs = QTabWidget()
        tabs.setObjectName("pressurePipeSummaryTabs")
        lay.addWidget(tabs, 1)

        result_tab = QWidget()
        result_lay = QVBoxLayout(result_tab)
        result_lay.setContentsMargins(0, 0, 0, 0)
        result_lay.setSpacing(8)

        filter_lay = QHBoxLayout()
        filter_lay.setContentsMargins(0, 0, 0, 0)
        filter_lay.setSpacing(8)
        filter_lay.addWidget(QLabel("搜索"))
        search_edit = LineEdit()
        search_edit.setObjectName("pressurePipeSummarySearchEdit")
        search_edit.setPlaceholderText("输入名称或流量段")
        search_edit.setClearButtonEnabled(True)
        filter_lay.addWidget(search_edit, 1)
        filter_lay.addWidget(QLabel("状态"))
        status_filter = ComboBox()
        status_filter.setObjectName("pressurePipeSummaryStatusFilter")
        status_filter.addItems(["全部", "正式计入", "失败"])
        filter_lay.addWidget(status_filter)
        filter_lay.addStretch()
        result_lay.addLayout(filter_lay)

        reference_tab = QWidget()
        reference_lay = QVBoxLayout(reference_tab)
        reference_lay.setContentsMargins(0, 0, 0, 0)
        reference_lay.setSpacing(8)
        reference_text = QTextEdit()
        reference_text.setObjectName("pressurePipeSummaryReferenceText")
        reference_text.setReadOnly(True)
        reference_text.setFont(QFont("Consolas", 10))
        if reference_records:
            reference_lines = [
                "参考结果仅供人工复核，不计入表3、连续链总损失或累计水损。"
            ]
            if official_records:
                reference_lines.append("同名整组值与逐段合计可能不同，因为计算范围不同；最终以“正式计入”逐段合计为准。")
            reference_lines.append("")
            for idx, rec in enumerate(reference_records, 1):
                reference_lines.append(f"{idx}. {format_pressure_pipe_record_detail(rec, precision=4)}")
                reference_lines.append("")
            reference_text.setPlainText("\n".join(reference_lines).rstrip())
        else:
            reference_text.setPlainText("暂无参考结果。")
        reference_lay.addWidget(reference_text, 1)

        chain_tab = QWidget()
        chain_lay = QVBoxLayout(chain_tab)
        chain_lay.setContentsMargins(0, 0, 0, 0)
        chain_lay.setSpacing(8)

        detail_tab = QWidget()
        detail_lay = QVBoxLayout(detail_tab)
        detail_lay.setContentsMargins(0, 0, 0, 0)
        detail_lay.setSpacing(8)

        chain_summaries = data.get("chain_summaries", []) or []
        chain_summary_box = QTextEdit()
        chain_summary_box.setObjectName("pressurePipeSummaryChainText")
        chain_summary_box.setReadOnly(True)
        chain_summary_box.setFont(QFont("Consolas", 10))
        if chain_summaries:
            chain_summary_box.setPlainText(
                "\n\n".join(
                    format_pressure_pipe_chain_summary(item, precision=4)
                    for item in chain_summaries
                )
            )
        else:
            chain_summary_box.setPlainText("暂无连续承压链总览")
        chain_lay.addWidget(chain_summary_box, 1)

        headers = [
            "查看", "流量段", "名称", "状态", "数据模式", "总损失(m)", "沿程(m)",
            "弯头(m)", "进口渐变(m)", "出口渐变(m)", "备注",
            "下限总损失（m）", "Δ总损(m)"
        ]
        table = QTableWidget(len(table_records), len(headers))
        table.setObjectName("pressurePipeSummaryTable")
        table.setHorizontalHeaderLabels(headers)
        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setFont(QFont("Microsoft YaHei", 10))
        table.setMinimumHeight(360)
        table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        table.setSortingEnabled(False)
        hh = table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.Stretch)
        for col in [3, 4, 5, 6, 7, 8, 9]:
            hh.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(10, QHeaderView.Stretch)
        hh.setSectionResizeMode(11, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(12, QHeaderView.ResizeToContents)

        has_sensitivity_data = any(
            rec.get("sensitivity_low_total_head_loss") is not None for rec in official_records
        )
        show_sensitivity = has_sensitivity_data

        # ---- 对比摘要卡片 ----
        compare_card = QFrame()
        compare_card.setStyleSheet(
            "QFrame#compareCard {"
            "  background: qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #f8fbff, stop:1 #f0f5fc);"
            "  border: 1px solid #c8dff5; border-radius: 6px;"
            "}"
        )
        compare_card.setObjectName("compareCard")
        compare_card_lay = QVBoxLayout(compare_card)
        compare_card_lay.setContentsMargins(14, 10, 14, 10)
        compare_card_lay.setSpacing(8)

        _card_title = QLabel("📊  球墨铸铁管 f 值对比摘要")
        _card_title.setStyleSheet("font-size:13px;font-weight:600;color:#0058a3;background:transparent;")
        compare_card_lay.addWidget(_card_title)

        compare_grid = QGridLayout()
        compare_grid.setSpacing(8)

        _compare_items_data = []
        for rec in official_records:
            if rec.get("status") != "success":
                continue
            main_val = rec.get("total_head_loss")
            low_val = rec.get("sensitivity_low_total_head_loss")
            delta_val = rec.get("sensitivity_delta_total_head_loss")
            if main_val is None or low_val is None:
                continue
            try:
                main_f = float(main_val)
                low_f = float(low_val)
                delta_f = float(delta_val) if delta_val is not None else (low_f - main_f)
                pct = (delta_f / main_f * 100) if main_f != 0 else 0
            except (TypeError, ValueError):
                continue
            _compare_items_data.append({
                "name": rec.get("name", "未命名"),
                "flow_section": rec.get("flow_section", "-"),
                "main": main_f, "low": low_f, "delta": delta_f, "pct": pct,
            })

        _item_style = (
            "QFrame { background: #fff; border: 1px solid #e0eaf5;"
            " border-radius: 4px; }"
        )
        for idx, citem in enumerate(_compare_items_data):
            row_label = QLabel(f"流量段 {citem['flow_section']}  —  {citem['name']}")
            row_label.setStyleSheet("font-size:12px;font-weight:600;color:#333;background:transparent;")
            compare_grid.addWidget(row_label, idx * 2, 0, 1, 3)

            for col, (label_t, val, color) in enumerate([
                (f"主值 (f=223200)", f"{citem['main']:.4f} m", "#1a1a1a"),
                (f"下限 (f=189900)", f"{citem['low']:.4f} m", "#1a1a1a"),
                (f"差值 (下限−主值)", f"{citem['delta']:+.4f} m  ({citem['pct']:+.1f}%)", "#0078d4"),
            ]):
                _f = QFrame()
                _f.setStyleSheet(_item_style)
                _fl = QVBoxLayout(_f)
                _fl.setContentsMargins(10, 6, 10, 6)
                _fl.setSpacing(2)
                _lbl = QLabel(label_t)
                _lbl.setStyleSheet("font-size:11px;color:#666;background:transparent;border:none;")
                _lbl.setAlignment(Qt.AlignCenter)
                _val = QLabel(val)
                _val.setStyleSheet(f"font-size:14px;font-weight:600;color:{color};background:transparent;border:none;")
                _val.setAlignment(Qt.AlignCenter)
                _fl.addWidget(_lbl)
                _fl.addWidget(_val)
                compare_grid.addWidget(_f, idx * 2 + 1, col)

        compare_card_lay.addLayout(compare_grid)

        if not _compare_items_data:
            _no_data = QLabel("暂无对比数据")
            _no_data.setStyleSheet("font-size:12px;color:#999;background:transparent;")
            _no_data.setAlignment(Qt.AlignCenter)
            compare_card_lay.addWidget(_no_data)

        compare_card.setVisible(show_sensitivity)

        def _set_sensitivity_columns_visible(visible: bool):
            table.setColumnHidden(11, not visible)
            table.setColumnHidden(12, not visible)
            compare_card.setVisible(visible)

        _set_sensitivity_columns_visible(show_sensitivity)

        detail_text = QTextEdit()
        detail_text.setObjectName("pressurePipeSummaryDetailText")
        detail_text.setReadOnly(True)
        detail_text.setFont(QFont("Consolas", 10))
        detail_lay.addWidget(detail_text, 1)

        def _fmt(v):
            try:
                return f"{float(v):.4f}"
            except (TypeError, ValueError):
                return "-"

        records_by_identity = {
            str(rec.get("identity", "") or ""): rec
            for rec in records
        }

        def _item(text, rec):
            """创建带记录身份的表格项，便于排序后仍能定位详情。"""
            item = QTableWidgetItem(text)
            item.setData(Qt.UserRole, str(rec.get("identity", "") or ""))
            return item

        def _show_record_detail(rec: dict, switch_to_detail: bool = False):
            detail_text.setPlainText(format_pressure_pipe_record_detail(rec, precision=4))
            # 与下方过程框保持同步（若未写入则追加，已写入则不重复）
            self._append_pressure_pipe_calc_details(data)
            if switch_to_detail:
                tabs.setCurrentWidget(detail_tab)

        def _show_detail_from_row(row: int):
            """按当前表格行打开对应记录详情。"""
            item = table.item(row, 2) or table.item(row, 1)
            if item is None:
                return
            rec = records_by_identity.get(str(item.data(Qt.UserRole) or ""))
            if rec is not None:
                _show_record_detail(rec, switch_to_detail=True)

        def _apply_table_filter():
            """只过滤当前表格视图，不改变底部应用范围。"""
            keyword = search_edit.text().strip().lower()
            status_choice = status_filter.currentText().strip()
            for row in range(table.rowCount()):
                flow_text = table.item(row, 1).text() if table.item(row, 1) else ""
                name_text = table.item(row, 2).text() if table.item(row, 2) else ""
                status_text = table.item(row, 3).text() if table.item(row, 3) else ""
                searchable = f"{flow_text} {name_text}".lower()
                keyword_ok = (not keyword) or (keyword in searchable)
                status_ok = status_choice == "全部" or status_text == status_choice
                table.setRowHidden(row, not (keyword_ok and status_ok))

        for i, rec in enumerate(table_records):
            btn = PushButton("查看详情")
            btn.clicked.connect(lambda checked=False, r=rec: _show_record_detail(r, switch_to_detail=True))
            table.setCellWidget(i, 0, btn)

            table.setItem(i, 1, _item(str(rec.get("flow_section", "") or "-"), rec))
            table.setItem(i, 2, _item(str(rec.get("name", "") or "未命名"), rec))

            status_ok = rec.get("status") == "success"
            status_text = "正式计入" if status_ok else "失败"
            status_color = "#2E7D32" if status_ok else "#C62828"
            status_item = _item(status_text, rec)
            status_item.setForeground(QColor(status_color))
            table.setItem(i, 3, status_item)
            default_mode = "平面模式" if status_ok else "-"
            table.setItem(i, 4, _item(str(rec.get("data_mode", "") or default_mode), rec))

            if status_ok:
                table.setItem(i, 5, _item(_fmt(rec.get("total_head_loss")), rec))
                table.setItem(i, 6, _item(_fmt(rec.get("friction_loss")), rec))
                table.setItem(i, 7, _item(_fmt(rec.get("total_bend_loss")), rec))
                table.setItem(i, 8, _item(_fmt(rec.get("inlet_transition_loss")), rec))
                table.setItem(i, 9, _item(_fmt(rec.get("outlet_transition_loss")), rec))
                note_text = (rec.get("note", "") or "").strip()
                table.setItem(i, 11, _item(_fmt(rec.get("sensitivity_low_total_head_loss")), rec))
                table.setItem(i, 12, _item(_fmt(rec.get("sensitivity_delta_total_head_loss")), rec))
            else:
                for col in [5, 6, 7, 8, 9, 11, 12]:
                    table.setItem(i, col, _item("-", rec))
                note_text = (rec.get("error", "") or "").strip() or "计算失败"
            table.setItem(i, 10, _item(note_text, rec))

        table.cellDoubleClicked.connect(lambda row, _col: _show_detail_from_row(row))
        table.doubleClicked.connect(lambda index: _show_detail_from_row(index.row()))
        search_edit.textChanged.connect(lambda _text: _apply_table_filter())
        status_filter.currentTextChanged.connect(lambda _text: _apply_table_filter())
        table.setSortingEnabled(True)
        result_lay.addWidget(table, 1)
        result_lay.addWidget(compare_card)

        tabs.addTab(result_tab, "结果汇总")
        tabs.addTab(reference_tab, "参考结果")
        tabs.addTab(chain_tab, "连续链总览")
        tabs.addTab(detail_tab, "计算详情")
        tabs.setCurrentWidget(result_tab)

        if table_records:
            _show_record_detail(table_records[0])
        elif reference_records:
            detail_text.setPlainText(format_pressure_pipe_record_detail(reference_records[0], precision=4))

        btn_lay = QHBoxLayout()
        btn_lay.addStretch()

        btn_apply = PrimaryPushButton("应用全部正式结果并关闭")

        def _apply_and_close():
            if not apply_results_by_identity:
                InfoBar.warning("提示", "没有可应用的计算结果",
                               parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 检查是否有成功的计算结果
            has_success = any(
                rec.get("status") == "success"
                and rec.get("writeback_enabled", True)
                and rec.get("total_head_loss") is not None
                for rec in records
            )
            if not has_success:
                InfoBar.warning("提示", "所有计算均失败，无法应用结果",
                               parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            self._apply_pressure_pipe_results(apply_results_by_identity, data)
            dlg._confirmed = True
            dlg.close()

        btn_apply.clicked.connect(_apply_and_close)
        btn_lay.addWidget(btn_apply)

        btn_cancel = PushButton("取消")
        btn_cancel.clicked.connect(dlg.close)
        btn_lay.addWidget(btn_cancel)

        lay.addLayout(btn_lay)
        dlg.show()

    def _open_pressure_pipe_calculator(self):
        """打开有压管道水力计算窗口"""
        debug_print("[DEBUG] _open_pressure_pipe_calculator 被调用")
        nodes = self._build_nodes_from_table()
        self._repair_pressure_pipe_open_state_from_table3(nodes)
        self._close_pressure_pipe_summary_dialog(force=True)
        if not self._ensure_downstream_ready("有压管道水力计算"):
            return
        if not CALCULATOR_AVAILABLE:
            debug_print("[DEBUG] CALCULATOR_AVAILABLE = False，返回")
            InfoBar.error("不可用", "核心计算引擎未加载",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        if not nodes:
            debug_print("[DEBUG] nodes 为空，返回")
            InfoBar.info("提示", "表格中没有数据，请先导入断面参数",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        # 检查是否已插入渐变段
        transition_topology_ready = self._has_transition_topology_ready(nodes)
        if not transition_topology_ready:
            debug_print("[DEBUG] transition_topology_ready = False，返回")
            InfoBar.warning("提示",
                           "请先点击工具栏的【插入渐变段】按钮，完成渐变段插入后再进行有压管道水力计算。\n"
                           "插入渐变段后，系统才能准确获取有压管道上下游流速、断面参数等信息。",
                           parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        # 检查是否有有压管道
        has_ppipe = any(
            self._is_pressure_pipe_like_node(n)
            for n in nodes if n.structure_type
        )
        debug_print(f"[DEBUG] has_ppipe = {has_ppipe}")
        if not has_ppipe:
            debug_print("[DEBUG] has_ppipe = False，返回")
            InfoBar.info("提示", "表格中没有有压管道同类数据，请确保存在结构形式为“有压管道/定向钻/顶管”的行",
                        parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return

        # 先提取pipe_groups和manager
        debug_print("[DEBUG] 开始提取有压管道分组")
        try:
            _water_profile_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                '推求水面线'
            )
            import sys as _sys
            if _water_profile_dir not in _sys.path:
                _sys.path.insert(0, _water_profile_dir)
            settings = self._build_settings()
            dialog_context = self._prepare_pressure_pipe_dialog_context(
                nodes,
                settings=settings,
                show_xxpipe_warning=True,
            )
            pipe_groups = dialog_context["pipe_groups"]
            chain_descriptors = dialog_context["chain_descriptors"]
            pressure_routes = dialog_context.get("pressure_routes", []) or []
            xxpipe_route_mode = bool(dialog_context.get("xxpipe_route_mode"))
            route_import_targets = dict(dialog_context.get("route_import_targets", {}) or {})
            if not pipe_groups and not chain_descriptors:
                if xxpipe_route_mode and dialog_context.get("blocked_route_names"):
                    return
                InfoBar.info("提示", "未找到有压管道数据组",
                            parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 使用共享的 PressurePipeManager（已绑定项目路径）
            if self._pressure_pipe_manager is not None:
                manager = self._pressure_pipe_manager
            else:
                from managers.pressure_pipe_manager import PressurePipeManager
                manager = PressurePipeManager()
                self._pressure_pipe_manager = manager
        except Exception as e:
            InfoBar.error("错误", f"初始化失败: {e}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
            return

        # 弹出配置对话框
        from app_渠系计算前端.water_profile.water_profile_dialogs import PressurePipeConfigDialog
        config_dlg = PressurePipeConfigDialog(
            parent=self,
            pipe_groups=pipe_groups,
            manager=manager,
            pressure_chains=chain_descriptors,
            xxpipe_route_mode=xxpipe_route_mode,
            route_import_targets=route_import_targets,
        )
        if config_dlg.exec() != QDialog.Accepted:
            debug_print("[DEBUG] 用户取消了配置对话框")
            return

        # 获取用户配置
        turn_radius_payload = config_dlg.get_turn_radius_payload() if hasattr(config_dlg, "get_turn_radius_payload") else {}
        d_override_payload = config_dlg.get_d_override_payload() if hasattr(config_dlg, "get_d_override_payload") else {}
        payload_apply_result = self._apply_pressure_pipe_dialog_payloads(
            pipe_groups,
            turn_radius_payload,
            d_override_payload,
        )
        if payload_apply_result.get("changed_cells", 0) > 0:
            self._recalculate_geometry()
        nodes = self._build_nodes_from_table()
        self._show_pressure_turn_radius_fallback_notice_if_needed()
        longitudinal_nodes_dict = config_dlg.get_longitudinal_nodes_dict()
        debug_print(f"[DEBUG] 纵断面数据: {list(longitudinal_nodes_dict.keys())}")

        debug_print("[DEBUG] 开始导入模块和提取有压管道分组")
        try:
            # 导入有压管道相关模块
            _water_profile_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                '推求水面线'
            )
            import sys as _sys
            if _water_profile_dir not in _sys.path:
                _sys.path.insert(0, _water_profile_dir)
            from core.pressure_pipe_calc import calc_total_head_loss, calc_total_head_loss_with_spatial, PIPE_MATERIALS
            from utils.pressure_pipe_common import resolve_pressure_pipe_material

            settings = self._build_settings()
            dialog_context = self._prepare_pressure_pipe_dialog_context(
                nodes,
                settings=settings,
                show_xxpipe_warning=False,
            )
            pipe_groups = dialog_context["pipe_groups"]
            chain_descriptors = dialog_context["chain_descriptors"]
            pressure_routes = dialog_context.get("pressure_routes", []) or []
            if not pipe_groups and not chain_descriptors:
                InfoBar.info("提示", "未找到有压管道数据组",
                            parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
                return

            # 使用共享的 PressurePipeManager（已绑定项目路径）
            if self._pressure_pipe_manager is not None:
                manager = self._pressure_pipe_manager
            else:
                from managers.pressure_pipe_manager import PressurePipeManager
                manager = PressurePipeManager()
                self._pressure_pipe_manager = manager

            self._hydrate_pressure_pipe_groups_from_manager(pipe_groups, manager)

            # 逐条有压管道计算总水头损失并记录完整过程（标准深度）
            results_by_identity = {}
            records = []
            default_material = "预应力钢筒混凝土管"
            run_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            SENSITIVITY_LOW_F = 1.899e5
            chain_member_lookup = {}
            handled_identities = set()
            route_profile_segments_by_key = self._build_pressure_pipe_route_profile_segments(
                pipe_groups,
                longitudinal_nodes_dict,
            )
            self._pressure_pipe_calc_longitudinal_nodes_dict = copy.deepcopy(longitudinal_nodes_dict or {})
            self._pressure_pipe_calc_route_profile_segments_by_key = copy.deepcopy(
                route_profile_segments_by_key or {}
            )
            for descriptor in chain_descriptors:
                for member in descriptor.get("members", []) or []:
                    member_identity = self._get_pressure_chain_member_identity(member)
                    if member_identity:
                        chain_member_lookup[member_identity] = member
            chain_source_lookup = self._build_pressure_chain_source_lookup(chain_descriptors)

            for group in pipe_groups:
                flow_section = self._get_pressure_pipe_group_flow_section(group)
                identity = self._build_pressure_pipe_group_identity(group)
                chain_member = chain_member_lookup.get(identity)
                split_row_members = list(chain_source_lookup.get(identity, []) or [])
                is_split_tail_named_group = bool(split_row_members) or self._is_pressure_pipe_group_split_to_row_members(group)
                pipe_name = (
                    str(getattr(chain_member, "display_name", "") or "").strip()
                    if chain_member is not None
                    else ""
                ) or self._get_pressure_pipe_group_display_name(group)
                storage_key = self._get_pressure_pipe_group_storage_key(group)
                route_key = self._get_pressure_pipe_group_route_key(group)
                route_display_name = self._get_pressure_pipe_group_route_display_name(group)
                route_profile_segments = copy.deepcopy(
                    route_profile_segments_by_key.get(route_key, []) or []
                )
                has_generated_tunnel_profile = any(
                    str(segment.get("source_kind", "") or "").strip() == "generated_tunnel"
                    for segment in route_profile_segments
                )
                if route_key:
                    persist_profile_segments = route_profile_segments if has_generated_tunnel_profile else None
                else:
                    persist_profile_segments = None
                route_long_nodes, pipe_long_nodes, spatial_fallback_reason = (
                    self._resolve_pressure_pipe_group_longitudinal_nodes(
                        group,
                        longitudinal_nodes_dict,
                        route_profile_segments_by_key=route_profile_segments_by_key,
                    )
                )
                base_record = {
                    "identity": identity,
                    "storage_key": storage_key,
                    "display_name": pipe_name,
                    "flow_section": flow_section,
                    "name": pipe_name,
                }
                if (
                    self._is_pressure_pipe_group_tunnel_segment(group)
                    and chain_member is not None
                    and self._is_pressure_chain_single_row_member(chain_member)
                ):
                    continue
                if chain_member is not None and self._is_pressure_chain_prefix_member(chain_member):
                    continue
                handled_identities.add(identity)

                if chain_member is not None and self._is_pressure_chain_anchor_member(chain_member):
                    records.append(self._build_pressure_chain_anchor_record(chain_member))
                    continue

                if self._is_pressure_pipe_row_segment_group(group):
                    if self._is_pressure_pipe_route_anchor_group(group):
                        records.append(self._build_pressure_pipe_route_anchor_record(group))
                        continue
                    record = self._calculate_unnamed_pressure_pipe_group_result(
                        group,
                        nodes,
                        pipe_long_nodes,
                        spatial_fallback_reason=spatial_fallback_reason,
                    )
                    records.append(record)
                    if record.get("status") == "success" and record.get("writeback_enabled", True):
                        results_by_identity[identity] = record
                        route_longitudinal_payload = route_long_nodes if route_long_nodes else None
                        manager.set_result(
                            storage_key,
                            total_head_loss=float(record.get("total_head_loss", 0.0) or 0.0),
                            friction_loss=float(record.get("friction_loss", 0.0) or 0.0),
                            total_bend_loss=float(record.get("total_bend_loss", 0.0) or 0.0),
                            inlet_transition_loss=float(record.get("inlet_transition_loss", 0.0) or 0.0),
                            outlet_transition_loss=float(record.get("outlet_transition_loss", 0.0) or 0.0),
                            pipe_velocity=float(record.get("pipe_velocity", 0.0) or 0.0),
                            plan_total_length=float(record.get("total_length", 0.0) or 0.0),
                            data_mode=str(record.get("data_mode", "") or ""),
                            longitudinal_nodes=route_longitudinal_payload,
                            route_key=route_key,
                            route_display_name=route_display_name,
                            profile_segments=persist_profile_segments,
                            computed_from_profile_source=str(
                                record.get("computed_from_profile_source", "") or ""
                            ).strip() or None,
                        )
                    continue

                if not group.is_valid():
                    msg = group.get_validation_message()
                    records.append({
                        **base_record,
                        "status": "failed",
                        "writeback_enabled": False,
                        "error": msg or f"{pipe_name}: 数据不完整，已跳过",
                    })
                    continue

                raw_material_key = str(getattr(group, "material_key", "") or "").strip()
                material_info = resolve_pressure_pipe_material(
                    raw_material_key,
                    PIPE_MATERIALS,
                    default_material=default_material,
                )
                material_key = str(material_info.get("canonical_key", default_material) or default_material)
                display_material = str(material_info.get("display_value", raw_material_key or material_key) or material_key)
                note = ""
                if raw_material_key and bool(material_info.get("used_default")):
                    note = f"未识别管材\"{raw_material_key}\"，已按\"{default_material}\"计算"
                if spatial_fallback_reason:
                    note = "；".join([item for item in [note, spatial_fallback_reason] if item])

                try:
                    if pipe_long_nodes:
                        # 使用平面/纵断面独立叠加口径计算
                        calc_res = calc_total_head_loss_with_spatial(
                            name=pipe_name,
                            Q=group.design_flow,
                            D=group.diameter,
                            material_key=material_key,
                            ip_points=group.ip_points,
                            longitudinal_nodes=pipe_long_nodes,
                            upstream_velocity=group.upstream_velocity,
                            downstream_velocity=group.downstream_velocity,
                            inlet_transition_form=group.inlet_transition_form,
                            outlet_transition_form=group.outlet_transition_form,
                            inlet_transition_zeta=group.inlet_transition_zeta,
                            outlet_transition_zeta=group.outlet_transition_zeta,
                            has_inlet_transition=group.has_inlet_transition,
                            has_outlet_transition=group.has_outlet_transition,
                            inlet_transition_reason=group.inlet_transition_reason,
                            outlet_transition_reason=group.outlet_transition_reason,
                        )
                    else:
                        # 使用仅平面独立计算口径计算
                        calc_res = calc_total_head_loss(
                            name=pipe_name,
                            Q=group.design_flow,
                            D=group.diameter,
                            material_key=material_key,
                            ip_points=group.ip_points,
                            upstream_velocity=group.upstream_velocity,
                            downstream_velocity=group.downstream_velocity,
                            inlet_transition_form=group.inlet_transition_form,
                            outlet_transition_form=group.outlet_transition_form,
                            inlet_transition_zeta=group.inlet_transition_zeta,
                            outlet_transition_zeta=group.outlet_transition_zeta,
                            has_inlet_transition=group.has_inlet_transition,
                            has_outlet_transition=group.has_outlet_transition,
                            inlet_transition_reason=group.inlet_transition_reason,
                            outlet_transition_reason=group.outlet_transition_reason,
                        )
                except Exception as ex:
                    transition_note = build_pressure_pipe_transition_note(
                        has_inlet_transition=group.has_inlet_transition,
                        inlet_transition_reason=group.inlet_transition_reason,
                        has_outlet_transition=group.has_outlet_transition,
                        outlet_transition_reason=group.outlet_transition_reason,
                    )
                    note_parts = [item for item in [note, transition_note] if item]
                    records.append({
                        **base_record,
                        "status": "failed",
                        "writeback_enabled": False,
                        "Q": group.design_flow,
                        "D": group.diameter,
                        "material_key": display_material,
                        "resolved_material_key": material_key,
                        "error": f"计算失败: {ex}",
                        "note": "；".join(note_parts),
                        "has_inlet_transition": group.has_inlet_transition,
                        "has_outlet_transition": group.has_outlet_transition,
                        "inlet_transition_reason": group.inlet_transition_reason,
                        "outlet_transition_reason": group.outlet_transition_reason,
                    })
                    continue

                transition_note = build_pressure_pipe_transition_note(
                    has_inlet_transition=calc_res.has_inlet_transition,
                    inlet_transition_reason=calc_res.inlet_transition_reason,
                    has_outlet_transition=calc_res.has_outlet_transition,
                    outlet_transition_reason=calc_res.outlet_transition_reason,
                )
                note_parts = [item for item in [note, transition_note] if item]
                record = {
                    **base_record,
                    "status": "success",
                    "writeback_enabled": not is_split_tail_named_group,
                    "Q": group.design_flow,
                    "D": group.diameter,
                    "material_key": display_material,
                    "resolved_material_key": material_key,
                    "total_length": calc_res.total_length,
                    "pipe_velocity": calc_res.pipe_velocity,
                    "friction_loss": calc_res.friction_loss,
                    "total_bend_loss": calc_res.total_bend_loss,
                    "inlet_transition_loss": calc_res.inlet_transition_loss,
                    "outlet_transition_loss": calc_res.outlet_transition_loss,
                    "total_head_loss": calc_res.total_head_loss,
                    "calc_steps": calc_res.calc_steps,
                    "data_mode": calc_res.data_mode,
                    "note": "；".join(note_parts),
                    "has_inlet_transition": calc_res.has_inlet_transition,
                    "has_outlet_transition": calc_res.has_outlet_transition,
                    "inlet_transition_reason": calc_res.inlet_transition_reason,
                    "outlet_transition_reason": calc_res.outlet_transition_reason,
                }
                if is_split_tail_named_group:
                    summary_only_note = "表3按逐段承压成员回写，本整组结果仅用于窗口汇总"
                    record["note"] = "；".join(
                        item for item in [record.get("note", ""), summary_only_note] if item
                    )
                if material_key == "球墨铸铁管":
                    fr = calc_res.friction_details or {}
                    main_f = fr.get("f")
                    q_m3h = fr.get("Q_m3h")
                    d_mm = fr.get("d_mm")
                    m_exp = fr.get("m")
                    b_exp = fr.get("b")
                    low_friction_loss = None
                    try:
                        if all(v is not None for v in [q_m3h, d_mm, m_exp, b_exp]) and float(d_mm) > 0:
                            low_friction_loss = (
                                SENSITIVITY_LOW_F * calc_res.total_length * (float(q_m3h) ** float(m_exp))
                                / (float(d_mm) ** float(b_exp))
                            )
                        elif main_f:
                            low_friction_loss = calc_res.friction_loss * (SENSITIVITY_LOW_F / float(main_f))
                    except Exception:
                        low_friction_loss = None

                    if low_friction_loss is not None:
                        low_total_head_loss = (
                            float(low_friction_loss)
                            + float(calc_res.total_bend_loss)
                            + float(calc_res.inlet_transition_loss)
                            + float(calc_res.outlet_transition_loss)
                        )
                        record.update({
                            "sensitivity_material": "球墨铸铁管",
                            "sensitivity_main_f": main_f,
                            "sensitivity_low_f": SENSITIVITY_LOW_F,
                            "sensitivity_low_friction_loss": low_friction_loss,
                            "sensitivity_low_total_head_loss": low_total_head_loss,
                            "sensitivity_delta_total_head_loss": low_total_head_loss - float(calc_res.total_head_loss),
                        })
                records.append(record)
                if record.get("writeback_enabled", True):
                    results_by_identity[identity] = record

                # 持久化计算结果，便于后续追溯
                route_longitudinal_payload = route_long_nodes if route_long_nodes else None
                manager.set_result(
                    storage_key,
                    total_head_loss=calc_res.total_head_loss,
                    friction_loss=calc_res.friction_loss,
                    total_bend_loss=calc_res.total_bend_loss,
                    inlet_transition_loss=calc_res.inlet_transition_loss,
                    outlet_transition_loss=calc_res.outlet_transition_loss,
                    pipe_velocity=calc_res.pipe_velocity,
                    plan_total_length=calc_res.total_length,
                    data_mode=calc_res.data_mode,
                    longitudinal_nodes=route_longitudinal_payload,
                    route_key=route_key,
                    route_display_name=route_display_name,
                    profile_segments=persist_profile_segments,
                    computed_from_profile_source=str(
                        record.get("computed_from_profile_source", "") or ""
                    ).strip() or None,
                )

            for descriptor in chain_descriptors:
                for member in descriptor.get("members", []) or []:
                    identity = self._get_pressure_chain_member_identity(member)
                    if not identity or identity in handled_identities:
                        continue
                    if self._is_pressure_chain_prefix_member(member):
                        record = self._calculate_pressure_chain_prefix_member_result(
                            member,
                            nodes,
                            settings,
                            chain_members=descriptor.get("members", []) or [],
                            longitudinal_nodes_dict=longitudinal_nodes_dict,
                            route_profile_segments_by_key=route_profile_segments_by_key,
                        )
                    else:
                        record = self._calculate_pressure_chain_single_row_member_result(member, nodes, settings)
                    records.append(record)
                    handled_identities.add(identity)
                    if record.get("status") == "success" and record.get("writeback_enabled", True):
                        results_by_identity[identity] = record

            self._pressure_pipe_calc_longitudinal_nodes_dict = {}
            self._pressure_pipe_calc_route_profile_segments_by_key = {}

            record_map = {
                str(record.get("identity", "") or "").strip(): record
                for record in records
                if str(record.get("identity", "") or "").strip()
            }
            chain_summaries = [
                self._build_pressure_pipe_chain_summary(descriptor, record_map)
                for descriptor in chain_descriptors
            ]

            batch_data = normalize_pressure_pipe_calc_records({
                "last_run_at": run_at,
                "records": records,
                "chain_summaries": chain_summaries,
            })
            self._pressure_pipe_calc_records = batch_data
            self._pressure_pipe_last_run_at = batch_data.get("last_run_at", "")

            save_pressure_routes = getattr(manager, "save_pressure_routes", None)
            if callable(save_pressure_routes):
                try:
                    pressure_routes, route_payloads = self._build_pressure_route_persist_payloads(
                        nodes,
                        settings,
                        longitudinal_nodes_dict or {},
                        route_profile_segments_by_key or {},
                    )
                    route_payloads_by_key = {
                        str(route.get("route_key", "") or "").strip(): route
                        for route in route_payloads
                        if str(route.get("route_key", "") or "").strip()
                    }
                    route_profiles = {
                        route_key: copy.deepcopy((longitudinal_nodes_dict or {}).get(route_key))
                        for route_key in route_payloads_by_key
                        if route_key in (longitudinal_nodes_dict or {})
                    }
                    segment_payloads = self._build_pressure_segment_persist_payloads(
                        pressure_routes,
                        record_map,
                        route_profiles,
                        route_profile_segments_by_key or {},
                        route_payloads_by_key,
                    )
                    save_pressure_routes(
                        route_payloads,
                        route_profiles=route_profiles,
                        segment_results=segment_payloads,
                    )
                except Exception:
                    import traceback
                    traceback.print_exc()

            self._update_pressure_pipe_last_result_button()

            # 追加结构化过程到下方详情框 + 弹出汇总对话框（不立即回写）
            self._append_pressure_pipe_calc_details(batch_data)
            self._show_pressure_pipe_calc_summary_dialog(batch_data, results_by_identity)

        except ImportError as e:
            import traceback
            traceback.print_exc()
            InfoBar.warning("提示", f"有压管道水力计算模块加载失败: {str(e)}",
                           parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)
        except Exception as e:
            import traceback
            traceback.print_exc()
            InfoBar.error("错误", f"打开有压管道计算窗口失败: {str(e)}",
                         parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _open_pressure_pipe_water_hammer_checker(self):
        """打开有压管道水锤验算专用窗口。"""
        enabled, reason = self._resolve_pressure_pipe_water_hammer_button_state()
        if not enabled:
            InfoBar.info(
                "操作已锁定",
                reason or "当前条件暂不支持有压管道水锤验算。",
                parent=self._info_parent(),
                duration=4500,
                position=InfoBarPosition.TOP,
            )
            self._refresh_pressure_pipe_water_hammer_controls()
            return

        if not CALCULATOR_AVAILABLE:
            InfoBar.error(
                "不可用",
                "核心计算引擎未加载",
                parent=self._info_parent(),
                duration=5000,
                position=InfoBarPosition.TOP,
            )
            return

        nodes = self._collect_pressure_pipe_water_hammer_nodes()
        if not nodes:
            InfoBar.info(
                "提示",
                "表格中没有可用于水锤验算的数据，请先执行计算。",
                parent=self._info_parent(),
                duration=3000,
                position=InfoBarPosition.TOP,
            )
            self._refresh_pressure_pipe_water_hammer_controls()
            return

        try:
            settings = self._build_settings()
            dialog_context = self._prepare_pressure_pipe_dialog_context(
                nodes,
                settings=settings,
                show_xxpipe_warning=True,
            )
            pipe_groups = dialog_context["pipe_groups"]
            chain_descriptors = dialog_context["chain_descriptors"]
            xxpipe_route_mode = bool(dialog_context.get("xxpipe_route_mode"))
            route_import_targets = dict(dialog_context.get("route_import_targets", {}) or {})
            if not pipe_groups and not chain_descriptors:
                if xxpipe_route_mode and dialog_context.get("blocked_route_names"):
                    return
                InfoBar.info(
                    "提示",
                    "未找到有压管道数据组，无法进行水锤验算。",
                    parent=self._info_parent(),
                    duration=3000,
                    position=InfoBarPosition.TOP,
                )
                return

            if self._pressure_pipe_manager is not None:
                manager = self._pressure_pipe_manager
            else:
                from managers.pressure_pipe_manager import PressurePipeManager
                manager = PressurePipeManager()
                self._pressure_pipe_manager = manager
        except Exception as e:
            InfoBar.error(
                "错误",
                f"初始化水锤验算窗口失败: {e}",
                parent=self._info_parent(),
                duration=5000,
                position=InfoBarPosition.TOP,
            )
            return

        from app_渠系计算前端.water_profile.water_profile_dialogs import PressurePipeWaterHammerDialog

        config_dlg = PressurePipeWaterHammerDialog(
            parent=self,
            pipe_groups=pipe_groups,
            manager=manager,
            pressure_chains=chain_descriptors,
            xxpipe_route_mode=xxpipe_route_mode,
            route_import_targets=route_import_targets,
        )
        if config_dlg.exec() == QDialog.Accepted:
            self._refresh_pressure_pipe_controls()

    # ================================================================
    # 导出
    # ================================================================
    def _export_excel(self):
        if not self.calculated_nodes:
            InfoBar.warning("提示", "无结果可导出，请先执行计算", parent=self._info_parent(), duration=2000, position=InfoBarPosition.TOP)
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            InfoBar.warning("缺少依赖", "需要: pip install openpyxl", parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            return
        ch_name = self.channel_name_edit.text().strip()
        ch_level = self.channel_level_combo.currentText()
        auto_name = f"{ch_name}{ch_level}_水面线计算结果.xlsx" if ch_name else "水面线计算结果.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(self, "导出Excel", auto_name, "Excel文件 (*.xlsx)")
        if not filepath: return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "水面线计算结果"
            ncols = len(NODE_EXPORT_HEADERS)
            # 第1行：标题
            ws['A1'] = f"{ch_name}{ch_level} 水面线计算结果"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
            ws['A1'].alignment = Alignment(horizontal='center')
            # 第2行：基础信息
            ws['A2'] = "渠道名称"; ws['A2'].font = Font(bold=True); ws['B2'] = ch_name
            ws['C2'] = "渠道级别"; ws['C2'].font = Font(bold=True); ws['D2'] = ch_level
            ws['E2'] = "起始水位(m)"; ws['E2'].font = Font(bold=True); ws['F2'] = self.start_wl_edit.text()
            ws['G2'] = "起始桩号"; ws['G2'].font = Font(bold=True); ws['H2'] = self.start_station_edit.text()
            ws['I2'] = "设计流量"; ws['I2'].font = Font(bold=True); ws['J2'] = self.design_flow_edit.text()
            ws['K2'] = "加大流量"; ws['K2'].font = Font(bold=True); ws['L2'] = self.max_flow_edit.text()
            # 第3行：表头
            hdr_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            for c, h in enumerate(NODE_EXPORT_HEADERS, 1):
                cell = ws.cell(row=3, column=c, value=h)
                cell.font = Font(bold=True)
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
            # 第4行起：从统一表格读取数据
            for r in range(self.node_table.rowCount()):
                for c in range(self.node_table.columnCount()):
                    item = self.node_table.item(r, c)
                    ws.cell(row=r+4, column=c+1, value=item.text() if item else "")

            note_lines = self._build_terminal_gate_backfill_report_lines(self.calculated_nodes)
            if note_lines:
                note_start_row = self.node_table.rowCount() + 6
                ws.cell(row=note_start_row, column=1, value="末尾闸行高程回推说明").font = Font(bold=True)
                ws.merge_cells(
                    start_row=note_start_row,
                    start_column=1,
                    end_row=note_start_row,
                    end_column=ncols,
                )
                for offset, line in enumerate(note_lines[3:], start=1):
                    ws.cell(row=note_start_row + offset, column=1, value=line)
                    ws.merge_cells(
                        start_row=note_start_row + offset,
                        start_column=1,
                        end_row=note_start_row + offset,
                        end_column=ncols,
                    )
            # 自动列宽
            for col_num in range(1, ncols + 1):
                max_len = len(str(NODE_EXPORT_HEADERS[col_num-1]))
                for row_num in range(4, ws.max_row + 1):
                    cv = ws.cell(row=row_num, column=col_num).value
                    if cv: max_len = max(max_len, len(str(cv)))
                ws.column_dimensions[get_column_letter(col_num)].width = min(max_len + 3, 30)
            wb.save(filepath)
            InfoBar.success("导出成功", f"已保存: {filepath}", parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            ask_open_file(filepath, self._info_parent())
        except PermissionError:
            InfoBar.error("文件被占用", f"无法写入文件，请先关闭已打开的同名文件（如Excel等），然后重新操作。", parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("导出失败", str(e), parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _export_word(self):
        if not WORD_EXPORT_AVAILABLE:
            InfoBar.warning("缺少依赖",
                "Word导出需要安装 python-docx、latex2mathml、lxml。请执行: pip install python-docx latex2mathml lxml",
                parent=self._info_parent(), duration=6000, position=InfoBarPosition.TOP)
            return
        if not self.calculated_nodes:
            InfoBar.warning("提示", "请先进行计算。", parent=self._info_parent(), duration=3000, position=InfoBarPosition.TOP)
            return
        ch_name = self.channel_name_edit.text().strip()
        meta = load_meta()
        auto_purpose = build_calc_purpose('water_profile', project=meta.project_name, name=ch_name, section_type='')
        dlg = ExportConfirmDialog('water_profile', '推求水面线计算书', auto_purpose, parent=self._info_parent())
        from PySide6.QtWidgets import QDialog
        if dlg.exec() != QDialog.Accepted:
            return
        self._word_export_meta = dlg.get_meta()
        self._word_export_purpose = dlg.get_calc_purpose()
        self._word_export_refs = dlg.get_references()
        ch_level = self.channel_level_combo.currentText()
        auto_name = f"{ch_name}_水面线计算书.docx" if ch_name else "水面线计算书.docx"
        filepath, _ = QFileDialog.getSaveFileName(self, "保存Word报告", auto_name, "Word文档 (*.docx);;所有文件 (*.*)")
        if not filepath: return
        try:
            self._build_word_report(filepath)
            InfoBar.success("导出成功", f"Word报告已保存到: {filepath}", parent=self._info_parent(), duration=4000, position=InfoBarPosition.TOP)
            ask_open_file(filepath, self._info_parent())
        except PermissionError:
            InfoBar.error("文件被占用", "无法写入文件，请先关闭已打开的同名Word文档，然后重新操作。", parent=self._info_parent(), duration=8000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("导出失败", f"Word导出失败: {str(e)}", parent=self._info_parent(), duration=5000, position=InfoBarPosition.TOP)

    def _build_word_report(self, filepath):
        """构建水面线计算Word报告（工程产品运行卡格式）"""
        settings = self._settings
        nodes = self.calculated_nodes
        ch_name = settings.channel_name if settings else ""
        ch_level = settings.channel_level if settings else ""
        prefix = settings.get_station_prefix() if settings else ""
        meta = getattr(self, '_word_export_meta', load_meta())
        purpose = getattr(self, '_word_export_purpose', '')
        refs = getattr(self, '_word_export_refs', REFERENCES_BASE.get('water_profile', []))

        doc = create_engineering_report_doc(
            meta=meta,
            calc_title='推求水面线计算书',
            calc_content_desc=f'{ch_name}水面线推求计算' if ch_name else '水面线推求计算',
            calc_purpose=purpose,
            references=refs,
            calc_program_text=f'渠系建筑物水力计算系统 V1.0\n推求水面线',
        )
        doc.add_page_break()

        # 5. 基本计算参数
        doc_add_eng_h(doc, '5、基本计算参数')
        params = []
        params.append(("渠道名称", ch_name or "-"))
        params.append(("渠道级别", ch_level or "-"))
        if settings:
            params.append(("起始水位", f"{settings.start_water_level} m"))
            params.append(("起始桩号", ProjectSettings.format_station(settings.start_station, prefix) if settings.start_station else "-"))
            if getattr(settings, 'design_flows', None):
                params.append(("设计流量", ", ".join(f"{q:.3f}" for q in settings.design_flows) + " m³/s"))
            else:
                params.append(("设计流量", f"{settings.design_flow} m³/s"))
            if getattr(settings, 'max_flows', None):
                params.append(("加大流量", ", ".join(f"{q:.3f}" for q in settings.max_flows) + " m³/s"))
            else:
                params.append(("加大流量", f"{settings.max_flow} m³/s"))
            params.append(("糙率", str(settings.roughness)))
            if getattr(settings, 'siphon_roughness', None) is not None:
                params.append(("倒虹吸糙率", str(settings.siphon_roughness)))
            params.append(("转弯半径", self._describe_turn_radius_entry_usage(settings)))
            params.append(("渡槽/隧洞渐变段(进口)", f"{settings.transition_inlet_form}(ζ={settings.transition_inlet_zeta:.2f})"))
            params.append(("渡槽/隧洞渐变段(出口)", f"{settings.transition_outlet_form}(ζ={settings.transition_outlet_zeta:.2f})"))
            params.append(("倒虹吸渐变段(进口)", f"{settings.siphon_transition_inlet_form}(ζ={settings.siphon_transition_inlet_zeta:.2f})"))
            params.append(("倒虹吸渐变段(出口)", f"{settings.siphon_transition_outlet_form}(ζ={settings.siphon_transition_outlet_zeta:.2f})"))
        params.append(("总节点数", str(len(nodes))))
        doc_add_param_table(doc, params)

        # 6. 详细计算过程
        doc_add_eng_h(doc, '6、详细计算过程')
        self._append_pressure_pipe_calc_details(getattr(self, "_pressure_pipe_calc_records", None))
        calc_text = self.detail_text.toPlainText()
        doc_render_calc_text_eng(doc, calc_text, skip_title_keyword='详细计算结果')

        gate_backfill_lines = self._build_terminal_gate_backfill_report_lines(nodes)
        if gate_backfill_lines and "末尾闸行高程回推说明" not in calc_text:
            doc_add_eng_h(doc, '末尾闸行高程回推说明')
            for line in gate_backfill_lines[3:]:
                if line.strip():
                    doc_add_eng_body(doc, line.strip())

        # 7. 建筑物长度汇总
        if hasattr(self, '_last_building_lengths') and self._last_building_lengths:
            doc_add_eng_h(doc, '7、建筑物长度汇总')
            headers = ['序号', '名称', '结构形式', '长度(m)', '起始桩号', '终止桩号']
            data = []
            for i, bl in enumerate(self._last_building_lengths, 1):
                s_s = bl.get('start_station', 0.0)
                s_e = bl.get('end_station', 0.0)
                data.append([
                    str(i),
                    bl.get('name', '-'),
                    bl.get('structure_type', '-'),
                    f"{bl.get('length', 0.0):.3f}",
                    ProjectSettings.format_station(s_s, prefix),
                    ProjectSettings.format_station(s_e, prefix),
                ])
            total_length = sum(bl.get('length', 0.0) for bl in self._last_building_lengths)
            data.append(['合计', '', '', f"{total_length:.3f}", '', ''])
            doc_add_table_caption(doc, '表 1  建筑物长度汇总表')
            doc_add_styled_table(doc, headers, data, with_full_border=True)

        doc.save(filepath)

    # ================================================================
    # CAD 工具
    # ================================================================
    def _cad_combined_dxf(self):
        """导出全部DXF（纵断面+断面汇总+IP表合并）"""
        try:
            from app_渠系计算前端.water_profile.cad_tools import export_combined_dxf
            export_combined_dxf(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from app_渠系计算前端.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"合并DXF导出时发生错误:\n{e}")

    def _cad_longitudinal_profile(self):
        """生成纵断面表格（DXF 文件，也支持 TXT）"""
        try:
            from app_渠系计算前端.water_profile.cad_tools import export_longitudinal_profile_dxf
            export_longitudinal_profile_dxf(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from app_渠系计算前端.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"生成纵断面表格时发生错误:\n{e}")

    def _cad_section_summary(self):
        """生成断面汇总表"""
        try:
            from app_渠系计算前端.water_profile.cad_tools import open_section_summary_table
            open_section_summary_table(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from app_渠系计算前端.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"生成断面汇总表时发生错误:\n{e}")

    def _cad_bzzh2(self):
        """生成bzzh2命令内容（ZDM用）"""
        try:
            from app_渠系计算前端.water_profile.cad_tools import extract_bzzh2_data
            extract_bzzh2_data(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from app_渠系计算前端.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"生成bzzh2命令时发生错误:\n{e}")

    def _cad_building_plan(self):
        """建筑物名称上平面图（AutoCAD -TEXT 命令）"""
        try:
            from app_渠系计算前端.water_profile.cad_tools import export_building_name_plan
            export_building_name_plan(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from app_渠系计算前端.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"生成建筑物平面图时发生错误:\n{e}")

    def _cad_ip_table(self):
        """IP坐标及弯道参数表导出DXF/Excel"""
        try:
            from app_渠系计算前端.water_profile.cad_tools import export_ip_plan_table
            export_ip_plan_table(self)
        except Exception as e:
            import traceback; traceback.print_exc()
            from app_渠系计算前端.styles import fluent_error
            fluent_error(self.window(), "操作失败", f"生成IP坐标表时发生错误:\n{e}")

    def showEvent(self, event):
        super().showEvent(event)
        self._schedule_adjust_splitter_for_settings()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        auto_resize_table(self.node_table)
        width = event.size().width()
        if abs(width - self._last_layout_width) >= 8:
            self._last_layout_width = width
            self._schedule_adjust_splitter_for_settings()

    # ================================================================
    # 项目文件序列化/反序列化（用于 .qxproj 项目保存功能）
    # ================================================================
    def to_project_dict(self) -> dict:
        """
        将水面线推求面板数据序列化为字典
        
        用于 .qxproj 项目文件保存，包含所有设置、节点数据和计算结果。
        
        Returns:
            包含所有数据的字典
        """
        from 推求水面线.models.data_models import ChannelNode, ProjectSettings
        
        # 收集 UI 设置（控件当前值）
        ui_settings = {
            "channel_name": self.channel_name_edit.text().strip(),
            "channel_level": self.channel_level_combo.currentText(),
            "start_water_level": self.start_wl_edit.text().strip(),
            "design_flows_text": self.design_flow_edit.text().strip(),
            "max_flows_text": self.max_flow_edit.text().strip(),
            "start_station_text": self.start_station_edit.text().strip(),
            "roughness": self.roughness_edit.text().strip(),
            "turn_radius": self.turn_radius_edit.text().strip(),
            # 渐变段设置
            "trans_inlet_form": self.trans_inlet_combo.currentText(),
            "trans_inlet_zeta": self.trans_inlet_zeta.text().strip(),
            "trans_outlet_form": self.trans_outlet_combo.currentText(),
            "trans_outlet_zeta": self.trans_outlet_zeta.text().strip(),
            "oc_trans_form": self.oc_trans_combo.currentText(),
            "oc_trans_zeta": self.oc_trans_zeta.text().strip(),
            "siphon_inlet_form": self.siphon_inlet_combo.currentText(),
            "siphon_inlet_zeta": self.siphon_inlet_zeta.text().strip(),
            "siphon_outlet_form": self.siphon_outlet_combo.currentText(),
            "siphon_outlet_zeta": self.siphon_outlet_zeta.text().strip(),
            "length_rule_nudge_seen": bool(getattr(self, "_length_rule_nudge_seen", False)),
        }
        
        # 序列化 ProjectSettings
        project_settings = {}
        current_settings = self._build_settings() if CALCULATOR_AVAILABLE else self._settings
        if current_settings:
            self._settings = current_settings
            project_settings = current_settings.to_dict()
        
        # 序列化节点列表
        nodes_data = []
        for node in self.nodes:
            if hasattr(node, 'to_project_dict'):
                nodes_data.append(node.to_project_dict())
        
        # 序列化计算结果节点列表
        calculated_nodes_data = []
        for node in self.calculated_nodes:
            if hasattr(node, 'to_project_dict'):
                calculated_nodes_data.append(node.to_project_dict())
        
        # 收集额外缓存数据（key 转为字符串，JSON 要求）
        _pp_cache = getattr(self, "_custom_pressurized_pipe_params", {}) or {}
        from app_渠系计算前端.water_profile.cad_tools import _serialize_pressurized_cache_rows

        _pp_siphon = _serialize_pressurized_cache_rows(_pp_cache.get("siphon", []), "siphon")
        _pp_pressure = _serialize_pressurized_cache_rows(_pp_cache.get("pressure_pipe", []), "pressure_pipe")
        extra_caches = {
            "node_structure_heights": {str(k): v for k, v in self._node_structure_heights.items()},
            "node_chamfer_params": {str(k): v for k, v in self._node_chamfer_params.items()},
            "node_u_params": {str(k): v for k, v in self._node_u_params.items()},
            "node_velocity_increased": {str(k): v for k, v in self._node_velocity_increased.items()},
            "text_export_settings": self._text_export_settings.copy(),
            "plan_text_settings": self._plan_text_settings.copy(),
            "custom_pressurized_pipe_params": {
                "siphon": _pp_siphon,
                "pressure_pipe": _pp_pressure,
            },
            "custom_struct_thickness": dict(getattr(self, "_custom_struct_thickness", {}) or {}),
            "custom_rock_lining": dict(getattr(self, "_custom_rock_lining", {}) or {}),
            "custom_tunnel_unified": dict(getattr(self, "_custom_tunnel_unified", {}) or {}),
        }
        
        # 收集倒虹吸糙率数据
        siphon_roughness_data = []
        if hasattr(self.siphon_roughness_chips, '_pairs'):
            siphon_roughness_data = list(self.siphon_roughness_chips._pairs)
        pressure_pipe_roughness_data = []
        if hasattr(self, 'pressure_pipe_roughness_chips') and hasattr(self.pressure_pipe_roughness_chips, '_pairs'):
            pressure_pipe_roughness_data = list(self.pressure_pipe_roughness_chips._pairs)

        # 合并面板附加状态（表1/表2 + 下游门禁）
        section_result_rows = []
        if self._section_result_table:
            for row in range(self._section_result_table.rowCount()):
                row_data = []
                for col in range(self._section_result_table.columnCount()):
                    item = self._section_result_table.item(row, col)
                    row_data.append(item.text() if item else "")
                section_result_rows.append(row_data)
        batch_panel_compat = {}
        if self._batch_backend and hasattr(self._batch_backend, "to_project_dict"):
            try:
                batch_panel_compat = self._batch_backend.to_project_dict()
            except Exception:
                batch_panel_compat = {}
        merged_section = {
            "sync_ready": bool(getattr(self, "_section_sync_ready", False)),
            "transition_topology_prepared": bool(getattr(self, "_transition_topology_prepared", False)),
            "state_text": self._section_state_label.text().strip() if self._section_state_label else "",
            "flow_segments_text": self._section_flow_segments_edit.text().strip() if hasattr(self, "_section_flow_segments_edit") and self._section_flow_segments_edit else "",
            "result_rows": section_result_rows,
            "first_success_auto_jump_done": bool(getattr(self, "_section_first_success_switched", False)),
            "settings_collapsed": bool(self._settings_group.is_collapsed()) if self._settings_group else False,
            "transition_settings_collapsed": bool(self._transition_group.is_collapsed()) if self._transition_group else False,
        }
        
        # 表3原始快照（用于重开后逐单元格一致恢复）
        table3_rows = []
        if self.node_table:
            for row in range(self.node_table.rowCount()):
                row_data = []
                for col in range(self.node_table.columnCount()):
                    item = self.node_table.item(row, col)
                    row_data.append(item.text() if item else "")
                table3_rows.append(row_data)

        return {
            "version": "1.0",
            "ui_settings": ui_settings,
            "project_settings": project_settings,
            "nodes": nodes_data,
            "calculated_nodes": calculated_nodes_data,
            "node_table_rows": table3_rows,
            "extra_caches": extra_caches,
            "siphon_roughness_data": siphon_roughness_data,
            "pressure_pipe_roughness_data": pressure_pipe_roughness_data,
            "pressure_pipe_calc_records": normalize_pressure_pipe_calc_records(
                getattr(self, "_pressure_pipe_calc_records", None)
            ),
            "batch_panel_compat": batch_panel_compat,
            "merged_section": merged_section,
        }
    
    def from_project_dict(self, d: dict, skip_dirty_signal: bool = False):
        """
        从字典恢复水面线推求面板数据
        
        用于 .qxproj 项目文件加载。
        
        Args:
            d: 序列化的字典数据
            skip_dirty_signal: 是否跳过脏状态信号（加载时应为True）
        """
        from 推求水面线.models.data_models import ChannelNode, ProjectSettings
        
        # 设置守卫标志，防止加载时触发脏状态和 cellChanged
        old_updating = self._updating_cells
        old_loading = getattr(self, '_loading_project', False)
        self._updating_cells = True
        self._loading_project = True
        
        try:
            self._clear_transition_length_rule_nudge()
            # 默认重置有压管道计算记录（兼容旧项目缺失字段）
            self._pressure_pipe_calc_records = empty_pressure_pipe_calc_records()
            self._pressure_pipe_last_run_at = ""
            self._section_failure_auto_expanded_once = False
            self._section_failure_records = []

            # 恢复 UI 设置
            ui = d.get("ui_settings", {})
            self._length_rule_nudge_seen = bool(ui.get("length_rule_nudge_seen", False))
            
            if ui.get("channel_name"):
                self.channel_name_edit.setText(ui["channel_name"])
            
            if ui.get("channel_level"):
                idx = self.channel_level_combo.findText(ui["channel_level"])
                if idx >= 0:
                    self.channel_level_combo.setCurrentIndex(idx)
            
            if ui.get("start_water_level"):
                self.start_wl_edit.setText(ui["start_water_level"])
            
            if ui.get("design_flows_text"):
                self.design_flow_edit.setText(ui["design_flows_text"])
            
            if ui.get("max_flows_text"):
                self.max_flow_edit.setText(ui["max_flows_text"])
            
            if ui.get("start_station_text"):
                self.start_station_edit.setText(ui["start_station_text"])
            
            if ui.get("roughness"):
                self.roughness_edit.setText(ui["roughness"])
            
            if "turn_radius" in ui:
                self.turn_radius_edit.setText(ui.get("turn_radius") or "")
            
            # 恢复渐变段设置
            if ui.get("trans_inlet_form"):
                idx = self.trans_inlet_combo.findText(ui["trans_inlet_form"])
                if idx >= 0:
                    self.trans_inlet_combo.setCurrentIndex(idx)
            if ui.get("trans_inlet_zeta"):
                self.trans_inlet_zeta.setText(ui["trans_inlet_zeta"])
            
            if ui.get("trans_outlet_form"):
                idx = self.trans_outlet_combo.findText(ui["trans_outlet_form"])
                if idx >= 0:
                    self.trans_outlet_combo.setCurrentIndex(idx)
            if ui.get("trans_outlet_zeta"):
                self.trans_outlet_zeta.setText(ui["trans_outlet_zeta"])
            
            if ui.get("oc_trans_form"):
                idx = self.oc_trans_combo.findText(ui["oc_trans_form"])
                if idx >= 0:
                    self.oc_trans_combo.setCurrentIndex(idx)
            if ui.get("oc_trans_zeta"):
                self.oc_trans_zeta.setText(ui["oc_trans_zeta"])
            
            if ui.get("siphon_inlet_form"):
                idx = self.siphon_inlet_combo.findText(ui["siphon_inlet_form"])
                if idx >= 0:
                    self.siphon_inlet_combo.setCurrentIndex(idx)
            if ui.get("siphon_inlet_zeta"):
                self.siphon_inlet_zeta.setText(ui["siphon_inlet_zeta"])
            
            if ui.get("siphon_outlet_form"):
                idx = self.siphon_outlet_combo.findText(ui["siphon_outlet_form"])
                if idx >= 0:
                    self.siphon_outlet_combo.setCurrentIndex(idx)
            if ui.get("siphon_outlet_zeta"):
                self.siphon_outlet_zeta.setText(ui["siphon_outlet_zeta"])
            
            # 恢复 ProjectSettings
            proj_settings = d.get("project_settings", {})
            if proj_settings:
                self._settings = ProjectSettings.from_dict(proj_settings)
                self._load_transition_length_rules(
                    getattr(self._settings, "transition_length_rules", []) or []
                )
                design_flows = list(getattr(self._settings, "design_flows", []) or [])
                if not design_flows and getattr(self._settings, "design_flow", 0) > 0:
                    design_flows = [float(self._settings.design_flow)]
                max_flows = list(getattr(self._settings, "max_flows", []) or [])
                if not max_flows and getattr(self._settings, "max_flow", 0) > 0:
                    max_flows = [float(self._settings.max_flow)]

                if not self.design_flow_edit.text().strip() and design_flows:
                    self.design_flow_edit.setText(format_flow_values_text(design_flows))
                if not self.max_flow_edit.text().strip() and max_flows:
                    self.max_flow_edit.setText(format_flow_values_text(max_flows))
            else:
                self._settings = None
                self._load_transition_length_rules([])
            self._sync_flow_segment_widgets(reset_index=True)
            
            # 恢复节点列表
            nodes_data = d.get("nodes", [])
            self.nodes = []
            for nd in nodes_data:
                self.nodes.append(ChannelNode.from_project_dict(nd))
            migrated_named_group_result = self._migrate_named_pressure_pipe_group_results(self.nodes)
            
            # 恢复计算结果节点列表
            calc_nodes_data = d.get("calculated_nodes", [])
            self.calculated_nodes = []
            for nd in calc_nodes_data:
                self.calculated_nodes.append(ChannelNode.from_project_dict(nd))
            if self._migrate_named_pressure_pipe_group_results(self.calculated_nodes):
                migrated_named_group_result = True
            
            # 恢复额外缓存数据（key 从字符串转回 int）
            extra = d.get("extra_caches", {})
            
            struct_heights = extra.get("node_structure_heights", {})
            self._node_structure_heights = {int(k): v for k, v in struct_heights.items()}
            
            chamfer_params = extra.get("node_chamfer_params", {})
            self._node_chamfer_params = {int(k): v for k, v in chamfer_params.items()}
            
            u_params = extra.get("node_u_params", {})
            self._node_u_params = {int(k): v for k, v in u_params.items()}

            vi_params = extra.get("node_velocity_increased", {})
            self._node_velocity_increased = {int(k): v for k, v in vi_params.items()}

            text_settings = extra.get("text_export_settings", {})
            if text_settings:
                self._text_export_settings.update(text_settings)
            
            plan_settings = extra.get("plan_text_settings", {})
            if plan_settings:
                self._plan_text_settings.update(plan_settings)

            # 恢复 CAD 导出复用参数缓存
            pp_cache = extra.get("custom_pressurized_pipe_params", {})
            if isinstance(pp_cache, dict):
                from app_渠系计算前端.water_profile.cad_tools import (
                    _extract_pressurized_param_entities,
                    _merge_pressurized_param_defaults,
                    _serialize_pressurized_cache_rows,
                )

                source_nodes = self.calculated_nodes or self.nodes
                siphon_entities, _ = _extract_pressurized_param_entities(source_nodes, "siphon")
                pressure_entities, _ = _extract_pressurized_param_entities(source_nodes, "pressure_pipe")
                self._custom_pressurized_pipe_params = {
                    "siphon": _serialize_pressurized_cache_rows(
                        _merge_pressurized_param_defaults(siphon_entities, pp_cache.get("siphon", [])),
                        "siphon",
                    ),
                    "pressure_pipe": _serialize_pressurized_cache_rows(
                        _merge_pressurized_param_defaults(pressure_entities, pp_cache.get("pressure_pipe", [])),
                        "pressure_pipe",
                    ),
                }
            else:
                self._custom_pressurized_pipe_params = {"siphon": [], "pressure_pipe": []}

            struct_t = extra.get("custom_struct_thickness", {})
            self._custom_struct_thickness = dict(struct_t) if isinstance(struct_t, dict) else {}
            rock_lining = extra.get("custom_rock_lining", {})
            self._custom_rock_lining = dict(rock_lining) if isinstance(rock_lining, dict) else {}
            tunnel_unified = extra.get("custom_tunnel_unified", {})
            self._custom_tunnel_unified = dict(tunnel_unified) if isinstance(tunnel_unified, dict) else {}
            
            # 恢复倒虹吸糙率数据
            siphon_data = d.get("siphon_roughness_data", [])
            if hasattr(self, 'siphon_roughness_chips'):
                if siphon_data:
                    self.siphon_roughness_chips.set_siphon_data(siphon_data)
                else:
                    self.siphon_roughness_chips.clear()
            ppipe_data = d.get("pressure_pipe_roughness_data", [])
            if hasattr(self, 'pressure_pipe_roughness_chips'):
                if ppipe_data:
                    self.pressure_pipe_roughness_chips.set_pairs(ppipe_data)
                else:
                    self.pressure_pipe_roughness_chips.clear()
            self._refresh_roughness_overview_visibility()

            # 恢复有压管道计算记录
            self._pressure_pipe_calc_records = normalize_pressure_pipe_calc_records(
                d.get("pressure_pipe_calc_records", None)
            )
            self._pressure_pipe_last_run_at = self._pressure_pipe_calc_records.get("last_run_at", "")

            # 恢复合并面板附加状态（表1/表2）
            merged_section = d.get("merged_section", {})
            if not isinstance(merged_section, dict):
                merged_section = {}
            self._section_first_success_switched = bool(
                merged_section.get("first_success_auto_jump_done", False)
            )
            settings_collapsed = bool(merged_section.get("settings_collapsed", False))
            transition_collapsed = bool(merged_section.get("transition_settings_collapsed", False))
            if self._settings_group:
                self._settings_group.set_collapsed(settings_collapsed)
            if self._transition_group:
                self._transition_group.set_collapsed(transition_collapsed)
            batch_panel_compat = d.get("batch_panel_compat", {})
            if not isinstance(batch_panel_compat, dict):
                batch_panel_compat = {}
            if self._batch_backend:
                try:
                    self._batch_backend._clear_input(force=True)
                    self._batch_backend._clear_results()
                except Exception:
                    pass
                if batch_panel_compat and hasattr(self._batch_backend, "from_project_dict"):
                    self._batch_backend.from_project_dict(batch_panel_compat, skip_dirty_signal=True)
            flow_segments_text = str(
                merged_section.get("flow_segments_text", "") or batch_panel_compat.get("flow_segments", "")
            ).strip()
            if flow_segments_text and hasattr(self, "_section_flow_segments_edit") and self._section_flow_segments_edit:
                self._section_flow_segments_edit.setText(flow_segments_text)
            result_rows = merged_section.get("result_rows", [])
            if self._section_result_table and isinstance(result_rows, list) and result_rows:
                with self._table_batch_update(self._section_result_table):
                    self._section_result_table.setRowCount(len(result_rows))
                    for row_idx, row_data in enumerate(result_rows):
                        if not isinstance(row_data, (list, tuple)):
                            continue
                        for col_idx, cell_value in enumerate(row_data):
                            if col_idx >= self._section_result_table.columnCount():
                                break
                            item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                            self._section_result_table.setItem(row_idx, col_idx, item)
            self._refresh_section_failure_feedback(auto_focus=False, auto_expand_once=True)
            
            # 刷新节点表格显示（先按节点模型恢复）
            if self.calculated_nodes:
                # 有计算结果，显示计算后的数据
                self._update_table_from_nodes_full(self.calculated_nodes)
            elif self.nodes:
                # 无计算结果，显示原始节点数据
                self._update_table_from_nodes_full(self.nodes)
            else:
                # 清空表格
                self.node_table.setRowCount(0)

            # 再按原始快照覆盖文本，确保“保存→重开”逐单元格一致
            table3_rows = d.get("node_table_rows", [])
            if isinstance(table3_rows, list) and table3_rows:
                self._apply_node_table_text_snapshot(table3_rows)
            if self.calculated_nodes:
                self._repair_missing_transition_length_details(self.calculated_nodes)
                self._repair_missing_transition_loss_details(self.calculated_nodes)
            self._refresh_all_transition_length_presentations(self.calculated_nodes or self.nodes)
            if migrated_named_group_result and self.node_table.rowCount() > 1:
                self._recalculate_silent()

            # 恢复下游门禁状态（默认锁定）
            state_text = str(merged_section.get("state_text", "")).strip()
            sync_ready, legacy_loaded = self._resolve_loaded_section_gate_state(
                merged_section,
                self.node_table.rowCount(),
                calculated_node_count=len(self.calculated_nodes or []),
                result_row_count=len(result_rows) if isinstance(result_rows, list) else 0,
            )
            self._transition_topology_prepared = bool(
                merged_section.get("transition_topology_prepared", False)
            )
            if sync_ready and self.node_table.rowCount() > 0:
                self._section_sync_ready = True
                self._set_downstream_actions_enabled(
                    True,
                    state_text=state_text or "状态：断面全成功，表1+表2已同步到表3"
                )
            elif legacy_loaded:
                self._section_sync_ready = True
                self._set_downstream_actions_enabled(
                    True,
                    state_text="状态：已加载旧版项目结果，可继续计算（建议先执行断面批量计算）"
                )
            else:
                self._mark_section_results_stale(
                    state_text or "状态：断面结果未就绪，请先执行断面批量计算"
                )
            self._refresh_pressure_pipe_controls()
            self._switch_workspace_tab(self._tab_section_input)
            self._rebuild_calculation_summary_state(self.calculated_nodes or self.nodes)
            QTimer.singleShot(0, self._adjust_splitter_for_settings)
            
        finally:
            self._updating_cells = old_updating
            self._loading_project = old_loading

    def _apply_node_table_text_snapshot(self, table_rows):
        """将保存的表3文本快照回填到当前node_table，保持单元格文本一致。"""
        if not self.node_table or not isinstance(table_rows, list):
            return

        with self._table_batch_update(self.node_table):
            # 若当前行数与快照不一致，按快照调整（尽量保留现有 item 的样式/只读状态）
            self.node_table.setRowCount(len(table_rows))
            col_count = self.node_table.columnCount()
            for row_idx, row_data in enumerate(table_rows):
                if not isinstance(row_data, (list, tuple)):
                    continue
                for col_idx in range(col_count):
                    text = ""
                    if col_idx < len(row_data):
                        cell_value = row_data[col_idx]
                        text = str(cell_value) if cell_value is not None else ""
                    item = self.node_table.item(row_idx, col_idx)
                    if item is None:
                        item = QTableWidgetItem("")
                        item.setTextAlignment(Qt.AlignCenter)
                        if col_idx not in EDITABLE_COLS:
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        if row_idx == 0 and col_idx in FIRST_ROW_LOCKED_LOSS_COLS:
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        self.node_table.setItem(row_idx, col_idx, item)
                    if item.text() != text:
                        item.setText(text)
            self._apply_table1_source_row_lock_flags()
            self._refresh_source_coord_payloads_from_table()
