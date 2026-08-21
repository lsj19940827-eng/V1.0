# -*- coding: utf-8 -*-
"""七一水库充水渠真实 Excel 的专项内核链式水面线回归测试。"""

import copy
import os
import shutil
import sys
import tempfile
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
os.environ.setdefault("QTWEBENGINE_DISABLE_SANDBOX", "1")
os.environ.setdefault(
    "MPLCONFIGDIR",
    str(Path(tempfile.gettempdir()) / "codex-mplconfig"),
)

ROOT = Path(__file__).resolve().parents[1]
WATER_PROFILE_ROOT = ROOT / "推求水面线"
for _path in (str(ROOT), str(WATER_PROFILE_ROOT), str(ROOT / "calc_渠系计算算法内核")):
    if _path not in sys.path:
        sys.path.insert(0, _path)

from PySide6.QtWidgets import QApplication

from app_渠系计算前端.batch.panel import BatchPanel
import app_渠系计算前端.batch.panel as batch_panel_mod
from app_渠系计算前端.water_profile import cad_tools
from core.calculator import WaterProfileCalculator
from core.spillway_steep_chute_adapter import (
    SPILLWAY_STEEP_CHUTE_PARAM_KEY,
    is_spillway_steep_chute_node,
)
from models.data_models import ChannelNode, ProjectSettings
from models.enums import StructureType


QIYI_EXCEL = ROOT / "data" / "七一水库充水渠批量计算用表.xlsx"
FILL_CHANNEL_ALIASES = ("充水渠", "泄水渠", "陡坡", "泄水渠与陡坡")
TAIL_EXCEL_ROWS = (17, 18, 19, 20)


class _TextStub:
    """提供导出函数需要的 text() 接口。"""

    def __init__(self, value):
        self._value = value

    def text(self):
        """返回固定文本。"""
        return self._value


class _ComboStub:
    """提供导出函数需要的 currentText() 接口。"""

    def __init__(self, value):
        self._value = value

    def currentText(self):
        """返回固定下拉值。"""
        return self._value


class _InfoBarStub:
    """屏蔽界面提示，避免测试依赖真实弹窗。"""

    @staticmethod
    def success(*_args, **_kwargs):
        """忽略成功提示。"""
        return None

    @staticmethod
    def warning(*_args, **_kwargs):
        """忽略警告提示。"""
        return None

    @staticmethod
    def error(*_args, **_kwargs):
        """忽略错误提示。"""
        return None

    @staticmethod
    def info(*_args, **_kwargs):
        """忽略普通提示。"""
        return None


class _AcceptedTextDialog:
    """让 DXF 文本导出参数弹窗自动确认。"""

    def __init__(self, *_args, **_kwargs):
        self.result = {}

    def exec(self):
        """模拟用户点击确定。"""
        return cad_tools.QDialog.Accepted


def _get_qapp():
    """获取或创建 Qt 应用。"""
    return QApplication.instance() or QApplication([])


def _flush_events(rounds=4):
    """处理少量 Qt 事件，确保表格状态稳定。"""
    app = _get_qapp()
    for _ in range(max(1, rounds)):
        app.processEvents()


def _prepare_batch_panel(monkeypatch):
    """创建批量计算面板并屏蔽弹窗。"""
    _get_qapp()
    monkeypatch.setattr(batch_panel_mod, "InfoBar", _InfoBarStub)
    monkeypatch.setattr(batch_panel_mod, "fluent_batch_result", lambda *args, **kwargs: None)
    monkeypatch.setattr(batch_panel_mod, "fluent_info", lambda *args, **kwargs: None)
    monkeypatch.setattr(batch_panel_mod, "fluent_question", lambda *args, **kwargs: True)

    panel = BatchPanel()
    panel.resize(1400, 900)
    panel.show()
    _flush_events(6)
    panel._clear_input(force=True)
    _flush_events(2)
    return panel


def _load_qiyi_panel(monkeypatch):
    """把七一 Excel 导入批量计算面板。"""
    assert QIYI_EXCEL.exists(), "七一水库充水渠批量计算用表.xlsx 不存在"
    panel = _prepare_batch_panel(monkeypatch)
    panel._do_load_from_filepath(str(QIYI_EXCEL), is_sample=False)
    _flush_events(4)
    return panel


def _run_qiyi_batch_calculation(monkeypatch):
    """执行七一 Excel 批量计算并捕获共享注册数据。"""
    panel = _load_qiyi_panel(monkeypatch)
    registered_rows = []

    class _SharedManagerStub:
        """记录批量计算同步给水面线的数据。"""

        def clear_batch_results(self):
            """清空上一次注册数据。"""
            registered_rows.clear()

        def register_batch_results(self, rows):
            """保存注册数据并返回注册数量。"""
            registered_rows[:] = copy.deepcopy(rows)
            return len(rows)

    monkeypatch.setattr(batch_panel_mod, "get_shared_data_manager", lambda: _SharedManagerStub())
    panel._batch_calculate()
    _flush_events(4)
    return panel, registered_rows


def _sf(value, default=0.0):
    """把单元格文本转成浮点数。"""
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _assert_close(actual, expected, *, rel=1e-9, abs=1e-9):
    """比较两个数值字段。"""
    assert float(actual) == pytest.approx(float(expected), rel=rel, abs=abs)


def _build_settings_from_batch(panel, registered_rows):
    """从七一批量结果构造水面线项目参数。"""
    max_flow = 0.0
    if registered_rows:
        max_flow = float(
            registered_rows[0].get("Q_increased")
            or registered_rows[0].get("Q_inc")
            or registered_rows[0].get("Q")
            or 0.0
        )
    return ProjectSettings(
        channel_name=panel.channel_name_edit.text().strip(),
        channel_level=panel.channel_level_combo.currentText(),
        start_station=0.0,
        start_water_level=float(panel.start_wl_edit.text()),
        design_flow=0.7,
        max_flow=max_flow,
        design_flows=[0.7],
        max_flows=[max_flow] if max_flow > 0 else [0.7],
        roughness=0.014,
    )


def _node_from_registered_row(row):
    """把批量结果行转换成水面线节点。"""
    node = ChannelNode()
    node.flow_section = str(row.get("flow_section", "1") or "1")
    node.name = str(row.get("building_name", "") or "")
    node.structure_type = StructureType.from_string(str(row.get("section_type", "明渠-矩形") or "明渠-矩形"))
    node.x = float(row.get("coord_X", 0.0) or 0.0)
    node.y = float(row.get("coord_Y", 0.0) or 0.0)
    node.turn_radius = float(row.get("turn_radius", 0.0) or 0.0)
    node.flow = float(row.get("Q", 0.0) or 0.0)
    node.roughness = float(row.get("n", 0.014) or 0.014)
    slope_inv = float(row.get("slope_inv", 0.0) or 0.0)
    node.slope_i = 1.0 / slope_inv if slope_inv > 0 else 0.0

    b_value = float(row.get("b_design", row.get("B", 0.0)) or 0.0)
    h_total = float(row.get("h_prime", row.get("H_total", 0.0)) or 0.0)
    node.water_depth = float(row.get("h_design", row.get("y_d", 0.0)) or 0.0)
    node.velocity = float(row.get("V_design", row.get("V_d", 0.0)) or 0.0)
    node.velocity_increased = float(row.get("V_increased", row.get("V_i", 0.0)) or 0.0)
    node.structure_height = h_total
    node.section_params.update(
        {
            "B": b_value,
            "m": float(row.get("m", 0.0) or 0.0),
            "A": float(row.get("A_design", row.get("A_d", 0.0)) or 0.0),
            "X": float(row.get("X_design", row.get("P_d", 0.0)) or 0.0),
            "R": float(row.get("R_design", row.get("R_d", 0.0)) or 0.0),
            "H_total": h_total,
            "use_increase": bool(row.get("_use_increase", True)),
        }
    )
    spillway_payload = row.get(SPILLWAY_STEEP_CHUTE_PARAM_KEY)
    if isinstance(spillway_payload, dict) and spillway_payload:
        node.section_params[SPILLWAY_STEEP_CHUTE_PARAM_KEY] = copy.deepcopy(spillway_payload)
    return node


def _calculate_qiyi_nodes(panel, registered_rows):
    """用批量结果构造并计算水面线节点。"""
    settings = _build_settings_from_batch(panel, registered_rows)
    nodes = [_node_from_registered_row(row) for row in registered_rows]
    calculated = WaterProfileCalculator(settings).calculate_all(nodes)
    return settings, calculated


def _tail_batch_indexes(panel):
    """返回七一末尾 4 行在 Qt 表格中的索引。"""
    row_count = panel.input_table.rowCount()
    return range(row_count - 4, row_count)


def test_qiyi_excel_fixture_has_fill_channel_required_inputs():
    """七一真实 Excel 的末尾 4 行应已改为充水渠，手工 Q加大应保持空白。"""
    wb = load_workbook(QIYI_EXCEL, data_only=False, read_only=True)
    ws = wb.active

    assert [ws.cell(1, c).value for c in range(1, 9)] == [
        "渠道名称",
        "七一",
        "渠道级别",
        "充水渠",
        "渠道起始水位高程(m)",
        400,
        "起始桩号",
        0,
    ]
    assert ws.cell(1, 9).value in (None, "")
    assert ws.cell(1, 10).value in (None, "")

    for excel_row in TAIL_EXCEL_ROWS:
        assert ws.cell(excel_row, 4).value == "充水渠"
        assert str(ws.cell(excel_row, 19).value).strip() == "0.1"
        assert str(ws.cell(excel_row, 20).value).strip() == "100"


def test_qiyi_excel_imports_fill_channel_rows(monkeypatch):
    """七一真实 Excel 应导入 18 行，末尾 4 行识别为专项类型。"""
    panel = _load_qiyi_panel(monkeypatch)

    assert panel.channel_name_edit.text().strip() == "七一"
    assert panel.channel_level_combo.currentText() == "充水渠"
    assert float(panel.start_wl_edit.text()) == pytest.approx(400.0)
    assert panel.start_station_edit.text().strip() == "0+000.000"
    assert panel.input_table.rowCount() == 18
    assert getattr(panel, "_manual_qmax_by_segment", {}) == {}

    slopes = []
    for row in _tail_batch_indexes(panel):
        assert panel.input_table.item(row, batch_panel_mod.COL_SECTION_TYPE).text() == "泄水渠与陡坡"
        assert panel.input_table.item(row, batch_panel_mod.COL_BUILDING_NAME).text().strip() == "-"
        assert panel.input_table.item(row, batch_panel_mod.COL_Q).text().strip() == "0.7"
        assert panel.input_table.item(row, batch_panel_mod.COL_N).text().strip() == "0.014"
        assert panel.input_table.item(row, batch_panel_mod.COL_M).text().strip() == "0"
        assert panel.input_table.item(row, batch_panel_mod.COL_B).text().strip() == "1"
        assert panel.input_table.item(row, batch_panel_mod.COL_V_MIN).text().strip() == "0.1"
        assert panel.input_table.item(row, batch_panel_mod.COL_V_MAX).text().strip() == "100"
        assert panel.input_table.item(row, batch_panel_mod.COL_TURN_RADIUS).text().strip() == "10"
        slopes.append(float(panel.input_table.item(row, batch_panel_mod.COL_SLOPE).text()))

    assert slopes == [100.0, 10.0, 24.5, 24.5]


def test_qiyi_batch_calculation_registers_spillway_chain_rows(monkeypatch):
    """七一批量计算应把末尾 4 行作为泄水渠与陡坡专项参数注册。"""
    panel, registered_rows = _run_qiyi_batch_calculation(monkeypatch)

    assert panel.result_table.rowCount() == 18
    assert len(panel.batch_results) == 18
    assert len(registered_rows) == 18
    assert all(item["result"].get("success") is True for item in panel.batch_results)
    assert all("manual_qmax_from_excel" not in item["result"] for item in panel.batch_results)

    for row in registered_rows[-4:]:
        assert row.get("section_type") == "泄水渠与陡坡"
        assert row.get("building_name") == "-"
        assert row.get("is_spillway_steep_chute") is True
        assert row.get("Q") == pytest.approx(0.7)
        assert row.get("B") == pytest.approx(1.0)
        assert row.get("m") == pytest.approx(0.0)
        assert SPILLWAY_STEEP_CHUTE_PARAM_KEY in row
        assert "h_design" not in row
        assert "V_design" not in row


def test_fill_channel_aliases_import_as_spillway_type(monkeypatch, tmp_path):
    """Excel 中充水渠类名称都应导入为泄水渠与陡坡专项类型。"""
    source = tmp_path / "qiyi_aliases.xlsx"
    shutil.copy2(QIYI_EXCEL, source)
    wb = load_workbook(source)
    ws = wb.active
    for excel_row, alias in zip(TAIL_EXCEL_ROWS[:3], FILL_CHANNEL_ALIASES):
        ws.cell(excel_row, 4).value = alias
    wb.save(source)

    panel = _prepare_batch_panel(monkeypatch)
    panel._do_load_from_filepath(str(source), is_sample=False)
    _flush_events(4)

    assert [panel.input_table.item(i, batch_panel_mod.COL_SECTION_TYPE).text() for i in range(14, 18)] == [
        "泄水渠与陡坡",
        "泄水渠与陡坡",
        "泄水渠与陡坡",
        "泄水渠与陡坡",
    ]


def test_qiyi_fill_channel_water_profile_and_combined_dxf(monkeypatch, tmp_path):
    """七一充水渠应按变坡专项链完成表3水面线并导出可打开的全部 DXF。"""
    ezdxf = pytest.importorskip("ezdxf")
    panel, registered_rows = _run_qiyi_batch_calculation(monkeypatch)
    settings, calculated_nodes = _calculate_qiyi_nodes(panel, registered_rows)
    spillway_nodes = [node for node in calculated_nodes if is_spillway_steep_chute_node(node)]

    assert len(spillway_nodes) == 4
    assert all(node.water_level > 0 and node.water_depth > 0 and node.velocity > 0 for node in spillway_nodes)
    payloads = [node.section_params[SPILLWAY_STEEP_CHUTE_PARAM_KEY] for node in spillway_nodes]
    assert [item["slope_inv"] for item in payloads[0]["chain_subsegments"]] == pytest.approx([100.0, 10.0, 24.5])
    assert [item["node_count"] for item in payloads[0]["chain_subsegments"]] == [2, 2, 2]
    assert all(payload["chain_segment_count"] == 3 for payload in payloads)
    assert payloads[2]["segment_key"] == payloads[3]["segment_key"]

    out_path = tmp_path / "qiyi_fill_channel_combined.dxf"
    messages = []
    export_panel = SimpleNamespace(
        calculated_nodes=calculated_nodes,
        nodes=calculated_nodes,
        _text_export_settings={},
        _custom_pressurized_pipe_params={},
        channel_name_edit=_TextStub("七一"),
        channel_level_combo=_ComboStub("充水渠"),
    )
    export_panel.window = lambda: None
    export_panel._build_settings = lambda: settings
    export_panel._build_nodes_from_table = lambda: calculated_nodes

    monkeypatch.setattr(cad_tools, "TextExportSettingsDialog", _AcceptedTextDialog)
    monkeypatch.setattr(cad_tools.QFileDialog, "getSaveFileName", staticmethod(lambda *_args, **_kwargs: (str(out_path), "DXF")))
    monkeypatch.setattr(cad_tools, "fluent_question", lambda *_args, **_kwargs: False)
    monkeypatch.setattr(cad_tools, "fluent_info", lambda *_args, **_kwargs: messages.append(("info", _args)))
    monkeypatch.setattr(cad_tools, "fluent_error", lambda *_args, **_kwargs: messages.append(("error", _args)))

    cad_tools.export_combined_dxf(export_panel)

    assert out_path.exists()
    assert [kind for kind, _ in messages if kind == "error"] == []
    doc = ezdxf.readfile(out_path)
    layer_names = {layer.dxf.name for layer in doc.layers}
    assert "断面汇总表" in layer_names
    assert "IP坐标表" in layer_names
    assert any(name.startswith("纵断面_") for name in layer_names)

    modelspace = doc.modelspace()
    assert len(modelspace) > 0
    text_values = [entity.dxf.text for entity in modelspace.query("TEXT")]
    joined_text = "\n".join(text_values)
    assert "七一充" in joined_text
    assert "泄水渠与陡坡" in joined_text
    assert "-泄水渠与陡坡" not in joined_text
    assert "泄陡进" not in joined_text
    assert "泄陡出" not in joined_text
    assert "IP坐标及弯道参数表" in joined_text
    assert "断面尺寸及水力要素表" in joined_text
