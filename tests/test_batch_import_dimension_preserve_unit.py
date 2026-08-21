# -*- coding: utf-8 -*-
"""Regression coverage for preserving explicit open-channel bottom widths in batch mode."""

import os
import sys
import tempfile
import types
from pathlib import Path

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
os.environ.setdefault("QTWEBENGINE_DISABLE_SANDBOX", "1")
os.environ.setdefault(
    "MPLCONFIGDIR",
    str(Path(tempfile.gettempdir()) / "codex-mplconfig"),
)

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from PySide6.QtWidgets import QApplication

from app_渠系计算前端.batch.panel import BatchPanel, INPUT_HEADERS
import app_渠系计算前端.batch.panel as batch_panel_mod


TARGET_BATCH_INPUT_HEADERS = [
    "序号", "流量段", "建筑物名称", "结构形式", "X", "Y",
    "Q(m³/s)", "糙率n", "比降(1/)",
    "边坡系数m", "底宽B(m)", "暗涵高宽比H/B", "暗涵高度H(m)", "宽深比β",
    "半径R(m)", "直径D(m)",
    "矩形渡槽深宽比", "拉杆高度(m)", "倒角角度(°)", "倒角底边(m)", "圆心角(°)", "直墙高度H直(m)",
    "不淤流速", "不冲流速", "转弯半径(m)", "管材",
    "左上坡m1", "平台宽B1(m)", "左下坡m2", "渠底宽B2(m)", "右坡m3", "平台高差h1(m)",
    "泄水渠入口宽度(m)", "泄水渠堰上总水头(m)", "泄水渠入口连接形式", "泄水渠手动流量系数",
    "泄水渠侧收缩系数", "泄水渠动能修正系数", "泄水渠掺气系数",
    "泄水渠侧墙安全超高(m)", "泄水渠池深系数", "泄水渠整流长度系数",
]

OLD_BATCH_INPUT_HEADERS_BEFORE_AE_AF_REORDER = [
    "序号", "流量段", "建筑物名称", "结构形式", "X", "Y",
    "Q(m³/s)", "糙率n", "比降(1/)",
    "边坡系数m", "底宽B(m)", "暗涵高宽比H/B", "暗涵高度H(m)", "宽深比β",
    "半径R(m)", "直径D(m)",
    "矩形渡槽深宽比", "倒角角度(°)", "倒角底边(m)", "圆心角(°)",
    "不淤流速", "不冲流速", "转弯半径(m)", "管材",
    "左上坡m1", "平台宽B1(m)", "左下坡m2", "渠底宽B2(m)", "右坡m3", "平台高差h1(m)",
    "直墙高度H直(m)", "拉杆高度(m)",
]


def _get_qapp():
    return QApplication.instance() or QApplication([])


def _flush_events(rounds=4):
    app = _get_qapp()
    for _ in range(max(1, rounds)):
        app.processEvents()


def _build_row(*, q="1", slope_inv="2.2", b="1.5"):
    row = [""] * len(INPUT_HEADERS)
    row[batch_panel_mod.COL_SEQ] = "1"
    row[batch_panel_mod.COL_SEGMENT] = "1"
    row[batch_panel_mod.COL_BUILDING_NAME] = "-"
    row[batch_panel_mod.COL_SECTION_TYPE] = "明渠-矩形"
    row[batch_panel_mod.COL_Q] = q
    row[batch_panel_mod.COL_N] = "0.014"
    row[batch_panel_mod.COL_SLOPE] = slope_inv
    row[batch_panel_mod.COL_M] = "0"
    row[batch_panel_mod.COL_B] = b
    row[batch_panel_mod.COL_V_MIN] = "0.1"
    row[batch_panel_mod.COL_V_MAX] = "100"
    return row


def _build_rect_aqueduct_row(*, q="1", slope_inv="2000", b="1.5", ratio="0.8", tie_rod_height=""):
    row = [""] * len(INPUT_HEADERS)
    row[batch_panel_mod.COL_SEQ] = "1"
    row[batch_panel_mod.COL_SEGMENT] = "1"
    row[batch_panel_mod.COL_BUILDING_NAME] = "测试渡槽"
    row[batch_panel_mod.COL_SECTION_TYPE] = "渡槽-矩形"
    row[batch_panel_mod.COL_Q] = q
    row[batch_panel_mod.COL_N] = "0.014"
    row[batch_panel_mod.COL_SLOPE] = slope_inv
    row[batch_panel_mod.COL_B] = b
    row[batch_panel_mod.COL_DUCAO_DEPTH_RATIO] = ratio
    row[batch_panel_mod.COL_V_MIN] = "0.1"
    row[batch_panel_mod.COL_V_MAX] = "100"
    if tie_rod_height != "":
        row[batch_panel_mod.COL_TIE_ROD_HEIGHT] = tie_rod_height
    return row


def _build_rect_culvert_row(*, q="17", slope_inv="4000", b="4.2"):
    row = [""] * len(INPUT_HEADERS)
    row[batch_panel_mod.COL_SEQ] = "1"
    row[batch_panel_mod.COL_SEGMENT] = "1"
    row[batch_panel_mod.COL_BUILDING_NAME] = "测试暗涵"
    row[batch_panel_mod.COL_SECTION_TYPE] = "暗涵-矩形"
    row[batch_panel_mod.COL_Q] = q
    row[batch_panel_mod.COL_N] = "0.014"
    row[batch_panel_mod.COL_SLOPE] = slope_inv
    row[batch_panel_mod.COL_B] = b
    row[batch_panel_mod.COL_V_MIN] = "0.1"
    row[batch_panel_mod.COL_V_MAX] = "100"
    return row


def _build_u_aqueduct_row(*, q="1", slope_inv="2000", radius="1.5", tie_rod_height=""):
    row = [""] * len(INPUT_HEADERS)
    row[batch_panel_mod.COL_SEQ] = "1"
    row[batch_panel_mod.COL_SEGMENT] = "1"
    row[batch_panel_mod.COL_BUILDING_NAME] = "测试U形渡槽"
    row[batch_panel_mod.COL_SECTION_TYPE] = "渡槽-U形"
    row[batch_panel_mod.COL_Q] = q
    row[batch_panel_mod.COL_N] = "0.014"
    row[batch_panel_mod.COL_SLOPE] = slope_inv
    row[batch_panel_mod.COL_R] = radius
    row[batch_panel_mod.COL_V_MIN] = "0.1"
    row[batch_panel_mod.COL_V_MAX] = "100"
    if tie_rod_height != "":
        row[batch_panel_mod.COL_TIE_ROD_HEIGHT] = tie_rod_height
    return row


def _build_spillway_row(*, building_name="泄槽A", section_type="泄水渠与陡坡", inlet_head="2.8", alpha_profile="1.18"):
    row = [""] * len(INPUT_HEADERS)
    row[batch_panel_mod.COL_SEQ] = "1"
    row[batch_panel_mod.COL_SEGMENT] = "1"
    row[batch_panel_mod.COL_BUILDING_NAME] = building_name
    row[batch_panel_mod.COL_SECTION_TYPE] = section_type
    row[batch_panel_mod.COL_X] = "10"
    row[batch_panel_mod.COL_Y] = "20"
    row[batch_panel_mod.COL_Q] = "12"
    row[batch_panel_mod.COL_N] = "0.014"
    row[batch_panel_mod.COL_SLOPE] = "18"
    row[batch_panel_mod.COL_M] = "0.5"
    row[batch_panel_mod.COL_B] = "2.4"
    row[batch_panel_mod.COL_SPILLWAY_INLET_WEIR_WIDTH] = "2.1"
    row[batch_panel_mod.COL_SPILLWAY_INLET_HEAD] = inlet_head
    row[batch_panel_mod.COL_SPILLWAY_INLET_CONNECTION_TYPE] = "八字墙连接"
    row[batch_panel_mod.COL_SPILLWAY_WEIR_COEFFICIENT] = ""
    row[batch_panel_mod.COL_SPILLWAY_CONTRACTION_COEFFICIENT] = "0.96"
    row[batch_panel_mod.COL_SPILLWAY_ALPHA_PROFILE] = alpha_profile
    row[batch_panel_mod.COL_SPILLWAY_AERATION_COEFFICIENT] = "1.25"
    row[batch_panel_mod.COL_SPILLWAY_SIDEWALL_FREEBOARD] = "0.45"
    row[batch_panel_mod.COL_SPILLWAY_POOL_DEPTH_FACTOR] = "1.12"
    row[batch_panel_mod.COL_SPILLWAY_OUTLET_RECTIFICATION_FACTOR] = "9.5"
    return row


def _prepare_panel(monkeypatch):
    _get_qapp()

    class _InfoBarStub:
        @staticmethod
        def success(*args, **kwargs):
            return None

        @staticmethod
        def warning(*args, **kwargs):
            return None

        @staticmethod
        def error(*args, **kwargs):
            return None

        @staticmethod
        def info(*args, **kwargs):
            return None

    monkeypatch.setattr(batch_panel_mod, "InfoBar", _InfoBarStub)
    monkeypatch.setattr(batch_panel_mod, "fluent_batch_result", lambda *args, **kwargs: None)
    monkeypatch.setattr(batch_panel_mod, "fluent_info", lambda *args, **kwargs: None)

    panel = BatchPanel()
    panel.resize(1400, 900)
    panel.show()
    _flush_events(6)
    panel._clear_input(force=True)
    _flush_events(2)
    return panel


def _set_single_row(panel, row_data):
    panel._add_row(row_data)
    panel._renumber()
    _flush_events(2)


def _install_fake_openpyxl(monkeypatch, *, b_value="1.5"):
    legacy_headers = [
        "序号", "流量段", "建筑物名称", "结构形式", "X", "Y",
        "Q(m³/s)", "糙率n", "比降(1/)", "边坡系数m",
        "底宽B(m)", "明渠宽深比", "半径R(m)", "直径D(m)",
        "矩形渡槽深宽比", "倒角角度(°)", "倒角底边(m)", "圆心角(°)",
        "不淤流速", "不冲流速", "转弯半径(m)", "管材",
    ]
    cells = {
        (1, 1): "渠道名称",
        (1, 2): "测试渠道",
        (1, 3): "渠道级别",
        (1, 4): "支渠",
        (1, 5): "起始水位",
        (1, 6): "369.5",
        (1, 8): "0+000.000",
        (3, 1): "1",
        (3, 2): "1",
        (3, 3): "-",
        (3, 4): "明渠-矩形",
        (3, 7): "1",
        (3, 8): "0.014",
        (3, 9): "2.2",
        (3, 10): "0",
        (3, 11): b_value,
        (3, 19): "0.1",
        (3, 20): "100",
    }
    for idx, header in enumerate(legacy_headers, start=1):
        cells[(2, idx)] = header

    class _Cell:
        def __init__(self, value):
            self.value = value

    class _Sheet:
        max_row = 3
        max_column = len(legacy_headers)

        def cell(self, row, column):
            return _Cell(cells.get((row, column)))

    class _Workbook:
        active = _Sheet()

    fake_openpyxl = types.ModuleType("openpyxl")
    fake_openpyxl.load_workbook = lambda *args, **kwargs: _Workbook()
    monkeypatch.setitem(sys.modules, "openpyxl", fake_openpyxl)


def _install_fake_openpyxl_table(monkeypatch, headers, row_values):
    cells = {
        (1, 1): "渠道名称",
        (1, 2): "测试渠道",
        (1, 3): "渠道级别",
        (1, 4): "支渠",
        (1, 5): "起始水位",
        (1, 6): "369.5",
        (1, 8): "0+000.000",
    }
    for idx, header in enumerate(headers, start=1):
        cells[(2, idx)] = header
    for idx, value in enumerate(row_values, start=1):
        cells[(3, idx)] = value

    class _Cell:
        def __init__(self, value):
            self.value = value

    class _Sheet:
        max_row = 3
        max_column = max(len(headers), len(row_values))

        def cell(self, row, column):
            return _Cell(cells.get((row, column)))

    class _Workbook:
        active = _Sheet()

    fake_openpyxl = types.ModuleType("openpyxl")
    fake_openpyxl.load_workbook = lambda *args, **kwargs: _Workbook()
    monkeypatch.setitem(sys.modules, "openpyxl", fake_openpyxl)


def test_batch_calculate_preserves_explicit_bottom_width_and_registers_result(monkeypatch):
    registered_rows = []

    class _SharedManager:
        def clear_batch_results(self):
            return None

        def register_batch_results(self, rows):
            registered_rows.extend(rows)
            return len(rows)

    monkeypatch.setattr(batch_panel_mod, "SHARED_DATA_AVAILABLE", True)
    monkeypatch.setattr(batch_panel_mod, "get_shared_data_manager", lambda: _SharedManager())

    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_row())
    panel._set_excel_import_session_active(True)
    panel._mark_row_as_excel_imported(0, True)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    status_col = panel.result_table.columnCount() - 1
    result = panel.batch_results[0]["result"]

    assert panel.result_table.item(0, 4).text() == "1.500"
    assert panel.result_table.item(0, status_col).text() == "⚠ 按显式底宽计算"
    assert result["preserved_manual_b"] is True
    assert result["used_manual_b"] is True
    assert any("宽深比" in warning for warning in result["constraint_warnings"])
    assert panel._has_calc_errors is False
    assert registered_rows
    assert registered_rows[0]["b_design"] == 1.5

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_batch_calculate_registers_explicit_zero_turn_radius_text(monkeypatch):
    registered_rows = []

    class _SharedManager:
        def clear_batch_results(self):
            return None

        def register_batch_results(self, rows):
            registered_rows.extend(rows)
            return len(rows)

    monkeypatch.setattr(batch_panel_mod, "SHARED_DATA_AVAILABLE", True)
    monkeypatch.setattr(batch_panel_mod, "get_shared_data_manager", lambda: _SharedManager())

    panel = _prepare_panel(monkeypatch)
    row = _build_row()
    row[batch_panel_mod.COL_TURN_RADIUS] = "0"
    _set_single_row(panel, row)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    assert registered_rows
    assert registered_rows[0]["turn_radius"] == 0.0
    assert registered_rows[0]["turn_radius_text"] == "0"

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_batch_calculate_treats_spillway_alias_as_special_payload(monkeypatch):
    """Excel 填泄水渠别名时，计算走专项，DXF 显示保留原始名称。"""
    registered_rows = []

    class _SharedManager:
        def clear_batch_results(self):
            return None

        def register_batch_results(self, rows):
            registered_rows.extend(rows)
            return len(rows)

    monkeypatch.setattr(batch_panel_mod, "SHARED_DATA_AVAILABLE", True)
    monkeypatch.setattr(batch_panel_mod, "get_shared_data_manager", lambda: _SharedManager())

    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_spillway_row(section_type="充水渠"))
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    result = panel.batch_results[0]["result"]

    assert result["section_type"] == "泄水渠与陡坡"
    assert result["display_structure_type"] == "充水渠"
    assert result["is_spillway_steep_chute"] is True
    assert "spillway_steep_chute" in result
    assert result["B"] == pytest.approx(2.4)
    assert result["m"] == pytest.approx(0.5)
    assert "V_design" not in result
    assert "A_design" not in result
    advanced = result["spillway_steep_chute"]["advanced_params"]
    assert advanced["inlet_head"] == pytest.approx(2.8)
    assert advanced["alpha_profile"] == pytest.approx(1.18)
    assert advanced["inlet_connection_type_label"] == "八字墙连接"
    assert registered_rows[0]["section_type"] == "泄水渠与陡坡"
    assert registered_rows[0]["display_structure_type"] == "充水渠"
    assert registered_rows[0]["spillway_steep_chute"]["display_structure_type"] == "充水渠"
    assert registered_rows[0]["spillway_steep_chute"]["advanced_params"] == advanced

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_real_excel_import_marks_rows_as_imported(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _install_fake_openpyxl(monkeypatch)

    panel._do_load_from_filepath("import.xlsx", is_sample=False)
    _flush_events(4)

    assert panel._excel_import_session_active is True
    assert panel._is_row_excel_imported(0) is True
    assert panel.input_table.item(0, 10).text() == "1.5"

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_real_excel_import_reads_rect_culvert_visible_h_and_hb_columns(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    headers = list(batch_panel_mod.INPUT_HEADERS)
    row = [""] * len(headers)
    row[0] = "1"
    row[1] = "1"
    row[2] = "测试暗涵"
    row[3] = "暗涵-矩形"
    row[batch_panel_mod.COL_Q] = "17"
    row[batch_panel_mod.COL_N] = "0.014"
    row[batch_panel_mod.COL_SLOPE] = "4000"
    row[batch_panel_mod.COL_B] = "4.2"
    row[batch_panel_mod.COL_RECT_CULVERT_HB_RATIO] = "0.976"
    row[batch_panel_mod.COL_RECT_CULVERT_H] = "4.1"
    row[batch_panel_mod.COL_V_MIN] = "0.1"
    row[batch_panel_mod.COL_V_MAX] = "100"
    _install_fake_openpyxl_table(monkeypatch, headers, row)

    panel._do_load_from_filepath("new-template.xlsx", is_sample=False)
    _flush_events(4)

    assert panel.input_table.item(0, batch_panel_mod.COL_B).text() == "4.2"
    assert panel.input_table.item(0, batch_panel_mod.COL_RECT_CULVERT_HB_RATIO).text() == "0.976"
    assert panel.input_table.item(0, batch_panel_mod.COL_RECT_CULVERT_H).text() == "4.1"

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_real_excel_import_normalizes_spillway_alias_to_spillway(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    headers = list(batch_panel_mod.INPUT_HEADERS)
    row = _build_spillway_row(inlet_head="3.1", alpha_profile="1.22")
    _install_fake_openpyxl_table(monkeypatch, headers, row)

    panel._do_load_from_filepath("spillway-template.xlsx", is_sample=False)
    _flush_events(4)

    assert panel.input_table.item(0, batch_panel_mod.COL_SECTION_TYPE).text() == "泄水渠与陡坡"
    assert panel.input_table.item(0, batch_panel_mod.COL_SPILLWAY_INLET_HEAD).text() == "3.1"
    assert panel.input_table.item(0, batch_panel_mod.COL_SPILLWAY_ALPHA_PROFILE).text() == "1.22"
    assert panel.input_table.item(0, batch_panel_mod.COL_SPILLWAY_INLET_CONNECTION_TYPE).text() == "八字墙连接"

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_official_blank_template_contains_rect_culvert_visible_size_columns():
    """正式空模板应包含矩形暗涵 H/B 与 H 明列，且紧跟底宽 B。"""
    openpyxl = pytest.importorskip("openpyxl")
    template_path = ROOT / "data" / "多流量段批量计算_导入Excel（模板）.xlsx"

    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]

    b_idx = headers.index("底宽B(m)")
    assert headers[b_idx + 1] == "暗涵高宽比H/B"
    assert headers[b_idx + 2] == "暗涵高度H(m)"
    assert headers[b_idx + 3] == "宽深比β"


def test_batch_input_headers_follow_user_friendly_parameter_order():
    """批量输入列顺序应按参数关联排布，而不是把 H直/拉杆高度追加到最后。"""

    assert list(batch_panel_mod.INPUT_HEADERS) == TARGET_BATCH_INPUT_HEADERS
    assert batch_panel_mod.COL_TIE_ROD_HEIGHT == TARGET_BATCH_INPUT_HEADERS.index("拉杆高度(m)")
    assert batch_panel_mod.COL_ARCH_H_STRAIGHT == TARGET_BATCH_INPUT_HEADERS.index("直墙高度H直(m)")
    assert TARGET_BATCH_INPUT_HEADERS.index("矩形渡槽深宽比") < batch_panel_mod.COL_TIE_ROD_HEIGHT < TARGET_BATCH_INPUT_HEADERS.index("倒角角度(°)")
    assert TARGET_BATCH_INPUT_HEADERS.index("圆心角(°)") < batch_panel_mod.COL_ARCH_H_STRAIGHT < TARGET_BATCH_INPUT_HEADERS.index("不淤流速")
    assert batch_panel_mod.COL_SPILLWAY_INLET_WEIR_WIDTH == TARGET_BATCH_INPUT_HEADERS.index("泄水渠入口宽度(m)")
    assert batch_panel_mod.COL_SPILLWAY_OUTLET_RECTIFICATION_FACTOR == TARGET_BATCH_INPUT_HEADERS.index("泄水渠整流长度系数")


def test_real_excel_import_maps_legacy_mingqu_beta_header_without_shift(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    legacy_headers = [
        "序号", "流量段", "建筑物名称", "结构形式", "X", "Y",
        "Q(m³/s)", "糙率n", "比降(1/)", "边坡系数m",
        "底宽B(m)", "明渠宽深比", "半径R(m)", "直径D(m)",
        "矩形渡槽深宽比", "倒角角度(°)", "倒角底边(m)", "圆心角(°)",
        "不淤流速", "不冲流速", "转弯半径(m)", "管材",
    ]
    row = [
        "1", "1", "测试暗涵", "暗涵-矩形", "100", "200",
        "17", "0.014", "4000", "",
        "4.2", "1.2", "", "",
        "", "", "", "180",
        "0.1", "100", "0", "",
    ]
    _install_fake_openpyxl_table(monkeypatch, legacy_headers, row)

    panel._do_load_from_filepath("legacy-template.xlsx", is_sample=False)
    _flush_events(4)

    assert panel.input_table.item(0, batch_panel_mod.COL_B).text() == "4.2"
    assert panel.input_table.item(0, batch_panel_mod.COL_BETA).text() == "1.2"
    assert panel.input_table.item(0, batch_panel_mod.COL_RECT_CULVERT_HB_RATIO).text() == ""
    assert panel.input_table.item(0, batch_panel_mod.COL_RECT_CULVERT_H).text() == ""
    assert panel.input_table.item(0, batch_panel_mod.COL_THETA).text() == "180"
    assert panel.input_table.item(0, batch_panel_mod.COL_TURN_RADIUS).text() == "0"

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_legacy_excel_without_xy_and_reliable_headers_maps_optional_columns(monkeypatch):
    """旧无X/Y且表头不可识别时，倒角等参数不应被新增列挤偏。"""
    panel = _prepare_panel(monkeypatch)
    weak_headers = [""] * 20
    weak_headers[4] = "Q(m³/s)"
    row = [
        "1", "1", "旧无表头渡槽", "渡槽-矩形",
        "5", "0.014", "2000", "",
        "2.5", "1.2", "", "",
        "0.8", "45", "0.2", "180",
        "0.1", "100", "0", "PCCP管0.013",
    ]
    _install_fake_openpyxl_table(monkeypatch, weak_headers, row)

    panel._do_load_from_filepath("legacy-no-xy-weak-header.xlsx", is_sample=False)
    _flush_events(4)

    assert panel.input_table.item(0, batch_panel_mod.COL_Q).text() == "5"
    assert panel.input_table.item(0, batch_panel_mod.COL_B).text() == "2.5"
    assert panel.input_table.item(0, batch_panel_mod.COL_BETA).text() == "1.2"
    assert panel.input_table.item(0, batch_panel_mod.COL_DUCAO_DEPTH_RATIO).text() == "0.8"
    assert panel.input_table.item(0, batch_panel_mod.COL_TIE_ROD_HEIGHT).text() == ""
    assert panel.input_table.item(0, batch_panel_mod.COL_CHAMFER_ANGLE).text() == "45"
    assert panel.input_table.item(0, batch_panel_mod.COL_CHAMFER_LENGTH).text() == "0.2"
    assert panel.input_table.item(0, batch_panel_mod.COL_THETA).text() == "180"
    assert panel.input_table.item(0, batch_panel_mod.COL_ARCH_H_STRAIGHT).text() == ""
    assert panel.input_table.item(0, batch_panel_mod.COL_V_MIN).text() == "0.1"
    assert panel.input_table.item(0, batch_panel_mod.COL_V_MAX).text() == "100"
    assert panel.input_table.item(0, batch_panel_mod.COL_TURN_RADIUS).text() == "0"
    assert panel.input_table.item(0, batch_panel_mod.COL_PIPE_MATERIAL).text() == "PCCP管0.013"

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_legacy_project_rows_before_ae_af_reorder_restore_to_new_column_order(monkeypatch):
    """旧工程没有表头顺序时，应把旧 AE/AF 的 H直/拉杆高度迁移到新位置。"""
    panel = _prepare_panel(monkeypatch)
    legacy_row = [""] * len(OLD_BATCH_INPUT_HEADERS_BEFORE_AE_AF_REORDER)
    legacy_index = {header: idx for idx, header in enumerate(OLD_BATCH_INPUT_HEADERS_BEFORE_AE_AF_REORDER)}
    legacy_values = {
        "序号": "1",
        "流量段": "1",
        "建筑物名称": "旧工程行",
        "结构形式": "渡槽-矩形",
        "Q(m³/s)": "5",
        "糙率n": "0.014",
        "比降(1/)": "2000",
        "底宽B(m)": "2.5",
        "矩形渡槽深宽比": "0.8",
        "倒角角度(°)": "45",
        "倒角底边(m)": "0.2",
        "不淤流速": "0.1",
        "不冲流速": "100",
        "直墙高度H直(m)": "1.2",
        "拉杆高度(m)": "0.35",
        "平台高差h1(m)": "0.6",
    }
    for header, value in legacy_values.items():
        legacy_row[legacy_index[header]] = value

    panel.from_project_dict(
        {
            "input_rows": [legacy_row],
            "inc_checked": True,
            "detail_checked": False,
        },
        skip_dirty_signal=True,
    )
    _flush_events(4)

    assert panel.input_table.item(0, batch_panel_mod.COL_TIE_ROD_HEIGHT).text() == "0.35"
    assert panel.input_table.item(0, batch_panel_mod.COL_CHAMFER_ANGLE).text() == "45"
    assert panel.input_table.item(0, batch_panel_mod.COL_CHAMFER_LENGTH).text() == "0.2"
    assert panel.input_table.item(0, batch_panel_mod.COL_ARCH_H_STRAIGHT).text() == "1.2"
    assert panel.input_table.item(0, batch_panel_mod.COL_COMPOUND_H1).text() == "0.6"

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_sample_load_does_not_enable_excel_import_lock(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _install_fake_openpyxl(monkeypatch)

    panel._do_load_from_filepath("sample.xlsx", is_sample=True)
    _flush_events(4)

    assert panel._excel_import_session_active is False
    assert panel._is_row_excel_imported(0) is False

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_batch_calculate_preserves_explicit_bottom_width_without_import_lock(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_row())
    panel._set_excel_import_session_active(False)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    status_col = panel.result_table.columnCount() - 1
    result = panel.batch_results[0]["result"]

    assert panel.result_table.item(0, 4).text() == "1.500"
    assert panel.result_table.item(0, status_col).text() == "⚠ 按显式底宽计算"
    assert result["preserved_manual_b"] is True
    assert result["used_manual_b"] is True
    assert any("宽深比" in warning for warning in result["constraint_warnings"])

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_batch_calculate_without_explicit_bottom_width_keeps_appendix_e_auto_design(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_row(b=""))
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    status_col = panel.result_table.columnCount() - 1
    result = panel.batch_results[0]["result"]

    assert panel.result_table.item(0, 4).text() == "0.429"
    assert panel.result_table.item(0, status_col).text() == "✓ 成功"
    assert result["preserved_manual_b"] is False
    assert result["used_manual_b"] is False

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_sample_load_preserves_explicit_bottom_width_during_batch_calculate(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _install_fake_openpyxl(monkeypatch)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._do_load_from_filepath("sample.xlsx", is_sample=True)
    _flush_events(4)
    panel._batch_calculate()
    _flush_events(6)

    status_col = panel.result_table.columnCount() - 1
    result = panel.batch_results[0]["result"]

    assert panel._excel_import_session_active is False
    assert panel._is_row_excel_imported(0) is False
    assert panel.result_table.item(0, 4).text() == "1.500"
    assert panel.result_table.item(0, status_col).text() == "⚠ 按显式底宽计算"
    assert result["preserved_manual_b"] is True
    assert result["used_manual_b"] is True

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_project_roundtrip_preserves_explicit_bottom_width_on_recalculate(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_row())
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    saved = panel.to_project_dict()

    restored = _prepare_panel(monkeypatch)
    restored.from_project_dict(saved, skip_dirty_signal=True)
    _flush_events(4)
    restored.inc_cb.setChecked(False)
    restored.detail_cb.setChecked(False)
    restored._batch_calculate()
    _flush_events(6)

    status_col = restored.result_table.columnCount() - 1
    result = restored.batch_results[0]["result"]

    assert restored._excel_import_session_active is False
    assert restored._is_row_excel_imported(0) is False
    assert restored.result_table.item(0, 4).text() == "1.500"
    assert restored.result_table.item(0, status_col).text() == "⚠ 按显式底宽计算"
    assert result["preserved_manual_b"] is True
    assert result["used_manual_b"] is True

    panel.close()
    panel.deleteLater()
    restored.close()
    restored.deleteLater()
    _flush_events(4)


def test_rect_culvert_visible_height_survives_snapshot_restore(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_rect_culvert_row())
    panel.input_table.setItem(0, batch_panel_mod.COL_RECT_CULVERT_H, batch_panel_mod.QTableWidgetItem("4.1"))

    snapshot = panel._snapshot_table()
    panel.input_table.setItem(0, batch_panel_mod.COL_RECT_CULVERT_H, batch_panel_mod.QTableWidgetItem(""))
    panel._restore_table(snapshot)
    _flush_events(4)

    assert panel.input_table.item(0, batch_panel_mod.COL_RECT_CULVERT_H).text() == "4.1"

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_rect_culvert_visible_height_survives_copy_row(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_rect_culvert_row())
    panel.input_table.setItem(0, batch_panel_mod.COL_RECT_CULVERT_H, batch_panel_mod.QTableWidgetItem("4.1"))

    panel.input_table.selectRow(0)
    panel._copy_row()
    _flush_events(4)

    assert panel.input_table.rowCount() == 2
    assert panel.input_table.item(1, batch_panel_mod.COL_RECT_CULVERT_H).text() == "4.1"

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_project_roundtrip_preserves_rect_culvert_visible_height(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_rect_culvert_row())
    panel.input_table.setItem(0, batch_panel_mod.COL_RECT_CULVERT_H, batch_panel_mod.QTableWidgetItem("4.1"))

    saved = panel.to_project_dict()

    restored = _prepare_panel(monkeypatch)
    restored.from_project_dict(saved, skip_dirty_signal=True)
    _flush_events(4)

    assert saved["input_rows"][0][batch_panel_mod.COL_RECT_CULVERT_H] == "4.1"
    assert saved["input_header_order"] == list(batch_panel_mod.INPUT_HEADERS)
    assert saved["input_row_meta"][0].get("rect_culvert_manual_H") is None
    assert restored.input_table.item(0, batch_panel_mod.COL_RECT_CULVERT_H).text() == "4.1"

    panel.close()
    panel.deleteLater()
    restored.close()
    restored.deleteLater()
    _flush_events(4)


def test_legacy_project_hidden_rect_culvert_height_migrates_to_visible_column(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    old_len = len(batch_panel_mod.INPUT_HEADERS) - 2
    legacy_row = [""] * old_len
    legacy_row[0] = "1"
    legacy_row[1] = "1"
    legacy_row[2] = "测试暗涵"
    legacy_row[3] = "暗涵-矩形"
    legacy_row[6] = "17"
    legacy_row[7] = "0.014"
    legacy_row[8] = "4000"
    legacy_row[10] = "4.2"
    legacy_row[18] = "0.1"
    legacy_row[19] = "100"
    saved = {
        "input_rows": [legacy_row],
        "input_row_meta": [{"rect_culvert_manual_H": 4.1}],
        "inc_checked": True,
        "detail_checked": False,
    }

    panel.from_project_dict(saved, skip_dirty_signal=True)
    _flush_events(4)

    assert panel.input_table.item(0, batch_panel_mod.COL_B).text() == "4.2"
    assert panel.input_table.item(0, batch_panel_mod.COL_RECT_CULVERT_H).text() == "4.1"
    assert panel.input_table.item(0, batch_panel_mod.COL_RECT_CULVERT_HB_RATIO).text() == ""

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_batch_calculate_import_lock_failure_disables_exports(monkeypatch):
    registered_rows = []

    class _SharedManager:
        def clear_batch_results(self):
            return None

        def register_batch_results(self, rows):
            registered_rows.extend(rows)
            return len(rows)

    monkeypatch.setattr(batch_panel_mod, "SHARED_DATA_AVAILABLE", True)
    monkeypatch.setattr(batch_panel_mod, "get_shared_data_manager", lambda: _SharedManager())

    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_row(q="1000", slope_inv="2000", b="0.01"))
    panel._set_excel_import_session_active(True)
    panel._mark_row_as_excel_imported(0, True)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    status_col = panel.result_table.columnCount() - 1
    status_text = panel.result_table.item(0, status_col).text()

    assert panel._has_calc_errors is True
    assert "未自动改写" in status_text
    assert panel._btn_export_excel.isEnabled() is False
    assert panel._btn_export_word.isEnabled() is False
    assert registered_rows == []

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_rect_aqueduct_import_preserves_bottom_width_and_warns_on_ratio_conflict(monkeypatch):
    registered_rows = []
    captured_kwargs = []
    original_calc = batch_panel_mod.ducao_rect_calculate

    class _SharedManager:
        def clear_batch_results(self):
            return None

        def register_batch_results(self, rows):
            registered_rows.extend(rows)
            return len(rows)

    def _capture_calc(*args, **kwargs):
        captured_kwargs.append(dict(kwargs))
        return original_calc(*args, **kwargs)

    monkeypatch.setattr(batch_panel_mod, "SHARED_DATA_AVAILABLE", True)
    monkeypatch.setattr(batch_panel_mod, "get_shared_data_manager", lambda: _SharedManager())
    monkeypatch.setattr(batch_panel_mod, "ducao_rect_calculate", _capture_calc)

    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_rect_aqueduct_row(ratio="0.8"))
    panel._set_excel_import_session_active(True)
    panel._mark_row_as_excel_imported(0, True)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    status_col = panel.result_table.columnCount() - 1
    result = panel.batch_results[0]["result"]

    assert captured_kwargs
    assert captured_kwargs[0]["manual_B"] == 1.5
    assert panel.result_table.item(0, 4).text() == "1.500"
    assert panel.result_table.item(0, status_col).text() == "⚠ 按导入尺寸计算"
    assert result["B"] == 1.5
    assert result["preserved_manual_b"] is True
    assert result["used_manual_b"] is True
    assert result["imported_depth_width_ratio"] == 0.8
    assert any("H/B" in warning for warning in result["constraint_warnings"])
    assert panel._has_calc_errors is False
    assert panel._btn_export_excel.isEnabled() is True
    assert panel._btn_export_word.isEnabled() is True
    assert registered_rows
    assert registered_rows[0]["B"] == 1.5

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_u_aqueduct_batch_passes_tie_rod_height(monkeypatch):
    captured_kwargs = []
    original_calc = batch_panel_mod.ducao_u_calculate

    def _capture_calc(*args, **kwargs):
        captured_kwargs.append(dict(kwargs))
        return original_calc(*args, **kwargs)

    monkeypatch.setattr(batch_panel_mod, "ducao_u_calculate", _capture_calc)

    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_u_aqueduct_row(tie_rod_height="0.30"))
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    result = panel.batch_results[0]["result"]

    assert captured_kwargs
    assert captured_kwargs[0]["tie_rod_height"] == pytest.approx(0.30)
    assert result["tie_rod_height"] == pytest.approx(0.30)
    assert result["H_total"] == pytest.approx(result["tie_bottom_height"] + 0.30)
    assert result["Fb"] == pytest.approx(result["tie_bottom_height"] - result["h_increased"])

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_u_aqueduct_batch_detail_report_handles_tie_rod_height(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    row = _build_u_aqueduct_row(q="5.0", slope_inv="3000", radius="2.4", tie_rod_height="0.30")
    result = panel._calculate_single(
        "渡槽-U形",
        Q=5.0,
        n=0.014,
        slope_inv=3000,
        v_min=0.1,
        v_max=100.0,
        R=2.4,
        tie_rod_height=0.30,
        manual_increase_percent=0,
    )

    report = panel._gen_detail_report(row, result)

    assert "生成详细报告时出错" not in report
    assert "设计拉杆底净距" in report

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_batch_result_hides_effective_freeboard_when_increase_disabled(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_rect_aqueduct_row(ratio="", tie_rod_height="0.25"))
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    fb_col = batch_panel_mod.RESULT_HEADERS.index("加大有效超高Fb(m)")

    assert panel.result_table.item(0, fb_col).text() == "—"

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_batch_detail_report_hides_increase_section_when_disabled(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    row = _build_rect_aqueduct_row(ratio="", tie_rod_height="0.25")
    result = panel._calculate_single(
        "渡槽-矩形",
        Q=1.0,
        n=0.014,
        slope_inv=2000,
        v_min=0.1,
        v_max=100.0,
        b=1.5,
        tie_rod_height=0.25,
        manual_increase_percent=0,
        preserve_imported_dimensions=True,
    )
    result["_use_increase"] = False

    report = panel._gen_detail_report(row, result)

    assert "加大流量工况" not in report
    assert "加大有效超高" not in report
    assert "设计拉杆底净距" in report

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_rect_aqueduct_batch_passes_tie_rod_height(monkeypatch):
    captured_kwargs = []
    original_calc = batch_panel_mod.ducao_rect_calculate

    def _capture_calc(*args, **kwargs):
        captured_kwargs.append(dict(kwargs))
        return original_calc(*args, **kwargs)

    monkeypatch.setattr(batch_panel_mod, "ducao_rect_calculate", _capture_calc)

    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_rect_aqueduct_row(ratio="", tie_rod_height="0.25"))
    panel._set_excel_import_session_active(True)
    panel._mark_row_as_excel_imported(0, True)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    result = panel.batch_results[0]["result"]

    assert captured_kwargs
    assert captured_kwargs[0]["tie_rod_height"] == pytest.approx(0.25)
    assert result["tie_rod_height"] == pytest.approx(0.25)
    assert result["H_total"] == pytest.approx(result["tie_bottom_height"] + 0.25)
    assert result["Fb"] == pytest.approx(result["tie_bottom_height"] - result["h_increased"])

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_rect_aqueduct_import_preserves_bottom_width_without_warning_when_ratio_missing(monkeypatch):
    captured_kwargs = []
    original_calc = batch_panel_mod.ducao_rect_calculate

    def _capture_calc(*args, **kwargs):
        captured_kwargs.append(dict(kwargs))
        return original_calc(*args, **kwargs)

    monkeypatch.setattr(batch_panel_mod, "ducao_rect_calculate", _capture_calc)

    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_rect_aqueduct_row(ratio=""))
    panel._set_excel_import_session_active(True)
    panel._mark_row_as_excel_imported(0, True)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    status_col = panel.result_table.columnCount() - 1
    result = panel.batch_results[0]["result"]

    assert captured_kwargs
    assert captured_kwargs[0]["manual_B"] == 1.5
    assert panel.result_table.item(0, 4).text() == "1.500"
    assert panel.result_table.item(0, status_col).text() == "✓ 成功"
    assert result["preserved_manual_b"] is True
    assert result["used_manual_b"] is True
    assert result.get("imported_depth_width_ratio") is None
    assert result["constraint_warnings"] == []

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_rect_aqueduct_import_ratio_within_tolerance_keeps_success_status(monkeypatch):
    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_rect_aqueduct_row(ratio="0.61"))
    panel._set_excel_import_session_active(True)
    panel._mark_row_as_excel_imported(0, True)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    status_col = panel.result_table.columnCount() - 1
    result = panel.batch_results[0]["result"]

    assert panel.result_table.item(0, 4).text() == "1.500"
    assert panel.result_table.item(0, status_col).text() == "✓ 成功"
    assert result["preserved_manual_b"] is True
    assert result["used_manual_b"] is True
    assert result["constraint_warnings"] == []

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_rect_aqueduct_import_lock_failure_disables_exports(monkeypatch):
    registered_rows = []
    captured_kwargs = []
    original_calc = batch_panel_mod.ducao_rect_calculate

    class _SharedManager:
        def clear_batch_results(self):
            return None

        def register_batch_results(self, rows):
            registered_rows.extend(rows)
            return len(rows)

    def _capture_calc(*args, **kwargs):
        captured_kwargs.append(dict(kwargs))
        return original_calc(*args, **kwargs)

    monkeypatch.setattr(batch_panel_mod, "SHARED_DATA_AVAILABLE", True)
    monkeypatch.setattr(batch_panel_mod, "get_shared_data_manager", lambda: _SharedManager())
    monkeypatch.setattr(batch_panel_mod, "ducao_rect_calculate", _capture_calc)

    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_rect_aqueduct_row(q="1000", b="0.01", ratio="0.8"))
    panel._set_excel_import_session_active(True)
    panel._mark_row_as_excel_imported(0, True)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    status_col = panel.result_table.columnCount() - 1
    status_text = panel.result_table.item(0, status_col).text()

    assert captured_kwargs
    assert captured_kwargs[0]["manual_B"] == 0.01
    assert panel._has_calc_errors is True
    assert "无法计算水深" in status_text
    assert panel._btn_export_excel.isEnabled() is False
    assert panel._btn_export_word.isEnabled() is False
    assert registered_rows == []

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_rect_aqueduct_without_import_lock_keeps_ratio_search_behavior(monkeypatch):
    captured_kwargs = []
    original_calc = batch_panel_mod.ducao_rect_calculate

    def _capture_calc(*args, **kwargs):
        captured_kwargs.append(dict(kwargs))
        return original_calc(*args, **kwargs)

    monkeypatch.setattr(batch_panel_mod, "ducao_rect_calculate", _capture_calc)

    panel = _prepare_panel(monkeypatch)
    _set_single_row(panel, _build_rect_aqueduct_row(ratio="0.8"))
    panel._set_excel_import_session_active(False)
    panel.inc_cb.setChecked(False)
    panel.detail_cb.setChecked(False)

    panel._batch_calculate()
    _flush_events(6)

    status_col = panel.result_table.columnCount() - 1
    result = panel.batch_results[0]["result"]

    assert captured_kwargs
    assert captured_kwargs[0].get("manual_B") is None
    assert panel.result_table.item(0, 4).text() == "1.310"
    assert panel.result_table.item(0, status_col).text() == "✓ 成功"
    assert result.get("preserved_manual_b") in (None, False)

    panel.close()
    panel.deleteLater()
    _flush_events(4)


def test_batch_excel_template_contains_tie_rod_height_header():
    from openpyxl import load_workbook

    template_path = ROOT / "data" / "多流量段批量计算_导入Excel（模板）.xlsx"
    workbook = load_workbook(template_path, read_only=True, data_only=True)
    worksheet = workbook.active

    header_values = []
    for row in worksheet.iter_rows(min_row=1, max_row=8, values_only=True):
        header_values.extend(str(value).strip() for value in row if value is not None)

    assert "拉杆高度(m)" in header_values


def test_official_blank_template_contains_beginner_guide_sheets():
    """正式空模板应在导入表后提供按结构分类的新手填写说明。"""
    from openpyxl import load_workbook

    template_path = ROOT / "data" / "多流量段批量计算_导入Excel（模板）.xlsx"
    workbook = load_workbook(template_path, read_only=True, data_only=True)

    expected_sheets = [
        "导入模板",
        "0-先看这里",
        "1-公共字段",
        "2-明渠",
        "3-渡槽",
        "4-隧洞",
        "5-暗涵",
        "6-承压与闸类",
        "7-字段速查",
    ]

    assert workbook.sheetnames[: len(expected_sheets)] == expected_sheets
    assert workbook.active.title == "导入模板"


def test_official_blank_template_guides_cover_all_section_types():
    """说明页应覆盖批量面板支持的全部结构形式，避免新手选型后找不到填写规则。"""
    from openpyxl import load_workbook

    template_path = ROOT / "data" / "多流量段批量计算_导入Excel（模板）.xlsx"
    workbook = load_workbook(template_path, read_only=True, data_only=True)
    guide_text_parts = []
    for sheet_name in workbook.sheetnames:
        if sheet_name == "导入模板":
            continue
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(values_only=True):
            for value in row:
                if value is not None:
                    guide_text_parts.append(str(value))
    guide_text = "\n".join(guide_text_parts)

    for section_type in batch_panel_mod.SECTION_TYPES:
        assert section_type in guide_text
    assert "黄色=必须填" in guide_text
    assert "5.5,,7.25" in guide_text
    assert "倒角角度(°) 表示外倾角 α" in guide_text


def test_official_blank_template_field_lookup_covers_all_input_headers():
    """字段速查页应逐项解释当前全部导入表头。"""
    from openpyxl import load_workbook

    template_path = ROOT / "data" / "多流量段批量计算_导入Excel（模板）.xlsx"
    workbook = load_workbook(template_path, read_only=True, data_only=True)
    worksheet = workbook["7-字段速查"]

    first_col_values = {
        str(row[0]).strip()
        for row in worksheet.iter_rows(min_row=1, values_only=True)
        if row and row[0] is not None
    }

    for header in batch_panel_mod.INPUT_HEADERS:
        assert header in first_col_values


def test_official_blank_template_section_type_column_has_dropdown_hint():
    """导入模板的结构形式列应带下拉提示，但不能改变原有表头顺序。"""
    from openpyxl import load_workbook

    template_path = ROOT / "data" / "多流量段批量计算_导入Excel（模板）.xlsx"
    workbook = load_workbook(template_path, read_only=False, data_only=True)
    worksheet = workbook["导入模板"]
    headers = [worksheet.cell(row=2, column=col).value for col in range(1, len(batch_panel_mod.INPUT_HEADERS) + 1)]

    assert headers == batch_panel_mod.INPUT_HEADERS
    assert any(
        "D3:D500" in str(validation.sqref)
        for validation in worksheet.data_validations.dataValidation
    )


def _rgb(cell):
    color = cell.fill.fgColor
    return color.rgb if color.type == "rgb" else str(color.indexed)


def _border_styles(cell):
    return (
        cell.border.left.style,
        cell.border.right.style,
        cell.border.top.style,
        cell.border.bottom.style,
    )


def test_official_blank_template_import_sheet_formatting_is_consistent():
    """导入模板应统一表头、数据区和打开体验，避免重排列后出现局部断层。"""
    from openpyxl import load_workbook

    template_path = ROOT / "data" / "多流量段批量计算_导入Excel（模板）.xlsx"
    workbook = load_workbook(template_path, read_only=False, data_only=True)
    worksheet = workbook["导入模板"]
    headers = [worksheet.cell(row=2, column=col).value for col in range(1, len(batch_panel_mod.INPUT_HEADERS) + 1)]

    assert workbook.active.title == "导入模板"
    assert worksheet.freeze_panes == "E3"
    assert worksheet.sheet_view.showGridLines is False
    assert worksheet.auto_filter.ref == "A2:AP500"
    assert headers == batch_panel_mod.INPUT_HEADERS
    assert "D3:D500" in {
        str(validation.sqref)
        for validation in worksheet.data_validations.dataValidation
    }

    for col in range(1, len(batch_panel_mod.INPUT_HEADERS) + 1):
        cell = worksheet.cell(row=2, column=col)
        assert _rgb(cell) == "FF4472C4"
        assert cell.font.bold is True
        assert cell.font.color is not None
        assert cell.font.color.type == "rgb"
        assert cell.font.color.rgb == "FFFFFFFF"
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"
        assert cell.alignment.wrap_text is True
        assert _border_styles(cell) == ("thin", "thin", "thin", "thin")

    for row in range(3, 501):
        expected_fill = "FFFFFFFF" if row % 2 == 1 else "FFF2F7FB"
        fills = {
            _rgb(worksheet.cell(row=row, column=col))
            for col in range(1, len(batch_panel_mod.INPUT_HEADERS) + 1)
        }
        assert fills == {expected_fill}
        for col in range(1, len(batch_panel_mod.INPUT_HEADERS) + 1):
            cell = worksheet.cell(row=row, column=col)
            assert _border_styles(cell) == ("thin", "thin", "thin", "thin")
            assert cell.alignment.vertical == "center"


def test_official_blank_template_guide_sheet_formatting_is_consistent():
    """说明页应隐藏网格线、冻结表头，并保持字段速查顺序。"""
    from openpyxl import load_workbook

    template_path = ROOT / "data" / "多流量段批量计算_导入Excel（模板）.xlsx"
    workbook = load_workbook(template_path, read_only=False, data_only=True)
    expected_sheet_names = [
        "导入模板",
        "0-先看这里",
        "1-公共字段",
        "2-明渠",
        "3-渡槽",
        "4-隧洞",
        "5-暗涵",
        "6-承压与闸类",
        "7-字段速查",
    ]

    assert workbook.sheetnames[: len(expected_sheet_names)] == expected_sheet_names
    for sheet_name in expected_sheet_names[1:]:
        worksheet = workbook[sheet_name]
        assert worksheet.sheet_view.showGridLines is False
        assert worksheet.freeze_panes == "A7"
        assert _rgb(worksheet["A1"]) == "FF1F4E79"
        assert _rgb(worksheet["A6"]) == "FF4472C4"

    lookup = workbook["7-字段速查"]
    lookup_fields = [
        lookup.cell(row=row, column=1).value
        for row in range(7, 7 + len(batch_panel_mod.INPUT_HEADERS))
    ]
    assert lookup_fields == list(batch_panel_mod.INPUT_HEADERS)
    assert lookup.auto_filter.ref == "A6:G48"
