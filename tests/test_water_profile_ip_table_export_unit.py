# -*- coding: utf-8 -*-
"""IP 坐标及弯道参数表导出回归测试。"""

from pathlib import Path
import importlib.util
import sys
from types import SimpleNamespace

import openpyxl


def _load_cad_tools():
    """加载 cad_tools 模块，供独立单元测试复用。"""
    root = Path(__file__).resolve().parents[1]
    root_str = str(root)
    if root_str not in sys.path:
        sys.path.insert(0, root_str)
    matches = list(root.glob("*/water_profile/cad_tools.py"))
    assert matches, "未找到 cad_tools.py"
    spec = importlib.util.spec_from_file_location("cad_tools_ip_table_test_mod", matches[0])
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


cad_tools = _load_cad_tools()


class _FakeTextEntity:
    """记录文字实体，避免真实依赖 ezdxf。"""

    def set_placement(self, *_args, **_kwargs):
        """保持与 ezdxf 文本实体一致的调用口径。"""
        return self


class _FakeModelspace:
    """最小化 modelspace 桩对象。"""

    def __init__(self):
        self.texts = []
        self.lines = []

    def add_line(self, start, end, dxfattribs=None):
        """记录线段绘制请求。"""
        self.lines.append((start, end, dict(dxfattribs or {})))
        return None

    def add_text(self, text, dxfattribs=None):
        """记录文字绘制请求。"""
        self.texts.append((str(text), dict(dxfattribs or {})))
        return _FakeTextEntity()


def _make_node(**overrides):
    """构造一个满足 IP 表导出所需字段的节点。"""
    base = {
        "ip_number": 3,
        "name": "江支",
        "x": 634292.4718,
        "y": 3445320.3641,
        "station_BC": 0.0,
        "station_MC": 39.497,
        "station_EC": 39.497,
        "turn_angle": 8.66,
        "turn_radius": 0.0,
        "tangent_length": 0.0,
        "arc_length": 0.0,
        "bottom_elevation": 408.5,
        "water_level": 408.9,
        "structure_type": SimpleNamespace(value="明渠-矩形"),
        "in_out": "",
        "is_transition": False,
        "is_auto_inserted_channel": False,
    }
    base.update(overrides)
    return SimpleNamespace(**base)


def test_ip_table_preview_headers_include_design_water_level():
    assert cad_tools._get_ip_table_preview_headers() == [
        "IP点",
        "E（m）",
        "N（m）",
        "弯前(千米+米)",
        "里程(千米+米)",
        "弯末(千米+米)",
        "转角",
        "半径",
        "切线长",
        "弧长",
        "底高程(m)",
        "设计水位(m)",
    ]


def test_compute_ip_preview_data_appends_design_water_level_after_bottom_elevation():
    node = _make_node()

    preview_data, real_nodes = cad_tools._compute_ip_preview_data(
        [node],
        "",
        {"station_decimals": 2},
    )

    assert real_nodes == [node]
    assert len(preview_data[0]) == 12
    assert preview_data[0][3:6] == ["0+000.00", "0+039.50", "0+039.50"]
    assert preview_data[0][-2:] == ["408.500", "408.900"]


def test_compute_ip_preview_data_uses_dash_when_design_water_level_is_zero():
    node = _make_node(water_level=0.0)

    preview_data, _ = cad_tools._compute_ip_preview_data([node], "", {"station_decimals": 2})

    assert len(preview_data[0]) == 12
    assert preview_data[0][-1] == "-"


def test_compute_ip_preview_data_respects_custom_station_decimals():
    node = _make_node()

    preview_data, _ = cad_tools._compute_ip_preview_data([node], "", {"station_decimals": 3})

    assert preview_data[0][3:6] == ["0+000.000", "0+039.497", "0+039.497"]


def test_draw_ip_table_on_msp_writes_design_water_level_header_and_value(monkeypatch):
    fake_ezdxf = SimpleNamespace(
        enums=SimpleNamespace(
            TextEntityAlignment=SimpleNamespace(MIDDLE_CENTER="MIDDLE_CENTER")
        )
    )
    monkeypatch.setitem(sys.modules, "ezdxf", fake_ezdxf)
    msp = _FakeModelspace()
    preview_data, _ = cad_tools._compute_ip_preview_data([_make_node()], "", {"station_decimals": 2})

    cad_tools._draw_ip_table_on_msp(msp, 0.0, 0.0, preview_data)

    texts = [text for text, _attrs in msp.texts]
    assert "里程(千米+米)" in texts
    assert "0+039.50" in texts
    assert "设计水位(m)" in texts
    assert "408.900" in texts


def test_write_ip_table_excel_sheet_keeps_design_water_level_column():
    wb = openpyxl.Workbook()
    ws = wb.active

    cad_tools._write_ip_table_excel_sheet(ws, [_make_node()], "", {"station_decimals": 2})

    assert ws.max_column == 12
    assert ws["E2"].value == "里程(千米+米)"
    assert ws["D3"].value == "0+000.00"
    assert ws["E3"].value == "0+039.50"
    assert ws["F3"].value == "0+039.50"
    assert ws["L1"].value == "设计水位(m)"
    assert ws["L3"].value == 408.9
    assert ws["L3"].number_format == "0.000"


def test_write_ip_table_excel_sheet_respects_custom_station_decimals():
    wb = openpyxl.Workbook()
    ws = wb.active

    cad_tools._write_ip_table_excel_sheet(ws, [_make_node()], "", {"station_decimals": 3})

    assert ws["E3"].value == "0+039.497"


def test_compute_ip_preview_data_uses_display_ip_number_after_special_entry_exit_rows():
    special = _make_node(
        ip_number=5,
        display_ip_number=None,
        name="黄角坝",
        structure_type=SimpleNamespace(value="隧洞-圆形"),
        in_out="进",
    )
    normal = _make_node(
        ip_number=6,
        display_ip_number=5,
        name="",
        structure_type=SimpleNamespace(value="明渠-矩形"),
        in_out="",
    )

    preview_data, _ = cad_tools._compute_ip_preview_data([special, normal], "", {"station_decimals": 2})

    assert preview_data[0][0] == "黄角坝隧进"
    assert preview_data[1][0] == "IP5"
