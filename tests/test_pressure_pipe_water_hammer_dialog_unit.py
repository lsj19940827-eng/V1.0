# -*- coding: utf-8 -*-
"""基础水锤验算弹窗单元测试。"""

import os
import shutil
import sys
import tempfile
from pathlib import Path
from types import SimpleNamespace

import pytest
from PySide6.QtCore import Qt
from PySide6.QtTest import QTest
from PySide6.QtWidgets import QApplication

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(ROOT / "推求水面线") not in sys.path:
    sys.path.insert(0, str(ROOT / "推求水面线"))

from app_渠系计算前端.water_profile.water_profile_dialogs import (  # noqa: E402
    PressurePipeConfigDialog,
    PressurePipeWaterHammerDialog,
)
import app_渠系计算前端.water_profile.water_hammer_principle as principle_mod  # noqa: E402
from app_渠系计算前端.water_profile.water_hammer_export import (  # noqa: E402
    build_water_hammer_export_workbook,
    water_hammer_exportable_detail_count,
)
import app_渠系计算前端.water_profile.water_profile_dialogs as dialogs_mod  # noqa: E402
from 推求水面线.managers.pressure_pipe_manager import PressurePipeConfig, PressurePipeManager  # noqa: E402
from 推求水面线.models.enums import StructureType  # noqa: E402


def _get_qapp():
    """返回测试可复用的 QApplication。"""
    return QApplication.instance() or QApplication([])


def _flush_events(rounds: int = 4):
    """推动事件循环，确保控件状态刷新。"""
    app = _get_qapp()
    for _ in range(max(1, rounds)):
        app.processEvents()


def _make_group():
    """构造最小可用的有压管道分组。"""
    rows = [
        SimpleNamespace(
            section_params={"D": 1.2},
            turn_radius=0.0,
            flow_section="1",
            water_level=101.25,
        ),
    ]
    return SimpleNamespace(
        name="",
        display_name="流量段1 第5行有压管道",
        storage_key="flow1-row5",
        identity="flow1-row5",
        group_mode="unnamed_row_segment",
        design_flow=1.52,
        diameter=1.2,
        material_key="钢管",
        ip_points=[{"x": 0.0, "y": 0.0}, {"x": 210.0, "y": 0.0}],
        rows=rows,
        row_indices=[4],
        target_row_index=4,
        upstream_row_index=3,
    )


def _make_manager(project_path: Path, group) -> PressurePipeManager:
    """创建带已有专项结果的管理器。"""
    manager = PressurePipeManager(str(project_path))
    cfg = PressurePipeConfig(
        name=group.display_name,
        Q=group.design_flow,
        D=group.diameter,
        material_key=group.material_key,
        pipe_velocity=1.34,
        plan_total_length=210.0,
    )
    manager.set_pipe_config(group.storage_key, cfg)
    return manager


def _make_route_group(
    key: str,
    row_index: int,
    structure_type: str = "有压管道",
    *,
    route_key: str = "flow1-route1",
    start_mc: float = 0.0,
    end_mc: float = 10.0,
    diameter: float = 1.0,
    design_flow: float = 1.0,
    material_key: str = "钢管",
    pipe_velocity: float = 1.0,
):
    """构造整线水锤测试用分组。"""
    row = SimpleNamespace(
        structure_type=StructureType.from_string(structure_type),
        section_params={"D": diameter, "pipe_material": material_key},
        flow=design_flow,
        water_level=100.0 + row_index,
    )
    group = SimpleNamespace(
        name=key,
        display_name=key,
        storage_key=key,
        identity=key,
        route_key=route_key,
        route_display_name="测试整线",
        route_start_row_index=0,
        route_end_row_index=99,
        route_start_mc=0.0,
        route_end_mc=100.0,
        structure_type=structure_type,
        rows=[row],
        row_indices=[row_index],
        target_row_index=row_index,
        upstream_row_index=max(row_index - 1, 0),
        segment_start_mc=start_mc,
        segment_end_mc=end_mc,
        group_mode="unnamed_row_segment",
        design_flow=design_flow,
        diameter=diameter,
        material_key=material_key,
        ip_points=[{"x": start_mc, "y": 0.0}, {"x": end_mc, "y": 0.0}],
        plan_total_length=max(0.0, end_mc - start_mc),
    )
    group.pipe_velocity = pipe_velocity
    return group


def _make_route_dialog(groups, manager=None):
    """创建整线模式弹窗并刷新事件。"""
    _get_qapp()
    dialog = PressurePipeWaterHammerDialog(
        pipe_groups=groups,
        manager=manager,
        pressure_chains=[],
        xxpipe_route_mode=True,
    )
    dialog.show()
    _flush_events(6)
    return dialog


def _read_float(widget) -> float:
    """把输入框文本安全转成浮点数。"""
    return float(str(widget.text() or "0").strip())


def _set_route_member_a0(widgets, member_key: str, value: float):
    """在整线成员表中填写指定成员的a0。"""
    table = widgets["water_hammer_member_table"]
    rows = widgets.get("water_hammer_member_a0_rows", {})
    column = widgets.get("water_hammer_member_a0_column", -1)
    row = rows[str(member_key)]
    item = table.item(row, column)
    assert item is not None
    item.setText(str(value))


def _set_route_profile_and_water_levels(dialog, groups, *, centerline_elevation: float, water_level: float):
    """给整线测试补入纵断面中心线和表3水位。"""
    route_key = "flow1-route1"
    start_mc = min(float(getattr(group, "segment_start_mc", 0.0)) for group in groups)
    end_mc = max(float(getattr(group, "segment_end_mc", 0.0)) for group in groups)
    dialog._longitudinal_data[route_key] = [
        {"chainage": start_mc, "elevation": centerline_elevation},
        {"chainage": end_mc, "elevation": centerline_elevation},
    ]
    for group in groups:
        for row in group.rows:
            row.water_level = water_level


def test_longitudinal_import_rejects_plan_coordinate_y_as_elevation():
    """纵断面导入应拦截明显来自平面坐标的 Y 值。"""
    nodes = [
        SimpleNamespace(chainage=14193.795, elevation=3376075.2),
        SimpleNamespace(chainage=19945.706, elevation=3376021.6),
    ]

    with pytest.raises(ValueError) as excinfo:
        PressurePipeConfigDialog._raise_if_import_stays_in_raw_coordinate_space(
            "文家梁连续整线",
            [],
            nodes,
        )

    message = str(excinfo.value)
    assert "文家梁连续整线" in message
    assert "疑似平面坐标 Y 值" in message
    assert "不是真实高程" in message
    assert "X=桩号、Y=管道中心线真实高程" in message


def test_longitudinal_import_accepts_realistic_centerline_elevation():
    """合理高程的纵断面节点不应被平面坐标校验拦截。"""
    nodes = [
        SimpleNamespace(chainage=14193.795, elevation=422.356),
        SimpleNamespace(chainage=19945.706, elevation=414.419),
    ]

    PressurePipeConfigDialog._raise_if_import_stays_in_raw_coordinate_space(
        "文家梁连续整线",
        [],
        nodes,
    )


def test_dialog_prefills_and_persists_basic_water_hammer_inputs_and_results():
    """弹窗应能预填、验算、保存并在重开后恢复基础水锤数据。"""
    _get_qapp()
    case_dir = Path(tempfile.mkdtemp(prefix="wh_dialog_"))
    project_path = case_dir / "demo.qxproj"
    group = _make_group()

    try:
        manager = _make_manager(project_path, group)
        dialog = PressurePipeWaterHammerDialog(pipe_groups=[group], manager=manager)
        dialog.show()
        _flush_events(6)

        widgets = dialog._card_widgets[group.storage_key]
        assert _read_float(widgets["water_hammer_length_edit"]) == pytest.approx(210.0)
        assert _read_float(widgets["water_hammer_velocity_edit"]) == pytest.approx(1.34)
        assert _read_float(widgets["water_hammer_head_edit"]) == pytest.approx(101.25)
        assert _read_float(widgets["water_hammer_elastic_modulus_edit"]) > 0
        assert _read_float(widgets["water_hammer_wall_thickness_edit"]) == pytest.approx(0.06)
        assert _read_float(widgets["water_hammer_closing_time_edit"]) == pytest.approx(300.0)

        widgets["water_hammer_wall_thickness_edit"].setText("0.016")
        widgets["water_hammer_head_edit"].setText("102.4")
        widgets["water_hammer_closing_time_edit"].setText("0.25")
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        assert "可计算" in widgets["water_hammer_status_label"].text()
        assert widgets["water_hammer_result_hmax_label"].text() != "-"
        assert widgets["water_hammer_result_hmin_label"].text() != "-"
        assert widgets["water_hammer_result_control_type_label"].text()
        assert widgets["water_hammer_result_diagram_type_label"].text() != "-"
        assert "图1-3-3" in widgets["water_hammer_result_diagram_type_label"].text()

        dialog.accept()
        _flush_events(2)

        reloaded_manager = PressurePipeManager(str(project_path))
        loaded = reloaded_manager.get_pipe_config(group.storage_key)
        assert loaded is not None
        assert loaded.wall_thickness_m == pytest.approx(0.016)
        assert loaded.water_hammer_basic["status"] == "可计算"
        assert loaded.water_hammer_basic["inputs"]["closing_time_s"] == pytest.approx(0.25)

        dialog_reopen = PressurePipeWaterHammerDialog(pipe_groups=[group], manager=reloaded_manager)
        dialog_reopen.show()
        _flush_events(6)

        widgets_reopen = dialog_reopen._card_widgets[group.storage_key]
        assert _read_float(widgets_reopen["water_hammer_wall_thickness_edit"]) == pytest.approx(0.016)
        assert _read_float(widgets_reopen["water_hammer_head_edit"]) == pytest.approx(102.4)
        assert _read_float(widgets_reopen["water_hammer_closing_time_edit"]) == pytest.approx(0.25)
        assert "可计算" in widgets_reopen["water_hammer_status_label"].text()
        assert widgets_reopen["water_hammer_result_hmax_label"].text() != "-"
        assert widgets_reopen["water_hammer_result_hmin_label"].text() != "-"
        assert widgets_reopen["water_hammer_result_diagram_type_label"].text() != "-"

        dialog.close()
        dialog_reopen.close()
    finally:
        shutil.rmtree(case_dir, ignore_errors=True)


def test_basic_water_hammer_exempt_result_shows_blank_values():
    """单管满足免验算时界面只显示结论，不展示水击数值。"""
    _get_qapp()
    group = _make_group()
    dialog = PressurePipeWaterHammerDialog(pipe_groups=[group])
    try:
        dialog.show()
        _flush_events(6)
        widgets = dialog._card_widgets[group.storage_key]
        widgets["water_hammer_wall_thickness_edit"].setText("0.016")
        widgets["water_hammer_head_edit"].setText("102.4")
        widgets["water_hammer_closing_time_edit"].setText("300")

        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        assert "可不验算" in widgets["water_hammer_status_label"].text()
        assert widgets["water_hammer_result_delta_h_label"].text() == "-"
        assert widgets["water_hammer_result_hmax_label"].text() == "-"
        assert widgets["water_hammer_result_hmin_label"].text() == "-"
        assert widgets["water_hammer_result_control_type_label"].text() == "-"
        assert widgets["water_hammer_result_diagram_type_label"].text() == "-"
    finally:
        dialog.close()


def test_basic_water_hammer_pccp_defaults_cp_without_a0_and_persists_a0():
    """PCCP单管缺少a0时按cp=1继续验算，补充后应保存并能重开恢复。"""
    _get_qapp()
    case_dir = Path(tempfile.mkdtemp(prefix="wh_basic_a0_"))
    project_path = case_dir / "demo.qxproj"
    group = _make_group()
    group.material_key = "PCCP管"

    try:
        manager = _make_manager(project_path, group)
        cfg = manager.get_pipe_config(group.storage_key)
        cfg.material_key = "PCCP管"
        manager.set_pipe_config(group.storage_key, cfg)

        dialog = PressurePipeWaterHammerDialog(pipe_groups=[group], manager=manager)
        dialog.show()
        _flush_events(6)
        widgets = dialog._card_widgets[group.storage_key]
        assert "water_hammer_a0_edit" in widgets
        widgets["water_hammer_wall_thickness_edit"].setText("0.08")
        widgets["water_hammer_head_edit"].setText("120")
        widgets["water_hammer_closing_time_edit"].setText("0.2")

        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)
        assert "可计算" in widgets["water_hammer_status_label"].text()
        assert "未填写 a0" in widgets["water_hammer_status_label"].text()
        assert "cp=1" in widgets["water_hammer_status_label"].text()
        assert widgets["water_hammer_result_cp_label"].text() == "1.0000"
        assert widgets["water_hammer_result_delta_h_label"].text() != "-"

        widgets["water_hammer_a0_edit"].setText("0.2")
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)
        assert "可计算" in widgets["water_hammer_status_label"].text()
        assert "未填写 a0" not in widgets["water_hammer_status_label"].text()
        assert widgets["water_hammer_result_cp_label"].text() != "-"
        assert widgets["water_hammer_result_cp_label"].text() != "1.0000"
        assert widgets["water_hammer_result_gbt_positive_label"].text() != "-"
        assert widgets["water_hammer_result_linear_positive_label"].text() != "-"
        assert widgets["water_hammer_result_governing_method_label"].text() != "-"

        dialog.accept()
        reloaded_manager = PressurePipeManager(str(project_path))
        loaded = reloaded_manager.get_pipe_config(group.storage_key)
        assert loaded.water_hammer_basic["inputs"]["reinforcement_ratio_a0"] == pytest.approx(0.2)

        reopened = PressurePipeWaterHammerDialog(pipe_groups=[group], manager=reloaded_manager)
        reopened.show()
        _flush_events(6)
        reopened_widgets = reopened._card_widgets[group.storage_key]
        assert _read_float(reopened_widgets["water_hammer_a0_edit"]) == pytest.approx(0.2)
        reopened.close()
        dialog.close()
    finally:
        shutil.rmtree(case_dir, ignore_errors=True)


def test_pressure_pipe_config_dialog_no_longer_shows_water_hammer_panel():
    """有压管道水力计算窗口不再承载水锤验算入口。"""
    _get_qapp()
    group = _make_group()

    dialog = PressurePipeConfigDialog(pipe_groups=[group])
    try:
        dialog.show()
        _flush_events(6)

        widgets = dialog._card_widgets[group.storage_key]
        assert "water_hammer_panel" not in widgets
    finally:
        dialog.close()


def test_pressure_pipe_water_hammer_dialog_shows_only_water_hammer_panel():
    """水锤专用窗口应显示水锤验算区，不显示水力计算的R/D设置区。"""
    _get_qapp()
    group = _make_group()

    dialog = PressurePipeWaterHammerDialog(pipe_groups=[group])
    try:
        dialog.show()
        _flush_events(6)

        widgets = dialog._card_widgets[group.storage_key]
        assert "water_hammer_panel" in widgets
        assert "turn_n_edit" not in widgets
        assert widgets["water_hammer_calc_btn"].property("waterHammerAction") is True
        assert widgets["water_hammer_calc_btn"].cursor().shape() == Qt.PointingHandCursor
        assert dialog.windowTitle() == "有压管道水锤验算"
    finally:
        dialog.close()


def test_route_water_hammer_segments_split_at_tunnel_and_keep_pressure_runs_whole():
    """整线水锤段应按连续有压段生成，并被无压隧洞切断。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0, pipe_velocity=0.8),
        _make_route_group("pipe-b", 1, "定向钻", start_mc=10.0, end_mc=30.0, pipe_velocity=1.2),
        _make_route_group("tunnel", 2, "隧洞-圆形", start_mc=30.0, end_mc=80.0),
        _make_route_group("pipe-c", 3, "有压管道", start_mc=80.0, end_mc=120.0, pipe_velocity=0.9),
        _make_route_group("pipe-d", 4, "顶管", start_mc=120.0, end_mc=150.0, pipe_velocity=1.6),
    ]

    dialog = _make_route_dialog(groups)
    try:
        segments = dialog._build_route_water_hammer_segments(dialog._route_contexts["flow1-route1"])

        assert [segment["segment_key"] for segment in segments] == [
            "flow1-route1::pipe-a::pipe-b",
            "flow1-route1::pipe-c::pipe-d",
        ]
        assert [segment["member_keys"] for segment in segments] == [
            ["pipe-a", "pipe-b"],
            ["pipe-c", "pipe-d"],
        ]
        assert [segment["length_m"] for segment in segments] == pytest.approx([30.0, 70.0])
    finally:
        dialog.close()


def test_route_water_hammer_mixed_params_show_member_table_without_representative_fields():
    """混合参数连续有压段应通过成员表展示参数，不再显示代表行。"""
    groups = [
        _make_route_group(
            "pipe-a",
            0,
            "有压管道",
            start_mc=0.0,
            end_mc=10.0,
            diameter=0.8,
            design_flow=0.8,
            material_key="HDPE管",
            pipe_velocity=0.9,
        ),
        _make_route_group(
            "pipe-b",
            1,
            "有压管道",
            start_mc=10.0,
            end_mc=25.0,
            diameter=0.5,
            design_flow=0.8,
            material_key="钢管",
            pipe_velocity=1.7,
        ),
    ]

    dialog = _make_route_dialog(groups)
    try:
        segments = dialog._build_route_water_hammer_segments(dialog._route_contexts["flow1-route1"])

        assert len(segments) == 1
        segment = segments[0]
        assert segment["mixed_params"] is True
        assert "D" in segment["param_summary"]
        assert "管材" in segment["param_summary"]

        route_refs = dialog._route_widgets["flow1-route1"]
        widgets = route_refs["water_hammer_segment_widgets"][0]
        assert "water_hammer_representative_combo" not in widgets
        assert "water_hammer_diameter_edit" not in widgets
        assert "water_hammer_velocity_edit" not in widgets
        assert "water_hammer_head_edit" not in widgets
        assert "water_hammer_elastic_modulus_edit" not in widgets
        assert widgets["water_hammer_segment_title_label"].text() == "整线连续有压段 | 第1行 ~ 第2行 | 2 个计算成员"
        assert "每个成员使用自己的管径、流量、管材、弹性模量和流速" in (
            widgets["water_hammer_segment_hint_label"].text()
        )
        member_table = widgets["water_hammer_member_table"]
        assert member_table.rowCount() == 2
        headers = [member_table.horizontalHeaderItem(i).text() for i in range(member_table.columnCount())]
        assert headers == [
            "序号",
            "行范围",
            "桩号范围",
            "L(m)",
            "Q(m³/s)",
            "D(m)",
            "管材",
            "E(N/m²)",
            "v0(m/s)",
            "a0",
            "计算用途",
        ]
        assert member_table.item(0, 1).text() == "第1行"
        assert member_table.item(1, 1).text() == "第2行"
        assert member_table.item(0, 10).text() == "参与整线水锤验算"
    finally:
        dialog.close()


def test_route_water_hammer_mixed_material_members_use_own_elastic_modulus():
    """混合管材整线仍按一段验算，但每个成员使用自己的管材 E。"""
    groups = [
        _make_route_group(
            "pipe-pccp",
            0,
            "有压管道",
            start_mc=0.0,
            end_mc=10.0,
            diameter=1.0,
            design_flow=1.0,
            material_key="PCCP管",
            pipe_velocity=1.0,
        ),
        _make_route_group(
            "pipe-ductile",
            1,
            "有压管道",
            start_mc=10.0,
            end_mc=20.0,
            diameter=1.0,
            design_flow=1.0,
            material_key="球墨铸铁管",
            pipe_velocity=1.2,
        ),
    ]

    dialog = _make_route_dialog(groups)
    try:
        segments = dialog._build_route_water_hammer_segments(dialog._route_contexts["flow1-route1"])

        assert len(segments) == 1
        assert segments[0]["member_keys"] == ["pipe-pccp", "pipe-ductile"]
        candidates = {candidate["key"]: candidate for candidate in segments[0]["representative_candidates"]}
        assert candidates["pipe-pccp"]["material_key"] == "PCCP管"
        assert candidates["pipe-pccp"]["resolved_material_key"] == "预应力钢筒混凝土管"
        assert candidates["pipe-pccp"]["elastic_modulus_pa"] == pytest.approx(20.6e9)
        assert candidates["pipe-ductile"]["resolved_material_key"] == "球墨铸铁管"
        assert candidates["pipe-ductile"]["elastic_modulus_pa"] == pytest.approx(160.0e9)

        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=400.0)
        widgets = dialog._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        members = dialog._build_route_water_hammer_distribution_members(widgets["segment_key"])
        assert [member["elastic_modulus_pa"] for member in members] == pytest.approx([20.6e9, 160.0e9])
        assert members[0]["reinforcement_ratio_a0"] is None

        widgets["water_hammer_wall_thickness_edit"].setText("0.02")
        widgets["water_hammer_closing_time_edit"].setText("0.01")
        dialog._calculate_route_water_hammer_segment(widgets["segment_key"])
        _flush_events(6)

        assert "缺少有效弹性模量 E" not in widgets["water_hammer_status_label"].text()
        assert "未填写 a0" in widgets["water_hammer_status_label"].text()
        assert widgets["water_hammer_result_sample_count_label"].text() != "0"
        assert widgets["water_hammer_result_gbt_positive_label"].text() != "-"
        assert widgets["water_hammer_result_linear_positive_label"].text() != "-"
        assert widgets["water_hammer_result_governing_method_label"].text() != "-"
        details = widgets["water_hammer_result"].get("details", [])
        pccp_detail = next(item for item in details if item.get("member_key") == "pipe-pccp")
        assert pccp_detail["pipe_coefficient_cp"] == pytest.approx(1.0)
        assert pccp_detail["reinforcement_ratio_a0"] is None
    finally:
        dialog.close()


def test_route_water_hammer_exempt_result_shows_blank_values_and_no_details():
    """整线满足免验算时界面不展示水击数值和采样明细。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0),
    ]

    dialog = _make_route_dialog(groups)
    try:
        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=400.0)
        widgets = dialog._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        widgets["water_hammer_wall_thickness_edit"].setText("0.02")
        widgets["water_hammer_closing_time_edit"].setText("100")

        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        assert "可不验算" in widgets["water_hammer_status_label"].text()
        assert widgets["water_hammer_result_control_delta_label"].text() == "-"
        assert widgets["water_hammer_result_negative_delta_label"].text() == "-"
        assert widgets["water_hammer_result_min_margin_label"].text() == "-"
        assert widgets["water_hammer_result_exceed_count_label"].text() == "-"
        assert widgets["water_hammer_result_negative_risk_count_label"].text() == "-"
        assert widgets["water_hammer_result_sample_count_label"].text() == "-"
        assert widgets["water_hammer_result_conclusion_label"].text() == "可不验算"
        assert widgets["water_hammer_detail_table"].rowCount() == 0
    finally:
        dialog.close()


def test_route_only_dialog_shows_water_hammer_segment_panel_without_child_pipe_cards():
    """route-only 模式下整线卡应直接显示水锤段入口。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0),
        _make_route_group("pipe-b", 1, "有压管道", start_mc=10.0, end_mc=20.0),
    ]

    dialog = _make_route_dialog(groups)
    try:
        route_refs = dialog._route_widgets["flow1-route1"]
        segment_widgets = route_refs["water_hammer_segment_widgets"]

        assert dialog._card_widgets == {}
        assert len(segment_widgets) == 1
        assert segment_widgets[0]["segment_key"] == "flow1-route1::pipe-a::pipe-b"
        assert segment_widgets[0]["water_hammer_calc_btn"].text() == "验算/刷新"
        assert _read_float(segment_widgets[0]["water_hammer_wall_thickness_edit"]) == pytest.approx(0.06)
        assert _read_float(segment_widgets[0]["water_hammer_closing_time_edit"]) == pytest.approx(300.0)
        assert _read_float(segment_widgets[0]["water_hammer_allowable_pressure_edit"]) == pytest.approx(1.0)
        assert _read_float(route_refs["water_hammer_bulk_e_edit"]) == pytest.approx(0.06)
        assert _read_float(route_refs["water_hammer_bulk_ts_edit"]) == pytest.approx(300.0)
        assert _read_float(route_refs["water_hammer_bulk_pressure_edit"]) == pytest.approx(1.0)
        assert route_refs["water_hammer_bulk_hint_label"].text() == "e、Ts 和允许压力为整线级输入；每个水锤段共用一套。"
        assert route_refs["water_hammer_bulk_all_btn"].text() == "填到全部水锤段"
        assert "water_hammer_bulk_same_material_btn" not in route_refs
        assert route_refs["water_hammer_export_excel_btn"].text() == "导出Excel"
        assert route_refs["water_hammer_principle_btn"].text() == "验算原理"
    finally:
        dialog.close()


def test_route_water_hammer_principle_dialog_shows_formula_and_control_point_values():
    """验算原理窗口应展示公式，并在有结果时展示控制采样点代入值。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0, diameter=1.0),
    ]

    dialog = _make_route_dialog(groups)
    try:
        route_refs = dialog._route_widgets["flow1-route1"]
        widgets = route_refs["water_hammer_segment_widgets"][0]

        dialog._show_route_water_hammer_principle("flow1-route1")
        _flush_events(4)
        principle_dialog = dialog._water_hammer_principle_dialog
        assert "水锤波速" in principle_dialog.html
        assert 'class="latex-source"' not in principle_dialog.html
        assert r"a=\frac{1435}" not in principle_dialog.html
        assert "1425" in principle_dialog.html or "GB/T 20203-2017 口径" in principle_dialog.html
        assert "GB/T 20203-2017 式(21)" in principle_dialog.html
        assert "双重取大" in principle_dialog.html
        assert r"\mu_s=2\sum_i" not in principle_dialog.html
        assert "记作 μ<sub>s</sub>" in principle_dialog.html
        assert "H<sub>0</sub> 为初始压强水头" in principle_dialog.html
        assert "T<sub>s</sub> 不大于水锤相时" in principle_dialog.html
        assert "先验算后可看到控制采样点代入值" in principle_dialog.html
        principle_dialog.close()

        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=400.0)
        widgets["water_hammer_wall_thickness_edit"].setText("0.02")
        widgets["water_hammer_closing_time_edit"].setText("0.01")
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        dialog._show_route_water_hammer_principle("flow1-route1")
        _flush_events(4)
        principle_dialog = dialog._water_hammer_principle_dialog
        assert "控制采样点代入值" in principle_dialog.html
        assert "整线统一输入" in principle_dialog.html
        assert "所属成员" in principle_dialog.html
        assert "GB/T ΔH<sup>+</sup>" in principle_dialog.html
        assert "线性启闭 ΔH<sup>+</sup>" in principle_dialog.html
        assert "ΔH<sup>+</sup>" in principle_dialog.html
        assert "H<sub>max</sub>" in principle_dialog.html
    finally:
        try:
            dialog._water_hammer_principle_dialog.close()
        except Exception:
            pass
        dialog.close()


def test_water_hammer_principle_formula_fallback_shows_source_when_svg_missing(monkeypatch):
    """公式渲染失败时应显示原公式兜底，不能显示 None。"""
    monkeypatch.setattr(principle_mod, "render_latex_svg", lambda *args, **kwargs: None)

    html = principle_mod.build_water_hammer_principle_html(
        route_name="",
        segment_name="",
        inputs={},
        result={},
    )

    assert "None" not in html
    assert r"a=\frac{1425}" in html
    assert "GB/T 20203-2017 式(21)" in html


def test_route_water_hammer_excel_exports_summary_and_calculated_segment_details(monkeypatch):
    """Excel应导出全部段汇总，并只为已验算段生成明细Sheet。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0, material_key="PCCP管"),
        _make_route_group("tunnel", 1, "隧洞-圆形", start_mc=10.0, end_mc=20.0),
        _make_route_group("pipe-b", 2, "有压管道", start_mc=20.0, end_mc=30.0),
    ]
    export_path = Path(tempfile.mkdtemp(prefix="wh_export_")) / "水锤明细.xlsx"

    dialog = _make_route_dialog(groups)
    try:
        from app_渠系计算前端 import export_utils as export_utils_mod

        monkeypatch.setattr(dialogs_mod, "fluent_info", lambda *args, **kwargs: None)
        monkeypatch.setattr(dialogs_mod, "fluent_error", lambda *args, **kwargs: None)
        monkeypatch.setattr(export_utils_mod, "ask_open_file", lambda *args, **kwargs: None)
        monkeypatch.setattr(
            dialogs_mod.QFileDialog,
            "getSaveFileName",
            lambda *args, **kwargs: (str(export_path), "Excel文件 (*.xlsx)"),
        )
        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=400.0)
        route_refs = dialog._route_widgets["flow1-route1"]
        first_segment = route_refs["water_hammer_segment_widgets"][0]
        first_segment["water_hammer_wall_thickness_edit"].setText("0.02")
        first_segment["water_hammer_closing_time_edit"].setText("0.01")
        QTest.mouseClick(first_segment["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        dialog._export_route_water_hammer_excel()
        _flush_events(2)

        import openpyxl

        workbook = openpyxl.load_workbook(export_path)
        assert workbook.sheetnames[0] == "汇总"
        assert len(workbook.sheetnames) == 2
        summary = workbook["汇总"]
        assert summary.cell(row=2, column=4).value == "已验算"
        assert summary.cell(row=3, column=4).value == "未验算"
        assert summary.cell(row=2, column=6).value == pytest.approx(1.0)
        assert summary.cell(row=2, column=7).value == pytest.approx(101.9368, rel=1e-4)
        assert summary.cell(row=2, column=8).value == "管底"
        assert summary.cell(row=2, column=15).value == pytest.approx(1.0)
        assert summary.cell(row=2, column=16).value is None
        assert "未填写 a0" in str(summary.cell(row=2, column=21).value or "")
        detail = workbook[workbook.sheetnames[1]]
        assert detail.cell(row=1, column=1).value == "桩号(m)"
        assert detail.cell(row=1, column=2).value == "所属成员"
        assert detail.cell(row=1, column=3).value == "D(m)"
        assert detail.cell(row=1, column=4).value == "v0(m/s)"
        assert detail.cell(row=1, column=5).value == "a(m/s)"
        assert detail.cell(row=1, column=6).value == "cp"
        assert detail.cell(row=1, column=7).value == "a0"
        assert detail.cell(row=1, column=12).value == "Hmax(m)"
        assert detail.cell(row=1, column=13).value == "Hmin(m)"
        assert detail.cell(row=1, column=14).value == "GB/T ΔH+(m)"
        assert detail.cell(row=1, column=15).value == "线性启闭 ΔH+(m)"
        assert detail.cell(row=1, column=22).value == "控制来源"
        assert str(detail.cell(row=2, column=2).value or "").startswith("pipe-a")
        assert detail.cell(row=2, column=3).value == pytest.approx(1.0)
        assert detail.cell(row=2, column=4).value == pytest.approx(1.0)
        assert detail.cell(row=2, column=5).value > 0
        assert detail.cell(row=2, column=6).value == pytest.approx(1.0)
        assert detail.cell(row=2, column=7).value is None
        assert detail.cell(row=2, column=12).value is not None
        assert detail.cell(row=2, column=13).value is not None
        assert detail.cell(row=2, column=22).value in {
            "GB/T 20203-2017 式(19)",
            "GB/T 20203-2017 式(21)",
            "线性启闭理论",
            "双重验算一致",
        }
        assert detail.max_row > 1
    finally:
        dialog.close()
        shutil.rmtree(export_path.parent, ignore_errors=True)


def test_route_water_hammer_excel_exports_exempt_segment_summary_without_detail_sheet():
    """全为免验算水锤段时也应导出汇总表，不生成明细Sheet。"""
    segments = [
        {
            "route_name": "整线A",
            "segment_name": "水锤段1",
            "result": {
                "status": "可不验算",
                "reason": "启闭时间 Ts 满足 GB/T 20203-2017 5.1.7.4：Ts >= 40L/a，可不验算关阀水击压力",
                "is_exempt": True,
                "exemption_threshold_s": 12.3,
                "details": [],
            },
        }
    ]

    assert water_hammer_exportable_detail_count(segments) == 1
    workbook = build_water_hammer_export_workbook(segments)

    assert workbook.sheetnames == ["汇总"]
    summary = workbook["汇总"]
    assert summary.cell(row=2, column=4).value == "已处理"
    assert summary.cell(row=2, column=5).value == "可不验算"
    summary_headers = [summary.cell(row=1, column=col).value for col in range(1, summary.max_column + 1)]
    assert "GB/T ΔH+(m)" in summary_headers
    assert "线性启闭 ΔH+(m)" in summary_headers
    assert "控制来源" in summary_headers
    note_col = summary_headers.index("备注") + 1
    assert "GB/T 20203-2017 5.1.7.4" in summary.cell(row=2, column=note_col).value


def test_route_water_hammer_excel_export_offers_open_file_after_save(monkeypatch):
    """导出水锤Excel成功后应让用户直接打开文件。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0),
    ]
    export_path = Path(tempfile.mkdtemp(prefix="wh_export_open_")) / "水锤明细.xlsx"
    opened_paths = []

    dialog = _make_route_dialog(groups)
    try:
        from app_渠系计算前端 import export_utils as export_utils_mod

        monkeypatch.setattr(dialogs_mod, "fluent_info", lambda *args, **kwargs: None)
        monkeypatch.setattr(dialogs_mod, "fluent_error", lambda *args, **kwargs: None)
        monkeypatch.setattr(
            dialogs_mod.QFileDialog,
            "getSaveFileName",
            lambda *args, **kwargs: (str(export_path), "Excel文件 (*.xlsx)"),
        )
        monkeypatch.setattr(
            export_utils_mod,
            "ask_open_file",
            lambda filepath, parent=None: opened_paths.append((Path(filepath), parent)),
        )
        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=400.0)
        widgets = dialog._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        widgets["water_hammer_wall_thickness_edit"].setText("0.02")
        widgets["water_hammer_closing_time_edit"].setText("0.01")
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        dialog._export_route_water_hammer_excel()
        _flush_events(2)

        assert opened_paths == [(export_path, dialog)]
    finally:
        dialog.close()
        shutil.rmtree(export_path.parent, ignore_errors=True)


def test_route_water_hammer_excel_skips_empty_export_without_save_dialog(monkeypatch):
    """全部水锤段都未验算时不应生成空Excel，也不应弹出保存路径。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0),
    ]
    notices = []

    dialog = _make_route_dialog(groups)
    try:
        monkeypatch.setattr(dialogs_mod, "fluent_info", lambda _parent, title, content: notices.append((title, content)))
        monkeypatch.setattr(
            dialogs_mod.QFileDialog,
            "getSaveFileName",
            lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("未验算时不应请求保存路径")),
        )

        dialog._export_route_water_hammer_excel()

        assert notices
        assert "没有可导出的水锤验算明细" in notices[-1][1]
    finally:
        dialog.close()


def test_route_water_hammer_result_layout_stays_stable_after_long_result_text():
    """验算后的长文本结果不应撑宽水锤窗口内容区。"""
    groups = [
        _make_route_group(
            "pipe-a",
            0,
            "有压管道",
            start_mc=0.0,
            end_mc=10436.427,
            diameter=0.4,
            design_flow=0.1,
            material_key="HDPE管",
            pipe_velocity=0.7958,
        ),
    ]

    dialog = _make_route_dialog(groups)
    try:
        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=397.16)
        widgets = dialog._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        widgets["water_hammer_wall_thickness_edit"].setText("0.1")
        widgets["water_hammer_closing_time_edit"].setText("300")
        _flush_events(6)

        before_width = dialog._resolve_content_size().width()
        calc_btn = widgets["water_hammer_calc_btn"]
        assert calc_btn.text() == "验算/刷新"
        assert calc_btn.property("waterHammerAction") is True
        assert calc_btn.cursor().shape() == Qt.PointingHandCursor

        QTest.mouseClick(calc_btn, Qt.LeftButton)
        _flush_events(8)

        after_width = dialog._resolve_content_size().width()
        assert after_width <= before_width + 80
        assert widgets["water_hammer_result_control_type_label"].text() != "-"
        assert "图1-3-3" in widgets["water_hammer_result_diagram_type_label"].text()
        assert widgets["water_hammer_result_control_type_label"].wordWrap() is True
        assert widgets["water_hammer_result_diagram_type_label"].wordWrap() is True
    finally:
        dialog.close()


def test_route_water_hammer_bulk_inputs_apply_all_segments_without_same_material_button():
    """整线水锤段只保留全部水锤段批量填写入口。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0, material_key="钢管"),
        _make_route_group("tunnel-a", 1, "隧洞-圆形", start_mc=10.0, end_mc=20.0),
        _make_route_group("pipe-b", 2, "有压管道", start_mc=20.0, end_mc=30.0, material_key="HDPE管"),
        _make_route_group("tunnel-b", 3, "隧洞-圆形", start_mc=30.0, end_mc=40.0),
        _make_route_group("pipe-c", 4, "有压管道", start_mc=40.0, end_mc=50.0, material_key="钢管"),
    ]

    dialog = _make_route_dialog(groups)
    try:
        route_refs = dialog._route_widgets["flow1-route1"]
        segment_widgets = route_refs["water_hammer_segment_widgets"]
        assert len(segment_widgets) == 3
        assert "water_hammer_bulk_same_material_btn" not in route_refs
        assert route_refs["water_hammer_bulk_all_btn"].text() == "填到全部水锤段"

        route_refs["water_hammer_bulk_e_edit"].setText("0.011")
        route_refs["water_hammer_bulk_ts_edit"].setText("0.02")
        QTest.mouseClick(route_refs["water_hammer_bulk_all_btn"], Qt.LeftButton)
        _flush_events(4)

        for widgets in segment_widgets:
            assert _read_float(widgets["water_hammer_wall_thickness_edit"]) == pytest.approx(0.011)
            assert _read_float(widgets["water_hammer_closing_time_edit"]) == pytest.approx(0.02)
    finally:
        dialog.close()


def test_route_water_hammer_calculation_does_not_mutate_table3_loss_chain_fields():
    """整线水锤验算只做独立校核，不改表3损失和水位字段。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0),
        _make_route_group("pipe-b", 1, "有压管道", start_mc=10.0, end_mc=20.0),
    ]
    for idx, group in enumerate(groups):
        row = group.rows[0]
        row.head_loss_total = 1.0 + idx
        row.head_loss_cumulative = 2.0 + idx
        row.water_level = 100.0 + idx
        row.pressure_pipe_window_override = {"total_head_loss": 9.9}
    before = [
        (
            group.rows[0].head_loss_total,
            group.rows[0].head_loss_cumulative,
            group.rows[0].water_level,
            dict(group.rows[0].pressure_pipe_window_override),
        )
        for group in groups
    ]

    dialog = _make_route_dialog(groups)
    try:
        widgets = dialog._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        widgets["water_hammer_wall_thickness_edit"].setText("0.012")
        widgets["water_hammer_closing_time_edit"].setText("0.01")
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        after = [
            (
                group.rows[0].head_loss_total,
                group.rows[0].head_loss_cumulative,
                group.rows[0].water_level,
                dict(group.rows[0].pressure_pipe_window_override),
            )
            for group in groups
        ]
        assert after == before
    finally:
        dialog.close()


def test_route_water_hammer_distribution_check_shows_pass_and_detail_rows():
    """整线水锤验算应按管顶余量判定通过并展示全线采样明细。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0, diameter=1.0),
        _make_route_group("pipe-b", 1, "有压管道", start_mc=10.0, end_mc=20.0, diameter=0.5, pipe_velocity=2.0),
    ]

    dialog = _make_route_dialog(groups)
    try:
        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=400.0)
        widgets = dialog._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        widgets["water_hammer_wall_thickness_edit"].setText("0.02")
        widgets["water_hammer_closing_time_edit"].setText("0.01")
        widgets["water_hammer_allowable_pressure_edit"].setText("5")
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        assert "通过" in widgets["water_hammer_status_label"].text()
        assert widgets["water_hammer_result_conclusion_label"].text() == "通过"
        assert widgets["water_hammer_result_exceed_count_label"].text() == "0"
        assert _read_float(widgets["water_hammer_result_min_margin_label"]) > 0
        assert widgets["water_hammer_result_diagram_type_label"].text() != "-"
        assert "图1-3-3" in widgets["water_hammer_result_diagram_type_label"].text()

        assert widgets["water_hammer_detail_table"].isVisible() is False
        QTest.mouseClick(widgets["water_hammer_detail_btn"], Qt.LeftButton)
        _flush_events(4)
        assert widgets["water_hammer_detail_table"].isVisible() is True
        assert widgets["water_hammer_detail_table"].rowCount() == 6
        headers = [
            widgets["water_hammer_detail_table"].horizontalHeaderItem(i).text()
            for i in range(widgets["water_hammer_detail_table"].columnCount())
        ]
        assert headers[:7] == ["桩号(m)", "所属成员", "D(m)", "v0(m/s)", "a(m/s)", "cp", "a0"]
        assert headers[11] == "Hmax(m)"
        assert headers[12] == "Hmin(m)"
        assert headers[13] == "GB/T ΔH+(m)"
        assert headers[14] == "线性启闭 ΔH+(m)"
        assert headers[15] == "控制ΔH+(m)"
        assert headers[18] == "承压余量(m)"
        assert headers[20] == "管顶最低压力水头(m)"
        assert headers[21] == "控制来源"
        assert headers[25] == "图1-3-3对照"
        assert widgets["water_hammer_detail_table"].item(0, 1).text().startswith("pipe-a")
        assert _read_float(widgets["water_hammer_detail_table"].item(0, 2)) == pytest.approx(1.0)
        assert _read_float(widgets["water_hammer_detail_table"].item(0, 3)) == pytest.approx(1.0)
        assert _read_float(widgets["water_hammer_detail_table"].item(0, 4)) > 0
        assert _read_float(widgets["water_hammer_detail_table"].item(0, 5)) == pytest.approx(1.0)
        assert _read_float(widgets["water_hammer_detail_table"].item(0, 11)) > 0
        assert _read_float(widgets["water_hammer_detail_table"].item(0, 13)) > 0
        assert widgets["water_hammer_detail_table"].item(0, 21).text()
    finally:
        dialog.close()


def test_route_water_hammer_distribution_skips_zero_length_anchor_members():
    """整线水锤验算应跳过起终点相同的锚点行，继续计算真实管段。"""
    groups = [
        _make_route_group("anchor", 0, "有压管道", start_mc=0.0, end_mc=0.0, diameter=1.0),
        _make_route_group("pipe-a", 1, "有压管道", start_mc=0.0, end_mc=10.0, diameter=1.0),
    ]

    dialog = _make_route_dialog(groups)
    try:
        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=400.0)
        widgets = dialog._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        widgets["water_hammer_wall_thickness_edit"].setText("0.02")
        widgets["water_hammer_closing_time_edit"].setText("0.01")

        members = dialog._build_route_water_hammer_distribution_members(widgets["segment_key"])
        assert [member["key"] for member in members] == ["pipe-a"]

        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        assert "成员长度必须大于0" not in widgets["water_hammer_status_label"].text()
        assert "通过" in widgets["water_hammer_status_label"].text()
        assert widgets["water_hammer_detail_table"].rowCount() > 0
    finally:
        dialog.close()


def test_route_water_hammer_distribution_reports_missing_table3_water_level():
    """表3水位未生成时应提示先执行计算，不得把0水位当有效压力线。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0, diameter=1.0),
    ]

    dialog = _make_route_dialog(groups)
    try:
        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=0.0)
        widgets = dialog._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        widgets["water_hammer_wall_thickness_edit"].setText("0.02")
        widgets["water_hammer_closing_time_edit"].setText("0.01")
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        status_text = widgets["water_hammer_status_label"].text()
        assert "数据缺失" in status_text
        assert "缺少表3水位" in status_text
        assert "采样点超出数据覆盖范围" not in status_text
        assert widgets["water_hammer_detail_table"].rowCount() == 0
    finally:
        dialog.close()


def test_route_water_hammer_distribution_check_marks_failed_when_margin_negative():
    """整线任一采样点余量为负时应显示不通过。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0, diameter=1.0),
    ]

    dialog = _make_route_dialog(groups)
    try:
        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=120.0)
        widgets = dialog._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        widgets["water_hammer_wall_thickness_edit"].setText("0.02")
        widgets["water_hammer_closing_time_edit"].setText("0.01")
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        assert "不通过" in widgets["water_hammer_status_label"].text()
        assert widgets["water_hammer_result_conclusion_label"].text() == "不通过"
        assert int(widgets["water_hammer_result_exceed_count_label"].text()) > 0
        assert _read_float(widgets["water_hammer_result_min_margin_label"]) < 0
    finally:
        dialog.close()


def test_route_water_hammer_distribution_uses_members_without_single_value_inputs():
    """整线分布正式判定应使用成员自身参数，不提供单值代表输入。"""
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0, diameter=1.0, pipe_velocity=1.0),
    ]

    dialog = _make_route_dialog(groups)
    try:
        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=130.0)
        widgets = dialog._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        assert "water_hammer_diameter_edit" not in widgets
        assert "water_hammer_velocity_edit" not in widgets
        assert "water_hammer_elastic_modulus_edit" not in widgets
        widgets["water_hammer_wall_thickness_edit"].setText("0.02")
        widgets["water_hammer_closing_time_edit"].setText("0.01")
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)

        assert "不通过" in widgets["water_hammer_status_label"].text()
        assert widgets["water_hammer_result_sample_count_label"].text() != "0"
    finally:
        dialog.close()


def test_route_water_hammer_segments_persist_and_restore_from_manager():
    """route 级水锤结果应随项目保存并在重开弹窗后恢复。"""
    case_dir = Path(tempfile.mkdtemp(prefix="wh_route_"))
    project_path = case_dir / "demo.qxproj"
    groups = [
        _make_route_group("pipe-a", 0, "有压管道", start_mc=0.0, end_mc=10.0),
        _make_route_group("pipe-b", 1, "有压管道", start_mc=10.0, end_mc=20.0),
    ]

    try:
        manager = PressurePipeManager(str(project_path))
        dialog = _make_route_dialog(groups, manager=manager)
        _set_route_profile_and_water_levels(dialog, groups, centerline_elevation=100.0, water_level=400.0)
        route_refs = dialog._route_widgets["flow1-route1"]
        widgets = route_refs["water_hammer_segment_widgets"][0]
        widgets["water_hammer_wall_thickness_edit"].setText("0.02")
        widgets["water_hammer_closing_time_edit"].setText("0.01")
        widgets["water_hammer_allowable_pressure_edit"].setText("5")
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(6)
        dialog._persist_route_water_hammer_segments()
        _flush_events(2)

        reloaded_manager = PressurePipeManager(str(project_path))
        snapshot = reloaded_manager.get_route_config("flow1-route1")
        assert snapshot["water_hammer_segments"][0]["segment_key"] == "flow1-route1::pipe-a::pipe-b"
        assert snapshot["water_hammer_segments"][0]["inputs"]["wall_thickness_m"] == pytest.approx(0.02)
        assert snapshot["water_hammer_segments"][0]["inputs"]["closing_time_s"] == pytest.approx(0.01)
        assert snapshot["water_hammer_segments"][0]["inputs"]["allowable_pressure_mpa"] == pytest.approx(5.0)
        assert snapshot["water_hammer_segments"][0]["result"]["status"] == "通过"
        assert snapshot["water_hammer_segments"][0]["result"]["details"] == []
        assert snapshot["water_hammer_segments"][0]["result"]["sample_count"] > 0

        dialog_reopen = _make_route_dialog(groups, manager=reloaded_manager)
        widgets_reopen = dialog_reopen._route_widgets["flow1-route1"]["water_hammer_segment_widgets"][0]
        assert _read_float(widgets_reopen["water_hammer_wall_thickness_edit"]) == pytest.approx(0.02)
        assert _read_float(widgets_reopen["water_hammer_closing_time_edit"]) == pytest.approx(0.01)
        assert _read_float(widgets_reopen["water_hammer_allowable_pressure_edit"]) == pytest.approx(5.0)
        assert "通过" in widgets_reopen["water_hammer_status_label"].text()
        dialog_reopen.close()
    finally:
        try:
            dialog.close()
        except Exception:
            pass
        shutil.rmtree(case_dir, ignore_errors=True)


def test_real_nanfengsi_tail_route_keeps_three_members_and_rejects_plan_coordinate_profile_if_present():
    """南峰寺末端有压整线应保留三段参数，并避免缓存平面坐标高程。"""
    excel_path = ROOT / "data" / "有压管道改名-南峰寺支渠批量计算.xlsx"
    dxf_path = ROOT / "data" / "南峰寺文家梁有压管道纵断面dxf.dxf"
    if not excel_path.exists() or not dxf_path.exists():
        pytest.skip("缺少南峰寺真实样例文件")

    siphon_root = ROOT / "倒虹吸水力计算系统"
    if str(siphon_root) not in sys.path:
        sys.path.insert(0, str(siphon_root))
    try:
        from app_渠系计算前端.batch.panel import BatchPanel
        from app_渠系计算前端.water_profile.panel import WaterProfilePanel
        from dxf_parser import DxfParser
        from shared.shared_data_manager import get_shared_data_manager
    except Exception as exc:
        pytest.skip(f"真实样例依赖不可用：{exc}")

    _get_qapp()
    shared_data = get_shared_data_manager()
    shared_data.clear_batch_results()
    batch = BatchPanel()
    water_panel = WaterProfilePanel()
    dialog = None
    try:
        batch._do_load_from_filepath(str(excel_path), is_sample=True, sample_title="diag", sample_desc="diag")
        _flush_events(4)
        batch._batch_calculate()
        _flush_events(6)
        assert len(batch.batch_results) == 222

        water_panel._import_from_batch()
        _flush_events(8)
        settings = water_panel._build_settings()
        context = water_panel._prepare_pressure_pipe_dialog_context(
            water_panel.nodes,
            settings=settings,
            show_xxpipe_warning=False,
        )
        route_targets = context.get("route_import_targets") or {}
        assert len(route_targets) == 1
        route_key = next(iter(route_targets.keys()))

        dialog = PressurePipeWaterHammerDialog(
            pipe_groups=context.get("pipe_groups") or [],
            pressure_chains=context.get("chain_descriptors") or [],
            xxpipe_route_mode=bool(context.get("xxpipe_route_mode")),
            route_import_targets=route_targets,
        )
        _flush_events(6)

        segments = dialog._build_route_water_hammer_segments(dialog._route_contexts[route_key])
        assert len(segments) == 1
        candidates = segments[0]["representative_candidates"]
        assert len(candidates) == 3
        assert [candidate["design_flow"] for candidate in candidates] == pytest.approx([2.04, 1.46, 0.65])
        assert [candidate["diameter_m"] for candidate in candidates] == pytest.approx([1.6, 1.2, 0.8])
        assert [candidate["material_key"] for candidate in candidates] == [
            "PCCP管0.014",
            "球墨铸铁管",
            "球墨铸铁管",
        ]
        assert candidates[0]["resolved_material_key"] == "预应力钢筒混凝土管_n014"
        assert [candidate["elastic_modulus_pa"] for candidate in candidates] == pytest.approx(
            [20.6e9, 160.0e9, 160.0e9]
        )
        route_refs = dialog._route_widgets[route_key]
        widgets = route_refs["water_hammer_segment_widgets"][0]
        assert "water_hammer_representative_combo" not in widgets
        assert widgets["water_hammer_member_table"].rowCount() == 3

        route_import_result = None
        try:
            route_import_result = dialog._resolve_xxpipe_route_import_result(
                route_key,
                str(dxf_path),
                DxfParser,
                route_targets[route_key].get("targets") or [],
            )
        except ValueError as exc:
            message = str(exc)
            assert "文家梁连续整线" in message
            assert "疑似平面坐标 Y 值" in message
            assert "不是真实高程" in message
            assert route_key not in dialog._longitudinal_data
        else:
            elevations = [
                float(node.get("elevation", 0.0))
                for node in list(route_import_result.get("merged_nodes") or [])
                if isinstance(node, dict)
            ]
            assert elevations
            assert max(abs(value) for value in elevations) < 10000.0
    finally:
        if dialog is not None:
            dialog.close()
        water_panel.close()
        batch.close()
        shared_data.clear_batch_results()


def test_real_jiangjiaba_route_water_hammer_handles_profile_endpoint_tail_gap():
    """真实江家坝样例的DXF端点尾差不应导致水击覆盖不足误报。"""
    excel_path = ROOT / "data" / "江家坝支管批量计算用表.xlsx"
    dxf_path = ROOT / "data" / "江家坝支管纵剖面中心线.dxf"
    if not excel_path.exists() or not dxf_path.exists():
        pytest.skip("缺少江家坝真实样例文件")

    siphon_root = ROOT / "倒虹吸水力计算系统"
    if str(siphon_root) not in sys.path:
        sys.path.insert(0, str(siphon_root))
    try:
        from app_渠系计算前端.batch.panel import BatchPanel
        from app_渠系计算前端.water_profile.panel import WaterProfilePanel
        from dxf_parser import DxfParser
        from shared.shared_data_manager import get_shared_data_manager
    except Exception as exc:
        pytest.skip(f"真实样例依赖不可用：{exc}")

    _get_qapp()
    shared_data = get_shared_data_manager()
    shared_data.clear_batch_results()
    batch = BatchPanel()
    water_panel = WaterProfilePanel()
    dialog = None
    try:
        batch._do_load_from_filepath(str(excel_path), is_sample=True, sample_title="diag", sample_desc="diag")
        _flush_events(4)
        batch._batch_calculate()
        _flush_events(6)
        assert len(batch.batch_results) == 130

        water_panel._import_from_batch()
        _flush_events(8)
        settings = water_panel._build_settings()
        context = water_panel._prepare_pressure_pipe_dialog_context(
            water_panel.nodes,
            settings=settings,
            show_xxpipe_warning=False,
        )
        route_targets = context.get("route_import_targets") or {}
        route_key = next(iter(route_targets.keys()))
        dialog = PressurePipeWaterHammerDialog(
            pipe_groups=context.get("pipe_groups") or [],
            pressure_chains=context.get("chain_descriptors") or [],
            xxpipe_route_mode=bool(context.get("xxpipe_route_mode")),
            route_import_targets=route_targets,
        )
        _flush_events(6)
        route_refs = dialog._route_widgets[route_key]
        widgets = route_refs["water_hammer_segment_widgets"][0]
        route_import_result = dialog._resolve_xxpipe_route_import_result(
            route_key,
            str(dxf_path),
            DxfParser,
            route_targets[route_key].get("targets") or [],
        )
        dialog._longitudinal_data[route_key] = list(route_import_result.get("merged_nodes") or [])
        longitudinal_nodes_dict = {route_key: list(route_import_result.get("merged_nodes") or [])}
        route_profile_segments_by_key = WaterProfilePanel._build_pressure_pipe_route_profile_segments(
            context.get("pipe_groups") or [],
            longitudinal_nodes_dict,
        )
        valid_angle_rows = 0
        bend_rows = 0
        fold_rows = 0
        for group in context.get("pipe_groups") or []:
            if not WaterProfilePanel._is_pressure_pipe_row_segment_group(group):
                continue
            target_idx = WaterProfilePanel._coerce_pressure_pipe_row_index(
                getattr(group, "target_row_index", -1)
            )
            if target_idx < 0 or target_idx >= len(water_panel.nodes):
                continue
            target_node = water_panel.nodes[target_idx]
            turn_angle = float(getattr(target_node, "turn_angle", 0.0) or 0.0)
            if not (0.1 <= turn_angle < 180.0):
                continue

            valid_angle_rows += 1
            _, pipe_long_nodes, fallback_reason = water_panel._resolve_pressure_pipe_group_longitudinal_nodes(
                group,
                longitudinal_nodes_dict,
                route_profile_segments_by_key=route_profile_segments_by_key,
            )
            record = water_panel._calculate_unnamed_pressure_pipe_group_result(
                group,
                water_panel.nodes,
                pipe_long_nodes,
                spatial_fallback_reason=fallback_reason,
            )
            assert record["status"] == "success"
            if float(record.get("total_bend_loss", 0.0) or 0.0) > 0:
                bend_rows += 1
            if (record.get("bend_details", {}) or {}).get("method") == "pressure_pipe_fold":
                fold_rows += 1

        assert valid_angle_rows == 121
        assert bend_rows == valid_angle_rows
        assert fold_rows == valid_angle_rows
        widgets["water_hammer_wall_thickness_edit"].setText("0.1")
        widgets["water_hammer_closing_time_edit"].setText("60")

        for group in context.get("pipe_groups") or []:
            for row in getattr(group, "rows", []) or []:
                row.water_level = 0.0
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(8)
        missing_status = widgets["water_hammer_status_label"].text()
        assert "缺少表3水位" in missing_status
        assert "采样点超出数据覆盖范围" not in missing_status

        for group in context.get("pipe_groups") or []:
            for row in getattr(group, "rows", []) or []:
                row.water_level = 1000.0
        QTest.mouseClick(widgets["water_hammer_calc_btn"], Qt.LeftButton)
        _flush_events(8)
        valid_status = widgets["water_hammer_status_label"].text()
        assert "采样点超出数据覆盖范围" not in valid_status
        assert "覆盖不足" not in valid_status
        assert "缺少表3水位" not in valid_status
        assert widgets["water_hammer_detail_table"].rowCount() > 0
    finally:
        if dialog is not None:
            dialog.close()
        water_panel.close()
        batch.close()
        shared_data.clear_batch_results()
