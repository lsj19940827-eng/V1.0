# -*- coding: utf-8 -*-
"""渐变段 .qxproj 往返与重开后关键链路回归测试。"""

import importlib
import importlib.util
import json
import os
import sys
import tempfile
from pathlib import Path

import openpyxl

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
os.environ.setdefault("QTWEBENGINE_DISABLE_SANDBOX", "1")

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from PySide6.QtWidgets import QApplication, QTableWidgetItem

from 推求水面线.models.data_models import ChannelNode, ProjectSettings, TransitionLengthRule
from 推求水面线.models.enums import InOutType, StructureType


def _get_qapp():
    return QApplication.instance() or QApplication([])


def _flush_events(rounds: int = 4):
    app = _get_qapp()
    for _ in range(max(1, rounds)):
        app.processEvents()


def _load_panel_module():
    panel_path = (ROOT / "app_渠系计算前端" / "water_profile" / "panel.py").resolve()
    spec = importlib.util.spec_from_file_location("wp_transition_qxproj_roundtrip", panel_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _build_panel(module):
    _get_qapp()
    for attr in ("_open_transition_length_rules", "_show_node_table_context_menu"):
        if not hasattr(module.WaterProfilePanel, attr):
            setattr(module.WaterProfilePanel, attr, lambda self, *args, **kwargs: None)
    panel = module.WaterProfilePanel()
    panel.resize(1400, 900)
    return panel


def _set_cell(table, row, col, text):
    item = table.item(row, col)
    if item is None:
        item = QTableWidgetItem("")
        table.setItem(row, col, item)
    item.setText(str(text))


def _make_nodes():
    upstream = ChannelNode()
    upstream.flow_section = "1"
    upstream.name = "上游明渠"
    upstream.structure_type = StructureType.from_string("明渠-梯形")
    upstream.section_params = {"B": 2.1, "m": 1.4, "h": 1.3, "A": 5.096, "X": 5.986, "R": 0.851}
    upstream.water_depth = 1.3
    upstream.velocity = 1.35
    upstream.roughness = 0.014
    upstream.flow = 4.8
    upstream.water_level = 420.0
    upstream.bottom_elevation = 418.7
    upstream.top_elevation = 420.8
    upstream.head_loss_cumulative = 0.0
    upstream.head_loss_total = 0.0

    transition = ChannelNode()
    transition.flow_section = "1"
    transition.name = "-"
    transition.structure_type = StructureType.TRANSITION
    transition.is_transition = True
    transition.transition_type = "进口"
    transition.transition_form = "曲线形反弯扭曲面"
    transition.transition_length = 5.0
    transition.transition_length_override_m = 5.0
    transition.transition_length_source = "override"
    transition.roughness = 0.014
    transition.flow = 4.8
    transition.head_loss_transition = 0.1
    transition.head_loss_cumulative = 0.1

    tunnel = ChannelNode()
    tunnel.flow_section = "1"
    tunnel.name = "隧洞1"
    tunnel.structure_type = StructureType.from_string("隧洞-圆形")
    tunnel.in_out = InOutType.INLET
    tunnel.section_params = {"D": 2.4, "A": 4.52, "X": 5.33, "R": 0.848}
    tunnel.water_depth = 1.8
    tunnel.velocity = 2.1
    tunnel.roughness = 0.014
    tunnel.flow = 4.8
    tunnel.structure_height = 2.4
    tunnel.head_loss_siphon = 0.3
    tunnel.head_loss_total = 0.3
    tunnel.head_loss_cumulative = 0.4
    tunnel.water_level = 419.6
    tunnel.bottom_elevation = 417.8
    tunnel.top_elevation = 420.2

    downstream = ChannelNode()
    downstream.flow_section = "1"
    downstream.name = "下游明渠"
    downstream.structure_type = StructureType.from_string("明渠-矩形")
    downstream.section_params = {"B": 2.6, "h": 1.5, "A": 3.9, "X": 5.6, "R": 0.696}
    downstream.water_depth = 1.5
    downstream.velocity = 1.4
    downstream.roughness = 0.014
    downstream.flow = 4.8
    downstream.structure_height = 2.0
    downstream.head_loss_siphon = 0.2
    downstream.head_loss_total = 0.2
    downstream.head_loss_cumulative = 0.6
    downstream.water_level = 419.4
    downstream.bottom_elevation = 417.9
    downstream.top_elevation = 419.9

    return [upstream, transition, tunnel, downstream]


def _make_settings():
    settings = ProjectSettings(
        channel_name="合作",
        channel_level="干渠",
        start_water_level=420.0,
        start_station=10097.309,
    )
    settings.transition_length_rules = [
        TransitionLengthRule(
            upstream_structure_type="明渠-梯形",
            downstream_structure_type="隧洞-圆形",
            transition_type="进口",
            rule_mode="step_up",
            step_size_m=1.0,
        )
    ]
    return settings


def _prepare_table_snapshot(panel):
    panel.channel_name_edit.setText("合作")
    idx = panel.channel_level_combo.findText("干渠")
    if idx >= 0:
        panel.channel_level_combo.setCurrentIndex(idx)
    panel.start_wl_edit.setText("420.000")
    panel.start_station_edit.setText("10+097.309")
    panel.design_flow_edit.setText("4.8")
    panel.max_flow_edit.setText("5.0")

    panel._node_structure_heights[0] = 2.1
    panel._node_structure_heights[2] = 2.4
    panel._node_structure_heights[3] = 2.0

    table = panel.node_table
    _set_cell(table, 0, 27, "1.300")
    _set_cell(table, 0, 41, "420.000")
    _set_cell(table, 0, 42, "418.700")
    _set_cell(table, 0, 43, "420.800")

    _set_cell(table, 1, 32, "5.000")
    _set_cell(table, 1, 33, "0.1000")
    _set_cell(table, 1, 40, "0.1000")

    _set_cell(table, 2, 27, "1.800")
    for col in (34, 35, 36, 37):
        _set_cell(table, 2, col, "0.0000")
    _set_cell(table, 2, 38, "0.3000")
    _set_cell(table, 2, 39, "0.3000")
    _set_cell(table, 2, 40, "0.4000")
    _set_cell(table, 2, 41, "419.600")
    _set_cell(table, 2, 42, "417.800")
    _set_cell(table, 2, 43, "420.200")

    _set_cell(table, 3, 27, "1.500")
    for col in (34, 35, 36, 37):
        _set_cell(table, 3, col, "0.0000")
    _set_cell(table, 3, 38, "0.2000")
    _set_cell(table, 3, 39, "0.2000")
    _set_cell(table, 3, 40, "0.6000")
    _set_cell(table, 3, 41, "419.400")
    _set_cell(table, 3, 42, "417.900")
    _set_cell(table, 3, 43, "419.900")


def test_qxproj_roundtrip_supports_immediate_export_details_and_table_edit(monkeypatch):
    module = _load_panel_module()
    with tempfile.TemporaryDirectory(dir=str(ROOT / "tests")) as temp_dir:
        tmp_path = Path(temp_dir)
        panel = _build_panel(module)
        try:
            panel._settings = _make_settings()
            nodes = _make_nodes()
            panel.calculated_nodes = nodes
            panel._update_table_from_nodes_full(nodes)
            _prepare_table_snapshot(panel)
            _flush_events()

            state = panel.to_project_dict()
            state["calculated_nodes"][1]["transition_length_calc_details"] = {}
            state["calculated_nodes"][1]["transition_calc_details"] = {}

            qxproj_path = tmp_path / "transition_roundtrip.qxproj"
            qxproj_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
            loaded_state = json.loads(qxproj_path.read_text(encoding="utf-8"))
        finally:
            panel.deleteLater()

        restored = _build_panel(module)
        try:
            restored.from_project_dict(loaded_state, skip_dirty_signal=True)
            _flush_events()
            assert restored.node_table.item(1, 32).text() == "5.000"
            assert restored.calculated_nodes[1].transition_length_override_m == 5.0
            assert len(restored._settings.transition_length_rules) == 1
            assert restored._settings.transition_length_rules[0].rule_mode == "step_up"

            dialog_calls = {"length": [], "loss": []}
            formula_module = importlib.import_module("app_渠系计算前端.water_profile.formula_dialog")
            monkeypatch.setattr(
                formula_module,
                "show_transition_length_dialog",
                lambda parent, node_name, details: dialog_calls["length"].append((node_name, details)),
            )
            monkeypatch.setattr(
                formula_module,
                "show_transition_loss_dialog",
                lambda parent, node_name, details: dialog_calls["loss"].append((node_name, details)),
            )

            restored._show_transition_length_details(1, restored.calculated_nodes[1])
            restored._show_transition_calc_details(1, restored.calculated_nodes[1])

            assert dialog_calls["length"], "重开项目后应可直接双击查看长度详情"
            assert dialog_calls["loss"], "重开项目后应可直接双击查看损失详情"
            length_details = dialog_calls["length"][0][1]
            loss_details = dialog_calls["loss"][0][1]
            assert length_details["actual_length"] == 5.0
            assert length_details["source"] == "override"
            assert "单条覆盖长度小于公式/规范值" in (length_details["warning"] or "")
            assert loss_details["length"] == 5.0
            assert loss_details["length_details"]["source"] == "override"

            export_path = tmp_path / "water_profile_roundtrip.xlsx"
            info_calls = {"success": [], "warning": [], "error": [], "open": []}
            monkeypatch.setattr(
                module.QFileDialog,
                "getSaveFileName",
                lambda *args, **kwargs: (str(export_path), "Excel文件 (*.xlsx)"),
            )
            monkeypatch.setattr(module, "ask_open_file", lambda path, parent=None: info_calls["open"].append(path))
            monkeypatch.setattr(module.InfoBar, "success", lambda *args, **kwargs: info_calls["success"].append((args, kwargs)))
            monkeypatch.setattr(module.InfoBar, "warning", lambda *args, **kwargs: info_calls["warning"].append((args, kwargs)))
            monkeypatch.setattr(module.InfoBar, "error", lambda *args, **kwargs: info_calls["error"].append((args, kwargs)))

            restored._export_excel()

            assert export_path.exists(), "重开项目后应可不重算直接导出"
            assert info_calls["success"], "导出成功后应走成功提示分支"
            assert info_calls["open"] == [str(export_path)]

            wb = openpyxl.load_workbook(export_path)
            ws = wb.active
            assert ws.cell(row=5, column=33).value == "5.000"
            assert ws.cell(row=6, column=39).value == "0.3000"
            assert ws.cell(row=7, column=39).value == "0.2000"

            restored.node_table.item(2, 38).setText("0.5000")
            restored._on_loss_cell_changed(2, 38)

            assert restored.node_table.item(2, 39).text() == "0.5000"
            assert restored.node_table.item(2, 40).text() == "0.6000"
            assert restored.node_table.item(2, 41).text() == "419.400"
            assert restored.node_table.item(3, 40).text() == "0.8000"
            assert restored.node_table.item(3, 41).text() == "419.200"
            assert restored.calculated_nodes[2].head_loss_total == 0.5
            assert restored.calculated_nodes[3].head_loss_cumulative == 0.8
        finally:
            restored.deleteLater()
