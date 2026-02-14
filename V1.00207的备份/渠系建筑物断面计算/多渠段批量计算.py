# -*- coding: utf-8 -*-
"""
渠系建筑物多流量段批量水力计算系统

功能说明:
1. 支持批量输入多个流量段的参数
2. 支持所有断面类型(明渠、渡槽、隧洞、矩形暗涵、圆形无压管道)
3. 可通过表格批量输入,支持新增、删除、复制、Excel粘贴
4. 一键批量计算所有流量段
5. 生成结果汇总表,支持导出Excel
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import List, Dict, Any
import sys
import os
import math
import subprocess

# 获取基础模块
try:
    # 明渠计算 (含梯形、矩形、圆形)
    from 明渠设计 import (
        quick_calculate as mingqu_calculate,
        quick_calculate_circular as circular_calculate
    )
    MINGQU_AVAILABLE = True
    CIRCULAR_AVAILABLE = True
except ImportError:
    MINGQU_AVAILABLE = False
    CIRCULAR_AVAILABLE = False

try:
    # 渡槽计算
    from 渡槽设计 import (
        quick_calculate_u as ducao_u_calculate,
        quick_calculate_rect as ducao_rect_calculate
    )
    DUCAO_AVAILABLE = True
except ImportError:
    DUCAO_AVAILABLE = False

try:
    # 隧洞计算
    from 隧洞设计 import (
        quick_calculate_circular as suidong_circular_calculate,
        quick_calculate_horseshoe as suidong_horseshoe_calculate,
        quick_calculate_horseshoe_std as suidong_horseshoe_std_calculate
    )
    SUIDONG_AVAILABLE = True
except ImportError:
    SUIDONG_AVAILABLE = False

# 矩形暗涵计算（独立模块）
try:
    from 矩形暗涵设计 import (
        quick_calculate_rectangular_culvert as suidong_rect_calculate
    )
    RECT_CULVERT_AVAILABLE = True
except ImportError:
    RECT_CULVERT_AVAILABLE = False

# 共享数据管理器（用于与推求水面线系统交互）
# 注意：必须使用与main_window.py相同的导入路径，以确保使用同一个单例实例
try:
    # 添加推求水面线目录到路径（从渠系建筑物断面计算/上溯到V1.0/再进入推求水面线/）
    _water_profile_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '推求水面线')
    if _water_profile_dir not in sys.path:
        sys.path.insert(0, _water_profile_dir)
    from ui.shared_data_manager import get_shared_data_manager
    SHARED_DATA_AVAILABLE = True
except ImportError:
    SHARED_DATA_AVAILABLE = False

# 断面类型映射表（用户输入名称 → 程序内部名称）
SECTION_TYPE_MAPPING = {
    "明渠-矩形": "明渠-矩形",
    "明渠-梯形": "明渠-梯形",
    "明渠-圆形": "明渠-圆形",
    "隧洞-圆拱直墙型": "隧洞-圆拱直墙型",
    "隧洞-马蹄形Ⅰ型": "隧洞-马蹄形Ⅰ型",
    "隧洞-马蹄形Ⅱ型": "隧洞-马蹄形Ⅱ型",
    "隧洞-圆形": "隧洞-圆形",
    "渡槽-矩形": "渡槽-矩形",
    "渡槽-U形": "渡槽-U形",
    "矩形暗涵": "矩形暗涵",
    "倒虹吸": "倒虹吸",  # 占位行，不参与计算
    "分水闸": "分水闸",  # 占位行，不参与计算（仅过闸水头损失）
    "分水口": "分水口",  # 占位行，不参与计算（仅过闸水头损失）
}

# 已废弃的断面类型（用于错误提示）
DEPRECATED_SECTION_TYPES = {
    "隧洞-标准1型": "隧洞-马蹄形Ⅰ型",
    "隧洞-标准2型": "隧洞-马蹄形Ⅱ型"
}


def format_station_display(value: float) -> str:
    """
    将桩号数值格式化为标准显示格式（如 12+111.222）
    
    参数:
        value: 桩号数值（米），如 12111.222
    
    返回:
        格式化后的桩号字符串，如 "12+111.222"
    """
    if value < 0:
        value = abs(value)
    
    # 分离公里数和米数
    km = int(value // 1000)
    meters = value % 1000
    
    # 格式化为 km+xxx.xxx 格式，米数部分保持3位整数+3位小数
    return f"{km}+{meters:07.3f}"


def parse_station_input(input_str: str) -> float:
    """
    解析桩号输入字符串为数值
    
    支持两种输入格式：
    1. 纯数字：12111.222 -> 12111.222
    2. 带加号格式：12+111.222 -> 12111.222
    
    参数:
        input_str: 输入字符串
    
    返回:
        桩号数值（米）
    """
    input_str = str(input_str).strip()
    if not input_str:
        return 0.0
    
    # 检查是否包含加号
    if '+' in input_str:
        parts = input_str.split('+')
        if len(parts) == 2:
            try:
                km = int(parts[0])
                meters = float(parts[1])
                return km * 1000 + meters
            except ValueError:
                pass
    
    # 尝试直接解析为数字
    try:
        return float(input_str)
    except ValueError:
        return 0.0


class BatchCalculationPanel(ttk.Frame):
    """渠系建筑物多流量段批量水力计算系统面板"""
    
    def __init__(self, parent):
        super().__init__(parent)
        
        self.batch_results = []  # 存储批量计算结果
        self._freeze_enabled = False
        self._input_data = []
        self._input_table_host = None
        self._input_view_container = None
        self.frozen_sheet = None
        self._ensure_tksheet()
        
        self._create_ui()

    def _destroy_input_view(self) -> None:
        for widget in (getattr(self, "frozen_sheet", None), getattr(self, "input_sheet", None), getattr(self, "_input_view_container", None)):
            try:
                if widget is not None:
                    widget.destroy()
            except Exception:
                pass
        self.frozen_sheet = None
        self._input_view_container = None

    def _setup_structure_type_dropdown(self, sheet, col_index: int) -> None:
        section_types = [
            "明渠-梯形",
            "明渠-矩形",
            "明渠-圆形",
            "渡槽-U形",
            "渡槽-矩形",
            "隧洞-圆形",
            "隧洞-圆拱直墙型",
            "隧洞-马蹄形Ⅰ型",
            "隧洞-马蹄形Ⅱ型",
            "矩形暗涵",
            "倒虹吸",
        ]
        try:
            deprecated = list(DEPRECATED_SECTION_TYPES.keys())
        except Exception:
            deprecated = []

        try:
            sheet.dropdown_column(
                col_index,
                values=section_types + deprecated,
                state="readonly",
                validate_input=False,
                redraw=False,
            )
        except Exception:
            pass

    def _create_input_sheet_widget(self, parent, headers: List[str], col_widths: List[int], *, show_x: bool, show_y: bool, col_offset: int) -> Any:
        sheet = self._Sheet(
            parent,
            headers=headers,
            show_header=True,
            show_row_index=False,
            show_x_scrollbar=show_x,
            show_y_scrollbar=show_y,
            to_clipboard_delimiter="\t",
            from_clipboard_delimiters=["\t"],
            expand_sheet_if_paste_too_big=True,
        )
        try:
            sheet.change_theme("light blue")
        except Exception:
            pass

        for idx, width in enumerate(col_widths):
            applied = False
            for call in (
                lambda: sheet.column_width(column=idx, width=width, redraw=False),
                lambda: sheet.column_width(idx, width=width, redraw=False),
            ):
                if applied:
                    break
                try:
                    call()
                    applied = True
                except Exception:
                    pass

        try:
            sheet.enable_bindings(
                (
                    "single_select",
                    "row_select",
                    "column_select",
                    "drag_select",
                    "arrowkeys",
                    "edit_cell",
                    "select_all",
                    "copy",
                    "cut",
                    "paste",
                    "delete",
                    "undo",
                    "right_click_popup_menu",
                    "column_width_resize",
                    "row_height_resize",
                )
            )
        except Exception:
            try:
                sheet.enable_bindings("all")
            except Exception:
                pass

        sheet.bind("<Double-Button-1>", lambda e, off=col_offset: self._on_input_sheet_double_click(e, off))
        sheet.bind("<Control-a>", lambda e: (self._select_all_input(), "break"))
        sheet.bind("<Control-A>", lambda e: (self._select_all_input(), "break"))
        sheet.bind("<Control-Shift-C>", lambda e: (self._copy_selection_to_clipboard(include_header=True), "break"))
        sheet.bind("<Control-Shift-c>", lambda e: (self._copy_selection_to_clipboard(include_header=True), "break"))
        sheet.bind("<Control-d>", lambda e: (self._fill_down_input(), "break"))
        sheet.bind("<Control-D>", lambda e: (self._fill_down_input(), "break"))
        sheet.bind("<Control-Delete>", lambda e: (self._delete_row(), "break"))
        sheet.bind("<Control-Shift-Delete>", lambda e: (self._delete_row(), "break"))

        try:
            sheet.extra_bindings("end_edit_cell", func=lambda ev, off=col_offset: self._on_input_end_edit_cell(ev, off))
        except Exception:
            pass
        try:
            sheet.extra_bindings("end_paste", func=lambda ev, off=col_offset: self._on_input_end_paste(ev, off))
        except Exception:
            pass

        try:
            sheet.popup_menu_add_command("全选", lambda: self._select_all_input(), table_menu=True, header_menu=False, index_menu=False)
            sheet.popup_menu_add_command("复制(含表头)", lambda: self._copy_selection_to_clipboard(include_header=True), table_menu=True, header_menu=False, index_menu=False)
            sheet.popup_menu_add_command("填充向下(Ctrl+D)", lambda: self._fill_down_input(), table_menu=True, header_menu=False, index_menu=False)
            sheet.popup_menu_add_command("清空内容", lambda: sheet.delete(), table_menu=True, header_menu=False, index_menu=False)
            sheet.popup_menu_add_command("撤销", lambda: sheet.undo(), table_menu=True, header_menu=False, index_menu=False)
            sheet.popup_menu_add_command("重做", lambda: sheet.redo(), table_menu=True, header_menu=False, index_menu=False)
            sheet.popup_menu_add_command("插入行", lambda: self._insert_row(), table_menu=True, header_menu=False, index_menu=False)
            sheet.popup_menu_add_command("删除行", lambda: self._delete_row(), table_menu=True, header_menu=False, index_menu=False)
            sheet.popup_menu_add_command("复制行", lambda: self._copy_row(), table_menu=True, header_menu=False, index_menu=False)
            sheet.popup_menu_add_command("参数设置", lambda: self._open_parameter_dialog(), table_menu=True, header_menu=False, index_menu=False)
        except Exception:
            pass

        return sheet

    def _rebuild_input_view(self) -> None:
        if self._input_table_host is None:
            return

        self._destroy_input_view()

        if not self._freeze_enabled:
            self.input_sheet = self._create_input_sheet_widget(
                self._input_table_host,
                self._input_headers,
                self._input_col_widths,
                show_x=True,
                show_y=True,
                col_offset=0,
            )
            self.input_sheet.pack(fill=tk.BOTH, expand=True)
            self._setup_structure_type_dropdown(self.input_sheet, 3)
            self._set_input_sheet_data(self._input_data, redraw=True, reset_row_positions=True)
            return

        container = ttk.Frame(self._input_table_host)
        container.pack(fill=tk.BOTH, expand=True)
        self._input_view_container = container

        left_headers = self._input_headers[:4]
        right_headers = self._input_headers[4:]
        left_widths = self._input_col_widths[:4]
        right_widths = self._input_col_widths[4:]
        left_width_sum = sum(int(w) for w in left_widths if w)

        left_frame = ttk.Frame(container, width=left_width_sum + 4)
        right_frame = ttk.Frame(container)
        left_frame.grid(row=0, column=0, sticky="ns")
        right_frame.grid(row=0, column=1, sticky="nsew")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(1, weight=1)
        try:
            left_frame.grid_propagate(False)
        except Exception:
            pass

        self.frozen_sheet = self._create_input_sheet_widget(left_frame, left_headers, left_widths, show_x=False, show_y=False, col_offset=0)
        self.input_sheet = self._create_input_sheet_widget(right_frame, right_headers, right_widths, show_x=True, show_y=True, col_offset=4)
        self.frozen_sheet.pack(fill=tk.BOTH, expand=True)
        self.input_sheet.pack(fill=tk.BOTH, expand=True)

        self._setup_structure_type_dropdown(self.frozen_sheet, 3)

        try:
            self.frozen_sheet.sync_scroll(self.input_sheet)
            self.input_sheet.sync_scroll(self.frozen_sheet)
        except Exception:
            pass

        self._set_input_sheet_data(self._input_data, redraw=True, reset_row_positions=True)

    def _toggle_freeze_input_columns(self) -> None:
        try:
            self._sync_input_data_from_views()
        except Exception:
            pass
        self._freeze_enabled = not self._freeze_enabled
        if hasattr(self, "freeze_button") and self.freeze_button is not None:
            try:
                self.freeze_button.configure(text="取消冻结" if self._freeze_enabled else "冻结前4列")
            except Exception:
                pass
        self._rebuild_input_view()

    def _ensure_tksheet(self) -> None:
        try:
            from tksheet import Sheet as _Sheet
            self._Sheet = _Sheet
            return
        except Exception:
            pass

        install = messagebox.askyesno(
            "缺少依赖",
            "未安装 tksheet。\n是否自动安装？\n\n将执行: python -m pip install tksheet",
        )
        if not install:
            raise ImportError("tksheet is required for BatchCalculationPanel")

        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "tksheet"])
            from tksheet import Sheet as _Sheet
            self._Sheet = _Sheet
        except Exception as e:
            messagebox.showerror("安装失败", f"自动安装 tksheet 失败：{e}\n\n请手动执行: pip install tksheet")
            raise
    
    def _create_ui(self):
        """创建UI"""
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 上方:输入表格区域
        self._create_input_area(main_frame)
        
        # 下方:结果显示区域
        self._create_result_area(main_frame)
    
    def _create_input_area(self, parent):
        """创建输入表格区域"""
        input_container = ttk.Frame(parent)
        input_container.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 标题和操作按钮
        header_frame = ttk.Frame(input_container)
        header_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(header_frame, text="输入数据表格 - 流量段参数", 
                 font=('', 11, 'bold')).pack(side=tk.LEFT)
        
        # 操作按钮
        btn_frame = ttk.Frame(header_frame)
        btn_frame.pack(side=tk.RIGHT)
        
        ttk.Button(btn_frame, text="新增行", command=self._add_row, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="插入行", command=self._insert_row, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="删除行", command=self._delete_row, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="复制行", command=self._copy_row, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="导入Excel", command=self._import_from_excel, width=10).pack(side=tk.LEFT, padx=2)
        self.freeze_button = ttk.Button(btn_frame, text="冻结前4列", command=self._toggle_freeze_input_columns, width=12)
        self.freeze_button.pack(side=tk.LEFT, padx=2)
        
        # 流量段输入区域
        flow_frame = ttk.LabelFrame(input_container, text="流量段设置 (各流量段的设计流量Q)", padding="5")
        flow_frame.pack(fill=tk.X, pady=(5, 5))
        
        # 第一行：流量值输入
        flow_row1 = ttk.Frame(flow_frame)
        flow_row1.pack(fill=tk.X, pady=(0, 3))
        
        ttk.Label(flow_row1, text="流量值 (m³/s):").pack(side=tk.LEFT, padx=(0, 5))
        self.flow_segments_var = tk.StringVar(value="5.0, 4.0, 3.0")
        self.flow_segments_entry = ttk.Entry(flow_row1, textvariable=self.flow_segments_var, width=40)
        self.flow_segments_entry.pack(side=tk.LEFT, padx=5)
        
        # 格式提示放在"应用到表格"按钮前面，使用蓝色高对比度
        ttk.Label(flow_row1, text="格式:Q1,Q2,Q3...其中Q1代表第一流量段流量、Q2代表第二流量段流量...", 
                 foreground='#0066CC').pack(side=tk.LEFT, padx=5)
        
        ttk.Button(flow_row1, text="应用到表格", command=self._apply_flow_segments, width=10).pack(side=tk.LEFT, padx=5)
        
        # 第二行：渠道名称、类型、起始桩号、起始水位
        flow_row2 = ttk.Frame(flow_frame)
        flow_row2.pack(fill=tk.X, pady=(0, 0))
        
        # 渠道名称输入
        ttk.Label(flow_row2, text="渠道名称:").pack(side=tk.LEFT, padx=(0, 5))
        self.var_channel_name = tk.StringVar(value="")
        self.channel_name_entry = ttk.Entry(flow_row2, textvariable=self.var_channel_name, width=15)
        self.channel_name_entry.pack(side=tk.LEFT, padx=5)
        
        # 渠道类型下拉框
        ttk.Label(flow_row2, text="渠道类型:").pack(side=tk.LEFT, padx=(20, 5))
        self.var_channel_level = tk.StringVar(value="支渠")
        # 渠道类型选项（与推求水面线程序保持一致）
        channel_level_options = [
            "总干渠", "总干管", "分干渠", "分干管", 
            "干渠", "干管", "支渠", "支管", "分支渠", "分支管"
        ]
        self.channel_level_combo = ttk.Combobox(
            flow_row2, 
            textvariable=self.var_channel_level, 
            values=channel_level_options, 
            width=10, 
            state="readonly"
        )
        self.channel_level_combo.pack(side=tk.LEFT, padx=5)
        
        # 起始桩号输入
        ttk.Label(flow_row2, text="起始桩号:").pack(side=tk.LEFT, padx=(20, 5))
        self.var_start_station = tk.StringVar(value="0+000.000")
        self._entry_start_station = ttk.Entry(flow_row2, textvariable=self.var_start_station, width=15)
        self._entry_start_station.pack(side=tk.LEFT, padx=(0, 3))
        # 起始桩号输入提示（蓝色高对比度）
        ttk.Label(flow_row2, text="(如12+111.222,输入12111.222)", 
                  foreground='#0066CC', font=('', 8)).pack(side=tk.LEFT, padx=(0, 5))
        # 绑定焦点事件：失去焦点时自动格式化显示
        self._entry_start_station.bind('<FocusOut>', self._on_start_station_focus_out)
        self._entry_start_station.bind('<FocusIn>', self._on_start_station_focus_in)
        # 绑定Enter键：按Enter键也触发格式化
        self._entry_start_station.bind('<Return>', self._on_start_station_enter)
        self._entry_start_station.bind('<KP_Enter>', self._on_start_station_enter)
        
        # 渠道起始水位输入
        ttk.Label(flow_row2, text="渠道起始水位 (m):").pack(side=tk.LEFT, padx=(20, 5))
        self.var_start_water_level = tk.StringVar(value="")
        self.start_water_level_entry = ttk.Entry(flow_row2, textvariable=self.var_start_water_level, width=12)
        self.start_water_level_entry.pack(side=tk.LEFT, padx=5)
        
        # 输入表格 - 重新设计列结构（已删除桩号列）
        # 列顺序: 序号 | 流量段 | 建筑物名称 | 断面类型 | Q | 糙率n | 比降 | 边坡系数m | 底宽B | 宽深比 | 半径R | 直径D | 倒角角度 | 倒角底边 | 圆心角 | 不淤流速 | 不冲流速
        
        # 添加表格操作提示
        hint_frame = ttk.Frame(input_container)
        hint_frame.pack(fill=tk.X, pady=(0, 2))
        ttk.Label(hint_frame, text="提示: 双击单元格编辑 | Ctrl+V 粘贴到选中位置 | Ctrl+C 复制选中行", 
                 foreground='black', font=('', 8)).pack(side=tk.LEFT)
        
        table_frame = ttk.Frame(input_container)
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        # 表格列定义（已删除桩号列，新增X、Y列）
        columns = ("序号", "流量段", "建筑物名称", "结构形式", "X", "Y", "Q(m³/s)", "糙率n", "比降(1/)",
                   "边坡系数m", "底宽B(m)", "明渠宽深比", "半径R(m)", "直径D(m)", 
                   "矩形渡槽深宽比", "倒角角度(°)", "倒角底边(m)", "圆心角(°)", "不淤流速", "不冲流速")

        self._input_headers = list(columns)
        self._input_col_widths = [50, 60, 120, 110, 180, 180, 70, 60, 70, 80, 70, 80, 70, 70, 90, 80, 80, 70, 70, 70]
        self._input_table_host = table_frame
        if not self._input_data:
            self._input_data = []
        self._rebuild_input_view()
        
        # 添加几行示例数据
        self._add_sample_data()
        
        # 初始化后自动调整输入表格列宽
        self.after(100, self._auto_fit_input_columns)
        
        # 计算按钮
        calc_frame = ttk.Frame(input_container)
        calc_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.calc_button = ttk.Button(calc_frame, text="开始批量计算", command=self._batch_calculate, 
                  style='Accent.TButton')
        self.calc_button.pack(side=tk.LEFT, padx=5)
        
        # 详细输出选项
        self.show_detail_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(calc_frame, text="启用详细计算过程输出", 
                       variable=self.show_detail_var).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(calc_frame, text="参数设置", command=self._open_parameter_dialog).pack(side=tk.LEFT, padx=5)
        ttk.Button(calc_frame, text="清空输入", command=self._clear_input).pack(side=tk.LEFT, padx=5)
        
        # 提示标签
        ttk.Label(calc_frame, text="提示: 双击参数列打开参数设置弹窗; 双击断面类型列可选择类型", 
                 foreground='black').pack(side=tk.RIGHT, padx=10)
    
    def _create_result_area(self, parent):
        """创建结果显示区域"""
        result_container = ttk.LabelFrame(parent, text="计算结果汇总与详细过程", padding="5")
        result_container.pack(fill=tk.BOTH, expand=True)
        
        # --- 导出和操作按钮 (移到上方先pack，侧重底部停靠) ---
        op_frame = ttk.Frame(result_container)
        op_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(5, 0))
        
        ttk.Button(op_frame, text="导出Excel报告", command=self._export_to_excel, 
                  width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(op_frame, text="导出详细过程(TXT)", command=self._export_detailed_to_txt,
                  width=18).pack(side=tk.LEFT, padx=5)
        ttk.Button(op_frame, text="清空结果", command=self._clear_results).pack(side=tk.LEFT, padx=5)
        
        # --- 创建选项卡 (后pack并expand=True占据剩余空间) ---
        self.result_notebook = ttk.Notebook(result_container)
        self.result_notebook.pack(fill=tk.BOTH, expand=True)
        
        # 选项卡1: 汇总表格
        summary_frame = ttk.Frame(self.result_notebook, padding="5")
        self.result_notebook.add(summary_frame, text="结果汇总表")
        
        # 选项卡2: 详细过程
        detail_frame = ttk.Frame(self.result_notebook, padding="5")
        self.result_notebook.add(detail_frame, text="详细计算过程")
        
        # --- 汇总表格区域 ---（结构形式列，已删除桩号列）
        result_columns = ("序号", "流量段", "建筑物名称", "结构形式", "底宽B(m)", "直径D(m)", "半径R(m)", "h设计(m)", "V设计(m/s)", 
                         "A设计(m²)", "R水力(m)", "湿周χ(m)", "Q加大(m³/s)", "h加大(m)", 
                         "V加大(m/s)", "超高Fb(m)", "设计净空高度(m)", "加大净空高度(m)", "设计净空比例(%)", "加大净空比例(%)", "状态")

        self._result_headers = list(result_columns)
        self._result_col_widths = [70, 90, 130, 110, 90, 90, 80, 100, 110, 100, 100, 95, 120, 100, 110, 105, 175, 175, 165, 165, 200]
        self.result_sheet = self._Sheet(
            summary_frame,
            headers=self._result_headers,
            show_header=True,
            show_row_index=False,
            show_x_scrollbar=True,
            show_y_scrollbar=True,
            to_clipboard_delimiter="\t",
            from_clipboard_delimiters=["\t"],
        )
        self.result_sheet.grid(row=0, column=0, sticky="nsew")

        try:
            self.result_sheet.change_theme("light blue")
        except Exception:
            pass

        for idx, width in enumerate(self._result_col_widths):
            applied = False
            for call in (
                lambda: self.result_sheet.column_width(column=idx, width=width, redraw=False),
                lambda: self.result_sheet.column_width(idx, width=width, redraw=False),
            ):
                if applied:
                    break
                try:
                    call()
                    applied = True
                except Exception:
                    pass

        try:
            self.result_sheet.enable_bindings(
                (
                    "single_select",
                    "row_select",
                    "column_select",
                    "drag_select",
                    "arrowkeys",
                    "select_all",
                    "copy",
                    "right_click_popup_menu",
                    "column_width_resize",
                    "row_height_resize",
                )
            )
        except Exception:
            try:
                self.result_sheet.enable_bindings("all")
            except Exception:
                pass

        try:
            self.result_sheet.readonly_columns(columns=list(range(len(self._result_headers))), readonly=True, redraw=False)
        except Exception:
            pass

        self.result_sheet.bind("<Control-a>", lambda e: (self._select_all_result(), "break"))
        self.result_sheet.bind("<Control-A>", lambda e: (self._select_all_result(), "break"))
        self.result_sheet.bind("<Control-Shift-C>", lambda e: (self._copy_result_selection_to_clipboard(include_header=True), "break"))
        self.result_sheet.bind("<Control-Shift-c>", lambda e: (self._copy_result_selection_to_clipboard(include_header=True), "break"))

        try:
            self.result_sheet.popup_menu_add_command("全选", lambda: self._select_all_result(), table_menu=True, header_menu=False, index_menu=False)
            self.result_sheet.popup_menu_add_command("复制(含表头)", lambda: self._copy_result_selection_to_clipboard(include_header=True), table_menu=True, header_menu=False, index_menu=False)
        except Exception:
            pass

        summary_frame.grid_rowconfigure(0, weight=1)
        summary_frame.grid_columnconfigure(0, weight=1)
        
        # --- 详细过程区域 ---
        self.detail_text = tk.Text(detail_frame, wrap=tk.NONE, undo=True, font=('Consolas', 10))
        text_v_scroll = ttk.Scrollbar(detail_frame, orient=tk.VERTICAL, command=self.detail_text.yview)
        text_h_scroll = ttk.Scrollbar(detail_frame, orient=tk.HORIZONTAL, command=self.detail_text.xview)
        self.detail_text.configure(yscrollcommand=text_v_scroll.set, xscrollcommand=text_h_scroll.set)
        
        self.detail_text.grid(row=0, column=0, sticky='nsew')
        text_v_scroll.grid(row=0, column=1, sticky='ns')
        text_h_scroll.grid(row=1, column=0, sticky='ew')
        
        detail_frame.grid_rowconfigure(0, weight=1)
        detail_frame.grid_columnconfigure(0, weight=1)
        
        # 初始化后自动调整结果表格列宽
        self.after(100, self._auto_fit_result_columns)

    def _normalize_row(self, row: Any, total_cols: int) -> List[Any]:
        row_list = list(row) if isinstance(row, (list, tuple)) else [row]
        if len(row_list) < total_cols:
            row_list.extend([""] * (total_cols - len(row_list)))
        elif len(row_list) > total_cols:
            row_list = row_list[:total_cols]
        return row_list

    def _get_input_sheet_data(self) -> List[List[Any]]:
        if isinstance(getattr(self, "_input_data", None), list) and self._input_data:
            return [self._normalize_row(r, len(self._input_headers)) for r in self._input_data]

        try:
            data = self.input_sheet.get_sheet_data(return_copy=True)
        except TypeError:
            data = self.input_sheet.get_sheet_data()
        except Exception:
            data = []

        normalized = [self._normalize_row(r, len(self._input_headers)) for r in (data or [])]
        self._input_data = normalized
        return normalized

    def _set_input_sheet_data(self, rows: List[List[Any]], *, redraw: bool = True, reset_row_positions: bool = False) -> None:
        normalized = [self._normalize_row(r, len(self._input_headers)) for r in (rows or [])]
        self._input_data = normalized

        if getattr(self, "_freeze_enabled", False) and getattr(self, "frozen_sheet", None) is not None:
            left_rows = [r[:4] for r in normalized]
            right_rows = [r[4:] for r in normalized]
            try:
                self.frozen_sheet.set_sheet_data(
                    left_rows,
                    reset_col_positions=False,
                    reset_row_positions=reset_row_positions,
                    redraw=redraw,
                )
            except TypeError:
                try:
                    self.frozen_sheet.set_sheet_data(left_rows)
                except Exception:
                    pass
            try:
                self.input_sheet.set_sheet_data(
                    right_rows,
                    reset_col_positions=False,
                    reset_row_positions=reset_row_positions,
                    redraw=redraw,
                )
            except TypeError:
                try:
                    self.input_sheet.set_sheet_data(right_rows)
                except Exception:
                    pass
            if redraw:
                try:
                    self.frozen_sheet.refresh()
                except Exception:
                    pass
                try:
                    self.input_sheet.refresh()
                except Exception:
                    pass
            return

        try:
            self.input_sheet.set_sheet_data(
                normalized,
                reset_col_positions=False,
                reset_row_positions=reset_row_positions,
                redraw=redraw,
            )
        except TypeError:
            try:
                self.input_sheet.set_sheet_data(normalized)
            except Exception:
                pass
            if redraw:
                try:
                    self.input_sheet.refresh()
                except Exception:
                    pass

    def _get_result_sheet_data(self) -> List[List[Any]]:
        try:
            data = self.result_sheet.get_sheet_data(return_copy=True)
        except TypeError:
            data = self.result_sheet.get_sheet_data()
        except Exception:
            data = []
        return [self._normalize_row(r, len(self._result_headers)) for r in (data or [])]

    def _set_result_sheet_data(self, rows: List[List[Any]], *, redraw: bool = True, reset_row_positions: bool = False) -> None:
        normalized = [self._normalize_row(r, len(self._result_headers)) for r in (rows or [])]
        try:
            self.result_sheet.set_sheet_data(
                normalized,
                reset_col_positions=False,
                reset_row_positions=reset_row_positions,
                redraw=redraw,
            )
        except TypeError:
            self.result_sheet.set_sheet_data(normalized)
            if redraw:
                try:
                    self.result_sheet.refresh()
                except Exception:
                    pass

    def _select_all_input(self) -> None:
        if getattr(self, "_freeze_enabled", False) and getattr(self, "frozen_sheet", None) is not None:
            try:
                self.frozen_sheet.select_all(redraw=False)
            except Exception:
                pass
            try:
                self.input_sheet.select_all(redraw=False)
            except Exception:
                pass
            try:
                self.frozen_sheet.redraw()
            except Exception:
                pass
            try:
                self.input_sheet.redraw()
            except Exception:
                pass
            return

        try:
            self.input_sheet.select_all(redraw=True)
        except Exception:
            try:
                self.input_sheet.select_all()
            except Exception:
                pass

    def _select_all_result(self) -> None:
        try:
            self.result_sheet.select_all(redraw=True)
        except Exception:
            try:
                self.result_sheet.select_all()
            except Exception:
                pass

    def _setup_input_sheet_dropdowns(self) -> None:
        section_types = [
            "明渠-梯形",
            "明渠-矩形",
            "明渠-圆形",
            "渡槽-U形",
            "渡槽-矩形",
            "隧洞-圆形",
            "隧洞-圆拱直墙型",
            "隧洞-马蹄形Ⅰ型",
            "隧洞-马蹄形Ⅱ型",
            "矩形暗涵",
            "倒虹吸",
        ]
        try:
            deprecated = list(DEPRECATED_SECTION_TYPES.keys())
        except Exception:
            deprecated = []

        try:
            self.input_sheet.dropdown_column(
                3,
                values=section_types + deprecated,
                state="readonly",
                validate_input=False,
                redraw=False,
            )
        except Exception:
            pass

    def _get_input_active_cell(self) -> Any:
        if getattr(self, "_freeze_enabled", False) and getattr(self, "frozen_sheet", None) is not None:
            for sheet, off in ((self.input_sheet, 4), (self.frozen_sheet, 0)):
                for getter in (lambda s=sheet: s.get_currently_selected(), lambda s=sheet: s.get_selected_cells()):
                    try:
                        sel = getter()
                    except Exception:
                        continue
                    if isinstance(sel, tuple) and len(sel) >= 2 and all(isinstance(x, int) for x in sel[:2]):
                        return sel[0], off + sel[1]
                    if isinstance(sel, (list, set, tuple)):
                        for item in sel:
                            if isinstance(item, tuple) and len(item) >= 2 and all(isinstance(x, int) for x in item[:2]):
                                return item[0], off + item[1]
            return None

        for getter in (lambda: self.input_sheet.get_currently_selected(), lambda: self.input_sheet.get_selected_cells()):
            try:
                sel = getter()
            except Exception:
                continue
            if isinstance(sel, tuple) and len(sel) >= 2 and all(isinstance(x, int) for x in sel[:2]):
                return sel[0], sel[1]
            if isinstance(sel, (list, set, tuple)):
                for item in sel:
                    if isinstance(item, tuple) and len(item) >= 2 and all(isinstance(x, int) for x in item[:2]):
                        return item[0], item[1]
        return None

    def _get_selected_input_rows(self) -> List[int]:
        rows = set()
        if getattr(self, "_freeze_enabled", False) and getattr(self, "frozen_sheet", None) is not None:
            for sheet in (self.frozen_sheet, self.input_sheet):
                for getter in (lambda s=sheet: s.get_selected_rows(), lambda s=sheet: s.get_selected_row_indices()):
                    try:
                        selected = getter()
                    except Exception:
                        continue
                    if isinstance(selected, (list, tuple, set)):
                        rows.update(int(r) for r in selected if isinstance(r, int))
                        break
        else:
            for getter in (lambda: self.input_sheet.get_selected_rows(), lambda: self.input_sheet.get_selected_row_indices()):
                try:
                    selected = getter()
                except Exception:
                    continue
                if isinstance(selected, (list, tuple, set)):
                    rows.update(int(r) for r in selected if isinstance(r, int))
                    break

        if not rows:
            active = self._get_input_active_cell()
            if active:
                rows.add(active[0])
        return sorted(rows)

    def _update_input_row(self, row_idx: int, values: List[Any], *, redraw: bool = False) -> None:
        values = self._normalize_row(values, len(self._input_headers))
        applied = False
        for call in (
            lambda: self.input_sheet.set_row_data(row_idx, values=values, add_columns=True, redraw=redraw),
            lambda: self.input_sheet.set_row_data(row_idx, values, add_columns=True, redraw=redraw),
        ):
            if applied:
                break
            try:
                call()
                applied = True
            except Exception:
                pass
        if not applied:
            data = self._get_input_sheet_data()
            if 0 <= row_idx < len(data):
                data[row_idx] = values
                self._set_input_sheet_data(data, redraw=redraw, reset_row_positions=False)

    def _show_diversion_gate_hint(self, seq_no):
        """显示分水闸流量段分界提示（非模态，2秒后自动消失）"""
        try:
            hint_win = tk.Toplevel(self)
            hint_win.overrideredirect(True)  # 无边框窗口
            hint_win.attributes('-topmost', True)
            
            msg = f"提示：序号{seq_no}已设为分水闸（流量段分界点）\n请确认下一行的流量段编号已正确递增"
            label = tk.Label(hint_win, text=msg, bg='#FFFACD', fg='#333333',
                           font=('Microsoft YaHei', 9), padx=12, pady=8,
                           relief='solid', borderwidth=1)
            label.pack()
            
            # 定位到鼠标附近
            x = self.winfo_pointerx() + 15
            y = self.winfo_pointery() + 15
            hint_win.geometry(f"+{x}+{y}")
            
            # 2.5秒后自动消失
            self.after(2500, lambda: hint_win.destroy() if hint_win.winfo_exists() else None)
        except Exception:
            pass  # 提示失败不影响主流程

    def _apply_section_type_change_to_row(self, row_values: List[Any], new_type: str) -> List[Any]:
        values = self._normalize_row(row_values, len(self._input_headers))
        values[3] = new_type

        if "明渠" in new_type or new_type == "矩形暗涵":
            values[2] = "-"
        elif "分水" in new_type:
            # 分水闸/分水口：保留或设置建筑物名称（不设为"-"）
            if str(values[2]).strip() == "-" or not str(values[2]).strip():
                values[2] = f"分水闸{values[0]}"
        elif str(values[2]).strip() == "-":
            values[2] = f"建筑物{values[0]}"

        # 分水闸/分水口：清空所有计算参数（不需要断面计算）
        if "分水" in new_type:
            for i in range(8, 18):
                values[i] = ""
            return values

        for i in range(9, 18):
            values[i] = ""

        if not values[8] or str(values[8]).strip() == "":
            values[8] = "3000"

        if "明渠-梯形" in new_type:
            values[9] = "1.0"

        return values

    def _on_input_sheet_double_click(self, event, col_offset: int = 0):
        sheet = getattr(event, "widget", None)
        if sheet is None or not hasattr(sheet, "identify_region"):
            sheet = self.input_sheet

        region = None
        try:
            region = sheet.identify_region(event)
        except Exception:
            pass

        if region != "table":
            try:
                return sheet.edit_cell(event)
            except Exception:
                return "break"

        try:
            row = sheet.identify_row(event)
            col = sheet.identify_column(event)
        except Exception:
            return "break"

        if row is None or col is None:
            return "break"

        try:
            sheet.set_currently_selected(row, col)
        except Exception:
            pass

        full_col = col_offset + col

        if full_col == 3:
            try:
                sheet.open_dropdown(row, col)
            except Exception:
                try:
                    sheet.edit_cell(event, dropdown=True)
                except Exception:
                    pass
            return "break"

        if 6 <= full_col <= 19:
            self._open_parameter_dialog_for_row(row)
            return "break"

        try:
            sheet.edit_cell(event)
        except Exception:
            pass
        return "break"

    def _parse_event_rc(self, event) -> Any:
        if isinstance(event, dict):
            r = event.get("row")
            c = event.get("column")
            if isinstance(r, int) and isinstance(c, int):
                return (r, c)
            loc = event.get("loc")
            if isinstance(loc, (tuple, list)) and len(loc) >= 2 and all(isinstance(x, int) for x in loc[:2]):
                return (loc[0], loc[1])
        if isinstance(event, (tuple, list)) and len(event) >= 2:
            if isinstance(event[0], int) and isinstance(event[1], int):
                return (event[0], event[1])
            if isinstance(event[1], (tuple, list)) and len(event[1]) >= 2 and all(isinstance(x, int) for x in event[1][:2]):
                return (event[1][0], event[1][1])
        return None

    def _sync_input_data_from_views(self) -> List[List[Any]]:
        if getattr(self, "_freeze_enabled", False) and getattr(self, "frozen_sheet", None) is not None:
            try:
                left = self.frozen_sheet.get_sheet_data(return_copy=True)
            except TypeError:
                left = self.frozen_sheet.get_sheet_data()
            except Exception:
                left = []
            try:
                right = self.input_sheet.get_sheet_data(return_copy=True)
            except TypeError:
                right = self.input_sheet.get_sheet_data()
            except Exception:
                right = []

            rows = max(len(left or []), len(right or []))
            merged = []
            for i in range(rows):
                lrow = self._normalize_row((left[i] if i < len(left or []) else []), 4)
                rrow = list(right[i] if i < len(right or []) else [])
                merged.append(self._normalize_row(lrow + rrow, len(self._input_headers)))
            self._input_data = merged
            return merged

        try:
            data = self.input_sheet.get_sheet_data(return_copy=True)
        except TypeError:
            data = self.input_sheet.get_sheet_data()
        except Exception:
            data = []
        normalized = [self._normalize_row(r, len(self._input_headers)) for r in (data or [])]
        self._input_data = normalized
        return normalized

    def _on_input_end_edit_cell(self, event, col_offset: int = 0):
        rc = self._parse_event_rc(event)
        if not rc:
            return
        r, c = rc
        if r < 0 or c < 0:
            return

        full_c = col_offset + c
        data = self._sync_input_data_from_views()
        if r >= len(data):
            return
        row = data[r]

        if full_c == 3:
            raw = str(row[3]).strip()
            mapped, error_msg = self._validate_section_type(raw)
            if error_msg:
                messagebox.showwarning("提示", error_msg)
            if not mapped:
                return
            data[r] = self._apply_section_type_change_to_row(row, mapped)
            self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)
            return

        if full_c == 1:
            try:
                seg = int(str(row[1]).strip())
                row[6] = str(self.get_flow_for_segment(seg))
                data[r] = row
                self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)
            except Exception:
                return

    def _on_input_end_paste(self, event, col_offset: int = 0):
        data = self._sync_input_data_from_views()
        if not data:
            return

        changed_rows = set()
        try:
            if isinstance(event, (tuple, list)) and len(event) >= 2:
                maybe_rc = event[1]
                if isinstance(maybe_rc, (tuple, list)) and len(maybe_rc) >= 2 and all(isinstance(x, int) for x in maybe_rc[:2]):
                    start_r = maybe_rc[0]
                    pasted = event[2] if len(event) >= 3 else None
                    if isinstance(pasted, (list, tuple)):
                        for r_off, prow in enumerate(pasted):
                            if not isinstance(prow, (list, tuple)):
                                continue
                            changed_rows.add(start_r + r_off)
        except Exception:
            pass

        if not changed_rows:
            active = self._get_input_active_cell()
            if active:
                changed_rows.add(active[0])

        for r in sorted(changed_rows):
            if r < 0 or r >= len(data):
                continue
            row = data[r]
            raw_type = str(row[3]).strip()
            mapped, error_msg = self._validate_section_type(raw_type)
            if error_msg:
                messagebox.showwarning("提示", error_msg)
            if mapped and mapped != raw_type:
                row = self._apply_section_type_change_to_row(row, mapped)
                data[r] = row
            if not row[8] or str(row[8]).strip() == "":
                row[8] = "3000"
                data[r] = row

        self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)
        self._auto_detect_flow_segments()

    
    def _export_detailed_to_txt(self):
        """将详细计算过程导出到TXT文件"""
        content = self.detail_text.get(1.0, tk.END).strip()
        if not content:
            messagebox.showwarning("提示", "目前没有可导出的详细计算过程")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            title="保存详细计算过程"
        )
        
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                messagebox.showinfo("成功", f"详细过程已成功导出至:\n{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"保存失败: {str(e)}")
    
    def _apply_flow_segments(self):
        """将流量段设置应用到表格"""
        try:
            # 解析流量值
            flow_str = self.flow_segments_var.get().strip()
            if not flow_str:
                messagebox.showwarning("提示", "请输入流量值")
                return
            
            # 支持逗号或空格分隔
            flow_str = flow_str.replace('，', ',')  # 中文逗号转英文
            flow_values = [float(q.strip()) for q in flow_str.split(',') if q.strip()]
            
            if not flow_values:
                messagebox.showwarning("提示", "请输入有效的流量值")
                return
            
            data = self._get_input_sheet_data()
            updated_count = 0
            for r, values in enumerate(data):
                try:
                    segment = int(str(values[1]).strip())  # 流量段编号
                    if 1 <= segment <= len(flow_values):
                        values[6] = str(flow_values[segment - 1])  # 更新Q列(索引6，X、Y之后)
                        updated_count += 1
                except (ValueError, IndexError):
                    continue

            self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)
            
            messagebox.showinfo("完成", f"已将流量值应用到 {updated_count} 行数据\n"
                              f"流量段1~{len(flow_values)}: {flow_values}")
            
        except ValueError as e:
            messagebox.showerror("错误", f"流量值格式错误: {str(e)}\n请使用逗号分隔的数字")
    
    def get_flow_for_segment(self, segment: int) -> float:
        """获取指定流量段的流量值 (供其他模块调用)"""
        try:
            flow_str = self.flow_segments_var.get().strip()
            if not flow_str:
                return 5.0  # 默认值
            
            flow_str = flow_str.replace('，', ',')
            flow_values = [float(q.strip()) for q in flow_str.split(',') if q.strip()]
            
            if 1 <= segment <= len(flow_values):
                return flow_values[segment - 1]
            else:
                return flow_values[-1] if flow_values else 5.0
        except:
            return 5.0
    
    def _auto_detect_flow_segments(self):
        """自动从表格数据中识别流量段及其对应的流量值，并更新流量设置区域"""
        data = self._get_input_sheet_data()
        if not data:
            return
        
        # 收集各流量段的流量值（每个流量段取第一个有效Q值）
        segment_flow_map = {}  # {流量段编号: Q值}
        
        for values in data:
            if len(values) < 7:
                continue
            
            try:
                segment = int(str(values[1]).strip())  # 流量段编号在索引1
                q_value = str(values[6]).strip()  # Q值在索引6
                
                # 如果该流量段还没有记录，且Q值有效，则记录
                if segment not in segment_flow_map and q_value:
                    segment_flow_map[segment] = float(q_value)
            except (ValueError, TypeError, IndexError):
                continue
        
        # 如果没有识别到任何流量段，不更新
        if not segment_flow_map:
            return
        
        # 按流量段编号排序，构建流量值字符串
        max_segment = max(segment_flow_map.keys())
        flow_values = []
        for seg in range(1, max_segment + 1):
            if seg in segment_flow_map:
                flow_values.append(str(segment_flow_map[seg]))
            else:
                # 如果某个流量段缺失，使用前一个流量段的值或默认值
                flow_values.append(flow_values[-1] if flow_values else "5.0")
        
        # 更新流量设置输入框
        flow_str = ", ".join(flow_values)
        self.flow_segments_var.set(flow_str)
    
    def _add_sample_data(self):
        """添加示例数据 - 来自多流量段表格填写示例.xlsx的完整数据"""
        # 新列顺序（已删除桩号，新增X、Y）: 序号, 流量段, 建筑物名称, 结构形式, X, Y, Q, 糙率n, 比降, 边坡系数m, 底宽B, 明渠宽深比, 半径R, 直径D, 矩形渡槽深宽比, 倒角角度, 倒角底边, 圆心角, 不淤流速, 不冲流速
        # 完整示例数据（含X、Y坐标），与多流量段表格填写示例.xlsx模板一致
        samples = [
            ("1", "1", "-", "明渠-矩形", "649606.177086", "3377745.982674", "5", "0.014", "3000", "0", "2", "", "", "", "", "", "", "", "0.1", "100"),
            ("2", "1", "-", "明渠-矩形", "649534.180449", "3377664.854614", "5", "0.014", "3000", "0", "2", "", "", "", "", "", "", "", "0.1", "100"),
            ("3", "1", "-", "明渠-矩形", "649480.482814", "3377634.277101", "5", "0.014", "3000", "0", "2", "", "", "", "", "", "", "", "0.1", "100"),
            ("4", "1", "土地垭", "隧洞-马蹄形Ⅰ型", "649478.323235", "3377610.807806", "5", "0.014", "2000", "", "", "", "1.5", "", "", "", "", "", "0.1", "100"),
            ("5", "1", "土地垭", "隧洞-马蹄形Ⅰ型", "649441.884821", "3377556.331275", "5", "0.014", "2000", "", "", "", "1.5", "", "", "", "", "", "0.1", "100"),
            ("6", "1", "-", "明渠-梯形", "649440.214195", "3377528.976904", "5", "0.014", "3000", "1", "1.8", "", "", "", "", "", "", "", "0.1", "100"),
            ("7", "1", "-", "明渠-梯形", "649419.568825", "3377522.66441", "5", "0.014", "3000", "1", "1.8", "", "", "", "", "", "", "", "0.1", "100"),
            ("8", "1", "磨尔滩", "分水闸", "649402.139216", "3377539.733849", "5", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("9", "2", "-", "明渠-梯形", "649310.705602", "3377545.834305", "4", "0.014", "3000", "1", "1.8", "", "", "", "", "", "", "", "0.1", "100"),
            ("10", "2", "沪蓉", "倒虹吸", "649264.563059", "3377548.912938", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("11", "2", "沪蓉", "倒虹吸", "649244.41293", "3377550.257356", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("12", "2", "沪蓉", "倒虹吸", "649220.867829", "3377563.964679", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("13", "2", "沪蓉", "倒虹吸", "649184.732272", "3377556.518614", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("14", "2", "沪蓉", "倒虹吸", "649146.2872", "3377588.1779", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("15", "2", "宋家沟", "隧洞-圆拱直墙型", "649104.399377", "3377613.995873", "4", "0.014", "2000", "", "2.5", "", "", "", "", "", "", "120", "0.1", "100"),
            ("16", "2", "宋家沟", "隧洞-圆拱直墙型", "649098.598741", "3377595.290122", "4", "0.014", "2000", "", "2.5", "", "", "", "", "", "", "120", "0.1", "100"),
            ("17", "2", "广岳", "倒虹吸", "649086.007282", "3377582.009467", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("18", "2", "广岳", "倒虹吸", "649066.369061", "3377577.838164", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("19", "2", "广岳", "倒虹吸", "649033.673293", "3377532.99707", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("20", "2", "广岳", "倒虹吸", "649018.983612", "3377482.347484", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("21", "2", "伍家沟", "隧洞-圆拱直墙型", "648991.829093", "3377453.363461", "4", "0.014", "2000", "", "2.5", "", "", "", "", "", "", "120", "0.1", "100"),
            ("22", "2", "伍家沟", "隧洞-圆拱直墙型", "648969.028473", "3377444.44637", "4", "0.014", "2000", "", "2.5", "", "", "", "", "", "", "120", "0.1", "100"),
            ("23", "2", "伍家沟", "隧洞-圆拱直墙型", "648918.161636", "3377447.833438", "4", "0.014", "2000", "", "2.5", "", "", "", "", "", "", "120", "0.1", "100"),
            ("24", "2", "刘家沟", "渡槽-矩形", "648879.873566", "3377424.400731", "4", "0.014", "2000", "", "", "", "", "", "0.8", "30", "0.3", "120", "0.1", "100"),
            ("25", "2", "刘家沟", "渡槽-矩形", "648873.319207", "3377389.201113", "4", "0.014", "2000", "", "", "", "", "", "0.8", "30", "0.3", "120", "0.1", "100"),
            ("26", "2", "广高路", "倒虹吸", "648875.83158", "3377349.478728", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("27", "2", "广高路", "倒虹吸", "648859.515404", "3377325.867714", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("28", "2", "广高路", "倒虹吸", "648823.413217", "3377328.752934", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("29", "2", "广高路", "倒虹吸", "648778.747964", "3377306.056947", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("30", "2", "广高路", "倒虹吸", "648742.967801", "3377279.279514", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("31", "2", "广高路", "倒虹吸", "648740.770589", "3377262.028944", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("32", "2", "广高路", "倒虹吸", "648704.058844", "3377256.358551", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("33", "2", "-", "明渠-圆形", "648687.241348", "3377234.03888", "4", "0.014", "3000", "0", "", "", "", "2.8", "", "", "", "", "0.1", "100"),
            ("34", "2", "-", "明渠-圆形", "648677.298141", "3377230.295614", "4", "0.014", "3000", "0", "", "", "", "2.8", "", "", "", "", "0.1", "100"),
            ("35", "2", "台儿沟", "隧洞-圆形", "648610.458063", "3377205.132683", "4", "0.014", "2000", "", "", "", "", "2.6", "", "", "", "", "0.1", "100"),
            ("36", "2", "美团沟", "分水闸", "648588.193106", "3377182.717782", "4", "0.014", "", "", "", "", "", "", "", "", "", "", "0.1", "100"),
            ("37", "3", "台儿沟", "隧洞-圆形", "648359.767433", "3377105.690753", "3", "0.014", "2000", "", "", "", "", "2.6", "", "", "", "", "0.1", "100"),
            ("38", "3", "台儿沟", "隧洞-圆形", "648259.932461", "3376966.162254", "3", "0.014", "2000", "", "", "", "", "2.6", "", "", "", "", "0.1", "100"),
            ("39", "3", "梨子园", "渡槽-U形", "647962.330045", "3376909.650621", "3", "0.014", "1500", "", "", "", "1.6", "", "", "", "", "", "0.1", "100"),
            ("40", "3", "梨子园", "渡槽-U形", "647644.538898", "3376595.606329", "3", "0.014", "1500", "", "", "", "1.6", "", "", "", "", "", "0.1", "100"),
            ("41", "3", "油房垭", "隧洞-圆拱直墙型", "647641.215709", "3376559.422576", "3", "0.014", "2000", "", "2.5", "", "", "", "", "", "", "120", "0.1", "100"),
            ("42", "3", "油房垭", "隧洞-圆拱直墙型", "647597.709292", "3376537.187663", "3", "0.014", "2000", "", "2.5", "", "", "", "", "", "", "120", "0.1", "100"),
            ("43", "3", "-", "矩形暗涵", "647506.778347", "3376513.531331", "3", "0.014", "3000", "0", "2.4", "", "", "", "", "", "", "", "0.1", "100"),
            ("44", "3", "-", "矩形暗涵", "647387.9806", "3376403.8971", "3", "0.014", "3000", "", "2.4", "", "", "", "", "", "", "", "", ""),
        ]
        self._set_input_sheet_data([list(r) for r in samples], redraw=True, reset_row_positions=True)
        
        # 设置默认起始水位
        self.var_start_water_level.set("400")
    
    def _on_tree_click(self, event):
        """处理表格单击事件，记录选中位置用于粘贴"""
        region = self.input_tree.identify_region(event.x, event.y)
        column = self.input_tree.identify_column(event.x)
        
        if not column:
            return
        
        # 获取列索引（从#1开始，需转换为0-based）
        col_idx = int(column.replace("#", "")) - 1
        
        if region == "heading":
            # 点击表头，记录列索引用于整列粘贴
            self._header_click_col_idx = col_idx
            self._selected_cell_item = None
            self._selected_cell_col_idx = None
        else:
            # 点击单元格
            item = self.input_tree.identify_row(event.y)
            if item:
                self._selected_cell_item = item
                self._selected_cell_col_idx = col_idx
                self._header_click_col_idx = None
                # 选中该行
                self.input_tree.selection_set(item)
    
    def _validate_section_type(self, input_type: str):
        """验证并映射断面类型
        
        返回:
            (mapped_type, error_msg) 元组
            - 有效类型: (映射后的类型, None)
            - 废弃类型: (None, 错误提示)
            - 无效类型: (None, None)
        """
        if not input_type:
            return (None, None)
        input_type = str(input_type).strip()
        
        # 检查是否是有效类型
        if input_type in SECTION_TYPE_MAPPING:
            return (SECTION_TYPE_MAPPING[input_type], None)
        
        # 检查是否是废弃类型
        if input_type in DEPRECATED_SECTION_TYPES:
            correct_type = DEPRECATED_SECTION_TYPES[input_type]
            return (None, f'"{input_type}"已废弃，请使用"{correct_type}"')
        
        return (None, None)
    
    def _paste_to_cells(self, event=None):
        """从剪贴板粘贴数据到选中位置（Excel风格）"""
        try:
            clipboard_data = self.clipboard_get()
        except tk.TclError:
            messagebox.showwarning("提示", "剪贴板为空或无法读取")
            return
        
        if not clipboard_data or not clipboard_data.strip():
            messagebox.showwarning("提示", "剪贴板为空")
            return
        
        # 解析剪贴板内容（\n分行，\t分列）
        lines = clipboard_data.strip().split('\n')
        
        # 确定粘贴起始位置
        if self._header_click_col_idx is not None:
            start_col = self._header_click_col_idx
            start_row_idx = 0
        elif self._selected_cell_item and self._selected_cell_col_idx is not None:
            start_col = self._selected_cell_col_idx
            start_row_idx = self.input_tree.index(self._selected_cell_item)
        else:
            messagebox.showwarning("提示", "请先单击要粘贴的起始位置（表头或单元格）")
            return
        
        # 获取当前所有行
        items = list(self.input_tree.get_children())
        
        # 收集无效断面类型警告
        invalid_types = []
        deprecated_types = []  # 废弃类型错误
        pasted_rows = 0
        
        for row_offset, line in enumerate(lines):
            if not line.strip():
                continue
            
            cols = line.split('\t')
            target_row_idx = start_row_idx + row_offset
            
            # 自动新增行（如果需要）
            while target_row_idx >= len(items):
                # 获取行数以确定新行序号
                existing_count = len(items)
                new_seq = existing_count + 1
                new_segment = 1  # 默认流量段1
                new_Q = str(self.get_flow_for_segment(new_segment))
                new_values = self._create_default_row(new_seq, new_segment, new_Q)
                self.input_tree.insert('', 'end', values=new_values)
                items = list(self.input_tree.get_children())
            
            target_item = items[target_row_idx]
            values = list(self.input_tree.item(target_item)['values'])
            
            # 确保values长度足够
            while len(values) < 20:
                values.append("")
            
            for col_offset, cell_value in enumerate(cols):
                target_col = start_col + col_offset
                if target_col < len(values):
                    cell_value = cell_value.strip()
                    
                    # 断面类型列(索引3)特殊处理
                    if target_col == 3:
                        mapped, error_msg = self._validate_section_type(cell_value)
                        if mapped:
                            values[target_col] = mapped
                        elif error_msg:
                            # 废弃类型错误
                            deprecated_types.append(f"行{target_row_idx + 1}: {error_msg}")
                        elif cell_value:
                            # 无效类型
                            invalid_types.append(f"行{target_row_idx + 1}: {cell_value}")
                    else:
                        values[target_col] = cell_value
            
            self.input_tree.item(target_item, values=values)
            pasted_rows += 1
        
        # 刷新表格显示
        self.input_tree.update_idletasks()
        
        # 显示结果
        has_errors = deprecated_types or invalid_types
        if has_errors:
            error_msgs = []
            
            # 废弃类型错误（优先显示）
            if deprecated_types:
                error_msgs.append("【已废弃的类型】")
                for msg in deprecated_types[:5]:
                    error_msgs.append(msg)
                if len(deprecated_types) > 5:
                    error_msgs.append(f"... 共{len(deprecated_types)}条")
            
            # 无效类型
            if invalid_types:
                if deprecated_types:
                    error_msgs.append("")
                error_msgs.append("【无效的类型】")
                for msg in invalid_types[:5]:
                    error_msgs.append(msg)
                if len(invalid_types) > 5:
                    error_msgs.append(f"... 共{len(invalid_types)}条")
            
            error_msgs.append("")
            error_msgs.append("有效类型：明渠-矩形、明渠-梯形、明渠-圆形、隧洞-圆拱直墙型、隧洞-马蹄形Ⅰ型、隧洞-马蹄形Ⅱ型、隧洞-圆形、渡槽-矩形、渡槽-U形、矩形暗涵、倒虹吸")
            
            messagebox.showwarning("断面类型错误", 
                f"已粘贴 {pasted_rows} 行数据\n\n" + "\n".join(error_msgs))
        else:
            messagebox.showinfo("成功", f"已粘贴 {pasted_rows} 行数据")
        
        # 粘贴后自动识别流量段并更新流量设置
        self._auto_detect_flow_segments()
        
        # 粘贴后自动调整列宽以适应新数据
        self._auto_fit_input_columns()
    
    def _add_row(self):
        """新增一行 - 在表格末尾添加新行，流量段编号递增"""
        data = self._get_input_sheet_data()
        if data:
            try:
                last_segment = int(str(data[-1][1]).strip())
            except Exception:
                last_segment = 1
            new_segment = last_segment + 1
            new_seq = len(data) + 1
        else:
            new_seq = 1
            new_segment = 1

        new_Q = str(self.get_flow_for_segment(new_segment))
        new_values = list(self._create_default_row(new_seq, new_segment, new_Q))
        data.append(new_values)
        self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)
        self._renumber_rows()

        try:
            self.input_sheet.see(len(data) - 1, 0, keep_yscroll=False, keep_xscroll=True, redraw=False)
        except Exception:
            pass
        try:
            self.input_sheet.set_currently_selected(len(data) - 1, 0)
            self.input_sheet.select_row(len(data) - 1, redraw=False)
        except Exception:
            pass
    
    def _insert_row(self):
        """插入行 - 在选中行之前插入新行，继承上一行的流量段"""
        data = self._get_input_sheet_data()
        if not data:
            messagebox.showwarning("警告", "请先在表格中选择插入位置")
            return

        selected_rows = self._get_selected_input_rows()
        if selected_rows:
            insert_at = min(selected_rows)
        else:
            messagebox.showwarning("警告", "请先选择要在其之前插入行的位置")
            return

        if insert_at > 0 and insert_at - 1 < len(data):
            prev_values = data[insert_at - 1]
            try:
                new_segment = int(str(prev_values[1]).strip())
            except Exception:
                new_segment = 1
        else:
            new_segment = 1

        new_Q = str(self.get_flow_for_segment(new_segment))
        new_seq = len(data) + 1
        new_values = list(self._create_default_row(new_seq, new_segment, new_Q))
        data.insert(insert_at, new_values)
        self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)

        self._renumber_rows()
        try:
            self.input_sheet.set_currently_selected(insert_at, 0)
            self.input_sheet.select_row(insert_at, redraw=False)
            self.input_sheet.see(insert_at, 0, keep_yscroll=False, keep_xscroll=True, redraw=False)
        except Exception:
            pass
    
    def _add_same_segment_building(self):
        """新增同段建筑物 - 复用选中行的流量段编号和Q参数"""
        data = self._get_input_sheet_data()
        if not data:
            messagebox.showwarning("警告", "请先在表格中选择一行作为参考")
            return

        ref_row = None
        selected_rows = self._get_selected_input_rows()

        if selected_rows:
            r = min(selected_rows)
            if 0 <= r < len(data):
                ref_row = data[r]

        if not ref_row:
            messagebox.showwarning("警告", "请先选择一行作为参考")
            return

        try:
            ref_segment = int(str(ref_row[1]).strip())
        except Exception:
            ref_segment = 1
        ref_Q = str(ref_row[6]) if len(ref_row) > 6 else str(self.get_flow_for_segment(ref_segment))

        new_seq = len(data) + 1
        new_values = list(self._create_default_row(new_seq, ref_segment, ref_Q))
        data.append(new_values)
        self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)
        self._renumber_rows()
        try:
            self.input_sheet.set_currently_selected(len(data) - 1, 0)
            self.input_sheet.select_row(len(data) - 1, redraw=False)
            self.input_sheet.see(len(data) - 1, 0, keep_yscroll=False, keep_xscroll=True, redraw=False)
        except Exception:
            pass
    
    def _create_default_row(self, seq: int, segment: int, Q: str) -> tuple:
        """创建默认行数据 - 内部辅助方法"""
        # 新列顺序（已删除桩号，新增X、Y）: 序号, 流量段, 建筑物名称, 结构形式, X, Y, Q, 糙率n, 比降, 边坡系数m, 底宽B, 明渠宽深比, 半径R, 直径D, 矩形渡槽深宽比, 倒角角度, 倒角底边, 圆心角, 不淤流速, 不冲流速
        return (
            str(seq),           # 0: 序号
            str(segment),       # 1: 流量段
            "-",                # 2: 建筑物名称 (明渠默认为-)
            "明渠-梯形",        # 3: 结构形式
            "",                 # 4: X (坐标)
            "",                 # 5: Y (坐标)
            Q,                  # 6: Q(m³/s)
            "0.014",            # 7: 糙率n
            "3000",             # 8: 比降(1/)
            "1.0",              # 9: 边坡系数m
            "",                 # 10: 底宽B(m)
            "",                 # 11: 明渠宽深比
            "",                 # 12: 半径R(m)
            "",                 # 13: 直径D(m)
            "",                 # 14: 矩形渡槽深宽比
            "",                 # 15: 倒角角度
            "",                 # 16: 倒角底边
            "",                 # 17: 圆心角
            "0.1",              # 18: 不淤流速
            "100"               # 19: 不冲流速
        )
    
    def _delete_row(self):
        """删除选中行 - 带确认对话框"""
        data = self._get_input_sheet_data()
        if not data:
            messagebox.showwarning("警告", "表格为空")
            return

        selected_rows = self._get_selected_input_rows()

        selected_rows = sorted({r for r in selected_rows if isinstance(r, int) and 0 <= r < len(data)})
        if not selected_rows:
            messagebox.showwarning("警告", "请先选择要删除的行")
            return

        count = len(selected_rows)
        if count == 1:
            item_values = data[selected_rows[0]]
            confirm_msg = f"确定要删除行 '{item_values[2]}' 吗?"
        else:
            confirm_msg = f"确定要删除选中的 {count} 行吗?"

        if not messagebox.askyesno("确认删除", confirm_msg, icon="warning"):
            return

        for r in reversed(selected_rows):
            if 0 <= r < len(data):
                data.pop(r)
        self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)

        self._renumber_rows()
    
    def _copy_row(self):
        """复制选中行"""
        data = self._get_input_sheet_data()
        if not data:
            messagebox.showwarning("警告", "表格为空")
            return

        selected_rows = self._get_selected_input_rows()
        if not selected_rows:
            messagebox.showwarning("警告", "请先选择要复制的行")
            return

        for r in selected_rows:
            if r < 0 or r >= len(data):
                continue
            values = list(data[r])
            values[0] = str(len(data) + 1)
            data.append(values)

        self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)
        self._renumber_rows()
    
    def _copy_selection_to_clipboard(self, event=None, include_header: bool = False):
        """将当前选区复制到剪贴板（TSV，Excel兼容）"""
        data = self._sync_input_data_from_views()
        if not data:
            return "break"

        ranges = []
        if getattr(self, "_freeze_enabled", False) and getattr(self, "frozen_sheet", None) is not None:
            for sheet, off in ((self.frozen_sheet, 0), (self.input_sheet, 4)):
                try:
                    has_sel = sheet.anything_selected()
                except Exception:
                    has_sel = True
                if not has_sel:
                    continue
                try:
                    r1, c1, r2, c2 = sheet.get_selected_min_max()
                    ranges.append((r1, off + c1, r2, off + c2))
                except Exception:
                    pass
        else:
            try:
                r1, c1, r2, c2 = self.input_sheet.get_selected_min_max()
                ranges.append((r1, c1, r2, c2))
            except Exception:
                pass

        if ranges:
            min_r = min(r[0] for r in ranges)
            min_c = min(r[1] for r in ranges)
            max_r = max(r[2] for r in ranges)
            max_c = max(r[3] for r in ranges)
        else:
            active = self._get_input_active_cell()
            if not active:
                try:
                    active = self.frozen_sheet.get_currently_selected()
                except Exception:
                    active = None
            if not active:
                return "break"
            min_r = max_r = int(active[0])
            min_c = max_c = int(active[1])

        min_r = max(0, min_r)
        min_c = max(0, min_c)
        max_r = min(len(data), max_r)
        max_c = min(len(self._input_headers), max_c)

        lines = []
        if include_header:
            header_line = "\t".join(str(h) for h in self._input_headers[min_c:max_c])
            lines.append(header_line)

        for r in range(min_r, max_r):
            row = data[r]
            line = "\t".join("" if v is None else str(v) for v in row[min_c:max_c])
            lines.append(line)

        clipboard_text = "\n".join(lines)
        try:
            self.clipboard_clear()
            self.clipboard_append(clipboard_text)
        except Exception:
            pass
        return "break"

    def _copy_result_selection_to_clipboard(self, event=None, include_header: bool = False):
        """将结果表当前选区复制到剪贴板（TSV，Excel兼容）"""
        try:
            min_r, min_c, max_r, max_c = self.result_sheet.get_selected_min_max()
        except Exception:
            try:
                active = self.result_sheet.get_currently_selected()
                if active:
                    min_r = max_r = int(active[0])
                    min_c = max_c = int(active[1])
                else:
                    return "break"
            except Exception:
                return "break"

        data = self._get_result_sheet_data()
        if not data:
            return "break"

        min_r = max(0, min_r)
        min_c = max(0, min_c)
        max_r = min(len(data), max_r)
        max_c = min(len(self._result_headers), max_c)

        lines = []
        if include_header:
            header_line = "\t".join(str(h) for h in self._result_headers[min_c:max_c])
            lines.append(header_line)

        for r in range(min_r, max_r):
            row = data[r]
            line = "\t".join("" if v is None else str(v) for v in row[min_c:max_c])
            lines.append(line)

        clipboard_text = "\n".join(lines)
        try:
            self.clipboard_clear()
            self.clipboard_append(clipboard_text)
        except Exception:
            pass
        return "break"

    def _fill_down_input(self) -> None:
        data = self._sync_input_data_from_views()
        if not data:
            return

        cells = []
        if getattr(self, "_freeze_enabled", False) and getattr(self, "frozen_sheet", None) is not None:
            try:
                for r, c in (self.frozen_sheet.get_selected_cells() or []):
                    if isinstance(r, int) and isinstance(c, int):
                        cells.append((r, c))
            except Exception:
                pass
            try:
                for r, c in (self.input_sheet.get_selected_cells() or []):
                    if isinstance(r, int) and isinstance(c, int):
                        cells.append((r, 4 + c))
            except Exception:
                pass
        else:
            try:
                for r, c in (self.input_sheet.get_selected_cells() or []):
                    if isinstance(r, int) and isinstance(c, int):
                        cells.append((r, c))
            except Exception:
                pass

        if not cells:
            active = self._get_input_active_cell()
            if not active:
                return
            cells = [active]

        by_col: Dict[int, List[int]] = {}
        for r, c in cells:
            by_col.setdefault(c, []).append(r)

        changed = False
        for c, rows in by_col.items():
            if c < 0 or c >= len(self._input_headers):
                continue
            min_row = min(rows)
            if min_row < 0 or min_row >= len(data):
                continue
            fill_value = data[min_row][c]
            for r in rows:
                if r == min_row:
                    continue
                if r < 0 or r >= len(data):
                    continue
                if data[r][c] != fill_value:
                    data[r][c] = fill_value
                    changed = True

        if changed:
            self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)
    
    def _paste_from_excel(self):
        """从Excel粘贴数据"""
        try:
            target = self.input_sheet
            if getattr(self, "_freeze_enabled", False) and getattr(self, "frozen_sheet", None) is not None:
                try:
                    if self.frozen_sheet.anything_selected():
                        target = self.frozen_sheet
                except Exception:
                    pass
            target.paste()
        except Exception as e:
            messagebox.showerror("错误", f"粘贴失败: {str(e)}")
    
    def _import_from_excel(self):
        """从Excel文件导入数据（支持多流量段表格填写示例格式）
        
        Excel模板格式要求：
        - 第1行: A1=渠道名称(标签), B1=渠道名称(值), C1=渠道级别(标签), D1=渠道级别(值), 
                 E1=渠道起始水位高程(标签), F1=水位(值), G1=起始桩号(标签), H1=桩号(值)
        - 第2行: 表头
        - 第3行起: 数据
        """
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("错误", "需要安装openpyxl库\n请运行: pip install openpyxl")
            return
        
        filepath = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not filepath:
            return
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # 读取第1行信息（新格式）
            # A1=渠道名称(标签), B1=渠道名称(值)
            # C1=渠道级别(标签), D1=渠道级别(值)
            # E1=渠道起始水位高程(标签), F1=水位(值)
            # G1=起始桩号(标签), H1=桩号(值)
            import_info_parts = []
            
            # 读取渠道名称 (B1)
            channel_name_label = ws.cell(row=1, column=1).value
            channel_name_value = ws.cell(row=1, column=2).value
            if channel_name_label and "渠道名称" in str(channel_name_label):
                if channel_name_value is not None:
                    self.var_channel_name.set(str(channel_name_value))
                    import_info_parts.append(f"渠道名称: {channel_name_value}")
            
            # 读取渠道级别 (D1) - Excel下拉框选择的值可以正常读取
            channel_level_label = ws.cell(row=1, column=3).value
            channel_level_value = ws.cell(row=1, column=4).value
            if channel_level_label and "渠道级别" in str(channel_level_label):
                if channel_level_value is not None:
                    level_str = str(channel_level_value).strip()
                    # 验证是否为有效的渠道级别选项
                    valid_levels = [
                        "总干渠", "总干管", "分干渠", "分干管", 
                        "干渠", "干管", "支渠", "支管", "分支渠", "分支管"
                    ]
                    if level_str in valid_levels:
                        self.var_channel_level.set(level_str)
                        import_info_parts.append(f"渠道级别: {level_str}")
                    else:
                        import_info_parts.append(f"渠道级别: {level_str} (无效值，已忽略)")
            
            # 读取渠道起始水位高程 (F1)
            start_level_label = ws.cell(row=1, column=5).value
            start_level_value = ws.cell(row=1, column=6).value
            if start_level_label and ("起始水位" in str(start_level_label) or "水位" in str(start_level_label)):
                if start_level_value is not None:
                    self.var_start_water_level.set(str(start_level_value))
                    import_info_parts.append(f"起始水位: {start_level_value} m")
            
            # 读取起始桩号 (H1) - 格式为 "0+000.000" 或纯数字
            start_station_label = ws.cell(row=1, column=7).value
            start_station_value = ws.cell(row=1, column=8).value
            if start_station_value is not None:
                station_str = str(start_station_value).strip()
                if station_str:
                    # 解析桩号值并格式化显示
                    station_value = parse_station_input(station_str)
                    self.set_start_station_value(station_value)
                    formatted_station = format_station_display(station_value)
                    import_info_parts.append(f"起始桩号: {formatted_station}")
            
            # 读取数据（第3行起）
            data_rows = []
            for row_idx in range(3, ws.max_row + 1):
                row_data = []
                for col_idx in range(1, 21):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                if any(v.strip() for v in row_data):
                    data_rows.append(row_data)
            
            if not data_rows:
                messagebox.showwarning("提示", "Excel文件中没有数据")
                return
            
            self._set_input_sheet_data(data_rows, redraw=True, reset_row_positions=True)
            self._auto_detect_flow_segments()
            
            # 构建导入结果消息
            import_info_msg = "\n".join(import_info_parts) if import_info_parts else ""
            if import_info_msg:
                import_info_msg = "\n" + import_info_msg
            
            messagebox.showinfo("成功", f"已成功导入 {len(data_rows)} 行数据{import_info_msg}")
            
        except Exception as e:
            messagebox.showerror("错误", f"导入失败: {str(e)}")
    
    def _edit_cell(self, event):
        """双击编辑单元格"""
        # 使用 identify_row 直接获取双击位置的行，避免选择状态滞后问题
        item = self.input_tree.identify_row(event.y)
        if not item:
            return
        
        # 确保双击的行被选中，保持选择状态一致
        self.input_tree.selection_set(item)
        self.input_tree.focus(item)
        
        column = self.input_tree.identify_column(event.x)
        col_index = int(column.replace('#', '')) - 1
        
        # 获取当前行的结构形式
        values = self.input_tree.item(item)['values']
        section_type = values[3]
        
        # 新列索引（新增X、Y列）: 0序号, 1流量段, 2建筑物名称, 3结构形式, 4X, 5Y, 6Q, 7糙率n, 8比降, 
        #          9边坡系数m, 10底宽B, 11明渠宽深比, 12半径R, 13直径D, 
        #          14矩形渡槽深宽比, 15倒角角度, 16倒角底边, 17圆心角, 18不淤流速, 19不冲流速
        
        # 根据列索引决定编辑方式
        if col_index == 3:
            # 结构形式列(索引3),使用下拉框
            self._edit_cell_with_combobox(item, column, col_index)
        elif col_index == 2:
            # 建筑物名称列 - 明渠和矩形暗涵无需输入
            if "明渠" in section_type or section_type == "矩形暗涵":
                messagebox.showinfo("提示", f"{section_type}类型无需输入建筑物名称")
                return
            self._edit_cell_with_entry(item, column, col_index)
        elif col_index in [0, 1, 4, 5]:
            # 序号、流量段、X、Y列,使用文本框直接编辑
            self._edit_cell_with_entry(item, column, col_index)
        elif col_index >= 6 and col_index <= 19:
            # 参数列(6-19),打开参数设置弹窗
            self._open_parameter_dialog()
    
    def _edit_cell_with_combobox(self, item, column, col_index):
        """使用下拉框编辑断面类型列"""
        # 获取单元格位置
        x, y, width, height = self.input_tree.bbox(item, column)
        
        # 获取当前值
        value = self.input_tree.item(item)['values'][col_index]
        
        # 定义所有可用的断面类型
        section_types = [
            "明渠-梯形",
            "明渠-矩形",
            "明渠-圆形",
            "渡槽-U形",
            "渡槽-矩形",
            "隧洞-圆形",
            "隧洞-圆拱直墙型",
            "隧洞-马蹄形Ⅰ型",
            "隧洞-马蹄形Ⅱ型",
            "矩形暗涵",
            "倒虹吸",  # 占位行，不参与计算
            "分水闸",  # 占位行，不参与计算（仅过闸水头损失）
            "分水口",  # 占位行，不参与计算（仅过闸水头损失）
        ]
        
        # 创建下拉框
        edit_var = tk.StringVar(value=str(value))
        edit_combo = ttk.Combobox(self.input_tree, textvariable=edit_var, 
                                  values=section_types, state='readonly')
        edit_combo.place(x=x, y=y, width=width, height=height)
        
        # 延迟设置焦点和自动展开，确保组件已完全加载且避免与 Treeview 双击事件冲突
        def auto_open():
            if edit_combo.winfo_exists():
                edit_combo.focus_set()
                # 使用 <Down> 信号触发展开，在 readonly 模式下比 <Button-1> 更稳定
                edit_combo.event_generate('<Down>')
        
        self.after(100, auto_open)
        
        def save_edit(event=None):
            new_value = edit_var.get()
            old_values = self.input_tree.item(item)['values']
            old_type = old_values[3]  # 结构形式在索引3
            
            # 如果类型没有变化，直接退出
            if new_value == old_type:
                edit_combo.destroy()
                return

            values = list(old_values)
            values[col_index] = new_value
            
            # 1. 处理建筑物名称
            if "明渠" in new_value or new_value == "矩形暗涵":
                values[2] = "-"  # 建筑物名称设为"-"表示无需命名
            elif "分水" in new_value:
                # 分水闸/分水口：保留或设置建筑物名称
                if str(values[2]).strip() == "-" or not str(values[2]).strip():
                    values[2] = f"分水闸{values[0]}"
            elif values[2] == "-":
                # 从明渠切换回其他需要名称的类型时，给予默认名称
                values[2] = f"建筑物{values[0]}"
            
            # 2. 分水闸/分水口特殊处理：清空所有断面参数
            if "分水" in new_value:
                for i in range(8, 18):
                    values[i] = ""
                self.input_tree.item(item, values=values)
                edit_combo.destroy()
                # 显示友好提示
                self._show_diversion_gate_hint(values[0])
                return
                
            # 3. 清理和重置参数列 (索引9-17: 可选参数，新增X、Y后索引后移)
            # 索引映射: 8:比降, 9:m, 10:b, 11:明渠宽深比, 12:R, 13:D, 14:矩形渡槽深宽比, 15:倒角角度, 16:倒角底边, 17:圆心角
            for i in range(9, 18):
                values[i] = ""
            
            # 确保比降参数保持合理默认值
            if not values[8] or str(values[8]).strip() == "":
                values[8] = "3000"
                
            # 4. 根据新类型设置特定默认值
            if "明渠-梯形" in new_value:
                values[9] = "1.0"  # 梯形断面默认给1.0的边坡系数
                
            self.input_tree.item(item, values=values)
            edit_combo.destroy()
        
        edit_combo.bind('<<ComboboxSelected>>', save_edit)
        # 延迟绑定 FocusOut，防止在双击创建瞬间因事件冒泡导致误触发销毁，提升“丝滑”感
        self.after(200, lambda: edit_combo.bind('<FocusOut>', save_edit) if edit_combo.winfo_exists() else None)
        edit_combo.bind('<Escape>', lambda e: edit_combo.destroy())
    
    def _edit_cell_with_entry(self, item, column, col_index):
        """使用文本框编辑普通列"""
        # 获取单元格位置
        x, y, width, height = self.input_tree.bbox(item, column)
        
        # 检查是否为个性化参数列,且当前结构形式不支持该参数
        values = self.input_tree.item(item)['values']
        section_type = values[3]  # 结构形式在第4列(索引3)
        
        # 定义每种结构形式支持的个性化参数
        # col_index: 14=矩形渡槽深宽比, 15=倒角角度, 16=倒角底边, 17=圆心角（新增X、Y后索引后移）
        param_support = {
            "明渠-梯形": [],  # 不支持任何个性化参数
            "明渠-矩形": [],  # 不支持任何个性化参数
            "明渠-圆形": [],  # 不支持任何个性化参数
            "渡槽-U形": [],  # 不支持任何个性化参数
            "渡槽-矩形": [14, 15, 16],  # 支持矩形渡槽深宽比、倒角角度和倒角底边
            "隧洞-圆形": [],  # 不支持任何个性化参数
            "隧洞-圆拱直墙型": [17],  # 支持圆心角
            "隧洞-马蹄形Ⅰ型": [],  # 不支持任何个性化参数
            "隧洞-马蹄形Ⅱ型": [],  # 不支持任何个性化参数
            "矩形暗涵": []  # 不支持任何个性化参数
        }
        
        # 检查当前列是否为不支持的个性化参数
        if col_index in [14, 15, 16, 17]:  # 个性化参数列
            supported_cols = param_support.get(section_type, [])
            if col_index not in supported_cols:
                # 该结构形式不支持此参数,禁止编辑
                messagebox.showwarning("提示", f"{section_type}不支持此参数,该列无需填写")
                return
        
        # 创建编辑框
        value = self.input_tree.item(item)['values'][col_index]
        
        edit_var = tk.StringVar(value=str(value))
        edit_entry = ttk.Entry(self.input_tree, textvariable=edit_var)
        edit_entry.place(x=x, y=y, width=width, height=height)
        edit_entry.focus()
        edit_entry.select_range(0, tk.END)
        
        def save_edit(event=None):
            new_value = edit_var.get()
            values = list(self.input_tree.item(item)['values'])
            values[col_index] = new_value
            
            # 如果修改的是流量段列(col_index=1),需要同步更新Q值
            if col_index == 1:
                try:
                    segment_num = int(new_value)
                    new_Q = self.get_flow_for_segment(segment_num)
                    values[6] = str(new_Q)  # 更新Q列(索引6，新增X、Y后)
                except (ValueError, TypeError):
                    pass  # 如果流量段不是有效数字,不更新Q
            
            self.input_tree.item(item, values=values)
            edit_entry.destroy()
        
        edit_entry.bind('<Return>', save_edit)
        edit_entry.bind('<FocusOut>', save_edit)
        edit_entry.bind('<Escape>', lambda e: edit_entry.destroy())
    
    def _open_parameter_dialog(self, event=None):
        """打开参数设置弹窗"""
        row_idx = None
        selected_rows = self._get_selected_input_rows()
        if selected_rows:
            row_idx = min(selected_rows)
        else:
            active = self._get_input_active_cell()
            if active:
                row_idx = active[0]

        if row_idx is None:
            messagebox.showwarning("提示", "请先选择要设置参数的行")
            return

        self._open_parameter_dialog_for_row(row_idx)

    def _open_parameter_dialog_for_row(self, row_idx: int) -> None:
        data = self._sync_input_data_from_views()
        if row_idx < 0 or row_idx >= len(data):
            messagebox.showwarning("提示", "请选择有效的行")
            return

        values = data[row_idx]
        section_type = str(values[3]).strip()

        current_values = {
            "Q": values[6],
            "n": values[7],
            "slope_inv": values[8],
            "m": values[9],
            "b": values[10],
            "b_h_ratio": values[11],
            "R": values[12],
            "D": values[13],
            "ducao_depth_ratio": values[14],
            "chamfer_angle": values[15],
            "chamfer_length": values[16],
            "theta": values[17],
            "v_min": values[18],
            "v_max": values[19],
        }

        dialog = SectionParameterDialog(self.winfo_toplevel(), section_type, current_values)
        self.winfo_toplevel().wait_window(dialog)

        result = dialog.get_result()
        if result is not None:
            self._update_table_row(row_idx, result, section_type)
    
    def _update_table_row(self, row_idx: int, params: dict, section_type: str):
        """将弹窗参数回填到表格，并确保表格立即刷新显示"""
        data = self._get_input_sheet_data()
        if row_idx < 0 or row_idx >= len(data):
            return
        values = list(data[row_idx])
        
        # 确保values列表有足够的长度(20列，新增X、Y)
        values = self._normalize_row(values, 20)
        
        # 新列索引映射（新增X、Y列）:
        # 6:Q, 7:糙率n, 8:比降, 9:边坡系数m, 10:底宽B, 11:明渠宽深比, 
        # 12:半径R, 13:直径D, 14:矩形渡槽深宽比, 15:倒角角度, 16:倒角底边, 17:圆心角, 18:不淤流速, 19:不冲流速
        
        # 更新基础参数
        values[6] = str(params.get('Q', values[6]))
        values[7] = str(params.get('n', values[7]))
        values[8] = str(params.get('slope_inv', values[8]))
        values[18] = str(params.get('v_min', values[18]))
        values[19] = str(params.get('v_max', values[19]))
        
        # 清空所有可选参数列，然后根据结构形式填充
        for i in range(9, 18):
            values[i] = ""
        
        # 根据结构形式更新对应的参数列
        if "明渠-梯形" in section_type or "明渠-矩形" in section_type:
            m_val = params.get('m', "")
            values[9] = str(m_val) if m_val != "" else ""
            b_val = params.get('b', "")
            values[10] = str(b_val) if b_val != "" else ""
            ratio_val = params.get('b_h_ratio', "")
            values[11] = str(ratio_val) if ratio_val != "" else ""
            
        elif "明渠-圆形" in section_type:
            D_val = params.get('D', "")
            values[13] = str(D_val) if D_val != "" else ""
            
        elif "渡槽-U形" in section_type:
            R_val = params.get('R', "")
            values[12] = str(R_val) if R_val != "" else ""
            
        elif "渡槽-矩形" in section_type:
            # 深宽比存储在矩形渡槽深宽比列（索引14）
            h_b_ratio = params.get('h_b_ratio', "")
            values[14] = str(h_b_ratio) if h_b_ratio != "" else ""
            chamfer_angle = params.get('chamfer_angle', "")
            chamfer_length = params.get('chamfer_length', "")
            values[15] = str(chamfer_angle) if chamfer_angle != "" else ""
            values[16] = str(chamfer_length) if chamfer_length != "" else ""
            
        elif "隧洞-圆形" in section_type:
            D_val = params.get('D', "")
            values[13] = str(D_val) if D_val != "" else ""
            
        elif "隧洞-圆拱直墙型" in section_type:
            B_val = params.get('B', "")
            values[10] = str(B_val) if B_val != "" else ""
            theta_val = params.get('theta', "")
            values[17] = str(theta_val) if theta_val != "" else ""
            
        elif "隧洞-马蹄形" in section_type:
            r_val = params.get('r', "")
            values[12] = str(r_val) if r_val != "" else ""
            
        elif section_type == "矩形暗涵":
            # 宽深比存储在宽深比列（索引11）
            BH_ratio_val = params.get('BH_ratio_rect', "")
            values[11] = str(BH_ratio_val) if BH_ratio_val != "" else ""
            B_rect_val = params.get('B_rect', "")
            values[10] = str(B_rect_val) if B_rect_val != "" else ""
        
        data[row_idx] = values
        self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)
        try:
            self.input_sheet.set_currently_selected(row_idx, 6)
            self.input_sheet.select_row(row_idx, redraw=False)
            self.input_sheet.see(row_idx, 0, keep_yscroll=False, keep_xscroll=True, redraw=False)
        except Exception:
            pass
    
    def _renumber_rows(self):
        """重新编号行"""
        data = self._get_input_sheet_data()
        for i, row in enumerate(data, 1):
            row[0] = str(i)
        self._set_input_sheet_data(data, redraw=True, reset_row_positions=False)
    
    def _on_start_station_focus_out(self, event=None):
        """起始桩号输入框失去焦点时格式化显示"""
        self._format_start_station_display()
    
    def _on_start_station_focus_in(self, event=None):
        """起始桩号输入框获得焦点时转换为纯数字便于编辑"""
        current = self.var_start_station.get()
        value = parse_station_input(current)
        # 显示纯数字格式便于编辑
        self.var_start_station.set(str(value))
    
    def _on_start_station_enter(self, event=None):
        """起始桩号输入框按Enter键时格式化显示"""
        self._format_start_station_display()
        return "break"  # 阻止Enter键的默认行为
    
    def _format_start_station_display(self):
        """格式化起始桩号显示"""
        current = self.var_start_station.get()
        value = parse_station_input(current)
        formatted = format_station_display(value)
        self.var_start_station.set(formatted)
    
    def get_start_station_value(self) -> float:
        """获取起始桩号的数值（米）"""
        return parse_station_input(self.var_start_station.get())
    
    def set_start_station_value(self, value: float) -> None:
        """
        设置起始桩号值
        
        参数:
            value: 桩号数值（米）
        """
        formatted = format_station_display(value)
        self.var_start_station.set(formatted)
    
    def _clear_input(self):
        """清空输入表格"""
        if messagebox.askyesno("确认", "确定要清空所有输入数据吗?"):
            self._set_input_sheet_data([], redraw=True, reset_row_positions=True)
    
    def _batch_calculate(self):
        """批量计算所有流量段"""
        # 清空之前的结果
        self._clear_results()

        input_rows = self._sync_input_data_from_views()
        if not input_rows:
            messagebox.showwarning("警告", "请先输入数据")
            return

        self.batch_results = []
        success_count = 0
        fail_count = 0
        skip_count = 0  # 统计跳过的倒虹吸数量
        error_details = []  # 收集详细错误信息
        total_count = 0
        result_rows: List[List[Any]] = []
        
        # 遍历每一行进行计算(每一行代表一个建筑物)
        for values in input_rows:
            values = self._normalize_row(values, 20)
            if not any(str(v).strip() for v in values):
                continue

            section_type = str(values[3]).strip()
            if not section_type:
                continue

            total_count += 1
            
            try:
                # 解析输入参数 - 新列索引（新增X、Y列）
                # 0:序号, 1:流量段, 2:建筑物名称, 3:结构形式, 4:X, 5:Y, 6:Q, 7:糙率n, 8:比降
                # 9:边坡系数m, 10:底宽B, 11:明渠宽深比, 12:半径R, 13:直径D
                # 14:矩形渡槽深宽比, 15:倒角角度, 16:倒角底边, 17:圆心角, 18:不淤流速, 19:不冲流速
                seq = str(values[0]).strip() if values[0] is not None else ""
                segment = str(values[1]).strip() if values[1] is not None else ""
                building_name = str(values[2]).strip() if values[2] is not None else ""
                section_type = str(values[3]).strip() if values[3] is not None else ""
                
                # 倒虹吸特殊处理：不参与计算，作为占位行
                if section_type == "倒虹吸":
                    row_out = ["-"] * len(self._result_headers)
                    row_out[0] = seq
                    row_out[1] = segment
                    row_out[2] = building_name
                    row_out[3] = "倒虹吸"
                    row_out[-1] = "⏭ 占位行(不参与计算)"
                    result_rows.append(row_out)
                    
                    # 创建倒虹吸占位结果，用于导入推求水面线
                    def safe_float_local(val, default=0.0):
                        if not val:
                            return default
                        s = str(val).strip()
                        return float(s) if s else default
                    
                    siphon_result = {
                        'success': True,
                        'section_type': "倒虹吸",
                        'is_siphon': True,
                        'flow_section': segment,
                        'building_name': building_name,
                        'coord_X': safe_float_local(values[4], 0.0),
                        'coord_Y': safe_float_local(values[5], 0.0),
                        'Q': safe_float_local(values[6], 0.0),
                        'n': safe_float_local(values[7], 0.014),
                        'slope_inv': 0,
                    }
                    self.batch_results.append({
                        'input': values,
                        'result': siphon_result
                    })
                    
                    skip_count += 1  # 统计跳过数量
                    continue  # 跳过后续计算
                
                # 分水闸/分水口特殊处理：不参与断面计算，仅标记流量段分界
                if "分水" in section_type:
                    row_out = ["-"] * len(self._result_headers)
                    row_out[0] = seq
                    row_out[1] = segment
                    row_out[2] = building_name
                    row_out[3] = section_type
                    row_out[-1] = "⏭ 分水闸(不参与断面计算)"
                    result_rows.append(row_out)
                    
                    # 创建分水闸占位结果，用于导入推求水面线
                    def safe_float_gate(val, default=0.0):
                        if not val:
                            return default
                        s = str(val).strip()
                        return float(s) if s else default
                    
                    gate_result = {
                        'success': True,
                        'section_type': section_type,
                        'is_diversion_gate': True,
                        'flow_section': segment,
                        'building_name': building_name,
                        'coord_X': safe_float_gate(values[4], 0.0),
                        'coord_Y': safe_float_gate(values[5], 0.0),
                        'Q': safe_float_gate(values[6], 0.0),
                        'n': safe_float_gate(values[7], 0.014),
                        'slope_inv': 0,
                    }
                    self.batch_results.append({
                        'input': values,
                        'result': gate_result
                    })
                    
                    skip_count += 1  # 统计跳过数量
                    continue  # 跳过后续计算
                
                # 验证必填参数并给出明确错误提示
                if not values[6] or str(values[6]).strip() == "":
                    raise ValueError(f"流量Q不能为空")
                if not values[7] or str(values[7]).strip() == "":
                    raise ValueError(f"糙率n不能为空")
                if not values[8] or str(values[8]).strip() == "":
                    raise ValueError(f"比降不能为空")
                
                Q = float(values[6])
                n = float(values[7])
                slope_inv = float(values[8])
                
                # 可选参数：使用更简洁的处理方式
                def safe_float(val, default=0):
                    """安全地将值转换为浮点数，空值返回默认值"""
                    if not val:  # None, 0, False, 空字符串
                        return default
                    s = str(val).strip()
                    if not s:  # 空白字符串
                        return default
                    return float(s)
                
                m = safe_float(values[9], 0)           # 边坡系数
                b = safe_float(values[10], 0)           # 底宽
                b_h_ratio = safe_float(values[11], 0)   # 明渠宽深比
                R = safe_float(values[12], 0)         # 半径
                D = safe_float(values[13], 0)         # 直径
                ducao_depth_ratio = safe_float(values[14], 0)  # 矩形渡槽深宽比
                chamfer_angle = safe_float(values[15], 0)
                chamfer_length = safe_float(values[16], 0)
                theta_deg = safe_float(values[17], 0)
                v_min = safe_float(values[18], 0.1)
                v_max = safe_float(values[19], 100)
                
                # 根据断面类型确定geo_param
                # 确保section_type是字符串类型
                section_type_str = str(section_type) if section_type else ""
                
                if "明渠-梯形" in section_type_str or "明渠-矩形" in section_type_str:
                    geo_param = m
                elif "明渠-圆形" in section_type_str or "隧洞-圆形" in section_type_str:
                    geo_param = D
                elif "渡槽-U形" in section_type_str or "隧洞-马蹄形" in section_type_str:
                    geo_param = R
                elif "隧洞-圆拱直墙型" in section_type_str or section_type_str == "矩形暗涵":
                    geo_param = b  # 底宽
                else:
                    geo_param = 0
                
                # 根据断面类型调用不同的计算函数
                # 确保使用字符串类型的section_type
                result = self._calculate_single(section_type_str, Q, geo_param, n, slope_inv, v_min, v_max,
                                                chamfer_angle, chamfer_length, theta_deg, b_h_ratio,
                                                manual_b=b, manual_beta=b_h_ratio, ducao_depth_ratio=ducao_depth_ratio)
                
                if result and result.get('success'):
                    # 提取关键结果（已删除桩号参数）
                    result_row = self._extract_result_data(seq, segment, building_name, section_type, result)
                    result_rows.append(list(result_row))
                    self.batch_results.append({
                        'input': values,
                        'result': result
                    })
                    success_count += 1
                else:
                    error_msg = result.get('error_message', '计算失败') if result else '计算失败'
                    # 记录详细错误（如果是管径/底宽/半径过小等特定错误）
                    if "计算失败：" in error_msg:
                        error_details.append(f"序号{seq} ({building_name}): {error_msg}")

                    row_out = ["-"] * len(self._result_headers)
                    row_out[0] = seq
                    row_out[1] = segment
                    row_out[2] = building_name
                    row_out[3] = section_type
                    row_out[-1] = f"失败: {error_msg}"
                    result_rows.append(row_out)
                    fail_count += 1
                    
            except Exception as e:
                error_msg = str(e)
                row_out = ["-"] * len(self._result_headers)
                row_out[0] = str(values[0]).strip() if values[0] is not None else ""
                row_out[1] = str(values[1]).strip() if values[1] is not None else ""
                row_out[2] = str(values[2]).strip() if values[2] is not None else ""
                row_out[3] = str(values[3]).strip() if values[3] is not None else ""
                row_out[-1] = f"错误: {error_msg}"
                result_rows.append(row_out)
                fail_count += 1

        self._set_result_sheet_data(result_rows, redraw=True, reset_row_positions=True)
        
        # 处理详细输出
        if self.show_detail_var.get() and self.batch_results:
            self.detail_text.configure(state=tk.NORMAL)
            self.detail_text.delete(1.0, tk.END)
            
            full_report = []
            full_report.append("=" * 80)
            full_report.append(" " * 25 + "渠系建筑物多流量段批量水力计算系统 - 详细计算过程报告")
            full_report.append("=" * 80)
            full_report.append(f"计算时间: {os.popen('date /t').read().strip()} {os.popen('time /t').read().strip()}")
            full_report.append(f"总计建筑物数量: {len(self.batch_results)}")
            full_report.append("-" * 80)
            full_report.append("")
            
            for i, item_data in enumerate(self.batch_results, 1):
                input_vals = item_data['input']
                result_data = item_data['result']
                
                # 生成单个建筑物的报告
                report = self._generate_detailed_report(input_vals, result_data)
                full_report.append(f"【项目 {i}】")
                full_report.append(report)
                full_report.append("\n" + "*" * 80 + "\n")
            
            self.detail_text.insert(tk.END, "\n".join(full_report))
            self.detail_text.configure(state=tk.DISABLED)
            
            # 自动切换到详细过程选项卡
            self.result_notebook.select(1)
        
        # 自动调整结果表格列宽
        self._auto_fit_result_columns()
        
        # 显示汇总信息
        summary_msg = f"总计: {total_count}条\n成功: {success_count}条\n失败: {fail_count}条"
        
        # 如果有跳过的倒虹吸/分水闸，添加说明
        if skip_count > 0:
            summary_msg += f"\n跳过: {skip_count}条 (倒虹吸/分水闸占位行不参与断面计算)"
        
        if error_details:
            # 如果有详细错误信息，拼接到汇总信息中
            detailed_errors = "\n\n" + "="*30 + "\n详细失败原因及建议：\n" + "\n\n".join(error_details)
            # 使用较宽的警告弹窗显示
            messagebox.showwarning("批量计算完成 (存在异常)", summary_msg + detailed_errors)
        else:
            messagebox.showinfo("批量计算完成", summary_msg)
        
        # 注册批量计算结果到共享数据管理器，供"从计算结果导入"功能使用
        if SHARED_DATA_AVAILABLE and self.batch_results:
            try:
                shared_data = get_shared_data_manager()
                results_for_register = [item['result'] for item in self.batch_results]
                shared_data.register_batch_results(results_for_register)
            except Exception as e:
                print(f"注册批量结果到共享数据管理器失败: {e}")
    
    def _calculate_single(self, section_type: str, Q: float, geo_param: float, 
                          n: float, slope_inv: float, v_min: float, v_max: float,
                          chamfer_angle: float = 0, chamfer_length: float = 0, theta_deg: float = 0,
                          b_h_ratio: float = 0, manual_b: float = 0, manual_beta: float = 0,
                          ducao_depth_ratio: float = 0) -> Dict[str, Any]:
        """计算单个流量段"""
        # 根据断面类型选择计算函数
        if "明渠-梯形" in section_type:
            if not MINGQU_AVAILABLE:
                return {'success': False, 'error_message': '明渠计算模块未加载'}
            m = geo_param
            b_val = manual_b if manual_b > 0 else None
            beta_val = manual_beta if manual_beta > 0 else None
            return mingqu_calculate(Q=Q, m=m, n=n, slope_inv=slope_inv, 
                                   v_min=v_min, v_max=v_max,
                                   manual_b=b_val, manual_beta=beta_val)
        
        elif "明渠-矩形" in section_type:
            if not MINGQU_AVAILABLE:
                return {'success': False, 'error_message': '明渠计算模块未加载'}
            b_val = manual_b if manual_b > 0 else None
            beta_val = manual_beta if manual_beta > 0 else None
            return mingqu_calculate(Q=Q, m=0, n=n, slope_inv=slope_inv,
                                   v_min=v_min, v_max=v_max,
                                   manual_b=b_val, manual_beta=beta_val)
        
        elif "明渠-圆形" in section_type:
            if not CIRCULAR_AVAILABLE:
                return {'success': False, 'error_message': '圆形断面计算模块未加载'}
            D = geo_param if geo_param > 0 else None
            return circular_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                     v_min=v_min, v_max=v_max, manual_D=D)
        
        elif "渡槽-U形" in section_type:
            if not DUCAO_AVAILABLE:
                return {'success': False, 'error_message': '渡槽计算模块未加载'}
            R = geo_param if geo_param > 0 else None
            return ducao_u_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                    v_min=v_min, v_max=v_max, manual_R=R)
        
        elif "渡槽-矩形" in section_type:
            if not DUCAO_AVAILABLE:
                return {'success': False, 'error_message': '渡槽计算模块未加载'}
            # 传递深宽比和倒角参数
            # 深宽比: 使用ducao_depth_ratio参数，若未指定则默认0.8
            depth_ratio = ducao_depth_ratio if ducao_depth_ratio > 0 else 0.8
            return ducao_rect_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                       v_min=v_min, v_max=v_max,
                                       depth_width_ratio=depth_ratio,
                                       chamfer_angle=chamfer_angle if chamfer_angle > 0 else 0,
                                       chamfer_length=chamfer_length if chamfer_length > 0 else 0)
        
        elif "隧洞-圆形" in section_type:
            if not SUIDONG_AVAILABLE:
                return {'success': False, 'error_message': '隧洞计算模块未加载'}
            D = geo_param if geo_param > 0 else None
            return suidong_circular_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                             v_min=v_min, v_max=v_max, manual_D=D)
        
        elif "隧洞-圆拱直墙型" in section_type:
            if not SUIDONG_AVAILABLE:
                return {'success': False, 'error_message': '隧洞计算模块未加载'}
            B = geo_param if geo_param > 0 else None
            # 传递圆心角参数
            return suidong_horseshoe_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                              v_min=v_min, v_max=v_max, manual_B=B,
                                              theta_deg=theta_deg if theta_deg > 0 else None)
        
        elif "隧洞-马蹄形" in section_type:
            if not SUIDONG_AVAILABLE:
                return {'success': False, 'error_message': '隧洞计算模块未加载'}
            r = geo_param if geo_param > 0 else None
            # 根据Ⅰ型或Ⅱ型选择函数
            return suidong_horseshoe_std_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                                  v_min=v_min, v_max=v_max, manual_r=r,
                                                  section_type=2 if "Ⅱ型" in section_type else 1)
        
        elif "隧洞-矩形暗涵" in section_type or section_type == "矩形暗涵":
            if not SUIDONG_AVAILABLE:
                return {'success': False, 'error_message': '隧洞计算模块未加载'}
            B = geo_param if geo_param > 0 else None
            BH_ratio = b_h_ratio if b_h_ratio > 0 else None
            return suidong_rect_calculate(Q=Q, n=n, slope_inv=slope_inv,
                                         v_min=v_min, v_max=v_max, manual_B=B,
                                         target_BH_ratio=BH_ratio)
        
        else:
            return {'success': False, 'error_message': f'不支持的断面类型: {section_type}'}
    
    def _generate_detailed_report(self, input_vals, result) -> str:
        """生成单个建筑物的详细计算报告"""
        # input_vals 索引映射（新增X、Y列）:
        # 0:序号, 1:流量段, 2:建筑物名称, 3:结构形式, 4:X, 5:Y, 6:Q, 7:糙率n, 8:比降
        # 9:边坡系数m, 10:底宽B, 11:明渠宽深比, 12:半径R, 13:直径D
        # 14:矩形渡槽深宽比, 15:倒角角度, 16:倒角底边, 17:圆心角, 18:不淤流速, 19:不冲流速
        
        section_type = input_vals[3]  # 结构形式在索引3
        building_name = input_vals[2]
        seq = input_vals[0]
        
        header = f"序号: {seq} | 建筑物: {building_name} | 类型: {section_type}\n"
        header += "-" * 50 + "\n"
        
        try:
            # 倒虹吸：占位行，不参与断面计算
            if "倒虹吸" in section_type:
                return header + self._format_placeholder_report(input_vals, result, "倒虹吸")
            # 分水闸/分水口：占位行，流量段分界点
            if "分水" in section_type:
                return header + self._format_diversion_gate_report(input_vals, result)
            if "明渠" in section_type:
                return header + self._format_mingqu_report(input_vals, result)
            elif "渡槽" in section_type:
                return header + self._format_ducao_report(input_vals, result)
            elif "隧洞" in section_type:
                return header + self._format_suidong_report(input_vals, result)
            elif "矩形暗涵" in section_type:
                return header + self._format_suidong_report(input_vals, result)
            else:
                return header + "不支持的断面类型，无法生成详细报告。"
        except Exception as e:
            return header + f"生成详细报告时出错: {str(e)}"

    def _format_placeholder_report(self, input_vals, result, type_name: str) -> str:
        """格式化占位行报告（倒虹吸等不参与断面计算的行）"""
        output = []
        output.append(f"  ⏭ {type_name}占位行，不参与断面计算。")
        coord_x = input_vals[4] if len(input_vals) > 4 else ""
        coord_y = input_vals[5] if len(input_vals) > 5 else ""
        if coord_x and coord_y:
            output.append(f"  坐标: X={coord_x}, Y={coord_y}")
        return "\n".join(output)

    def _format_diversion_gate_report(self, input_vals, result) -> str:
        """格式化分水闸/分水口详细报告"""
        output = []
        section_type = str(input_vals[3])
        segment = str(input_vals[1]).strip()
        
        output.append(f"  ⏭ {section_type} — 流量段分界点（不参与断面计算）")
        output.append("")
        output.append("【基本信息】")
        output.append(f"  所在流量段: 第{segment}段")
        
        # 坐标信息
        coord_x = input_vals[4] if len(input_vals) > 4 else ""
        coord_y = input_vals[5] if len(input_vals) > 5 else ""
        if coord_x and coord_y:
            output.append(f"  坐标: X={coord_x}, Y={coord_y}")
        
        # 流量信息
        q_val = str(input_vals[6]).strip() if len(input_vals) > 6 and input_vals[6] else ""
        if q_val:
            output.append(f"  本段设计流量 Q = {q_val} m³/s")
        
        output.append("")
        output.append("【说明】")
        output.append(f"  {section_type}为流量段分界构筑物，其下游为新的流量段（流量段{int(segment)+1}）。")
        output.append("  该行不参与断面尺寸计算，在推求水面线模块中将产生过闸水头损失（默认0.2m）。")
        
        return "\n".join(output)

    def _format_mingqu_report(self, input_vals, result) -> str:
        """格式化明渠报告"""
        Q = float(input_vals[6])       # Q在索引6（新增X、Y后）
        n = float(input_vals[7])       # 糙率n在索引7
        slope_inv = float(input_vals[8])  # 比降在索引8
        v_min = float(input_vals[18])  # 不淤流速在索引18
        v_max = float(input_vals[19])  # 不冲流速在索引19
        section_type = input_vals[3]   # 结构形式在索引3
        i = 1.0 / slope_inv
        
        output = []
        
        if "圆形" in section_type:
            # 圆形明渠
            D_design = result.get('D_design', 0)
            h_d = result.get('y_d', 0)
            V_d = result.get('V_d', 0)
            A_d = result.get('A_d', 0)
            P_d = result.get('P_d', 0)
            R_d = result.get('R_d', 0)
            PA_d = result.get('PA_d', 0)
            FB_d = result.get('FB_d', 0)
            Q_check_d = result.get('Q_check_d', 0)
            
            output.append("【一、输入参数】")
            output.append(f"  断面类型 = 圆形")
            output.append(f"  设计流量 Q = {Q:.3f} m³/s")
            output.append(f"  糙率 n = {n}")
            output.append(f"  水力坡降 1/{int(slope_inv)}")
            output.append(f"  不淤流速 = {v_min} m/s")
            output.append(f"  不冲流速 = {v_max} m/s")
            output.append("")
            
            output.append("【二、管径确定】")
            output.append(f"  1. 设计管径: D = {D_design:.2f} m")
            pipe_area = math.pi * D_design**2 / 4
            output.append(f"  2. 管道总断面积计算:")
            output.append(f"     A总 = π × D² / 4 = {math.pi:.4f} × {D_design:.2f}² / 4 = {pipe_area:.3f} m²")
            output.append("")
            
            output.append("【三、设计流量工况计算】")
            output.append(f"  Q = {Q:.3f} m³/s")
            output.append(f"  3. 设计水深: h_d = {h_d:.3f} m")
            if h_d > 0 and D_design > 0:
                R_radius = D_design / 2
                theta = 2 * math.acos((R_radius - h_d) / R_radius)
                output.append(f"  4. 圆心角计算:")
                output.append(f"     θ = 2 × arccos((R-h)/R) = {math.degrees(theta):.2f}°")
            output.append(f"  5. 过水面积 A_d = {A_d:.3f} m²")
            output.append(f"  6. 湿周 χ_d = {P_d:.3f} m")
            output.append(f"  7. 水力半径 R_d = {R_d:.3f} m")
            output.append(f"  8. 设计流速 V_d = (1/n) × R^(2/3) × i^(1/2) = {V_d:.3f} m/s")
            output.append(f"  9. 流量校核 Q计算 = {Q_check_d:.3f} m³/s")
            output.append(f"  10. 净空高度 Fb_d = D - h_d = {D_design:.2f} - {h_d:.3f} = {FB_d:.3f} m")
            output.append(f"  11. 净空面积百分比 PA_d = {PA_d:.1f}%")
            output.append("")
            
            # 加大流量
            increase_info = result.get('increase_percent', '20%')
            Q_inc = result.get('Q_inc', 0)
            h_i = result.get('y_i', 0)
            V_i = result.get('V_i', 0)
            A_i = result.get('A_i', 0)
            P_i = result.get('P_i', 0)
            R_i = result.get('R_i', 0)
            Q_check_i = result.get('Q_check_i', 0)
            PA_i = result.get('PA_i', 0)
            FB_i = result.get('FB_i', 0)
            output.append("【四、加大流量工况计算】")
            output.append(f"  流量加大比例 = {increase_info}")
            output.append(f"  Q_i = {Q_inc:.3f} m³/s")
            output.append(f"  1. 加大水深: h_i = {h_i:.3f} m")
            if h_i > 0 and D_design > 0:
                R_radius = D_design / 2
                theta_i = 2 * math.acos((R_radius - h_i) / R_radius)
                output.append(f"  2. 圆心角计算:")
                output.append(f"     θ_i = 2 × arccos((R-h_i)/R)")
                output.append(f"        = 2 × arccos(({R_radius:.3f} - {h_i:.3f}) / {R_radius:.3f})")
                output.append(f"        = 2 × arccos({(R_radius - h_i)/R_radius:.4f})")
                output.append(f"        = {math.degrees(theta_i):.2f}° ({theta_i:.4f} rad)")
                output.append("")
                output.append(f"  3. 过水面积计算:")
                output.append(f"     A_i = (D²/8) × (θ_i - sinθ_i)")
                output.append(f"         = ({D_design:.3f}²/8) × ({theta_i:.4f} - sin{theta_i:.4f})")
                output.append(f"         = {A_i:.3f} m²")
                output.append("")
                output.append(f"  4. 湿周计算:")
                output.append(f"     χ_i = D × θ_i / 2")
                output.append(f"         = {D_design:.3f} × {theta_i:.4f} / 2")
                output.append(f"         = {P_i:.3f} m")
            else:
                output.append("  2. 圆心角及几何参数计算: 条件不足，无法计算")
            output.append("")
            output.append(f"  5. 水力半径计算:")
            output.append(f"     R_i = A_i / χ_i")
            output.append(f"         = {A_i:.3f} / {P_i:.3f}")
            output.append(f"         = {R_i:.3f} m")
            output.append("")
            output.append(f"  6. 加大流速计算 (曼宁公式):")
            output.append(f"     V_i = (1/n) × R_i^(2/3) × i^(1/2)")
            output.append(f"         = (1/{n}) × {R_i:.3f}^(2/3) × {i:.6f}^(1/2)")
            output.append(f"         = {1/n:.2f} × {R_i**(2/3):.4f} × {math.sqrt(i):.6f}")
            output.append(f"         = {V_i:.3f} m/s")
            output.append("")
            output.append(f"  7. 流量校核:")
            output.append(f"     Q计算 = V_i × A_i")
            output.append(f"          = {V_i:.3f} × {A_i:.3f}")
            output.append(f"          = {Q_check_i:.3f} m³/s")
            output.append("")
            output.append(f"  8. 净空高度计算:")
            output.append(f"     Fb_i = D - h_i")
            output.append(f"          = {D_design:.2f} - {h_i:.3f}")
            output.append(f"          = {FB_i:.3f} m")
            output.append("")
            output.append(f"  9. 净空面积计算:")
            output.append(f"     PA_i = (A总 - A_i) / A总 × 100%")
            output.append(f"          = ({pipe_area:.3f} - {A_i:.3f}) / {pipe_area:.3f} × 100%")
            output.append(f"          = {PA_i:.1f}%")
            output.append("")

            output.append("【五、验证】")
            velocity_ok = v_min <= V_d <= v_max
            output.append(f"  1. 流速验证（规范要求）")
            output.append(f"     规范要求: {v_min} ≤ V ≤ {v_max} m/s")
            output.append(f"     计算结果: V = {V_d:.3f} m/s")
            output.append(f"     验证结果: {'通过 ✓' if velocity_ok else '未通过 ✗'}")
            output.append("")
            
            fb_ok = FB_i >= 0.4 # 参考 CircularChannelPanel
            output.append(f"  2. 净空高度验证")
            output.append(f"     规范要求: Fb加大 ≥ 0.4 m")
            output.append(f"     计算结果: Fb加大 = {FB_i:.3f} m")
            output.append(f"     验证结果: {'通过 ✓' if fb_ok else '未通过 ✗'}")
            output.append("")
            
            pa_ok = PA_i >= 15 # 参考 CircularChannelPanel
            output.append(f"  3. 净空面积验证")
            output.append(f"     规范要求: PA加大 ≥ 15 %")
            output.append(f"     计算结果: PA加大 = {PA_i:.1f} %")
            output.append(f"     验证结果: {'通过 ✓' if pa_ok else '未通过 ✗'}")
        else:
            # 梯形/矩形明渠
            m = float(input_vals[9]) if input_vals[9] else 0
            b = result.get('b_design', 0)
            h = result.get('h_design', 0)
            V = result.get('V_design', 0)
            A = result.get('A_design', 0)
            chi = result.get('X_design', 0)
            R = result.get('R_design', 0)
            beta = result.get('Beta_design', 0)
            Q_calc = result.get('Q_calc', 0)
            
            output.append("【一、输入参数】")
            output.append(f"  断面类型 = {section_type}")
            output.append(f"  设计流量 Q = {Q:.3f} m³/s")
            if "梯形" in section_type:
                output.append(f"  边坡系数 m = {m}")
            output.append(f"  糙率 n = {n}")
            output.append(f"  水力坡降 i = 1/{int(slope_inv)} = {i:.6f}")
            output.append(f"  不淤流速 = {v_min} m/s")
            output.append(f"  不冲流速 = {v_max} m/s")
            output.append("")
            
            output.append("【二、设计方法】")
            output.append(f"  采用方法: {result.get('design_method', '自动计算')}")
            output.append("")

            output.append("【三、设计结果】")
            output.append(f"  1. 设计底宽 B = {b:.2f} m")
            output.append(f"  2. 设计水深 h = {h:.3f} m")
            output.append(f"  3. 宽深比 β = B/h = {b:.2f}/{h:.3f} = {beta:.3f}")
            output.append("")
            output.append("  4. 过水面积计算:")
            output.append(f"     A = (B + m×h) × h = ({b:.2f} + {m}×{h:.3f}) × {h:.3f} = {A:.3f} m²")
            output.append("")
            output.append("  5. 湿周计算:")
            sqrt_1_m2 = math.sqrt(1 + m*m)
            output.append(f"     χ = B + 2×h×√(1+m²) = {b:.2f} + 2×{h:.3f}×{sqrt_1_m2:.4f} = {chi:.3f} m")
            output.append("")
            output.append("  6. 水力半径计算:")
            output.append(f"     R = A/χ = {A:.3f}/{chi:.3f} = {R:.3f} m")
            output.append("")
            output.append("  7. 设计流速计算 (曼宁公式):")
            output.append(f"     V = (1/n) × R^(2/3) × i^(1/2) = (1/{n}) × {R:.3f}^(2/3) × {i:.6f}^(1/2) = {V:.3f} m/s")
            output.append("")
            output.append("  8. 流量校核:")
            output.append(f"      Q计算 = V × A = {V:.3f} × {A:.3f} = {Q_calc:.3f} m³/s")
            output.append(f"      误差 = {abs(Q_calc-Q)/Q*100:.2f}%")
            
            output.append("")
            output.append("【四、加大流量工况】")
            inc_pct = result.get('increase_percent', 0)
            Q_inc = result.get('Q_increased', 0)
            h_inc = result.get('h_increased', 0)
            V_inc = result.get('V_increased', 0)
            A_inc = result.get('A_increased', 0)
            X_inc = result.get('X_increased', 0)
            R_inc = result.get('R_increased', 0)
            Fb = result.get('Fb', 0)
            H = result.get('h_prime', 0)

            output.append(f"  加大比例 = {inc_pct:.1f}%")
            output.append(f"  1. 加大流量 Q加大 = Q × (1 + {inc_pct/100:.2f}) = {Q_inc:.3f} m³/s")
            output.append(f"  2. 加大水深 h加大 = {h_inc:.3f} m")
            output.append("")
            output.append("  3. 过水面积计算:")
            output.append(f"     A加大 = (B + m×h加大) × h加大 = ({b:.2f} + {m}×{h_inc:.3f}) × {h_inc:.3f} = {A_inc:.3f} m²")
            output.append("")
            output.append("  4. 湿周计算:")
            sqrt_1_m2 = math.sqrt(1 + m*m)
            output.append(f"     χ加大 = B + 2×h加大×√(1+m²) = {b:.2f} + 2×{h_inc:.3f}×{sqrt_1_m2:.4f} = {X_inc:.3f} m")
            output.append("")
            output.append("  5. 水力半径计算:")
            output.append(f"     R加大 = A加大/χ加大 = {A_inc:.3f}/{X_inc:.3f} = {R_inc:.3f} m")
            output.append("")
            output.append("  6. 加大流速计算 (曼宁公式):")
            output.append(f"     V加大 = (1/n) × R加大^(2/3) × i^(1/2) = (1/{n}) × {R_inc:.3f}^(2/3) × {i:.6f}^(1/2) = {V_inc:.3f} m/s")
            output.append("")
            output.append("  7. 超高和渠道高度:")
            output.append(f"     超高 Fb = (1/4)×h加大 + 0.2 = {Fb:.3f} m")
            output.append(f"     渠道高度 H = h加大 + Fb = {H:.3f} m")
            
            output.append("")
            output.append("【五、验证】")
            velocity_ok = v_min <= V <= v_max
            output.append(f"  1. 流速验证")
            output.append(f"     范围要求: {v_min} ≤ V ≤ {v_max} m/s")
            output.append(f"     计算结果: V = {V:.3f} m/s")
            output.append(f"     验证结果: {'通过 ✓' if velocity_ok else '未通过 ✗'}")
            output.append("")
            
            beta_ok = 0 < beta <= 8.0 # 假设最大宽深比8
            output.append(f"  2. 宽深比验证")
            output.append(f"     规范要求: 0 < β ≤ 8.0")
            output.append(f"     计算结果: β = {beta:.3f}")
            output.append(f"     验证结果: {'通过 ✓' if beta_ok else '未通过 ✗'}")
            output.append("")
            
            # 3. 超高复核（规范 6.4.8-2）
            fb_req = 0.25 * h_inc + 0.2
            fb_ok = Fb >= (fb_req - 0.001) # 允许极小误差
            output.append(f"  3. 超高复核（规范 6.4.8-2）")
            output.append(f"     规范要求: Fb ≥ (1/4)×h加大 + 0.2 = {fb_req:.3f} m")
            output.append(f"     计算结果: Fb = {Fb:.3f} m")
            output.append(f"     验证结果: {Fb:.3f} {'≥' if fb_ok else '<'} {fb_req:.3f} → {'通过 ✓' if fb_ok else '未通过 ✗'}")

        return "\n".join(output)


    def _format_ducao_report(self, input_vals, result) -> str:
        """格式化渡槽报告"""
        Q = float(input_vals[6])       # Q在索引6（新增X、Y后）
        n = float(input_vals[7])       # 糙率n在索引7
        slope_inv = float(input_vals[8])  # 比降在索引8
        v_min = float(input_vals[18])  # 不淤流速在索引18
        v_max = float(input_vals[19])  # 不冲流速在索引19
        section_type = input_vals[3]   # 结构形式在索引3
        i = 1.0 / slope_inv
        
        output = []
        output.append("【一、输入参数】")
        output.append(f"  设计流量 Q = {Q:.3f} m³/s")
        output.append(f"  糙率 n = {n}")
        output.append(f"  水力坡降 i = 1/{int(slope_inv)} = {i:.6f}")
        output.append("")
        
        if "U形" in section_type:
            R = result.get('R', 0)
            f = result.get('f', 0)
            B = result.get('B', 0)
            f_R = result.get('f_R', 0)
            H_total = result.get('H_total', 0)
            output.append("【二、断面尺寸】")
            output.append(f"  内半径 R = {R:.2f} m")
            output.append("")
            output.append("  1. 槽宽计算:")
            output.append(f"     B = 2 × R = 2 × {R:.2f} = {B:.2f} m")
            output.append("")
            output.append("  2. 直段高度:")
            output.append(f"     f = {f:.2f} m")
            output.append(f"     f/R = {f:.2f} / {R:.2f} = {f_R:.3f}")
            output.append("")
            output.append("  3. 槽身总高计算:")
            output.append(f"     H = R + f = {R:.2f} + {f:.2f} = {H_total:.2f} m")
            output.append("")
            output.append("  4. H/B比值计算:")
            H_B_ratio = H_total / B if B > 0 else 0
            output.append(f"     H/B = 槽身总高 ÷ 槽宽 = {H_total:.2f} ÷ {B:.2f} = {H_B_ratio:.3f}")
        else:
            B = result.get('B', 0)
            H_total = result.get('H_total', 0)
            ratio = result.get('depth_width_ratio', 0)
            output.append("【二、断面尺寸】")
            output.append(f"  槽宽 B = {B:.2f} m")
            output.append(f"  深宽比 = {ratio:.3f}")
            output.append("")
            output.append("  1. 槽高计算:")
            output.append(f"     H = B × 深宽比 = {B:.2f} × {ratio:.3f} = {H_total:.2f} m")
            output.append("")
            output.append("  2. H/B比值计算:")
            H_B_ratio = H_total / B if B > 0 else 0
            output.append(f"     H/B = 槽身总高 ÷ 槽宽 = {H_total:.2f} ÷ {B:.2f} = {H_B_ratio:.3f}")
            if result.get('has_chamfer', False):
                output.append("")
                output.append(f"  3. 倒角参数: 角度 {result.get('chamfer_angle', 0):.1f}°, 底边 {result.get('chamfer_length', 0):.2f} m")
        
        output.append("")
        output.append("【三、设计流量工况】")
        h_design = result.get('h_design', 0)
        A_design = result.get('A_design', 0)
        P_design = result.get('P_design', 0)
        R_hyd = result.get('R_hyd_design', 0)
        V_design = result.get('V_design', 0)
        Q_calc = result.get('Q_calc', 0)
        
        output.append(f"  设计水深 h = {h_design:.3f} m")
        output.append("")
        
        if "U形" in section_type:
            R = result.get('R', 0)
            output.append("  2. 过水面积计算 (U形断面):")
            if h_design <= R:
                theta_val = math.acos((R - h_design) / R) if R > 0 else 0
                output.append(f"     当 h ≤ R 时: θ = arccos((R-h)/R) = {math.degrees(theta_val):.2f}°")
                output.append(f"     A = R² × (θ - sinθ×cosθ) = {A_design:.3f} m²")
            else:
                output.append(f"     当 h > R 时: A = πR²/2 + 2R×(h-R) = {A_design:.3f} m²")
            output.append("")
            output.append("  3. 湿周计算 (U形断面):")
            if h_design <= R:
                output.append(f"     P = 2Rθ = {P_design:.3f} m")
            else:
                output.append(f"     P = πR + 2×(h-R) = {P_design:.3f} m")
        else:
            output.append("  2. 过水面积计算 (矩形断面):")
            output.append(f"     A = B × h = {B:.2f} × {h_design:.3f} = {A_design:.3f} m²")
            output.append("")
            output.append("  3. 湿周计算 (矩形断面):")
            output.append(f"     P = B + 2×h = {B:.2f} + 2×{h_design:.3f} = {P_design:.3f} m")
            
        output.append("")
        output.append("  4. 水力半径计算:")
        output.append(f"     R = A / P = {A_design:.3f} / {P_design:.3f} = {R_hyd:.3f} m")
        output.append("")
        output.append("  5. 设计流速计算 (曼宁公式):")
        output.append(f"     V = (1/n) × R^(2/3) × i^(1/2) = (1/{n}) × {R_hyd:.3f}^(2/3) × {i:.6f}^(1/2) = {V_design:.3f} m/s")
        output.append("")
        output.append(f"  6. 流量校核 Q计算 = A × V = {A_design:.3f} × {V_design:.3f} = {Q_calc:.3f} m³/s")
        
        output.append("")
        output.append("【四、加大流量工况】")
        inc_pct = result.get('increase_percent', 0)
        Q_inc = result.get('Q_increased', 0)
        h_inc = result.get('h_increased', 0)
        V_inc = result.get('V_increased', 0)
        A_inc = result.get('A_increased', 0)
        P_inc = result.get('P_increased', 0)
        R_inc = result.get('R_hyd_increased', 0)
        Fb = result.get('Fb', 0)
        Q_calc_inc = (1/n) * A_inc * (R_inc ** (2/3)) * (i ** 0.5) if R_inc > 0 else 0

        output.append(f"  1. 加大比例 = {inc_pct:.1f}%")
        output.append(f"  2. 加大流量 Q加大 = Q × (1 + {inc_pct/100:.2f}) = {Q_inc:.3f} m³/s")
        output.append(f"  3. 加大水深 h加大 = {h_inc:.3f} m")
        output.append("")

        if "U形" in section_type:
            R_val = result.get('R', 0)
            output.append("  4. 过水面积计算 (U形断面):")
            if h_inc <= R_val:
                theta_rad = 2 * math.acos((R_val - h_inc) / R_val)
                theta_deg = math.degrees(theta_rad)
                output.append(f"     A加大 = (R²/2) × (θ - sinθ)")
                output.append(f"          = ({R_val:.2f}²/2) × ({theta_deg:.2f}° - sin{theta_deg:.2f}°)")
                output.append(f"          = {A_inc:.3f} m²")
            else:
                output.append(f"     A加大 = πR²/2 + 2R×(h加大-R)")
                output.append(f"          = π×{R_val:.2f}²/2 + 2×{R_val:.2f}×({h_inc:.3f}-{R_val:.2f})")
                output.append(f"          = {math.pi*R_val**2/2:.3f} + {2*R_val*(h_inc-R_val):.3f}")
                output.append(f"          = {A_inc:.3f} m²")
            output.append("")
            output.append("  5. 湿周计算 (U形断面):")
            if h_inc <= R_val:
                theta_rad = 2 * math.acos((R_val - h_inc) / R_val)
                theta_deg = math.degrees(theta_rad)
                output.append(f"     P加大 = R × θ×π/180")
                output.append(f"          = {R_val:.2f} × {theta_deg:.2f}°×π/180")
                output.append(f"          = {P_inc:.3f} m")
            else:
                output.append(f"     P加大 = πR + 2×(h加大-R)")
                output.append(f"          = π×{R_val:.2f} + 2×({h_inc:.3f}-{R_val:.2f})")
                output.append(f"          = {math.pi*R_val:.3f} + {2*(h_inc-R_val):.3f}")
                output.append(f"          = {P_inc:.3f} m")
            output.append("")
        else:
            # 矩形断面
            B_val = result.get('B', 0)
            output.append("  4. 过水面积计算 (矩形断面):")
            output.append(f"     A加大 = B × h加大")
            output.append(f"          = {B_val:.2f} × {h_inc:.3f}")
            output.append(f"          = {A_inc:.3f} m²")
            output.append("")
            output.append("  5. 湿周计算 (矩形断面):")
            output.append(f"     P加大 = B + 2×h加大")
            output.append(f"          = {B_val:.2f} + 2×{h_inc:.3f}")
            output.append(f"          = {B_val:.2f} + {2*h_inc:.3f}")
            output.append(f"          = {P_inc:.3f} m")
            output.append("")

        output.append("  6. 水力半径计算:")
        output.append(f"     R加大 = A加大 / P加大 = {A_inc:.3f} / {P_inc:.3f} = {R_inc:.3f} m")
        output.append("")

        output.append("  7. 加大流速计算 (曼宁公式):")
        output.append(f"     V加大 = (1/n) × R加大^(2/3) × i^(1/2)")
        output.append(f"          = (1/{n}) × {R_inc:.3f}^(2/3) × {i:.6f}^(1/2) = {V_inc:.3f} m/s")
        output.append("")

        output.append("  8. 流量校核:")
        output.append(f"     Q计算 = A加大 × V加大 = {A_inc:.3f} × {V_inc:.3f} = {Q_calc_inc:.3f} m³/s")
        output.append("")

        output.append(f"  9. 超高 Fb = H - h加大 = {H_total:.2f} - {h_inc:.3f} = {Fb:.3f} m")

        
        output.append("")
        output.append("【五、验证】")
        
        # 1. 流速验证（规范 9.4.1-1）
        v_recommended_min = 1.0  # m/s
        v_recommended_max = 2.5  # m/s
        velocity_ok = v_recommended_min <= V_design <= v_recommended_max
        output.append(f"  1. 流速验证（规范 9.4.1-1）")
        output.append(f"     规范要求: 槽内设计流速宜为 1.0～2.5 m/s")
        output.append(f"     计算结果: V = {V_design:.3f} m/s")
        if velocity_ok:
            output.append(f"     验证结果: {v_recommended_min} ≤ {V_design:.3f} ≤ {v_recommended_max} → 通过 ✓")
        else:
            if V_design < v_recommended_min:
                output.append(f"     验证结果: {V_design:.3f} < {v_recommended_min} → 超出推荐范围 ⚠")
                output.append(f"     提示: 流速过小，可能造成泾积，建议调整断面尺寸")
            else:
                output.append(f"     验证结果: {V_design:.3f} > {v_recommended_max} → 超出推荐范围 ⚠")
                output.append(f"     提示: 流速过大，可能造成冲刷，建议调整断面尺寸")
        output.append("")
        
        # 2. 超高验证（规范 9.4.1-2）
        output.append(f"  2. 超高验证（规范 9.4.1-2）")
        
        if "U形" in section_type:
            # U形断面超高验证
            R_val = result.get('R', 0)
            Fb_design = H_total - h_design  # 设计流量超高
            Fb_design_min = R_val / 5  # 设计流量时最小超高：槽身直径的1/10 = R/5
            Fb_inc_min = 0.10  # 加大流量时最小超高
            
            output.append(f"     断面类型: U形")
            output.append(f"     规范要求:")
            output.append(f"       - 设计流量: 超高不应小于槽身直径的1/10 (即2R/10 = R/5 = {Fb_design_min:.3f} m)")
            output.append(f"       - 加大流量: 超高不应小于 0.10 m")
            output.append(f"")
            output.append(f"     计算结果:")
            output.append(f"       - 设计流量超高: Fb_设计 = H - h_设计 = {H_total:.2f} - {h_design:.3f} = {Fb_design:.3f} m")
            output.append(f"       - 加大流量超高: Fb_加大 = H - h_加大 = {H_total:.2f} - {h_inc:.3f} = {Fb:.3f} m")
            output.append(f"")
            
            fb_design_ok = Fb_design >= Fb_design_min
            fb_inc_ok = Fb >= Fb_inc_min
            
            output.append(f"     验证结果:")
            output.append(f"       - 设计流量: {Fb_design:.3f} {'\u2265' if fb_design_ok else '<'} {Fb_design_min:.3f} → {'\u901a\u8fc7 \u2713' if fb_design_ok else '\u672a\u901a\u8fc7 \u2717'}")
            output.append(f"       - 加大流量: {Fb:.3f} {'\u2265' if fb_inc_ok else '<'} {Fb_inc_min:.2f} → {'\u901a\u8fc7 \u2713' if fb_inc_ok else '\u672a\u901a\u8fc7 \u2717'}")
        else:
            # 矩形断面超高验证
            Fb_design = H_total - h_design  # 设计流量超高
            Fb_design_min = h_design / 12 + 0.05  # 设计流量时最小超高
            Fb_inc_min = 0.10  # 加大流量时最小超高
            
            output.append(f"     断面类型: 矩形")
            output.append(f"     规范要求:")
            output.append(f"       - 设计流量: 超高不应小于 h/12 + 0.05 = {h_design:.3f}/12 + 0.05 = {Fb_design_min:.3f} m")
            output.append(f"       - 加大流量: 超高不应小于 0.10 m")
            output.append(f"")
            output.append(f"     计算结果:")
            output.append(f"       - 设计流量超高: Fb_设计 = H - h_设计 = {H_total:.2f} - {h_design:.3f} = {Fb_design:.3f} m")
            output.append(f"       - 加大流量超高: Fb_加大 = H - h_加大 = {H_total:.2f} - {h_inc:.3f} = {Fb:.3f} m")
            output.append(f"")
            
            fb_design_ok = Fb_design >= Fb_design_min
            fb_inc_ok = Fb >= Fb_inc_min
            
            output.append(f"     验证结果:")
            output.append(f"       - 设计流量: {Fb_design:.3f} {'\u2265' if fb_design_ok else '<'} {Fb_design_min:.3f} → {'\u901a\u8fc7 \u2713' if fb_design_ok else '\u672a\u901a\u8fc7 \u2717'}")
            output.append(f"       - 加大流量: {Fb:.3f} {'\u2265' if fb_inc_ok else '<'} {Fb_inc_min:.3f} → {'\u901a\u8fc7 \u2713' if fb_inc_ok else '\u672a\u901a\u8fc7 \u2717'}")
        
        return "\n".join(output)


    def _format_suidong_report(self, input_vals, result) -> str:
        """格式化隧洞/矩形暗涵报告"""
        Q = float(input_vals[6])       # Q在索引6（新增X、Y后）
        n = float(input_vals[7])       # 糙率n在索引7
        slope_inv = float(input_vals[8])  # 比降在索引8
        v_min = float(input_vals[18])  # 不淤流速在索引18
        v_max = float(input_vals[19])  # 不冲流速在索引19
        section_type = input_vals[3]   # 结构形式在索引3
        i = 1.0 / slope_inv
        
        output = []
        output.append("【一、输入参数】")
        output.append(f"  设计流量 Q = {Q:.3f} m³/s")
        output.append(f"  糙率 n = {n}")
        output.append(f"  水力坡降 1/{int(slope_inv)}")
        output.append(f"  不淤流速 = {v_min} m/s")
        output.append(f"  不冲流速 = {v_max} m/s")
        output.append("")
        
        # 提取关键数据
        h_design = result.get('h_design', 0)
        A_design = result.get('A_design', 0)
        P_design = result.get('P_design', 0)
        R_hyd = result.get('R_hyd_design', 0)
        V_design = result.get('V_design', 0)
        Q_calc = result.get('Q_calc', 0)
        fb_pct_design = result.get('freeboard_pct_design', 0)
        fb_hgt_design = result.get('freeboard_hgt_design', 0)
        A_total = result.get('A_total', 0)

        output.append("【二、断面尺寸】")
        if "圆形" in section_type:
            D = result.get('D', 0)
            output.append(f"  1. 设计直径: D = {D:.2f} m")
            output.append(f"  2. 断面总面积: A总 = π × D² / 4 = {PI:.4f} × {D:.2f}² / 4 = {A_total:.3f} m²")
        elif "圆拱直墙型" in section_type:
            B = result.get('B', 0)
            H = result.get('H_total', 0)
            theta_deg = result.get('theta_deg', 0)
            output.append(f"  1. 设计宽度: B = {B:.2f} m")
            output.append(f"  2. 设计高度: H = {H:.2f} m")
            output.append(f"  3. 拱顶圆心角: θ = {theta_deg:.1f}°")
            output.append(f"  4. 高宽比: H/B = {H:.2f} / {B:.2f} = {H/B:.3f}" if B > 0 else "  4. 高宽比: N/A")
            output.append(f"  5. 断面总面积: A总 = {A_total:.3f} m²")
        elif "马蹄形" in section_type:
            r = result.get('r', 0)
            output.append(f"  1. 设计半径: R = {r:.2f} m")
            output.append(f"  2. 等效直径: 2R = {2*r:.2f} m")
            output.append(f"  3. 断面总面积: A总 = {A_total:.3f} m²")
        elif "矩形暗涵" in section_type or section_type == "矩形暗涵":
            B = result.get('B', 0)
            H = result.get('H', 0)
            BH_ratio = result.get('BH_ratio', 0)
            output.append(f"  1. 设计宽度: B = {B:.2f} m")
            output.append(f"  2. 设计高度: H = {H:.2f} m")
            output.append(f"  3. 宽深比 β = B / h_设计 = {B:.2f} / {h_design:.3f} = {BH_ratio:.3f}")
            output.append(f"  4. 总断面积: A总 = B × H = {B:.2f} × {H:.2f} = {A_total:.3f} m²")
        
        output.append("")
        output.append("【三、设计流量工况】")
        output.append(f"  1. 设计水深 h = {h_design:.3f} m")
        output.append(f"  2. 过水面积 A = {A_design:.3f} m²")
        output.append(f"  3. 湿周计算 χ = {P_design:.3f} m")
        output.append(f"  4. 水力半径 R = A / χ = {R_hyd:.3f} m")
        output.append(f"  5. 设计流速 V = (1/n) × R^(2/3) × i^(1/2) = (1/{n}) × {R_hyd:.3f}^(2/3) × {i:.6f}^(1/2) = {V_design:.3f} m/s")
        output.append(f"  6. 流量校核 Q计算 = A × V = {A_design:.3f} × {V_design:.3f} = {Q_calc:.3f} m³/s")
        output.append(f"  7. 净空高度 Fb = {fb_hgt_design:.3f} m")
        output.append(f"  8. 净空面积比例 = {fb_pct_design:.1f}%")
        
        output.append("")
        output.append("【四、加大流量工况】")
        inc_pct = result.get('increase_percent', 0)
        Q_inc = result.get('Q_increased', 0)
        h_inc = result.get('h_increased', 0)
        V_inc = result.get('V_increased', 0)
        A_inc = result.get('A_increased', 0)
        P_inc = result.get('P_increased', 0)
        R_inc = result.get('R_hyd_increased', 0)
        fb_pct_inc = result.get('freeboard_pct_inc', 0)
        fb_hgt_inc = result.get('freeboard_hgt_inc', 0)
        Q_calc_inc = (1/n) * A_inc * (R_inc ** (2/3)) * (i ** 0.5) if R_inc > 0 else 0

        # 提取用于展示公式的数据
        H_total_val = result.get('H_total', result.get('H', result.get('D', result.get('r', 0) * 2)))

        output.append(f"  9. 加大比例 = {inc_pct:.1f}%")
        output.append(f"  10. 加大流量 Q加大 = {Q_inc:.3f} m³/s")
        output.append(f"  11. 加大水深 h加大 = {h_inc:.3f} m")
        output.append("")

        output.append("  12. 过水面积计算:")
        output.append(f"     A加大 = {A_inc:.3f} m²")
        output.append("")

        output.append("  13. 湿周计算:")
        output.append(f"     P加大 = {P_inc:.3f} m")
        output.append("")

        output.append("  14. 水力半径计算:")
        output.append(f"     R加大 = A加大 / P加大 = {A_inc:.3f} / {P_inc:.3f} = {R_inc:.3f} m")
        output.append("")

        output.append("  15. 加大流速计算 (曼宁公式):")
        output.append(f"     V加大 = (1/n) × R加大^(2/3) × i^(1/2)")
        output.append(f"          = (1/{n}) × {R_inc:.3f}^(2/3) × {i:.6f}^(1/2) = {V_inc:.3f} m/s")
        output.append("")

        output.append("  16. 流量校核:")
        output.append(f"     Q计算 = A加大 × V加大 = {A_inc:.3f} × {V_inc:.3f} = {Q_calc_inc:.3f} m³/s")
        output.append("")

        output.append(f"  17. 净空高度 Fb加大 = H - h加大 = {H_total_val:.2f} - {h_inc:.3f} = {fb_hgt_inc:.3f} m")
        output.append(f"  18. 净空面积比例 = {fb_pct_inc:.1f}%")
        
        output.append("")
        output.append("【五、验证结论】")
        velocity_ok = v_min <= V_design <= v_max
        output.append(f"  1. 流速验证: {v_min} ≤ {V_design:.3f} ≤ {v_max} → {'通过 ✓' if velocity_ok else '未通过 ✗'}")
        
        # 隧洞/暗涵特有的净空验证
        min_fb_hgt = 0.4
        min_fb_pct = 15.0 if "矩形暗涵" not in section_type else 10.0
        fb_hgt_ok = fb_hgt_inc >= min_fb_hgt
        fb_pct_ok = fb_pct_inc >= min_fb_pct
        
        output.append(f"  2. 净空高度验证: Fb加大 = {fb_hgt_inc:.3f}m ≥ {min_fb_hgt}m → {'通过 ✓' if fb_hgt_ok else '需注意'}")
        output.append(f"  3. 净空比例验证: {fb_pct_inc:.1f}% ≥ {min_fb_pct}% → {'通过 ✓' if fb_pct_ok else '需注意'}")
        
        return "\n".join(output)


    def _extract_result_data(self, seq: str, segment: str, building_name: str, 
                            section_type: str, result: Dict[str, Any]) -> tuple:
        """从计算结果中提取关键数据（已删除桩号参数）"""
        
        # 格式化输出辅助函数
        def fmt(val, default="-"):
            if val is None or val == 0:
                return default
            if isinstance(val, (int, float)):
                return f"{val:.3f}"
            return str(val)
        
        def fmt_pct(val, default="-"):
            """格式化百分比"""
            if val is None or val == 0:
                return default
            if isinstance(val, (int, float)):
                return f"{val:.1f}"
            return str(val)
        
        # 初始化默认值
        B_val = "-"   # 底宽B（明渠梯形、明渠矩形、矩形暗涵、隧洞圆拱直墙型、渡槽矩形）
        D_val = "-"   # 直径D（明渠圆形、隧洞圆形）
        R_val = "-"   # 半径（渡槽-U形、隧洞-马蹄形Ⅰ型、隧洞-马蹄形Ⅱ型）
        h_design = "-"
        V_design = "-"
        A_design = "-"
        R_hyd = "-"
        chi = "-"
        Q_inc = "-"
        h_inc = "-"
        V_inc = "-"
        # 超高Fb（用于明渠/渡槽）、设计净空高度、加大净空高度（用于隧洞/暗渠）
        Fb_surcharge = "-"  # 超高，用于明渠和渡槽
        Fb_clearance_design = "-"  # 设计净空高度，用于隧洞和暗渠
        Fb_clearance_inc = "-"  # 加大净空高度，用于隧洞和暗渠
        Fb_pct_design = "-"
        Fb_pct_inc = "-"
        
        # 根据断面类型提取不同的字段
        if "明渠" in section_type:
            # 圆形明渠用直径D，其他用底宽B
            if "圆形" in section_type:
                D_val = result.get('D_design', 0) if 'D_design' in result else result.get('D', 0)
            else:
                B_val = result.get('b_design', 0) if 'b_design' in result else result.get('B', 0)
            h_design = result.get('h_design', 0) if 'h_design' in result else result.get('y_d', 0)
            V_design = result.get('V_design', 0) if 'V_design' in result else result.get('V_d', 0)
            A_design = result.get('A_design', 0) if 'A_design' in result else result.get('A_d', 0)
            R_hyd = result.get('R_design', 0) if 'R_design' in result else result.get('R_d', 0)
            chi = result.get('X_design', 0) if 'X_design' in result else result.get('P_d', 0)
            Q_inc = result.get('Q_increased', 0) if 'Q_increased' in result else result.get('Q_inc', 0)
            h_inc = result.get('h_increased', 0) if 'h_increased' in result else result.get('y_i', 0)
            V_inc = result.get('V_increased', 0) if 'V_increased' in result else result.get('V_i', 0)
            
            # 圆形明渠：超高/净空高度特殊处理
            if "圆形" in section_type:
                # 圆形明渠显示设计净空高度和加大净空高度
                Fb_surcharge = result.get('FB_d', 0)  # 设计工况净空高度
                Fb_clearance_design = result.get('FB_d', 0)  # 设计净空高度
                Fb_clearance_inc = result.get('FB_i', 0)  # 加大净空高度
                # 圆形明渠额外显示净空比例
                Fb_pct_design = result.get('PA_d', "-")
                Fb_pct_inc = result.get('PA_i', "-")
            else:
                # 梯形/矩形明渠：显示超高
                Fb_surcharge = result.get('Fb', 0) if 'Fb' in result else result.get('FB_d', 0)
                Fb_clearance_design = "-"  # 非圆形明渠不显示净空高度
                Fb_clearance_inc = "-"
                Fb_pct_design = "-"
                Fb_pct_inc = "-"
            
        elif "渡槽" in section_type:
            # U形渡槽显示半径，矩形渡槽显示底宽B
            if "U形" in section_type:
                R_val = result.get('R', 0)  # 在半径列显示
            else:
                B_val = result.get('B', 0) if 'B' in result else result.get('b_design', 0)
            h_design = result.get('h_design', 0)
            V_design = result.get('V_design', 0)
            A_design = result.get('A_design', 0)
            R_hyd = result.get('R_hyd_design', 0) if 'R_hyd_design' in result else result.get('R_hyd', 0)
            chi = result.get('P_design', 0) if 'P_design' in result else result.get('chi', 0)
            Q_inc = result.get('Q_increased', 0) if 'Q_increased' in result else result.get('Q_inc', 0)
            h_inc = result.get('h_increased', 0) if 'h_increased' in result else result.get('y_i', 0)
            V_inc = result.get('V_increased', 0) if 'V_increased' in result else result.get('V_i', 0)
            Fb_surcharge = result.get('Fb', 0) if 'Fb' in result else result.get('FB_d', 0)  # 渡槽超高
            Fb_clearance_design = "-"  # 渡槽不显示净空高度
            Fb_clearance_inc = "-"
            Fb_pct_design = "-"  # 渡槽不计算净空比例
            Fb_pct_inc = "-"
            
        elif "隧洞" in section_type:
            # 隧洞类型：圆形用直径D，圆拱直墙型用底宽B，马蹄形用半径
            if "圆形" in section_type:
                D_val = result.get('D_design', 0) if 'D_design' in result else result.get('D', 0)
            elif "圆拱直墙型" in section_type:
                B_val = result.get('B', 0) if 'B' in result else result.get('b_design', 0)
            elif "马蹄形" in section_type:
                R_val = result.get('r', 0) if 'r' in result else result.get('R_design', 0)
            else:
                D_val = result.get('D_design', 0) if 'D_design' in result else result.get('B', 0)
            
            h_design = result.get('y_d', 0) if 'y_d' in result else result.get('h_design', 0)
            V_design = result.get('V_d', 0) if 'V_d' in result else result.get('V_design', 0)
            A_design = result.get('A_d', 0) if 'A_d' in result else result.get('A_design', 0)
            R_hyd = result.get('R_d', 0) if 'R_d' in result else result.get('R_hyd_design', 0)
            chi = result.get('P_d', 0) if 'P_d' in result else result.get('P_design', 0)
            Q_inc = result.get('Q_inc', 0) if 'Q_inc' in result else result.get('Q_increased', 0)
            h_inc = result.get('y_i', 0) if 'y_i' in result else result.get('h_increased', 0)
            V_inc = result.get('V_i', 0) if 'V_i' in result else result.get('V_increased', 0)
            # 隧洞：不显示超高，显示设计和加大工况的净空高度和比例
            Fb_surcharge = "-"  # 隧洞不显示超高
            Fb_clearance_design = result.get('freeboard_hgt_design', 0) if 'freeboard_hgt_design' in result else \
                        result.get('FB_d', 0) if 'FB_d' in result else 0
            Fb_clearance_inc = result.get('freeboard_hgt_inc', 0) if 'freeboard_hgt_inc' in result else \
                     result.get('FB_i', 0) if 'FB_i' in result else 0
            Fb_pct_design = result.get('freeboard_pct_design', 0) if 'freeboard_pct_design' in result else \
                            result.get('FB_pct_d', 0) if 'FB_pct_d' in result else 0
            Fb_pct_inc = result.get('freeboard_pct_inc', 0) if 'freeboard_pct_inc' in result else \
                         result.get('FB_pct_i', 0) if 'FB_pct_i' in result else 0
        
        elif section_type == "矩形暗涵":
            # 矩形暗涵作为独立类型，使用底宽B
            B_val = result.get('B', 0) if 'B' in result else result.get('b_design', 0)
            h_design = result.get('h_design', 0)
            V_design = result.get('V_design', 0)
            A_design = result.get('A_design', 0)
            R_hyd = result.get('R_hyd_design', 0)
            chi = result.get('P_design', 0)
            Q_inc = result.get('Q_increased', 0)
            h_inc = result.get('h_increased', 0)
            V_inc = result.get('V_increased', 0)
            # 矩形暗涵：不显示超高，显示设计和加大工况的净空高度和比例
            Fb_surcharge = "-"  # 矩形暗涵不显示超高
            Fb_clearance_design = result.get('freeboard_hgt_design', 0)
            Fb_clearance_inc = result.get('freeboard_hgt_inc', 0)
            Fb_pct_design = result.get('freeboard_pct_design', 0)
            Fb_pct_inc = result.get('freeboard_pct_inc', 0)
        
        return (seq, segment, building_name, section_type, fmt(B_val), fmt(D_val), fmt(R_val), fmt(h_design), fmt(V_design), 
                fmt(A_design), fmt(R_hyd), fmt(chi), fmt(Q_inc), fmt(h_inc), 
                fmt(V_inc), fmt(Fb_surcharge), fmt(Fb_clearance_design), fmt(Fb_clearance_inc), 
                fmt_pct(Fb_pct_design), fmt_pct(Fb_pct_inc), "✓ 成功")
    
    def _clear_results(self):
        """清空结果表格"""
        self._set_result_sheet_data([], redraw=True, reset_row_positions=True)
        self.batch_results = []
        
        # 同时清空详细过程文本框
        if hasattr(self, 'detail_text'):
            self.detail_text.configure(state=tk.NORMAL)
            self.detail_text.delete(1.0, tk.END)
            self.detail_text.configure(state=tk.DISABLED)
        
        # 同时清除共享数据管理器中的批量结果
        if SHARED_DATA_AVAILABLE:
            try:
                get_shared_data_manager().clear_batch_results()
            except Exception:
                pass
    
    def _auto_fit_columns(self, tree):
        """自动调整Treeview表格列宽以完整显示表头和内容"""
        # 获取所有列
        columns = tree['columns']
        
        # 使用tkinter font来更精确计算文字宽度
        try:
            import tkinter.font as tkfont
            # 获取Treeview实际使用的字体
            style = ttk.Style()
            tree_font_config = style.lookup("Treeview.Heading", "font")
            if tree_font_config:
                heading_font = tkfont.Font(font=tree_font_config)
            else:
                heading_font = tkfont.nametofont("TkDefaultFont")
            default_font = tkfont.nametofont("TkDefaultFont")
        except:
            heading_font = None
            default_font = None
        
        # 计算每列需要的最小宽度
        for col in columns:
            # 获取表头文字
            header_text = tree.heading(col)['text']
            
            # 计算表头宽度 - 使用更大的填充值确保表头完整显示
            if heading_font:
                # 使用实际表头字体测量，并增加额外填充
                header_width = heading_font.measure(header_text) + 50  # 增加填充到50px
            else:
                # 回退方案：中文字符按2.5个英文字符宽度计算（更宽松）
                chinese_chars = sum(1 for c in header_text if ord(c) > 127)
                english_chars = len(header_text) - chinese_chars
                header_width = (chinese_chars * 18 + english_chars * 8) + 50
            
            max_width = header_width
            
            # 遍历所有行数据，找出最大内容宽度
            for item in tree.get_children():
                values = tree.item(item)['values']
                col_idx = list(columns).index(col)
                if col_idx < len(values):
                    cell_text = str(values[col_idx]) if values[col_idx] else ""
                    if default_font:
                        cell_width = default_font.measure(cell_text) + 20
                    else:
                        chinese_chars = sum(1 for c in cell_text if ord(c) > 127)
                        english_chars = len(cell_text) - chinese_chars
                        cell_width = (chinese_chars * 16 + english_chars * 8) + 20
                    max_width = max(max_width, cell_width)
            
            # 设置最小和最大宽度限制 - 提高最小宽度
            max_width = max(80, min(max_width, 500))
            # 设置列宽，同时设置 stretch=False 以支持水平滚动，minwidth 确保不会被压缩
            tree.column(col, width=max_width, minwidth=max_width, stretch=False)
        
        # 强制更新布局以确保滚动条正确显示
        tree.update_idletasks()
    
    def _auto_fit_input_columns(self):
        """自动调整输入表格列宽"""
        if getattr(self, "_freeze_enabled", False) and getattr(self, "frozen_sheet", None) is not None:
            try:
                self.frozen_sheet.set_all_column_widths(only_set_if_too_small=True, redraw=False)
            except Exception:
                pass
            try:
                self.input_sheet.set_all_column_widths(only_set_if_too_small=True, redraw=False)
            except Exception:
                pass
            try:
                self.input_sheet.set_xview(0.0)
            except Exception:
                pass
            try:
                self.frozen_sheet.redraw()
            except Exception:
                pass
            try:
                self.input_sheet.redraw()
            except Exception:
                pass
            return

        try:
            self.input_sheet.set_all_column_widths(only_set_if_too_small=True, redraw=False)
        except Exception:
            pass
        try:
            self.input_sheet.set_xview(0.0)
        except Exception:
            pass
        try:
            self.input_sheet.redraw()
        except Exception:
            try:
                self.input_sheet.refresh()
            except Exception:
                pass
    
    def _auto_fit_result_columns(self):
        """自动调整结果表格列宽"""
        try:
            self.result_sheet.set_all_column_widths(only_set_if_too_small=True, redraw=False)
        except Exception:
            pass
        try:
            self.result_sheet.set_xview(0.0)
        except Exception:
            pass
        try:
            self.result_sheet.redraw()
        except Exception:
            try:
                self.result_sheet.refresh()
            except Exception:
                pass
    
    def _export_to_excel(self):
        """导出结果到Excel"""
        result_rows = self._get_result_sheet_data()
        if not self.batch_results and not result_rows:
            messagebox.showwarning("警告", "没有可导出的结果")
            return
        
        # 验证4个必填字段是否已输入
        missing_fields = []
        
        channel_name = self.var_channel_name.get().strip()
        if not channel_name:
            missing_fields.append("渠道名称")
        
        channel_level = self.var_channel_level.get().strip()
        if not channel_level:
            missing_fields.append("渠道类型")
        
        start_station = self.var_start_station.get().strip()
        if not start_station:
            missing_fields.append("起始桩号")
        
        start_water_level = self.var_start_water_level.get().strip()
        if not start_water_level:
            missing_fields.append("渠道起始水位")
        
        if missing_fields:
            fields_str = "、".join(missing_fields)
            messagebox.showwarning("警告", f"以下必填项尚未填写，请补充后再导出Excel：\n\n  {fields_str}")
            return
        
        try:
            # 尝试导入openpyxl
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            messagebox.showerror("错误", "需要安装openpyxl库才能导出Excel\n请运行: pip install openpyxl")
            return
        
        # 选择保存路径
        filepath = filedialog.asksaveasfilename(
            title="保存Excel报告",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not filepath:
            return
        
        try:
            # 从输入表获取参数数据，建立映射（以序号为key，确保唯一匹配）
            # 列索引（新增X、Y列）: 0序号, 1流量段, 2建筑物名称, 3结构形式, 4X, 5Y, 6Q, 7糙率n, 8比降, 9边坡系数m, 10底宽B...
            input_params_map = {}  # key: 序号, value: (X, Y, Q, 糙率, 比降, 边坡系数)
            for input_values in self._get_input_sheet_data():
                if len(input_values) >= 12:
                    seq = str(input_values[0]).strip()
                    x_coord = input_values[4] if input_values[4] else ""
                    y_coord = input_values[5] if input_values[5] else ""
                    q_value = input_values[6] if input_values[6] else ""
                    roughness = input_values[7] if input_values[7] else ""
                    slope = input_values[8] if input_values[8] else ""
                    side_slope = input_values[9] if input_values[9] else ""
                    input_params_map[seq] = (x_coord, y_coord, q_value, roughness, slope, side_slope)
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "批量计算结果"
            
            # 写入标题
            ws['A1'] = "渠系建筑物多流量段批量水力计算系统结果报告"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:Z1')  # 扩展到Z列(26列，新增X、Y列)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # 写入第2行基础信息：渠道名称、渠道类型、起始水位、起始桩号
            # 获取基础信息
            channel_name = self.var_channel_name.get().strip()
            channel_level = self.var_channel_level.get().strip()
            start_station = self.var_start_station.get().strip()
            
            # A2-B2: 渠道名称
            ws['A2'] = "渠道名称"
            ws['A2'].font = Font(bold=True)
            ws['B2'] = channel_name if channel_name else ""
            
            # C2-D2: 渠道类型
            ws['C2'] = "渠道类型"
            ws['C2'].font = Font(bold=True)
            ws['D2'] = channel_level if channel_level else ""
            
            # E2-F2: 渠道起始水位
            ws['E2'] = "起始水位(m)"
            ws['E2'].font = Font(bold=True)
            ws['F2'] = start_water_level if start_water_level else ""
            
            # G2-H2: 起始桩号
            ws['G2'] = "起始桩号"
            ws['G2'].font = Font(bold=True)
            ws['H2'] = start_station if start_station else ""
            
            # 写入表头 - 新增X、Y列，添加Q(m³/s)、糙率n、比降(1/)、边坡系数m列（第3行）
            headers = ["序号", "流量段", "建筑物名称", "结构形式", "X", "Y",
                      "Q(m³/s)", "糙率n", "比降(1/)", "边坡系数m",  # 新增4列
                      "底宽B(m)", "直径D(m)", "半径R(m)", "h设计(m)", "V设计(m/s)", 
                      "A设计(m²)", "R水力(m)", "湿周χ(m)", "Q加大(m³/s)", "h加大(m)", 
                      "V加大(m/s)", "超高Fb(m)", "设计净空高度(m)", "加大净空高度(m)", "设计净空比例(%)", "加大净空比例(%)", "状态"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # 写入数据 - 从第4行开始（第1行标题，第2行起始水位，第3行表头）
            row_num = 4
            for result_values in result_rows:
                result_values = list(result_values)
                
                # 获取序号用于匹配
                seq = str(result_values[0]).strip() if len(result_values) > 0 else ""
                
                # 从映射获取X、Y、Q、糙率、比降、边坡系数（以序号为key）
                x_coord, y_coord, q_value, roughness, slope, side_slope = input_params_map.get(seq, ("", "", "", "", "", ""))
                
                # 构建新的数据行：在结构形式(索引3)之后插入X、Y、Q、糙率、比降、边坡系数
                # 原result_values: [序号, 流量段, 建筑物名称, 结构形式, 底宽B, 直径D, 半径, h设计, V设计, ...]
                # 新输出顺序: [序号, 流量段, 建筑物名称, 结构形式, X, Y, Q, 糙率, 比降, 边坡系数, 底宽B, 直径D, 半径, h设计, ...]
                new_values = result_values[:4] + [x_coord, y_coord, q_value, roughness, slope, side_slope] + result_values[4:]
                
                for col_num, value in enumerate(new_values, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                row_num += 1
            
            # 调整列宽 - 使用get_column_letter避免MergedCell问题
            from openpyxl.utils import get_column_letter
            for col_num in range(1, len(headers) + 1):
                max_length = len(str(headers[col_num - 1]))
                col_letter = get_column_letter(col_num)
                # 检查数据行的最大长度
                for row_num in range(4, ws.max_row + 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # 保存文件
            wb.save(filepath)
            messagebox.showinfo("成功", f"Excel报告已导出到:\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")


# ============================================================
# 断面参数输入弹窗
# ============================================================

class SectionParameterDialog(tk.Toplevel):
    """断面参数输入弹窗 - 根据断面类型显示不同的参数输入界面"""
    
    def __init__(self, parent, section_type: str, current_values: dict):
        """
        初始化弹窗
        
        参数:
            parent: 父窗口
            section_type: 断面类型 (如 "明渠-梯形", "渡槽-U形" 等)
            current_values: 当前行的参数值字典
        """
        super().__init__(parent)
        
        self.section_type = section_type
        self.current_values = current_values
        self.result = None  # 存储返回结果
        
        # 设置窗口属性
        self.title(f"参数设置 - {section_type}")
        self.transient(parent)  # 设置为父窗口的临时窗口
        self.grab_set()  # 模态对话框
        
        # 创建变量
        self._create_variables()
        
        # 创建UI
        self._create_ui()
        
        # 加载当前值
        self._load_current_values()
        
        # 居中显示 - 相对于父窗口居中
        self.update_idletasks()
        width = 450
        height = 550
        
        # 确保父窗口已经渲染完成
        parent.update_idletasks()
        
        # 获取父窗口的位置和尺寸
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        # 如果父窗口尺寸为0，使用屏幕尺寸
        if parent_width <= 1 or parent_height <= 1:
            parent_x = 0
            parent_y = 0
            parent_width = self.winfo_screenwidth()
            parent_height = self.winfo_screenheight()
        
        # 计算弹窗位置，使其在父窗口中居中
        x = parent_x + (parent_width // 2) - (width // 2)
        y = parent_y + (parent_height // 2) - (height // 2)
        # 确保弹窗不超出屏幕边界
        x = max(0, min(x, self.winfo_screenwidth() - width))
        y = max(0, min(y, self.winfo_screenheight() - height))
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # 绑定快捷键
        self.bind('<Return>', lambda e: self._on_confirm())
        self.bind('<Escape>', lambda e: self._on_cancel())
        
        # 设置焦点
        self.Q_entry.focus_set()
    
    def _create_variables(self):
        """创建输入变量"""
        # 通用参数
        self.Q_var = tk.DoubleVar(value=5.0)
        self.n_var = tk.DoubleVar(value=0.014)
        self.slope_inv_var = tk.DoubleVar(value=3000)
        self.v_min_var = tk.DoubleVar(value=0.1)
        self.v_max_var = tk.DoubleVar(value=100.0)
        
        # 明渠参数
        self.m_var = tk.StringVar(value="")  # 边坡系数
        self.b_var = tk.StringVar(value="")  # 底宽
        self.b_h_ratio_var = tk.StringVar(value="")  # 宽深比
        
        # 渡槽-U形参数
        self.R_var = tk.StringVar(value="")  # 内半径
        
        # 渡槽-矩形参数
        self.chamfer_angle_var = tk.StringVar(value="")  # 倒角角度
        self.chamfer_length_var = tk.StringVar(value="")  # 倒角底边
        self.h_b_ratio_var = tk.StringVar(value="")  # 深宽比(矩形渡槽)
        
        # 隧洞参数
        self.D_var = tk.StringVar(value="")  # 直径
        self.B_var = tk.StringVar(value="")  # 底宽
        self.theta_var = tk.StringVar(value="")  # 圆心角
        self.r_var = tk.StringVar(value="")  # 马蹄形半径
        self.B_rect_var = tk.StringVar(value="")  # 矩形暗涵底宽
        self.H_rect_var = tk.StringVar(value="")  # 矩形暗涵槽高
        self.BH_ratio_rect_var = tk.StringVar(value="")  # 矩形暗涵宽深比
    
    def _create_ui(self):
        """创建UI界面"""
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 基础参数区
        self._create_basic_inputs(main_frame)
        
        # 流速参数区
        self._create_velocity_inputs(main_frame)
        
        # 可选参数区（根据断面类型）
        self._create_optional_inputs(main_frame)
        
        # 按钮区
        self._create_buttons(main_frame)
    
    def _create_basic_inputs(self, parent):
        """创建基础参数输入区"""
        frame = ttk.LabelFrame(parent, text="基础参数", padding="10")
        frame.pack(fill=tk.X, pady=(0, 10))
        
        row = 0
        # 设计流量
        ttk.Label(frame, text="设计流量 Q (m³/s):").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.Q_entry = ttk.Entry(frame, textvariable=self.Q_var, width=18)
        self.Q_entry.grid(row=row, column=1, padx=5, pady=5)
        
        # 糙率
        row += 1
        ttk.Label(frame, text="糙率 n:").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.n_var, width=18).grid(row=row, column=1, padx=5, pady=5)
        
        # 水力坡降
        row += 1
        ttk.Label(frame, text="水力坡降 1/").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.slope_inv_var, width=18).grid(row=row, column=1, padx=5, pady=5)
    
    def _create_velocity_inputs(self, parent):
        """创建流速参数输入区"""
        frame = ttk.LabelFrame(parent, text="流速参数", padding="10")
        frame.pack(fill=tk.X, pady=(0, 10))
        
        row = 0
        # 不淤流速
        ttk.Label(frame, text="不淤流速 (m/s):").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.v_min_var, width=18).grid(row=row, column=1, padx=5, pady=5)
        
        # 不冲流速
        row += 1
        ttk.Label(frame, text="不冲流速 (m/s):").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.v_max_var, width=18).grid(row=row, column=1, padx=5, pady=5)
        
        # 提示
        row += 1
        ttk.Label(frame, text="(一般情况下保持默认数值即可)", font=('', 8), 
                 foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
    
    def _create_optional_inputs(self, parent):
        """根据断面类型创建可选参数输入区"""
        frame = ttk.LabelFrame(parent, text="可选参数", padding="10")
        frame.pack(fill=tk.X, pady=(0, 10))
        
        row = 0
        
        # 根据断面类型显示不同控件
        if "明渠-梯形" in self.section_type:
            # 边坡系数
            ttk.Label(frame, text="边坡系数 m:").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.m_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            # 手动底宽
            ttk.Label(frame, text="手动底宽 B (m):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.b_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            # 宽深比
            ttk.Label(frame, text="宽深比 B/h:").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.b_h_ratio_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(梯形断面边坡系数必填; 底宽B和宽深比可选)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
            
        elif "明渠-矩形" in self.section_type:
            # 矩形断面 m=0
            ttk.Label(frame, text="边坡系数 m = 0 (矩形断面)", 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=5)
            self.m_var.set("0")
            row += 1
            # 手动底宽
            ttk.Label(frame, text="手动底宽 B (m):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.b_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            # 宽深比
            ttk.Label(frame, text="宽深比 B/h:").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.b_h_ratio_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(底宽B和宽深比可选，留空则自动计算)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
            
        elif "明渠-圆形" in self.section_type:
            # 手动直径
            ttk.Label(frame, text="手动直径 D (m):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.D_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(留空则自动计算)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
            
        elif "渡槽-U形" in self.section_type:
            # 手动内半径
            ttk.Label(frame, text="手动内半径 R (m):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.R_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(留空则自动计算)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
            
        elif "渡槽-矩形" in self.section_type:
            # 深宽比
            ttk.Label(frame, text="深宽比 H/B:").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.h_b_ratio_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(推荐值0.6~0.8，留空则默认0.8)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
            row += 1
            # 倒角角度
            ttk.Label(frame, text="倒角角度 (度):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.chamfer_angle_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            # 倒角底边
            ttk.Label(frame, text="倒角底边 (m):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.chamfer_length_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(倒角两者需同时填写或同时留空)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
            
        elif "隧洞-圆形" in self.section_type:
            # 手动直径
            ttk.Label(frame, text="手动直径 D (m):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.D_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(留空则自动计算)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
            
        elif "隧洞-圆拱直墙型" in self.section_type:
            # 拱顶圆心角
            ttk.Label(frame, text="拱顶圆心角 (度):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.theta_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(留空则采用180°)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
            row += 1
            # 手动底宽
            ttk.Label(frame, text="手动底宽 B (m):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.B_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(手动底宽B留空则自动计算)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
            
        elif "隧洞-马蹄形" in self.section_type:
            # 手动半径
            ttk.Label(frame, text="手动半径 R (m):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.r_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(留空则自动计算)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
        elif self.section_type == "矩形暗涵":
            # 手动宽深比
            ttk.Label(frame, text="手动宽深比:").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.BH_ratio_rect_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            # 手动底宽
            ttk.Label(frame, text="手动底宽 B (m):").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.B_rect_var, width=18).grid(row=row, column=1, padx=5, pady=5)
            row += 1
            ttk.Label(frame, text="(二选一输入，留空则自动计算)", font=('', 8), 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
        else:
            ttk.Label(frame, text="(无额外参数)", 
                     foreground='black').grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=5)
    
    def _create_buttons(self, parent):
        """创建按钮区"""
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(btn_frame, text="确认", command=self._on_confirm, width=12).pack(side=tk.LEFT, padx=5, expand=True)
        ttk.Button(btn_frame, text="取消", command=self._on_cancel, width=12).pack(side=tk.LEFT, padx=5, expand=True)
    
    def _load_current_values(self):
        """从表格数据加载到输入控件"""
        try:
            # 基础参数
            if self.current_values.get('Q'):
                self.Q_var.set(float(self.current_values['Q']))
            if self.current_values.get('n'):
                self.n_var.set(float(self.current_values['n']))
            if self.current_values.get('slope_inv'):
                self.slope_inv_var.set(float(self.current_values['slope_inv']))
            if self.current_values.get('v_min'):
                self.v_min_var.set(float(self.current_values['v_min']))
            if self.current_values.get('v_max'):
                self.v_max_var.set(float(self.current_values['v_max']))
            
            # 根据断面类型加载特定参数 - 使用新的列映射
            if "明渠-梯形" in self.section_type or "明渠-矩形" in self.section_type:
                m = self.current_values.get('m', '')
                self.m_var.set(str(m) if m else "")
                b = self.current_values.get('b', '')
                self.b_var.set(str(b) if b else "")
                b_h_ratio = self.current_values.get('b_h_ratio', '')
                self.b_h_ratio_var.set(str(b_h_ratio) if b_h_ratio else "")
            elif "明渠-圆形" in self.section_type or "隧洞-圆形" in self.section_type:
                D = self.current_values.get('D', '')
                self.D_var.set(str(D) if D else "")
            elif "渡槽-U形" in self.section_type:
                R = self.current_values.get('R', '')
                self.R_var.set(str(R) if R else "")
            elif "渡槽-矩形" in self.section_type:
                chamfer_angle = self.current_values.get('chamfer_angle', '')
                chamfer_length = self.current_values.get('chamfer_length', '')
                self.chamfer_angle_var.set(str(chamfer_angle) if chamfer_angle else "")
                self.chamfer_length_var.set(str(chamfer_length) if chamfer_length else "")
                # 深宽比(使用矩形渡槽深宽比列的值)
                h_b_ratio = self.current_values.get('ducao_depth_ratio', '')
                self.h_b_ratio_var.set(str(h_b_ratio) if h_b_ratio else "")
            elif "隧洞-圆拱直墙型" in self.section_type:
                # 使用b列作为底宽
                b = self.current_values.get('b', '')
                self.B_var.set(str(b) if b else "")
                theta = self.current_values.get('theta', '')
                self.theta_var.set(str(theta) if theta else "")
            elif "隧洞-马蹄形" in self.section_type:
                R = self.current_values.get('R', '')
                self.r_var.set(str(R) if R else "")
            elif self.section_type == "矩形暗涵":
                # 宽深比使用b_h_ratio列
                BH_ratio = self.current_values.get('b_h_ratio', '')
                self.BH_ratio_rect_var.set(str(BH_ratio) if BH_ratio else "")
                b = self.current_values.get('b', '')
                self.B_rect_var.set(str(b) if b else "")
                
        except (ValueError, TypeError):
            pass  # 忽略转换错误，使用默认值
    
    def _on_confirm(self):
        """确认按钮处理"""
        try:
            # 收集并验证参数
            result = {}
            
            # 基础参数验证
            Q = self.Q_var.get()
            if Q <= 0:
                raise ValueError("设计流量必须大于0")
            result['Q'] = Q
            
            n = self.n_var.get()
            if n <= 0:
                raise ValueError("糙率必须大于0")
            result['n'] = n
            
            slope_inv = self.slope_inv_var.get()
            if slope_inv <= 0:
                raise ValueError("水力坡降倒数必须大于0")
            result['slope_inv'] = slope_inv
            
            v_min = self.v_min_var.get()
            v_max = self.v_max_var.get()
            if v_min >= v_max:
                raise ValueError("不淤流速必须小于不冲流速")
            result['v_min'] = v_min
            result['v_max'] = v_max
            
            # 根据断面类型收集特定参数
            if "明渠-梯形" in self.section_type:
                m_str = self.m_var.get().strip()
                if not m_str:
                    raise ValueError("梯形断面必须填写边坡系数m")
                m = float(m_str)
                if m < 0:
                    raise ValueError("边坡系数不能为负")
                result['m'] = m
                # 底宽和宽深比（可选）
                b_str = self.b_var.get().strip()
                result['b'] = float(b_str) if b_str else ""
                ratio_str = self.b_h_ratio_var.get().strip()
                result['b_h_ratio'] = float(ratio_str) if ratio_str else ""
                
            elif "明渠-矩形" in self.section_type:
                result['m'] = 0
                # 底宽和宽深比（可选）
                b_str = self.b_var.get().strip()
                result['b'] = float(b_str) if b_str else ""
                ratio_str = self.b_h_ratio_var.get().strip()
                result['b_h_ratio'] = float(ratio_str) if ratio_str else ""
                
            elif "明渠-圆形" in self.section_type:
                D_str = self.D_var.get().strip()
                result['D'] = float(D_str) if D_str else ""
                
            elif "渡槽-U形" in self.section_type:
                R_str = self.R_var.get().strip()
                result['R'] = float(R_str) if R_str else ""
                
            elif "渡槽-矩形" in self.section_type:
                # 深宽比
                h_b_ratio_str = self.h_b_ratio_var.get().strip()
                result['h_b_ratio'] = float(h_b_ratio_str) if h_b_ratio_str else ""
                # 倒角参数
                angle_str = self.chamfer_angle_var.get().strip()
                length_str = self.chamfer_length_var.get().strip()
                result['chamfer_angle'] = float(angle_str) if angle_str else ""
                result['chamfer_length'] = float(length_str) if length_str else ""
                
            elif "隧洞-圆形" in self.section_type:
                D_str = self.D_var.get().strip()
                result['D'] = float(D_str) if D_str else ""
                
            elif "隧洞-圆拱直墙型" in self.section_type:
                theta_str = self.theta_var.get().strip()
                B_str = self.B_var.get().strip()
                if theta_str:
                    theta = float(theta_str)
                    if theta < 90 or theta > 180:
                        raise ValueError("圆心角必须在90~180度之间")
                    result['theta'] = theta
                else:
                    result['theta'] = ""
                result['B'] = float(B_str) if B_str else ""
                
            elif "隧洞-马蹄形" in self.section_type:
                r_str = self.r_var.get().strip()
                result['r'] = float(r_str) if r_str else ""
            
            elif self.section_type == "矩形暗涵":
                BH_ratio_str = self.BH_ratio_rect_var.get().strip()
                B_str = self.B_rect_var.get().strip()
                result['BH_ratio_rect'] = float(BH_ratio_str) if BH_ratio_str else ""
                result['B_rect'] = float(B_str) if B_str else ""
            
            self.result = result
            self.destroy()
            
        except ValueError as e:
            messagebox.showerror("输入错误", str(e), parent=self)
        except Exception as e:
            messagebox.showerror("错误", f"参数处理错误: {str(e)}", parent=self)
    
    def _on_cancel(self):
        """取消按钮处理"""
        self.result = None
        self.destroy()
    
    def get_result(self) -> dict:
        """获取结果"""
        return self.result


# 测试代码
if __name__ == "__main__":
    root = tk.Tk()
    root.title("渠系建筑物多流量段批量水力计算系统")
    root.geometry("1400x900")
    
    panel = BatchCalculationPanel(root)
    panel.pack(fill=tk.BOTH, expand=True)
    
    root.mainloop()
