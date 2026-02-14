# -*- coding: utf-8 -*-
"""
Excel文件读写工具

提供Excel文件的导入导出功能。
"""

from typing import List, Dict, Any, Optional
from pathlib import Path

# 尝试导入pandas和openpyxl
HAS_PANDAS = False
HAS_OPENPYXL = False
IMPORT_ERROR_MSG = ""

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError as e:
    IMPORT_ERROR_MSG += f"pandas导入失败: {e}\n"
except Exception as e:
    IMPORT_ERROR_MSG += f"pandas导入异常: {e}\n"

try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except ImportError as e:
    IMPORT_ERROR_MSG += f"openpyxl导入失败: {e}\n"
except Exception as e:
    IMPORT_ERROR_MSG += f"openpyxl导入异常: {e}\n"


class ExcelIO:
    """
    Excel文件读写处理器
    
    支持读取和写入Excel文件，兼容xlsx和xls格式。
    """
    
    def __init__(self):
        """初始化Excel处理器"""
        if not HAS_PANDAS and not HAS_OPENPYXL:
            error_msg = "需要安装 pandas 或 openpyxl 库来处理Excel文件"
            if IMPORT_ERROR_MSG:
                error_msg += f"\n详细信息:\n{IMPORT_ERROR_MSG}"
            raise ImportError(error_msg)
    
    def read_excel(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        读取Excel文件
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，默认读取第一个工作表
            
        Returns:
            数据列表，每行为一个字典
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        if HAS_PANDAS:
            return self._read_with_pandas(file_path, sheet_name)
        else:
            return self._read_with_openpyxl(file_path, sheet_name)
    
    def _read_with_pandas(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """使用pandas读取Excel"""
        df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
        return df.to_dict('records')
    
    def _read_with_openpyxl(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """使用openpyxl读取Excel"""
        wb = load_workbook(file_path, data_only=True)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        data = []
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            return data
        
        # 第一行作为表头
        headers = [str(h) if h is not None else f"列{i}" for i, h in enumerate(rows[0])]
        
        # 读取数据行
        for row in rows[1:]:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            data.append(row_dict)
        
        wb.close()
        return data
    
    def write_excel(self, file_path: str, data: List[Dict[str, Any]], 
                    sheet_name: str = "Sheet1") -> None:
        """
        写入Excel文件
        
        Args:
            file_path: 输出文件路径
            data: 数据列表，每行为一个字典
            sheet_name: 工作表名称
        """
        if not data:
            raise ValueError("数据不能为空")
        
        if HAS_PANDAS:
            self._write_with_pandas(file_path, data, sheet_name)
        else:
            self._write_with_openpyxl(file_path, data, sheet_name)
    
    def _write_with_pandas(self, file_path: str, data: List[Dict[str, Any]], 
                           sheet_name: str) -> None:
        """使用pandas写入Excel"""
        df = pd.DataFrame(data)
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
    
    def _write_with_openpyxl(self, file_path: str, data: List[Dict[str, Any]], 
                             sheet_name: str) -> None:
        """使用openpyxl写入Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 写入表头
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
        
        wb.save(file_path)
        wb.close()
    
    def read_coordinates(self, file_path: str, x_col: str = "X", y_col: str = "Y") -> List[tuple]:
        """
        读取坐标数据
        
        Args:
            file_path: Excel文件路径
            x_col: X坐标列名
            y_col: Y坐标列名
            
        Returns:
            坐标列表 [(x1, y1), (x2, y2), ...]
        """
        data = self.read_excel(file_path)
        coordinates = []
        
        for row in data:
            x = row.get(x_col)
            y = row.get(y_col)
            if x is not None and y is not None:
                try:
                    coordinates.append((float(x), float(y)))
                except (ValueError, TypeError):
                    continue
        
        return coordinates
    
    def export_results(self, file_path: str, nodes: List[Any], 
                       settings: Optional[Any] = None) -> None:
        """
        导出计算结果到Excel
        
        Args:
            file_path: 输出文件路径
            nodes: ChannelNode节点列表
            settings: ProjectSettings项目设置（可选）
        """
        # 转换节点数据为字典列表
        data = []
        for node in nodes:
            if hasattr(node, 'to_dict'):
                data.append(node.to_dict())
            else:
                data.append(vars(node))
        
        self.write_excel(file_path, data, sheet_name="水面线计算结果")
