# -*- coding: utf-8 -*-
"""
Excel文件读写工具

提供Excel文件的导入导出功能。
"""

from typing import List, Dict, Any, Optional
from pathlib import Path

# 尝试导入pandas和openpyxl
HAS_PANDAS = False
HAS_OPENPYXL = False
IMPORT_ERROR_MSG = ""

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError as e:
    IMPORT_ERROR_MSG += f"pandas导入失败: {e}\n"
except Exception as e:
    IMPORT_ERROR_MSG += f"pandas导入异常: {e}\n"

try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except ImportError as e:
    IMPORT_ERROR_MSG += f"openpyxl导入失败: {e}\n"
except Exception as e:
    IMPORT_ERROR_MSG += f"openpyxl导入异常: {e}\n"


class ExcelIO:
    """
    Excel文件读写处理器
    
    支持读取和写入Excel文件，兼容xlsx和xls格式。
    """
    
    def __init__(self):
        """初始化Excel处理器"""
        if not HAS_PANDAS and not HAS_OPENPYXL:
            error_msg = "需要安装 pandas 或 openpyxl 库来处理Excel文件"
            if IMPORT_ERROR_MSG:
                error_msg += f"\n详细信息:\n{IMPORT_ERROR_MSG}"
            raise ImportError(error_msg)
    
    def read_excel(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        读取Excel文件
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，默认读取第一个工作表
            
        Returns:
            数据列表，每行为一个字典
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        if HAS_PANDAS:
            return self._read_with_pandas(file_path, sheet_name)
        else:
            return self._read_with_openpyxl(file_path, sheet_name)
    
    def _read_with_pandas(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """使用pandas读取Excel"""
        df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
        return df.to_dict('records')
    
    def _read_with_openpyxl(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """使用openpyxl读取Excel"""
        wb = load_workbook(file_path, data_only=True)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        data = []
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            return data
        
        # 第一行作为表头
        headers = [str(h) if h is not None else f"列{i}" for i, h in enumerate(rows[0])]
        
        # 读取数据行
        for row in rows[1:]:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            data.append(row_dict)
        
        wb.close()
        return data
    
    def write_excel(self, file_path: str, data: List[Dict[str, Any]], 
                    sheet_name: str = "Sheet1") -> None:
        """
        写入Excel文件
        
        Args:
            file_path: 输出文件路径
            data: 数据列表，每行为一个字典
            sheet_name: 工作表名称
        """
        if not data:
            raise ValueError("数据不能为空")
        
        if HAS_PANDAS:
            self._write_with_pandas(file_path, data, sheet_name)
        else:
            self._write_with_openpyxl(file_path, data, sheet_name)
    
    def _write_with_pandas(self, file_path: str, data: List[Dict[str, Any]], 
                           sheet_name: str) -> None:
        """使用pandas写入Excel"""
        df = pd.DataFrame(data)
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
    
    def _write_with_openpyxl(self, file_path: str, data: List[Dict[str, Any]], 
                             sheet_name: str) -> None:
        """使用openpyxl写入Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 写入表头
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
        
        wb.save(file_path)
        wb.close()
    
    def read_coordinates(self, file_path: str, x_col: str = "X", y_col: str = "Y") -> List[tuple]:
        """
        读取坐标数据
        
        Args:
            file_path: Excel文件路径
            x_col: X坐标列名
            y_col: Y坐标列名
            
        Returns:
            坐标列表 [(x1, y1), (x2, y2), ...]
        """
        data = self.read_excel(file_path)
        coordinates = []
        
        for row in data:
            x = row.get(x_col)
            y = row.get(y_col)
            if x is not None and y is not None:
                try:
                    coordinates.append((float(x), float(y)))
                except (ValueError, TypeError):
                    continue
        
        return coordinates
    
    def export_results(self, file_path: str, nodes: List[Any], 
                       settings: Optional[Any] = None) -> None:
        """
        导出计算结果到Excel
        
        Args:
            file_path: 输出文件路径
            nodes: ChannelNode节点列表
            settings: ProjectSettings项目设置（可选）
        """
        # 转换节点数据为字典列表
        data = []
        for node in nodes:
            if hasattr(node, 'to_dict'):
                data.append(node.to_dict())
            else:
                data.append(vars(node))
        
        self.write_excel(file_path, data, sheet_name="水面线计算结果")
    
    def export_ip_plan_table(self, file_path: str, nodes: List[Any],
                             settings: Any) -> None:
        """
        导出「IP点上平面图」Excel文件
        
        按照标准排版格式生成Excel文件，包含：
        - 两行合并表头
        - IP点名称、坐标、桩号、弯道参数、底高程
        
        Args:
            file_path: 输出文件路径（.xlsx）
            nodes: ChannelNode节点列表
            settings: ProjectSettings项目设置（用于桩号前缀等）
        """
        if not HAS_OPENPYXL:
            raise ImportError("需要安装 openpyxl 库来导出IP点上平面图Excel文件")
        
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        # 导入枚举类型（用于判断进出口）
        import sys, os
        parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if parent_dir not in sys.path:
            sys.path.insert(0, parent_dir)
        from models.enums import InOutType
        
        # 获取桩号前缀
        station_prefix = settings.get_station_prefix()
        
        # ========== 辅助函数 ==========
        
        def _safe_float(val, default=0.0):
            """安全转换为浮点数，防止属性为 None / tuple / str 等异常类型"""
            if val is None:
                return default
            try:
                return float(val)
            except (TypeError, ValueError):
                return default
        
        def format_ip_name(node):
            """格式化IP点名称
            
            普通IP点: {prefix}IP{number}
            结构物进出口: {建筑物名称}{类型缩写}{进/出}
            """
            try:
                if node.in_out in (InOutType.INLET, InOutType.OUTLET):
                    # 获取结构形式缩写
                    struct_abbr = ""
                    if node.structure_type:
                        struct_str = node.structure_type.value
                        if "隧洞" in struct_str:
                            struct_abbr = "隧"
                        elif "倒虹吸" in struct_str:
                            struct_abbr = "倒"
                        elif "渡槽" in struct_str:
                            struct_abbr = "渡"
                        elif "暗涵" in struct_str:
                            struct_abbr = "暗"
                    # 获取进出口简写
                    in_out_str = "进" if node.in_out == InOutType.INLET else "出"
                    return f"{node.name}{struct_abbr}{in_out_str}"
            except Exception:
                pass
            return f"{station_prefix}IP{getattr(node, 'ip_number', 0)}"
        
        def format_station(value):
            """格式化桩号"""
            return settings.format_station(_safe_float(value), station_prefix)
        
        # ========== 过滤节点 ==========
        # 排除渐变段和自动插入的明渠段
        real_nodes = [
            n for n in nodes
            if not getattr(n, 'is_transition', False)
            and not getattr(n, 'is_auto_inserted_channel', False)
        ]
        
        if not real_nodes:
            raise ValueError("没有有效的IP点数据可导出")
        
        # ========== 创建工作簿 ==========
        wb = Workbook()
        ws = wb.active
        ws.title = "IP点上平面图"
        
        # ========== 样式定义 ==========
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_font = Font(name='Microsoft YaHei', size=10, bold=True)
        data_font = Font(name='Microsoft YaHei', size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # ========== 写入表头（第1行：分组标题，第2行：列标题） ==========
        
        # 第1行分组标题
        ws.cell(row=1, column=1, value="IP点")         # A1
        ws.cell(row=1, column=2, value="坐标值")       # B1
        ws.cell(row=1, column=4, value="桩号")         # D1
        ws.cell(row=1, column=7, value="弯道参数")     # G1
        ws.cell(row=1, column=11, value="底高程\nm")   # K1
        
        # 第2行列标题
        ws.cell(row=2, column=2, value="E（m）")
        ws.cell(row=2, column=3, value="N（m）")
        ws.cell(row=2, column=4, value="弯前(千米+米)")
        ws.cell(row=2, column=5, value="弯中(千米+米)")
        ws.cell(row=2, column=6, value="弯末(千米+米)")
        ws.cell(row=2, column=7, value="转角")
        ws.cell(row=2, column=8, value="半径")
        ws.cell(row=2, column=9, value="切线长")
        ws.cell(row=2, column=10, value="弧长")
        
        # 合并单元格
        ws.merge_cells('A1:A2')   # IP点
        ws.merge_cells('B1:C1')   # 坐标值
        ws.merge_cells('D1:F1')   # 桩号
        ws.merge_cells('G1:J1')   # 弯道参数
        ws.merge_cells('K1:K2')   # 底高程
        
        # 设置表头样式
        for row in range(1, 3):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill
        
        # ========== 写入数据行（从第3行开始） ==========
        for row_idx, node in enumerate(real_nodes, start=3):
            # A列：IP点名称
            ws.cell(row=row_idx, column=1, value=format_ip_name(node))
            
            # B列：E坐标（6位小数）
            cell_b = ws.cell(row=row_idx, column=2, value=_safe_float(node.x))
            cell_b.number_format = '0.000000'
            
            # C列：N坐标（6位小数）
            cell_c = ws.cell(row=row_idx, column=3, value=_safe_float(node.y))
            cell_c.number_format = '0.000000'
            
            # D列：弯前桩号（字符串）
            ws.cell(row=row_idx, column=4, value=format_station(node.station_BC))
            
            # E列：弯中桩号（字符串）
            ws.cell(row=row_idx, column=5, value=format_station(node.station_MC))
            
            # F列：弯末桩号（字符串）
            ws.cell(row=row_idx, column=6, value=format_station(node.station_EC))
            
            # G列：转角（3位小数）
            cell_g = ws.cell(row=row_idx, column=7, value=round(_safe_float(node.turn_angle), 3))
            cell_g.number_format = '0.000'
            
            # H列：半径（3位小数）
            cell_h = ws.cell(row=row_idx, column=8, value=round(_safe_float(node.turn_radius), 3))
            cell_h.number_format = '0.000'
            
            # I列：切线长（3位小数）
            cell_i = ws.cell(row=row_idx, column=9, value=round(_safe_float(node.tangent_length), 3))
            cell_i.number_format = '0.000'
            
            # J列：弧长（3位小数）
            cell_j = ws.cell(row=row_idx, column=10, value=round(_safe_float(node.arc_length), 3))
            cell_j.number_format = '0.000'
            
            # K列：底高程（3位小数）
            cell_k = ws.cell(row=row_idx, column=11, value=round(_safe_float(node.bottom_elevation), 3))
            cell_k.number_format = '0.000'
            
            # 设置数据行样式
            for col in range(1, 12):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = data_font
                cell.border = thin_border
                if col == 1:
                    cell.alignment = left_align
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # ========== 设置列宽 ==========
        column_widths = {
            'A': 14,   # IP点
            'B': 14,   # E坐标
            'C': 14,   # N坐标
            'D': 18,   # 弯前桩号
            'E': 18,   # 弯中桩号
            'F': 18,   # 弯末桩号
            'G': 8,    # 转角
            'H': 8,    # 半径
            'I': 8,    # 切线长
            'J': 8,    # 弧长
            'K': 10,   # 底高程
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # ========== 保存文件 ==========
        wb.save(file_path)
        wb.close()
