# -*- coding: utf-8 -*-
"""
土石方计算 Excel 成果导出器

输出 5 张工作表（已在 PRD 3.4.3 中确认）：
1. 纵断面数据表       — 桩号/地面高程/设计底高程/挖深
2. 横断面面积汇总表   — 桩号/开挖面积(分层)/回填面积/断面宽度
3. 土石方工程量计算表 — 桩号段/开挖方量(分层)/回填方量/累计方量
4. 工程量汇总表       — 按渠段/全线汇总
5. 方法对比表         — 三种方法结果及差异百分比
"""

from __future__ import annotations
from typing import Optional

from 土石方计算.models.section import (
    LongitudinalData,
    CrossSectionData,
    VolumeResult,
)


class EarthworkExcelExporter:
    """
    土石方成果 Excel 导出器

    Usage
    -----
    >>> exporter = EarthworkExcelExporter()
    >>> exporter.export(
    ...     long_data=long_data,
    ...     sections=sections,
    ...     volume_result=volume_result,
    ...     output_path="土石方计算成果.xlsx"
    ... )
    """

    # 数字格式
    FMT_ELEV  = "0.000"    # 高程（m），3 位小数
    FMT_AREA  = "0.00"     # 面积（m²），2 位小数
    FMT_VOL   = "0.0"      # 体积（m³），1 位小数
    FMT_PCT   = "0.00%"    # 百分比
    FMT_STATION = "@"      # 桩号（文本格式）

    # 列宽（字符数）
    COL_W_STATION  = 14
    COL_W_ELEV     = 12
    COL_W_AREA     = 12
    COL_W_VOL      = 14

    def export(
        self,
        output_path: str,
        long_data: Optional[LongitudinalData] = None,
        sections: Optional[list[CrossSectionData]] = None,
        volume_result: Optional[VolumeResult] = None,
        project_name: str = "",
    ) -> None:
        """
        导出全部计算成果到 Excel 文件。

        Parameters
        ----------
        output_path : 输出路径（.xlsx）
        long_data : 纵断面数据
        sections : 横断面列表（含面积结果）
        volume_result : 工程量计算结果
        project_name : 项目名称（写入标题行）
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, Alignment, Border, Side, PatternFill
            )
        except ImportError as exc:
            raise ImportError("未安装 openpyxl 库，请执行: pip install openpyxl") from exc

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认 Sheet

        if long_data:
            self._write_longitudinal_sheet(wb, long_data, project_name)

        if sections:
            self._write_cross_section_sheet(wb, sections, project_name)

        if volume_result:
            self._write_volume_sheet(wb, volume_result, sections or [], project_name)
            self._write_summary_sheet(wb, volume_result, sections or [], project_name)
            self._write_comparison_sheet(wb, volume_result)

        wb.save(output_path)

    # ------------------------------------------------------------------
    # Sheet 1：纵断面数据表
    # ------------------------------------------------------------------

    def _write_longitudinal_sheet(
        self, wb, data: LongitudinalData, project_name: str
    ) -> None:
        ws = wb.create_sheet("纵断面数据")
        cut_depths = data.get_cut_depths()

        # 标题行
        self._write_title(ws, project_name + " — 纵断面数据表", col_span=4)

        # 表头
        headers = ["桩号", "地面高程(m)", "设计底高程(m)", "挖深/填高(m)"]
        self._write_headers(ws, headers, row=2)
        self._set_col_widths(ws, [self.COL_W_STATION, self.COL_W_ELEV,
                                   self.COL_W_ELEV, self.COL_W_ELEV])

        # 数据行
        for i, s in enumerate(data.stations):
            row = i + 3
            g = data.ground_elevations[i]
            d = data.design_elevations[i]
            cut = cut_depths[i]
            ws.cell(row, 1, _fmt_station(s))
            ws.cell(row, 2, round(g, 3) if g is not None else "")
            ws.cell(row, 3, round(d, 3) if d is not None else "")
            ws.cell(row, 4, round(cut, 3) if cut is not None else "")

        self._apply_basic_style(ws, start_row=2,
                                end_row=len(data.stations) + 2, cols=4)

    # ------------------------------------------------------------------
    # Sheet 2：横断面面积汇总表
    # ------------------------------------------------------------------

    def _write_cross_section_sheet(
        self, wb, sections: list[CrossSectionData], project_name: str
    ) -> None:
        ws = wb.create_sheet("横断面面积汇总")

        # 收集所有地质层名
        all_layers = _collect_layer_names(sections)
        n_layers = len(all_layers)

        # 标题
        total_cols = 4 + n_layers + 2
        self._write_title(ws, project_name + " — 横断面面积汇总表",
                          col_span=total_cols)

        # 表头
        headers = (["桩号", "地面高程(m)", "设计底高程(m)", "挖深(m)"]
                   + [f"{n}\n开挖面积(m²)" for n in all_layers]
                   + ["开挖总面积(m²)", "回填面积(m²)"])
        self._write_headers(ws, headers, row=2)

        col_widths = ([self.COL_W_STATION] + [self.COL_W_ELEV] * 3
                      + [self.COL_W_AREA] * n_layers
                      + [self.COL_W_AREA, self.COL_W_AREA])
        self._set_col_widths(ws, col_widths)

        for i, sec in enumerate(sections):
            row = i + 3
            ar = sec.area_result
            ws.cell(row, 1, _fmt_station(sec.station))
            if ar:
                ws.cell(row, 2, round(ar.ground_elevation_center, 3))
                ws.cell(row, 3, round(ar.design_invert_elevation, 3))
                ws.cell(row, 4, round(ar.cut_depth, 3))
                for j, ln in enumerate(all_layers):
                    ws.cell(row, 5 + j,
                            round(ar.excavation_by_layer.get(ln, 0.0), 2))
                ws.cell(row, 5 + n_layers,
                        round(ar.excavation_total, 2))
                ws.cell(row, 6 + n_layers,
                        round(ar.fill_area, 2))

        self._apply_basic_style(ws, start_row=2,
                                end_row=len(sections) + 2, cols=total_cols)

    # ------------------------------------------------------------------
    # Sheet 3：土石方工程量计算表
    # ------------------------------------------------------------------

    def _write_volume_sheet(
        self,
        wb,
        result: VolumeResult,
        sections: list[CrossSectionData],
        project_name: str,
    ) -> None:
        ws = wb.create_sheet("土石方工程量")

        all_layers = _collect_layer_names(sections)
        n_layers = len(all_layers)
        # 列布局：桩号段(3) + 总量平均/棱台(2) + 各层平均(n) + 各层棱台(n) + 回填平均/棱台(2) + 累计(1)
        total_cols = 3 + 2 + n_layers * 2 + 2 + 1

        self._write_title(ws, project_name + " — 土石方工程量计算表",
                          col_span=total_cols)

        layer_avg_headers   = [f"{n}\n平均断面(m³)" for n in all_layers]
        layer_prism_headers = [f"{n}\n棱台法(m³)"   for n in all_layers]
        headers = (["起始桩号", "终止桩号", "段长(m)",
                    "平均断面\n开挖总(m³)", "棱台法\n开挖总(m³)"]
                   + layer_avg_headers
                   + layer_prism_headers
                   + ["平均断面\n回填(m³)", "棱台法\n回填(m³)",
                      "开挖累计\n(平均断面,m³)"])
        self._write_headers(ws, headers, row=2)

        col_widths = [self.COL_W_STATION, self.COL_W_STATION, 10
                      ] + [self.COL_W_VOL] * (total_cols - 3)
        self._set_col_widths(ws, col_widths)

        cum_avg = 0.0
        for i, seg in enumerate(result.segments):
            row = i + 3
            ws.cell(row, 1, _fmt_station(seg.station_start))
            ws.cell(row, 2, _fmt_station(seg.station_end))
            ws.cell(row, 3, round(seg.length, 1))
            ws.cell(row, 4, round(seg.excavation_avg, 1))
            ws.cell(row, 5, round(seg.excavation_prismatoid, 1))
            # 各层平均断面法
            for j, ln in enumerate(all_layers):
                ws.cell(row, 6 + j,
                        round(seg.excavation_by_layer_avg.get(ln, 0.0), 1))
            # 各层棱台法
            for j, ln in enumerate(all_layers):
                ws.cell(row, 6 + n_layers + j,
                        round(seg.excavation_by_layer_prismatoid.get(ln, 0.0), 1))
            # 回填 + 累计
            ws.cell(row, 6 + n_layers * 2, round(seg.fill_avg, 1))
            ws.cell(row, 7 + n_layers * 2, round(seg.fill_prismatoid, 1))
            cum_avg += seg.excavation_avg
            ws.cell(row, 8 + n_layers * 2, round(cum_avg, 1))

        self._apply_basic_style(ws, start_row=2,
                                end_row=len(result.segments) + 2,
                                cols=total_cols)

    # ------------------------------------------------------------------
    # Sheet 4：工程量汇总表
    # ------------------------------------------------------------------

    def _write_summary_sheet(
        self,
        wb,
        result: VolumeResult,
        sections: list[CrossSectionData],
        project_name: str,
    ) -> None:
        ws = wb.create_sheet("工程量汇总")
        self._write_title(ws, project_name + " — 工程量汇总表", col_span=4)

        layer_totals = result.total_by_layer_avg()
        rows_data = [
            ("开挖总量（平均断面法）", round(result.total_excavation_avg, 1), "m³", ""),
            ("开挖总量（棱台法）", round(result.total_excavation_prismatoid, 1), "m³", ""),
            ("回填总量（平均断面法）", round(result.total_fill_avg, 1), "m³", ""),
        ]
        for ln, vol in layer_totals.items():
            rows_data.append((f"{ln}开挖量（平均断面法）", round(vol, 1), "m³", ""))

        if result.tin_volume_excavation is not None:
            rows_data.append(
                ("开挖总量（TIN体积法）", round(result.tin_volume_excavation, 1), "m³", "")
            )
        if result.tin_volume_fill is not None:
            rows_data.append(
                ("回填总量（TIN体积法）", round(result.tin_volume_fill, 1), "m³", "")
            )

        headers = ["项目", "工程量", "单位", "备注"]
        self._write_headers(ws, headers, row=2)
        self._set_col_widths(ws, [30, 16, 8, 20])

        for i, (name, val, unit, note) in enumerate(rows_data):
            row = i + 3
            ws.cell(row, 1, name)
            ws.cell(row, 2, val)
            ws.cell(row, 3, unit)
            ws.cell(row, 4, note)

        self._apply_basic_style(ws, start_row=2,
                                end_row=len(rows_data) + 2, cols=4)

    # ------------------------------------------------------------------
    # Sheet 5：方法对比表
    # ------------------------------------------------------------------

    def _write_comparison_sheet(self, wb, result: VolumeResult) -> None:
        ws = wb.create_sheet("方法对比")
        self._write_title(ws, "三种土石方计算方法对比", col_span=5)

        headers = ["计算方法", "开挖方量(m³)", "回填方量(m³)", "与平均断面法差异", "备注"]
        self._write_headers(ws, headers, row=2)
        self._set_col_widths(ws, [20, 16, 16, 16, 24])

        avg_exc = result.total_excavation_avg
        pris_exc = result.total_excavation_prismatoid
        tin_exc = result.tin_volume_excavation

        def diff_pct(val: Optional[float]) -> str:
            if val is None or avg_exc < 1e-6:
                return "-"
            return f"{(val - avg_exc) / avg_exc * 100:+.2f}%"

        data = [
            ("平均断面法", round(avg_exc, 1),
             round(result.total_fill_avg, 1), "基准", "水利行业常用"),
            ("棱台法", round(pris_exc, 1),
             round(result.total_fill_avg, 1), diff_pct(pris_exc), "面积变化大时精度更高"),
            ("TIN体积法",
             round(tin_exc, 1) if tin_exc is not None else "-",
             round(result.tin_volume_fill, 1) if result.tin_volume_fill else "-",
             diff_pct(tin_exc), "三维精确法，计算量大"),
        ]

        for i, row_data in enumerate(data):
            row = i + 3
            for j, val in enumerate(row_data):
                ws.cell(row, j + 1, val)

        self._apply_basic_style(ws, start_row=2, end_row=5, cols=5)

    # ------------------------------------------------------------------
    # 样式辅助
    # ------------------------------------------------------------------

    @staticmethod
    def _write_title(ws, title: str, col_span: int) -> None:
        from openpyxl.styles import Font, Alignment
        ws.cell(1, 1, title)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=col_span)
        cell = ws.cell(1, 1)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    @staticmethod
    def _write_headers(ws, headers: list[str], row: int) -> None:
        from openpyxl.styles import Font, Alignment, PatternFill
        fill = PatternFill("solid", fgColor="4472C4")
        for j, h in enumerate(headers):
            cell = ws.cell(row, j + 1, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        ws.row_dimensions[row].height = 30

    @staticmethod
    def _set_col_widths(ws, widths: list[float]) -> None:
        from openpyxl.utils import get_column_letter
        for j, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(j + 1)].width = w

    @staticmethod
    def _apply_basic_style(ws, start_row: int, end_row: int, cols: int) -> None:
        """对数据区域应用边框和居中对齐"""
        from openpyxl.styles import Border, Side, Alignment
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in range(start_row, end_row + 1):
            for c in range(1, cols + 1):
                cell = ws.cell(r, c)
                cell.border = border
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")


# ------------------------------------------------------------------
# 工具函数
# ------------------------------------------------------------------

def _fmt_station(station: float) -> str:
    """格式化桩号为 K0+000.000 字符串"""
    km = int(station // 1000)
    m = station - km * 1000.0
    return f"K{km}+{m:07.3f}"


def _collect_layer_names(sections: list[CrossSectionData]) -> list[str]:
    """从所有断面的地质分层结果中收集所有层名（保序去重）"""
    seen: set[str] = set()
    names: list[str] = []
    for sec in sections:
        if sec.area_result:
            for ln in sec.area_result.excavation_by_layer:
                if ln not in seen:
                    seen.add(ln)
                    names.append(ln)
    return names
