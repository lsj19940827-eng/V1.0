# -*- coding: utf-8 -*-
"""
Excel 数据读取器

支持：
- 地形点坐标表（X/Y/Z 列映射）
- 渠道中心线桩号坐标表
- 纵断面设计底高程表
- 地质分层厚度表
"""

from __future__ import annotations
from typing import Optional

from 土石方计算.models.terrain import TerrainPoint
from 土石方计算.models.alignment import Alignment
from 土石方计算.io.csv_reader import CSVTerrainReader


class ExcelTerrainReader:
    """
    Excel 地形/中心线数据读取器

    内部使用 openpyxl 读取 .xlsx，数据转换后复用 CSVTerrainReader 的解析逻辑。

    Usage
    -----
    >>> reader = ExcelTerrainReader("survey.xlsx")
    >>> pts = reader.read_terrain(sheet="高程点", col_x="X", col_y="Y", col_z="Z")
    >>> alignment = reader.read_centerline(sheet="中心线",
    ...                                     col_station="桩号", col_x="X", col_y="Y")
    """

    def __init__(self, path: str):
        import os
        if not os.path.exists(path):
            raise FileNotFoundError(f"Excel 文件不存在: {path}")
        self._path = path
        self._wb = None  # 懒加载
        self._is_xls = path.lower().endswith('.xls') and not path.lower().endswith('.xlsx')

    # ------------------------------------------------------------------
    # 地形点
    # ------------------------------------------------------------------

    def read_terrain(
        self,
        sheet: str = "Sheet1",
        col_x: str | int = "X",
        col_y: str | int = "Y",
        col_z: str | int = "Z",
        col_id: Optional[str | int] = None,
        header_row: int = 1,
    ) -> list[TerrainPoint]:
        """
        读取地形点坐标表。

        Parameters
        ----------
        sheet : 工作表名称
        col_x/y/z : 列名（字符串，如 "X"）或列索引（1-based int）
        col_id : 点号列（可选）
        header_row : 标题行所在的行号（1-based）

        Returns
        -------
        list[TerrainPoint]
        """
        rows, headers = self._read_sheet(sheet, header_row)
        ix = self._resolve_col(col_x, headers)
        iy = self._resolve_col(col_y, headers)
        iz = self._resolve_col(col_z, headers)
        iid = self._resolve_col(col_id, headers) if col_id is not None else None

        result: list[TerrainPoint] = []
        for i, row in enumerate(rows):
            try:
                x = float(row[ix])
                y = float(row[iy])
                z = float(row[iz])
                src = str(row[iid]) if iid is not None else f"excel_{i}"
                result.append(TerrainPoint(x=x, y=y, z=z, source=src))
            except (IndexError, TypeError, ValueError):
                continue
        return result

    # ------------------------------------------------------------------
    # 中心线
    # ------------------------------------------------------------------

    def read_centerline(
        self,
        sheet: str = "中心线",
        col_station: str | int = "桩号",
        col_x: str | int = "X",
        col_y: str | int = "Y",
        header_row: int = 1,
    ) -> Alignment:
        """
        读取中心线桩号坐标表，构建 Alignment 对象。
        """
        rows, headers = self._read_sheet(sheet, header_row)
        is_col = self._resolve_col(col_station, headers)
        ix = self._resolve_col(col_x, headers)
        iy = self._resolve_col(col_y, headers)

        _parser = CSVTerrainReader._parse_station
        records: list[tuple[float, float, float]] = []
        for row in rows:
            try:
                station = _parser(str(row[is_col]))
                x = float(row[ix])
                y = float(row[iy])
                records.append((station, x, y))
            except (IndexError, TypeError, ValueError):
                continue

        if len(records) < 2:
            raise ValueError(f"中心线有效数据不足（{len(records)} 行），至少需要 2 行")
        return Alignment.from_station_table(records)

    # ------------------------------------------------------------------
    # 纵断面设计底高程表
    # ------------------------------------------------------------------

    def read_design_elevations(
        self,
        sheet: str = "设计底高程",
        col_station: str | int = "桩号",
        col_elevation: str | int = "设计底高程",
        header_row: int = 1,
    ) -> dict[float, float]:
        """
        读取纵断面设计底高程表。

        Returns
        -------
        {station: invert_elevation}
        """
        rows, headers = self._read_sheet(sheet, header_row)
        is_col = self._resolve_col(col_station, headers)
        ie_col = self._resolve_col(col_elevation, headers)

        _parser = CSVTerrainReader._parse_station
        result: dict[float, float] = {}
        for row in rows:
            try:
                station = _parser(str(row[is_col]))
                elev = float(row[ie_col])
                result[station] = elev
            except (IndexError, TypeError, ValueError):
                continue
        return result

    # ------------------------------------------------------------------
    # 地质分层厚度表
    # ------------------------------------------------------------------

    def read_geology_depths(
        self,
        sheet: str = "地质分层",
        col_station: str | int = "桩号",
        layer_columns: Optional[dict[str, str | int]] = None,
        header_row: int = 1,
    ) -> dict[str, list[tuple[float, float]]]:
        """
        读取地质分层厚度表。

        Parameters
        ----------
        layer_columns : {地质层名称: 列名/列索引}
            如 {"残坡积层": "残坡积层厚(m)", "强风化层": "强风化层厚(m)"}

        Returns
        -------
        {layer_name: [(station, thickness_m), ...]}
        """
        if layer_columns is None:
            return {}

        rows, headers = self._read_sheet(sheet, header_row)
        is_col = self._resolve_col(col_station, headers)
        _parser = CSVTerrainReader._parse_station

        result: dict[str, list[tuple[float, float]]] = {
            name: [] for name in layer_columns
        }
        col_indices = {
            name: self._resolve_col(col, headers)
            for name, col in layer_columns.items()
        }

        for row in rows:
            try:
                station = _parser(str(row[is_col]))
            except (IndexError, TypeError, ValueError):
                continue
            for name, ci in col_indices.items():
                try:
                    thickness = float(row[ci])
                    result[name].append((station, thickness))
                except (IndexError, TypeError, ValueError):
                    pass

        return result

    # ------------------------------------------------------------------
    # 内部辅助
    # ------------------------------------------------------------------

    def _ensure_loaded(self):
        if self._wb is None:
            if self._is_xls:
                try:
                    import xlrd
                except ImportError as exc:
                    raise ImportError(
                        "未安装 xlrd 库，请执行: pip install xlrd"
                    ) from exc
                self._wb = xlrd.open_workbook(self._path)
            else:
                try:
                    import openpyxl
                except ImportError as exc:
                    raise ImportError(
                        "未安装 openpyxl 库，请执行: pip install openpyxl"
                    ) from exc
                self._wb = openpyxl.load_workbook(
                    self._path, read_only=True, data_only=True
                )

    def _read_sheet(
        self, sheet_name: str, header_row: int
    ) -> tuple[list[list], list[str]]:
        """
        读取工作表，返回 (数据行列表, 标题行)。

        数据行从 header_row + 1 开始，每行是 cell.value 的列表。
        """
        self._ensure_loaded()
        if self._is_xls:
            names = self._wb.sheet_names()
            if sheet_name not in names:
                ws = self._wb.sheet_by_index(0)
            else:
                ws = self._wb.sheet_by_name(sheet_name)
            all_rows = [
                tuple(ws.cell_value(r, c) for c in range(ws.ncols))
                for r in range(ws.nrows)
            ]
        else:
            if sheet_name not in self._wb.sheetnames:
                ws = self._wb.active
            else:
                ws = self._wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []

        header_idx = header_row - 1
        headers = [str(c) if c is not None else "" for c in all_rows[header_idx]]
        data_rows = [
            [c for c in row]
            for row in all_rows[header_idx + 1:]
        ]
        return data_rows, headers

    @staticmethod
    def _resolve_col(col: str | int, headers: list[str]) -> int:
        """
        将列名或列索引解析为 0-based 整数列索引。
        字符串列名按 headers 查找，int 按 1-based 列号转换（兼容 Excel 习惯）。
        """
        if isinstance(col, int):
            return col - 1  # 1-based → 0-based
        # 字符串列名
        col_str = str(col).strip()
        for i, h in enumerate(headers):
            if h.strip() == col_str:
                return i
        raise ValueError(f"列名 {col_str!r} 在表头中未找到，可用列：{headers}")
