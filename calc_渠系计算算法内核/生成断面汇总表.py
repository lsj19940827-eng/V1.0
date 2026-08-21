# -*- coding: utf-8 -*-
"""
断面尺寸及水力要素汇总表生成模块

根据各流量段参数，调用水力计算模块，按实际结构类型生成建筑物断面汇总表并导出为 Excel。
可直接用于 AutoCAD 制表（通过第三方插件粘贴）。

表格类型（按结果出现情况生成）:
  1. 矩形明渠断面尺寸及水力要素表
  2. 梯形明渠断面尺寸及水力要素表
  3. 圆拱直墙型隧洞断面尺寸及水力要素表（含 III/IV/V 类围岩）
  4. U形渡槽断面尺寸及水力要素表
  5. 矩形暗涵断面尺寸及水力要素表
  6. 圆管涵断面尺寸及水力要素表
  7. 倒虹吸断面尺寸及水力要素表（管道材质可选）
"""

import math
import re
import os
import sys
from typing import List, Dict, Any, Optional, Tuple

# ============================================================
# 导入同级计算模块
# ============================================================
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
if _THIS_DIR not in sys.path:
    sys.path.insert(0, _THIS_DIR)
_REPO_ROOT = os.path.dirname(_THIS_DIR)
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from 明渠设计 import (
    quick_calculate_rectangular,
    quick_calculate_trapezoidal,
    quick_calculate_circular as _calc_circular_pipe,
    quick_calculate_u_section as _calc_u_channel,
)
from 隧洞设计 import (
    quick_calculate_horseshoe as _calc_horseshoe,
    quick_calculate_circular as _calc_tunnel_circular,
    solve_water_depth_horseshoe,
    calculate_horseshoe_outputs,
    solve_water_depth_flat_bottom_circular,
    calculate_flat_bottom_circular_outputs,
    _build_flat_bottom_circular_geometry as _build_tunnel_flat_bottom_geometry,
    get_flow_increase_percent as _tunnel_inc_pct,
)
from 渡槽设计 import quick_calculate_u as _calc_aqueduct_u
from 矩形暗涵设计 import (
    quick_calculate_rectangular_culvert as _calc_rect_culvert,
    get_flow_increase_percent_rect as _culvert_inc_pct,
)
from 圆拱直墙型暗涵设计 import quick_calculate_arch_culvert as _calc_arch_culvert
from 有压管道设计 import (
    PIPE_MATERIALS as PRESSURE_PIPE_MATERIALS,
    get_flow_increase_percent as _pressure_pipe_inc_pct,
)
from 推求水面线.utils.pressure_pipe_common import normalize_pressure_pipe_material_key as _normalize_shared_pressure_pipe_material_key

try:
    from 推求水面线.core.pressure_pipe_calc import calc_total_head_loss as _calc_pressure_pipe_total_head_loss
except ImportError:
    _calc_pressure_pipe_total_head_loss = None

# ============================================================
# 常量
# ============================================================
PI = math.pi
V_MIN = 0.3
V_MAX = 6.0
U_CHANNEL_ALPHA_DEFAULT = 14.0
U_CHANNEL_THETA_DEFAULT = 152.0

SEGMENT_NAMES = [
    "第一流量段", "第二流量段", "第三流量段", "第四流量段",
    "第五流量段", "第六流量段", "第七流量段",
]


def _segment_name(idx: int) -> str:
    if idx <= 0:
        return "流量段"
    if idx <= len(SEGMENT_NAMES):
        return SEGMENT_NAMES[idx - 1]
    return f"第{idx}流量段"

# 隧洞围岩分类
ROCK_CLASSES = ["III类", "IV类", "V类"]
ROCK_LINING_DEFAULT = {
    "III类": {"t0": 0.35, "t": 0.30},
    "IV类":  {"t0": 0.40, "t": 0.40},
    "V类":   {"t0": 0.50, "t": 0.50},
}

_CULVERT_STRUCTURE_TYPE_ALIASES = {
    "暗涵": "暗涵-矩形",
    "暗渠": "暗涵-矩形",
    "矩形暗渠": "暗涵-矩形",
    "矩形暗涵": "暗涵-矩形",
    "暗涵-矩形": "暗涵-矩形",
    "圆拱直墙型暗涵": "暗涵-圆拱直墙型",
    "暗涵-圆拱直墙型": "暗涵-圆拱直墙型",
}

# 倒虹吸管道材质 → 糙率
SIPHON_MATERIALS = {
    "HDPE管":       0.010,
    "PCCP管":       0.012,
    "球墨铸铁管":   0.012,
    "钢管":         0.011,
    "钢筋混凝土管": 0.014,
    "玻璃钢夹砂管": 0.009,
}

def _normalize_pressure_pipe_material_key(material_key: str) -> str:
    return _normalize_shared_pressure_pipe_material_key(
        material_key,
        PRESSURE_PIPE_MATERIALS,
        default_material="球墨铸铁管",
    )


def _get_pressure_pipe_material_params(material_key: str) -> Tuple[str, Dict[str, Any]]:
    key = _normalize_pressure_pipe_material_key(material_key)
    return key, dict(PRESSURE_PIPE_MATERIALS.get(key, PRESSURE_PIPE_MATERIALS["球墨铸铁管"]))


def normalize_pressure_pipe_material_key(material_key: str) -> str:
    """导出/界面共用：将历史材质名或展示名归一化为 canonical key。"""
    return _normalize_pressure_pipe_material_key(material_key)


def get_pressure_pipe_material_display_name(material_key: str) -> str:
    """导出/界面共用：返回材质的人类可读全称。"""
    key, params = _get_pressure_pipe_material_params(material_key)
    return str(params.get("name") or key)


def _fmt_compact(value: float, digits: int = 3) -> str:
    return f"{float(value):.{digits}f}".rstrip("0").rstrip(".")


def _format_pressure_pipe_fmb(material_key: str) -> str:
    _, params = _get_pressure_pipe_material_params(material_key)
    return f"{int(round(params['f']))} / {_fmt_compact(params['m'])} / {_fmt_compact(params['b'])}"


def _format_pressure_pipe_total_head_loss(value):
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "-":
            return "-"
        try:
            value = float(text)
        except (TypeError, ValueError):
            return "-"
    if isinstance(value, (int, float)):
        try:
            number = float(value)
        except (TypeError, ValueError):
            return "-"
        if math.isfinite(number) and number >= 0:
            return round(number, 4)
    return "-"


def _format_pressure_pipe_total_length(value):
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "-":
            return "-"
        try:
            value = float(text)
        except (TypeError, ValueError):
            return "-"
    if isinstance(value, (int, float)):
        try:
            number = float(value)
        except (TypeError, ValueError):
            return "-"
        if math.isfinite(number) and number > 0:
            return round(number, 4)
    return "-"


def _pressure_pipe_length_km(value):
    length_m = _format_pressure_pipe_total_length(value)
    if length_m == "-":
        return "-"
    return round(float(length_m) / 1000.0, 3)


def _pressure_pipe_diameter_m(value):
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "-":
            return "-"
        try:
            value = float(text)
        except (TypeError, ValueError):
            return "-"
    if isinstance(value, (int, float)):
        try:
            number = float(value)
        except (TypeError, ValueError):
            return "-"
        if not math.isfinite(number) or number <= 0:
            return "-"
        if number > 20:
            number /= 1000.0
        return round(number, 3)
    return "-"


_PRESSURE_PIPE_BUILDING_GROUPS = [
    ("tunnel", "隧洞", "tunnel_count", "tunnel_length"),
    ("directional_drill", "定向钻", "directional_drill_count", "directional_drill_length"),
    ("jacking", "顶管", "jacking_count", "jacking_length"),
]


def _format_pressure_pipe_water_level(value):
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "-":
            return "-"
        try:
            value = float(text)
        except (TypeError, ValueError):
            return "-"
    if isinstance(value, (int, float)):
        try:
            number = float(value)
        except (TypeError, ValueError):
            return "-"
        if math.isfinite(number):
            return round(number, 3)
    return "-"


def _pressure_pipe_building_count(value):
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "-":
            return 0
        try:
            value = float(text)
        except (TypeError, ValueError):
            return 0
    if isinstance(value, (int, float)):
        try:
            number = float(value)
        except (TypeError, ValueError):
            return 0
        if math.isfinite(number) and number > 0:
            return int(round(number))
    return 0


def _pressure_pipe_building_length_km(value):
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "-":
            return "-"
        try:
            value = float(text)
        except (TypeError, ValueError):
            return "-"
    if isinstance(value, (int, float)):
        try:
            number = float(value)
        except (TypeError, ValueError):
            return "-"
        if math.isfinite(number) and number >= 0:
            return round(number / 1000.0, 3)
    return "-"


def _pressure_pipe_show_building_characteristics(data: List[Dict[str, Any]]) -> bool:
    for row in data or []:
        if bool(row.get("show_building_characteristics")):
            return True
        for _key, _label, count_key, _length_key in _PRESSURE_PIPE_BUILDING_GROUPS:
            if _pressure_pipe_building_count(row.get(count_key)) > 0:
                return True
    return False


def _format_pressure_pipe_velocity(value: Any) -> str:
    """将压力管道特性表里的设计流速统一格式化为两位小数。"""
    if value in (None, "", "-"):
        return "-"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "-"
    if not math.isfinite(number) or number < 0:
        return "-"
    return f"{number:.2f}"


def _pressure_pipe_row_values(row: Dict[str, Any], include_buildings: bool) -> List[Any]:
    values = [
        row["name"],
        row["Q"],
        row.get("Q_inc", ""),
        _pressure_pipe_length_km(row.get("total_length")),
        row.get("pipe_material", ""),
        _pressure_pipe_diameter_m(row.get("DN_mm", "")),
        _format_pressure_pipe_velocity(row.get("V", "")),
        _format_pressure_pipe_water_level(row.get("start_water_level")),
        _format_pressure_pipe_water_level(row.get("end_water_level")),
    ]
    if include_buildings:
        for _key, _label, count_key, length_key in _PRESSURE_PIPE_BUILDING_GROUPS:
            count_value = _pressure_pipe_building_count(row.get(count_key))
            if count_value <= 0:
                values.extend(["-", "-"])
                continue
            values.extend([count_value, _pressure_pipe_building_length_km(row.get(length_key))])
    return values


def _pressure_pipe_dxf_headers(include_buildings: bool) -> List[Tuple[str, Optional[str]]]:
    headers = [
        ("流量段", None),
        ("设计流量", "m³/s"),
        ("加大流量", "m³/s"),
        ("长度", "km"),
        ("管材", None),
        ("管径", "m"),
        ("设计流速", "m/s"),
        ("渠首水位", "m"),
        ("渠末水位", "m"),
    ]
    if include_buildings:
        headers.extend([
            ("隧洞座数", None),
            ("隧洞长度（km）", None),
            ("定向钻座数", None),
            ("定向钻长度（km）", None),
            ("顶管座数", None),
            ("顶管长度（km）", None),
        ])
    return headers


def _pressure_pipe_excel_col_widths(include_buildings: bool) -> List[float]:
    widths = [14, 12, 12, 12, 14, 10, 12, 11, 11]
    if include_buildings:
        widths.extend([8, 11, 8, 11, 8, 11])
    return widths


def _pressure_pipe_dxf_col_widths(include_buildings: bool) -> List[float]:
    widths = [16, 10, 10, 10, 14, 10, 10, 10, 10]
    if include_buildings:
        widths.extend([8, 10, 8, 10, 8, 10])
    return _dxf_col_widths(widths)


def _pressure_pipe_header_cells(include_buildings: bool) -> List[Dict[str, Any]]:
    cells = [
        {"row": 0, "col": 0, "rowspan": 3, "colspan": 1, "text": "流量段"},
        {"row": 0, "col": 1, "rowspan": 1, "colspan": 1, "text": "设计流量"},
        {"row": 1, "col": 1, "rowspan": 2, "colspan": 1, "text": "m³/s"},
        {"row": 0, "col": 2, "rowspan": 1, "colspan": 1, "text": "加大流量"},
        {"row": 1, "col": 2, "rowspan": 2, "colspan": 1, "text": "m³/s"},
        {"row": 0, "col": 3, "rowspan": 1, "colspan": 1, "text": "长度"},
        {"row": 1, "col": 3, "rowspan": 2, "colspan": 1, "text": "km"},
        {"row": 0, "col": 4, "rowspan": 3, "colspan": 1, "text": "管材"},
        {"row": 0, "col": 5, "rowspan": 1, "colspan": 1, "text": "管径"},
        {"row": 1, "col": 5, "rowspan": 2, "colspan": 1, "text": "m"},
        {"row": 0, "col": 6, "rowspan": 1, "colspan": 1, "text": "设计流速"},
        {"row": 1, "col": 6, "rowspan": 2, "colspan": 1, "text": "m/s"},
        {"row": 0, "col": 7, "rowspan": 1, "colspan": 2, "text": "设计压力线"},
        {"row": 1, "col": 7, "rowspan": 1, "colspan": 1, "text": "渠首水位"},
        {"row": 2, "col": 7, "rowspan": 1, "colspan": 1, "text": "m"},
        {"row": 1, "col": 8, "rowspan": 1, "colspan": 1, "text": "渠末水位"},
        {"row": 2, "col": 8, "rowspan": 1, "colspan": 1, "text": "m"},
    ]
    if not include_buildings:
        return cells

    start_col = 9
    cells.append({"row": 0, "col": start_col, "rowspan": 1, "colspan": 6, "text": "建筑物特性"})
    for offset, (_key, label, _count_key, _length_key) in enumerate(_PRESSURE_PIPE_BUILDING_GROUPS):
        col = start_col + offset * 2
        cells.append({"row": 1, "col": col, "rowspan": 1, "colspan": 2, "text": label})
        cells.append({"row": 2, "col": col, "rowspan": 1, "colspan": 1, "text": "座数"})
        cells.append({"row": 2, "col": col + 1, "rowspan": 1, "colspan": 1, "text": "长度（km）"})
    return cells


def _compute_pressure_pipe_total_head_loss(seg: Dict[str, Any], q_value: float, d_m: float, material_key: str):
    total_head_loss = seg.get("total_head_loss", "")
    if isinstance(total_head_loss, (int, float)) and total_head_loss >= 0:
        return float(total_head_loss)
    if _calc_pressure_pipe_total_head_loss is None:
        return total_head_loss

    ip_points = seg.get("ip_points") or []
    if len(ip_points) < 2:
        return total_head_loss

    try:
        result = _calc_pressure_pipe_total_head_loss(
            name=str(seg.get("name") or ""),
            Q=float(q_value),
            D=float(d_m),
            material_key=material_key,
            ip_points=ip_points,
            upstream_velocity=_to_float(seg.get("upstream_velocity", 0.0), 0.0),
            downstream_velocity=_to_float(seg.get("downstream_velocity", 0.0), 0.0),
            inlet_transition_form=str(seg.get("inlet_transition_form") or "反弯扭曲面"),
            outlet_transition_form=str(seg.get("outlet_transition_form") or "反弯扭曲面"),
            inlet_transition_zeta=_to_float(seg.get("inlet_transition_zeta", 0.0), 0.0) or None,
            outlet_transition_zeta=_to_float(seg.get("outlet_transition_zeta", 0.0), 0.0) or None,
        )
    except Exception:
        return total_head_loss

    return float(result.total_head_loss)

# ============================================================
# 推求水面线结果提取（尽可能复用计算结果）
# ============================================================

def _to_float(val, default: float = 0.0) -> float:
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip()
        if not s or s in ("-", "N/A", "nan"):
            return default
        return float(s)
    except Exception:
        return default


def _is_valid_num(val) -> bool:
    return isinstance(val, (int, float)) and val > 0


def _parse_flow_section_index(flow_section: str) -> Optional[int]:
    if not flow_section:
        return None
    s = str(flow_section).strip()
    if not s:
        return None
    m = re.search(r"\d+", s)
    if m:
        idx = int(m.group(0))
        if idx > 0:
            return idx
    for i, name in enumerate(SEGMENT_NAMES, start=1):
        if name in s:
            return i
    # 兼容中文数字（支持十位）
    cn_map = {"零": 0, "一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
              "六": 6, "七": 7, "八": 8, "九": 9}
    m_cn = re.search(r"[一二三四五六七八九十]+", s)
    if m_cn:
        cn = m_cn.group(0)
        if cn == "十":
            return 10
        if "十" in cn:
            parts = cn.split("十")
            tens = cn_map.get(parts[0], 1) if parts[0] else 1
            ones = cn_map.get(parts[1], 0) if len(parts) > 1 and parts[1] else 0
            return tens * 10 + ones
        if cn in cn_map:
            return cn_map[cn]
    return None


def _get_struct_name(node) -> str:
    def _normalize_culvert_structure_type(value: Any) -> str:
        """将旧暗涵别名归一为断面汇总使用的标准名称。"""
        text = str(value or "").strip()
        return _CULVERT_STRUCTURE_TYPE_ALIASES.get(text, text)

    params = getattr(node, "section_params", {}) or {}
    culvert_family_type = str(params.get("culvert_family_type", "") or "").strip()
    if culvert_family_type:
        return _normalize_culvert_structure_type(culvert_family_type)
    st = getattr(node, "structure_type", None)
    if hasattr(st, "value"):
        return _normalize_culvert_structure_type(st.value)
    return _normalize_culvert_structure_type(st)


def _parse_horseshoe_section_type(value: Any) -> Optional[int]:
    if isinstance(value, (int, float)):
        ivalue = int(value)
        if ivalue in (1, 2):
            return ivalue
    text = str(value or "").strip()
    if not text:
        return None
    if "Ⅱ" in text or "II" in text or "2" in text:
        return 2
    if "Ⅰ" in text or "I" in text or "1" in text:
        return 1
    return None


def _resolve_horseshoe_section_type(seg: Dict[str, Any], default: Optional[int] = 1) -> Optional[int]:
    for key in ("horseshoe_section_type", "section_type", "section_type_name"):
        parsed = _parse_horseshoe_section_type(seg.get(key))
        if parsed is not None:
            return parsed
    return default


def _group_horseshoe_segments(segments: List[Dict[str, Any]]) -> List[Tuple[int, List[Dict[str, Any]]]]:
    grouped: Dict[int, List[Dict[str, Any]]] = {}
    order: List[int] = []
    for seg in segments or []:
        section_type = _resolve_horseshoe_section_type(seg, default=1) or 1
        if section_type not in grouped:
            grouped[section_type] = []
            order.append(section_type)
        seg_copy = dict(seg)
        seg_copy["horseshoe_section_type"] = section_type
        grouped[section_type].append(seg_copy)
    return [(section_type, grouped[section_type]) for section_type in order]


def _expand_horseshoe_table_order(table_order: List[str], horseshoe_keys: List[str]) -> List[str]:
    if not horseshoe_keys:
        return [key for key in table_order if key != "tunnel_horseshoe"]
    expanded: List[str] = []
    inserted = False
    for key in table_order:
        if key == "tunnel_horseshoe":
            expanded.extend(horseshoe_keys)
            inserted = True
            continue
        expanded.append(key)
    if not inserted:
        expanded.extend(horseshoe_keys)
    return expanded


def _classify_structure(node) -> Optional[str]:
    name = _get_struct_name(node)
    params = getattr(node, "section_params", {}) or {}

    if getattr(node, "is_inverted_siphon", False) or "倒虹吸" in name:
        return "siphon"

    # 有压管道（与倒虹吸类似，但独立表格）
    if "有压管道" in name:
        return "pressure_pipe"

    # 隧洞细分：圆拱直墙型 / 圆形 / 马蹄形
    if "隧洞" in name or "隧" in name:
        if "平底圆形" in name:
            return "tunnel_flat_bottom_circular"
        if "圆形" in name:
            return "tunnel_circular"
        if "马蹄" in name:
            return "tunnel_horseshoe"
        if "圆拱直墙" in name:
            return "tunnel_arch"
        # 仅写"隧洞"时，依据参数判断
        d_val = _to_float(params.get("D", params.get("d", 0.0)), 0.0)
        r_val = _to_float(params.get("R_circle", params.get("R", 0.0)), 0.0)
        b_val = _to_float(params.get("B", params.get("b", 0.0)), 0.0)
        if d_val > 0 and b_val > 0:
            return "tunnel_flat_bottom_circular"
        if d_val > 0 and b_val <= 0:
            return "tunnel_circular"
        if r_val > 0 and b_val <= 0:
            return "tunnel_horseshoe"
        return "tunnel_arch"

    # 渡槽细分：U形 / 矩形
    if "渡槽" in name:
        if "矩形" in name:
            return "aqueduct_rect"
        if "U" in name or "u" in name or "U" in name:
            return "aqueduct_u"
        # 仅写"渡槽"时，依据参数判断
        r_val = _to_float(params.get("R_circle", params.get("R", 0.0)), 0.0)
        if r_val > 0:
            return "aqueduct_u"
        return "aqueduct_rect"

    if "暗涵" in name:
        if "圆拱直墙" in name or "圆弧直墙" in name:
            return "rect_culvert_arch"
        return "rect_culvert"

    # 明渠圆形 / 圆管涵
    if "明渠-圆形" in name or "圆形明渠" in name or "明渠圆形" in name or "圆管涵" in name:
        return "circular_channel"

    if "明渠-U形" in name or "U形明渠" in name or "明渠U形" in name:
        return "u_channel"

    # 明渠梯形 / 矩形
    if "明渠-梯形" in name or ("明渠" in name and "梯形" in name) or "梯形明渠" in name:
        return "trap_channel"
    if "明渠-矩形" in name or ("明渠" in name and "矩形" in name) or "矩形明渠" in name:
        return "rect_channel"

    # 兼容旧值：仅写“矩形/梯形/圆形”
    if "U形" in name and "明渠" in name:
        return "u_channel"
    if "梯形" in name:
        return "trap_channel"
    if "矩形" in name:
        return "rect_channel"
    if "圆形" in name:
        return "circular_channel"

    # 仅写“明渠/充水渠/泄水渠”时，依据参数判断圆形/U形/梯形/矩形
    if "明渠" in name or "充水渠" in name or "泄水渠" in name:
        d_val = _to_float(params.get("D", params.get("R_circle", 0.0)), 0.0)
        if d_val > 0:
            return "circular_channel"
        theta_val = _to_float(params.get("theta_deg", 0.0), 0.0)
        r_val = _to_float(params.get("R_circle", params.get("R", 0.0)), 0.0)
        if theta_val > 0 and r_val > 0:
            return "u_channel"
        m_val = _to_float(params.get("m", 0.0), 0.0)
        if m_val > 0:
            return "trap_channel"
        return "rect_channel"

    return None


def _assign_if_valid(target: Dict[str, Any], key: str, value: Any) -> None:
    if isinstance(value, (int, float)):
        if value <= 0:
            return
    else:
        if value is None:
            return
        if isinstance(value, str) and not value.strip():
            return
    if key not in target or (isinstance(target.get(key), (int, float)) and target.get(key, 0) <= 0):
        target[key] = value


def _extract_segment_defaults_from_nodes(nodes) -> Tuple[Dict[str, Dict[int, Dict[str, Any]]], Dict[int, float]]:
    defaults = {
        "rect_channel": {},
        "trap_channel": {},
        "circular_channel": {},
        "u_channel": {},
        "tunnel_arch": {},
        "tunnel_circular": {},
        "tunnel_flat_bottom_circular": {},
        "tunnel_horseshoe": {},
        "aqueduct_u": {},
        "aqueduct_rect": {},
        "rect_culvert": {},
        "rect_culvert_arch": {},
        "siphon": {},
        "pressure_pipe": {},
    }
    flow_qs: Dict[int, float] = {}

    if not nodes:
        return defaults, flow_qs

    for node in nodes:
        if getattr(node, "is_transition", False) or getattr(node, "is_auto_inserted_channel", False):
            continue

        seg_idx = _parse_flow_section_index(getattr(node, "flow_section", ""))
        if not seg_idx:
            continue

        q = _to_float(getattr(node, "flow", 0.0), 0.0)
        if q > 0:
            flow_qs[seg_idx] = max(flow_qs.get(seg_idx, 0.0), q)

        struct_key = _classify_structure(node)
        if not struct_key:
            continue

        target = defaults[struct_key].setdefault(seg_idx, {"name": _segment_name(seg_idx)})

        _assign_if_valid(target, "Q", q)
        _assign_if_valid(target, "n", _to_float(getattr(node, "roughness", 0.0), 0.0))

        slope_i = _to_float(getattr(node, "slope_i", 0.0), 0.0)
        if slope_i > 0:
            _assign_if_valid(target, "slope_inv", 1.0 / slope_i)

        params = getattr(node, "section_params", {}) or {}
        if "use_increase" in params or hasattr(node, "use_increase"):
            target["use_increase"] = _normalize_use_increase(
                params.get("use_increase", getattr(node, "use_increase", None)),
                True,
            )
        b_val = _to_float(params.get("B", params.get("b", params.get("b_design", 0.0))), 0.0)
        _assign_if_valid(target, "B", b_val)

        m_val = _to_float(params.get("m", 0.0), 0.0)
        _assign_if_valid(target, "m", m_val)

        d_val = _to_float(params.get("D", params.get("d", 0.0)), 0.0)
        _assign_if_valid(target, "D", d_val)

        r_val = _to_float(params.get("R_circle", params.get("R", 0.0)), 0.0)
        _assign_if_valid(target, "R", r_val)

        h_val = _to_float(getattr(node, "water_depth", 0.0), 0.0)
        if h_val <= 0:
            h_val = _to_float(params.get("h", params.get("water_depth", 0.0)), 0.0)
        _assign_if_valid(target, "H1", h_val)

        v_val = _to_float(getattr(node, "velocity", 0.0), 0.0)
        _assign_if_valid(target, "V", v_val)

        h_total = _to_float(getattr(node, "structure_height", 0.0), 0.0)
        if struct_key == "rect_culvert_arch":
            _assign_if_valid(target, "H_total", math.ceil(h_total * 100) / 100)
            if "H_straight" in params:
                h_straight_val = _to_float(params.get("H_straight"), None)
                if h_straight_val is not None and h_straight_val >= 0:
                    target["H_straight"] = h_straight_val
            if "manual_H_straight" in params:
                manual_h_straight = _to_float(params.get("manual_H_straight"), None)
                if manual_h_straight is not None and manual_h_straight >= 0:
                    target["manual_H_straight"] = manual_h_straight
            if "used_manual_H_straight" in params:
                target["used_manual_H_straight"] = bool(params.get("used_manual_H_straight"))
        elif struct_key not in {"rect_channel", "trap_channel"}:
            _assign_if_valid(target, "H", math.ceil(h_total * 100) / 100)

        theta_deg = _to_float(params.get("theta_deg", 0.0), 0.0)
        _assign_if_valid(target, "theta_deg", theta_deg)

        alpha_deg = _to_float(params.get("alpha_deg", params.get("chamfer_angle", 0.0)), 0.0)
        _assign_if_valid(target, "alpha_deg", alpha_deg)

        total_head_loss = _to_float(
            getattr(node, "head_loss_siphon", getattr(node, "external_head_loss", 0.0)),
            0.0,
        )
        _assign_if_valid(target, "total_head_loss", total_head_loss)

        pipe_material = str(params.get("pipe_material", "") or "").strip()
        if pipe_material:
            _assign_if_valid(target, "pipe_material", pipe_material)

        # 矩形渡槽倒角参数
        if struct_key == "aqueduct_rect":
            chamfer_angle = _to_float(params.get("chamfer_angle", 0.0), 0.0)
            _assign_if_valid(target, "chamfer_angle", chamfer_angle)
            chamfer_length = _to_float(params.get("chamfer_length", 0.0), 0.0)
            _assign_if_valid(target, "chamfer_length", chamfer_length)

        # 倒虹吸直径（优先D/结构高度）
        if struct_key == "siphon":
            dn_src = d_val if d_val > 0 else h_total
            if dn_src > 0:
                dn_mm = dn_src * 1000 if dn_src < 20 else dn_src
                _assign_if_valid(target, "DN_mm", dn_mm)

        # 有压管道直径（与倒虹吸类似）
        if struct_key == "pressure_pipe":
            dn_src = d_val if d_val > 0 else h_total
            if dn_src > 0:
                dn_mm = dn_src * 1000 if dn_src < 20 else dn_src
                _assign_if_valid(target, "DN_mm", dn_mm)

        if struct_key == "tunnel_horseshoe":
            struct_type = getattr(node, "structure_type", None)
            struct_name = str(getattr(struct_type, "value", struct_type) or "")
            horseshoe_section_type = (
                _parse_horseshoe_section_type(struct_name)
                or _parse_horseshoe_section_type(params.get("horseshoe_section_type"))
                or _parse_horseshoe_section_type(params.get("section_type"))
            )
            if horseshoe_section_type in (1, 2):
                _assign_if_valid(target, "horseshoe_section_type", horseshoe_section_type)

    return defaults, flow_qs


def _apply_overrides(row: Dict[str, Any], seg: Dict[str, Any], mapping: Dict[str, str]) -> None:
    for row_key, seg_key in mapping.items():
        if seg_key in seg:
            val = seg.get(seg_key)
            if isinstance(val, (int, float)):
                if val <= 0:
                    continue
            else:
                if val is None or (isinstance(val, str) and not val.strip()):
                    continue
            row[row_key] = val


def _normalize_use_increase(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if not text:
        return default
    if text in {"0", "false", "no", "off", "否", "不", "unchecked"}:
        return False
    if text in {"1", "true", "yes", "on", "是", "checked"}:
        return True
    return default


def _resolve_manual_increase_percent(seg: Dict[str, Any], use_increase: bool) -> Optional[float]:
    if not use_increase:
        return 0.0
    for key in ("manual_increase_percent", "manual_increase", "increase_percent"):
        if key not in seg:
            continue
        value = seg.get(key)
        try:
            percent = float(value)
        except (TypeError, ValueError):
            continue
        if percent >= 0:
            return percent
    return None


def _parse_tie_rod_height(value: Any) -> float:
    text = str(value or "").strip()
    if not text:
        return 0.0
    normalized = (
        text.replace("×", "x")
        .replace("X", "x")
        .replace("*", "x")
        .replace(" ", "")
    )
    parts = [part for part in normalized.split("x") if part]
    if len(parts) < 2:
        return 0.0
    try:
        height = float(parts[1])
    except (TypeError, ValueError):
        return 0.0
    return height if height > 0 else 0.0


def _compose_open_channel_height(base_height: Any, tie_rod: Any = "") -> Any:
    if not isinstance(base_height, (int, float)) or base_height <= 0:
        return base_height
    return round(float(base_height) + _parse_tie_rod_height(tie_rod), 3)


def _blank_increase_fields(row: Dict[str, Any]) -> Dict[str, Any]:
    if _normalize_use_increase(row.get("use_increase"), True):
        return row
    row["Q_inc"] = ""
    row["H2"] = ""
    return row


def _blank_open_channel_increase_fields(row: Dict[str, Any]) -> Dict[str, Any]:
    return _blank_increase_fields(row)


def _open_channel_include_increase_columns(data: List[Dict[str, Any]]) -> bool:
    if not data:
        return True
    return any(_normalize_use_increase((row or {}).get("use_increase"), True) for row in data)


def _format_slope_inv_text(value: Any) -> str:
    return f"1/{value:g}" if value else ""


def _open_channel_increase_value(row: Dict[str, Any], key: str) -> Any:
    if not _normalize_use_increase((row or {}).get("use_increase"), True):
        return ""
    return row.get(key, "")


def _filter_increase_columns(
    headers: List[Tuple[Any, Any]],
    col_widths: List[Any],
    rows: List[List[Any]],
    *,
    include_increase: bool,
    merge_groups: Optional[List[Tuple[List[int], int]]] = None,
) -> Tuple[List[Tuple[Any, Any]], List[Any], List[List[Any]], Optional[List[Tuple[List[int], int]]]]:
    if include_increase:
        return headers, col_widths, rows, merge_groups

    removable = {"加大流量", "加大水深H₂"}
    keep_indices = [idx for idx, (name, _unit) in enumerate(headers) if name not in removable]
    index_map = {old_idx: new_idx for new_idx, old_idx in enumerate(keep_indices)}

    filtered_headers = [headers[idx] for idx in keep_indices]
    filtered_widths = [col_widths[idx] for idx in keep_indices]
    filtered_rows = [[row[idx] for idx in keep_indices] for row in rows]

    filtered_merge = []
    for cols, span in merge_groups or []:
        mapped_cols = [index_map[idx] for idx in cols if idx in index_map]
        if mapped_cols:
            filtered_merge.append((mapped_cols, span))

    return filtered_headers, filtered_widths, filtered_rows, (filtered_merge or None)


def _positive_or_none(value: Any) -> Optional[float]:
    number = _to_float(value, 0.0)
    return number if number > 0 else None


def _first_positive(seg: Dict[str, Any], *keys: str) -> Optional[float]:
    for key in keys:
        if key in seg:
            value = _positive_or_none(seg.get(key))
            if value is not None:
                return value
    return None


def _has_locked_geometry(seg: Dict[str, Any], keys: Tuple[str, ...]) -> bool:
    return any(_positive_or_none(seg.get(key)) is not None for key in keys)


def _locked_geometry_present(segments: List[Dict[str, Any]], keys: Tuple[str, ...]) -> bool:
    return any(_has_locked_geometry(seg, keys) for seg in segments or [])


def _derive_tunnel_arch_geometry(
    *,
    b: Optional[float],
    h_total: Optional[float],
    h_straight: Optional[float],
    r_arch: Optional[float],
    theta_deg: Optional[float],
) -> Optional[Tuple[float, float, float, float, float]]:
    B = _positive_or_none(b)
    H_total = _positive_or_none(h_total)
    H_straight = _positive_or_none(h_straight)
    theta = _positive_or_none(theta_deg) or 180.0
    theta_rad = math.radians(theta)
    sin_half = math.sin(theta_rad / 2.0)
    one_minus_cos = 1.0 - math.cos(theta_rad / 2.0)

    R_arch = _positive_or_none(r_arch)
    if R_arch is None and B is not None and abs(sin_half) > 1e-9:
        R_arch = (B / 2.0) / sin_half
    if H_straight is None and H_total is not None and R_arch is not None:
        H_straight = H_total - R_arch * one_minus_cos
    if H_total is None and H_straight is not None and R_arch is not None:
        H_total = H_straight + R_arch * one_minus_cos
    if B is None and R_arch is not None and abs(sin_half) > 1e-9:
        B = 2.0 * R_arch * sin_half

    if not all(value is not None and value > 0 for value in (B, H_total, R_arch)):
        return None
    if H_straight is None or H_straight < 0:
        return None
    return float(B), float(H_total), float(H_straight), float(R_arch), float(theta)

# ============================================================
# 默认流量段参数（各表独立）
# ============================================================

def _default_segments_rect_channel():
    """矩形明渠默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    slopes = [3000, 3000, 3000, 3000, 5555, 6666, 7777]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": slopes[i], "n": 0.014,
             "wall_t": 0.3, "tie_rod": "0.2×0.2"} for i in range(7)]

def _default_segments_trap_channel():
    """梯形明渠默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    slopes = [3000, 3000, 3000, 3000, 5555, 6666, 7777]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": slopes[i], "n": 0.014,
             "m": 1.0, "wall_t": 0.3, "tie_rod": "0.2×0.2"} for i in range(7)]


def _default_segments_u_channel():
    """U形明渠默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    slopes = [3000, 3000, 3000, 3000, 5555, 6666, 7777]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": slopes[i], "n": 0.014,
             "R": 0.8, "alpha_deg": U_CHANNEL_ALPHA_DEFAULT, "theta_deg": U_CHANNEL_THETA_DEFAULT}
            for i in range(7)]

def _default_segments_tunnel():
    """隧洞（圆拱直墙型）默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    slopes = [2000, 2000, 2000, 2000, 2500, 2500, 2500]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": slopes[i], "n": 0.014}
            for i in range(7)]

# 向后兼容别名
_default_segments_tunnel_arch = _default_segments_tunnel

def _default_segments_tunnel_circular():
    """隧洞（圆形）默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    slopes = [2000, 2000, 2000, 2000, 2500, 2500, 2500]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": slopes[i], "n": 0.014}
            for i in range(7)]

def _default_segments_tunnel_flat_bottom_circular():
    """隧洞（平底圆形）默认参数。"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    slopes = [2000, 2000, 2000, 2000, 2500, 2500, 2500]
    return [
        {
            "name": _segment_name(i + 1),
            "Q": Qs[i],
            "slope_inv": slopes[i],
            "n": 0.014,
            "D": 4.0,
            "B": 2.0,
        }
        for i in range(7)
    ]

def _default_segments_tunnel_horseshoe():
    """隧洞（马蹄形）默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    slopes = [2000, 2000, 2000, 2000, 2500, 2500, 2500]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": slopes[i], "n": 0.014}
            for i in range(7)]

def _default_segments_aqueduct():
    """渡槽（U形）默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": 2000, "n": 0.014,
             "wall_t": 0.35} for i in range(7)]

# 向后兼容别名
_default_segments_aqueduct_u = _default_segments_aqueduct

def _default_segments_aqueduct_rect():
    """渡槽（矩形）默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": 2000, "n": 0.014,
             "wall_t": 0.35} for i in range(7)]

def _default_segments_rect_culvert():
    """矩形暗涵默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": 2500, "n": 0.014,
             "t0": 0.4, "t1": 0.4, "t2": 0.4} for i in range(7)]


def _default_segments_rect_culvert_arch():
    """圆拱直墙型暗涵默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": 2500, "n": 0.014,
             "theta_deg": 150.0, "t0": 0.4, "t": 0.4} for i in range(7)]

def _default_segments_circular_pipe():
    """圆管涵默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.6, 0.4, 0.2, 0.2]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "slope_inv": 3000, "n": 0.014,
             "pipe_material": "钢筋混凝土"} for i in range(7)]

def _default_segments_siphon():
    """倒虹吸默认参数"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "DN_mm": 1500} for i in range(7)]

def _default_segments_pressure_pipe():
    """有压管道默认参数（与倒虹吸类似）"""
    Qs = [2.0, 1.3, 0.8, 0.5, 0.4, 0.2, 0.5]
    return [{"name": _segment_name(i + 1), "Q": Qs[i], "DN_mm": 1500} for i in range(7)]


# ============================================================
# 1. 矩形明渠
# ============================================================

def compute_rect_channel(segments: List[Dict]) -> List[Dict]:
    rows = []
    for seg in segments:
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        wall_t = seg.get("wall_t", 0.3)
        tie_rod = seg.get("tie_rod", "0.2×0.2")
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        manual_b = _first_positive(seg, "B")

        res = quick_calculate_rectangular(
            Q=Q, n=n, slope_inv=slope_inv, v_min=V_MIN, v_max=V_MAX,
            manual_b=manual_b,
            manual_increase_percent=_resolve_manual_increase_percent(seg, use_increase),
            preserve_manual_b=manual_b is not None,
        )
        if not res.get("success"):
            row = {"name": seg["name"], "Q": Q, "Q_inc": "", "slope_inv": slope_inv,
                   "n": n, "B": "", "H": "", "t": wall_t, "tie_rod": tie_rod,
                   "H1": "", "H2": "", "V": "", "use_increase": use_increase}
            _apply_overrides(row, seg, {
                "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
                "B": "B", "H": "H", "H1": "H1", "H2": "H2", "V": "V",
                "t": "t", "tie_rod": "tie_rod",
            })
            rows.append(_blank_open_channel_increase_fields(row))
            continue

        row = {
            "name":      seg["name"],
            "Q":         Q,
            "Q_inc":     round(res["Q_increased"], 3),
            "slope_inv": slope_inv,
            "n":         n,
            "B":         round(res["b_design"], 2),
            "H":         _compose_open_channel_height(round(res["h_prime"], 3), tie_rod),
            "t":         wall_t,
            "tie_rod":   tie_rod,
            "H1":        round(res["h_design"], 3),
            "H2":        round(res["h_increased"], 3),
            "V":         round(res["V_design"], 3),
            "use_increase": use_increase,
        }
        _apply_overrides(row, seg, {
            "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
            "B": "B", "H": "H", "H1": "H1", "H2": "H2", "V": "V",
            "t": "t", "tie_rod": "tie_rod",
        })
        rows.append(_blank_open_channel_increase_fields(row))
    return rows


# ============================================================
# 2. 梯形明渠
# ============================================================

def compute_trapezoid_channel(segments: List[Dict]) -> List[Dict]:
    rows = []
    for seg in segments:
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        m = seg.get("m", 1.0)
        wall_t = seg.get("wall_t", 0.3)
        tie_rod = seg.get("tie_rod", "0.2×0.2")
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        manual_b = _first_positive(seg, "B")

        res = quick_calculate_trapezoidal(
            Q=Q, m=m, n=n, slope_inv=slope_inv, v_min=V_MIN, v_max=V_MAX,
            manual_beta=_first_positive(seg, "beta", "Beta_design"),
            manual_b=manual_b,
            manual_increase_percent=_resolve_manual_increase_percent(seg, use_increase),
            preserve_manual_b=manual_b is not None,
        )
        if not res.get("success"):
            row = {"name": seg["name"], "Q": Q, "Q_inc": "", "slope_inv": slope_inv,
                   "n": n, "m": m, "B": "", "H": "", "t": wall_t, "tie_rod": tie_rod,
                   "H1": "", "H2": "", "V": "", "beta": "", "use_increase": use_increase}
            _apply_overrides(row, seg, {
                "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
                "m": "m", "B": "B", "H": "H", "H1": "H1", "H2": "H2", "V": "V",
                "t": "t", "tie_rod": "tie_rod", "beta": "beta",
            })
            rows.append(_blank_open_channel_increase_fields(row))
            continue

        row = {
            "name":      seg["name"],
            "Q":         Q,
            "Q_inc":     round(res["Q_increased"], 3),
            "slope_inv": slope_inv,
            "n":         n,
            "m":         m,
            "B":         round(res["b_design"], 2),
            "H":         _compose_open_channel_height(round(res["h_prime"], 3), tie_rod),
            "t":         wall_t,
            "tie_rod":   tie_rod,
            "H1":        round(res["h_design"], 3),
            "H2":        round(res["h_increased"], 3),
            "V":         round(res["V_design"], 3),
            "beta":      round(res.get("Beta_design", 0) or 0, 3) if res.get("Beta_design", 0) else "",
            "use_increase": use_increase,
        }
        _apply_overrides(row, seg, {
            "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
            "m": "m", "B": "B", "H": "H", "H1": "H1", "H2": "H2", "V": "V",
            "beta": "beta", "t": "t", "tie_rod": "tie_rod",
        })
        rows.append(_blank_open_channel_increase_fields(row))
    return rows


# ============================================================
# 2b. U形明渠
# ============================================================

def compute_u_channel(segments: List[Dict]) -> List[Dict]:
    rows = []
    for seg in segments:
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        radius = _first_positive(seg, "R")
        alpha_deg = _first_positive(seg, "alpha_deg", "chamfer_angle") or U_CHANNEL_ALPHA_DEFAULT
        theta_deg = _first_positive(seg, "theta_deg") or U_CHANNEL_THETA_DEFAULT
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)

        if radius is None:
            row = {
                "name": seg["name"], "Q": Q, "Q_inc": "", "slope_inv": slope_inv, "n": n,
                "R": "", "alpha_deg": alpha_deg, "theta_deg": theta_deg,
                "H": "", "H1": "", "H2": "", "V": "", "use_increase": use_increase,
            }
            _apply_overrides(row, seg, {
                "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
                "R": "R", "alpha_deg": "alpha_deg", "theta_deg": "theta_deg",
                "H": "H", "H1": "H1", "H2": "H2", "V": "V",
            })
            rows.append(_blank_open_channel_increase_fields(row))
            continue

        res = _calc_u_channel(
            Q=Q,
            R=radius,
            alpha_deg=alpha_deg,
            theta_deg=theta_deg,
            n=n,
            slope_inv=slope_inv,
            v_min=V_MIN,
            v_max=V_MAX,
            manual_increase_percent=_resolve_manual_increase_percent(seg, use_increase),
        )
        if not res.get("success"):
            row = {
                "name": seg["name"], "Q": Q, "Q_inc": "", "slope_inv": slope_inv, "n": n,
                "R": round(radius, 2), "alpha_deg": alpha_deg, "theta_deg": theta_deg,
                "H": "", "H1": "", "H2": "", "V": "", "use_increase": use_increase,
            }
            _apply_overrides(row, seg, {
                "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
                "R": "R", "alpha_deg": "alpha_deg", "theta_deg": "theta_deg",
                "H": "H", "H1": "H1", "H2": "H2", "V": "V",
            })
            rows.append(_blank_open_channel_increase_fields(row))
            continue

        row = {
            "name": seg["name"],
            "Q": Q,
            "Q_inc": round(res["Q_increased"], 3),
            "slope_inv": slope_inv,
            "n": n,
            "R": round(res["R"], 2),
            "alpha_deg": round(res.get("alpha_deg", alpha_deg), 3),
            "theta_deg": round(res.get("theta_deg", theta_deg), 3),
            "H": round(res["h_prime"], 3),
            "H1": round(res["h_design"], 3),
            "H2": round(res["h_increased"], 3),
            "V": round(res["V_design"], 3),
            "use_increase": use_increase,
        }
        _apply_overrides(row, seg, {
            "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
            "R": "R", "alpha_deg": "alpha_deg", "theta_deg": "theta_deg",
            "H": "H", "H1": "H1", "H2": "H2", "V": "V",
        })
        rows.append(_blank_open_channel_increase_fields(row))
    return rows


# ============================================================
# 3. 隧洞（圆拱直墙型 — 统一断面 + 围岩分类）
# ============================================================

def compute_tunnel(segments: List[Dict],
                   rock_lining: Dict = None,
                   unified: bool = False) -> Tuple[List[Dict], Dict]:
    """
    返回 (rows, tunnel_info)
      rows: 每个 segment × 3 行
      tunnel_info: {"B", "H_total", "H_straight", "R_arch", "theta_deg"}
    unified=True:  按最大流量段设计统一断面，各段分别求水深
    unified=False: 各流量段独立设计断面尺寸
    """
    if rock_lining is None:
        rock_lining = ROCK_LINING_DEFAULT

    override_map = {
        "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
        "B": "B", "H_straight": "H_straight", "R_arch": "R_arch",
        "H1": "H1", "H2": "H2", "V": "V",
    }

    def _design_one_seg(seg, B, H_total, H_straight, R_arch, theta_rad):
        """用给定断面尺寸为单个流量段求水深，返回 rows 列表。"""
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        slope = 1.0 / slope_inv
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        inc_pct = _tunnel_inc_pct(Q) if use_increase else 0.0
        Q_inc = Q * (1 + inc_pct / 100)

        h_d, ok_d = solve_water_depth_horseshoe(B, H_total, theta_rad, n, slope, Q)
        if use_increase:
            h_i, ok_i = solve_water_depth_horseshoe(B, H_total, theta_rad, n, slope, Q_inc)
        else:
            h_i, ok_i = 0.0, False

        V_d = 0.0
        if ok_d and h_d > 0:
            out_d = calculate_horseshoe_outputs(B, H_total, theta_rad, h_d, n, slope)
            V_d = out_d["V"]

        seg_rows = []
        for rc in ROCK_CLASSES:
            row = {
                "name":       seg["name"],
                "Q":          Q,
                "Q_inc":      round(Q_inc, 3) if use_increase else "",
                "rock_class": rc,
                "slope_inv":  slope_inv,
                "n":          n,
                "B":          round(B, 2),
                "H_straight": round(H_straight, 2),
                "R_arch":     round(R_arch, 3),
                "t0":         rock_lining[rc]["t0"],
                "t":          rock_lining[rc]["t"],
                "H1":         round(h_d, 2) if ok_d else "",
                "H2":         round(h_i, 2) if use_increase and ok_i else "",
                "V":          round(V_d, 2) if V_d > 0 else "",
                "use_increase": use_increase,
            }
            _apply_overrides(row, seg, override_map)
            seg_rows.append(_blank_increase_fields(row))
        return seg_rows

    def _empty_rows_for_seg(seg):
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        seg_rows = []
        for rc in ROCK_CLASSES:
            row = {"name": seg["name"], "Q": seg["Q"], "Q_inc": "",
                   "rock_class": rc, "slope_inv": seg["slope_inv"],
                   "n": seg.get("n", 0.014),
                   "B": "", "H_straight": "", "R_arch": "",
                   "t0": rock_lining[rc]["t0"], "t": rock_lining[rc]["t"],
                   "H1": "", "H2": "", "V": "", "use_increase": use_increase}
            _apply_overrides(row, seg, override_map)
            seg_rows.append(_blank_increase_fields(row))
        return seg_rows

    def _resolve_geometry(seg):
        geom = _derive_tunnel_arch_geometry(
            b=seg.get("B"),
            h_total=seg.get("H_total", seg.get("H")),
            h_straight=seg.get("H_straight"),
            r_arch=seg.get("R_arch"),
            theta_deg=seg.get("theta_deg"),
        )
        if geom is not None:
            return geom

        res = _calc_horseshoe(
            Q=seg["Q"], n=seg.get("n", 0.014),
            slope_inv=seg["slope_inv"], v_min=V_MIN, v_max=V_MAX,
            theta_deg=_first_positive(seg, "theta_deg") or 180.0,
            manual_B=_first_positive(seg, "B"),
        )
        if not res.get("success"):
            return None

        B = res["B"]
        H_total = res["H_total"]
        H_straight = res["H_straight"]
        theta_deg = res.get("theta_deg", 180.0)
        theta_rad = math.radians(theta_deg)
        sin_half = math.sin(theta_rad / 2)
        R_arch = (B / 2) / sin_half if abs(sin_half) > 1e-9 else B / 2
        return B, H_total, H_straight, R_arch, theta_deg

    if unified and _locked_geometry_present(segments, ("B", "H", "H_total", "H_straight", "R_arch", "theta_deg")):
        unified = False

    if unified:
        # --- 统一断面：用最大 Q 设计 ---
        max_seg = max(segments, key=lambda s: s["Q"])
        geom_max = _resolve_geometry(max_seg)
        if geom_max is None:
            empty_info = {"B": 0, "H_total": 0, "H_straight": 0, "R_arch": 0, "theta_deg": 180}
            rows = []
            for seg in segments:
                rows.extend(_empty_rows_for_seg(seg))
            return rows, empty_info

        B, H_total, H_straight, R_arch, theta_deg = geom_max
        theta_rad = math.radians(theta_deg)

        tunnel_info = {"B": B, "H_total": H_total, "H_straight": H_straight,
                       "R_arch": R_arch, "theta_deg": theta_deg}

        rows = []
        for seg in segments:
            rows.extend(_design_one_seg(seg, B, H_total, H_straight, R_arch, theta_rad))
        return rows, tunnel_info
    else:
        # --- 独立断面：各流量段分别设计 ---
        rows = []
        first_info = None
        for seg in segments:
            geom = _resolve_geometry(seg)
            if geom is None:
                rows.extend(_empty_rows_for_seg(seg))
                continue

            B, H_total, H_straight, R_arch, theta_deg = geom
            theta_rad = math.radians(theta_deg)

            if first_info is None:
                first_info = {"B": B, "H_total": H_total, "H_straight": H_straight,
                              "R_arch": R_arch, "theta_deg": theta_deg}

            rows.extend(_design_one_seg(seg, B, H_total, H_straight, R_arch, theta_rad))

        if first_info is None:
            first_info = {"B": 0, "H_total": 0, "H_straight": 0, "R_arch": 0, "theta_deg": 180}
        return rows, first_info

# 向后兼容别名
compute_tunnel_arch = compute_tunnel


# ============================================================
# 3b. 隧洞（圆形 — 统一断面 + 围岩分类）
# ============================================================

def compute_tunnel_circular(segments: List[Dict],
                            rock_lining: Dict = None,
                            unified: bool = False) -> Tuple[List[Dict], Dict]:
    """
    圆形隧洞计算。
    返回 (rows, tunnel_info)
      rows: 每个 segment × 3 行（III/IV/V类围岩）
      tunnel_info: {"D"}
    unified=True:  按最大流量段设计统一断面，各段分别求水深
    unified=False: 各流量段独立设计断面尺寸
    """
    if rock_lining is None:
        rock_lining = ROCK_LINING_DEFAULT

    from 隧洞设计 import solve_water_depth_circular, calculate_circular_outputs

    override_map = {
        "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
        "D": "D", "H1": "H1", "H2": "H2", "V": "V",
    }

    def _design_one_seg(seg, D):
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        slope = 1.0 / slope_inv
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        inc_pct = _tunnel_inc_pct(Q) if use_increase else 0.0
        Q_inc = Q * (1 + inc_pct / 100)

        h_d, ok_d = solve_water_depth_circular(D, n, slope, Q)
        if use_increase:
            h_i, ok_i = solve_water_depth_circular(D, n, slope, Q_inc)
        else:
            h_i, ok_i = 0.0, False

        V_d = 0.0
        if ok_d and h_d > 0:
            out_d = calculate_circular_outputs(D, h_d, n, slope)
            V_d = out_d["V"]

        seg_rows = []
        for rc in ROCK_CLASSES:
            row = {
                "name":       seg["name"],
                "Q":          Q,
                "Q_inc":      round(Q_inc, 3) if use_increase else "",
                "rock_class": rc,
                "slope_inv":  slope_inv,
                "n":          n,
                "D":          round(D, 2),
                "t0":         rock_lining[rc]["t0"],
                "t":          rock_lining[rc]["t"],
                "H1":         round(h_d, 2) if ok_d else "",
                "H2":         round(h_i, 2) if use_increase and ok_i else "",
                "V":          round(V_d, 2) if V_d > 0 else "",
                "use_increase": use_increase,
            }
            _apply_overrides(row, seg, override_map)
            seg_rows.append(_blank_increase_fields(row))
        return seg_rows

    def _empty_rows_for_seg(seg):
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        seg_rows = []
        for rc in ROCK_CLASSES:
            row = {"name": seg["name"], "Q": seg["Q"], "Q_inc": "",
                   "rock_class": rc, "slope_inv": seg["slope_inv"],
                   "n": seg.get("n", 0.014),
                   "D": "", "t0": rock_lining[rc]["t0"], "t": rock_lining[rc]["t"],
                   "H1": "", "H2": "", "V": "", "use_increase": use_increase}
            _apply_overrides(row, seg, override_map)
            seg_rows.append(_blank_increase_fields(row))
        return seg_rows

    if unified and _locked_geometry_present(segments, ("D",)):
        unified = False

    if unified:
        # --- 统一断面：用最大 Q 设计 ---
        max_seg = max(segments, key=lambda s: s["Q"])
        res_max = _calc_tunnel_circular(
            Q=max_seg["Q"], n=max_seg.get("n", 0.014),
            slope_inv=max_seg["slope_inv"], v_min=V_MIN, v_max=V_MAX,
            manual_D=_first_positive(max_seg, "D"),
        )
        if not res_max.get("success"):
            empty_info = {"D": 0}
            rows = []
            for seg in segments:
                rows.extend(_empty_rows_for_seg(seg))
            return rows, empty_info

        D = res_max["D"]
        tunnel_info = {"D": D}

        rows = []
        for seg in segments:
            rows.extend(_design_one_seg(seg, D))
        return rows, tunnel_info
    else:
        # --- 独立断面：各流量段分别设计 ---
        rows = []
        first_info = None
        for seg in segments:
            D = _first_positive(seg, "D")
            if D is None:
                res = _calc_tunnel_circular(
                    Q=seg["Q"], n=seg.get("n", 0.014),
                    slope_inv=seg["slope_inv"], v_min=V_MIN, v_max=V_MAX,
                    manual_D=_first_positive(seg, "D"),
                )
                if not res.get("success"):
                    rows.extend(_empty_rows_for_seg(seg))
                    continue
                D = res["D"]
            if first_info is None:
                first_info = {"D": D}

            rows.extend(_design_one_seg(seg, D))

        if first_info is None:
            first_info = {"D": 0}
        return rows, first_info


# ============================================================
# 3c. 隧洞（平底圆形 — 固定 D/B + 围岩分类）
# ============================================================

def compute_tunnel_flat_bottom_circular(
    segments: List[Dict],
    rock_lining: Dict = None,
    unified: bool = False,
) -> Tuple[List[Dict], Dict]:
    """
    平底圆形隧洞计算。
    返回 (rows, tunnel_info)
      rows: 每个 segment × 3 行（III/IV/V类围岩）
      tunnel_info: {"D", "B", "H_total"}
    unified=True: 仅在段内未锁定 D/B 时按最大流量段统一取值。
    unified=False: 各流量段沿用各自 D/B。
    """
    if rock_lining is None:
        rock_lining = ROCK_LINING_DEFAULT

    override_map = {
        "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
        "D": "D", "B": "B", "H_total": "H_total", "H1": "H1", "H2": "H2", "V": "V",
    }

    def _resolve_geometry(seg):
        D = _first_positive(seg, "D")
        B = _first_positive(seg, "B")
        if D is None or B is None or B > D:
            return None
        geom = _build_tunnel_flat_bottom_geometry(D, B)
        return D, B, geom["H_total"]

    def _design_one_seg(seg, D, B, H_total):
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        slope = 1.0 / slope_inv
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        inc_pct = _tunnel_inc_pct(Q) if use_increase else 0.0
        Q_inc = Q * (1 + inc_pct / 100)

        h_d, ok_d = solve_water_depth_flat_bottom_circular(D, B, n, slope, Q)
        if use_increase:
            h_i, ok_i = solve_water_depth_flat_bottom_circular(D, B, n, slope, Q_inc)
        else:
            h_i, ok_i = 0.0, False

        V_d = 0.0
        if ok_d and h_d > 0:
            out_d = calculate_flat_bottom_circular_outputs(D, B, h_d, n, slope)
            V_d = out_d["V"]

        seg_rows = []
        for rc in ROCK_CLASSES:
            row = {
                "name": seg["name"],
                "Q": Q,
                "Q_inc": round(Q_inc, 3) if use_increase else "",
                "rock_class": rc,
                "slope_inv": slope_inv,
                "n": n,
                "D": round(D, 2),
                "B": round(B, 2),
                "H_total": round(H_total, 2),
                "t0": rock_lining[rc]["t0"],
                "t": rock_lining[rc]["t"],
                "H1": round(h_d, 2) if ok_d else "",
                "H2": round(h_i, 2) if use_increase and ok_i else "",
                "V": round(V_d, 2) if V_d > 0 else "",
                "use_increase": use_increase,
            }
            _apply_overrides(row, seg, override_map)
            seg_rows.append(_blank_increase_fields(row))
        return seg_rows

    def _empty_rows_for_seg(seg):
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        seg_rows = []
        for rc in ROCK_CLASSES:
            row = {
                "name": seg["name"],
                "Q": seg["Q"],
                "Q_inc": "",
                "rock_class": rc,
                "slope_inv": seg["slope_inv"],
                "n": seg.get("n", 0.014),
                "D": "",
                "B": "",
                "H_total": "",
                "t0": rock_lining[rc]["t0"],
                "t": rock_lining[rc]["t"],
                "H1": "",
                "H2": "",
                "V": "",
                "use_increase": use_increase,
            }
            _apply_overrides(row, seg, override_map)
            seg_rows.append(_blank_increase_fields(row))
        return seg_rows

    if unified and _locked_geometry_present(segments, ("D", "B")):
        unified = False

    if unified:
        max_seg = max(segments, key=lambda s: s["Q"])
        geom_max = _resolve_geometry(max_seg)
        if geom_max is None:
            empty_info = {"D": 0, "B": 0, "H_total": 0}
            rows = []
            for seg in segments:
                rows.extend(_empty_rows_for_seg(seg))
            return rows, empty_info

        D, B, H_total = geom_max
        tunnel_info = {"D": D, "B": B, "H_total": H_total}
        rows = []
        for seg in segments:
            rows.extend(_design_one_seg(seg, D, B, H_total))
        return rows, tunnel_info

    rows = []
    first_info = None
    for seg in segments:
        geom = _resolve_geometry(seg)
        if geom is None:
            rows.extend(_empty_rows_for_seg(seg))
            continue

        D, B, H_total = geom
        if first_info is None:
            first_info = {"D": D, "B": B, "H_total": H_total}

        rows.extend(_design_one_seg(seg, D, B, H_total))

    if first_info is None:
        first_info = {"D": 0, "B": 0, "H_total": 0}
    return rows, first_info


# ============================================================
# 3d. 隧洞（马蹄形 — 统一断面 + 围岩分类）
# ============================================================

def compute_tunnel_horseshoe(segments: List[Dict],
                             section_type: int = 1,
                             rock_lining: Dict = None,
                             unified: bool = False) -> Tuple[List[Dict], Dict]:
    """
    马蹄形隧洞计算。
    返回 (rows, tunnel_info)
      rows: 每个 segment × 3 行（III/IV/V类围岩）
      tunnel_info: {"R", "section_type_name"}
    section_type: 1=标准Ⅰ型, 2=标准Ⅱ型
    unified=True:  按最大流量段设计统一断面，各段分别求水深
    unified=False: 各流量段独立设计断面尺寸
    """
    if rock_lining is None:
        rock_lining = ROCK_LINING_DEFAULT

    from 隧洞设计 import (
        quick_calculate_horseshoe_std,
        solve_water_depth_horseshoe_std,
        calculate_horseshoe_std_outputs,
    )

    type_name = "马蹄形标准Ⅰ型" if section_type == 1 else "马蹄形标准Ⅱ型"

    override_map = {
        "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
        "R": "R", "H1": "H1", "H2": "H2", "V": "V",
    }

    def _design_one_seg(seg, R):
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        slope = 1.0 / slope_inv
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        inc_pct = _tunnel_inc_pct(Q) if use_increase else 0.0
        Q_inc = Q * (1 + inc_pct / 100)

        h_d, ok_d = solve_water_depth_horseshoe_std(section_type, R, n, slope, Q)
        if use_increase:
            h_i, ok_i = solve_water_depth_horseshoe_std(section_type, R, n, slope, Q_inc)
        else:
            h_i, ok_i = 0.0, False

        V_d = 0.0
        if ok_d and h_d > 0:
            out_d = calculate_horseshoe_std_outputs(section_type, R, h_d, n, slope)
            V_d = out_d["V"]

        seg_rows = []
        for rc in ROCK_CLASSES:
            row = {
                "name":       seg["name"],
                "Q":          Q,
                "Q_inc":      round(Q_inc, 3) if use_increase else "",
                "rock_class": rc,
                "slope_inv":  slope_inv,
                "n":          n,
                "R":          round(R, 2),
                "t0":         rock_lining[rc]["t0"],
                "t":          rock_lining[rc]["t"],
                "H1":         round(h_d, 2) if ok_d else "",
                "H2":         round(h_i, 2) if use_increase and ok_i else "",
                "V":          round(V_d, 2) if V_d > 0 else "",
                "use_increase": use_increase,
            }
            _apply_overrides(row, seg, override_map)
            seg_rows.append(_blank_increase_fields(row))
        return seg_rows

    def _empty_rows_for_seg(seg):
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        seg_rows = []
        for rc in ROCK_CLASSES:
            row = {"name": seg["name"], "Q": seg["Q"], "Q_inc": "",
                   "rock_class": rc, "slope_inv": seg["slope_inv"],
                   "n": seg.get("n", 0.014),
                   "R": "", "t0": rock_lining[rc]["t0"], "t": rock_lining[rc]["t"],
                   "H1": "", "H2": "", "V": "", "use_increase": use_increase}
            _apply_overrides(row, seg, override_map)
            seg_rows.append(_blank_increase_fields(row))
        return seg_rows

    if unified and _locked_geometry_present(segments, ("R",)):
        unified = False

    if unified:
        # --- 统一断面：用最大 Q 设计 ---
        max_seg = max(segments, key=lambda s: s["Q"])
        res_max = quick_calculate_horseshoe_std(
            Q=max_seg["Q"], n=max_seg.get("n", 0.014),
            slope_inv=max_seg["slope_inv"], v_min=V_MIN, v_max=V_MAX,
            section_type=section_type,
            manual_r=_first_positive(max_seg, "R"),
        )
        if not res_max.get("success"):
            empty_info = {"R": 0, "section_type_name": type_name}
            rows = []
            for seg in segments:
                rows.extend(_empty_rows_for_seg(seg))
            return rows, empty_info

        R = res_max["r"]
        tunnel_info = {"R": R, "section_type_name": type_name}

        rows = []
        for seg in segments:
            rows.extend(_design_one_seg(seg, R))
        return rows, tunnel_info
    else:
        # --- 独立断面：各流量段分别设计 ---
        rows = []
        first_info = None
        for seg in segments:
            R = _first_positive(seg, "R")
            if R is None:
                res = quick_calculate_horseshoe_std(
                    Q=seg["Q"], n=seg.get("n", 0.014),
                    slope_inv=seg["slope_inv"], v_min=V_MIN, v_max=V_MAX,
                    section_type=section_type,
                    manual_r=_first_positive(seg, "R"),
                )
                if not res.get("success"):
                    rows.extend(_empty_rows_for_seg(seg))
                    continue
                R = res["r"]
            if first_info is None:
                first_info = {"R": R, "section_type_name": type_name}

            rows.extend(_design_one_seg(seg, R))

        if first_info is None:
            first_info = {"R": 0, "section_type_name": type_name}
        return rows, first_info


def _build_horseshoe_export_entries(
    segments: List[Dict],
    rock_lining: Dict = None,
    unified: bool = False,
) -> List[Dict[str, Any]]:
    entries: List[Dict[str, Any]] = []
    grouped_segments = _group_horseshoe_segments(segments)
    mixed_types = len(grouped_segments) > 1
    for section_type, grouped in grouped_segments:
        rows, info = compute_tunnel_horseshoe(
            grouped,
            section_type=section_type,
            rock_lining=rock_lining,
            unified=unified,
        )
        section_type_name = info.get("section_type_name") or (
            "马蹄形标准Ⅰ型" if section_type == 1 else "马蹄形标准Ⅱ型"
        )
        table_key = f"tunnel_horseshoe_{section_type}" if mixed_types else "tunnel_horseshoe"
        entries.append({
            "key": table_key,
            "section_type": section_type,
            "rows": rows,
            "sheet_name": section_type_name + "隧洞",
            "title": section_type_name + "隧洞断面尺寸及水力要素表",
            "info": info,
        })
    return entries


# ============================================================
# 4a. 渡槽 (U 形)
# ============================================================

def compute_aqueduct_u(segments: List[Dict]) -> List[Dict]:
    rows = []
    for seg in segments:
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        wall_t = seg.get("wall_t", 0.35)
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)

        res = _calc_aqueduct_u(
            Q=Q, n=n, slope_inv=slope_inv, v_min=V_MIN, v_max=V_MAX,
            manual_R=_first_positive(seg, "R"),
            manual_increase_percent=_resolve_manual_increase_percent(seg, use_increase),
        )
        if not res.get("success"):
            row = {"name": seg["name"], "Q": Q, "Q_inc": "", "slope_inv": slope_inv,
                   "n": n, "R": "", "H": "", "t": wall_t,
                   "H1": "", "H2": "", "V": "", "HB_ratio": "", "use_increase": use_increase}
            _apply_overrides(row, seg, {
                "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
                "R": "R", "H": "H", "H1": "H1", "H2": "H2", "V": "V",
                "t": "t",
            })
            rows.append(_blank_increase_fields(row))
            continue

        R = res["R"]
        H_total = res["H_total"]
        hb_ratio = H_total / (2 * R) if R > 0 else 0

        row = {
            "name":      seg["name"],
            "Q":         Q,
            "Q_inc":     round(res["Q_increased"], 3),
            "slope_inv": slope_inv,
            "n":         n,
            "R":         round(R, 2),
            "H":         round(H_total, 2),
            "t":         wall_t,
            "H1":        round(res["h_design"], 2),
            "H2":        round(res["h_increased"], 2),
            "V":         round(res["V_design"], 3),
            "HB_ratio":  round(hb_ratio, 3),
            "use_increase": use_increase,
        }
        _apply_overrides(row, seg, {
            "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
            "R": "R", "H": "H", "H1": "H1", "H2": "H2", "V": "V",
            "t": "t", "HB_ratio": "HB_ratio",
        })
        rows.append(_blank_increase_fields(row))
    return rows


# ============================================================
# 4b. 渡槽 (矩形)
# ============================================================

def compute_aqueduct_rect(segments: List[Dict]) -> List[Dict]:
    from 渡槽设计 import quick_calculate_rect as _calc_aqueduct_rect
    rows = []
    for seg in segments:
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        wall_t = seg.get("wall_t", 0.35)
        chamfer_angle = seg.get("chamfer_angle", 0)
        chamfer_length = seg.get("chamfer_length", 0)
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)

        res = _calc_aqueduct_rect(
            Q=Q, n=n, slope_inv=slope_inv, v_min=V_MIN, v_max=V_MAX,
            chamfer_angle=chamfer_angle, chamfer_length=chamfer_length,
            manual_B=_first_positive(seg, "B"),
            manual_increase_percent=_resolve_manual_increase_percent(seg, use_increase),
        )
        if not res.get("success"):
            row = {"name": seg["name"], "Q": Q, "Q_inc": "", "slope_inv": slope_inv,
                   "n": n, "B": "", "H": "", "t": wall_t,
                   "chamfer_angle": chamfer_angle if chamfer_angle else "",
                   "chamfer_length": chamfer_length if chamfer_length else "",
                   "H1": "", "H2": "", "V": "", "use_increase": use_increase}
            _apply_overrides(row, seg, {
                "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
                "B": "B", "H": "H", "H1": "H1", "H2": "H2", "V": "V",
                "t": "t", "chamfer_angle": "chamfer_angle", "chamfer_length": "chamfer_length",
            })
            rows.append(_blank_increase_fields(row))
            continue

        row = {
            "name":           seg["name"],
            "Q":              Q,
            "Q_inc":          round(res["Q_increased"], 3),
            "slope_inv":      slope_inv,
            "n":              n,
            "B":              round(res["B"], 2),
            "H":              round(res["H_total"], 2),
            "t":              wall_t,
            "chamfer_angle":  res.get("chamfer_angle", 0) or "",
            "chamfer_length": res.get("chamfer_length", 0) or "",
            "H1":             round(res["h_design"], 2),
            "H2":             round(res["h_increased"], 2),
            "V":              round(res["V_design"], 3),
            "use_increase":   use_increase,
        }
        _apply_overrides(row, seg, {
            "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
            "B": "B", "H": "H", "H1": "H1", "H2": "H2", "V": "V",
            "t": "t", "chamfer_angle": "chamfer_angle", "chamfer_length": "chamfer_length",
        })
        rows.append(_blank_increase_fields(row))
    return rows


# ============================================================
# 5. 矩形暗涵
# ============================================================

def compute_rect_culvert(segments: List[Dict]) -> List[Dict]:
    rows = []
    for seg in segments:
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        t0 = seg.get("t0", 0.4)
        t1 = seg.get("t1", 0.4)
        t2 = seg.get("t2", 0.4)
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)

        res = _calc_rect_culvert(
            Q=Q, n=n, slope_inv=slope_inv, v_min=V_MIN, v_max=V_MAX,
            manual_B=_first_positive(seg, "B"),
            manual_increase_percent=_resolve_manual_increase_percent(seg, use_increase),
        )
        if not res.get("success"):
            row = {"name": seg["name"], "Q": Q, "Q_inc": "", "slope_inv": slope_inv,
                   "n": n, "B": "", "H": "", "t0": t0, "t1": t1, "t2": t2,
                   "H1": "", "H2": "", "V": "", "use_increase": use_increase}
            _apply_overrides(row, seg, {
                "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
                "B": "B", "H": "H", "H1": "H1", "H2": "H2", "V": "V",
                "t0": "t0", "t1": "t1", "t2": "t2",
            })
            rows.append(_blank_increase_fields(row))
            continue

        row = {
            "name":      seg["name"],
            "Q":         Q,
            "Q_inc":     round(res["Q_increased"], 3),
            "slope_inv": slope_inv,
            "n":         n,
            "B":         round(res["B"], 2),
            "H":         math.ceil(res["H"] * 100) / 100,
            "t0":        t0,
            "t1":        t1,
            "t2":        t2,
            "H1":        round(res["h_design"], 2),
            "H2":        round(res["h_increased"], 2),
            "V":         round(res["V_design"], 2),
            "use_increase": use_increase,
        }
        _apply_overrides(row, seg, {
            "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
            "B": "B", "H": "H", "H1": "H1", "H2": "H2", "V": "V",
            "t0": "t0", "t1": "t1", "t2": "t2",
        })
        rows.append(_blank_increase_fields(row))
    return rows


def compute_rect_culvert_arch(segments: List[Dict]) -> List[Dict]:
    """圆拱直墙型暗涵计算。"""
    rows = []
    override_map = {
        "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
        "B": "B", "H_total": "H_total", "H_straight": "H_straight", "R_arch": "R_arch",
        "theta_deg": "theta_deg", "t0": "t0", "t": "t",
        "H1": "H1", "H2": "H2", "V": "V",
    }

    def _resolve_geometry(seg):
        geom = _derive_tunnel_arch_geometry(
            b=seg.get("B"),
            h_total=seg.get("H_total", seg.get("H")),
            h_straight=seg.get("H_straight"),
            r_arch=seg.get("R_arch"),
            theta_deg=seg.get("theta_deg"),
        )
        if geom is not None:
            return geom

        manual_increase_percent = _resolve_manual_increase_percent(seg, True)
        if manual_increase_percent == 0:
            manual_increase_percent = None

        res = _calc_arch_culvert(
            Q=seg["Q"],
            n=seg.get("n", 0.014),
            slope_inv=seg["slope_inv"],
            v_min=V_MIN,
            v_max=V_MAX,
            theta_deg=_first_positive(seg, "theta_deg") or 150.0,
            manual_B=_first_positive(seg, "B"),
            manual_H_straight=seg.get("manual_H_straight", seg.get("H_straight")),
            manual_increase_percent=manual_increase_percent,
        )
        if not res.get("success"):
            return None

        B = res["B"]
        H_total = res["H_total"]
        H_straight = res["H_straight"]
        theta_deg = res.get("theta_deg", 150.0)
        theta_rad = math.radians(theta_deg)
        sin_half = math.sin(theta_rad / 2.0)
        R_arch = (B / 2.0) / sin_half if abs(sin_half) > 1e-9 else B / 2.0
        return B, H_total, H_straight, R_arch, theta_deg

    for seg in segments:
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        t0 = seg.get("t0", 0.4)
        t = seg.get("t", seg.get("t1", 0.4))
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)
        geom = _resolve_geometry(seg)
        if geom is None:
            row = {
                "name": seg["name"], "Q": Q, "Q_inc": "", "slope_inv": slope_inv, "n": n,
                "B": "", "H_straight": "", "R_arch": "", "t0": t0, "t": t,
                "H1": "", "H2": "", "V": "", "use_increase": use_increase,
            }
            _apply_overrides(row, seg, override_map)
            rows.append(_blank_increase_fields(row))
            continue

        B, H_total, H_straight, R_arch, theta_deg = geom
        slope = 1.0 / slope_inv
        manual_increase_percent = _resolve_manual_increase_percent(seg, use_increase)
        if use_increase:
            inc_pct = (
                manual_increase_percent
                if manual_increase_percent is not None
                else _culvert_inc_pct(Q)
            )
        else:
            inc_pct = 0.0
        Q_inc = Q * (1 + inc_pct / 100.0)
        theta_rad = math.radians(theta_deg)
        h_d, ok_d = solve_water_depth_horseshoe(B, H_total, theta_rad, n, slope, Q)
        if use_increase:
            h_i, ok_i = solve_water_depth_horseshoe(B, H_total, theta_rad, n, slope, Q_inc)
        else:
            h_i, ok_i = 0.0, False

        V_d = 0.0
        if ok_d and h_d > 0:
            out_d = calculate_horseshoe_outputs(B, H_total, theta_rad, h_d, n, slope)
            V_d = out_d["V"]

        row = {
            "name": seg["name"],
            "Q": Q,
            "Q_inc": round(Q_inc, 3) if use_increase else "",
            "slope_inv": slope_inv,
            "n": n,
            "B": round(B, 2),
            "H_total": round(H_total, 2),
            "H_straight": round(H_straight, 2),
            "R_arch": round(R_arch, 3),
            "theta_deg": round(theta_deg, 3),
            "t0": t0,
            "t": t,
            "H1": round(h_d, 2) if ok_d else "",
            "H2": round(h_i, 2) if use_increase and ok_i else "",
            "V": round(V_d, 2) if V_d > 0 else "",
            "use_increase": use_increase,
        }
        _apply_overrides(row, seg, override_map)
        rows.append(_blank_increase_fields(row))
    return rows


# ============================================================
# 5. 圆管涵（无压自由面流）
# ============================================================

def compute_circular_pipe(segments: List[Dict]) -> List[Dict]:
    """使用明渠-圆形计算模块（无最小直径 2m 限制，适合圆管涵）"""
    rows = []
    for seg in segments:
        Q = seg["Q"]
        slope_inv = seg["slope_inv"]
        n = seg.get("n", 0.014)
        pipe_mat = seg.get("pipe_material", "钢筋混凝土")
        use_increase = _normalize_use_increase(seg.get("use_increase"), True)

        res = _calc_circular_pipe(
            Q=Q, n=n, slope_inv=slope_inv, v_min=V_MIN, v_max=V_MAX,
            increase_percent=_resolve_manual_increase_percent(seg, use_increase),
            manual_D=_first_positive(seg, "D"),
        )
        if not res.get("success"):
            row = {"name": seg["name"], "Q": Q, "Q_inc": "", "slope_inv": slope_inv,
                   "n": n, "D": "", "pipe_material": pipe_mat,
                   "H1": "", "H2": "", "V": "", "use_increase": use_increase}
            _apply_overrides(row, seg, {
                "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
                "D": "D", "pipe_material": "pipe_material",
                "H1": "H1", "H2": "H2", "V": "V",
            })
            rows.append(_blank_open_channel_increase_fields(row))
            continue

        D_val = res.get("D_design", 0) or 0
        Q_inc = res.get("Q_inc", 0) or 0
        y_d = res.get("y_d", 0) or 0
        y_i = res.get("y_i", 0) or 0
        V_d = res.get("V_d", 0) or 0

        row = {
            "name":          seg["name"],
            "Q":             Q,
            "Q_inc":         round(Q_inc, 3) if Q_inc else "",
            "slope_inv":     slope_inv,
            "n":             n,
            "D":             round(D_val, 1) if D_val else "",
            "pipe_material": pipe_mat,
            "H1":            round(y_d, 3) if y_d else "",
            "H2":            round(y_i, 3) if y_i else "",
            "V":             round(V_d, 3) if V_d else "",
            "use_increase":  use_increase,
        }
        _apply_overrides(row, seg, {
            "Q": "Q", "Q_inc": "Q_inc", "slope_inv": "slope_inv", "n": "n",
            "D": "D", "pipe_material": "pipe_material",
            "H1": "H1", "H2": "H2", "V": "V",
        })
        rows.append(_blank_open_channel_increase_fields(row))
    return rows


# ============================================================
# 6. 倒虹吸（有压满管流）
# ============================================================

def compute_siphon(segments: List[Dict],
                   pipe_material: str = "球墨铸铁管") -> List[Dict]:
    rows = []
    for seg in segments:
        # 支持每段独立材质：优先使用段级 pipe_material，否则用全局参数
        seg_mat = seg.get("pipe_material", pipe_material)
        n = seg.get("n", SIPHON_MATERIALS.get(seg_mat, 0.012))

        Q = seg["Q"]
        DN_mm = seg.get("DN_mm", 1500)

        row = {
            "name":          seg["name"],
            "Q":             Q,
            "Q_inc":         round(Q * (1 + _tunnel_inc_pct(Q) / 100), 3),
            "n":             n,
            "DN_mm":         DN_mm,
            "pipe_material": seg_mat,
            "V":             "-",
        }
        _apply_overrides(row, seg, {
            "Q": "Q", "Q_inc": "Q_inc",
            "n": "n", "DN_mm": "DN_mm", "pipe_material": "pipe_material", "V": "V",
        })
        rows.append(row)
    return rows


# ============================================================
# 7. 有压管道（与倒虹吸类似，用于断面汇总表独立输出）
# ============================================================

def compute_pressure_pipe(segments: List[Dict],
                          pipe_material: str = "球墨铸铁管") -> List[Dict]:
    """有压管道断面汇总表计算（保留 f/m/b 与总损失数据，但新版表只显示特性字段）"""
    rows = []
    for seg in segments:
        # 支持每段独立材质：优先使用段级 pipe_material，否则用全局参数
        seg_mat = seg.get("pipe_material", pipe_material)
        material_key, material_params = _get_pressure_pipe_material_params(seg_mat)
        fmb_text = _format_pressure_pipe_fmb(seg_mat)

        Q = seg["Q"]
        DN_mm = seg.get("DN_mm", 1500)

        row = {
            "name":               seg["name"],
            "Q":                  Q,
            "Q_inc":              round(Q * (1 + _pressure_pipe_inc_pct(Q) / 100), 3),
            "friction_params":    fmb_text,
            "pressure_f":         material_params["f"],
            "pressure_m":         material_params["m"],
            "pressure_b":         material_params["b"],
            "DN_mm":              DN_mm,
            "pipe_material":      get_pressure_pipe_material_display_name(material_key),
            "pipe_material_key":  material_key,
            "V":                  "-",
            "plan_total_length":  seg.get("plan_total_length", "-"),
            "total_length":       seg.get("total_length", seg.get("plan_total_length", "-")),
            "total_head_loss":    "-",
            "start_water_level":  seg.get("start_water_level", "-"),
            "end_water_level":    seg.get("end_water_level", "-"),
            "tunnel_count":       seg.get("tunnel_count", 0),
            "tunnel_length":      seg.get("tunnel_length", 0.0),
            "directional_drill_count": seg.get("directional_drill_count", 0),
            "directional_drill_length": seg.get("directional_drill_length", 0.0),
            "jacking_count":      seg.get("jacking_count", 0),
            "jacking_length":     seg.get("jacking_length", 0.0),
            "show_building_characteristics": bool(seg.get("show_building_characteristics", False)),
        }
        _apply_overrides(row, seg, {
            "Q": "Q", "Q_inc": "Q_inc",
            "friction_params": "friction_params",
            "DN_mm": "DN_mm",
            "pipe_material": "pipe_material",
            "V": "V",
            "plan_total_length": "plan_total_length",
            "total_length": "total_length",
            "total_head_loss": "total_head_loss",
            "start_water_level": "start_water_level",
            "end_water_level": "end_water_level",
            "tunnel_count": "tunnel_count",
            "tunnel_length": "tunnel_length",
            "directional_drill_count": "directional_drill_count",
            "directional_drill_length": "directional_drill_length",
            "jacking_count": "jacking_count",
            "jacking_length": "jacking_length",
            "show_building_characteristics": "show_building_characteristics",
        })
        row["pipe_material"] = get_pressure_pipe_material_display_name(row.get("pipe_material"))
        row["pipe_material_key"] = normalize_pressure_pipe_material_key(
            row.get("pipe_material_key") or row.get("pipe_material")
        )
        if row.get("total_length") in (None, "", "-"):
            row["total_length"] = row.get("plan_total_length", "-")
        row["total_length"] = _format_pressure_pipe_total_length(row.get("total_length"))
        row["total_head_loss"] = _format_pressure_pipe_total_head_loss(row.get("total_head_loss"))
        row["start_water_level"] = _format_pressure_pipe_water_level(row.get("start_water_level"))
        row["end_water_level"] = _format_pressure_pipe_water_level(row.get("end_water_level"))
        row["show_building_characteristics"] = bool(row.get("show_building_characteristics"))
        rows.append(row)
    return rows


# ============================================================
# Excel 导出 — 公共辅助
# ============================================================

def _get_openpyxl():
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    thin = Side(style="thin")
    styles = {
        "border":     Border(left=thin, right=thin, top=thin, bottom=thin),
        "title_font": Font(name="宋体", bold=True, size=12),
        "hdr_font":   Font(name="宋体", bold=True, size=10),
        "cell_font":  Font(name="宋体", size=10),
        "center":     Alignment(horizontal="center", vertical="center", wrap_text=True),
        "hdr_fill":   PatternFill(fill_type=None),
    }
    return openpyxl, styles, get_column_letter


def _sc(ws, r, c, val, styles, font_key="cell_font"):
    """写入一个单元格并设置样式"""
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = styles[font_key]
    cell.alignment = styles["center"]
    cell.border = styles["border"]
    return cell


def _write_title(ws, row, col_start, col_end, title, styles):
    """写入标题行（合并居中）"""
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=title)
    cell.font = styles["title_font"]
    cell.alignment = styles["center"]
    cell.border = styles["border"]
    # 给合并区域每个格子加边框
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = styles["border"]


def _write_header_2row(ws, row_name, row_unit, col, name, unit, styles):
    """写入两行表头（名称行 + 单位行）"""
    if unit:
        _sc(ws, row_name, col, name, styles, "hdr_font").fill = styles["hdr_fill"]
        _sc(ws, row_unit, col, unit, styles, "hdr_font").fill = styles["hdr_fill"]
    else:
        # 无单位 → 合并两行
        ws.merge_cells(start_row=row_name, start_column=col,
                       end_row=row_unit, end_column=col)
        _sc(ws, row_name, col, name, styles, "hdr_font").fill = styles["hdr_fill"]
        ws.cell(row=row_unit, column=col).border = styles["border"]
        ws.cell(row=row_unit, column=col).fill = styles["hdr_fill"]


def _merge_vertical(ws, r_start, r_end, col, val, styles, font_key="cell_font"):
    """竖向合并多个单元格并写入值"""
    if r_start == r_end:
        _sc(ws, r_start, col, val, styles, font_key)
        return
    ws.merge_cells(start_row=r_start, start_column=col,
                   end_row=r_end, end_column=col)
    _sc(ws, r_start, col, val, styles, font_key)
    for r in range(r_start, r_end + 1):
        ws.cell(row=r, column=col).border = styles["border"]
        ws.cell(row=r, column=col).alignment = styles["center"]


def _merge_header_cell(ws, r_start, c_start, r_end, c_end, value, styles):
    if r_start != r_end or c_start != c_end:
        ws.merge_cells(
            start_row=r_start,
            start_column=c_start,
            end_row=r_end,
            end_column=c_end,
        )
    cell = ws.cell(row=r_start, column=c_start, value=value)
    cell.font = styles["hdr_font"]
    cell.alignment = styles["center"]
    cell.border = styles["border"]
    cell.fill = styles["hdr_fill"]
    for r in range(r_start, r_end + 1):
        for c in range(c_start, c_end + 1):
            hdr_cell = ws.cell(row=r, column=c)
            hdr_cell.border = styles["border"]
            hdr_cell.alignment = styles["center"]
            hdr_cell.fill = styles["hdr_fill"]


def _set_col_width(ws, col, width, gcl):
    ws.column_dimensions[gcl(col)].width = width


def _filter_table_columns(column_defs: List[Dict[str, Any]], include_increase: bool) -> List[Dict[str, Any]]:
    columns: List[Dict[str, Any]] = []
    for col in column_defs:
        if col.get("increase_only") and not include_increase:
            continue
        columns.append(col)
    return columns


def _table_headers(column_defs: List[Dict[str, Any]]) -> List[Tuple[str, Optional[str]]]:
    return [(str(col["header"]), col.get("unit")) for col in column_defs]


def _table_excel_widths(column_defs: List[Dict[str, Any]]) -> List[float]:
    return [float(col["excel_width"]) for col in column_defs]


def _table_dxf_widths(column_defs: List[Dict[str, Any]]) -> List[float]:
    return _dxf_col_widths([float(col["dxf_width"]) for col in column_defs])


def _table_row_values(row: Dict[str, Any], column_defs: List[Dict[str, Any]]) -> List[Any]:
    values: List[Any] = []
    for col in column_defs:
        getter = col["getter"]
        values.append(getter(row))
    return values


def _rect_channel_column_defs(include_increase: bool) -> List[Dict[str, Any]]:
    return _filter_table_columns([
        {"header": "流量段", "unit": None, "excel_width": 14, "dxf_width": 12, "getter": lambda d: d["name"]},
        {"header": "设计流量", "unit": "m³/s", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d["Q"]},
        {"header": "加大流量", "unit": "m³/s", "excel_width": 12, "dxf_width": 10, "increase_only": True, "getter": lambda d: _open_channel_increase_value(d, "Q_inc")},
        {"header": "1/底坡", "unit": None, "excel_width": 12, "dxf_width": 10, "getter": lambda d: _format_slope_inv_text(d.get("slope_inv"))},
        {"header": "糙率", "unit": None, "excel_width": 10, "dxf_width": 8, "getter": lambda d: d["n"]},
        {"header": "底宽B", "unit": "m", "excel_width": 10, "dxf_width": 8, "getter": lambda d: d.get("B", "")},
        {"header": "高度H", "unit": "m", "excel_width": 10, "dxf_width": 9, "getter": lambda d: d.get("H", "")},
        {"header": "壁厚t", "unit": "m", "excel_width": 10, "dxf_width": 8, "getter": lambda d: d.get("t", "")},
        {"header": "拉杆尺寸", "unit": "m", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d.get("tie_rod", "")},
        {"header": "设计水深H₁", "unit": "m", "excel_width": 13, "dxf_width": 12, "getter": lambda d: d.get("H1", "")},
        {"header": "加大水深H₂", "unit": "m", "excel_width": 13, "dxf_width": 12, "increase_only": True, "getter": lambda d: _open_channel_increase_value(d, "H2")},
        {"header": "设计流速", "unit": "m/s", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d.get("V", "")},
    ], include_increase)


def _trapezoid_channel_column_defs(include_increase: bool) -> List[Dict[str, Any]]:
    return _filter_table_columns([
        {"header": "流量段", "unit": None, "excel_width": 14, "dxf_width": 12, "getter": lambda d: d["name"]},
        {"header": "设计流量", "unit": "m³/s", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d["Q"]},
        {"header": "加大流量", "unit": "m³/s", "excel_width": 12, "dxf_width": 10, "increase_only": True, "getter": lambda d: _open_channel_increase_value(d, "Q_inc")},
        {"header": "1/坡降", "unit": None, "excel_width": 12, "dxf_width": 10, "getter": lambda d: _format_slope_inv_text(d.get("slope_inv"))},
        {"header": "糙率", "unit": None, "excel_width": 10, "dxf_width": 8, "getter": lambda d: d["n"]},
        {"header": "边坡系数m", "unit": None, "excel_width": 12, "dxf_width": 10, "getter": lambda d: d.get("m", "")},
        {"header": "底宽B", "unit": "m", "excel_width": 10, "dxf_width": 8, "getter": lambda d: d.get("B", "")},
        {"header": "高度H", "unit": "m", "excel_width": 10, "dxf_width": 9, "getter": lambda d: d.get("H", "")},
        {"header": "壁厚t", "unit": "m", "excel_width": 10, "dxf_width": 8, "getter": lambda d: d.get("t", "")},
        {"header": "拉杆尺寸", "unit": "m", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d.get("tie_rod", "")},
        {"header": "设计水深H₁", "unit": "m", "excel_width": 13, "dxf_width": 12, "getter": lambda d: d.get("H1", "")},
        {"header": "加大水深H₂", "unit": "m", "excel_width": 13, "dxf_width": 12, "increase_only": True, "getter": lambda d: _open_channel_increase_value(d, "H2")},
        {"header": "设计流速", "unit": "m/s", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d.get("V", "")},
    ], include_increase)


def _u_channel_column_defs(include_increase: bool) -> List[Dict[str, Any]]:
    return _filter_table_columns([
        {"header": "流量段", "unit": None, "excel_width": 14, "dxf_width": 14, "getter": lambda d: d["name"]},
        {"header": "设计流量", "unit": "m³/s", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d["Q"]},
        {"header": "加大流量", "unit": "m³/s", "excel_width": 12, "dxf_width": 10, "increase_only": True, "getter": lambda d: _open_channel_increase_value(d, "Q_inc")},
        {"header": "1/底坡", "unit": None, "excel_width": 12, "dxf_width": 10, "getter": lambda d: _format_slope_inv_text(d.get("slope_inv"))},
        {"header": "糙率", "unit": None, "excel_width": 10, "dxf_width": 8, "getter": lambda d: d["n"]},
        {"header": "半径R", "unit": "m", "excel_width": 10, "dxf_width": 8, "getter": lambda d: d.get("R", "")},
        {"header": "外倾角α", "unit": "°", "excel_width": 10, "dxf_width": 8, "getter": lambda d: d.get("alpha_deg", "")},
        {"header": "圆心角θ", "unit": "°", "excel_width": 10, "dxf_width": 8, "getter": lambda d: d.get("theta_deg", "")},
        {"header": "高度H", "unit": "m", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d.get("H", "")},
        {"header": "设计水深H1", "unit": "m", "excel_width": 13, "dxf_width": 10, "getter": lambda d: d.get("H1", "")},
        {"header": "加大水深H2", "unit": "m", "excel_width": 13, "dxf_width": 10, "increase_only": True, "getter": lambda d: _open_channel_increase_value(d, "H2")},
        {"header": "设计流速", "unit": "m/s", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d.get("V", "")},
    ], include_increase)


def _circular_channel_column_defs(include_increase: bool) -> List[Dict[str, Any]]:
    return _filter_table_columns([
        {"header": "流量段", "unit": None, "excel_width": 14, "dxf_width": 12, "getter": lambda d: d["name"]},
        {"header": "设计流量", "unit": "m³/s", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d["Q"]},
        {"header": "加大流量", "unit": "m³/s", "excel_width": 12, "dxf_width": 10, "increase_only": True, "getter": lambda d: _open_channel_increase_value(d, "Q_inc")},
        {"header": "1/底坡", "unit": None, "excel_width": 12, "dxf_width": 10, "getter": lambda d: _format_slope_inv_text(d.get("slope_inv"))},
        {"header": "糙率", "unit": None, "excel_width": 10, "dxf_width": 8, "getter": lambda d: d["n"]},
        {"header": "直径D", "unit": "m", "excel_width": 10, "dxf_width": 8, "getter": lambda d: d.get("D", "")},
        {"header": "管道材质", "unit": None, "excel_width": 15, "dxf_width": 14, "getter": lambda d: d.get("pipe_material", "")},
        {"header": "设计水深H₁", "unit": "m", "excel_width": 13, "dxf_width": 12, "getter": lambda d: d.get("H1", "")},
        {"header": "加大水深H₂", "unit": "m", "excel_width": 13, "dxf_width": 12, "increase_only": True, "getter": lambda d: _open_channel_increase_value(d, "H2")},
        {"header": "设计流速v", "unit": "m/s", "excel_width": 12, "dxf_width": 10, "getter": lambda d: d.get("V", "")},
    ], include_increase)


# ============================================================
# Sheet 1: 矩形明渠
# ============================================================

def _write_rect_channel(ws, data, styles, gcl, col_offset=0):
    """写入矩形明渠表到 ws，从 col_offset+1 列开始"""
    C = col_offset  # 列偏移
    R1 = 1  # 标题行
    columns = _rect_channel_column_defs(_open_channel_include_increase_columns(data))
    headers = _table_headers(columns)
    col_widths = _table_excel_widths(columns)
    NCOLS = len(columns)

    # 标题
    _write_title(ws, R1, C + 1, C + NCOLS, "矩形明渠断面尺寸及水力要素表", styles)
    # 表头
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    # 列宽
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)
    # 数据
    for ri, d in enumerate(data):
        r = R1 + 3 + ri
        vals = _table_row_values(d, columns)
        for ci, v in enumerate(vals):
            _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# Sheet 1b: 梯形明渠
# ============================================================

def _write_trapezoid_channel(ws, data, styles, gcl, col_offset=0):
    """写入梯形明渠表到 ws，从 col_offset+1 列开始"""
    C = col_offset
    R1 = 1
    columns = _trapezoid_channel_column_defs(_open_channel_include_increase_columns(data))
    headers = _table_headers(columns)
    col_widths = _table_excel_widths(columns)
    NCOLS = len(columns)

    _write_title(ws, R1, C + 1, C + NCOLS, "梯形明渠断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    for ri, d in enumerate(data):
        r = R1 + 3 + ri
        vals = _table_row_values(d, columns)
        for ci, v in enumerate(vals):
            _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# Sheet 1c: U形明渠
# ============================================================

def _write_u_channel(ws, data, styles, gcl, col_offset=0):
    C = col_offset
    R1 = 1
    columns = _u_channel_column_defs(_open_channel_include_increase_columns(data))
    headers = _table_headers(columns)
    col_widths = _table_excel_widths(columns)
    NCOLS = len(columns)

    _write_title(ws, R1, C + 1, C + NCOLS, "U形明渠断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    for ri, d in enumerate(data):
        r = R1 + 3 + ri
        vals = _table_row_values(d, columns)
        for ci, v in enumerate(vals):
            _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# Sheet 2: 隧洞
# ============================================================

def _write_tunnel(ws, data, styles, gcl, col_offset=0):
    C = col_offset
    R1 = 1

    headers_full = [
        ("流量段",      None),
        ("设计流量",    "m³/s"),
        ("加大流量",    "m³/s"),
        ("围岩类型",    None),
        ("1/底坡",      None),
        ("糙率",        None),
        ("底宽B",       "m"),
        ("直墙高H",     "m"),
        ("顶拱半径R",   "m"),
        ("底板厚t₀",    "m"),
        ("边墙顶拱厚t", "m"),
        ("设计水深H₁",  "m"),
        ("加大水深H₂",  "m"),
        ("设计流速",    "m/s"),
    ]
    col_widths_full = [14, 12, 12, 12, 12, 10, 10, 10, 12, 11, 13, 13, 13, 12]
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            d["rock_class"],
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"],
            d.get("B", ""), d.get("H_straight", ""), d.get("R_arch", ""),
            d["t0"], d["t"],
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, merge_groups = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
        merge_groups=[([0, 1, 2], 3)] if len(data) >= 3 else None,
    )
    NCOLS = len(headers)

    _write_title(ws, R1, C + 1, C + NCOLS, "圆拱直墙型隧洞断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    num_segments = len(rows) // 3 if rows else 0
    merged_cols = {col for cols, _span in merge_groups or [] for col in cols}
    for si in range(num_segments):
        base_idx = si * 3
        r_start = R1 + 3 + base_idx
        r_end = r_start + 2
        first_vals = rows[base_idx]

        for cols, _span in merge_groups or []:
            for col_idx in cols:
                _merge_vertical(ws, r_start, r_end, C + 1 + col_idx, first_vals[col_idx], styles)

        for j in range(3):
            r = r_start + j
            vals = rows[base_idx + j]
            for ci, v in enumerate(vals):
                if ci in merged_cols:
                    continue
                _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS

# 向后兼容别名
_write_tunnel_arch = _write_tunnel


# ============================================================
# Sheet 2b: 圆形隧洞
# ============================================================

def _write_tunnel_circular(ws, data, styles, gcl, col_offset=0):
    C = col_offset
    R1 = 1

    headers_full = [
        ("流量段",      None),
        ("设计流量",    "m³/s"),
        ("加大流量",    "m³/s"),
        ("围岩类型",    None),
        ("1/底坡",      None),
        ("糙率",        None),
        ("直径D",       "m"),
        ("底板厚t₀",    "m"),
        ("衬砌厚t",     "m"),
        ("设计水深H₁",  "m"),
        ("加大水深H₂",  "m"),
        ("设计流速",    "m/s"),
    ]
    col_widths_full = [14, 12, 12, 12, 12, 10, 10, 11, 11, 13, 13, 12]
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            d["rock_class"],
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"],
            d.get("D", ""), d["t0"], d["t"],
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, merge_groups = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
        merge_groups=[([0, 1, 2], 3)] if len(data) >= 3 else None,
    )
    NCOLS = len(headers)

    _write_title(ws, R1, C + 1, C + NCOLS, "圆形隧洞断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    num_segments = len(rows) // 3 if rows else 0
    merged_cols = {col for cols, _span in merge_groups or [] for col in cols}
    for si in range(num_segments):
        base_idx = si * 3
        r_start = R1 + 3 + base_idx
        r_end = r_start + 2
        first_vals = rows[base_idx]

        for cols, _span in merge_groups or []:
            for col_idx in cols:
                _merge_vertical(ws, r_start, r_end, C + 1 + col_idx, first_vals[col_idx], styles)

        for j in range(3):
            r = r_start + j
            vals = rows[base_idx + j]
            for ci, v in enumerate(vals):
                if ci in merged_cols:
                    continue
                _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# Sheet 2c: 平底圆形隧洞
# ============================================================

def _write_tunnel_flat_bottom_circular(ws, data, styles, gcl, col_offset=0):
    C = col_offset
    R1 = 1

    headers_full = [
        ("流量段", None),
        ("设计流量", "m³/s"),
        ("加大流量", "m³/s"),
        ("围岩类型", None),
        ("1/底坡", None),
        ("糙率", None),
        ("直径D", "m"),
        ("平底宽B", "m"),
        ("总高H", "m"),
        ("底板厚t₀", "m"),
        ("衬砌厚t", "m"),
        ("设计水深H₁", "m"),
        ("加大水深H₂", "m"),
        ("设计流速", "m/s"),
    ]
    col_widths_full = [14, 12, 12, 12, 12, 10, 10, 10, 10, 11, 11, 13, 13, 12]
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            d["rock_class"],
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"],
            d.get("D", ""), d.get("B", ""), d.get("H_total", ""),
            d["t0"], d["t"],
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, merge_groups = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
        merge_groups=[([0, 1, 2], 3)] if len(data) >= 3 else None,
    )
    NCOLS = len(headers)

    _write_title(ws, R1, C + 1, C + NCOLS, "平底圆形隧洞断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    num_segments = len(rows) // 3 if rows else 0
    merged_cols = {col for cols, _span in merge_groups or [] for col in cols}
    for si in range(num_segments):
        base_idx = si * 3
        r_start = R1 + 3 + base_idx
        r_end = r_start + 2
        first_vals = rows[base_idx]

        for cols, _span in merge_groups or []:
            for col_idx in cols:
                _merge_vertical(ws, r_start, r_end, C + 1 + col_idx, first_vals[col_idx], styles)

        for j in range(3):
            r = r_start + j
            vals = rows[base_idx + j]
            for ci, v in enumerate(vals):
                if ci in merged_cols:
                    continue
                _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# Sheet 2d: 马蹄形隧洞
# ============================================================

def _write_tunnel_horseshoe(ws, data, styles, gcl, col_offset=0, title=None):
    C = col_offset
    R1 = 1

    headers_full = [
        ("流量段",      None),
        ("设计流量",    "m³/s"),
        ("加大流量",    "m³/s"),
        ("围岩类型",    None),
        ("1/底坡",      None),
        ("糙率",        None),
        ("半径R",       "m"),
        ("底板厚t₀",    "m"),
        ("衬砌厚t",     "m"),
        ("设计水深H₁",  "m"),
        ("加大水深H₂",  "m"),
        ("设计流速",    "m/s"),
    ]
    col_widths_full = [14, 12, 12, 12, 12, 10, 10, 11, 11, 13, 13, 12]
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            d["rock_class"],
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"],
            d.get("R", ""), d["t0"], d["t"],
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, merge_groups = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
        merge_groups=[([0, 1, 2], 3)] if len(data) >= 3 else None,
    )
    NCOLS = len(headers)

    _write_title(ws, R1, C + 1, C + NCOLS, title or "马蹄形隧洞断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    num_segments = len(rows) // 3 if rows else 0
    merged_cols = {col for cols, _span in merge_groups or [] for col in cols}
    for si in range(num_segments):
        base_idx = si * 3
        r_start = R1 + 3 + base_idx
        r_end = r_start + 2
        first_vals = rows[base_idx]

        for cols, _span in merge_groups or []:
            for col_idx in cols:
                _merge_vertical(ws, r_start, r_end, C + 1 + col_idx, first_vals[col_idx], styles)

        for j in range(3):
            r = r_start + j
            vals = rows[base_idx + j]
            for ci, v in enumerate(vals):
                if ci in merged_cols:
                    continue
                _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# Sheet 3: 渡槽
# ============================================================

def _write_aqueduct(ws, data, styles, gcl, col_offset=0):
    C = col_offset
    R1 = 1

    headers_full = [
        ("流量段",     None),
        ("设计流量",   "m³/s"),
        ("加大流量",   "m³/s"),
        ("1/底坡",     None),
        ("糙率",       None),
        ("半径R",      None),
        ("槽深H",      "m"),
        ("壁厚t",      "m"),
        ("设计水深H₁", "m"),
        ("加大水深H₂", "m"),
        ("设计流速",   "m/s"),
        ("高宽比",     None),
    ]
    col_widths_full = [14, 12, 12, 12, 10, 10, 10, 10, 13, 13, 12, 10]
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"], d.get("R", ""), d.get("H", ""), d.get("t", ""),
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
            d.get("HB_ratio", ""),
        ])
    headers, col_widths, rows, _merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
    )
    NCOLS = len(headers)

    _write_title(ws, R1, C + 1, C + NCOLS, "U形渡槽断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    for ri, vals in enumerate(rows):
        r = R1 + 3 + ri
        for ci, v in enumerate(vals):
            _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS

# 向后兼容别名
_write_aqueduct_u = _write_aqueduct


# ============================================================
# Sheet 3b: 矩形渡槽
# ============================================================

def _write_aqueduct_rect(ws, data, styles, gcl, col_offset=0):
    # 动态判断是否有倒角数据
    has_chamfer = any(d.get("chamfer_angle") for d in data)

    C = col_offset
    R1 = 1

    if has_chamfer:
        headers_full = [
            ("流量段",     None),
            ("设计流量",   "m³/s"),
            ("加大流量",   "m³/s"),
            ("1/底坡",     None),
            ("糙率",       None),
            ("底宽B",      "m"),
            ("槽深H",      "m"),
            ("壁厚t",      "m"),
            ("倒角角度",   "°"),
            ("倒角底边长", "m"),
            ("设计水深H₁", "m"),
            ("加大水深H₂", "m"),
            ("设计流速",   "m/s"),
        ]
        col_widths_full = [14, 12, 12, 12, 10, 10, 10, 10, 11, 11, 13, 13, 12]
    else:
        headers_full = [
            ("流量段",     None),
            ("设计流量",   "m³/s"),
            ("加大流量",   "m³/s"),
            ("1/底坡",     None),
            ("糙率",       None),
            ("底宽B",      "m"),
            ("槽深H",      "m"),
            ("壁厚t",      "m"),
            ("设计水深H₁", "m"),
            ("加大水深H₂", "m"),
            ("设计流速",   "m/s"),
        ]
        col_widths_full = [14, 12, 12, 12, 10, 10, 10, 10, 13, 13, 12]

    rows_full = []
    for d in data:
        vals = [d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
                f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
                d["n"], d.get("B", ""), d.get("H", ""), d.get("t", "")]
        if has_chamfer:
            vals += [d.get("chamfer_angle", ""), d.get("chamfer_length", "")]
        vals += [d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", "")]
        rows_full.append(vals)

    headers, col_widths, rows, _merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
    )
    NCOLS = len(headers)

    _write_title(ws, R1, C + 1, C + NCOLS, "矩形渡槽断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    for ri, vals in enumerate(rows):
        r = R1 + 3 + ri
        for ci, v in enumerate(vals):
            _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# Sheet 4: 矩形暗涵
# ============================================================

def _write_rect_culvert(ws, data, styles, gcl, col_offset=0):
    C = col_offset
    R1 = 1

    headers_full = [
        ("流量段",     None),
        ("设计流量",   "m³/s"),
        ("加大流量",   "m³/s"),
        ("1/底坡",     None),
        ("糙率",       None),
        ("底宽B",      "m"),
        ("高度H",      "m"),
        ("底板厚t₀",   "m"),
        ("边墙厚t₁",   "m"),
        ("顶板厚t₂",   "m"),
        ("设计水深H₁", "m"),
        ("加大水深H₂", "m"),
        ("设计流速",   "m/s"),
    ]
    col_widths_full = [14, 12, 12, 12, 10, 10, 10, 11, 11, 11, 13, 13, 12]
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"], d.get("B", ""), d.get("H", ""),
            d.get("t0", ""), d.get("t1", ""), d.get("t2", ""),
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, _merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
    )
    NCOLS = len(headers)

    _write_title(ws, R1, C + 1, C + NCOLS, "矩形暗涵断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    for ri, vals in enumerate(rows):
        r = R1 + 3 + ri
        for ci, v in enumerate(vals):
            _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


def _write_rect_culvert_arch(ws, data, styles, gcl, col_offset=0):
    C = col_offset
    R1 = 1

    headers_full = [
        ("流量段", None),
        ("设计流量", "m³/s"),
        ("加大流量", "m³/s"),
        ("1/底坡", None),
        ("糙率", None),
        ("底宽B", "m"),
        ("直墙高H", "m"),
        ("顶拱半径R", "m"),
        ("底板厚t₀", "m"),
        ("边墙顶拱厚t", "m"),
        ("设计水深H₁", "m"),
        ("加大水深H₂", "m"),
        ("设计流速", "m/s"),
    ]
    col_widths_full = [14, 12, 12, 12, 10, 10, 10, 12, 11, 13, 13, 13, 12]
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"], d.get("B", ""), d.get("H_straight", ""), d.get("R_arch", ""),
            d.get("t0", ""), d.get("t", ""),
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, _merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
    )
    NCOLS = len(headers)

    _write_title(ws, R1, C + 1, C + NCOLS, "圆拱直墙型暗涵断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    for ri, vals in enumerate(rows):
        r = R1 + 3 + ri
        for ci, v in enumerate(vals):
            _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# Sheet 5: 圆形明渠（圆管涵）
# ============================================================

def _write_circular_pipe(ws, data, styles, gcl, col_offset=0):
    C = col_offset
    R1 = 1
    columns = _circular_channel_column_defs(_open_channel_include_increase_columns(data))
    headers = _table_headers(columns)
    col_widths = _table_excel_widths(columns)
    NCOLS = len(columns)

    _write_title(ws, R1, C + 1, C + NCOLS, "圆管涵断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    for ri, d in enumerate(data):
        r = R1 + 3 + ri
        vals = _table_row_values(d, columns)
        for ci, v in enumerate(vals):
            _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# Sheet 6: 倒虹吸
# ============================================================

def _write_siphon(ws, data, styles, gcl, col_offset=0):
    C = col_offset
    NCOLS = 7
    R1 = 1

    headers = [
        ("倒虹吸名称及流量段", None),
        ("设计流量",   "m³/s"),
        ("加大流量",   "m³/s"),
        ("糙率",       None),
        ("直径DN",     "mm"),
        ("管道材质",   None),
        ("设计流速v",  "m/s"),
    ]
    col_widths = [14, 12, 12, 10, 12, 15, 12]

    _write_title(ws, R1, C + 1, C + NCOLS, "倒虹吸断面尺寸及水力要素表", styles)
    for i, (name, unit) in enumerate(headers):
        _write_header_2row(ws, R1 + 1, R1 + 2, C + 1 + i, name, unit, styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    for ri, d in enumerate(data):
        r = R1 + 3 + ri
        vals = [d["name"], d["Q"], d.get("Q_inc", ""),
                d["n"], d.get("DN_mm", ""), d.get("pipe_material", ""),
                d.get("V", "")]
        for ci, v in enumerate(vals):
            _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# Sheet 7: 有压管道
# ============================================================

def _write_pressure_pipe(ws, data, styles, gcl, col_offset=0):
    """有压管道 Excel 导出（新版特性表，保留隐藏计算结果但不显示）"""
    C = col_offset
    R1 = 1
    include_buildings = _pressure_pipe_show_building_characteristics(data)
    headers = _pressure_pipe_dxf_headers(include_buildings)
    col_widths = _pressure_pipe_excel_col_widths(include_buildings)
    NCOLS = len(headers)

    _write_title(ws, R1, C + 1, C + NCOLS, "压力管道特性表", styles)
    for i, w in enumerate(col_widths):
        _set_col_width(ws, C + 1 + i, w, gcl)

    for cell_def in _pressure_pipe_header_cells(include_buildings):
        _merge_header_cell(
            ws,
            R1 + 1 + int(cell_def["row"]),
            C + 1 + int(cell_def["col"]),
            R1 + int(cell_def["row"]) + int(cell_def.get("rowspan", 1)),
            C + int(cell_def["col"]) + int(cell_def.get("colspan", 1)),
            cell_def["text"],
            styles,
        )

    for ri, d in enumerate(data):
        r = R1 + 4 + ri
        vals = _pressure_pipe_row_values(d, include_buildings)
        for ci, v in enumerate(vals):
            _sc(ws, r, C + 1 + ci, v, styles)

    return NCOLS


# ============================================================
# 汇总 Sheet: 所有表格水平排列
# ============================================================

def _write_all_on_one_sheet(ws, tables, styles, gcl):
    """将多个表格水平排列在同一 Sheet（中间隔 1 列空白）"""
    offset = 0
    for writer, data in tables:
        ncols = writer(ws, data, styles, gcl, col_offset=offset)
        offset += ncols + 1  # 空 1 列


# ============================================================
# 主入口：生成 Excel
# ============================================================

def generate_excel(
    filepath: str,
    rect_channel_segs: List[Dict] = None,
    trap_channel_segs: List[Dict] = None,
    u_channel_segs: List[Dict] = None,
    tunnel_segs: List[Dict] = None,
    tunnel_arch_segs: List[Dict] = None,
    tunnel_circular_segs: List[Dict] = None,
    tunnel_flat_bottom_circular_segs: List[Dict] = None,
    tunnel_horseshoe_segs: List[Dict] = None,
    aqueduct_segs: List[Dict] = None,
    aqueduct_u_segs: List[Dict] = None,
    aqueduct_rect_segs: List[Dict] = None,
    rect_culvert_segs: List[Dict] = None,
    rect_culvert_arch_segs: List[Dict] = None,
    circular_pipe_segs: List[Dict] = None,
    siphon_segs: List[Dict] = None,
    siphon_material: str = "球墨铸铁管",
    pressure_pipe_segs: List[Dict] = None,
    pressure_pipe_material: str = "球墨铸铁管",
    rock_lining: Dict = None,
    table_order: List[str] = None,
    tunnel_unified_arch: bool = False,
    tunnel_unified_circular: bool = False,
    tunnel_unified_flat_bottom_circular: bool = False,
    tunnel_unified_horseshoe: bool = False,
) -> str:
    """
    生成包含多种断面汇总表的 Excel 文件（按实际类型动态生成）。

    参数:
        filepath: 保存路径
        *_segs:   各表流量段参数列表（None 则用默认值）
        tunnel_segs: 旧参数（向后兼容，等同于 tunnel_arch_segs）
        aqueduct_segs: 旧参数（向后兼容，等同于 aqueduct_u_segs）
        siphon_material: 倒虹吸管道材质
        pressure_pipe_material: 有压管道材质
        rock_lining: 隧洞围岩衬砸厚度
        tunnel_unified_arch: 圆拱直墙型隧洞是否统一断面
        tunnel_unified_circular: 圆形隧洞是否统一断面
        tunnel_unified_horseshoe: 马蹄形隧洞是否统一断面

    返回:
        保存的文件路径
    """
    openpyxl, styles, gcl = _get_openpyxl()

    if rect_channel_segs is None:
        rect_channel_segs = _default_segments_rect_channel()
    if trap_channel_segs is None:
        trap_channel_segs = []
    if u_channel_segs is None:
        u_channel_segs = []

    # 隧洞：向后兼容旧 tunnel_segs 参数 → tunnel_arch_segs
    if tunnel_arch_segs is None and tunnel_segs is not None:
        tunnel_arch_segs = tunnel_segs
    if tunnel_arch_segs is None:
        tunnel_arch_segs = _default_segments_tunnel_arch()
    if tunnel_circular_segs is None:
        tunnel_circular_segs = []
    if tunnel_flat_bottom_circular_segs is None:
        tunnel_flat_bottom_circular_segs = []
    if tunnel_horseshoe_segs is None:
        tunnel_horseshoe_segs = []

    # 渡槽：向后兼容旧 aqueduct_segs 参数 → aqueduct_u_segs
    if aqueduct_u_segs is None and aqueduct_segs is not None:
        aqueduct_u_segs = aqueduct_segs
    if aqueduct_u_segs is None:
        aqueduct_u_segs = _default_segments_aqueduct_u()
    if aqueduct_rect_segs is None:
        aqueduct_rect_segs = []

    if rect_culvert_segs is None:
        rect_culvert_segs = _default_segments_rect_culvert()
    if rect_culvert_arch_segs is None:
        rect_culvert_arch_segs = []
    if circular_pipe_segs is None:
        circular_pipe_segs = _default_segments_circular_pipe()
    if siphon_segs is None:
        siphon_segs = _default_segments_siphon()
    if pressure_pipe_segs is None:
        pressure_pipe_segs = []

    # ---- 计算 ----
    d1 = compute_rect_channel(rect_channel_segs) if rect_channel_segs else []
    d1b = compute_trapezoid_channel(trap_channel_segs) if trap_channel_segs else []
    d1c = compute_u_channel(u_channel_segs) if u_channel_segs else []
    d2_arch, _ = compute_tunnel(tunnel_arch_segs, rock_lining, unified=tunnel_unified_arch) if tunnel_arch_segs else ([], {})
    d2_circ, _ = compute_tunnel_circular(tunnel_circular_segs, rock_lining, unified=tunnel_unified_circular) if tunnel_circular_segs else ([], {})
    d2_flat_bottom, _ = compute_tunnel_flat_bottom_circular(
        tunnel_flat_bottom_circular_segs,
        rock_lining,
        unified=tunnel_unified_flat_bottom_circular,
    ) if tunnel_flat_bottom_circular_segs else ([], {})
    horseshoe_entries = _build_horseshoe_export_entries(
        tunnel_horseshoe_segs,
        rock_lining=rock_lining,
        unified=tunnel_unified_horseshoe,
    ) if tunnel_horseshoe_segs else []
    d3_u = compute_aqueduct_u(aqueduct_u_segs) if aqueduct_u_segs else []
    d3_rect = compute_aqueduct_rect(aqueduct_rect_segs) if aqueduct_rect_segs else []
    d4 = compute_rect_culvert(rect_culvert_segs) if rect_culvert_segs else []
    d4_arch = compute_rect_culvert_arch(rect_culvert_arch_segs) if rect_culvert_arch_segs else []
    d5 = compute_circular_pipe(circular_pipe_segs) if circular_pipe_segs else []
    d6 = compute_siphon(siphon_segs, siphon_material) if siphon_segs else []
    d7 = compute_pressure_pipe(pressure_pipe_segs, pressure_pipe_material) if pressure_pipe_segs else []

    wb = openpyxl.Workbook()

    tables_map = {
        "u_channel":        ("U形明渠",              _write_u_channel,          d1c),
        "rect_channel":     ("矩形明渠",          _write_rect_channel,       d1),
        "trap_channel":     ("梯形明渠",          _write_trapezoid_channel,  d1b),
        "tunnel_arch":      ("圆拱直墙型隧洞",    _write_tunnel,             d2_arch),
        "tunnel_circular":  ("圆形隧洞",          _write_tunnel_circular,    d2_circ),
        "tunnel_flat_bottom_circular": ("平底圆形隧洞", _write_tunnel_flat_bottom_circular, d2_flat_bottom),
        "aqueduct_u":       ("U形渡槽",           _write_aqueduct,           d3_u),
        "aqueduct_rect":    ("矩形渡槽",          _write_aqueduct_rect,      d3_rect),
        "rect_culvert":     ("矩形暗涵",          _write_rect_culvert,       d4),
        "rect_culvert_arch": ("圆拱直墙型暗涵",   _write_rect_culvert_arch,  d4_arch),
        "circular_channel": ("圆形明渠(圆管涵)",   _write_circular_pipe,      d5),
        "siphon":           ("倒虹吸",            _write_siphon,             d6),
        "pressure_pipe":    ("有压管道",          _write_pressure_pipe,      d7),
        # 向后兼容旧 key
        "tunnel":           ("圆拱直墙型隧洞",    _write_tunnel,             d2_arch),
        "aqueduct":         ("U形渡槽",           _write_aqueduct,           d3_u),
    }
    horseshoe_keys = [entry["key"] for entry in horseshoe_entries]
    for entry in horseshoe_entries:
        title = entry["title"]
        tables_map[entry["key"]] = (
            entry["sheet_name"],
            lambda ws, data, styles, gcl, col_offset=0, _title=title: _write_tunnel_horseshoe(
                ws,
                data,
                styles,
                gcl,
                col_offset=col_offset,
                title=_title,
            ),
            entry["rows"],
        )
    
    if not table_order:
        table_order = ["rect_channel", "trap_channel", "u_channel",
                       "tunnel_arch", "tunnel_circular", "tunnel_flat_bottom_circular", "tunnel_horseshoe",
                       "aqueduct_u", "aqueduct_rect",
                       "rect_culvert", "rect_culvert_arch", "circular_channel", "siphon", "pressure_pipe"]
    table_order = _expand_horseshoe_table_order(table_order, horseshoe_keys)

    tables = []
    for key in table_order:
        info = tables_map.get(key)
        if not info:
            continue
        sheet_name, writer, data = info
        if data:
            tables.append((sheet_name, writer, data))

    # 如果没有有效表，至少保留矩形明渠
    if not tables:
        tables.append(("矩形明渠", _write_rect_channel, d1))

    # ---- 独立 Sheet ----
    first_name, first_writer, first_data = tables[0]
    ws_default = wb.active
    ws_default.title = first_name
    first_writer(ws_default, first_data, styles, gcl)

    for sheet_name, writer, data in tables[1:]:
        ws = wb.create_sheet(sheet_name)
        writer(ws, data, styles, gcl)

    # ---- 汇总 Sheet（水平排列） ----
    ws_all = wb.create_sheet("汇总(并列)", 0)
    _write_all_on_one_sheet(ws_all, [(w, d) for _, w, d in tables], styles, gcl)

    wb.save(filepath)
    return filepath


# ============================================================
# DXF 导出 — 通用表格绘制引擎
# ============================================================

# DXF 表格样式常量（单位: mm）
_DXF_ROW_H       = 7.0    # 普通行高
_DXF_HDR_ROW_H   = 10.0   # 表头行高
_DXF_TITLE_ROW_H = 10.0   # 标题行高
_DXF_TEXT_H       = 3.5    # 数据文字高度
_DXF_HDR_TEXT_H   = 3.5    # 表头文字高度
_DXF_TITLE_TEXT_H = 5.0    # 标题文字高度
_DXF_COL_PAD      = 3.5    # 单元格左右合计留白(mm)
_DXF_TABLE_GAP    = 8.0    # 多表格之间的纵向间距


_DXF_WIDTH_FACTOR = 0.7   # 全局宽度因子（仿宋 标准）
_DXF_FONT_NAME   = "仿宋"  # DXF 文字样式字体（Unicode版，支持下标字符）

# Unicode 下标字符 → 普通字符映射
_SUBSCRIPT_MAP = str.maketrans(
    '₀₁₂₃₄₅₆₇₈₉', '0123456789'
)
_SUBSCRIPT_CHARS = set('₀₁₂₃₄₅₆₇₈₉')

# Unicode 上标字符 → 普通字符映射
_SUPERSCRIPT_MAP = str.maketrans(
    '¹²³⁰⁴⁵⁶⁷⁸⁹', '1230456789'
)
_SUPERSCRIPT_CHARS = set('¹²³⁰⁴⁵⁶⁷⁸⁹')

# 上下标合并映射（用于 _dxf_sanitize）
_SCRIPT_MAP = {**dict(zip('₀₁₂₃₄₅₆₇₈₉', '0123456789')),
               **dict(zip('¹²³⁰⁴⁵⁶⁷⁸⁹', '1230456789'))}
_ALL_SCRIPT_CHARS = _SUBSCRIPT_CHARS | _SUPERSCRIPT_CHARS
_SANITIZE_MAP = str.maketrans(_SCRIPT_MAP)


def _dxf_sanitize(text):
    """将文本中的 Unicode 上下标字符转为普通数字（用于宽度估算等场景）。"""
    if text is None:
        return text
    return str(text).translate(_SANITIZE_MAP)


def _has_scripts(text):
    """检测文本是否包含 Unicode 上标或下标字符。"""
    if text is None:
        return False
    return any(c in _ALL_SCRIPT_CHARS for c in str(text))


def _to_mtext_script(text):
    """将含 Unicode 上下标的文本转换为 MTEXT 堆叠格式。

    下标示例: 'H₁' → 'H{\\H0.7x;\\S^ 1;}'    (上标留空，下标 = 1)
    上标示例: 'm³' → 'm{\\H0.7x;\\S3^ ;}' (上标 = 3，下标留空)

    AutoCAD MTEXT 的 \\S 堆叠命令: \\S上标^下标;
    用 {\\H0.7x; ...} 分组缩小字号。
    """
    if text is None:
        return ""
    result = []
    s = str(text)
    i = 0
    while i < len(s):
        if s[i] in _SUBSCRIPT_CHARS:
            # 收集连续的下标字符
            sub_chars = []
            while i < len(s) and s[i] in _SUBSCRIPT_CHARS:
                sub_chars.append(s[i].translate(_SUBSCRIPT_MAP))
                i += 1
            sub_text = ''.join(sub_chars)
            # MTEXT 下标: {\H0.7x;\S^ 下标;}  (上标为空)
            result.append("{\\H0.7x;\\S^ " + sub_text + ";}")
        elif s[i] in _SUPERSCRIPT_CHARS:
            # 收集连续的上标字符
            sup_chars = []
            while i < len(s) and s[i] in _SUPERSCRIPT_CHARS:
                sup_chars.append(s[i].translate(_SUPERSCRIPT_MAP))
                i += 1
            sup_text = ''.join(sup_chars)
            # MTEXT 上标: {\H0.7x;\S上标^ ;}  (下标为空)
            result.append("{\\H0.7x;\\S" + sup_text + "^ ;}")
        else:
            result.append(s[i])
            i += 1
    return ''.join(result)


def _dxf_text_width(text, text_height):
    """估算 DXF 文字渲染宽度(mm)。
    中文字符宽度 ≈ text_height（仿宋方块字，em 宽 = 字高），
    ASCII 字符宽度 ≈ text_height × 0.6。
    注: 宽度因子仅影响 AutoCAD 渲染，不压缩估算值，以防列宽不足。
    """
    if text is None:
        return 0.0
    s = _dxf_sanitize(text)
    w = 0.0
    for ch in s:
        if ord(ch) > 0x7F:  # CJK / 全角
            w += text_height
        else:
            w += text_height * 0.6
    return w


def _dxf_auto_col_widths(headers, data_rows):
    """根据表头和数据内容自动计算每列宽度(mm)。"""
    ncols = len(headers)
    widths = [0.0] * ncols
    for ci, (name, unit) in enumerate(headers):
        w_name = _dxf_text_width(name, _DXF_HDR_TEXT_H)
        w_unit = _dxf_text_width(unit, _DXF_HDR_TEXT_H) if unit else 0.0
        widths[ci] = max(w_name, w_unit)
    for row in data_rows:
        for ci, val in enumerate(row):
            if ci >= ncols:
                break
            w = _dxf_text_width(val, _DXF_TEXT_H)
            if w > widths[ci]:
                widths[ci] = w
    return [w + _DXF_COL_PAD for w in widths]


def _dxf_draw_table(msp, origin_x, origin_y, title, headers, col_widths_mm,
                    data_rows, merge_groups=None, layer="TABLE"):
    """
    在 DXF modelspace 中绘制一个完整表格。

    参数:
        msp:           ezdxf modelspace
        origin_x/y:    表格左上角坐标（Y 向下为负）
        title:         标题文字
        headers:       [(name, unit), ...] — unit=None 时名称行与单位行合并
        col_widths_mm: [float, ...] 每列最小宽度(mm)，实际宽度按文字自适应
        data_rows:     [[val, val, ...], ...] 每行的单元格值列表
        merge_groups:  可选，[(col_indices, group_size), ...]
                       col_indices: 需要纵向合并的列索引列表
                       group_size:  每组合并的行数（如隧洞每段3行）
        layer:         DXF 图层名

    返回:
        表格总高度（正值，用于计算下一个表格的起始Y）
    """
    import ezdxf

    data_merge_groups = merge_groups
    header_row_count = 2
    header_cells = None
    if isinstance(merge_groups, dict):
        data_merge_groups = merge_groups.get("data_merge_groups")
        header_row_count = int(merge_groups.get("header_row_count", 2) or 2)
        header_cells = merge_groups.get("header_cells") or None

    # 自适应列宽：取内容估算宽度与传入最小宽度的较大值
    auto_widths = _dxf_auto_col_widths(headers, data_rows)
    col_widths_mm = [max(a, m) for a, m in zip(auto_widths, col_widths_mm)]

    ncols = len(col_widths_mm)
    nrows = len(data_rows)

    # 计算列的 X 坐标（累加）
    col_x = [origin_x]
    for w in col_widths_mm:
        col_x.append(col_x[-1] + w)
    total_w = col_x[-1] - col_x[0]

    # Y 坐标（向下为负）
    y_title_top = origin_y
    y_title_bot = y_title_top - _DXF_TITLE_ROW_H
    header_row_tops = []
    header_row_bottoms = []
    current_y = y_title_bot
    for _ in range(max(header_row_count, 1)):
        header_row_tops.append(current_y)
        current_y -= _DXF_HDR_ROW_H
        header_row_bottoms.append(current_y)
    y_data_top = header_row_bottoms[-1]

    # 各数据行的 Y 坐标
    row_y = [y_data_top]
    for _ in range(nrows):
        row_y.append(row_y[-1] - _DXF_ROW_H)

    total_h = y_title_top - row_y[-1]

    dxfattribs_line = {"layer": layer}
    x_left, x_right = col_x[0], col_x[-1]

    # ---- 构建合并信息查找表 ----
    # merged_cells[ri][ci] = (group_start_row, group_size) 如果该单元格被合并
    merged_cells = {}
    if data_merge_groups:
        for merge_cols, group_size in data_merge_groups:
            if group_size <= 1:
                continue
            num_groups = nrows // group_size
            for gi in range(num_groups):
                r_start = gi * group_size
                for ci in merge_cols:
                    for offset in range(group_size):
                        ri = r_start + offset
                        if ri not in merged_cells:
                            merged_cells[ri] = {}
                        merged_cells[ri][ci] = (r_start, group_size)

    # ---- 绘制标题行 ----
    msp.add_line((x_left, y_title_top), (x_right, y_title_top), dxfattribs=dxfattribs_line)
    msp.add_line((x_left, y_title_bot), (x_right, y_title_bot), dxfattribs=dxfattribs_line)
    msp.add_line((x_left, y_title_top), (x_left, y_title_bot), dxfattribs=dxfattribs_line)
    msp.add_line((x_right, y_title_top), (x_right, y_title_bot), dxfattribs=dxfattribs_line)

    if header_cells:
        header_matrix = [[None for _ in range(ncols)] for _ in range(header_row_count)]
        for idx, cell in enumerate(header_cells):
            row = int(cell["row"])
            col = int(cell["col"])
            rowspan = int(cell.get("rowspan", 1))
            colspan = int(cell.get("colspan", 1))
            for ri in range(row, row + rowspan):
                for ci in range(col, col + colspan):
                    if 0 <= ri < header_row_count and 0 <= ci < ncols:
                        header_matrix[ri][ci] = idx

        for boundary_idx in range(header_row_count - 1):
            y_line = header_row_bottoms[boundary_idx]
            seg_start = None
            for ci in range(ncols):
                same_cell = header_matrix[boundary_idx][ci] == header_matrix[boundary_idx + 1][ci]
                if same_cell:
                    if seg_start is not None:
                        msp.add_line(
                            (col_x[seg_start], y_line),
                            (col_x[ci], y_line),
                            dxfattribs=dxfattribs_line,
                        )
                        seg_start = None
                elif seg_start is None:
                    seg_start = ci
            if seg_start is not None:
                msp.add_line(
                    (col_x[seg_start], y_line),
                    (col_x[ncols], y_line),
                    dxfattribs=dxfattribs_line,
                )

        msp.add_line((x_left, y_data_top), (x_right, y_data_top), dxfattribs=dxfattribs_line)

        for ci in range(1, ncols):
            x_line = col_x[ci]
            for ri in range(header_row_count):
                if header_matrix[ri][ci - 1] == header_matrix[ri][ci]:
                    continue
                msp.add_line(
                    (x_line, header_row_tops[ri]),
                    (x_line, header_row_bottoms[ri]),
                    dxfattribs=dxfattribs_line,
                )
        msp.add_line((x_left, y_title_bot), (x_left, y_data_top), dxfattribs=dxfattribs_line)
        msp.add_line((x_right, y_title_bot), (x_right, y_data_top), dxfattribs=dxfattribs_line)
    else:
        # ---- 绘制表头区 ----
        # 表头区有两行：名称行 + 单位行
        # 无单位的列需要合并两行（不画中间水平线）
        hdr_merged_cols = set()
        for ci, (_name, unit) in enumerate(headers):
            if not unit:
                hdr_merged_cols.add(ci)

        y_hdr1_bot = header_row_bottoms[0]
        y_hdr2_bot = header_row_bottoms[1] if header_row_count > 1 else header_row_bottoms[0]

        for ci in range(ncols):
            if ci not in hdr_merged_cols:
                msp.add_line((col_x[ci], y_hdr1_bot), (col_x[ci + 1], y_hdr1_bot),
                             dxfattribs=dxfattribs_line)
        msp.add_line((x_left, y_hdr2_bot), (x_right, y_hdr2_bot), dxfattribs=dxfattribs_line)
        for x in col_x:
            msp.add_line((x, y_title_bot), (x, y_hdr2_bot), dxfattribs=dxfattribs_line)

    # ---- 绘制数据区水平线（分段画，跳过合并单元格） ----
    # 第一行顶线和最后一行底线是完整的
    msp.add_line((x_left, y_data_top), (x_right, y_data_top), dxfattribs=dxfattribs_line)
    if nrows > 0:
        msp.add_line((x_left, row_y[-1]), (x_right, row_y[-1]), dxfattribs=dxfattribs_line)

    # 中间行分隔线（ri=1..nrows-1），跳过合并区域内部
    for ri in range(1, nrows):
        y_line = row_y[ri]
        # 标记每列是否需要跳过（处于合并区域的非首行）
        skip_col = set()
        if ri in merged_cells:
            for ci, (r_start, gs) in merged_cells[ri].items():
                if ri != r_start:  # 不是合并组的首行 → 跳过
                    skip_col.add(ci)
        # 分段画水平线：连续的非跳过列画一条线
        seg_start = None
        for ci in range(ncols):
            if ci in skip_col:
                if seg_start is not None:
                    msp.add_line((col_x[seg_start], y_line), (col_x[ci], y_line),
                                 dxfattribs=dxfattribs_line)
                    seg_start = None
            else:
                if seg_start is None:
                    seg_start = ci
        if seg_start is not None:
            msp.add_line((col_x[seg_start], y_line), (col_x[ncols], y_line),
                         dxfattribs=dxfattribs_line)

    # ---- 绘制数据区竖线 ----
    y_bottom = row_y[-1] if nrows > 0 else y_data_top
    for x in col_x:
        msp.add_line((x, y_data_top), (x, y_bottom), dxfattribs=dxfattribs_line)

    # ---- 局部辅助: 写入单元格文字（含下标自动识别） ----
    def _add_cell_text(text_str, cx, cy, h):
        """含上下标字符时用 MTEXT 堆叠，否则用普通 TEXT。"""
        if _has_scripts(text_str):
            mt = msp.add_mtext(
                _to_mtext_script(text_str),
                dxfattribs={"layer": layer, "char_height": h, "style": "Standard"}
            )
            mt.set_location(insert=(cx, cy), attachment_point=5)  # MIDDLE_CENTER
        else:
            msp.add_text(
                _dxf_sanitize(str(text_str)),
                dxfattribs={"layer": layer, "height": h,
                            "width": _DXF_WIDTH_FACTOR, "style": "Standard"}
            ).set_placement((cx, cy), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)

    # ---- 写入标题文字 ----
    title_cx = x_left + total_w / 2
    title_cy = (y_title_top + y_title_bot) / 2
    _add_cell_text(title, title_cx, title_cy, _DXF_TITLE_TEXT_H)

    # ---- 写入表头文字 ----
    if header_cells:
        for cell in header_cells:
            row = int(cell["row"])
            col = int(cell["col"])
            rowspan = int(cell.get("rowspan", 1))
            colspan = int(cell.get("colspan", 1))
            cx = (col_x[col] + col_x[col + colspan]) / 2
            cy = (header_row_tops[row] + header_row_bottoms[row + rowspan - 1]) / 2
            _add_cell_text(str(cell["text"]), cx, cy, _DXF_HDR_TEXT_H)
    else:
        y_hdr1_bot = header_row_bottoms[0]
        y_hdr2_bot = header_row_bottoms[1] if header_row_count > 1 else header_row_bottoms[0]
        for ci, (name, unit) in enumerate(headers):
            cx = (col_x[ci] + col_x[ci + 1]) / 2
            if unit:
                cy1 = (y_title_bot + y_hdr1_bot) / 2
                _add_cell_text(name, cx, cy1, _DXF_HDR_TEXT_H)
                cy2 = (y_hdr1_bot + y_hdr2_bot) / 2
                _add_cell_text(unit, cx, cy2, _DXF_HDR_TEXT_H)
            else:
                cy = (y_title_bot + y_hdr2_bot) / 2
                _add_cell_text(name, cx, cy, _DXF_HDR_TEXT_H)

    # ---- 写入数据文字（考虑合并） ----
    written_merged = set()  # 已写入的合并单元格 (r_start, ci)
    for ri, row_vals in enumerate(data_rows):
        for ci, val in enumerate(row_vals):
            if val is None or val == "":
                continue
            cx = (col_x[ci] + col_x[ci + 1]) / 2

            # 检查是否在合并区域
            if ri in merged_cells and ci in merged_cells[ri]:
                r_start, gs = merged_cells[ri][ci]
                key = (r_start, ci)
                if key in written_merged:
                    continue  # 该合并区域已写过文字
                written_merged.add(key)
                # 在合并区域的垂直中心写文字
                r_end = r_start + gs - 1
                cy = (row_y[r_start] + row_y[r_end + 1]) / 2
                # 使用首行的值
                merge_val = data_rows[r_start][ci]
                if merge_val is None or merge_val == "":
                    continue
                _add_cell_text(str(merge_val), cx, cy, _DXF_TEXT_H)
            else:
                cy = (row_y[ri] + row_y[ri + 1]) / 2
                _add_cell_text(str(val), cx, cy, _DXF_TEXT_H)

    return total_h


def _dxf_col_widths(excel_widths):
    """将 Excel 列宽列表转换为 DXF mm 最小宽度（作为下限参考）"""
    return [w * 0.8 for w in excel_widths]


# ============================================================
# DXF 各表类型数据构建
# ============================================================

def _dxf_build_rect_channel(data):
    title = "矩形明渠断面尺寸及水力要素表"
    columns = _rect_channel_column_defs(_open_channel_include_increase_columns(data))
    headers = _table_headers(columns)
    col_widths = _table_dxf_widths(columns)
    rows = []
    for d in data:
        rows.append(_table_row_values(d, columns))
    return title, headers, col_widths, rows, None


def _dxf_build_trapezoid_channel(data):
    title = "梯形明渠断面尺寸及水力要素表"
    columns = _trapezoid_channel_column_defs(_open_channel_include_increase_columns(data))
    headers = _table_headers(columns)
    col_widths = _table_dxf_widths(columns)
    rows = []
    for d in data:
        rows.append(_table_row_values(d, columns))
    return title, headers, col_widths, rows, None


def _dxf_build_u_channel(data):
    title = "U形明渠断面尺寸及水力要素表"
    columns = _u_channel_column_defs(_open_channel_include_increase_columns(data))
    headers = _table_headers(columns)
    col_widths = _table_dxf_widths(columns)
    rows = []
    for d in data:
        rows.append(_table_row_values(d, columns))
    return title, headers, col_widths, rows, None


def _dxf_build_tunnel(data):
    title = "圆拱直墙型隧洞断面尺寸及水力要素表"
    headers_full = [
        ("流量段", None), ("设计流量", "m³/s"), ("加大流量", "m³/s"),
        ("围岩类型", None), ("1/底坡", None), ("糙率", None),
        ("底宽B", "m"), ("直墙高H", "m"), ("顶拱半径R", "m"),
        ("底板厚t₀", "m"), ("边墙顶拱厚t", "m"),
        ("设计水深H₁", "m"), ("加大水深H₂", "m"), ("设计流速", "m/s"),
    ]
    col_widths_full = _dxf_col_widths([12, 10, 10, 10, 10, 8, 8, 9, 11, 10, 12, 12, 12, 10])
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            d["rock_class"],
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"],
            d.get("B", ""), d.get("H_straight", ""), d.get("R_arch", ""),
            d["t0"], d["t"],
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
        merge_groups=[([0, 1, 2], 3)] if len(data) >= 3 else None,
    )
    return title, headers, col_widths, rows, merge


def _dxf_build_tunnel_circular(data):
    title = "圆形隧洞断面尺寸及水力要素表"
    headers_full = [
        ("流量段", None), ("设计流量", "m³/s"), ("加大流量", "m³/s"),
        ("围岩类型", None), ("1/底坡", None), ("糙率", None),
        ("直径D", "m"), ("底板厚t₀", "m"), ("衬砌厚t", "m"),
        ("设计水深H₁", "m"), ("加大水深H₂", "m"), ("设计流速", "m/s"),
    ]
    col_widths_full = _dxf_col_widths([12, 10, 10, 10, 10, 8, 8, 10, 10, 12, 12, 10])
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            d["rock_class"],
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"],
            d.get("D", ""), d["t0"], d["t"],
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
        merge_groups=[([0, 1, 2], 3)] if len(data) >= 3 else None,
    )
    return title, headers, col_widths, rows, merge


def _dxf_build_tunnel_flat_bottom_circular(data):
    title = "平底圆形隧洞断面尺寸及水力要素表"
    headers_full = [
        ("流量段", None), ("设计流量", "m³/s"), ("加大流量", "m³/s"),
        ("围岩类型", None), ("1/底坡", None), ("糙率", None),
        ("直径D", "m"), ("平底宽B", "m"), ("总高H", "m"),
        ("底板厚t₀", "m"), ("衬砌厚t", "m"),
        ("设计水深H₁", "m"), ("加大水深H₂", "m"), ("设计流速", "m/s"),
    ]
    col_widths_full = _dxf_col_widths([12, 10, 10, 10, 10, 8, 8, 8, 8, 10, 10, 12, 12, 10])
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            d["rock_class"],
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"],
            d.get("D", ""), d.get("B", ""), d.get("H_total", ""),
            d["t0"], d["t"],
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
        merge_groups=[([0, 1, 2], 3)] if len(data) >= 3 else None,
    )
    return title, headers, col_widths, rows, merge


def _dxf_build_tunnel_horseshoe(data, title=None):
    title = title or "马蹄形隧洞断面尺寸及水力要素表"
    headers_full = [
        ("流量段", None), ("设计流量", "m³/s"), ("加大流量", "m³/s"),
        ("围岩类型", None), ("1/底坡", None), ("糙率", None),
        ("半径R", "m"), ("底板厚t₀", "m"), ("衬砌厚t", "m"),
        ("设计水深H₁", "m"), ("加大水深H₂", "m"), ("设计流速", "m/s"),
    ]
    col_widths_full = _dxf_col_widths([12, 10, 10, 10, 10, 8, 8, 10, 10, 12, 12, 10])
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            d["rock_class"],
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"],
            d.get("R", ""), d["t0"], d["t"],
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
        merge_groups=[([0, 1, 2], 3)] if len(data) >= 3 else None,
    )
    return title, headers, col_widths, rows, merge


def _dxf_build_aqueduct_u(data):
    title = "U形渡槽断面尺寸及水力要素表"
    headers_full = [
        ("流量段", None), ("设计流量", "m³/s"), ("加大流量", "m³/s"),
        ("1/底坡", None), ("糙率", None), ("半径R", None),
        ("槽深H", "m"), ("壁厚t", "m"),
        ("设计水深H₁", "m"), ("加大水深H₂", "m"), ("设计流速", "m/s"),
        ("高宽比", None),
    ]
    col_widths_full = _dxf_col_widths([12, 10, 10, 10, 8, 8, 8, 8, 12, 12, 10, 8])
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"], d.get("R", ""), d.get("H", ""), d.get("t", ""),
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
            d.get("HB_ratio", ""),
        ])
    headers, col_widths, rows, _merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
    )
    return title, headers, col_widths, rows, None


def _dxf_build_aqueduct_rect(data):
    title = "矩形渡槽断面尺寸及水力要素表"
    headers_full = [
        ("流量段", None), ("设计流量", "m³/s"), ("加大流量", "m³/s"),
        ("1/底坡", None), ("糙率", None), ("底宽B", "m"),
        ("槽深H", "m"), ("壁厚t", "m"),
        ("设计水深H₁", "m"), ("加大水深H₂", "m"), ("设计流速", "m/s"),
    ]
    col_widths_full = _dxf_col_widths([12, 10, 10, 10, 8, 8, 8, 8, 12, 12, 10])
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"], d.get("B", ""), d.get("H", ""), d.get("t", ""),
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, _merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
    )
    return title, headers, col_widths, rows, None


def _dxf_build_rect_culvert(data):
    title = "矩形暗涵断面尺寸及水力要素表"
    headers_full = [
        ("流量段", None), ("设计流量", "m³/s"), ("加大流量", "m³/s"),
        ("1/底坡", None), ("糙率", None), ("底宽B", "m"),
        ("高度H", "m"), ("底板厚t₀", "m"), ("边墙厚t₁", "m"), ("顶板厚t₂", "m"),
        ("设计水深H₁", "m"), ("加大水深H₂", "m"), ("设计流速", "m/s"),
    ]
    col_widths_full = _dxf_col_widths([12, 10, 10, 10, 8, 8, 8, 10, 10, 10, 12, 12, 10])
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"], d.get("B", ""), d.get("H", ""),
            d.get("t0", ""), d.get("t1", ""), d.get("t2", ""),
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, _merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
    )
    return title, headers, col_widths, rows, None


def _dxf_build_rect_culvert_arch(data):
    title = "圆拱直墙型暗涵断面尺寸及水力要素表"
    headers_full = [
        ("流量段", None), ("设计流量", "m³/s"), ("加大流量", "m³/s"),
        ("1/底坡", None), ("糙率", None), ("底宽B", "m"),
        ("直墙高H", "m"), ("顶拱半径R", "m"),
        ("底板厚t₀", "m"), ("边墙顶拱厚t", "m"),
        ("设计水深H₁", "m"), ("加大水深H₂", "m"), ("设计流速", "m/s"),
    ]
    col_widths_full = _dxf_col_widths([12, 10, 10, 10, 8, 8, 8, 10, 10, 12, 12, 12, 10])
    rows_full = []
    for d in data:
        rows_full.append([
            d["name"], d["Q"], _open_channel_increase_value(d, "Q_inc"),
            f'1/{d["slope_inv"]:g}' if d.get("slope_inv") else "",
            d["n"], d.get("B", ""), d.get("H_straight", ""), d.get("R_arch", ""),
            d.get("t0", ""), d.get("t", ""),
            d.get("H1", ""), _open_channel_increase_value(d, "H2"), d.get("V", ""),
        ])
    headers, col_widths, rows, _merge = _filter_increase_columns(
        headers_full,
        col_widths_full,
        rows_full,
        include_increase=_open_channel_include_increase_columns(data),
    )
    return title, headers, col_widths, rows, None


def _dxf_build_circular_pipe(data):
    title = "圆管涵断面尺寸及水力要素表"
    columns = _circular_channel_column_defs(_open_channel_include_increase_columns(data))
    headers = _table_headers(columns)
    col_widths = _table_dxf_widths(columns)
    rows = []
    for d in data:
        rows.append(_table_row_values(d, columns))
    return title, headers, col_widths, rows, None


def _dxf_build_siphon(data):
    title = "倒虹吸断面尺寸及水力要素表"
    headers = [
        ("倒虹吸名称及流量段", None), ("设计流量", "m³/s"), ("加大流量", "m³/s"),
        ("糙率", None), ("直径DN", "mm"), ("管道材质", None),
        ("设计流速v", "m/s"),
    ]
    col_widths = _dxf_col_widths([12, 10, 10, 8, 10, 14, 10])
    rows = []
    for d in data:
        rows.append([
            d["name"], d["Q"], d.get("Q_inc", ""),
            d["n"], d.get("DN_mm", ""), d.get("pipe_material", ""),
            d.get("V", ""),
        ])
    return title, headers, col_widths, rows, None


def _dxf_build_pressure_pipe(data):
    """有压管道断面汇总表（新版特性表，保留隐藏计算结果但不显示）"""
    title = "压力管道特性表"
    include_buildings = _pressure_pipe_show_building_characteristics(data)
    headers = _pressure_pipe_dxf_headers(include_buildings)
    col_widths = _pressure_pipe_dxf_col_widths(include_buildings)
    rows = []
    for d in data:
        rows.append(_pressure_pipe_row_values(d, include_buildings))
    return title, headers, col_widths, rows, {
        "data_merge_groups": None,
        "header_row_count": 3,
        "header_cells": _pressure_pipe_header_cells(include_buildings),
    }


# DXF 构建函数映射
_DXF_BUILDERS = {
    "rect_channel":     _dxf_build_rect_channel,
    "trap_channel":     _dxf_build_trapezoid_channel,
    "u_channel":        _dxf_build_u_channel,
    "tunnel_arch":      _dxf_build_tunnel,
    "tunnel_circular":  _dxf_build_tunnel_circular,
    "tunnel_flat_bottom_circular": _dxf_build_tunnel_flat_bottom_circular,
    "tunnel_horseshoe": _dxf_build_tunnel_horseshoe,
    "tunnel_horseshoe_1": _dxf_build_tunnel_horseshoe,
    "tunnel_horseshoe_2": _dxf_build_tunnel_horseshoe,
    "aqueduct_u":       _dxf_build_aqueduct_u,
    "aqueduct_rect":    _dxf_build_aqueduct_rect,
    "rect_culvert":     _dxf_build_rect_culvert,
    "rect_culvert_arch": _dxf_build_rect_culvert_arch,
    "circular_channel": _dxf_build_circular_pipe,
    "siphon":           _dxf_build_siphon,
    "pressure_pipe":    _dxf_build_pressure_pipe,
    # 向后兼容
    "tunnel":           _dxf_build_tunnel,
    "aqueduct":         _dxf_build_aqueduct_u,
}


def generate_dxf(
    filepath: str,
    rect_channel_segs: List[Dict] = None,
    trap_channel_segs: List[Dict] = None,
    u_channel_segs: List[Dict] = None,
    tunnel_segs: List[Dict] = None,
    tunnel_arch_segs: List[Dict] = None,
    tunnel_circular_segs: List[Dict] = None,
    tunnel_flat_bottom_circular_segs: List[Dict] = None,
    tunnel_horseshoe_segs: List[Dict] = None,
    aqueduct_segs: List[Dict] = None,
    aqueduct_u_segs: List[Dict] = None,
    aqueduct_rect_segs: List[Dict] = None,
    rect_culvert_segs: List[Dict] = None,
    rect_culvert_arch_segs: List[Dict] = None,
    circular_pipe_segs: List[Dict] = None,
    siphon_segs: List[Dict] = None,
    siphon_material: str = "球墨铸铁管",
    pressure_pipe_segs: List[Dict] = None,
    pressure_pipe_material: str = "球墨铸铁管",
    rock_lining: Dict = None,
    table_order: List[str] = None,
    tunnel_unified_arch: bool = False,
    tunnel_unified_circular: bool = False,
    tunnel_unified_flat_bottom_circular: bool = False,
    tunnel_unified_horseshoe: bool = False,
) -> str:
    """
    生成包含多种断面汇总表的 DXF 文件。
    参数与 generate_excel() 完全一致。
    所有表格纵向排列在模型空间中。

    返回:
        保存的文件路径
    """
    import ezdxf

    # ---- 参数默认值处理（与 generate_excel 一致） ----
    if rect_channel_segs is None:
        rect_channel_segs = _default_segments_rect_channel()
    if trap_channel_segs is None:
        trap_channel_segs = []
    if u_channel_segs is None:
        u_channel_segs = []
    if tunnel_arch_segs is None and tunnel_segs is not None:
        tunnel_arch_segs = tunnel_segs
    if tunnel_arch_segs is None:
        tunnel_arch_segs = _default_segments_tunnel_arch()
    if tunnel_circular_segs is None:
        tunnel_circular_segs = []
    if tunnel_flat_bottom_circular_segs is None:
        tunnel_flat_bottom_circular_segs = []
    if tunnel_horseshoe_segs is None:
        tunnel_horseshoe_segs = []
    if aqueduct_u_segs is None and aqueduct_segs is not None:
        aqueduct_u_segs = aqueduct_segs
    if aqueduct_u_segs is None:
        aqueduct_u_segs = _default_segments_aqueduct_u()
    if aqueduct_rect_segs is None:
        aqueduct_rect_segs = []
    if rect_culvert_segs is None:
        rect_culvert_segs = _default_segments_rect_culvert()
    if rect_culvert_arch_segs is None:
        rect_culvert_arch_segs = []
    if circular_pipe_segs is None:
        circular_pipe_segs = _default_segments_circular_pipe()
    if siphon_segs is None:
        siphon_segs = _default_segments_siphon()
    if pressure_pipe_segs is None:
        pressure_pipe_segs = []

    # ---- 计算 ----
    d1 = compute_rect_channel(rect_channel_segs) if rect_channel_segs else []
    d1b = compute_trapezoid_channel(trap_channel_segs) if trap_channel_segs else []
    d1c = compute_u_channel(u_channel_segs) if u_channel_segs else []
    d2_arch, _ = compute_tunnel(tunnel_arch_segs, rock_lining, unified=tunnel_unified_arch) if tunnel_arch_segs else ([], {})
    d2_circ, _ = compute_tunnel_circular(tunnel_circular_segs, rock_lining, unified=tunnel_unified_circular) if tunnel_circular_segs else ([], {})
    d2_flat_bottom, _ = compute_tunnel_flat_bottom_circular(
        tunnel_flat_bottom_circular_segs,
        rock_lining,
        unified=tunnel_unified_flat_bottom_circular,
    ) if tunnel_flat_bottom_circular_segs else ([], {})
    horseshoe_entries = _build_horseshoe_export_entries(
        tunnel_horseshoe_segs,
        rock_lining=rock_lining,
        unified=tunnel_unified_horseshoe,
    ) if tunnel_horseshoe_segs else []
    d3_u = compute_aqueduct_u(aqueduct_u_segs) if aqueduct_u_segs else []
    d3_rect = compute_aqueduct_rect(aqueduct_rect_segs) if aqueduct_rect_segs else []
    d4 = compute_rect_culvert(rect_culvert_segs) if rect_culvert_segs else []
    d4_arch = compute_rect_culvert_arch(rect_culvert_arch_segs) if rect_culvert_arch_segs else []
    d5 = compute_circular_pipe(circular_pipe_segs) if circular_pipe_segs else []
    d6 = compute_siphon(siphon_segs, siphon_material) if siphon_segs else []
    d7 = compute_pressure_pipe(pressure_pipe_segs, pressure_pipe_material) if pressure_pipe_segs else []

    data_map = {
        "rect_channel":     d1,
        "trap_channel":     d1b,
        "u_channel":        d1c,
        "tunnel_arch":      d2_arch,
        "tunnel_circular":  d2_circ,
        "tunnel_flat_bottom_circular": d2_flat_bottom,
        "aqueduct_u":       d3_u,
        "aqueduct_rect":    d3_rect,
        "rect_culvert":     d4,
        "rect_culvert_arch": d4_arch,
        "circular_channel": d5,
        "siphon":           d6,
        "pressure_pipe":    d7,
        "tunnel":           d2_arch,
        "aqueduct":         d3_u,
    }
    horseshoe_keys = [entry["key"] for entry in horseshoe_entries]
    horseshoe_titles = {entry["key"]: entry["title"] for entry in horseshoe_entries}
    for entry in horseshoe_entries:
        data_map[entry["key"]] = entry["rows"]

    if not table_order:
        table_order = ["rect_channel", "trap_channel", "u_channel",
                       "tunnel_arch", "tunnel_circular", "tunnel_flat_bottom_circular", "tunnel_horseshoe",
                       "aqueduct_u", "aqueduct_rect",
                       "rect_culvert", "rect_culvert_arch", "circular_channel", "siphon", "pressure_pipe"]
    table_order = _expand_horseshoe_table_order(table_order, horseshoe_keys)

    # 收集有数据的表格
    tables = []
    for key in table_order:
        d = data_map.get(key)
        builder = _DXF_BUILDERS.get(key)
        if d and builder:
            tables.append((key, builder, d))

    if not tables:
        tables.append(("rect_channel", _dxf_build_rect_channel, d1))

    # ---- 创建 DXF 文件 ----
    doc = ezdxf.new("R2010")
    msp = doc.modelspace()

    # 设置中文字体样式：TrueType 仿宋_GB2312，宽度因子0.7
    if "Standard" in doc.styles:
        _sty = doc.styles.get("Standard")
    else:
        _sty = doc.styles.add("Standard")
    _sty.dxf.font = ""            # 清除 SHX 引用
    _sty.dxf.width = _DXF_WIDTH_FACTOR
    try:
        if "ACAD" not in doc.appids:
            doc.appids.new("ACAD")
    except Exception:
        pass
    _sty.set_xdata("ACAD", [(1000, _DXF_FONT_NAME), (1071, 0)])

    # 绘制各表格（纵向排列）
    current_y = 0.0
    for key, builder, d in tables:
        title, headers, col_widths, rows, merge = builder(d)
        if key in horseshoe_titles:
            title = horseshoe_titles[key]
        h = _dxf_draw_table(
            msp, 0.0, current_y,
            title, headers, col_widths, rows,
            merge_groups=merge, layer="TABLE"
        )
        current_y -= (h + _DXF_TABLE_GAP)

    doc.saveas(filepath)
    return filepath


# ============================================================
# Tkinter GUI 已移除，请使用 PySide6 版本：app_渠系计算前端/water_profile/cad_tools.py
# ============================================================
